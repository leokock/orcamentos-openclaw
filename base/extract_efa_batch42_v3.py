#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract esquadrias_detail, fachada_detail, acabamentos_pus from batch 42-83.
V3: Complete re-extraction for ALL 42 projects. Handles:
  - GDI alt Ger_Executivo (C=code, D=desc, E=unit, F=qty, G=pu, H=total)
  - Standard Ger_Executivo (H=level_type, I=code, J=desc, K=unit, L=qty, M=pu, N=total)
  - Kirchner CUSTOS DIRETOS (A=etapa, B=qty, C=unit, D=pu, E=total)
  - Sienge Relatório N2 (A=code, B=desc, C=total, D=%)
  - Sienge Relatório full (A=code, B=desc, C=un, D=qty, E=pu, F=total)
  - GMF-style Sienge export (Planilha1)
  - Muller Relatório (Planilha2 with embedded code)
  - MG3 BIM tabela sheets
  - Summary-only ORÇAMENTO_EXECUTIVO (etapa+valor)
  - Sub-category extraction from 2-level Relatório codes
"""
import json, sys, re, os, traceback
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

# Process ALL projects 42-83
BATCH = [(i, ALL_PROJECTS[i]) for i in range(42, 84)]


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    if v is None:
        return ''
    return str(v).replace('\xa0', ' ').strip()


def normalize(text):
    t = text.lower().strip()
    replacements = {
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'è': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'õ': 'o', 'ô': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c', 'ñ': 'n',
        '\xa0': ' ',
    }
    for old, new in replacements.items():
        t = t.replace(old, new)
    return t


# ============================================================
# Classification helpers
# ============================================================

def classify_esquadria(desc):
    d = normalize(desc)
    if any(x in d for x in ['ensaio', 'laudo', 'projeto de esquadria', 'consultoria',
                              'impermeabilizacao de esquadria', 'estanqueidade',
                              'impermeabilizacao de peitoris']):
        return None
    if any(x in d for x in ['pele de vidro', 'curtain wall', 'muro cortina', 'fachada de vidro',
                              'fachada envidracada', 'fachada vidro', 'vitrine']):
        return 'pele_vidro'
    if 'brise' in d:
        return 'brises'
    if ('guarda corpo' in d or 'guarda-corpo' in d or 'guardacorpo' in d):
        if 'ensaio' not in d and 'laudo' not in d and 'teste' not in d:
            return 'guarda_corpo'
    if 'corrimao' in d and 'madeira' in d:
        return 'portas_madeira'  # corrimão de madeira -> portas_madeira group
    if 'corta fogo' in d or 'corta-fogo' in d or 'cortafogo' in d:
        return 'portas_corta_fogo'
    if ('porta' in d and ('madeira' in d or 'semi oca' in d or 'semi-oca' in d or 'semioca' in d or 'mdf' in d)):
        return 'portas_madeira'
    if 'porta de madeira' in d:
        return 'portas_madeira'
    if 'fechadura' in d and ('eletronica' in d or 'digital' in d):
        return 'portas_madeira'
    # Aluminio
    if any(x in d for x in ['esquadrias de aluminio', 'esquadria de aluminio',
                              'janela de aluminio', 'janelas aluminio',
                              'esquadrias aluminio', 'esquadrias em aluminio']):
        return 'aluminio'
    if 'aluminio' in d and ('esquadria' in d or 'janela' in d):
        return 'aluminio'
    if 'aluminio' in d and 'porta' in d and 'madeira' not in d and 'corta' not in d:
        return 'aluminio'
    if 'contramarco' in d and ('aluminio' in d or 'argamassa' in d):
        return 'aluminio'
    if 'portao de aluminio' in d:
        return 'aluminio'
    # Esquadrias metálicas (iron doors/gates excl. corta-fogo)
    if 'esquadrias metalicas' in d or 'esquadria metalica' in d:
        return 'aluminio'  # group with aluminio for simplicity
    if ('portao' in d and ('ferro' in d or 'metalic' in d)):
        return 'aluminio'
    # Vidros
    if any(x in d for x in ['vidro temperado', 'vidro laminado', 'vidro insulado',
                              'envidracamento', 'vidros']):
        if 'guarda' not in d and 'pele' not in d:
            return 'vidros'
    return None


def classify_fachada(desc):
    d = normalize(desc)
    is_ext = any(x in d for x in ['externo', 'externa', 'fachada', 'exterior'])
    if 'reboco' in d and is_ext:
        return 'reboco_externo'
    if 'massa unica' in d and is_ext:
        return 'reboco_externo'
    if 'chapisco' in d and is_ext:
        return 'reboco_externo'
    if ('argamassado' in d or 'argamassa' in d) and is_ext:
        return 'reboco_externo'
    if 'requadro' in d and is_ext:
        return 'reboco_externo'
    if ('textura' in d or 'pintura' in d) and is_ext:
        return 'textura_pintura_ext'
    if 'grafiato' in d and is_ext:
        return 'textura_pintura_ext'
    if 'selador' in d and is_ext:
        return 'textura_pintura_ext'
    if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d or 'ceramic' in d) and is_ext:
        return 'pastilha_porcelanato'
    if 'acm' in d and is_ext:
        return 'pastilha_porcelanato'
    if 'concreto aparente' in d and is_ext:
        return 'pastilha_porcelanato'  # treatment counts as finish
    if 'junta' in d and is_ext:
        return 'juntas'
    return None


def classify_acabamento(desc, unit):
    d = normalize(desc)
    u = normalize(str(unit)) if unit else ''
    if 'mao de obra' in d or 'empreitada' in d:
        return None
    if 'rodape' in d:
        return 'rodape'
    if 'porcelanato' in d and u in ('m2', 'm²', ''):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'porcelanato'
    if ('ceramica' in d or 'ceramico' in d) and ('revestimento' in d or 'piso' in d or u in ('m2', 'm²')):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'pastilha' not in d:
                return 'ceramica'
    if 'laminado' in d and ('piso' in d or 'revestimento' in d or u in ('m2', 'm²')):
        return 'laminado'
    if 'contrapiso' in d and u in ('m2', 'm²'):
        return 'contrapiso'
    if 'forro' in d and ('gesso' in d or 'mineral' in d):
        return 'forro_gesso'
    if ('pintura' in d or 'tinta' in d) and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'piso' not in d:
                return 'pintura'
    if 'textura' in d and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'pintura'
    if 'chapisco' in d and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'chapisco'
    if ('reboco' in d or 'massa unica' in d) and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'reboco'
    return None


def classify_esquadria_subetapa(desc):
    """Classify esquadria sub-categories from subetapa (2-level) descriptions."""
    d = normalize(desc)
    if 'esquadria' in d and 'alumin' in d:
        return 'aluminio'
    if 'esquadria' in d and 'metalic' in d:
        return 'aluminio'
    if 'esquadria' in d and ('extern' in d):
        return 'aluminio'
    if 'esquadria' in d and ('intern' in d or 'madeira' in d or 'mdf' in d):
        return 'portas_madeira'
    if 'guarda corpo' in d or 'guarda-corpo' in d:
        return 'guarda_corpo'
    if 'brise' in d:
        return 'brises'
    if 'corta fogo' in d or 'corta-fogo' in d:
        return 'portas_corta_fogo'
    if 'vidro' in d and 'pele' not in d and 'esquadria' not in d and 'alumin' not in d:
        return 'vidros'
    if 'pele de vidro' in d or 'pele vidro' in d:
        return 'pele_vidro'
    return None


# ============================================================
# Accumulators
# ============================================================

def _add_esq(esq, cat, total, qty, pu, desc):
    if cat not in esq:
        esq[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0}
    esq[cat]['valor'] += total
    esq[cat]['qtd'] += qty
    if pu > 0 and qty > 0:
        esq[cat]['pu_sum'] += pu * qty
        esq[cat]['pu_count'] += qty


def _add_fach(fach, cat, total):
    if cat not in fach:
        fach[cat] = {'valor': 0.0}
    fach[cat]['valor'] += total


def _add_acab(acab, cat, pu, qty, desc, unit):
    if cat not in acab:
        acab[cat] = []
    acab[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})


# ============================================================
# Format-specific extractors
# ============================================================

def extract_ger_executivo_standard(wb, sheet_name):
    """Standard Ger_Executivo: H=level_type, I=code, J=desc, K=unit, L=qty, M=pu, N=total"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}
    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=20, values_only=True):
        if len(row) < 14:
            continue
        lt = safe_str(row[7] if row[7] else '')
        desc = safe_str(row[9] if row[9] else '')
        unit = safe_str(row[10] if row[10] else '')
        qty = safe_float(row[11])
        pu = safe_float(row[12])
        total = safe_float(row[13])
        if not desc:
            continue

        if lt == 'SERVIÇO':
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        if lt == 'SUBETAPA' and total > 0:
            d = normalize(desc)
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total
            # Fachada subetapas
            if ('reboco' in d or 'argamassado' in d or 'argamassa' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_ger_executivo_alt(wb, sheet_name):
    """GDI-style: C=code, D=desc, E=unit, F=qty, G=pu, H=total (0-indexed: 2,3,4,5,6,7)"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}
    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=12, values_only=True):
        if len(row) < 8:
            continue
        item_code = safe_str(row[2] if row[2] else '')
        desc = safe_str(row[3] if row[3] else '')
        unit = safe_str(row[4] if row[4] else '')
        qty = safe_float(row[5])
        pu = safe_float(row[6])
        total = safe_float(row[7])
        if not desc:
            continue

        parts = [p for p in item_code.split('.') if p.strip()]

        # Service level (3+ parts like "20.001.001" or "20.001.001.001")
        if len(parts) >= 3 and (total > 0 or (pu > 0 and qty > 0)):
            item_total = total if total > 0 else pu * qty
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, item_total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, item_total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        # Subetapa (2 parts like "20.001")
        if len(parts) == 2 and total > 0:
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total
            d = normalize(desc)
            if ('reboco' in d or 'argamassado' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d or 'acabamento' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_kirchner_custos_diretos(wb, sheet_name):
    """Kirchner: A=etapa (hierarchical 1.2.17.1...), B=qty, C=unit, D=pu, E=total.
    Service items have non-dot headers as descriptions."""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=10, values_only=True):
        if len(row) < 5:
            continue
        etapa_raw = safe_str(row[0] if row[0] else '')
        qty = safe_float(row[1])
        unit = safe_str(row[2] if row[2] else '')
        pu = safe_float(row[3])
        total = safe_float(row[4])

        if not etapa_raw:
            continue

        # Kirchner codes use "1.2.17.1.1." format
        # Strip trailing dots and split
        code_clean = etapa_raw.rstrip('.')
        parts = [p.strip() for p in code_clean.split('.') if p.strip()]

        # Determine hierarchy level - but Kirchner also has text descriptions
        # Text descriptions (not code-like) are actual service items
        is_code = all(p.replace(' ', '').isdigit() or len(p) < 5 for p in parts[:4]) and len(parts) >= 3

        if not is_code:
            # This is a service item description line
            desc = etapa_raw
            if total > 0 or (pu > 0 and qty > 0):
                item_total = total if total > 0 else pu * qty
                cat = classify_esquadria(desc)
                if cat:
                    _add_esq(esq, cat, item_total, qty, pu, desc)
                cat = classify_fachada(desc)
                if cat:
                    _add_fach(fach, cat, item_total)
                if pu > 0 and qty > 0:
                    cat = classify_acabamento(desc, unit)
                    if cat:
                        _add_acab(acab, cat, pu, qty, desc, unit)
        else:
            # Code line - check subetapa level for sub-categories
            # Get description from the text after the code
            desc = etapa_raw
            if total > 0 and len(parts) >= 4:
                # This is a subetapa with total - use for sub-classification
                subcat = classify_esquadria_subetapa(desc)
                if subcat:
                    esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                    esq[f'{subcat}_sub']['valor'] += total

    return esq, fach, acab


def extract_relatorio_sienge(wb, sheet_name):
    """Sienge Relatório: A=code, B=desc, C=total, D=%  or  A=code, B=desc, C=un, D=qty, E=pu, F=total"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    # Detect column layout from header row - scan wider range
    col_map = None
    header_row = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=35, values_only=True)):
        if not row:
            continue
        for ci, cell in enumerate(row[:10]):
            cs = normalize(safe_str(cell))
            if cs in ('codigo', 'código'):
                # Found header row
                header_row = i + 1
                un_idx = desc_idx = qty_idx = pu_idx = total_idx = pct_idx = None
                for cj, h in enumerate(row[:35]):
                    hs = normalize(safe_str(h))
                    if 'descri' in hs:
                        desc_idx = cj
                    elif hs in ('un.', 'un', 'und', 'unid', 'unidade'):
                        un_idx = cj
                    elif 'quantidade' in hs:
                        qty_idx = cj
                    elif 'unitario' in hs or ('unit' in hs and 'prec' in hs):
                        pu_idx = cj
                    elif 'total' in hs and ('prec' in hs or pct_idx is not None):
                        total_idx = cj
                    elif '% total' in hs or 'total' in hs:
                        if total_idx is None:
                            total_idx = cj
                        else:
                            pct_idx = cj
                col_map = {
                    'code': ci,
                    'desc': desc_idx or ci + 1,
                    'unit': un_idx,
                    'qty': qty_idx,
                    'pu': pu_idx,
                    'total': total_idx or ci + 2,
                }
                break
        if col_map:
            break

    if not col_map:
        col_map = {'code': 0, 'desc': 1, 'unit': None, 'qty': None, 'pu': None, 'total': 2}

    rel_max_col = max(c for c in [col_map['code'], col_map['desc'],
                                   col_map.get('unit'), col_map.get('qty'),
                                   col_map.get('pu'), col_map['total']]
                      if c is not None) + 2

    for row in ws.iter_rows(min_row=header_row + 1, max_row=10000, max_col=rel_max_col, values_only=True):
        if not row:
            continue
        code = safe_str(row[col_map['code']] if col_map['code'] < len(row) else '')
        desc = safe_str(row[col_map['desc']] if col_map['desc'] < len(row) else '')
        unit = safe_str(row[col_map['unit']] if col_map['unit'] is not None and col_map['unit'] < len(row) else '')
        qty = safe_float(row[col_map['qty']] if col_map['qty'] is not None and col_map['qty'] < len(row) else 0)
        pu = safe_float(row[col_map['pu']] if col_map['pu'] is not None and col_map['pu'] < len(row) else 0)
        total = safe_float(row[col_map['total']] if col_map['total'] < len(row) else 0)

        if not desc and not code:
            continue

        code_parts = [p for p in code.strip().split('.') if p.strip()]

        # Service level (4+ parts or 3+ parts depending on format)
        is_service = len(code_parts) >= 4 or (len(code_parts) == 3 and pu > 0)
        is_subetapa = len(code_parts) in (2, 3) and total > 0

        if is_service and (total > 0 or (pu > 0 and qty > 0)):
            item_total = total if total > 0 else pu * qty
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, item_total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, item_total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        if is_subetapa:
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total
            d = normalize(desc)
            if ('reboco' in d or 'argamassado' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_relatorio_embedded(wb, sheet_name, code_col=10, desc_col=11, total_col=12):
    """Embedded Relatório in Gerenciamento_Exec (cols K-N typically)"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=max(code_col, desc_col, total_col) + 2, values_only=True):
        if len(row) <= total_col:
            continue
        code = safe_str(row[code_col] if row[code_col] else '')
        desc = safe_str(row[desc_col] if row[desc_col] else '')
        total = safe_float(row[total_col] if row[total_col] else 0)

        if not code and not desc:
            continue

        code_parts = [p for p in code.strip().split('.') if p.strip()]

        # 2-level (subetapas like "09", "10")
        if len(code_parts) == 1 and total > 0:
            d = normalize(desc)
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total
            if ('revestimento' in d or 'acabamento' in d) and ('externo' in d or 'fachada' in d):
                fach.setdefault('_fachada_sub_total', 0.0)
                fach['_fachada_sub_total'] += total

        # 2-level like "18.001"
        if len(code_parts) == 2 and total > 0:
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total

    return esq, fach, acab


def extract_sienge_planilha(wb, sheet_name):
    """Sienge full planilha export: variable format, detect code column.
    Handles wide layouts (GMF: code=A, desc=B, un=N, qty=Q, pu=W, total=col28)."""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    # Detect layout by finding header with Código/Descrição - scan wider columns
    col_map = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, max_col=35, values_only=True)):
        if not row:
            continue
        for ci, cell in enumerate(row[:15]):
            cs = normalize(safe_str(cell))
            if cs in ('codigo', 'código'):
                un_idx = desc_idx = qty_idx = pu_idx = total_idx = None
                for cj, h in enumerate(row[:35]):
                    hs = normalize(safe_str(h))
                    if 'descri' in hs:
                        desc_idx = cj
                    elif hs in ('un.', 'un', 'und', 'unid', 'unidade'):
                        un_idx = cj
                    elif 'quantidade' in hs or 'quant' in hs:
                        qty_idx = cj
                    elif 'unitario' in hs or ('unit' in hs and 'prec' in hs):
                        pu_idx = cj
                    elif 'total' in hs and 'prec' in hs:
                        total_idx = cj
                    elif total_idx is None and 'total' in hs:
                        total_idx = cj
                col_map = {
                    'code': ci,
                    'desc': desc_idx or ci + 1,
                    'unit': un_idx,
                    'qty': qty_idx,
                    'pu': pu_idx,
                    'total': total_idx or ci + 2,
                    'header_row': i + 1,
                }
                break
        if col_map:
            break

    if not col_map:
        return esq, fach, acab

    max_needed_col = max(c for c in [col_map['code'], col_map['desc'],
                                      col_map.get('unit'), col_map.get('qty'),
                                      col_map.get('pu'), col_map['total']]
                         if c is not None) + 2

    for row in ws.iter_rows(min_row=col_map['header_row'] + 1, max_row=10000, max_col=max_needed_col, values_only=True):
        if not row:
            continue
        code = safe_str(row[col_map['code']] if col_map['code'] < len(row) else '')
        desc = safe_str(row[col_map['desc']] if col_map['desc'] < len(row) else '')
        unit = safe_str(row[col_map['unit']] if col_map['unit'] is not None and col_map['unit'] < len(row) else '')
        qty = safe_float(row[col_map['qty']] if col_map['qty'] is not None and col_map['qty'] < len(row) else 0)
        pu = safe_float(row[col_map['pu']] if col_map['pu'] is not None and col_map['pu'] < len(row) else 0)
        total = safe_float(row[col_map['total']] if col_map['total'] < len(row) else 0)

        if not desc and not code:
            continue

        code_parts = [p for p in code.strip().split('.') if p.strip()]
        is_service = len(code_parts) >= 4 or (len(code_parts) >= 3 and qty > 0 and pu > 0)

        if is_service and (total > 0 or (pu > 0 and qty > 0)):
            item_total = total if total > 0 else pu * qty
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, item_total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, item_total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        # Subetapa level
        is_subetapa = len(code_parts) in (2, 3) and total > 0
        if is_subetapa:
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total
            d = normalize(desc)
            if ('reboco' in d or 'argamassado' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_summary_etapas(wb, sheet_name):
    """Extract esquadrias/fachada totals from summary sheet (ETAPA + VALOR)."""
    ws = wb[sheet_name]
    esq, fach = {}, {}

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        if len(row) < 2:
            continue
        desc = safe_str(row[0] if row[0] else '')
        total = safe_float(row[1])

        if not desc or total <= 0:
            continue

        d = normalize(desc)
        if 'esquadria' in d or 'vidros e ferragens' in d:
            esq['_total'] = esq.get('_total', 0) + total
        if 'fachada' in d:
            fach['_total'] = fach.get('_total', 0) + total

    return esq, fach


def extract_muller_planilha2_relatorio(wb, sheet_name):
    """Muller Planilha2 embedded Relatório (code in col C, desc in col E, total in col G)"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=15, values_only=True):
        if len(row) < 7:
            continue
        # Muller format: col C (idx 2) = code, col E (idx 4) = desc, col G (idx 6) = total
        code = safe_str(row[2] if row[2] else '')
        desc = safe_str(row[4] if row[4] else '')
        total = safe_float(row[6])

        if not code and not desc:
            continue

        code_parts = [p for p in code.strip().split('.') if p.strip()]

        # Service level
        if len(code_parts) >= 3 and total > 0:
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, total, 0, 0, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, total)

        # Subetapa level
        if len(code_parts) == 2 and total > 0:
            subcat = classify_esquadria_subetapa(desc)
            if subcat:
                esq.setdefault(f'{subcat}_sub', {'valor': 0.0})
                esq[f'{subcat}_sub']['valor'] += total

        # Single code = etapa summary
        if len(code_parts) == 1 and total > 0:
            d = normalize(desc)
            if 'esquadria' in d:
                esq['_total_esquadrias'] = esq.get('_total_esquadrias', 0) + total
            if 'fachada' in d:
                fach['_total_fachada'] = fach.get('_total_fachada', 0) + total

    return esq, fach, acab


def extract_mg3_esquadrias(wb):
    """Extract from MG3 BIM tabela sheets."""
    esq = {}
    sheets = wb.sheetnames

    for sname in sheets:
        sn = normalize(sname)

        if 'tabela de portas' in sn:
            ws = wb[sname]
            total_madeira = 0
            total_pcf = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                desc = safe_str(row[1] if len(row) > 1 and row[1] else '')
                qty = safe_float(row[4] if len(row) > 4 else 0)
                if qty <= 0:
                    continue
                d = normalize(desc)
                if 'corta-fogo' in d or 'corta fogo' in d:
                    total_pcf += qty
                elif 'mdf' in d or 'madeira' in d or d.startswith('p'):
                    total_madeira += qty
            if total_madeira > 0:
                esq['portas_madeira'] = {'valor': None, 'qtd': round(total_madeira, 2)}
            if total_pcf > 0:
                esq['portas_corta_fogo'] = {'valor': None, 'qtd': round(total_pcf, 2)}

        if 'tabela de janelas' in sn:
            ws = wb[sname]
            total_janelas = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                qty = safe_float(row[5] if len(row) > 5 else 0)
                if qty > 0:
                    code = safe_str(row[0] if row[0] else '')
                    if code and code[0] in ('J', 'G', 'V', 'C', 'P') and code != 'Marca de tipo':
                        total_janelas += qty
            if total_janelas > 0:
                esq.setdefault('aluminio', {'valor': None, 'qtd': 0})
                esq['aluminio']['qtd'] = esq['aluminio'].get('qtd', 0) + round(total_janelas, 2)

        if 'tabela de brises' in sn:
            ws = wb[sname]
            total_brises = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                qty = safe_float(row[5] if len(row) > 5 else 0)
                if qty > 0:
                    code = safe_str(row[0] if row[0] else '')
                    if code and not code.startswith('Marca') and not code.startswith('Código'):
                        total_brises += qty
            if total_brises > 0:
                esq['brises'] = {'valor': None, 'qtd': round(total_brises, 2)}

        if 'tabela de painel de vidro' in sn:
            ws = wb[sname]
            total_pele = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                qty = safe_float(row[5] if len(row) > 5 else 0)
                if qty > 0:
                    code = safe_str(row[0] if row[0] else '')
                    if code and not code.startswith('Marca') and not code.startswith('Código'):
                        total_pele += qty
            if total_pele > 0:
                esq['pele_vidro'] = {'valor': None, 'qtd': round(total_pele, 2)}

    return esq


def extract_mg3_fachada(wb):
    """Extract from MG3 B05. TABELA REVEST FACHADAS"""
    fach = {}
    ws = None
    for sname in wb.sheetnames:
        if 'tabela revest fachada' in normalize(sname):
            ws = wb[sname]
            break
    if not ws:
        return fach

    total_area = 0
    for row in ws.iter_rows(min_row=1, max_row=1000, max_col=10, values_only=True):
        area = safe_float(row[3] if len(row) > 3 else 0)
        if area > 0:
            total_area += area
    if total_area > 0:
        fach['textura_pintura_ext'] = {'valor': None, 'qtd_m2': round(total_area, 2)}
    return fach


# ============================================================
# Build final output
# ============================================================

def build_esquadrias_detail(raw):
    result = {}
    categories = ['aluminio', 'guarda_corpo', 'portas_madeira', 'portas_corta_fogo',
                   'vidros', 'pele_vidro', 'brises']
    for cat in categories:
        if cat in raw:
            data = raw[cat]
            entry = {}
            if isinstance(data, dict):
                val = data.get('valor')
                if val is not None and val > 0:
                    entry['valor'] = round(val, 2)
                elif val is None:
                    pass  # No valor but has qtd
                qtd = data.get('qtd') or 0
                if qtd > 0:
                    entry['qtd'] = round(qtd, 2)
                pu_count = data.get('pu_count') or 0
                if pu_count > 0:
                    entry['pu'] = round((data.get('pu_sum') or 0) / pu_count, 2)
                qtd_m2 = data.get('qtd_m2') or 0
                if qtd_m2 > 0:
                    entry['qtd_m2'] = round(qtd_m2, 2)
            if entry:
                result[cat] = entry
        elif f'{cat}_sub' in raw:
            data = raw[f'{cat}_sub']
            if isinstance(data, dict) and data.get('valor', 0) > 0:
                result[cat] = {'valor': round(data['valor'], 2)}
    if '_total_esquadrias' in raw and not result:
        result['_total'] = round(raw['_total_esquadrias'], 2)
    return result if result else None


def build_fachada_detail(raw):
    result = {}
    categories = ['reboco_externo', 'textura_pintura_ext', 'pastilha_porcelanato', 'juntas']
    for cat in categories:
        if cat in raw:
            entry = {}
            val = raw[cat].get('valor', 0) or 0
            if val > 0:
                entry['valor'] = round(val, 2)
            qty = raw[cat].get('qtd', 0) or 0
            if qty > 0:
                entry['qtd_m2'] = round(qty, 2)
            qtm = raw[cat].get('qtd_m2', 0) or 0
            if qtm > 0:
                entry['qtd_m2'] = round(qtm, 2)
            if entry:
                result[cat] = entry
        elif f'{cat}_sub' in raw:
            data = raw[f'{cat}_sub']
            val = data.get('valor', 0) or 0 if isinstance(data, dict) else 0
            if val > 0:
                result[cat] = {'valor': round(val, 2)}
    if '_total_fachada' in raw and not result:
        result['_total'] = round(raw['_total_fachada'], 2)
    if '_brises_m2' in raw:
        result['_brises_m2'] = raw['_brises_m2']
    return result if result else None


def build_acabamentos_pus(raw):
    result = {}
    for cat, items in raw.items():
        if not items:
            continue
        total_weighted = sum(i['pu'] * i['qty'] for i in items)
        total_qty = sum(i['qty'] for i in items)
        if total_qty > 0:
            result[cat] = round(total_weighted / total_qty, 2)
    return result if result else None


# ============================================================
# Merge helpers
# ============================================================

def _merge(target, source):
    for k, v in source.items():
        if k not in target:
            target[k] = v
        elif isinstance(v, dict) and isinstance(target[k], dict):
            for dk, dv in v.items():
                if dk in target[k] and isinstance(dv, (int, float)):
                    target[k][dk] = target[k].get(dk, 0) + dv
                elif dk not in target[k]:
                    target[k][dk] = dv


def _merge_fach(target, source):
    for k, v in source.items():
        if k not in target:
            target[k] = v
        elif isinstance(v, dict) and isinstance(target[k], dict):
            for dk, dv in v.items():
                if dk in target[k] and isinstance(dv, (int, float)):
                    target[k][dk] = target[k].get(dk, 0) + dv
                elif dk not in target[k]:
                    target[k][dk] = dv
        elif isinstance(v, (int, float)):
            target[k] = max(target.get(k, 0), v)


def _merge_acab(target, source):
    for k, v in source.items():
        if k not in target:
            target[k] = v
        else:
            target[k].extend(v)


# ============================================================
# Process each project
# ============================================================

def detect_ger_exec_format(wb, sheet_name):
    """Detect whether Ger_Executivo uses standard (H=type) or alt (C=code) format."""
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=1, max_row=50, max_col=20, values_only=True):
        if len(row) > 7:
            lt = safe_str(row[7] if row[7] else '')
            if lt in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
                return 'standard'
        # Check for code pattern in col C (index 2)
        if len(row) > 2:
            code = safe_str(row[2] if row[2] else '')
            parts = [p for p in code.split('.') if p.strip()]
            if len(parts) >= 3 and all(p.strip().isdigit() or len(p.strip()) <= 4 for p in parts):
                return 'alt'
    return 'unknown'


def process_project(idx, proj):
    slug = proj['slug']
    path = proj['path']

    if not os.path.exists(path):
        return slug, None, None, None, f"File not found"

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return slug, None, None, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    all_esq, all_fach, all_acab = {}, {}, {}
    methods = []

    try:
        # === Strategy 1: Ger_Executivo (standard or alt) ===
        for sname in sheets:
            sn_lower = sname.lower().replace(' ', '_').strip('_')
            if sn_lower.startswith('ger_executivo') or sn_lower == 'ger_executivo':
                fmt = detect_ger_exec_format(wb, sname)
                if fmt == 'standard':
                    e, f, a = extract_ger_executivo_standard(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                    if e or f or a:
                        methods.append(f'GerExec-std:{sname}')
                elif fmt == 'alt':
                    e, f, a = extract_ger_executivo_alt(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                    if e or f or a:
                        methods.append(f'GerExec-alt:{sname}')

        # === Strategy 2: Relatório sheets ===
        for sname in sheets:
            if 'relat' in normalize(sname):
                e, f, a = extract_relatorio_sienge(wb, sname)
                _merge(all_esq, e)
                _merge_fach(all_fach, f)
                _merge_acab(all_acab, a)
                if e or f or a:
                    methods.append(f'Relatorio:{sname}')

        # === Strategy 3: ORÇAMENTO_EXECUTIVO / similar summary sheets ===
        for sname in sheets:
            sn = normalize(sname)
            if 'orcamento_executivo' in sn or 'orcamento executivo' in sn:
                # Check if detail or summary
                ws = wb[sname]
                is_detail = False
                for row in ws.iter_rows(min_row=1, max_row=20, max_col=15, values_only=True):
                    if row and len(row) > 7 and row[7]:
                        lt = safe_str(row[7])
                        if lt in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
                            is_detail = True
                            break
                if is_detail:
                    e, f, a = extract_ger_executivo_standard(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                    if e or f or a:
                        methods.append(f'OrcExec-detail:{sname}')
                else:
                    e2, f2 = extract_summary_etapas(wb, sname)
                    _merge(all_esq, e2)
                    _merge_fach(all_fach, f2)
                    if e2 or f2:
                        methods.append(f'OrcExec-summary:{sname}')

        # === Strategy 4: Gerenciamento executivo (with trailing space) ===
        for sname in sheets:
            sn = sname.strip().lower()
            if sn in ('gerenciamento executivo', 'executivo'):
                ws = wb[sname]
                is_detail = False
                for row in ws.iter_rows(min_row=1, max_row=20, max_col=15, values_only=True):
                    if row and len(row) > 7 and row[7]:
                        lt = safe_str(row[7])
                        if lt in ('SERVIÇO', 'SUBETAPA', 'ETAPA'):
                            is_detail = True
                            break
                if is_detail:
                    e, f, a = extract_ger_executivo_standard(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                    if e or f or a:
                        methods.append(f'GerExecAlt:{sname}')
                else:
                    e2, f2 = extract_summary_etapas(wb, sname)
                    _merge(all_esq, e2)
                    _merge_fach(all_fach, f2)
                    if e2 or f2:
                        methods.append(f'Summary:{sname}')

        # === Strategy 5: Gerenciamento_Exec with embedded Relatório ===
        for sname in sheets:
            if 'gerenciamento_exec' in sname.lower().strip():
                # Check for embedded Relatório in columns K-N
                ws = wb[sname]
                has_embedded = False
                for row in ws.iter_rows(min_row=1, max_row=5, max_col=15, values_only=True):
                    if len(row) > 10 and row[10]:
                        cs = safe_str(row[10])
                        if cs and (cs.strip().isdigit() or 'Código' in cs):
                            has_embedded = True
                            break
                if has_embedded:
                    e, f, a = extract_relatorio_embedded(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                    if e or f or a:
                        methods.append(f'EmbeddedRel:{sname}')

                # Also extract summary from cols A-B
                e2, f2 = extract_summary_etapas(wb, sname)
                _merge(all_esq, e2)
                _merge_fach(all_fach, f2)

        # === Strategy 6: Kirchner CUSTOS DIRETOS ===
        if 'CUSTOS DIRETOS' in sheets:
            e, f, a = extract_kirchner_custos_diretos(wb, 'CUSTOS DIRETOS')
            _merge(all_esq, e)
            _merge_fach(all_fach, f)
            _merge_acab(all_acab, a)
            if e or f or a:
                methods.append('Kirchner:CUSTOS DIRETOS')

        # === Strategy 7: Sienge Planilha exports ===
        for sname in sheets:
            sn = normalize(sname)
            if sn in ('planilha1', 'planilha2', 'sheet1'):
                e, f, a = extract_sienge_planilha(wb, sname)
                _merge(all_esq, e)
                _merge_fach(all_fach, f)
                _merge_acab(all_acab, a)
                if e or f or a:
                    methods.append(f'Planilha:{sname}')

                # Also try Muller format for Planilha2
                if 'planilha2' in sn:
                    e2, f2, a2 = extract_muller_planilha2_relatorio(wb, sname)
                    _merge(all_esq, e2)
                    _merge_fach(all_fach, f2)
                    _merge_acab(all_acab, a2)

        # === Strategy 8: MG3 tabela sheets ===
        has_mg3 = any('tabela de portas' in normalize(s) for s in sheets)
        if has_mg3:
            e = extract_mg3_esquadrias(wb)
            _merge(all_esq, e)
            if e:
                methods.append('MG3:esquadrias')
            f = extract_mg3_fachada(wb)
            _merge_fach(all_fach, f)
            if f:
                methods.append('MG3:fachada')

        # === Strategy 9: Orçamento Resumo / Resumo (summary fallback) ===
        if not all_esq and not all_fach:
            for sname in sheets:
                sn = normalize(sname)
                if sn in ('orcamento resumo', 'resumo', 'resumo geral'):
                    e, f = extract_summary_etapas(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    if e or f:
                        methods.append(f'Resumo:{sname}')

    except Exception as ex:
        wb.close()
        return slug, None, None, None, f"Error: {traceback.format_exc()}"

    wb.close()

    esq_detail = build_esquadrias_detail(all_esq)
    fach_detail = build_fachada_detail(all_fach)
    acab_pus = build_acabamentos_pus(all_acab)

    return slug, esq_detail, fach_detail, acab_pus, None, methods


def update_json(slug, esq, fach, acab, methods):
    idx_path = INDICES_DIR / f"{slug}.json"
    if idx_path.exists():
        with open(idx_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {'slug': slug}

    updated = False

    if esq is not None:
        existing = data.get('esquadrias_detail', {})
        if existing:
            existing_cats = [k for k in existing.keys() if not k.startswith('_')]
            new_cats = [k for k in esq.keys() if not k.startswith('_')]
            # Replace if new has more categories, or existing only has _total
            if not existing_cats and new_cats:
                data['esquadrias_detail'] = esq
                updated = True
            elif len(new_cats) > len(existing_cats):
                data['esquadrias_detail'] = esq
                updated = True
            else:
                # Merge: add new keys, keep existing
                for k, v in esq.items():
                    if k not in existing:
                        existing[k] = v
                        updated = True
                data['esquadrias_detail'] = existing
        else:
            data['esquadrias_detail'] = esq
            updated = True

    if fach is not None:
        existing = data.get('fachada_detail', {})
        if existing:
            existing_cats = [k for k in existing.keys() if not k.startswith('_')]
            new_cats = [k for k in fach.keys() if not k.startswith('_')]
            if not existing_cats and new_cats:
                data['fachada_detail'] = fach
                updated = True
            elif len(new_cats) > len(existing_cats):
                data['fachada_detail'] = fach
                updated = True
            else:
                for k, v in fach.items():
                    if k not in existing:
                        existing[k] = v
                        updated = True
                data['fachada_detail'] = existing
        else:
            data['fachada_detail'] = fach
            updated = True

    if acab is not None:
        existing = data.get('acabamentos_pus', {})
        if existing:
            for k, v in acab.items():
                if k not in existing:
                    existing[k] = v
                    updated = True
            data['acabamentos_pus'] = existing
        else:
            data['acabamentos_pus'] = acab
            updated = True

    if updated:
        with open(idx_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    return updated


def main():
    results = {'updated': 0, 'no_change': 0, 'error': 0}
    details = []

    for idx, proj in BATCH:
        slug = proj['slug']
        print(f"Processing {idx}: {slug}...", end=' ')

        result = process_project(idx, proj)
        slug_r, esq, fach, acab, err = result[:5]
        methods = result[5] if len(result) > 5 else []

        if err:
            print(f"ERROR: {err}")
            results['error'] += 1
            details.append({'slug': slug, 'status': 'error', 'error': err})
            continue

        if not esq and not fach and not acab:
            print("NO DATA EXTRACTED")
            results['no_change'] += 1
            details.append({'slug': slug, 'status': 'no_data'})
            continue

        was_updated = update_json(slug, esq, fach, acab, methods)

        if was_updated:
            results['updated'] += 1
            parts = []
            if esq:
                parts.append(f"esq={list(esq.keys())}")
            if fach:
                parts.append(f"fach={list(fach.keys())}")
            if acab:
                parts.append(f"acab={list(acab.keys())}")
            print(f"UPDATED ({'; '.join(parts)}) via {','.join(methods)}")
        else:
            results['no_change'] += 1
            print(f"NO CHANGE (existing >= new)")

        details.append({
            'slug': slug,
            'status': 'updated' if was_updated else 'no_change',
            'esquadrias': list(esq.keys()) if esq else [],
            'fachada': list(fach.keys()) if fach else [],
            'acabamentos': list(acab.keys()) if acab else [],
            'methods': methods,
        })

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total processed: {len(BATCH)}")
    print(f"Updated: {results['updated']}")
    print(f"No change: {results['no_change']}")
    print(f"Errors: {results['error']}")

    # Coverage analysis
    print("\n--- Final coverage (after update) ---")
    for idx, proj in BATCH:
        slug = proj['slug']
        json_path = INDICES_DIR / f"{slug}.json"
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        esq = data.get('esquadrias_detail', {})
        fach = data.get('fachada_detail', {})
        acab = data.get('acabamentos_pus', {})
        esq_cats = [k for k in esq.keys() if not k.startswith('_')]
        fach_cats = [k for k in fach.keys() if not k.startswith('_')]
        acab_cats = list(acab.keys()) if acab else []

        missing = []
        if not esq_cats and not esq.get('_total'):
            missing.append('NO_ESQ')
        elif not esq_cats:
            missing.append('esq=_total_only')
        if not fach_cats and not fach.get('_total'):
            missing.append('NO_FACH')
        elif not fach_cats:
            missing.append('fach=_total_only')
        if not acab_cats:
            missing.append('NO_ACAB')

        if missing:
            print(f"  {idx}: {slug} -> {', '.join(missing)}")

    print("\n--- Esquadrias detail category coverage ---")
    esq_cats_count = {}
    for idx, proj in BATCH:
        slug = proj['slug']
        json_path = INDICES_DIR / f"{slug}.json"
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for cat in data.get('esquadrias_detail', {}).keys():
            esq_cats_count[cat] = esq_cats_count.get(cat, 0) + 1
    for cat, count in sorted(esq_cats_count.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}/{len(BATCH)} projects")

    print("\n--- Fachada detail category coverage ---")
    fach_cats_count = {}
    for idx, proj in BATCH:
        slug = proj['slug']
        json_path = INDICES_DIR / f"{slug}.json"
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for cat in data.get('fachada_detail', {}).keys():
            fach_cats_count[cat] = fach_cats_count.get(cat, 0) + 1
    for cat, count in sorted(fach_cats_count.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}/{len(BATCH)} projects")

    print("\n--- Acabamentos PU category coverage ---")
    acab_cats_count = {}
    for idx, proj in BATCH:
        slug = proj['slug']
        json_path = INDICES_DIR / f"{slug}.json"
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for cat in data.get('acabamentos_pus', {}).keys():
            acab_cats_count[cat] = acab_cats_count.get(cat, 0) + 1
    for cat, count in sorted(acab_cats_count.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}/{len(BATCH)} projects")


if __name__ == '__main__':
    main()
