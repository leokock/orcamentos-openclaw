#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract structural indices from construction budget spreadsheets.
Batch: projects 42-83 (0-based indices 41..82).

Extracts: concrete (m3), steel (kg), formwork (m2), foundation/pile data.
Updates each indices-executivo/{slug}.json adding indices_estruturais.
"""
import json, sys, re, os, traceback
from pathlib import Path
from collections import Counter
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

# Load project mapping and metadata
with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

with open(BASE_DIR / "projetos-metadados.json", "r", encoding="utf-8") as f:
    METADATA = json.load(f)

BATCH = ALL_PROJECTS[41:83]  # projects 42-83


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def get_ac(slug):
    m = METADATA.get(slug, {})
    ac = m.get('ac')
    if ac and isinstance(ac, (int, float)) and ac > 100:
        return float(ac)
    # Also check indices-executivo
    idx_path = INDICES_DIR / f"{slug}.json"
    if idx_path.exists():
        with open(idx_path, encoding='utf-8') as f:
            idx = json.load(f)
        ac2 = idx.get('ac')
        if ac2 and isinstance(ac2, (int, float)) and ac2 > 100:
            return float(ac2)
    return None


# -- Classification helpers ------------------------------------------------

def is_concreto(desc):
    """Check if description is a concrete item."""
    d = desc.lower().strip()
    if 'concreto' not in d:
        return False
    # Exclude: concreto magro (lastro), controle tecnologico, argamassa, mao de obra
    if any(x in d for x in ['magro', 'controle', 'tecnológ', 'tecnolog', 'argamassa',
                             'mão de obra', 'mao de obra']):
        return False
    return True


def is_aco(desc):
    """Check if description is a steel/rebar item."""
    d = desc.lower().strip()
    if any(x in d for x in ['armação', 'armacao', 'armadura']):
        if any(x in d for x in ['aço', 'aco', 'ca-50', 'ca-60', 'ca50', 'ca60']):
            return True
        if 'tela' in d or 'treliç' in d or 'trelic' in d:
            return True
        return True
    if 'aço ca' in d or 'aco ca' in d:
        return True
    if re.search(r'tela\s+(q|soldada)', d):
        return True
    if 'treliça' in d or 'trelica' in d:
        return True
    return False


def is_forma(desc):
    """Check if description is a formwork item."""
    d = desc.lower().strip()
    if any(x in d for x in ['fôrma', 'forma', 'formas']):
        if 'eps' in d:
            return False
        if 'informaç' in d or 'informac' in d:
            return False
        return True
    return False


def is_estaca(desc):
    """Check if description is a pile item."""
    d = desc.lower().strip()
    if any(x in d for x in ['estaca', 'perfuração', 'perfuracao']):
        if any(x in d for x in ['hélice', 'helice', 'franki', 'raiz', 'mega', 'pré-moldada',
                                 'pre-moldada', 'escavada', 'perfuração de estaca',
                                 'perfuracao de estaca', 'cravação', 'cravacao']):
            return True
        if 'perfuração' in d or 'perfuracao' in d:
            return True
    return False


def extract_fck(desc):
    """Extract fck value from concrete description."""
    d = desc.upper()
    m = re.search(r'FCK\s*[=:]?\s*(\d+)', d)
    if m:
        return int(m.group(1))
    m = re.search(r'(\d+)\s*MPA', d)
    if m:
        return int(m.group(1))
    return None


def extract_estaca_tipo(desc):
    """Extract pile type from description."""
    d = desc.lower()
    if 'hélice contínua' in d or 'helice continua' in d:
        return 'hélice contínua'
    if 'franki' in d:
        return 'Franki'
    if 'raiz' in d:
        return 'raiz'
    if 'mega' in d:
        return 'mega'
    if 'pré-moldada' in d or 'pre-moldada' in d:
        return 'pré-moldada'
    if 'escavada' in d:
        return 'escavada'
    if 'hélice' in d or 'helice' in d:
        return 'hélice contínua'
    return None


# -- Sheet-specific extractors -------------------------------------------

def extract_ger_executivo(wb, sheet_name='Ger_Executivo'):
    """Extract from Ger_Executivo format (Cartesian standard)."""
    ws = wb[sheet_name]
    result = {
        'concreto_items': [], 'aco_items': [], 'forma_items': [],
        'estaca_items': [], 'fck_values': [], 'estaca_tipos': [],
        'mo_estrutural_total': 0.0, 'infra_total': 0.0, 'supra_total': 0.0,
    }

    desc_col = 9; unit_col = 10; qty_col = 11; total_col = 13
    code_col = 8; level_col = 7
    current_section = None

    for row in ws.iter_rows(min_row=1, max_row=800, max_col=20, values_only=True):
        level = str(row[level_col]).strip() if row[level_col] else ''
        code = str(row[code_col]).strip() if row[code_col] else ''
        desc = str(row[desc_col]).strip() if row[desc_col] else ''
        unit = str(row[unit_col]).strip().lower() if row[unit_col] else ''
        qty = safe_float(row[qty_col])
        total = safe_float(row[total_col])

        if not desc:
            continue

        if level in ['CÉLULA CONSTRUTIVA', 'CELULA CONSTRUTIVA']:
            d_lower = desc.lower()
            if 'infraestrutura' in d_lower or 'fundaç' in d_lower:
                current_section = 'infra'
            elif 'supraestrutura' in d_lower or 'superestrutura' in d_lower:
                current_section = 'supra'
            else:
                current_section = None
            if current_section == 'infra':
                result['infra_total'] = total
            elif current_section == 'supra':
                result['supra_total'] = total
            continue

        if code.startswith('02'):
            current_section = 'infra'
        elif code.startswith('03'):
            current_section = 'supra'

        if level != 'SERVIÇO' and level != 'SERVICO':
            continue

        if qty <= 0:
            continue

        desc_lower = desc.lower()

        if is_concreto(desc) and unit in ['m3', 'm³']:
            fck = extract_fck(desc)
            result['concreto_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'section': current_section, 'fck': fck
            })
            if fck:
                result['fck_values'].append(fck)
        elif is_aco(desc) and unit in ['kg', 'kgf', 'ton', 't']:
            mult = 1000.0 if unit in ['ton', 't'] else 1.0
            result['aco_items'].append({
                'desc': desc, 'qty': qty * mult, 'total': total,
                'section': current_section
            })
        elif is_forma(desc) and unit in ['m2', 'm²']:
            result['forma_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'section': current_section
            })
        elif is_estaca(desc) and unit in ['m', 'ml', 'un', 'und', 'pç', 'pc']:
            tipo = extract_estaca_tipo(desc)
            if tipo:
                result['estaca_tipos'].append(tipo)
            result['estaca_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'unit': unit, 'tipo': tipo
            })
        if 'mão de obra' in desc_lower or 'mao de obra' in desc_lower:
            if any(x in desc_lower for x in ['infra', 'supra', 'estrutura', 'fundaç', 'fundac']):
                result['mo_estrutural_total'] += total

    return result


def extract_relatorio(wb, sheet_name='Relatório'):
    """Extract from Sienge Relatorio N2 format."""
    ws = wb[sheet_name]
    result = {
        'concreto_items': [], 'aco_items': [], 'forma_items': [],
        'estaca_items': [], 'fck_values': [], 'estaca_tipos': [],
        'mo_estrutural_total': 0.0, 'infra_total': 0.0, 'supra_total': 0.0,
    }

    unit_col = None; qty_col = None; total_col = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=35, values_only=True), 1):
        for j, v in enumerate(row):
            if v is not None:
                vs = str(v).strip().lower()
                if vs in ['un.', 'unidade', 'un', 'und.']:
                    unit_col = j
                elif 'quantidade' in vs:
                    qty_col = j
                elif 'preço total' in vs or 'preco total' in vs or vs == 'total':
                    if total_col is None:
                        total_col = j

    if unit_col is None: unit_col = 13
    if qty_col is None: qty_col = 16
    if total_col is None: total_col = 28

    current_section = None

    for row in ws.iter_rows(min_row=1, max_row=1500, max_col=35, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if row[1] else ''
        unit = str(row[unit_col]).strip().lower() if len(row) > unit_col and row[unit_col] else ''
        qty = safe_float(row[qty_col]) if len(row) > qty_col else 0.0
        total = safe_float(row[total_col]) if len(row) > total_col else 0.0

        if not desc:
            continue

        desc_lower = desc.lower()
        code_clean = code.strip()

        if re.match(r'^02\s*$', code_clean):
            current_section = 'infra'
            result['infra_total'] = total
            continue
        elif re.match(r'^03\s*$', code_clean):
            current_section = 'supra'
            result['supra_total'] = total
            continue
        elif re.match(r'^0[4-9]\s*$', code_clean) or re.match(r'^[1-9][0-9]*\s*$', code_clean):
            if not code_clean.startswith('02') and not code_clean.startswith('03'):
                current_section = None
                continue

        if code_clean.startswith('02'):
            current_section = 'infra'
        elif code_clean.startswith('03'):
            current_section = 'supra'

        if not re.match(r'^\d+\.\d+\.\d+\.\d+', code_clean):
            continue

        if qty <= 0:
            continue

        if is_concreto(desc) and unit in ['m3', 'm³']:
            fck = extract_fck(desc)
            result['concreto_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'section': current_section, 'fck': fck
            })
            if fck:
                result['fck_values'].append(fck)
        elif is_aco(desc) and unit in ['kg', 'kgf', 'ton', 't']:
            mult = 1000.0 if unit in ['ton', 't'] else 1.0
            result['aco_items'].append({
                'desc': desc, 'qty': qty * mult, 'total': total,
                'section': current_section
            })
        elif is_forma(desc) and unit in ['m2', 'm²']:
            result['forma_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'section': current_section
            })
        elif is_estaca(desc) and unit in ['m', 'ml', 'un', 'und', 'pç', 'pc']:
            tipo = extract_estaca_tipo(desc)
            if tipo:
                result['estaca_tipos'].append(tipo)
            result['estaca_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'unit': unit, 'tipo': tipo
            })
        if 'mão de obra' in desc_lower or 'mao de obra' in desc_lower:
            if any(x in desc_lower for x in ['infra', 'supra', 'estrutura', 'fundaç', 'fundac']):
                result['mo_estrutural_total'] += total

    return result


def extract_orcamento_executivo(wb, sheet_name):
    """Extract from ORCAMENTO_EXECUTIVO / ORCAMENTO sheets."""
    ws = wb[sheet_name]
    result = {
        'concreto_items': [], 'aco_items': [], 'forma_items': [],
        'estaca_items': [], 'fck_values': [], 'estaca_tipos': [],
        'mo_estrutural_total': 0.0, 'infra_total': 0.0, 'supra_total': 0.0,
    }

    desc_col = None; unit_col = None; qty_col = None; total_col = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, max_col=30, values_only=True), 1):
        for j, v in enumerate(row):
            if v is not None:
                vs = str(v).strip().lower()
                if vs in ['descrição', 'descricao', 'descrição do serviço']:
                    desc_col = j
                elif vs in ['un.', 'unidade', 'un', 'und', 'und.']:
                    unit_col = j
                elif vs in ['quantidade', 'quant.', 'quant', 'qtde', 'qtd', 'quantidade orçada']:
                    qty_col = j
                elif vs in ['total', 'preço total', 'valor total', 'subtotal']:
                    if total_col is None:
                        total_col = j

    if desc_col is None: desc_col = 1
    if unit_col is None: unit_col = desc_col + 1
    if qty_col is None: qty_col = unit_col + 1
    if total_col is None: total_col = qty_col + 2

    current_section = None
    code_col = 0

    for row in ws.iter_rows(min_row=1, max_row=1500, max_col=max(total_col + 1, 30), values_only=True):
        code = str(row[code_col]).strip() if row[code_col] else ''
        desc = str(row[desc_col]).strip() if len(row) > desc_col and row[desc_col] else ''
        unit = str(row[unit_col]).strip().lower() if len(row) > unit_col and row[unit_col] else ''
        qty = safe_float(row[qty_col]) if len(row) > qty_col else 0.0
        total = safe_float(row[total_col]) if len(row) > total_col else 0.0

        if not desc:
            continue

        desc_lower = desc.lower()

        if 'infraestrutura' in desc_lower and len(code.replace('.', '')) <= 3:
            current_section = 'infra'
            if total > 0:
                result['infra_total'] = total
            continue
        elif 'supraestrutura' in desc_lower and len(code.replace('.', '')) <= 3:
            current_section = 'supra'
            if total > 0:
                result['supra_total'] = total
            continue

        if code.startswith('02'):
            current_section = 'infra'
        elif code.startswith('03'):
            current_section = 'supra'

        if qty <= 0:
            continue

        if is_concreto(desc) and unit in ['m3', 'm³']:
            fck = extract_fck(desc)
            result['concreto_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'section': current_section, 'fck': fck
            })
            if fck:
                result['fck_values'].append(fck)
        elif is_aco(desc) and unit in ['kg', 'kgf', 'ton', 't']:
            mult = 1000.0 if unit in ['ton', 't'] else 1.0
            result['aco_items'].append({
                'desc': desc, 'qty': qty * mult, 'total': total,
                'section': current_section
            })
        elif is_forma(desc) and unit in ['m2', 'm²']:
            result['forma_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'section': current_section
            })
        elif is_estaca(desc) and unit in ['m', 'ml', 'un', 'und', 'pç', 'pc']:
            tipo = extract_estaca_tipo(desc)
            if tipo:
                result['estaca_tipos'].append(tipo)
            result['estaca_items'].append({
                'desc': desc, 'qty': qty, 'total': total,
                'unit': unit, 'tipo': tipo
            })
        if 'mão de obra' in desc_lower or 'mao de obra' in desc_lower:
            if any(x in desc_lower for x in ['infra', 'supra', 'estrutura', 'fundaç', 'fundac']):
                result['mo_estrutural_total'] += total

    return result


def extract_resumo_gerenciamento(wb, sheet_name):
    """Extract from summary / gerenciamento sheets."""
    ws = wb[sheet_name]
    result = {
        'concreto_items': [], 'aco_items': [], 'forma_items': [],
        'estaca_items': [], 'fck_values': [], 'estaca_tipos': [],
        'mo_estrutural_total': 0.0, 'infra_total': 0.0, 'supra_total': 0.0,
    }

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=25, values_only=True):
        for j, v in enumerate(row):
            if v is not None:
                vs = str(v).strip()
                vs_lower = vs.lower()
                if len(vs) > 10 and any(k in vs_lower for k in ['concreto', 'armação', 'armacao',
                    'fôrma', 'forma', 'estaca', 'perfuração', 'perfuracao', 'aço ca', 'aco ca']):
                    unit = ''
                    qty = 0.0
                    total_val = 0.0
                    for jj in range(max(0, j-3), min(len(row), j+8)):
                        if jj == j:
                            continue
                        cell_v = row[jj]
                        if cell_v is not None:
                            cell_s = str(cell_v).strip().lower()
                            if cell_s in ['m3', 'm³', 'kg', 'kgf', 'm2', 'm²', 'm', 'ml',
                                          'un', 'und', 'pç', 'pc', 'ton', 't']:
                                unit = cell_s
                            elif isinstance(cell_v, (int, float)) and cell_v > 0:
                                if qty == 0.0:
                                    qty = float(cell_v)
                                elif total_val == 0.0:
                                    total_val = float(cell_v)

                    if qty <= 0:
                        continue

                    if is_concreto(vs) and unit in ['m3', 'm³']:
                        fck = extract_fck(vs)
                        result['concreto_items'].append({'desc': vs, 'qty': qty, 'total': total_val, 'section': None, 'fck': fck})
                        if fck:
                            result['fck_values'].append(fck)
                    elif is_aco(vs) and unit in ['kg', 'kgf', 'ton', 't']:
                        mult = 1000.0 if unit in ['ton', 't'] else 1.0
                        result['aco_items'].append({'desc': vs, 'qty': qty * mult, 'total': total_val, 'section': None})
                    elif is_forma(vs) and unit in ['m2', 'm²']:
                        result['forma_items'].append({'desc': vs, 'qty': qty, 'total': total_val, 'section': None})
                    elif is_estaca(vs) and unit in ['m', 'ml', 'un', 'und']:
                        tipo = extract_estaca_tipo(vs)
                        if tipo:
                            result['estaca_tipos'].append(tipo)
                        result['estaca_items'].append({'desc': vs, 'qty': qty, 'total': total_val, 'unit': unit, 'tipo': tipo})
                    break

    return result


def merge_results(results_list):
    merged = {
        'concreto_items': [], 'aco_items': [], 'forma_items': [],
        'estaca_items': [], 'fck_values': [], 'estaca_tipos': [],
        'mo_estrutural_total': 0.0, 'infra_total': 0.0, 'supra_total': 0.0,
    }
    for r in results_list:
        merged['concreto_items'].extend(r['concreto_items'])
        merged['aco_items'].extend(r['aco_items'])
        merged['forma_items'].extend(r['forma_items'])
        merged['estaca_items'].extend(r['estaca_items'])
        merged['fck_values'].extend(r['fck_values'])
        merged['estaca_tipos'].extend(r['estaca_tipos'])
        merged['mo_estrutural_total'] += r['mo_estrutural_total']
        if r['infra_total'] > merged['infra_total']:
            merged['infra_total'] = r['infra_total']
        if r['supra_total'] > merged['supra_total']:
            merged['supra_total'] = r['supra_total']
    return merged


def compute_indices(raw, ac):
    concreto_total = sum(i['qty'] for i in raw['concreto_items'])
    aco_total = sum(i['qty'] for i in raw['aco_items'])
    forma_total = sum(i['qty'] for i in raw['forma_items'])

    fundacao_concreto = sum(i['qty'] for i in raw['concreto_items'] if i.get('section') == 'infra')
    fundacao_aco = sum(i['qty'] for i in raw['aco_items'] if i.get('section') == 'infra')

    estacas_m = [i for i in raw['estaca_items'] if i.get('unit') in ['m', 'ml']]
    estacas_un = [i for i in raw['estaca_items'] if i.get('unit') in ['un', 'und', 'pç', 'pc']]
    estacas_comprimento = sum(i['qty'] for i in estacas_m)
    estacas_qtd = sum(i['qty'] for i in estacas_un)

    fck_predominante = None
    if raw['fck_values']:
        fck_counter = Counter(raw['fck_values'])
        fck_predominante = fck_counter.most_common(1)[0][0]

    tipo_estaca = None
    if raw['estaca_tipos']:
        tipo_counter = Counter(raw['estaca_tipos'])
        tipo_estaca = tipo_counter.most_common(1)[0][0]

    indices = {
        'concreto_total_m3': round(concreto_total, 2) if concreto_total > 0 else None,
        'concreto_m3_por_m2_ac': round(concreto_total / ac, 4) if concreto_total > 0 and ac else None,
        'aco_total_kg': round(aco_total, 2) if aco_total > 0 else None,
        'aco_kg_por_m3_concreto': round(aco_total / concreto_total, 2) if aco_total > 0 and concreto_total > 0 else None,
        'forma_total_m2': round(forma_total, 2) if forma_total > 0 else None,
        'forma_m2_por_m2_ac': round(forma_total / ac, 4) if forma_total > 0 and ac else None,
        'fck_predominante': fck_predominante,
        'tipo_estaca': tipo_estaca,
        'estacas_comprimento_total_m': round(estacas_comprimento, 2) if estacas_comprimento > 0 else None,
        'estacas_qtd': int(estacas_qtd) if estacas_qtd > 0 else None,
        'fundacao_concreto_m3': round(fundacao_concreto, 2) if fundacao_concreto > 0 else None,
        'fundacao_aco_kg': round(fundacao_aco, 2) if fundacao_aco > 0 else None,
        'mo_estrutural_rsm2': round(raw['mo_estrutural_total'] / ac, 2) if raw['mo_estrutural_total'] > 0 and ac else None,
    }
    return indices


def process_project(proj):
    slug = proj['slug']
    path = proj['path']
    ac = get_ac(slug)

    if not os.path.exists(path):
        return slug, None, f"File not found: {path}"

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return slug, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    results = []

    try:
        # Strategy 1: Ger_Executivo (most complete)
        for sname in ['Ger_Executivo', 'Ger_executivo', 'GER_EXECUTIVO']:
            if sname in sheets:
                results.append(extract_ger_executivo(wb, sname))
                break

        # Strategy 2: Relatorio (Sienge format)
        for sname in ['Relatório', 'Relatorio', 'RELATÓRIO']:
            if sname in sheets:
                results.append(extract_relatorio(wb, sname))
                break

        # Strategy 3: ORCAMENTO_EXECUTIVO
        for sname in sheets:
            sn_lower = sname.lower().strip()
            if 'orçamento_executivo' in sn_lower or 'orcamento_executivo' in sn_lower:
                results.append(extract_orcamento_executivo(wb, sname))
                break
            if sn_lower == 'orçamento' or sn_lower == 'orcamento':
                results.append(extract_orcamento_executivo(wb, sname))
                break

        # Strategy 4: ORCAMENTO sheet
        if 'ORÇAMENTO' in sheets and 'ORÇAMENTO_EXECUTIVO' not in '|'.join(sheets):
            if 'ORÇAMENTO' not in [s for s in sheets if 'EXECUTIVO' in s.upper()]:
                try:
                    results.append(extract_orcamento_executivo(wb, 'ORÇAMENTO'))
                except:
                    pass

        # Strategy 5: Summary "Gerenciamento executivo" / "Executivo"
        for sname in sheets:
            sn_lower = sname.lower().strip()
            if 'gerenciamento executivo' in sn_lower or sn_lower in ['executivo', 'executivo ']:
                if not any('Ger_Executivo' in s for s in sheets):
                    results.append(extract_resumo_gerenciamento(wb, sname))
                break

        # Strategy 6: Direct structural sheets
        for sname in sheets:
            sn_lower = sname.lower().strip()
            if sn_lower in ['supraestrutura', 'infraestrutura', 'resumo estrutura']:
                try:
                    results.append(extract_orcamento_executivo(wb, sname))
                except:
                    pass

        # Fallback: first sheet
        if not results:
            try:
                results.append(extract_orcamento_executivo(wb, sheets[0]))
            except:
                pass

    finally:
        wb.close()

    if not results:
        return slug, None, "No extractable sheets found"

    merged = merge_results(results)

    # Deduplicate
    seen_concreto = set()
    deduped_concreto = []
    for item in merged['concreto_items']:
        key = (round(item['qty'], 1), item.get('section', ''))
        if key not in seen_concreto:
            seen_concreto.add(key)
            deduped_concreto.append(item)
    merged['concreto_items'] = deduped_concreto

    seen_aco = set()
    deduped_aco = []
    for item in merged['aco_items']:
        key = (round(item['qty'], 1), item.get('section', ''))
        if key not in seen_aco:
            seen_aco.add(key)
            deduped_aco.append(item)
    merged['aco_items'] = deduped_aco

    seen_forma = set()
    deduped_forma = []
    for item in merged['forma_items']:
        key = (round(item['qty'], 1), item.get('section', ''))
        if key not in seen_forma:
            seen_forma.add(key)
            deduped_forma.append(item)
    merged['forma_items'] = deduped_forma

    seen_estaca = set()
    deduped_estaca = []
    for item in merged['estaca_items']:
        key = (round(item['qty'], 1), item.get('unit', ''))
        if key not in seen_estaca:
            seen_estaca.add(key)
            deduped_estaca.append(item)
    merged['estaca_items'] = deduped_estaca

    indices = compute_indices(merged, ac)

    has_data = any(v is not None for v in indices.values())
    if not has_data:
        return slug, None, "No structural data found in sheets"

    return slug, indices, None


# -- Main -----------------------------------------------------------------

def main():
    print(f"Processing {len(BATCH)} projects (42-83)...\n")

    results_summary = []
    success = 0
    partial = 0
    failed = 0

    for i, proj in enumerate(BATCH):
        idx = i + 42
        slug = proj['slug']
        print(f"[{idx}] {slug}...", end=" ", flush=True)

        slug_result, indices, error = process_project(proj)

        if error:
            print(f"SKIP: {error}")
            failed += 1
            results_summary.append({'idx': idx, 'slug': slug, 'status': 'failed', 'error': error})
            continue

        if indices is None:
            print("NO DATA")
            failed += 1
            results_summary.append({'idx': idx, 'slug': slug, 'status': 'no_data'})
            continue

        non_null = sum(1 for v in indices.values() if v is not None)
        total_fields = len(indices)

        # Update JSON
        idx_path = INDICES_DIR / f"{slug}.json"
        if idx_path.exists():
            with open(idx_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
        else:
            existing = {'projeto': slug}

        existing['indices_estruturais'] = indices

        with open(idx_path, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)

        if non_null >= 3:
            success += 1
            status = 'ok'
        else:
            partial += 1
            status = 'partial'

        conc = indices.get('concreto_total_m3')
        aco = indices.get('aco_total_kg')
        forma = indices.get('forma_total_m2')
        ratio = indices.get('aco_kg_por_m3_concreto')
        fck = indices.get('fck_predominante')
        estaca = indices.get('tipo_estaca')

        print(f"{status.upper()} [{non_null}/{total_fields}] "
              f"conc={conc} aco={aco} forma={forma} "
              f"kg/m3={ratio} fck={fck} estaca={estaca}")

        results_summary.append({
            'idx': idx, 'slug': slug, 'status': status,
            'non_null': non_null, 'indices': indices
        })

    # -- Summary -----------------------------------------------------------
    print(f"\n{'='*70}")
    print(f"SUMMARY: {len(BATCH)} projects")
    print(f"  Success (>=3 fields): {success}")
    print(f"  Partial (<3 fields):  {partial}")
    print(f"  Failed/No data:       {failed}")
    print(f"{'='*70}")

    # Field coverage
    field_counts = {}
    for r in results_summary:
        if r.get('indices'):
            for k, v in r['indices'].items():
                if k not in field_counts:
                    field_counts[k] = 0
                if v is not None:
                    field_counts[k] += 1

    print("\nField coverage:")
    for k, v in sorted(field_counts.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}/{success + partial}")

    # Key stats
    conc_vals = [r['indices']['concreto_total_m3'] for r in results_summary
                 if r.get('indices') and r['indices'].get('concreto_total_m3')]
    aco_vals = [r['indices']['aco_total_kg'] for r in results_summary
                if r.get('indices') and r['indices'].get('aco_total_kg')]
    ratio_vals = [r['indices']['aco_kg_por_m3_concreto'] for r in results_summary
                  if r.get('indices') and r['indices'].get('aco_kg_por_m3_concreto')]
    conc_m2_vals = [r['indices']['concreto_m3_por_m2_ac'] for r in results_summary
                    if r.get('indices') and r['indices'].get('concreto_m3_por_m2_ac')]

    print("\nKey statistics:")
    if conc_vals:
        print(f"  Concrete total m3: avg={sum(conc_vals)/len(conc_vals):.1f}, "
              f"min={min(conc_vals):.1f}, max={max(conc_vals):.1f} (n={len(conc_vals)})")
    if aco_vals:
        print(f"  Steel total kg: avg={sum(aco_vals)/len(aco_vals):.0f}, "
              f"min={min(aco_vals):.0f}, max={max(aco_vals):.0f} (n={len(aco_vals)})")
    if ratio_vals:
        print(f"  Steel kg/m3 concrete: avg={sum(ratio_vals)/len(ratio_vals):.1f}, "
              f"min={min(ratio_vals):.1f}, max={max(ratio_vals):.1f} (n={len(ratio_vals)})")
    if conc_m2_vals:
        print(f"  Concrete m3/m2 AC: avg={sum(conc_m2_vals)/len(conc_m2_vals):.4f}, "
              f"min={min(conc_m2_vals):.4f}, max={max(conc_m2_vals):.4f} (n={len(conc_m2_vals)})")


if __name__ == '__main__':
    main()
