#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract ESQUADRIAS, FACHADA, and ACABAMENTOS detail from construction budget spreadsheets.
Batch: projects 0-41 (0-based indices).

Updates each indices-executivo/{slug}.json adding:
  - esquadrias_detail
  - fachada_detail
  - acabamentos_pus
"""
import json, sys, re, os, traceback
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

BATCH = ALL_PROJECTS[0:42]  # projects 0-41


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def norm(s):
    """Normalize string for comparison: lowercase, strip, handle encoding."""
    if s is None:
        return ''
    return str(s).strip().lower()


def contains_any(text, keywords):
    t = norm(text)
    return any(k in t for k in keywords)


def resolve_path(project):
    """Convert Drive path to local accessible path."""
    p = project['path']
    # Try Windows path directly
    if os.path.exists(p):
        return p
    # Try converting G: drive path
    return p


def find_sheet(wb, candidates):
    """Find a sheet by trying multiple name candidates (case-insensitive, partial match)."""
    sheets_lower = {s.lower(): s for s in wb.sheetnames}
    for c in candidates:
        cl = c.lower()
        if cl in sheets_lower:
            return wb[sheets_lower[cl]]
        # Partial match
        for sl, sn in sheets_lower.items():
            if cl in sl or sl in cl:
                return wb[sn]
    return None


def detect_format(wb):
    """Detect spreadsheet format type."""
    sheets_lower = [s.lower() for s in wb.sheetnames]

    # Cartesian discipline sheets format
    has_esquadrias_sheet = any('esquadria' in s for s in sheets_lower)
    has_fachada_sheet = any('fachada' in s for s in sheets_lower)
    has_acabamentos_sheet = any('acabamento' in s for s in sheets_lower)

    if has_esquadrias_sheet or has_fachada_sheet or has_acabamentos_sheet:
        return 'cartesian'

    # Sienge Relatório format
    if any(s in ['relatório', 'relatorio', 'relatã³rio'] for s in sheets_lower):
        return 'relatorio'

    # Ger_Executivo format
    if any('ger_executivo' in s or 'gerenciamento' in s for s in sheets_lower):
        return 'ger_executivo'

    # Has Orçamento sheet (Sienge export)
    if any(s in ['orçamento', 'orcamento', 'orã§amento'] for s in sheets_lower):
        return 'orcamento_sienge'

    # Resumo-based format
    if any(s in ['resumo'] for s in sheets_lower):
        return 'resumo'

    return 'unknown'


# ============================================================================
# ESQUADRIAS EXTRACTION
# ============================================================================

def extract_esquadrias_cartesian(wb):
    """Extract from dedicated ESQUADRIAS sheet (Cartesian format)."""
    ws = find_sheet(wb, ['ESQUADRIAS', 'Esquadrias', 'ESQUADRIAS E VIDROS', 'Esquadrias e Vidros'])
    if ws is None:
        return None

    result = {
        'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
        'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
        'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
        'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
        'vidros': {'valor': None},
        'pele_vidro': {'valor': None},
        'brises': {'valor': None},
        'total': None
    }

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=20, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc:
            continue

        qty = safe_float(row[3]) if len(row) > 3 else 0
        unit = norm(row[4]) if len(row) > 4 and row[4] else ''
        pu = safe_float(row[5]) if len(row) > 5 else 0
        total = safe_float(row[6]) if len(row) > 6 else 0

        if total == 0:
            continue

        if 'esquadria' in desc and 'alumín' in desc or 'alumin' in desc and 'esquadria' in desc:
            result['aluminio'] = {'valor': total, 'area_m2': qty if 'm' in unit else None,
                                  'pu_m2': pu if pu > 0 else None}
        elif ('alumínio' in desc or 'aluminio' in desc) and ('esquadria' not in desc):
            if result['aluminio']['valor'] is None:
                result['aluminio'] = {'valor': total, 'area_m2': qty if 'm' in unit else None,
                                      'pu_m2': pu if pu > 0 else None}
        elif 'guarda' in desc and ('corpo' in desc or 'vidro' in desc):
            result['guarda_corpo'] = {'valor': total, 'comprimento_m': qty if 'm' in unit else None,
                                      'pu_m': pu if pu > 0 else None}
        elif 'porta' in desc and ('madeira' in desc or 'mad' in desc):
            result['portas_madeira'] = {'valor': total, 'qtd': qty if 'un' in unit else None,
                                        'pu_un': pu if pu > 0 else None}
        elif ('pcf' in desc or 'corta' in desc and 'fogo' in desc or 'corta-fogo' in desc):
            result['portas_corta_fogo'] = {'valor': total, 'qtd': qty if 'un' in unit else None,
                                           'pu_un': pu if pu > 0 else None}
        elif 'vidro' in desc and 'guarda' not in desc:
            if 'pele' in desc:
                result['pele_vidro'] = {'valor': total}
            else:
                result['vidros'] = {'valor': total}
        elif 'brise' in desc:
            result['brises'] = {'valor': total}
        elif 'pele' in desc and 'vidro' in desc:
            result['pele_vidro'] = {'valor': total}

        if desc.startswith('total esquadria') or desc == 'total esquadrias':
            result['total'] = total

    return result


def extract_esquadrias_relatorio(wb):
    """Extract from Sienge Relatório format."""
    ws = find_sheet(wb, ['Relatório', 'Relatorio'])
    if ws is None:
        return None

    result = {
        'aluminio': {'valor': 0, 'area_m2': 0, 'pu_m2': None},
        'guarda_corpo': {'valor': 0, 'comprimento_m': 0, 'pu_m': None},
        'portas_madeira': {'valor': 0, 'qtd': 0, 'pu_un': None},
        'portas_corta_fogo': {'valor': 0, 'qtd': 0, 'pu_un': None},
        'vidros': {'valor': 0},
        'pele_vidro': {'valor': 0},
        'brises': {'valor': 0},
        'total': None
    }

    # Find the N2 header for esquadrias and accumulate
    in_esquadrias = False
    esquadrias_n2_code = None

    # Detect column layout
    total_col = 28  # default
    unit_col = 13
    qty_col = 16

    for row in ws.iter_rows(min_row=1, max_row=15, max_col=35, values_only=True):
        for j, v in enumerate(row):
            if v is not None:
                vs = norm(v)
                if 'preço total' in vs or 'preco total' in vs:
                    total_col = j

    aluminio_items = []
    guarda_corpo_items = []
    portas_madeira_items = []
    portas_pcf_items = []
    vidros_items = []
    brise_items = []

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=35, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if row[1] else ''
        if not desc:
            continue

        d = norm(desc)
        total = safe_float(row[total_col]) if len(row) > total_col else 0
        unit = norm(row[unit_col]) if len(row) > unit_col and row[unit_col] else ''
        qty = safe_float(row[qty_col]) if len(row) > qty_col else 0

        # Detect N2 esquadrias section
        if re.match(r'^\d{2}$', code):
            if 'esquadria' in d or 'vidro' in d and 'ferragens' in d:
                in_esquadrias = True
                esquadrias_n2_code = code
                result['total'] = total
            elif in_esquadrias:
                in_esquadrias = False
            continue

        if not in_esquadrias:
            # Also scan for items outside esquadrias section if they match
            continue

        # Only process leaf items (N4 codes: XX.YYY.ZZZ.NNN)
        if not re.match(r'^\d+\.\d+\.\d+\.\d+', code):
            # Check N3 for sub-totals
            if re.match(r'^\d+\.\d+\.\d+$', code):
                if 'esquadria' in d and 'alumín' in d.replace('í', 'i') or 'alumin' in d:
                    pass  # N3 subtotal
            continue

        if total <= 0 or qty <= 0:
            continue

        # Classify item
        if contains_any(desc, ['esquadria de alumín', 'esquadria de alumin', 'esquadrias de alumín',
                               'esquadrias de alumin', 'contramarco de alumín', 'contramarco de alumin']):
            aluminio_items.append({'qty': qty, 'total': total, 'unit': unit, 'desc': desc})
        elif contains_any(desc, ['guarda-corpo', 'guarda corpo', 'corrimão', 'corrimao']):
            guarda_corpo_items.append({'qty': qty, 'total': total, 'unit': unit, 'desc': desc})
        elif contains_any(desc, ['porta de madeira', 'portas de madeira', 'porta madeira']):
            portas_madeira_items.append({'qty': qty, 'total': total, 'unit': unit, 'desc': desc})
        elif contains_any(desc, ['corta-fogo', 'corta fogo', 'porta corta']):
            portas_pcf_items.append({'qty': qty, 'total': total, 'unit': unit, 'desc': desc})
        elif contains_any(desc, ['pele de vidro']):
            result['pele_vidro']['valor'] = result['pele_vidro'].get('valor', 0) + total
        elif contains_any(desc, ['brise']):
            brise_items.append({'qty': qty, 'total': total, 'unit': unit, 'desc': desc})
        elif contains_any(desc, ['vidro', 'vidraç', 'vidrac']):
            vidros_items.append({'qty': qty, 'total': total, 'unit': unit, 'desc': desc})

    # Aggregate
    if aluminio_items:
        total_val = sum(i['total'] for i in aluminio_items)
        total_area = sum(i['qty'] for i in aluminio_items if 'm2' in i['unit'] or 'm²' in i['unit'])
        result['aluminio'] = {
            'valor': round(total_val, 2),
            'area_m2': round(total_area, 2) if total_area > 0 else None,
            'pu_m2': round(total_val / total_area, 2) if total_area > 0 else None
        }
    else:
        result['aluminio'] = {'valor': None, 'area_m2': None, 'pu_m2': None}

    if guarda_corpo_items:
        total_val = sum(i['total'] for i in guarda_corpo_items)
        total_m = sum(i['qty'] for i in guarda_corpo_items if i['unit'] in ['m', 'ml'])
        result['guarda_corpo'] = {
            'valor': round(total_val, 2),
            'comprimento_m': round(total_m, 2) if total_m > 0 else None,
            'pu_m': round(total_val / total_m, 2) if total_m > 0 else None
        }
    else:
        result['guarda_corpo'] = {'valor': None, 'comprimento_m': None, 'pu_m': None}

    if portas_madeira_items:
        total_val = sum(i['total'] for i in portas_madeira_items)
        total_un = sum(i['qty'] for i in portas_madeira_items if i['unit'] in ['un', 'und', 'pç', 'pc', 'un.'])
        result['portas_madeira'] = {
            'valor': round(total_val, 2),
            'qtd': round(total_un) if total_un > 0 else None,
            'pu_un': round(total_val / total_un, 2) if total_un > 0 else None
        }
    else:
        result['portas_madeira'] = {'valor': None, 'qtd': None, 'pu_un': None}

    if portas_pcf_items:
        total_val = sum(i['total'] for i in portas_pcf_items)
        total_un = sum(i['qty'] for i in portas_pcf_items if i['unit'] in ['un', 'und', 'pç', 'pc', 'un.'])
        result['portas_corta_fogo'] = {
            'valor': round(total_val, 2),
            'qtd': round(total_un) if total_un > 0 else None,
            'pu_un': round(total_val / total_un, 2) if total_un > 0 else None
        }
    else:
        result['portas_corta_fogo'] = {'valor': None, 'qtd': None, 'pu_un': None}

    if vidros_items:
        result['vidros'] = {'valor': round(sum(i['total'] for i in vidros_items), 2)}
    else:
        result['vidros'] = {'valor': None}

    if brise_items:
        result['brises'] = {'valor': round(sum(i['total'] for i in brise_items), 2)}
    else:
        result['brises'] = {'valor': None}

    if not result['pele_vidro'].get('valor'):
        result['pele_vidro'] = {'valor': None}

    return result


def extract_esquadrias_ger_executivo(wb):
    """Extract from Ger_Executivo sheet."""
    ws = find_sheet(wb, ['Ger_Executivo', 'GER_EXECUTIVO', 'Ger Executivo'])
    if ws is None:
        return None

    result = {
        'aluminio': {'valor': 0, 'area_m2': 0, 'pu_m2': None},
        'guarda_corpo': {'valor': 0, 'comprimento_m': 0, 'pu_m': None},
        'portas_madeira': {'valor': 0, 'qtd': 0, 'pu_un': None},
        'portas_corta_fogo': {'valor': 0, 'qtd': 0, 'pu_un': None},
        'vidros': {'valor': 0},
        'pele_vidro': {'valor': 0},
        'brises': {'valor': 0},
        'total': None
    }

    in_esquadrias = False
    desc_col = 9; unit_col = 10; qty_col = 11; total_col = 13
    level_col = 7

    for row in ws.iter_rows(min_row=1, max_row=1500, max_col=20, values_only=True):
        level = str(row[level_col]).strip() if len(row) > level_col and row[level_col] else ''
        desc = str(row[desc_col]).strip() if len(row) > desc_col and row[desc_col] else ''
        unit = norm(row[unit_col]) if len(row) > unit_col and row[unit_col] else ''
        qty = safe_float(row[qty_col]) if len(row) > qty_col else 0
        total = safe_float(row[total_col]) if len(row) > total_col else 0

        if not desc:
            continue
        d = norm(desc)

        if level in ['CÉLULA CONSTRUTIVA', 'CELULA CONSTRUTIVA']:
            if 'esquadria' in d:
                in_esquadrias = True
                result['total'] = total
            elif in_esquadrias:
                in_esquadrias = False
            continue

        if not in_esquadrias:
            continue

        if level != 'SERVIÇO' and level != 'SERVICO':
            continue

        if total <= 0:
            continue

        if contains_any(desc, ['esquadria de alumín', 'esquadria de alumin', 'esquadrias de alumín',
                               'esquadrias de alumin']):
            result['aluminio']['valor'] += total
            if 'm' in unit:
                result['aluminio']['area_m2'] += qty
        elif contains_any(desc, ['guarda-corpo', 'guarda corpo', 'corrimão', 'corrimao']):
            result['guarda_corpo']['valor'] += total
            if unit in ['m', 'ml']:
                result['guarda_corpo']['comprimento_m'] += qty
        elif contains_any(desc, ['porta de madeira', 'portas de madeira', 'porta madeira',
                                  'esquadria de madeira', 'esquadrias de madeira']):
            result['portas_madeira']['valor'] += total
            if unit in ['un', 'und', 'pç', 'pc']:
                result['portas_madeira']['qtd'] += qty
        elif contains_any(desc, ['corta-fogo', 'corta fogo', 'pcf']):
            result['portas_corta_fogo']['valor'] += total
            if unit in ['un', 'und', 'pç', 'pc']:
                result['portas_corta_fogo']['qtd'] += qty
        elif 'pele de vidro' in d or 'pele vidro' in d:
            result['pele_vidro']['valor'] += total
        elif 'brise' in d:
            result['brises']['valor'] += total
        elif 'vidro' in d and 'guarda' not in d:
            result['vidros']['valor'] += total

    # Clean up zeros
    for key in ['aluminio', 'guarda_corpo', 'portas_madeira', 'portas_corta_fogo']:
        if result[key].get('valor', 0) == 0:
            result[key] = {k: None for k in result[key]}
        else:
            result[key]['valor'] = round(result[key]['valor'], 2)
            for subkey in list(result[key].keys()):
                if subkey != 'valor' and result[key][subkey] == 0:
                    result[key][subkey] = None
                elif subkey != 'valor' and result[key][subkey] is not None:
                    result[key][subkey] = round(result[key][subkey], 2)

    for key in ['vidros', 'pele_vidro', 'brises']:
        if result[key].get('valor', 0) == 0:
            result[key] = {'valor': None}
        else:
            result[key]['valor'] = round(result[key]['valor'], 2)

    # Compute PU if we have area/qty
    if result['aluminio']['valor'] and result['aluminio'].get('area_m2'):
        result['aluminio']['pu_m2'] = round(result['aluminio']['valor'] / result['aluminio']['area_m2'], 2)
    if result['guarda_corpo']['valor'] and result['guarda_corpo'].get('comprimento_m'):
        result['guarda_corpo']['pu_m'] = round(result['guarda_corpo']['valor'] / result['guarda_corpo']['comprimento_m'], 2)
    if result['portas_madeira']['valor'] and result['portas_madeira'].get('qtd'):
        result['portas_madeira']['pu_un'] = round(result['portas_madeira']['valor'] / result['portas_madeira']['qtd'], 2)
    if result['portas_corta_fogo']['valor'] and result['portas_corta_fogo'].get('qtd'):
        result['portas_corta_fogo']['pu_un'] = round(result['portas_corta_fogo']['valor'] / result['portas_corta_fogo']['qtd'], 2)

    return result


def extract_totals_from_resumo(wb):
    """Extract esquadrias/fachada totals from Resumo sheet (Sienge export format)."""
    ws = find_sheet(wb, ['Resumo', 'RESUMO', 'EAP_SINTETICO', 'EAP Sintético'])
    if ws is None:
        return None, None

    esq_total = None
    fac_total = None

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=15, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc:
            continue

        # Find value in first numeric column after description
        val = None
        for ci in range(1, min(10, len(row))):
            v = safe_float(row[ci])
            if v > 0:
                val = v
                break

        if val is None:
            continue

        if 'esquadria' in desc:
            esq_total = val
        elif 'fachada' in desc:
            fac_total = val

    return esq_total, fac_total


# ============================================================================
# FACHADA EXTRACTION
# ============================================================================

def extract_fachada_cartesian(wb):
    """Extract from dedicated fachada sheet (Rev. Fachada, FACHADA, etc.)."""
    ws = find_sheet(wb, ['Rev. Fachada', 'FACHADA', 'Fachada', 'Rev Fachada',
                          'Revestimento de Fachada', 'REVESTIMENTO FACHADA'])
    if ws is None:
        return None

    result = {
        'reboco_externo': {'valor': None},
        'textura_pintura_ext': {'valor': None},
        'pastilha_porcelanato': {'valor': None},
        'juntas': {'valor': None},
        'total': None
    }

    for row in ws.iter_rows(min_row=1, max_row=80, max_col=20, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc:
            continue

        total = safe_float(row[6]) if len(row) > 6 else 0
        if total == 0:
            # Some sheets have total in different column
            for ci in [5, 7, 8]:
                if len(row) > ci:
                    t = safe_float(row[ci])
                    if t > 0:
                        total = t
                        break

        if 'chapisco' in desc or 'reboco' in desc or 'massa única' in desc or 'massa unica' in desc:
            if 'total' not in desc:
                result['reboco_externo'] = {'valor': total}
        elif 'textura' in desc or ('pintura' in desc and 'ext' in desc):
            result['textura_pintura_ext'] = {'valor': total}
        elif 'pastilha' in desc or ('porcelanato' in desc and 'fachada' not in desc):
            # Pastilha or porcelanato on facade
            result['pastilha_porcelanato'] = {'valor': total}
        elif 'junta' in desc:
            result['juntas'] = {'valor': total}
        elif desc.startswith('total fachada'):
            result['total'] = total

    return result


def extract_fachada_relatorio(wb):
    """Extract fachada from Sienge Relatório format."""
    ws = find_sheet(wb, ['Relatório', 'Relatorio'])
    if ws is None:
        return None

    result = {
        'reboco_externo': {'valor': 0},
        'textura_pintura_ext': {'valor': 0},
        'pastilha_porcelanato': {'valor': 0},
        'juntas': {'valor': 0},
        'total': None
    }

    total_col = 28
    unit_col = 13
    qty_col = 16

    for row in ws.iter_rows(min_row=1, max_row=15, max_col=35, values_only=True):
        for j, v in enumerate(row):
            if v is not None:
                vs = norm(v)
                if 'preço total' in vs or 'preco total' in vs:
                    total_col = j

    in_fachada = False

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=35, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if row[1] else ''
        if not desc:
            continue

        d = norm(desc)
        total = safe_float(row[total_col]) if len(row) > total_col else 0

        # Detect N2 fachada section
        if re.match(r'^\d{2}$', code):
            if 'fachada' in d:
                in_fachada = True
                result['total'] = total
            elif in_fachada:
                in_fachada = False
            continue

        if not in_fachada:
            continue

        # N3 subtotals
        if re.match(r'^\d+\.\d+\.\d+$', code):
            if 'revestimento argamassado' in d or 'chapisco' in d and 'reboco' in d:
                result['reboco_externo']['valor'] += total
            elif 'pintura' in d or 'textura' in d:
                result['textura_pintura_ext']['valor'] += total
            elif 'acabamento' in d and 'especial' in d or 'porcelanato' in d or 'pastilha' in d:
                result['pastilha_porcelanato']['valor'] += total
            elif 'junta' in d or 'tratamento' in d:
                result['juntas']['valor'] += total
            continue

        # N4 items - classify individually
        if not re.match(r'^\d+\.\d+\.\d+\.\d+', code):
            continue

        if total <= 0:
            continue

        if contains_any(desc, ['chapisco ext', 'reboco ext', 'massa única ext', 'massa unica ext',
                               'massa externa', 'revestimento argamassado']):
            pass  # Already captured in N3
        elif contains_any(desc, ['textura', 'pintura ext', 'pintura acríl', 'pintura acril',
                                  'selador ext', 'fundo selador']):
            pass  # Already captured in N3

    # Clean up zeros
    for key in ['reboco_externo', 'textura_pintura_ext', 'pastilha_porcelanato', 'juntas']:
        if result[key]['valor'] == 0:
            result[key] = {'valor': None}
        else:
            result[key]['valor'] = round(result[key]['valor'], 2)

    return result


# ============================================================================
# ACABAMENTOS PUs EXTRACTION
# ============================================================================

def extract_acabamentos_pus_cartesian(wb):
    """Extract acabamentos PUs from dedicated sheet.

    Cartesian sheets have two layouts:
    - Acabamentos: [Desc, Qty, unit, PU_or_param, param_unit, Ref, Total]
      When param_unit is 'R$ / m²' -> col[3] is PU per m2
      When param_unit is 'R$ / AC' -> col[3] is PU per AC (not per m2)
    - PINTURA/TETO: [Desc, Param_value, param_unit, Qty, qty_unit, PU, Total]
      col[5] is PU
    """
    ws = find_sheet(wb, ['Acabamentos de Piso e Parede', 'ACABAMENTOS', 'Acabamentos',
                          'REVESTIMENTO PISO PAREDE', 'Revestimento Piso e Parede'])
    if ws is None:
        return None

    result = {
        'porcelanato_pu_m2': None,
        'ceramica_pu_m2': None,
        'laminado_pu_m2': None,
        'rodape_pu_m': None,
        'contrapiso_pu_m2': None,
        'forro_gesso_pu_m2': None,
        'pintura_parede_pu_m2': None,
        'chapisco_pu_m2': None,
        'reboco_pu_m2': None
    }

    for row in ws.iter_rows(min_row=1, max_row=80, max_col=20, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc or desc.startswith('total') or desc.startswith('descri'):
            continue

        qty = safe_float(row[1]) if len(row) > 1 else 0
        unit = norm(row[2]) if len(row) > 2 and row[2] else ''
        param_val = safe_float(row[3]) if len(row) > 3 else 0
        param_unit = norm(row[4]) if len(row) > 4 and row[4] else ''
        total = safe_float(row[6]) if len(row) > 6 else 0

        # Determine the real PU per unit area
        pu = None
        if 'r$ / m' in param_unit and 'ac' not in param_unit:
            # Direct R$/m2 or R$/m price
            pu = param_val
        elif qty > 0 and total > 0 and ('m' in unit):
            # Compute PU from total/qty when qty is in m2
            pu = total / qty
        # Skip R$/AC items - those are not per-m2 PUs

        if pu is None or pu <= 0:
            continue

        if 'porcelanato' in desc:
            if result['porcelanato_pu_m2'] is None:
                result['porcelanato_pu_m2'] = round(pu, 2)
        elif 'cerâmica' in desc or 'ceramica' in desc or 'cerâmico' in desc or 'ceramico' in desc:
            if result['ceramica_pu_m2'] is None:
                result['ceramica_pu_m2'] = round(pu, 2)
        elif 'laminado' in desc or 'vinílico' in desc or 'vinilico' in desc:
            if result['laminado_pu_m2'] is None:
                result['laminado_pu_m2'] = round(pu, 2)
        elif 'rodapé' in desc or 'rodape' in desc:
            if result['rodape_pu_m'] is None:
                result['rodape_pu_m'] = round(pu, 2)

    return result


def extract_acabamentos_pus_relatorio(wb):
    """Extract acabamentos PUs from Sienge Relatório format."""
    ws = find_sheet(wb, ['Relatório', 'Relatorio'])
    if ws is None:
        return None

    result = {
        'porcelanato_pu_m2': None,
        'ceramica_pu_m2': None,
        'laminado_pu_m2': None,
        'rodape_pu_m': None,
        'contrapiso_pu_m2': None,
        'forro_gesso_pu_m2': None,
        'pintura_parede_pu_m2': None,
        'chapisco_pu_m2': None,
        'reboco_pu_m2': None
    }

    total_col = 28
    unit_col = 13
    qty_col = 16

    for row in ws.iter_rows(min_row=1, max_row=15, max_col=35, values_only=True):
        for j, v in enumerate(row):
            if v is not None:
                vs = norm(v)
                if 'preço total' in vs or 'preco total' in vs:
                    total_col = j

    # Collect items with qty and total to compute PU
    porcelanato_items = []
    ceramica_items = []
    laminado_items = []
    rodape_items = []
    contrapiso_items = []
    forro_items = []
    pintura_items = []
    chapisco_items = []
    reboco_items = []

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=35, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if row[1] else ''
        if not desc:
            continue

        d = norm(desc)
        total = safe_float(row[total_col]) if len(row) > total_col else 0
        unit = norm(row[unit_col]) if len(row) > unit_col and row[unit_col] else ''
        qty = safe_float(row[qty_col]) if len(row) > qty_col else 0

        if not re.match(r'^\d+\.\d+\.\d+\.\d+', code):
            continue
        if total <= 0 or qty <= 0:
            continue

        # Exclude MO items for material PU
        if 'mão de obra' in d or 'mao de obra' in d:
            continue

        if 'porcelanato' in d and ('piso' in d or 'revestimento' in d) and unit in ['m2', 'm²']:
            porcelanato_items.append({'qty': qty, 'total': total})
        elif ('cerâmica' in d or 'ceramica' in d or 'cerâmico' in d or 'ceramico' in d) and unit in ['m2', 'm²']:
            ceramica_items.append({'qty': qty, 'total': total})
        elif ('laminado' in d or 'vinílico' in d or 'vinilico' in d) and unit in ['m2', 'm²']:
            laminado_items.append({'qty': qty, 'total': total})
        elif ('rodapé' in d or 'rodape' in d) and unit in ['m', 'ml', 'm2', 'm²']:
            rodape_items.append({'qty': qty, 'total': total})
        elif 'contrapiso' in d and unit in ['m2', 'm²']:
            contrapiso_items.append({'qty': qty, 'total': total})
        elif ('forro' in d and 'gesso' in d) and unit in ['m2', 'm²']:
            forro_items.append({'qty': qty, 'total': total})
        elif ('pintura' in d and ('parede' in d or 'interna' in d)) and unit in ['m2', 'm²']:
            pintura_items.append({'qty': qty, 'total': total})
        elif 'chapisco' in d and 'interno' in d and unit in ['m2', 'm²']:
            chapisco_items.append({'qty': qty, 'total': total})
        elif ('reboco' in d and 'interno' in d) and unit in ['m2', 'm²']:
            reboco_items.append({'qty': qty, 'total': total})

    def compute_pu(items):
        if not items:
            return None
        total_qty = sum(i['qty'] for i in items)
        total_val = sum(i['total'] for i in items)
        if total_qty > 0:
            return round(total_val / total_qty, 2)
        return None

    result['porcelanato_pu_m2'] = compute_pu(porcelanato_items)
    result['ceramica_pu_m2'] = compute_pu(ceramica_items)
    result['laminado_pu_m2'] = compute_pu(laminado_items)
    result['rodape_pu_m'] = compute_pu(rodape_items)
    result['contrapiso_pu_m2'] = compute_pu(contrapiso_items)
    result['forro_gesso_pu_m2'] = compute_pu(forro_items)
    result['pintura_parede_pu_m2'] = compute_pu(pintura_items)
    result['chapisco_pu_m2'] = compute_pu(chapisco_items)
    result['reboco_pu_m2'] = compute_pu(reboco_items)

    return result


def extract_pus_from_rev_internos(wb):
    """Extract PUs from Rev. Internos Piso e Parede sheet (Cartesian).

    Layout: [Desc, Qty, unit, Param_value, param_unit, Ref, Total]
    Param_unit could be 'R$ / m²' (direct PU) or 'm² / AC' (parametric index, not PU).
    When param_unit is m2/AC, we compute PU = total/qty.
    """
    ws = find_sheet(wb, ['Rev. Internos Piso e Parede', 'Rev Internos', 'REVESTIMENTO INTERNO',
                          'Revestimento Interno'])
    if ws is None:
        return None

    result = {
        'porcelanato_pu_m2': None,
        'ceramica_pu_m2': None,
        'laminado_pu_m2': None,
        'rodape_pu_m': None,
        'contrapiso_pu_m2': None,
        'forro_gesso_pu_m2': None,
        'pintura_parede_pu_m2': None,
        'chapisco_pu_m2': None,
        'reboco_pu_m2': None
    }

    for row in ws.iter_rows(min_row=1, max_row=80, max_col=20, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc or desc.startswith('total') or desc.startswith('descri'):
            continue

        qty = safe_float(row[1]) if len(row) > 1 else 0
        unit = norm(row[2]) if len(row) > 2 and row[2] else ''
        param_val = safe_float(row[3]) if len(row) > 3 else 0
        param_unit = norm(row[4]) if len(row) > 4 and row[4] else ''
        total = safe_float(row[6]) if len(row) > 6 else 0

        pu = None
        if 'r$ / m' in param_unit and 'ac' not in param_unit:
            pu = param_val
        elif qty > 0 and total > 0 and ('m' in unit):
            pu = total / qty

        if pu is None or pu <= 0:
            continue

        if 'contrapiso' in desc:
            result['contrapiso_pu_m2'] = round(pu, 2)
        elif 'chapisco' in desc and 'reboco' in desc:
            result['chapisco_pu_m2'] = round(pu, 2)
            result['reboco_pu_m2'] = round(pu, 2)
        elif 'chapisco' in desc:
            result['chapisco_pu_m2'] = round(pu, 2)
        elif 'reboco' in desc:
            result['reboco_pu_m2'] = round(pu, 2)

    return result


def extract_pus_from_pintura(wb):
    """Extract pintura PU from PINTURA sheet.

    Layout: [Desc, Param_value, param_unit, Qty, qty_unit, PU, Total]
    col[5] is PU (preço unitário).
    Only use when param_unit is 'R$ / m²'.
    """
    ws = find_sheet(wb, ['PINTURA INTERNA', 'PINTURA', 'Pintura', 'Pintura Interna'])
    if ws is None:
        return None

    for row in ws.iter_rows(min_row=1, max_row=60, max_col=20, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc or desc.startswith('total') or desc.startswith('descri'):
            continue

        param_val = safe_float(row[1]) if len(row) > 1 else 0
        param_unit = norm(row[2]) if len(row) > 2 and row[2] else ''
        qty = safe_float(row[3]) if len(row) > 3 else 0
        qty_unit = norm(row[4]) if len(row) > 4 and row[4] else ''
        pu = safe_float(row[5]) if len(row) > 5 else 0
        total = safe_float(row[6]) if len(row) > 6 else 0

        # Only accept R$/m2 PUs
        if 'r$ / m' in param_unit and 'ac' not in param_unit:
            if pu <= 0 and qty > 0 and total > 0:
                pu = total / qty
        elif 'r$ / ac' in param_unit:
            # R$/AC is not a per-m2 PU - skip
            continue
        else:
            # Fallback: compute from total/qty if qty is in m2
            if pu <= 0 and qty > 0 and total > 0 and 'm' in qty_unit:
                pu = total / qty

        if pu <= 0:
            continue

        # Filter: wall paint only, exclude floor/anti-slip/epoxy
        if 'piso' in desc or 'epoxi' in desc or 'epox' in desc or 'antiderrapant' in desc:
            continue
        if 'pintura' in desc and ('parede' in desc or 'interna' in desc or 'latex' in desc or
                                   'látex' in desc or 'acríl' in desc or 'acril' in desc):
            return round(pu, 2)

    return None


def extract_pus_from_forro(wb):
    """Extract forro gesso PU from TETO/FORRO sheet.

    Layout: [Desc, Param_value, param_unit, Qty, qty_unit, PU, Total]
    col[5] is PU.
    """
    ws = find_sheet(wb, ['TETO', 'FORRO', 'Forro', 'Teto'])
    if ws is None:
        return None

    forro_items = []

    for row in ws.iter_rows(min_row=1, max_row=60, max_col=20, values_only=True):
        desc = norm(row[0]) if row[0] else ''
        if not desc or desc.startswith('total') or desc.startswith('descri'):
            continue

        param_unit = norm(row[2]) if len(row) > 2 and row[2] else ''
        qty = safe_float(row[3]) if len(row) > 3 else 0
        pu = safe_float(row[5]) if len(row) > 5 else 0
        total = safe_float(row[6]) if len(row) > 6 else 0

        if pu <= 0 and qty > 0 and total > 0:
            pu = total / qty

        if pu <= 0:
            continue

        if ('gesso' in desc or 'forro' in desc) and 'negativo' not in desc and 'estucamento' not in desc:
            forro_items.append({'qty': qty, 'total': total, 'pu': pu})

    if forro_items:
        # Weighted average PU
        total_qty = sum(i['qty'] for i in forro_items)
        total_val = sum(i['total'] for i in forro_items)
        if total_qty > 0:
            return round(total_val / total_qty, 2)
    return None


# ============================================================================
# MAIN EXTRACTION LOGIC
# ============================================================================

def extract_from_ger_executivo_sheet(wb):
    """Extract esquadrias total and fachada total from Ger_Executivo summary rows."""
    ws = find_sheet(wb, ['Ger_Executivo', 'GER_EXECUTIVO', 'Ger Executivo'])
    if ws is None:
        return None, None

    esq_total = None
    fac_total = None

    desc_col = 9
    total_col = 13
    level_col = 7

    for row in ws.iter_rows(min_row=1, max_row=1500, max_col=20, values_only=True):
        level = str(row[level_col]).strip() if len(row) > level_col and row[level_col] else ''
        desc = str(row[desc_col]).strip() if len(row) > desc_col and row[desc_col] else ''
        total = safe_float(row[total_col]) if len(row) > total_col else 0

        if not desc:
            continue
        d = norm(desc)

        if level in ['CÉLULA CONSTRUTIVA', 'CELULA CONSTRUTIVA']:
            if 'esquadria' in d:
                esq_total = total
            elif 'fachada' in d:
                fac_total = total

    return esq_total, fac_total


def process_project(project):
    """Process a single project and return the extracted data."""
    slug = project['slug']
    path = resolve_path(project)

    if not os.path.exists(path):
        return slug, None, None, None, f"File not found: {path}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return slug, None, None, None, f"Cannot open: {e}"

    try:
        fmt = detect_format(wb)

        esquadrias = None
        fachada = None
        acabamentos = None

        if fmt == 'cartesian':
            # Try dedicated sheets first
            esquadrias = extract_esquadrias_cartesian(wb)
            fachada = extract_fachada_cartesian(wb)
            acabamentos = extract_acabamentos_pus_cartesian(wb)

            # Try to supplement PUs from other sheets
            if acabamentos is None:
                acabamentos = {
                    'porcelanato_pu_m2': None, 'ceramica_pu_m2': None, 'laminado_pu_m2': None,
                    'rodape_pu_m': None, 'contrapiso_pu_m2': None, 'forro_gesso_pu_m2': None,
                    'pintura_parede_pu_m2': None, 'chapisco_pu_m2': None, 'reboco_pu_m2': None
                }

            # Fill missing PUs from Rev. Internos sheet
            rev_pus = extract_pus_from_rev_internos(wb)
            if rev_pus:
                for k, v in rev_pus.items():
                    if acabamentos.get(k) is None and v is not None:
                        acabamentos[k] = v

            # Fill pintura PU from PINTURA sheet
            if acabamentos.get('pintura_parede_pu_m2') is None:
                pin_pu = extract_pus_from_pintura(wb)
                if pin_pu:
                    acabamentos['pintura_parede_pu_m2'] = pin_pu

            # Fill forro PU from TETO sheet
            if acabamentos.get('forro_gesso_pu_m2') is None:
                forro_pu = extract_pus_from_forro(wb)
                if forro_pu:
                    acabamentos['forro_gesso_pu_m2'] = forro_pu

            # If esquadrias sheet didn't have data, try Ger_Executivo
            if esquadrias is None:
                esquadrias = extract_esquadrias_ger_executivo(wb)

            # Get totals from Ger_Executivo if missing
            if (esquadrias and esquadrias.get('total') is None) or (fachada and fachada.get('total') is None):
                esq_t, fac_t = extract_from_ger_executivo_sheet(wb)
                if esquadrias and esquadrias.get('total') is None:
                    esquadrias['total'] = esq_t
                if fachada and fachada.get('total') is None:
                    fachada['total'] = fac_t

            # Final fallback: Resumo totals
            if (esquadrias is None or esquadrias.get('total') is None) or (fachada is None or fachada.get('total') is None):
                esq_t_r, fac_t_r = extract_totals_from_resumo(wb)
                if esquadrias is None and esq_t_r:
                    esquadrias = {
                        'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                        'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                        'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                        'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                        'vidros': {'valor': None}, 'pele_vidro': {'valor': None},
                        'brises': {'valor': None}, 'total': esq_t_r
                    }
                elif esquadrias and esquadrias.get('total') is None and esq_t_r:
                    esquadrias['total'] = esq_t_r
                if fachada is None and fac_t_r:
                    fachada = {
                        'reboco_externo': {'valor': None},
                        'textura_pintura_ext': {'valor': None},
                        'pastilha_porcelanato': {'valor': None},
                        'juntas': {'valor': None}, 'total': fac_t_r
                    }
                elif fachada and fachada.get('total') is None and fac_t_r:
                    fachada['total'] = fac_t_r

        elif fmt == 'relatorio':
            esquadrias = extract_esquadrias_relatorio(wb)
            fachada = extract_fachada_relatorio(wb)
            acabamentos = extract_acabamentos_pus_relatorio(wb)

        elif fmt == 'ger_executivo':
            esquadrias = extract_esquadrias_ger_executivo(wb)
            fachada = None  # Ger_Executivo does not have fachada detail inline
            esq_t, fac_t = extract_from_ger_executivo_sheet(wb)
            if fac_t:
                fachada = {
                    'reboco_externo': {'valor': None},
                    'textura_pintura_ext': {'valor': None},
                    'pastilha_porcelanato': {'valor': None},
                    'juntas': {'valor': None},
                    'total': fac_t
                }
            acabamentos = {
                'porcelanato_pu_m2': None, 'ceramica_pu_m2': None, 'laminado_pu_m2': None,
                'rodape_pu_m': None, 'contrapiso_pu_m2': None, 'forro_gesso_pu_m2': None,
                'pintura_parede_pu_m2': None, 'chapisco_pu_m2': None, 'reboco_pu_m2': None
            }

        elif fmt == 'orcamento_sienge':
            # Try Relatório sheet if it exists, else limited extraction
            esquadrias = extract_esquadrias_relatorio(wb)
            fachada = extract_fachada_relatorio(wb)
            acabamentos = extract_acabamentos_pus_relatorio(wb)

            # Fallback: get totals from Resumo
            esq_t_r, fac_t_r = extract_totals_from_resumo(wb)
            if esquadrias is None or esquadrias.get('total') is None:
                if esq_t_r:
                    if esquadrias is None:
                        esquadrias = {
                            'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                            'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                            'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                            'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                            'vidros': {'valor': None}, 'pele_vidro': {'valor': None},
                            'brises': {'valor': None}, 'total': esq_t_r
                        }
                    else:
                        esquadrias['total'] = esq_t_r
            if fachada is None or fachada.get('total') is None:
                if fac_t_r:
                    if fachada is None:
                        fachada = {
                            'reboco_externo': {'valor': None},
                            'textura_pintura_ext': {'valor': None},
                            'pastilha_porcelanato': {'valor': None},
                            'juntas': {'valor': None}, 'total': fac_t_r
                        }
                    else:
                        fachada['total'] = fac_t_r

        elif fmt == 'resumo':
            # Resumo-only format - get totals
            esq_t_r, fac_t_r = extract_totals_from_resumo(wb)
            if esq_t_r:
                esquadrias = {
                    'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                    'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                    'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                    'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                    'vidros': {'valor': None}, 'pele_vidro': {'valor': None},
                    'brises': {'valor': None}, 'total': esq_t_r
                }
            if fac_t_r:
                fachada = {
                    'reboco_externo': {'valor': None},
                    'textura_pintura_ext': {'valor': None},
                    'pastilha_porcelanato': {'valor': None},
                    'juntas': {'valor': None}, 'total': fac_t_r
                }

        else:
            # Unknown format - try all extractors
            esquadrias = extract_esquadrias_cartesian(wb)
            if esquadrias is None:
                esquadrias = extract_esquadrias_relatorio(wb)
            if esquadrias is None:
                esquadrias = extract_esquadrias_ger_executivo(wb)

            fachada = extract_fachada_cartesian(wb)
            if fachada is None:
                fachada = extract_fachada_relatorio(wb)

            acabamentos = extract_acabamentos_pus_cartesian(wb)
            if acabamentos is None:
                acabamentos = extract_acabamentos_pus_relatorio(wb)

            # Final fallback: Resumo totals
            esq_t_r, fac_t_r = extract_totals_from_resumo(wb)
            if esquadrias is None or esquadrias.get('total') is None:
                if esq_t_r:
                    if esquadrias is None:
                        esquadrias = {
                            'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                            'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                            'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                            'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                            'vidros': {'valor': None}, 'pele_vidro': {'valor': None},
                            'brises': {'valor': None}, 'total': esq_t_r
                        }
                    else:
                        esquadrias['total'] = esq_t_r
            if fachada is None or fachada.get('total') is None:
                if fac_t_r:
                    if fachada is None:
                        fachada = {
                            'reboco_externo': {'valor': None},
                            'textura_pintura_ext': {'valor': None},
                            'pastilha_porcelanato': {'valor': None},
                            'juntas': {'valor': None}, 'total': fac_t_r
                        }
                    else:
                        fachada['total'] = fac_t_r

        # Provide defaults for None results
        if esquadrias is None:
            esquadrias = {
                'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                'vidros': {'valor': None},
                'pele_vidro': {'valor': None},
                'brises': {'valor': None},
                'total': None
            }

        if fachada is None:
            fachada = {
                'reboco_externo': {'valor': None},
                'textura_pintura_ext': {'valor': None},
                'pastilha_porcelanato': {'valor': None},
                'juntas': {'valor': None},
                'total': None
            }

        if acabamentos is None:
            acabamentos = {
                'porcelanato_pu_m2': None, 'ceramica_pu_m2': None, 'laminado_pu_m2': None,
                'rodape_pu_m': None, 'contrapiso_pu_m2': None, 'forro_gesso_pu_m2': None,
                'pintura_parede_pu_m2': None, 'chapisco_pu_m2': None, 'reboco_pu_m2': None
            }

        wb.close()
        return slug, esquadrias, fachada, acabamentos, f"OK ({fmt})"

    except Exception as e:
        wb.close()
        return slug, None, None, None, f"Error: {traceback.format_exc()}"


def update_json(slug, esquadrias, fachada, acabamentos):
    """Update the indices-executivo JSON file."""
    idx_path = INDICES_DIR / f"{slug}.json"
    if not idx_path.exists():
        print(f"  WARNING: {idx_path} not found, skipping")
        return False

    with open(idx_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if esquadrias is not None:
        data['esquadrias_detail'] = esquadrias
    if fachada is not None:
        data['fachada_detail'] = fachada
    if acabamentos is not None:
        data['acabamentos_pus'] = acabamentos

    with open(idx_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return True


# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    print(f"Processing {len(BATCH)} projects (batch 0-41)...")
    print("=" * 80)

    stats = {
        'ok': 0, 'error': 0, 'not_found': 0,
        'esquadrias_found': 0, 'fachada_found': 0, 'acabamentos_found': 0,
        'esquadrias_with_data': 0, 'fachada_with_data': 0, 'acabamentos_with_data': 0
    }

    for i, project in enumerate(BATCH):
        slug = project['slug']
        print(f"\n[{i+1}/{len(BATCH)}] {slug}")

        slug, esquadrias, fachada, acabamentos, status = process_project(project)
        print(f"  Status: {status}")

        if 'not found' in status.lower():
            stats['not_found'] += 1
            # Still update with null defaults
            update_json(slug, esquadrias or {
                'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                'vidros': {'valor': None}, 'pele_vidro': {'valor': None},
                'brises': {'valor': None}, 'total': None
            }, fachada or {
                'reboco_externo': {'valor': None}, 'textura_pintura_ext': {'valor': None},
                'pastilha_porcelanato': {'valor': None}, 'juntas': {'valor': None}, 'total': None
            }, acabamentos or {
                'porcelanato_pu_m2': None, 'ceramica_pu_m2': None, 'laminado_pu_m2': None,
                'rodape_pu_m': None, 'contrapiso_pu_m2': None, 'forro_gesso_pu_m2': None,
                'pintura_parede_pu_m2': None, 'chapisco_pu_m2': None, 'reboco_pu_m2': None
            })
            continue

        if 'error' in status.lower() or 'cannot open' in status.lower():
            stats['error'] += 1
            # Update with null defaults
            update_json(slug, {
                'aluminio': {'valor': None, 'area_m2': None, 'pu_m2': None},
                'guarda_corpo': {'valor': None, 'comprimento_m': None, 'pu_m': None},
                'portas_madeira': {'valor': None, 'qtd': None, 'pu_un': None},
                'portas_corta_fogo': {'valor': None, 'qtd': None, 'pu_un': None},
                'vidros': {'valor': None}, 'pele_vidro': {'valor': None},
                'brises': {'valor': None}, 'total': None
            }, {
                'reboco_externo': {'valor': None}, 'textura_pintura_ext': {'valor': None},
                'pastilha_porcelanato': {'valor': None}, 'juntas': {'valor': None}, 'total': None
            }, {
                'porcelanato_pu_m2': None, 'ceramica_pu_m2': None, 'laminado_pu_m2': None,
                'rodape_pu_m': None, 'contrapiso_pu_m2': None, 'forro_gesso_pu_m2': None,
                'pintura_parede_pu_m2': None, 'chapisco_pu_m2': None, 'reboco_pu_m2': None
            })
            continue

        stats['ok'] += 1

        # Check data richness
        if esquadrias:
            stats['esquadrias_found'] += 1
            has_data = any(v is not None for k, v in esquadrias.items()
                         if k != 'total' and isinstance(v, dict) and v.get('valor') is not None)
            if has_data or esquadrias.get('total') is not None:
                stats['esquadrias_with_data'] += 1
                print(f"  Esquadrias: total={esquadrias.get('total')}")
                if esquadrias.get('aluminio', {}).get('valor'):
                    print(f"    Alumínio: R${esquadrias['aluminio']['valor']:,.2f}")
                if esquadrias.get('portas_madeira', {}).get('valor'):
                    print(f"    Portas madeira: R${esquadrias['portas_madeira']['valor']:,.2f}")

        if fachada:
            stats['fachada_found'] += 1
            has_data = any(isinstance(v, dict) and v.get('valor') is not None
                         for k, v in fachada.items() if k != 'total')
            if has_data or fachada.get('total') is not None:
                stats['fachada_with_data'] += 1
                print(f"  Fachada: total={fachada.get('total')}")

        if acabamentos:
            stats['acabamentos_found'] += 1
            has_data = any(v is not None for v in acabamentos.values())
            if has_data:
                stats['acabamentos_with_data'] += 1
                filled = [k for k, v in acabamentos.items() if v is not None]
                print(f"  Acabamentos PUs: {len(filled)} PUs found ({', '.join(filled)})")

        updated = update_json(slug, esquadrias, fachada, acabamentos)
        if updated:
            print(f"  JSON updated")

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total projects: {len(BATCH)}")
    print(f"OK: {stats['ok']}")
    print(f"Errors: {stats['error']}")
    print(f"Not found: {stats['not_found']}")
    print(f"Esquadrias extracted: {stats['esquadrias_with_data']}/{stats['esquadrias_found']}")
    print(f"Fachada extracted: {stats['fachada_with_data']}/{stats['fachada_found']}")
    print(f"Acabamentos PUs extracted: {stats['acabamentos_with_data']}/{stats['acabamentos_found']}")
