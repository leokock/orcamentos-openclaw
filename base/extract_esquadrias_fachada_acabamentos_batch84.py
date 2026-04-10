#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract ESQUADRIAS, FACHADA, and ACABAMENTOS detail from construction budget spreadsheets.

Extracts:
  - esquadrias_detail: aluminio, guarda_corpo, portas_madeira, portas_corta_fogo,
    vidros, pele_vidro, brises (valor/qtd/pu where available)
  - fachada_detail: reboco_externo, textura_pintura_ext, pastilha_porcelanato, juntas (valor)
  - acabamentos_pus: porcelanato R$/m2, ceramica R$/m2, laminado R$/m2, rodape R$/m,
    contrapiso R$/m2, forro_gesso R$/m2, pintura R$/m2, chapisco R$/m2, reboco R$/m2

Sources: Relatório (Sienge N2/N3), Ger_Executivo/ORÇAMENTO sections,
         dedicated Esquadrias/Fachada/Acabamentos parametric sheets.

Batch: projects 84-125.
"""
import json
import sys
import os
import re
import traceback

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

BASE_DIR = r"C:\Users\leona\orcamentos-openclaw\base"
INDICES_DIR = os.path.join(BASE_DIR, "indices-executivo")
MAPPING_FILE = os.path.join(BASE_DIR, "_all_projects_mapping.json")
BATCH_START = 84
BATCH_END = 126

# ============================================================
# Helpers (same as other extraction scripts)
# ============================================================
def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def safe_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val, 2) if val != 0 else None
    s = str(val).strip().replace(",", ".")
    m = re.search(r'[\d]+[.]?[\d]*', s)
    if m:
        try:
            v = float(m.group())
            return round(v, 2) if v != 0 else None
        except:
            pass
    return None

def normalize(s):
    if not s:
        return ""
    s = s.lower().strip()
    for a, b in {'á':'a','à':'a','â':'a','ã':'a','é':'e','ê':'e','è':'e',
                  'í':'i','î':'i','ó':'o','ô':'o','õ':'o','ò':'o',
                  'ú':'u','û':'u','ù':'u','ç':'c','ñ':'n'}.items():
        s = s.replace(a, b)
    return s

def find_sheet(wb, candidates):
    sheets_lower = {s.lower().strip(): s for s in wb.sheetnames}
    for c in candidates:
        c_lower = c.lower().strip()
        if c_lower in sheets_lower:
            return wb[sheets_lower[c_lower]]
    for c in candidates:
        c_lower = c.lower().strip()
        for k, v in sheets_lower.items():
            if c_lower in k:
                return wb[v]
    return None

def read_all_rows(ws, max_rows=1500):
    rows = []
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        rows.append(list(row))
    return rows

def is_valid_value(v):
    if v is None:
        return False
    if isinstance(v, (int, float)):
        if v > 50000000 and v == int(v):
            return False
        return v > 0
    return False

def detect_value_col(rows, max_scan=20):
    for rn, row in enumerate(rows[:max_scan]):
        if not row:
            continue
        for ci, cell in enumerate(row):
            s = normalize(safe_str(cell))
            if s in ['preco total', 'custo total', 'valor total', 'total r$',
                      'valor orcado', 'valor orcado', 'valor']:
                return ci
    return None

def detect_wide_value_col(rows, max_scan=20):
    """For Sienge Relatório format where value column can be col 28+."""
    for rn, row in enumerate(rows[:max_scan]):
        if not row:
            continue
        for ci, cell in enumerate(row):
            s = normalize(safe_str(cell))
            if s in ['preco total', 'custo total', 'valor total', 'total r$',
                      'valor orcado', 'valor orcado', 'valor']:
                return ci
    # Fallback: scan first data rows for rightmost numeric column
    for rn, row in enumerate(rows[10:25]):
        if not row:
            continue
        code = safe_str(row[0]).strip()
        if re.match(r'^\d{2}\s*$', code):
            # N1 row found, look for rightmost value
            for ci in range(len(row)-1, 5, -1):
                v = safe_num(row[ci])
                if v and is_valid_value(v) and v > 100:
                    return ci
    return None

def get_row_desc(row, max_col=10):
    for ci in range(min(len(row), max_col)):
        s = safe_str(row[ci])
        if len(s) > 3 and not re.match(r'^[\d.,\s]+$', s) and s not in (
            'CÉLULA CONSTRUTIVA', 'ETAPA', 'SUBETAPA', 'SERVIÇO',
            'vb', 'un', 'mes', 'm', 'm2', 'm3', 'kg', 'm²', 'm³'):
            return s
    return ""

def get_pu_from_row(row, val_col=None, qty_col=None):
    """Try to extract unit price from a row. Returns (pu, qty, total)."""
    total = None
    qty = None
    pu = None

    # If we know columns, use them
    if val_col is not None and val_col < len(row):
        total = safe_num(row[val_col])

    # Look for qty and pu columns near each other
    # In Sienge format: col layout is typically code, desc, ..., un, qty, pu, total
    if qty_col is not None and qty_col < len(row):
        qty = safe_num(row[qty_col])

    if total and qty and qty > 0:
        pu = round(total / qty, 2)

    return pu, qty, total


# ============================================================
# Extract from Relatório (Sienge format with N1/N2/N3 codes)
# ============================================================
def extract_from_relatorio(wb):
    ws = find_sheet(wb, ['Relatório', 'Relatorio'])
    if not ws:
        return None, None, None

    rows = read_all_rows(ws, max_rows=1500)
    val_col = detect_wide_value_col(rows)

    if val_col is None:
        return None, None, None

    # Detect qty column (usually val_col - 2 or so)
    qty_col = None
    un_col = None
    for rn, row in enumerate(rows[:15]):
        if not row:
            continue
        for ci, cell in enumerate(row):
            s = normalize(safe_str(cell))
            if s in ['quantidade orcada', 'quantidade', 'qtde', 'qtd']:
                qty_col = ci
            if s in ['un.', 'un', 'unidade']:
                un_col = ci

    # Detect code and desc columns
    code_col = 0
    desc_col = 1

    esquadrias = {}
    fachada = {}
    acabamentos_pus = {}

    current_n1 = ""
    current_n1_code = ""

    for rn, row in enumerate(rows):
        if not row or len(row) <= max(code_col, desc_col):
            continue

        code = safe_str(row[code_col]).strip()
        desc = safe_str(row[desc_col]) if desc_col < len(row) else ""
        val = safe_num(row[val_col]) if val_col and val_col < len(row) else None
        qty = safe_num(row[qty_col]) if qty_col and qty_col < len(row) else None

        desc_norm = normalize(desc)

        is_n1 = bool(re.match(r'^\d{2}\s*$', code))
        is_n2 = bool(re.match(r'^\d{2}\.\d{3}\s*$', code))
        is_n3 = bool(re.match(r'^\d{2}\.\d{3}\.\d{3}\s*$', code))
        is_n4 = bool(re.match(r'^\d{2}\.\d{3}\.\d{3}\.\d{3}\s*$', code))

        if is_n1:
            current_n1 = desc_norm
            current_n1_code = code.strip()

        # ──────────────────────────────────────────────
        # ESQUADRIAS extraction (N3 level under ESQUADRIAS N1)
        # ──────────────────────────────────────────────
        if 'esquadria' in current_n1 or 'vidro' in current_n1 or 'ferragem' in current_n1:
            if is_n3 and val and is_valid_value(val):
                # Alumínio
                if any(k in desc_norm for k in ['esquadrias de aluminio', 'esquadria de aluminio',
                                                   'aluminio', 'caixilhos de aluminio']):
                    if 'aluminio' not in esquadrias or val > (esquadrias['aluminio'].get('valor') or 0):
                        entry = {'valor': val}
                        if qty:
                            entry['qtd'] = qty
                            entry['pu'] = round(val / qty, 2)
                        esquadrias.setdefault('aluminio', {})
                        esquadrias['aluminio']['valor'] = esquadrias.get('aluminio', {}).get('valor', 0) + val
                        if qty:
                            esquadrias['aluminio']['qtd'] = esquadrias.get('aluminio', {}).get('qtd', 0) + qty

                # Guarda corpo
                elif any(k in desc_norm for k in ['guarda corpo', 'guarda-corpo', 'gradil', 'corrimao']):
                    esquadrias.setdefault('guarda_corpo', {})
                    esquadrias['guarda_corpo']['valor'] = esquadrias.get('guarda_corpo', {}).get('valor', 0) + val
                    if qty:
                        esquadrias['guarda_corpo']['qtd'] = esquadrias.get('guarda_corpo', {}).get('qtd', 0) + qty

                # Portas de madeira
                elif any(k in desc_norm for k in ['esquadrias de madeira', 'esquadria de madeira',
                                                     'portas de madeira', 'porta de madeira',
                                                     'porta interna', 'kit porta']):
                    esquadrias.setdefault('portas_madeira', {})
                    esquadrias['portas_madeira']['valor'] = esquadrias.get('portas_madeira', {}).get('valor', 0) + val
                    if qty:
                        esquadrias['portas_madeira']['qtd'] = esquadrias.get('portas_madeira', {}).get('qtd', 0) + qty

                # Portas corta fogo
                elif any(k in desc_norm for k in ['corta fogo', 'corta-fogo', 'pcf']):
                    esquadrias.setdefault('portas_corta_fogo', {})
                    esquadrias['portas_corta_fogo']['valor'] = esquadrias.get('portas_corta_fogo', {}).get('valor', 0) + val
                    if qty:
                        esquadrias['portas_corta_fogo']['qtd'] = esquadrias.get('portas_corta_fogo', {}).get('qtd', 0) + qty

                # Vidros (pele de vidro, vidro temperado, etc.)
                elif any(k in desc_norm for k in ['pele de vidro', 'pele vidro']):
                    esquadrias.setdefault('pele_vidro', {})
                    esquadrias['pele_vidro']['valor'] = esquadrias.get('pele_vidro', {}).get('valor', 0) + val
                    if qty:
                        esquadrias['pele_vidro']['qtd'] = esquadrias.get('pele_vidro', {}).get('qtd', 0) + qty
                elif any(k in desc_norm for k in ['vidro', 'envidracamento']):
                    esquadrias.setdefault('vidros', {})
                    esquadrias['vidros']['valor'] = esquadrias.get('vidros', {}).get('valor', 0) + val
                    if qty:
                        esquadrias['vidros']['qtd'] = esquadrias.get('vidros', {}).get('qtd', 0) + qty

                # Brises
                elif any(k in desc_norm for k in ['brise', 'brisol']):
                    esquadrias.setdefault('brises', {})
                    esquadrias['brises']['valor'] = esquadrias.get('brises', {}).get('valor', 0) + val
                    if qty:
                        esquadrias['brises']['qtd'] = esquadrias.get('brises', {}).get('qtd', 0) + qty

                # Serralheria / metálicas
                elif any(k in desc_norm for k in ['metalica', 'serralheria']):
                    esquadrias.setdefault('metalicas_serralheria', {})
                    esquadrias['metalicas_serralheria']['valor'] = esquadrias.get('metalicas_serralheria', {}).get('valor', 0) + val

                # Ferragens
                elif any(k in desc_norm for k in ['ferragem', 'ferragens']):
                    esquadrias.setdefault('ferragens', {})
                    esquadrias['ferragens']['valor'] = esquadrias.get('ferragens', {}).get('valor', 0) + val

            # N4 level items for PU extraction
            if is_n4 and val and is_valid_value(val) and qty and qty > 0:
                pu = round(val / qty, 2)
                # Aluminium PU
                if any(k in desc_norm for k in ['esquadrias de aluminio', 'aluminio']):
                    if 'aluminio' in esquadrias and 'pu' not in esquadrias['aluminio']:
                        esquadrias['aluminio']['pu'] = pu
                # Porta madeira PU
                elif any(k in desc_norm for k in ['kit porta', 'porta de madeira']):
                    if 'portas_madeira' in esquadrias and 'pu' not in esquadrias['portas_madeira']:
                        esquadrias['portas_madeira']['pu'] = pu
                # Guarda corpo PU
                elif any(k in desc_norm for k in ['guarda-corpo', 'guarda corpo']):
                    if 'guarda_corpo' in esquadrias and 'pu' not in esquadrias['guarda_corpo']:
                        esquadrias['guarda_corpo']['pu'] = pu

        # ──────────────────────────────────────────────
        # FACHADA extraction (N3 level under FACHADA N1)
        # ──────────────────────────────────────────────
        if 'fachada' in current_n1 or ('revestimento' in current_n1 and 'acabamento' in current_n1 and 'fachada' in current_n1):
            if is_n3 and val and is_valid_value(val):
                # Reboco externo / Revestimento argamassado
                if any(k in desc_norm for k in ['revestimento argamassado', 'reboco externo',
                                                   'reboco de fachada', 'emboço']):
                    fachada['reboco_externo'] = fachada.get('reboco_externo', 0) + val

                # Textura / Pintura externa
                elif any(k in desc_norm for k in ['pintura e textura', 'textura em fachada',
                                                     'pintura em fachada', 'pintura externa',
                                                     'textura externa']):
                    fachada['textura_pintura_ext'] = fachada.get('textura_pintura_ext', 0) + val

                # Pastilha / Porcelanato de fachada
                elif any(k in desc_norm for k in ['acabamentos especiais', 'pastilha', 'porcelanato',
                                                     'pedra natural', 'revestimento ceramico ext']):
                    fachada['pastilha_porcelanato'] = fachada.get('pastilha_porcelanato', 0) + val

                # Juntas / Tratamentos
                elif any(k in desc_norm for k in ['tratamento', 'junta', 'selante']):
                    fachada['juntas'] = fachada.get('juntas', 0) + val

            # Also catch N4 items with pastilha/porcelanato for more detail
            if is_n4 and val and is_valid_value(val) and ('fachada' in current_n1):
                if any(k in desc_norm for k in ['pastilha ceramica', 'pastilha']):
                    fachada.setdefault('pastilha_ceramica_valor', 0)
                    fachada['pastilha_ceramica_valor'] += val
                elif any(k in desc_norm for k in ['porcelanato amadeirado', 'porcelanato']):
                    fachada.setdefault('porcelanato_fachada_valor', 0)
                    fachada['porcelanato_fachada_valor'] += val

        # ──────────────────────────────────────────────
        # ACABAMENTOS PUs extraction (N4 items)
        # ──────────────────────────────────────────────
        if is_n4 and val and is_valid_value(val) and qty and qty > 0:
            pu = round(val / qty, 2)
            un = safe_str(row[un_col]).lower().strip() if un_col and un_col < len(row) else ""

            # Porcelanato piso (R$/m2)
            if any(k in desc_norm for k in ['porcelanato']) and ('piso' in desc_norm or 'piso' in current_n1.lower()):
                if un in ['m²', 'm2', 'm2', ''] or pu < 500:
                    if 'porcelanato' not in acabamentos_pus or pu > acabamentos_pus.get('porcelanato', 0):
                        acabamentos_pus['porcelanato'] = pu

            # Cerâmica parede (R$/m2)
            if any(k in desc_norm for k in ['ceramica', 'ceramico']):
                if un in ['m²', 'm2', ''] or pu < 500:
                    if 'ceramica' not in acabamentos_pus or pu > acabamentos_pus.get('ceramica', 0):
                        acabamentos_pus['ceramica'] = pu

            # Laminado (R$/m2)
            if any(k in desc_norm for k in ['laminado', 'vinilico']):
                if un in ['m²', 'm2', ''] or pu < 500:
                    if 'laminado' not in acabamentos_pus:
                        acabamentos_pus['laminado'] = pu

            # Contrapiso (R$/m2)
            if any(k in desc_norm for k in ['contrapiso']):
                if un in ['m²', 'm2', ''] or pu < 200:
                    if 'contrapiso' not in acabamentos_pus or pu > acabamentos_pus.get('contrapiso', 0):
                        acabamentos_pus['contrapiso'] = pu

            # Forro de gesso (R$/m2)
            if any(k in desc_norm for k in ['forro de gesso', 'forro gesso']):
                if un in ['m²', 'm2', ''] or pu < 200:
                    if 'forro_gesso' not in acabamentos_pus:
                        acabamentos_pus['forro_gesso'] = pu

            # Chapisco (R$/m2)
            if any(k in desc_norm for k in ['chapisco']):
                if un in ['m²', 'm2', ''] or pu < 100:
                    if 'chapisco' not in acabamentos_pus or pu > acabamentos_pus.get('chapisco', 0):
                        acabamentos_pus['chapisco'] = pu

            # Reboco interno (R$/m2)
            if any(k in desc_norm for k in ['reboco', 'massa unica']) and 'externo' not in desc_norm and 'fachada' not in current_n1:
                if un in ['m²', 'm2', ''] or pu < 200:
                    if 'reboco' not in acabamentos_pus or pu > acabamentos_pus.get('reboco', 0):
                        acabamentos_pus['reboco'] = pu

            # Pintura interna (R$/m2)
            if any(k in desc_norm for k in ['pintura com tinta', 'tinta latex', 'tinta acrilica']):
                if 'externo' not in desc_norm and 'fachada' not in current_n1:
                    if un in ['m²', 'm2', ''] or pu < 100:
                        if 'pintura' not in acabamentos_pus or pu > acabamentos_pus.get('pintura', 0):
                            acabamentos_pus['pintura'] = pu

            # Rodapé (R$/m)
            if any(k in desc_norm for k in ['rodape']):
                if un in ['m', 'm', ''] or pu < 100:
                    if 'rodape' not in acabamentos_pus:
                        acabamentos_pus['rodape'] = pu

    # Compute PUs for esquadrias where possible
    for key in ['aluminio', 'guarda_corpo', 'portas_madeira']:
        if key in esquadrias and 'valor' in esquadrias[key] and 'qtd' in esquadrias[key]:
            q = esquadrias[key]['qtd']
            if q and q > 0:
                esquadrias[key]['pu'] = round(esquadrias[key]['valor'] / q, 2)

    # Round all values
    for k, v in esquadrias.items():
        if isinstance(v, dict):
            for sk in v:
                if isinstance(v[sk], float):
                    v[sk] = round(v[sk], 2)
    for k in fachada:
        if isinstance(fachada[k], float):
            fachada[k] = round(fachada[k], 2)

    esq_result = esquadrias if esquadrias else None
    fac_result = fachada if fachada else None
    acb_result = acabamentos_pus if acabamentos_pus else None

    return esq_result, fac_result, acb_result


# ============================================================
# Extract from ORÇAMENTO / EAP / Ger_Executivo (Cartesian format)
# ============================================================
def detect_eap_format(rows):
    """Detect Cartesian EAP format where col 7 = level type, col 8 = code, col 9 = desc."""
    for rn, row in enumerate(rows[:20]):
        if not row or len(row) < 10:
            continue
        for ci in range(5, min(len(row), 12)):
            s = safe_str(row[ci]).upper()
            if s in ('CÉLULA CONSTRUTIVA', 'CELULA CONSTRUTIVA', 'ETAPA', 'SUBETAPA', 'SERVIÇO', 'SERVICO'):
                # Found level column, desc is likely ci+2
                return ci, ci+1, ci+2
            # Header row: look for 'Atalho' / 'NIVEL' / 'Item' / 'Descrição'
            if s in ('NIVEL', 'NÍVEL'):
                return ci, ci+1, ci+2
    return None, None, None

def extract_from_orc_executivo(wb):
    ws = find_sheet(wb, ['orçamento', 'ORÇAMENTO_EXECUTIVO', 'ORÇAMENTO_EXECUTIVO_COMENTADO',
                          'EAP', 'Executivo', 'Executivo ', 'Gerenciamento executivo',
                          'Gerenciamento executivo ', 'ORÇAMENTO_PARAMETRICO',
                          'ORÇAMENTO_PARAMÉTRICO', 'ORÇAMENTO_PRIVATIVA'])
    if not ws:
        return None, None, None

    rows = read_all_rows(ws, max_rows=1500)

    # Detect format: Cartesian EAP (level col, code col, desc col) or standard
    level_col, code_col_eap, desc_col_eap = detect_eap_format(rows)

    val_col = detect_value_col(rows)

    if val_col is None:
        val_col = detect_wide_value_col(rows)

    if val_col is None:
        # For Cartesian EAP: look for "Total" or "Preço" headers
        for rn, row in enumerate(rows[:10]):
            if not row:
                continue
            for ci in range(10, min(len(row), 20)):
                s = normalize(safe_str(row[ci]))
                if s in ['total', 'preco total', 'custo total', 'valor total']:
                    val_col = ci
                    break
            if val_col:
                break

    if val_col is None:
        # Fallback: rightmost column with values > 1000
        for row in rows[:30]:
            if not row:
                continue
            for ci in range(len(row)-1, 5, -1):
                if isinstance(row[ci], (int, float)) and row[ci] > 1000 and is_valid_value(row[ci]):
                    val_col = ci
                    break
            if val_col:
                break

    if val_col is None:
        return None, None, None

    # Detect qty column (usually 2 before val_col in Cartesian format)
    qty_col_eap = None
    un_col_eap = None
    if val_col:
        for rn, row in enumerate(rows[:10]):
            if not row:
                continue
            for ci in range(max(0, val_col-5), val_col):
                s = normalize(safe_str(row[ci]))
                if s in ['quant.', 'quant', 'quantidade', 'qtde', 'qtd']:
                    qty_col_eap = ci
                if s in ['unidade', 'un.', 'un']:
                    un_col_eap = ci

    esquadrias = {}
    fachada = {}
    acabamentos_pus = {}

    current_section = ""
    is_n3_level = False

    for row in rows:
        if not row or len(row) <= val_col:
            continue

        # Extract code, desc, val based on detected format
        if level_col is not None and code_col_eap is not None and desc_col_eap is not None:
            level_type = normalize(safe_str(row[level_col])) if level_col < len(row) else ""
            code = safe_str(row[code_col_eap]).strip() if code_col_eap < len(row) else ""
            desc = safe_str(row[desc_col_eap]) if desc_col_eap < len(row) else ""
        else:
            level_type = ""
            code = safe_str(row[0]).strip()
            desc = get_row_desc(row, 10)

        val = safe_num(row[val_col])
        desc_norm = normalize(desc)
        qty = safe_num(row[qty_col_eap]) if qty_col_eap and qty_col_eap < len(row) else None
        un = safe_str(row[un_col_eap]).lower() if un_col_eap and un_col_eap < len(row) else ""

        if not desc_norm:
            continue

        # Detect section level
        is_n1 = 'celula construtiva' in level_type or bool(re.match(r'^\d{2}$', code))
        is_n2 = 'etapa' in level_type or bool(re.match(r'^\d{2}\.\d{3}$', code))
        is_n3 = 'subetapa' in level_type or bool(re.match(r'^\d{2}\.\d{3}\.\d{3}$', code))
        is_n4 = 'servico' in level_type or 'serviço' in safe_str(row[level_col] if level_col and level_col < len(row) else "").lower() or bool(re.match(r'^\d{2}\.\d{3}\.\d{3}\.\d{3}$', code))

        if is_n1:
            if 'esquadria' in desc_norm:
                current_section = 'esquadrias'
            elif 'fachada' in desc_norm:
                current_section = 'fachada'
            elif 'acabament' in desc_norm and ('piso' in desc_norm or 'parede' in desc_norm):
                current_section = 'acabamentos'
            elif 'revestiment' in desc_norm and 'interno' in desc_norm:
                current_section = 'revestimentos'
            elif 'pintura' in desc_norm and 'intern' in desc_norm:
                current_section = 'pintura'
            else:
                current_section = ""

        if not val or not is_valid_value(val):
            continue

        # ESQUADRIAS section (N3 = SUBETAPA level items)
        if current_section == 'esquadrias' and (is_n3 or is_n2):
            if any(k in desc_norm for k in ['esquadrias de aluminio', 'aluminio', 'janela']):
                if 'porta' not in desc_norm:
                    esquadrias.setdefault('aluminio', {})
                    esquadrias['aluminio']['valor'] = esquadrias.get('aluminio', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['guarda corpo', 'guarda-corpo', 'guarda-corpos', 'gradil', 'corrimao', 'corrimoes']):
                esquadrias.setdefault('guarda_corpo', {})
                esquadrias['guarda_corpo']['valor'] = esquadrias.get('guarda_corpo', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['porta de madeira', 'portas de madeira', 'esquadrias de madeira',
                                                 'porta interna', 'kit porta', 'portas']):
                if 'corta fogo' not in desc_norm and 'aluminio' not in desc_norm:
                    esquadrias.setdefault('portas_madeira', {})
                    esquadrias['portas_madeira']['valor'] = esquadrias.get('portas_madeira', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['corta fogo', 'corta-fogo', 'pcf']):
                esquadrias.setdefault('portas_corta_fogo', {})
                esquadrias['portas_corta_fogo']['valor'] = esquadrias.get('portas_corta_fogo', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['pele de vidro', 'pele vidro']):
                esquadrias.setdefault('pele_vidro', {})
                esquadrias['pele_vidro']['valor'] = esquadrias.get('pele_vidro', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['vidro', 'envidracamento']):
                esquadrias.setdefault('vidros', {})
                esquadrias['vidros']['valor'] = esquadrias.get('vidros', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['brise']):
                esquadrias.setdefault('brises', {})
                esquadrias['brises']['valor'] = esquadrias.get('brises', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['metalica', 'serralheria', 'contramarco']):
                esquadrias.setdefault('metalicas_serralheria', {})
                esquadrias['metalicas_serralheria']['valor'] = esquadrias.get('metalicas_serralheria', {}).get('valor', 0) + val
            elif any(k in desc_norm for k in ['ferragem', 'ferragens']):
                esquadrias.setdefault('ferragens', {})
                esquadrias['ferragens']['valor'] = esquadrias.get('ferragens', {}).get('valor', 0) + val

        # FACHADA section (N3 level)
        elif current_section == 'fachada' and (is_n3 or is_n2):
            if any(k in desc_norm for k in ['reboco', 'argamassado', 'emboco', 'revestimento argamassado']):
                fachada['reboco_externo'] = fachada.get('reboco_externo', 0) + val
            elif any(k in desc_norm for k in ['textura', 'pintura']):
                fachada['textura_pintura_ext'] = fachada.get('textura_pintura_ext', 0) + val
            elif any(k in desc_norm for k in ['pastilha', 'porcelanato', 'pedra', 'acabamentos especiais']):
                fachada['pastilha_porcelanato'] = fachada.get('pastilha_porcelanato', 0) + val
            elif any(k in desc_norm for k in ['junta', 'selante', 'tratamento']):
                fachada['juntas'] = fachada.get('juntas', 0) + val

        # ACABAMENTOS PUs from N4 items
        if is_n4 and val and is_valid_value(val) and qty and qty > 0:
            pu = round(val / qty, 2)
            # Porcelanato
            if 'porcelanato' in desc_norm and ('piso' in desc_norm or current_section in ('acabamentos', 'revestimentos')):
                if pu < 500 and ('porcelanato' not in acabamentos_pus or pu > acabamentos_pus.get('porcelanato', 0)):
                    acabamentos_pus['porcelanato'] = pu
            # Cerâmica
            if any(k in desc_norm for k in ['ceramica', 'ceramico']):
                if pu < 500 and ('ceramica' not in acabamentos_pus or pu > acabamentos_pus.get('ceramica', 0)):
                    acabamentos_pus['ceramica'] = pu
            # Contrapiso
            if 'contrapiso' in desc_norm:
                if pu < 200 and ('contrapiso' not in acabamentos_pus or pu > acabamentos_pus.get('contrapiso', 0)):
                    acabamentos_pus['contrapiso'] = pu
            # Forro gesso
            if any(k in desc_norm for k in ['forro de gesso', 'forro gesso']):
                if pu < 200 and 'forro_gesso' not in acabamentos_pus:
                    acabamentos_pus['forro_gesso'] = pu
            # Chapisco
            if 'chapisco' in desc_norm:
                if pu < 100 and ('chapisco' not in acabamentos_pus or pu > acabamentos_pus.get('chapisco', 0)):
                    acabamentos_pus['chapisco'] = pu
            # Reboco
            if any(k in desc_norm for k in ['reboco', 'massa unica']) and 'externo' not in desc_norm and current_section != 'fachada':
                if pu < 200 and ('reboco' not in acabamentos_pus or pu > acabamentos_pus.get('reboco', 0)):
                    acabamentos_pus['reboco'] = pu
            # Pintura
            if any(k in desc_norm for k in ['pintura', 'tinta latex', 'tinta acrilica']) and current_section != 'fachada':
                if pu < 100 and ('pintura' not in acabamentos_pus or pu > acabamentos_pus.get('pintura', 0)):
                    acabamentos_pus['pintura'] = pu
            # Rodapé
            if 'rodape' in desc_norm:
                if pu < 100 and 'rodape' not in acabamentos_pus:
                    acabamentos_pus['rodape'] = pu

    # Round all values
    for k, v in esquadrias.items():
        if isinstance(v, dict):
            for sk in v:
                if isinstance(v[sk], float):
                    v[sk] = round(v[sk], 2)
    for k in fachada:
        if isinstance(fachada[k], float):
            fachada[k] = round(fachada[k], 2)

    esq_result = esquadrias if esquadrias else None
    fac_result = fachada if fachada else None
    acb_result = acabamentos_pus if acabamentos_pus else None

    return esq_result, fac_result, acb_result


# ============================================================
# Extract from Parametric sheets (Esquadrias, Fachada, Acabamentos)
# ============================================================
def extract_from_parametric_sheets(wb):
    esquadrias = {}
    fachada = {}
    acabamentos_pus = {}

    # ── Parametric Esquadrias ──
    ws = find_sheet(wb, ['Esquadrias'])
    if ws:
        rows = read_all_rows(ws, max_rows=60)
        for row in rows:
            if not row:
                continue
            desc = safe_str(row[0])
            desc_norm = normalize(desc)

            # Parametric format: desc, _, _, qty, un, pu, un, ref, total
            val = safe_num(row[8]) if len(row) > 8 else None
            qty = safe_num(row[3]) if len(row) > 3 else None
            pu = safe_num(row[5]) if len(row) > 5 else None

            if not val or not is_valid_value(val):
                continue

            entry = {'valor': val}
            if qty:
                entry['qtd'] = qty
            if pu:
                entry['pu'] = pu

            if any(k in desc_norm for k in ['esquadrias de aluminio', 'aluminio']):
                if 'porta' not in desc_norm:
                    esquadrias['aluminio'] = entry
            elif any(k in desc_norm for k in ['guarda corpo', 'guarda-corpo']):
                esquadrias['guarda_corpo'] = entry
            elif any(k in desc_norm for k in ['esquadrias internas de madeira', 'portas de madeira',
                                                 'esquadria de madeira', 'porta interna']):
                esquadrias['portas_madeira'] = entry
            elif any(k in desc_norm for k in ['pcf', 'porta corta fogo', 'corta fogo']):
                esquadrias['portas_corta_fogo'] = entry
            elif any(k in desc_norm for k in ['pele de vidro']):
                esquadrias['pele_vidro'] = entry
            elif any(k in desc_norm for k in ['esquadrias de vidro', 'vidro']):
                esquadrias['vidros'] = entry
            elif any(k in desc_norm for k in ['brise']):
                esquadrias['brises'] = entry
            elif any(k in desc_norm for k in ['serralheria', 'metalica']):
                esquadrias['metalicas_serralheria'] = entry
            elif any(k in desc_norm for k in ['portao', 'portoes']):
                esquadrias.setdefault('aluminio', {})
                esquadrias['aluminio']['valor'] = esquadrias.get('aluminio', {}).get('valor', 0) + val

    # ── Parametric Fachada (Rev. Fachada) ──
    ws = find_sheet(wb, ['Rev. Fachada', 'Fachada', 'FACHADA'])
    if ws:
        rows = read_all_rows(ws, max_rows=60)
        for row in rows:
            if not row:
                continue
            desc = safe_str(row[0])
            desc_norm = normalize(desc)

            val = safe_num(row[8]) if len(row) > 8 else None
            qty = safe_num(row[3]) if len(row) > 3 else None
            pu = safe_num(row[5]) if len(row) > 5 else None

            if not val or not is_valid_value(val):
                continue

            if any(k in desc_norm for k in ['reboco', 'argamassa', 'chapisco ext']):
                fachada['reboco_externo'] = fachada.get('reboco_externo', 0) + val
            elif any(k in desc_norm for k in ['textura', 'pintura ext']):
                fachada['textura_pintura_ext'] = fachada.get('textura_pintura_ext', 0) + val
            elif any(k in desc_norm for k in ['pastilha', 'porcelanato', 'pedra', 'ceramica']):
                fachada['pastilha_porcelanato'] = fachada.get('pastilha_porcelanato', 0) + val
            elif any(k in desc_norm for k in ['junta', 'selante', 'tratamento']):
                fachada['juntas'] = fachada.get('juntas', 0) + val

    # ── Parametric Acabamentos ──
    ws = find_sheet(wb, ['Acabamentos de Piso e Parede', 'Acabamentos Piso Parede',
                          'ACABAMENTOS', 'Acabamentos'])
    if ws:
        rows = read_all_rows(ws, max_rows=60)
        for row in rows:
            if not row:
                continue
            desc = safe_str(row[0])
            desc_norm = normalize(desc)

            # Parametric format has PU in col 5 and unit in col 6
            pu = safe_num(row[5]) if len(row) > 5 else None
            un = safe_str(row[6]).lower() if len(row) > 6 else ""

            if not pu or pu <= 0:
                continue

            if 'porcelanato' in desc_norm and ('piso' in desc_norm or 'm2' in un or 'm²' in un):
                acabamentos_pus['porcelanato'] = pu
            elif 'ceramica' in desc_norm:
                acabamentos_pus['ceramica'] = pu
            elif 'laminado' in desc_norm or 'vinilico' in desc_norm:
                acabamentos_pus['laminado'] = pu
            elif 'rodape' in desc_norm:
                acabamentos_pus['rodape'] = pu

    # ── Parametric Rev. Internos Piso e Parede ──
    ws = find_sheet(wb, ['Rev. Internos Piso e Parede', 'Revestimentos Internos'])
    if ws:
        rows = read_all_rows(ws, max_rows=60)
        for row in rows:
            if not row:
                continue
            desc = safe_str(row[0])
            desc_norm = normalize(desc)

            pu = safe_num(row[5]) if len(row) > 5 else None
            if not pu or pu <= 0:
                continue

            if 'contrapiso' in desc_norm:
                acabamentos_pus['contrapiso'] = pu
            elif 'chapisco' in desc_norm:
                acabamentos_pus['chapisco'] = pu
            elif 'reboco' in desc_norm or 'massa unica' in desc_norm:
                acabamentos_pus['reboco'] = pu

    # ── Parametric Acabamentos de Teto ──
    ws = find_sheet(wb, ['Acabamentos de Teto', 'Acabamentos Teto'])
    if ws:
        rows = read_all_rows(ws, max_rows=60)
        for row in rows:
            if not row:
                continue
            desc = safe_str(row[0])
            desc_norm = normalize(desc)

            pu = safe_num(row[5]) if len(row) > 5 else None
            if not pu or pu <= 0:
                continue

            if any(k in desc_norm for k in ['forro de gesso', 'gesso', 'forro']):
                acabamentos_pus['forro_gesso'] = pu

    # ── Parametric Pintura Interna ──
    ws = find_sheet(wb, ['Pintura Interna', 'Pintura'])
    if ws:
        rows = read_all_rows(ws, max_rows=60)
        for row in rows:
            if not row:
                continue
            desc = safe_str(row[0])
            desc_norm = normalize(desc)

            pu = safe_num(row[5]) if len(row) > 5 else None
            if not pu or pu <= 0:
                continue

            if any(k in desc_norm for k in ['pintura', 'tinta']):
                if 'pintura' not in acabamentos_pus:
                    acabamentos_pus['pintura'] = pu

    # Round fachada values
    for k in fachada:
        if isinstance(fachada[k], float):
            fachada[k] = round(fachada[k], 2)

    esq_result = esquadrias if esquadrias else None
    fac_result = fachada if fachada else None
    acb_result = acabamentos_pus if acabamentos_pus else None

    return esq_result, fac_result, acb_result


# ============================================================
# Extract from Orçamento Resumo (high-level subtotals only)
# ============================================================
def extract_from_resumo(wb):
    ws = find_sheet(wb, ['Resumo', 'Orçamento Resumo'])
    if not ws:
        return None, None

    rows = read_all_rows(ws, max_rows=60)

    esquadrias_total = None
    fachada_total = None

    for row in rows:
        if not row:
            continue
        desc = safe_str(row[0])
        desc_norm = normalize(desc)

        # Find value in row
        val = None
        for ci in range(1, min(len(row), 15)):
            v = safe_num(row[ci])
            if v and is_valid_value(v) and v > 100:
                val = v
                break

        if not val:
            continue

        if any(k in desc_norm for k in ['esquadrias']):
            esquadrias_total = val
        elif any(k in desc_norm for k in ['fachada', 'revestimento e acabamento de fachada',
                                            'revestimentos e acabamentos de fachada']):
            fachada_total = val

    esq = {'total': esquadrias_total} if esquadrias_total else None
    fac = {'total': fachada_total} if fachada_total else None
    return esq, fac


# ============================================================
# Main processing per project
# ============================================================
def process_project(project_info, idx):
    slug = project_info['slug']
    path = project_info['path']

    result = {
        'esquadrias_detail': None,
        'fachada_detail': None,
        'acabamentos_pus': None,
    }

    if not os.path.exists(path):
        print(f"  [{idx}] {slug}: FILE NOT FOUND")
        return result

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets_lower = [s.lower().strip() for s in wb.sheetnames]

        has_rel = any('relatório' in s or 'relatorio' in s for s in sheets_lower)
        has_orc = any(s in ('orçamento_executivo', 'orcamento_executivo', 'orçamento_executivo_comentado',
                             'orçamento', 'orcamento', 'orçamento_parametrico', 'orçamento_paramétrico',
                             'orçamento_privativa') for s in sheets_lower)
        has_eap = any(s.strip() == 'eap' for s in sheets_lower)
        has_ger_exec = any('gerenciamento executivo' in s or 'ger_executivo' in s for s in sheets_lower)
        has_parametric_esq = any('esquadria' in s for s in sheets_lower)
        has_parametric_fac = any('fachada' in s or 'rev. fachada' in s for s in sheets_lower)
        has_parametric_acab = any('acabamento' in s for s in sheets_lower)
        has_resumo = any('resumo' in s for s in sheets_lower)

        # Strategy 1: Relatório (best source - has N1/N2/N3/N4 with values)
        if has_rel:
            esq, fac, acb = extract_from_relatorio(wb)
            if esq:
                result['esquadrias_detail'] = esq
            if fac:
                result['fachada_detail'] = fac
            if acb:
                result['acabamentos_pus'] = acb

        # Strategy 2: ORÇAMENTO_EXECUTIVO / EAP sheets
        if has_orc or has_eap or has_ger_exec:
            esq_orc, fac_orc, acb_orc = extract_from_orc_executivo(wb)
            if esq_orc and not result['esquadrias_detail']:
                result['esquadrias_detail'] = esq_orc
            elif esq_orc and result['esquadrias_detail']:
                # Merge
                for k, v in esq_orc.items():
                    if k not in result['esquadrias_detail']:
                        result['esquadrias_detail'][k] = v
            if fac_orc and not result['fachada_detail']:
                result['fachada_detail'] = fac_orc
            elif fac_orc and result['fachada_detail']:
                for k, v in fac_orc.items():
                    if k not in result['fachada_detail']:
                        result['fachada_detail'][k] = v
            if acb_orc and not result['acabamentos_pus']:
                result['acabamentos_pus'] = acb_orc

        # Strategy 3: Parametric dedicated sheets
        if has_parametric_esq or has_parametric_fac or has_parametric_acab:
            esq_p, fac_p, acb_p = extract_from_parametric_sheets(wb)
            if esq_p and not result['esquadrias_detail']:
                result['esquadrias_detail'] = esq_p
            elif esq_p and result['esquadrias_detail']:
                for k, v in esq_p.items():
                    if k not in result['esquadrias_detail']:
                        result['esquadrias_detail'][k] = v
            if fac_p and not result['fachada_detail']:
                result['fachada_detail'] = fac_p
            elif fac_p and result['fachada_detail']:
                for k, v in fac_p.items():
                    if k not in result['fachada_detail']:
                        result['fachada_detail'][k] = v
            if acb_p and not result['acabamentos_pus']:
                result['acabamentos_pus'] = acb_p
            elif acb_p and result['acabamentos_pus']:
                for k, v in acb_p.items():
                    if k not in result['acabamentos_pus']:
                        result['acabamentos_pus'][k] = v

        # Strategy 4: Resumo (fallback for totals only)
        if has_resumo:
            esq_r, fac_r = extract_from_resumo(wb)
            if esq_r and not result['esquadrias_detail']:
                result['esquadrias_detail'] = esq_r
            if fac_r and not result['fachada_detail']:
                result['fachada_detail'] = fac_r

        wb.close()

    except Exception as e:
        print(f"  [{idx}] {slug}: ERROR - {e}")
        traceback.print_exc()

    return result


# ============================================================
# Main
# ============================================================
def main():
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        all_projects = json.load(f)

    batch = all_projects[BATCH_START:BATCH_END]

    stats = {
        'total': len(batch),
        'esquadrias_extracted': 0,
        'fachada_extracted': 0,
        'acabamentos_extracted': 0,
        'errors': 0,
    }

    for i, project in enumerate(batch):
        idx = BATCH_START + i
        slug = project['slug']
        print(f"\n[{idx}] Processing: {slug}")

        try:
            extracted = process_project(project, idx)

            json_path = os.path.join(INDICES_DIR, f"{slug}.json")
            if not os.path.exists(json_path):
                print(f"  [{idx}] {slug}: JSON not found, skipping")
                continue

            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            updated = False

            if extracted['esquadrias_detail']:
                data['esquadrias_detail'] = extracted['esquadrias_detail']
                updated = True
                stats['esquadrias_extracted'] += 1
                e_str = json.dumps(extracted['esquadrias_detail'], ensure_ascii=False)
                print(f"  Esquadrias: {e_str[:200]}")

            if extracted['fachada_detail']:
                data['fachada_detail'] = extracted['fachada_detail']
                updated = True
                stats['fachada_extracted'] += 1
                f_str = json.dumps(extracted['fachada_detail'], ensure_ascii=False)
                print(f"  Fachada: {f_str[:200]}")

            if extracted['acabamentos_pus']:
                data['acabamentos_pus'] = extracted['acabamentos_pus']
                updated = True
                stats['acabamentos_extracted'] += 1
                a_str = json.dumps(extracted['acabamentos_pus'], ensure_ascii=False)
                print(f"  Acabamentos PUs: {a_str[:200]}")

            if updated:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"  -> Updated")
            else:
                print(f"  -> No new data extracted")

        except Exception as e:
            print(f"  [{idx}] {slug}: FATAL ERROR - {e}")
            traceback.print_exc()
            stats['errors'] += 1

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total projects: {stats['total']}")
    print(f"Esquadrias extracted: {stats['esquadrias_extracted']}")
    print(f"Fachada extracted: {stats['fachada_extracted']}")
    print(f"Acabamentos PUs extracted: {stats['acabamentos_extracted']}")
    print(f"Errors: {stats['errors']}")

if __name__ == '__main__':
    main()
