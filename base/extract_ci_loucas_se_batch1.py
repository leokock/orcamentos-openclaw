#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract detailed cost data for CI, Louças/Metais, and Sistemas Especiais
from construction budget spreadsheets.
Batch 1: entries 0-41

Updates each indices-executivo/{slug}.json with:
  - ci_detalhado
  - loucas_metais_detail
  - sistemas_especiais_detail
"""
import json
import os
import re
import sys
import traceback

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# Config
# ============================================================
BASE_DIR = r"C:\Users\leona\orcamentos-openclaw\base"
MAPPING_FILE = os.path.join(BASE_DIR, "_all_projects_mapping.json")
INDICES_DIR = os.path.join(BASE_DIR, "indices-executivo")
BATCH_START = 0
BATCH_END = 42  # exclusive


# ============================================================
# Helpers
# ============================================================
def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def safe_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val == 0:
            return None
        return round(val, 2)
    s = str(val).strip().replace(",", ".").replace("R$", "").replace(" ", "")
    m = re.search(r'[\d]+[.]?[\d]*', s)
    if m:
        try:
            v = float(m.group())
            if v == 0:
                return None
            return round(v, 2)
        except:
            pass
    return None

def safe_int(val):
    n = safe_num(val)
    if n is not None:
        return int(n)
    return None

def normalize(s):
    if not s:
        return ""
    s = str(s).lower().strip()
    for a, b in {'á':'a','à':'a','â':'a','ã':'a','é':'e','ê':'e','è':'e',
                  'í':'i','î':'i','ó':'o','ô':'o','õ':'o','ò':'o',
                  'ú':'u','û':'u','ù':'u','ç':'c','ñ':'n'}.items():
        s = s.replace(a, b)
    return s

def read_rows(ws, max_rows=200, max_cols=30):
    """Safely read rows from a worksheet."""
    rows = []
    try:
        for row in ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
            rows.append(tuple(row))
    except:
        pass
    return rows

def classify_sheet(name):
    """Classify a sheet by type for extraction purposes."""
    n = normalize(name)
    # CI-related
    if "ger_tec" in n or "ger tec" in n or "gerenciamento tecnico" in n:
        return "ger_tec_adm"
    if "projetos e gerenciamento" in n:
        return "ger_tec_adm"
    if n.strip() in ("epc's", "epcs", "epc"):
        return "epcs"
    if "canteiro" in n:
        return "canteiro"
    if "ensaio" in n:
        return "ensaios"
    if "custos_indiretos" in n or "custo indireto" in n:
        return "custos_indiretos"
    # Ger Executivo (has CI totals + discipline totals)
    if "ger_executivo" in n or "ger executivo" in n or "gerenciamento executivo" in n:
        return "ger_executivo"
    if "gerenciamento_exec" in n:
        return "ger_executivo"
    # Loucas e metais
    if "louca" in n or "metai" in n:
        return "loucas_metais"
    # Sistemas especiais / Equipamentos especiais
    if "sistema" in n and "especiai" in n:
        return "sistemas_especiais"
    if "equipamento" in n and "especiai" in n:
        return "sistemas_especiais"
    if "equipamento" in n and n.strip() == "equipamentos":
        return "equipamentos"
    # Piscina
    if "piscina" in n:
        return "piscina"
    # Relatorio (Sienge)
    if "relatorio" in n:
        return "relatorio"
    # Apresentacao / Orcamento Executivo summary
    if "orcamento_executivo" in n or "orcamento executivo" in n:
        return "apresentacao"
    if "apresentacao" in n or "resumo" in n:
        return "apresentacao"
    if "analise_custo" in n or "analise custo" in n or "analise_custos" in n:
        return "analise_custos"
    return "other"


# ============================================================
# CI Extraction
# ============================================================
def extract_ci_from_ger_tec_adm(rows, ci):
    """Extract CI from Ger_Tec e Adm sheets (description-based structure)."""
    # These sheets have hierarchical items with descriptions.
    # We look for section totals or aggregate from item-level values.
    for i, row in enumerate(rows):
        if not row or len(row) < 10:
            continue
        desc = normalize(safe_str(row[9]) if len(row) > 9 else "")
        if not desc:
            desc = normalize(safe_str(row[1]) if len(row) > 1 else "")

        # Look for "Descrição" column - the actual data columns vary
        # In many files: col I (8) = Item, col J (9) = Descrição,
        # values in later columns (K onwards)
        # Try to find total values in columns after description
        vals = []
        for j in range(10, min(len(row), 25)):
            v = safe_num(row[j])
            if v is not None and v > 100:
                vals.append((j, v))

        if not vals:
            continue

        # Get the largest value as likely total
        total = max(v for _, v in vals) if vals else None

        # Match CI categories
        if not total:
            continue

        # Level: check NIVEL column (usually col H = index 7)
        nivel = safe_str(row[7]) if len(row) > 7 else ""

        if "servicos tecnicos" in desc or "projetos" in desc and "consultoria" in desc:
            if nivel in ("CÉLULA CONSTRUTIVA", "C\u00c9LULA CONSTRUTIVA", "CELULA CONSTRUTIVA"):
                if ci.get("projetos_consultorias") is None or total > (ci.get("projetos_consultorias") or 0):
                    ci["projetos_consultorias"] = total
            elif "estudos" in desc and "projetos" in desc and "consultoria" in desc:
                if ci.get("projetos_consultorias") is None or total > (ci.get("projetos_consultorias") or 0):
                    ci["projetos_consultorias"] = total


def extract_ci_from_relatorio(rows, ci):
    """Extract CI from Sienge Relatório sheets."""
    # Format: col 0 = código, col 1 = descrição, col 28 = preço total
    current_n1 = None
    current_n1_val = None

    # Track subtotals
    projetos_total = None
    taxas_total = None
    equipe_adm_total = None
    epcs_total = None
    equipamentos_total = None
    ensaios_total = None
    canteiro_total = None

    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue
        code = safe_str(row[0])
        desc_raw = safe_str(row[1])
        desc = normalize(desc_raw)

        # Get total value - try column 28, then scan other columns
        total = None
        if len(row) > 28:
            total = safe_num(row[28])
        if total is None:
            for j in range(20, min(len(row), 35)):
                v = safe_num(row[j])
                if v is not None and v > 50:
                    total = v
                    break

        if not code or not desc:
            continue

        # Identify N1 sections (e.g., "01", "02")
        code_clean = code.strip()
        parts = code_clean.split(".")

        # N2 level (e.g., "01.001")
        if len(parts) == 2 and total:
            if any(k in desc for k in ["estudos", "projetos", "consultoria"]):
                if "ensaio" not in desc:
                    projetos_total = (projetos_total or 0) + total
            elif any(k in desc for k in ["taxa", "documento", "licenca"]):
                taxas_total = (taxas_total or 0) + total
            elif any(k in desc for k in ["administracao", "administrativo", "canteiro"]):
                # Try to split canteiro vs admin
                if "canteiro" in desc:
                    canteiro_total = (canteiro_total or 0) + total
                else:
                    equipe_adm_total = (equipe_adm_total or 0) + total
            elif any(k in desc for k in ["equipamento"]):
                if "protecao" in desc:
                    epcs_total = (epcs_total or 0) + total
                else:
                    equipamentos_total = (equipamentos_total or 0) + total
            elif "ensaio" in desc or "laudo" in desc:
                ensaios_total = (ensaios_total or 0) + total
            elif "epc" in desc or "protecao coletiva" in desc:
                epcs_total = (epcs_total or 0) + total

        # N3 level as fallback (e.g., "01.001.003" = Ensaios)
        elif len(parts) == 3 and total:
            if "ensaio" in desc or "laudo" in desc:
                if ensaios_total is None:
                    ensaios_total = total
            elif "canteiro" in desc:
                if canteiro_total is None:
                    canteiro_total = total
            elif "epc" in desc or "protecao coletiva" in desc or "equipamentos de protecao" in desc:
                if epcs_total is None:
                    epcs_total = total
            elif "equipamento" in desc and ("carga" in desc or "transporte" in desc or "obra" in desc):
                equipamentos_total = (equipamentos_total or 0) + total

        # N1 level totals (e.g., "01" = GERENCIAMENTO TÉCNICO)
        elif len(parts) == 1 and len(code_clean) <= 3 and total:
            if "gerenciamento tecnico" in desc or "servicos tecnicos" in desc:
                ci["projetos_consultorias"] = total  # entire Ger Técnico as proxy
            elif "gerenciamento administrativo" in desc or "administracao" in desc:
                if ci.get("equipe_adm") is None:
                    ci["equipe_adm"] = total

    # Apply extracted values
    if projetos_total and ci.get("projetos_consultorias") is None:
        ci["projetos_consultorias"] = projetos_total
    if taxas_total and ci.get("taxas_licencas") is None:
        ci["taxas_licencas"] = taxas_total
    if equipe_adm_total and ci.get("equipe_adm") is None:
        ci["equipe_adm"] = equipe_adm_total
    if epcs_total and ci.get("epcs") is None:
        ci["epcs"] = epcs_total
    if equipamentos_total and ci.get("equipamentos_carga") is None:
        ci["equipamentos_carga"] = equipamentos_total
    if ensaios_total and ci.get("ensaios") is None:
        ci["ensaios"] = ensaios_total
    if canteiro_total and ci.get("canteiro") is None:
        ci["canteiro"] = canteiro_total


def extract_ci_from_apresentacao(rows, ci):
    """Extract CI from presentation/summary sheets."""
    for i, row in enumerate(rows):
        if not row:
            continue
        for j, cell in enumerate(row[:6]):
            desc = normalize(safe_str(cell))
            if not desc:
                continue

            # Find associated value
            val = None
            for k in range(max(0, j-1), min(len(row), j+4)):
                if k == j:
                    continue
                v = safe_num(row[k])
                if v is not None and v > 100:
                    val = v
                    break

            if val is None:
                continue

            if "gerenciamento tecnico" in desc or "ger. tecnico" in desc:
                if ci.get("projetos_consultorias") is None:
                    ci["projetos_consultorias"] = val


def extract_ci_from_ger_executivo_detail(rows, ci):
    """Extract CI totals from Ger_Executivo sheets with hierarchy data."""
    # These sheets have structural data but values may be in later columns
    # Look for CÉLULA CONSTRUTIVA level entries with totals
    for i, row in enumerate(rows):
        if not row or len(row) < 10:
            continue
        desc = normalize(safe_str(row[9]) if len(row) > 9 else "")
        nivel = safe_str(row[7]) if len(row) > 7 else ""

        # Find total values
        total = None
        for j in range(10, min(len(row), 30)):
            v = safe_num(row[j])
            if v is not None and v > 1000:
                if total is None or v > total:
                    total = v

        if not total or not desc:
            continue

        # Match CI N1 categories at ETAPA level
        if nivel in ("ETAPA", "SUBETAPA"):
            if any(k in desc for k in ["estudos", "projetos", "consultoria"]):
                if "ensaio" not in desc:
                    ci["projetos_consultorias"] = (ci.get("projetos_consultorias") or 0) + total
            elif any(k in desc for k in ["taxa", "documentos"]):
                ci["taxas_licencas"] = (ci.get("taxas_licencas") or 0) + total
            elif any(k in desc for k in ["ensaio", "controle tecnologico"]):
                ci["ensaios"] = (ci.get("ensaios") or 0) + total
            elif any(k in desc for k in ["administracao", "equipe"]):
                ci["equipe_adm"] = (ci.get("equipe_adm") or 0) + total
            elif "canteiro" in desc or "instalacoes provisorias" in desc:
                ci["canteiro"] = (ci.get("canteiro") or 0) + total
            elif "equipamento" in desc:
                if "protecao" in desc or "epc" in desc:
                    ci["epcs"] = (ci.get("epcs") or 0) + total
                else:
                    ci["equipamentos_carga"] = (ci.get("equipamentos_carga") or 0) + total


# ============================================================
# Sistemas Especiais Extraction
# ============================================================
def extract_se_from_dedicated_sheet(rows, se):
    """Extract from SISTEMAS ESPECIAIS or EQUIPAMENTOS ESPECIAIS sheet.

    Handles two patterns:
    1. Line items: Desc | Un | Qtd | PU | Total
    2. Section headers: SECTION_NAME | '' | '' | '' | SubTotal  (col 4)
    """
    # First pass: identify section headers and their subtotals
    section_totals = {}
    current_section = None
    for i, row in enumerate(rows):
        if not row:
            continue
        desc = normalize(safe_str(row[0]))
        if not desc:
            continue

        # Section header: has text in col 0, empty cols 1-3, value in col 4
        is_header = (len(row) > 4
                     and safe_str(row[1]).strip() == ""
                     and safe_str(row[2]).strip() == ""
                     and safe_num(row[4]) is not None
                     and safe_num(row[4]) > 100)

        if is_header:
            section_totals[desc] = safe_num(row[4])
            current_section = desc

    # Match section totals to SE categories
    for desc, total in section_totals.items():
        if "grupo de gerador" in desc:
            se["gerador"] = total
        elif "piscina" in desc:
            se["piscina"] = (se.get("piscina") or 0) + total
        elif "subesta" in desc:
            # Subestação - not a separate SE category, skip
            pass
        elif "sauna" in desc or "spa" in desc:
            # Not a separate SE category
            pass
        elif "ete" in desc or "tratamento de esgoto" in desc:
            se["ete"] = total

    # Second pass: extract line items
    for i, row in enumerate(rows):
        if not row:
            continue

        desc = normalize(safe_str(row[0]))
        if not desc:
            desc = normalize(safe_str(row[1]) if len(row) > 1 else "")
        if not desc:
            continue

        # Skip section headers (already processed) and column headers
        if desc.startswith("descric") or desc == "total":
            continue
        is_header = (len(row) > 4
                     and safe_str(row[1]).strip() == ""
                     and safe_str(row[2]).strip() == ""
                     and safe_num(row[4]) is not None)
        if is_header:
            continue

        # Line item: Desc | Un | Qtd | PU | Total (or Desc | Qtd | PU | Total)
        qtd = None
        pu = None
        total = None

        if len(row) > 4:
            total = safe_num(row[4])
            qtd = safe_int(row[2])
            pu = safe_num(row[3])
        if total is None and len(row) > 3:
            total = safe_num(row[3])
            qtd = safe_int(row[1])
            pu = safe_num(row[2])

        if total is None or total <= 0:
            continue

        # Match items - be specific to avoid false positives
        if "elevador" in desc and "cremalheira" not in desc and "manutencao" not in desc:
            if se.get("elevadores") is None or total > (se.get("elevadores") or 0):
                se["elevadores"] = total
                if qtd and 0 < qtd < 20:
                    se["elevadores_qtd"] = qtd
                    if total > 0 and qtd > 0:
                        se["elevadores_pu"] = round(total / qtd, 2)
        elif ("grupo de gerador" in desc or desc == "gerador" or
              ("gerador" in desc and "cloro" not in desc and "calor" not in desc)):
            if se.get("gerador") is None or total > (se.get("gerador") or 0):
                se["gerador"] = total
        elif "piscina" in desc and "total" not in desc and "revestimento" not in desc:
            se["piscina"] = (se.get("piscina") or 0) + total
        elif "ete" in desc or "tratamento de esgoto" in desc or "estacao de tratamento" in desc:
            se["ete"] = total
        elif "automacao" in desc or "telecomunicacao" in desc or "sistema logico" in desc:
            if se.get("automacao") is None or total > (se.get("automacao") or 0):
                se["automacao"] = total


def extract_se_from_adore_style(rows, se):
    """Extract from Adore-style SISTEMAS ESPECIAIS (sections: CLIMATIZAÇÃO, COMUNICAÇÃO, EQUIPAMENTOS)."""
    current_section = None
    for i, row in enumerate(rows):
        if not row:
            continue
        desc = normalize(safe_str(row[0]))

        # Detect sections
        if "climatizac" in desc or "exaust" in desc:
            current_section = "climatizacao"
            continue
        elif "comunicac" in desc:
            current_section = "comunicacao"
            continue
        elif "equipamento" in desc and "total" not in desc:
            current_section = "equipamentos"
            continue
        elif "outros sistema" in desc:
            current_section = "outros"
            continue

        if not desc or desc.startswith("descric") or desc == "total":
            # Skip headers; capture TOTAL values
            if desc == "total":
                val = safe_num(row[3]) if len(row) > 3 else None
                if val is None:
                    val = safe_num(row[2]) if len(row) > 2 else None
                # Don't overwrite with section totals
            continue

        # Parse line items in equipamentos section
        if current_section == "equipamentos":
            qtd = safe_int(row[1]) if len(row) > 1 else None
            pu = safe_num(row[2]) if len(row) > 2 else None
            total = safe_num(row[3]) if len(row) > 3 else None

            if total is None or total == 0:
                continue

            if "elevador" in desc:
                se["elevadores"] = total
                if qtd:
                    se["elevadores_qtd"] = qtd
                if pu:
                    se["elevadores_pu"] = pu
            elif "gerador" in desc and "cloro" not in desc and "calor" not in desc:
                se["gerador"] = total
            elif "piscina" in desc or "equipamento de piscina" in desc:
                se["piscina"] = (se.get("piscina") or 0) + total
            elif "bomba" in desc:
                # Bombs are often part of piscina or general
                pass


def extract_se_from_relatorio(rows, se):
    """Extract Sistemas Especiais from Sienge Relatório."""
    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue
        code = safe_str(row[0])
        desc = normalize(safe_str(row[1]))
        total = None
        if len(row) > 28:
            total = safe_num(row[28])
        if total is None:
            for j in range(20, min(len(row), 35)):
                v = safe_num(row[j])
                if v is not None and v > 50:
                    total = v
                    break

        if not code or not total:
            continue

        parts = code.strip().split(".")

        # Item level (N4) with specific equipment
        if "elevador" in desc and "cremalheira" not in desc and "manutencao" not in desc:
            if len(parts) >= 3:
                se["elevadores"] = (se.get("elevadores") or 0) + total
        elif ("gerador" in desc and "cloro" not in desc and "calor" not in desc) and len(parts) >= 3:
            se["gerador"] = (se.get("gerador") or 0) + total
        elif ("piscina" in desc or ("aquecimento" in desc and "piscina" in desc)) and len(parts) >= 3:
            se["piscina"] = (se.get("piscina") or 0) + total
        elif ("ete" in desc or "tratamento de esgoto" in desc) and len(parts) >= 3:
            se["ete"] = (se.get("ete") or 0) + total
        elif ("automacao" in desc or "sistema logico" in desc or "telecomunicacao" in desc) and len(parts) <= 3:
            if se.get("automacao") is None or total > (se.get("automacao") or 0):
                se["automacao"] = total

        # N2 level totals for EQUIPAMENTOS E SISTEMAS ESPECIAIS
        if len(parts) == 2:
            if "equipamento" in desc and "protecao" not in desc:
                if "elevador" in desc:
                    se["elevadores"] = total
                elif any(k in desc for k in ["automacao", "sistema logico", "telecomunicacao"]):
                    se["automacao"] = total
                elif "sistema especiai" in desc:
                    # Don't overwrite specifics
                    pass


def extract_se_from_piscina_sheet(rows, se):
    """Extract piscina costs from dedicated PISCINA sheet."""
    total = None
    for i, row in enumerate(rows):
        if not row:
            continue
        for j in range(len(row)):
            desc = normalize(safe_str(row[j]))
            if "total" in desc:
                for k in range(j+1, min(len(row), j+5)):
                    v = safe_num(row[k])
                    if v is not None and v > 100:
                        if total is None or v > total:
                            total = v
                        break
    if total:
        se["piscina"] = total


# ============================================================
# Louças e Metais Extraction
# ============================================================
def extract_loucas_from_dedicated_sheet(rows, lm):
    """Extract from LOUÇAS E METAIS sheet."""
    # Detect column layout from header row
    # Common patterns:
    #   [?, ?, Aplicação, Descrição, Un., Total, ...]  (offset=2)
    #   [?, Aplicação, Descrição, Un., Total, ...]     (offset=1)
    #   [Aplicação, Descrição, Un., Total, ...]        (offset=0)
    desc_col = None
    qtd_col = None

    for i, row in enumerate(rows[:10]):
        if not row:
            continue
        for j in range(min(len(row), 10)):
            h = normalize(safe_str(row[j]))
            if h == "total" or h == "total ":
                if qtd_col is None:
                    qtd_col = j
            if h in ("descricao do material",):
                desc_col = j
            elif h == "descricao" and desc_col is None:
                desc_col = j
            if h == "aplicacao" or h == "aplicac" or "aplicac" in h:
                # Aplicação column: description is next, total after un.
                if desc_col is None:
                    desc_col = j + 1

    # If we found desc but not qtd, look for "Total" again
    if desc_col is not None and qtd_col is None:
        for i, row in enumerate(rows[:10]):
            if not row:
                continue
            for j in range(desc_col + 1, min(len(row), 10)):
                h = normalize(safe_str(row[j]))
                if h.startswith("total") and "total geral" not in h:
                    qtd_col = j
                    break
            if qtd_col is not None:
                break

    # Fallback: try common layouts based on observed patterns
    if desc_col is None:
        # Pattern: [?, ?, Aplicação, Descrição, Un., Total, ...]
        for i, row in enumerate(rows[:10]):
            if not row:
                continue
            for j in range(min(len(row), 6)):
                h = normalize(safe_str(row[j]))
                if "louca" in h or "aparelho" in h or "metai" in h:
                    desc_col = j + 1 if j + 1 < len(row) else j
                    break
            if desc_col is not None:
                break

    if desc_col is None:
        desc_col = 3
    if qtd_col is None:
        qtd_col = 5

    for i, row in enumerate(rows):
        if not row:
            continue

        desc = normalize(safe_str(row[desc_col]) if desc_col < len(row) else "")
        qtd = safe_int(row[qtd_col]) if qtd_col < len(row) else None

        # Skip header rows and empty rows
        if not desc or desc.startswith("descric") or desc.startswith("un") or desc.startswith("ambiente"):
            continue
        # Skip rows that are clearly non-item rows
        if desc in ("total", "total geral", ""):
            continue
        # Skip rows where the "quantity" is unreasonably large for a fixture count
        # Single row should not exceed ~2000 units (large building might have 200 apts * 5 each)
        if qtd is not None and qtd > 5000:
            qtd = None

        if not desc or qtd is None:
            # Try finding totals in summary rows
            for j, cell in enumerate(row):
                txt = normalize(safe_str(cell))
                if "loucas" in txt and "aparelhos" in txt:
                    # Summary row with total
                    for k in range(j+1, min(len(row), j+5)):
                        v = safe_num(row[k])
                        if v is not None and v > 10:
                            lm["loucas_total"] = (lm.get("loucas_total") or 0) + v
                            break
                elif "metais" in txt and "acabamento" in txt:
                    for k in range(j+1, min(len(row), j+5)):
                        v = safe_num(row[k])
                        if v is not None and v > 10:
                            lm["metais_total"] = (lm.get("metais_total") or 0) + v
                            break
                elif "artefatos em pedra" in txt:
                    for k in range(j+1, min(len(row), j+5)):
                        v = safe_num(row[k])
                        if v is not None and v > 10:
                            lm["bancadas_granito"] = (lm.get("bancadas_granito") or 0) + v
                            break
            continue

        if qtd <= 0:
            continue

        # Match specific items
        if "bacia sanitaria" in desc or "bacia sanit" in desc:
            if "pne" not in desc:
                lm["bacias_sanitarias_qtd"] = (lm.get("bacias_sanitarias_qtd") or 0) + qtd
            else:
                lm["bacias_sanitarias_qtd"] = (lm.get("bacias_sanitarias_qtd") or 0) + qtd
        elif "cuba" in desc:
            lm["cubas_qtd"] = (lm.get("cubas_qtd") or 0) + qtd
        elif "lavatorio" in desc:
            lm["cubas_qtd"] = (lm.get("cubas_qtd") or 0) + qtd
        elif "torneira" in desc or "misturador" in desc:
            lm["torneiras_qtd"] = (lm.get("torneiras_qtd") or 0) + qtd
        elif "registro" in desc or "acabamento de registro" in desc:
            lm["registros_qtd"] = (lm.get("registros_qtd") or 0) + qtd
        elif "chuveiro" in desc:
            lm["chuveiros_qtd"] = (lm.get("chuveiros_qtd") or 0) + qtd
        elif "ducha" in desc:
            lm["chuveiros_qtd"] = (lm.get("chuveiros_qtd") or 0) + qtd
        elif "bancada" in desc and ("granito" in desc or "cozinha" in desc or "banheiro" in desc or "tanque" in desc):
            lm["bancadas_granito_m2"] = (lm.get("bancadas_granito_m2") or 0) + qtd
        elif "tanque" in desc:
            lm["cubas_qtd"] = (lm.get("cubas_qtd") or 0) + qtd


def extract_loucas_from_relatorio(rows, lm):
    """Extract louças totals from Sienge Relatório (summing aparelhos sanitários)."""
    total_loucas = 0
    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue
        code = safe_str(row[0])
        desc = normalize(safe_str(row[1]))
        total = None
        if len(row) > 28:
            total = safe_num(row[28])

        if not total:
            continue

        parts = code.strip().split(".")
        # Sum all "Aparelhos sanitários, louças, metais" N3 level entries
        if len(parts) == 3 and ("aparelhos sanitarios" in desc or "loucas" in desc and "metais" in desc):
            total_loucas += total

    if total_loucas > 0:
        lm["loucas_total"] = total_loucas


# ============================================================
# Main extraction per project
# ============================================================
def extract_project(slug, filepath):
    ci = {}
    lm = {}
    se = {}

    if not os.path.exists(filepath):
        return ci, lm, se, "FILE NOT FOUND"

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return ci, lm, se, f"OPEN ERROR: {str(e)[:60]}"

    try:
        sheetnames = wb.sheetnames
        sheet_classes = {}
        for name in sheetnames:
            sheet_classes[name] = classify_sheet(name)

        for name in sheetnames:
            stype = sheet_classes[name]
            if stype == "other":
                continue
            try:
                ws = wb[name]
                rows = read_rows(ws, max_rows=2500, max_cols=30)

                if stype == "relatorio":
                    extract_ci_from_relatorio(rows, ci)
                    extract_se_from_relatorio(rows, se)
                    extract_loucas_from_relatorio(rows, lm)

                elif stype == "ger_tec_adm":
                    extract_ci_from_ger_tec_adm(rows, ci)
                    extract_ci_from_ger_executivo_detail(rows, ci)

                elif stype == "ger_executivo":
                    extract_ci_from_ger_executivo_detail(rows, ci)

                elif stype == "sistemas_especiais":
                    # Check format
                    has_sections = False
                    for row in rows[:10]:
                        if row:
                            txt = normalize(safe_str(row[0]))
                            if "climatizac" in txt or "equipamento" in txt:
                                has_sections = True
                                break

                    if has_sections:
                        extract_se_from_adore_style(rows, se)
                    else:
                        extract_se_from_dedicated_sheet(rows, se)

                elif stype == "piscina":
                    extract_se_from_piscina_sheet(rows, se)

                elif stype == "loucas_metais":
                    extract_loucas_from_dedicated_sheet(rows, lm)

                elif stype == "apresentacao" or stype == "analise_custos":
                    extract_ci_from_apresentacao(rows, ci)

                elif stype == "custos_indiretos":
                    # Treat like ger_tec_adm
                    extract_ci_from_ger_tec_adm(rows, ci)

                elif stype == "epcs":
                    # EPCs sheet usually has quantitative data but not always totals
                    # Scan for total
                    for row in rows:
                        if not row:
                            continue
                        for j, cell in enumerate(row):
                            if normalize(safe_str(cell)) in ("total", "total epcs", "total epc"):
                                for k in range(j+1, min(len(row), j+5)):
                                    v = safe_num(row[k])
                                    if v is not None and v > 100:
                                        ci["epcs"] = v
                                        break

                elif stype == "canteiro":
                    # Scan for canteiro total
                    for row in rows:
                        if not row:
                            continue
                        for j, cell in enumerate(row):
                            if "total" in normalize(safe_str(cell)):
                                for k in range(j+1, min(len(row), j+5)):
                                    v = safe_num(row[k])
                                    if v is not None and v > 100:
                                        # Take the largest total found
                                        if ci.get("canteiro") is None or v > ci["canteiro"]:
                                            ci["canteiro"] = v
                                        break

                elif stype == "ensaios":
                    for row in rows:
                        if not row:
                            continue
                        for j, cell in enumerate(row):
                            if "total" in normalize(safe_str(cell)):
                                for k in range(j+1, min(len(row), j+5)):
                                    v = safe_num(row[k])
                                    if v is not None and v > 100:
                                        if ci.get("ensaios") is None or v > ci["ensaios"]:
                                            ci["ensaios"] = v
                                        break

                elif stype == "equipamentos":
                    extract_se_from_dedicated_sheet(rows, se)

            except Exception as e:
                pass

    except Exception as e:
        return ci, lm, se, f"PROCESSING ERROR: {str(e)[:60]}"
    finally:
        try:
            wb.close()
        except:
            pass

    # Count found fields
    ci_count = sum(1 for v in ci.values() if v is not None)
    lm_count = sum(1 for v in lm.values() if v is not None)
    se_count = sum(1 for v in se.values() if v is not None)
    total = ci_count + lm_count + se_count

    return ci, lm, se, f"OK ({ci_count}ci/{lm_count}lm/{se_count}se = {total} fields)"


def format_ci_output(ci):
    """Format CI data into standard output structure."""
    return {
        "projetos_consultorias": {"valor": ci.get("projetos_consultorias")},
        "taxas_licencas": {"valor": ci.get("taxas_licencas")},
        "equipe_adm": {"valor": ci.get("equipe_adm"), "prazo_meses": ci.get("prazo_meses")},
        "epcs": {"valor": ci.get("epcs")},
        "equipamentos_carga": {"valor": ci.get("equipamentos_carga")},
        "ensaios": {"valor": ci.get("ensaios")},
        "canteiro": {"valor": ci.get("canteiro")}
    }

def format_lm_output(lm):
    """Format louças/metais data into standard output structure."""
    return {
        "bacias_sanitarias": {"qtd": lm.get("bacias_sanitarias_qtd"), "valor": lm.get("bacias_sanitarias_valor"), "pu": lm.get("bacias_sanitarias_pu")},
        "cubas": {"qtd": lm.get("cubas_qtd"), "valor": lm.get("cubas_valor"), "pu": lm.get("cubas_pu")},
        "torneiras": {"qtd": lm.get("torneiras_qtd"), "valor": lm.get("torneiras_valor"), "pu": lm.get("torneiras_pu")},
        "registros": {"qtd": lm.get("registros_qtd"), "valor": lm.get("registros_valor")},
        "chuveiros": {"qtd": lm.get("chuveiros_qtd"), "valor": lm.get("chuveiros_valor")},
        "bancadas_granito": {"m2": lm.get("bancadas_granito_m2"), "valor": lm.get("bancadas_granito")},
        "metais_total": {"valor": lm.get("metais_total")},
        "loucas_total": {"valor": lm.get("loucas_total")}
    }

def format_se_output(se):
    """Format sistemas especiais data into standard output structure."""
    return {
        "elevadores": {"valor": se.get("elevadores"), "qtd": se.get("elevadores_qtd"), "pu_un": se.get("elevadores_pu")},
        "gerador": {"valor": se.get("gerador")},
        "piscina": {"valor": se.get("piscina")},
        "ete": {"valor": se.get("ete")},
        "automacao": {"valor": se.get("automacao")}
    }

def is_empty_detail(detail):
    """Check if a detail dict has all null values."""
    for key, sub in detail.items():
        if isinstance(sub, dict):
            for v in sub.values():
                if v is not None:
                    return False
        elif sub is not None:
            return False
    return True


# ============================================================
# Main
# ============================================================
def main():
    with open(MAPPING_FILE, "r", encoding="utf-8") as f:
        mapping = json.load(f)

    batch = mapping[BATCH_START:BATCH_END]
    print(f"Processing batch 1: {len(batch)} projects (indices {BATCH_START}-{BATCH_END-1})")
    print("=" * 100)

    results = []
    updated_files = 0
    total_fields = 0

    for i, entry in enumerate(batch):
        slug = entry["slug"]
        filepath = entry["path"]
        fname = os.path.basename(filepath)

        print(f"\n[{BATCH_START + i:3d}] {slug}")
        print(f"     File: {fname[:60]}")

        ci, lm, se, status = extract_project(slug, filepath)
        results.append((slug, ci, lm, se, status))

        print(f"     Status: {status}")
        if ci:
            print(f"     CI: {ci}")
        if se:
            print(f"     SE: {se}")
        if lm:
            lm_summary = {k: v for k, v in lm.items() if v is not None}
            print(f"     LM: {lm_summary}")

    # Update JSON files
    print("\n" + "=" * 100)
    print("Updating indices-executivo JSON files...")

    for slug, ci, lm, se, status in results:
        json_path = os.path.join(INDICES_DIR, f"{slug}.json")

        if not os.path.exists(json_path):
            print(f"  SKIP {slug}: JSON file not found")
            continue

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        changed = False

        ci_detail = format_ci_output(ci)
        lm_detail = format_lm_output(lm)
        se_detail = format_se_output(se)

        if not is_empty_detail(ci_detail):
            data["ci_detalhado"] = ci_detail
            changed = True

        if not is_empty_detail(lm_detail):
            data["loucas_metais_detail"] = lm_detail
            changed = True

        if not is_empty_detail(se_detail):
            data["sistemas_especiais_detail"] = se_detail
            changed = True

        if changed:
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            updated_files += 1
            n = sum(1 for v in ci.values() if v is not None) + \
                sum(1 for v in lm.values() if v is not None) + \
                sum(1 for v in se.values() if v is not None)
            total_fields += n
            print(f"  UPDATED {slug} ({n} fields)")
        else:
            print(f"  NO DATA {slug}")

    # Summary
    print("\n" + "=" * 100)
    print("SUMMARY")
    print("=" * 100)

    ci_projects = sum(1 for _, ci, _, _, _ in results if ci)
    lm_projects = sum(1 for _, _, lm, _, _ in results if lm)
    se_projects = sum(1 for _, _, _, se, _ in results if se)
    errors = sum(1 for _, _, _, _, s in results if "ERROR" in s or "NOT FOUND" in s)

    print(f"Total projects processed: {len(results)}")
    print(f"Projects with CI data:    {ci_projects}")
    print(f"Projects with LM data:    {lm_projects}")
    print(f"Projects with SE data:    {se_projects}")
    print(f"Files updated:            {updated_files}")
    print(f"Total fields extracted:   {total_fields}")
    print(f"Errors/Not found:         {errors}")

    # Detailed table
    print(f"\n{'Slug':<45} {'CI':>3} {'LM':>3} {'SE':>3} {'Status'}")
    print("-" * 80)
    for slug, ci, lm, se, status in results:
        ci_n = sum(1 for v in ci.values() if v is not None)
        lm_n = sum(1 for v in lm.values() if v is not None)
        se_n = sum(1 for v in se.values() if v is not None)
        print(f"{slug:<45} {ci_n:>3} {lm_n:>3} {se_n:>3} {status}")


if __name__ == "__main__":
    main()
