#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract project metadata from construction budget spreadsheets.
Batch 1: entries 0-41 (adore-cacupe to cn-brava-valley)
"""
import json
import os
import re
import sys
import traceback
from datetime import datetime, date

import openpyxl

# Force UTF-8 output on Windows
sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# Config
# ============================================================
BASE_DIR = r"C:\Users\leona\orcamentos-openclaw\base"
MAPPING_FILE = os.path.join(BASE_DIR, "_all_projects_mapping.json")
METADADOS_FILE = os.path.join(BASE_DIR, "projetos-metadados.json")
BATCH_START = 0
BATCH_END = 42  # exclusive

FIELDS = [
    "ur", "np", "npt", "npg", "elev", "vag", "at", "prazo_meses",
    "tipo_laje", "tipo_fundacao", "padrao_acabamento",
    "cub_valor", "cub_data_base", "cidade", "n_torres"
]

# ============================================================
# Helpers
# ============================================================

def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def safe_lower(val):
    return safe_str(val).lower()

def safe_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val == 0:
            return None
        return val
    s = str(val).strip().replace(",", ".")
    m = re.search(r'[\d]+[.]?[\d]*', s)
    if m:
        try:
            v = float(m.group())
            return int(v) if v == int(v) else v
        except:
            pass
    return None

def safe_int(val):
    n = safe_num(val)
    if n is not None:
        return int(n)
    return None

def is_date(val):
    return isinstance(val, (datetime, date))

def format_date(val):
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val)

def normalize(s):
    if not s:
        return ""
    s = s.lower().strip()
    for a, b in {'á':'a','à':'a','â':'a','ã':'a','é':'e','ê':'e','è':'e',
                  'í':'i','î':'i','ó':'o','ô':'o','õ':'o','ò':'o',
                  'ú':'u','û':'u','ù':'u','ç':'c','ñ':'n'}.items():
        s = s.replace(a, b)
    return s

def read_rows(ws, max_rows=60):
    """Safely read rows from a worksheet."""
    rows = []
    try:
        for row in ws.iter_rows(max_row=max_rows, values_only=True):
            rows.append(tuple(row))
    except:
        pass
    return rows

# ============================================================
# Known SC cities for fuzzy matching
# ============================================================
SC_CITIES = [
    "Balneário Camboriú", "Itapema", "Florianópolis", "Itajaí",
    "Navegantes", "Joinville", "Blumenau", "Brusque", "Camboriú",
    "Penha", "Piçarras", "São José", "Palhoça", "Tijucas",
    "Porto Belo", "Bombinhas", "Governador Celso Ramos", "Biguaçu",
    "Gaspar", "Jaraguá do Sul", "Chapecó", "Criciúma", "Tubarão",
    "Gravatal", "Lages", "São Bento do Sul", "Praia Brava",
    "Cacupé", "Ingleses", "Canasvieiras", "Jurerê",
]

# ============================================================
# Extraction strategies
# ============================================================

def extract_dados_iniciais(rows, meta):
    """
    Extract from DADOS_INICIAIS / OBRA style sheets.
    Pattern: col B = description, col C = identifier, col D = unit, col E = value
    OR: col B = description, col C = unit, col D = value
    """
    for i, row in enumerate(rows):
        if not row or len(row) < 3:
            continue

        # Build searchable text from first few columns
        texts = [normalize(safe_str(c)) for c in row[:6]]
        joined = " ".join(texts)

        # Find the value: scan columns for the numeric value
        def find_num(min_val=None, max_val=None):
            for j in range(1, min(len(row), 8)):
                n = safe_num(row[j])
                if n is not None:
                    if min_val is not None and n < min_val:
                        continue
                    if max_val is not None and n > max_val:
                        continue
                    return n
            return None

        def find_int(min_val=None, max_val=None):
            v = find_num(min_val, max_val)
            return int(v) if v is not None else None

        # AT - Área do terreno
        if meta.get("at") is None:
            if any(x in joined for x in ["area do terreno", "área do terreno"]):
                v = find_num(10, 100000)
                if v: meta["at"] = v
            elif "at" in texts and "m2" in joined:
                v = find_num(10, 100000)
                if v: meta["at"] = v

        # UR - Unidades residenciais
        if meta.get("ur") is None:
            if any(x in joined for x in ["unidades residenciais", "numero de aptos", "n de aptos", "unidades residencias"]):
                v = find_int(1, 2000)
                if v: meta["ur"] = v
            elif "ur" in texts and "un" in joined:
                v = find_int(1, 2000)
                if v: meta["ur"] = v

        # NP - Total pavimentos
        if meta.get("np") is None:
            if any(x in joined for x in ["numero total de pavimentos", "numero de pavimentos"]):
                if not any(x in joined for x in ["tipo", "garagem", "embasamento", "duplex", "subsolo", "diferenciado"]):
                    v = find_int(1, 100)
                    if v: meta["np"] = v
            elif "np" in texts and "un" in joined and "npt" not in texts and "npg" not in texts and "npd" not in texts:
                v = find_int(1, 100)
                if v: meta["np"] = v

        # NPT - Pavimentos tipo
        if meta.get("npt") is None:
            if "pavimentos tipo" in joined or "pavimento tipo" in joined:
                if "diferenciado" not in joined:
                    v = find_int(1, 80)
                    if v: meta["npt"] = v
            elif "npt" in texts:
                v = find_int(1, 80)
                if v: meta["npt"] = v

        # NPG - Pavimentos garagem / subsolos
        if meta.get("npg") is None:
            if any(x in joined for x in ["pavimentos garagem", "numero de subsolos", "numero de subsolo", "pavimento garagem"]):
                v = find_int(0, 20)
                if v: meta["npg"] = v
            elif "npg" in texts or "ns" in texts:
                v = find_int(0, 20)
                if v: meta["npg"] = v

        # ELEV - Elevadores
        if meta.get("elev") is None:
            if "elevador" in joined:
                v = find_int(1, 20)
                if v: meta["elev"] = v
            elif "elev" in texts:
                v = find_int(1, 20)
                if v: meta["elev"] = v

        # VAG - Vagas
        if meta.get("vag") is None:
            if "vaga" in joined:
                v = find_int(1, 5000)
                if v: meta["vag"] = v
            elif "vag" in texts:
                v = find_int(1, 5000)
                if v: meta["vag"] = v

        # Prazo
        if meta.get("prazo_meses") is None:
            if any(x in joined for x in ["meses de obra", "prazo"]):
                v = find_int(6, 120)
                if v: meta["prazo_meses"] = v

        # Torres
        if meta.get("n_torres") is None:
            if any(x in joined for x in ["numero de torres", "numero torres", "torre"]):
                if "tipo" not in joined and "projecao" not in joined and "area" not in joined:
                    v = find_int(1, 20)
                    if v: meta["n_torres"] = v


def extract_apresentacao(rows, meta):
    """
    Extract from Apresentação-style summary sheets.
    These have: ETAPA | VALOR | % | VALOR/m2 | ... | ÁREA CONSTRUÍDA | value | m2
    And CUB row at bottom.
    """
    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue

        texts = [safe_lower(c) for c in row[:15]]
        joined = " ".join(texts)

        # ÁREA CONSTRUÍDA in the header area (usually row 2, cols F-G)
        for j, cell in enumerate(row):
            s = normalize(safe_str(cell))
            if "area construida" in s or "área construída" in s.lower():
                # Value is in next column
                if j + 1 < len(row):
                    n = safe_num(row[j + 1])
                    if n and n > 100:
                        # Not directly a field we extract but useful context
                        pass

        # CUB row: typically "CUB/SC" or "CUB" | value | date
        if "cub" in joined:
            for j, cell in enumerate(row):
                s = normalize(safe_str(cell))
                if "cub" in s:
                    # Scan remaining cells for value and date
                    for k in range(j + 1, min(len(row), j + 6)):
                        if k < len(row):
                            n = safe_num(row[k])
                            if n and 1000 < n < 10000 and meta.get("cub_valor") is None:
                                meta["cub_valor"] = round(n, 2)
                            if is_date(row[k]) and meta.get("cub_data_base") is None:
                                meta["cub_data_base"] = format_date(row[k])
                    break

        # UNIDADES from "AREAS PRIVATIVAS" or "Quadro de áreas" style
        if meta.get("ur") is None:
            if "unidades" in joined and ("residencial" in joined or "total" in joined):
                for j, cell in enumerate(row):
                    s = normalize(safe_str(cell))
                    if "unidades" in s:
                        for k in range(j + 1, min(len(row), j + 5)):
                            if k < len(row):
                                n = safe_int(row[k])
                                if n and 1 <= n <= 2000:
                                    meta["ur"] = n
                                    break
                        break

        # Look for N° UNIDADES or similar
        if meta.get("ur") is None:
            for j, cell in enumerate(row):
                s = normalize(safe_str(cell))
                if "n unidades" in s or "total de unidades" in s or "unidades" == s.strip():
                    for k in range(j + 1, min(len(row), j + 5)):
                        if k < len(row):
                            n = safe_int(row[k])
                            if n and 1 <= n <= 2000:
                                meta["ur"] = n
                                break


def extract_calculo_medias(rows, meta):
    """
    Extract from CÁLCULO_MÉDIAS sheets.
    Pattern: Row with "CUB" label followed by CUB values,
             Row with "DATA BASE" followed by dates.
    """
    for i, row in enumerate(rows):
        if not row:
            continue

        texts = [safe_lower(c) for c in row[:15]]
        joined = " ".join(texts)

        # CUB row
        if "cub" in texts and meta.get("cub_valor") is None:
            # Find the rightmost CUB value (most recent)
            best_cub = None
            for j, cell in enumerate(row):
                if safe_lower(cell) == "cub":
                    continue
                n = safe_num(cell)
                if n and 1000 < n < 10000:
                    best_cub = round(n, 2)
            if best_cub:
                meta["cub_valor"] = best_cub

        # DATA BASE row
        if "data base" in joined and meta.get("cub_data_base") is None:
            best_date = None
            for j, cell in enumerate(row):
                if is_date(cell):
                    best_date = format_date(cell)
            if best_date:
                meta["cub_data_base"] = best_date

        # PAVIMENTOS row
        if "pavimentos" in texts and meta.get("np") is None:
            for j, cell in enumerate(row):
                n = safe_int(cell)
                if n and 1 <= n <= 100:
                    meta["np"] = n
                    break

        # M2 row (area construída for reference)
        if "m2" in texts:
            for j, cell in enumerate(row):
                n = safe_num(cell)
                if n and n > 500:
                    pass  # AC is not in our target fields


def extract_relatorio_sienge(rows, meta):
    """
    Extract from Sienge-style Relatório sheets.
    Has: Obra, Cliente, Endereço da obra, etc.
    """
    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue

        texts = [normalize(safe_str(c)) for c in row[:15]]
        joined = " ".join(texts)

        # Endereço da obra - extract city from address
        if "endereco da obra" in joined or ("endereco" in joined and meta.get("cidade") is None):
            for cell in row:
                s = safe_str(cell)
                if len(s) > 15:
                    # Try to extract "Cidade/UF" from address
                    # Pattern: "... - Bairro - Cidade/UF - CEP" or "... Cidade/UF"
                    city_match = re.search(r'[-–]\s*([\w\sáéíóúâêîôûãõç]+?)\s*/\s*(?:SC|PR|RS|RJ|SP|MG)', s, re.IGNORECASE)
                    if city_match and meta.get("cidade") is None:
                        cidade = city_match.group(1).strip()
                        # Take last segment after any remaining dashes
                        if " - " in cidade:
                            cidade = cidade.split(" - ")[-1].strip()
                        if len(cidade) > 2:
                            meta["cidade"] = cidade.title()


def scan_all_text(rows, meta):
    """
    Generic keyword scan across all cell text for tipo_laje, tipo_fundacao,
    padrao_acabamento, cidade, n_torres.
    """
    all_text = ""
    for row in rows:
        if not row:
            continue
        for cell in row:
            if cell is not None:
                all_text += " " + str(cell)

    all_norm = normalize(all_text)
    all_lower = all_text.lower()

    # tipo_laje
    if meta.get("tipo_laje") is None:
        if "protendida" in all_norm:
            meta["tipo_laje"] = "protendida"
        elif "nervurada" in all_norm:
            meta["tipo_laje"] = "nervurada"
        elif "cubeta" in all_norm:
            meta["tipo_laje"] = "cubetas"
        elif "macica" in all_norm and "laje" in all_norm:
            meta["tipo_laje"] = "maciça"
        elif "trelicada" in all_norm or "trelica" in all_norm:
            meta["tipo_laje"] = "treliçada"

    # tipo_fundacao
    if meta.get("tipo_fundacao") is None:
        if "helice" in all_norm:
            meta["tipo_fundacao"] = "hélice contínua"
        elif "estaca raiz" in all_norm:
            meta["tipo_fundacao"] = "estaca raiz"
        elif "franki" in all_norm:
            meta["tipo_fundacao"] = "estaca Franki"
        elif "estaca" in all_norm and "fundac" in all_norm:
            meta["tipo_fundacao"] = "estaca"
        elif "sapata" in all_norm:
            meta["tipo_fundacao"] = "sapata"
        elif "radier" in all_norm:
            meta["tipo_fundacao"] = "radier"
        elif "tubulao" in all_norm:
            meta["tipo_fundacao"] = "tubulão"

    # padrao_acabamento
    if meta.get("padrao_acabamento") is None:
        if "alto padrao" in all_norm or "alto padrão" in all_lower:
            meta["padrao_acabamento"] = "alto"
        elif "padrao medio" in all_norm:
            meta["padrao_acabamento"] = "médio"

    # n_torres
    if meta.get("n_torres") is None:
        torre_match = re.search(r'(\d+)\s*torre', all_norm)
        if torre_match:
            n = int(torre_match.group(1))
            if 1 <= n <= 20:
                meta["n_torres"] = n
        elif "torre unica" in all_norm:
            meta["n_torres"] = 1

    # cidade from known cities
    if meta.get("cidade") is None:
        for city in SC_CITIES:
            cn = normalize(city)
            # Ensure word boundary match (avoid matching "Itajaí" inside "ESQUADRIAS")
            # Use a simple check: the city must appear as a separate word/phrase
            idx = all_norm.find(cn)
            if idx >= 0:
                # Verify it's not in the middle of another word
                before = all_norm[idx-1] if idx > 0 else " "
                after = all_norm[idx+len(cn)] if idx+len(cn) < len(all_norm) else " "
                if not before.isalpha() and not after.isalpha():
                    meta["cidade"] = city
                    break

    # Also try address pattern extraction from "Cidade/SC" or "Cidade/PR" etc.
    if meta.get("cidade") is None:
        # Pattern: "- Cidade/UF" or "Cidade/UF" in addresses
        city_match = re.search(r'[-–]\s*([\w\sáéíóúâêîôûãõç]{3,30}?)\s*/\s*(?:SC|PR|RS|RJ|SP|MG)', all_text, re.IGNORECASE)
        if city_match:
            candidate = city_match.group(1).strip()
            # Filter out obvious non-cities (short names, numbers, common noise words)
            noise = {"vidros", "ferragens", "cub", "etapa", "obra", "valor", "total", "área"}
            if len(candidate) > 2 and not any(c.isdigit() for c in candidate) and normalize(candidate) not in noise:
                meta["cidade"] = candidate.title()
        # Fallback: "Cidade/UF" without dash
        if meta.get("cidade") is None:
            city_match2 = re.search(r'([\w\sáéíóúâêîôûãõç]{3,30}?)\s*/\s*(?:SC|PR|RS|RJ|SP|MG)', all_text, re.IGNORECASE)
            if city_match2:
                candidate = city_match2.group(1).strip()
                # Take last meaningful segment (after comma or dash)
                parts = re.split(r'[,\-–]', candidate)
                candidate = parts[-1].strip()
                if len(candidate) > 2 and not any(c.isdigit() for c in candidate) and normalize(candidate) not in noise:
                    meta["cidade"] = candidate.title()


def extract_from_executivo_sheet(rows, meta):
    """
    Extract from 'Executivo' or 'Gerenciamento_Exec' style sheets.
    These sometimes have UNIDADES count in a header/side area.
    """
    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue

        texts = [safe_lower(c) for c in row[:15]]
        joined = " ".join(texts)

        # Look for VALOR / UNIDADE header which implies unit count nearby
        # Or look for a number that could be unit count in specific cells

        # Some sheets have ÁREA CONSTRUÍDA | value | m² in the header rows
        for j, cell in enumerate(row):
            s = normalize(safe_str(cell))
            if "area construida" in s:
                if j + 1 < len(row):
                    n = safe_num(row[j + 1])
                    # Not a target field but could help context

        # Some sheets have number of units embedded
        if meta.get("ur") is None and "unidade" in joined:
            for j, cell in enumerate(row):
                s = normalize(safe_str(cell))
                if "unidade" in s and "valor" not in s:
                    for k in range(j + 1, min(len(row), j + 4)):
                        if k < len(row):
                            n = safe_int(row[k])
                            if n and 1 <= n <= 2000:
                                meta["ur"] = n
                                break


def extract_from_areas_privativas(rows, meta):
    """
    Extract from AREAS PRIVATIVAS sheets.
    Often has: ÁREA PRIVATIVA | value | m² and UNIDADES | value | UN
    """
    for i, row in enumerate(rows):
        if not row or len(row) < 2:
            continue

        for j, cell in enumerate(row):
            s = normalize(safe_str(cell))

            if meta.get("ur") is None:
                if "unidades" in s or "unidade" in s:
                    if "valor" not in s:
                        for k in range(j + 1, min(len(row), j + 4)):
                            if k < len(row):
                                n = safe_int(row[k])
                                if n and 1 <= n <= 2000:
                                    meta["ur"] = n
                                    break


# ============================================================
# Sheet classification
# ============================================================

def get_sheet_type(name):
    """Classify a sheet by its name."""
    n = normalize(name)
    if any(k in n for k in ["dados_iniciais", "dados iniciais"]):
        return "dados"
    if n.strip() == "obra" or n.strip() == "dados":
        return "dados"
    if n.strip() == "capa":
        return "capa"
    if any(k in n for k in ["calculo_medias", "calculo medias", "cálculo"]):
        return "medias"
    if "relatorio" in n:
        return "relatorio"
    if any(k in n for k in ["orcamento_executivo", "orcamento executivo", "orçamento executivo"]):
        return "apresentacao"
    if any(k in n for k in ["apresentacao", "apresentação"]):
        return "apresentacao"
    if any(k in n for k in ["areas privativas", "areas_privativas", "áreas privativas"]):
        return "areas"
    if "quadro de area" in n:
        return "areas"
    if any(k in n for k in ["resumo"]):
        return "resumo"
    if any(k in n for k in ["gerenciamento"]):
        return "gerenciamento"
    if "executivo" in n or "orcamento resumo" in n:
        return "executivo"
    if any(k in n for k in ["versoes", "versões"]):
        return "versoes"
    if "painel" in n:
        return "painel"
    return "other"


# ============================================================
# Main extraction per project
# ============================================================

def extract_project(slug, filepath):
    meta = {f: None for f in FIELDS}

    if not os.path.exists(filepath):
        return meta, "FILE NOT FOUND"

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return meta, f"OPEN ERROR: {e}"

    try:
        sheetnames = wb.sheetnames

        # Classify all sheets
        sheet_types = {}
        for name in sheetnames:
            sheet_types[name] = get_sheet_type(name)

        # Process sheets in priority order
        priority_order = ["dados", "capa", "areas", "apresentacao", "resumo",
                         "gerenciamento", "executivo", "medias", "relatorio",
                         "versoes", "painel", "other"]

        for stype in priority_order:
            for name in sheetnames:
                if sheet_types[name] != stype:
                    continue
                try:
                    ws = wb[name]
                    rows = read_rows(ws, max_rows=60)

                    if stype in ("dados", "capa"):
                        extract_dados_iniciais(rows, meta)
                        scan_all_text(rows, meta)

                    elif stype == "areas":
                        extract_from_areas_privativas(rows, meta)
                        extract_apresentacao(rows, meta)
                        scan_all_text(rows[:30], meta)

                    elif stype in ("apresentacao", "resumo", "versoes", "painel"):
                        extract_apresentacao(rows, meta)
                        scan_all_text(rows[:20], meta)

                    elif stype in ("gerenciamento", "executivo"):
                        extract_from_executivo_sheet(rows, meta)
                        extract_apresentacao(rows, meta)
                        # Also check deeper rows for CUB in these sheets
                        extract_apresentacao(rows[15:30], meta)
                        scan_all_text(rows[:20], meta)

                    elif stype == "medias":
                        extract_calculo_medias(rows, meta)

                    elif stype == "relatorio":
                        extract_relatorio_sienge(rows, meta)
                        scan_all_text(rows[:20], meta)

                    else:
                        # For unknown sheets, just scan first rows
                        if len(rows) < 50:
                            extract_dados_iniciais(rows[:30], meta)
                            scan_all_text(rows[:20], meta)

                except Exception as e:
                    pass

        # Second pass: scan infrastructure/supraestrutura sheets for fundacao/laje keywords
        if meta.get("tipo_fundacao") is None or meta.get("tipo_laje") is None:
            for name in sheetnames:
                n = normalize(name)
                if any(k in n for k in ["infraestrutura", "supraestrutura", "fundac", "suprae"]):
                    try:
                        ws = wb[name]
                        rows = read_rows(ws, max_rows=30)
                        scan_all_text(rows, meta)
                    except:
                        pass

    except Exception as e:
        return meta, f"PROCESSING ERROR: {e}"
    finally:
        try:
            wb.close()
        except:
            pass

    found = sum(1 for v in meta.values() if v is not None)
    return meta, f"OK ({found}/{len(FIELDS)} fields)"


def resolve_path(entry):
    return entry.get("path", "").replace("\\", "/")


# ============================================================
# Main
# ============================================================

def main():
    with open(MAPPING_FILE, "r", encoding="utf-8") as f:
        mapping = json.load(f)
    with open(METADADOS_FILE, "r", encoding="utf-8") as f:
        metadados = json.load(f)

    batch = mapping[BATCH_START:BATCH_END]
    print(f"Processing batch 1: {len(batch)} projects (indices {BATCH_START}-{BATCH_END-1})")
    print("=" * 100)

    results = []

    for i, entry in enumerate(batch):
        slug = entry["slug"]
        filepath = resolve_path(entry)
        fname = os.path.basename(filepath)

        print(f"\n[{BATCH_START + i:3d}] {slug}")
        print(f"     File: {fname}")

        meta, status = extract_project(slug, filepath)
        results.append((slug, meta, status))

        found_fields = {k: v for k, v in meta.items() if v is not None}
        print(f"     Status: {status}")
        if found_fields:
            for k, v in found_fields.items():
                print(f"     {k}: {v}")

    # Update metadados
    print("\n" + "=" * 100)
    print("Updating projetos-metadados.json...")

    updated_count = 0
    for slug, meta, status in results:
        if slug not in metadados:
            metadados[slug] = {}

        for field in FIELDS:
            new_val = meta[field]
            old_val = metadados[slug].get(field)
            # For our target fields: always write the new extraction result.
            # If we found a value, use it. If not, set null (cleans up garbage from prior runs).
            # This does NOT touch other existing fields like ac, total, rsm2, etc.
            if new_val != old_val:
                updated_count += 1
            metadados[slug][field] = new_val

    with open(METADADOS_FILE, "w", encoding="utf-8") as f:
        json.dump(metadados, f, indent=2, ensure_ascii=False)

    print(f"Updated {updated_count} field values across {len(results)} projects.")

    # Summary table
    print("\n" + "=" * 100)
    print("SUMMARY TABLE")
    print("=" * 100)

    hdr = f"{'Slug':<45} {'UR':>4} {'NP':>4} {'NPT':>4} {'NPG':>4} {'ELEV':>4} {'VAG':>5} {'AT':>10} {'Prazo':>5} {'CUB':>8} {'Torres':>6} {'Laje':<12} {'Fund.':<15} {'Cidade':<22}"
    print(hdr)
    print("-" * len(hdr))

    total_found = {f: 0 for f in FIELDS}

    for slug, meta, status in results:
        def fmt(v, w=4):
            if v is None:
                return "-".rjust(w)
            return str(v)[:w].rjust(w)

        for f in FIELDS:
            if meta[f] is not None:
                total_found[f] += 1

        line = (
            f"{slug:<45} "
            f"{fmt(meta['ur'], 4)} "
            f"{fmt(meta['np'], 4)} "
            f"{fmt(meta['npt'], 4)} "
            f"{fmt(meta['npg'], 4)} "
            f"{fmt(meta['elev'], 4)} "
            f"{fmt(meta['vag'], 5)} "
            f"{fmt(meta['at'], 10)} "
            f"{fmt(meta['prazo_meses'], 5)} "
            f"{fmt(meta['cub_valor'], 8)} "
            f"{fmt(meta['n_torres'], 6)} "
            f"{str(meta['tipo_laje'] or '-'):<12} "
            f"{str(meta['tipo_fundacao'] or '-'):<15} "
            f"{str(meta['cidade'] or '-'):<22}"
        )
        print(line)

    print("-" * len(hdr))
    print(f"\nField coverage ({len(results)} projects):")
    for f in FIELDS:
        pct = total_found[f] / len(results) * 100
        filled = int(pct / 5)
        bar = "#" * filled + "." * (20 - filled)
        print(f"  {f:<20} {total_found[f]:>3}/{len(results)} ({pct:5.1f}%) [{bar}]")


if __name__ == "__main__":
    main()
