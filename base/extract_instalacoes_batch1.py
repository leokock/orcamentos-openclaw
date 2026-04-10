#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract installation breakdown data from construction budget spreadsheets.
Batch 1: projects 0-41.

Splits "Instalações" macrogroup into 5 sub-disciplines:
  - hidrossanitarias
  - eletricas
  - preventivas
  - gas
  - telecom

Also extracts MO/Material splits where available.
"""
import json, sys, re, os
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

# Load project mapping
with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

BATCH = ALL_PROJECTS[0:42]  # projects 0-41

# ── Keyword matchers ──────────────────────────────────────────────────────
def classify_discipline(text):
    """Classify a description text into a discipline category."""
    t = text.lower().strip()

    # Telecom / communications (check first, more specific)
    if any(k in t for k in ['telecom', 'comunicaç', 'comunicac', 'interfon', 'cftv',
                            'telefon', 'lógic', 'logic', 'automação', 'automaç',
                            'rede estruturada', 'dados', 'tv vhf', 'tv/cabo',
                            'som e wifi', 'wifi']):
        return 'telecom'

    # Gas
    if any(k in t for k in ['gás', 'gas', 'glp', 'gn ']):
        # But not "gas + preventiva" combined
        if 'prevent' not in t and 'incêndio' not in t and 'incendio' not in t:
            return 'gas'

    # Preventivas (fire prevention, PPCI, SPDA)
    if any(k in t for k in ['prevent', 'ppci', 'pci', 'incêndio', 'incendio',
                            'hidrante', 'sprinkler', 'alarme', 'spda',
                            'proteção e aterramento', 'protecao e aterramento']):
        return 'preventivas'

    # Hidrossanitárias
    if any(k in t for k in ['hidro', 'sanitár', 'sanitar', 'água fria', 'agua fria',
                            'água quente', 'agua quente', 'esgoto', 'pluvial',
                            'drenagem', 'bombas', 'bomba', 'ligações prediais',
                            'ligacoes prediais', 'reuso', 'tratamento de esgoto',
                            'hidrometro', 'hidrômetro']):
        return 'hidrossanitarias'

    # Elétricas
    if any(k in t for k in ['elétr', 'eletr', 'eletrodut', 'eletrocalh',
                            'fios e cabos', 'cabos e fiaç', 'cabos e fiac',
                            'quadros elétric', 'quadros eletric', 'disjuntor',
                            'tomada', 'interruptor', 'luminár', 'luminar',
                            'lâmpada', 'lampada', 'gerador', 'energia']):
        return 'eletricas'

    return None

def is_mo_line(text):
    """Check if this is a labor (MO) line."""
    t = text.lower().strip()
    return any(k in t for k in ['mão de obra', 'mao de obra', 'mo ', 'mão-de-obra',
                                 'm.o.', 'm.o ', '- mo', '(mo)'])

def is_material_line(text):
    """Check if this is a material line."""
    t = text.lower().strip()
    return any(k in t for k in ['material', 'mat.', 'mat '])


def safe_float(v):
    """Safely convert a value to float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


# ── Format-specific extractors ───────────────────────────────────────────

def extract_from_ger_executivo(wb, sheet_name='Ger_Executivo'):
    """
    Extract from Ger_Executivo sheet format (CTN standard).
    Looks for installation discipline headers and their totals.
    """
    ws = wb[sheet_name]
    result = {}
    mo_material = {}

    # Collect all rows
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=30, values_only=True):
        rows.append(row)

    in_instalacoes = False
    current_discipline = None
    discipline_total = 0
    discipline_mo = 0

    for i, row in enumerate(rows):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        row_text = ' '.join(str(v) for v in vals)
        row_lower = row_text.lower()

        # Detect Instalações section
        if 'instalações' in row_lower or 'instalacoes' in row_lower:
            in_instalacoes = True

        # Detect end of Instalações (next macrogroup)
        if in_instalacoes and not current_discipline:
            if any(k in row_lower for k in ['impermeabiliza', 'louças', 'loucas',
                                              'esquadria', 'revestiment', 'forro',
                                              'pintura', 'fachada', 'complementar',
                                              'paisagismo', 'cobertura', 'sistemas especiais']):
                in_instalacoes = False
                continue

        # Look for discipline headers
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                disc = classify_discipline(v)
                if disc and (in_instalacoes or disc in ['hidrossanitarias', 'eletricas', 'preventivas', 'gas', 'telecom']):
                    # Save previous discipline
                    if current_discipline and discipline_total > 0:
                        if current_discipline not in result or discipline_total > result[current_discipline]:
                            result[current_discipline] = discipline_total
                        if discipline_mo > 0:
                            mo_material[current_discipline] = {
                                'mo_valor': round(discipline_mo, 2),
                                'material_valor': round(discipline_total - discipline_mo, 2),
                                'mo_pct': round(discipline_mo / discipline_total * 100, 1) if discipline_total > 0 else 0
                            }

                    current_discipline = disc
                    discipline_total = 0
                    discipline_mo = 0

                    # Check if total is on the same row
                    for v2 in vals:
                        if isinstance(v2, (int, float)) and v2 > 1000:
                            discipline_total = float(v2)
                    break

        # Track totals and MO within current discipline
        if current_discipline:
            if any(isinstance(v, str) and 'total' in str(v).lower() for v in vals):
                for v in vals:
                    if isinstance(v, (int, float)) and v > 1000:
                        if v > discipline_total:
                            discipline_total = float(v)
                        break

            if any(isinstance(v, str) and is_mo_line(str(v)) for v in vals):
                for v in vals:
                    if isinstance(v, (int, float)) and v > 100:
                        discipline_mo += float(v)
                        break

    # Save last discipline
    if current_discipline and discipline_total > 0:
        if current_discipline not in result or discipline_total > result[current_discipline]:
            result[current_discipline] = discipline_total
        if discipline_mo > 0:
            mo_material[current_discipline] = {
                'mo_valor': round(discipline_mo, 2),
                'material_valor': round(discipline_total - discipline_mo, 2),
                'mo_pct': round(discipline_mo / discipline_total * 100, 1) if discipline_total > 0 else 0
            }

    return result, mo_material


def extract_from_relatorio_sienge(wb):
    """
    Extract from Sienge 'Relatório' format.
    Uses hierarchical codes like 05.001, 06.001, 07.001 etc.
    """
    ws = wb['Relatório']
    result = {}
    mo_material = {}

    rows_data = []
    for row in ws.iter_rows(min_row=1, max_row=3000, max_col=30, values_only=True):
        rows_data.append(row)

    # First pass: identify N1-level codes and their disciplines
    n1_codes = {}

    for row in rows_data:
        if row[0] is None or not isinstance(row[0], str):
            continue
        code = row[0].strip()
        desc = ''
        for v in row[1:]:
            if isinstance(v, str) and len(v) > 2:
                desc = v
                break

        # N1 level: XX or XX.XXX
        if re.match(r'^\d{2}\s*$', code) or re.match(r'^\d{2}\.\d{3}\s*$', code):
            disc = classify_discipline(desc)
            if disc:
                prefix = code.strip()[:2]
                n1_codes[prefix] = disc

    # Second pass: collect totals
    disc_totals = {}
    disc_mo = {}

    for row in rows_data:
        if row[0] is None or not isinstance(row[0], str):
            continue
        code = row[0].strip()
        desc = ''
        for v in row[1:]:
            if isinstance(v, str) and len(v) > 2:
                desc = v
                break

        # Get total value - try multiple columns (different formats use different cols)
        total_val = 0
        for col_idx in [28, 27, 26, 25, 20, 15, 10, 8, 6, 5, 4, 3]:
            if len(row) > col_idx and isinstance(row[col_idx], (int, float)) and row[col_idx] > 0:
                total_val = float(row[col_idx])
                break

        prefix = code[:2] if len(code) >= 2 else ''
        if prefix in n1_codes:
            disc = n1_codes[prefix]

            if re.match(r'^\d{2}\s*$', code):
                if total_val > 0:
                    disc_totals[disc] = total_val
            elif re.match(r'^\d{2}\.\d{3}\s*$', code):
                if total_val > 0 and disc not in disc_totals:
                    disc_totals[disc] = disc_totals.get(disc, 0) + total_val

            # Check for MO
            if re.match(r'^\d{2}\.\d{3}\.\d{3}', code):
                if is_mo_line(desc) and total_val > 0:
                    disc_mo[disc] = disc_mo.get(disc, 0) + total_val

    result = {disc: round(val, 2) for disc, val in disc_totals.items()}

    for disc, mo_val in disc_mo.items():
        if disc in result and result[disc] > 0:
            mo_material[disc] = {
                'mo_valor': round(mo_val, 2),
                'material_valor': round(result[disc] - mo_val, 2),
                'mo_pct': round(mo_val / result[disc] * 100, 1) if result[disc] > 0 else 0
            }

    return result, mo_material


def extract_from_obra_sheet(wb, sheet_name='Obra'):
    """
    Extract from 'Obra' summary sheet.
    Look for subsection rows under "Instalações".
    """
    ws = wb[sheet_name]
    result = {}

    in_instalacoes = False
    for row in ws.iter_rows(min_row=1, max_row=500, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        if isinstance(vals[0], str):
            t = vals[0].lower().strip()
            if 'instalações' in t or 'instalacoes' in t:
                in_instalacoes = True
                disc = classify_discipline(vals[0])
                if disc:
                    for v in vals[1:]:
                        if isinstance(v, (int, float)) and v > 1000:
                            result[disc] = round(float(v), 2)
                            break
                continue

            # If we hit next macrogroup, stop
            if in_instalacoes and not any(k in t for k in ['hidro', 'eletr', 'prevent', 'gas', 'gás', 'telecom', 'ppci', 'pci', 'spda', 'comunic', 'instal']):
                if any(isinstance(v, (int, float)) and v > 1000 for v in vals):
                    break

        if in_instalacoes and isinstance(vals[0], str):
            disc = classify_discipline(vals[0])
            if disc:
                for v in vals[1:]:
                    if isinstance(v, (int, float)) and v > 1000:
                        result[disc] = result.get(disc, 0) + round(float(v), 2)
                        break

    return result, {}


def extract_from_instalacoes_sheet(wb, sheet_name):
    """
    Extract from a dedicated INSTALAÇÕES summary sheet.
    """
    ws = wb[sheet_name]
    result = {}
    mo_material = {}

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        if isinstance(vals[0], str):
            desc = vals[0]
            disc = classify_discipline(desc)
            if disc:
                total = 0
                for v in vals[1:]:
                    if isinstance(v, (int, float)) and v > 1000:
                        total = float(v)
                        break

                if total > 0:
                    is_mo = is_mo_line(desc)
                    if is_mo:
                        if disc not in mo_material:
                            mo_material[disc] = {'mo_valor': 0, 'material_valor': 0, 'mo_pct': 0}
                        mo_material[disc]['mo_valor'] += round(total, 2)
                    else:
                        result[disc] = result.get(disc, 0) + round(total, 2)

    for disc in mo_material:
        if disc in result:
            total = result[disc] + mo_material[disc]['mo_valor']
            result[disc] = round(total, 2)
            mo_material[disc]['material_valor'] = round(total - mo_material[disc]['mo_valor'], 2)
            mo_material[disc]['mo_pct'] = round(mo_material[disc]['mo_valor'] / total * 100, 1) if total > 0 else 0

    return result, mo_material


def extract_from_separate_sheets(wb, sheets):
    """
    Extract from separate discipline sheets (HIDRO, ELÉTRICA, PPCI, etc.)
    """
    result = {}
    mo_material = {}

    for sname in sheets:
        sname_upper = sname.strip().upper()
        disc = None

        if 'HIDRO' in sname_upper or 'SANITAR' in sname_upper:
            disc = 'hidrossanitarias'
        elif 'ELETR' in sname_upper or 'ELETRIC' in sname_upper:
            disc = 'eletricas'
        elif 'PCI' in sname_upper or 'PPCI' in sname_upper or 'PREV' in sname_upper or 'SPDA' in sname_upper:
            disc = 'preventivas'
        elif 'GAS' in sname_upper or 'GÁS' in sname_upper or 'GLP' in sname_upper:
            if 'PCI' not in sname_upper and 'PREV' not in sname_upper:
                disc = 'gas'
            else:
                disc = 'preventivas'
        elif 'INTERF' in sname_upper or 'TELEC' in sname_upper or 'TELEF' in sname_upper or 'COMUNIC' in sname_upper:
            disc = 'telecom'

        if not disc:
            continue

        ws = wb[sname]
        sheet_total = 0
        sheet_mo = 0
        last_big_val = 0

        for row in ws.iter_rows(min_row=1, max_row=500, max_col=20, values_only=True):
            vals = [v for v in row if v is not None]
            if not vals:
                continue

            row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()

            # Track the biggest value in "total" rows
            if 'total' in row_text and 'subtotal' not in row_text:
                for v in vals:
                    if isinstance(v, (int, float)) and v > 1000:
                        if v > sheet_total:
                            sheet_total = float(v)
                        break

            # Track any large value (fallback if no "total" row)
            for v in vals:
                if isinstance(v, (int, float)) and v > last_big_val:
                    last_big_val = float(v)

            if is_mo_line(row_text):
                for v in vals:
                    if isinstance(v, (int, float)) and v > 100:
                        sheet_mo += float(v)
                        break

        # If no explicit total row found, use the largest value seen
        if sheet_total == 0 and last_big_val > 10000:
            sheet_total = last_big_val

        if sheet_total > 0:
            result[disc] = result.get(disc, 0) + round(sheet_total, 2)
            if sheet_mo > 0:
                if disc not in mo_material:
                    mo_material[disc] = {'mo_valor': 0, 'material_valor': 0, 'mo_pct': 0}
                mo_material[disc]['mo_valor'] += round(sheet_mo, 2)

    for disc in mo_material:
        if disc in result and result[disc] > 0:
            mo_material[disc]['material_valor'] = round(result[disc] - mo_material[disc]['mo_valor'], 2)
            mo_material[disc]['mo_pct'] = round(mo_material[disc]['mo_valor'] / result[disc] * 100, 1)

    return result, mo_material


def extract_from_apresentacao_sheet(wb, sheet_name):
    """
    Extract from 'Apresentação' or 'APRESENTAÇÃO' sheets which have
    a summary table with all macrogroups and sub-disciplines.
    """
    ws = wb[sheet_name]
    result = {}
    mo_material = {}

    in_instalacoes = False

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=20, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        desc = ''
        total_val = 0
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                if not desc:
                    desc = v
            if isinstance(v, (int, float)) and abs(v) > 100:
                if total_val == 0:
                    total_val = float(v)

        if not desc:
            continue

        dl = desc.lower()

        if 'instalações' in dl or 'instalacoes' in dl:
            in_instalacoes = True
            continue

        # Exit section
        if in_instalacoes and any(k in dl for k in ['impermeabiliza', 'louças', 'loucas',
                                                      'esquadria', 'revestiment', 'forro',
                                                      'pintura', 'fachada', 'complementar',
                                                      'sistemas especiais', 'cobertura', 'paisagismo']):
            in_instalacoes = False

        if in_instalacoes or disc_is_installation_keyword(dl):
            disc = classify_discipline(desc)
            if disc and total_val > 0:
                if is_mo_line(desc):
                    if disc not in mo_material:
                        mo_material[disc] = {'mo_valor': 0, 'material_valor': 0, 'mo_pct': 0}
                    mo_material[disc]['mo_valor'] += round(abs(total_val), 2)
                else:
                    result[disc] = result.get(disc, 0) + round(abs(total_val), 2)

    for disc in mo_material:
        if disc in result:
            total = result[disc] + mo_material[disc]['mo_valor']
            result[disc] = round(total, 2)
            mo_material[disc]['material_valor'] = round(total - mo_material[disc]['mo_valor'], 2)
            mo_material[disc]['mo_pct'] = round(mo_material[disc]['mo_valor'] / total * 100, 1) if total > 0 else 0

    return result, mo_material


def disc_is_installation_keyword(text):
    """Check if text contains installation-related keywords."""
    return any(k in text for k in ['hidrossanit', 'elétric', 'eletric', 'preventiv',
                                    'telecom', 'incêndio', 'incendio', 'ppci', 'spda'])


def extract_from_resumo_sheet(wb, sheet_name):
    """
    Extract from summary/resumo sheets.
    """
    ws = wb[sheet_name]
    result = {}

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        for v in vals:
            if isinstance(v, str):
                disc = classify_discipline(v)
                if disc:
                    for v2 in vals:
                        if isinstance(v2, (int, float)) and v2 > 1000:
                            result[disc] = result.get(disc, 0) + round(float(v2), 2)
                            break
                    break

    return result, {}


def extract_from_eap_sheet(wb, sheet_name='EAP'):
    """
    Extract from EAP (WBS) sheets.
    """
    ws = wb[sheet_name]
    result = {}
    in_instalacoes = False

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=20, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        desc = ''
        total_val = 0
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                if not desc:
                    desc = v
            if isinstance(v, (int, float)) and v > 1000:
                total_val = float(v)

        if not desc:
            continue

        dl = desc.lower()
        if 'instalações' in dl or 'instalacoes' in dl:
            in_instalacoes = True

        if in_instalacoes and any(k in dl for k in ['impermeabiliza', 'louças', 'loucas', 'esquadria',
                                                      'revestiment', 'teto', 'forro', 'piso', 'pintura',
                                                      'fachada', 'complementar']):
            in_instalacoes = False

        if in_instalacoes:
            disc = classify_discipline(desc)
            if disc and total_val > 0:
                result[disc] = result.get(disc, 0) + round(total_val, 2)

    return result, {}


def extract_generic_scan(wb):
    """
    Generic scan: go through ALL sheets looking for installation-related rows.
    Last resort strategy.
    """
    result = {}
    mo_material = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        in_instalacoes = False

        try:
            for row in ws.iter_rows(min_row=1, max_row=1000, max_col=20, values_only=True):
                vals = [v for v in row if v is not None]
                if not vals:
                    continue

                desc = ''
                total_val = 0
                for v in vals:
                    if isinstance(v, str) and len(v) > 3:
                        if not desc:
                            desc = v
                    if isinstance(v, (int, float)) and abs(v) > 1000:
                        if total_val == 0:
                            total_val = abs(float(v))

                if not desc:
                    continue

                dl = desc.lower()

                if 'instalações' in dl or 'instalacoes' in dl:
                    in_instalacoes = True
                    continue

                if in_instalacoes and any(k in dl for k in ['impermeabiliza', 'louças', 'esquadria',
                                                              'revestiment', 'forro', 'pintura', 'fachada',
                                                              'complementar', 'sistemas especiais', 'cobertura']):
                    in_instalacoes = False

                if in_instalacoes:
                    disc = classify_discipline(desc)
                    if disc and total_val > 0:
                        # Only keep the largest value per discipline (likely the total)
                        if disc not in result or total_val > result[disc]:
                            result[disc] = round(total_val, 2)
        except Exception:
            continue

    return result, mo_material


# ── Main extraction logic per project ────────────────────────────────────

def extract_project(slug, path):
    """Extract installation breakdown from a project's xlsx."""
    path = path.replace('\\', '/')

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    sheets_upper = [s.strip().upper() for s in sheets]

    result = {}
    mo_material = {}
    method = ''

    try:
        # Strategy 1: Dedicated discipline sheets (most reliable)
        disc_sheets = [s for s in sheets if any(k in s.upper() for k in
                      ['HIDRO', 'ELETR', 'ELETRIC', 'PCI', 'PPCI', 'PREV', 'GÁS', 'GAS', 'GLP',
                       'INTERF', 'TELEC', 'TELEF', 'COMUNIC', 'SPDA'])]
        disc_sheets = [s for s in disc_sheets if 'PROVISÓR' not in s.upper() and
                      'INSUMO' not in s.upper() and 'PREÇO' not in s.upper() and
                      'PRECO' not in s.upper()]

        if disc_sheets:
            r, m = extract_from_separate_sheets(wb, disc_sheets)
            if r:
                result = r
                mo_material = m
                method = f'Separate sheets ({", ".join(disc_sheets[:3])})'

        # Strategy 2: Ger_Executivo
        ger_exec_sheets = [s for s in sheets if 'ger_executivo' in s.lower() or 'gerexecutivo' in s.lower()]
        if not result and ger_exec_sheets:
            r, m = extract_from_ger_executivo(wb, ger_exec_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Ger_Executivo ({ger_exec_sheets[0]})'

        # Strategy 3: Sienge Relatório
        if not result and 'Relatório' in sheets:
            r, m = extract_from_relatorio_sienge(wb)
            if r:
                result = r
                mo_material = m
                method = 'Relatório (Sienge)'

        # Strategy 4: Dedicated INSTALAÇÕES sheet
        inst_sheets = [s for s in sheets if 'INSTALAÇÕES' in s.upper() or 'INSTALACOES' in s.upper()]
        inst_sheets = [s for s in inst_sheets if 'PROVISÓRI' not in s.upper() and 'COMPLEMENT' not in s.upper()]
        if not result and inst_sheets:
            r, m = extract_from_instalacoes_sheet(wb, inst_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'INSTALAÇÕES sheet ({inst_sheets[0]})'

        # Strategy 5: Apresentação sheet
        apres_sheets = [s for s in sheets if 'apresenta' in s.lower() or 'APRESENTA' in s.upper()]
        if not result and apres_sheets:
            r, m = extract_from_apresentacao_sheet(wb, apres_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Apresentação ({apres_sheets[0]})'

        # Strategy 6: Obra sheet
        obra_sheets = [s for s in sheets if s.strip().lower() in ['obra', 'obra ']]
        if not result and obra_sheets:
            r, m = extract_from_obra_sheet(wb, obra_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Obra sheet ({obra_sheets[0]})'

        # Strategy 7: Resumo / Orçamento Resumo
        resumo_sheets = [s for s in sheets if 'resumo' in s.lower()]
        if not result and resumo_sheets:
            r, m = extract_from_resumo_sheet(wb, resumo_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Resumo ({resumo_sheets[0]})'

        # Strategy 8: EAP sheet
        eap_sheets = [s for s in sheets if s.strip().upper() == 'EAP']
        if not result and eap_sheets:
            r, m = extract_from_eap_sheet(wb, eap_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'EAP ({eap_sheets[0]})'

        # Strategy 9: Generic scan (last resort)
        if not result:
            r, m = extract_generic_scan(wb)
            if r:
                result = r
                mo_material = m
                method = 'Generic scan'

        # HYBRID: If primary strategy found some, try separate sheets to supplement
        if result and disc_sheets and 'Separate sheets' not in method:
            r2, m2 = extract_from_separate_sheets(wb, disc_sheets)
            for disc in r2:
                if disc not in result:
                    result[disc] = r2[disc]
                    if disc in m2:
                        mo_material[disc] = m2[disc]
                    method += f' + {disc} from sheets'

    except Exception as e:
        wb.close()
        return None, None, f"Error extracting: {e}"

    wb.close()

    if not result:
        return None, None, f"No breakdown found. Sheets: {sheets}"

    return result, mo_material, method


# ── Process all projects ─────────────────────────────────────────────────

def main():
    stats = {
        'total': len(BATCH),
        'extracted': 0,
        'already_had': 0,
        'failed': 0,
        'no_breakdown': 0,
        'disciplines_found': {
            'hidrossanitarias': 0,
            'eletricas': 0,
            'preventivas': 0,
            'gas': 0,
            'telecom': 0
        },
        'mo_splits_found': 0
    }

    details = []

    for p in BATCH:
        slug = p['slug']
        path = p['path'].replace('\\', '/')

        print(f"\n{'='*60}")
        print(f"Processing: {slug}")

        # Check if already has data
        json_path = INDICES_DIR / f"{slug}.json"
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
            existing_ib = existing.get('instalacoes_breakdown', {})
            if existing_ib and len(existing_ib) >= 2:
                print(f"  SKIP: already has {list(existing_ib.keys())}")
                stats['already_had'] += 1
                for disc in existing_ib:
                    if disc in stats['disciplines_found']:
                        stats['disciplines_found'][disc] += 1
                details.append({'slug': slug, 'status': 'already_had', 'disciplines': list(existing_ib.keys())})
                continue

        result, mo_material, msg = extract_project(slug, path)

        if result is None:
            print(f"  FAILED: {msg}")
            stats['failed'] += 1
            details.append({'slug': slug, 'status': 'failed', 'reason': msg})
            continue

        if not result:
            print(f"  NO BREAKDOWN: {msg}")
            stats['no_breakdown'] += 1
            details.append({'slug': slug, 'status': 'no_breakdown', 'reason': msg})
            continue

        print(f"  Method: {msg}")
        print(f"  Disciplines found: {list(result.keys())}")
        for disc, val in sorted(result.items()):
            mo_info = ''
            if mo_material and disc in mo_material:
                mo_info = f"  (MO: R${mo_material[disc]['mo_valor']:,.2f} = {mo_material[disc]['mo_pct']}%)"
                stats['mo_splits_found'] += 1
            print(f"    {disc}: R${val:,.2f}{mo_info}")
            if disc in stats['disciplines_found']:
                stats['disciplines_found'][disc] += 1

        stats['extracted'] += 1

        # Update JSON file
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            print(f"  WARNING: {json_path} does not exist, creating")
            data = {'slug': slug}

        # Build instalacoes_breakdown
        breakdown = {}
        for disc in ['hidrossanitarias', 'eletricas', 'preventivas', 'gas', 'telecom']:
            if disc in result:
                breakdown[disc] = round(result[disc], 2)

        data['instalacoes_breakdown'] = breakdown

        # Build split_mo_material for installation disciplines
        if mo_material:
            # Preserve existing split_mo_material for non-installation disciplines
            existing_splits = data.get('split_mo_material', {})
            inst_splits = {}
            for disc, info in mo_material.items():
                if info.get('mo_valor', 0) > 0:
                    inst_splits[disc] = {
                        'mo_valor': round(info['mo_valor'], 2),
                        'material_valor': round(info['material_valor'], 2),
                        'mo_pct': info['mo_pct']
                    }
            if inst_splits:
                # Merge: keep existing non-installation splits, add new installation splits
                if isinstance(existing_splits, dict):
                    for k, v in existing_splits.items():
                        if k not in ['hidrossanitarias', 'eletricas', 'preventivas', 'gas', 'telecom']:
                            inst_splits[k] = v
                data['split_mo_material'] = inst_splits

        data['instalacoes_extraction_method'] = msg

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        details.append({'slug': slug, 'status': 'extracted', 'disciplines': list(result.keys()), 'method': msg})

    # Print summary
    print(f"\n\n{'='*60}")
    print(f"SUMMARY — Batch 1 (projects 0-41)")
    print(f"{'='*60}")
    print(f"Total projects in batch: {stats['total']}")
    print(f"Already had data (skipped): {stats['already_had']}")
    print(f"Newly extracted: {stats['extracted']}")
    print(f"Failed to open: {stats['failed']}")
    print(f"No breakdown found: {stats['no_breakdown']}")
    print(f"\nDisciplines found across ALL projects (including pre-existing):")
    for disc, count in sorted(stats['disciplines_found'].items()):
        total_with = stats['extracted'] + stats['already_had']
        print(f"  {disc}: {count}/{total_with} projects")
    print(f"\nMO/Material splits found: {stats['mo_splits_found']} discipline-project combinations")

    # List failures
    failures = [d for d in details if d['status'] not in ('extracted', 'already_had')]
    if failures:
        print(f"\nProjects without breakdown ({len(failures)}):")
        for d in failures:
            print(f"  {d['slug']}: {d.get('reason', 'unknown')}")

    # List what was extracted
    extracted = [d for d in details if d['status'] == 'extracted']
    if extracted:
        print(f"\nNewly extracted ({len(extracted)}):")
        for d in extracted:
            print(f"  {d['slug']}: {d['disciplines']} via {d.get('method', '?')}")


if __name__ == '__main__':
    main()
