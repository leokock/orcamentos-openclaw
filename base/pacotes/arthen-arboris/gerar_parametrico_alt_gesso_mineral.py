#!/usr/bin/env python
"""
Arthen Arboris — gera versão alternativa do paramétrico com GESSO MINERAL
no forro das áreas secas (substituindo gesso acartonado).

Estratégia:
- Áreas secas (70% do forro): gesso mineral R$ 53/m² (material + instalação)
  → Fonte: Supabase indices-cartesian, cluster_id=3781, mediana n_obs=16
- Áreas molhadas BWC (15%): mantém gesso acartonado RU + perfil + MO
- Perfil estrutural: só nas BWC (gesso mineral é aplicado direto na laje)
- MO empreitada: só nas BWC (mineral tem instalação inclusa no PU)

Output: arthen-arboris-parametrico-alt-gesso-mineral-v01.xlsx
"""
from __future__ import annotations

import shutil
from pathlib import Path
from openpyxl import load_workbook

BASE = Path(r"C:\Users\leona\orcamentos\parametricos\arthen-arboris")
SRC = BASE / "arthen-arboris-parametrico-v00-final.xlsx"
DST = BASE / "arthen-arboris-parametrico-alt-gesso-mineral-v01.xlsx"

# PUs do Supabase indices-cartesian (cluster_id)
PU_GESSO_MINERAL = 53   # 3781: Forro mineral (material + instalação), mediana
PU_GESSO_RU = 35        # paramétrico original (material RU)
PU_PERFIL = 15          # paramétrico original
PU_MO = 25              # paramétrico original


def main():
    print(f"> Copiando {SRC.name} → {DST.name}")
    shutil.copy(SRC, DST)

    wb = load_workbook(DST, data_only=False)

    # ------------------------------------------------------------------
    # Aba Teto — substitui 4 linhas pela versão alt com gesso mineral
    # ------------------------------------------------------------------
    ws = wb["Teto"]
    # Qtd referencia INDICES!D16 (índice forro) × DADOS_PROJETO!B5 (AC)
    # L4: Forro áreas secas → MINERAL (PU 53, sem perfil, MO inclusa)
    ws["A4"] = "Forro Mineral"
    ws["B4"] = "Secas (mineral in natura)"
    ws["C4"] = "=ROUND(INDICES!D16*DADOS_PROJETO!B5*0.7,0)"
    ws["D4"] = "m²"
    ws["E4"] = PU_GESSO_MINERAL
    ws["F4"] = "=C4*E4"
    ws["G4"] = "Supabase 3781 (n=16) | 70% áreas secas"

    # L5: Forro RU BWC → mantém acartonado (mineral não serve em wet area)
    ws["A5"] = "Forro RU"
    ws["B5"] = "BWC (acartonado mantido)"
    ws["C5"] = "=ROUND(INDICES!D16*DADOS_PROJETO!B5*0.15,0)"
    ws["D5"] = "m²"
    ws["E5"] = PU_GESSO_RU
    ws["F5"] = "=C5*E5"
    ws["G5"] = "PU base | 15% BWC"

    # L6: Perfis → agora SÓ nas BWC (mineral aplica direto, não precisa estrutura)
    ws["A6"] = "Perfis"
    ws["B6"] = "Estrutura BWC"
    ws["C6"] = "=ROUND(INDICES!D16*DADOS_PROJETO!B5*0.15,0)"  # era 0.85, agora 0.15
    ws["D6"] = "m²"
    ws["E6"] = PU_PERFIL
    ws["F6"] = "=C6*E6"
    ws["G6"] = "PU base | só BWC (mineral não usa perfil)"

    # L7: MO → agora SÓ nas BWC (mineral tem MO inclusa no PU 53)
    ws["A7"] = "MO forro"
    ws["B7"] = "Empreitada BWC"
    ws["C7"] = "=ROUND(INDICES!D16*DADOS_PROJETO!B5*0.15,0)"  # era 1.00, agora 0.15
    ws["D7"] = "m²"
    ws["E7"] = PU_MO
    ws["F7"] = "=C7*E7"
    ws["G7"] = "MO acartonado BWC | mineral já tem MO inclusa"

    # ------------------------------------------------------------------
    # PREMISSAS — adiciona nota sobre a alternativa
    # ------------------------------------------------------------------
    ws_prem = wb["PREMISSAS"]
    linha_livre = ws_prem.max_row + 2
    ws_prem.cell(row=linha_livre, column=1, value="")
    ws_prem.cell(row=linha_livre + 1, column=1, value="═══ ALTERNATIVA: GESSO MINERAL (v01) ═══")
    ws_prem.cell(row=linha_livre + 2, column=1, value="Versão que substitui forro de gesso acartonado por gesso mineral nas áreas secas.")
    ws_prem.cell(row=linha_livre + 3, column=1, value=f"Gesso mineral: R$ {PU_GESSO_MINERAL}/m² (Supabase cluster_id=3781, mediana n_obs=16, n_proj=2, cv=0.049).")
    ws_prem.cell(row=linha_livre + 4, column=1, value="Benefício: menor custo (forro secas); instalação mais rápida; melhor acústica; monolítico.")
    ws_prem.cell(row=linha_livre + 5, column=1, value="Limitação: não aplicável em áreas molhadas (BWC mantém acartonado RU); não permite passagem de instalações/sprinklers embutidos.")
    ws_prem.cell(row=linha_livre + 6, column=1, value="Confiança: MÉDIA — base Supabase tem só 2 projetos. Validar com fornecedor local (Itapema/SC) antes de fechar.")

    # ------------------------------------------------------------------
    # DADOS_PROJETO — muda o título
    # ------------------------------------------------------------------
    ws_dp = wb["DADOS_PROJETO"]
    ws_dp["A1"] = "ARTHEN ARBORIS - PARAMETRICO ALT-GESSO-MINERAL v01"

    ws_painel = wb["PAINEL"]
    ws_painel["A1"] = "ARTHEN ARBORIS - PARAMETRICO ALT-GESSO-MINERAL v01"
    ws_painel["A2"] = "Versao alternativa: forro gesso MINERAL (secas) vs acartonado (v00-final)"
    # PAINEL B4-B7 tem totais hardcoded — vamos deixar comentário pro Leo recalcular ao abrir
    # (não mexer — o Excel vai recalcular as fórmulas ligadas a CUSTOS_MACROGRUPO.Teto)

    wb.save(DST)
    print(f"> Salvo em {DST}")
    print()

    # ------------------------------------------------------------------
    # Simulação em Python pra preview do delta (sem precisar abrir Excel)
    # ------------------------------------------------------------------
    ac = 12472.98         # DADOS_PROJETO!B5
    idx_forro = 1.16      # INDICES!D16 (entrega Completa, shell=100%)
    qtd_total = round(idx_forro * ac, 0)
    qtd_secas_70 = round(idx_forro * ac * 0.7, 0)
    qtd_bwc_15 = round(idx_forro * ac * 0.15, 0)

    # v00-final (atual)
    v00_ST = qtd_secas_70 * 28
    v00_RU = qtd_bwc_15 * 35
    v00_perfil = round(idx_forro * ac * 0.85, 0) * 15
    v00_mo = qtd_total * 25
    v00_total = v00_ST + v00_RU + v00_perfil + v00_mo

    # v01-alt (mineral)
    alt_mineral = qtd_secas_70 * PU_GESSO_MINERAL
    alt_RU = qtd_bwc_15 * PU_GESSO_RU
    alt_perfil = qtd_bwc_15 * PU_PERFIL
    alt_mo = qtd_bwc_15 * PU_MO
    alt_total = alt_mineral + alt_RU + alt_perfil + alt_mo

    custo_total_v00 = 39728474.50
    delta = alt_total - v00_total
    custo_total_alt = custo_total_v00 + delta

    print("=" * 70)
    print("SIMULACAO - MACROGRUPO TETO")
    print("=" * 70)
    print(f"AC: {ac:,.2f} m2 | Indice forro: {idx_forro} | Qtd total: {qtd_total:,.0f} m2")
    print()
    print(f"  {'ITEM':<45} {'V00-FINAL':>12} {'ALT-MINERAL':>12}")
    print(f"  {'-'*45} {'-'*12} {'-'*12}")
    print(f"  {'Forro secas (70%)':<45} R$ {v00_ST:>8,.0f} R$ {alt_mineral:>8,.0f}")
    print(f"    v00: acartonado ST PU R$ 28  |  alt: mineral PU R$ {PU_GESSO_MINERAL}")
    print(f"  {'Forro BWC (15%)':<45} R$ {v00_RU:>8,.0f} R$ {alt_RU:>8,.0f}")
    print(f"  {'Perfis (v00: 85%  alt: 15% BWC)':<45} R$ {v00_perfil:>8,.0f} R$ {alt_perfil:>8,.0f}")
    print(f"  {'MO (v00: 100%  alt: 15% BWC)':<45} R$ {v00_mo:>8,.0f} R$ {alt_mo:>8,.0f}")
    print(f"  {'-'*45} {'-'*12} {'-'*12}")
    print(f"  {'TOTAL TETO':<45} R$ {v00_total:>8,.0f} R$ {alt_total:>8,.0f}")
    print()
    print(f"  Delta TETO:     R$ {delta:>+11,.0f} ({(delta/v00_total)*100:+.1f}%)")
    print()
    print("=" * 70)
    print("IMPACTO NO CUSTO TOTAL DO PARAMETRICO")
    print("=" * 70)
    print(f"  Custo Total v00-final:      R$ {custo_total_v00:>14,.2f}")
    print(f"  Custo Total alt-mineral:    R$ {custo_total_alt:>14,.2f}")
    print(f"  Delta:                      R$ {delta:>+14,.2f} ({(delta/custo_total_v00)*100:+.2f}% do total)")
    print()
    print(f"  R$/m2 AC v00:   R$ {custo_total_v00/ac:,.2f}")
    print(f"  R$/m2 AC alt:   R$ {custo_total_alt/ac:,.2f}")
    print(f"  Economia por UR (98 UR):  R$ {-delta/98:,.2f}")


if __name__ == "__main__":
    main()
