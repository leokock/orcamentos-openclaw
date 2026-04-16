"""
Rateio de hidrossanitário por pavimento do Thozen-Electra.

Deriva o total de hidro do preliminar paramétrico (R$ 2.583.129,65 em PREMISSAS L38)
distribuindo pelos pavimentos físicos via m² ponderado por fator de intensidade hidro.

NÃO é levantamento BIM — é sanity check / ordem de grandeza.
Fatores, split material/MO e distribuição de AC ficam editáveis na aba Premissas.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PKG = Path(__file__).resolve().parent.parent
STATE = PKG / "state.json"
CONFIG = PKG / "parametrico-v2-config.json"
PRELIMINAR = PKG / "preliminar-thozen-electra.xlsx"
OUT = Path(__file__).resolve().parent / "hidro-por-pavimento.xlsx"

FMT_BRL = '"R$" #,##0.00'
FMT_PCT = "0.00%"
FMT_NUM = "#,##0.00"
FMT_INT = "#,##0"

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
HEADER_FONT = Font(bold=True, size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FFD7D7")
WARN_FONT = Font(bold=True, color="9C0006", size=11)


def read_hidro_total(xlsx: Path) -> float:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["PREMISSAS"]
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and "HIDROSSANITÁRIAS" in str(row[0]).upper():
            raw = str(row[1]).replace(".", "").replace(",", ".")
            return float(raw)
    raise RuntimeError("Linha hidrossanitário não encontrada em PREMISSAS")


def load_sources():
    state = json.loads(STATE.read_text(encoding="utf-8"))
    config = json.loads(CONFIG.read_text(encoding="utf-8"))
    hidro_total = read_hidro_total(PRELIMINAR)
    return state, config, hidro_total


def style_title(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = TITLE_FILL
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def style_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_premissas(ws, state, config, hidro_total):
    style_title(ws, 1, "PREMISSAS DO RATEIO HIDROSSANITÁRIO — THOZEN ELECTRA", 5)

    ws.merge_cells("A3:E5")
    c = ws.cell(
        row=3, column=1,
        value=(
            "⚠ ESTE É UM RATEIO DERIVADO DO PRELIMINAR PARAMÉTRICO — NÃO É LEVANTAMENTO BIM.\n"
            "Use para sanity check e ordem de grandeza. Quantidades reais por pavimento virão "
            "do Visus/IFC quando o levantamento BIM estiver pronto.\n"
            "Células amarelas são editáveis — altere e o rateio recalcula."
        ),
    )
    c.fill = WARN_FILL
    c.font = WARN_FONT
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    row = 7
    ws.cell(row=row, column=1, value="TOTAL HIDROSSANITÁRIO DO PRELIMINAR").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value="Total hidro (R$)")
    b = ws.cell(row=row, column=2, value=hidro_total)
    b.number_format = FMT_BRL
    b.fill = INPUT_FILL
    ref_total = f"Premissas!$B${row}"
    row += 1
    ws.cell(row=row, column=1, value="% Material")
    b = ws.cell(row=row, column=2, value=0.35)
    b.number_format = FMT_PCT
    b.fill = INPUT_FILL
    ref_pct_mat = f"Premissas!$B${row}"
    row += 1
    ws.cell(row=row, column=1, value="% Mão de obra")
    b = ws.cell(row=row, column=2, value=0.65)
    b.number_format = FMT_PCT
    b.fill = INPUT_FILL
    ref_pct_mo = f"Premissas!$B${row}"
    row += 1
    ws.cell(row=row, column=1, value="Material calculado (R$)")
    b = ws.cell(row=row, column=2, value=f"={ref_total}*{ref_pct_mat}")
    b.number_format = FMT_BRL
    ref_mat = f"Premissas!$B${row}"
    row += 1
    ws.cell(row=row, column=1, value="MO calculada (R$)")
    b = ws.cell(row=row, column=2, value=f"={ref_total}*{ref_pct_mo}")
    b.number_format = FMT_BRL
    ref_mo = f"Premissas!$B${row}"

    row += 2
    ws.cell(row=row, column=1, value="ESTRUTURA DO EMPREENDIMENTO").font = Font(bold=True)
    row += 1
    estrutura = [
        ("Número de torres", config.get("briefing", {}).get("n_torres", "2")),
        ("Pavimentos totais por torre (np)", config.get("np", 30)),
        ("Pavimentos tipo por torre (npt)", config.get("npt", 24)),
        ("Apartamentos totais (UR)", config.get("ur", state.get("ur", 348))),
        ("Banheiros por apartamento", config.get("briefing", {}).get("n_banheiros", "3")),
        ("AC total (m²)", config.get("ac", state.get("ac", 37893.89))),
        ("Índice hidro (m tubulação / m² AC)", 1.08),
    ]
    refs = {}
    for label, value in estrutura:
        ws.cell(row=row, column=1, value=label)
        b = ws.cell(row=row, column=2, value=float(value) if isinstance(value, (int, float)) else float(value))
        b.fill = INPUT_FILL
        if "m²" in label or "Índice" in label:
            b.number_format = FMT_NUM
        else:
            b.number_format = FMT_INT
        refs[label] = f"Premissas!$B${row}"
        row += 1

    ref_torres = refs["Número de torres"]
    ref_np = refs["Pavimentos totais por torre (np)"]
    ref_npt = refs["Pavimentos tipo por torre (npt)"]
    ref_ur = refs["Apartamentos totais (UR)"]
    ref_ac = refs["AC total (m²)"]
    ref_idx = refs["Índice hidro (m tubulação / m² AC)"]
    ref_banh = refs["Banheiros por apartamento"]

    row += 1
    ws.cell(row=row, column=1, value="DISTRIBUIÇÃO DE AC POR TIPO DE PAVIMENTO").font = Font(bold=True)
    row += 1
    style_header(ws, row, ["Tipo", "Qtd/torre", "% AC", "Fator hidro", "AC unit (m²)"])
    row += 1

    tipos = [
        ("Subsolo",  2, 0.08, 0.15),
        ("Térreo",   1, 0.02, 0.50),
        ("Lazer",    1, 0.03, 0.75),
        ("Tipo",    24, 0.85, 1.00),
        ("Técnico",  2, 0.02, 0.25),
    ]
    tipo_refs = {}
    tipo_start = row
    for nome, qtd, pct, fator in tipos:
        ws.cell(row=row, column=1, value=nome).font = Font(bold=True)
        c_qtd = ws.cell(row=row, column=2, value=qtd)
        c_qtd.fill = INPUT_FILL
        c_qtd.number_format = FMT_INT
        c_pct = ws.cell(row=row, column=3, value=pct)
        c_pct.fill = INPUT_FILL
        c_pct.number_format = FMT_PCT
        c_fat = ws.cell(row=row, column=4, value=fator)
        c_fat.fill = INPUT_FILL
        c_fat.number_format = "0.00"
        c_ac = ws.cell(
            row=row, column=5,
            value=f"=IFERROR(({ref_ac}*C{row})/(B{row}*{ref_torres}),0)",
        )
        c_ac.number_format = FMT_NUM
        tipo_refs[nome] = {
            "row": row,
            "qtd": f"Premissas!$B${row}",
            "pct": f"Premissas!$C${row}",
            "fator": f"Premissas!$D${row}",
            "ac_unit": f"Premissas!$E${row}",
        }
        row += 1
    tipo_end = row - 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B{tipo_start}:B{tipo_end})").number_format = FMT_INT
    ws.cell(row=row, column=3, value=f"=SUM(C{tipo_start}:C{tipo_end})").number_format = FMT_PCT
    ws.cell(row=row, column=5, value=f"=SUMPRODUCT(B{tipo_start}:B{tipo_end},E{tipo_start}:E{tipo_end})*{ref_torres}").number_format = FMT_NUM
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = Font(bold=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    return {
        "total": ref_total,
        "mat": ref_mat,
        "mo": ref_mo,
        "torres": ref_torres,
        "np": ref_np,
        "npt": ref_npt,
        "ur": ref_ur,
        "ac": ref_ac,
        "idx": ref_idx,
        "banh": ref_banh,
        "tipos": tipo_refs,
    }


def build_pavimentos_list(config):
    torres = int(float(config.get("briefing", {}).get("n_torres", "2")))
    np_ = int(config.get("np", 30))
    npt = int(config.get("npt", 24))
    n_subsolo = 2
    n_terreo = 1
    n_lazer = 1
    n_tecnico = np_ - (n_subsolo + n_terreo + n_lazer + npt)
    if n_tecnico < 0:
        n_tecnico = 0

    pavs = []
    for torre in range(1, torres + 1):
        for i in range(1, n_subsolo + 1):
            pavs.append((torre, f"Subsolo {i}", "Subsolo"))
        pavs.append((torre, "Térreo", "Térreo"))
        pavs.append((torre, "Lazer", "Lazer"))
        for i in range(1, npt + 1):
            pavs.append((torre, f"Tipo {i:02d}", "Tipo"))
        for i in range(1, n_tecnico + 1):
            pavs.append((torre, f"Técnico {i}", "Técnico"))
    return pavs


def write_pavimentos(ws, pavs, refs):
    style_title(ws, 1, "PAVIMENTOS — AC, FATOR, PESO, % PESO", 7)
    style_header(ws, 3, ["Torre", "Nº seq", "Pavimento", "Tipo", "AC (m²)", "Fator", "Peso"])

    row_start = 4
    for i, (torre, nome, tipo) in enumerate(pavs):
        r = row_start + i
        ws.cell(row=r, column=1, value=torre).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=nome)
        ws.cell(row=r, column=4, value=tipo).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5, value=f"={refs['tipos'][tipo]['ac_unit']}").number_format = FMT_NUM
        ws.cell(row=r, column=6, value=f"={refs['tipos'][tipo]['fator']}").number_format = "0.00"
        ws.cell(row=r, column=7, value=f"=E{r}*F{r}").number_format = FMT_NUM
    row_end = row_start + len(pavs) - 1

    total_row = row_end + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=5, value=f"=SUM(E{row_start}:E{row_end})").number_format = FMT_NUM
    ws.cell(row=total_row, column=7, value=f"=SUM(G{row_start}:G{row_end})").number_format = FMT_NUM
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = HEADER_FILL
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    return row_start, row_end, total_row


def write_rateio_valores(ws, pavs, refs, pav_start, pav_end, pav_total_row):
    style_title(ws, 1, "RATEIO DE VALORES (R$) POR PAVIMENTO", 8)
    style_header(ws, 3, ["Torre", "Nº seq", "Pavimento", "Tipo", "% Peso", "R$ Material", "R$ MO", "R$ Total"])

    soma_pesos = f"Pavimentos!$G${pav_total_row}"
    for i, (torre, nome, tipo) in enumerate(pavs):
        r = 4 + i
        pav_row = pav_start + i
        ws.cell(row=r, column=1, value=torre).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=nome)
        ws.cell(row=r, column=4, value=tipo).alignment = Alignment(horizontal="center")
        peso_ref = f"Pavimentos!$G${pav_row}"
        ws.cell(row=r, column=5, value=f"={peso_ref}/{soma_pesos}").number_format = FMT_PCT
        ws.cell(row=r, column=6, value=f"=E{r}*{refs['mat']}").number_format = FMT_BRL
        ws.cell(row=r, column=7, value=f"=E{r}*{refs['mo']}").number_format = FMT_BRL
        ws.cell(row=r, column=8, value=f"=F{r}+G{r}").number_format = FMT_BRL

    row_end = 4 + len(pavs) - 1
    total_row = row_end + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    ws.cell(row=total_row, column=5, value=f"=SUM(E4:E{row_end})").number_format = FMT_PCT
    ws.cell(row=total_row, column=6, value=f"=SUM(F4:F{row_end})").number_format = FMT_BRL
    ws.cell(row=total_row, column=7, value=f"=SUM(G4:G{row_end})").number_format = FMT_BRL
    ws.cell(row=total_row, column=8, value=f"=SUM(H4:H{row_end})").number_format = FMT_BRL
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = HEADER_FILL
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 18


def write_rateio_quantidades(ws, pavs, refs, pav_start, pav_end):
    style_title(ws, 1, "RATEIO DE QUANTIDADES FÍSICAS POR PAVIMENTO", 10)
    style_header(
        ws, 3,
        ["Torre", "Pavimento", "Tipo", "AC (m²)", "Tubulação\n(m)", "Vasos", "Lavatórios", "Chuveiros", "Pias\ncozinha", "Tanques\nAS"],
    )

    apts_por_tipo = f"IFERROR({refs['ur']}/({refs['tipos']['Tipo']['qtd']}*{refs['torres']}),0)"

    for i, (torre, nome, tipo) in enumerate(pavs):
        r = 4 + i
        pav_row = pav_start + i
        ws.cell(row=r, column=1, value=torre).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=nome)
        ws.cell(row=r, column=3, value=tipo).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=f"=Pavimentos!$E${pav_row}").number_format = FMT_NUM
        ws.cell(row=r, column=5, value=f"=D{r}*{refs['idx']}*Pavimentos!$F${pav_row}").number_format = FMT_NUM

        if tipo == "Tipo":
            ws.cell(row=r, column=6, value=f"=ROUND({apts_por_tipo}*{refs['banh']},0)").number_format = FMT_INT
            ws.cell(row=r, column=7, value=f"=ROUND({apts_por_tipo}*{refs['banh']},0)").number_format = FMT_INT
            ws.cell(row=r, column=8, value=f"=ROUND({apts_por_tipo}*{refs['banh']},0)").number_format = FMT_INT
            ws.cell(row=r, column=9, value=f"=ROUND({apts_por_tipo},0)").number_format = FMT_INT
            ws.cell(row=r, column=10, value=f"=ROUND({apts_por_tipo},0)").number_format = FMT_INT
        elif tipo == "Lazer":
            ws.cell(row=r, column=6, value=6).number_format = FMT_INT
            ws.cell(row=r, column=7, value=6).number_format = FMT_INT
            ws.cell(row=r, column=8, value=4).number_format = FMT_INT
            ws.cell(row=r, column=9, value=2).number_format = FMT_INT
            ws.cell(row=r, column=10, value=0).number_format = FMT_INT
        elif tipo == "Térreo":
            ws.cell(row=r, column=6, value=3).number_format = FMT_INT
            ws.cell(row=r, column=7, value=3).number_format = FMT_INT
            ws.cell(row=r, column=8, value=0).number_format = FMT_INT
            ws.cell(row=r, column=9, value=1).number_format = FMT_INT
            ws.cell(row=r, column=10, value=0).number_format = FMT_INT
        else:
            for c in range(6, 11):
                ws.cell(row=r, column=c, value=0).number_format = FMT_INT

    row_end = 4 + len(pavs) - 1
    total_row = row_end + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    for col, letter in [(4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J")]:
        fmt = FMT_NUM if letter in ("D", "E") else FMT_INT
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}4:{letter}{row_end})").number_format = fmt
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).fill = HEADER_FILL
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 12
    ws.row_dimensions[3].height = 30


def write_resumo_por_tipo(ws, refs):
    style_title(ws, 1, "RESUMO POR TIPO DE PAVIMENTO", 6)
    style_header(ws, 3, ["Tipo", "Qtd total", "AC total (m²)", "R$ total", "% R$", "Tubulação (m)"])

    tipos_ordem = ["Subsolo", "Térreo", "Lazer", "Tipo", "Técnico"]
    row = 4
    for nome in tipos_ordem:
        tref = refs["tipos"][nome]
        ws.cell(row=row, column=1, value=nome).font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"={tref['qtd']}*{refs['torres']}").number_format = FMT_INT
        ws.cell(row=row, column=3, value=f"={refs['ac']}*{tref['pct']}").number_format = FMT_NUM
        ws.cell(
            row=row, column=4,
            value=f'=SUMIF(\'Rateio-valores\'!D:D,"{nome}",\'Rateio-valores\'!H:H)',
        ).number_format = FMT_BRL
        ws.cell(row=row, column=5, value=f"=D{row}/{refs['total']}").number_format = FMT_PCT
        ws.cell(row=row, column=6, value=f"=C{row}*{refs['idx']}*{tref['fator']}").number_format = FMT_NUM
        row += 1

    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=SUM(B4:B{row-1})").number_format = FMT_INT
    ws.cell(row=row, column=3, value=f"=SUM(C4:C{row-1})").number_format = FMT_NUM
    ws.cell(row=row, column=4, value=f"=SUM(D4:D{row-1})").number_format = FMT_BRL
    ws.cell(row=row, column=5, value=f"=SUM(E4:E{row-1})").number_format = FMT_PCT
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = Font(bold=True)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 16


def main():
    state, config, hidro_total = load_sources()
    print(f"AC total (state): {state['ac']}")
    print(f"Hidro total (preliminar PREMISSAS L38): R$ {hidro_total:,.2f}")
    print(f"Config: torres={config['briefing']['n_torres']}, np={config['np']}, npt={config['npt']}, ur={config['ur']}")

    wb = Workbook()
    ws_pre = wb.active
    ws_pre.title = "Premissas"
    ws_pav = wb.create_sheet("Pavimentos")
    ws_val = wb.create_sheet("Rateio-valores")
    ws_qtd = wb.create_sheet("Rateio-quantidades")
    ws_res = wb.create_sheet("Resumo-por-tipo")

    refs = write_premissas(ws_pre, state, config, hidro_total)
    pavs = build_pavimentos_list(config)
    print(f"Total pavimentos físicos: {len(pavs)}")

    pav_start, pav_end, pav_total_row = write_pavimentos(ws_pav, pavs, refs)
    write_rateio_valores(ws_val, pavs, refs, pav_start, pav_end, pav_total_row)
    write_rateio_quantidades(ws_qtd, pavs, refs, pav_start, pav_end)
    write_resumo_por_tipo(ws_res, refs)

    wb.save(OUT)
    print(f"OK: {OUT}")


if __name__ == "__main__":
    main()
