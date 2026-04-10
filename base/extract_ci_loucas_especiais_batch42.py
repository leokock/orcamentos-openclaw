#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract DETAILED COST DATA from construction budget spreadsheets:
  - ci_detalhado: Custos Indiretos breakdown (projetos, taxas, equipe_adm, epcs, equipamentos, ensaios, canteiro)
  - loucas_metais_detail: Louças e Metais summary (bacias, cubas, torneiras, registros, chuveiros, bancadas, totals)
  - sistemas_especiais_detail: Sistemas Especiais (elevadores, gerador, piscina, ete, automacao/climatizacao)

Batch: projects 42-83 (0-indexed) from _all_projects_mapping.json

Strategies per format:
  1. Ger_Tec e Adm sheet (CTN standard) → CI from code structure (01=Técnicos, 02=Adm, 03=Equip)
  2. Sienge Relatório → CI from second set of codes after obra (01-05)
  3. Ger_Executivo → CI from CÉLULA CONSTRUTIVA with Gerenciamento label
  4. Gerenciamento executivo / Obra summary → macrogroup-only (limited detail)
  5. LOUÇAS sheet / Louças e metais → item-level counts aggregated by category
  6. SISTEMAS ESPECIAIS sheet → item-level values
  7. Relatório N2 items → Louças (09) and Sistemas Especiais (11)
"""
import json, sys, re, os
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

BATCH = ALL_PROJECTS[42:84]  # projects 42-83


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


# ═══════════════════════════════════════════════════════════════════════════
# CI DETALHADO EXTRACTORS
# ═══════════════════════════════════════════════════════════════════════════

def extract_ci_from_ger_tec_adm(wb, sheet_name='Ger_Tec e Adm'):
    """Extract CI from Ger_Tec e Adm sheet (CTN standard with codes 01/02/03)."""
    ws = wb[sheet_name]
    ci = {}

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=16, values_only=True):
        vals = list(row)
        nivel = safe_str(vals[7]) if len(vals) > 7 else ''
        item = safe_str(vals[8]) if len(vals) > 8 else ''
        desc = safe_str(vals[9]) if len(vals) > 9 else ''
        total = safe_float(vals[13]) if len(vals) > 13 else 0.0
        dl = desc.lower()

        if not item or total <= 0:
            continue

        # CÉLULA CONSTRUTIVA level (01, 02, 03)
        if nivel in ('CÉLULA CONSTRUTIVA', 'C\u00c9LULA CONSTRUTIVA') or 'LULA CONSTRUTIVA' in nivel:
            if item == '01':
                ci['servicos_tecnicos_total'] = round(total, 2)
            elif item == '02':
                ci['gerenciamento_adm_total'] = round(total, 2)
            elif item == '03':
                ci['equipamentos_total'] = round(total, 2)

        # ETAPA level (01.001, 01.002, etc.)
        if nivel == 'ETAPA' or 'ETAPA' in nivel:
            if item.startswith('01.'):
                if any(k in dl for k in ['projeto', 'consultoria', 'estudo', 'ensaio']):
                    ci['projetos_consultorias'] = ci.get('projetos_consultorias', 0) + round(total, 2)
                elif any(k in dl for k in ['taxa', 'licen', 'documento', 'consumo']):
                    ci['taxas_licencas'] = ci.get('taxas_licencas', 0) + round(total, 2)
            elif item.startswith('02.'):
                if any(k in dl for k in ['segurança', 'seguranc', 'meio ambiente', 'saúde', 'saude', 'epc']):
                    ci['epcs'] = ci.get('epcs', 0) + round(total, 2)
                elif any(k in dl for k in ['administra', 'canteiro', 'equipe', 'gestão', 'gestao']):
                    ci['equipe_adm_canteiro'] = ci.get('equipe_adm_canteiro', 0) + round(total, 2)
            elif item.startswith('03.'):
                ci['equipamentos_carga'] = ci.get('equipamentos_carga', 0) + round(total, 2)

        # SUBETAPA level for finer detail
        if nivel == 'SUBETAPA' or 'SUBETAPA' in nivel:
            if item.startswith('01.001.'):
                if any(k in dl for k in ['estudo']):
                    ci['estudos'] = ci.get('estudos', 0) + round(total, 2)
                elif any(k in dl for k in ['projeto']):
                    ci['projetos'] = ci.get('projetos', 0) + round(total, 2)
                elif any(k in dl for k in ['consultoria']):
                    ci['consultorias'] = ci.get('consultorias', 0) + round(total, 2)
                elif any(k in dl for k in ['ensaio', 'laudo', 'controle tecnol']):
                    ci['ensaios'] = ci.get('ensaios', 0) + round(total, 2)
            elif item.startswith('01.002.'):
                if any(k in dl for k in ['licen', 'taxa', 'imposto', 'seguro', 'aprova']):
                    ci['licencas_taxas'] = ci.get('licencas_taxas', 0) + round(total, 2)
                elif any(k in dl for k in ['documento']):
                    ci['documentos'] = ci.get('documentos', 0) + round(total, 2)
            elif item.startswith('02.001.'):
                if any(k in dl for k in ['equipamento', 'epc', 'proteç']):
                    ci['epcs_equipamentos'] = ci.get('epcs_equipamentos', 0) + round(total, 2)
                elif any(k in dl for k in ['meio ambiente']):
                    ci['meio_ambiente'] = ci.get('meio_ambiente', 0) + round(total, 2)
            elif item.startswith('02.002.'):
                if any(k in dl for k in ['equipe', 'gestão', 'gestao', 'apoio']):
                    ci['equipe_gestao'] = ci.get('equipe_gestao', 0) + round(total, 2)
                elif any(k in dl for k in ['serviços iniciais', 'servicos iniciais']):
                    ci['servicos_iniciais'] = ci.get('servicos_iniciais', 0) + round(total, 2)
                elif any(k in dl for k in ['despesa', 'consumo', 'manutenç']):
                    ci['despesas_consumo'] = ci.get('despesas_consumo', 0) + round(total, 2)
                elif any(k in dl for k in ['instalações provisórias', 'instalacoes provisor', 'canteiro']):
                    ci['canteiro_instalacoes'] = ci.get('canteiro_instalacoes', 0) + round(total, 2)
                elif any(k in dl for k in ['ligaç', 'ligac', 'provisóri', 'provisori']):
                    ci['ligacoes_provisorias'] = ci.get('ligacoes_provisorias', 0) + round(total, 2)

    if ci:
        ci['_metodo'] = f'Ger_Tec_Adm:{sheet_name}'
    return ci


def extract_ci_from_relatorio(wb, sheet_name='Relatório'):
    """Extract CI from Sienge Relatório second section (after obra items)."""
    ws = wb[sheet_name]
    ci = {}

    # In Relatório format, CI codes restart from 01 after the obra section
    # We need to detect the second occurrence of code "01"
    first_01_found = False
    in_ci_section = False

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=30, values_only=True):
        vals = list(row)
        col0 = safe_str(vals[0]) if vals[0] is not None else ''
        col1 = safe_str(vals[1]) if len(vals) > 1 else ''
        col28 = safe_float(vals[28]) if len(vals) > 28 else 0.0

        code = col0.strip()
        dl = col1.lower()

        # Detect the CI section
        if re.match(r'^01$', code):
            if first_01_found:
                in_ci_section = True
            first_01_found = True

        if not in_ci_section:
            continue

        if not re.match(r'^\d{2}(\.\d{3}){0,2}$', code):
            continue

        total = col28

        # N1 level in CI section
        if re.match(r'^\d{2}$', code):
            if code == '01':
                ci['projetos_consultorias'] = round(total, 2)
            elif code == '02':
                ci['taxas_licencas'] = round(total, 2)
            elif code == '03':
                ci['epcs'] = round(total, 2)
            elif code == '04':
                ci['equipe_adm_canteiro'] = round(total, 2)
            elif code == '05':
                ci['equipamentos_carga'] = round(total, 2)

        # N2 level (XX.XXX) — finer subcategories
        if re.match(r'^\d{2}\.\d{3}$', code):
            pass  # Already captured at N1 level above

        # N3 level (XX.XXX.XXX) for even more detail
        if re.match(r'^\d{2}\.\d{3}\.\d{3}$', code) and total > 0:
            if code.startswith('01.'):
                if any(k in dl for k in ['estudo']):
                    ci['estudos'] = ci.get('estudos', 0) + round(total, 2)
                elif any(k in dl for k in ['projeto']):
                    ci['projetos'] = ci.get('projetos', 0) + round(total, 2)
                elif any(k in dl for k in ['consultoria']):
                    ci['consultorias'] = ci.get('consultorias', 0) + round(total, 2)
                elif any(k in dl for k in ['ensaio', 'laudo']):
                    ci['ensaios'] = ci.get('ensaios', 0) + round(total, 2)
            elif code.startswith('03.'):
                if any(k in dl for k in ['segurança', 'seguranc']):
                    ci['seguranca_trabalho'] = ci.get('seguranca_trabalho', 0) + round(total, 2)
                elif any(k in dl for k in ['equipamento', 'proteç', 'protec']):
                    ci['epcs_equipamentos'] = ci.get('epcs_equipamentos', 0) + round(total, 2)
            elif code.startswith('04.'):
                if any(k in dl for k in ['equipe', 'gestão', 'gestao', 'apoio']):
                    ci['equipe_gestao'] = ci.get('equipe_gestao', 0) + round(total, 2)
                elif any(k in dl for k in ['canteiro']):
                    ci['canteiro_instalacoes'] = ci.get('canteiro_instalacoes', 0) + round(total, 2)
                elif any(k in dl for k in ['despesa', 'consumo']):
                    ci['despesas_consumo'] = ci.get('despesas_consumo', 0) + round(total, 2)
            elif code.startswith('05.'):
                if any(k in dl for k in ['carga', 'transporte']):
                    ci['equipamentos_carga_transporte'] = ci.get('equipamentos_carga_transporte', 0) + round(total, 2)
                elif any(k in dl for k in ['obra']):
                    ci['equipamentos_obra'] = ci.get('equipamentos_obra', 0) + round(total, 2)

    if ci:
        # Calculate total
        n1_keys = ['projetos_consultorias', 'taxas_licencas', 'epcs', 'equipe_adm_canteiro', 'equipamentos_carga']
        ci['ci_total'] = round(sum(ci.get(k, 0) for k in n1_keys), 2)
        ci['_metodo'] = f'Relatório:{sheet_name}'
    return ci


def extract_ci_from_orcamento_executivo(wb, sheet_name):
    """Extract CI from ORÇAMENTO_EXECUTIVO sheet with hierarchical codes (01.01.001 format)
    or flat macrogroup summary format."""
    ws = wb[sheet_name]
    ci = {}

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=10, values_only=True):
        vals = list(row)
        # Format 1 (mabrem): code in col0, desc in col1, total in col5
        code_col0 = safe_str(vals[0]) if vals[0] is not None else ''
        desc_col1 = safe_str(vals[1]) if len(vals) > 1 else ''
        total_col5 = safe_float(vals[5]) if len(vals) > 5 else 0.0
        # Format 2 (macom): desc in col1, total in col2
        desc_col1_alt = safe_str(vals[1]) if len(vals) > 1 else ''
        total_col2 = safe_float(vals[2]) if len(vals) > 2 else 0.0

        # === Format 1: hierarchical codes in col0 ===
        code = code_col0.strip()
        dl = desc_col1.lower()

        if re.match(r'^01$', code) and total_col5 > 0:
            if any(k in dl for k in ['gerenciamento', 'custos indiretos', 'serviços técnicos']):
                ci['ci_total'] = round(total_col5, 2)
        elif re.match(r'^01\.01$', code) and total_col5 > 0:
            if any(k in dl for k in ['gerenciamento técnico', 'gerenciamento tec', 'serviços técnicos']):
                ci['servicos_tecnicos_total'] = round(total_col5, 2)
        elif re.match(r'^01\.01\.001$', code) and total_col5 > 0:
            if any(k in dl for k in ['estudo', 'projeto', 'consultoria']):
                ci['projetos_consultorias'] = round(total_col5, 2)
        elif re.match(r'^01\.01\.001\.\d{4}$', code) and total_col5 > 0:
            if any(k in dl for k in ['estudo']):
                ci['estudos'] = ci.get('estudos', 0) + round(total_col5, 2)
            elif any(k in dl for k in ['projeto']):
                ci['projetos'] = ci.get('projetos', 0) + round(total_col5, 2)
            elif any(k in dl for k in ['consultoria']):
                ci['consultorias'] = ci.get('consultorias', 0) + round(total_col5, 2)
        elif re.match(r'^01\.01\.002$', code) and total_col5 > 0:
            ci['taxas_licencas'] = round(total_col5, 2)
        elif re.match(r'^01\.01\.003$', code) and total_col5 > 0:
            ci['ensaios'] = round(total_col5, 2)
        elif re.match(r'^01\.02$', code) and total_col5 > 0:
            if any(k in dl for k in ['gerenciamento adm', 'administrat']):
                ci['gerenciamento_adm_total'] = round(total_col5, 2)
        elif re.match(r'^01\.02\.001$', code) and total_col5 > 0:
            if any(k in dl for k in ['segurança', 'seguranc', 'epc', 'proteç']):
                ci['epcs'] = round(total_col5, 2)
        elif re.match(r'^01\.02\.002$', code) and total_col5 > 0:
            if any(k in dl for k in ['administra', 'canteiro', 'equipe']):
                ci['equipe_adm_canteiro'] = round(total_col5, 2)
        elif re.match(r'^01\.03$', code) and total_col5 > 0:
            if any(k in dl for k in ['equipamento']):
                ci['equipamentos_total'] = round(total_col5, 2)

        # === Format 2: flat macrogroup summary ===
        # desc in col0 or col1, value in col1 or col2
        for desc_check, val_check in [(code_col0, safe_float(vals[1]) if len(vals) > 1 else 0.0),
                                        (desc_col1_alt, total_col2)]:
            if not desc_check or not isinstance(desc_check, str):
                continue
            dc = desc_check.lower()
            if any(k in dc for k in ['gerenciamento técnico e administrativo', 'despesas indiretas',
                                       'custos indiretos']):
                if val_check > 1000:
                    ci['ci_total'] = round(val_check, 2)

    if ci:
        ci['_metodo'] = f'ORÇAMENTO_EXECUTIVO:{sheet_name}'
    return ci


def extract_ci_from_executivo_summary(wb, sheet_name):
    """Extract CI from Executivo/Gerenciamento executivo summary sheets."""
    ws = wb[sheet_name]
    ci = {}

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        vals = list(row)
        # Try col0=desc, col1=value
        desc0 = safe_str(vals[0]) if len(vals) > 0 else ''
        val1 = safe_float(vals[1]) if len(vals) > 1 else 0.0
        # Try col1=desc, col2=value
        desc1 = safe_str(vals[1]) if len(vals) > 1 else ''
        val2 = safe_float(vals[2]) if len(vals) > 2 else 0.0

        for desc, value in [(desc0, val1), (desc1, val2)]:
            if not desc or value <= 1000:
                continue
            dl = desc.lower()
            if any(k in dl for k in ['gerenciamento técnico e administrativo',
                                       'gerenciamento tec', 'custos indiretos', 'despesas indiretas']):
                ci['ci_total'] = round(value, 2)
                break

    if ci:
        ci['_metodo'] = f'Executivo_summary:{sheet_name}'
    return ci


def extract_ci_from_relatorio_alt(wb, sheet_name):
    """Extract CI from Relatório with alternate column layout (total in col5 instead of col28)."""
    ws = wb[sheet_name]
    ci = {}

    first_01_found = False
    in_ci_section = False

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=10, values_only=True):
        vals = list(row)
        code = safe_str(vals[0]) if vals[0] is not None else ''
        desc = safe_str(vals[1]) if len(vals) > 1 else ''
        # Try column 5 for total
        total = safe_float(vals[5]) if len(vals) > 5 else 0.0
        if total == 0:
            total = safe_float(vals[4]) if len(vals) > 4 else 0.0
        if total == 0:
            total = safe_float(vals[3]) if len(vals) > 3 else 0.0
        dl = desc.lower()

        code = code.strip()

        # Check if first code "01" refers to CI-like content (CANTEIRO, PRELIMINARES)
        if re.match(r'^01\s*$', code):
            if any(k in dl for k in ['canteiro', 'preliminar', 'serviços técnicos', 'servicos tecnicos',
                                       'gerenciamento', 'custos indiretos']):
                ci['ci_total'] = round(total, 2)
                ci['_metodo'] = f'Relatório_alt:{sheet_name}'

    return ci


def extract_ci_from_resumo_orcamento(wb, sheet_name):
    """Extract CI from Orçamento Resumo summary sheet."""
    ws = wb[sheet_name]
    ci = {}

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        desc = ''
        value = 0
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                desc = v
            elif isinstance(v, (int, float)) and v > 1000:
                if value == 0:
                    value = float(v)

        if not desc or value <= 0:
            continue

        dl = desc.lower()
        if any(k in dl for k in ['gerenciamento', 'custos indiretos', 'despesas indiretas']):
            ci['ci_total'] = round(value, 2)

    if ci:
        ci['_metodo'] = f'Resumo:{sheet_name}'
    return ci


def extract_ci_from_obra_sheet(wb, sheet_name):
    """Extract CI total from Obra summary sheet."""
    ws = wb[sheet_name]
    ci = {}

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        if isinstance(vals[0], str):
            dl = vals[0].lower()
            if any(k in dl for k in ['gerenciamento', 'custos indiretos', 'ger_tec', 'serviços técnicos',
                                       'servicos tecnicos', 'administra']):
                for v in vals[1:]:
                    if isinstance(v, (int, float)) and v > 1000:
                        if 'gerenciamento' in dl or 'custos indiretos' in dl:
                            ci['ci_total'] = round(float(v), 2)
                        elif 'serviços técnicos' in dl or 'servicos tecnicos' in dl:
                            ci['projetos_consultorias'] = round(float(v), 2)
                        elif 'administra' in dl:
                            ci['equipe_adm_canteiro'] = round(float(v), 2)
                        break
            elif any(k in dl for k in ['equipamento']) and 'especiai' not in dl:
                for v in vals[1:]:
                    if isinstance(v, (int, float)) and v > 1000:
                        ci['equipamentos_carga'] = round(float(v), 2)
                        break

    if ci:
        ci['_metodo'] = f'Obra:{sheet_name}'
    return ci


def extract_ci_from_epcs_sheet(wb, sheet_name):
    """Extract EPCs total from dedicated EPCs sheet."""
    ws = wb[sheet_name]
    total = 0

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()
        if 'total' in row_text and 'subtotal' not in row_text:
            for v in vals:
                if isinstance(v, (int, float)) and v > 100:
                    if v > total:
                        total = float(v)
                    break

    return round(total, 2) if total > 0 else 0


def extract_ci_from_canteiro_sheet(wb, sheet_name):
    """Extract canteiro total from dedicated CANTEIRO sheet."""
    ws = wb[sheet_name]
    total = 0

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()
        if 'total' in row_text and 'subtotal' not in row_text:
            for v in vals:
                if isinstance(v, (int, float)) and v > 100:
                    if v > total:
                        total = float(v)
                    break

    return round(total, 2) if total > 0 else 0


def extract_ci_from_projetos_sheet(wb, sheet_name):
    """Extract projetos/consultorias total from PROJETOS sheet."""
    ws = wb[sheet_name]
    total = 0

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()
        if 'total' in row_text and 'subtotal' not in row_text:
            for v in vals:
                if isinstance(v, (int, float)) and v > 100:
                    if v > total:
                        total = float(v)
                    break

    return round(total, 2) if total > 0 else 0


def extract_ci_from_ensaios_sheet(wb, sheet_name):
    """Extract ensaios total from Ensaios sheet."""
    ws = wb[sheet_name]
    total = 0

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()
        if 'total' in row_text and 'subtotal' not in row_text:
            for v in vals:
                if isinstance(v, (int, float)) and v > 100:
                    if v > total:
                        total = float(v)
                    break

    return round(total, 2) if total > 0 else 0


# ═══════════════════════════════════════════════════════════════════════════
# LOUÇAS E METAIS EXTRACTORS
# ═══════════════════════════════════════════════════════════════════════════

def classify_louca(desc):
    """Classify a louça/metal item into a category."""
    dl = desc.lower()
    if any(k in dl for k in ['vaso sanitário', 'vaso sanitario', 'bacia sanitária', 'bacia sanitaria',
                              'bacia convencional', 'bacia com caixa', 'bacia suspensa']):
        return 'bacias'
    if any(k in dl for k in ['cuba de embutir', 'cuba de sobrepor', 'cuba de cozinha', 'cuba de apoio',
                              'cuba inox', 'cuba semi-encaixe', 'cuba redonda', 'cuba oval']):
        return 'cubas'
    if any(k in dl for k in ['torneira', 'misturador', 'monocomando']):
        return 'torneiras'
    if any(k in dl for k in ['registro', 'acabamento metálico', 'acabamento metalico', 'acabamento met']):
        return 'registros'
    if any(k in dl for k in ['chuveiro', 'ducha', 'ducha higiênica', 'ducha higien']):
        return 'chuveiros'
    if any(k in dl for k in ['bancada', 'granito', 'mármore', 'marmore', 'pedra']):
        return 'bancadas'
    if any(k in dl for k in ['anel de vedação', 'anel de vedac', 'parafuso', 'sifão', 'sifao',
                              'válvula', 'valvula', 'engate', 'rabicho', 'flexível', 'flexivel',
                              'barra de apoio']):
        return 'acessorios'
    return 'outros'


def extract_loucas_from_sheet(wb, sheet_name):
    """Extract louças e metais from dedicated sheet.
    Two common formats:
    A) Quantity format: col3=qty/unit, col4=repetitions, col5=total_qty
       Header: QUANTIDADE | REPETIÇÃO | TOTAL
    B) Pricing format: col3=unit, col4=preço unitário, col5=preço total
       Header: unidade | Preço unitário | Preço total
    We detect via headers and extract accordingly.
    """
    ws = wb[sheet_name]
    counts = {}  # category -> total quantity
    total_valor = 0
    is_pricing_format = False
    in_pricing_section = False  # For mixed-format sheets

    # First pass: detect format from headers
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=10, values_only=True):
        vals = list(row)
        row_text = ' '.join(safe_str(v).lower() for v in vals)
        if 'preço unitário' in row_text or 'preco unitario' in row_text or 'preço total' in row_text:
            is_pricing_format = True
            break
        if 'repetição' in row_text or 'repeticao' in row_text:
            is_pricing_format = False
            break

    # Second pass: extract data
    for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
        vals = list(row)
        desc = safe_str(vals[1]) if len(vals) > 1 else ''
        col3 = safe_float(vals[3]) if len(vals) > 3 else 0
        col4 = safe_float(vals[4]) if len(vals) > 4 else 0
        col5 = safe_float(vals[5]) if len(vals) > 5 else 0

        if not desc or col5 <= 0:
            continue

        if is_pricing_format:
            # Pricing format: col5 = total monetary value, derive qty = col5/col4
            cat = classify_louca(desc)
            if cat != 'outros' and cat != 'acessorios' and col4 > 0:
                qty = round(col5 / col4)
                if 0 < qty <= 2000:
                    counts[cat] = counts.get(cat, 0) + int(qty)
            # Capture total value from header rows (e.g., "LOUÇAS" total)
            if desc.upper().strip() in ('LOUÇAS', 'LOUCAS', 'LOUÇAS E METAIS'):
                total_valor = round(col5, 2)
        else:
            # Quantity format: col5 = total quantity
            # Detect transition to pricing section
            if col4 > 100 and col5 > 1000 and col3 > 0 and abs(col5 - col3 * col4) < col5 * 0.01:
                in_pricing_section = True

            if in_pricing_section:
                continue

            # Skip unreasonably large "quantities" (these are monetary values)
            if col5 > 2000:
                continue

            cat = classify_louca(desc)
            if cat != 'outros' and cat != 'acessorios':
                counts[cat] = counts.get(cat, 0) + int(col5)

    result = {}
    if total_valor > 0:
        result['total_valor'] = total_valor
    if counts:
        result.update(counts)
    if result:
        result['_metodo'] = f'Louças_sheet:{sheet_name}'
    return result


def extract_loucas_from_relatorio(wb, sheet_name='Relatório'):
    """Extract louças e metais from Sienge Relatório (code 09)."""
    ws = wb[sheet_name]
    counts = {}
    total_value = 0

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=30, values_only=True):
        vals = list(row)
        code = safe_str(vals[0]) if vals[0] is not None else ''
        desc = safe_str(vals[1]) if len(vals) > 1 else ''
        col28 = safe_float(vals[28]) if len(vals) > 28 else 0.0

        if not code:
            continue

        # N1 total for Louças
        if code.strip() == '09' and col28 > 0:
            total_value = round(col28, 2)
            continue

        # N3/N4 items under code 09
        if code.strip().startswith('09.') and desc:
            cat = classify_louca(desc)
            qty = safe_float(vals[4]) if len(vals) > 4 else 0.0  # quantity column varies
            # In relatório, try column with qty
            for ci in [4, 5, 6]:
                if len(vals) > ci and isinstance(vals[ci], (int, float)) and 0 < vals[ci] < 10000:
                    qty = float(vals[ci])
                    break

            if cat != 'outros' and cat != 'acessorios' and qty > 0:
                counts[cat] = counts.get(cat, 0) + int(qty)

    result = {}
    if total_value > 0:
        result['total_valor'] = total_value
    if counts:
        result.update(counts)
    if result:
        result['_metodo'] = f'Relatório:09'
    return result


def extract_loucas_from_ger_executivo(wb, sheet_name):
    """Extract louças e metais total from Ger_Executivo sheet."""
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=16, values_only=True):
        vals = list(row)
        nivel = safe_str(vals[7]) if len(vals) > 7 else ''
        desc = safe_str(vals[9]) if len(vals) > 9 else ''
        total = safe_float(vals[13]) if len(vals) > 13 else 0.0
        dl = desc.lower()

        if ('LULA CONSTRUTIVA' in nivel or nivel in ('CÉLULA CONSTRUTIVA', 'C\u00c9LULA CONSTRUTIVA')) and total > 0:
            if any(k in dl for k in ['louça', 'louca', 'louças', 'loucas', 'metais']):
                return {'total_valor': round(total, 2), '_metodo': f'Ger_Executivo:{sheet_name}'}

        if 'ETAPA' in nivel and total > 0:
            if any(k in dl for k in ['louça', 'louca', 'louças', 'loucas', 'metais']):
                return {'total_valor': round(total, 2), '_metodo': f'Ger_Executivo:{sheet_name}'}

    return {}


def extract_loucas_from_consolidacao(wb, sheet_name):
    """Extract louças from Consolidação/Pavimentos sheet (muller format)."""
    ws = wb[sheet_name]
    counts = {}
    total_valor = 0

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=12, values_only=True):
        vals = list(row)
        nivel = safe_str(vals[1]) if len(vals) > 1 else ''
        desc = safe_str(vals[2]) if len(vals) > 2 else ''
        if not desc:
            desc = safe_str(vals[3]) if len(vals) > 3 else ''
        qtd = safe_float(vals[6]) if len(vals) > 6 else 0
        valor = safe_float(vals[8]) if len(vals) > 8 else 0

        if nivel == 'Serviço' or nivel == 'Servico':
            # This is a service line with quantity and total
            cat = classify_louca(desc)
            if cat != 'outros' and cat != 'acessorios' and qtd > 0:
                counts[cat] = counts.get(cat, 0) + int(qtd)
            if valor > 0:
                total_valor += valor
        elif nivel == 'Etapa':
            # Total for the section
            pass

    result = {}
    if total_valor > 0:
        result['total_valor'] = round(total_valor, 2)
    if counts:
        result.update(counts)
    if result:
        result['_metodo'] = f'Consolidação:{sheet_name}'
    return result


def extract_loucas_from_orcamento_executivo(wb, sheet_name):
    """Extract louças total from ORÇAMENTO_EXECUTIVO sheet."""
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=10, values_only=True):
        vals = list(row)
        # Check multiple column layouts
        for desc_col, val_col in [(1, 5), (0, 5), (1, 2), (0, 1)]:
            desc = safe_str(vals[desc_col]) if len(vals) > desc_col else ''
            total = safe_float(vals[val_col]) if len(vals) > val_col else 0.0
            dl = desc.lower()

            if any(k in dl for k in ['louça', 'louca', 'louças', 'loucas']) and total > 0:
                if not any(k in dl for k in ['insumo', 'serviço', 'servico']):
                    return {'total_valor': round(total, 2), '_metodo': f'ORÇAMENTO_EXECUTIVO:{sheet_name}'}

    return {}


def extract_loucas_from_relatorio_altcol(wb, sheet_name):
    """Extract louças from Relatório with total in alternate column."""
    ws = wb[sheet_name]
    total_value = 0

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=10, values_only=True):
        vals = list(row)
        code = safe_str(vals[0]) if vals[0] is not None else ''
        desc = safe_str(vals[1]) if len(vals) > 1 else ''
        dl = desc.lower()

        if not code.strip():
            continue

        # Find N1 for louças
        if re.match(r'^\d{2}\s*$', code.strip()):
            if any(k in dl for k in ['louça', 'louca', 'louças', 'loucas', 'metais']):
                total = safe_float(vals[5]) if len(vals) > 5 else 0.0
                if total == 0:
                    total = safe_float(vals[4]) if len(vals) > 4 else 0.0
                if total == 0:
                    total = safe_float(vals[3]) if len(vals) > 3 else 0.0
                if total > 0:
                    return {'total_valor': round(total, 2), '_metodo': f'Relatório_alt:{sheet_name}'}

    return {}


# ═══════════════════════════════════════════════════════════════════════════
# SISTEMAS ESPECIAIS EXTRACTORS
# ═══════════════════════════════════════════════════════════════════════════

def classify_sistema_especial(desc):
    """Classify a sistema especial item."""
    dl = desc.lower()
    if any(k in dl for k in ['elevador']):
        return 'elevadores'
    if any(k in dl for k in ['gerador', 'grupo gerador']):
        return 'gerador'
    if any(k in dl for k in ['piscina']):
        return 'piscina'
    if any(k in dl for k in ['ete', 'estação de tratamento', 'estacao de tratamento', 'tratamento de esgoto']):
        return 'ete'
    if any(k in dl for k in ['automação', 'automac', 'domotica', 'domótica']):
        return 'automacao'
    if any(k in dl for k in ['climatiza', 'ar condicionado', 'ar-condicionado', 'split', 'vrf']):
        return 'climatizacao'
    if any(k in dl for k in ['aqueciment', 'boiler', 'solar térmico', 'solar termico']):
        return 'aquecimento'
    if any(k in dl for k in ['churrasqueira', 'exaust', 'ventokit']):
        return 'churrasqueiras_exaustao'
    if any(k in dl for k in ['pressuriza']):
        return 'pressurizacao'
    return 'outros'


def extract_sistemas_from_sheet(wb, sheet_name):
    """Extract from dedicated SISTEMAS ESPECIAIS sheet."""
    ws = wb[sheet_name]
    items = {}

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        desc = ''
        qtd = 0
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                desc = v
            elif isinstance(v, (int, float)) and 0 < v < 10000 and qtd == 0:
                qtd = int(v)

        if desc:
            cat = classify_sistema_especial(desc)
            if cat != 'outros' and qtd > 0:
                if cat not in items:
                    items[cat] = {'qtd': 0}
                items[cat]['qtd'] += qtd

    if items:
        items['_metodo'] = f'Sistemas_sheet:{sheet_name}'
    return items


def extract_sistemas_from_relatorio(wb, sheet_name='Relatório'):
    """Extract sistemas especiais from Relatório (code 11 or similar)."""
    ws = wb[sheet_name]
    items = {}
    total_value = 0

    # Find N1 code for Sistemas Especiais (usually 11)
    sistemas_code = None

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=30, values_only=True):
        vals = list(row)
        code = safe_str(vals[0]) if vals[0] is not None else ''
        desc = safe_str(vals[1]) if len(vals) > 1 else ''
        col28 = safe_float(vals[28]) if len(vals) > 28 else 0.0
        dl = desc.lower()

        if not code.strip():
            continue

        # Find N1 code for sistemas especiais
        if re.match(r'^\d{2}$', code.strip()):
            if any(k in dl for k in ['equipamento', 'sistemas especiais', 'sist']):
                if 'especiai' in dl or 'equipamento' in dl:
                    sistemas_code = code.strip()
                    total_value = round(col28, 2)
                    continue

        # N3/N4 items under the sistemas especiais code
        if sistemas_code and code.strip().startswith(sistemas_code + '.'):
            if re.match(r'^\d{2}\.\d{3}\.\d{3}$', code.strip()) or re.match(r'^\d{2}\.\d{3}\.\d{3}\.\d{3}$', code.strip()):
                cat = classify_sistema_especial(desc)
                if cat != 'outros':
                    if cat not in items:
                        items[cat] = {'valor': 0}
                    if col28 > 0:
                        items[cat]['valor'] = items[cat].get('valor', 0) + round(col28, 2)
                    # Try to get qtd
                    for ci in range(2, 10):
                        if len(vals) > ci and isinstance(vals[ci], (int, float)):
                            v = vals[ci]
                            if 0 < v < 1000:
                                items[cat]['qtd'] = items[cat].get('qtd', 0) + int(v)
                                break

    result = {}
    if total_value > 0:
        result['total_valor'] = total_value
    for cat, data in items.items():
        if cat != '_metodo':
            result[cat] = data
    if result:
        result['_metodo'] = f'Relatório:{sistemas_code}'
    return result


def extract_sistemas_from_ger_executivo(wb, sheet_name):
    """Extract sistemas especiais total and sub-items from Ger_Executivo."""
    ws = wb[sheet_name]
    result = {}
    in_sistemas = False
    sistemas_cell_code = None

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=16, values_only=True):
        vals = list(row)
        nivel = safe_str(vals[7]) if len(vals) > 7 else ''
        item = safe_str(vals[8]) if len(vals) > 8 else ''
        desc = safe_str(vals[9]) if len(vals) > 9 else ''
        total = safe_float(vals[13]) if len(vals) > 13 else 0.0
        dl = desc.lower()

        if not item:
            continue

        # Find CÉLULA CONSTRUTIVA for Sistemas Especiais
        if ('LULA CONSTRUTIVA' in nivel or nivel in ('CÉLULA CONSTRUTIVA', 'C\u00c9LULA CONSTRUTIVA')):
            if any(k in dl for k in ['sistema', 'especiai', 'equipamento']):
                if 'especiai' in dl or ('sistema' in dl and 'elétric' not in dl and 'eletric' not in dl):
                    in_sistemas = True
                    sistemas_cell_code = item
                    result['total_valor'] = round(total, 2)
                    continue
            else:
                in_sistemas = False

        if in_sistemas and total > 0:
            if 'ETAPA' in nivel or 'SUBETAPA' in nivel:
                cat = classify_sistema_especial(desc)
                if cat != 'outros':
                    if cat not in result or not isinstance(result.get(cat), dict):
                        result[cat] = {'valor': 0}
                    result[cat]['valor'] = round(total, 2)

    if result and 'total_valor' in result:
        result['_metodo'] = f'Ger_Executivo:{sheet_name}'
    return result


def extract_sistemas_from_orcamento_executivo(wb, sheet_name):
    """Extract sistemas especiais from ORÇAMENTO_EXECUTIVO sheet."""
    ws = wb[sheet_name]
    result = {}
    in_sistemas = False

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=10, values_only=True):
        vals = list(row)
        # Try multiple column layouts
        for desc_col, val_col in [(1, 5), (0, 5), (1, 2), (0, 1)]:
            desc = safe_str(vals[desc_col]) if len(vals) > desc_col else ''
            total = safe_float(vals[val_col]) if len(vals) > val_col else 0.0
            dl = desc.lower()

            if any(k in dl for k in ['equipamentos e sistemas', 'sistemas especiais', 'sistemas e equipamentos']):
                if total > 1000:
                    result['total_valor'] = round(total, 2)
                    in_sistemas = True
                    break
            elif in_sistemas and total > 0 and desc:
                cat = classify_sistema_especial(desc)
                if cat != 'outros':
                    if cat not in result or not isinstance(result.get(cat), dict):
                        result[cat] = {'valor': 0}
                    result[cat]['valor'] = round(total, 2)
                    break

    if result:
        result['_metodo'] = f'ORÇAMENTO_EXECUTIVO:{sheet_name}'
    return result


def extract_sistemas_from_relatorio_altcol(wb, sheet_name):
    """Extract sistemas from Relatório with total in alternate column."""
    ws = wb[sheet_name]
    result = {}

    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=10, values_only=True):
        vals = list(row)
        code = safe_str(vals[0]) if vals[0] is not None else ''
        desc = safe_str(vals[1]) if len(vals) > 1 else ''
        dl = desc.lower()

        if not code.strip():
            continue

        if re.match(r'^\d{2}\s*$', code.strip()):
            if any(k in dl for k in ['equipamento', 'sistemas especiais', 'sist']):
                if 'especiai' in dl or 'equipamento' in dl:
                    total = safe_float(vals[5]) if len(vals) > 5 else 0.0
                    if total == 0:
                        total = safe_float(vals[4]) if len(vals) > 4 else 0.0
                    if total == 0:
                        total = safe_float(vals[3]) if len(vals) > 3 else 0.0
                    if total > 0:
                        result['total_valor'] = round(total, 2)
                        result['_metodo'] = f'Relatório_alt:{sheet_name}'

    return result


def extract_sistemas_from_executivo_summary(wb, sheet_name):
    """Extract sistemas total from Executivo/Gerenciamento summary sheets."""
    ws = wb[sheet_name]
    result = {}

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        vals = list(row)
        # Try col0=desc, col1=value AND col1=desc, col2=value
        desc0 = safe_str(vals[0]) if len(vals) > 0 else ''
        val1 = safe_float(vals[1]) if len(vals) > 1 else 0.0
        desc1 = safe_str(vals[1]) if len(vals) > 1 else ''
        val2 = safe_float(vals[2]) if len(vals) > 2 else 0.0

        for desc, value in [(desc0, val1), (desc1, val2)]:
            if not desc or value <= 1000:
                continue
            dl = desc.lower()
            if any(k in dl for k in ['equipamentos e sistemas', 'sistemas especiais',
                                       'sistemas e equipamentos', 'equipamentos especiais']):
                result['total_valor'] = round(value, 2)
                break

    if result:
        result['_metodo'] = f'Executivo_summary:{sheet_name}'
    return result


def extract_sistemas_from_piscina_sheet(wb, sheet_name):
    """Extract piscina total from dedicated sheet."""
    ws = wb[sheet_name]
    total = 0

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()
        if 'total' in row_text:
            for v in vals:
                if isinstance(v, (int, float)) and v > 100:
                    if v > total:
                        total = float(v)
                    break

    return round(total, 2) if total > 0 else 0


# ═══════════════════════════════════════════════════════════════════════════
# MAIN EXTRACTION LOGIC
# ═══════════════════════════════════════════════════════════════════════════

def extract_project(slug, path):
    """Extract CI, Louças, and Sistemas Especiais from a project."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, None, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    sheets_lower = {s.lower().strip(): s for s in sheets}

    ci = {}
    loucas = {}
    sistemas = {}
    methods = []

    try:
        # ── Common sheet lists ──────────────────────────────────────────
        relatorio_sheets = [s for s in sheets if 'Relatório' in s or 'relatorio' in s.lower() or 'Relatório' in s]
        ger_exec_sheets = [s for s in sheets if 'Ger_Executivo' in s or 'ger_executivo' in s.lower()]
        orc_exec_sheets = [s for s in sheets if 'ORÇAMENTO_EXECUTIVO' in s.upper() or 'ORCAMENTO_EXECUTIVO' in s.upper()]
        obra_sheets = [s for s in sheets if s.strip().lower() in ('obra', 'obra ')]
        executivo_sheets = [s for s in sheets if s.strip().lower() in ('executivo', 'executivo ')]
        ger_sum_sheets = [s for s in sheets if 'gerenciamento executivo' in s.lower() or 'gerenciamento_exec' in s.lower()]
        resumo_sheets = [s for s in sheets if 'resumo' in s.lower() and 'esquadria' not in s.lower()]

        # ── CI EXTRACTION ──────────────────────────────────────────────

        # Strategy 1: Ger_Tec e Adm (most detailed)
        ger_tec_sheets = [s for s in sheets if 'Ger_Tec' in s or 'ger_tec' in s.lower()]
        if ger_tec_sheets:
            ci = extract_ci_from_ger_tec_adm(wb, ger_tec_sheets[0])
            if ci:
                methods.append(f'CI:Ger_Tec_Adm')

        # Strategy 2: Relatório (Sienge) with col28
        if not ci and relatorio_sheets:
            ci = extract_ci_from_relatorio(wb, relatorio_sheets[0])
            if ci:
                methods.append(f'CI:Relatório')

        # Strategy 3: ORÇAMENTO_EXECUTIVO (mabrem-style with hierarchical codes)
        if not ci and orc_exec_sheets:
            ci = extract_ci_from_orcamento_executivo(wb, orc_exec_sheets[0])
            if ci:
                methods.append(f'CI:ORC_EXEC')

        # Strategy 4: Obra sheet
        if not ci and obra_sheets:
            ci = extract_ci_from_obra_sheet(wb, obra_sheets[0])
            if ci:
                methods.append(f'CI:Obra')

        # Strategy 5: Relatório with alternate column (total in col5)
        if not ci and relatorio_sheets:
            ci = extract_ci_from_relatorio_alt(wb, relatorio_sheets[0])
            if ci:
                methods.append(f'CI:Relatório_alt')

        # Strategy 6: Executivo / Gerenciamento summary — try ALL matching sheets
        if not ci:
            for slist_name, slist in [('Executivo', executivo_sheets), ('Ger_sum', ger_sum_sheets)]:
                for sn in slist:
                    ci = extract_ci_from_executivo_summary(wb, sn)
                    if ci:
                        methods.append(f'CI:{slist_name}:{sn}')
                        break
                if ci:
                    break

        # Strategy 8: Orçamento Resumo
        if not ci and resumo_sheets:
            ci = extract_ci_from_resumo_orcamento(wb, resumo_sheets[0])
            if ci:
                methods.append(f'CI:Resumo')

        # Supplement from dedicated sheets
        if ci:
            # EPCs
            epc_sheets = [s for s in sheets if any(k in s.upper() for k in ['EPC'])]
            epc_sheets = [s for s in epc_sheets if 'ESPECIAI' not in s.upper()]
            if epc_sheets and 'epcs' not in ci:
                epc_val = extract_ci_from_epcs_sheet(wb, epc_sheets[0])
                if epc_val > 0:
                    ci['epcs_from_sheet'] = epc_val

            # CANTEIRO
            canteiro_sheets = [s for s in sheets if 'CANTEIRO' in s.upper()]
            if canteiro_sheets and 'canteiro_instalacoes' not in ci:
                cant_val = extract_ci_from_canteiro_sheet(wb, canteiro_sheets[0])
                if cant_val > 0:
                    ci['canteiro_from_sheet'] = cant_val

            # PROJETOS
            proj_sheets = [s for s in sheets if s.strip().upper() == 'PROJETOS']
            if proj_sheets and 'projetos' not in ci:
                proj_val = extract_ci_from_projetos_sheet(wb, proj_sheets[0])
                if proj_val > 0:
                    ci['projetos_from_sheet'] = proj_val

            # Ensaios
            ensaio_sheets = [s for s in sheets if 'Ensaio' in s or 'ENSAIO' in s.upper()]
            if ensaio_sheets and 'ensaios' not in ci:
                ens_val = extract_ci_from_ensaios_sheet(wb, ensaio_sheets[0])
                if ens_val > 0:
                    ci['ensaios_from_sheet'] = ens_val

        # ── LOUÇAS E METAIS EXTRACTION ──────────────────────────────────

        # Strategy 1: Dedicated Louças sheet
        louca_sheets = [s for s in sheets if any(k in s.lower() for k in ['louça', 'louca', 'louças', 'loucas'])]
        # Prefer Consolidação over Pavimentos
        consolidacao_sheets = [s for s in louca_sheets if 'consolid' in s.lower()]
        if consolidacao_sheets:
            loucas = extract_loucas_from_consolidacao(wb, consolidacao_sheets[0])
            if loucas:
                methods.append(f'Louças:{consolidacao_sheets[0]}')

        if not loucas and louca_sheets:
            loucas = extract_loucas_from_sheet(wb, louca_sheets[0])
            if loucas:
                methods.append(f'Louças:{louca_sheets[0]}')

        # Strategy 2: Relatório (code 09 with col28)
        if not loucas and relatorio_sheets:
            loucas = extract_loucas_from_relatorio(wb, relatorio_sheets[0])
            if loucas:
                methods.append(f'Louças:Relatório')

        # Strategy 3: Relatório alternate columns
        if not loucas and relatorio_sheets:
            loucas = extract_loucas_from_relatorio_altcol(wb, relatorio_sheets[0])
            if loucas:
                methods.append(f'Louças:Relatório_alt')

        # Strategy 4: Ger_Executivo
        if not loucas and ger_exec_sheets:
            loucas = extract_loucas_from_ger_executivo(wb, ger_exec_sheets[0])
            if loucas:
                methods.append(f'Louças:Ger_Executivo')

        # Strategy 5: ORÇAMENTO_EXECUTIVO
        if not loucas and orc_exec_sheets:
            loucas = extract_loucas_from_orcamento_executivo(wb, orc_exec_sheets[0])
            if loucas:
                methods.append(f'Louças:ORC_EXEC')

        # ── SISTEMAS ESPECIAIS EXTRACTION ─────────────────────────────

        # Strategy 1: Dedicated sheet
        sist_sheets = [s for s in sheets if any(k in s.upper() for k in ['SISTEMAS ESPECIAIS', 'SISTEMA ESPECIAL'])]
        if sist_sheets:
            sistemas = extract_sistemas_from_sheet(wb, sist_sheets[0])
            if sistemas:
                methods.append(f'Sistemas:{sist_sheets[0]}')

        # Strategy 2: Relatório (col28)
        if not sistemas and relatorio_sheets:
            sistemas = extract_sistemas_from_relatorio(wb, relatorio_sheets[0])
            if sistemas:
                methods.append(f'Sistemas:Relatório')

        # Strategy 3: Relatório alternate columns
        if not sistemas and relatorio_sheets:
            sistemas = extract_sistemas_from_relatorio_altcol(wb, relatorio_sheets[0])
            if sistemas:
                methods.append(f'Sistemas:Relatório_alt')

        # Strategy 4: Ger_Executivo
        if not sistemas and ger_exec_sheets:
            sistemas = extract_sistemas_from_ger_executivo(wb, ger_exec_sheets[0])
            if sistemas:
                methods.append(f'Sistemas:Ger_Executivo')

        # Strategy 5: ORÇAMENTO_EXECUTIVO
        if not sistemas and orc_exec_sheets:
            sistemas = extract_sistemas_from_orcamento_executivo(wb, orc_exec_sheets[0])
            if sistemas:
                methods.append(f'Sistemas:ORC_EXEC')

        # Strategy 6: Executivo/Gerenciamento summary
        if not sistemas:
            for summary_sheet_list in [executivo_sheets, ger_sum_sheets, resumo_sheets]:
                if summary_sheet_list:
                    sistemas = extract_sistemas_from_executivo_summary(wb, summary_sheet_list[0])
                    if sistemas:
                        methods.append(f'Sistemas:{summary_sheet_list[0]}')
                        break

        # Supplement from dedicated sub-sheets
        if not sistemas.get('piscina'):
            piscina_sheets = [s for s in sheets if 'PISCINA' in s.upper()]
            if piscina_sheets:
                pisc_val = extract_sistemas_from_piscina_sheet(wb, piscina_sheets[0])
                if pisc_val > 0:
                    sistemas['piscina'] = {'valor': pisc_val}
                    if '_metodo' not in sistemas:
                        sistemas['_metodo'] = 'PISCINA_sheet'
                    methods.append('Piscina:sheet')

    except Exception as e:
        wb.close()
        return None, None, None, f"Error: {e}"

    wb.close()
    return ci, loucas, sistemas, '; '.join(methods) if methods else 'no_data'


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    stats = {
        'total': len(BATCH),
        'ci_extracted': 0,
        'loucas_extracted': 0,
        'sistemas_extracted': 0,
        'failed': 0,
        'no_data': 0,
    }

    details = []

    for p in BATCH:
        slug = p['slug']
        path = p['path']

        print(f"\n{'='*60}")
        print(f"Processing: {slug}")

        ci, loucas, sistemas, msg = extract_project(slug, path)

        if ci is None:
            print(f"  FAILED: {msg}")
            stats['failed'] += 1
            details.append({'slug': slug, 'status': 'failed', 'reason': msg})
            continue

        has_data = False

        if ci:
            print(f"  CI: {list(ci.keys())}")
            stats['ci_extracted'] += 1
            has_data = True

        if loucas:
            print(f"  Louças: {list(loucas.keys())}")
            stats['loucas_extracted'] += 1
            has_data = True

        if sistemas:
            print(f"  Sistemas: {list(sistemas.keys())}")
            stats['sistemas_extracted'] += 1
            has_data = True

        if not has_data:
            print(f"  NO DATA: {msg}")
            stats['no_data'] += 1
            details.append({'slug': slug, 'status': 'no_data', 'reason': msg})
            continue

        print(f"  Methods: {msg}")

        # Update JSON file
        json_path = INDICES_DIR / f"{slug}.json"
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            print(f"  WARNING: {json_path} does not exist, creating")
            data = {'slug': slug}

        if ci:
            # Clean up _metodo
            metodo = ci.pop('_metodo', '')
            data['ci_detalhado'] = ci
            data['ci_detalhado']['_metodo'] = metodo

        if loucas:
            metodo = loucas.pop('_metodo', '')
            data['loucas_metais_detail'] = loucas
            data['loucas_metais_detail']['_metodo'] = metodo

        if sistemas:
            metodo = sistemas.pop('_metodo', '')
            # Convert any nested dicts for JSON serialization
            clean_sistemas = {}
            for k, v in sistemas.items():
                clean_sistemas[k] = v
            data['sistemas_especiais_detail'] = clean_sistemas
            data['sistemas_especiais_detail']['_metodo'] = metodo

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        details.append({'slug': slug, 'status': 'extracted',
                        'ci': bool(ci), 'loucas': bool(loucas), 'sistemas': bool(sistemas),
                        'methods': msg})

    # ── SUMMARY ──────────────────────────────────────────────────────────
    print(f"\n\n{'='*60}")
    print(f"SUMMARY — Batch 42-83")
    print(f"{'='*60}")
    print(f"Total projects: {stats['total']}")
    print(f"CI extracted: {stats['ci_extracted']}")
    print(f"Louças extracted: {stats['loucas_extracted']}")
    print(f"Sistemas Especiais extracted: {stats['sistemas_extracted']}")
    print(f"Failed to open: {stats['failed']}")
    print(f"No data found: {stats['no_data']}")

    # Per-project summary
    print(f"\nPer-project results:")
    for d in details:
        status = d['status']
        if status == 'extracted':
            parts = []
            if d.get('ci'):
                parts.append('CI')
            if d.get('loucas'):
                parts.append('Louças')
            if d.get('sistemas'):
                parts.append('Sistemas')
            print(f"  {d['slug']}: {', '.join(parts)}")
        else:
            print(f"  {d['slug']}: {status} — {d.get('reason', '')}")


if __name__ == '__main__':
    main()
