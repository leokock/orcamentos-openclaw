#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract detailed cost data from construction budget spreadsheets:
  - ci_detalhado: Custos Indiretos (projetos, taxas, equipe_adm, epcs, equipamentos, ensaios, canteiro)
  - loucas_metais_detail: Louças e Metais (bacias, cubas, torneiras, registros, chuveiros, bancadas, total)
  - sistemas_especiais_detail: Sistemas Especiais (elevadores, gerador, piscina, ete, automacao)

Batch: projects 84-125.
"""
import json
import sys
import os
import re
import traceback

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl

BASE_DIR = r"C:\Users\leona\orcamentos-openclaw\base"
INDICES_DIR = os.path.join(BASE_DIR, "indices-executivo")
MAPPING_FILE = os.path.join(BASE_DIR, "_all_projects_mapping.json")
BATCH_START = 84
BATCH_END = 126

# ============================================================
# Helpers
# ============================================================
def safe_str(val):
    if val is None:
        return ""
    return str(val).strip()

def safe_num(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val, 2) if val != 0 else None
    s = str(val).strip().replace(",", ".")
    m = re.search(r'[\d]+[.]?[\d]*', s)
    if m:
        try:
            v = float(m.group())
            return round(v, 2) if v != 0 else None
        except:
            pass
    return None

def normalize(s):
    if not s:
        return ""
    s = s.lower().strip()
    for a, b in {'á':'a','à':'a','â':'a','ã':'a','é':'e','ê':'e','è':'e',
                  'í':'i','î':'i','ó':'o','ô':'o','õ':'o','ò':'o',
                  'ú':'u','û':'u','ù':'u','ç':'c','ñ':'n'}.items():
        s = s.replace(a, b)
    return s

def find_sheet(wb, candidates):
    sheets_lower = {s.lower().strip(): s for s in wb.sheetnames}
    # First pass: exact matches only
    for c in candidates:
        c_lower = c.lower().strip()
        if c_lower in sheets_lower:
            return wb[sheets_lower[c_lower]]
    # Second pass: partial matches
    for c in candidates:
        c_lower = c.lower().strip()
        for k, v in sheets_lower.items():
            if c_lower in k:
                return wb[v]
    return None

def read_all_rows(ws, max_rows=500):
    rows = []
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        rows.append(list(row))
    return rows

def detect_value_col(rows, max_scan=20):
    """Auto-detect the column containing 'Preço total' / 'Custo total' / 'Valor total'."""
    for rn, row in enumerate(rows[:max_scan]):
        if not row:
            continue
        for ci, cell in enumerate(row):
            s = normalize(safe_str(cell))
            if s in ['preco total', 'custo total', 'valor total', 'total r$',
                      'valor orcado', 'valor orçado', 'valor']:
                return ci
    return None

def is_valid_value(v):
    """Filter out item codes (huge integers like 1001000001) from actual monetary values."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        # Codes are typically > 100M and are integers
        if v > 50000000 and v == int(v):
            return False
        return v > 0
    return False

def get_row_value(row, val_col=None):
    """Get the monetary value from a row, using known column or rightmost valid number."""
    if val_col is not None and val_col < len(row):
        v = safe_num(row[val_col])
        if v and is_valid_value(v):
            return v
    # Fallback: rightmost valid number
    for ci in range(len(row)-1, 0, -1):
        v = safe_num(row[ci])
        if v and is_valid_value(v):
            return v
    return None

def get_row_desc(row, max_col=10):
    """Get description text from a row."""
    for ci in range(min(len(row), max_col)):
        s = safe_str(row[ci])
        if len(s) > 3 and not re.match(r'^[\d.,\s]+$', s) and s not in ('CÉLULA CONSTRUTIVA', 'ETAPA', 'SUBETAPA', 'SERVIÇO', 'vb', 'un', 'mes', 'm', 'm2', 'm3', 'kg'):
            return s
    return ""

# ============================================================
# CI from Relatório (auto-detect value column)
# ============================================================
def extract_ci_from_relatorio(wb):
    ws = find_sheet(wb, ['Relatório', 'Relatorio'])
    if not ws:
        return None, None, None

    rows = read_all_rows(ws, max_rows=1500)
    val_col = detect_value_col(rows)

    if val_col is None:
        return None, None, None

    # Auto-detect code column (usually 0, but sometimes 1 when col 0 has row numbers)
    code_col = 0
    desc_col = 1
    for row in rows[:20]:
        if not row:
            continue
        s0 = normalize(safe_str(row[0]))
        s1 = normalize(safe_str(row[1])) if len(row) > 1 else ""
        if s0 == 'codigo' or re.match(r'^\d{2}\s*$', safe_str(row[0]).strip()):
            code_col = 0
            desc_col = 1
            break
        if s1 == 'codigo' or (len(row) > 1 and re.match(r'^\d{2}\s*$', safe_str(row[1]).strip())):
            code_col = 1
            desc_col = 2
            break

    ci = {}
    sistemas = {}
    loucas_total = None

    # Track unidade construtiva sections
    current_uc = ""
    in_ci_uc = False

    for rn, row in enumerate(rows):
        if not row or len(row) <= val_col:
            continue

        # Detect unidade construtiva
        text_full = ' '.join(safe_str(c) for c in row[:8])
        text_norm = normalize(text_full)
        if 'unidade construtiva' in text_norm:
            current_uc = text_full
            in_ci_uc = 'gerenciamento' in text_norm or 'administrativ' in text_norm
            continue

        code = safe_str(row[code_col]).strip() if code_col < len(row) else ""
        desc = safe_str(row[desc_col]) if desc_col < len(row) else ""
        val = safe_num(row[val_col])

        if not code or not desc:
            continue

        desc_norm = normalize(desc)

        # N2 level (XX.XXX) or N1 level (XX)
        is_n2 = bool(re.match(r'^\d{2}\.\d{3}\s*$', code))
        is_n1 = bool(re.match(r'^\d{2}\s*$', code))
        is_n3 = bool(re.match(r'^\d{2}\.\d{3}\.\d{3}\s*$', code))

        if not val or not is_valid_value(val):
            continue

        # CI extraction
        if is_n2 or is_n1 or is_n3:
            # Projetos e Consultorias
            if any(k in desc_norm for k in ['estudos, projetos e consultoria', 'projetos e consultoria', 'estudos, projetos, consultoria']):
                if not ci.get('projetos_consultorias') or (is_n2 and val > 1000):
                    ci['projetos_consultorias'] = val

            # Taxas e Licenças
            elif any(k in desc_norm for k in ['consumos, taxas', 'taxas e documentos']):
                ci['taxas_licencas'] = val

            # Equipe ADM
            elif any(k in desc_norm for k in ['equipe administrativa', 'equipe de obra', 'equipe administrat']):
                ci['equipe_adm'] = val

            # EPCs
            elif any(k in desc_norm for k in ['seguranca, meio ambiente', 'equipamentos de protecao']):
                ci['epcs'] = val

            # Equipamentos de carga
            elif any(k in desc_norm for k in ['equipamentos de carga', 'maquinas e equipamentos', 'equipamentos']):
                if is_n2 or is_n1:
                    if 'sistemas' not in desc_norm:
                        ci['equipamentos_carga'] = val

            # Ensaios
            elif any(k in desc_norm for k in ['ensaios e laudos', 'controle tecnologico', 'ensaios']):
                if is_n2 or is_n3:
                    ci['ensaios'] = val

            # Canteiro
            elif any(k in desc_norm for k in ['canteiro de obra', 'instalacoes de canteiro', 'administracao e canteiro', 'administração e canteiro']):
                ci['canteiro'] = val

            # Licenciamentos (N3)
            elif any(k in desc_norm for k in ['licenciamentos, taxas']):
                if not ci.get('taxas_licencas'):
                    ci['taxas_licencas'] = val

            # N1 level CI sections in compact Relatório
            if is_n1:
                if 'administracao e canteiro' in desc_norm or 'administração e canteiro' in desc_norm:
                    ci['canteiro'] = val
                elif 'gerenciamento tecnico' in desc_norm or 'gerenciamento administrativo' in desc_norm:
                    if not ci.get('projetos_consultorias'):
                        # This is a broader category, store as overall CI
                        pass

            # Sistemas Especiais
            if any(k in desc_norm for k in ['equipamentos e sistemas especiais', 'sistemas especiais',
                                              'sistemas e equipamentos especiais']):
                sistemas['total'] = val
            elif 'elevador' in desc_norm and 'porta' not in desc_norm and 'cremalheira' not in desc_norm and val > 10000:
                # Take highest value, not accumulated
                if 'elevadores' not in sistemas or val > sistemas['elevadores']['valor']:
                    sistemas['elevadores'] = {'valor': val}
            elif 'gerador' in desc_norm:
                sistemas['gerador'] = val
            elif 'piscina' in desc_norm and val > 1000:
                sistemas['piscina'] = val
            elif 'automacao' in desc_norm or 'automação' in desc_norm:
                if 'sistemas logicos' in desc_norm or val > 10000:
                    sistemas['automacao'] = val

            # Louças
            if any(k in desc_norm for k in ['loucas e metais', 'aparelhos sanitarios loucas']):
                loucas_total = val

    ci_result = ci if ci else None
    sist_result = sistemas if sistemas else None
    loucas_result = {'total': loucas_total} if loucas_total else None

    return ci_result, loucas_result, sist_result

# ============================================================
# CI from Ger_Tec e Adm (Cartesian EAP format)
# ============================================================
def extract_ci_from_ger_tec(wb):
    ws = find_sheet(wb, ['Ger_Tec e Adm', 'Ger_Administrativo', 'Gerenciamento'])
    if not ws:
        return None

    rows = read_all_rows(ws, max_rows=300)

    ci = {}

    # Detect format: Cartesian (text codes like 01.001) or Neuhaus (numeric codes like 1001000001)
    # Also detect value column
    val_col = detect_value_col(rows)

    for rn, row in enumerate(rows):
        if not row:
            continue

        desc = get_row_desc(row)
        desc_norm = normalize(desc)

        if not desc_norm:
            continue

        val = get_row_value(row, val_col)
        if not val or not is_valid_value(val):
            continue

        # Check for N2 subtotal patterns
        # Text-code format: look for ETAPA indicator
        is_subtotal = False
        text_parts = [safe_str(c) for c in row[:10] if c is not None]
        for t in text_parts:
            if t in ('ETAPA',):
                is_subtotal = True
                break
            # Numeric code subtotals: 4-digit codes like 1001, 2004
            if re.match(r'^\d{1,4}$', t) and int(t) < 10000 and int(t) > 0:
                is_subtotal = True
                break

        # Also check if row has a section-level code
        code = safe_str(row[0]) if row[0] else ""
        if re.match(r'^\d{1,2}$', code.strip()) and val > 10000:
            is_subtotal = True

        if any(k in desc_norm for k in ['estudos, projetos e consultoria', 'projetos e consultoria', 'estudos, projetos, consultoria', 'estudos, ensaios, laudos, memoriais']):
            ci['projetos_consultorias'] = val
        elif any(k in desc_norm for k in ['consumos, taxas', 'taxas e documentos', 'taxas, impostos, aprovacoes']):
            ci['taxas_licencas'] = val
        elif any(k in desc_norm for k in ['equipe administrativa', 'equipe de obra', 'equipe adm']):
            ci['equipe_adm'] = val
        elif any(k in desc_norm for k in ['ensaios e laudos', 'controle tecnologico']) and is_subtotal:
            ci['ensaios'] = val
        elif desc_norm == 'ensaios' and is_subtotal:
            ci['ensaios'] = val
        elif any(k in desc_norm for k in ['terreno - incorporacao', 'terreno']):
            if val > 100000 and is_subtotal:
                ci['taxas_licencas'] = val
        elif any(k in desc_norm for k in ['projetos']):
            if is_subtotal and not ci.get('projetos_consultorias') and val > 50000:
                ci['projetos_consultorias'] = val
        elif any(k in desc_norm for k in ['consultorias']):
            if is_subtotal and val > 10000:
                if ci.get('projetos_consultorias'):
                    ci['projetos_consultorias'] += val
                else:
                    ci['projetos_consultorias'] = val
        elif any(k in desc_norm for k in ['aprovacoes e legalizacoes', 'aprovacao e legalizacao']):
            if is_subtotal:
                ci['taxas_licencas'] = val

    return ci if ci else None

# ============================================================
# EPC and CANTEIRO totals
# ============================================================
def extract_sheet_total(wb, sheet_names, max_rows=150):
    """Extract total from a sheet by finding the largest value or a TOTAL row."""
    ws = find_sheet(wb, sheet_names)
    if not ws:
        return None

    rows = read_all_rows(ws, max_rows)
    val_col = detect_value_col(rows)

    total = None
    total_from_label = None

    for rn, row in enumerate(rows):
        if not row:
            continue
        text = normalize(' '.join(safe_str(c) for c in row[:5]))

        for ci_idx in range(len(row)):
            v = safe_num(row[ci_idx])
            if v and is_valid_value(v) and v > 100:
                if 'total' in text and ('geral' in text or rn > 10):
                    total_from_label = v
                if total is None or v > total:
                    total = v

    return total_from_label if total_from_label else total

# ============================================================
# Louças e Metais
# ============================================================
def classify_louca(desc):
    d = normalize(desc)
    if any(k in d for k in ['bacia', 'vaso sanitario', 'vaso sanit', 'caixa acoplada']):
        return 'bacias'
    if any(k in d for k in ['cuba']):
        return 'cubas'
    if any(k in d for k in ['torneira', 'misturador']):
        return 'torneiras'
    if any(k in d for k in ['registro', 'acabamento metalico para registro', 'acabamento para registro']):
        return 'registros'
    if any(k in d for k in ['chuveiro', 'ducha']):
        return 'chuveiros'
    if any(k in d for k in ['bancada']):
        return 'bancadas'
    return 'outros'

def is_quantity_only_sheet(rows):
    """Check if a sheet is quantity-only (no pricing columns)."""
    for row in rows[:15]:
        if not row:
            continue
        for cell in row[:20]:
            s = normalize(safe_str(cell))
            if any(k in s for k in ['preco unitario', 'custo unitario', 'preco total', 'custo total', 'valor total', 'r$']):
                return False
    return True

def extract_loucas_metais(wb):
    ws = find_sheet(wb, ['LOUÇAS E METAIS', 'Louças e metais', 'Louças e Metais', 'LOUÇAS E METAIS RESUMO'])
    if not ws:
        return None

    rows = read_all_rows(ws, max_rows=500)

    # Skip quantity-only sheets (no pricing data)
    if is_quantity_only_sheet(rows):
        return None

    detail = {
        'bacias': 0, 'cubas': 0, 'torneiras': 0,
        'registros': 0, 'chuveiros': 0, 'bancadas': 0,
        'outros': 0, 'total': 0,
    }

    # Detect format and value column
    val_col = None
    desc_col = None

    for rn, row in enumerate(rows[:15]):
        if not row:
            continue
        for ci, cell in enumerate(row[:15]):
            s = normalize(safe_str(cell))
            if s in ['custo total', 'valor total', 'preco total', 'total r$']:
                val_col = ci
            elif s == 'total' and ci > 3:
                # Check if this is a real value column (not quantity)
                # Look at the neighboring column - if there's 'Preço unitário' nearby, it's cost
                has_price_neighbor = False
                for ci2, cell2 in enumerate(row[:15]):
                    s2 = normalize(safe_str(cell2))
                    if any(k in s2 for k in ['preco', 'custo', 'r$']):
                        has_price_neighbor = True
                        break
                if has_price_neighbor:
                    val_col = ci
            if s in ['descricao', 'descrição', 'descricao do material']:
                desc_col = ci

    # Strategy 1: Value column identified, extract per-item
    if val_col is not None and desc_col is not None:
        for row in rows[5:]:
            if not row or len(row) <= max(val_col, desc_col):
                continue
            desc = safe_str(row[desc_col])
            val = safe_num(row[val_col])
            if desc and val and val > 0 and is_valid_value(val):
                cat = classify_louca(desc)
                detail[cat] += val
                detail['total'] += val

    # Strategy 2: Look for per-line costs (Custo Total in col after Qty)
    elif val_col is not None:
        for row in rows[5:]:
            if not row or len(row) <= val_col:
                continue
            desc = get_row_desc(row, 8)
            val = safe_num(row[val_col])
            if desc and val and val > 0 and is_valid_value(val):
                cat = classify_louca(desc)
                detail[cat] += val
                detail['total'] += val

    # Strategy 3: No value column - look for subtotals at bottom
    if detail['total'] == 0:
        grand_total = None
        for row in rows:
            if not row:
                continue
            text = normalize(' '.join(safe_str(c) for c in row[:5]))
            if 'total' in text:
                for ci in range(len(row)):
                    v = safe_num(row[ci])
                    if v and is_valid_value(v) and v > 100:
                        if grand_total is None or v > grand_total:
                            grand_total = v
        if grand_total:
            detail['total'] = grand_total

    # Clean up
    for k in list(detail.keys()):
        if detail[k] == 0 and k != 'total':
            detail[k] = None
        elif isinstance(detail[k], float):
            detail[k] = round(detail[k], 2)

    if detail['total'] == 0:
        return None

    return detail

# ============================================================
# Sistemas Especiais
# ============================================================
def extract_sistemas_especiais(wb):
    ws = find_sheet(wb, ['SISTEMAS ESPECIAIS', 'Sistemas especiais', 'Sistemas Especiais'])
    if not ws:
        return None

    rows = read_all_rows(ws, max_rows=200)
    val_col = detect_value_col(rows)

    detail = {
        'elevadores': None, 'gerador': None, 'piscina': None,
        'ete': None, 'automacao': None, 'cftv': None,
        'interfone': None, 'som_ambiente': None, 'total': 0,
    }

    for row in rows:
        if not row:
            continue

        desc = get_row_desc(row, 10)
        val = get_row_value(row, val_col)

        if not desc or not val:
            continue

        desc_norm = normalize(desc)

        # Match known categories - take highest value for elevadores (avoid sub-item accumulation)
        if 'elevador' in desc_norm and 'porta' not in desc_norm and 'cremalheira' not in desc_norm:
            if detail['elevadores'] is None:
                detail['elevadores'] = {'valor': val}
            elif val > detail['elevadores']['valor']:
                detail['elevadores']['valor'] = val
        elif 'gerador' in desc_norm:
            detail['gerador'] = (detail['gerador'] or 0) + val
        elif 'piscina' in desc_norm:
            detail['piscina'] = (detail['piscina'] or 0) + val
        elif any(k in desc_norm for k in ['ete', 'tratamento de esgoto', 'estacao de tratamento']):
            detail['ete'] = (detail['ete'] or 0) + val
        elif any(k in desc_norm for k in ['automacao', 'automação']):
            detail['automacao'] = (detail['automacao'] or 0) + val
        elif any(k in desc_norm for k in ['cftv', 'circuito fechado', 'sistema cftv']):
            detail['cftv'] = (detail['cftv'] or 0) + val
        elif any(k in desc_norm for k in ['interfone', 'video porteiro', 'videoporteiro']):
            detail['interfone'] = (detail['interfone'] or 0) + val
        elif any(k in desc_norm for k in ['som de ambiente', 'som ambiente']):
            detail['som_ambiente'] = (detail['som_ambiente'] or 0) + val

        if 'total' == desc_norm.strip() and val > detail['total']:
            detail['total'] = val

    # Calculate total if not found
    if detail['total'] == 0:
        s = 0
        for k, v in detail.items():
            if k == 'total':
                continue
            if isinstance(v, dict):
                s += v.get('valor', 0) or 0
            elif isinstance(v, (int, float)) and v:
                s += v
        detail['total'] = round(s, 2) if s else None
    else:
        detail['total'] = round(detail['total'], 2)

    # Round all values
    if detail['elevadores']:
        detail['elevadores']['valor'] = round(detail['elevadores']['valor'], 2)
    for k in ['gerador', 'piscina', 'ete', 'automacao', 'cftv', 'interfone', 'som_ambiente']:
        if detail[k]:
            detail[k] = round(detail[k], 2)

    if all(v is None or v == 0 for k, v in detail.items() if k != 'total') and not detail.get('total'):
        return None

    return detail

# ============================================================
# ORÇAMENTO_EXECUTIVO / EAP extraction
# ============================================================
def extract_from_orc_executivo(wb):
    # Try exact matches first, then partial; prioritize sheets more likely to have data
    ws = find_sheet(wb, ['orçamento', 'ORÇAMENTO_EXECUTIVO', 'ORÇAMENTO_EXECUTIVO_COMENTADO',
                          'EAP', 'Executivo', 'Executivo ', 'Gerenciamento executivo',
                          'Gerenciamento executivo ', 'ORÇAMENTO_PARAMETRICO',
                          'ORÇAMENTO_PARAMÉTRICO'])
    if not ws:
        return None, None, None

    rows = read_all_rows(ws, max_rows=1500)
    val_col = detect_value_col(rows)

    if val_col is None:
        # Try to find from data - rightmost column with values > 1000
        for row in rows[:30]:
            if not row:
                continue
            for ci in range(len(row)-1, 0, -1):
                if isinstance(row[ci], (int, float)) and row[ci] > 1000 and is_valid_value(row[ci]):
                    val_col = ci
                    break
            if val_col:
                break

    if val_col is None:
        return None, None, None

    ci = {}
    sistemas = {}
    loucas_total = None

    for row in rows:
        if not row or len(row) <= val_col:
            continue

        desc = get_row_desc(row, 10)
        val = safe_num(row[val_col])
        if not desc or not val or not is_valid_value(val):
            continue

        desc_norm = normalize(desc)

        # CI
        if any(k in desc_norm for k in ['estudos, projetos e consultoria', 'projetos e consultoria']):
            ci['projetos_consultorias'] = val
        elif any(k in desc_norm for k in ['consumos, taxas', 'taxas e documentos']):
            ci['taxas_licencas'] = val
        elif any(k in desc_norm for k in ['equipe administrativa', 'equipe adm']):
            ci['equipe_adm'] = val
        elif any(k in desc_norm for k in ['seguranca, meio ambiente', 'equipamentos de protecao']):
            ci['epcs'] = val
        elif any(k in desc_norm for k in ['equipamentos de carga', 'maquinas e equipamentos']):
            ci['equipamentos_carga'] = val
        elif any(k in desc_norm for k in ['ensaios e laudos', 'controle tecnologico']):
            ci['ensaios'] = val
        elif any(k in desc_norm for k in ['canteiro de obra', 'instalacoes de canteiro', 'operacao do canteiro', 'administracao e canteiro']):
            ci['canteiro'] = val

        # Sistemas - only assign, don't accumulate (avoid double-counting)
        if any(k in desc_norm for k in ['equipamentos e sistemas especiais', 'sistemas especiais',
                                          'sistemas e equipamentos especiais']):
            sistemas['total'] = val
        elif 'elevador' in desc_norm and 'porta' not in desc_norm and 'cremalheira' not in desc_norm and val > 10000:
            # Only take the first/largest elevador entry (avoid accumulating sub-items)
            if 'elevadores' not in sistemas:
                sistemas['elevadores'] = {'valor': val}
            elif val > sistemas['elevadores']['valor']:
                sistemas['elevadores'] = {'valor': val}
        elif 'gerador' in desc_norm and 'energia' not in desc_norm:
            sistemas['gerador'] = val
        elif 'piscina' in desc_norm and val > 1000:
            sistemas['piscina'] = val
        elif any(k in desc_norm for k in ['automacao', 'automação', 'sistemas logicos']):
            sistemas['automacao'] = val

        # Louças
        if any(k in desc_norm for k in ['loucas e metais', 'aparelhos sanitarios loucas']):
            loucas_total = val

    ci_result = ci if ci else None
    sist_result = sistemas if sistemas else None
    loucas_result = {'total': loucas_total} if loucas_total else None

    return ci_result, loucas_result, sist_result

# ============================================================
# Compact Relatório (N1 only, multiple UC sections)
# ============================================================
def extract_from_compact_relatorio(wb):
    """Handle compact Relatório with N1-only rows and multiple unidade construtiva sections."""
    ws = find_sheet(wb, ['Relatório', 'Relatorio'])
    if not ws:
        return None, None, None

    rows = read_all_rows(ws, max_rows=200)
    val_col = detect_value_col(rows)

    if val_col is None:
        return None, None, None

    ci = {}
    sistemas = {}
    loucas_total = None

    in_ci_section = False

    for rn, row in enumerate(rows):
        if not row or len(row) <= val_col:
            continue

        text_full = ' '.join(safe_str(c) for c in row[:8])
        text_norm = normalize(text_full)

        # Section detection
        if 'total da unidade construtiva' in text_norm:
            in_ci_section = False
            continue

        if 'unidade construtiva' in text_norm or 'obra' in text_norm[:10].lower():
            if 'gerenciamento' in text_norm or 'administrativ' in text_norm:
                in_ci_section = True
            else:
                in_ci_section = False
            continue

        code = safe_str(row[0]).strip()
        desc = safe_str(row[1]) if len(row) > 1 else ""
        val = safe_num(row[val_col])
        desc_norm = normalize(desc)

        if not val or not is_valid_value(val):
            continue

        is_n1 = bool(re.match(r'^\d{2}\s*$', code))

        if is_n1:
            # CI items in CI section
            if in_ci_section:
                if 'seguranca' in desc_norm or 'epc' in desc_norm:
                    ci['epcs'] = val
                elif 'administracao' in desc_norm or 'canteiro' in desc_norm:
                    ci['canteiro'] = val
                elif 'equipamentos' in desc_norm:
                    ci['equipamentos_carga'] = val

            # Sistemas
            if any(k in desc_norm for k in ['equipamentos e sistemas especiais', 'sistemas especiais']):
                sistemas['total'] = val

    ci_result = ci if ci else None
    sist_result = sistemas if sistemas else None

    return ci_result, None, sist_result

# ============================================================
# Piscina from dedicated sheet
# ============================================================
def extract_piscina_total(wb):
    ws = find_sheet(wb, ['PISCINAS', 'PISCINA', 'Piscina | Sauna | Hidro', 'Piscina'])
    if not ws:
        return None
    return extract_sheet_total(wb, ['PISCINAS', 'PISCINA', 'Piscina | Sauna | Hidro', 'Piscina'], 100)

# ============================================================
# Main processing per project
# ============================================================
def process_project(project_info, idx):
    slug = project_info['slug']
    path = project_info['path']

    result = {
        'ci_detalhado': None,
        'loucas_metais_detail': None,
        'sistemas_especiais_detail': None,
    }

    if not os.path.exists(path):
        print(f"  [{idx}] {slug}: FILE NOT FOUND")
        return result

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets_lower = [s.lower().strip() for s in wb.sheetnames]

        has_ger = any('ger_tec' in s or 'ger_administrativ' in s for s in sheets_lower)
        has_loucas = any('louça' in s or 'louca' in s or 'metais' in s for s in sheets_lower)
        has_sist = any('sistema' in s and 'especiai' in s for s in sheets_lower)
        has_rel = any('relatório' in s or 'relatorio' in s for s in sheets_lower)
        has_epc = any('epc' in s for s in sheets_lower)
        has_cant = any('canteiro' in s for s in sheets_lower)
        has_orc_exec = any(s in ('orçamento_executivo', 'orcamento_executivo', 'orçamento_executivo_comentado',
                                  'orçamento', 'orcamento', 'orçamento_parametrico', 'orçamento_paramétrico') for s in sheets_lower)
        has_ger_exec = any('gerenciamento executivo' in s for s in sheets_lower)
        has_eap = any(s.strip() == 'eap' for s in sheets_lower)

        # ── CI ────────────────────────────────────────────────
        ci = None

        if has_ger:
            ci = extract_ci_from_ger_tec(wb)

        if has_rel:
            ci_rel, loucas_rel, sist_rel = extract_ci_from_relatorio(wb)
            if ci_rel and not ci:
                ci = ci_rel
            elif ci_rel and ci:
                # Merge
                for k, v in ci_rel.items():
                    if v and not ci.get(k):
                        ci[k] = v
            if loucas_rel and not result['loucas_metais_detail']:
                result['loucas_metais_detail'] = loucas_rel
            if sist_rel and not result['sistemas_especiais_detail']:
                result['sistemas_especiais_detail'] = sist_rel

        if has_orc_exec or has_ger_exec or has_eap:
            ci_orc, loucas_orc, sist_orc = extract_from_orc_executivo(wb)
            if ci_orc and not ci:
                ci = ci_orc
            elif ci_orc and ci:
                for k, v in ci_orc.items():
                    if v and not ci.get(k):
                        ci[k] = v
            if loucas_orc and not result['loucas_metais_detail']:
                result['loucas_metais_detail'] = loucas_orc
            if sist_orc and not result['sistemas_especiais_detail']:
                result['sistemas_especiais_detail'] = sist_orc

        # If Relatório but still no CI, try compact format
        if has_rel and not ci:
            ci_comp, _, sist_comp = extract_from_compact_relatorio(wb)
            if ci_comp:
                ci = ci_comp
            if sist_comp and not result['sistemas_especiais_detail']:
                result['sistemas_especiais_detail'] = sist_comp

        # Add EPC and CANTEIRO totals
        if ci:
            if has_epc and not ci.get('epcs'):
                v = extract_sheet_total(wb, ["EPC's", "EPCs", "EPC"], 100)
                if v:
                    ci['epcs'] = v
            if has_cant and not ci.get('canteiro'):
                v = extract_sheet_total(wb, ['CANTEIRO', 'Canteiro'], 100)
                if v:
                    ci['canteiro'] = v
        elif has_epc or has_cant:
            ci = {}
            if has_epc:
                v = extract_sheet_total(wb, ["EPC's", "EPCs", "EPC"], 100)
                if v:
                    ci['epcs'] = v
            if has_cant:
                v = extract_sheet_total(wb, ['CANTEIRO', 'Canteiro'], 100)
                if v:
                    ci['canteiro'] = v
            if not ci:
                ci = None

        result['ci_detalhado'] = ci

        # ── Louças ────────────────────────────────────────────
        if has_loucas and not result.get('loucas_metais_detail'):
            loucas = extract_loucas_metais(wb)
            if loucas:
                result['loucas_metais_detail'] = loucas

        # ── Sistemas ──────────────────────────────────────────
        if has_sist and not result.get('sistemas_especiais_detail'):
            sist = extract_sistemas_especiais(wb)
            if sist:
                result['sistemas_especiais_detail'] = sist

        # Try Resumo sheet for sistemas (direct subtotals)
        ws_resumo = find_sheet(wb, ['Resumo', 'Orçamento Resumo'])
        if ws_resumo and not result.get('sistemas_especiais_detail'):
            resumo_rows = read_all_rows(ws_resumo, max_rows=50)
            sist_resumo = {}
            for row in resumo_rows:
                if not row:
                    continue
                desc = safe_str(row[0])
                desc_norm = normalize(desc)
                val = None
                for ci_idx in range(1, min(len(row), 10)):
                    v = safe_num(row[ci_idx])
                    if v and is_valid_value(v) and v > 1000:
                        val = v
                        break
                if not val:
                    continue
                if 'elevador' in desc_norm and 'porta' not in desc_norm:
                    sist_resumo['elevadores'] = {'valor': val}
                elif 'gerador' in desc_norm:
                    sist_resumo['gerador'] = val
                elif 'piscina' in desc_norm:
                    sist_resumo['piscina'] = val
                elif any(k in desc_norm for k in ['equipamentos e sistemas especiais', 'sistemas especiais']):
                    sist_resumo['total'] = val
            if sist_resumo:
                result['sistemas_especiais_detail'] = sist_resumo

        # Piscina from dedicated sheet
        piscina_val = extract_piscina_total(wb)
        if piscina_val:
            if result['sistemas_especiais_detail'] is None:
                result['sistemas_especiais_detail'] = {'piscina': piscina_val, 'total': piscina_val}
            elif not result['sistemas_especiais_detail'].get('piscina'):
                result['sistemas_especiais_detail']['piscina'] = piscina_val

        wb.close()

    except Exception as e:
        print(f"  [{idx}] {slug}: ERROR - {e}")
        traceback.print_exc()

    return result

# ============================================================
# Main
# ============================================================
def main():
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        all_projects = json.load(f)

    batch = all_projects[BATCH_START:BATCH_END]

    stats = {
        'total': len(batch),
        'ci_extracted': 0,
        'loucas_extracted': 0,
        'sistemas_extracted': 0,
        'errors': 0,
    }

    for i, project in enumerate(batch):
        idx = BATCH_START + i
        slug = project['slug']
        print(f"\n[{idx}] Processing: {slug}")

        try:
            extracted = process_project(project, idx)

            json_path = os.path.join(INDICES_DIR, f"{slug}.json")
            if not os.path.exists(json_path):
                print(f"  [{idx}] {slug}: JSON not found, skipping")
                continue

            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            updated = False

            if extracted['ci_detalhado']:
                data['ci_detalhado'] = extracted['ci_detalhado']
                updated = True
                stats['ci_extracted'] += 1
                ci_str = json.dumps(extracted['ci_detalhado'], ensure_ascii=False)
                print(f"  CI: {ci_str[:150]}")

            if extracted['loucas_metais_detail']:
                data['loucas_metais_detail'] = extracted['loucas_metais_detail']
                updated = True
                stats['loucas_extracted'] += 1
                l_str = json.dumps(extracted['loucas_metais_detail'], ensure_ascii=False)
                print(f"  Loucas: {l_str[:150]}")

            if extracted['sistemas_especiais_detail']:
                data['sistemas_especiais_detail'] = extracted['sistemas_especiais_detail']
                updated = True
                stats['sistemas_extracted'] += 1
                s_str = json.dumps(extracted['sistemas_especiais_detail'], ensure_ascii=False)
                print(f"  Sistemas: {s_str[:150]}")

            if updated:
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                print(f"  -> Updated")
            else:
                print(f"  -> No new data extracted")

        except Exception as e:
            print(f"  [{idx}] {slug}: FATAL ERROR - {e}")
            traceback.print_exc()
            stats['errors'] += 1

    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total projects: {stats['total']}")
    print(f"CI extracted: {stats['ci_extracted']}")
    print(f"Louças extracted: {stats['loucas_extracted']}")
    print(f"Sistemas extracted: {stats['sistemas_extracted']}")
    print(f"Errors: {stats['errors']}")

if __name__ == '__main__':
    main()
