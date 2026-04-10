#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract esquadrias, fachada, and acabamentos detail from construction budget spreadsheets.
Batch: projects 42-83 (0-based indices 42..83).

Extracts:
  - esquadrias_detail: aluminio, guarda_corpo, portas_madeira, portas_corta_fogo, vidros, pele_vidro, brises
  - fachada_detail: reboco_externo, textura_pintura_ext, pastilha_porcelanato, juntas
  - acabamentos_pus: porcelanato, ceramica, laminado, rodape, contrapiso, forro_gesso, pintura, chapisco, reboco

Updates each indices-executivo/{slug}.json adding these fields.
"""
import json, sys, re, os, traceback
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

BATCH = ALL_PROJECTS[42:84]  # projects 42-83 inclusive


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def normalize(text):
    """Lowercase + remove accents for matching."""
    t = text.lower().strip()
    replacements = {
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'è': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'õ': 'o', 'ô': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c', 'ñ': 'n',
        '\xa0': ' ',
    }
    for old, new in replacements.items():
        t = t.replace(old, new)
    return t


# ============================================================
# Classification helpers
# ============================================================

def classify_esquadria(desc):
    """Classify a description into esquadrias sub-category."""
    d = normalize(desc)

    # Skip non-esquadria items that mention esquadria keywords incidentally
    if any(x in d for x in ['ensaio', 'laudo', 'projeto de esquadria', 'consultoria',
                              'impermeabilizacao de esquadria', 'estanqueidade']):
        return None

    # Pele de vidro / curtain wall
    if any(x in d for x in ['pele de vidro', 'curtain wall', 'muro cortina', 'fachada de vidro',
                              'fachada envidracada', 'fachada vidro']):
        return 'pele_vidro'

    # Brise
    if 'brise' in d:
        return 'brises'

    # Guarda corpo (exclude ensaio/laudo/teste)
    if ('guarda corpo' in d or 'guarda-corpo' in d or 'guardacorpo' in d):
        if 'ensaio' not in d and 'laudo' not in d and 'teste' not in d:
            return 'guarda_corpo'

    # Porta corta fogo
    if 'corta fogo' in d or 'corta-fogo' in d or 'cortafogo' in d:
        return 'portas_corta_fogo'

    # Portas de madeira
    if ('porta' in d and ('madeira' in d or 'semi oca' in d or 'semi-oca' in d or 'semioca' in d)):
        return 'portas_madeira'
    if 'porta de madeira' in d:
        return 'portas_madeira'
    if 'instalacao de portas de madeira' in d or 'instalacao de porta de madeira' in d:
        return 'portas_madeira'

    # Aluminio (esquadrias)
    if any(x in d for x in ['esquadrias de aluminio', 'esquadria de aluminio',
                              'janela de aluminio', 'janelas aluminio',
                              'esquadrias aluminio']):
        return 'aluminio'
    if 'aluminio' in d and ('esquadria' in d or 'janela' in d):
        return 'aluminio'
    if 'aluminio' in d and 'porta' in d and 'madeira' not in d and 'corta' not in d:
        return 'aluminio'
    if 'contramarco' in d and 'aluminio' in d:
        return 'aluminio'
    if 'portao' in d and 'aluminio' in d:
        return 'aluminio'
    if 'moldura' in d and 'aluminio' in d:
        return 'aluminio'

    # Generic aluminio esquadrias
    if d.startswith('esquadrias de aluminio') or d.startswith('esquadria aluminio'):
        return 'aluminio'

    # Vidros
    if any(x in d for x in ['vidro temperado', 'vidro laminado', 'vidro insulado',
                              'envidracamento', 'vidros']):
        if 'guarda' not in d:
            return 'vidros'

    return None


def classify_fachada(desc):
    """Classify a description into fachada sub-category."""
    d = normalize(desc)

    # Must be external/fachada context
    is_ext = any(x in d for x in ['externo', 'externa', 'fachada', 'exterior'])

    if 'reboco' in d and is_ext:
        return 'reboco_externo'
    if 'massa unica' in d and is_ext:
        return 'reboco_externo'
    if ('argamassado' in d or 'argamassa' in d) and 'reboco' in d and is_ext:
        return 'reboco_externo'

    if ('textura' in d or 'pintura' in d) and is_ext:
        return 'textura_pintura_ext'
    if 'grafiato' in d and is_ext:
        return 'textura_pintura_ext'

    if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d or 'ceramic' in d) and is_ext:
        return 'pastilha_porcelanato'
    if 'pecas ceramicas' in d and is_ext:
        return 'pastilha_porcelanato'
    if 'acabamento' in d and ('ceramica' in d or 'ceramic' in d) and is_ext:
        return 'pastilha_porcelanato'

    if 'junta' in d and is_ext:
        return 'juntas'
    if 'junta de dilatacao' in d:
        return 'juntas'
    if 'junta' in d and 'fachada' in d:
        return 'juntas'

    return None


def classify_acabamento(desc, unit):
    """Classify a description into acabamentos category and return category name."""
    d = normalize(desc)
    u = normalize(str(unit)) if unit else ''

    # Skip mão de obra items for PU extraction
    if 'mao de obra' in d or 'empreitada' in d:
        return None

    # Rodapé (check before porcelanato to avoid "rodapé de porcelanato" being classified as porcelanato)
    if 'rodape' in d:
        return 'rodape'

    # Porcelanato (internal walls and floors)
    if 'porcelanato' in d and u == 'm2':
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'porcelanato'
    if 'porcelanato' in d and ('piso' in d or 'revestimento' in d or 'parede' in d):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if u in ('m2', 'm²', ''):
                return 'porcelanato'

    # Ceramica
    if ('ceramica' in d or 'ceramico' in d) and ('revestimento' in d or 'piso' in d or u == 'm2'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'pastilha' not in d:
                return 'ceramica'

    # Laminado
    if 'laminado' in d and ('piso' in d or 'revestimento' in d or u == 'm2'):
        return 'laminado'

    # Contrapiso
    if 'contrapiso' in d and (u == 'm2' or u == 'm²'):
        return 'contrapiso'

    # Forro de gesso
    if 'forro' in d and ('gesso' in d or 'mineral' in d):
        if u in ('m2', 'm²', 'vb', ''):
            return 'forro_gesso'

    # Pintura interna
    if ('pintura' in d or 'tinta' in d) and (u == 'm2' or u == 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'piso' not in d:
                return 'pintura'
    # Textura interna (also pintura category)
    if 'textura' in d and (u == 'm2' or u == 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'pintura'

    # Chapisco interno
    if 'chapisco' in d and (u == 'm2' or u == 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'chapisco'

    # Reboco interno
    if ('reboco' in d or 'massa unica' in d) and (u == 'm2' or u == 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'reboco'

    return None


# ============================================================
# Ger_Executivo extractor
# ============================================================

def extract_from_ger_executivo(wb, sheet_name):
    """Extract esquadrias/fachada/acabamentos from Ger_Executivo format.
    Columns: H=level_type, I=code, J=desc, K=unit, L=qty, M=pu, N=total, B=celula
    """
    ws = wb[sheet_name]
    esquadrias = {}  # {category: {'valor': float, 'qtd': float, 'pu': float, 'items': [...]}}
    fachada = {}     # {category: {'valor': float}}
    acabamentos = {} # {category: [{'pu': float, 'qty': float, 'desc': str}]}

    for row in ws.iter_rows(min_row=1, max_row=8000, max_col=20, values_only=True):
        if len(row) < 14:
            continue
        level_type = safe_str(row[7] if row[7] else '')   # H (index 7)
        desc = safe_str(row[9] if row[9] else '')          # J (index 9)
        unit = safe_str(row[10] if row[10] else '')        # K (index 10)
        qty = safe_float(row[11])                          # L (index 11)
        pu = safe_float(row[12])                           # M (index 12)
        total = safe_float(row[13])                        # N (index 13)
        code = safe_str(row[8] if row[8] else '')          # I (index 8)
        celula = row[1]                                    # B (index 1)

        if not desc:
            continue

        # --- Esquadrias ---
        # Handle both SERVIÇO (item-level) and SUBETAPA/ETAPA (subtotal-level)
        if level_type == 'SERVIÇO' and desc:
            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                esquadrias[cat]['valor'] += total
                esquadrias[cat]['qtd'] += qty
                if pu > 0:
                    esquadrias[cat]['pu_sum'] += pu * qty  # weighted
                    esquadrias[cat]['pu_count'] += qty
                esquadrias[cat]['items'].append({'desc': desc, 'qty': qty, 'pu': pu, 'total': total})

        # Also check SUBETAPA for esquadrias sections we might miss at SERVIÇO level
        if level_type == 'SUBETAPA' and total > 0:
            d = normalize(desc)
            # Check for whole section values (use only if no SERVICO items found)
            if 'esquadrias de aluminio' in d and 'aluminio' not in esquadrias:
                esquadrias.setdefault('aluminio_subetapa', {'valor': 0.0})
                esquadrias['aluminio_subetapa']['valor'] += total
            if 'guarda corpo' in d and 'guarda_corpo' not in esquadrias:
                esquadrias.setdefault('guarda_corpo_subetapa', {'valor': 0.0})
                esquadrias['guarda_corpo_subetapa']['valor'] += total
            if 'brise' in d and 'brises' not in esquadrias:
                esquadrias.setdefault('brises_subetapa', {'valor': 0.0})
                esquadrias['brises_subetapa']['valor'] += total

        # --- Fachada ---
        if level_type == 'SERVIÇO' and desc:
            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += total

        # Also catch SUBETAPA-level fachada sections
        if level_type == 'SUBETAPA' and total > 0:
            d = normalize(desc)
            if 'revestimento argamassado' in d and ('fachada' in d or 'externo' in d):
                if 'reboco_externo' not in fachada:
                    fachada.setdefault('reboco_externo_sub', {'valor': 0.0})
                    fachada['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and 'fachada' in d:
                if 'textura_pintura_ext' not in fachada:
                    fachada.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                    fachada['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('fachada' in d or 'externo' in d):
                if 'pastilha_porcelanato' not in fachada:
                    fachada.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                    fachada['pastilha_porcelanato_sub']['valor'] += total

        # Check CÉLULA CONSTRUTIVA or ETAPA for fachada section total
        if level_type in ('CÉLULA CONSTRUTIVA', 'ETAPA') and total > 0:
            d = normalize(desc)
            if 'revestimentos' in d and 'acabamentos' in d and 'fachada' in d:
                fachada.setdefault('_total_fachada', total)

        # --- Acabamentos PUs ---
        if level_type == 'SERVIÇO' and desc and pu > 0 and qty > 0:
            cat = classify_acabamento(desc, unit)
            if cat:
                if cat not in acabamentos:
                    acabamentos[cat] = []
                acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

    return esquadrias, fachada, acabamentos


# ============================================================
# Relatório N2 extractor
# ============================================================

def extract_from_relatorio(wb, sheet_name):
    """Extract from Sienge Relatório N2 format.
    Columns: A=code, B=desc, N=unit, Q=qty, W=pu, AC=total
    Auto-detects simplified format (A=code, B=desc, C=total) and delegates.
    """
    ws = wb[sheet_name]

    # Auto-detect format by checking header row
    # Auto-detect format: Simple (A=code, B=desc, C=total) vs Full N2
    is_simple = False
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=30, values_only=True):
        if row and row[0] and str(row[0]).strip().lower() in ('código', 'codigo'):
            # Full N2 has 'Preço unitário' at col W (index 22) and 'Preço total' at col AC (index 28)
            has_pu_col = len(row) > 22 and row[22] and 'unit' in str(row[22]).lower()
            has_total_ac = len(row) > 28 and row[28] and 'total' in str(row[28]).lower()
            # Simple format has 'Preço total' at col C (index 2)
            has_total_c = len(row) > 2 and row[2] and 'total' in str(row[2]).lower()
            if has_total_c and not has_pu_col:
                is_simple = True
            elif not has_pu_col and not has_total_ac:
                is_simple = True
            break

    if is_simple:
        return extract_from_relatorio_simple(wb, sheet_name)

    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=8000, max_col=30, values_only=True):
        if len(row) < 23:
            continue
        code = safe_str(row[0] if row[0] else '')    # A (index 0)
        desc = safe_str(row[1] if row[1] else '')     # B (index 1)
        unit = safe_str(row[13] if row[13] else '')   # N (index 13)
        qty = safe_float(row[16])                     # Q (index 16)
        pu = safe_float(row[22])                      # W (index 22)
        total = safe_float(row[28] if len(row) > 28 else 0) # AC (index 28)

        if not desc:
            continue

        # Detect if this is a service-level item (4-part code like XX.XXX.XXX.XXX)
        code_clean = code.strip()
        parts = [p for p in code_clean.split('.') if p.strip()]
        is_service = len(parts) >= 4
        is_subetapa = len(parts) == 3
        is_etapa = len(parts) <= 2

        if is_service:
            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                item_total = total if total > 0 else (pu * qty if pu > 0 and qty > 0 else 0)
                esquadrias[cat]['valor'] += item_total
                esquadrias[cat]['qtd'] += qty
                if pu > 0:
                    esquadrias[cat]['pu_sum'] += pu * qty
                    esquadrias[cat]['pu_count'] += qty
                esquadrias[cat]['items'].append({'desc': desc, 'qty': qty, 'pu': pu, 'total': item_total})

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                item_total = total if total > 0 else (pu * qty if pu > 0 and qty > 0 else 0)
                fachada[cat]['valor'] += item_total

            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

    return esquadrias, fachada, acabamentos


def extract_from_relatorio_simple(wb, sheet_name):
    """Extract from simplified Relatório format (A=code, B=desc, C=total).
    No unit price or quantity - just totals at section level.
    """
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=8000, max_col=10, values_only=True):
        if len(row) < 3:
            continue
        code = safe_str(row[0] if row[0] else '')
        desc = safe_str(row[1] if row[1] else '')
        total = safe_float(row[2])

        if not desc or total <= 0:
            continue

        d = normalize(desc)
        code_clean = code.strip()
        parts = [p for p in code_clean.split('.') if p.strip()]

        # Etapa level (2 parts like "02.015")
        if len(parts) == 2:
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d:
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

        # Sub-etapa level (3 parts)
        if len(parts) == 3:
            if 'esquadria' in d and 'alumin' in d:
                esquadrias.setdefault('aluminio_subetapa', {'valor': 0.0})
                esquadrias['aluminio_subetapa']['valor'] += total
            elif 'guarda corpo' in d:
                esquadrias.setdefault('guarda_corpo_subetapa', {'valor': 0.0})
                esquadrias['guarda_corpo_subetapa']['valor'] += total
            elif 'brise' in d:
                esquadrias.setdefault('brises_subetapa', {'valor': 0.0})
                esquadrias['brises_subetapa']['valor'] += total
            elif 'esquadria' in d and 'madeira' in d:
                esquadrias.setdefault('portas_madeira', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_madeira']['valor'] += total
            elif ('corta fogo' in d or 'corta-fogo' in d) or ('metalica' in d and 'esquadria' in d):
                esquadrias.setdefault('portas_corta_fogo', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_corta_fogo']['valor'] += total

            # Fachada sub-sections
            if 'reboco' in d and ('externo' in d or 'fachada' in d):
                fachada.setdefault('reboco_externo', {'valor': 0.0})
                fachada['reboco_externo']['valor'] += total
            if ('textura' in d or 'pintura' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('textura_pintura_ext', {'valor': 0.0})
                fachada['textura_pintura_ext']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('pastilha_porcelanato', {'valor': 0.0})
                fachada['pastilha_porcelanato']['valor'] += total

    return esquadrias, fachada, acabamentos


# ============================================================
# ORÇAMENTO sheet extractor - auto-detects column layout
# ============================================================

def extract_from_orcamento_sheet(wb, sheet_name):
    """Extract from ORÇAMENTO/Executivo sheets. Auto-detects format:
    1) Ger_Executivo column layout (H=level_type, I=code, J=desc, K=unit, L=qty, M=pu, N=total)
    2) Flat ORÇAMENTO_EXECUTIVO (A=code, B=desc, C=un, D=qty, E=pu, F=total)
    3) Summary format (A=etapa, B=valor_orcado)
    """
    ws = wb[sheet_name]

    # Detect format by checking first rows
    format_type = 'unknown'
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=20, values_only=True):
        if not row:
            continue
        # Check for Ger_Executivo format: column H has level types
        if len(row) > 7 and row[7] and str(row[7]).strip() in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
            format_type = 'ger_executivo'
            break
        # Check for flat format: header row with 'Código', 'Descrição', 'Un.', etc
        if row[0] and str(row[0]).strip().lower() in ('código', 'codigo'):
            if len(row) > 2 and row[2]:
                r2 = str(row[2]).strip().lower()
                if r2 in ('un.', 'un', 'und', 'unid', 'unidade'):
                    format_type = 'flat'
                    break
                if 'total' in r2:
                    format_type = 'flat_simple'
                    break
        # Check for summary format: 'ETAPA' header (may be in col A or col B)
        for ci in range(min(3, len(row))):
            if row[ci] and str(row[ci]).strip() == 'ETAPA':
                next_ci = ci + 1
                if next_ci < len(row) and row[next_ci] and 'VALOR' in str(row[next_ci]).upper():
                    format_type = 'summary'
                    break
        if format_type != 'unknown':
            break

    if format_type == 'ger_executivo':
        return extract_from_ger_executivo(wb, sheet_name)
    elif format_type == 'flat':
        return extract_from_flat_orcamento(wb, sheet_name)
    elif format_type == 'flat_simple':
        return extract_from_relatorio_simple(wb, sheet_name)
    elif format_type == 'summary':
        return extract_from_summary(wb, sheet_name)
    else:
        # Try Ger_Executivo as default
        return extract_from_ger_executivo(wb, sheet_name)


def extract_from_flat_orcamento(wb, sheet_name):
    """Extract from flat ORÇAMENTO_EXECUTIVO format.
    Columns: A=code, B=desc, C=un, D=qty, E=pu, F=total, G=comments
    """
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=8000, max_col=10, values_only=True):
        if len(row) < 6:
            continue
        code = safe_str(row[0] if row[0] else '')
        desc = safe_str(row[1] if row[1] else '')
        unit = safe_str(row[2] if row[2] else '')
        qty = safe_float(row[3])
        pu = safe_float(row[4])
        total = safe_float(row[5])

        if not desc:
            continue

        # Detect service-level items by code structure
        code_clean = code.strip()
        parts = [p for p in code_clean.split('.') if p.strip()]
        is_service = len(parts) >= 4

        if is_service and total > 0:
            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                esquadrias[cat]['valor'] += total
                esquadrias[cat]['qtd'] += qty
                if pu > 0 and qty > 0:
                    esquadrias[cat]['pu_sum'] += pu * qty
                    esquadrias[cat]['pu_count'] += qty
                esquadrias[cat]['items'].append({'desc': desc, 'qty': qty, 'pu': pu, 'total': total})

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += total

            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

        # Also get section totals from 2-part codes
        if len(parts) == 1 and total > 0:
            d = normalize(desc)
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d:
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    return esquadrias, fachada, acabamentos


# ============================================================
# Summary-only extractor (ORÇAMENTO_EXECUTIVO with just macro rows)
# ============================================================

def extract_from_summary(wb, sheet_name):
    """Extract from summary-only sheets (just etapa names + values, no detail).
    Handles data in col A+B or col B+C (some sheets have empty col A)."""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=10, values_only=True):
        if len(row) < 2:
            continue
        # Try col A (index 0) as desc, col B (index 1) as total
        desc = safe_str(row[0] if row[0] else '')
        total = safe_float(row[1])
        # If A is empty or B is not numeric, try B as desc, C as total
        if (not desc or total <= 0) and len(row) > 2:
            desc = safe_str(row[1] if row[1] else '')
            total = safe_float(row[2])

        if not desc or total <= 0:
            continue

        d = normalize(desc)
        if 'esquadria' in d:
            esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
        if 'fachada' in d:
            fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    return esquadrias, fachada, acabamentos


# ============================================================
# Gerenciamento executivo (alternative sheet name)
# ============================================================

def extract_from_gerenciamento(wb, sheet_name):
    """Extract from 'Gerenciamento executivo' or 'Executivo' sheet.
    May have different column layout - need to detect."""
    ws = wb[sheet_name]

    # Check if it follows Ger_Executivo layout (col H = level type)
    sample_h = None
    for row in ws.iter_rows(min_row=1, max_row=50, max_col=15, values_only=True):
        if len(row) < 8:
            continue
        h_val = safe_str(row[7] if row[7] else '')
        if h_val in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
            sample_h = h_val
            break

    if sample_h:
        return extract_from_ger_executivo(wb, sheet_name)

    # Check for flat format (A=code, B=desc, C=total or A=code, B=desc, C=un, D=qty, E=pu, F=total)
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=10, values_only=True):
        if row and row[0]:
            r0 = str(row[0]).strip().lower()
            if r0 in ('atalho', 'código', 'codigo', 'item'):
                # Flat format with header
                return extract_from_flat_orcamento(wb, sheet_name)

    # Check for alternate gerenciamento format: A=code ("01 "), B=desc, C=total
    esquadrias = {}
    fachada = {}
    acabamentos = {}
    for row in ws.iter_rows(min_row=1, max_row=200, max_col=10, values_only=True):
        if len(row) < 3:
            continue
        code_raw = safe_str(row[0] if row[0] else '')
        desc = safe_str(row[1] if row[1] else '')
        total = safe_float(row[2])
        if not desc or total <= 0:
            continue
        d = normalize(desc)
        if 'esquadria' in d:
            esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
        if 'fachada' in d:
            fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    if esquadrias or fachada:
        return esquadrias, fachada, acabamentos

    # Fallback: summary format (A=desc, B=total)
    return extract_from_summary(wb, sheet_name)


# ============================================================
# Orçamento Resumo extractor
# ============================================================

def extract_from_orcamento_resumo(wb, sheet_name):
    """Extract from 'Orçamento Resumo' sheets (summary format)."""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True):
        if len(row) < 2:
            continue
        # Try various column positions for desc and total (0-indexed)
        for desc_idx in [0, 1, 2]:
            desc = safe_str(row[desc_idx] if desc_idx < len(row) and row[desc_idx] else '')
            if desc and len(desc) > 5:
                d = normalize(desc)
                # Find numeric column for total
                for total_idx in [1, 2, 3, 4, 5]:
                    if total_idx == desc_idx:
                        continue
                    total = safe_float(row[total_idx] if total_idx < len(row) else 0)
                    if total > 1000:  # reasonable minimum for a discipline total
                        if 'esquadria' in d:
                            esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
                        if 'fachada' in d:
                            fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total
                        break
                break

    return esquadrias, fachada, acabamentos


# ============================================================
# Build final output
# ============================================================

def build_esquadrias_detail(raw_esquadrias):
    """Convert raw extracted data into final esquadrias_detail structure."""
    result = {}
    categories = ['aluminio', 'guarda_corpo', 'portas_madeira', 'portas_corta_fogo',
                   'vidros', 'pele_vidro', 'brises']

    for cat in categories:
        if cat in raw_esquadrias:
            data = raw_esquadrias[cat]
            entry = {'valor': round(data['valor'], 2) if data['valor'] else None}
            if data.get('qtd', 0) > 0:
                entry['qtd'] = round(data['qtd'], 2)
            if data.get('pu_count', 0) > 0:
                entry['pu'] = round(data['pu_sum'] / data['pu_count'], 2)
            result[cat] = entry
        # Check subetapa fallbacks
        elif f'{cat}_subetapa' in raw_esquadrias:
            data = raw_esquadrias[f'{cat}_subetapa']
            result[cat] = {'valor': round(data['valor'], 2)}

    # Add total if available
    if '_total_esquadrias' in raw_esquadrias:
        result['_total'] = round(raw_esquadrias['_total_esquadrias'], 2)

    return result if result else None


def build_fachada_detail(raw_fachada):
    """Convert raw extracted data into final fachada_detail structure."""
    result = {}
    categories = ['reboco_externo', 'textura_pintura_ext', 'pastilha_porcelanato', 'juntas']

    for cat in categories:
        if cat in raw_fachada:
            result[cat] = {'valor': round(raw_fachada[cat]['valor'], 2)}
        elif f'{cat}_sub' in raw_fachada:
            result[cat] = {'valor': round(raw_fachada[f'{cat}_sub']['valor'], 2)}

    if '_total_fachada' in raw_fachada:
        result['_total'] = round(raw_fachada['_total_fachada'], 2)

    return result if result else None


def build_acabamentos_pus(raw_acabamentos):
    """Convert raw extracted data into final acabamentos_pus structure (weighted avg PU)."""
    result = {}
    # Max reasonable PU for each category (sanity check)
    pu_max = {
        'porcelanato': 800,    # R$/m2
        'ceramica': 400,       # R$/m2
        'laminado': 400,       # R$/m2
        'rodape': 200,         # R$/m
        'contrapiso': 200,     # R$/m2
        'forro_gesso': 300,    # R$/m2
        'pintura': 100,        # R$/m2
        'chapisco': 50,        # R$/m2
        'reboco': 100,         # R$/m2
    }

    for cat, items in raw_acabamentos.items():
        if not items:
            continue
        # Filter out items with unreasonable PUs (e.g., vb items with total as PU)
        max_pu = pu_max.get(cat, 500)
        valid_items = [i for i in items if i['pu'] <= max_pu and i['qty'] > 0
                       and i.get('unit', '').lower().strip() not in ('vb', 'un', 'cj', 'conjunto')]
        if not valid_items:
            # Fallback: try all items but filter extreme outliers
            valid_items = [i for i in items if i['pu'] <= max_pu * 5 and i['qty'] > 0]
        if not valid_items:
            continue
        total_weighted = sum(i['pu'] * i['qty'] for i in valid_items)
        total_qty = sum(i['qty'] for i in valid_items)
        if total_qty > 0:
            avg_pu = total_weighted / total_qty
            if avg_pu <= max_pu * 2:  # final sanity check
                result[cat] = round(avg_pu, 2)

    return result if result else None


# ============================================================
# Process one project
# ============================================================

def process_project(proj):
    slug = proj['slug']
    path = proj['path']

    if not os.path.exists(path):
        return slug, None, None, None, f"File not found: {path}"

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return slug, None, None, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    all_esq = {}
    all_fach = {}
    all_acab = {}

    try:
        extracted = False

        # Strategy 1: Ger_Executivo (most complete, has all detail)
        for sname in sheets:
            if sname.lower().replace(' ', '_').startswith('ger_executivo'):
                try:
                    e, f, a = extract_from_orcamento_sheet(wb, sname)  # auto-detects format
                    if e or f or a:
                        merge_results(all_esq, e)
                        merge_results_fachada(all_fach, f)
                        merge_results_acabamentos(all_acab, a)
                        extracted = True
                    else:
                        # Try alternate format: gerenciamento handler
                        e, f, a = extract_from_gerenciamento(wb, sname)
                        if e or f or a:
                            merge_results(all_esq, e)
                            merge_results_fachada(all_fach, f)
                            merge_results_acabamentos(all_acab, a)
                            extracted = True
                except Exception as ex:
                    pass
                break

        # Strategy 2: ORÇAMENTO sheet (same column layout as Ger_Executivo)
        if not extracted:
            for sname in sheets:
                sn = sname.strip()
                if sn in ('ORÇAMENTO', 'Orçamento', 'ORCAMENTO'):
                    try:
                        e, f, a = extract_from_orcamento_sheet(wb, sname)
                        merge_results(all_esq, e)
                        merge_results_fachada(all_fach, f)
                        merge_results_acabamentos(all_acab, a)
                        extracted = True
                    except Exception as ex:
                        pass
                    break

        # Strategy 3: Relatório (Sienge format)
        if not extracted:
            for sname in sheets:
                sn_lower = sname.lower().strip()
                if 'relat' in sn_lower:
                    try:
                        e, f, a = extract_from_relatorio(wb, sname)
                        if e or f or a:
                            merge_results(all_esq, e)
                            merge_results_fachada(all_fach, f)
                            merge_results_acabamentos(all_acab, a)
                            extracted = True
                    except Exception as ex:
                        pass
                    break

        # Strategy 4: ORÇAMENTO_EXECUTIVO sheets
        if not extracted:
            for sname in sheets:
                sn_lower = sname.lower().strip()
                if ('orcamento' in sn_lower or 'orçamento' in sn_lower) and 'executiv' in sn_lower:
                    try:
                        e, f, a = extract_from_orcamento_sheet(wb, sname)
                        if e or f or a:
                            merge_results(all_esq, e)
                            merge_results_fachada(all_fach, f)
                            merge_results_acabamentos(all_acab, a)
                            extracted = True
                    except:
                        pass
                    # Try summary format too
                    if not extracted:
                        try:
                            e, f, a = extract_from_summary(wb, sname)
                            merge_results(all_esq, e)
                            merge_results_fachada(all_fach, f)
                            extracted = True
                        except:
                            pass
                    break

        # Strategy 5: Gerenciamento executivo / Executivo sheets
        if not extracted:
            for sname in sheets:
                sn_lower = sname.lower().strip()
                if 'gerenciamento executivo' in sn_lower or sn_lower == 'executivo':
                    try:
                        e, f, a = extract_from_gerenciamento(wb, sname)
                        merge_results(all_esq, e)
                        merge_results_fachada(all_fach, f)
                        merge_results_acabamentos(all_acab, a)
                        extracted = True
                    except Exception as ex:
                        pass
                    break

        # Strategy 6: Gerenciamento_Exec
        if not extracted:
            for sname in sheets:
                sn_lower = sname.lower().strip()
                if 'gerenciamento_exec' in sn_lower:
                    try:
                        e, f, a = extract_from_gerenciamento(wb, sname)
                        merge_results(all_esq, e)
                        merge_results_fachada(all_fach, f)
                        merge_results_acabamentos(all_acab, a)
                        extracted = True
                    except:
                        pass
                    break

        # Strategy 7: Orçamento Resumo
        if not extracted:
            for sname in sheets:
                sn_lower = sname.lower().strip()
                if 'orcamento resumo' in sn_lower or 'orçamento resumo' in sn_lower:
                    try:
                        e, f, a = extract_from_orcamento_resumo(wb, sname)
                        merge_results(all_esq, e)
                        merge_results_fachada(all_fach, f)
                        extracted = True
                    except:
                        pass
                    break

        # Strategy 8: 'Obra' sheet (summary format with area quadro)
        if not extracted:
            for sname in sheets:
                if sname.strip().lower() in ('obra', 'obra '):
                    try:
                        e, f, a = extract_from_summary(wb, sname)
                        if e or f:
                            merge_results(all_esq, e)
                            merge_results_fachada(all_fach, f)
                            extracted = True
                    except:
                        pass
                    break

        # Strategy 9: First sheet fallback
        if not extracted and sheets:
            try:
                # Try auto-detecting format via extract_from_orcamento_sheet
                e, f, a = extract_from_orcamento_sheet(wb, sheets[0])
                if e or f or a:
                    merge_results(all_esq, e)
                    merge_results_fachada(all_fach, f)
                    merge_results_acabamentos(all_acab, a)
                    extracted = True
            except:
                pass
            if not extracted:
                try:
                    e, f, a = extract_from_summary(wb, sheets[0])
                    merge_results(all_esq, e)
                    merge_results_fachada(all_fach, f)
                    extracted = True
                except:
                    pass

    except Exception as ex:
        wb.close()
        return slug, None, None, None, f"Error: {traceback.format_exc()}"

    wb.close()

    esq_detail = build_esquadrias_detail(all_esq)
    fach_detail = build_fachada_detail(all_fach)
    acab_pus = build_acabamentos_pus(all_acab)

    return slug, esq_detail, fach_detail, acab_pus, None


def merge_results(target, source):
    """Merge esquadrias results."""
    for k, v in source.items():
        if k not in target:
            target[k] = v
        elif isinstance(v, dict) and isinstance(target[k], dict):
            if 'valor' in v and 'valor' in target[k]:
                target[k]['valor'] = target[k].get('valor', 0) + v.get('valor', 0)
                if 'qtd' in v:
                    target[k]['qtd'] = target[k].get('qtd', 0) + v.get('qtd', 0)
                if 'pu_sum' in v:
                    target[k]['pu_sum'] = target[k].get('pu_sum', 0) + v.get('pu_sum', 0)
                    target[k]['pu_count'] = target[k].get('pu_count', 0) + v.get('pu_count', 0)
                if 'items' in v:
                    target[k].setdefault('items', []).extend(v.get('items', []))


def merge_results_fachada(target, source):
    """Merge fachada results."""
    for k, v in source.items():
        if k not in target:
            target[k] = v
        elif isinstance(v, dict) and isinstance(target[k], dict):
            if 'valor' in v:
                target[k]['valor'] = target[k].get('valor', 0) + v.get('valor', 0)
        elif isinstance(v, (int, float)) and isinstance(target[k], (int, float)):
            target[k] = max(target[k], v)  # for _total, keep larger


def merge_results_acabamentos(target, source):
    """Merge acabamentos results."""
    for k, v in source.items():
        if k not in target:
            target[k] = v
        else:
            target[k].extend(v)


# ============================================================
# Main
# ============================================================

def main():
    results = {'success': 0, 'skipped': 0, 'error': 0, 'no_data': 0}
    details = []

    for proj in BATCH:
        slug = proj['slug']
        print(f"Processing {slug}...", end=' ')

        slug_r, esq, fach, acab, err = process_project(proj)

        if err:
            print(f"ERROR: {err}")
            results['error'] += 1
            details.append({'slug': slug, 'status': 'error', 'error': err})
            continue

        if not esq and not fach and not acab:
            print("NO DATA found")
            results['no_data'] += 1
            details.append({'slug': slug, 'status': 'no_data'})
            # Still update JSON to mark as processed
            update_json(slug, esq, fach, acab)
            continue

        update_json(slug, esq, fach, acab)
        results['success'] += 1

        summary_parts = []
        if esq:
            summary_parts.append(f"esq={len(esq)} cats")
        if fach:
            summary_parts.append(f"fach={len(fach)} cats")
        if acab:
            summary_parts.append(f"acab={len(acab)} cats")
        print(f"OK ({', '.join(summary_parts)})")
        details.append({'slug': slug, 'status': 'ok',
                        'esquadrias': list(esq.keys()) if esq else [],
                        'fachada': list(fach.keys()) if fach else [],
                        'acabamentos': list(acab.keys()) if acab else []})

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total: {len(BATCH)}")
    print(f"Success (with data): {results['success']}")
    print(f"No data found: {results['no_data']}")
    print(f"Errors: {results['error']}")

    # Detailed breakdown
    print("\n--- Esquadrias coverage ---")
    esq_cats = {}
    for d in details:
        for cat in d.get('esquadrias', []):
            esq_cats[cat] = esq_cats.get(cat, 0) + 1
    for cat, count in sorted(esq_cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count} projects")

    print("\n--- Fachada coverage ---")
    fach_cats = {}
    for d in details:
        for cat in d.get('fachada', []):
            fach_cats[cat] = fach_cats.get(cat, 0) + 1
    for cat, count in sorted(fach_cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count} projects")

    print("\n--- Acabamentos PU coverage ---")
    acab_cats = {}
    for d in details:
        for cat in d.get('acabamentos', []):
            acab_cats[cat] = acab_cats.get(cat, 0) + 1
    for cat, count in sorted(acab_cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count} projects")


def update_json(slug, esq, fach, acab):
    """Update the indices-executivo JSON file for this project."""
    idx_path = INDICES_DIR / f"{slug}.json"
    if idx_path.exists():
        with open(idx_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {'projeto': slug}

    if esq is not None:
        data['esquadrias_detail'] = esq
    if fach is not None:
        data['fachada_detail'] = fach
    if acab is not None:
        data['acabamentos_pus'] = acab

    with open(idx_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


if __name__ == '__main__':
    main()
