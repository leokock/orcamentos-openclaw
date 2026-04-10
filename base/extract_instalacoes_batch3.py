#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract installation breakdown data from construction budget spreadsheets.
Batch 3: projects 84-125 (0-indexed 83-124).

Splits "Instalações" macrogroup into 5 sub-disciplines:
  - hidrossanitarias
  - eletricas
  - preventivas
  - gas
  - telecom

Also extracts MO/Material splits where available.
"""
import json, sys, re, os
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

# Load project mapping
with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

BATCH = ALL_PROJECTS[83:125]  # projects 84-125

# ── Keyword matchers ──────────────────────────────────────────────────────
def classify_discipline(text):
    """Classify a description text into a discipline category."""
    t = text.lower().strip()

    # Telecom / communications (check first, more specific)
    if any(k in t for k in ['telecom', 'comunicaç', 'comunicac', 'interfon', 'cftv',
                            'telefon', 'lógic', 'logic', 'automação', 'automaç',
                            'rede estruturada', 'dados', 'tv vhf', 'tv/cabo',
                            'som e wifi', 'wifi']):
        return 'telecom'

    # Gas
    if any(k in t for k in ['gás', 'gas', 'glp', 'gn ']):
        # But not "gas + preventiva" combined
        if 'prevent' not in t and 'incêndio' not in t and 'incendio' not in t:
            return 'gas'

    # Preventivas (fire prevention, PPCI, SPDA)
    if any(k in t for k in ['prevent', 'ppci', 'pci', 'incêndio', 'incendio',
                            'hidrante', 'sprinkler', 'alarme', 'spda',
                            'proteção e aterramento', 'protecao e aterramento']):
        return 'preventivas'

    # Hidrossanitárias
    if any(k in t for k in ['hidro', 'sanitár', 'sanitar', 'água fria', 'agua fria',
                            'água quente', 'agua quente', 'esgoto', 'pluvial',
                            'drenagem', 'bombas', 'bomba', 'ligações prediais',
                            'ligacoes prediais', 'reuso', 'tratamento de esgoto',
                            'hidrometro', 'hidrômetro']):
        return 'hidrossanitarias'

    # Elétricas
    if any(k in t for k in ['elétr', 'eletr', 'eletrodut', 'eletrocalh',
                            'fios e cabos', 'cabos e fiaç', 'cabos e fiac',
                            'quadros elétric', 'quadros eletric', 'disjuntor',
                            'tomada', 'interruptor', 'luminár', 'luminar',
                            'lâmpada', 'lampada', 'gerador', 'energia']):
        return 'eletricas'

    return None

def is_mo_line(text):
    """Check if this is a labor (MO) line."""
    t = text.lower().strip()
    return any(k in t for k in ['mão de obra', 'mao de obra', 'mo ', 'mão-de-obra',
                                 'm.o.', 'm.o ', '- mo', '(mo)'])

def is_material_line(text):
    """Check if this is a material line."""
    t = text.lower().strip()
    return any(k in t for k in ['material', 'mat.', 'mat '])


def safe_float(v):
    """Safely convert a value to float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


# ── Format-specific extractors ───────────────────────────────────────────

def extract_from_ger_executivo(wb, sheet_name='Ger_Executivo'):
    """
    Extract from Ger_Executivo sheet format (CTN standard).
    Uses 'CÉLULA CONSTRUTIVA' rows with codes like 06, 07, 08, 09.
    """
    ws = wb[sheet_name]
    result = {}
    mo_material = {}

    current_cell = None  # current cell-constructive code (06, 07, etc.)
    current_discipline = None
    cell_total = 0
    cell_mo = 0

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=30, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        row_text = ' '.join(str(v) for v in vals)

        # Look for CÉLULA CONSTRUTIVA headers
        if 'CÉLULA CONSTRUTIVA' in row_text or 'CELULA CONSTRUTIVA' in row_text:
            # Save previous cell data
            if current_discipline and current_cell:
                result[current_discipline] = cell_total
                if cell_mo > 0:
                    mo_material[current_discipline] = {
                        'mo_valor': round(cell_mo, 2),
                        'material_valor': round(cell_total - cell_mo, 2),
                        'mo_pct': round(cell_mo / cell_total * 100, 1) if cell_total > 0 else 0
                    }

            # Parse new cell
            for v in vals:
                if isinstance(v, str):
                    disc = classify_discipline(v)
                    if disc:
                        current_discipline = disc
                        break
            else:
                current_discipline = None

            # Find the total value (usually a large number in the row)
            cell_total = 0
            for v in vals:
                if isinstance(v, (int, float)) and v > 1000:
                    cell_total = float(v)
                    break
            cell_mo = 0
            current_cell = True
            continue

        # Track MO within current cell
        if current_discipline and current_cell:
            if any(isinstance(v, str) and is_mo_line(v) for v in vals):
                for v in vals:
                    if isinstance(v, (int, float)) and v > 100:
                        cell_mo += float(v)
                        break

    # Save last cell
    if current_discipline and current_cell:
        result[current_discipline] = cell_total
        if cell_mo > 0:
            mo_material[current_discipline] = {
                'mo_valor': round(cell_mo, 2),
                'material_valor': round(cell_total - cell_mo, 2),
                'mo_pct': round(cell_mo / cell_total * 100, 1) if cell_total > 0 else 0
            }

    return result, mo_material


def extract_from_relatorio_sienge(wb):
    """
    Extract from Sienge 'Relatório' format.
    Uses hierarchical codes like 05.001, 06.001, 07.001 etc.
    Total values in column index 28 (AC column).
    """
    ws = wb['Relatório']
    result = {}
    mo_material = {}

    # First pass: identify N1-level codes and their disciplines
    n1_codes = {}  # code -> discipline

    # Collect all rows with data
    rows_data = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3000, max_col=30, values_only=True), 1):
        rows_data.append(row)

    for row in rows_data:
        if row[0] is None or not isinstance(row[0], str):
            continue
        code = row[0].strip()
        desc = row[1] if len([v for v in row if v is not None]) > 1 and isinstance(row[1], str) else ''

        # N1 level: XX. or XX (2 digits)
        if re.match(r'^\d{2}\s*$', code) or re.match(r'^\d{2}\.\d{3}\s*$', code):
            disc = classify_discipline(desc)
            if disc:
                prefix = code.strip()[:2]
                n1_codes[prefix] = disc

    # Second pass: collect totals for each discipline
    disc_totals = {}
    disc_mo = {}

    for row in rows_data:
        if row[0] is None or not isinstance(row[0], str):
            continue
        code = row[0].strip()
        desc = row[1] if row[1] and isinstance(row[1], str) else ''

        # Get total from column 28
        total_val = safe_float(row[28]) if len(row) > 28 else 0

        # N1 total (XX or XX.XXX level)
        prefix = code[:2]
        if prefix in n1_codes:
            disc = n1_codes[prefix]

            # Only take N1 level total (XX or XX.XXX, not XX.XXX.XXX)
            if re.match(r'^\d{2}\s*$', code) or re.match(r'^\d{2}\.\d{3}\s*$', code):
                if total_val > 0:
                    # Use N1 (XX) total if available, otherwise N2 (XX.XXX)
                    if re.match(r'^\d{2}\s*$', code):
                        disc_totals[disc] = total_val
                    elif disc not in disc_totals:
                        disc_totals[disc] = total_val

            # Check for MO subetapas
            if re.match(r'^\d{2}\.\d{3}\.\d{3}\s*$', code):
                if 'mão de obra' in desc.lower() or 'mao de obra' in desc.lower():
                    if total_val > 0:
                        disc_mo[disc] = disc_mo.get(disc, 0) + total_val

            # Check for MO service lines
            if re.match(r'^\d{2}\.\d{3}\.\d{3}\.\d{3}\s*$', code):
                if is_mo_line(desc):
                    if total_val > 0:
                        disc_mo[disc] = disc_mo.get(disc, 0) + total_val

    # Handle "Preventivas e GLP" combined sections
    # If we have a combined preventivas+gas category, keep as preventivas
    # unless gas was also found separately

    result = {}
    for disc, val in disc_totals.items():
        result[disc] = round(val, 2)

    for disc, mo_val in disc_mo.items():
        if disc in result and result[disc] > 0:
            mo_material[disc] = {
                'mo_valor': round(mo_val, 2),
                'material_valor': round(result[disc] - mo_val, 2),
                'mo_pct': round(mo_val / result[disc] * 100, 1) if result[disc] > 0 else 0
            }

    return result, mo_material


def extract_from_obra_sheet(wb, sheet_name='Obra'):
    """
    Extract from 'Obra' summary sheet.
    Look for subsection rows under "Instalações".
    """
    ws = wb[sheet_name]
    result = {}

    in_instalacoes = False
    for row in ws.iter_rows(min_row=1, max_row=500, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        row_text = ' '.join(str(v) for v in vals)

        # Check if we're entering the Instalações section
        if isinstance(vals[0], str):
            t = vals[0].lower().strip()
            if 'instalações' in t or 'instalacoes' in t:
                in_instalacoes = True
                # Check if this is a combined line
                disc = classify_discipline(vals[0])
                if disc:
                    for v in vals[1:]:
                        if isinstance(v, (int, float)) and v > 1000:
                            result[disc] = round(float(v), 2)
                            break
                continue

            # If we hit next macrogroup, stop
            if in_instalacoes and not any(k in t for k in ['hidro', 'eletr', 'prevent', 'gas', 'gás', 'telecom', 'ppci', 'pci', 'spda', 'comunic', 'instal']):
                if any(isinstance(v, (int, float)) and v > 1000 for v in vals):
                    # This is a new macrogroup section
                    break

        if in_instalacoes and isinstance(vals[0], str):
            disc = classify_discipline(vals[0])
            if disc:
                for v in vals[1:]:
                    if isinstance(v, (int, float)) and v > 1000:
                        result[disc] = result.get(disc, 0) + round(float(v), 2)
                        break

    return result, {}


def extract_from_instalacoes_sheet(wb, sheet_name):
    """
    Extract from a dedicated INSTALAÇÕES summary sheet that lists sub-disciplines.
    """
    ws = wb[sheet_name]
    result = {}
    mo_material = {}

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=15, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        if isinstance(vals[0], str):
            desc = vals[0]
            disc = classify_discipline(desc)
            if disc:
                # Find the total value
                # Usually in column 3 (index 3) or a nearby column
                total = 0
                for v in vals[1:]:
                    if isinstance(v, (int, float)) and v > 1000:
                        total = float(v)
                        break

                if total > 0:
                    is_mo = is_mo_line(desc)
                    if is_mo:
                        if disc not in mo_material:
                            mo_material[disc] = {'mo_valor': 0, 'material_valor': 0, 'mo_pct': 0}
                        mo_material[disc]['mo_valor'] += round(total, 2)
                    else:
                        result[disc] = result.get(disc, 0) + round(total, 2)

    # Update MO percentages
    for disc in mo_material:
        if disc in result:
            total = result[disc] + mo_material[disc]['mo_valor']
            result[disc] = round(total, 2)
            mo_material[disc]['material_valor'] = round(total - mo_material[disc]['mo_valor'], 2)
            mo_material[disc]['mo_pct'] = round(mo_material[disc]['mo_valor'] / total * 100, 1) if total > 0 else 0

    return result, mo_material


def extract_from_separate_sheets(wb, sheets):
    """
    Extract from separate discipline sheets (HIDRO, ELÉTRICA, PPCI, etc.)
    Look for total/subtotal rows.
    """
    result = {}
    mo_material = {}

    sheet_discipline_map = {
        'HIDRO': 'hidrossanitarias',
        'HIDROSANITÁRIO': 'hidrossanitarias',
        'HIDROSSANITÁRIO': 'hidrossanitarias',
        'HIDROSSANITARIO': 'hidrossanitarias',
        'Hidrossanitário': 'hidrossanitarias',
        'HIDROSSANITÁRIO_Qtvos': 'hidrossanitarias',
        'HIDROSSANITÁRIO_ESTIMATIVA': 'hidrossanitarias',
        'ELÉTRICA': 'eletricas',
        'ELÉTRICA_Qtvos': 'eletricas',
        'ELÉTRICA_ESTIMATIVA': 'eletricas',
        'Elétrica': 'eletricas',
        'Eletrico': 'eletricas',
        'ELÉTRICO': 'eletricas',
        'ELETR': 'eletricas',
        'Elétrico': 'eletricas',
        'PREV': 'preventivas',
        'PCI': 'preventivas',
        'PPCI': 'preventivas',
        'PPCI-ESTIMATIVA': 'preventivas',
        'PPCI ': 'preventivas',
        'PCI | GLP': 'preventivas',  # combined
        'PREVENTIVA': 'preventivas',
        'GÁS': 'gas',
        'GAS': 'gas',
        'INTERFONIA': 'telecom',
        'INTERFONIA E ACESSOS': 'telecom',
        'Telefonico': 'telecom',
        'Telefone': 'telecom',
        'TELECOMUNICAÇÕES': 'telecom',
        'COMUNICAÇÕES': 'telecom',
        'SOM E WIFI': 'telecom',
        'SPDA': 'preventivas',
        'ESTIMATIVA - SPDA': 'preventivas',
    }

    for sname in sheets:
        # Match sheet name to discipline
        disc = None
        for pattern, d in sheet_discipline_map.items():
            if sname.strip() == pattern or sname.strip().upper() == pattern.upper():
                disc = d
                break

        if not disc:
            # Try partial matching
            sname_upper = sname.strip().upper()
            if 'HIDRO' in sname_upper:
                disc = 'hidrossanitarias'
            elif 'ELETR' in sname_upper or 'ELETRIC' in sname_upper:
                disc = 'eletricas'
            elif 'PCI' in sname_upper or 'PPCI' in sname_upper or 'PREV' in sname_upper:
                disc = 'preventivas'
            elif 'GAS' in sname_upper or 'GÁS' in sname_upper or 'GLP' in sname_upper:
                if 'PCI' not in sname_upper and 'PREV' not in sname_upper:
                    disc = 'gas'
                else:
                    disc = 'preventivas'
            elif 'INTERF' in sname_upper or 'TELEC' in sname_upper or 'TELEF' in sname_upper or 'COMUNIC' in sname_upper:
                disc = 'telecom'
            elif 'SPDA' in sname_upper:
                disc = 'preventivas'

        if not disc:
            continue

        # Don't overwrite if we already have this discipline from a better source
        # (Qtvos sheets are quantitative, ESTIMATIVA sheets have estimates)

        ws = wb[sname]
        sheet_total = 0
        sheet_mo = 0

        # Scan for total rows
        for row in ws.iter_rows(min_row=1, max_row=500, max_col=20, values_only=True):
            vals = [v for v in row if v is not None]
            if not vals:
                continue

            row_text = ' '.join(str(v) for v in vals if isinstance(v, str)).lower()

            # Look for "total" or "subtotal" rows
            if 'total' in row_text and 'subtotal' not in row_text:
                for v in vals:
                    if isinstance(v, (int, float)) and v > 1000:
                        if v > sheet_total:
                            sheet_total = float(v)
                        break

            # Look for MO lines
            if is_mo_line(row_text):
                for v in vals:
                    if isinstance(v, (int, float)) and v > 100:
                        sheet_mo += float(v)
                        break

        if sheet_total > 0:
            result[disc] = result.get(disc, 0) + round(sheet_total, 2)
            if sheet_mo > 0:
                if disc not in mo_material:
                    mo_material[disc] = {'mo_valor': 0, 'material_valor': 0, 'mo_pct': 0}
                mo_material[disc]['mo_valor'] += round(sheet_mo, 2)

    # Update MO percentages
    for disc in mo_material:
        if disc in result and result[disc] > 0:
            mo_material[disc]['material_valor'] = round(result[disc] - mo_material[disc]['mo_valor'], 2)
            mo_material[disc]['mo_pct'] = round(mo_material[disc]['mo_valor'] / result[disc] * 100, 1)

    return result, mo_material


def extract_from_orçamento_executivo_sheet(wb, sheet_name):
    """
    Extract from ORÇAMENTO_EXECUTIVO or similar summary sheets.
    These have a flat list of items with N1/N2 codes.
    """
    ws = wb[sheet_name]
    result = {}
    mo_material = {}

    in_instalacoes = False
    current_discipline = None

    for row in ws.iter_rows(min_row=1, max_row=3000, max_col=20, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        # Find text descriptions
        desc = ''
        total_val = 0
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                desc = v
            if isinstance(v, (int, float)) and abs(v) > 100:
                if total_val == 0:
                    total_val = float(v)

        if not desc:
            continue

        dl = desc.lower()

        # Check if we're in the Instalações section
        if 'instalações' in dl or 'instalacoes' in dl:
            in_instalacoes = True

        # Check if we've left (next macrogroup)
        if in_instalacoes and any(k in dl for k in ['impermeabiliza', 'rev. interno', 'revestimento interno',
                                                      'teto', 'forro', 'piso', 'pintura', 'esquadria',
                                                      'fachada', 'complementar', 'louças', 'loucas']):
            in_instalacoes = False

        if in_instalacoes or any(k in dl for k in ['hidrossanit', 'elétric', 'eletric', 'preventiv', 'telecom']):
            disc = classify_discipline(desc)
            if disc and total_val > 0:
                if is_mo_line(desc):
                    if disc not in mo_material:
                        mo_material[disc] = {'mo_valor': 0, 'material_valor': 0, 'mo_pct': 0}
                    mo_material[disc]['mo_valor'] += round(abs(total_val), 2)
                else:
                    result[disc] = result.get(disc, 0) + round(abs(total_val), 2)

    # Consolidate MO
    for disc in mo_material:
        if disc in result:
            total = result[disc] + mo_material[disc]['mo_valor']
            result[disc] = round(total, 2)
            mo_material[disc]['material_valor'] = round(total - mo_material[disc]['mo_valor'], 2)
            mo_material[disc]['mo_pct'] = round(mo_material[disc]['mo_valor'] / total * 100, 1) if total > 0 else 0

    return result, mo_material


def extract_from_resumo_sheet(wb, sheet_name):
    """
    Extract from summary/resumo sheets like 'Orçamento Resumo'.
    """
    ws = wb[sheet_name]
    result = {}

    for row in ws.iter_rows(min_row=1, max_row=200, max_col=20, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        for v in vals:
            if isinstance(v, str):
                disc = classify_discipline(v)
                if disc:
                    # Find the total value
                    for v2 in vals:
                        if isinstance(v2, (int, float)) and v2 > 1000:
                            result[disc] = result.get(disc, 0) + round(float(v2), 2)
                            break
                    break

    return result, {}


def extract_from_eap_sheet(wb, sheet_name='EAP'):
    """
    Extract from EAP (WBS) sheets.
    """
    ws = wb[sheet_name]
    result = {}

    in_instalacoes = False

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=20, values_only=True):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        desc = ''
        total_val = 0
        for v in vals:
            if isinstance(v, str) and len(v) > 3:
                if not desc:
                    desc = v
            if isinstance(v, (int, float)) and v > 1000:
                total_val = float(v)

        if not desc:
            continue

        dl = desc.lower()
        if 'instalações' in dl or 'instalacoes' in dl:
            in_instalacoes = True

        # Next section
        if in_instalacoes and any(k in dl for k in ['impermeabiliza', 'louças', 'loucas', 'esquadria',
                                                      'revestiment', 'teto', 'forro', 'piso', 'pintura',
                                                      'fachada', 'complementar']):
            in_instalacoes = False

        if in_instalacoes:
            disc = classify_discipline(desc)
            if disc and total_val > 0:
                result[disc] = result.get(disc, 0) + round(total_val, 2)

    return result, {}


# ── Main extraction logic per project ────────────────────────────────────

def extract_project(slug, path):
    """Extract installation breakdown from a project's xlsx."""
    path = path.replace('\\', '/')

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return None, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    sheets_upper = [s.strip().upper() for s in sheets]

    result = {}
    mo_material = {}
    method = ''

    try:
        # Strategy 1: Ger_Executivo (most detailed)
        ger_exec_sheets = [s for s in sheets if 'Ger_Executivo' in s or 'ger_executivo' in s.lower()]
        if ger_exec_sheets:
            r, m = extract_from_ger_executivo(wb, ger_exec_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Ger_Executivo ({ger_exec_sheets[0]})'

        # Strategy 2: Sienge Relatório
        if not result and 'Relatório' in sheets:
            r, m = extract_from_relatorio_sienge(wb)
            if r:
                result = r
                mo_material = m
                method = 'Relatório (Sienge)'

        # Strategy 3: Dedicated INSTALAÇÕES sheet
        inst_sheets = [s for s in sheets if 'INSTALAÇÕES' in s.upper() and 'PROVISÓRI' not in s.upper() and 'COMPLEMENT' not in s.upper()]
        if not result and inst_sheets:
            r, m = extract_from_instalacoes_sheet(wb, inst_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'INSTALAÇÕES sheet ({inst_sheets[0]})'

        # Strategy 4: Separate discipline sheets
        disc_sheets = [s for s in sheets if any(k in s.upper() for k in
                      ['HIDRO', 'ELETR', 'ELETRIC', 'PCI', 'PPCI', 'PREV', 'GÁS', 'GAS', 'GLP',
                       'INTERF', 'TELEC', 'TELEF', 'COMUNIC', 'SPDA'])]
        # Filter out non-installation sheets
        disc_sheets = [s for s in disc_sheets if 'PROVISÓR' not in s.upper() and
                      'INSUMO' not in s.upper() and 'PREÇO' not in s.upper() and
                      'Qtvos' not in s]

        if not result and disc_sheets:
            r, m = extract_from_separate_sheets(wb, disc_sheets)
            if r:
                result = r
                mo_material = m
                method = f'Separate sheets ({", ".join(disc_sheets[:3])})'

        # Strategy 5: ORÇAMENTO_EXECUTIVO sheet
        orc_exec_sheets = [s for s in sheets if 'ORÇAMENTO_EXECUTIVO' in s.upper() or 'ORCAMENTO_EXECUTIVO' in s.upper()]
        if not result and orc_exec_sheets:
            r, m = extract_from_orçamento_executivo_sheet(wb, orc_exec_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'ORÇAMENTO_EXECUTIVO ({orc_exec_sheets[0]})'

        # Strategy 6: Obra sheet
        obra_sheets = [s for s in sheets if s.strip().lower() in ['obra', 'obra ']]
        if not result and obra_sheets:
            r, m = extract_from_obra_sheet(wb, obra_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Obra sheet ({obra_sheets[0]})'

        # Strategy 7: Resumo / Orçamento Resumo
        resumo_sheets = [s for s in sheets if 'resumo' in s.lower() or 'Resumo' in s]
        if not result and resumo_sheets:
            r, m = extract_from_resumo_sheet(wb, resumo_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'Resumo ({resumo_sheets[0]})'

        # Strategy 8: EAP sheet
        eap_sheets = [s for s in sheets if s.strip().upper() == 'EAP']
        if not result and eap_sheets:
            r, m = extract_from_eap_sheet(wb, eap_sheets[0])
            if r:
                result = r
                mo_material = m
                method = f'EAP ({eap_sheets[0]})'

        # HYBRID: If Ger_Executivo found some but separate sheets might add more
        if result and disc_sheets:
            r2, m2 = extract_from_separate_sheets(wb, disc_sheets)
            for disc in r2:
                if disc not in result:
                    result[disc] = r2[disc]
                    if disc in m2:
                        mo_material[disc] = m2[disc]
                    method += f' + {disc} from sheets'

    except Exception as e:
        wb.close()
        return None, None, f"Error extracting: {e}"

    wb.close()

    if not result:
        return None, None, f"No breakdown found. Sheets: {sheets}"

    return result, mo_material, method


# ── Handle combined "Preventivas e GLP" ─────────────────────────────────

def split_preventivas_gas(result, mo_material, wb_path, wb_sheets):
    """
    If we only have 'preventivas' and it includes gas, try to split them.
    """
    # This is best-effort; some projects combine them
    return result, mo_material


# ── Process all projects ─────────────────────────────────────────────────

def main():
    stats = {
        'total': len(BATCH),
        'extracted': 0,
        'failed': 0,
        'no_breakdown': 0,
        'disciplines_found': {
            'hidrossanitarias': 0,
            'eletricas': 0,
            'preventivas': 0,
            'gas': 0,
            'telecom': 0
        },
        'mo_splits_found': 0
    }

    details = []

    for p in BATCH:
        slug = p['slug']
        path = p['path'].replace('\\', '/')

        print(f"\n{'='*60}")
        print(f"Processing: {slug}")

        result, mo_material, msg = extract_project(slug, path)

        if result is None:
            print(f"  FAILED: {msg}")
            stats['failed'] += 1
            details.append({'slug': slug, 'status': 'failed', 'reason': msg})
            continue

        if not result:
            print(f"  NO BREAKDOWN: {msg}")
            stats['no_breakdown'] += 1
            details.append({'slug': slug, 'status': 'no_breakdown', 'reason': msg})
            continue

        print(f"  Method: {msg}")
        print(f"  Disciplines found: {list(result.keys())}")
        for disc, val in sorted(result.items()):
            mo_info = ''
            if disc in mo_material:
                mo_info = f"  (MO: R${mo_material[disc]['mo_valor']:,.2f} = {mo_material[disc]['mo_pct']}%)"
                stats['mo_splits_found'] += 1
            print(f"    {disc}: R${val:,.2f}{mo_info}")
            stats['disciplines_found'][disc] += 1

        stats['extracted'] += 1

        # Update JSON file
        json_path = INDICES_DIR / f"{slug}.json"
        if json_path.exists():
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        else:
            print(f"  WARNING: {json_path} does not exist, creating")
            data = {'slug': slug}

        # Build instalacoes_breakdown
        breakdown = {}
        for disc in ['hidrossanitarias', 'eletricas', 'preventivas', 'gas', 'telecom']:
            if disc in result:
                breakdown[disc] = round(result[disc], 2)

        data['instalacoes_breakdown'] = breakdown

        # Build split_mo_material
        if mo_material:
            splits = {}
            for disc, info in mo_material.items():
                if info['mo_valor'] > 0:
                    splits[disc] = {
                        'mo_valor': round(info['mo_valor'], 2),
                        'material_valor': round(info['material_valor'], 2),
                        'mo_pct': info['mo_pct']
                    }
            if splits:
                data['split_mo_material'] = splits

        data['instalacoes_extraction_method'] = msg

        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        details.append({'slug': slug, 'status': 'extracted', 'disciplines': list(result.keys()), 'method': msg})

    # Print summary
    print(f"\n\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total projects in batch: {stats['total']}")
    print(f"Extracted: {stats['extracted']}")
    print(f"Failed to open: {stats['failed']}")
    print(f"No breakdown found: {stats['no_breakdown']}")
    print(f"\nDisciplines found across projects:")
    for disc, count in sorted(stats['disciplines_found'].items()):
        print(f"  {disc}: {count}/{stats['extracted']} projects")
    print(f"\nMO/Material splits found: {stats['mo_splits_found']} discipline-project combinations")

    # List failures
    failures = [d for d in details if d['status'] != 'extracted']
    if failures:
        print(f"\nProjects without breakdown ({len(failures)}):")
        for d in failures:
            print(f"  {d['slug']}: {d.get('reason', 'unknown')}")


if __name__ == '__main__':
    main()
