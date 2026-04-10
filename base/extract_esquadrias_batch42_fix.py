#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Fix script for 12 remaining no-data projects in batch 42-83.
Handles additional spreadsheet formats not covered by the main script.

Formats handled:
1. Compact A-F (code/desc/unit/qty/pu/total) - eze-eilat, mabrem-gran-torino, grandezza-gran-tower
2. Summary ETAPA sheets (A or B for name, next col for value) - grandezza-gran-royal, macom-adda, macom-villa-lobos, gmf-moradas-do-atalaia, mtf-m-village-jardim
3. Ger_Executivo with B-I columns (code in B, desc in C) - gdi-playa-negra, muller-empreendimentos-elleva
4. Planilha de Orçamento Sienge (Grupo format) - mabrem-san-marino
5. Quantitative tables without prices - mg3-la-vie (limited extraction)
"""
import json, sys, os, traceback
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def normalize(text):
    t = text.lower().strip()
    replacements = {
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'è': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'õ': 'o', 'ô': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c', 'ñ': 'n',
        '\xa0': ' ', '\t': ' ',
    }
    for old, new in replacements.items():
        t = t.replace(old, new)
    return t


def classify_esquadria(desc):
    d = normalize(desc)
    if any(x in d for x in ['pele de vidro', 'curtain wall', 'muro cortina', 'fachada de vidro',
                              'fachada envidracada', 'fachada vidro']):
        return 'pele_vidro'
    if 'brise' in d:
        return 'brises'
    if 'guarda corpo' in d or 'guarda-corpo' in d or 'guardacorpo' in d:
        return 'guarda_corpo'
    if 'corta fogo' in d or 'corta-fogo' in d or 'cortafogo' in d:
        return 'portas_corta_fogo'
    if ('porta' in d and ('madeira' in d or 'semi oca' in d or 'semi-oca' in d or 'semioca' in d)):
        return 'portas_madeira'
    if 'porta de madeira' in d:
        return 'portas_madeira'
    if any(x in d for x in ['esquadrias de aluminio', 'esquadria de aluminio',
                              'janela de aluminio', 'janelas aluminio',
                              'esquadrias aluminio', 'esquadrias de pvc']):
        return 'aluminio'
    if 'aluminio' in d and ('esquadria' in d or 'janela' in d or 'porta' in d):
        return 'aluminio'
    if 'pvc' in d and ('esquadria' in d or 'janela' in d):
        return 'aluminio'
    if 'contramarco' in d and ('aluminio' in d or 'alumínio' in d.lower()):
        return 'aluminio'
    if any(x in d for x in ['vidro temperado', 'vidro laminado', 'vidro insulado',
                              'envidracamento', 'vidros']):
        if 'guarda' not in d:
            return 'vidros'
    return None


def classify_fachada(desc):
    d = normalize(desc)
    is_ext = any(x in d for x in ['externo', 'externa', 'fachada', 'exterior'])
    if 'reboco' in d and is_ext:
        return 'reboco_externo'
    if 'massa unica' in d and is_ext:
        return 'reboco_externo'
    if ('textura' in d or 'pintura' in d) and is_ext:
        return 'textura_pintura_ext'
    if 'grafiato' in d and is_ext:
        return 'textura_pintura_ext'
    if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and is_ext:
        return 'pastilha_porcelanato'
    if 'junta' in d and is_ext:
        return 'juntas'
    return None


def classify_acabamento(desc, unit):
    d = normalize(desc)
    u = normalize(str(unit)) if unit else ''
    if 'mao de obra' in d or 'empreitada' in d:
        return None
    if 'porcelanato' in d and ('piso' in d or 'revestimento' in d or u == 'm2'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'porcelanato'
    if ('ceramica' in d or 'ceramico' in d) and ('revestimento' in d or 'piso' in d or u == 'm2'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'pastilha' not in d:
                return 'ceramica'
    if 'laminado' in d and ('piso' in d or 'revestimento' in d or u == 'm2'):
        if 'vidro' not in d:
            return 'laminado'
    if 'rodape' in d:
        return 'rodape'
    if 'contrapiso' in d and u == 'm2':
        return 'contrapiso'
    if 'forro' in d and ('gesso' in d or 'mineral' in d):
        return 'forro_gesso'
    if ('pintura' in d or 'tinta' in d) and u == 'm2':
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'piso' not in d:
                return 'pintura'
    if 'chapisco' in d and u == 'm2':
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'chapisco'
    if ('reboco' in d or 'massa unica' in d) and u == 'm2':
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'reboco'
    return None


# ============================================================
# Compact A-F extractor (code/desc/unit/qty/pu/total in cols 0-5)
# ============================================================

def extract_compact_af(wb, sheet_name):
    """Extract from compact A-F format: A=code, B=desc, C=unit, D=qty, E=pu, F=total"""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=10, values_only=True):
        if len(row) < 6:
            continue
        code = safe_str(row[0])
        desc = safe_str(row[1])
        unit = safe_str(row[2])
        qty = safe_float(row[3])
        pu = safe_float(row[4])
        total = safe_float(row[5])

        if not desc:
            continue

        parts = [p for p in code.replace(' ', '').split('.') if p.strip()]
        is_service = len(parts) >= 4
        is_subetapa = len(parts) == 3
        is_etapa = len(parts) <= 2

        if is_service and pu > 0 and qty > 0:
            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                item_total = total if total > 0 else pu * qty
                esquadrias[cat]['valor'] += item_total
                esquadrias[cat]['qtd'] += qty
                esquadrias[cat]['pu_sum'] += pu * qty
                esquadrias[cat]['pu_count'] += qty
                esquadrias[cat]['items'].append({'desc': desc, 'qty': qty, 'pu': pu, 'total': item_total})

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += (total if total > 0 else pu * qty)

            # Only collect acabamentos PU from m2/m items (not VB/un lump sums)
            unit_n = normalize(unit)
            if unit_n in ('m2', 'm'):
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

        # Subetapa fallbacks
        if is_subetapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'alumin' in d and 'aluminio' not in esquadrias:
                esquadrias.setdefault('aluminio_subetapa', {'valor': 0.0})
                esquadrias['aluminio_subetapa']['valor'] += total
            if 'guarda corpo' in d and 'guarda_corpo' not in esquadrias:
                esquadrias.setdefault('guarda_corpo_subetapa', {'valor': 0.0})
                esquadrias['guarda_corpo_subetapa']['valor'] += total
            if 'brise' in d and 'brises' not in esquadrias:
                esquadrias.setdefault('brises_subetapa', {'valor': 0.0})
                esquadrias['brises_subetapa']['valor'] += total
            if ('madeira' in d or 'porta' in d) and 'esquadria' in d:
                esquadrias.setdefault('portas_madeira', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_madeira']['valor'] += total
            if 'corta fogo' in d or 'corta-fogo' in d:
                esquadrias.setdefault('portas_corta_fogo', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_corta_fogo']['valor'] += total

            if ('reboco' in d or 'argamassa' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('reboco_externo', {'valor': 0.0})
                fachada['reboco_externo']['valor'] += total
            if ('textura' in d or 'pintura' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('textura_pintura_ext', {'valor': 0.0})
                fachada['textura_pintura_ext']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('pastilha_porcelanato', {'valor': 0.0})
                fachada['pastilha_porcelanato']['valor'] += total

        # Etapa-level totals
        if is_etapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d and ('revestiment' in d or 'acabament' in d):
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    return esquadrias, fachada, acabamentos


# ============================================================
# Code-in-B extractor (B=code, C=desc, D=unit, E=qty, F=pu, G=total)
# ============================================================

def extract_code_in_b(wb, sheet_name):
    """Extract from Ger_Executivo variant: B=code, C=desc, D=unit, E=qty, F=pu, G=total"""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=10, values_only=True):
        if len(row) < 8:
            continue
        code = safe_str(row[1])   # B
        desc = safe_str(row[2])   # C
        unit = safe_str(row[3])   # D
        qty = safe_float(row[4])  # E
        pu = safe_float(row[5])   # F
        total = safe_float(row[6])  # G

        if not desc:
            continue

        parts = [p for p in code.replace(' ', '').split('.') if p.strip()]
        is_service = len(parts) >= 4
        is_subetapa = len(parts) == 3
        is_etapa = len(parts) <= 2

        if is_service:
            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                item_total = total if total > 0 else (pu * qty if pu > 0 and qty > 0 else 0)
                esquadrias[cat]['valor'] += item_total
                esquadrias[cat]['qtd'] += qty
                if pu > 0 and qty > 0:
                    esquadrias[cat]['pu_sum'] += pu * qty
                    esquadrias[cat]['pu_count'] += qty
                esquadrias[cat]['items'].append({'desc': desc, 'qty': qty, 'pu': pu, 'total': item_total})

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += (total if total > 0 else (pu * qty if pu > 0 and qty > 0 else 0))

            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

        # Subetapa
        if is_subetapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'alumin' in d and 'aluminio' not in esquadrias:
                esquadrias.setdefault('aluminio_subetapa', {'valor': 0.0})
                esquadrias['aluminio_subetapa']['valor'] += total
            if 'guarda corpo' in d and 'guarda_corpo' not in esquadrias:
                esquadrias.setdefault('guarda_corpo_subetapa', {'valor': 0.0})
                esquadrias['guarda_corpo_subetapa']['valor'] += total
            if 'brise' in d and 'brises' not in esquadrias:
                esquadrias.setdefault('brises_subetapa', {'valor': 0.0})
                esquadrias['brises_subetapa']['valor'] += total
            if ('madeira' in d) and ('esquadria' in d or 'porta' in d) and 'portas_madeira' not in esquadrias:
                esquadrias.setdefault('portas_madeira', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_madeira']['valor'] += total
            if ('corta fogo' in d or 'corta-fogo' in d) and 'portas_corta_fogo' not in esquadrias:
                esquadrias.setdefault('portas_corta_fogo', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_corta_fogo']['valor'] += total
            if ('metalica' in d) and ('esquadria' in d or 'serralheria' in d) and 'portas_corta_fogo' not in esquadrias:
                esquadrias.setdefault('portas_corta_fogo', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_corta_fogo']['valor'] += total

            if ('reboco' in d or 'argamassa' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('reboco_externo', {'valor': 0.0})
                fachada['reboco_externo']['valor'] += total
            if ('textura' in d or 'pintura' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('textura_pintura_ext', {'valor': 0.0})
                fachada['textura_pintura_ext']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('pastilha_porcelanato', {'valor': 0.0})
                fachada['pastilha_porcelanato']['valor'] += total

        if is_etapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d and ('revestiment' in d or 'acabament' in d):
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    return esquadrias, fachada, acabamentos


# ============================================================
# Code-in-C extractor (C=code, D=desc, E=unit, F=qty, G=pu, H=total)
# ============================================================

def extract_code_in_c(wb, sheet_name):
    """Extract from Ger_Executivo variant: C=code, D=desc, E=unit, F=qty, G=pu, H=total"""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=5000, max_col=10, values_only=True):
        if len(row) < 8:
            continue
        code = safe_str(row[2])   # C
        desc = safe_str(row[3])   # D
        unit = safe_str(row[4])   # E
        qty = safe_float(row[5])  # F
        pu = safe_float(row[6])   # G
        total = safe_float(row[7])  # H

        if not desc:
            continue

        parts = [p for p in code.replace(' ', '').split('.') if p.strip()]
        is_service = len(parts) >= 4
        is_subetapa = len(parts) == 3
        is_etapa = len(parts) <= 2

        if is_service:
            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                item_total = total if total > 0 else (pu * qty if pu > 0 and qty > 0 else 0)
                esquadrias[cat]['valor'] += item_total
                esquadrias[cat]['qtd'] += qty
                if pu > 0 and qty > 0:
                    esquadrias[cat]['pu_sum'] += pu * qty
                    esquadrias[cat]['pu_count'] += qty
                esquadrias[cat]['items'].append({'desc': desc, 'qty': qty, 'pu': pu, 'total': item_total})

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += (total if total > 0 else (pu * qty if pu > 0 and qty > 0 else 0))

            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

        # Subetapa
        if is_subetapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'alumin' in d and 'aluminio' not in esquadrias:
                esquadrias.setdefault('aluminio_subetapa', {'valor': 0.0})
                esquadrias['aluminio_subetapa']['valor'] += total
            if 'guarda corpo' in d and 'guarda_corpo' not in esquadrias:
                esquadrias.setdefault('guarda_corpo_subetapa', {'valor': 0.0})
                esquadrias['guarda_corpo_subetapa']['valor'] += total
            if 'brise' in d and 'brises' not in esquadrias:
                esquadrias.setdefault('brises_subetapa', {'valor': 0.0})
                esquadrias['brises_subetapa']['valor'] += total
            if ('madeira' in d) and ('esquadria' in d or 'porta' in d) and 'portas_madeira' not in esquadrias:
                esquadrias.setdefault('portas_madeira', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_madeira']['valor'] += total
            if ('corta fogo' in d or 'corta-fogo' in d) and 'portas_corta_fogo' not in esquadrias:
                esquadrias.setdefault('portas_corta_fogo', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_corta_fogo']['valor'] += total
            if ('metalica' in d) and ('esquadria' in d or 'serralheria' in d) and 'portas_corta_fogo' not in esquadrias:
                esquadrias.setdefault('portas_corta_fogo', {'valor': 0.0, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []})
                esquadrias['portas_corta_fogo']['valor'] += total

            if ('reboco' in d or 'argamassa' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('reboco_externo', {'valor': 0.0})
                fachada['reboco_externo']['valor'] += total
            if ('textura' in d or 'pintura' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('textura_pintura_ext', {'valor': 0.0})
                fachada['textura_pintura_ext']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('externo' in d or 'fachada' in d):
                fachada.setdefault('pastilha_porcelanato', {'valor': 0.0})
                fachada['pastilha_porcelanato']['valor'] += total

        if is_etapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d and ('revestiment' in d or 'acabament' in d):
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    return esquadrias, fachada, acabamentos


# ============================================================
# Sienge Planilha de Orçamento do Grupo format
# B=code/desc mixed, D=desc, F=unit, J=qty, K=pu, M=total
# ============================================================

def extract_sienge_grupo(wb, sheet_name):
    """Extract from Sienge Planilha de Orçamento do Grupo (mabrem-san-marino format)."""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=15, values_only=True):
        if len(row) < 14:
            continue
        col_a = safe_str(row[0])
        col_b = safe_str(row[1])
        col_c = safe_str(row[2])
        col_d = safe_str(row[3])
        col_e = safe_str(row[4])
        col_f = safe_str(row[6]) if len(row) > 6 else ''  # unit (index 6)
        col_j = safe_float(row[10])  # Quantidade (index 10)
        col_k = safe_float(row[11])  # Custo Unit. (index 11)
        col_m = safe_float(row[13]) if len(row) > 13 else 0  # Total (index 13)

        # Service rows have code in col_b like '01.01.001.0001.0001'
        # Section headers have code in col_b like '01.01' and desc in col_c
        code = col_b.strip()
        parts = [p for p in code.split('.') if p.strip()]

        # Try to get desc from col_e (service detail) or col_c (section header)
        desc = col_e if col_e else col_c.strip()
        desc = desc.strip()

        # Total from col N (index 13)
        total = col_m

        if not desc or not code:
            continue

        is_service = len(parts) >= 5 or (len(parts) >= 4 and col_k > 0)
        is_subetapa = len(parts) == 3
        is_etapa = len(parts) <= 2

        if is_service and col_k > 0 and col_j > 0:
            unit = col_f
            unit_n = normalize(unit)
            is_vb = unit_n in ('vb', 'un', 'cj')

            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                item_total = total if total > 0 else col_k * col_j
                esquadrias[cat]['valor'] += item_total
                if not is_vb:
                    esquadrias[cat]['qtd'] += col_j
                    esquadrias[cat]['pu_sum'] += col_k * col_j
                    esquadrias[cat]['pu_count'] += col_j

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += (total if total > 0 else col_k * col_j)

            if unit_n in ('m2', 'm'):
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': col_k, 'qty': col_j, 'desc': desc, 'unit': unit})

        # Section headers for fallback
        if is_subetapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'alumin' in d:
                esquadrias.setdefault('aluminio_subetapa', {'valor': 0.0})
                esquadrias['aluminio_subetapa']['valor'] += total

        if is_etapa and total > 0:
            d = normalize(desc)
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d:
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    return esquadrias, fachada, acabamentos


# ============================================================
# Summary-only etapa extractor (flexible column detection)
# ============================================================

def extract_summary_flexible(wb, sheet_name):
    """Extract etapa-level totals from summary sheets with flexible column detection."""
    ws = wb[sheet_name]
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=100, max_col=15, values_only=True):
        if len(row) < 2:
            continue

        # Find description and value columns dynamically
        for desc_idx in range(min(5, len(row))):
            desc = safe_str(row[desc_idx])
            if not desc or len(desc) < 5:
                continue
            d = normalize(desc)

            # Skip header rows
            if d in ('etapa', 'total', 'cub/sc', 'cub/estado'):
                continue

            # Find the first numeric column after desc_idx
            for val_idx in range(desc_idx + 1, min(desc_idx + 5, len(row))):
                val = safe_float(row[val_idx]) if val_idx < len(row) else 0
                if val > 1000:  # Reasonable minimum for an etapa total
                    if 'esquadria' in d and ('vidro' in d or 'ferragem' in d or 'ferragens' in d):
                        esquadrias['_total_esquadrias'] = val
                    elif 'fachada' in d and ('revestiment' in d or 'acabament' in d):
                        fachada['_total_fachada'] = val
                    break
            break  # Only use first desc column found per row

    return esquadrias, fachada, acabamentos


# ============================================================
# Build output helpers
# ============================================================

def build_esquadrias_detail(raw_esquadrias):
    result = {}
    categories = ['aluminio', 'guarda_corpo', 'portas_madeira', 'portas_corta_fogo',
                   'vidros', 'pele_vidro', 'brises']
    for cat in categories:
        if cat in raw_esquadrias:
            data = raw_esquadrias[cat]
            entry = {'valor': round(data['valor'], 2) if data['valor'] else None}
            if data.get('qtd', 0) > 0:
                entry['qtd'] = round(data['qtd'], 2)
            if data.get('pu_count', 0) > 0:
                entry['pu'] = round(data['pu_sum'] / data['pu_count'], 2)
            result[cat] = entry
        elif f'{cat}_subetapa' in raw_esquadrias:
            data = raw_esquadrias[f'{cat}_subetapa']
            result[cat] = {'valor': round(data['valor'], 2)}
    if '_total_esquadrias' in raw_esquadrias:
        result['_total'] = round(raw_esquadrias['_total_esquadrias'], 2)
    return result if result else None


def build_fachada_detail(raw_fachada):
    result = {}
    categories = ['reboco_externo', 'textura_pintura_ext', 'pastilha_porcelanato', 'juntas']
    for cat in categories:
        if cat in raw_fachada:
            result[cat] = {'valor': round(raw_fachada[cat]['valor'], 2)}
        elif f'{cat}_sub' in raw_fachada:
            result[cat] = {'valor': round(raw_fachada[f'{cat}_sub']['valor'], 2)}
    if '_total_fachada' in raw_fachada:
        result['_total'] = round(raw_fachada['_total_fachada'], 2)
    return result if result else None


def build_acabamentos_pus(raw_acabamentos):
    result = {}
    for cat, items in raw_acabamentos.items():
        if not items:
            continue
        total_weighted = sum(i['pu'] * i['qty'] for i in items)
        total_qty = sum(i['qty'] for i in items)
        if total_qty > 0:
            avg_pu = total_weighted / total_qty
            result[cat] = round(avg_pu, 2)
    return result if result else None


def update_json(slug, esq, fach, acab):
    idx_path = INDICES_DIR / f"{slug}.json"
    if idx_path.exists():
        with open(idx_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {'projeto': slug}
    if esq is not None:
        data['esquadrias_detail'] = esq
    if fach is not None:
        data['fachada_detail'] = fach
    if acab is not None:
        data['acabamentos_pus'] = acab
    with open(idx_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ============================================================
# Project-specific handlers
# ============================================================

def process_eze_eilat():
    """Compact A-F format in 'Relatório' sheet."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'eze-eilat')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_compact_af(wb, 'Relatório')
    wb.close()
    return e, f, a


def process_grandezza_gran_tower():
    """Compact A-F format in 'ORÇAMENTO_EXECUTIVO' sheet."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'grandezza-gran-tower')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_compact_af(wb, 'ORÇAMENTO_EXECUTIVO')
    wb.close()
    return e, f, a


def process_mabrem_gran_torino():
    """Compact A-F format in 'ORÇAMENTO_EXECUTIVO' sheet."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'mabrem-gran-torino')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_compact_af(wb, 'ORÇAMENTO_EXECUTIVO')
    wb.close()
    return e, f, a


def process_grandezza_gran_royal():
    """Summary ETAPA in A, VALOR in B."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'grandezza-gran-royal')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_summary_flexible(wb, 'ORÇAMENTO_EXECUTIVO')
    wb.close()
    return e, f, a


def process_macom_adda():
    """Summary in ORÇAMENTO_EXECUTIVO: ETAPA in B, VALOR in C."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'macom-adda')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_summary_flexible(wb, 'ORÇAMENTO_EXECUTIVO')
    wb.close()
    return e, f, a


def process_macom_villa_lobos():
    """Summary in ORÇAMENTO_EXECUTIVO: ETAPA in B, VALOR in C."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'macom-villa-lobos')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_summary_flexible(wb, 'ORÇAMENTO_EXECUTIVO')
    wb.close()
    return e, f, a


def process_gmf_moradas_do_atalaia():
    """Summary in ORÇAMENTO_EXECUTIVO: ETAPA in A, VALOR_EXECUTIVO in col E."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'gmf-moradas-do-atalaia')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    # This one has parametric vs executivo columns. Executivo total is in col F (index 5)
    ws = wb['ORÇAMENTO_EXECUTIVO']
    esquadrias = {}
    fachada = {}

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        if len(row) < 6:
            continue
        desc = safe_str(row[0])
        d = normalize(desc)
        val = safe_float(row[5])  # Executivo VALOR ORÇADO column
        if val <= 1000:
            continue
        if 'esquadria' in d and ('vidro' in d or 'ferragem' in d or 'ferragens' in d):
            esquadrias['_total_esquadrias'] = val
        if 'fachada' in d and ('revestiment' in d or 'acabament' in d):
            fachada['_total_fachada'] = val

    wb.close()
    esq = build_esquadrias_detail(esquadrias)
    fach = build_fachada_detail(fachada)
    return esquadrias, fachada, {}


def process_mtf_m_village_jardim():
    """Summary in 'Gerenciamento executivo ': ETAPA in B, VALOR in C."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'mtf-m-village-jardim')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    e, f, a = extract_summary_flexible(wb, 'Gerenciamento executivo ')
    wb.close()
    return e, f, a


def process_gdi_playa_negra():
    """Ger_Executivo with B-I columns + dedicated Esquadrias sheet."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'gdi-playa-negra')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)

    # Extract from Ger_Executivo (B=code, C=desc, D=unit, E=qty, F=pu, G=total)
    e, f, a = extract_code_in_b(wb, 'Ger_Executivo')

    # Also extract from dedicated Esquadrias sheet
    if 'Esquadrias' in wb.sheetnames:
        ws = wb['Esquadrias']
        aluminio_total = 0
        portas_madeira_total = 0
        portas_corta_fogo_total = 0
        guarda_corpo_total = 0

        for row in ws.iter_rows(min_row=1, max_row=500, max_col=12, values_only=True):
            if len(row) < 11:
                continue
            code = safe_str(row[0])
            desc = safe_str(row[1])
            material = safe_str(row[2])
            tipo = safe_str(row[3])
            qty = safe_float(row[6])
            area = safe_float(row[8])
            cu = safe_float(row[9])
            ct = safe_float(row[10])

            d = normalize(desc)
            m = normalize(material)
            t = normalize(tipo)

            if ct > 0:
                if 'madeira' in m and ('porta' in t or 'porta' in d):
                    portas_madeira_total += ct
                elif 'metalica' in m and ('pcf' in t or 'corta fogo' in d):
                    portas_corta_fogo_total += ct
                elif 'aluminio' in m and ('janela' in t or 'esquadria' in d):
                    aluminio_total += ct

        # Only add esquadrias from dedicated sheet if not already found
        if aluminio_total > 0 and 'aluminio' not in e:
            e['aluminio'] = {'valor': aluminio_total, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []}
        if portas_madeira_total > 0 and 'portas_madeira' not in e:
            e['portas_madeira'] = {'valor': portas_madeira_total, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []}
        if portas_corta_fogo_total > 0 and 'portas_corta_fogo' not in e:
            e['portas_corta_fogo'] = {'valor': portas_corta_fogo_total, 'qtd': 0, 'pu_sum': 0, 'pu_count': 0, 'items': []}

    wb.close()
    return e, f, a


def process_muller_elleva():
    """Resumo sheet with C-H columns (code in C, desc in D, unit in E, qty in F, pu in G, total in H)."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'muller-empreendimentos-elleva')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    # Columns are shifted: C=code, D=desc, E=unit, F=qty, G=pu, H=total
    e, f, a = extract_code_in_c(wb, 'Resumo')
    wb.close()
    return e, f, a


def process_mabrem_san_marino():
    """Sienge Planilha do Grupo format in Sheet1."""
    proj = next(p for p in ALL_PROJECTS if p['slug'] == 'mabrem-san-marino')
    wb = openpyxl.load_workbook(proj['path'], read_only=True, data_only=True)
    # Format: col_a=code (in col 1), col_b=desc/insumo (col 2), col_d=desc detail (col 4),
    # col_f=unit (col 6), col_j=qty (col 10), col_k=pu (col 11), col_m=total (col 13)
    ws = wb['Sheet1']
    esquadrias = {}
    fachada = {}
    acabamentos = {}

    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=15, values_only=True):
        if len(row) < 14:
            continue
        code = safe_str(row[1])  # col B (index 1)
        desc_header = safe_str(row[2])  # col C (index 2) - section headers
        desc_detail = safe_str(row[4])  # col E (index 4) - service detail
        unit = safe_str(row[6])  # col G (index 6)
        qty = safe_float(row[10])  # col K (index 10)
        pu = safe_float(row[11])  # col L (index 11)
        total = safe_float(row[13])  # col N (index 13)

        desc = desc_detail if desc_detail else desc_header.strip()
        if not desc:
            continue

        parts = [p for p in code.replace(' ', '').split('.') if p.strip()]
        is_service = len(parts) >= 5 or (len(parts) >= 4 and pu > 0)
        is_subetapa = len(parts) == 3
        is_etapa = len(parts) <= 2 and total > 0

        if is_service and pu > 0 and qty > 0:
            unit_n = normalize(unit)
            is_vb = unit_n in ('vb', 'un', 'cj')

            cat = classify_esquadria(desc)
            if cat:
                if cat not in esquadrias:
                    esquadrias[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0, 'items': []}
                item_total = total if total > 0 else pu * qty
                esquadrias[cat]['valor'] += item_total
                if not is_vb:
                    esquadrias[cat]['qtd'] += qty
                    esquadrias[cat]['pu_sum'] += pu * qty
                    esquadrias[cat]['pu_count'] += qty

            cat = classify_fachada(desc)
            if cat:
                if cat not in fachada:
                    fachada[cat] = {'valor': 0.0}
                fachada[cat]['valor'] += (total if total > 0 else pu * qty)

            if unit_n in ('m2', 'm'):
                cat = classify_acabamento(desc, unit)
                if cat:
                    if cat not in acabamentos:
                        acabamentos[cat] = []
                    acabamentos[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})

        if is_etapa:
            d = normalize(desc_header)
            if 'esquadria' in d:
                esquadrias['_total_esquadrias'] = esquadrias.get('_total_esquadrias', 0) + total
            if 'fachada' in d:
                fachada['_total_fachada'] = fachada.get('_total_fachada', 0) + total

    wb.close()
    return esquadrias, fachada, acabamentos


def process_mg3_la_vie():
    """Quantitative tables -- limited extraction (no prices in most sheets)."""
    # This is an architecture quantitative file, not a budget spreadsheet.
    # It has areas (m2) for windows, doors, facades, but no prices.
    # Return None to leave current state.
    return {}, {}, {}


# ============================================================
# Main
# ============================================================

HANDLERS = {
    'eze-eilat': process_eze_eilat,
    'gdi-playa-negra': process_gdi_playa_negra,
    'gmf-moradas-do-atalaia': process_gmf_moradas_do_atalaia,
    'grandezza-gran-royal': process_grandezza_gran_royal,
    'grandezza-gran-tower': process_grandezza_gran_tower,
    'mabrem-gran-torino': process_mabrem_gran_torino,
    'mabrem-san-marino': process_mabrem_san_marino,
    'macom-adda': process_macom_adda,
    'macom-villa-lobos': process_macom_villa_lobos,
    'mg3-la-vie': process_mg3_la_vie,
    'mtf-m-village-jardim': process_mtf_m_village_jardim,
    'muller-empreendimentos-elleva': process_muller_elleva,
}


def main():
    results = {'success': 0, 'no_data': 0, 'error': 0}

    for slug, handler in HANDLERS.items():
        print(f"Processing {slug}...", end=' ')
        try:
            raw_esq, raw_fach, raw_acab = handler()
            esq = build_esquadrias_detail(raw_esq)
            fach = build_fachada_detail(raw_fach)
            acab = build_acabamentos_pus(raw_acab)

            if not esq and not fach and not acab:
                print("NO DATA")
                results['no_data'] += 1
                continue

            update_json(slug, esq, fach, acab)
            parts = []
            if esq:
                parts.append(f"esq={list(esq.keys())}")
            if fach:
                parts.append(f"fach={list(fach.keys())}")
            if acab:
                parts.append(f"acab={list(acab.keys())}")
            print(f"OK ({', '.join(parts)})")
            results['success'] += 1

        except Exception as ex:
            print(f"ERROR: {traceback.format_exc()}")
            results['error'] += 1

    print(f"\n{'='*60}")
    print(f"FIX RESULTS: {results['success']} success, {results['no_data']} no data, {results['error']} errors")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
