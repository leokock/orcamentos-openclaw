#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extract esquadrias_detail, fachada_detail, acabamentos_pus from batch 42-83.
V2: Deeper extraction for projects that only got _total or nothing from V1.

Handles diverse formats:
  - Ger_Executivo (standard H/I/J/K/L/M/N columns)
  - ORÇAMENTO_EXECUTIVO summary (A=etapa, B=valor)
  - Relatório / Sienge N2 (A=code, B=desc, C/E/F=un/pu/total)
  - Named detail sheets (Esquadrias, Fachada, Piso, etc.)
  - Sienge full export (B=code, D=desc, F=unit, J/K=values)
  - MG3 tabela sheets (BIM-based quantitativos)
  - Kirchner-style itemized sheets per discipline
"""
import json, sys, re, os, traceback
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"

with open(BASE_DIR / "_all_projects_mapping.json", "r", encoding="utf-8") as f:
    ALL_PROJECTS = json.load(f)

# Only process projects 42-83 that still need data
NEED_INDICES = [44,45,48,49,50,51,55,56,57,61,65,66,69,70,71,72,74,75,76,77,79,81,82,83]
BATCH = [ALL_PROJECTS[i] for i in NEED_INDICES]


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()


def normalize(text):
    t = text.lower().strip()
    replacements = {
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'è': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'õ': 'o', 'ô': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c', 'ñ': 'n',
        '\xa0': ' ',
    }
    for old, new in replacements.items():
        t = t.replace(old, new)
    return t


# ============================================================
# Classification helpers (same as V1, enhanced)
# ============================================================

def classify_esquadria(desc):
    d = normalize(desc)
    if any(x in d for x in ['ensaio', 'laudo', 'projeto de esquadria', 'consultoria',
                              'impermeabilizacao de esquadria', 'estanqueidade']):
        return None
    if any(x in d for x in ['pele de vidro', 'curtain wall', 'muro cortina', 'fachada de vidro',
                              'fachada envidracada', 'fachada vidro']):
        return 'pele_vidro'
    if 'brise' in d:
        return 'brises'
    if ('guarda corpo' in d or 'guarda-corpo' in d or 'guardacorpo' in d):
        if 'ensaio' not in d and 'laudo' not in d and 'teste' not in d:
            return 'guarda_corpo'
    if 'corta fogo' in d or 'corta-fogo' in d or 'cortafogo' in d:
        return 'portas_corta_fogo'
    if ('porta' in d and ('madeira' in d or 'semi oca' in d or 'semi-oca' in d or 'semioca' in d or 'mdf' in d)):
        return 'portas_madeira'
    if 'porta de madeira' in d:
        return 'portas_madeira'
    # Aluminio
    if any(x in d for x in ['esquadrias de aluminio', 'esquadria de aluminio',
                              'janela de aluminio', 'janelas aluminio',
                              'esquadrias aluminio']):
        return 'aluminio'
    if 'aluminio' in d and ('esquadria' in d or 'janela' in d):
        return 'aluminio'
    if 'aluminio' in d and 'porta' in d and 'madeira' not in d and 'corta' not in d:
        return 'aluminio'
    if 'contramarco' in d and 'aluminio' in d:
        return 'aluminio'
    # Vidros
    if any(x in d for x in ['vidro temperado', 'vidro laminado', 'vidro insulado',
                              'envidracamento', 'vidros']):
        if 'guarda' not in d:
            return 'vidros'
    return None


def classify_fachada(desc):
    d = normalize(desc)
    is_ext = any(x in d for x in ['externo', 'externa', 'fachada', 'exterior'])
    if 'reboco' in d and is_ext:
        return 'reboco_externo'
    if 'massa unica' in d and is_ext:
        return 'reboco_externo'
    if ('argamassado' in d or 'argamassa' in d) and is_ext:
        return 'reboco_externo'
    if ('textura' in d or 'pintura' in d) and is_ext:
        return 'textura_pintura_ext'
    if 'grafiato' in d and is_ext:
        return 'textura_pintura_ext'
    if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d or 'ceramic' in d) and is_ext:
        return 'pastilha_porcelanato'
    if 'junta' in d and is_ext:
        return 'juntas'
    return None


def classify_acabamento(desc, unit):
    d = normalize(desc)
    u = normalize(str(unit)) if unit else ''
    if 'mao de obra' in d or 'empreitada' in d:
        return None
    # Reject VB (verba/lump sum) - PU is not meaningful as R$/m2
    if u in ('vb', 'verba', 'cj', 'conjunto', 'gl', 'global', 'un', 'und', 'unid', 'unidade'):
        return None
    if 'rodape' in d:
        return 'rodape'
    if 'porcelanato' in d and u in ('m2', 'm²', ''):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'porcelanato'
    if ('ceramica' in d or 'ceramico' in d) and ('revestimento' in d or 'piso' in d or u in ('m2', 'm²')):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'pastilha' not in d:
                return 'ceramica'
    if 'laminado' in d and ('piso' in d or 'revestimento' in d or u in ('m2', 'm²')):
        return 'laminado'
    if 'contrapiso' in d and u in ('m2', 'm²'):
        return 'contrapiso'
    if 'forro' in d and ('gesso' in d or 'mineral' in d):
        return 'forro_gesso'
    if ('pintura' in d or 'tinta' in d) and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            if 'piso' not in d:
                return 'pintura'
    if 'textura' in d and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'pintura'
    if 'chapisco' in d and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'chapisco'
    if ('reboco' in d or 'massa unica' in d) and u in ('m2', 'm²'):
        if 'externo' not in d and 'externa' not in d and 'fachada' not in d:
            return 'reboco'
    return None


# ============================================================
# Format-specific extractors
# ============================================================

def extract_ger_executivo(wb, sheet_name):
    """Ger_Executivo standard: H=level_type, I=code, J=desc, K=unit, L=qty, M=pu, N=total"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}
    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=20, values_only=True):
        if len(row) < 14:
            continue
        lt = safe_str(row[7] if row[7] else '')
        desc = safe_str(row[9] if row[9] else '')
        unit = safe_str(row[10] if row[10] else '')
        qty = safe_float(row[11])
        pu = safe_float(row[12])
        total = safe_float(row[13])
        if not desc:
            continue

        if lt == 'SERVIÇO':
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        if lt == 'SUBETAPA' and total > 0:
            d = normalize(desc)
            if 'esquadrias de aluminio' in d:
                esq.setdefault('aluminio_sub', {'valor': 0.0})
                esq['aluminio_sub']['valor'] += total
            if 'guarda corpo' in d or 'guarda-corpo' in d:
                esq.setdefault('guarda_corpo_sub', {'valor': 0.0})
                esq['guarda_corpo_sub']['valor'] += total
            if 'brise' in d:
                esq.setdefault('brises_sub', {'valor': 0.0})
                esq['brises_sub']['valor'] += total
            if 'porta' in d and ('madeira' in d or 'mdf' in d):
                esq.setdefault('portas_madeira_sub', {'valor': 0.0})
                esq['portas_madeira_sub']['valor'] += total
            if 'corta fogo' in d or 'corta-fogo' in d:
                esq.setdefault('portas_corta_fogo_sub', {'valor': 0.0})
                esq['portas_corta_fogo_sub']['valor'] += total
            if 'vidro' in d and 'pele' not in d and 'esquadria' not in d and 'alumin' not in d:
                esq.setdefault('vidros_sub', {'valor': 0.0})
                esq['vidros_sub']['valor'] += total
            if 'pele de vidro' in d or 'pele vidro' in d:
                esq.setdefault('pele_vidro_sub', {'valor': 0.0})
                esq['pele_vidro_sub']['valor'] += total
            # Fachada subetapa
            if ('reboco' in d or 'argamassado' in d or 'argamassa' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_ger_executivo_alt(wb, sheet_name):
    """GDI-style Ger_Executivo: B=Atalho, C=Item, D=Descrição, E=Unidade, F=Quant, G=Preço un, H=Total.
    Also handles similar layouts where Item code is in col C (index 2).
    """
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}
    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=12, values_only=True):
        if len(row) < 8:
            continue
        item_code = safe_str(row[2] if row[2] else '')
        desc = safe_str(row[3] if row[3] else '')
        unit = safe_str(row[4] if row[4] else '')
        qty = safe_float(row[5])
        pu = safe_float(row[6])
        total = safe_float(row[7])
        if not desc:
            continue

        parts = [p for p in item_code.split('.') if p.strip()]

        # Service level (3+ parts like "08.001.001")
        if len(parts) >= 3 and (total > 0 or (pu > 0 and qty > 0)):
            item_total = total if total > 0 else pu * qty
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, item_total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, item_total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        # Subetapa (2 parts like "08.001")
        if len(parts) == 2 and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'alumin' in d:
                esq.setdefault('aluminio_sub', {'valor': 0.0})
                esq['aluminio_sub']['valor'] += total
            if 'guarda corpo' in d or 'guarda-corpo' in d:
                esq.setdefault('guarda_corpo_sub', {'valor': 0.0})
                esq['guarda_corpo_sub']['valor'] += total
            if 'porta' in d and ('madeira' in d or 'mdf' in d):
                esq.setdefault('portas_madeira_sub', {'valor': 0.0})
                esq['portas_madeira_sub']['valor'] += total
            if 'corta fogo' in d or 'corta-fogo' in d:
                esq.setdefault('portas_corta_fogo_sub', {'valor': 0.0})
                esq['portas_corta_fogo_sub']['valor'] += total
            if ('reboco' in d or 'argamassado' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total

    return esq, fach, acab


def extract_relatorio_flat(wb, sheet_name):
    """Relatório/Sienge flat: A=code, B=desc, C=un, D=qty, E=pu, F=total"""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}
    # Detect column layout from header row
    col_map = None
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=30, values_only=True):
        if not row or not row[0]:
            continue
        r0 = str(row[0]).strip().lower()
        if r0 in ('código', 'codigo'):
            # Find which columns have what
            headers = [safe_str(c).lower() for c in row[:30] if c]
            # Try to find 'Un.', 'Quantidade', 'Preço unitário', 'Preço total'
            un_idx = desc_idx = qty_idx = pu_idx = total_idx = None
            for ci, h in enumerate(row[:30]):
                hs = normalize(safe_str(h))
                if 'descri' in hs:
                    desc_idx = ci
                elif hs in ('un.', 'un', 'und', 'unid', 'unidade'):
                    un_idx = ci
                elif 'quantidade' in hs:
                    qty_idx = ci
                elif 'unitario' in hs or ('unit' in hs and 'prec' in hs):
                    pu_idx = ci
                elif 'total' in hs and 'prec' in hs:
                    total_idx = ci
                elif total_idx is None and 'total' in hs:
                    total_idx = ci
            col_map = {
                'code': 0,
                'desc': desc_idx or 1,
                'unit': un_idx,
                'qty': qty_idx,
                'pu': pu_idx,
                'total': total_idx or 2,
            }
            break

    if not col_map:
        # Default: A=code, B=desc, C=un, D=qty, E=pu, F=total
        col_map = {'code': 0, 'desc': 1, 'unit': 2, 'qty': 3, 'pu': 4, 'total': 5}

    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=30, values_only=True):
        code = safe_str(row[col_map['code']] if col_map['code'] is not None and col_map['code'] < len(row) else '')
        desc = safe_str(row[col_map['desc']] if col_map['desc'] is not None and col_map['desc'] < len(row) else '')
        unit = safe_str(row[col_map['unit']] if col_map['unit'] is not None and col_map['unit'] < len(row) else '') if col_map['unit'] is not None else ''
        qty = safe_float(row[col_map['qty']] if col_map['qty'] is not None and col_map['qty'] < len(row) else 0)
        pu = safe_float(row[col_map['pu']] if col_map['pu'] is not None and col_map['pu'] < len(row) else 0)
        total = safe_float(row[col_map['total']] if col_map['total'] is not None and col_map['total'] < len(row) else 0)
        if not desc:
            continue

        code_parts = [p for p in code.strip().split('.') if p.strip()]
        is_service = len(code_parts) >= 4

        if is_service and (total > 0 or (pu > 0 and qty > 0)):
            item_total = total if total > 0 else pu * qty
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, item_total, qty, pu, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, item_total)
            if pu > 0 and qty > 0:
                cat = classify_acabamento(desc, unit)
                if cat:
                    _add_acab(acab, cat, pu, qty, desc, unit)

        # Subetapa level (3 parts)
        if len(code_parts) == 3 and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'alumin' in d:
                esq.setdefault('aluminio_sub', {'valor': 0.0})
                esq['aluminio_sub']['valor'] += total
            elif 'guarda corpo' in d or 'guarda-corpo' in d:
                esq.setdefault('guarda_corpo_sub', {'valor': 0.0})
                esq['guarda_corpo_sub']['valor'] += total
            elif 'brise' in d:
                esq.setdefault('brises_sub', {'valor': 0.0})
                esq['brises_sub']['valor'] += total
            elif 'esquadria' in d and ('madeira' in d or 'mdf' in d):
                esq.setdefault('portas_madeira_sub', {'valor': 0.0})
                esq['portas_madeira_sub']['valor'] += total
            elif 'corta fogo' in d or 'corta-fogo' in d:
                esq.setdefault('portas_corta_fogo_sub', {'valor': 0.0})
                esq['portas_corta_fogo_sub']['valor'] += total

            if ('reboco' in d or 'argamassado' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('pastilha' in d or 'porcelanato' in d or 'ceramica' in d or 'acabamento' in d) and ('fachada' in d or 'externo' in d):
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_sienge_full(wb, sheet_name):
    """Sienge full export: B=code (with spaces), D=desc, F=unit, J or K=value.
    Lines alternate: service line then insumo line."""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    for row in ws.iter_rows(min_row=1, max_row=10000, max_col=15, values_only=True):
        if len(row) < 7:
            continue
        code_raw = safe_str(row[1] if row[1] else '')
        desc = safe_str(row[4] if len(row) > 4 and row[4] else '')
        if not desc:
            desc = safe_str(row[3] if len(row) > 3 and row[3] else '')
        unit = safe_str(row[6] if len(row) > 6 and row[6] else '')
        val1 = safe_float(row[10] if len(row) > 10 else 0)
        val2 = safe_float(row[11] if len(row) > 11 else 0)
        total = max(val1, val2) if val1 > 0 or val2 > 0 else 0

        if not desc or not code_raw:
            continue

        # Parse code structure
        code_clean = code_raw.strip()
        # Sienge codes: "02.13.001.0001" with spaces
        parts = [p for p in code_clean.split('.') if p.strip()]

        if len(parts) >= 4 and total > 0:
            cat = classify_esquadria(desc)
            if cat:
                _add_esq(esq, cat, total, 0, 0, desc)
            cat = classify_fachada(desc)
            if cat:
                _add_fach(fach, cat, total)

        # Subetapa (3 parts)
        if len(parts) == 3 and total > 0:
            d = normalize(desc)
            if 'esquadria' in d and 'metalic' in d:
                esq.setdefault('aluminio_sub', {'valor': 0.0})
                esq['aluminio_sub']['valor'] += total
            if 'esquadria' in d and 'madeira' in d:
                esq.setdefault('portas_madeira_sub', {'valor': 0.0})
                esq['portas_madeira_sub']['valor'] += total
            if ('reboco' in d or 'argamassado' in d) and 'fachada' in d:
                fach.setdefault('reboco_externo_sub', {'valor': 0.0})
                fach['reboco_externo_sub']['valor'] += total
            if ('pintura' in d or 'textura' in d) and 'fachada' in d:
                fach.setdefault('textura_pintura_ext_sub', {'valor': 0.0})
                fach['textura_pintura_ext_sub']['valor'] += total
            if ('acabamento' in d or 'pastilha' in d or 'ceramica' in d) and 'fachada' in d:
                fach.setdefault('pastilha_porcelanato_sub', {'valor': 0.0})
                fach['pastilha_porcelanato_sub']['valor'] += total

    return esq, fach, acab


def extract_named_esquadrias_sheet(wb, sheet_name):
    """Extract from a dedicated Esquadrias sheet (like GDI/Muller format).
    Rows: code, desc, material, type, width, height, qty, contramarco, area, cu
    """
    ws = wb[sheet_name]
    esq = {}
    in_section = None

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=15, values_only=True):
        if len(row) < 6:
            continue

        # Detect section headers
        r0 = safe_str(row[0] if row[0] else '')
        r1 = safe_str(row[1] if row[1] else '')
        d_r0 = normalize(r0)

        # Section headers often in col A or B
        if 'portas unitarias' in d_r0 or 'geral portas' in d_r0:
            in_section = 'portas'
            # Row may also have totals
            area = safe_float(row[8] if len(row) > 8 else 0)
            cu = safe_float(row[9] if len(row) > 9 else 0)
            total_val = safe_float(row[10] if len(row) > 10 else 0)
            continue

        if 'esquadrias de aluminio' in d_r0 or 'aluminio' in d_r0 and 'esquadria' in d_r0:
            in_section = 'aluminio'
            area = safe_float(row[8] if len(row) > 8 else 0)
            cu = safe_float(row[9] if len(row) > 9 else 0)
            total_val = safe_float(row[10] if len(row) > 10 else 0)
            if total_val > 0:
                _add_esq(esq, 'aluminio', total_val, area, cu, 'Esquadrias de alumínio (seção)')
            continue

        if 'pele' in d_r0 and ('vidro' in d_r0 or len(d_r0) < 10):
            in_section = 'pele'
            area = safe_float(row[8] if len(row) > 8 else 0)
            cu = safe_float(row[9] if len(row) > 9 else 0)
            total_val = safe_float(row[10] if len(row) > 10 else 0)
            if total_val > 0:
                _add_esq(esq, 'pele_vidro', total_val, area, cu, 'Pele de vidro (seção)')
            continue

        if 'guarda corpo' in d_r0 or 'guarda-corpo' in d_r0 or 'corrimao' in d_r0:
            in_section = 'guarda_corpo'
            area = safe_float(row[8] if len(row) > 8 else 0)
            total_val = safe_float(row[10] if len(row) > 10 else 0)
            if total_val > 0:
                _add_esq(esq, 'guarda_corpo', total_val, area, 0, 'Guarda corpo (seção)')
            continue

        # Individual items
        material = safe_str(row[2] if len(row) > 2 and row[2] else '')
        if not material and len(row) > 5:
            material = safe_str(row[5] if row[5] else '')
        desc = safe_str(row[1] if row[1] else '')
        qty = safe_float(row[6] if len(row) > 6 else 0)
        area = safe_float(row[8] if len(row) > 8 else 0)
        cu = safe_float(row[9] if len(row) > 9 else 0)

        if not desc or qty <= 0:
            continue

        mat_n = normalize(material)
        desc_n = normalize(desc)

        if 'corta' in desc_n and 'fogo' in desc_n:
            item_total = cu * area if cu > 0 and area > 0 else (cu * qty if cu > 0 else 0)
            _add_esq(esq, 'portas_corta_fogo', item_total, qty, cu, desc)
        elif 'madeira' in mat_n or 'mdf' in mat_n:
            item_total = cu * area if cu > 0 and area > 0 else (cu * qty if cu > 0 else 0)
            _add_esq(esq, 'portas_madeira', item_total, qty, cu, desc)
        elif 'aluminio' in mat_n or 'alumínio' in mat_n or 'alumin' in mat_n:
            # Already counted in section total, don't double count
            pass
        elif 'vidro' in mat_n:
            item_total = cu * area if cu > 0 and area > 0 else (cu * qty if cu > 0 else 0)
            _add_esq(esq, 'vidros', item_total, qty, cu, desc)

    return esq


def extract_muller_esquadrias(wb):
    """Extract from Muller-style sheets: Esquadrias, Esquadrias de Alum, Esquadrias de Made, Porta corta fogo, Corrimão"""
    esq = {}
    sheets = wb.sheetnames

    # Esquadrias de Alum
    for sname in sheets:
        sn = normalize(sname)
        if 'esquadrias de alum' in sn or 'esquadria alum' in sn:
            total_m2 = 0
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                if len(row) < 4:
                    continue
                elem = safe_str(row[1] if row[1] else '')
                unit = safe_str(row[2] if row[2] else '')
                val = safe_float(row[3])
                en = normalize(elem)
                if 'esquadria' in en and 'alumin' in en and 'm' in normalize(unit) and val > 0:
                    total_m2 += val
            if total_m2 > 0:
                esq['aluminio'] = {'valor': None, 'qtd': round(total_m2, 2)}

    # Esquadrias de Madeira
    for sname in sheets:
        sn = normalize(sname)
        if 'esquadrias de made' in sn or 'esquadria madeira' in sn:
            total_un = 0
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                if len(row) < 4:
                    continue
                elem = safe_str(row[1] if row[1] else '')
                unit = safe_str(row[2] if row[2] else '')
                val = safe_float(row[3])
                if 'porta' in normalize(elem) and val > 0:
                    total_un += val
            if total_un > 0:
                esq['portas_madeira'] = {'valor': None, 'qtd': round(total_un, 2)}

    # Porta corta fogo
    for sname in sheets:
        sn = normalize(sname)
        if 'corta fogo' in sn or 'corta-fogo' in sn:
            total_un = 0
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                if len(row) < 4:
                    continue
                elem = safe_str(row[1] if row[1] else '')
                val = safe_float(row[3])
                if 'corta fogo' in normalize(elem) and val > 0:
                    total_un += val
            if total_un > 0:
                esq['portas_corta_fogo'] = {'valor': None, 'qtd': round(total_un, 2)}

    # Corrimão / Guarda corpo
    for sname in sheets:
        sn = normalize(sname)
        if sn in ('corrimao', 'guarda corpo', 'guarda-corpo'):
            ws = wb[sname]
            # Just note existence
            esq.setdefault('guarda_corpo', {'valor': None, 'qtd': None})

    return esq


def extract_muller_fachada(wb):
    """Extract from Muller Fachada sheet: BIM quantitativos format"""
    fach = {}
    ws = None
    for sname in wb.sheetnames:
        if normalize(sname) == 'fachada':
            ws = wb[sname]
            break
    if not ws:
        return fach

    for row in ws.iter_rows(min_row=1, max_row=500, max_col=15, values_only=True):
        if len(row) < 5:
            continue
        desc = safe_str(row[1] if row[1] else '')
        unit = safe_str(row[2] if row[2] else '')
        qty = safe_float(row[3])
        rev = safe_str(row[4] if len(row) > 4 and row[4] else '')

        d = normalize(desc)
        if qty > 0 and ('m' in normalize(unit)):
            if 'textura' in d or 'pintura' in d:
                fach.setdefault('textura_pintura_ext', {'valor': 0.0, 'qtd': 0.0})
                fach['textura_pintura_ext']['qtd'] += qty
            elif 'reboco' in d or 'chapisco' in d or 'argamassa' in d:
                fach.setdefault('reboco_externo', {'valor': 0.0, 'qtd': 0.0})
                fach['reboco_externo']['qtd'] += qty
            elif 'pastilha' in d or 'porcelanato' in d or 'ceramica' in d:
                fach.setdefault('pastilha_porcelanato', {'valor': 0.0, 'qtd': 0.0})
                fach['pastilha_porcelanato']['qtd'] += qty

    return fach


def extract_muller_acabamentos(wb):
    """Extract acabamentos from Muller-style Piso/Acabamento sheets"""
    acab = {}
    sheets = wb.sheetnames

    for sname in sheets:
        sn = normalize(sname)
        ws = wb[sname]
        if sn in ('piso', 'soleira', 'rodape', 'acabamento parede', 'acabamento teto',
                   'pintura', 'forro', 'drywall'):
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                if len(row) < 4:
                    continue
                elem = safe_str(row[1] if row[1] else '')
                unit = safe_str(row[2] if row[2] else '')
                qty = safe_float(row[3])
                d = normalize(elem)
                if qty <= 0:
                    continue
                # We only have quantities, no PUs, so skip PU extraction
                # But record what we find as notes
    return acab


def extract_kirchner_fachada(wb):
    """Extract from Kirchner FACHADA sheet (matrix format)"""
    fach = {}
    ws = None
    for sname in wb.sheetnames:
        if normalize(sname) == 'fachada':
            ws = wb[sname]
            break
    if not ws:
        return fach

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        r1 = safe_str(row[1] if len(row) > 1 and row[1] else '')
        if r1 == 'TOTAL':
            # Columns: C=ceramico_amadeirado, D=pastilha, E=ceramico_marmorizado, F=brise
            c_val = safe_float(row[2] if len(row) > 2 else 0)
            d_val = safe_float(row[3] if len(row) > 3 else 0)
            e_val = safe_float(row[4] if len(row) > 4 else 0)
            f_val = safe_float(row[5] if len(row) > 5 else 0)
            if c_val + d_val + e_val > 0:
                fach['pastilha_porcelanato'] = {'valor': None, 'qtd_m2': round(c_val + d_val + e_val, 2)}
            if f_val > 0:
                # Brises on fachada - record as note
                fach['_brises_m2'] = round(f_val, 2)
            break
    return fach


def extract_kirchner_revestimentos(wb):
    """Extract acabamentos from Kirchner REVESTIMENTOS sheet (summary of areas)"""
    acab = {}
    ws = None
    for sname in wb.sheetnames:
        if normalize(sname) == 'revestimentos':
            ws = wb[sname]
            break
    if not ws:
        return acab

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        r1 = safe_str(row[1] if len(row) > 1 and row[1] else '')
        if r1 == 'TOTAL':
            # C = reboco alvenaria, D = reboco estrutura, E = total reboco, F = azulejo, G = pintura
            total_reboco = safe_float(row[4] if len(row) > 4 else 0)
            azulejo = safe_float(row[5] if len(row) > 5 else 0)
            pintura = safe_float(row[6] if len(row) > 6 else 0)
            # Only m2 quantities, no PUs
            break
    return acab


def extract_kirchner_pisos(wb):
    """Extract from Kirchner PISOS sheet"""
    acab = {}
    ws = None
    for sname in wb.sheetnames:
        if normalize(sname) == 'pisos':
            ws = wb[sname]
            break
    if not ws:
        return acab
    # Usually a BIM quantitativo - m2 quantities, no PUs
    return acab


def extract_mg3_esquadrias(wb):
    """Extract from MG3 tabela sheets: portas, janelas, brises, painel vidro"""
    esq = {}
    sheets = wb.sheetnames

    for sname in sheets:
        sn = normalize(sname)

        # Portas
        if 'tabela de portas' in sn:
            ws = wb[sname]
            total_madeira = 0
            total_alum = 0
            total_pcf = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                desc = safe_str(row[1] if len(row) > 1 and row[1] else '')
                qty = safe_float(row[4] if len(row) > 4 else 0)
                if qty <= 0:
                    continue
                d = normalize(desc)
                if 'corta-fogo' in d or 'corta fogo' in d:
                    total_pcf += qty
                elif 'mdf' in d or 'madeira' in d:
                    total_madeira += qty
                elif 'alumin' in d:
                    total_alum += qty
            if total_madeira > 0:
                esq['portas_madeira'] = {'valor': None, 'qtd': round(total_madeira, 2)}
            if total_pcf > 0:
                esq['portas_corta_fogo'] = {'valor': None, 'qtd': round(total_pcf, 2)}

        # Janelas
        if 'tabela de janelas' in sn:
            ws = wb[sname]
            total_janelas = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                qty = safe_float(row[5] if len(row) > 5 else 0)
                if qty > 0:
                    code = safe_str(row[0] if row[0] else '')
                    if code and code[0] in ('J', 'G', 'V', 'C', 'P') and code != 'Marca de tipo':
                        total_janelas += qty
            if total_janelas > 0:
                esq.setdefault('aluminio', {'valor': None, 'qtd': 0})
                esq['aluminio']['qtd'] = esq['aluminio'].get('qtd', 0) + round(total_janelas, 2)

        # Brises
        if 'tabela de brises' in sn:
            ws = wb[sname]
            total_brises = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                qty = safe_float(row[5] if len(row) > 5 else 0)
                if qty > 0:
                    code = safe_str(row[0] if row[0] else '')
                    if code and not code.startswith('Marca') and not code.startswith('Código'):
                        total_brises += qty
            if total_brises > 0:
                esq['brises'] = {'valor': None, 'qtd': round(total_brises, 2)}

        # Painel de vidro
        if 'tabela de painel de vidro' in sn:
            ws = wb[sname]
            total_pele = 0
            for row in ws.iter_rows(min_row=1, max_row=500, max_col=10, values_only=True):
                qty = safe_float(row[5] if len(row) > 5 else 0)
                if qty > 0:
                    code = safe_str(row[0] if row[0] else '')
                    if code and not code.startswith('Marca') and not code.startswith('Código'):
                        total_pele += qty
            if total_pele > 0:
                esq['pele_vidro'] = {'valor': None, 'qtd': round(total_pele, 2)}

    return esq


def extract_mg3_fachada(wb):
    """Extract from MG3 B05. TABELA REVEST FACHADAS"""
    fach = {}
    ws = None
    for sname in wb.sheetnames:
        if 'tabela revest fachada' in normalize(sname):
            ws = wb[sname]
            break
    if not ws:
        return fach

    total_area = 0
    for row in ws.iter_rows(min_row=1, max_row=1000, max_col=10, values_only=True):
        area = safe_float(row[3] if len(row) > 3 else 0)
        if area > 0:
            total_area += area
    if total_area > 0:
        fach['textura_pintura_ext'] = {'valor': None, 'qtd_m2': round(total_area, 2)}
    return fach


def extract_summary_only(wb, sheet_name):
    """Extract from summary-only sheets (etapa + value). Only gets _total."""
    ws = wb[sheet_name]
    esq, fach, acab = {}, {}, {}

    for row in ws.iter_rows(min_row=1, max_row=50, max_col=10, values_only=True):
        if len(row) < 2:
            continue
        # Try various column combos
        for desc_idx in range(min(5, len(row))):
            desc = safe_str(row[desc_idx] if row[desc_idx] else '')
            if desc and len(desc) > 5:
                d = normalize(desc)
                for val_idx in range(min(6, len(row))):
                    if val_idx == desc_idx:
                        continue
                    total = safe_float(row[val_idx])
                    if total > 1000:
                        if 'esquadria' in d:
                            esq['_total_esquadrias'] = esq.get('_total_esquadrias', 0) + total
                        if 'fachada' in d or ('revestimento' in d and 'externo' in d):
                            fach['_total_fachada'] = fach.get('_total_fachada', 0) + total
                        break
                break
    return esq, fach, acab


# ============================================================
# Accumulator helpers
# ============================================================

def _add_esq(esq, cat, total, qty, pu, desc):
    if cat not in esq:
        esq[cat] = {'valor': 0.0, 'qtd': 0.0, 'pu_sum': 0.0, 'pu_count': 0}
    esq[cat]['valor'] += total
    esq[cat]['qtd'] += qty
    if pu > 0 and qty > 0:
        esq[cat]['pu_sum'] += pu * qty
        esq[cat]['pu_count'] += qty


def _add_fach(fach, cat, total):
    if cat not in fach:
        fach[cat] = {'valor': 0.0}
    fach[cat]['valor'] += total


def _add_acab(acab, cat, pu, qty, desc, unit):
    if cat not in acab:
        acab[cat] = []
    acab[cat].append({'pu': pu, 'qty': qty, 'desc': desc, 'unit': unit})


# ============================================================
# Build final output
# ============================================================

def build_esquadrias_detail(raw):
    result = {}
    categories = ['aluminio', 'guarda_corpo', 'portas_madeira', 'portas_corta_fogo',
                   'vidros', 'pele_vidro', 'brises']
    for cat in categories:
        if cat in raw:
            data = raw[cat]
            entry = {}
            if isinstance(data, dict):
                val = data.get('valor') or 0
                if val > 0:
                    entry['valor'] = round(val, 2)
                qtd = data.get('qtd') or 0
                if qtd > 0:
                    entry['qtd'] = round(qtd, 2)
                pu_count = data.get('pu_count') or 0
                if pu_count > 0:
                    entry['pu'] = round((data.get('pu_sum') or 0) / pu_count, 2)
                qtd_m2 = data.get('qtd_m2') or 0
                if qtd_m2 > 0:
                    entry['qtd_m2'] = round(qtd_m2, 2)
            if entry:
                result[cat] = entry
        elif f'{cat}_sub' in raw:
            data = raw[f'{cat}_sub']
            if isinstance(data, dict) and data.get('valor', 0) > 0:
                result[cat] = {'valor': round(data['valor'], 2)}
    if '_total_esquadrias' in raw:
        result['_total'] = round(raw['_total_esquadrias'], 2)
    return result if result else None


def build_fachada_detail(raw):
    result = {}
    categories = ['reboco_externo', 'textura_pintura_ext', 'pastilha_porcelanato', 'juntas']
    for cat in categories:
        if cat in raw:
            entry = {}
            val = raw[cat].get('valor', 0) or 0
            if val > 0:
                entry['valor'] = round(val, 2)
            qty = raw[cat].get('qtd', 0) or 0
            if qty > 0:
                entry['qtd_m2'] = round(qty, 2)
            qtm = raw[cat].get('qtd_m2', 0) or 0
            if qtm > 0:
                entry['qtd_m2'] = round(qtm, 2)
            if entry:
                result[cat] = entry
        elif f'{cat}_sub' in raw:
            data = raw[f'{cat}_sub']
            val = data.get('valor', 0) or 0 if isinstance(data, dict) else 0
            if val > 0:
                result[cat] = {'valor': round(val, 2)}
    if '_total_fachada' in raw:
        result['_total'] = round(raw['_total_fachada'], 2)
    if '_brises_m2' in raw:
        result['_brises_m2'] = raw['_brises_m2']
    return result if result else None


def build_acabamentos_pus(raw):
    result = {}
    for cat, items in raw.items():
        if not items:
            continue
        total_weighted = sum(i['pu'] * i['qty'] for i in items)
        total_qty = sum(i['qty'] for i in items)
        if total_qty > 0:
            result[cat] = round(total_weighted / total_qty, 2)
    return result if result else None


# ============================================================
# Process each project
# ============================================================

def process_project(proj):
    slug = proj['slug']
    path = proj['path']

    if not os.path.exists(path):
        return slug, None, None, None, f"File not found"

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return slug, None, None, None, f"Cannot open: {e}"

    sheets = wb.sheetnames
    all_esq, all_fach, all_acab = {}, {}, {}

    try:
        # === Strategy 1: Ger_Executivo ===
        for sname in sheets:
            if sname.lower().replace(' ', '_').startswith('ger_executivo'):
                try:
                    # Check if standard layout (col H has level types) or alt (col C has Item code)
                    ws = wb[sname]
                    is_standard = False
                    is_alt = False
                    for row in ws.iter_rows(min_row=1, max_row=30, max_col=15, values_only=True):
                        if row and len(row) > 7 and row[7]:
                            lt = safe_str(row[7])
                            if lt in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
                                is_standard = True
                                break
                        # Check for alt layout: row[2] has code like "01.001"
                        if row and len(row) > 3 and row[2]:
                            item = safe_str(row[2])
                            parts = [p for p in item.split('.') if p.strip()]
                            if len(parts) >= 2 and all(1 <= len(p.strip()) <= 4 and p.strip().isdigit() for p in parts):
                                is_alt = True
                    if is_standard:
                        e, f, a = extract_ger_executivo(wb, sname)
                    elif is_alt:
                        e, f, a = extract_ger_executivo_alt(wb, sname)
                    else:
                        e, f, a = extract_ger_executivo(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                except:
                    pass

        # === Strategy 2: Relatório / Sienge N2 ===
        for sname in sheets:
            if 'relat' in sname.lower().strip():
                try:
                    e, f, a = extract_relatorio_flat(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                except:
                    pass

        # === Strategy 3: ORÇAMENTO_EXECUTIVO ===
        for sname in sheets:
            sn = sname.strip()
            if sn in ('ORÇAMENTO_EXECUTIVO', 'Orçamento_Executivo', 'ORCAMENTO_EXECUTIVO',
                       'ORÇAMENTO_EXECUTIVO_TOTAL', 'ORÇAMENTO_EXECUTIVO_EDIFICAÇÕES'):
                try:
                    # Check if this is summary-only
                    ws = wb[sname]
                    first_rows = []
                    for row in ws.iter_rows(min_row=1, max_row=5, max_col=10, values_only=True):
                        first_rows.append(row)
                    is_summary = False
                    for row in first_rows:
                        for ci in range(min(3, len(row))):
                            if row[ci] and str(row[ci]).strip() == 'ETAPA':
                                is_summary = True
                                break
                        if is_summary:
                            break
                    if is_summary:
                        e, f, a = extract_summary_only(wb, sname)
                    else:
                        e, f, a = extract_relatorio_flat(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                except:
                    pass

        # === Strategy 4: Gerenciamento executivo / Executivo ===
        for sname in sheets:
            sn = sname.strip().lower()
            if 'gerenciamento executivo' in sn or sn == 'executivo':
                try:
                    ws = wb[sname]
                    is_standard = False
                    is_alt = False
                    for row in ws.iter_rows(min_row=1, max_row=30, max_col=15, values_only=True):
                        if row and len(row) > 7 and row[7]:
                            lt = safe_str(row[7])
                            if lt in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
                                is_standard = True
                                break
                        if row and len(row) > 3 and row[2]:
                            item = safe_str(row[2])
                            parts = [p for p in item.split('.') if p.strip()]
                            if len(parts) >= 2 and all(1 <= len(p.strip()) <= 4 and p.strip().isdigit() for p in parts):
                                is_alt = True
                    if is_standard:
                        e, f, a = extract_ger_executivo(wb, sname)
                    elif is_alt:
                        e, f, a = extract_ger_executivo_alt(wb, sname)
                    else:
                        e, f, a = extract_summary_only(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                except:
                    pass

        # === Strategy 5: Gerenciamento_Exec ===
        for sname in sheets:
            if 'gerenciamento_exec' in sname.lower().strip():
                try:
                    ws = wb[sname]
                    is_standard = False
                    is_alt = False
                    for row in ws.iter_rows(min_row=1, max_row=30, max_col=15, values_only=True):
                        if row and len(row) > 7 and row[7]:
                            lt = safe_str(row[7])
                            if lt in ('SERVIÇO', 'SUBETAPA', 'ETAPA', 'CÉLULA CONSTRUTIVA'):
                                is_standard = True
                                break
                        if row and len(row) > 3 and row[2]:
                            item = safe_str(row[2])
                            parts = [p for p in item.split('.') if p.strip()]
                            if len(parts) >= 2 and all(1 <= len(p.strip()) <= 4 and p.strip().isdigit() for p in parts):
                                is_alt = True
                    if is_standard:
                        e, f, a = extract_ger_executivo(wb, sname)
                    elif is_alt:
                        e, f, a = extract_ger_executivo_alt(wb, sname)
                    else:
                        e, f, a = extract_summary_only(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                except:
                    pass

        # === Strategy 6: Named Esquadrias sheets ===
        for sname in sheets:
            sn = normalize(sname)
            if sn == 'esquadrias' or sn.startswith('esquadrias '):
                try:
                    e = extract_named_esquadrias_sheet(wb, sname)
                    _merge(all_esq, e)
                except:
                    pass

        # === Strategy 7: Muller-style named sheets ===
        has_muller_sheets = any('esquadrias de alum' in normalize(s) for s in sheets)
        if has_muller_sheets:
            try:
                e = extract_muller_esquadrias(wb)
                _merge(all_esq, e)
            except:
                pass
            try:
                f = extract_muller_fachada(wb)
                _merge_fach(all_fach, f)
            except:
                pass

        # === Strategy 8: MG3 tabela sheets ===
        has_mg3 = any('tabela de portas' in normalize(s) for s in sheets)
        if has_mg3:
            try:
                e = extract_mg3_esquadrias(wb)
                _merge(all_esq, e)
            except:
                pass
            try:
                f = extract_mg3_fachada(wb)
                _merge_fach(all_fach, f)
            except:
                pass

        # === Strategy 9: Kirchner-style named sheets ===
        has_kirchner = any(normalize(s) == 'fachada' for s in sheets) and any(normalize(s) == 'revestimentos' for s in sheets)
        if has_kirchner:
            try:
                f = extract_kirchner_fachada(wb)
                _merge_fach(all_fach, f)
            except:
                pass

        # === Strategy 10: Sienge full export (Sheet1 with B=code) ===
        if 'Sheet1' in sheets and not all_esq and not all_fach:
            try:
                e, f, a = extract_sienge_full(wb, 'Sheet1')
                _merge(all_esq, e)
                _merge_fach(all_fach, f)
                _merge_acab(all_acab, a)
            except:
                pass

        # === Strategy 11: Orçamento Resumo ===
        for sname in sheets:
            if 'orcamento resumo' in normalize(sname) or 'orçamento resumo' in sname.lower():
                try:
                    e, f, a = extract_summary_only(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                except:
                    pass

        # === Strategy 12: CUSTOS DIRETOS (Kirchner) ===
        for sname in sheets:
            if normalize(sname) == 'custos diretos':
                try:
                    e, f, a = extract_relatorio_flat(wb, sname)
                    _merge(all_esq, e)
                    _merge_fach(all_fach, f)
                    _merge_acab(all_acab, a)
                except:
                    pass

        # === Strategy 13: Resumo sheet as fallback ===
        if not all_esq and not all_fach:
            for sname in sheets:
                if normalize(sname) in ('resumo', 'resumo geral'):
                    try:
                        e, f, a = extract_summary_only(wb, sname)
                        _merge(all_esq, e)
                        _merge_fach(all_fach, f)
                    except:
                        pass

    except Exception as ex:
        wb.close()
        return slug, None, None, None, f"Error: {traceback.format_exc()}"

    wb.close()

    esq_detail = build_esquadrias_detail(all_esq)
    fach_detail = build_fachada_detail(all_fach)
    acab_pus = build_acabamentos_pus(all_acab)

    return slug, esq_detail, fach_detail, acab_pus, None


def _merge(target, source):
    for k, v in source.items():
        if k not in target:
            target[k] = v
        elif isinstance(v, dict) and isinstance(target[k], dict):
            for dk, dv in v.items():
                if dk in target[k] and isinstance(dv, (int, float)):
                    target[k][dk] = target[k].get(dk, 0) + dv
                elif dk not in target[k]:
                    target[k][dk] = dv


def _merge_fach(target, source):
    for k, v in source.items():
        if k not in target:
            target[k] = v
        elif isinstance(v, dict) and isinstance(target[k], dict):
            for dk, dv in v.items():
                if dk in target[k] and isinstance(dv, (int, float)):
                    target[k][dk] = target[k].get(dk, 0) + dv
                elif dk not in target[k]:
                    target[k][dk] = dv
        elif isinstance(v, (int, float)):
            target[k] = max(target.get(k, 0), v)


def _merge_acab(target, source):
    for k, v in source.items():
        if k not in target:
            target[k] = v
        else:
            target[k].extend(v)


def update_json(slug, esq, fach, acab):
    idx_path = INDICES_DIR / f"{slug}.json"
    if idx_path.exists():
        with open(idx_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    else:
        data = {'slug': slug}

    # Always write all 3 fields (empty dict if no data found)
    data['esquadrias_detail'] = esq if esq is not None else data.get('esquadrias_detail', {})
    data['fachada_detail'] = fach if fach is not None else data.get('fachada_detail', {})
    if acab is not None:
        existing = data.get('acabamentos_pus', {})
        if existing:
            for k, v in acab.items():
                if k not in existing:
                    existing[k] = v
            data['acabamentos_pus'] = existing
        else:
            data['acabamentos_pus'] = acab
    elif 'acabamentos_pus' not in data:
        data['acabamentos_pus'] = {}

    with open(idx_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    results = {'success': 0, 'no_new': 0, 'error': 0}
    details = []

    for proj in BATCH:
        slug = proj['slug']
        print(f"Processing {slug}...", end=' ')

        slug_r, esq, fach, acab, err = process_project(proj)

        if err:
            print(f"ERROR: {err}")
            results['error'] += 1
            details.append({'slug': slug, 'status': 'error', 'error': err})
            continue

        if not esq and not fach and not acab:
            print("NO NEW DATA (writing empty fields)")
            results['no_new'] += 1
            details.append({'slug': slug, 'status': 'no_new'})
            update_json(slug, esq, fach, acab)  # Still write empty fields
            continue

        update_json(slug, esq, fach, acab)
        results['success'] += 1

        parts = []
        if esq:
            parts.append(f"esq={list(esq.keys())}")
        if fach:
            parts.append(f"fach={list(fach.keys())}")
        if acab:
            parts.append(f"acab={list(acab.keys())}")
        print(f"OK ({'; '.join(parts)})")
        details.append({'slug': slug, 'status': 'ok',
                        'esquadrias': list(esq.keys()) if esq else [],
                        'fachada': list(fach.keys()) if fach else [],
                        'acabamentos': list(acab.keys()) if acab else []})

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total processed: {len(BATCH)}")
    print(f"Success (new data): {results['success']}")
    print(f"No new data: {results['no_new']}")
    print(f"Errors: {results['error']}")

    print("\n--- Esquadrias detail coverage ---")
    esq_cats = {}
    for d in details:
        for cat in d.get('esquadrias', []):
            esq_cats[cat] = esq_cats.get(cat, 0) + 1
    for cat, count in sorted(esq_cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count} projects")

    print("\n--- Fachada detail coverage ---")
    fach_cats = {}
    for d in details:
        for cat in d.get('fachada', []):
            fach_cats[cat] = fach_cats.get(cat, 0) + 1
    for cat, count in sorted(fach_cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count} projects")

    print("\n--- Acabamentos PU coverage ---")
    acab_cats = {}
    for d in details:
        for cat in d.get('acabamentos', []):
            acab_cats[cat] = acab_cats.get(cat, 0) + 1
    for cat, count in sorted(acab_cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count} projects")


if __name__ == '__main__':
    main()
