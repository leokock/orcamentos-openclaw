#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract STRUCTURAL INDICES from construction budget spreadsheets.
Batch 1: projects 0-41.

Extracts: concrete (m3), steel (kg), formwork (m2), piles data
from multiple spreadsheet formats (CTN parametric, CTN Ger_Executivo,
Sienge Relatório, dedicated structural sheets).
"""
import json
import sys
import re
import os
import traceback
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# ============================================================
# Config
# ============================================================
BASE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base")
INDICES_DIR = BASE_DIR / "indices-executivo"
MAPPING_FILE = BASE_DIR / "_all_projects_mapping.json"
METADADOS_FILE = BASE_DIR / "projetos-metadados.json"

BATCH_START = 0
BATCH_END = 42  # exclusive

# ============================================================
# Helpers
# ============================================================

def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return 0.0

def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()

def normalize(s):
    if not s:
        return ""
    s = s.lower().strip()
    for a, b in {'á':'a','à':'a','â':'a','ã':'a','é':'e','ê':'e','è':'e',
                  'í':'i','î':'i','ó':'o','ô':'o','õ':'o','ò':'o',
                  'ú':'u','û':'u','ù':'u','ç':'c','ñ':'n'}.items():
        s = s.replace(a, b)
    return s

def row_text(row):
    """Build searchable text from a row."""
    return ' '.join(safe_str(v).lower() for v in row if v is not None)

def find_numeric_in_row(row, min_val=None, max_val=None):
    """Find first numeric value in row within bounds."""
    for v in row:
        if v is None:
            continue
        if isinstance(v, (int, float)) and v != 0:
            n = float(v)
            if min_val is not None and n < min_val:
                continue
            if max_val is not None and n > max_val:
                continue
            return n
    return None


# ============================================================
# Format 1: CTN Parametric Sheets (INFRAESTRUTURA / SUPRAESTRUTURA)
# ============================================================

def extract_from_parametric_sheets(wb, sheets):
    """
    Extract from dedicated parametric sheets like INFRAESTRUTURA, SUPRAESTRUTURA.
    These have descriptive rows with parameters and quantities.
    """
    result = {
        'concreto_infra_m3': 0,
        'concreto_supra_m3': 0,
        'aco_infra_kg': 0,
        'aco_supra_kg': 0,
        'forma_supra_m2': 0,
        'estacas_m': 0,
        'estacas_qtd': 0,
        'tipo_estaca': None,
        'fck_list': [],
        'mo_estrutural_total': 0,
    }

    for sname in sheets:
        ws = wb[sname]
        sn = normalize(sname)
        is_infra = 'infraestrutura' in sn or 'fundac' in sn
        is_supra = 'supraestrutura' in sn or 'estrutura' in sn and not is_infra

        rows = []
        try:
            for row in ws.iter_rows(max_row=200, max_col=20, values_only=True):
                rows.append(tuple(row))
        except:
            continue

        for i, row in enumerate(rows):
            rt = row_text(row)

            # Estacas / piles
            if any(k in rt for k in ['estaca', 'helice', 'hélice']):
                # Try to get pile type
                if result['tipo_estaca'] is None:
                    if 'helice' in rt or 'hélice' in rt:
                        result['tipo_estaca'] = 'hélice contínua'
                    elif 'raiz' in rt:
                        result['tipo_estaca'] = 'raiz'
                    elif 'franki' in rt:
                        result['tipo_estaca'] = 'franki'
                    elif 'pre-moldada' in rt or 'pré-moldada' in rt or 'pre moldada' in rt:
                        result['tipo_estaca'] = 'pré-moldada'

                # Look for estaca quantities
                for j, v in enumerate(row):
                    sv = safe_str(v).lower()
                    if 'un' in sv and j > 0:
                        # previous cell might be quantity
                        n = safe_float(row[j-1]) if j > 0 else 0
                        if 1 < n < 2000:
                            result['estacas_qtd'] += int(n)
                    if sv.endswith('m') or 'm ' in sv:
                        n = safe_float(row[j-1]) if j > 0 else 0
                        if 10 < n < 50000:
                            result['estacas_m'] += n

            # Concreto
            if 'concreto' in rt and ('m³' in rt or 'm3' in rt):
                for j, v in enumerate(row):
                    sv = safe_str(v)
                    if 'm³' in sv or 'm3' in sv.lower():
                        # The quantity is likely the cell before
                        n = safe_float(row[j-1]) if j > 0 else 0
                        if n <= 0:
                            # Or the cell with m³ might have a number nearby
                            for k in range(max(0, j-3), min(len(row), j+3)):
                                if isinstance(row[k], (int, float)) and row[k] > 1:
                                    n = float(row[k])
                                    break
                        if n > 0:
                            if is_infra:
                                result['concreto_infra_m3'] += n
                            else:
                                result['concreto_supra_m3'] += n

                # Extract fck
                for fck in re.findall(r'(\d+)\s*(?:mpa|MPa)', rt):
                    fck_val = f"{fck}MPa"
                    if fck_val not in result['fck_list']:
                        result['fck_list'].append(fck_val)

            # Armação / Aço
            if any(k in rt for k in ['armação', 'armacao', 'armadura', 'aço', 'aco']) and 'kg' in rt:
                for j, v in enumerate(row):
                    sv = safe_str(v)
                    if 'kg' in sv.lower():
                        n = safe_float(row[j-1]) if j > 0 else 0
                        if n <= 0:
                            for k in range(max(0, j-3), min(len(row), j+3)):
                                if isinstance(row[k], (int, float)) and row[k] > 100:
                                    n = float(row[k])
                                    break
                        if n > 0:
                            if is_infra:
                                result['aco_infra_kg'] += n
                            else:
                                result['aco_supra_kg'] += n

            # Forma
            if any(k in rt for k in ['forma', 'fôrma']) and ('m²' in rt or 'm2' in rt):
                for j, v in enumerate(row):
                    sv = safe_str(v)
                    if 'm²' in sv or 'm2' in sv.lower():
                        n = safe_float(row[j-1]) if j > 0 else 0
                        if n <= 0:
                            for k in range(max(0, j-3), min(len(row), j+3)):
                                if isinstance(row[k], (int, float)) and row[k] > 10:
                                    n = float(row[k])
                                    break
                        if n > 0:
                            result['forma_supra_m2'] += n

            # Mão de obra estrutural
            if any(k in rt for k in ['mão de obra', 'mao de obra', 'mo -']) and any(k in rt for k in ['estrut', 'concreto armado', 'supraestrutura']):
                for v in row:
                    if isinstance(v, (int, float)) and v > 10000:
                        result['mo_estrutural_total'] += float(v)
                        break

    return result


# ============================================================
# Format 2: CTN Ger_Executivo (detailed line items)
# ============================================================

def detect_ger_executivo_format(rows):
    """
    Detect column layout for Ger_Executivo sheets.
    CTN standard: col[1]=section_code, col[7]=type, col[8]=code, col[9]=desc, col[10]=unit, col[11]=qty, col[13]=total
    But column positions may vary. Detect by finding 'SERVIÇO' or 'ETAPA' type markers.
    """
    for i, row in enumerate(rows[:50]):
        for j, v in enumerate(row):
            if safe_str(v) == 'SERVIÇO' and j > 3:
                # Found the type column. Standard layout:
                # type=j, code=j+1, desc=j+2, unit=j+3, qty=j+4, price=j+5, total=j+6
                # section code is typically at j-6 or column 1
                return {
                    'section_col': 1,  # section code (1=movterra, 2=infra, 3=supra, etc.)
                    'type_col': j,
                    'code_col': j + 1,
                    'desc_col': j + 2,
                    'unit_col': j + 3,
                    'qty_col': j + 4,
                    'price_col': j + 5,
                    'total_col': j + 6,
                }
    return None


def extract_from_ger_executivo(wb, sheet_name):
    """
    Extract structural data from Ger_Executivo/ORCAMENTO style sheets.
    CTN format has fixed columns with section codes, type, code, description, unit, qty, price, total.
    Section 02 = INFRAESTRUTURA, Section 03 = SUPRAESTRUTURA.
    """
    ws = wb[sheet_name]
    result = {
        'concreto_infra_m3': 0,
        'concreto_supra_m3': 0,
        'aco_infra_kg': 0,
        'aco_supra_kg': 0,
        'forma_infra_m2': 0,
        'forma_supra_m2': 0,
        'estacas_m': 0,
        'estacas_qtd': 0,
        'tipo_estaca': None,
        'fck_list': [],
        'mo_estrutural_total': 0,
    }

    rows = []
    try:
        for row in ws.iter_rows(max_row=5000, max_col=30, values_only=True):
            rows.append(tuple(row))
    except:
        return result

    # Detect column layout
    fmt = detect_ger_executivo_format(rows)

    if fmt:
        # ── CTN structured format ──
        sc = fmt['section_col']
        dc = fmt['desc_col']
        uc = fmt['unit_col']
        qc = fmt['qty_col']
        tc = fmt['total_col']
        typc = fmt['type_col']

        for i, row in enumerate(rows):
            if len(row) <= max(dc, uc, qc):
                continue

            section_code = row[sc] if sc < len(row) else None
            row_type = safe_str(row[typc]) if typc < len(row) else ''
            desc = safe_str(row[dc]).lower() if dc < len(row) else ''
            unit = safe_str(row[uc]).lower() if uc < len(row) else ''
            qty = safe_float(row[qc]) if qc < len(row) else 0
            total_val = safe_float(row[tc]) if tc < len(row) else 0

            # Determine section
            try:
                sec_num = int(section_code) if section_code is not None else 0
            except (ValueError, TypeError):
                sec_num = 0

            is_infra = sec_num == 2
            is_supra = sec_num == 3

            if not is_infra and not is_supra:
                continue

            if row_type != 'SERVIÇO' and row_type != 'SERVICO':
                # For ETAPA/SUBETAPA, track MO totals
                if row_type in ['ETAPA', 'SUBETAPA']:
                    if any(k in desc for k in ['mão de obra', 'mao de obra']):
                        if total_val > 0:
                            result['mo_estrutural_total'] += total_val
                continue

            if qty <= 0:
                continue

            # Classify by unit and description
            is_concreto = any(k in desc for k in ['concreto', 'conc.', 'bombeável', 'bombeavel'])
            is_aco = any(k in desc for k in ['aço', 'aco', 'armação', 'armacao', 'armadura',
                                              'ca-50', 'ca-60', 'ca50', 'ca60', 'protensão', 'protensao'])
            is_forma = any(k in desc for k in ['forma', 'fôrma'])
            is_estaca = any(k in desc for k in ['estaca', 'hélice', 'helice'])
            is_mo = any(k in desc for k in ['mão de obra', 'mao de obra', 'moe -', 'moe –'])
            is_laje_trelicada = 'laje' in desc and ('treliç' in desc or 'trelic' in desc or 'pré' in desc or 'pre' in desc)

            if unit in ['m3', 'm³']:
                if is_concreto:
                    if is_infra:
                        result['concreto_infra_m3'] += qty
                    else:
                        result['concreto_supra_m3'] += qty
                    # Extract fck
                    for fck in re.findall(r'(\d+)\s*(?:mpa|MPa)', desc):
                        fck_val = f"{fck}MPa"
                        if fck_val not in result['fck_list']:
                            result['fck_list'].append(fck_val)

            elif unit == 'kg':
                if is_aco:
                    if is_infra:
                        result['aco_infra_kg'] += qty
                    else:
                        result['aco_supra_kg'] += qty

            elif unit in ['m2', 'm²']:
                if is_forma:
                    if is_infra:
                        result['forma_infra_m2'] += qty
                    else:
                        result['forma_supra_m2'] += qty
                elif is_laje_trelicada:
                    # Laje treliçada area contributes to formwork
                    result['forma_supra_m2'] += qty
                elif is_mo:
                    # MO line with m2 unit and R$/m2 price = structural labor
                    result['mo_estrutural_total'] += total_val

            elif unit == 'm':
                if is_estaca:
                    result['estacas_m'] += qty
                    if result['tipo_estaca'] is None:
                        if 'helice' in desc or 'hélice' in desc:
                            result['tipo_estaca'] = 'hélice contínua'
                        elif 'raiz' in desc:
                            result['tipo_estaca'] = 'raiz'
                        elif 'franki' in desc:
                            result['tipo_estaca'] = 'franki'

            elif unit in ['un', 'und']:
                if is_estaca:
                    result['estacas_qtd'] += int(qty)

    else:
        # ── Fallback: generic row scanning ──
        in_infra = False
        in_supra = False

        for i, row in enumerate(rows):
            vals = [v for v in row if v is not None]
            if not vals:
                continue
            rt = row_text(row)

            # Section detection
            if 'infraestrutura' in rt:
                in_infra = True
                in_supra = False
            elif 'supraestrutura' in rt:
                in_infra = False
                in_supra = True
            elif any(k in rt for k in ['alvenaria', 'instalações', 'instalacoes',
                                        'impermeabiliz', 'fachada', 'esquadria']):
                in_infra = False
                in_supra = False

            if not in_infra and not in_supra:
                continue

            # Scan for unit and quantity
            desc = ''
            unit = None
            qty = None

            for j, v in enumerate(row):
                sv = safe_str(v)
                svl = sv.lower()
                if isinstance(v, str) and len(v) > 10:
                    desc = svl
                if svl in ['m3', 'm³']:
                    unit = 'm3'
                    for k in [j+1, j-1, j+2]:
                        if 0 <= k < len(row) and isinstance(row[k], (int, float)) and row[k] > 0:
                            qty = float(row[k])
                            break
                elif svl == 'kg':
                    unit = 'kg'
                    for k in [j+1, j-1, j+2]:
                        if 0 <= k < len(row) and isinstance(row[k], (int, float)) and row[k] > 0:
                            qty = float(row[k])
                            break
                elif svl in ['m2', 'm²']:
                    unit = 'm2'
                    for k in [j+1, j-1, j+2]:
                        if 0 <= k < len(row) and isinstance(row[k], (int, float)) and row[k] > 0:
                            qty = float(row[k])
                            break
                elif svl == 'm' and any(k in desc for k in ['estaca', 'hélice', 'helice']):
                    unit = 'm_estaca'
                    for k in [j+1, j-1, j+2]:
                        if 0 <= k < len(row) and isinstance(row[k], (int, float)) and row[k] > 0:
                            qty = float(row[k])
                            break

            if unit and qty and qty > 0:
                is_concreto = any(k in desc for k in ['concreto', 'conc.', 'bombeável', 'bombeavel'])
                is_aco = any(k in desc for k in ['aço', 'aco', 'armação', 'armacao', 'armadura', 'ca-50', 'ca-60'])
                is_forma = any(k in desc for k in ['forma', 'fôrma'])
                is_estaca = any(k in desc for k in ['estaca', 'hélice', 'helice'])

                if is_concreto and unit == 'm3':
                    if in_infra:
                        result['concreto_infra_m3'] += qty
                    else:
                        result['concreto_supra_m3'] += qty
                    for fck in re.findall(r'(\d+)\s*(?:mpa|MPa)', desc):
                        fck_val = f"{fck}MPa"
                        if fck_val not in result['fck_list']:
                            result['fck_list'].append(fck_val)
                elif is_aco and unit == 'kg':
                    if in_infra:
                        result['aco_infra_kg'] += qty
                    else:
                        result['aco_supra_kg'] += qty
                elif is_forma and unit == 'm2':
                    if in_infra:
                        result['forma_infra_m2'] += qty
                    else:
                        result['forma_supra_m2'] += qty
                elif is_estaca and unit == 'm_estaca':
                    result['estacas_m'] += qty
                    if result['tipo_estaca'] is None:
                        if 'helice' in desc or 'hélice' in desc:
                            result['tipo_estaca'] = 'hélice contínua'

    return result


# ============================================================
# Format 3: Dedicated Structural Sheets (Estacas, Vigas, etc.)
# ============================================================

def extract_from_structural_sheets(wb, sheets):
    """
    Extract from dedicated structural sheets like Estacas, Blocos,
    Supraestrutura (steel table), Indices de Bloco, etc.
    """
    result = {
        'concreto_infra_m3': 0,
        'concreto_supra_m3': 0,
        'aco_infra_kg': 0,
        'aco_supra_kg': 0,
        'forma_infra_m2': 0,
        'forma_supra_m2': 0,
        'estacas_m': 0,
        'estacas_qtd': 0,
        'tipo_estaca': None,
        'fck_list': [],
        'aco_total_from_table_kg': 0,
    }

    for sname in sheets:
        ws = wb[sname]
        sn = normalize(sname)

        rows = []
        try:
            for row in ws.iter_rows(max_row=500, max_col=20, values_only=True):
                rows.append(tuple(row))
        except:
            continue

        # ── Estacas sheet ──
        if 'estaca' in sn and 'arm' not in sn:
            for i, row in enumerate(rows):
                rt = row_text(row)

                # Detect pile type
                if result['tipo_estaca'] is None:
                    if 'helice' in rt or 'hélice' in rt:
                        result['tipo_estaca'] = 'hélice contínua'
                    elif 'raiz' in rt:
                        result['tipo_estaca'] = 'raiz'
                    elif 'franki' in rt:
                        result['tipo_estaca'] = 'franki'
                    elif 'pre-moldada' in rt or 'pré-moldada' in rt:
                        result['tipo_estaca'] = 'pré-moldada'

                # Look for TOTAL GERAL row
                if 'total geral' in rt or 'total' in rt:
                    # Find the total meters (usually a large number in the row)
                    nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 10]
                    if nums:
                        # First large number is often total qty, second is total meters
                        for n in nums:
                            if 10 < n < 5000 and result['estacas_qtd'] == 0:
                                result['estacas_qtd'] = int(n)
                            elif n > 100 and result['estacas_m'] == 0:
                                result['estacas_m'] = n

                # Look for CONCRETO column values
                if 'concreto' in rt and ('m3' in rt or 'm³' in rt):
                    nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 1]
                    if nums:
                        result['concreto_infra_m3'] += max(nums)

                # Individual pile rows: detect by diameter pattern
                if any(k in rt for k in ['diam', 'diâmetro', 'ø']):
                    # Count quantity (usually 2nd or 3rd number)
                    pass

            # Second pass: look for summary row with total m and concreto
            for i, row in enumerate(rows):
                rt = row_text(row)
                if 'total' in rt:
                    vals = [v for v in row if isinstance(v, (int, float))]
                    # In estacas sheets, TOTAL GERAL typically has:
                    # [qty, total_m_linear, escav_m3, concreto_m3, ...]
                    if len(vals) >= 2:
                        for v in vals:
                            if 50 < v < 50000 and result['estacas_m'] < v:
                                result['estacas_m'] = v
                                break

        # ── Supraestrutura (steel armature) sheet ──
        if 'supraestrutura' in sn or 'estrutura' in sn:
            for i, row in enumerate(rows):
                rt = row_text(row)

                # Look for TOTAL - kg row
                if 'total' in rt and 'kg' in rt:
                    nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 100]
                    if nums:
                        total_kg = sum(nums)
                        if total_kg > result['aco_supra_kg']:
                            result['aco_supra_kg'] = total_kg

                # Look for RESUMO - CONCRETO section
                if 'resumo' in rt and 'concreto' in rt:
                    # Following rows have m3 values
                    for j in range(i+1, min(i+15, len(rows))):
                        r2 = rows[j]
                        nums = [float(v) for v in r2 if isinstance(v, (int, float)) and v > 0]
                        rt2 = row_text(r2)
                        if 'conferir' in rt2 or 'total' in rt2:
                            if nums:
                                result['concreto_supra_m3'] = max(result['concreto_supra_m3'], max(nums))
                            break
                        if nums:
                            result['concreto_supra_m3'] += nums[0]
                        # Extract fck from "tipo" column
                        for v in r2:
                            if isinstance(v, (int, float)) and 20 <= v <= 60:
                                fck_val = f"{int(v)}MPa"
                                if fck_val not in result['fck_list']:
                                    result['fck_list'].append(fck_val)

        # ── Indices de Bloco / Fundação Rasa sheets ──
        if any(k in sn for k in ['bloco', 'fundacao rasa', 'fundação rasa', 'infraestrutura']):
            for i, row in enumerate(rows):
                rt = row_text(row)
                if 'total' in rt:
                    # Look for concreto and aco totals
                    if 'concreto' in rt or 'm3' in rt or 'm³' in rt:
                        nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 1]
                        if nums:
                            result['concreto_infra_m3'] += max(nums)
                    if 'aco' in rt or 'aço' in rt or 'kg' in rt:
                        nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 10]
                        if nums:
                            result['aco_infra_kg'] += max(nums)
                    if 'forma' in rt or 'fôrma' in rt:
                        nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 1]
                        if nums:
                            result['forma_infra_m2'] += max(nums)

        # ── Aço_Infraestrutura sheet ──
        if 'aco_infraestrutura' in sn or 'aço_infraestrutura' in sn:
            for i, row in enumerate(rows):
                rt = row_text(row)
                if 'total' in rt and 'kg' in rt:
                    nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 100]
                    if nums:
                        result['aco_infra_kg'] = max(result['aco_infra_kg'], sum(nums))

        # ── Vigas / Lajes / Pilares sheets (for formwork m2) ──
        if any(k in sn for k in ['vigas', 'laje', 'pilar']):
            for i, row in enumerate(rows):
                rt = row_text(row)
                # These sheets often have form areas
                if 'total' in rt and ('forma' in rt or 'm2' in rt or 'm²' in rt):
                    nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 10]
                    if nums:
                        if 'fund' in sn:
                            result['forma_infra_m2'] += max(nums)
                        else:
                            result['forma_supra_m2'] += max(nums)

        # ── Quantitativo aço sheet ──
        if 'quantitativo aco' in sn or 'quantitativo aço' in sn:
            for i, row in enumerate(rows):
                rt = row_text(row)
                if 'total' in rt and 'geral' in rt:
                    nums = [float(v) for v in row if isinstance(v, (int, float)) and v > 100]
                    if nums:
                        result['aco_total_from_table_kg'] = max(nums)

    return result


# ============================================================
# Format 4: Sienge Relatório
# ============================================================

def extract_from_relatorio(wb, sheet_name='Relatório'):
    """
    Extract structural data from Sienge Relatório format.
    Format: col[0]=code, col[1]=description, col[2]=unit, col[3]=qty (or similar).
    Section detected from N1 codes (XX = INFRAESTRUTURA/SUPRAESTRUTURA) or description.
    """
    result = {
        'concreto_infra_m3': 0,
        'concreto_supra_m3': 0,
        'aco_infra_kg': 0,
        'aco_supra_kg': 0,
        'forma_infra_m2': 0,
        'forma_supra_m2': 0,
        'estacas_m': 0,
        'estacas_qtd': 0,
        'tipo_estaca': None,
        'fck_list': [],
        'mo_estrutural_total': 0,
    }

    ws = wb[sheet_name]
    rows = []
    try:
        for row in ws.iter_rows(max_row=5000, max_col=30, values_only=True):
            rows.append(tuple(row))
    except:
        return result

    # First pass: identify which N1 codes correspond to infra/supra
    infra_prefixes = set()
    supra_prefixes = set()

    for row in rows:
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        rt = row_text(row)

        # Look for N1 headers like "01 " "INFRAESTRUTURA" or "02 " "SUPRAESTRUTURA"
        code = safe_str(row[0] if row[0] is not None else '').strip()
        desc_candidates = [safe_str(v).lower() for v in row[1:6] if isinstance(v, str) and len(safe_str(v)) > 3]
        desc_joined = ' '.join(desc_candidates)

        # Match N1 code pattern (2 digits, possibly with trailing space or .XXX)
        m = re.match(r'^(\d{2})\s*$', code)
        if m:
            prefix = m.group(1)
            if 'infraestrutura' in desc_joined or 'fundaç' in desc_joined or 'fundac' in desc_joined:
                infra_prefixes.add(prefix)
            elif 'supraestrutura' in desc_joined:
                supra_prefixes.add(prefix)

    # Detect column layout: find where code, desc, unit, qty are
    # Strategy: look for header row with "Un." or "Quantidade" or find first data row with a known unit
    desc_col = None
    unit_col = None
    qty_col = None

    # First try: find header row
    for row in rows[:30]:
        for j, v in enumerate(row):
            sv = safe_str(v).lower()
            if sv in ['un.', 'un', 'und.', 'unid.', 'unidade']:
                unit_col = j
            elif any(k in sv for k in ['quantidade', 'qtd']):
                qty_col = j
            elif sv in ['descrição', 'descricao', 'descrição do serviço']:
                desc_col = j
        if unit_col is not None:
            break

    # Second try: find from data rows
    if unit_col is None:
        for row in rows:
            for j, v in enumerate(row):
                sv = safe_str(v).lower()
                if sv in ['m3', 'm³', 'kg', 'm2', 'm²', 'vb', 'un', 'und', 'm']:
                    unit_col = j
                    break
            if unit_col is not None:
                break

    # Find qty column: look for numbers near the unit column
    if unit_col is not None and qty_col is None:
        for row in rows:
            if unit_col < len(row) and row[unit_col] is not None:
                uv = safe_str(row[unit_col]).lower()
                if uv in ['m3', 'm³', 'kg', 'm2', 'm²', 'vb', 'un', 'und', 'm']:
                    # Scan nearby columns for a number
                    for k in range(unit_col + 1, min(len(row), unit_col + 6)):
                        if isinstance(row[k], (int, float)) and row[k] > 0:
                            qty_col = k
                            break
                    if qty_col is not None:
                        break

    # Find desc column
    if desc_col is None:
        # Usually column 1 for Relatório format
        for row in rows[10:30]:
            if len(row) > 1 and isinstance(row[1], str) and len(row[1]) > 10:
                desc_col = 1
                break
        if desc_col is None:
            # Try column before unit
            if unit_col is not None:
                for k in range(unit_col - 1, -1, -1):
                    for row in rows[10:30]:
                        if k < len(row) and isinstance(row[k], str) and len(row[k]) > 5:
                            desc_col = k
                            break
                    if desc_col is not None:
                        break

    if unit_col is None or qty_col is None:
        return result

    # Second pass: extract data
    section = None

    for i, row in enumerate(rows):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        rt = row_text(row)
        code = safe_str(row[0] if row[0] is not None else '').strip()

        # Detect section from code prefix
        code_prefix = code[:2] if len(code) >= 2 else ''
        if code_prefix in infra_prefixes:
            section = 'infra'
        elif code_prefix in supra_prefixes:
            section = 'supra'
        elif not infra_prefixes and not supra_prefixes:
            # No clear N1 headers found - detect from description
            if 'infraestrutura' in rt or 'fundação' in rt or 'fundacao' in rt:
                section = 'infra'
            elif 'supraestrutura' in rt:
                section = 'supra'
            elif any(k in rt for k in ['alvenaria', 'instalações', 'instalacoes',
                                        'impermeabiliz', 'esquadria', 'revestiment',
                                        'fachada', 'complementar', 'pintura']):
                section = None

        if section is None:
            continue

        # Get unit and quantity
        unit = safe_str(row[unit_col]).lower() if unit_col < len(row) and row[unit_col] is not None else ''
        qty = safe_float(row[qty_col]) if qty_col < len(row) else 0

        # Get description
        desc = ''
        if desc_col is not None and desc_col < len(row):
            desc = safe_str(row[desc_col]).lower()
        if not desc:
            # Try to find longest string in row
            for v in row:
                if isinstance(v, str) and len(v) > len(desc):
                    desc = v.lower()

        if qty <= 0:
            continue

        # Classify
        is_concreto = any(k in desc for k in ['concreto', 'conc.', 'bombeável', 'bombeavel'])
        is_aco = any(k in desc for k in ['aço', 'aco', 'armação', 'armacao', 'armadura',
                                          'ca-50', 'ca-60', 'ca50', 'ca60'])
        is_forma = any(k in desc for k in ['forma', 'fôrma'])
        is_estaca = any(k in desc for k in ['estaca', 'hélice', 'helice'])
        is_mo = any(k in desc for k in ['mão de obra', 'mao de obra'])

        if unit in ['m3', 'm³']:
            if is_concreto:
                if section == 'infra':
                    result['concreto_infra_m3'] += qty
                else:
                    result['concreto_supra_m3'] += qty
                for fck in re.findall(r'(\d+)\s*[Mm][Pp][Aa]', desc):
                    fck_val = f"{fck}MPa"
                    if fck_val not in result['fck_list']:
                        result['fck_list'].append(fck_val)

        elif unit == 'kg':
            if is_aco:
                if section == 'infra':
                    result['aco_infra_kg'] += qty
                else:
                    result['aco_supra_kg'] += qty

        elif unit in ['m2', 'm²']:
            if is_forma:
                if section == 'infra':
                    result['forma_infra_m2'] += qty
                else:
                    result['forma_supra_m2'] += qty
            elif is_mo:
                # MO measured in m2 - track total
                pass  # We'd need unit price for this

        elif unit == 'm':
            if is_estaca:
                result['estacas_m'] += qty
                if result['tipo_estaca'] is None:
                    if 'helice' in desc or 'hélice' in desc:
                        result['tipo_estaca'] = 'hélice contínua'
                    elif 'raiz' in desc:
                        result['tipo_estaca'] = 'raiz'

        elif unit in ['un', 'und']:
            if is_estaca:
                result['estacas_qtd'] += int(qty)

    return result


# ============================================================
# Format 5: ORÇAMENTO / ORÇAMENTO_EXECUTIVO sheets
# ============================================================

def extract_from_orcamento_sheet(wb, sheet_name):
    """
    Extract from generic ORÇAMENTO/ORÇAMENTO_EXECUTIVO sheets.
    These have a table format with descriptions, units, quantities.
    """
    result = {
        'concreto_infra_m3': 0,
        'concreto_supra_m3': 0,
        'aco_infra_kg': 0,
        'aco_supra_kg': 0,
        'forma_infra_m2': 0,
        'forma_supra_m2': 0,
        'estacas_m': 0,
        'estacas_qtd': 0,
        'tipo_estaca': None,
        'fck_list': [],
        'mo_estrutural_total': 0,
    }

    ws = wb[sheet_name]
    rows = []
    try:
        for row in ws.iter_rows(max_row=5000, max_col=30, values_only=True):
            rows.append(tuple(row))
    except:
        return result

    section = None

    for i, row in enumerate(rows):
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        rt = row_text(row)

        # Section detection
        if 'infraestrutura' in rt or 'fundação' in rt.replace('ã', 'a') or 'fundacao' in rt:
            section = 'infra'
        elif 'supraestrutura' in rt:
            section = 'supra'
        elif any(k in rt for k in ['alvenaria', 'instalações', 'instalacoes', 'impermeabiliz',
                                    'esquadria', 'revestiment', 'fachada', 'complementar',
                                    'pintura']):
            section = None

        if section is None:
            continue

        # Find description, unit, quantity columns by scanning
        desc = ''
        unit = None
        qty = None

        for j, v in enumerate(row):
            sv = safe_str(v)
            svl = sv.lower()

            if isinstance(v, str) and len(v) > 8:
                desc = svl

            if svl in ['m³', 'm3'] and unit is None:
                unit = 'm3'
            elif svl == 'kg' and unit is None:
                unit = 'kg'
            elif svl in ['m²', 'm2'] and unit is None:
                unit = 'm2'
            elif svl == 'm' and unit is None and any(k in desc for k in ['estaca', 'helice', 'hélice']):
                unit = 'm_estaca'

        if unit:
            # Find quantity: scan for the unit cell and look nearby
            for j, v in enumerate(row):
                svl = safe_str(v).lower()
                if (unit == 'm3' and svl in ['m³', 'm3']) or \
                   (unit == 'kg' and svl == 'kg') or \
                   (unit == 'm2' and svl in ['m²', 'm2']) or \
                   (unit == 'm_estaca' and svl == 'm'):
                    # Qty is usually in the next column or 2 columns after
                    for k in [j+1, j+2, j-1]:
                        if 0 <= k < len(row) and isinstance(row[k], (int, float)) and row[k] > 0:
                            qty = float(row[k])
                            break
                    break

        if qty and qty > 0 and unit:
            is_concreto = any(k in desc for k in ['concreto', 'conc.'])
            is_aco = any(k in desc for k in ['aço', 'aco', 'armação', 'armacao', 'armadura', 'ca-50', 'ca-60'])
            is_forma = any(k in desc for k in ['forma', 'fôrma'])

            if is_concreto and unit == 'm3':
                if section == 'infra':
                    result['concreto_infra_m3'] += qty
                else:
                    result['concreto_supra_m3'] += qty
            elif is_aco and unit == 'kg':
                if section == 'infra':
                    result['aco_infra_kg'] += qty
                else:
                    result['aco_supra_kg'] += qty
            elif is_forma and unit == 'm2':
                if section == 'infra':
                    result['forma_infra_m2'] += qty
                else:
                    result['forma_supra_m2'] += qty

    return result


# ============================================================
# Merge results from multiple extractors
# ============================================================

def merge_results(*results):
    """Merge multiple extraction results, preferring non-zero values."""
    merged = {
        'concreto_infra_m3': 0,
        'concreto_supra_m3': 0,
        'aco_infra_kg': 0,
        'aco_supra_kg': 0,
        'forma_infra_m2': 0,
        'forma_supra_m2': 0,
        'estacas_m': 0,
        'estacas_qtd': 0,
        'tipo_estaca': None,
        'fck_list': [],
        'mo_estrutural_total': 0,
        'aco_total_from_table_kg': 0,
    }

    for r in results:
        for key in ['concreto_infra_m3', 'concreto_supra_m3', 'aco_infra_kg', 'aco_supra_kg',
                     'forma_infra_m2', 'forma_supra_m2', 'estacas_m', 'estacas_qtd',
                     'mo_estrutural_total', 'aco_total_from_table_kg']:
            if r.get(key, 0) > merged[key]:
                merged[key] = r[key]

        if r.get('tipo_estaca') and not merged['tipo_estaca']:
            merged['tipo_estaca'] = r['tipo_estaca']

        for fck in r.get('fck_list', []):
            if fck not in merged['fck_list']:
                merged['fck_list'].append(fck)

    return merged


def build_indices(merged, ac):
    """Build the final indices_estruturais dict from merged data."""
    concreto_total = merged['concreto_infra_m3'] + merged['concreto_supra_m3']
    aco_total = merged['aco_infra_kg'] + merged['aco_supra_kg']
    if merged.get('aco_total_from_table_kg', 0) > aco_total:
        aco_total = merged['aco_total_from_table_kg']
    forma_total = merged['forma_infra_m2'] + merged['forma_supra_m2']

    if concreto_total == 0 and aco_total == 0 and forma_total == 0:
        return None

    indices = {}

    if concreto_total > 0:
        indices['concreto_total_m3'] = round(concreto_total, 2)
        if ac and ac > 0:
            indices['concreto_m3_por_m2_ac'] = round(concreto_total / ac, 4)

    if aco_total > 0:
        indices['aco_total_kg'] = round(aco_total, 2)
        if concreto_total > 0:
            indices['aco_kg_por_m3_concreto'] = round(aco_total / concreto_total, 2)

    if forma_total > 0:
        indices['forma_total_m2'] = round(forma_total, 2)
        if ac and ac > 0:
            indices['forma_m2_por_m2_ac'] = round(forma_total / ac, 4)

    # fck predominante (highest fck)
    if merged['fck_list']:
        try:
            fck_vals = sorted([int(f.replace('MPa', '')) for f in merged['fck_list']], reverse=True)
            indices['fck_predominante'] = f"{fck_vals[0]}MPa"
        except:
            pass

    # Pile data
    if merged['tipo_estaca']:
        indices['tipo_estaca'] = merged['tipo_estaca']
    if merged['estacas_m'] > 0:
        indices['estacas_comprimento_total_m'] = round(merged['estacas_m'], 2)
    if merged['estacas_qtd'] > 0:
        indices['estacas_qtd'] = merged['estacas_qtd']

    # Foundation breakdown
    if merged['concreto_infra_m3'] > 0:
        indices['fundacao_concreto_m3'] = round(merged['concreto_infra_m3'], 2)
    if merged['aco_infra_kg'] > 0:
        indices['fundacao_aco_kg'] = round(merged['aco_infra_kg'], 2)

    # MO structural
    if merged['mo_estrutural_total'] > 0 and ac and ac > 0:
        indices['mo_estrutural_rsm2'] = round(merged['mo_estrutural_total'] / ac, 2)

    return indices


# ============================================================
# Main
# ============================================================

def main():
    with open(MAPPING_FILE, encoding='utf-8') as f:
        all_projects = json.load(f)

    with open(METADADOS_FILE, encoding='utf-8') as f:
        metadados = json.load(f)

    # Build AC lookup - metadados is a dict keyed by slug
    ac_map = {}
    if isinstance(metadados, dict):
        for slug, pm in metadados.items():
            if isinstance(pm, dict):
                ac_map[slug] = pm.get('ac')
    elif isinstance(metadados, list):
        for pm in metadados:
            s = pm.get('slug') or pm.get('projeto')
            if s:
                ac_map[s] = pm.get('ac')

    # Also get AC from indices-executivo files (they already have it)
    for idx_file in INDICES_DIR.glob("*.json"):
        slug = idx_file.stem
        if slug not in ac_map or not ac_map[slug]:
            try:
                with open(idx_file, encoding='utf-8') as f:
                    idx = json.load(f)
                if idx.get('ac'):
                    ac_map[slug] = idx['ac']
            except:
                pass

    batch = all_projects[BATCH_START:BATCH_END]

    results_summary = {
        'processed': 0,
        'extracted': 0,
        'skipped_no_file': 0,
        'skipped_no_data': 0,
        'errors': 0,
        'details': [],
    }

    for proj in batch:
        slug = proj['slug']
        path = proj['path']
        results_summary['processed'] += 1

        if not os.path.exists(path):
            print(f"[SKIP] {slug}: file not found")
            results_summary['skipped_no_file'] += 1
            continue

        ac = ac_map.get(slug, 0) or 0
        print(f"\n[PROC] {slug} (AC={ac})")

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheets = wb.sheetnames

            all_extracted = []

            # ── Strategy 1: Parametric sheets ──
            parametric_sheets = [s for s in sheets if normalize(s) in
                ['infraestrutura', 'supraestrutura']]
            if parametric_sheets:
                print(f"  Trying parametric sheets: {parametric_sheets}")
                r = extract_from_parametric_sheets(wb, parametric_sheets)
                all_extracted.append(r)

            # ── Strategy 2: Dedicated structural sheets ──
            structural_sheets = [s for s in sheets if any(k in normalize(s) for k in
                ['vigas', 'laje', 'pilar', 'estaca', 'bloco', 'fundacao rasa', 'fundação rasa',
                 'aco_infraestrutura', 'aço_infraestrutura', 'tabela de aco', 'tabela de aço',
                 'quantitativo aco', 'quantitativo aço', 'indices de bloco',
                 'resumo estrutura', 'laje_trelicada', 'laje_treliçada']) and
                normalize(s) not in ['infraestrutura', 'supraestrutura']]
            # Also include Supraestrutura-like dedicated sheets that have steel data
            for s in sheets:
                sn = normalize(s)
                if sn == 'supraestrutura' and s not in parametric_sheets:
                    structural_sheets.append(s)

            if structural_sheets:
                print(f"  Trying structural sheets: {structural_sheets}")
                r = extract_from_structural_sheets(wb, structural_sheets)
                all_extracted.append(r)

            # ── Strategy 3: Ger_Executivo ──
            ger_sheets = [s for s in sheets if normalize(s) in
                ['ger_executivo', 'gerenciamento executivo']]
            if ger_sheets:
                print(f"  Trying Ger_Executivo: {ger_sheets[0]}")
                r = extract_from_ger_executivo(wb, ger_sheets[0])
                all_extracted.append(r)

            # ── Strategy 4: Relatório ──
            rel_sheets = [s for s in sheets if 'relatório' in normalize(s) or 'relatorio' in normalize(s)]
            if rel_sheets:
                print(f"  Trying Relatório: {rel_sheets[0]}")
                r = extract_from_relatorio(wb, rel_sheets[0])
                all_extracted.append(r)

            # ── Strategy 5: ORÇAMENTO_EXECUTIVO / Orçamento sheets ──
            orc_sheets = [s for s in sheets if any(k in normalize(s) for k in
                ['orcamento_executivo', 'orçamento_executivo', 'orcamento executivo',
                 'orçamento executivo']) and s not in rel_sheets]
            if orc_sheets:
                print(f"  Trying Orçamento: {orc_sheets[0]}")
                r = extract_from_orcamento_sheet(wb, orc_sheets[0])
                all_extracted.append(r)

            # Also check "Orçamento" or "Orcamento" sheets
            for s in sheets:
                sn = normalize(s)
                if sn in ['orcamento', 'orçamento'] and s not in orc_sheets and s not in rel_sheets:
                    print(f"  Trying Orçamento generic: {s}")
                    r = extract_from_orcamento_sheet(wb, s)
                    all_extracted.append(r)

            wb.close()

            if not all_extracted:
                print(f"  No structural data sources found")
                results_summary['skipped_no_data'] += 1
                continue

            # Merge all results
            merged = merge_results(*all_extracted)
            indices = build_indices(merged, ac)

            if indices is None:
                print(f"  No structural quantities found (all zeros)")
                results_summary['skipped_no_data'] += 1
                continue

            print(f"  FOUND: concreto={indices.get('concreto_total_m3', 0)} m³, "
                  f"aço={indices.get('aco_total_kg', 0)} kg, "
                  f"forma={indices.get('forma_total_m2', 0)} m², "
                  f"estacas={indices.get('estacas_comprimento_total_m', 0)} m")

            # Update the indices-executivo JSON
            idx_file = INDICES_DIR / f"{slug}.json"
            if idx_file.exists():
                with open(idx_file, encoding='utf-8') as f:
                    idx_data = json.load(f)
            else:
                idx_data = {"projeto": slug}

            idx_data['indices_estruturais'] = indices

            with open(idx_file, 'w', encoding='utf-8') as f:
                json.dump(idx_data, f, ensure_ascii=False, indent=2)

            results_summary['extracted'] += 1
            results_summary['details'].append({
                'slug': slug,
                'concreto_total': indices.get('concreto_total_m3', 0),
                'aco_total': indices.get('aco_total_kg', 0),
                'forma_total': indices.get('forma_total_m2', 0),
                'estacas_m': indices.get('estacas_comprimento_total_m', 0),
                'tipo_estaca': indices.get('tipo_estaca'),
            })

        except Exception as e:
            print(f"  ERROR: {e}")
            traceback.print_exc()
            results_summary['errors'] += 1

    # ── Summary ──
    print("\n" + "=" * 70)
    print("SUMMARY - Structural Indices Extraction (Batch 1: projects 0-41)")
    print("=" * 70)
    print(f"Total processed:     {results_summary['processed']}")
    print(f"Successfully extracted: {results_summary['extracted']}")
    print(f"No file found:       {results_summary['skipped_no_file']}")
    print(f"No structural data:  {results_summary['skipped_no_data']}")
    print(f"Errors:              {results_summary['errors']}")
    print()

    if results_summary['details']:
        print("Projects with structural data:")
        print(f"{'Slug':<45} {'Concreto m³':>12} {'Aço kg':>12} {'Forma m²':>12} {'Estacas m':>12} {'Tipo'}")
        print("-" * 110)
        for d in results_summary['details']:
            print(f"{d['slug']:<45} {d['concreto_total']:>12.1f} {d['aco_total']:>12.1f} "
                  f"{d['forma_total']:>12.1f} {d['estacas_m']:>12.1f} {d.get('tipo_estaca', '-')}")


if __name__ == '__main__':
    main()
