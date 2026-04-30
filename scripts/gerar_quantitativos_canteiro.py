"""Gera quantitativos.xlsx final pra disciplina Canteiro do Electra Towers.

Diferente da extracao.xlsx (que tem fórmulas e linka pra CAPA hidden), este arquivo:
- Tem valores fechados (calculados) prontos pra entrega/consolidação
- Sinaliza pendências com cor amarela + coluna Status
- Resumo no topo, detalhamento por bloco, pendências no final

Uso:
    py -3.10 -X utf8 ~/orcamentos-openclaw/scripts/gerar_quantitativos_canteiro.py
"""

from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJETO_JSON = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra\orcamento\00-projeto\projeto.json"
)
DEST = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra\orcamento\04-disciplinas\Canteiro\quantitativos.xlsx"
)

# Estilos
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_TOTAL = Font(name="Calibri", size=11, bold=True)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_SMALL = Font(name="Calibri", size=9, italic=True, color="555555")

FILL_TITLE = PatternFill("solid", fgColor="1F4E79")  # azul escuro
FILL_HEADER = PatternFill("solid", fgColor="4472C4")  # azul médio
FILL_BLOCK = PatternFill("solid", fgColor="DDEBF7")  # azul claro
FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")  # amarelo claro (linha total)
FILL_PEND = PatternFill("solid", fgColor="FFE699")  # amarelo (pendência)
FILL_ALERT = PatternFill("solid", fgColor="F8CBAD")  # laranja (alerta forte)

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


def fmt_brl(v):
    if v is None:
        return ""
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def main():
    projeto = json.loads(PROJETO_JSON.read_text(encoding="utf-8"))

    # Drivers do Electra
    area_construida = projeto["areas"]["construida_total_m2"]  # 37893.89
    prazo_meses = projeto["projeto"]["prazo_obra_meses"]  # 48

    # Constantes paramétricas Cartesian R00 (do template)
    Pm = 47  # h/m² — produtividade média (D9; T5 usa 49, divergência)
    horas_mes = 180  # divisor (8h × 22.5 dias úteis)
    equipe_admin = 9  # constante (D11; P8 usa 5, divergência)
    epi_per_capita = 750
    ferramentas_per_capita = 600

    # Cálculos primários
    n_funcionarios = math.ceil((Pm * area_construida) / prazo_meses / horas_mes)
    produtividade_mes = area_construida / prazo_meses
    n_banheiros_obra = math.ceil(n_funcionarios / 20)

    # Bloco 2 — Barracão
    refeitorio_m2 = math.ceil((n_funcionarios + equipe_admin) * 1)
    vestiario_m2 = math.ceil((n_funcionarios + equipe_admin) * 1.5)
    if area_construida < 20000:
        almoxarifado_m2 = 100
    else:
        almoxarifado_m2 = 100 + ((area_construida - 20000) / 1000) * 5
    escritorio_m2 = math.ceil(equipe_admin * 8)
    baias_m2 = 5  # constante
    sala_cliente_m2 = 2  # constante

    # Custos unitários (template — valor base × 1.1 fator regional)
    custo_un = {
        "refeitorio": 266.76 * 1.1,
        "vestiario": 322.87 * 1.1,
        "almoxarifado": 272.88 * 1.1,
        "escritorio": 274.01 * 1.1,
        "baias": 115.48 * 1.1,
    }
    sala_cliente_un = custo_un["escritorio"] + (200 * 10) / 12 + 1200

    barracao_itens = [
        ("Refeitório", "m²", refeitorio_m2, custo_un["refeitorio"], "1 m²/funcionário (op + adm)", False),
        ("Vestiário/banheiro", "m²", vestiario_m2, custo_un["vestiario"], "1,5 m²/funcionário (vestiário) + lavatório/chuveiro NR-18", False),
        ("Almoxarifado", "m²", almoxarifado_m2, custo_un["almoxarifado"], "100 m² base + 5 m²/1.000 m² acima de 20.000", False),
        ("Escritório", "m²", escritorio_m2, custo_un["escritorio"], "8 m²/funcionário administrativo", False),
        ("Baias de material", "m²", baias_m2, custo_un["baias"], "5 m²/bloco (constante)", False),
        ("Sala do cliente", "m²", sala_cliente_m2, sala_cliente_un, "Container c/ mobiliário + decoração + EPIs", False),
    ]

    # Bloco 3 — Mobiliário
    n_cadeiras = math.ceil(equipe_admin * 1.5)
    mobiliario_itens = [
        ("Cadeira", "un", n_cadeiras, 200, "1,5 × Equipe admin", False),
        ("Mesa", "un", 1, 850, "Constante (1 mesa)", False),
        ("Ar condicionado", "un", 1, 3000, "Constante (1 unidade)", False),
        ("Armário", "m", 4, 900, "Constante (4 m lineares)", False),
        ("Escrivaninha", "un", equipe_admin, 800, "1 / funcionário admin", False),
    ]

    # Bloco 4 — Eletrônicos
    eletronicos_itens = [
        ("Computador pessoal", "un", equipe_admin, 3500, "1 / funcionário admin", False),
        ("Impressora", "un", 1, 900, "Constante", False),
        ("Roteador", "un", 2, 500, "Constante", False),
    ]

    # Bloco 1 — EPI/Ferramentas (R$ direto)
    bloco1_itens = [
        ("Ferramentas", "R$", n_funcionarios, ferramentas_per_capita, f"R$ {ferramentas_per_capita}/func operário", False),
        ("EPI", "R$", n_funcionarios, epi_per_capita, f"R$ {epi_per_capita}/func operário", False),
    ]

    # ---------------------- Workbook ----------------------
    wb = Workbook()

    # ==== Aba 1: Resumo ====
    ws = wb.active
    ws.title = "Resumo"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60

    r = 1
    ws.cell(r, 1, "QUANTITATIVOS — CANTEIRO").font = FONT_TITLE
    ws.cell(r, 1).fill = FILL_TITLE
    ws.cell(r, 1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 28
    r += 1

    ws.cell(r, 1, "Empreendimento").font = FONT_NORMAL
    ws.cell(r, 2, projeto["projeto"]["nome"]).font = FONT_TOTAL
    ws.cell(r, 3, "Electra Towers — Thozen Empreendimentos").font = FONT_SMALL
    r += 1

    ws.cell(r, 1, "Endereço").font = FONT_NORMAL
    e = projeto["projeto"]["endereco"]
    ws.cell(r, 2, f"{e['rua']}, {e['numero']}, {e['bairro']}, {e['cidade']}/{e['uf']}").font = FONT_NORMAL
    r += 1

    ws.cell(r, 1, "Revisão").font = FONT_NORMAL
    ws.cell(r, 2, projeto["revisao"]).font = FONT_NORMAL
    ws.cell(r, 3, projeto["data_atualizacao"]).font = FONT_SMALL
    r += 2

    # Drivers
    ws.cell(r, 1, "DRIVERS DO DIMENSIONAMENTO").font = FONT_HEADER
    ws.cell(r, 1).fill = FILL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    drivers = [
        ("Área construída total", f"{area_construida:,.2f} m²".replace(",", "X").replace(".", ",").replace("X", "."), "00-projeto/areas.md → CAPA!C10"),
        ("Prazo da obra", f"{prazo_meses} meses", "00-projeto/projeto.md → CAPA!C6"),
        ("Produtividade média (Pm)", f"{Pm} h/m²", "⚠ DIVERGÊNCIA: D9=47, T5=49 — usado 47"),
        ("Horas-mês por funcionário", f"{horas_mes} h", "Constante (8h × 22,5 dias úteis)"),
        ("Equipe administrativa", f"{equipe_admin} pessoas", "⚠ DIVERGÊNCIA: D11=9, P8=5 — usado 9"),
    ]
    for label, val, src in drivers:
        ws.cell(r, 1, label).font = FONT_NORMAL
        ws.cell(r, 2, val).font = FONT_TOTAL
        ws.cell(r, 3, src).font = FONT_SMALL
        if "DIVERGÊNCIA" in src:
            for c in range(1, 4):
                ws.cell(r, c).fill = FILL_PEND
        r += 1
    r += 1

    # Resultado dimensionamento
    ws.cell(r, 1, "RESULTADO DO DIMENSIONAMENTO").font = FONT_HEADER
    ws.cell(r, 1).fill = FILL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    ws.cell(r, 1, "Produtividade média (m²/mês)").font = FONT_NORMAL
    ws.cell(r, 2, f"{produtividade_mes:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")).font = FONT_TOTAL
    ws.cell(r, 3, "Área Construída ÷ Prazo").font = FONT_SMALL
    r += 1

    ws.cell(r, 1, "Nº de funcionários (operários)").font = FONT_NORMAL
    ws.cell(r, 2, n_funcionarios).font = FONT_TOTAL
    ws.cell(r, 2).alignment = Alignment(horizontal="right")
    ws.cell(r, 3, "ROUNDUP((Pm × Área) ÷ Prazo ÷ 180)").font = FONT_SMALL
    r += 1

    ws.cell(r, 1, "Equipe administrativa").font = FONT_NORMAL
    ws.cell(r, 2, equipe_admin).font = FONT_TOTAL
    ws.cell(r, 2).alignment = Alignment(horizontal="right")
    ws.cell(r, 3, "Constante R00").font = FONT_SMALL
    r += 1

    ws.cell(r, 1, "Banheiros (NR-18)").font = FONT_NORMAL
    ws.cell(r, 2, n_banheiros_obra).font = FONT_TOTAL
    ws.cell(r, 2).alignment = Alignment(horizontal="right")
    ws.cell(r, 3, "1 banheiro / 20 funcionários").font = FONT_SMALL
    r += 2

    # Custos por bloco — totais
    custos = {
        "Bloco 1 — Ferramentas": ferramentas_per_capita * n_funcionarios,
        "Bloco 1 — EPI": epi_per_capita * n_funcionarios,
        "Bloco 2 — Barracão (refeitório, vestiário, almox., escritório, baias, sala cliente)": sum(q * c for _, _, q, c, _, _ in barracao_itens),
        "Bloco 3 — Mobiliário (cadeira, mesa, AC, armário, escrivaninha)": sum(q * c for _, _, q, c, _, _ in mobiliario_itens),
        "Bloco 4 — Eletrônicos (PC, impressora, roteador)": sum(q * c for _, _, q, c, _, _ in eletronicos_itens),
    }
    total_geral = sum(custos.values())

    ws.cell(r, 1, "CUSTOS POR BLOCO").font = FONT_HEADER
    ws.cell(r, 1).fill = FILL_HEADER
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    for label, val in custos.items():
        ws.cell(r, 1, label).font = FONT_NORMAL
        ws.cell(r, 1).alignment = ALIGN_LEFT
        ws.cell(r, 2, fmt_brl(val)).font = FONT_TOTAL
        ws.cell(r, 2).alignment = ALIGN_RIGHT
        r += 1

    ws.cell(r, 1, "TOTAL GERAL — CANTEIRO").font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(r, 1).fill = FILL_TITLE
    ws.cell(r, 2, fmt_brl(total_geral)).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(r, 2).fill = FILL_TITLE
    ws.cell(r, 2).alignment = ALIGN_RIGHT
    ws.cell(r, 3, "").fill = FILL_TITLE
    ws.row_dimensions[r].height = 22

    # ==== Aba 2: Detalhamento ====
    ws2 = wb.create_sheet("Detalhamento")
    cols = [
        ("A", 32, "Item"),
        ("B", 8, "Unid."),
        ("C", 12, "Quant."),
        ("D", 16, "Custo unit."),
        ("E", 18, "Custo total"),
        ("F", 50, "Critério / Fórmula"),
        ("G", 14, "Status"),
    ]
    for letter, width, _ in cols:
        ws2.column_dimensions[letter].width = width

    r = 1
    ws2.cell(r, 1, "DETALHAMENTO POR BLOCO").font = FONT_TITLE
    ws2.cell(r, 1).fill = FILL_TITLE
    ws2.cell(r, 1).alignment = ALIGN_CENTER
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws2.row_dimensions[r].height = 28
    r += 2

    def write_block_header(row, label):
        ws2.cell(row, 1, label).font = FONT_HEADER
        for c in range(1, 8):
            ws2.cell(row, c).fill = FILL_BLOCK
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

    def write_table_header(row):
        for col_idx, (_, _, header) in enumerate(cols, start=1):
            ws2.cell(row, col_idx, header).font = FONT_HEADER
            ws2.cell(row, col_idx).fill = FILL_HEADER
            ws2.cell(row, col_idx).alignment = ALIGN_CENTER
            ws2.cell(row, col_idx).border = BORDER

    def write_item(row, item, unid, qtd, custo_un_v, criterio, pend, alerta=None):
        ws2.cell(row, 1, item).font = FONT_NORMAL
        ws2.cell(row, 1).border = BORDER
        ws2.cell(row, 2, unid).font = FONT_NORMAL
        ws2.cell(row, 2).alignment = ALIGN_CENTER
        ws2.cell(row, 2).border = BORDER
        ws2.cell(row, 3, qtd).font = FONT_NORMAL
        ws2.cell(row, 3).alignment = ALIGN_RIGHT
        ws2.cell(row, 3).border = BORDER
        ws2.cell(row, 3).number_format = "#,##0.00"
        if custo_un_v is not None:
            ws2.cell(row, 4, custo_un_v).font = FONT_NORMAL
            ws2.cell(row, 4).alignment = ALIGN_RIGHT
            ws2.cell(row, 4).number_format = '"R$" #,##0.00'
            ws2.cell(row, 5, qtd * custo_un_v).font = FONT_TOTAL
            ws2.cell(row, 5).alignment = ALIGN_RIGHT
            ws2.cell(row, 5).number_format = '"R$" #,##0.00'
        ws2.cell(row, 4).border = BORDER
        ws2.cell(row, 5).border = BORDER
        ws2.cell(row, 6, criterio).font = FONT_SMALL
        ws2.cell(row, 6).alignment = ALIGN_LEFT
        ws2.cell(row, 6).border = BORDER
        status_text = "OK"
        if alerta:
            status_text = alerta
        elif pend:
            status_text = "Revisar"
        ws2.cell(row, 7, status_text).font = FONT_NORMAL
        ws2.cell(row, 7).alignment = ALIGN_CENTER
        ws2.cell(row, 7).border = BORDER
        if pend or alerta:
            for c in range(1, 8):
                ws2.cell(row, c).fill = FILL_PEND if not alerta else FILL_ALERT

    def write_subtotal(row, label, total):
        ws2.cell(row, 1, label).font = FONT_TOTAL
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws2.cell(row, 1).fill = FILL_TOTAL
        ws2.cell(row, 5, total).font = FONT_TOTAL
        ws2.cell(row, 5).alignment = ALIGN_RIGHT
        ws2.cell(row, 5).number_format = '"R$" #,##0.00'
        ws2.cell(row, 5).fill = FILL_TOTAL
        for c in range(1, 8):
            ws2.cell(row, c).border = BORDER
            if c >= 6:
                ws2.cell(row, c).fill = FILL_TOTAL

    # Bloco 1
    write_block_header(r, "BLOCO 1 — Dimensionamento de Pessoal (Ferramentas + EPI)"); r += 1
    write_table_header(r); r += 1
    for item, unid, qtd, custo, criterio, pend in bloco1_itens:
        write_item(r, item, unid, qtd, custo, criterio, pend); r += 1
    subtotal_b1 = sum(q * c for _, _, q, c, _, _ in bloco1_itens)
    write_subtotal(r, "Subtotal Bloco 1", subtotal_b1); r += 2

    # Bloco 2
    write_block_header(r, "BLOCO 2 — Barracão de Obra"); r += 1
    write_table_header(r); r += 1
    for item, unid, qtd, custo, criterio, pend in barracao_itens:
        # Sala do cliente fórmula composta — sinaliza
        alerta = None
        if "Sala do cliente" in item:
            criterio = criterio + " — fórmula composta híbrida (revisar com equipe)"
            pend = True
        write_item(r, item, unid, qtd, custo, criterio, pend, alerta); r += 1
    subtotal_b2 = sum(q * c for _, _, q, c, _, _ in barracao_itens)
    write_subtotal(r, "Subtotal Bloco 2", subtotal_b2); r += 2

    # Bloco 3
    write_block_header(r, "BLOCO 3 — Mobiliário de Canteiro"); r += 1
    write_table_header(r); r += 1
    for item, unid, qtd, custo, criterio, pend in mobiliario_itens:
        write_item(r, item, unid, qtd, custo, criterio, pend); r += 1
    subtotal_b3 = sum(q * c for _, _, q, c, _, _ in mobiliario_itens)
    write_subtotal(r, "Subtotal Bloco 3", subtotal_b3); r += 2

    # Bloco 4
    write_block_header(r, "BLOCO 4 — Equipamentos Eletrônicos"); r += 1
    write_table_header(r); r += 1
    for item, unid, qtd, custo, criterio, pend in eletronicos_itens:
        write_item(r, item, unid, qtd, custo, criterio, pend); r += 1
    subtotal_b4 = sum(q * c for _, _, q, c, _, _ in eletronicos_itens)
    write_subtotal(r, "Subtotal Bloco 4", subtotal_b4); r += 2

    # Bloco "Outros NR-18" — nº banheiros (operários)
    write_block_header(r, "BLOCO 5 — Banheiros para operários (NR-18)"); r += 1
    write_table_header(r); r += 1
    write_item(r, "Banheiro de obra (operários)", "un", n_banheiros_obra, None, "1 / 20 funcionários (NR-18)",
               True, alerta="Custo unit. NÃO orçado nesta aba — vai junto com 'Vestiário/banheiro' do Bloco 2"); r += 2

    # TOTAL
    ws2.cell(r, 1, "TOTAL GERAL CANTEIRO").font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws2.cell(r, 1).fill = FILL_TITLE
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws2.cell(r, 5, total_geral).font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws2.cell(r, 5).fill = FILL_TITLE
    ws2.cell(r, 5).alignment = ALIGN_RIGHT
    ws2.cell(r, 5).number_format = '"R$" #,##0.00'
    for c in range(1, 8):
        ws2.cell(r, c).fill = FILL_TITLE
    ws2.row_dimensions[r].height = 24

    # ==== Aba 3: Pendências ====
    ws3 = wb.create_sheet("Pendências")
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 38
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 70
    ws3.column_dimensions["E"].width = 14

    r = 1
    ws3.cell(r, 1, "PENDÊNCIAS / DECISÕES EM ABERTO").font = FONT_TITLE
    ws3.cell(r, 1).fill = FILL_TITLE
    ws3.cell(r, 1).alignment = ALIGN_CENTER
    ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws3.row_dimensions[r].height = 28
    r += 2

    headers = ["#", "Pendência", "Severidade", "Descrição / Ação", "Status"]
    for col_idx, h in enumerate(headers, start=1):
        ws3.cell(r, col_idx, h).font = FONT_HEADER
        ws3.cell(r, col_idx).fill = FILL_HEADER
        ws3.cell(r, col_idx).alignment = ALIGN_CENTER
        ws3.cell(r, col_idx).border = BORDER
    r += 1

    pendencias = [
        ("Padronizar Pm", "Alta",
         "D9 usa 47 h/m², T5 (bloco auxiliar lateral) usa 49 h/m². Definir qual é o oficial Cartesian R00. "
         "Diferença gera 207 vs 215 funcionários — impacta cascata em tudo.", "Aberto"),
        ("Padronizar equipe administrativa", "Alta",
         "D11=9 (usado) vs P8=5 (auxiliar). Confirmar tamanho real da equipe Electra. Provavelmente 9 inclui "
         "engenheiro + mestre + estagiário + técnico segurança + 4-5 administrativo + cliente.", "Aberto"),
        ("Atualizar P7 (pavimentos)", "Baixa",
         "Bloco auxiliar tem P7=10 (template Elizabeth II). Electra tem 53 pavtos físicos. "
         "Não impacta cálculos atuais, mas confunde quem ler.", "Aberto"),
        ("Bloco auxiliar P12 = SUM(P11:P11)", "Baixa",
         "Fórmula incompleta (soma só uma célula). Provavelmente faltam termos do barracão completo. "
         "Verificar intenção original.", "Aberto"),
        ("Validar fator × 1,1 dos custos unitários", "Média",
         "E19-E22 multiplicam custo base por 1,1 (10%). Verificar se é fator de mercado regional Itapema/Porto Belo "
         "ou markup. Pra 2026 pode estar baixo (CUB/SC subiu).", "Aberto"),
        ("Sala do cliente — fórmula composta", "Média",
         "E24 = E22 + (200×10)/12 + 1200. Híbrida (custo escritório + diluição mobiliário/decoração + valor fixo). "
         "Revisar se faz sentido pra Electra ou se precisa cotar container específico.", "Aberto"),
        ("Central de armazenamento entre pavimentos", "Média",
         "B27 do template tem nota '* Central de armazenamento entre pavimentos' mas SEM orçamento. "
         "Pra Electra (24 pavtos × 2 torres) é provável que precise — confirmar com equipe.", "Aberto"),
        ("Cruzamento com EPCs", "Baixa",
         "Confirmar que tapume e proteções estão SÓ em EPCs (não duplicados aqui). "
         "Validar que Canteiro cobre apenas instalações fixas + pessoal.", "Aberto"),
        ("Bloco auxiliar lateral desconectado do principal", "Média",
         "O bloco lateral O3:Q46 calcula NR-18 paralelo (refeitório, lavatório, chuveiros, sanitários) "
         "mas o bloco principal usa constantes próprias. Refatorar pra um único critério.", "Aberto"),
    ]

    for i, (titulo, sev, desc, status) in enumerate(pendencias, start=1):
        ws3.cell(r, 1, i).font = FONT_NORMAL
        ws3.cell(r, 1).alignment = ALIGN_CENTER
        ws3.cell(r, 1).border = BORDER
        ws3.cell(r, 2, titulo).font = FONT_TOTAL
        ws3.cell(r, 2).alignment = ALIGN_LEFT
        ws3.cell(r, 2).border = BORDER
        ws3.cell(r, 3, sev).font = FONT_NORMAL
        ws3.cell(r, 3).alignment = ALIGN_CENTER
        ws3.cell(r, 3).border = BORDER
        ws3.cell(r, 4, desc).font = FONT_NORMAL
        ws3.cell(r, 4).alignment = ALIGN_LEFT
        ws3.cell(r, 4).border = BORDER
        ws3.cell(r, 5, status).font = FONT_NORMAL
        ws3.cell(r, 5).alignment = ALIGN_CENTER
        ws3.cell(r, 5).border = BORDER
        # cor por severidade
        if sev == "Alta":
            for c in range(1, 6):
                ws3.cell(r, c).fill = FILL_ALERT
        elif sev == "Média":
            for c in range(1, 6):
                ws3.cell(r, c).fill = FILL_PEND
        ws3.row_dimensions[r].height = 32
        r += 1

    r += 1
    ws3.cell(r, 1, "Legenda:").font = FONT_TOTAL
    r += 1
    ws3.cell(r, 1, "").fill = FILL_ALERT
    ws3.cell(r, 2, "Severidade Alta — bloqueia entrega ou impacta significativamente o orçamento").font = FONT_SMALL
    r += 1
    ws3.cell(r, 1, "").fill = FILL_PEND
    ws3.cell(r, 2, "Severidade Média — impacta valor mas não bloqueia").font = FONT_SMALL
    r += 1
    ws3.cell(r, 1, "Sem cor").font = FONT_SMALL
    ws3.cell(r, 2, "Severidade Baixa — observação ou ajuste cosmético").font = FONT_SMALL

    # ordem das abas: Resumo, Detalhamento, Pendências
    wb.active = 0

    # save (staging local + move pra Drive — evita corrupção)
    import tempfile, shutil
    staging = Path(tempfile.gettempdir()) / "canteiro_quantitativos_staging.xlsx"
    if staging.exists():
        staging.unlink()
    wb.save(staging)
    if DEST.exists():
        DEST.unlink()
    shutil.move(str(staging), str(DEST))

    print(f"✓ Gerado: {DEST}")
    print(f"  Tamanho: {DEST.stat().st_size/1024:.1f} KB")
    print(f"  TOTAL CANTEIRO: {fmt_brl(total_geral)}")
    print(f"  Funcionários: {n_funcionarios} | Equipe admin: {equipe_admin}")
    print(f"  Pendências sinalizadas: {len(pendencias)}")


if __name__ == "__main__":
    main()
