#!/usr/bin/env python3
"""Pré-popula um gate.xlsx com respostas do briefing + análise arquitetônica.

Lê:
- base/pacotes/{slug}/gate-{slug}.xlsx (gate base)
- base/pacotes/{slug}/analise-arquitetura.json (decisões inferidas do Bloco 0)
- briefing inline via --respostas (dict JSON)

Gera:
- base/pacotes/{slug}/gate-{slug}-validado.xlsx
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
PACOTES = BASE / "pacotes"


def aplicar_respostas(gate_xlsx: Path, respostas: dict, decisoes_arq: dict, output: Path) -> dict:
    wb = load_workbook(gate_xlsx)

    if "GATE" not in wb.sheetnames:
        raise RuntimeError(f"Aba GATE não encontrada em {gate_xlsx}")

    ws = wb["GATE"]
    decisoes_aplicadas = []
    decisoes_n_encontradas = []

    for row in ws.iter_rows(min_row=5):
        if not row or not row[0].value:
            continue
        decisao_label = str(row[0].value).strip()
        if decisao_label in respostas:
            row[1].value = respostas[decisao_label]
            decisoes_aplicadas.append((decisao_label, respostas[decisao_label]))
        else:
            current = row[1].value
            if current:
                decisoes_n_encontradas.append((decisao_label, current))

    if "PREMISSAS_PROPOSTAS" in wb.sheetnames:
        ws_pre = wb["PREMISSAS_PROPOSTAS"]
        n_premissas = sum(1 for r in ws_pre.iter_rows(min_row=5) if r and r[0].value)

    if "SUB_DISCIPLINAS" in wb.sheetnames:
        ws_sd = wb["SUB_DISCIPLINAS"]
        n_subdisc = sum(1 for r in ws_sd.iter_rows(min_row=5) if r and r[0].value)

    if decisoes_arq:
        if "ANALISE_ARQUITETONICA" not in wb.sheetnames:
            ws_arq = wb.create_sheet("ANALISE_ARQUITETONICA")
            ws_arq["A1"] = "ANÁLISE ARQUITETÔNICA — Bloco 0"
            ws_arq["A1"].font = ws_arq["A1"].font.copy(bold=True, size=12)
            ws_arq["A2"] = "Itens de lazer/sistemas inferidos do projeto arquitetônico (IFC + DXF + PDF)"
            ws_arq["A3"] = "Item"
            ws_arq["B3"] = "Detectado?"
            for c in [ws_arq["A3"], ws_arq["B3"]]:
                c.font = c.font.copy(bold=True)
            row = 4
            for k, v in decisoes_arq.items():
                ws_arq.cell(row, 1, k)
                ws_arq.cell(row, 2, "✓ Sim" if v else "—")
                row += 1

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        "output": str(output),
        "decisoes_aplicadas": len(decisoes_aplicadas),
        "decisoes_default": len(decisoes_n_encontradas),
        "lista_aplicadas": decisoes_aplicadas,
    }


# Respostas do briefing duvidas-projetos-noturno-2026-04-13.md
RESPOSTAS_PROJETOS = {
    "arthen-arboris": {
        "Tipo de Laje": "Convencional",
        "Tipo de Fundação": "Hélice contínua",
        "Subsolos": "0",
        "Padrão de Acabamento": "Médio",
        "Tipo de Fachada": "Textura projetada",
        "Tipo de Esquadria": "Alumínio anodizado",
        "Tipo de Piso predominante": "Misto",
        "Pintura padrão": "Acrílica + selador",
        "Climatização": "Infra apenas",
        "Sistema de Pressurização": "Não",
        "Gerador Dedicado": "Sim",
        "Piscina": "Sim",
        "Perdas estimadas (concreto)": "10%",
        "Prazo de obra (meses)": "36",
        "BDI total estimado": "25%",
        "Encargos (sobre MO)": "85%",
        "Tempo de protensão (meses)": "0",
        "Tipologia predominante": "Misto",
        "Garagem (vagas/UR)": "1.0",
        "Tipo de Contenção": "Não",
    },
    "placon-arminio-tavares": {
        "Tipo de Laje": "Convencional",
        "Tipo de Fundação": "Hélice contínua",
        "Subsolos": "1",
        "Padrão de Acabamento": "Médio",
        "Tipo de Fachada": "Textura projetada",
        "Tipo de Esquadria": "Alumínio anodizado",
        "Tipo de Piso predominante": "Porcelanato",
        "Pintura padrão": "Acrílica + selador",
        "Climatização": "Infra apenas",
        "Sistema de Pressurização": "Não",
        "Gerador Dedicado": "Sim",
        "Piscina": "Não",
        "Perdas estimadas (concreto)": "13%",
        "Prazo de obra (meses)": "24",
        "BDI total estimado": "25%",
        "Encargos (sobre MO)": "85%",
        "Tempo de protensão (meses)": "0",
        "Tipologia predominante": "Studios",
        "Garagem (vagas/UR)": "1.0",
        "Tipo de Contenção": "Cortina",
    },
    "thozen-electra": {
        "Tipo de Laje": "Protendida",
        "Tipo de Fundação": "Hélice contínua",
        "Subsolos": "1",
        "Padrão de Acabamento": "Alto",
        "Tipo de Fachada": "Pastilha",
        "Tipo de Esquadria": "Alumínio anodizado",
        "Tipo de Piso predominante": "Porcelanato",
        "Pintura padrão": "Acrílica + selador",
        "Climatização": "Infra + equipamentos",
        "Sistema de Pressurização": "Sim",
        "Gerador Dedicado": "Sim",
        "Piscina": "Sim",
        "Perdas estimadas (concreto)": "13%",
        "Prazo de obra (meses)": "36",
        "BDI total estimado": "25%",
        "Encargos (sobre MO)": "85%",
        "Tempo de protensão (meses)": "12",
        "Tipologia predominante": "1-2 dorms",
        "Garagem (vagas/UR)": "1.5",
        "Tipo de Contenção": "Cortina",
    },
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slug", required=True)
    args = ap.parse_args()

    pasta = PACOTES / args.slug
    gate_xlsx = pasta / f"gate-{args.slug}.xlsx"
    output = pasta / f"gate-{args.slug}-validado.xlsx"

    if not gate_xlsx.exists():
        raise FileNotFoundError(f"Gate base não encontrado: {gate_xlsx}")

    respostas = RESPOSTAS_PROJETOS.get(args.slug, {})
    if not respostas:
        print(f"[WARN] Sem respostas pré-definidas para {args.slug} — gerando cópia do default")

    decisoes_arq = {}
    arq_path = pasta / "analise-arquitetura.json"
    if arq_path.exists():
        try:
            arq = json.loads(arq_path.read_text(encoding="utf-8"))
            decisoes_arq = arq.get("decisoes_inferidas", {})
        except Exception as e:
            print(f"[WARN] erro lendo análise arquitetônica: {e}")

    result = aplicar_respostas(gate_xlsx, respostas, decisoes_arq, output)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
