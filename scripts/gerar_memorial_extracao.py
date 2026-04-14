#!/usr/bin/env python3
"""Phase 17 — Memorial de Extração por projeto.

Gera um documento (.md) detalhando, item por item do orçamento, QUAL foi
a lógica de extração:
- Descrição do item
- Qtd + unidade
- Fonte do dado (IFC path / DXF path / PDF / calibration / base qualitativa)
- PU aplicado + fonte do PU
- Total
- Nível de confiança

Entradas:
- base/quantitativos-consolidados/[projeto].json (BIM + DXF + PDF)
- base/pacotes/[projeto]/executivo-[projeto].xlsx (o executivo gerado)
- base/indices-derivados-v2.json (29 índices calibrados)
- base/itens-pus-agregados.json (4.210 PUs cross-projeto)

Saída: base/pacotes/[projeto]/memorial-extracao-[projeto].md
"""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
PACOTES = BASE / "pacotes"
CONS = BASE / "quantitativos-consolidados"


def fmt_money(v) -> str:
    if v is None or v == 0:
        return "—"
    try:
        return f"R$ {float(v):,.0f}".replace(",", ".")
    except Exception:
        return str(v)


def fmt_num(v, dec=2) -> str:
    if v is None:
        return "—"
    try:
        return f"{float(v):,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


STOPWORDS = {"de", "da", "do", "e", "em", "para", "com", "no", "na", "a", "o",
              "as", "os", "kg", "m", "m2", "m3", "un"}


def canonicalize(desc):
    if not desc:
        return ""
    s = strip_accents(desc.lower())
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b\d+([,\.]\d+)?\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def token_hash_key(canon):
    tokens = [t for t in canon.split() if t not in STOPWORDS and len(t) > 2]
    if not tokens:
        return ""
    tokens.sort(key=lambda t: -len(t))
    top = tokens[:4]
    top.sort()
    return "|".join(top)


def load_pus_lookup():
    data = json.loads((BASE / "itens-pus-agregados.json").read_text(encoding="utf-8"))
    return {p["key"]: p for p in data.get("pus_agregados", []) if p.get("key")}


def load_derivados():
    return json.loads((BASE / "indices-derivados-v2.json").read_text(encoding="utf-8")).get("indices", {})


def confidence_tag(n_obs):
    if n_obs >= 10:
        return "🟢 Alta"
    if n_obs >= 3:
        return "🟡 Média"
    return "🔴 Baixa"


def gerar_memorial(slug: str) -> Path | None:
    cons_path = CONS / f"{slug}.json"
    if not cons_path.exists():
        print(f"  {slug}: sem quantitativos consolidados")
        return None

    cons = json.loads(cons_path.read_text(encoding="utf-8"))
    exec_xlsx = PACOTES / slug / f"executivo-{slug}.xlsx"
    if not exec_xlsx.exists():
        print(f"  {slug}: sem executivo gerado")
        return None

    state_path = PACOTES / slug / "state.json"
    state = json.loads(state_path.read_text(encoding="utf-8")) if state_path.exists() else {}

    pus_lookup = load_pus_lookup()
    derivados = load_derivados()

    wb = load_workbook(exec_xlsx, data_only=True)

    ac = state.get("ac") or cons.get("ac_referencia_m2", 0)
    ur = state.get("ur") or cons.get("ur_referencia", 0)
    padrao = state.get("padrao") or "—"

    lines = [
        f"# Memorial de Extração — {slug}",
        "",
        f"_Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — Fase 17 (base enriquecida)_",
        "",
        "Este documento detalha a **lógica de extração** de cada item do orçamento executivo, rastreando fontes de dados e lógica de precificação.",
        "",
        "## 📊 Dados do Projeto",
        "",
        "| Campo | Valor |",
        "|---|---|",
        f"| **Slug** | `{slug}` |",
        f"| **AC (Área Construída)** | {fmt_num(ac, 2)} m² |",
        f"| **UR (Unidades)** | {ur} |",
        f"| **Padrão** | {padrao} |",
        "",
        "## 🏗 Quantitativos Extraídos do BIM",
        "",
        f"**Fontes:** {cons['fontes']['bim']} IFCs, {cons['fontes']['dxf']} DXFs, {cons['fontes']['pdf']} PDFs",
        "",
    ]

    q = cons.get("quantitativos", {})

    alv = q.get("alvenaria_e_paredes", {})
    if alv.get("n_unidades"):
        lines.append("### Alvenaria e Paredes (IfcWall)")
        lines.append("")
        lines.append(f"- **Total de elementos:** {alv['n_unidades']} paredes")
        lines.append(f"- **Área total:** {fmt_num(alv['area_total_m2'], 0)} m²")
        lines.append(f"- **Comprimento total:** {fmt_num(alv['comprimento_total_m'], 0)} m")
        lines.append("")
        lines.append("**Por tipo:**")
        lines.append("")
        lines.append("| Tipo (material + espessura) | Qtd | Área m² | Comp m | Exemplos no IFC |")
        lines.append("|---|---|---|---|---|")
        for key, data in sorted(alv.get("por_tipo", {}).items(),
                                 key=lambda x: -x[1].get("area_m2", 0))[:15]:
            exemplos = " / ".join(data.get("exemplos", [])[:2])[:80]
            lines.append(f"| {key} | {data['qtd']} | {fmt_num(data['area_m2'], 0)} | {fmt_num(data['comprimento_m'], 0)} | {exemplos} |")
        lines.append("")

    estr = q.get("estrutura", {})
    if estr:
        lines.append("### Estrutura (lajes + vigas + pilares)")
        lines.append("")
        lines.append("| Elemento | Quantidade | Volume m³ | Comprimento/Altura m | Fonte |")
        lines.append("|---|---|---|---|---|")
        lajes = estr.get("lajes", {})
        lines.append(f"| Lajes (IfcSlab) | {lajes.get('n', 0)} | {fmt_num(lajes.get('volume_m3', 0), 1)} | área {fmt_num(lajes.get('area_m2', 0), 0)} m² | {lajes.get('fonte', '')} |")
        vigas = estr.get("vigas", {})
        lines.append(f"| Vigas (IfcBeam) | {vigas.get('n', 0)} | {fmt_num(vigas.get('volume_m3', 0), 1)} | {fmt_num(vigas.get('comprimento_m', 0), 0)} | {vigas.get('fonte', '')} |")
        pil = estr.get("pilares", {})
        lines.append(f"| Pilares (IfcColumn) | {pil.get('n', 0)} | {fmt_num(pil.get('volume_m3', 0), 1)} | {fmt_num(pil.get('altura_total_m', 0), 0)} | {pil.get('fonte', '')} |")
        lines.append("")

        conc = estr.get("concreto_total_m3_estimado")
        if conc:
            lines.append(f"- **Concreto total estimado (BIM):** {fmt_num(conc, 1)} m³")
            por_m2 = estr.get("concreto_por_m2_ac_estimado")
            if por_m2:
                mediana_calibrada = 0.34
                delta_pct = (por_m2 - mediana_calibrada) / mediana_calibrada * 100
                status = "✅" if abs(delta_pct) < 20 else ("⚠️" if abs(delta_pct) < 50 else "🔴")
                lines.append(f"- **Índice concreto/m² AC:** {fmt_num(por_m2, 3)} m³/m² (vs mediana base 0,34) {status} delta {delta_pct:+.0f}%")
        lines.append("")

    esq = q.get("esquadrias_e_aberturas", {})
    if esq:
        lines.append("### Esquadrias e Aberturas")
        lines.append("")
        portas = esq.get("portas", {})
        janelas = esq.get("janelas", {})
        cw = esq.get("pele_de_vidro", {})
        gc = esq.get("guarda_corpos", {})
        lines.append(f"- **Portas (IfcDoor):** {portas.get('n', 0)} unidades")
        lines.append(f"- **Janelas (IfcWindow):** {janelas.get('n', 0)} unidades")
        if cw.get("n"):
            lines.append(f"- **Pele de vidro (IfcCurtainWall):** {cw['n']} elementos, {fmt_num(cw.get('area_m2', 0), 0)} m²")
        if gc.get("n"):
            lines.append(f"- **Guarda-corpos (IfcRailing):** {gc['n']} elementos, {fmt_num(gc.get('comprimento_m', 0), 0)} m")
        lines.append("")

        if portas.get("top_tipos"):
            lines.append("**Top 10 tipos de porta:**")
            lines.append("")
            for k, v in list(portas["top_tipos"].items())[:10]:
                lines.append(f"- **{k}**: {v} un")
            lines.append("")

        if janelas.get("top_tipos"):
            lines.append("**Top 10 tipos de janela:**")
            lines.append("")
            for k, v in list(janelas["top_tipos"].items())[:10]:
                lines.append(f"- **{k}**: {v} un")
            lines.append("")

    amb = q.get("ambientes", {})
    if amb.get("n"):
        lines.append("### Ambientes (IfcSpace)")
        lines.append("")
        lines.append(f"- **Total:** {amb['n']} ambientes")
        lines.append(f"- **Área total:** {fmt_num(amb['area_total_m2'], 0)} m²")
        if amb.get("volume_total_m3"):
            lines.append(f"- **Volume total:** {fmt_num(amb['volume_total_m3'], 0)} m³")
        lines.append("")
        if amb.get("top_tipos"):
            lines.append("**Top 15 tipos de ambiente por área:**")
            lines.append("")
            lines.append("| Ambiente | Qtd | Área m² |")
            lines.append("|---|---|---|")
            for k, v in list(amb["top_tipos"].items())[:15]:
                lines.append(f"| {k} | {v.get('n', 0)} | {fmt_num(v.get('area_m2', 0), 0)} |")
            lines.append("")

    lines.append("## 💰 Orçamento Executivo — Lógica de Extração por Item")
    lines.append("")
    lines.append("Cada item abaixo mostra: descrição, PU usado, PU base cross-projeto (se disponível), faixa P10-P90 e fonte.")
    lines.append("")

    total_items = 0
    matched_items = 0

    for sn in wb.sheetnames:
        if sn in ("RESUMO", "REFERENCIAS", "PREMISSAS", "ANALISE_ARQUITETONICA"):
            continue
        ws = wb[sn]

        items = []
        for r in ws.iter_rows(min_row=5, max_row=50, values_only=True):
            if not r or not r[0] or not r[1]:
                continue
            desc = str(r[1]).strip()
            un = r[2] if len(r) > 2 else ""
            qtd = r[3] if len(r) > 3 else None
            pu = r[4] if len(r) > 4 else None
            total = r[5] if len(r) > 5 else None
            items.append({"desc": desc, "un": un, "qtd": qtd, "pu": pu, "total": total})

        if not items:
            continue

        lines.append(f"### {sn}")
        lines.append("")
        lines.append("| Descrição | Un | Qtd | PU usado | PU base mediano | Faixa P10-P90 | n obs | Conf |")
        lines.append("|---|---|---|---|---|---|---|---|")

        for it in items[:20]:
            total_items += 1
            desc = it["desc"][:70]
            un = it["un"] or "—"
            qtd = fmt_num(it["qtd"], 2) if it["qtd"] else "—"
            pu = fmt_money(it["pu"]) if it["pu"] else "—"

            canon = canonicalize(it["desc"])
            key = token_hash_key(canon)
            match = pus_lookup.get(key)

            if match:
                matched_items += 1
                pu_med = f"R$ {match['pu_mediana']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                faixa = f"R$ {match['pu_p10']:.0f} - {match['pu_p90']:.0f}"
                n_obs = match["n_observacoes"]
                conf = confidence_tag(n_obs)
            else:
                pu_med = "—"
                faixa = "—"
                n_obs = "—"
                conf = "🔴 Sem ref"

            lines.append(f"| {desc} | {un} | {qtd} | {pu} | {pu_med} | {faixa} | {n_obs} | {conf} |")

        lines.append("")

    wb.close()

    lines.append("## 📈 Comparação com Índices Derivados (Fase 13)")
    lines.append("")
    lines.append("Totais esperados por macrogrupo baseados em 29 índices derivados de 126 projetos:")
    lines.append("")
    lines.append("| Índice | Mediana | × AC = Esperado | Status |")
    lines.append("|---|---|---|---|")

    targets = [
        ("custo_concreto_rsm2", "Concreto"),
        ("custo_aco_rsm2", "Aço"),
        ("custo_forma_rsm2", "Forma"),
        ("custo_escoramento_rsm2", "Escoramento"),
        ("custo_impermeabilizacao_rsm2", "Impermeabilização"),
        ("custo_elevador_rsm2", "Elevadores"),
        ("custo_pintura_rsm2", "Pintura"),
        ("custo_esquadrias_rsm2", "Esquadrias"),
        ("custo_loucas_rsm2", "Louças"),
    ]
    for key, label in targets:
        idx = derivados.get(key)
        if not idx:
            continue
        med = idx.get("mediana", 0)
        esperado = med * ac
        n = idx.get("n", 0)
        lines.append(f"| {label} | R$ {med:.0f}/m² | R$ {esperado:,.0f} | n={n} |".replace(",", "."))
    lines.append("")

    lines.append("## 📋 Resumo de Rastreabilidade")
    lines.append("")
    lines.append(f"- **Itens do executivo analisados:** {total_items}")
    lines.append(f"- **Com match na base de 4.210 PUs cross-projeto:** {matched_items} ({matched_items/total_items*100:.0f}% se total > 0)")
    lines.append(f"- **Sem referência direta:** {total_items - matched_items}")
    lines.append("")
    lines.append("## 🔗 Fontes e Arquivos")
    lines.append("")
    lines.append(f"- **Quantitativos BIM consolidados:** `base/quantitativos-consolidados/{slug}.json`")
    lines.append(f"- **BIM raw:** `base/quantitativos-bim/{slug}.json`")
    lines.append(f"- **DXF raw:** `base/quantitativos-dxf/{slug}.json`")
    lines.append(f"- **PDF raw:** `base/quantitativos-pdf/{slug}.json`")
    lines.append(f"- **Executivo:** `base/pacotes/{slug}/executivo-{slug}.xlsx`")
    lines.append(f"- **PUs cross-projeto:** `base/itens-pus-agregados.json` (4.210 clusters)")
    lines.append(f"- **Índices derivados:** `base/indices-derivados-v2.json` (29 índices)")
    lines.append(f"- **Base master:** `base/base-indices-master-2026-04-13.json`")
    lines.append("")

    out = PACOTES / slug / f"memorial-extracao-{slug}.md"
    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"  {slug}: saved {out.name} ({out.stat().st_size // 1024} KB)  items={total_items} matched={matched_items}")
    return out


def main():
    for slug in ["arthen-arboris", "placon-arminio-tavares", "thozen-electra"]:
        print(f"\n=== {slug} ===")
        gerar_memorial(slug)


if __name__ == "__main__":
    main()
