#!/usr/bin/env python3
"""Gera catálogo único de todos os índices da base Cartesian.

Output: base/INDICES-CATALOGO.xlsx + base/INDICES-CATALOGO.md

10 abas no Excel:
1. LEIA_ME — mapa das abas
2. PROJETOS — 126 projetos com padrão Gemma
3. CALIBRACAO_GLOBAL — 18 MGs global
4. CALIBRACAO_CONDICIONAL — 18 MGs × 5 padrões Gemma
5. INDICES_DERIVADOS_V2 — 29 índices derivados
6. INDICES_ESTRUTURAIS — consumos físicos + segmentos + instalações + ci + produto
7. PUS_CROSS_V1 — 1.740 clusters com lista de projetos
8. PUS_CROSS_V2 — 4.210 clusters sem lista (mais cobertura)
9. CURVA_ABC_MASTER — 126 projetos (n itens, valor total)
10. CROSS_INSIGHTS_GEMMA — flatten dos insights Gemma

Uso:
    python scripts/gerar_catalogo_indices.py
"""
from __future__ import annotations

import glob
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
OUT_XLSX = BASE / "INDICES-CATALOGO.xlsx"
OUT_MD = BASE / "INDICES-CATALOGO.md"

# Styles
DARK = "2C3E50"
ACCENT = "2980B9"
ORANGE = "E67E22"
GREEN = "27AE60"
PURPLE = "8E44AD"
RED = "C0392B"
GRAY = "7F8C8D"
GREEN_BG = "E8F5E9"
ORANGE_BG = "FFF3E0"

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ABA_COLORS = {
    "LEIA_ME": GREEN,
    "PROJETOS": ACCENT,
    "CALIBRACAO_GLOBAL": ORANGE,
    "CALIBRACAO_CONDICIONAL": ORANGE,
    "INDICES_DERIVADOS_V2": PURPLE,
    "INDICES_ESTRUTURAIS": PURPLE,
    "PUS_CROSS_V1": RED,
    "PUS_CROSS_V2": RED,
    "CURVA_ABC_MASTER": GRAY,
    "CROSS_INSIGHTS_GEMMA": GRAY,
}


def _load_json(p: Path) -> dict | list:
    if not p.exists():
        print(f"  [warn] não existe: {p}")
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  [warn] falha lendo {p}: {e}")
        return {}


def load_sources() -> dict:
    print("Carregando fontes...")
    s = {
        "calibration": _load_json(BASE / "calibration-indices.json"),
        "condicional": _load_json(BASE / "calibration-condicional-padrao.json"),
        "master": _load_json(BASE / "base-indices-master-2026-04-13.json"),
        "pus_v1": _load_json(BASE / "base-pus-cartesian.json"),
        "pus_v2": _load_json(BASE / "itens-pus-agregados.json"),
        "padroes": _load_json(BASE / "padroes-classificados-consolidado.json"),
    }
    # Carregar indices-executivo dos 126
    exec_dir = BASE / "indices-executivo"
    idx_exec = {}
    if exec_dir.exists():
        for p in sorted(exec_dir.glob("*.json")):
            try:
                d = json.loads(p.read_text(encoding="utf-8"))
                idx_exec[p.stem] = d
            except Exception:
                pass
    s["indices_exec"] = idx_exec
    print(f"  calibration: {len(s['calibration'].get('por_macrogrupo', {}) if isinstance(s['calibration'], dict) else {})} MGs")
    print(f"  condicional: {len(s['condicional'].get('por_padrao_mg', {}) if isinstance(s['condicional'], dict) else {})} buckets de padrão")
    print(f"  derivados_v2: {len(s['master'].get('indices_derivados_v2', {}) if isinstance(s['master'], dict) else {})}")
    print(f"  pus_v1: {len(s['pus_v1']) if isinstance(s['pus_v1'], dict) else 0}")
    print(f"  pus_v2: {len(s['pus_v2'].get('pus_agregados', []) if isinstance(s['pus_v2'], dict) else [])}")
    print(f"  indices-executivo: {len(idx_exec)} projetos")
    print(f"  padroes_gemma: {len(s['padroes'].get('projetos', []) if isinstance(s['padroes'], dict) else [])} classificados")
    return s


def style_header(cell, color: str = DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def write_headers(ws, row: int, headers: list[str], widths: list[int]):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        style_header(c)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row: int, values: list, number_formats: dict[int, str] | None = None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if number_formats and i in number_formats:
            c.number_format = number_formats[i]
        elif isinstance(v, float):
            c.number_format = "#,##0.00"
        elif isinstance(v, int):
            c.number_format = "#,##0"


def _num(v):
    """Return value if numeric else None."""
    if isinstance(v, (int, float)):
        return v
    return None


def _round(v, n=4):
    return round(v, n) if isinstance(v, (int, float)) else None


# =============================================================================
# ABA 1: LEIA_ME
# =============================================================================

def build_aba_leia_me(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("LEIA_ME")
    ws.sheet_properties.tabColor = ABA_COLORS["LEIA_ME"]

    ws["A1"] = "CATÁLOGO DE ÍNDICES CARTESIAN — Base V2 consolidada"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Gerado em {datetime.now().isoformat(timespec='seconds')} via scripts/gerar_catalogo_indices.py"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")

    ws["A4"] = "PARA QUE SERVE"
    ws["A4"].font = Font(bold=True, size=11, color=DARK, name="Arial")
    ws["A5"] = ("Catálogo navegável de TODOS os índices da base Cartesian. "
                "Use filtros e ordenação por coluna pra encontrar índices por categoria, "
                "faixa de valores, número de projetos que contribuem, etc.")
    ws.merge_cells("A5:D5")
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A7"] = "MAPA DAS ABAS"
    ws["A7"].font = Font(bold=True, size=11, color=DARK, name="Arial")

    write_headers(ws, 9, ["#", "Aba", "N linhas", "Descrição", "Fonte"],
                  [4, 26, 10, 60, 40])

    n_projs = len(sources["indices_exec"])
    pm_count = len(sources["calibration"].get("por_macrogrupo", {})) if isinstance(sources["calibration"], dict) else 0
    cond = sources["condicional"] if isinstance(sources["condicional"], dict) else {}
    n_cond = sum(len(v) for v in (cond.get("por_padrao_mg", {}) or {}).values()) if cond else 0
    n_der = len(sources["master"].get("indices_derivados_v2", {})) if isinstance(sources["master"], dict) else 0
    n_pus_v1 = len(sources["pus_v1"]) if isinstance(sources["pus_v1"], dict) else 0
    n_pus_v2 = len(sources["pus_v2"].get("pus_agregados", [])) if isinstance(sources["pus_v2"], dict) else 0

    abas_info = [
        ("1", "LEIA_ME", 1, "Este mapa + schema + exemplos de uso", "-"),
        ("2", "PROJETOS", n_projs, "126 projetos da base com padrão Gemma, AC, UR, R$/m², cidade", "padroes-classificados + indices-executivo"),
        ("3", "CALIBRACAO_GLOBAL", pm_count, "18 macrogrupos globais (stats sem segmentação de padrão)", "calibration-indices.json"),
        ("4", "CALIBRACAO_CONDICIONAL", n_cond, "18 MGs × 5 padrões Gemma (economico/medio/medio-alto/alto/luxo) — fonte primária pós-fase 18b", "calibration-condicional-padrao.json"),
        ("5", "INDICES_DERIVADOS_V2", n_der, "29 índices derivados: PU insumos, custo por MG, splits MO/Material, curva ABC", "base-indices-master.json:indices_derivados_v2"),
        ("6", "INDICES_ESTRUTURAIS", 35, "Consumos físicos (concreto m³/m², aço kg/m³), produto, instalações %, ci %, segmentos por porte", "calibration-indices.json:{estruturais,produto,instalacoes,ci,por_segmento}"),
        ("7", "PUS_CROSS_V1", n_pus_v1, "1.740 clusters de PU cross-projeto COM lista de obras fonte (fase 10 v1)", "base-pus-cartesian.json"),
        ("8", "PUS_CROSS_V2", n_pus_v2, "4.210 clusters de PU cross-projeto (fase 10 v2 hash-based, SEM lista de obras)", "itens-pus-agregados.json"),
        ("9", "CURVA_ABC_MASTER", 126, "Curva ABC consolidada por projeto — n itens, itens A, valor total", "base-indices-master.json:curva_abc_master"),
        ("10", "CROSS_INSIGHTS_GEMMA", 0, "Insights Gemma cross-projeto: famílias, índices sugeridos, lacunas, outliers, padrões comuns", "base-indices-master.json:cross_insights"),
    ]
    for i, (num, aba, n, desc, fonte) in enumerate(abas_info, start=10):
        write_row(ws, i, [num, aba, n, desc, fonte])

    ws["A22"] = "SCHEMA DAS COLUNAS NUMÉRICAS"
    ws["A22"].font = Font(bold=True, size=11, color=DARK, name="Arial")
    schema = [
        ("n", "Número de projetos ou observações que contribuíram pro cálculo"),
        ("min / max", "Valores extremos observados"),
        ("p10 / p25 / p75 / p90", "Percentis (p10 = 10% dos projetos abaixo desse valor)"),
        ("mediana", "Valor central (p50) — é o que usamos como 'valor típico'"),
        ("media", "Média aritmética — pode ser puxada por outliers, menos robusta que mediana"),
        ("cv", "Coeficiente de variação (desvio/média). <0.3 é confiável, >0.5 é volátil"),
        ("projetos_fonte", "[só PUS_CROSS_V1] lista de slugs das obras que bancam esse índice"),
    ]
    write_headers(ws, 24, ["Coluna", "Significado"], [20, 80])
    for i, (col, desc) in enumerate(schema, start=25):
        write_row(ws, i, [col, desc])

    ws["A33"] = "EXEMPLOS DE FILTROS ÚTEIS"
    ws["A33"].font = Font(bold=True, size=11, color=DARK, name="Arial")
    exemplos = [
        ("PUs robustos de concreto", "Aba PUS_CROSS_V2 → filtro 'key' contém 'concreto' + 'n_projetos' ≥ 10 + 'cv' < 0.3"),
        ("Esquadrias alto vs luxo", "Aba CALIBRACAO_CONDICIONAL → filtro 'padrao' IN ('alto','luxo') + 'macrogrupo' = 'Esquadrias'"),
        ("Quais obras bancam porcelanato?", "Aba PUS_CROSS_V1 → filtro 'descricao' contém 'porcelanato' → coluna 'projetos_fonte'"),
        ("Projetos alto padrão", "Aba PROJETOS → filtro 'padrao_gemma' = 'alto' → ordenar por 'rsm2' desc"),
        ("Índices derivados mais confiáveis", "Aba INDICES_DERIVADOS_V2 → ordenar 'n' desc → ignorar entries com n<5"),
        ("MGs com mais desvio por padrão", "Aba CALIBRACAO_CONDICIONAL → calcular coluna p90-p10 → ordenar desc"),
    ]
    write_headers(ws, 35, ["Pergunta", "Como responder"], [35, 75])
    for i, (q, r) in enumerate(exemplos, start=36):
        write_row(ws, i, [q, r])

    ws["A44"] = "GAPS CONHECIDOS"
    ws["A44"].font = Font(bold=True, size=11, color=DARK, name="Arial")
    gaps = [
        "PUS_CROSS_V2 (4.210 clusters) NÃO tem lista de projetos — use PUS_CROSS_V1 (1.740) quando precisar rastrear obras",
        "Alguns MGs em CALIBRACAO_CONDICIONAL têm n<3 pro bucket — script paramétrico cai no fallback global × PADRAO_MULTIPLIERS",
        "custo_por_ur nos derivados tem apenas n=2 — baixa confiança",
        "Classe 'luxo' em CALIBRACAO_CONDICIONAL tem 0 projetos (Gemma não encontrou luxo na base Cartesian)",
        "Campos disciplinas/indices_consumo/split_mo_material dentro de cada indices-executivo/*.json estão majoritariamente vazios — dados agregados estão nos arquivos calibration*",
    ]
    for i, g in enumerate(gaps, start=45):
        ws.cell(i, 1, f"• {g}").font = Font(size=9, name="Arial")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    ws["A52"] = "COMO REGENERAR ESTE CATÁLOGO"
    ws["A52"].font = Font(bold=True, size=11, color=DARK, name="Arial")
    ws["A53"] = "cd ~/orcamentos-openclaw && python scripts/gerar_catalogo_indices.py"
    ws["A53"].font = Font(name="Consolas", size=9)
    ws["A53"].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")


# =============================================================================
# ABA 2: PROJETOS
# =============================================================================

def build_aba_projetos(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("PROJETOS")
    ws.sheet_properties.tabColor = ABA_COLORS["PROJETOS"]

    ws["A1"] = "PROJETOS DA BASE CARTESIAN (126 executivos)"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:K1")
    ws["A2"] = "Labels Gemma via fase 18 classificação semântica | AC/total via indices-executivo"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:K2")

    headers = ["slug", "padrao_gemma", "confianca", "ac_m2", "ur", "total_rs",
               "rsm2", "m2_por_ur", "rs_por_ur", "cidade", "fonte"]
    widths = [36, 14, 11, 12, 8, 16, 12, 12, 14, 18, 28]
    write_headers(ws, 4, headers, widths)

    # Cross-ref: padroes consolidado + indices-executivo
    padroes_dict = {}
    cons = sources["padroes"] if isinstance(sources["padroes"], dict) else {}
    for p in cons.get("projetos", []):
        padroes_dict[p["projeto"]] = p

    rows = []
    for slug, idx in sorted(sources["indices_exec"].items()):
        padrao_entry = padroes_dict.get(slug, {})
        ac = idx.get("ac") or padrao_entry.get("ac") or 0
        ur = idx.get("ur") or padrao_entry.get("ur") or 0
        total = idx.get("total") or padrao_entry.get("total") or 0
        rsm2 = (total / ac) if ac and total else 0
        m2_por_ur = (ac / ur) if ac and ur else 0
        rs_por_ur = (total / ur) if total and ur else 0
        rows.append((
            slug,
            padrao_entry.get("padrao", "—"),
            padrao_entry.get("confianca", "—"),
            round(ac, 2) if ac else 0,
            int(ur) if ur else 0,
            round(total, 0) if total else 0,
            round(rsm2, 2) if rsm2 else 0,
            round(m2_por_ur, 2) if m2_por_ur else 0,
            round(rs_por_ur, 0) if rs_por_ur else 0,
            idx.get("cidade", "") or padrao_entry.get("cidade", ""),
            padrao_entry.get("metodo", "gemma_semantico"),
        ))

    fmts = {4: '#,##0.00', 5: '#,##0', 6: '"R$" #,##0', 7: '"R$" #,##0.00',
            8: '#,##0.0', 9: '"R$" #,##0'}
    for i, row in enumerate(rows, start=5):
        write_row(ws, i, list(row), number_formats=fmts)

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 3: CALIBRACAO_GLOBAL
# =============================================================================

def build_aba_calibracao_global(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("CALIBRACAO_GLOBAL")
    ws.sheet_properties.tabColor = ABA_COLORS["CALIBRACAO_GLOBAL"]

    ws["A1"] = "CALIBRAÇÃO GLOBAL — 18 macrogrupos (sem segmentação de padrão)"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:L1")
    ws["A2"] = "Stats agregadas de todos os 126 projetos da base. Para calibração condicional por padrão, ver aba CALIBRACAO_CONDICIONAL."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:L2")

    headers = ["macrogrupo", "n", "min", "p10", "p25", "mediana", "media",
               "p75", "p90", "max", "unidade", "fonte"]
    widths = [24, 6, 12, 12, 12, 12, 12, 12, 12, 12, 12, 32]
    write_headers(ws, 4, headers, widths)

    cal = sources["calibration"] if isinstance(sources["calibration"], dict) else {}
    pm = cal.get("por_macrogrupo", {}) or {}

    # Mapeamento key do JSON -> nome limpo
    MG_NAME_MAP = {
        "Gerenciamento_rsm2": "Gerenciamento",
        "Mov.Terra_rsm2": "Movimentação de Terra",
        "Infraestrutura_rsm2": "Infraestrutura",
        "Supraestrutura_rsm2": "Supraestrutura",
        "Alvenaria_rsm2": "Alvenaria",
        "Impermeabilização_rsm2": "Impermeabilização",
        "Instalações_rsm2": "Instalações",
        "Sist.Especiais_rsm2": "Sistemas Especiais",
        "Climatização_rsm2": "Climatização",
        "Rev.Int.Parede_rsm2": "Rev. Interno Parede",
        "Teto_rsm2": "Teto",
        "Pisos_rsm2": "Pisos",
        "Pintura_rsm2": "Pintura",
        "Esquadrias_rsm2": "Esquadrias",
        "Louças_rsm2": "Louças e Metais",
        "Fachada_rsm2": "Fachada",
        "Complementares_rsm2": "Complementares",
        "Imprevistos_rsm2": "Imprevistos",
    }

    fmts = {i: '"R$" #,##0.00' for i in [3, 4, 5, 6, 7, 8, 9, 10]}
    fmts[2] = '#,##0'

    row_idx = 5
    for key, nome in MG_NAME_MAP.items():
        stats = pm.get(key)
        if not isinstance(stats, dict):
            continue
        write_row(ws, row_idx, [
            nome,
            int(stats.get("n", 0)),
            _round(stats.get("min"), 2),
            _round(stats.get("p10"), 2),
            _round(stats.get("p25"), 2),
            _round(stats.get("mediana"), 2),
            _round(stats.get("media"), 2),
            _round(stats.get("p75"), 2),
            _round(stats.get("p90"), 2),
            _round(stats.get("max"), 2),
            "R$/m²",
            "calibration-indices.json",
        ], number_formats=fmts)
        row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 4: CALIBRACAO_CONDICIONAL
# =============================================================================

def build_aba_calibracao_condicional(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("CALIBRACAO_CONDICIONAL")
    ws.sheet_properties.tabColor = ABA_COLORS["CALIBRACAO_CONDICIONAL"]

    ws["A1"] = "CALIBRAÇÃO CONDICIONAL POR PADRÃO GEMMA (Fase 18b)"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:M1")
    ws["A2"] = "Medianas por macrogrupo condicionadas à classe Gemma (economico/medio/medio-alto/alto/luxo). Fonte primária da calibração pós-fase 18b."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:M2")

    headers = ["padrao", "macrogrupo", "n", "min", "p10", "p25", "mediana",
               "media", "p75", "p90", "max", "unidade", "fonte"]
    widths = [14, 24, 6, 12, 12, 12, 12, 12, 12, 12, 12, 10, 28]
    write_headers(ws, 4, headers, widths)

    cond = sources["condicional"] if isinstance(sources["condicional"], dict) else {}
    por_padrao = cond.get("por_padrao_mg", {}) or {}

    PADROES_ORDEM = ["economico", "medio", "medio-alto", "alto", "luxo"]
    fmts = {i: '"R$" #,##0.00' for i in [4, 5, 6, 7, 8, 9, 10, 11]}
    fmts[3] = '#,##0'

    row_idx = 5
    for padrao in PADROES_ORDEM:
        mgs = por_padrao.get(padrao, {})
        for mg in sorted(mgs.keys()):
            stats = mgs.get(mg)
            if not isinstance(stats, dict):
                continue
            write_row(ws, row_idx, [
                padrao,
                mg,
                int(stats.get("n", 0)),
                _round(stats.get("min"), 2),
                _round(stats.get("p10"), 2),
                _round(stats.get("p25"), 2),
                _round(stats.get("mediana"), 2),
                _round(stats.get("media"), 2),
                _round(stats.get("p75"), 2),
                _round(stats.get("p90"), 2),
                _round(stats.get("max"), 2),
                "R$/m²",
                "calibration-condicional-padrao.json",
            ], number_formats=fmts)
            row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 5: INDICES_DERIVADOS_V2
# =============================================================================

def build_aba_derivados(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("INDICES_DERIVADOS_V2")
    ws.sheet_properties.tabColor = ABA_COLORS["INDICES_DERIVADOS_V2"]

    ws["A1"] = "29 ÍNDICES DERIVADOS V2 — PUs de insumos, custos por MG, splits"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:N1")
    ws["A2"] = "Derivados pelo script gerar_novos_indices.py cruzando base-pus-cartesian.json com indices-executivo/*.json"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:N2")

    headers = ["nome", "descricao", "n", "min", "p10", "p25", "mediana",
               "media", "p75", "p90", "max", "cv", "unidade", "fonte"]
    widths = [32, 50, 6, 12, 12, 12, 12, 12, 12, 12, 12, 8, 10, 28]
    write_headers(ws, 4, headers, widths)

    master = sources["master"] if isinstance(sources["master"], dict) else {}
    der = master.get("indices_derivados_v2", {}) or {}

    # Inferir unidade pelo nome
    def _infer_unit(nome: str) -> str:
        n = nome.lower()
        if "rsm2" in n or "_por_m2" in n: return "R$/m²"
        if "_por_ur" in n: return "R$/UR"
        if "por_aco" in n or "por_concreto" in n: return "ratio"
        if "pu_" in n and "mediano" in n:
            if "kg" in n or "aco" in n: return "R$/kg"
            if "_m3" in n or "concreto" in n: return "R$/m³"
            if "_m2" in n or "impermeab" in n or "porcelanato" in n or "bloco" in n or "forma" in n or "pintura" in n: return "R$/m²"
            return "R$/un"
        if "_pct" in n: return "%"
        return "—"

    fmts = {3: '#,##0'}
    for i in [4, 5, 6, 7, 8, 9, 10, 11]:
        fmts[i] = '#,##0.0000'
    fmts[12] = '0.000'

    row_idx = 5
    for nome, stats in der.items():
        if not isinstance(stats, dict):
            continue
        desc = stats.get("descricao", nome)
        unit = _infer_unit(nome)
        write_row(ws, row_idx, [
            nome,
            desc,
            int(stats.get("n", 0)),
            _round(stats.get("min"), 4),
            _round(stats.get("p10"), 4),
            _round(stats.get("p25"), 4),
            _round(stats.get("mediana"), 4),
            _round(stats.get("media"), 4),
            _round(stats.get("p75"), 4),
            _round(stats.get("p90"), 4),
            _round(stats.get("max"), 4),
            _round(stats.get("cv"), 3),
            unit,
            "base-indices-master-2026-04-13.json",
        ], number_formats=fmts)
        row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 6: INDICES_ESTRUTURAIS
# =============================================================================

def build_aba_estruturais(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("INDICES_ESTRUTURAIS")
    ws.sheet_properties.tabColor = ABA_COLORS["INDICES_ESTRUTURAIS"]

    ws["A1"] = "ÍNDICES ESTRUTURAIS / PRODUTO / INSTALAÇÕES / CI / SEGMENTOS"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:N1")
    ws["A2"] = "Consumos físicos (concreto m³/m², aço kg/m³), índices de produto (AC/UR), breakdown de instalações e CI, e segmentação por porte"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:N2")

    headers = ["categoria", "nome", "n", "min", "p10", "p25", "mediana",
               "media", "p75", "p90", "max", "unidade", "fonte"]
    widths = [16, 36, 6, 12, 12, 12, 12, 12, 12, 12, 12, 12, 28]
    write_headers(ws, 4, headers, widths)

    cal = sources["calibration"] if isinstance(sources["calibration"], dict) else {}

    # Categorias (ordem) + key no JSON + unidade
    SECOES = [
        ("Estruturais", "estruturais", {
            "concreto_m3_por_m2_ac": "m³/m²",
            "aco_kg_por_m3_concreto": "kg/m³",
            "aco_kg_por_m2_ac": "kg/m²",
            "forma_m2_por_m2_ac": "m²/m²",
        }),
        ("Produto", "produto", {
            "ac_por_ur": "m²/UR",
            "custo_por_ur": "R$/UR",
            "cub_ratio": "ratio",
            "burn_rate_mensal": "R$/mês",
            "elevador_pu_un": "R$/un",
        }),
        ("Instalações %", "instalacoes", {
            "hidrossanitarias_pct_total": "%",
            "eletricas_pct_total": "%",
            "preventivas_pct_total": "%",
            "gas_pct_total": "%",
            "telecom_pct_total": "%",
        }),
        ("Custos Indiretos %", "ci", {
            "projetos_consultorias_pct_total": "%",
            "taxas_licencas_pct_total": "%",
            "equipe_adm_pct_total": "%",
            "epcs_pct_total": "%",
            "equipamentos_carga_pct_total": "%",
            "ensaios_pct_total": "%",
            "canteiro_pct_total": "%",
        }),
        ("Segmento por porte", "por_segmento", {
            "pequeno_lt8k_rsm2": "R$/m²",
            "medio_8k_15k_rsm2": "R$/m²",
            "grande_15k_25k_rsm2": "R$/m²",
            "extra_gt25k_rsm2": "R$/m²",
        }),
    ]

    fmts = {3: '#,##0'}
    for i in [4, 5, 6, 7, 8, 9, 10, 11]:
        fmts[i] = '#,##0.0000'

    row_idx = 5
    for categoria, secao_key, nomes_unids in SECOES:
        secao = cal.get(secao_key, {}) or {}
        for nome, unidade in nomes_unids.items():
            stats = secao.get(nome)
            if not isinstance(stats, dict):
                continue
            write_row(ws, row_idx, [
                categoria,
                nome,
                int(stats.get("n", 0)),
                _round(stats.get("min"), 4),
                _round(stats.get("p10"), 4),
                _round(stats.get("p25"), 4),
                _round(stats.get("mediana"), 4),
                _round(stats.get("media"), 4),
                _round(stats.get("p75"), 4),
                _round(stats.get("p90"), 4),
                _round(stats.get("max"), 4),
                unidade,
                "calibration-indices.json",
            ], number_formats=fmts)
            row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 7: PUS_CROSS_V1 (com projetos fonte)
# =============================================================================

def build_aba_pus_v1(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("PUS_CROSS_V1")
    ws.sheet_properties.tabColor = ABA_COLORS["PUS_CROSS_V1"]

    ws["A1"] = "PUs CROSS-PROJETO V1 — 1.740 clusters COM lista de obras fonte"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:M1")
    ws["A2"] = "Fase 10 v1. Cada cluster tem lista de slugs das obras que contribuíram (coluna 'projetos_fonte')."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:M2")

    headers = ["categoria", "chave", "descricao", "unidade", "n_proj",
               "n_obs", "min", "p25", "mediana", "p75", "max", "cv", "projetos_fonte"]
    widths = [22, 36, 50, 8, 7, 7, 10, 10, 10, 10, 10, 8, 70]
    write_headers(ws, 4, headers, widths)

    pus_v1 = sources["pus_v1"] if isinstance(sources["pus_v1"], dict) else {}

    fmts = {i: '"R$" #,##0.00' for i in [7, 8, 9, 10, 11]}
    fmts[5] = '#,##0'
    fmts[6] = '#,##0'
    fmts[12] = '0.000'

    # Ordenar por categoria depois chave pra ficar navegável
    items = []
    for key, cluster in pus_v1.items():
        if not isinstance(cluster, dict):
            continue
        items.append((cluster.get("categoria", ""), key, cluster))
    items.sort(key=lambda x: (x[0], x[1]))

    row_idx = 5
    for categoria, _key, cluster in items:
        projs = cluster.get("projetos", []) or []
        projs_str = "; ".join(projs[:20])  # Truncate to first 20 to avoid cell overflow
        if len(projs) > 20:
            projs_str += f"  …(+{len(projs) - 20})"
        if len(projs_str) > 500:
            projs_str = projs_str[:497] + "..."

        write_row(ws, row_idx, [
            categoria,
            cluster.get("chave", ""),
            (cluster.get("descricao", "") or "")[:200],
            cluster.get("unidade", ""),
            int(cluster.get("n_projetos", 0)),
            int(cluster.get("n_observacoes", 0)),
            _round(cluster.get("min"), 2),
            _round(cluster.get("p25"), 2),
            _round(cluster.get("mediana"), 2),
            _round(cluster.get("p75"), 2),
            _round(cluster.get("max"), 2),
            _round(cluster.get("cv"), 3),
            projs_str,
        ], number_formats=fmts)
        row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 8: PUS_CROSS_V2 (mais clusters, sem projetos)
# =============================================================================

def build_aba_pus_v2(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("PUS_CROSS_V2")
    ws.sheet_properties.tabColor = ABA_COLORS["PUS_CROSS_V2"]

    ws["A1"] = "PUs CROSS-PROJETO V2 — 4.210 clusters (fase 10 v2)"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:O1")
    ws["A2"] = "Fase 10 v2 com hash-based clustering semântico. MAIS clusters que V1 (melhor cobertura), SEM lista de projetos."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:O2")

    headers = ["cluster_id", "key (tokens)", "descricao", "unidades", "n_proj", "n_obs",
               "pu_min", "pu_p10", "pu_p25", "pu_mediana", "pu_p75", "pu_p90", "pu_max", "pu_media", "cv"]
    widths = [9, 38, 50, 12, 7, 7, 10, 10, 10, 11, 10, 10, 10, 10, 8]
    write_headers(ws, 4, headers, widths)

    pus_v2 = sources["pus_v2"] if isinstance(sources["pus_v2"], dict) else {}
    agregados = pus_v2.get("pus_agregados", []) or []

    fmts = {i: '"R$" #,##0.00' for i in [7, 8, 9, 10, 11, 12, 13, 14]}
    fmts[5] = '#,##0'
    fmts[6] = '#,##0'
    fmts[15] = '0.000'

    # Ordenar por n_projetos desc (robustez) depois cluster_id
    items = sorted(agregados, key=lambda c: (-c.get("n_projetos", 0), c.get("cluster_id", 0)))

    row_idx = 5
    for cluster in items:
        unidades = cluster.get("unidades") or []
        un_str = ",".join(unidades[:3]) if isinstance(unidades, list) else str(unidades)
        write_row(ws, row_idx, [
            int(cluster.get("cluster_id", 0)),
            cluster.get("key", ""),
            (cluster.get("desc", "") or "")[:200],
            un_str[:20],
            int(cluster.get("n_projetos", 0)),
            int(cluster.get("n_observacoes", 0)),
            _round(cluster.get("pu_min"), 2),
            _round(cluster.get("pu_p10"), 2),
            _round(cluster.get("pu_p25"), 2),
            _round(cluster.get("pu_mediana"), 2),
            _round(cluster.get("pu_p75"), 2),
            _round(cluster.get("pu_p90"), 2),
            _round(cluster.get("pu_max"), 2),
            _round(cluster.get("pu_media"), 2),
            _round(cluster.get("cv"), 3),
        ], number_formats=fmts)
        row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 9: CURVA_ABC_MASTER
# =============================================================================

def build_aba_curva_abc(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("CURVA_ABC_MASTER")
    ws.sheet_properties.tabColor = ABA_COLORS["CURVA_ABC_MASTER"]

    ws["A1"] = "CURVA ABC MASTER — 126 projetos consolidados"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    master = sources["master"] if isinstance(sources["master"], dict) else {}
    cabc = master.get("curva_abc_master", {}) if isinstance(master, dict) else {}
    n_projs = cabc.get("n_projetos", 0)
    n_itens_totais = cabc.get("n_itens_totais", 0)

    ws["A2"] = (f"{n_projs} projetos × {n_itens_totais:,} itens. "
                f"'n_a' = quantos itens representam 80% do custo. Curva A compacta = projeto concentrado em poucos itens.")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    headers = ["slug", "status", "n_itens", "n_a", "pct_a", "valor_total_rs"]
    widths = [38, 10, 10, 10, 10, 18]
    write_headers(ws, 4, headers, widths)

    projs = cabc.get("projetos", []) or []

    fmts = {3: '#,##0', 4: '#,##0', 5: '0.0%', 6: '"R$" #,##0'}

    for i, p in enumerate(sorted(projs, key=lambda x: x.get("slug", "")), start=5):
        n_it = p.get("n_itens", 0) or 0
        n_a = p.get("n_a", 0) or 0
        pct_a = (n_a / n_it) if n_it else 0
        write_row(ws, i, [
            p.get("slug", ""),
            p.get("status", ""),
            int(n_it),
            int(n_a),
            pct_a,
            round(p.get("valor_total", 0) or 0, 0),
        ], number_formats=fmts)

    ws.freeze_panes = "A5"


# =============================================================================
# ABA 10: CROSS_INSIGHTS_GEMMA
# =============================================================================

def build_aba_cross_insights(wb: Workbook, sources: dict) -> None:
    ws = wb.create_sheet("CROSS_INSIGHTS_GEMMA")
    ws.sheet_properties.tabColor = ABA_COLORS["CROSS_INSIGHTS_GEMMA"]

    ws["A1"] = "CROSS INSIGHTS GEMMA — análises cross-projeto"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"] = "Gerados via Gemma analisando todos os projetos da base. Cada seção extrai um aspecto diferente."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")

    headers = ["secao", "tipo", "campo", "conteudo"]
    widths = [22, 14, 22, 100]
    write_headers(ws, 4, headers, widths)

    master = sources["master"] if isinstance(sources["master"], dict) else {}
    cross = master.get("cross_insights", {}) or {}

    row_idx = 5
    for secao_nome, secao_data in cross.items():
        if not isinstance(secao_data, dict):
            continue
        parsed = secao_data.get("parsed", {})
        if not isinstance(parsed, dict):
            parsed = {}

        # parsed tem tipicamente 1 key que é uma lista de items
        for key, value in parsed.items():
            if isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        for subk, subv in item.items():
                            write_row(ws, row_idx, [
                                secao_nome,
                                key,
                                subk,
                                str(subv)[:300],
                            ])
                            row_idx += 1
                    else:
                        write_row(ws, row_idx, [secao_nome, key, "", str(item)[:300]])
                        row_idx += 1
            elif isinstance(value, dict):
                for subk, subv in value.items():
                    write_row(ws, row_idx, [secao_nome, key, subk, str(subv)[:300]])
                    row_idx += 1
            else:
                write_row(ws, row_idx, [secao_nome, key, "", str(value)[:300]])
                row_idx += 1

    ws.freeze_panes = "A5"


# =============================================================================
# MD MAPA
# =============================================================================

def write_md_mapa(sources: dict) -> None:
    n_projs = len(sources["indices_exec"])
    pm = sources["calibration"].get("por_macrogrupo", {}) if isinstance(sources["calibration"], dict) else {}
    pm_count = len(pm)
    cond = sources["condicional"] if isinstance(sources["condicional"], dict) else {}
    n_cond = sum(len(v) for v in (cond.get("por_padrao_mg", {}) or {}).values()) if cond else 0
    n_der = len(sources["master"].get("indices_derivados_v2", {}) if isinstance(sources["master"], dict) else {})
    n_pus_v1 = len(sources["pus_v1"]) if isinstance(sources["pus_v1"], dict) else 0
    n_pus_v2 = len(sources["pus_v2"].get("pus_agregados", []) if isinstance(sources["pus_v2"], dict) else [])
    xlsx_kb = OUT_XLSX.stat().st_size // 1024 if OUT_XLSX.exists() else '?'

    md = f"""# Catálogo de Índices Cartesian — Guia de Leitura

_Gerado em {datetime.now().isoformat(timespec='seconds')} via `scripts/gerar_catalogo_indices.py`_

**Planilha:** [INDICES-CATALOGO.xlsx](INDICES-CATALOGO.xlsx) — {xlsx_kb} KB, 10 abas, ~6.500 linhas de dados

---

## Bem-vindo

Se você é engenheiro civil e está abrindo essa planilha pela primeira vez, este documento é seu mapa. Ele explica **o que é cada coisa, como ler os números, e como usar o catálogo no dia-a-dia** de um orçamentista. Leia do começo ao fim da primeira vez — leva 10 minutos e te economiza horas depois.

A planilha é longa (quase 6.500 linhas de dados), mas é **toda filtável e ordenável**. Você não precisa "ler" ela linha por linha. Você **pergunta** (ex: "quanto custa o PU médio de porcelanato?") e **filtra** até achar a resposta.

---

## Por que esse catálogo existe

A Cartesian já fez mais de **{n_projs} orçamentos executivos** ao longo dos últimos anos. Cada um deles tem centenas de itens com quantidades, unidades, preços unitários, memoriais, premissas. Isso é um tesouro de informação, mas até agora estava **espalhado** em arquivos separados: cada projeto um .xlsx, cada calibração um .json, cada análise qualitativa um .md. Impossível de consultar rápido.

Nos últimos meses consolidamos tudo isso numa **base paramétrica V2** com vários tipos de índice derivados dos 126 projetos reais. Este catálogo é o **mapa único** dessa base — tudo o que a gente sabe sobre custo de construção consolidado num xlsx filtrável.

**Pra que serve na prática:**
- Estimar custo de um projeto novo sem precisar fazer orçamento detalhado
- Validar um orçamento contra o histórico (é razoável? tá fora da curva?)
- Justificar decisões técnicas pro cliente com dados ("nossa base de 23 projetos similares mostra que...")
- Identificar qual insumo tá caro/barato na sua proposta
- Comparar padrões (médio vs alto vs luxo) por macrogrupo
- Entender qual obra da base mais se parece com a que você está orçando

---

## Os 3 tipos de índice que você vai encontrar

Antes de entrar nas abas, saiba que temos **3 famílias diferentes** de índice, e elas respondem perguntas diferentes:

### 1. Índices de preço unitário (R$/unidade)

**"Quanto custa 1 m³ de concreto? 1 kg de aço? 1 m² de porcelanato?"**

Esses são os PUs (preços unitários) tradicionais. Em moeda por unidade física. Eles estão nas abas **PUS_CROSS_V1** (1.740 clusters com lista de obras-fonte) e **PUS_CROSS_V2** (4.210 clusters, mais cobertura mas sem lista). São bons pra compor um orçamento bottom-up — multiplicar quantidade × PU.

Exemplo: "Concreto usinado FCK 30 MPa bombeável, base da Cartesian mostra mediana de R$ 590/m³ (n=25 projetos)."

### 2. Índices de custo por área (R$/m² AC)

**"Quanto um m² de alvenaria custa no orçamento total? E um m² de pintura? E esquadrias?"**

Esses são os índices por **macrogrupo** expressos em R$/m² de AC (área construída). Úteis pra top-down — você multiplica R$/m² × AC do projeto e tem o custo do macrogrupo direto, sem precisar compor item a item. Estão nas abas **CALIBRACAO_GLOBAL** (média dos 126 sem distinção de padrão) e **CALIBRACAO_CONDICIONAL** (segregada por padrão Gemma: econômico, médio, médio-alto, alto, luxo).

Exemplo: "Esquadrias padrão alto, mediana R$ 395/m² AC (n=23 projetos classificados como 'alto')."

### 3. Índices de consumo físico (quantidade/m² AC)

**"Quantos m³ de concreto por m² de AC? Quantos kg de aço por m³ de concreto?"**

Esses são os índices **estruturais** que traduzem geometria em material. Ajudam a dimensionar rapidamente o **quantitativo** antes de ter projeto executivo. Estão na aba **INDICES_ESTRUTURAIS**.

Exemplo: "Concreto mediano 0,34 m³ por m² de AC (n=52 projetos). Se seu AC é 13.000 m², espere ~4.400 m³ de concreto estrutural."

**Regra prática:** num paramétrico novo você usa **consumo físico × AC × PU** pra bottom-up, ou **R$/m² × AC** pra top-down. As duas abordagens devem chegar em números próximos — se divergirem muito, é sinal de que algum índice tá fora da curva e vale investigar.

---

## Entendendo as colunas estatísticas (pense como quem lê um curva de distribuição)

Quase toda aba tem essas colunas:

| Coluna | O que significa | Como interpretar |
|---|---|---|
| **`n`** | Quantos projetos contribuíram pro índice | `n ≥ 10` é robusto, `n=5-9` razoável, `n<5` frágil (use com cuidado) |
| **`min` / `max`** | O menor e o maior valor observado | Dão a ideia do extremo — mas podem ser outliers. Não use pra calibração direta. |
| **`p10` / `p90`** | Valores de corte: 10% dos projetos ficam abaixo do p10 e 10% acima do p90 | A **faixa P10-P90 contém 80% dos projetos** — é o intervalo "normal" |
| **`p25` / `p75`** | Idem, mas 50% dos projetos ficam entre eles (faixa inter-quartil) | Intervalo **mais estreito, mais típico** — 50% dos projetos se encaixam aqui |
| **`mediana` (`p50`)** | O valor do meio — 50% dos projetos abaixo, 50% acima | **É o valor que a calibração V2 usa como "típico".** Mais robusto que a média |
| **`media`** | Média aritmética (soma / n) | Pode ser puxada por outliers (1 projeto com PU absurdo distorce). **Prefira mediana.** |
| **`cv`** | Coeficiente de variação = desvio-padrão / média | Grau de dispersão: `cv<0.3` = **confiável** (projetos parecidos). `0.3-0.5` = **variável** (depende do padrão/cidade). `>0.5` = **volátil** (investigar por quê) |
| **`projetos_fonte`** | _(só em PUS_CROSS_V1)_ | Lista de slugs das obras que contribuíram — permite rastrear "quem bancou esse número" |

### Leitura prática de uma linha

Exemplo real: **concreto usinado FCK 30**, na aba INDICES_DERIVADOS_V2:

```
n=25  min=420  p10=490  p25=560  mediana=590  media=605  p75=650  p90=710  max=820  cv=0.14
```

Como ler isso?
- **25 projetos** bancam esse índice — dados robustos ✅
- A **mediana é R$ 590/m³** — é o valor típico, use esse como default
- **80% dos projetos** ficam entre **R$ 490 e R$ 710/m³** (p10-p90) — essa é sua faixa razoável
- **50% dos projetos** ficam entre **R$ 560 e R$ 650** (p25-p75) — faixa mais apertada, sem extremos
- **CV = 0,14** → baixíssima dispersão, todos os projetos têm preço parecido → **confiança alta** nesse índice
- Se seu orçamento tem concreto a R$ 900, você está acima do max (**R$ 820**) — investiga. Pode ser erro, localização remota, fornecedor único.
- Se seu orçamento tem concreto a R$ 400, abaixo do min (**R$ 420**) — também investiga.

### Regra rápida de "confio ou não?"

| n | cv | Confiança |
|---|---|---|
| ≥ 10 | < 0.3 | 🟢 **Alta** — use direto |
| 5-9 | < 0.4 | 🟡 **Média** — use mas valide contra o contexto |
| < 5 | qualquer | 🔴 **Baixa** — use só como ponto de partida, não pra calibração final |
| qualquer | > 0.5 | 🔴 **Baixa** — a dispersão indica que o "típico" não se aplica. Precisa investigar o por quê (padrão? cidade? método construtivo?) |

---

## Visão guiada aba por aba

### 📋 Aba 1 — **LEIA_ME** (dentro do xlsx)

Versão resumida deste guia dentro da própria planilha. Serve pra você se orientar sem precisar sair do Excel. Contém o mapa das 10 abas, schema das colunas, exemplos de filtros, gaps conhecidos.

### 🏗 Aba 2 — **PROJETOS** ({n_projs} linhas)

**É o seu ponto de partida.** Lista completa dos 126 projetos executivos da base, com metadados principais.

Colunas: `slug | padrao_gemma | confianca | ac_m2 | ur | total_rs | rsm2 | m2_por_ur | rs_por_ur | cidade | fonte`

**Quando usar:**
- **"Quais obras são similares à minha?"** → filtre por `padrao_gemma` igual ao seu, `ac_m2` dentro de ±25%, `ur` próximo. Os que sobrarem são sua referência.
- **"Qual o R$/m² típico de alto padrão?"** → filtre `padrao_gemma = alto` → ordene `rsm2` → olhe a faixa.
- **"Meu projeto tem custo/UR razoável?"** → compare o valor do seu projeto com a coluna `rs_por_ur` dos similares.
- **Cross-reference:** quando uma outra aba citar um slug (ex: `adore-cacupe`), volte aqui pra ver quem é esse projeto, onde fica, qual o porte.

**Importante:** a coluna `padrao_gemma` não é um rótulo dado pelo orçamentista original — é uma **classificação feita automaticamente pelo Gemma (LLM)** analisando os itens de acabamento (porcelanato, mármore, ACM, elevador panorâmico, etc.) de cada projeto. 94% dos projetos têm confiança alta/média.

### 🧱 Aba 3 — **CALIBRACAO_GLOBAL** (18 linhas)

**Os 18 macrogrupos da Cartesian em R$/m² AC, agregados de todos os 126 projetos sem distinção de padrão.**

Colunas: `macrogrupo | n | min | p10 | p25 | mediana | media | p75 | p90 | max | unidade | fonte`

**Quando usar:**
- **Primeiro rascunho de um paramétrico:** pegue a mediana de cada macrogrupo × seu AC e some — você tem uma estimativa grosseira do custo total em 2 minutos.
- **Sanity check:** abriu um orçamento executivo novo e quer ver se o R$/m² de Esquadrias (por exemplo) está dentro do esperado? Essa aba te dá a faixa p10-p90 onde 80% dos projetos caem.

**Cuidado:** essa aba mistura projetos de padrão econômico com padrão alto. A dispersão é grande (cv geralmente 0.5+). **Use a aba CALIBRACAO_CONDICIONAL quando souber o padrão do seu projeto** — é mais preciso.

### 🎯 Aba 4 — **CALIBRACAO_CONDICIONAL** ({n_cond} linhas)

**A estrela da base V2.** Mesmos 18 macrogrupos, mas agora **segregados por padrão Gemma** (econômico / médio / médio-alto / alto / luxo). É a fonte primária da calibração que a Cartesian usa pós-fase 18b.

Colunas: `padrao | macrogrupo | n | min | p10 | p25 | mediana | media | p75 | p90 | max | unidade | fonte`

**Por que é melhor que a global:** um projeto alto padrão tem Esquadrias muito mais caro que um econômico. Se você usa a mediana "global", subestima alto e superestima econômico. Aqui a mediana é **específica do padrão**, então é sempre mais aderente.

**Quando usar:**
- **Paramétrico de projeto novo:** você sabe o padrão pretendido → pega a mediana do seu padrão em cada MG × AC → soma.
- **Comparar padrões:** quanto Esquadrias custa a mais em "alto" vs "médio"? Filtre `macrogrupo=Esquadrias` → ordene por `padrao` → calcule a diferença. Isso te dá dados pra negociar com cliente.
- **Entender o "salto de padrão":** de econômico pra médio as diferenças estão em Pisos, Rev.Parede, Louças. De alto pra luxo estão em Fachada, Sistemas Especiais, Complementares. Os dados estão aí, filtre e veja.

**Pegadinha:** alguns MGs têm `n<3` em certos padrões — nesses casos o sistema de calibração cai num **fallback global** (usa o valor da CALIBRACAO_GLOBAL multiplicado por um fator de padrão). Ou seja, um MG com n=2 é **referência fraca** — o próprio sistema sabe disso e não confia direto.

**Observação importante:** a classe **luxo** tem 0 projetos na base. Isso não é bug — significa que a Cartesian nunca orçou um luxo-luxo (casa Alphaville, cobertura linear Vieira Souto) no histórico. O "alto" da base é o topo real dos empreendimentos multifamiliares residenciais que a Cartesian atende.

### 📊 Aba 5 — **INDICES_DERIVADOS_V2** ({n_der} linhas)

**29 índices calculados combinando os dados brutos.** Inclui PUs consolidados de insumos principais (concreto, aço, porcelanato, pintura), custos de macrogrupo em R$/m², ratios técnicos (aço/concreto, fôrma/concreto) e proporções de curva ABC.

Colunas: `nome | descricao | n | min | p10 | p25 | mediana | media | p75 | p90 | max | cv | unidade | fonte`

**Quando usar:**
- **Composição rápida de custo estrutural:** `pu_concreto_usinado_mediano`, `pu_aco_ca50_mediano`, `pu_forma_madeira_mediano` — mediana dos PUs consolidados de todos os projetos.
- **Ratios pra dimensionar:** `concreto_por_aco_ratio` te dá a razão concreto:aço típica. `forma_por_concreto_ratio` idem pra fôrma.
- **Custos por macrogrupo já computados:** `custo_concreto_rsm2`, `custo_esquadrias_rsm2`, `custo_loucas_rsm2` — R$/m² já pronto pra macrogrupos específicos.
- **Indicadores de curva ABC:** `curva_abc_a_pct` mostra o % de itens que compõem 80% do custo nos projetos da base. Baixo = projeto concentrado em poucos itens; alto = projeto pulverizado.

**Dica:** ordene a aba por `n` desc e ignore tudo com `n<5`. O que sobrar é o que tem base estatística sólida.

### 🔩 Aba 6 — **INDICES_ESTRUTURAIS** (~22 linhas)

**Consumo físico + produto + instalações % + custos indiretos % + segmentos por porte.** É a aba dos engenheiros de verdade — quantidades, não preços.

Seções:
- **Estruturais:** `concreto_m3_por_m2_ac` (0,34 mediano), `aco_kg_por_m3_concreto` (~106), `aco_kg_por_m2_ac`, `forma_m2_por_m2_ac`
- **Produto:** `ac_por_ur` (quantos m² por unidade residencial), `custo_por_ur`, `cub_ratio`, `burn_rate_mensal`, `elevador_pu_un`
- **Instalações %:** hidrossanitárias, elétricas, preventivas, gás, telecom — cada uma como % do custo total do projeto
- **Custos Indiretos %:** projetos/consultorias, taxas/licenças, equipe ADM, EPCs, equipamentos, ensaios, canteiro
- **Segmento por porte:** R$/m² total separado em 4 faixas de AC (pequeno <8k, médio 8-15k, grande 15-25k, extra >25k)

**Quando usar:**
- **Dimensionar quantitativo antes do projeto executivo:** multiplica o consumo mediano × seu AC e você tem uma estimativa de material. `concreto_m3_por_m2_ac × 13.000 m²` = ~4.400 m³ de concreto.
- **Validar % de instalações:** seu projeto tem elétrica representando 10% do total, mas a base mostra mediana 6%? Investigue — pode ter projeto muito complexo (garage full automatizada, automação geral) ou pode ser erro de BDI.
- **Validar % de CI:** se Gerenciamento tá acima da faixa P75, provavelmente tem equipe superdimensionada ou prazo esticado.
- **Segmentação por porte:** escolha o segmento do seu projeto (AC <8k → pequeno, etc) e veja a faixa de R$/m² que projetos daquele porte apresentaram.

### 💰 Aba 7 — **PUS_CROSS_V1** ({n_pus_v1} linhas) ⭐ TESOURO DA BASE

**A aba mais preciosa pra quem quer rastrear "quem bancou esse número".** 1.740 clusters de PU cross-projeto, **com lista de obras-fonte em cada linha**.

Colunas: `categoria | chave | descricao | unidade | n_proj | n_obs | min | p25 | mediana | p75 | max | cv | projetos_fonte`

**O que é um "cluster":** itens semanticamente iguais de projetos diferentes agrupados. Ex: "Concreto usinado FCK 30 bombeado" pode aparecer em 30 projetos com pequenas variações de descrição — um cluster consolida todos esses num único registro com mediana robusta.

**Quando usar:**
- **"Quais obras sustentam esse PU mediano?"** → filtre a descrição, olhe a coluna `projetos_fonte`. Se são 3 obras só da mesma construtora, pode ser viés. Se são 15 obras de 5 construtoras diferentes, é uma mediana forte.
- **Auditar um PU específico do seu orçamento:** seu concreto tá a R$ 750. Nessa aba, concreto FCK 30 tem mediana R$ 590, p90 R$ 710. Você está acima do p90 — justifique (obra remota? usina única?).
- **Validar insumos caros:** ordene por `mediana` desc, veja os clusters mais caros da base. Se seu projeto tem um desses, precisa ter atenção especial pra não estourar.

**Dicas de filtro:**
- `n_proj ≥ 5` + `cv < 0.3` = PUs robustos
- Busca por palavra-chave na `chave` ou `descricao` (ex: "porcelanato", "bloco", "vidro")
- Ordene por `categoria` pra navegar macrogrupo por macrogrupo

### 📚 Aba 8 — **PUS_CROSS_V2** ({n_pus_v2} linhas) — mais cobertura, sem fontes

Versão nova (fase 10 v2) do cluster de PUs, com **2.4x mais clusters** (4.210 vs 1.740). A diferença: usa **hash-based clustering semântico** que consegue agrupar itens com descrições mais diversas.

Colunas: `cluster_id | key (tokens) | descricao | unidades | n_proj | n_obs | pu_min | pu_p10 | pu_p25 | pu_mediana | pu_p75 | pu_p90 | pu_max | pu_media | cv`

**Limitação:** o algoritmo V2 não preservou a **lista de projetos** fonte em cada cluster. Então você tem a quantidade (n_proj) mas não sabe **quais** obras bancam o número.

**Quando usar:**
- **Cobertura ampla:** quando a V1 não tem o item que você procura, provavelmente a V2 tem (quase 3x mais itens).
- **Comparação rápida de PUs:** ordene por `n_proj` desc pra ver os itens mais presentes na base.
- **`key`** mostra os tokens semânticos usados pro clustering (ex: `porcelanato|60x60|retificado`) — útil pra filtrar por palavra-chave.

**Regra prática:** use V1 quando precisar justificar/rastrear, V2 quando só quer um número de referência.

### 🅰 Aba 9 — **CURVA_ABC_MASTER** (126 linhas)

**1 linha por projeto da base.** Mostra quantos itens tem, quantos desses itens compõem a curva A (80% do custo), e o valor total.

Colunas: `slug | status | n_itens | n_a | pct_a | valor_total_rs`

**Pra que serve:**
- **Identificar projetos "concentrados" vs "pulverizados":** ordene por `pct_a` asc. Projetos com pct_a ≈ 5% são altamente concentrados (poucos itens dominam); pct_a ≈ 25% são pulverizados (muitos itens pequenos).
- **Achar projetos de referência por porte total:** filtre `valor_total_rs` próximo ao do seu projeto novo.
- **Validar sanidade do seu orçamento:** se seu projeto tem 2.000 itens e a curva A pega 400 (20%), compare com a base — projetos Cartesian costumam ter curva A na faixa 5-15% do total.

### 🧠 Aba 10 — **CROSS_INSIGHTS_GEMMA** (varia)

**Análises qualitativas cross-projeto feitas pelo Gemma (LLM local).** Leu todos os 126 projetos e destilou observações gerais da base.

Colunas: `secao | tipo | campo | conteudo`

Seções:
- **`familias`** — agrupamento de projetos por tipologia/característica
- **`indices_sugeridos`** — índices que o Gemma sugeriu criar a partir dos padrões que identificou
- **`lacunas`** — coisas que faltam na base (tipologias não cobertas, regiões, padrões)
- **`outliers`** — projetos que destoam do resto (pra cima ou pra baixo)
- **`padroes_comuns`** — padrões recorrentes que o Gemma identificou

**Quando usar:** menos pra cálculo, mais pra **entender a base qualitativamente**. Útil antes de tomar uma decisão estratégica ("tem projeto similar na base ou estamos no escuro?").

---

## Cenários reais de uso no dia-a-dia

### Cenário 1: "Preciso estimar um novo projeto de 13.000 m² alto padrão"

1. Abrir **PROJETOS** → filtrar `padrao_gemma = alto` + `ac_m2 entre 10.000 e 16.000` → anotar os ~5 slugs mais próximos
2. Abrir **CALIBRACAO_CONDICIONAL** → filtrar `padrao = alto` → pegar mediana de cada MG → multiplicar por 13.000 → somar → tem o **custo top-down**
3. Abrir **INDICES_ESTRUTURAIS** → seção "Segmento por porte" → linha `medio_8k_15k_rsm2` → ver se sua soma tá dentro da faixa p25-p75 daquele porte
4. Se tudo bate, seu paramétrico está em linha com a base. Se não bate, descobrir por qual MG especificamente divergiu.

### Cenário 2: "O orçamentista me mandou um PU de concreto R$ 780. Isso tá ok?"

1. Abrir **PUS_CROSS_V2** → filtrar `key` contém "concreto" + "usinado" → ver faixa `pu_p10` a `pu_p90`
2. Se o concreto mediano da base é R$ 590 e p90 é R$ 710, seu orçamentista está **acima do p90** → tem que justificar (localização remota? concreto especial? preço hoje diferente da data-base da média?)
3. Volte pra **PUS_CROSS_V1** → procure o mesmo cluster → coluna `projetos_fonte` → olhe quais obras bancam essa mediana → avalie se são obras em contexto similar ao seu

### Cenário 3: "Quanto vou cobrar a mais pra passar de médio-alto pra alto no mesmo projeto?"

1. Abrir **CALIBRACAO_CONDICIONAL** → filtrar `macrogrupo in (Pisos, Rev. Interno Parede, Esquadrias, Louças e Metais, Fachada)` (os 5 mais sensíveis ao padrão) → colunas `mediana`
2. Para cada MG: subtrair `mediana medio-alto` de `mediana alto`
3. Somar as diferenças × AC do projeto → tem a **ordem de grandeza do "salto de padrão"**
4. Use pra negociar com o cliente: "esse upgrade custa +R$ X/m², distribuído nesses 5 macrogrupos"

### Cenário 4: "Meu projeto tem Sistemas Especiais puxando o custo. Isso é comum?"

1. Abrir **CALIBRACAO_CONDICIONAL** → filtrar `padrao = alto` + `macrogrupo = Sistemas Especiais` → olhar mediana e p90
2. Abrir **PROJETOS** → identificar 2-3 projetos da base que são de alto padrão próximo ao seu
3. Abrir **CROSS_INSIGHTS_GEMMA** → procurar na seção `padroes_comuns` se tem observação sobre Sistemas Especiais em projetos alto padrão (ex: "piscina aquecida é comum em 60% dos alto padrão")
4. Cross-checar: se seu projeto tem piscina aquecida + sauna + spa + gerador dedicado, está dentro do padrão alto. Se tem automação total + heliponto + elevador panorâmico, está extrapolando pra luxo.

### Cenário 5: "Tô orçando e preciso validar o consumo físico de concreto"

1. Abrir **INDICES_ESTRUTURAIS** → seção "Estruturais" → linha `concreto_m3_por_m2_ac`
2. Mediana 0,34 m³/m² → multiplicar por seu AC → quantitativo esperado
3. Se seu projeto tem laje protendida, o número deve ser **menor** que a mediana (protendida usa ~0,22 m³/m²)
4. Se tem laje convencional pesada com pé-direito alto, **maior** (até 0,45)
5. Se seu quantitativo está fora da faixa p10-p90 (0,21-0,57), revisar modelagem estrutural

### Cenário 6: "Preciso saber se BDI da minha proposta tá agressivo ou não"

1. Abrir **INDICES_ESTRUTURAIS** → seção "Custos Indiretos %" → linhas `projetos_consultorias_pct_total`, `taxas_licencas_pct_total`, `equipe_adm_pct_total`, `canteiro_pct_total`
2. Somar as medianas → tem o **CI típico como % do total** nos projetos da base
3. Comparar com o CI da sua proposta → se estiver 20% abaixo da mediana, você está **agressivo** (pode dar prejuízo); 20% acima, **folgado** (pode perder o contrato)

---

## Gaps e pegadinhas

### 1. PUS_CROSS_V2 não tem lista de projetos

A versão 2 do clustering (4.210 clusters) ganhou cobertura mas perdeu rastreabilidade. Quando você quiser **auditar** um PU específico ou citar "fontes" pro cliente, use **PUS_CROSS_V1** (1.740 clusters com lista completa de obras). Quando quiser **cobertura ampla** (procurar um item raro), use V2.

### 2. Classe "luxo" tem 0 projetos

A base Cartesian não tem casos de luxo-luxo (cobertura linear, casa Alphaville premium). O "alto" é o topo real. Se você está orçando um projeto que claramente é luxo absoluto, use o "alto" como referência inferior e aplique um multiplicador — não confie no "luxo" da aba CALIBRACAO_CONDICIONAL porque ela tá vazia.

### 3. `custo_por_ur` tem n=2

Esse índice específico (R$ por unidade residencial) só tem 2 projetos com dado válido — confiança estatística **péssima**. Se precisar do custo/UR, melhor calcular direto da aba PROJETOS filtrando por similares e tirando a média dos 5 mais próximos.

### 4. Alguns MGs em CALIBRACAO_CONDICIONAL têm n baixo

Classes econômico e luxo têm poucos projetos — alguns macrogrupos específicos dentro delas podem ter `n<3`. Nesses casos o sistema de calibração cai num **fallback global × multiplicador de padrão**. Na prática: se você está orçando econômico e viu uma mediana estranha, confira a coluna `n` — se for baixo, desconfie.

### 5. CV alto num índice = não confiável

`cv > 0.5` significa que a dispersão é tão grande que o "valor típico" não tem muito significado. Isso acontece em macrogrupos muito dependentes do padrão (Louças, Fachada, Sistemas Especiais). A solução é usar a **CALIBRACAO_CONDICIONAL** em vez da global — ao segregar por padrão, o CV dentro de cada bucket cai drasticamente.

### 6. Os 126 projetos da base são TODOS da Cartesian

Essa base representa o universo de projetos que a Cartesian já orçou. Ela é **fortemente viesada para multifamiliar residencial alto-padrão em SC** (Itajaí, Balneário Camboriú, Florianópolis, Navegantes). Se você estiver orçando um galpão industrial, um hospital, ou um residencial em Pernambuco, os índices aqui são **ponto de partida**, não verdade absoluta.

---

## Glossário rápido

| Termo | O que é |
|---|---|
| **AC** | Área Construída (m²) — denominador padrão de todos os índices R$/m² |
| **UR** | Unidades Residenciais — apartamentos/casas no empreendimento |
| **Macrogrupo (MG)** | Uma das 18 categorias canônicas Cartesian (Gerenciamento, Mov.Terra, Infra, Supra, Alvenaria, etc) |
| **R$/m² AC** | Custo em reais dividido pela área construída total do projeto |
| **Paramétrico** | Orçamento rápido baseado em índices (top-down), sem BoQ detalhado |
| **Preliminar** | Versão mais detalhada do paramétrico, com itens de referência (pós-fase 18b da Cartesian) |
| **Executivo** | Orçamento completo item a item com quantitativos do projeto executivo — **NÃO é o que a base aqui tem** |
| **PU** | Preço Unitário (R$ por unidade física: m², m³, kg, un) |
| **CUB** | Custo Unitário Básico — índice mensal do Sinduscon usado pra normalizar data-base |
| **CUB Ratio** | Relação entre o R$/m² do projeto e o CUB da data-base (>1 = projeto mais caro que CUB) |
| **Bottom-up** | Compor custo item a item (Qtd × PU = Total) — o que a aba de detalhe do paramétrico faz |
| **Top-down** | Estimar custo agregado a partir de índices R$/m² × AC — o que CALIBRACAO_CONDICIONAL permite |
| **Mediana (p50)** | Valor central — 50% dos projetos abaixo, 50% acima. Mais robusta que média contra outliers |
| **Percentis p10/p25/p75/p90** | Cortes da distribuição. P10-P90 contém 80% dos projetos; P25-P75 contém 50% |
| **CV** | Coeficiente de Variação = desvio-padrão / média. Mede dispersão normalizada |
| **Curva ABC** | Ordenação de itens pelo valor acumulado. Itens "A" somam 80% do custo total |
| **Padrão Gemma** | Classificação (econômico/médio/médio-alto/alto/luxo) feita automaticamente pelo LLM Gemma analisando itens de acabamento |
| **Fase XX** | Marcos do desenvolvimento da base V2 (ver `base/FASES-FUTURAS.md` pra timeline completa) |

---

## Como regenerar este catálogo

Quando você atualizar a base (novo projeto processado, nova calibração, novos dados Gemma), rode:

```bash
cd ~/orcamentos-openclaw
python scripts/gerar_catalogo_indices.py
```

O script lê dinamicamente todos os JSONs da base e regenera **o xlsx + este md**. Rode sem argumentos.

## JSONs fonte (pra quem quer abrir o dado bruto)

- `base/calibration-indices.json` — 18 MGs global + estruturais + instalações + CI + produto + segmentos
- `base/calibration-condicional-padrao.json` — 18 MGs × 5 padrões Gemma (fase 18b, **fonte primária da calibração atual**)
- `base/base-indices-master-2026-04-13.json` — consolidado 322 KB (29 derivados + curva ABC + cross insights)
- `base/itens-pus-agregados.json` — 4.210 clusters PU V2 (fase 10 v2)
- `base/base-pus-cartesian.json` — 1.740 clusters PU V1 **com lista de projetos**
- `base/padroes-classificados-consolidado.json` — labels Gemma fase 18 (125/126 projetos classificados)
- `base/indices-executivo/*.json` — 126 arquivos individuais por projeto

## Onde encontrar mais contexto

- **[FASES-FUTURAS.md](FASES-FUTURAS.md)** — linha do tempo completa das fases de desenvolvimento da base V2 (1-19)
- **[SESSAO-2026-04-14-REVISAO-3-PACOTES.md](SESSAO-2026-04-14-REVISAO-3-PACOTES.md)** — narrativa da sessão que fechou o modelo V2 + 18b + 19
- **[PARAMETRICO-V2-HIBRIDO.md](PARAMETRICO-V2-HIBRIDO.md)** — como o gerador paramétrico usa esses índices na prática
- **[CAMADA-QUALITATIVA-GEMMA.md](CAMADA-QUALITATIVA-GEMMA.md)** — detalhes de como a camada qualitativa Gemma foi construída
"""
    OUT_MD.write_text(md, encoding="utf-8")
    print(f"MD mapa salvo: {OUT_MD}")


def main():
    print("=" * 70)
    print("GERANDO CATÁLOGO DE ÍNDICES CARTESIAN")
    print("=" * 70)

    sources = load_sources()

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    print("\nConstruindo abas...")
    build_aba_leia_me(wb, sources)
    print("  [1/10] LEIA_ME")
    build_aba_projetos(wb, sources)
    print("  [2/10] PROJETOS")
    build_aba_calibracao_global(wb, sources)
    print("  [3/10] CALIBRACAO_GLOBAL")
    build_aba_calibracao_condicional(wb, sources)
    print("  [4/10] CALIBRACAO_CONDICIONAL")
    build_aba_derivados(wb, sources)
    print("  [5/10] INDICES_DERIVADOS_V2")
    build_aba_estruturais(wb, sources)
    print("  [6/10] INDICES_ESTRUTURAIS")
    build_aba_pus_v1(wb, sources)
    print("  [7/10] PUS_CROSS_V1 (1.740)")
    build_aba_pus_v2(wb, sources)
    print("  [8/10] PUS_CROSS_V2 (4.210)")
    build_aba_curva_abc(wb, sources)
    print("  [9/10] CURVA_ABC_MASTER")
    build_aba_cross_insights(wb, sources)
    print("  [10/10] CROSS_INSIGHTS_GEMMA")

    print(f"\nSalvando {OUT_XLSX}...")
    wb.save(str(OUT_XLSX))
    size_kb = OUT_XLSX.stat().st_size // 1024
    print(f"  salvo ({size_kb} KB)")

    write_md_mapa(sources)

    print("\nCatalogo gerado com sucesso")


if __name__ == "__main__":
    main()
