#!/usr/bin/env python3
"""
Exportar EAP do Memorial para Excel com múltiplas abas
Thozen Electra Towers
"""
import os, sys, json, requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv('.env.sensitive')

SUPABASE_URL = os.getenv('MEMORIAL_SUPABASE_URL')
ANON_KEY = os.getenv('MEMORIAL_SUPABASE_ANON')
EMAIL = os.getenv('MEMORIAL_EMAIL')
PASSWORD = os.getenv('MEMORIAL_PASSWORD')

PROJECT_ID = "cdff0592-fb3c-4c13-9516-efde6b56b336"
BUDGET_ID = "3dc996e6-7bdd-4523-80c4-bb391cf7060d"

print("🔐 Autenticando...", flush=True)
auth = requests.post(
    f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
    headers={"apikey": ANON_KEY, "Content-Type": "application/json"},
    json={"email": EMAIL, "password": PASSWORD},
    timeout=10
).json()

token = auth['access_token']
headers = {'apikey': ANON_KEY, 'Authorization': f'Bearer {token}'}
print("✅ Autenticado\n", flush=True)

# Buscar dados do projeto
print("📊 Buscando dados do projeto...", flush=True)
project_response = requests.get(
    f"{SUPABASE_URL}/rest/v1/projects?select=*&id=eq.{PROJECT_ID}",
    headers=headers,
    timeout=10
)
project = project_response.json()[0] if project_response.json() else {}

# Buscar EAP completa
print("📋 Buscando EAP do Memorial...", flush=True)
items_response = requests.get(
    f"{SUPABASE_URL}/rest/v1/budget_items?select=*&budget_id=eq.{BUDGET_ID}&order=code,sequence",
    headers=headers,
    timeout=30
)
items = items_response.json()
print(f"✅ {len(items)} itens encontrados\n", flush=True)

# Criar workbook
wb = Workbook()
wb.remove(wb.active)  # Remove aba padrão

# ============================================================
# ABA 1: DADOS DE PROJETOS
# ============================================================
print("📄 Criando aba 'Dados de Projetos'...", flush=True)
ws_dados = wb.create_sheet("Dados de Projetos")

# Header style
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

ws_dados['A1'] = "DADOS DO PROJETO"
ws_dados['A1'].font = Font(bold=True, size=14)
ws_dados['A1'].fill = header_fill

dados_info = [
    ("Nome do Projeto", project.get('name', 'Edifício Electra Towers')),
    ("Cliente", project.get('client_name', 'Thozen')),
    ("Localização", project.get('location', 'São Paulo/SP')),
    ("Área Total (m²)", project.get('total_area', '')),
    ("Torres", "1 (Torre A)"),
    ("Pavimentos", project.get('floor_count', '')),
    ("Status", "Em Orçamentação"),
    ("Data", "20/03/2026")
]

row = 3
for label, value in dados_info:
    ws_dados[f'A{row}'] = label
    ws_dados[f'B{row}'] = value
    ws_dados[f'A{row}'].font = Font(bold=True)
    row += 1

ws_dados.column_dimensions['A'].width = 25
ws_dados.column_dimensions['B'].width = 40

# ============================================================
# ABA 2: GERENCIAMENTO EXECUTIVO (UC 01)
# ============================================================
print("📄 Criando aba 'Ger_Exec'...", flush=True)
ws_ger = wb.create_sheet("Ger_Exec")

# Header
headers_ger = ["Código", "Descrição", "UN", "QTD", "Preço Unit. (R$)", "Total (R$)", "Level"]
for col, header in enumerate(headers_ger, 1):
    cell = ws_ger.cell(1, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Filtrar UC 01
uc01_items = [i for i in items if i['code'].startswith('01')]

row = 2
for item in uc01_items:
    ws_ger[f'A{row}'] = item['code']
    ws_ger[f'B{row}'] = item['name'] or item['description']
    ws_ger[f'C{row}'] = item.get('unit', '')
    ws_ger[f'D{row}'] = item.get('quantity', '')
    ws_ger[f'E{row}'] = item.get('unit_price', '')
    ws_ger[f'F{row}'] = item.get('total_price', 0)
    ws_ger[f'G{row}'] = f"N{item['level']}"
    
    # Indent por level
    indent = (item['level'] - 1) * 2
    ws_ger[f'B{row}'].alignment = Alignment(indent=indent)
    
    # Bold nos níveis superiores
    if item['level'] < 4:
        ws_ger[f'A{row}'].font = Font(bold=True)
        ws_ger[f'B{row}'].font = Font(bold=True)
    
    row += 1

# Ajustar colunas
ws_ger.column_dimensions['A'].width = 15
ws_ger.column_dimensions['B'].width = 60
ws_ger.column_dimensions['C'].width = 8
ws_ger.column_dimensions['D'].width = 10
ws_ger.column_dimensions['E'].width = 15
ws_ger.column_dimensions['F'].width = 15
ws_ger.column_dimensions['G'].width = 8

# ============================================================
# ABA 3: TORRE A (EAP COMPLETA)
# ============================================================
print("📄 Criando aba 'Torre A'...", flush=True)
ws_torre = wb.create_sheet("Torre A")

# Header
headers_torre = ["Código", "Descrição", "UN", "QTD", "Preço Unit. (R$)", "Total (R$)", "Level", "CPU"]
for col, header in enumerate(headers_torre, 1):
    cell = ws_torre.cell(1, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Todos os itens
row = 2
for item in items:
    ws_torre[f'A{row}'] = item['code']
    ws_torre[f'B{row}'] = item['name'] or item['description']
    ws_torre[f'C{row}'] = item.get('unit', '')
    ws_torre[f'D{row}'] = item.get('quantity', '')
    ws_torre[f'E{row}'] = item.get('unit_price', '')
    ws_torre[f'F{row}'] = item.get('total_price', 0)
    ws_torre[f'G{row}'] = f"N{item['level']}"
    ws_torre[f'H{row}'] = "Sim" if item.get('composicao_id') else ""
    
    # Indent por level
    indent = (item['level'] - 1) * 2
    ws_torre[f'B{row}'].alignment = Alignment(indent=indent)
    
    # Bold nos níveis superiores
    if item['level'] < 4:
        ws_torre[f'A{row}'].font = Font(bold=True)
        ws_torre[f'B{row}'].font = Font(bold=True)
    
    # Highlight subetapas com CPU
    if item.get('composicao_id'):
        ws_torre[f'H{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    row += 1

# Ajustar colunas
ws_torre.column_dimensions['A'].width = 15
ws_torre.column_dimensions['B'].width = 60
ws_torre.column_dimensions['C'].width = 8
ws_torre.column_dimensions['D'].width = 10
ws_torre.column_dimensions['E'].width = 15
ws_torre.column_dimensions['F'].width = 15
ws_torre.column_dimensions['G'].width = 8
ws_torre.column_dimensions['H'].width = 8

# Adicionar totais
total_row = row + 1
ws_torre[f'E{total_row}'] = "TOTAL:"
ws_torre[f'E{total_row}'].font = Font(bold=True)
ws_torre[f'F{total_row}'] = f"=SUM(F2:F{row-1})"
ws_torre[f'F{total_row}'].font = Font(bold=True)

# Salvar
output_path = "output/Thozen_Electra_Towers_Orcamento.xlsx"
wb.save(output_path)

print(f"\n{'='*60}")
print(f"✅ Planilha criada com sucesso!")
print(f"{'='*60}")
print(f"📁 {output_path}")
print(f"\nAbas criadas:")
print(f"  1. Dados de Projetos")
print(f"  2. Ger_Exec ({len(uc01_items)} itens)")
print(f"  3. Torre A ({len(items)} itens)")
print(f"{'='*60}")
