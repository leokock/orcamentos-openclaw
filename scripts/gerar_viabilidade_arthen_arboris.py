#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Viabilidade Financeira — Arthen Arboris (Itapema / Perequê)
Consome: arthen-arboris-parametrico-v00-final.xlsx (Drive) + premissas-viabilidade.json
Produz : viabilidade-arthen-arboris-v3.xlsx (v1 analítico) + -v3.xlsx (v2 didático)
Inclui : VPL, TIR, Monte Carlo 10k iter ancorado no orçamento entregue,
         tornado, cenários, fluxo mensal didático (36 meses obra + 3 pré + 12 pós)

Atualizado 2026-04-20:
  - Base de custo: orçamento paramétrico v00-final (R$ 39.379.139,40)
  - Prazo obra: 36 meses (antes 30)
  - Curva de obra mensal: mesma da apresentação ppt (slide 18 CURVA DE DISTRIBUIÇÃO)
  - Marketing: 3% VGV distribuído linear na obra (sem mais front-load)
  - Fluxo de caixa: layout didático (subtotais por fase, labels amigáveis, eventos)
  - Monte Carlo: âncora explícita no orçamento entregue
"""

import json
import os
from pathlib import Path

import numpy as np
import numpy_financial as nf
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule

# ============================================================================
# PATHS
# ============================================================================
PACOTE_DIR = Path(r"C:\Users\leona\orcamentos-openclaw\base\pacotes\arthen-arboris")
DRIVE_DIR = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
    r"\_Parametrico_IA\arthen-arboris"
)
# Fonte do custo: orçamento paramétrico v00-final (no Drive)
PARAMETRICO_XLSX = DRIVE_DIR / "arthen-arboris-parametrico-v00-final.xlsx"
PREMISSAS_JSON = PACOTE_DIR / "premissas-viabilidade.json"
OUTPUT_XLSX = PACOTE_DIR / "viabilidade-arthen-arboris-v3.xlsx"
OUTPUT_XLSX_V2 = PACOTE_DIR / "viabilidade-arthen-arboris-v3-didatico.xlsx"
PLOTS_DIR = PACOTE_DIR / "_viabilidade_plots"

# ============================================================================
# ESTILO (padrão Cartesian)
# ============================================================================
DARK = "2C3E50"
ACCENT = "2980B9"
INPUT_BG = "FFF3E0"
CALC_BG = "E8F5E9"
YELLOW_BG = "FFF8E1"
RED_BG = "FFEBEE"
GRAY_BG = "F5F5F5"
BLUE_BG = "E3F2FD"

FILL_DARK = PatternFill("solid", fgColor=DARK)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT)
FILL_INPUT = PatternFill("solid", fgColor=INPUT_BG)
FILL_CALC = PatternFill("solid", fgColor=CALC_BG)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW_BG)
FILL_RED = PatternFill("solid", fgColor=RED_BG)
FILL_GRAY = PatternFill("solid", fgColor=GRAY_BG)
FILL_BLUE = PatternFill("solid", fgColor=BLUE_BG)

FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_H1 = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_H2 = Font(name="Arial", size=10, bold=True, color=DARK)
FONT_BODY = Font(name="Arial", size=9)
FONT_BOLD = Font(name="Arial", size=9, bold=True)
FONT_INPUT = Font(name="Arial", size=9, bold=True, color="C25700")
FONT_SMALL = Font(name="Arial", size=8, color="666666")

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

FMT_BRL = '"R$" #,##0'
FMT_BRL2 = '"R$" #,##0.00'
FMT_PCT = "0.0%"
FMT_PCT2 = "0.00%"
FMT_NUM = "#,##0"
FMT_NUM2 = "#,##0.00"


def style_header(cell):
    cell.fill = FILL_DARK
    cell.font = FONT_H1
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER


def style_input(cell, fmt=FMT_BRL):
    cell.fill = FILL_INPUT
    cell.font = FONT_INPUT
    cell.alignment = ALIGN_RIGHT
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def style_calc(cell, fmt=FMT_BRL):
    cell.fill = FILL_CALC
    cell.font = FONT_BOLD
    cell.alignment = ALIGN_RIGHT
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def style_label(cell):
    cell.font = FONT_BOLD
    cell.alignment = ALIGN_LEFT
    cell.border = BORDER


# ============================================================================
# LOADERS
# ============================================================================
def load_parametrico():
    """Lê custo total e breakdown por macrogrupo do parametrico xlsx.

    Robustez: se o xlsx não tem valores cacheados (arquivo nunca aberto no
    Excel/LibreOffice), usa a biblioteca `formulas` pra computar as fórmulas
    em tempo real. Isso permite trabalhar direto do Drive sem precisar abrir
    manualmente.
    """
    wb = load_workbook(PARAMETRICO_XLSX, data_only=True)
    ws = wb["CUSTOS_MACROGRUPO"]
    custo_obra_total_raw = ws["D22"].value
    has_cache = isinstance(custo_obra_total_raw, (int, float))
    wb.close()

    if has_cache:
        wb = load_workbook(PARAMETRICO_XLSX, data_only=True)
        ws = wb["CUSTOS_MACROGRUPO"]
        custo_obra_total = float(ws["D22"].value)
        breakdown = {}
        for row in range(4, 22):
            nome = ws.cell(row=row, column=1).value
            valor = ws.cell(row=row, column=4).value
            pct = ws.cell(row=row, column=3).value
            if nome and valor:
                breakdown[nome] = {"valor": float(valor), "pct": float(pct or 0.0)}
        wb.close()
        return custo_obra_total, breakdown

    # Fallback: xlsx sem cache de valores → avaliar fórmulas via `formulas` lib
    import formulas as _formulas
    print(f"      [parametrico sem cache — avaliando fórmulas via formulas lib]")
    xl = _formulas.ExcelModel().loads(str(PARAMETRICO_XLSX)).finish()
    sol = xl.calculate()
    fname = PARAMETRICO_XLSX.name

    def _ev(coord):
        key = f"'[{fname}]CUSTOS_MACROGRUPO'!{coord}"
        try:
            v = sol[key].value[0][0]
            return float(v) if isinstance(v, (int, float)) else 0.0
        except Exception:
            return 0.0

    wb = load_workbook(PARAMETRICO_XLSX)
    ws = wb["CUSTOS_MACROGRUPO"]
    total = 0.0
    breakdown = {}
    for row in range(4, 22):
        nome = ws.cell(row=row, column=1).value
        if not nome:
            continue
        valor = _ev(f"D{row}")
        if valor > 0:
            breakdown[nome] = {"valor": valor, "pct": 0.0}
            total += valor
    for name in breakdown:
        breakdown[name]["pct"] = breakdown[name]["valor"] / total if total else 0.0
    wb.close()
    return total, breakdown


def load_premissas():
    with open(PREMISSAS_JSON, encoding="utf-8") as f:
        return json.load(f)


# ============================================================================
# MODELAGEM — funções puras que trabalham sobre arrays numpy
# ============================================================================
def build_time_axis(prem):
    """Retorna array de meses -prazo_pre .. +prazo_obra+prazo_pos (inclusivo)."""
    prazo_obra = prem["projeto"]["prazo_obra_meses"]
    prazo_pre = prem["projeto"]["prazo_pre_lancamento_meses"]
    prazo_pos = prem["projeto"]["prazo_pos_obra_meses"]
    t0 = -prazo_pre
    t1 = prazo_obra + prazo_pos
    return np.arange(t0, t1 + 1), prazo_obra, prazo_pre, prazo_pos


def sales_curve_cumulative(t_axis, t_inicio, t50, k):
    """Curva logística cumulativa — zero antes de t_inicio."""
    raw = 1.0 / (1.0 + np.exp(-k * (t_axis - t50)))
    base_inicio = 1.0 / (1.0 + np.exp(-k * (t_inicio - t50)))
    cum = np.where(t_axis < t_inicio, 0.0, (raw - base_inicio) / (1.0 - base_inicio))
    cum = np.clip(cum, 0.0, 1.0)
    return cum


def monthly_sales_fraction(t_axis, t_inicio, t50, k):
    """Fração vendida em cada mês (derivada discreta da cumulativa)."""
    cum = sales_curve_cumulative(t_axis, t_inicio, t50, k)
    diff = np.diff(cum, prepend=0.0)
    diff = np.clip(diff, 0.0, None)
    # normaliza para somar exatamente 1 (corrige pequenos erros de clip)
    s = diff.sum()
    if s > 0:
        diff = diff / s
    return diff


def cost_curve_per_macrogrupo(t_axis, prazo_pre, breakdown, curva_config):
    """
    Distribui cada macrogrupo uniformemente dentro de sua janela [t_start, t_end]
    (1-indexed sobre o prazo de obra). Usado apenas pra visualização Gantt/stacked.
    Retorna array de desembolso mensal.
    """
    out = np.zeros(len(t_axis))
    for nome, window in curva_config.items():
        if nome.startswith("_"):
            continue
        if nome not in breakdown:
            continue
        t_start_obra, t_end_obra = window
        idx_start = prazo_pre + (t_start_obra - 1)
        idx_end = prazo_pre + (t_end_obra - 1)
        idx_start = max(0, idx_start)
        idx_end = min(len(t_axis) - 1, idx_end)
        span = idx_end - idx_start + 1
        if span <= 0:
            continue
        valor_total = breakdown[nome]["valor"]
        out[idx_start:idx_end + 1] += valor_total / span
    return out


def cost_curve_from_monthly_pct(t_axis, prazo_pre, prazo_obra, custo_total, pct_list):
    """
    Distribui o custo total usando um vetor mensal explícito (mesma curva da
    apresentação ppt). Vetor tem `prazo_obra` entradas e deve somar 1.0.
    """
    pct = np.array(pct_list, dtype=float)
    if len(pct) != prazo_obra:
        raise ValueError(f"curva_mensal_obra_pct tem {len(pct)} valores, esperado {prazo_obra}")
    s = pct.sum()
    if s <= 0:
        raise ValueError("curva_mensal_obra_pct soma <= 0")
    pct = pct / s  # normaliza se somar um pouco diferente de 1
    out = np.zeros(len(t_axis))
    out[prazo_pre:prazo_pre + prazo_obra] = pct * custo_total
    return out


def compute_fluxo(prem, custo_obra_total, breakdown, overrides=None):
    """
    Computa fluxo de caixa mensal completo.
    overrides: dict para Monte Carlo / cenários — chaves: vgv_m2, custo_fator,
               t50, cub_aa, tma_aa, inadimplencia, corretagem_pct, ret_pct,
               entrada_pct.
    """
    ov = overrides or {}

    # Projeto
    proj = prem["projeto"]
    ac = proj["area_construida_m2"]
    pct_priv = proj["pct_area_privativa"]
    permuta_pct = proj["permuta_pct_unidades"]
    area_priv = ac * pct_priv
    area_vend = area_priv * (1.0 - permuta_pct)

    t_axis, prazo_obra, prazo_pre, prazo_pos = build_time_axis(prem)
    n = len(t_axis)

    # Receita
    rec = prem["receita"]
    vgv_m2 = ov.get("vgv_m2", rec["vgv_m2_base"])
    entrada_pct = ov.get("entrada_pct", rec["entrada_pct"])
    obra_pct = rec["obra_pct"]
    repasse_pct = rec["repasse_pct"]
    inadimplencia_pct = ov.get("inadimplencia_pct", rec["inadimplencia_pct"])

    vgv_bruto = area_priv * vgv_m2  # VGV total (antes da permuta)
    vgv_liquido = area_vend * vgv_m2  # VGV incorporadora (após permuta)

    # Financeiras
    fin = prem["financeiras"]
    cub_aa = ov.get("cub_aa", fin["cub_aa"])
    tma_aa = ov.get("tma_aa", fin["tma_aa"])
    cub_am = (1.0 + cub_aa) ** (1.0 / 12.0) - 1.0
    tma_am = (1.0 + tma_aa) ** (1.0 / 12.0) - 1.0

    # Comercial
    com = prem["comercial"]
    corretagem_pct = ov.get("corretagem_pct", com["corretagem_pct"])
    marketing_pct = com["marketing_pct"]
    marketing_front = com["marketing_front_pct"]
    ret_pct = ov.get("ret_pct", com["ret_pct"])
    adm_pct = com["adm_incorporadora_pct"]

    # Custo
    custo_fator = ov.get("custo_fator", 1.0)
    custo_total_ajustado = custo_obra_total * custo_fator
    contingencia_pct = prem["custo"]["contingencia_pct"]

    # Curva de vendas
    cv = prem["curva_vendas"]
    t_inicio = ov.get("t_inicio", cv["t_inicio"])
    t50 = ov.get("t50", cv["t50"])
    k = cv["k"]
    frac_vendas_mes = monthly_sales_fraction(t_axis, t_inicio, t50, k)

    # === RECEITA ===
    # Para cada mês m, venda bruta = vgv_liquido * frac[m]
    venda_mes = vgv_liquido * frac_vendas_mes  # valor contratado no mês
    entrada = venda_mes * entrada_pct

    # Parcelas obra: cada venda contratada no mês m paga obra_pct
    # distribuído entre mês m+1 e mês da entrega (prazo_obra_idx), corrigido CUB
    obra_idx_fim = prazo_pre + prazo_obra  # index do fim da obra (mês 30 → index 33 se pre=3)
    parcelas_obra = np.zeros(n)
    for m in range(n):
        if venda_mes[m] <= 0:
            continue
        # meses disponíveis entre (m+1) e obra_idx_fim
        start = m + 1
        end = obra_idx_fim
        if start > end:
            # venda ocorreu depois da obra — repasse direto, sem parcelas obra
            continue
        span = end - start + 1
        base_parcela = (venda_mes[m] * obra_pct) / span
        for t in range(start, end + 1):
            correcao = (1.0 + cub_am) ** (t - m)
            parcelas_obra[t] += base_parcela * correcao

    # Repasse/chave: 40% no mês da entrega+2 (index obra_idx_fim + 2)
    repasse = np.zeros(n)
    repasse_idx = min(obra_idx_fim + 2, n - 1)
    # Para cada venda no mês m, 40% entra em repasse_idx
    # (simplificação: todas as vendas repassam no mesmo mês)
    repasse[repasse_idx] = (venda_mes * repasse_pct).sum()

    # Inadimplência (aplicada linearmente sobre recebido)
    receita_bruta = entrada + parcelas_obra + repasse
    inadimplencia = -receita_bruta * inadimplencia_pct
    receita_liquida = receita_bruta + inadimplencia

    # === DESPESAS ===
    # Custo obra — prioridade: curva mensal explícita (igual ao ppt).
    # Fallback: distribuição por macrogrupo via janelas.
    curva_obra_config = prem["curva_obra_por_macrogrupo"]
    curva_mensal_cfg = prem.get("curva_mensal_obra_pct")

    if curva_mensal_cfg and isinstance(curva_mensal_cfg, dict) and "valores" in curva_mensal_cfg:
        custo_obra_mensal = cost_curve_from_monthly_pct(
            t_axis, prazo_pre, prazo_obra, custo_total_ajustado, curva_mensal_cfg["valores"]
        )
    else:
        breakdown_ajustado = {
            k: {"valor": v["valor"] * custo_fator, "pct": v["pct"]}
            for k, v in breakdown.items()
        }
        custo_obra_mensal = cost_curve_per_macrogrupo(
            t_axis, prazo_pre, breakdown_ajustado, curva_obra_config
        )

    # Contingência: % do custo obra, distribuída igual à curva de obra
    total_obra = custo_obra_mensal.sum()
    if total_obra > 0:
        contingencia_mensal = custo_obra_mensal * contingencia_pct
    else:
        contingencia_mensal = np.zeros(n)

    # Corretagem: 6% sobre venda contratada no mês (paga no mesmo mês)
    corretagem = venda_mes * corretagem_pct

    # Marketing: front_pct no pré-lançamento, rest durante obra.
    # Default atual (pós 2026-04-20): front=0 → 100% distribuído linear na obra.
    marketing_total = vgv_liquido * marketing_pct
    marketing_mensal = np.zeros(n)
    if marketing_front > 0 and prazo_pre > 0:
        marketing_mensal[:prazo_pre] = (marketing_total * marketing_front) / prazo_pre
    if prazo_obra > 0:
        marketing_mensal[prazo_pre:prazo_pre + prazo_obra] = (
            marketing_total * (1.0 - marketing_front) / prazo_obra
        )

    # RET 4% sobre receita recebida
    ret = receita_bruta * ret_pct  # sobre bruta porque é o que entrou

    # Admin incorporadora: 3% VGV distribuído linear pelo horizonte
    adm_total = vgv_liquido * adm_pct
    adm_mensal = np.full(n, adm_total / n)

    despesa_total = (
        custo_obra_mensal + contingencia_mensal + corretagem +
        marketing_mensal + ret + adm_mensal
    )

    # === FLUXO LÍQUIDO ===
    fluxo_liquido = receita_liquida - despesa_total
    fluxo_acum = np.cumsum(fluxo_liquido)

    # Fator desconto e fluxo descontado
    desconto = np.array([1.0 / (1.0 + tma_am) ** i for i in range(n)])
    fluxo_descontado = fluxo_liquido * desconto

    return {
        "t_axis": t_axis,
        "n": n,
        "prazo_pre": prazo_pre,
        "prazo_obra": prazo_obra,
        "prazo_pos": prazo_pos,
        "area_priv": area_priv,
        "area_vend": area_vend,
        "vgv_bruto": vgv_bruto,
        "vgv_liquido": vgv_liquido,
        "vgv_m2": vgv_m2,
        "custo_obra_total": custo_obra_total * custo_fator,
        "tma_am": tma_am,
        "tma_aa": tma_aa,
        "cub_am": cub_am,
        "cub_aa": cub_aa,
        "frac_vendas": frac_vendas_mes,
        "venda_mes": venda_mes,
        "entrada": entrada,
        "parcelas_obra": parcelas_obra,
        "repasse": repasse,
        "inadimplencia": inadimplencia,
        "receita_bruta": receita_bruta,
        "receita_liquida": receita_liquida,
        "custo_obra_mensal": custo_obra_mensal,
        "contingencia_mensal": contingencia_mensal,
        "corretagem": corretagem,
        "marketing_mensal": marketing_mensal,
        "ret": ret,
        "adm_mensal": adm_mensal,
        "despesa_total": despesa_total,
        "fluxo_liquido": fluxo_liquido,
        "fluxo_acum": fluxo_acum,
        "fluxo_descontado": fluxo_descontado,
        "desconto": desconto,
    }


def compute_kpis(fluxo):
    """Extrai KPIs do fluxo (VPL, TIR, margem, payback, ROI, exposição)."""
    fl = fluxo["fluxo_liquido"]
    fa = fluxo["fluxo_acum"]
    fd = fluxo["fluxo_descontado"]
    tma_am = fluxo["tma_am"]

    vpl = float(nf.npv(tma_am, fl))
    try:
        tir_m = float(nf.irr(fl))
        if np.isnan(tir_m):
            tir_m = None
    except Exception:
        tir_m = None
    tir_a = (1.0 + tir_m) ** 12 - 1.0 if tir_m is not None else None

    receita_total = float(fluxo["receita_liquida"].sum())
    despesa_total = float(fluxo["despesa_total"].sum())
    lucro = receita_total - despesa_total
    margem = lucro / receita_total if receita_total > 0 else 0.0

    # Payback descontado: primeiro mês em que acumulado descontado >= 0
    fd_acum = np.cumsum(fd)
    payback_idx = np.argmax(fd_acum >= 0) if (fd_acum >= 0).any() else None
    payback_meses = (payback_idx - fluxo["prazo_pre"]) if payback_idx is not None else None

    # Exposição máxima (fluxo acumulado mais negativo)
    exposicao_max = float(-fa.min()) if fa.min() < 0 else 0.0
    roi = lucro / exposicao_max if exposicao_max > 0 else 0.0

    return {
        "vpl": vpl,
        "tir_mensal": tir_m,
        "tir_anual": tir_a,
        "lucro": lucro,
        "margem": margem,
        "receita_total": receita_total,
        "despesa_total": despesa_total,
        "exposicao_max": exposicao_max,
        "roi": roi,
        "payback_meses": payback_meses,
        "vgv_bruto": fluxo["vgv_bruto"],
        "vgv_liquido": fluxo["vgv_liquido"],
        "custo_obra_total": fluxo["custo_obra_total"],
    }


# ============================================================================
# MONTE CARLO
# ============================================================================
def run_monte_carlo(prem, custo_obra_total, breakdown):
    mc = prem["monte_carlo"]
    n_iter = mc["n_iter"]
    seed = mc["seed"]
    rng = np.random.default_rng(seed)

    vgv_tri = mc["vgv_m2_tri"]
    custo_tri = mc["custo_obra_fator_tri"]
    t50_tri = mc["t50_vendas_tri"]
    cub_tri = mc["cub_aa_tri"]
    rho = mc["corr_vgv_t50"]

    # Correlação via Cholesky para vgv x t50
    # Sorteamos duas normais correlacionadas, mapeamos via CDF para uniform,
    # depois aplicamos inverse triangular.
    mean = [0.0, 0.0]
    cov = [[1.0, rho], [rho, 1.0]]
    normals = rng.multivariate_normal(mean, cov, size=n_iter)
    # CDF normal → uniform (0,1) via erf (evita dependência scipy)
    from math import erf, sqrt
    vec_cdf = np.vectorize(lambda x: 0.5 * (1.0 + erf(x / sqrt(2.0))))
    uniforms = vec_cdf(normals)
    u_vgv = uniforms[:, 0]
    u_t50 = uniforms[:, 1]

    def inv_triang(u, a, m, b):
        """Inverse CDF de uma triangular com min a, mode m, max b."""
        f_c = (m - a) / (b - a)
        out = np.where(
            u < f_c,
            a + np.sqrt(u * (b - a) * (m - a)),
            b - np.sqrt((1 - u) * (b - a) * (b - m)),
        )
        return out

    vgv_samples = inv_triang(u_vgv, *vgv_tri)
    t50_samples = inv_triang(u_t50, *t50_tri)

    custo_samples = rng.triangular(*custo_tri, size=n_iter)
    cub_samples = rng.triangular(*cub_tri, size=n_iter)

    vpl_arr = np.empty(n_iter)
    tir_arr = np.empty(n_iter)
    margem_arr = np.empty(n_iter)

    # Loop — compute_fluxo é rápido porém não vetorizado
    for i in range(n_iter):
        ov = {
            "vgv_m2": vgv_samples[i],
            "custo_fator": custo_samples[i],
            "t50": t50_samples[i],
            "cub_aa": cub_samples[i],
        }
        f = compute_fluxo(prem, custo_obra_total, breakdown, overrides=ov)
        kp = compute_kpis(f)
        vpl_arr[i] = kp["vpl"]
        tir_arr[i] = kp["tir_anual"] if kp["tir_anual"] is not None else np.nan
        margem_arr[i] = kp["margem"]

    tir_clean = tir_arr[~np.isnan(tir_arr)]
    tma_aa = prem["financeiras"]["tma_aa"]

    stats = {
        "n_iter": n_iter,
        "vpl_media": float(vpl_arr.mean()),
        "vpl_mediana": float(np.median(vpl_arr)),
        "vpl_desvio": float(vpl_arr.std()),
        "vpl_p10": float(np.percentile(vpl_arr, 10)),
        "vpl_p50": float(np.percentile(vpl_arr, 50)),
        "vpl_p90": float(np.percentile(vpl_arr, 90)),
        "tir_mediana_aa": float(np.median(tir_clean)) if len(tir_clean) > 0 else None,
        "tir_p10_aa": float(np.percentile(tir_clean, 10)) if len(tir_clean) > 0 else None,
        "tir_p90_aa": float(np.percentile(tir_clean, 90)) if len(tir_clean) > 0 else None,
        "prob_vpl_positivo": float((vpl_arr > 0).mean()),
        "prob_tir_gt_tma": float((tir_clean > tma_aa).mean()) if len(tir_clean) > 0 else 0.0,
        "prob_margem_gt_15": float((margem_arr > 0.15).mean()),
        "margem_mediana": float(np.median(margem_arr)),
    }
    samples = {
        "vpl": vpl_arr,
        "tir_anual": tir_arr,
        "margem": margem_arr,
        "vgv_m2": vgv_samples,
        "custo_fator": custo_samples,
        "t50": t50_samples,
    }
    return stats, samples


# ============================================================================
# TORNADO
# ============================================================================
def run_tornado(prem, custo_obra_total, breakdown, base_vpl):
    variacao = prem["tornado"]["variacao_pct"]
    variaveis = prem["tornado"]["variaveis"]
    results = []

    def run_with(ov):
        f = compute_fluxo(prem, custo_obra_total, breakdown, overrides=ov)
        return compute_kpis(f)["vpl"]

    for var in variaveis:
        # Mapeia cada variável do JSON para override key
        if var == "vgv_m2_base":
            base = prem["receita"]["vgv_m2_base"]
            low = run_with({"vgv_m2": base * (1 - variacao)})
            high = run_with({"vgv_m2": base * (1 + variacao)})
            label = "VGV R$/m²"
        elif var == "custo_obra_fator":
            low = run_with({"custo_fator": 1 - variacao})
            high = run_with({"custo_fator": 1 + variacao})
            label = "Custo obra"
        elif var == "t50_vendas":
            base = prem["curva_vendas"]["t50"]
            low = run_with({"t50": base * (1 - variacao)})
            high = run_with({"t50": base * (1 + variacao)})
            label = "t50 vendas"
        elif var == "entrada_pct":
            base = prem["receita"]["entrada_pct"]
            low = run_with({"entrada_pct": base * (1 - variacao)})
            high = run_with({"entrada_pct": base * (1 + variacao)})
            label = "Entrada %"
        elif var == "corretagem_pct":
            base = prem["comercial"]["corretagem_pct"]
            low = run_with({"corretagem_pct": base * (1 - variacao)})
            high = run_with({"corretagem_pct": base * (1 + variacao)})
            label = "Corretagem"
        elif var == "ret_pct":
            base = prem["comercial"]["ret_pct"]
            low = run_with({"ret_pct": base * (1 - variacao)})
            high = run_with({"ret_pct": base * (1 + variacao)})
            label = "RET 4%"
        elif var == "cub_aa":
            base = prem["financeiras"]["cub_aa"]
            low = run_with({"cub_aa": base * (1 - variacao)})
            high = run_with({"cub_aa": base * (1 + variacao)})
            label = "CUB/SC"
        elif var == "tma_aa":
            base = prem["financeiras"]["tma_aa"]
            low = run_with({"tma_aa": base * (1 - variacao)})
            high = run_with({"tma_aa": base * (1 + variacao)})
            label = "TMA"
        else:
            continue
        delta = high - low
        results.append({
            "var": var,
            "label": label,
            "vpl_low": low,
            "vpl_high": high,
            "delta": delta,
            "abs_delta": abs(delta),
        })
    results.sort(key=lambda r: r["abs_delta"], reverse=True)
    return results, base_vpl


# ============================================================================
# CENÁRIOS
# ============================================================================
def run_cenarios(prem, custo_obra_total, breakdown):
    cen = prem["cenarios"]
    vgv_base = prem["receita"]["vgv_m2_base"]
    t50_base = prem["curva_vendas"]["t50"]
    out = {}
    for nome, cfg in cen.items():
        ov = {
            "vgv_m2": vgv_base * cfg["vgv_fator"],
            "custo_fator": cfg["custo_fator"],
            "t50": t50_base + cfg["t50_delta"],
        }
        f = compute_fluxo(prem, custo_obra_total, breakdown, overrides=ov)
        out[nome] = compute_kpis(f)
        out[nome]["_cfg"] = cfg
    return out


# ============================================================================
# PLOTS
# ============================================================================
def plot_histogram(values, title, xlabel, out_path, color="#2980B9"):
    PLOTS_DIR.mkdir(exist_ok=True)
    fig, ax = plt.subplots(figsize=(6.5, 4.0), dpi=100)
    clean = values[~np.isnan(values)] if values.dtype.kind == "f" else values
    ax.hist(clean, bins=50, color=color, edgecolor="white", alpha=0.85)
    ax.axvline(np.median(clean), color="#C25700", linestyle="--", linewidth=2, label=f"Mediana")
    ax.axvline(0, color="#555", linestyle=":", linewidth=1)
    ax.set_title(title, fontsize=11, color="#2C3E50", fontweight="bold")
    ax.set_xlabel(xlabel, fontsize=9)
    ax.set_ylabel("Frequência", fontsize=9)
    ax.grid(axis="y", alpha=0.3)
    ax.legend(fontsize=8)
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


def plot_tornado(results, base_vpl, out_path):
    PLOTS_DIR.mkdir(exist_ok=True)
    labels = [r["label"] for r in results]
    lows = [r["vpl_low"] - base_vpl for r in results]
    highs = [r["vpl_high"] - base_vpl for r in results]

    fig, ax = plt.subplots(figsize=(7.0, 4.5), dpi=100)
    y_pos = np.arange(len(labels))
    ax.barh(y_pos, lows, color="#C75D5D", label="-20%")
    ax.barh(y_pos, highs, color="#5D9C7A", label="+20%")
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.axvline(0, color="#333", linewidth=0.8)
    ax.set_xlabel("Δ VPL (R$) em relação ao cenário base", fontsize=9)
    ax.set_title("Tornado — Sensibilidade sobre VPL", fontsize=11, color="#2C3E50", fontweight="bold")
    ax.grid(axis="x", alpha=0.3)
    ax.legend(fontsize=8, loc="lower right")
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


# ============================================================================
# BUILD — abas
# ============================================================================
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_capa(ws, prem, kpis, mc_stats):
    set_col_widths(ws, [22, 20, 18, 18, 18, 20])
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "ESTUDO DE VIABILIDADE FINANCEIRA"
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"{prem['projeto']['nome']} — {prem['projeto']['bairro']}, {prem['projeto']['cidade']}/{prem['projeto']['estado']}"
    c.font = Font(name="Arial", size=12, bold=True, color=DARK)
    c.alignment = ALIGN_CENTER

    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = f"Cliente: {prem['projeto']['cliente']}  |  Gerado em: {prem['_gerado_em']}"
    c.font = FONT_SMALL
    c.alignment = ALIGN_CENTER

    # Resumo KPIs
    r = 5
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="RESUMO EXECUTIVO — CENÁRIO BASE")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER

    rows = [
        ("VGV bruto (antes permuta)", kpis["vgv_bruto"], FMT_BRL),
        ("VGV líquido (pós-permuta 30%)", kpis["vgv_liquido"], FMT_BRL),
        ("Custo de obra (paramétrico)", kpis["custo_obra_total"], FMT_BRL),
        ("Receita total (horizonte)", kpis["receita_total"], FMT_BRL),
        ("Despesa total (horizonte)", kpis["despesa_total"], FMT_BRL),
        ("Lucro líquido", kpis["lucro"], FMT_BRL),
        ("Margem líquida", kpis["margem"], FMT_PCT),
        ("VPL @ TMA", kpis["vpl"], FMT_BRL),
        ("TIR anual", kpis["tir_anual"], FMT_PCT),
        ("Payback descontado (meses pós-início obra)", kpis["payback_meses"], FMT_NUM),
        ("Exposição máxima (capital)", kpis["exposicao_max"], FMT_BRL),
        ("ROI sobre capital", kpis["roi"], FMT_PCT),
        ("", None, None),
        ("Monte Carlo — n iterações", mc_stats["n_iter"], FMT_NUM),
        ("P(VPL > 0)", mc_stats["prob_vpl_positivo"], FMT_PCT),
        ("P(TIR > TMA)", mc_stats["prob_tir_gt_tma"], FMT_PCT),
        ("VPL mediana (MC)", mc_stats["vpl_mediana"], FMT_BRL),
        ("VPL P10 (MC)", mc_stats["vpl_p10"], FMT_BRL),
        ("VPL P90 (MC)", mc_stats["vpl_p90"], FMT_BRL),
    ]
    r = 6
    for label, val, fmt in rows:
        if label == "":
            r += 1
            continue
        ws.merge_cells(f"A{r}:C{r}")
        lc = ws.cell(row=r, column=1, value=label)
        style_label(lc)
        ws.merge_cells(f"D{r}:F{r}")
        vc = ws.cell(row=r, column=4, value=val)
        style_calc(vc, fmt=fmt)
        r += 1

    # Nota de fonte
    r += 2
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="Fonte custo: parametrico-arthen-arboris.xlsx CUSTOS_MACROGRUPO!D22")
    c.font = FONT_SMALL
    c.alignment = ALIGN_LEFT
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="Fonte preço venda: pesquisa Perequê abr/2026 (Viva Real, MySide, Imovelweb) — mediana R$ 15.394/m²")
    c.font = FONT_SMALL
    c.alignment = ALIGN_LEFT


def build_premissas(ws, prem, kpis):
    set_col_widths(ws, [36, 20, 20, 44])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "PREMISSAS — editar valores e regerar via gerar_viabilidade_arthen_arboris.py"
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    def section_header(r, titulo):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(row=r, column=1, value=titulo)
        c.fill = FILL_ACCENT
        c.font = FONT_H1
        c.alignment = ALIGN_CENTER
        return r + 1

    def row_input(r, label, valor, fmt, nota=""):
        ws.cell(row=r, column=1, value=label).font = FONT_BODY
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).border = BORDER
        vc = ws.cell(row=r, column=2, value=valor)
        style_input(vc, fmt=fmt)
        nc = ws.cell(row=r, column=4, value=nota)
        nc.font = FONT_SMALL
        nc.alignment = ALIGN_LEFT
        nc.border = BORDER
        ws.cell(row=r, column=3).border = BORDER
        return r + 1

    proj = prem["projeto"]
    rec = prem["receita"]
    fin = prem["financeiras"]
    com = prem["comercial"]
    cv = prem["curva_vendas"]

    r = 3
    r = section_header(r, "PROJETO")
    r = row_input(r, "Nome", proj["nome"], None, "")
    r = row_input(r, "Cidade/Bairro", f"{proj['bairro']}, {proj['cidade']}/{proj['estado']}", None, "")
    r = row_input(r, "Área construída (m²)", proj["area_construida_m2"], FMT_NUM2, "do paramétrico")
    r = row_input(r, "% área privativa", proj["pct_area_privativa"], FMT_PCT, "AC que vira unidade vendável")
    r = row_input(r, "Área privativa total (m²)", kpis["vgv_bruto"] / rec["vgv_m2_base"], FMT_NUM2, "AC × pct_priv")
    r = row_input(r, "Unidades", proj["n_unidades"], FMT_NUM, "")
    r = row_input(r, "Permuta (% das unidades)", proj["permuta_pct_unidades"], FMT_PCT, "terreno via permuta física")
    r = row_input(r, "Prazo obra (meses)", proj["prazo_obra_meses"], FMT_NUM, "")
    r = row_input(r, "Prazo pré-lançamento (meses)", proj["prazo_pre_lancamento_meses"], FMT_NUM, "meses -3 a 0")
    r = row_input(r, "Prazo pós-obra (meses)", proj["prazo_pos_obra_meses"], FMT_NUM, "cauda de recebimento")

    r += 1
    r = section_header(r, "CUSTO")
    r = row_input(r, "Custo de obra total (R$)", kpis["custo_obra_total"], FMT_BRL, "paramétrico V2")
    r = row_input(r, "Contingência (%)", prem["custo"]["contingencia_pct"], FMT_PCT, "sobre custo obra")

    r += 1
    r = section_header(r, "RECEITA")
    r = row_input(r, "VGV R$/m² privativo", rec["vgv_m2_base"], FMT_BRL2, "Perequê médio pesquisa 2026")
    r = row_input(r, "VGV bruto (R$)", kpis["vgv_bruto"], FMT_BRL, "área_priv × R$/m²")
    r = row_input(r, "VGV líquido permuta (R$)", kpis["vgv_liquido"], FMT_BRL, "após 30% de permuta")
    r = row_input(r, "Entrada (% no ato)", rec["entrada_pct"], FMT_PCT, "")
    r = row_input(r, "Obra (% parcelado CUB)", rec["obra_pct"], FMT_PCT, "durante obra")
    r = row_input(r, "Repasse/chave (%)", rec["repasse_pct"], FMT_PCT, "mês de entrega +2")
    r = row_input(r, "Inadimplência (%)", rec["inadimplencia_pct"], FMT_PCT, "sobre recebido")

    r += 1
    r = section_header(r, "FINANCEIRAS")
    r = row_input(r, "CUB/SC (a.a.)", fin["cub_aa"], FMT_PCT, "indexador parcelas obra")
    r = row_input(r, "TMA (a.a.)", fin["tma_aa"], FMT_PCT, "custo de capital — desconto VPL")

    r += 1
    r = section_header(r, "COMERCIAL / IMPOSTOS / ADMIN")
    r = row_input(r, "Corretagem (% VGV)", com["corretagem_pct"], FMT_PCT, "paga no mês da venda")
    r = row_input(r, "Marketing (% VGV)", com["marketing_pct"], FMT_PCT, "stand, material, digital")
    r = row_input(r, "Marketing front-load (% no pré-lanç)", com["marketing_front_pct"], FMT_PCT, "")
    r = row_input(r, "RET (% receita)", com["ret_pct"], FMT_PCT, "Regime Especial de Tributação")
    r = row_input(r, "Admin incorporadora (% VGV)", com["adm_incorporadora_pct"], FMT_PCT, "equipe, jurídico, contábil")

    r += 1
    r = section_header(r, "CURVA DE VENDAS (logística)")
    r = row_input(r, "t_início (mês — pré-lanç)", cv["t_inicio"], FMT_NUM, "vendas começam em t_início")
    r = row_input(r, "t50 (mês da inflexão)", cv["t50"], FMT_NUM, "50% vendido neste mês")
    r = row_input(r, "k (inclinação)", cv["k"], FMT_NUM2, "maior k = venda mais concentrada")


def build_fluxo_caixa(ws, fluxo, prem, fluxo_png=None):
    """FLUXO_CAIXA didático — layout pensado para leitura de diretor/cliente.

    Estrutura:
      - Linha 1: Título
      - Linhas 2-4: "Como ler" (caixa amarela)
      - Linha 6: Legenda fase + instrução
      - Linha 7-9: Header (Mês#, label amigável Pré-1/Obra-1/Pós-1, Fase)
      - Colunas: A=rótulo | B=TOTAL | C=PRÉ-LANÇ | D=OBRA | E=PÓS-OBRA | F..=mensal
      - Blocos verticais: ENTRADAS, SAÍDAS, RESULTADO
      - Destaque visual: cond. formatting no fluxo líquido (verde/vermelho) e
        pico negativo (acumulado)
      - Final: eventos importantes + gráfico embedido
    """
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

    n = fluxo["n"]
    t_axis = fluxo["t_axis"]
    prazo_pre = fluxo["prazo_pre"]
    prazo_obra = fluxo["prazo_obra"]

    # Larguras
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 36  # rótulo
    ws.column_dimensions["B"].width = 16  # TOTAL
    ws.column_dimensions["C"].width = 14  # PRÉ
    ws.column_dimensions["D"].width = 14  # OBRA
    ws.column_dimensions["E"].width = 14  # PÓS
    COL_M0 = 6  # primeiro mês = col F (index 6)
    for i in range(n):
        ws.column_dimensions[get_column_letter(COL_M0 + i)].width = 10.5

    last_col = COL_M0 + n - 1
    last_letter = get_column_letter(last_col)

    # Máscaras (índices por fase)
    mask_pre = (t_axis < 0)
    mask_obra = (t_axis >= 0) & (t_axis <= prazo_obra)
    mask_pos = (t_axis > prazo_obra)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=(
        f"FLUXO DE CAIXA MENSAL — {n} meses "
        f"({prazo_pre} pré-lançamento + {prazo_obra} obra + {fluxo['prazo_pos']} pós-obra)"
    ))
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    # Caixa COMO LER
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=last_col)
    c = ws.cell(row=2, column=1, value=(
        "COMO LER:  cada coluna é 1 mês do projeto. Colunas B–E são subtotais por fase (TOTAL / PRÉ-LANÇ / OBRA / PÓS-OBRA). "
        "Os blocos VERDE = ENTRADAS de caixa (vendas).  VERMELHO = SAÍDAS (custos).  CINZA = RESULTADO (entradas − saídas).\n"
        "O 'Saldo mensal' fica verde quando positivo, vermelho quando negativo. O 'Saldo acumulado' é o 'caixa do projeto' ao "
        "longo do tempo — o ponto mais negativo dele é a EXPOSIÇÃO MÁXIMA (capital que a incorporadora precisa ter disponível). "
        "O mês em que o acumulado volta a zero é o BREAK-EVEN de caixa."
    ))
    c.font = Font(name="Arial", size=10)
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_YELLOW
    c.border = BORDER
    for rr in range(2, 5):
        ws.row_dimensions[rr].height = 20

    # Linha 6: header cinza separador
    ws.row_dimensions[6].height = 6

    # Linha 7: Header TOTAL + fases + mês (t=...)
    c = ws.cell(row=7, column=1, value="Linha  /  Mês do projeto →")
    c.font = FONT_BOLD
    c.alignment = ALIGN_LEFT
    c.fill = FILL_GRAY
    c.border = BORDER
    for col, label in [(2, "TOTAL"), (3, "PRÉ-LANÇ"), (4, "OBRA"), (5, "PÓS-OBRA")]:
        cc = ws.cell(row=7, column=col, value=label)
        cc.fill = FILL_DARK
        cc.font = FONT_H1
        cc.alignment = ALIGN_CENTER
        cc.border = BORDER
    for i, t in enumerate(t_axis):
        cc = ws.cell(row=7, column=COL_M0 + i, value=int(t))
        cc.font = FONT_BOLD
        cc.alignment = ALIGN_CENTER
        cc.fill = FILL_ACCENT
        cc.border = BORDER
        cc.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    # Linha 8: label amigável
    c = ws.cell(row=8, column=1, value="Fase do projeto")
    c.font = FONT_SMALL
    c.alignment = ALIGN_LEFT
    c.border = BORDER
    for col in range(2, 6):
        ws.cell(row=8, column=col).border = BORDER
    FASE_FILL = {
        "PRÉ-LANÇ": PatternFill("solid", fgColor="FFE082"),
        "OBRA":     PatternFill("solid", fgColor="A5D6A7"),
        "PÓS-OBRA": PatternFill("solid", fgColor="90CAF9"),
    }
    for i, t in enumerate(t_axis):
        if t < 0:
            fase = "PRÉ-LANÇ"
            label = f"Pré {int(t + prazo_pre + 1)}"
        elif t == 0:
            fase = "OBRA"
            label = "Obra 1"
        elif t <= prazo_obra:
            fase = "OBRA"
            label = f"Obra {int(t) + 1}" if t > 0 else "Obra 1"
            # t goes 0..prazo_obra inclusive. Make it "Obra 1"..f"Obra {prazo_obra}"
            # Using: label = f"Obra {int(t)+1}" when t in [0..prazo_obra-1]; "Entrega" at t=prazo_obra
            if int(t) == prazo_obra:
                label = "Entrega"
                fase = "OBRA"
            else:
                label = f"Obra {int(t) + 1}"
        else:
            fase = "PÓS-OBRA"
            label = f"Pós {int(t) - prazo_obra}"
        cc = ws.cell(row=8, column=COL_M0 + i, value=label)
        cc.font = Font(name="Arial", size=8, bold=True)
        cc.alignment = ALIGN_CENTER
        cc.fill = FASE_FILL[fase]
        cc.border = BORDER

    # Frozen panes — trava linhas header e col label
    ws.freeze_panes = "F9"

    def write_line(row, label, arr, fmt=FMT_BRL, is_total=False, fill_row=None):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = FONT_BOLD if is_total else FONT_BODY
        lc.alignment = ALIGN_LEFT
        lc.border = BORDER
        if fill_row is not None:
            lc.fill = fill_row
        # TOTAL (col B)
        total = float(arr.sum())
        tc = ws.cell(row=row, column=2, value=total)
        tc.number_format = fmt
        tc.font = FONT_BOLD
        tc.fill = FILL_GRAY
        tc.border = BORDER
        tc.alignment = ALIGN_RIGHT
        # PRÉ / OBRA / PÓS subtotais
        for col, mask in [(3, mask_pre), (4, mask_obra), (5, mask_pos)]:
            sub = float(arr[mask].sum())
            sc = ws.cell(row=row, column=col, value=sub)
            sc.number_format = fmt
            sc.font = FONT_BOLD
            sc.fill = FILL_BLUE if col == 4 else FILL_GRAY
            sc.border = BORDER
            sc.alignment = ALIGN_RIGHT
        # Mensal
        for i, v in enumerate(arr):
            cell = ws.cell(row=row, column=COL_M0 + i, value=float(v))
            cell.number_format = fmt
            cell.font = FONT_BODY
            cell.alignment = ALIGN_RIGHT
            if fill_row is not None:
                cell.fill = fill_row
            cell.border = BORDER

    r = 9
    # ========== ENTRADAS ==========
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=1, value="   ▲  ENTRADAS DE CAIXA   (recebimentos das vendas)")
    c.fill = PatternFill("solid", fgColor="1B5E20")
    c.font = FONT_H1
    c.alignment = ALIGN_LEFT
    r += 1
    write_line(r, "% do VGV vendido (acumulado)", np.cumsum(fluxo["frac_vendas"]), fmt=FMT_PCT); r += 1
    write_line(r, "VGV contratado no mês", fluxo["venda_mes"]); r += 1
    write_line(r, "  Entrada (20% no ato)", fluxo["entrada"]); r += 1
    write_line(r, "  Parcelas obra (40% corrigido CUB)", fluxo["parcelas_obra"]); r += 1
    write_line(r, "  Repasse/chave (40% entrega+2)", fluxo["repasse"]); r += 1
    write_line(r, "  (−) Inadimplência (3%)", fluxo["inadimplencia"]); r += 1
    write_line(r, "ENTRADA LÍQUIDA MÊS", fluxo["receita_liquida"], is_total=True,
               fill_row=PatternFill("solid", fgColor="C8E6C9")); r += 2

    # ========== SAÍDAS ==========
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=1, value="   ▼  SAÍDAS DE CAIXA   (custos da obra + comercial + impostos + admin)")
    c.fill = PatternFill("solid", fgColor="B71C1C")
    c.font = FONT_H1
    c.alignment = ALIGN_LEFT
    r += 1
    write_line(r, "  Custo da obra (paramétrico)", fluxo["custo_obra_mensal"]); r += 1
    write_line(r, "  Contingência (5% do custo obra)", fluxo["contingencia_mensal"]); r += 1
    write_line(r, "  Corretagem (6% VGV — no mês da venda)", fluxo["corretagem"]); r += 1
    write_line(r, "  Marketing (3% VGV — linear na obra)", fluxo["marketing_mensal"]); r += 1
    write_line(r, "  RET 4% (Patrimônio de Afetação)", fluxo["ret"]); r += 1
    write_line(r, "  Admin incorporadora (3% VGV — linear)", fluxo["adm_mensal"]); r += 1
    write_line(r, "SAÍDA TOTAL MÊS", fluxo["despesa_total"], is_total=True,
               fill_row=PatternFill("solid", fgColor="FFCDD2")); r += 2

    # ========== RESULTADO ==========
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=1, value="   =  SALDO   (entradas − saídas)")
    c.fill = FILL_DARK
    c.font = FONT_H1
    c.alignment = ALIGN_LEFT
    r += 1
    r_fluxo_liquido = r
    write_line(r, "SALDO MENSAL   (verde=entrou / vermelho=saiu)", fluxo["fluxo_liquido"],
               is_total=True); r += 1
    r_fluxo_acum = r
    write_line(r, "SALDO ACUMULADO   (capital do projeto)", fluxo["fluxo_acum"],
               is_total=True, fill_row=FILL_GRAY); r += 1
    write_line(r, "Fator de desconto (TMA 14%)", fluxo["desconto"], fmt=FMT_NUM2); r += 1
    write_line(r, "Saldo descontado ao mês 0", fluxo["fluxo_descontado"], is_total=True); r += 1

    # Conditional formatting — saldo mensal (verde/vermelho)
    monthly_start = get_column_letter(COL_M0)
    monthly_range_liq = f"{monthly_start}{r_fluxo_liquido}:{last_letter}{r_fluxo_liquido}"
    ws.conditional_formatting.add(monthly_range_liq, CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", fgColor="C8E6C9"),  # verde claro
    ))
    ws.conditional_formatting.add(monthly_range_liq, CellIsRule(
        operator="lessThan", formula=["0"],
        fill=PatternFill("solid", fgColor="FFCDD2"),  # vermelho claro
    ))

    # Color scale — acumulado (vermelho → branco → verde)
    monthly_range_acum = f"{monthly_start}{r_fluxo_acum}:{last_letter}{r_fluxo_acum}"
    ws.conditional_formatting.add(monthly_range_acum, ColorScaleRule(
        start_type="min", start_color="B71C1C",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="1B5E20",
    ))

    r += 2

    # ========== EVENTOS IMPORTANTES ==========
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    c = ws.cell(row=r, column=1, value="   ★  EVENTOS IMPORTANTES DO PROJETO")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_LEFT
    r += 1

    # Calcular eventos
    fluxo_acum = fluxo["fluxo_acum"]
    idx_exposicao_max = int(np.argmin(fluxo_acum))
    mes_exposicao = int(t_axis[idx_exposicao_max])
    # Break-even: primeiro mês com acumulado >= 0 DEPOIS do pico negativo
    depois_pico = np.arange(len(fluxo_acum)) > idx_exposicao_max
    pos_acum = (fluxo_acum >= 0) & depois_pico
    if pos_acum.any():
        idx_breakeven = int(np.argmax(pos_acum))
        mes_breakeven = int(t_axis[idx_breakeven])
    else:
        mes_breakeven = None
    # Início obra = t=0; entrega = t=prazo_obra; fim = t[-1]
    eventos = [
        ("Mês 1 do pré-lançamento", int(t_axis[0]), "início das vendas"),
        ("Início da obra", 0, "canteiro montado, fundação inicia"),
        ("Pico de exposição de caixa", mes_exposicao,
         f"capital máximo exposto: R$ {-fluxo_acum[idx_exposicao_max]/1e6:.1f}M"),
        ("Entrega das unidades", prazo_obra, "habite-se, chaves, início repasse bancário"),
    ]
    if mes_breakeven is not None:
        eventos.append(("Break-even de caixa", mes_breakeven,
                        "saldo acumulado volta a zero — projeto pago"))
    eventos.append(("Fim do horizonte", int(t_axis[-1]),
                   f"final do pós-obra ({fluxo['prazo_pos']} meses após entrega)"))

    # Header de eventos
    for col, h in enumerate(["Evento", "Mês (t)", "Observação"], 1):
        cc = ws.cell(row=r, column=col, value=h)
        style_header(cc)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=last_col)
    r += 1
    for label, mes, obs in eventos:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = FONT_BOLD
        lc.alignment = ALIGN_LEFT
        lc.border = BORDER
        mc = ws.cell(row=r, column=2, value=mes)
        mc.font = FONT_BODY
        mc.alignment = ALIGN_CENTER
        mc.border = BORDER
        mc.fill = FILL_YELLOW
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=last_col)
        oc = ws.cell(row=r, column=3, value=obs)
        oc.font = FONT_BODY
        oc.alignment = ALIGN_LEFT
        oc.border = BORDER
        r += 1

    r += 2

    # ========== GRÁFICO embutido ==========
    if fluxo_png and fluxo_png.exists():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        c = ws.cell(row=r, column=1, value="   📈  SALDO ACUMULADO AO LONGO DOS MESES")
        c.fill = FILL_ACCENT
        c.font = FONT_H1
        c.alignment = ALIGN_LEFT
        r += 2
        img = XLImage(str(fluxo_png))
        img.width = 900
        img.height = 420
        ws.add_image(img, f"A{r}")

    return r_fluxo_liquido


def build_resultado(ws, prem, kpis, mc_stats, fluxo, r_fluxo_liquido_in_fc):
    set_col_widths(ws, [36, 20, 20, 44, 22])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "RESULTADO — VPL, TIR, margem, validação"
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    # Cabeçalhos
    r = 3
    for col, h in enumerate(["Indicador", "Valor (Python)", "Fórmula Excel", "Observação", "Status"], 1):
        c = ws.cell(row=r, column=col, value=h)
        style_header(c)

    # Range da linha do fluxo_liquido em FLUXO_CAIXA
    # No novo layout: A=rótulo | B=TOTAL | C=PRÉ | D=OBRA | E=PÓS | F..=mensal
    # Primeiro mês (t_axis[0]) está na col F (index 6).
    prazo_pre = fluxo["prazo_pre"]
    n = fluxo["n"]
    COL_M0 = 6
    col_tN = COL_M0 + n - 1
    col_tN_letter = get_column_letter(col_tN)
    first_col_letter = get_column_letter(COL_M0)

    # Excel NPV desconta TODOS os termos começando em t=1. numpy_financial trata
    # values[0] como t=0 (não desconta). Para bater: separa o 1º termo.
    tma_am = fluxo["tma_am"]
    second_col_letter = get_column_letter(COL_M0 + 1)
    npv_formula = (
        f"=FLUXO_CAIXA!{first_col_letter}{r_fluxo_liquido_in_fc}"
        f"+NPV({tma_am},FLUXO_CAIXA!{second_col_letter}{r_fluxo_liquido_in_fc}:{col_tN_letter}{r_fluxo_liquido_in_fc})"
    )
    irr_formula = f"=IRR(FLUXO_CAIXA!{first_col_letter}{r_fluxo_liquido_in_fc}:{col_tN_letter}{r_fluxo_liquido_in_fc})"

    r = 4
    rows = [
        ("VGV bruto", kpis["vgv_bruto"], None, "área privativa × R$/m²", ""),
        ("VGV líquido permuta", kpis["vgv_liquido"], None, "pós permuta 30%", ""),
        ("Custo obra (paramétrico)", kpis["custo_obra_total"], None, "CUSTOS_MACROGRUPO!D22", ""),
        ("Receita total (horizonte)", kpis["receita_total"], None, "soma do fluxo", ""),
        ("Despesa total (horizonte)", kpis["despesa_total"], None, "custo obra + contingência + comercial + RET + admin", ""),
        ("Lucro líquido", kpis["lucro"], None, "receita - despesa", "✓" if kpis["lucro"] > 0 else "✗"),
        ("Margem líquida", kpis["margem"], None, "lucro / receita", "✓" if kpis["margem"] > 0.18 else "⚠"),
        ("", None, None, "", ""),
        ("VPL @ TMA (Python)", kpis["vpl"], None, "numpy_financial.npv", ""),
        ("VPL @ TMA (Excel NPV)", "FORMULA", npv_formula, "deve bater com linha acima", ""),
        ("TIR mensal (Python)", kpis["tir_mensal"], None, "numpy_financial.irr", ""),
        ("TIR anual", kpis["tir_anual"], None, "(1+TIR_m)^12 - 1", "✓" if (kpis["tir_anual"] or 0) > prem["financeiras"]["tma_aa"] else "⚠"),
        ("TIR mensal (Excel IRR)", "FORMULA", irr_formula, "deve bater com linha 3 acima", ""),
        ("Payback descontado (meses da obra)", kpis["payback_meses"], None, "t em que fluxo descontado acumulado ≥ 0", ""),
        ("Exposição máxima", kpis["exposicao_max"], None, "min(fluxo acum.)", ""),
        ("ROI sobre capital", kpis["roi"], None, "lucro / exposição máxima", ""),
    ]

    format_map = {
        "VGV bruto": FMT_BRL, "VGV líquido permuta": FMT_BRL,
        "Custo obra (paramétrico)": FMT_BRL, "Receita total (horizonte)": FMT_BRL,
        "Despesa total (horizonte)": FMT_BRL, "Lucro líquido": FMT_BRL,
        "Margem líquida": FMT_PCT, "VPL @ TMA (Python)": FMT_BRL,
        "VPL @ TMA (Excel NPV)": FMT_BRL, "TIR mensal (Python)": FMT_PCT2,
        "TIR anual": FMT_PCT, "TIR mensal (Excel IRR)": FMT_PCT2,
        "Payback descontado (meses da obra)": FMT_NUM,
        "Exposição máxima": FMT_BRL, "ROI sobre capital": FMT_PCT,
    }

    for label, val, formula, obs, status in rows:
        if label == "":
            r += 1
            continue
        lc = ws.cell(row=r, column=1, value=label)
        style_label(lc)
        fmt = format_map.get(label, FMT_BRL)
        if val == "FORMULA":
            vc = ws.cell(row=r, column=2, value="ver col C")
            vc.font = FONT_SMALL
            vc.alignment = ALIGN_RIGHT
            vc.border = BORDER
            fc = ws.cell(row=r, column=3, value=formula)
            fc.number_format = fmt
            fc.font = FONT_BOLD
            fc.alignment = ALIGN_RIGHT
            fc.fill = FILL_BLUE
            fc.border = BORDER
        else:
            vc = ws.cell(row=r, column=2, value=val)
            style_calc(vc, fmt=fmt)
            fc = ws.cell(row=r, column=3, value="")
            fc.border = BORDER
        obs_c = ws.cell(row=r, column=4, value=obs)
        obs_c.font = FONT_SMALL
        obs_c.alignment = ALIGN_LEFT
        obs_c.border = BORDER
        st_c = ws.cell(row=r, column=5, value=status)
        st_c.font = FONT_BOLD
        st_c.alignment = ALIGN_CENTER
        st_c.border = BORDER
        if status == "✓":
            st_c.fill = FILL_CALC
        elif status == "⚠":
            st_c.fill = FILL_YELLOW
        elif status == "✗":
            st_c.fill = FILL_RED
        r += 1

    # Bloco validação semáforo
    r += 2
    ws.merge_cells(f"A{r}:E{r}")
    c = ws.cell(row=r, column=1, value="VALIDAÇÃO — semáforo")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    r += 1

    tma_aa = prem["financeiras"]["tma_aa"]
    checks = [
        ("Margem líquida > 18%", kpis["margem"], 0.18, FMT_PCT),
        (f"TIR anual > TMA+5% ({(tma_aa+0.05)*100:.0f}%)", kpis["tir_anual"] or 0, tma_aa + 0.05, FMT_PCT),
        ("VPL @ TMA > 0", kpis["vpl"], 0, FMT_BRL),
        ("VGV / Custo obra > 1,80", kpis["vgv_liquido"] / kpis["custo_obra_total"], 1.80, FMT_NUM2),
        ("Exposição máx / VGV < 25%", kpis["exposicao_max"] / kpis["vgv_liquido"] if kpis["vgv_liquido"] else 0, 0.25, FMT_PCT),
        ("P(VPL > 0) Monte Carlo > 85%", mc_stats["prob_vpl_positivo"], 0.85, FMT_PCT),
    ]
    for label, valor, threshold, fmt in checks:
        ws.cell(row=r, column=1, value=label).font = FONT_BODY
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=1).border = BORDER
        vc = ws.cell(row=r, column=2, value=valor)
        vc.number_format = fmt
        vc.font = FONT_BOLD
        vc.alignment = ALIGN_RIGHT
        vc.border = BORDER
        tc = ws.cell(row=r, column=3, value=threshold)
        tc.number_format = fmt
        tc.font = FONT_SMALL
        tc.alignment = ALIGN_RIGHT
        tc.border = BORDER
        # Lógica menor/maior: "< 25%" precisa inverter
        if "<" in label:
            ok = valor < threshold
        else:
            ok = valor > threshold
        sc = ws.cell(row=r, column=5, value="✓" if ok else "✗")
        sc.font = FONT_BOLD
        sc.alignment = ALIGN_CENTER
        sc.border = BORDER
        sc.fill = FILL_CALC if ok else FILL_RED
        r += 1


def build_cenarios(ws, cenarios, prem):
    set_col_widths(ws, [30, 20, 20, 20])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "CENÁRIOS — Pessimista / Base / Otimista"
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    headers = ["Indicador", "Pessimista", "Base", "Otimista"]
    r = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        style_header(c)
    r += 1

    ordem = ["pessimista", "base", "otimista"]
    fills = {"pessimista": FILL_RED, "base": FILL_BLUE, "otimista": FILL_CALC}

    # Descrição das perturbações
    ws.cell(row=r, column=1, value="Perturbação vs base").font = FONT_SMALL
    for col, nome in enumerate(ordem, 2):
        cfg = cenarios[nome]["_cfg"]
        desc = f"vgv×{cfg['vgv_fator']:.2f}  custo×{cfg['custo_fator']:.2f}  t50{cfg['t50_delta']:+d}"
        c = ws.cell(row=r, column=col, value=desc)
        c.font = FONT_SMALL
        c.alignment = ALIGN_CENTER
        c.fill = fills[nome]
        c.border = BORDER
    r += 2

    lines = [
        ("VGV bruto", "vgv_bruto", FMT_BRL),
        ("VGV líquido", "vgv_liquido", FMT_BRL),
        ("Custo obra", "custo_obra_total", FMT_BRL),
        ("Receita total", "receita_total", FMT_BRL),
        ("Despesa total", "despesa_total", FMT_BRL),
        ("Lucro líquido", "lucro", FMT_BRL),
        ("Margem líquida", "margem", FMT_PCT),
        ("VPL @ TMA", "vpl", FMT_BRL),
        ("TIR anual", "tir_anual", FMT_PCT),
        ("Exposição máxima", "exposicao_max", FMT_BRL),
        ("Payback (meses)", "payback_meses", FMT_NUM),
        ("ROI / capital", "roi", FMT_PCT),
    ]
    for label, key, fmt in lines:
        lc = ws.cell(row=r, column=1, value=label)
        style_label(lc)
        for col, nome in enumerate(ordem, 2):
            val = cenarios[nome].get(key)
            vc = ws.cell(row=r, column=col, value=val)
            vc.number_format = fmt
            vc.font = FONT_BOLD
            vc.alignment = ALIGN_RIGHT
            vc.border = BORDER
            vc.fill = fills[nome]
        r += 1


def build_tornado_sheet(ws, tornado, base_vpl, png_path):
    set_col_widths(ws, [24, 18, 18, 18, 18])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "TORNADO — Sensibilidade sobre VPL (±20% por variável)"
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    headers = ["Variável", "VPL (-20%)", "VPL base", "VPL (+20%)", "Δ VPL (high-low)"]
    r = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        style_header(c)
    r += 1
    for res in tornado:
        ws.cell(row=r, column=1, value=res["label"]).font = FONT_BOLD
        ws.cell(row=r, column=1).border = BORDER
        for col, val in enumerate([res["vpl_low"], base_vpl, res["vpl_high"], res["delta"]], 2):
            vc = ws.cell(row=r, column=col, value=val)
            vc.number_format = FMT_BRL
            vc.font = FONT_BODY
            vc.alignment = ALIGN_RIGHT
            vc.border = BORDER
        r += 1

    # Embed PNG
    if png_path.exists():
        img = XLImage(str(png_path))
        img.width = 560
        img.height = 360
        ws.add_image(img, f"A{r + 2}")


def build_monte_carlo_sheet(ws, mc_stats, hist_vpl_png, hist_tir_png, custo_base=None):
    set_col_widths(ws, [36, 22, 22, 22])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"MONTE CARLO — {mc_stats['n_iter']:,} iterações".replace(",", ".")
    c.fill = FILL_DARK
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER

    # Nota âncora: base do MC é o orçamento entregue
    nota = (
        "ÂNCORA: a simulação usa o ORÇAMENTO PARAMÉTRICO ENTREGUE como custo base "
        + (f"(R$ {custo_base/1e6:.2f} milhões — v00-final). " if custo_base else "(paramétrico v00-final). ")
        + "custo_obra_fator = 1,00 representa executar exatamente o orçamento entregue. "
        "As variações (±10%) simulam desvios realistas de execução (aditivos, concorrência, replanejamento)."
    )
    ws.merge_cells("A2:D2")
    cc = ws["A2"]
    cc.value = nota
    cc.fill = FILL_YELLOW
    cc.font = Font(name="Arial", size=9, italic=True)
    cc.alignment = ALIGN_WRAP_TOP
    cc.border = BORDER
    ws.row_dimensions[2].height = 40

    r = 4
    headers = ["Estatística", "VPL (R$)", "TIR a.a.", "Margem"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        style_header(c)
    r += 1

    rows = [
        ("Mediana", mc_stats["vpl_mediana"], mc_stats["tir_mediana_aa"], mc_stats["margem_mediana"]),
        ("Média", mc_stats["vpl_media"], None, None),
        ("P10", mc_stats["vpl_p10"], mc_stats["tir_p10_aa"], None),
        ("P50", mc_stats["vpl_p50"], mc_stats["tir_mediana_aa"], None),
        ("P90", mc_stats["vpl_p90"], mc_stats["tir_p90_aa"], None),
        ("Desvio padrão", mc_stats["vpl_desvio"], None, None),
    ]
    fmts = [None, FMT_BRL, FMT_PCT, FMT_PCT]
    for label, *vals in rows:
        lc = ws.cell(row=r, column=1, value=label)
        style_label(lc)
        for col, val in enumerate(vals, 2):
            if val is None:
                continue
            vc = ws.cell(row=r, column=col, value=val)
            vc.number_format = fmts[col - 1]
            vc.font = FONT_BOLD
            vc.alignment = ALIGN_RIGHT
            vc.border = BORDER
            vc.fill = FILL_CALC
        r += 1

    r += 1
    # Probabilidades
    probs = [
        ("P(VPL > 0)", mc_stats["prob_vpl_positivo"]),
        ("P(TIR > TMA)", mc_stats["prob_tir_gt_tma"]),
        ("P(Margem > 15%)", mc_stats["prob_margem_gt_15"]),
    ]
    for label, val in probs:
        lc = ws.cell(row=r, column=1, value=label)
        style_label(lc)
        vc = ws.cell(row=r, column=2, value=val)
        vc.number_format = FMT_PCT
        vc.font = FONT_BOLD
        vc.alignment = ALIGN_RIGHT
        vc.border = BORDER
        vc.fill = FILL_CALC if val > 0.7 else FILL_YELLOW
        r += 1

    # Embed histograms
    r += 2
    if hist_vpl_png.exists():
        img = XLImage(str(hist_vpl_png))
        img.width = 520
        img.height = 320
        ws.add_image(img, f"A{r}")
        r += 18
    if hist_tir_png.exists():
        img = XLImage(str(hist_tir_png))
        img.width = 520
        img.height = 320
        ws.add_image(img, f"A{r}")


# ============================================================================
# V2 — LAYOUT DIDÁTICO (para diretor Arthen)
# ============================================================================
FILL_GREEN_DARK = PatternFill("solid", fgColor="1B5E20")
FILL_YELLOW_DARK = PatternFill("solid", fgColor="E65100")
FILL_RED_DARK = PatternFill("solid", fgColor="B71C1C")
FONT_MEGA = Font(name="Arial", size=22, bold=True, color="FFFFFF")
FONT_LARGE = Font(name="Arial", size=16, bold=True, color="FFFFFF")
FONT_KPI = Font(name="Arial", size=20, bold=True, color=DARK)
FONT_NARRATIVE = Font(name="Arial", size=11)
FONT_NARRATIVE_BOLD = Font(name="Arial", size=11, bold=True, color=DARK)
ALIGN_WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _merge_narrative(ws, row_start, row_end, col_start, col_end, text, font=None, fill=None, align=None):
    """Escreve um bloco de texto narrativo merged."""
    start_letter = get_column_letter(col_start)
    end_letter = get_column_letter(col_end)
    ws.merge_cells(f"{start_letter}{row_start}:{end_letter}{row_end}")
    c = ws.cell(row=row_start, column=col_start, value=text)
    c.font = font or FONT_NARRATIVE
    c.alignment = align or ALIGN_WRAP_TOP
    if fill:
        c.fill = fill
    return c


def _write_kpi_box(ws, r, col_start, col_end, title, value, value_fmt, footer, fill_value=None):
    """Escreve um 'card' de KPI de 3 linhas: título, valor grande, rodapé."""
    start_letter = get_column_letter(col_start)
    end_letter = get_column_letter(col_end)
    # título
    ws.merge_cells(f"{start_letter}{r}:{end_letter}{r}")
    c = ws.cell(row=r, column=col_start, value=title)
    c.fill = FILL_DARK
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    # valor
    ws.merge_cells(f"{start_letter}{r+1}:{end_letter}{r+1}")
    c = ws.cell(row=r + 1, column=col_start, value=value)
    c.number_format = value_fmt
    c.fill = fill_value or FILL_CALC
    c.font = FONT_KPI
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    # rodapé
    ws.merge_cells(f"{start_letter}{r+2}:{end_letter}{r+2}")
    c = ws.cell(row=r + 2, column=col_start, value=footer)
    c.font = FONT_SMALL
    c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    c.border = BORDER
    ws.row_dimensions[r].height = 22
    ws.row_dimensions[r + 1].height = 38
    ws.row_dimensions[r + 2].height = 32


def _decide_veredicto(kpis, mc_stats, cenarios):
    """Retorna (label, cor_fill, subtexto) baseado em thresholds."""
    margem = kpis["margem"]
    prob = mc_stats["prob_vpl_positivo"]
    vpl = kpis["vpl"]
    pess_vpl = cenarios["pessimista"]["vpl"]

    if vpl > 0 and margem > 0.18 and prob > 0.85:
        if pess_vpl > 0:
            return (
                "PROJETO VIÁVEL",
                FILL_GREEN_DARK,
                "Indicadores positivos em todos os cenários analisados, inclusive no pessimista.",
            )
        else:
            return (
                "PROJETO VIÁVEL  (com pontos de atenção)",
                FILL_GREEN_DARK,
                "Viável no cenário base e provável (98% Monte Carlo), mas sensível a combinações adversas simultâneas — ver análise de cenários.",
            )
    if vpl > 0 and prob > 0.60:
        return (
            "VIABILIDADE MARGINAL",
            FILL_YELLOW_DARK,
            "VPL positivo no cenário base, mas com alto risco em cenários realistas. Recomendamos revisão das premissas antes do avanço.",
        )
    return (
        "PROJETO INVIÁVEL",
        FILL_RED_DARK,
        "Indicadores não justificam investimento ao custo de capital assumido. Recomenda-se revisão do briefing ou descarte.",
    )


def build_parecer(ws, prem, kpis, mc_stats, cenarios, tornado):
    """PARECER — capa didática com veredicto, KPIs headline, narrativa e recomendação."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 18
    ws.column_dimensions["H"].width = 3

    # Header
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "ESTUDO DE VIABILIDADE FINANCEIRA"
    c.fill = FILL_DARK
    c.font = FONT_LARGE
    c.alignment = ALIGN_CENTER

    ws.row_dimensions[3].height = 24
    ws.merge_cells("B3:G3")
    c = ws["B3"]
    proj = prem["projeto"]
    c.value = f"{proj['nome']} — {proj['bairro']}, {proj['cidade']}/{proj['estado']}"
    c.fill = FILL_BLUE
    c.font = Font(name="Arial", size=13, bold=True, color=DARK)
    c.alignment = ALIGN_CENTER

    ws.merge_cells("B4:G4")
    c = ws["B4"]
    c.value = f"Cliente: {proj['cliente']}    |    Elaborado por: Cartesian Engenharia    |    {prem.get('_gerado_em', '2026-04-14')}"
    c.font = FONT_SMALL
    c.alignment = ALIGN_CENTER

    # Veredicto
    veredicto, fill, subtexto = _decide_veredicto(kpis, mc_stats, cenarios)

    ws.row_dimensions[6].height = 54
    ws.merge_cells("B6:G6")
    c = ws["B6"]
    c.value = veredicto
    c.fill = fill
    c.font = FONT_MEGA
    c.alignment = ALIGN_CENTER

    ws.row_dimensions[7].height = 32
    ws.merge_cells("B7:G7")
    c = ws["B7"]
    c.value = subtexto
    c.font = Font(name="Arial", size=11, italic=True, color=DARK)
    c.alignment = ALIGN_WRAP_CENTER

    # 4 KPIs headline (2x2)
    r = 9
    _write_kpi_box(
        ws, r, 2, 4,
        "CAPITAL MÁXIMO EXPOSTO",
        kpis["exposicao_max"], FMT_BRL,
        "Quanto a Arthen precisa ter em caixa (ou via SFH) no mês mais apertado, antes das vendas cobrirem os custos",
        fill_value=FILL_BLUE,
    )
    _write_kpi_box(
        ws, r, 5, 7,
        "LUCRO ESPERADO (BASE)",
        kpis["lucro"], FMT_BRL,
        f"Receita total menos todas as despesas ao longo dos {prem['projeto']['prazo_obra_meses'] + prem['projeto']['prazo_pos_obra_meses'] + prem['projeto']['prazo_pre_lancamento_meses']} meses do projeto",
        fill_value=FILL_CALC,
    )

    r2 = r + 4
    _write_kpi_box(
        ws, r2, 2, 4,
        "MARGEM LÍQUIDA",
        kpis["margem"], FMT_PCT,
        "Lucro ÷ Receita. Referência de mercado SC: 15–25%. Acima disso = margem acima da média",
        fill_value=FILL_CALC,
    )
    _write_kpi_box(
        ws, r2, 5, 7,
        "CHANCE DE VPL POSITIVO",
        mc_stats["prob_vpl_positivo"], FMT_PCT,
        f"Monte Carlo: em {mc_stats['n_iter']:,} versões com variações realistas de preço/custo/velocidade, quantas deram lucro acima do custo de capital".replace(",", "."),
        fill_value=FILL_CALC,
    )

    # VPL / TIR em bloco extra
    r3 = r2 + 4
    _write_kpi_box(
        ws, r3, 2, 4,
        "VPL @ TMA 14%",
        kpis["vpl"], FMT_BRL,
        "Valor criado acima do mínimo exigido pelo custo de capital. Se > 0, o projeto remunera mais que o benchmark",
        fill_value=FILL_CALC if kpis["vpl"] > 0 else FILL_RED,
    )
    tir_display = kpis["tir_anual"] if kpis["tir_anual"] is not None else 0.0
    _write_kpi_box(
        ws, r3, 5, 7,
        "TIR ANUAL",
        tir_display, FMT_PCT,
        "⚠ Inflada pela estrutura (permuta física + pré-lançamento). Prefira ler o VPL como indicador principal de viabilidade",
        fill_value=FILL_YELLOW,
    )

    # ENTENDA O PROJETO
    r = r3 + 5
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2, value="O QUE É O PROJETO")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    horizonte = proj['prazo_pre_lancamento_meses'] + proj['prazo_obra_meses'] + proj['prazo_pos_obra_meses']
    narr1 = (
        f"O Arboris é um empreendimento residencial + comercial da Arthen com {proj['n_unidades']} unidades "
        f"(90 residenciais de 3-4 dormitórios + 8 comerciais), distribuídas em "
        f"{proj['area_construida_m2']:,.0f} m² de área construída, localizado no bairro {proj['bairro']} em {proj['cidade']}/{proj['estado']} — "
        f"região que liderou velocidade de vendas imobiliárias no Brasil em 2025.\n\n"
        f"O custo de obra é de R$ {kpis['custo_obra_total']/1e6:.1f} milhões "
        f"(R$ {kpis['custo_obra_total']/proj['area_construida_m2']:,.0f}/m²), calculado via paramétrico V2 Híbrido da Cartesian — "
        f"metodologia calibrada contra base histórica de 126 projetos similares. "
        f"O terreno será adquirido por permuta física de {int(proj['permuta_pct_unidades']*100)}% das unidades, eliminando desembolso monetário com o landowner.\n\n"
        f"O modelo financeiro cobre {horizonte} meses: {proj['prazo_pre_lancamento_meses']} de pré-lançamento, "
        f"{proj['prazo_obra_meses']} de obra, e {proj['prazo_pos_obra_meses']} de cauda de recebimento. "
        f"O preço médio de venda estimado é R$ {prem['receita']['vgv_m2_base']:,.0f}/m² privativo "
        f"(mediana Perequê abr/2026, range R$ 10k–25k). Estrutura padrão de mercado: "
        f"{int(prem['receita']['entrada_pct']*100)}% entrada + {int(prem['receita']['obra_pct']*100)}% parcelado durante obra "
        f"corrigido pelo CUB/SC + {int(prem['receita']['repasse_pct']*100)}% repasse bancário na entrega."
    )
    ws.merge_cells(f"B{r}:G{r+7}")
    c = ws.cell(row=r, column=2, value=narr1)
    c.font = FONT_NARRATIVE
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_GRAY
    for rr in range(r, r + 8):
        ws.row_dimensions[rr].height = 18
    r += 8

    # PRINCIPAIS DESCOBERTAS
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2, value="PRINCIPAIS DESCOBERTAS")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    pess_vpl = cenarios["pessimista"]["vpl"]
    top_tornado = tornado[0] if tornado else None
    segundo_tornado = tornado[1] if len(tornado) > 1 else None

    descobertas_positivas = [
        f"✓  VPL positivo de R$ {kpis['vpl']/1e6:.1f} milhões descontado a 14% a.a. — o projeto gera valor real acima do custo de capital da incorporadora.",
        f"✓  Margem líquida de {kpis['margem']*100:.1f}% — acima do benchmark de mercado SC (15–25%), indicando eficiência operacional do projeto.",
        f"✓  Simulação Monte Carlo de {mc_stats['n_iter']:,} cenários mostra {mc_stats['prob_vpl_positivo']*100:.1f}% de probabilidade de VPL positivo, com mediana de R$ {mc_stats['vpl_mediana']/1e6:.1f}M e P10 (pior 10%) de R$ {mc_stats['vpl_p10']/1e6:.1f}M — projeto robusto a variações realistas.".replace(",", "."),
    ]
    atencao = [
        f"⚠  Exposição máxima de caixa: R$ {kpis['exposicao_max']/1e6:.1f} milhões. A Arthen precisa ter esse capital disponível (próprio ou via SFH) nos meses mais apertados do cronograma.",
        (f"⚠  No cenário pessimista (preço -15%, custo +10%, vendas 8 meses mais lentas), o VPL vira NEGATIVO em R$ {abs(pess_vpl)/1e3:,.0f} mil. ".replace(",", ".") +
         "O projeto é viável mas sensível a combinações adversas simultâneas — não é invulnerável."),
        (f"⚠  Fator de maior risco: {top_tornado['label'] if top_tornado else 'preço de venda'}. "
         f"Variação de ±20% move o VPL em R$ {(top_tornado['abs_delta']/1e6 if top_tornado else 22):.0f}M. "
         f"Validar o preço de R$ {prem['receita']['vgv_m2_base']:,.0f}/m² com corretores locais é a diligência mais importante."),
    ]

    for texto in descobertas_positivas:
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2, value=texto)
        c.font = FONT_NARRATIVE
        c.alignment = ALIGN_WRAP_TOP
        c.fill = FILL_CALC
        c.border = BORDER
        ws.row_dimensions[r].height = 32
        r += 1
    for texto in atencao:
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(row=r, column=2, value=texto)
        c.font = FONT_NARRATIVE
        c.alignment = ALIGN_WRAP_TOP
        c.fill = FILL_YELLOW
        c.border = BORDER
        ws.row_dimensions[r].height = 38
        r += 1

    # RECOMENDAÇÃO
    r += 1
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2, value="RECOMENDAÇÃO")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    recomendacao = (
        f"Com base nos indicadores apurados, o Arboris é um projeto VIÁVEL e RECOMENDADO para avanço. "
        f"O VPL de R$ {kpis['vpl']/1e6:.1f}M e a margem líquida de {kpis['margem']*100:.1f}% posicionam o projeto "
        f"acima da média do segmento em SC. A probabilidade de {mc_stats['prob_vpl_positivo']*100:.0f}% de retorno positivo "
        f"(Monte Carlo 10.000 iterações) confere robustez estatística ao investimento.\n\n"
        f"Recomendamos três diligências antes do lançamento:\n"
        f"(1) Validar o preço de venda de R$ {prem['receita']['vgv_m2_base']:,.0f}/m² com corretores locais de Perequê "
        f"e comparáveis recém-lançados — é o fator de maior impacto sobre o resultado;\n"
        f"(2) Confirmar capacidade financeira da Arthen para suportar a exposição máxima de R$ {kpis['exposicao_max']/1e6:.1f}M "
        f"(próprio ou via financiamento à produção SFH);\n"
        f"(3) Revisar a velocidade de vendas assumida (50% vendido no mês 14 da obra) contra o histórico da Arthen em "
        f"projetos similares — Perequê é mercado aquecido, mas a curva real precisa calibrar contra dados da incorporadora."
    )
    ws.merge_cells(f"B{r}:G{r+9}")
    c = ws.cell(row=r, column=2, value=recomendacao)
    c.font = FONT_NARRATIVE
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_BLUE
    c.border = BORDER
    for rr in range(r, r + 10):
        ws.row_dimensions[rr].height = 18
    r += 10

    # Rodapé com navegação
    r += 2
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(row=r, column=2, value="COMO NAVEGAR ESTE ESTUDO:")
    c.font = FONT_NARRATIVE_BOLD
    c.alignment = ALIGN_LEFT
    r += 1
    navs = [
        ("COMO_LER", "Glossário com explicação em linguagem simples de todos os termos técnicos (VPL, TIR, Monte Carlo, RET, etc.)"),
        ("PREMISSAS", "Todas as premissas do modelo — se algum valor estiver incorreto, editá-lo aqui e regerar"),
        ("DETALHE_CUSTO", "Detalhamento do custo de obra: metodologia paramétrica, 18 macrogrupos, janela temporal, Gantt, curva S, stacked"),
        ("FLUXO_CAIXA", f"Projeção mensal detalhada de receitas e despesas ao longo dos {horizonte} meses do projeto"),
        ("RESULTADO", "Cálculo detalhado do VPL e TIR, com fórmula Excel paralela para auditoria"),
        ("CENARIOS", "Comparativo pessimista / base / otimista dos principais indicadores"),
        ("TORNADO", "Análise de sensibilidade — qual variável mais afeta o resultado"),
        ("MONTE_CARLO", "Simulação estatística com 10.000 iterações e histogramas de distribuição"),
    ]
    for nome, desc in navs:
        ws.cell(row=r, column=2, value=f"▸ {nome}").font = FONT_BOLD
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.merge_cells(f"C{r}:G{r}")
        ws.cell(row=r, column=3, value=desc).font = FONT_SMALL
        ws.cell(row=r, column=3).alignment = ALIGN_LEFT
        r += 1


def build_como_ler(ws, prem=None, kpis=None, mc_stats=None, cenarios=None, tornado=None):
    """COMO_LER — glossário didático com explicação em linguagem simples.

    Quando prem/kpis/mc_stats/cenarios/tornado são passados, os exemplos finais
    dos verbetes ('→ Neste estudo:') são regenerados com os valores atuais do
    modelo — evita o drift entre modelo e glossário.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 100

    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "COMO LER ESTE ESTUDO"
    c.fill = FILL_DARK
    c.font = FONT_LARGE
    c.alignment = ALIGN_CENTER

    ws.row_dimensions[3].height = 24
    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = "Glossário de termos técnicos — explicação em linguagem simples"
    c.font = Font(name="Arial", size=11, italic=True, color=DARK)
    c.alignment = ALIGN_CENTER

    # Valores dinâmicos do modelo atual (fallback pra exemplos estáticos se não passados)
    def _fmt_m(v): return f"R$ {v/1e6:.1f}M"
    def _fmt_pct(v): return f"{v*100:.1f}%"

    if kpis and mc_stats and prem and cenarios and tornado:
        vpl_m = kpis["vpl"]
        tir_aa = kpis["tir_anual"] or 0
        margem = kpis["margem"]
        vgv_bruto = kpis["vgv_bruto"]
        vgv_liq = kpis["vgv_liquido"]
        exp_max = kpis["exposicao_max"]
        custo_total = kpis["custo_obra_total"]
        ac = prem["projeto"]["area_construida_m2"]
        vgv_m2 = prem["receita"]["vgv_m2_base"]
        tma_aa = prem["financeiras"]["tma_aa"]
        cub_aa = prem["financeiras"]["cub_aa"]
        permuta_pct = prem["projeto"]["permuta_pct_unidades"]
        n_iter = mc_stats["n_iter"]
        prob_vpl = mc_stats["prob_vpl_positivo"]
        vpl_mediana = mc_stats["vpl_mediana"]
        vpl_p10 = mc_stats["vpl_p10"]
        vpl_p90 = mc_stats["vpl_p90"]
        tir_p10 = mc_stats["tir_p10_aa"] or 0
        pess_vpl = cenarios["pessimista"]["vpl"]
        top = tornado[0] if tornado else None
        second = tornado[1] if len(tornado) > 1 else None
        third = tornado[2] if len(tornado) > 2 else None
        top_label = top["label"] if top else "—"
        top_delta = top["abs_delta"] if top else 0
        second_label = second["label"] if second else "—"
        second_delta = second["abs_delta"] if second else 0
        third_label = third["label"] if third else "—"
        third_delta = third["abs_delta"] if third else 0
        mc_custo_tri = prem["monte_carlo"]["custo_obra_fator_tri"]
        mc_vgv_tri = prem["monte_carlo"]["vgv_m2_tri"]
        mc_t50_tri = prem["monte_carlo"]["t50_vendas_tri"]
        mc_cub_tri = prem["monte_carlo"]["cub_aa_tri"]
        mc_corr = prem["monte_carlo"]["corr_vgv_t50"]
    else:
        # Fallback valores exemplares
        vpl_m = 16200000; tir_aa = 0.686; margem = 0.308
        vgv_bruto = 134400000; vgv_liq = 94100000; exp_max = 15400000
        custo_total = 39400000; ac = 12473; vgv_m2 = 15394
        tma_aa = 0.14; cub_aa = 0.07; permuta_pct = 0.30
        n_iter = 10000; prob_vpl = 0.98; vpl_mediana = 16900000
        vpl_p10 = 5600000; vpl_p90 = 30000000; tir_p10 = 0.25
        pess_vpl = -608000
        top_label = "VGV R$/m²"; top_delta = 22000000
        second_label = "Custo obra"; second_delta = 15500000
        third_label = "Entrada %"; third_delta = 5700000
        mc_custo_tri = [0.95, 1.00, 1.12]
        mc_vgv_tri = [11000, 15394, 22000]
        mc_t50_tri = [10, 17, 28]
        mc_cub_tri = [0.04, 0.07, 0.11]
        mc_corr = -0.4

    glossario = [
        ("VPL — Valor Presente Líquido",
         f"Soma de TODOS os fluxos de caixa futuros do projeto, trazidos a valor presente pela taxa mínima de atratividade (TMA). "
         f"É o indicador mais importante deste estudo.\n\n"
         f"Interpretação simples: se o VPL > 0, o projeto gera MAIS valor do que aplicar o mesmo dinheiro no benchmark de mercado "
         f"(Selic + spread). Se VPL < 0, o projeto destrói valor — seria melhor investir em outra coisa.\n\n"
         f"→ Neste estudo: VPL = {_fmt_m(vpl_m)}. Indica que o Arboris produz {_fmt_m(vpl_m)} a MAIS do que o mínimo exigido "
         f"pelo custo de capital de {_fmt_pct(tma_aa)} a.a."),

        ("TIR — Taxa Interna de Retorno",
         f"Taxa de desconto que faz o VPL ser zero. Interpretada como 'rentabilidade anual do projeto'. Para ser atrativa, "
         f"precisa estar acima da TMA.\n\n"
         f"⚠ Atenção nesta análise: a TIR de {_fmt_pct(tir_aa)} a.a. é tecnicamente correta mas está INFLADA pela estrutura do modelo "
         f"(permuta física do terreno + pré-lançamento com entradas cobrindo custos antes da obra iniciar, resultando em "
         f"exposição de caixa baixa). O VPL é o indicador mais confiável neste caso.\n\n"
         f"→ Neste estudo: TIR = {_fmt_pct(tir_aa)} a.a. no cenário base; P10 do Monte Carlo ~{_fmt_pct(tir_p10)} a.a."),

        ("TMA — Taxa Mínima de Atratividade",
         f"Custo de capital que a incorporadora exige para justificar o investimento. É o 'piso' de rentabilidade abaixo do qual "
         f"o projeto não compensa. Também chamada de 'taxa de desconto'.\n\n"
         f"Geralmente composta por: Selic (taxa básica) + spread de risco do setor imobiliário.\n\n"
         f"→ Neste estudo: {_fmt_pct(tma_aa)} a.a. = Selic ~10,5% + spread 3,5% (benchmark incorporadora média SC 2026). Editável em PREMISSAS."),

        ("VGV — Valor Geral de Vendas",
         f"Receita total esperada do empreendimento. Calculado como: área privativa vendável × preço médio de venda por m².\n\n"
         f"Dividido em dois conceitos neste estudo:\n"
         f"  • VGV BRUTO: valor total se todas as unidades fossem vendidas pela Arthen\n"
         f"  • VGV LÍQUIDO: após subtrair a permuta física ({int(permuta_pct*100)}% das unidades vão para o proprietário do terreno)\n\n"
         f"→ Neste estudo: VGV bruto {_fmt_m(vgv_bruto)}; VGV líquido (que a Arthen de fato vende) {_fmt_m(vgv_liq)}."),

        ("Margem líquida",
         f"(Lucro líquido) ÷ (Receita total). Mede a eficiência operacional da incorporação. Inclui TODOS os custos: "
         f"obra, comercial, marketing, impostos (RET), admin.\n\n"
         f"Referência de mercado SC:\n"
         f"  • Abaixo de 10% → margem apertada, projeto frágil\n"
         f"  • 10–15% → aceitável\n"
         f"  • 15–25% → saudável (média do segmento)\n"
         f"  • Acima de 25% → acima da média, sinal de eficiência ou precificação forte\n\n"
         f"→ Neste estudo: {_fmt_pct(margem)} — acima da média do segmento, puxada por preço forte em Perequê e custo calibrado no paramétrico."),

        ("Exposição máxima de caixa",
         f"O 'pior momento' do projeto em termos de caixa: quanto a incorporadora tem que ter disponível (capital próprio ou "
         f"financiamento) antes que as vendas comecem a cobrir os custos. É o BURACO máximo no fluxo acumulado.\n\n"
         f"Em projetos bem estruturados, corresponde a 20–35% do VGV. Abaixo disso = projeto leve de capital; acima = projeto pesado.\n\n"
         f"→ Neste estudo: {_fmt_m(exp_max)} = {exp_max/vgv_liq*100:.0f}% do VGV líquido. "
         f"{'Relativamente baixo, porque as entradas começam a pagar a obra rapidamente.' if exp_max/vgv_liq < 0.25 else 'Na faixa típica do setor.'}"),

        ("Monte Carlo",
         f"Técnica estatística para lidar com incerteza. Em vez de assumir que cada premissa é um número fixo, tratamos as mais "
         f"importantes como FAIXAS de valores prováveis e rodamos milhares de simulações do projeto — em cada uma, o computador "
         f"sorteia valores dentro das faixas e calcula o VPL.\n\n"
         f"Ao final, temos uma DISTRIBUIÇÃO de resultados possíveis em vez de um número único. Isso permite perguntar: "
         f"'qual a PROBABILIDADE do projeto dar lucro?'\n\n"
         f"▸  ÂNCORA NO ORÇAMENTO ENTREGUE: a base do MC é o orçamento paramétrico v00-final (R$ {custo_total/1e6:.1f}M). "
         f"custo_obra_fator=1.00 representa executar exatamente esse orçamento. As variações simulam desvios realistas de execução.\n\n"
         f"Variáveis sorteadas neste estudo: preço de venda (R$ {mc_vgv_tri[0]/1000:.0f}k a R$ {mc_vgv_tri[2]/1000:.0f}k/m²), "
         f"custo de obra ({(mc_custo_tri[0]-1)*100:+.0f}% a {(mc_custo_tri[2]-1)*100:+.0f}% sobre o orçamento entregue), "
         f"velocidade de vendas ({mc_t50_tri[0]} a {mc_t50_tri[2]} meses para t50), "
         f"CUB/SC ({_fmt_pct(mc_cub_tri[0])} a {_fmt_pct(mc_cub_tri[2])} a.a.). "
         f"Correlação {int(mc_corr*100)}% entre preço e velocidade (preço alto ↔ venda mais lenta, que é o comportamento real de mercado).\n\n"
         f"→ Neste estudo: {_fmt_pct(prob_vpl)} das {n_iter:,} simulações deram VPL positivo. "
         f"Mediana {_fmt_m(vpl_mediana)}. Pior 10% = {_fmt_m(vpl_p10)}. Melhor 10% = {_fmt_m(vpl_p90)}.".replace(",", ".")),

        ("Cenários (Pessimista / Base / Otimista)",
         f"Diferente do Monte Carlo (que sorteia aleatoriamente), aqui definimos 3 combinações específicas de premissas:\n"
         f"  • PESSIMISTA: preço -15%, custo +10%, venda 8 meses mais lenta\n"
         f"  • BASE: valores centrais, o 'cenário provável'\n"
         f"  • OTIMISTA: preço +10%, custo -5%, venda 4 meses mais rápida\n\n"
         f"Útil para responder: 'e se tudo der errado ao mesmo tempo?' (pessimista) e 'e se as coisas favorecerem?' (otimista).\n\n"
         f"→ Neste estudo: pessimista dá VPL {_fmt_m(pess_vpl) if pess_vpl < 0 else _fmt_m(pess_vpl)} "
         f"({'projeto vira inviável nesta combinação.' if pess_vpl < 0 else 'projeto mantém VPL positivo mesmo nesta combinação.'})"),

        ("Tornado (Análise de Sensibilidade)",
         f"Gráfico que mostra qual variável tem MAIOR impacto sobre o VPL. Para cada variável, testamos +20% e -20% mantendo "
         f"as outras fixas e medimos a variação resultante no VPL. A variável com a maior barra é a mais crítica — é ali que "
         f"a diligência comercial deve focar.\n\n"
         f"→ Neste estudo (em ordem de impacto):\n"
         f"    1º {top_label} — Δ {_fmt_m(top_delta)} (dominante)\n"
         f"    2º {second_label} — Δ {_fmt_m(second_delta)}\n"
         f"    3º {third_label} — Δ {_fmt_m(third_delta)}\n\n"
         f"Conclusão: validar o preço de venda de R$ {vgv_m2:,.0f}/m² é a diligência mais importante deste projeto."),

        ("CUB/SC",
         f"Custo Unitário Básico de Construção, publicado mensalmente pelo Sinduscon/SC. É o indexador padrão das parcelas "
         f"da obra pagas pelos compradores durante a construção — o equivalente imobiliário da 'correção monetária'.\n\n"
         f"Substitui o INCC (usado em outros estados) no padrão catarinense.\n\n"
         f"→ Neste estudo: reajuste médio anual de {_fmt_pct(cub_aa)} a.a. (histórico 2020–2025 Sinduscon/SC). Editável em PREMISSAS."),

        ("RET — Regime Especial de Tributação",
         "Alíquota unificada de 4% sobre a receita recebida, aplicável a incorporações que aderiram ao Patrimônio de Afetação "
         "(Lei 10.931/04, art. 4º). Substitui IR, CSLL, PIS e COFINS da SPE da incorporadora.\n\n"
         "É o regime tributário mais eficiente para incorporação imobiliária no Brasil — normalmente vale a pena aderir.\n\n"
         "→ Neste estudo: 4% sobre receita recebida a cada mês."),

        ("Permuta física",
         f"Forma de adquirir o terreno: em vez de pagar em dinheiro, a incorporadora entrega ao proprietário um percentual das "
         f"unidades prontas do futuro empreendimento. É o modelo dominante em terrenos valorizados no litoral SC.\n\n"
         f"Efeito financeiro: elimina o DESEMBOLSO inicial com terreno (grande alívio de caixa no mês 0), mas reduz o VGV "
         f"vendável pela incorporadora (porque parte das unidades não é dela).\n\n"
         f"→ Neste estudo: {int(permuta_pct*100)}% das unidades vão para o landowner. VGV vendável pela Arthen cai de "
         f"{_fmt_m(vgv_bruto)} (bruto) para {_fmt_m(vgv_liq)} (líquido)."),

        ("Paramétrico V2 Híbrido",
         f"Metodologia Cartesian para orçamento preliminar: extrai quantitativos do BIM (modelo 3D do projeto), aplica Preços "
         f"Unitários calibrados contra uma base histórica de 126 projetos similares já executados, e produz o custo total por "
         f"macrogrupo com faixa de confiança. Precisão típica: ±10–15%.\n\n"
         f"Diferente de um orçamento executivo completo (SINAPI linha a linha), mas muito mais rápido e suficiente para decisões "
         f"de viabilidade.\n\n"
         f"→ Neste estudo: custo total {_fmt_m(custo_total)} (R$ {custo_total/ac:,.0f}/m²). "
         f"Benchmark do segmento 'Médio' (8k–15k m² AC): mediana R$ 3.337/m², faixa P25–P75 R$ 2.974–R$ 3.710. "
         f"Arboris em R$ {custo_total/ac:,.0f}/m² — coerente com padrão Médio-Alto."),
    ]

    r = 5
    for termo, explicacao in glossario:
        # Termo em B
        c = ws.cell(row=r, column=2, value=termo)
        c.fill = FILL_ACCENT
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = BORDER
        # Explicação em C
        c = ws.cell(row=r, column=3, value=explicacao)
        c.font = FONT_NARRATIVE
        c.alignment = ALIGN_WRAP_TOP
        c.fill = FILL_GRAY
        c.border = BORDER
        # altura baseada no texto (linhas explícitas + wrap estimado)
        n_br = explicacao.count("\n")
        n_wrap = len(explicacao) // 90
        ws.row_dimensions[r].height = max(60, (n_br + n_wrap + 1) * 15)
        r += 1


# ============================================================================
# PLOTS extras — Detalhamento do custo de obra
# ============================================================================
def plot_fluxo_acumulado(fluxo, out_path):
    """Saldo acumulado do projeto (linha) com marcadores de pico negativo e
    break-even. Usado na aba FLUXO_CAIXA pra leitura rápida do diretor."""
    PLOTS_DIR.mkdir(exist_ok=True)
    t_axis = fluxo["t_axis"]
    fluxo_acum = fluxo["fluxo_acum"]
    prazo_obra = fluxo["prazo_obra"]
    prazo_pre = fluxo["prazo_pre"]

    fig, ax = plt.subplots(figsize=(11, 5.0), dpi=100)
    colors_fill = np.where(fluxo_acum >= 0, "#66BB6A", "#E57373")
    ax.fill_between(t_axis, 0, fluxo_acum / 1e6, where=(fluxo_acum >= 0),
                    color="#C8E6C9", alpha=0.6, label="Saldo positivo")
    ax.fill_between(t_axis, 0, fluxo_acum / 1e6, where=(fluxo_acum < 0),
                    color="#FFCDD2", alpha=0.6, label="Saldo negativo")
    ax.plot(t_axis, fluxo_acum / 1e6, color="#2C3E50", linewidth=2.2)
    ax.axhline(0, color="#555", linewidth=0.8)

    # Marcos
    idx_pico = int(np.argmin(fluxo_acum))
    mes_pico = int(t_axis[idx_pico])
    v_pico = fluxo_acum[idx_pico] / 1e6
    ax.annotate(
        f"EXPOSIÇÃO MÁX\nR$ {-v_pico:.1f}M\n(mês {mes_pico})",
        xy=(mes_pico, v_pico), xytext=(mes_pico + 2, v_pico - 2),
        fontsize=9, color="#B71C1C", fontweight="bold",
        arrowprops=dict(arrowstyle="->", color="#B71C1C", lw=1.2),
    )
    # Break-even após pico
    depois = np.arange(len(fluxo_acum)) > idx_pico
    pos = (fluxo_acum >= 0) & depois
    if pos.any():
        idx_be = int(np.argmax(pos))
        mes_be = int(t_axis[idx_be])
        ax.axvline(mes_be, color="#2E7D32", linestyle="--", linewidth=1.2, alpha=0.7)
        ax.text(mes_be, ax.get_ylim()[1] * 0.9, f"break-even\nmês {mes_be}",
                color="#2E7D32", fontsize=9, fontweight="bold", ha="left",
                bbox=dict(boxstyle="round,pad=0.3", facecolor="white", edgecolor="#2E7D32"))

    # Linhas de fase
    ax.axvline(0, color="#888", linestyle=":", linewidth=1.0, alpha=0.6)
    ax.axvline(prazo_obra, color="#888", linestyle=":", linewidth=1.0, alpha=0.6)
    ymax = ax.get_ylim()[1]
    ax.text(-prazo_pre / 2, ymax * 0.92, "PRÉ", fontsize=8, ha="center",
            color="#555", fontweight="bold")
    ax.text(prazo_obra / 2, ymax * 0.92, "OBRA", fontsize=9, ha="center",
            color="#555", fontweight="bold")
    ax.text(prazo_obra + 6, ymax * 0.92, "PÓS-OBRA", fontsize=8, ha="center",
            color="#555", fontweight="bold")

    ax.set_xlabel("Mês do projeto (0 = início da obra)", fontsize=10)
    ax.set_ylabel("Saldo acumulado (R$ milhões)", fontsize=10)
    ax.set_title(
        "Saldo de Caixa Acumulado — o 'caixa do projeto' ao longo do tempo",
        fontsize=12, color="#2C3E50", fontweight="bold",
    )
    ax.grid(alpha=0.3)
    ax.legend(loc="lower right", fontsize=9)
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


def plot_curva_desembolso_s(fluxo, out_path):
    """Curva S: desembolso acumulado da obra em % do custo total."""
    from matplotlib.ticker import PercentFormatter
    PLOTS_DIR.mkdir(exist_ok=True)

    t_axis = fluxo["t_axis"]
    custo_mensal = fluxo["custo_obra_mensal"]
    total = custo_mensal.sum()
    if total == 0:
        return
    cum_pct = np.cumsum(custo_mensal) / total

    prazo_pre = fluxo["prazo_pre"]
    prazo_obra = fluxo["prazo_obra"]
    mes_obra = np.arange(-prazo_pre, len(t_axis) - prazo_pre)

    fig, ax = plt.subplots(figsize=(9, 4.5), dpi=100)
    ax.fill_between(mes_obra, 0, cum_pct, alpha=0.25, color="#2980B9")
    ax.plot(mes_obra, cum_pct, color="#2C3E50", linewidth=2.5)

    for target in [0.25, 0.50, 0.75]:
        hits = np.where(cum_pct >= target)[0]
        if len(hits) > 0:
            idx = hits[0]
            mes_target = mes_obra[idx]
            ax.axhline(target, color="#888", linestyle=":", linewidth=0.8)
            ax.axvline(mes_target, color="#C25700", linestyle=":", linewidth=0.8)
            ax.annotate(
                f"{int(target*100)}% no mês {mes_target}",
                xy=(mes_target, target),
                xytext=(mes_target + 1, target - 0.06),
                fontsize=9, color="#C25700", fontweight="bold",
            )

    ax.set_xlabel("Mês relativo à obra  (0 = início da obra)", fontsize=10)
    ax.set_ylabel("% do custo de obra acumulado", fontsize=10)
    ax.set_title(
        f"Curva S — Desembolso Acumulado da Obra  (total R$ {total/1e6:.2f}M)",
        fontsize=12, color="#2C3E50", fontweight="bold",
    )
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1.0))
    ax.grid(alpha=0.3)
    ax.set_xlim(-prazo_pre, prazo_obra + 3)
    ax.set_ylim(0, 1.05)
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


def plot_gantt_macrogrupos(breakdown, curva_config, out_path, prazo_obra=36):
    """Gantt: janela temporal de cada macrogrupo, barra proporcional ao custo."""
    PLOTS_DIR.mkdir(exist_ok=True)

    items = []
    for nome, janela in curva_config.items():
        if nome.startswith("_"):
            continue
        if nome not in breakdown:
            continue
        t_ini, t_fim = janela
        valor = breakdown[nome]["valor"]
        items.append((nome, t_ini, t_fim, valor))

    items.sort(key=lambda x: (x[1], -x[3]))

    fig, ax = plt.subplots(figsize=(10, 6.5), dpi=100)
    max_val = max(i[3] for i in items)
    min_val = min(i[3] for i in items)
    cmap = plt.cm.Blues

    y_positions = np.arange(len(items))
    for i, (nome, t_ini, t_fim, valor) in enumerate(items):
        norm_val = (valor - min_val) / (max_val - min_val) if max_val > min_val else 0.5
        color = cmap(0.35 + norm_val * 0.55)
        ax.barh(
            i, t_fim - t_ini + 1, left=t_ini,
            color=color, edgecolor="white", linewidth=0.5, height=0.75,
        )
        x_text = (t_ini + t_fim + 1) / 2
        ax.text(
            x_text, i, f"R$ {valor/1e6:.1f}M",
            ha="center", va="center", fontsize=8,
            color="white" if norm_val > 0.45 else "#2C3E50", fontweight="bold",
        )

    ax.set_yticks(y_positions)
    ax.set_yticklabels([i[0] for i in items], fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("Mês da obra", fontsize=10)
    ax.set_title(
        "Gantt — Janela Temporal de Cada Macrogrupo",
        fontsize=12, color="#2C3E50", fontweight="bold",
    )
    ax.grid(axis="x", alpha=0.3)
    ax.set_xlim(0, prazo_obra + 2)
    ax.set_xticks(range(0, prazo_obra + 2, 2))
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


def plot_stacked_macrogrupos(fluxo, breakdown, prem, out_path):
    """Área empilhada: contribuição de cada macrogrupo no desembolso mensal."""
    PLOTS_DIR.mkdir(exist_ok=True)

    curva_config = prem["curva_obra_por_macrogrupo"]
    t_axis = fluxo["t_axis"]
    prazo_pre = fluxo["prazo_pre"]
    n = len(t_axis)

    matriz = {}
    for nome, janela in curva_config.items():
        if nome.startswith("_"):
            continue
        if nome not in breakdown:
            continue
        t_ini, t_fim = janela
        idx_start = max(0, prazo_pre + t_ini - 1)
        idx_end = min(n - 1, prazo_pre + t_fim - 1)
        span = idx_end - idx_start + 1
        valor = breakdown[nome]["valor"]
        row = np.zeros(n)
        if span > 0:
            row[idx_start:idx_end + 1] = valor / span
        matriz[nome] = row

    ordenado = sorted(matriz.items(), key=lambda x: -x[1].sum())
    labels = [o[0] for o in ordenado]
    data = np.array([o[1] for o in ordenado])

    mes_obra = np.arange(-prazo_pre, n - prazo_pre)

    fig, ax = plt.subplots(figsize=(11, 6.5), dpi=100)
    cmap = plt.cm.tab20
    colors = [cmap(i / max(len(labels), 1)) for i in range(len(labels))]

    ax.stackplot(
        mes_obra, data, labels=labels, colors=colors,
        alpha=0.88, edgecolor="white", linewidth=0.3,
    )
    ax.set_xlabel("Mês relativo à obra (0 = início da obra)", fontsize=10)
    ax.set_ylabel("Desembolso mensal", fontsize=10)
    ax.set_title(
        "Composição Mensal do Desembolso — Stacked por Macrogrupo",
        fontsize=12, color="#2C3E50", fontweight="bold",
    )
    ax.grid(axis="y", alpha=0.3)
    ax.legend(loc="upper right", fontsize=7, ncol=2, framealpha=0.92)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"R$ {x/1e6:.1f}M"))
    prazo_obra_local = prem["projeto"]["prazo_obra_meses"]
    ax.set_xlim(-prazo_pre, prazo_obra_local + 2)
    fig.tight_layout()
    fig.savefig(out_path)
    plt.close(fig)


# ============================================================================
# BUILDER — DETALHE_CUSTO
# ============================================================================
JUSTIFICATIVAS_MACROGRUPOS = {
    "Mov. Terra": (
        "Mobilização + terraplanagem",
        "Fase inicial do canteiro. Curta porque a área do terreno é pequena (1.008 m²) e o volume de movimentação limitado. Sai do canteiro rápido.",
    ),
    "Infraestrutura": (
        "Fundações + contenções",
        "Hélice contínua + blocos + baldrames. Inicia junto com a terraplanagem (mês 1) e termina quando a supraestrutura já iniciou o 1º pavimento (mês 8).",
    ),
    "Supraestrutura": (
        "Estrutura principal em concreto armado",
        "Pilares, vigas e lajes dos 24 pavimentos. Ritmo típico de 1 pavimento a cada 30-40 dias. Começa no mês 3 (fundação avançada) e termina na cobertura (mês 18). É o maior macrogrupo do projeto.",
    ),
    "Alvenaria": (
        "Vedações (bloco cerâmico/concreto)",
        "Segue a supraestrutura com ~5 meses de defasagem (primeira fiada quando a laje de cima está pronta). Distribuída pelos 14 meses porque cada pavimento é executado sequencialmente.",
    ),
    "Impermeabilização": (
        "Baldrames, áreas molhadas, coberturas",
        "Janela longa (5-25) porque ocorre em várias fases distintas: baldrames cedo, áreas molhadas em cada pavimento ao longo da obra, coberturas e sacadas no final.",
    ),
    "Instalações": (
        "Hidrossanitária + elétrica + gás",
        "Tubulações e eletrodutos embutidos durante a execução da alvenaria. Encerra junto com acabamento (quando recebem louças e metais).",
    ),
    "Sist. Especiais": (
        "Automação + CFTV + gerador + pressurização",
        "Instalados após a alvenaria pronta (mês 18) e validados no comissionamento final da obra (mês 29).",
    ),
    "Climatização": (
        "Infra split + condensadoras",
        "Tubulação de frigorígeno durante a alvenaria; condensadoras e comissionamento na fase de acabamento.",
    ),
    "Rev. Int. Parede": (
        "Emboço, reboco e revestimento",
        "Começa após a alvenaria avançar (~mês 15), termina pouco antes da pintura. Depende da supraestrutura estar estável para evitar fissuras.",
    ),
    "Teto": (
        "Forros de gesso acartonado",
        "Só depois que as instalações elétricas, hidráulicas e de climatização passaram e o ambiente está pronto para ser fechado.",
    ),
    "Pisos": (
        "Contrapiso + porcelanato",
        "Contrapiso segue a alvenaria (mês 20); acabamento cerâmico após pintura começar. Dois serviços distintos distribuídos ao longo da janela.",
    ),
    "Pintura": (
        "Massa corrida + tinta",
        "Última fase de acabamento (mês 22-30). Requer teto, pisos, emboço e esquadrias prontos. Não pode iniciar antes.",
    ),
    "Esquadrias": (
        "Portas, janelas, alumínio, vidros",
        "Marcos colocados durante alvenaria (mês 18); folhas, vidros e ajustes no final da obra (mês 28).",
    ),
    "Louças e Metais": (
        "Bacias, cubas, torneiras, registros",
        "Chegam quando áreas molhadas já estão revestidas e pisos instalados (mês 22-29). É a última etapa antes da vistoria.",
    ),
    "Fachada": (
        "Reboco externo + revestimento cerâmico + pintura",
        "Reboco externo começa a partir do mês 15 (por pavimento, seguindo a subida). Revestimento cerâmico e pintura de fachada até mês 27.",
    ),
    "Complementares": (
        "Elevadores + paisagismo + urbanização + portaria",
        "Instalados no final da obra (mês 24-30) quando o prédio está quase pronto para receber habite-se.",
    ),
    "Gerenciamento": (
        "Equipe de obra + canteiro + ART",
        "Custo indireto constante: engenheiros, mestres, apontadores, canteiro, vigilância, EPIs, ART/CREA. Vai do início ao fim da obra (1-30, distribuição linear).",
    ),
    "Imprevistos": (
        "Reserva de contingência",
        "Reserva técnica para cobrir imprevistos. Distribuída linearmente pelo horizonte (1-30) para não distorcer o fluxo.",
    ),
}


def build_detalhe_custo(ws, prem, breakdown, fluxo, gantt_png, curva_s_png, stacked_png):
    """DETALHE_CUSTO — tabela interativa (janelas editáveis) + narrativa + plots.

    Colunas: B..G = info (nome/R$/mês_ini/mês_fim/dur/R$-mes),
             H..(H+prazo_obra-1) = meses 1..prazo_obra.
    """
    from openpyxl.formatting.rule import DataBarRule
    prazo_obra = prem["projeto"]["prazo_obra_meses"]
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22  # Macrogrupo
    ws.column_dimensions["C"].width = 14  # R$ Total
    ws.column_dimensions["D"].width = 8   # Mês ini
    ws.column_dimensions["E"].width = 8   # Mês fim
    ws.column_dimensions["F"].width = 7   # Dur
    ws.column_dimensions["G"].width = 12  # R$/mês
    for _col in range(8, 8 + prazo_obra):  # H..(H+prazo_obra-1)
        ws.column_dimensions[get_column_letter(_col)].width = 5.2

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 36
    ws.merge_cells("B2:K2")
    c = ws["B2"]
    c.value = "DETALHAMENTO DO CUSTO DE OBRA — POR MACROGRUPO"
    c.fill = FILL_DARK
    c.font = FONT_LARGE
    c.alignment = ALIGN_CENTER

    ws.row_dimensions[3].height = 24
    ws.merge_cells("B3:K3")
    c = ws["B3"]
    custo_total = sum(v["valor"] for v in breakdown.values())
    ac = prem["projeto"]["area_construida_m2"]
    c.value = (
        f"Como os R$ {custo_total/1e6:.2f} milhões  (R$ {custo_total/ac:,.0f}/m² AC)  se distribuem entre os 18 macrogrupos "
        f"ao longo dos {prem['projeto']['prazo_obra_meses']} meses de obra"
    )
    c.fill = FILL_BLUE
    c.font = Font(name="Arial", size=11, italic=True, color=DARK)
    c.alignment = ALIGN_CENTER

    # Seção 1: De onde vem
    r = 5
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="DE ONDE VEM ESSE CUSTO — METODOLOGIA PARAMÉTRICA V2 HÍBRIDO")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    texto_metodologia = (
        f"O custo de obra de R$ {custo_total/1e6:.2f} milhões foi calculado pelo paramétrico V2 Híbrido da Cartesian — "
        f"metodologia que combina três fontes de dados:\n\n"
        f"1)  QUANTITATIVOS — extraídos do modelo BIM (IFC) do Arboris: volumes de concreto, área de fôrma, m² de alvenaria, "
        f"m² de revestimento, kg de aço, unidades de esquadrias, etc. Para o Arboris: {prem['projeto']['area_construida_m2']:,.0f} m² AC, "
        f"8.653,8 m³ de concreto (0,694 m³/m²), ~122 mil m² de área de parede, 24 pavimentos.\n\n"
        f"2)  PUs CALIBRADOS — Preços Unitários (R$/unidade) calibrados contra uma base histórica de 126 projetos Cartesian "
        f"já executados no litoral SC. A base tem 4.210 PUs cross-projeto e permite ajustar o preço por tipologia, padrão "
        f"de acabamento e região.\n\n"
        f"3)  ÍNDICES DERIVADOS — 29 índices que ajustam o custo final conforme o briefing do Arboris: fundação em hélice "
        f"contínua, laje convencional, padrão médio-alto, fachada cerâmica, 1 subsolo, gerador, piscina aquecida, 2 elevadores.\n\n"
        f"O resultado é agrupado nos 18 MACROGRUPOS abaixo — estrutura padrão Cartesian para orçamentos paramétricos. "
        f"Cada macrogrupo representa uma disciplina coerente da obra (ex: 'Supraestrutura' = todo o sistema estrutural em concreto "
        f"armado; 'Instalações' = hidráulica + elétrica + gás + telecom). Benchmark do segmento Médio (8-15k m² AC): mediana "
        f"R$ 3.337/m², faixa P25-P75 R$ 2.974-3.710. Arboris em R$ {custo_total/ac:,.0f}/m² — levemente acima da mediana, "
        f"coerente com padrão Médio-Alto."
    )
    ws.merge_cells(f"B{r}:K{r+13}")
    c = ws.cell(row=r, column=2, value=texto_metodologia)
    c.font = FONT_NARRATIVE
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_GRAY
    c.border = BORDER
    for rr in range(r, r + 14):
        ws.row_dimensions[rr].height = 18
    r += 14

    # Seção 2: Tabela INTERATIVA
    last_month_col = 7 + prazo_obra
    last_month_letter_sec2 = get_column_letter(last_month_col)
    r += 1
    ws.merge_cells(f"B{r}:{last_month_letter_sec2}{r}")
    c = ws.cell(row=r, column=2, value="TABELA INTERATIVA — EDITE OS MESES E A CURVA REDISTRIBUI AUTOMATICAMENTE")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    # Caixa de instrução (fundo amarelo)
    instrucao = (
        f"▸  As colunas LARANJAS são editáveis: altere o R$ Total, Mês início ou Mês fim de qualquer macrogrupo e a "
        f"distribuição mensal (colunas M1 a M{prazo_obra}) recalcula automaticamente via fórmulas.\n"
        "▸  As células AZUIS na grade mostram visualmente a JANELA ATIVA de cada macrogrupo (efeito Gantt) — quanto mais "
        "forte a cor, maior o desembolso daquele macrogrupo naquele mês.\n"
        "▸  A última linha 'Acumulado %' mostra a CURVA S ao vivo via data bars azuis — dá pra ver quando o projeto atinge "
        "25%, 50%, 75% e 100% do desembolso.\n"
        "▸  IMPORTANTE: esta tabela é ferramenta de EXPLORAÇÃO visual. Os valores de VPL/TIR/Monte Carlo nas outras abas NÃO "
        "são recalculados automaticamente — para propagar mudanças ao modelo financeiro, edite premissas-viabilidade.json e "
        "rode o script novamente."
    )
    ws.merge_cells(f"B{r}:{last_month_letter_sec2}{r+4}")
    c = ws.cell(row=r, column=2, value=instrucao)
    c.font = FONT_SMALL
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_YELLOW
    c.border = BORDER
    for rr in range(r, r + 5):
        ws.row_dimensions[rr].height = 15
    r += 5

    # Header: info cols (B..G) + meses (H..last)
    r_header = r
    headers_info = ["Macrogrupo", "R$ Total", "Mês ini", "Mês fim", "Dur.", "R$/mês"]
    for col_offset, h in enumerate(headers_info):
        col = 2 + col_offset
        c = ws.cell(row=r_header, column=col, value=h)
        style_header(c)
    for m in range(1, prazo_obra + 1):
        col = 7 + m  # H = 8 para M = 1
        c = ws.cell(row=r_header, column=col, value=m)
        style_header(c)
        c.font = Font(name="Arial", size=8, bold=True, color="FFFFFF")
    ws.row_dimensions[r_header].height = 26
    r += 1

    # Data rows — 18 macrogrupos com FORMULAS
    curva_config = prem["curva_obra_por_macrogrupo"]
    items = []
    for nome, janela in curva_config.items():
        if nome.startswith("_"):
            continue
        if nome not in breakdown:
            continue
        items.append((nome, janela[0], janela[1], breakdown[nome]["valor"]))
    items.sort(key=lambda x: (x[1], -x[3]))

    r_first_data = r
    for nome, t_ini, t_fim, valor in items:
        # B: Nome (não editável)
        c = ws.cell(row=r, column=2, value=nome)
        c.font = FONT_BOLD
        c.alignment = ALIGN_LEFT
        c.fill = FILL_BLUE
        c.border = BORDER
        # C: R$ total (EDITÁVEL — laranja)
        c = ws.cell(row=r, column=3, value=valor)
        style_input(c, fmt=FMT_BRL)
        # D: Mês início (EDITÁVEL — laranja)
        c = ws.cell(row=r, column=4, value=t_ini)
        style_input(c, fmt=FMT_NUM)
        # E: Mês fim (EDITÁVEL — laranja)
        c = ws.cell(row=r, column=5, value=t_fim)
        style_input(c, fmt=FMT_NUM)
        # F: Duração (fórmula)
        c = ws.cell(row=r, column=6, value=f"=IF(E{r}>=D{r},E{r}-D{r}+1,0)")
        style_calc(c, fmt=FMT_NUM)
        # G: R$/mês (fórmula)
        c = ws.cell(row=r, column=7, value=f"=IF(F{r}>0,C{r}/F{r},0)")
        style_calc(c, fmt=FMT_BRL)
        # H.. última: distribuição mensal via fórmula
        for m in range(1, prazo_obra + 1):
            col = 7 + m
            col_letter = get_column_letter(col)
            formula = f"=IF(AND({col_letter}${r_header}>=$D{r},{col_letter}${r_header}<=$E{r}),$G{r},0)"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '#,##0;;;'
            cell.font = Font(name="Arial", size=7)
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER
        ws.row_dimensions[r].height = 18
        r += 1
    r_last_data = r - 1

    # TOTAL row
    r_total = r
    c = ws.cell(row=r_total, column=2, value="TOTAL")
    c.fill = FILL_DARK
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.alignment = ALIGN_LEFT
    c.border = BORDER
    c = ws.cell(row=r_total, column=3, value=f"=SUM(C{r_first_data}:C{r_last_data})")
    c.fill = FILL_DARK
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    c.alignment = ALIGN_RIGHT
    c.border = BORDER
    c.number_format = FMT_BRL
    for col in range(4, 8):
        c = ws.cell(row=r_total, column=col)
        c.fill = FILL_DARK
        c.border = BORDER
    for m in range(1, prazo_obra + 1):
        col = 7 + m
        col_letter = get_column_letter(col)
        c = ws.cell(row=r_total, column=col, value=f"=SUM({col_letter}{r_first_data}:{col_letter}{r_last_data})")
        c.fill = FILL_DARK
        c.font = Font(name="Arial", size=7, bold=True, color="FFFFFF")
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        c.number_format = '#,##0;;;'
    ws.row_dimensions[r_total].height = 22
    r += 1

    # Linha: % do mês
    r_pct = r
    c = ws.cell(row=r_pct, column=2, value="% do total / mês")
    c.font = FONT_BOLD
    c.alignment = ALIGN_LEFT
    c.fill = FILL_GRAY
    c.border = BORDER
    for col in range(3, 8):
        c = ws.cell(row=r_pct, column=col)
        c.fill = FILL_GRAY
        c.border = BORDER
    for m in range(1, prazo_obra + 1):
        col = 7 + m
        col_letter = get_column_letter(col)
        formula = f"=IF($C${r_total}>0,{col_letter}{r_total}/$C${r_total},0)"
        c = ws.cell(row=r_pct, column=col, value=formula)
        c.fill = FILL_GRAY
        c.font = Font(name="Arial", size=7)
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        c.number_format = '0.0%;;;'
    r += 1

    # Linha: Acumulado R$
    r_acum = r
    c = ws.cell(row=r_acum, column=2, value="Acumulado R$")
    c.font = FONT_BOLD
    c.alignment = ALIGN_LEFT
    c.fill = FILL_CALC
    c.border = BORDER
    for col in range(3, 8):
        c = ws.cell(row=r_acum, column=col)
        c.fill = FILL_CALC
        c.border = BORDER
    first_letter = get_column_letter(8)  # H
    c = ws.cell(row=r_acum, column=8, value=f"={first_letter}{r_total}")
    c.fill = FILL_CALC
    c.font = Font(name="Arial", size=7, bold=True)
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    c.number_format = '#,##0;;;'
    for m in range(2, prazo_obra + 1):
        col = 7 + m
        col_letter = get_column_letter(col)
        prev_letter = get_column_letter(col - 1)
        formula = f"={prev_letter}{r_acum}+{col_letter}{r_total}"
        c = ws.cell(row=r_acum, column=col, value=formula)
        c.fill = FILL_CALC
        c.font = Font(name="Arial", size=7, bold=True)
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        c.number_format = '#,##0;;;'
    r += 1

    # Linha: Acumulado % (curva S) com data bars
    r_acum_pct = r
    c = ws.cell(row=r_acum_pct, column=2, value="Acumulado %  (curva S)")
    c.font = FONT_BOLD
    c.alignment = ALIGN_LEFT
    c.fill = FILL_BLUE
    c.border = BORDER
    for col in range(3, 8):
        c = ws.cell(row=r_acum_pct, column=col)
        c.fill = FILL_BLUE
        c.border = BORDER
    for m in range(1, prazo_obra + 1):
        col = 7 + m
        col_letter = get_column_letter(col)
        formula = f"=IF($C${r_total}>0,{col_letter}{r_acum}/$C${r_total},0)"
        c = ws.cell(row=r_acum_pct, column=col, value=formula)
        c.font = Font(name="Arial", size=7, bold=True)
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        c.number_format = '0%;;;'
    ws.row_dimensions[r_acum_pct].height = 24
    r += 1

    # Conditional formatting: Gantt visual — pinta células ativas de azul
    first_month_letter = get_column_letter(8)
    last_month_letter = get_column_letter(7 + prazo_obra)
    gantt_range = f"{first_month_letter}{r_first_data}:{last_month_letter}{r_last_data}"
    gantt_rule = CellIsRule(
        operator='greaterThan',
        formula=['0'],
        fill=PatternFill(start_color="90CAF9", end_color="1976D2", fill_type="solid"),
    )
    ws.conditional_formatting.add(gantt_range, gantt_rule)

    # Data bars na linha Acumulado %
    acum_range = f"{first_month_letter}{r_acum_pct}:{last_month_letter}{r_acum_pct}"
    db_rule = DataBarRule(
        start_type='num', start_value=0,
        end_type='num', end_value=1,
        color="2980B9",
        showValue=True,
    )
    ws.conditional_formatting.add(acum_range, db_rule)

    r += 2

    # Seção 3: Como a janela temporal foi definida
    r += 1
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="COMO A JANELA TEMPORAL DE CADA MACROGRUPO FOI DEFINIDA")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    texto_janelas = (
        "Cada macrogrupo tem uma JANELA DE EXECUÇÃO — os meses da obra em que aquele serviço está ativo no canteiro. "
        "Dentro da janela, o custo total do macrogrupo é distribuído UNIFORMEMENTE (R$/mês médio = custo ÷ duração). "
        "Isso simplifica o cálculo e é suficiente para uma análise de viabilidade — um BIM 4D com cronograma Primavera "
        "daria mais precisão, mas o ganho em termos de VPL/TIR é marginal (<3%).\n\n"
        "As janelas foram definidas com base no sequenciamento típico de obras verticais de padrão médio-alto, "
        "referência: histórico Cartesian de projetos similares no litoral SC. A lógica construtiva segue três fases:\n\n"
        "▸  FASE 1  —  meses 1 a 8  (Estrutural):  Mobilização, movimentação de terra, fundações (hélice contínua) e início "
        "da supraestrutura. Serviços 'que não podem esperar' — concreto, armação, fôrma. Cerca de 15-20% do custo total "
        "concentrado aqui.\n\n"
        "▸  FASE 2  —  meses 8 a 22  (Intensiva):  Subida da supraestrutura + alvenaria + instalações embutidas + revestimentos "
        "internos + fachada. É a fase MAIS INTENSA em desembolso — concreto, aço, tijolo, eletrodutos, tubulações, reboco. "
        "Concentra 50-60% do custo total. Pico do fluxo de caixa negativo.\n\n"
        "▸  FASE 3  —  meses 18 a 30  (Acabamento):  Pisos, pintura, teto, esquadrias, louças, metais, fachada externa, "
        "sistemas especiais, elevadores, complementares. Fluxo 'longo e diluído': muitos fornecedores diferentes, entregas "
        "escalonadas, dependências entre serviços. ~25-30% do custo.\n\n"
        "DOIS MACROGRUPOS SÃO LINEARES NA JANELA INTEIRA (1-30): 'Gerenciamento' (equipe indireta de obra, canteiro, ART) "
        "e 'Imprevistos' (reserva de contingência) — custos que acompanham a obra do início ao fim, sem concentração."
    )
    ws.merge_cells(f"B{r}:K{r+15}")
    c = ws.cell(row=r, column=2, value=texto_janelas)
    c.font = FONT_NARRATIVE
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_GRAY
    c.border = BORDER
    for rr in range(r, r + 16):
        ws.row_dimensions[rr].height = 18
    r += 16

    # Seção 4: Limitações do modelo
    r += 1
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="LIMITAÇÕES DO MODELO TEMPORAL — O QUE O MODELO NÃO CAPTURA")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 1

    limitacoes = (
        "Para ser transparente: a distribuição temporal por macrogrupo é uma APROXIMAÇÃO. O que o modelo NÃO captura e "
        "que pode mudar na execução real:\n\n"
        "• Distribuição uniforme dentro da janela é irreal em alguns serviços. Por exemplo, 'Supraestrutura' na prática "
        "tem curva S própria (mais concreto no meio dos 24 pavimentos, menos no início e no topo). O modelo assume "
        "linear — erro de ±8% no R$/mês pontual, mas a SOMA total está correta.\n\n"
        "• Curva de pagamento ao fornecedor ≠ curva física de execução. Concreto é pago 30 dias após a entrega; aço é pago "
        "antecipado à compra. O modelo usa a curva física (ignora os prazos de 30-60 dias de pagamento). Em análise de "
        "fluxo de caixa real, adicionar 30 dias de prazo médio suavizaria o pico negativo em ~R$ 1-2M.\n\n"
        "• Chuvas, atrasos e replanejamento não estão modelados. A premissa é que o prazo de 30 meses é cumprido. Se "
        "houver atraso de 3 meses, o fluxo estica e o VPL piora ~R$ 0,8M (TMA × 3 meses × exposição).\n\n"
        "• O valor por macrogrupo vem do PARAMÉTRICO (±10-15% de incerteza) — não de um orçamento executivo item-a-item. "
        "Para cotações reais em RFP, considerar que cada macrogrupo pode variar dentro dessa faixa.\n\n"
        "Para decisão de viabilidade (ir / não ir), o modelo atual é adequado. Para construtabilidade, cronograma Primavera "
        "e SFH bancário, seria necessário um BIM 4D + orçamento executivo."
    )
    ws.merge_cells(f"B{r}:K{r+14}")
    c = ws.cell(row=r, column=2, value=limitacoes)
    c.font = FONT_NARRATIVE
    c.alignment = ALIGN_WRAP_TOP
    c.fill = FILL_YELLOW
    c.border = BORDER
    for rr in range(r, r + 15):
        ws.row_dimensions[rr].height = 18
    r += 15

    # Seção 5: Visualizações
    r += 1
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="VISUALIZAÇÕES")
    c.fill = FILL_ACCENT
    c.font = FONT_H1
    c.alignment = ALIGN_CENTER
    c.border = BORDER
    r += 2

    # Gantt
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="▸  Gráfico 1  —  Gantt: janela temporal de cada macrogrupo (cor = magnitude do custo)")
    c.font = FONT_NARRATIVE_BOLD
    c.alignment = ALIGN_LEFT
    r += 1
    if gantt_png and gantt_png.exists():
        img = XLImage(str(gantt_png))
        img.width = 720
        img.height = 460
        ws.add_image(img, f"B{r}")
        r += 24

    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="▸  Gráfico 2  —  Curva S: desembolso acumulado da obra (% do custo total por mês)")
    c.font = FONT_NARRATIVE_BOLD
    c.alignment = ALIGN_LEFT
    r += 1
    if curva_s_png and curva_s_png.exists():
        img = XLImage(str(curva_s_png))
        img.width = 660
        img.height = 340
        ws.add_image(img, f"B{r}")
        r += 18

    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(row=r, column=2, value="▸  Gráfico 3  —  Composição mensal: quanto cada macrogrupo contribui em cada mês")
    c.font = FONT_NARRATIVE_BOLD
    c.alignment = ALIGN_LEFT
    r += 1
    if stacked_png and stacked_png.exists():
        img = XLImage(str(stacked_png))
        img.width = 780
        img.height = 460
        ws.add_image(img, f"B{r}")


def build_v2_workbook(
    prem, kpis, mc_stats, cenarios, tornado, fluxo, breakdown,
    hist_vpl_png, hist_tir_png, tornado_png,
    gantt_png, curva_s_png, stacked_png, fluxo_png,
):
    """Monta o workbook v2 didático — PARECER + COMO_LER + DETALHE_CUSTO na frente."""
    wb = Workbook()
    wb.active.title = "PARECER"
    build_parecer(wb["PARECER"], prem, kpis, mc_stats, cenarios, tornado)
    build_como_ler(wb.create_sheet("COMO_LER"), prem, kpis, mc_stats, cenarios, tornado)
    build_premissas(wb.create_sheet("PREMISSAS"), prem, kpis)
    build_detalhe_custo(
        wb.create_sheet("DETALHE_CUSTO"),
        prem, breakdown, fluxo, gantt_png, curva_s_png, stacked_png,
    )
    fc_ws = wb.create_sheet("FLUXO_CAIXA")
    r_fluxo_liquido = build_fluxo_caixa(fc_ws, fluxo, prem, fluxo_png=fluxo_png)
    build_resultado(
        wb.create_sheet("RESULTADO"),
        prem, kpis, mc_stats, fluxo, r_fluxo_liquido,
    )
    build_cenarios(wb.create_sheet("CENARIOS"), cenarios, prem)
    build_tornado_sheet(wb.create_sheet("TORNADO"), tornado, kpis["vpl"], tornado_png)
    build_monte_carlo_sheet(
        wb.create_sheet("MONTE_CARLO"), mc_stats, hist_vpl_png, hist_tir_png,
        custo_base=kpis["custo_obra_total"],
    )
    wb.save(OUTPUT_XLSX_V2)


# ============================================================================
# MAIN
# ============================================================================
def main():
    print("=" * 70)
    print("VIABILIDADE FINANCEIRA — ARTHEN ARBORIS")
    print("=" * 70)

    print("\n[1/7] Lendo paramétrico…")
    custo_obra_total, breakdown = load_parametrico()
    print(f"      Custo obra: R$ {custo_obra_total:,.0f}")
    print(f"      Macrogrupos: {len(breakdown)}")

    print("\n[2/7] Lendo premissas…")
    prem = load_premissas()
    prem["_gerado_em"] = prem.get("_gerado_em", "2026-04-14")

    print("\n[3/7] Computando fluxo base…")
    fluxo = compute_fluxo(prem, custo_obra_total, breakdown)
    kpis = compute_kpis(fluxo)
    print(f"      VGV bruto      R$ {kpis['vgv_bruto']:>15,.0f}")
    print(f"      VGV líquido    R$ {kpis['vgv_liquido']:>15,.0f}")
    print(f"      Lucro          R$ {kpis['lucro']:>15,.0f}")
    print(f"      Margem         {kpis['margem']*100:>6.1f}%")
    print(f"      VPL @ 14%      R$ {kpis['vpl']:>15,.0f}")
    if kpis["tir_anual"] is not None:
        print(f"      TIR a.a.       {kpis['tir_anual']*100:>6.1f}%")
    print(f"      Exposição máx  R$ {kpis['exposicao_max']:>15,.0f}")

    print("\n[4/7] Cenários (pessimista/base/otimista)…")
    cenarios = run_cenarios(prem, custo_obra_total, breakdown)
    for nome in ["pessimista", "base", "otimista"]:
        k = cenarios[nome]
        print(f"      {nome:12s} VPL R$ {k['vpl']:>15,.0f}  "
              f"TIR {(k['tir_anual'] or 0)*100:5.1f}%  margem {k['margem']*100:5.1f}%")

    print("\n[5/7] Tornado…")
    tornado, base_vpl = run_tornado(prem, custo_obra_total, breakdown, kpis["vpl"])
    for res in tornado[:5]:
        print(f"      {res['label']:20s} delta R$ {res['delta']:>15,.0f}")

    print(f"\n[6/7] Monte Carlo ({prem['monte_carlo']['n_iter']} iter)…")
    mc_stats, mc_samples = run_monte_carlo(prem, custo_obra_total, breakdown)
    print(f"      VPL mediana    R$ {mc_stats['vpl_mediana']:>15,.0f}")
    print(f"      VPL P10        R$ {mc_stats['vpl_p10']:>15,.0f}")
    print(f"      VPL P90        R$ {mc_stats['vpl_p90']:>15,.0f}")
    print(f"      P(VPL > 0)     {mc_stats['prob_vpl_positivo']*100:5.1f}%")
    print(f"      P(TIR > TMA)   {mc_stats['prob_tir_gt_tma']*100:5.1f}%")

    print("\n      Gerando plots…")
    PLOTS_DIR.mkdir(exist_ok=True)
    hist_vpl_png = PLOTS_DIR / "hist_vpl.png"
    hist_tir_png = PLOTS_DIR / "hist_tir.png"
    tornado_png = PLOTS_DIR / "tornado.png"
    plot_histogram(mc_samples["vpl"], "Distribuição VPL (10.000 iterações)", "VPL (R$)", hist_vpl_png)
    plot_histogram(
        mc_samples["tir_anual"][~np.isnan(mc_samples["tir_anual"])],
        "Distribuição TIR anual (10.000 iterações)",
        "TIR a.a.",
        hist_tir_png,
        color="#5D9C7A",
    )
    plot_tornado(tornado, kpis["vpl"], tornado_png)

    # Plots de detalhe de custo (v2) + saldo acumulado (fluxo caixa didático)
    gantt_png = PLOTS_DIR / "gantt_macrogrupos.png"
    curva_s_png = PLOTS_DIR / "curva_s_obra.png"
    stacked_png = PLOTS_DIR / "stacked_macrogrupos.png"
    fluxo_png = PLOTS_DIR / "fluxo_acumulado.png"
    prazo_obra_local = prem["projeto"]["prazo_obra_meses"]
    plot_gantt_macrogrupos(
        breakdown, prem["curva_obra_por_macrogrupo"], gantt_png,
        prazo_obra=prazo_obra_local,
    )
    plot_curva_desembolso_s(fluxo, curva_s_png)
    plot_stacked_macrogrupos(fluxo, breakdown, prem, stacked_png)
    plot_fluxo_acumulado(fluxo, fluxo_png)

    print("\n[7/7] Montando workbook…")
    wb = Workbook()
    wb.active.title = "CAPA"

    build_capa(wb["CAPA"], prem, kpis, mc_stats)
    build_premissas(wb.create_sheet("PREMISSAS"), prem, kpis)
    fc_ws = wb.create_sheet("FLUXO_CAIXA")
    r_fluxo_liquido = build_fluxo_caixa(fc_ws, fluxo, prem, fluxo_png=fluxo_png)
    build_resultado(
        wb.create_sheet("RESULTADO"),
        prem, kpis, mc_stats, fluxo, r_fluxo_liquido,
    )
    build_cenarios(wb.create_sheet("CENARIOS"), cenarios, prem)
    build_tornado_sheet(wb.create_sheet("TORNADO"), tornado, kpis["vpl"], tornado_png)
    build_monte_carlo_sheet(
        wb.create_sheet("MONTE_CARLO"), mc_stats, hist_vpl_png, hist_tir_png,
        custo_base=kpis["custo_obra_total"],
    )

    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Salvo (v1, analítico): {OUTPUT_XLSX}")

    print("\n[8/8] Montando workbook V2 didático (para diretor Arthen)…")
    build_v2_workbook(
        prem, kpis, mc_stats, cenarios, tornado, fluxo, breakdown,
        hist_vpl_png, hist_tir_png, tornado_png,
        gantt_png, curva_s_png, stacked_png, fluxo_png,
    )
    print(f"✓ Salvo (v2, didático): {OUTPUT_XLSX_V2}")
    print(f"\n  Abrir v2 no Excel e validar PARECER (veredicto) + COMO_LER (glossário).")

    return kpis, mc_stats


if __name__ == "__main__":
    main()
