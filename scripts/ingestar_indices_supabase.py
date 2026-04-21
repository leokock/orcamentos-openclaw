#!/usr/bin/env python
"""
Ingestão INDICES-CATALOGO.xlsx → Supabase (projeto indices-cartesian).

Lê as 9 abas de dados do xlsx e popula as 9 tabelas correspondentes.
Idempotente — pode rodar múltiplas vezes sem duplicar (upsert via constraints únicas).

Uso:
    py -3.10 scripts/ingestar_indices_supabase.py [--dry-run] [--only TABELA]

Pré-requisitos:
    - .env.indices-cartesian na raiz do repo (com SUPABASE_URL, SUPABASE_SECRET_KEY)
    - supabase-py, openpyxl, python-dotenv (py 3.10)
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterator

from dotenv import dotenv_values
from openpyxl import load_workbook
from supabase import create_client, Client

REPO_ROOT = Path(__file__).resolve().parent.parent
ENV_FILE = REPO_ROOT / ".env.indices-cartesian"
XLSX_PATH = REPO_ROOT / "base" / "INDICES-CATALOGO.xlsx"
LOG_PATH = REPO_ROOT / "base" / "log-ingestao-supabase.md"

# Header das tabelas de dados está na linha 4 do xlsx (L1-L3 = titulo/subtitulo/branco).
HEADER_ROW = 4
DATA_START_ROW = HEADER_ROW + 1

# Batch size pra reduzir número de requests HTTP (Supabase aceita upserts em lote).
BATCH_SIZE = 500

# Valores vazios do xlsx que devem virar None no Postgres.
EMPTY_SENTINELS = {"", None, "—", "-"}

PADRAO_ENUM_VALIDOS = {
    "economico", "medio", "medio-alto", "alto", "luxo", "nao_classificado",
}


# ----------------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------------

def to_numeric(value: Any) -> float | None:
    if value in EMPTY_SENTINELS:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # Às vezes vem string com vírgula (raro, mas defensivo)
    s = str(value).strip().replace(",", ".")
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> int | None:
    if value in EMPTY_SENTINELS:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def to_text(value: Any) -> str | None:
    if value in EMPTY_SENTINELS:
        return None
    return str(value).strip()


def to_padrao_enum(value: Any) -> str | None:
    t = to_text(value)
    if t is None:
        return None
    t_lower = t.lower().strip()
    # Map de variantes pras labels canônicas do enum
    alias = {
        "econômico": "economico",
        "médio": "medio",
        "médio-alto": "medio-alto",
        "medio alto": "medio-alto",
    }
    t_lower = alias.get(t_lower, t_lower)
    return t_lower if t_lower in PADRAO_ENUM_VALIDOS else "nao_classificado"


# ----------------------------------------------------------------------------
# Leitura de abas com mapping de colunas xlsx → campos Supabase
# ----------------------------------------------------------------------------

def iter_rows(ws, headers: list[str]) -> Iterator[tuple[int, dict[str, Any]]]:
    """
    Itera linhas de dados (L5+), retorna (num_linha_xlsx, {header: valor}).
    """
    for row_num, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        # Pula linhas totalmente vazias
        if not any(cell is not None and cell != "" for cell in row):
            continue
        # Primeira célula vazia = linha inválida ou rodapé
        if row[0] in (None, ""):
            continue
        yield row_num, dict(zip(headers, row))


def ler_header(ws) -> list[str]:
    header_row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))
    return [str(c).strip() if c is not None else "" for c in header_row]


# ----------------------------------------------------------------------------
# Transformadores por aba (xlsx → dict p/ Supabase)
# ----------------------------------------------------------------------------

def t_projetos(row_num: int, row: dict) -> dict:
    return {
        "slug": to_text(row.get("slug")),
        "padrao_gemma": to_padrao_enum(row.get("padrao_gemma")),
        "confianca": to_text(row.get("confianca")),
        "ac_m2": to_numeric(row.get("ac_m2")),
        "ur": to_int(row.get("ur")),
        "total_rs": to_numeric(row.get("total_rs")),
        "rsm2": to_numeric(row.get("rsm2")),
        "m2_por_ur": to_numeric(row.get("m2_por_ur")),
        "rs_por_ur": to_numeric(row.get("rs_por_ur")),
        "cidade": to_text(row.get("cidade")),
        "fonte": to_text(row.get("fonte")),
        "fonte_xlsx_linha": row_num,
    }


def t_calibracao_global(row_num: int, row: dict) -> dict:
    return {
        "macrogrupo": to_text(row.get("macrogrupo")),
        "n": to_int(row.get("n")),
        "min_val": to_numeric(row.get("min")),
        "p10": to_numeric(row.get("p10")),
        "p25": to_numeric(row.get("p25")),
        "mediana": to_numeric(row.get("mediana")),
        "media": to_numeric(row.get("media")),
        "p75": to_numeric(row.get("p75")),
        "p90": to_numeric(row.get("p90")),
        "max_val": to_numeric(row.get("max")),
        "unidade": to_text(row.get("unidade")),
        "fonte": to_text(row.get("fonte")),
        "fonte_xlsx_linha": row_num,
    }


def t_calibracao_condicional(row_num: int, row: dict) -> dict:
    return {
        "padrao": to_padrao_enum(row.get("padrao")),
        "macrogrupo": to_text(row.get("macrogrupo")),
        "n": to_int(row.get("n")),
        "min_val": to_numeric(row.get("min")),
        "p10": to_numeric(row.get("p10")),
        "p25": to_numeric(row.get("p25")),
        "mediana": to_numeric(row.get("mediana")),
        "media": to_numeric(row.get("media")),
        "p75": to_numeric(row.get("p75")),
        "p90": to_numeric(row.get("p90")),
        "max_val": to_numeric(row.get("max")),
        "unidade": to_text(row.get("unidade")),
        "fonte": to_text(row.get("fonte")),
        "fonte_xlsx_linha": row_num,
    }


def t_indices_derivados_v2(row_num: int, row: dict) -> dict:
    return {
        "nome": to_text(row.get("nome")),
        "descricao": to_text(row.get("descricao")),
        "n": to_int(row.get("n")),
        "min_val": to_numeric(row.get("min")),
        "p10": to_numeric(row.get("p10")),
        "p25": to_numeric(row.get("p25")),
        "mediana": to_numeric(row.get("mediana")),
        "media": to_numeric(row.get("media")),
        "p75": to_numeric(row.get("p75")),
        "p90": to_numeric(row.get("p90")),
        "max_val": to_numeric(row.get("max")),
        "cv": to_numeric(row.get("cv")),
        "unidade": to_text(row.get("unidade")),
        "fonte": to_text(row.get("fonte")),
        "fonte_xlsx_linha": row_num,
    }


def t_indices_estruturais(row_num: int, row: dict) -> dict:
    return {
        "categoria": to_text(row.get("categoria")),
        "nome": to_text(row.get("nome")),
        "n": to_int(row.get("n")),
        "min_val": to_numeric(row.get("min")),
        "p10": to_numeric(row.get("p10")),
        "p25": to_numeric(row.get("p25")),
        "mediana": to_numeric(row.get("mediana")),
        "media": to_numeric(row.get("media")),
        "p75": to_numeric(row.get("p75")),
        "p90": to_numeric(row.get("p90")),
        "max_val": to_numeric(row.get("max")),
        "unidade": to_text(row.get("unidade")),
        "fonte": to_text(row.get("fonte")),
        "fonte_xlsx_linha": row_num,
    }


def t_pus_cross_v1(row_num: int, row: dict) -> dict:
    return {
        "categoria": to_text(row.get("categoria")),
        "chave": to_text(row.get("chave")),
        "descricao": to_text(row.get("descricao")),
        "unidade": to_text(row.get("unidade")),
        "n_proj": to_int(row.get("n_proj")),
        "n_obs": to_int(row.get("n_obs")),
        "min_val": to_numeric(row.get("min")),
        "p25": to_numeric(row.get("p25")),
        "mediana": to_numeric(row.get("mediana")),
        "p75": to_numeric(row.get("p75")),
        "max_val": to_numeric(row.get("max")),
        "cv": to_numeric(row.get("cv")),
        "projetos_fonte": to_text(row.get("projetos_fonte")),
        "fonte_xlsx_linha": row_num,
    }


def t_pus_cross_v2(row_num: int, row: dict) -> dict:
    return {
        "cluster_id": to_int(row.get("cluster_id")),
        "key_tokens": to_text(row.get("key (tokens)")),
        "descricao": to_text(row.get("descricao")),
        "unidades": to_text(row.get("unidades")),
        "n_proj": to_int(row.get("n_proj")),
        "n_obs": to_int(row.get("n_obs")),
        "pu_min": to_numeric(row.get("pu_min")),
        "pu_p10": to_numeric(row.get("pu_p10")),
        "pu_p25": to_numeric(row.get("pu_p25")),
        "pu_mediana": to_numeric(row.get("pu_mediana")),
        "pu_p75": to_numeric(row.get("pu_p75")),
        "pu_p90": to_numeric(row.get("pu_p90")),
        "pu_max": to_numeric(row.get("pu_max")),
        "pu_media": to_numeric(row.get("pu_media")),
        "cv": to_numeric(row.get("cv")),
        "fonte_xlsx_linha": row_num,
    }


def t_curva_abc_master(row_num: int, row: dict) -> dict:
    return {
        "slug": to_text(row.get("slug")),
        "status": to_text(row.get("status")),
        "n_itens": to_int(row.get("n_itens")),
        "n_a": to_int(row.get("n_a")),
        "pct_a": to_numeric(row.get("pct_a")),
        "valor_total_rs": to_numeric(row.get("valor_total_rs")),
        "fonte_xlsx_linha": row_num,
    }


def t_cross_insights_gemma(row_num: int, row: dict) -> dict:
    return {
        "secao": to_text(row.get("secao")),
        "tipo": to_text(row.get("tipo")),
        "campo": to_text(row.get("campo")),
        "conteudo": to_text(row.get("conteudo")),
        "fonte_xlsx_linha": row_num,
    }


# ----------------------------------------------------------------------------
# Configuração das 9 ingestões
# ----------------------------------------------------------------------------

ABAS = [
    # (nome_aba_xlsx, nome_tabela_supabase, transformador, on_conflict_cols)
    ("PROJETOS", "projetos", t_projetos, "slug"),
    ("CALIBRACAO_GLOBAL", "calibracao_global", t_calibracao_global, "macrogrupo,fonte"),
    ("CALIBRACAO_CONDICIONAL", "calibracao_condicional", t_calibracao_condicional, "padrao,macrogrupo,fonte"),
    ("INDICES_DERIVADOS_V2", "indices_derivados_v2", t_indices_derivados_v2, "nome"),
    ("INDICES_ESTRUTURAIS", "indices_estruturais", t_indices_estruturais, "categoria,nome,tipo_obra"),
    ("PUS_CROSS_V1", "pus_cross_v1", t_pus_cross_v1, "categoria,chave"),
    ("PUS_CROSS_V2", "pus_cross_v2", t_pus_cross_v2, "cluster_id"),
    ("CURVA_ABC_MASTER", "curva_abc_master", t_curva_abc_master, "slug"),
    ("CROSS_INSIGHTS_GEMMA", "cross_insights_gemma", t_cross_insights_gemma, None),  # append (sem unique)
]


# ----------------------------------------------------------------------------
# Ingestão
# ----------------------------------------------------------------------------

def chunked(iterable, size):
    buf = []
    for item in iterable:
        buf.append(item)
        if len(buf) >= size:
            yield buf
            buf = []
    if buf:
        yield buf


def ingerir_aba(
    client: Client,
    wb,
    aba_nome: str,
    tabela: str,
    transformador: Callable,
    on_conflict: str | None,
    dry_run: bool,
) -> tuple[int, int]:
    """Retorna (n_lidas, n_enviadas)."""
    ws = wb[aba_nome]
    headers = ler_header(ws)
    print(f"\n=== {aba_nome} → {tabela} ===")
    print(f"  Header: {headers}")

    rows_validas = []
    for row_num, row in iter_rows(ws, headers):
        registro = transformador(row_num, row)
        # Sanitiza: remove None values opcionais? Não — Supabase aceita NULL, trata defaults.
        # Valida que a chave primária natural não é None (senão upsert falha)
        rows_validas.append(registro)

    print(f"  Linhas válidas lidas: {len(rows_validas)}")
    if dry_run:
        print(f"  [dry-run] pularia envio. Amostra[0]: {rows_validas[0] if rows_validas else '(vazio)'}")
        return len(rows_validas), 0

    if not rows_validas:
        return 0, 0

    # CROSS_INSIGHTS_GEMMA não tem unique constraint → limpa tabela antes (append puro)
    if on_conflict is None:
        print(f"  Limpando {tabela} antes do insert (sem unique)...")
        client.table(tabela).delete().gte("created_at", "1900-01-01").execute()

    total_enviado = 0
    for batch in chunked(rows_validas, BATCH_SIZE):
        if on_conflict:
            resp = client.table(tabela).upsert(batch, on_conflict=on_conflict).execute()
        else:
            resp = client.table(tabela).insert(batch).execute()
        total_enviado += len(resp.data) if resp.data else len(batch)
        print(f"    batch {total_enviado}/{len(rows_validas)}")

    return len(rows_validas), total_enviado


def verificar_contagens(client: Client, alvos: list[tuple[str, int]]) -> list[str]:
    """Compara contagem real no Supabase vs esperado. Retorna lista de alertas."""
    alertas = []
    print("\n=== VERIFICAÇÃO DE CONTAGENS ===")
    for tabela, esperado in alvos:
        resp = client.table(tabela).select("id", count="exact").limit(1).execute()
        real = resp.count if hasattr(resp, "count") and resp.count is not None else "?"
        flag = "OK" if real == esperado else "!!"
        msg = f"  {flag} {tabela}: real={real} esperado={esperado}"
        print(msg)
        if real != esperado:
            alertas.append(msg.strip())
    return alertas


def append_log(mensagem: str) -> None:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S BRT")
    separador = f"\n## {timestamp}\n\n"
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not LOG_PATH.exists():
        LOG_PATH.write_text("# Log de Ingestão — indices-cartesian no Supabase\n", encoding="utf-8")
    with LOG_PATH.open("a", encoding="utf-8") as f:
        f.write(separador + mensagem + "\n")


# ----------------------------------------------------------------------------
# main
# ----------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Só lê e imprime, não envia pro Supabase.")
    parser.add_argument("--only", type=str, help="Processa só a aba especificada (nome do xlsx).")
    args = parser.parse_args()

    if not ENV_FILE.exists():
        print(f"ERRO: .env não encontrado em {ENV_FILE}", file=sys.stderr)
        return 1
    env = dotenv_values(ENV_FILE)
    url = env.get("SUPABASE_URL")
    key = env.get("SUPABASE_SECRET_KEY")
    if not url or not key or key.startswith("COLE_AQUI"):
        print("ERRO: SUPABASE_URL ou SUPABASE_SECRET_KEY não preenchidos em .env", file=sys.stderr)
        return 1

    if not XLSX_PATH.exists():
        print(f"ERRO: xlsx não encontrado em {XLSX_PATH}", file=sys.stderr)
        return 1

    print(f"> Conectando no Supabase: {url}")
    client = create_client(url, key)

    print(f"> Abrindo xlsx: {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

    totais: list[tuple[str, int, int]] = []
    for aba_nome, tabela, transformador, on_conflict in ABAS:
        if args.only and args.only != aba_nome:
            continue
        n_lidas, n_enviadas = ingerir_aba(
            client, wb, aba_nome, tabela, transformador, on_conflict, args.dry_run
        )
        totais.append((tabela, n_lidas, n_enviadas))

    # Verificação end-to-end (só se não for dry-run)
    alertas = []
    if not args.dry_run and not args.only:
        alvos_contagem = [(t, n) for (t, n, _) in totais]
        alertas = verificar_contagens(client, alvos_contagem)

    # Log em markdown
    bloco = ["Modo: " + ("dry-run" if args.dry_run else "apply")]
    if args.only:
        bloco.append(f"Só aba: {args.only}")
    bloco.append("")
    bloco.append("| Tabela | Lidas | Enviadas |")
    bloco.append("|---|---:|---:|")
    for t, n_l, n_e in totais:
        bloco.append(f"| {t} | {n_l} | {n_e} |")
    if alertas:
        bloco.append("")
        bloco.append("**Divergências detectadas:**")
        for a in alertas:
            bloco.append(f"- {a}")
    append_log("\n".join(bloco))

    print("\n> Log atualizado em", LOG_PATH)
    return 0 if not alertas else 2


if __name__ == "__main__":
    sys.exit(main())
