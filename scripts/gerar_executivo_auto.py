#!/usr/bin/env python3
"""Gerador de Orçamento Executivo Automatizado.

Recebe:
  - slug do projeto novo
  - briefing (AC, UR, etc.)
  - gate validado pelo Leo (xlsx ou JSON com decisões)

Gera:
  - executivo.xlsx com 18 abas (uma por macrogrupo) + RESUMO + REFERENCIAS
  - cada item com confidence tag (verde/amarelo/vermelho)
  - PUs vindos da base qualitativa (mediana dos similares) com fallback pra base V2
  - rastreabilidade: cada linha cita projetos-fonte
  - log-execucao.md inicial
  - memorial-executivo.md inicial

Uso:
  python gerar_executivo_auto.py --slug projeto-novo --gate gate-validado.xlsx \\
                                 --ac 15000 --ur 90 -o executivo.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
import consulta_similares as cs  # noqa: E402

DARK = "2C3E50"
ACCENT = "2980B9"
GREEN_BG = "D4EDDA"
YELLOW_BG = "FFF3CD"
RED_BG = "F8D7DA"
HEADER_FILL = "2C3E50"

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

MACROGRUPOS_CANONICOS = [
    "Gerenciamento", "Movimentação de Terra", "Infraestrutura", "Supraestrutura",
    "Alvenaria", "Impermeabilização", "Instalações", "Sistemas Especiais",
    "Climatização", "Rev. Interno Parede", "Teto", "Pisos", "Pintura",
    "Esquadrias", "Louças e Metais", "Fachada", "Complementares", "Imprevistos",
]


def hdr(ws, row, cols, widths=None):
    for i, t in enumerate(cols, 1):
        c = ws.cell(row, i, t)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def cell(ws, r, c, value, bold=False, fill=None, fmt=None, align="left"):
    ce = ws.cell(r, c, value)
    ce.font = Font(bold=bold, size=9, name="Arial")
    ce.border = THIN
    ce.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align == "left"))
    if fill:
        ce.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    if fmt:
        ce.number_format = fmt
    return ce


def parse_gate_xlsx(path: Path) -> dict:
    """Lê o xlsx do gate validado e devolve dict com decisões."""
    if not path.exists():
        return {}
    wb = load_workbook(path, data_only=True)
    if "GATE" not in wb.sheetnames:
        return {}
    ws = wb["GATE"]
    decisoes = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        decisao = str(row[0]).strip()
        resposta = str(row[1]).strip() if row[1] is not None else ""
        if decisao and resposta:
            decisoes[decisao] = resposta

    premissas_aprovadas = []
    if "PREMISSAS_PROPOSTAS" in wb.sheetnames:
        ws_pre = wb["PREMISSAS_PROPOSTAS"]
        for row in ws_pre.iter_rows(min_row=5, values_only=True):
            if not row or not row[0]:
                continue
            usar = str(row[4] or "SIM").strip().upper()
            if usar == "SIM":
                premissas_aprovadas.append({
                    "area": str(row[0] or "").strip(),
                    "premissa": str(row[1] or "").strip(),
                })

    sub_disciplinas_aprovadas = []
    if "SUB_DISCIPLINAS" in wb.sheetnames:
        ws_sd = wb["SUB_DISCIPLINAS"]
        for row in ws_sd.iter_rows(min_row=5, values_only=True):
            if not row or not row[0]:
                continue
            usar = str(row[5] or "SIM").strip().upper()
            if usar == "SIM":
                sub_disciplinas_aprovadas.append({
                    "macrogrupo": str(row[0] or "").strip(),
                    "sub": str(row[1] or "").strip(),
                    "freq": str(row[2] or "").strip(),
                    "fontes": str(row[3] or "").strip(),
                    "itens_exemplo": str(row[4] or "").strip(),
                })

    wb.close()
    return {
        "decisoes": decisoes,
        "premissas_aprovadas": premissas_aprovadas,
        "sub_disciplinas_aprovadas": sub_disciplinas_aprovadas,
    }


def confidence_color(item: dict) -> str:
    """Retorna fill color baseado em confidence:
    verde = freq >=3, amarelo = freq 1-2, vermelho = sem dados.
    """
    freq = item.get("freq_projetos", 0)
    if freq >= 3:
        return GREEN_BG
    elif freq >= 1:
        return YELLOW_BG
    return RED_BG


def confidence_label(item: dict) -> str:
    freq = item.get("freq_projetos", 0)
    if freq >= 3:
        return "🟢 Alta"
    elif freq >= 1:
        return "🟡 Média"
    return "🔴 Baixa"


def aba_resumo(wb: Workbook, slug: str, ac: float, ur: int | None,
                similares: list[dict], macrogrupos_data: dict) -> None:
    ws = wb.create_sheet("RESUMO", 0)
    ws.sheet_properties.tabColor = ACCENT

    ws["A1"] = f"{slug.upper()} — ORÇAMENTO EXECUTIVO AUTOMATIZADO"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"Gerado em {datetime.now().isoformat(timespec='seconds')} | "  \
                f"AC={ac:.0f} m² | UR={ur or '?'} | "  \
                f"baseado em {len(similares)} projetos similares"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    hdr(ws, 4, ["#", "Macrogrupo", "Total R$", "R$/m²", "% do total", "N itens", "Confiança"],
        [4, 28, 18, 14, 12, 10, 14])

    grand_total = 0.0
    rows_data = []
    for i, mg in enumerate(MACROGRUPOS_CANONICOS, start=1):
        d = macrogrupos_data.get(mg, {})
        total = d.get("total", 0)
        n = d.get("n_itens", 0)
        confianca = d.get("confianca_media", "—")
        rows_data.append((i, mg, total, n, confianca))
        grand_total += total

    for i, (idx, mg, total, n, conf) in enumerate(rows_data, start=5):
        cell(ws, i, 1, idx, align="center")
        cell(ws, i, 2, mg, bold=True)
        cell(ws, i, 3, total, fmt='"R$" #,##0', align="right")
        cell(ws, i, 4, total / ac if ac else 0, fmt='"R$" #,##0.00', align="right")
        cell(ws, i, 5, total / grand_total if grand_total else 0, fmt='0.0%', align="right")
        cell(ws, i, 6, n, align="center")
        cell(ws, i, 7, conf, align="center")

    total_row = 5 + len(MACROGRUPOS_CANONICOS)
    cell(ws, total_row, 2, "TOTAL", bold=True, fill="EAEDED")
    cell(ws, total_row, 3, grand_total, bold=True, fmt='"R$" #,##0', fill="EAEDED", align="right")
    cell(ws, total_row, 4, grand_total / ac if ac else 0, bold=True, fmt='"R$" #,##0.00', fill="EAEDED", align="right")
    cell(ws, total_row, 5, 1.0, bold=True, fmt='0.0%', fill="EAEDED", align="right")


def aba_macrogrupo(wb: Workbook, mg: str, itens: list[dict], ac: float,
                    macro_v2: dict | None = None) -> dict:
    """Cria uma aba pra um macrogrupo. Total vem dos similares V2,
    itens vêm da camada qualitativa pra granularizar."""
    titulo = mg.replace(" ", "_")[:25]
    ws = wb.create_sheet(titulo)
    ws.sheet_properties.tabColor = ACCENT

    ws["A1"] = f"{mg.upper()} — DETALHAMENTO"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")

    if macro_v2:
        rsm2 = macro_v2.get("rsm2_mediano", 0)
        total_est = macro_v2.get("total_estimado", 0)
        n_am = macro_v2.get("n_amostras", 0)
        ws["A2"] = (f"Total estimado (mediana de {n_am} similares): "
                    f"R$ {total_est:,.0f}  |  R$/m²: {rsm2:,.2f}").replace(",", ".")
        ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
        ws.merge_cells("A2:H2")
        confianca = "🟢 Alta" if n_am >= 3 else ("🟡 Média" if n_am >= 1 else "🔴 Baixa")
        total_mg = total_est
    else:
        confianca = "🔴 Baixa"
        total_mg = 0

    hdr(ws, 4, ["#", "Descrição", "Un", "Qtd", "PU R$", "Total ref", "Freq", "Fontes"],
        [4, 50, 8, 12, 14, 16, 8, 50])

    if not itens:
        cell(ws, 5, 1, "(sem detalhamento granular nos similares — usar valor agregado acima)",
             bold=False)
        cell(ws, 5, 2, "")
        return {"total": total_mg, "n_itens": 0, "confianca_media": confianca}

    for i, it in enumerate(itens[:30], start=1):
        row = i + 4
        pu = it.get("pu_mediano") or 0
        qtd_med = it.get("qtd_mediana") or 0
        total_med = pu * qtd_med if pu and qtd_med else (it.get("total_mediano") or 0)
        fill = confidence_color(it)

        cell(ws, row, 1, i, align="center", fill=fill)
        cell(ws, row, 2, it.get("descricao", ""), fill=fill)
        cell(ws, row, 3, it.get("unidade", ""), align="center", fill=fill)
        cell(ws, row, 4, qtd_med, fmt='#,##0.00', align="right", fill=fill)
        cell(ws, row, 5, pu, fmt='"R$" #,##0.00', align="right", fill=fill)
        cell(ws, row, 6, total_med, fmt='"R$" #,##0', align="right", fill=fill)
        cell(ws, row, 7, it.get("freq_projetos", 0), align="center", fill=fill)
        cell(ws, row, 8, ", ".join(it.get("fontes", [])[:3]), fill=fill)

    return {"total": total_mg, "n_itens": min(30, len(itens)), "confianca_media": confianca}


def aba_referencias(wb: Workbook, similares: list[dict]) -> None:
    ws = wb.create_sheet("REFERENCIAS")
    ws.sheet_properties.tabColor = "8E44AD"

    ws["A1"] = "PROJETOS DE REFERÊNCIA usados na geração automatizada"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    hdr(ws, 3, ["#", "Projeto", "AC (m²)", "UR", "R$/m²", "Total R$"],
        [4, 38, 14, 8, 14, 18])

    for i, p in enumerate(similares, start=1):
        cell(ws, i + 3, 1, i, align="center")
        cell(ws, i + 3, 2, p["_slug"], bold=True)
        cell(ws, i + 3, 3, p.get("ac"), fmt='#,##0', align="right")
        cell(ws, i + 3, 4, p.get("ur"), align="center")
        cell(ws, i + 3, 5, p.get("rsm2") or 0, fmt='"R$" #,##0.00', align="right")
        cell(ws, i + 3, 6, p.get("total"), fmt='"R$" #,##0', align="right")


def aba_premissas(wb: Workbook, premissas_aprovadas: list[dict], decisoes: dict) -> None:
    ws = wb.create_sheet("PREMISSAS")
    ws.sheet_properties.tabColor = "C0392B"

    ws["A1"] = "PREMISSAS APROVADAS NO GATE"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:C1")

    hdr(ws, 3, ["Decisão / Área", "Valor / Premissa", "Origem"], [30, 60, 22])

    for k, v in (decisoes or {}).items():
        cell(ws, ws.max_row + 1, 1, k, bold=True)
        ws.cell(ws.max_row, 2, v)
        ws.cell(ws.max_row, 3, "Gate validado por Leo")

    for pr in premissas_aprovadas:
        row = ws.max_row + 1
        cell(ws, row, 1, pr.get("area", ""))
        cell(ws, row, 2, pr.get("premissa", ""))
        cell(ws, row, 3, "Base qualitativa")


def gerar_executivo(slug: str, ac: float, ur: int | None, padrao: str | None,
                     gate_path: Path | None, output: str) -> dict:
    similares = cs.projetos_similares(ac=ac, ur=ur, padrao=padrao, n=5)
    if not similares:
        raise RuntimeError("Nenhum projeto similar encontrado na base.")

    gate = parse_gate_xlsx(gate_path) if gate_path and gate_path.exists() else {}

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    valores_calibrados = cs.valores_macrogrupos_calibrados(ac, padrao=padrao)
    valores_similares = cs.valores_macrogrupos_v2(similares, ac)

    macrogrupos_data: dict[str, dict] = {}
    for mg in MACROGRUPOS_CANONICOS:
        enriq = cs.enriquecer_executivo(similares, mg, top_n=40, min_frequency=1)
        itens = enriq.get("itens_agregados", [])
        macro_source = valores_calibrados.get(mg) or valores_similares.get(mg)
        stats = aba_macrogrupo(wb, mg, itens, ac, macro_source)
        macrogrupos_data[mg] = stats

    aba_resumo(wb, slug, ac, ur, similares, macrogrupos_data)
    aba_referencias(wb, similares)
    aba_premissas(wb, gate.get("premissas_aprovadas", []), gate.get("decisoes", {}))

    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    grand_total = sum(d.get("total", 0) for d in macrogrupos_data.values())
    return {
        "output": output,
        "n_similares": len(similares),
        "similares": [p["_slug"] for p in similares],
        "grand_total": round(grand_total, 2),
        "rsm2": round(grand_total / ac, 2) if ac else 0,
        "macrogrupos_com_itens": sum(1 for d in macrogrupos_data.values() if d.get("n_itens", 0) > 0),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slug", required=True)
    ap.add_argument("--ac", type=float, required=True)
    ap.add_argument("--ur", type=int, default=None)
    ap.add_argument("--padrao", default=None)
    ap.add_argument("--gate", default=None, help="caminho do xlsx gate validado")
    ap.add_argument("-o", "--output", default=None)
    args = ap.parse_args()

    output = args.output or f"executivo-{args.slug}.xlsx"
    gate_path = Path(args.gate) if args.gate else None
    result = gerar_executivo(args.slug, args.ac, args.ur, args.padrao, gate_path, output)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
