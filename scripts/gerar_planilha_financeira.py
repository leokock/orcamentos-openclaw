#!/usr/bin/env python3
"""Phase 20f — Gera planilha de Analise Financeira dos Executivos.

12 abas:
 1. LEIA_ME
 2. RESUMO (numeros-chave)
 3. DISTRIBUICAO_MG (% por MG x padrao - Curva S)
 4. RSM2_POR_MG (R$/m² por MG x padrao)
 5. MO_MATERIAL (Split por disciplina x padrao)
 6. CUSTO_INDIRETO (CI/Total por padrao)
 7. QUARTIS_EFICIENCIA (top/bottom quartil por padrao)
 8. TOP_PROJETOS (top 10 eficientes/caros/maiores)
 9. FACHADA (envidracada vs convencional)
10. SISTEMAS_ESPECIAIS (elevadores, gerador, piscina)
11. INSTALACOES (eletricas/hidro/preventivas rsm2)
12. REGRESSAO_RSM2 (analise parametrica)

Uso: python scripts/gerar_planilha_financeira.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
AGG_FILE = BASE / "analise-financeira-agregada.json"
OUT_XLSX = BASE / "analise-financeira-cartesian.xlsx"

DARK, ACCENT, ORANGE, GREEN, PURPLE, RED, GRAY, TEAL, YELLOW_D = (
    "2C3E50", "2980B9", "E67E22", "27AE60", "8E44AD", "C0392B", "7F8C8D", "16A085", "F39C12"
)

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

GREEN_FILL = PatternFill(start_color="E8F8E0", end_color="E8F8E0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
RED_FILL = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")

TAB_COLORS = {
    "LEIA_ME": GREEN, "RESUMO": DARK, "DISTRIBUICAO_MG": ORANGE,
    "RSM2_POR_MG": ACCENT, "MO_MATERIAL": PURPLE, "CUSTO_INDIRETO": RED,
    "QUARTIS_EFICIENCIA": TEAL, "TOP_PROJETOS": YELLOW_D,
    "FACHADA": GRAY, "SISTEMAS_ESPECIAIS": PURPLE,
    "INSTALACOES": TEAL, "REGRESSAO_RSM2": DARK,
}


def _load(p: Path) -> dict:
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def style_header(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def write_headers(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        style_header(c, color)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, nf=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if nf and i in nf:
            c.number_format = nf[i]
        elif isinstance(v, float):
            c.number_format = "#,##0.00"
        elif isinstance(v, int):
            c.number_format = "#,##0"


# ═══════════════════════════════════════════════════════════════════════
def build_leia_me(wb, data):
    ws = wb.create_sheet("LEIA_ME")
    ws.sheet_properties.tabColor = TAB_COLORS["LEIA_ME"]
    ws["A1"] = "ANALISE FINANCEIRA - Orcamentos Executivos Cartesian"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    ws["A2"] = (f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')} . "
                f"{data['n_projetos']} projetos . {data['n_com_total']} com total . {data['n_com_rsm2']} com rsm2")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")

    row = 4
    ws.cell(row, 1, "FOCO DESTA PLANILHA").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    body = ("Analise FINANCEIRA (R$, % do total, R$/m2). Complementar a 'analise-produto' (indicadores fisicos). "
            "Base: 126 orcamentos executivos entregues. "
            "Filtro de validade: AC >= 1000 m2 e R$/m² entre 500 e 10.000 (elimina projetos com AC de 1 unidade apenas).")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 1).font = Font(size=9, name="Arial"); row += 3

    ws.cell(row, 1, "MAPA DAS ABAS").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 2
    write_headers(ws, row, ["#", "Aba", "Descricao"], [4, 22, 70]); row += 1
    abas = [
        ("1", "LEIA_ME", "Este guia"),
        ("2", "RESUMO", "Numeros-chave: R$/m2 por padrao, top eficientes/caros, mg com mais variancia"),
        ("3", "DISTRIBUICAO_MG", "% de cada MG no total, por padrao (Curva S reconstruida)"),
        ("4", "RSM2_POR_MG", "R$/m2 por MG x padrao (comparacao absoluta)"),
        ("5", "MO_MATERIAL", "Split % MO x % Material por disciplina x padrao"),
        ("6", "CUSTO_INDIRETO", "Custo Indireto (CI) / Total % por padrao"),
        ("7", "QUARTIS_EFICIENCIA", "Top 25% eficientes vs Bottom 25% caros por padrao"),
        ("8", "TOP_PROJETOS", "Top 10 mais eficientes (R$/m2 baixo) e top 10 mais caros"),
        ("9", "FACHADA", "Fachada envidracada vs convencional - custo / AC"),
        ("10", "SISTEMAS_ESPECIAIS", "Elevadores (qtd, PU), gerador, piscina"),
        ("11", "INSTALACOES", "Eletricas/Hidrossanitarias/Preventivas rsm2 por padrao"),
        ("12", "REGRESSAO_RSM2", "Correlacoes R$/m² com AC, UR, padrao"),
    ]
    for num, aba, desc in abas:
        write_row(ws, row, [num, aba, desc]); row += 1

    row += 1
    ws.cell(row, 1, "DECISOES METODOLOGICAS").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 2
    meth = [
        "1. Canonicalizacao de 147 variacoes de nomes de MG em 20 MG canonicos (ex: 'SUPRAESTRUTURA' = 'Supraestrutura').",
        "2. Filtro AC >= 1000 m2: elimina projetos com AC definido como uma unidade (gera R$/m² absurdo).",
        "3. Filtro R$/m² entre 500 e 10000: elimina outliers de tipo de projeto (fora da faixa normal).",
        "4. Stats padrao: percentis (p10/p25/med/p75/p90), CV, n.",
        "5. % MO + % Material nem sempre soma 100%: podem ter custos que nao sao nem MO nem material puro.",
        "6. Regressao: Pearson r entre R$/m² e AC/UR. Nao e modelo preditivo, so correlacao.",
    ]
    for m in meth:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, m).font = Font(size=9, name="Arial"); row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40


def build_resumo(wb, data):
    ws = wb.create_sheet("RESUMO")
    ws.sheet_properties.tabColor = TAB_COLORS["RESUMO"]
    ws["A1"] = "RESUMO EXECUTIVO - Analise Financeira"
    ws["A1"].font = Font(bold=True, size=13, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")

    row = 3
    ws.cell(row, 1, "R$/m² POR PADRAO (mediana)").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    regr = data.get("regressao_rsm2", {})
    med_pad = regr.get("mediana_rsm2_por_padrao", {})
    write_headers(ws, row, ["Padrao", "R$/m² mediana", "Projetos validos"], [16, 18, 18]); row += 1
    for pad in ["economico", "medio", "medio-alto", "alto"]:
        v = med_pad.get(pad)
        q = data.get("quartis_eficiencia", {}).get(pad, {})
        n = q.get("n_projetos", "-")
        write_row(ws, row, [pad, v, n], {2: "R$ #,##0"}); row += 1

    row += 2
    ws.cell(row, 1, "RATIO CARO/EFICIENTE POR PADRAO").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    write_headers(ws, row, ["Padrao", "N", "Mediana", "Ratio caro/eficiente"], [16, 6, 18, 22]); row += 1
    for pad, q in data.get("quartis_eficiencia", {}).items():
        write_row(ws, row, [pad, q["n_projetos"], q["mediana_rsm2"], q["ratio_caro_eficiente"]],
                 {3: "R$ #,##0", 4: "0.00"}); row += 1

    row += 2
    ws.cell(row, 1, "CORRELACOES-CHAVE").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    write_headers(ws, row, ["Variavel", "r (Pearson)", "Interpretacao"], [22, 14, 55]); row += 1
    kvs = [
        ("R$/m² vs AC (escala)", regr.get("correlacao_rsm2_vs_ac"),
         regr.get("interpretacao", {}).get("ac_effect", "")),
        ("R$/m² vs UR", regr.get("correlacao_rsm2_vs_ur") or "N/D",
         "Correlacao com n de unidades residenciais"),
    ]
    for k, r, interp in kvs:
        write_row(ws, row, [k, r, interp], {2: "0.000"}); row += 1

    row += 2
    ws.cell(row, 1, "CUSTO INDIRETO POR PADRAO").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    write_headers(ws, row, ["Padrao", "N", "CI/Total med %", "p25-p75 %"], [16, 6, 18, 20]); row += 1
    for pad, ci in data.get("custo_indireto_por_padrao", {}).items():
        s = ci["stats_pct"]
        write_row(ws, row, [pad, s["n"], s["mediana"], f"{s['p25']:.1f}-{s['p75']:.1f}"],
                 {3: "0.0"}); row += 1

    row += 2
    ws.cell(row, 1, "FACHADA: ENVIDRACADA vs CONVENCIONAL").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    fa = data.get("fachada_analysis", {})
    write_headers(ws, row, ["Tipo", "N projetos", "R$/m² AC mediana"], [22, 14, 20]); row += 1
    ev_s = fa.get("stats_envidracada_rsm2")
    cv_s = fa.get("stats_convencional_rsm2")
    write_row(ws, row, ["Envidracada", len(fa.get("com_envidracada", [])),
                        ev_s.get("mediana") if ev_s else None], {3: "R$ #,##0.00"}); row += 1
    write_row(ws, row, ["Convencional", len(fa.get("sem_envidracada", [])),
                        cv_s.get("mediana") if cv_s else None], {3: "R$ #,##0.00"}); row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = [20, 14, 20, 22, 18][col-1]


def build_distribuicao_mg(wb, data):
    ws = wb.create_sheet("DISTRIBUICAO_MG")
    ws.sheet_properties.tabColor = TAB_COLORS["DISTRIBUICAO_MG"]
    ws["A1"] = "DISTRIBUICAO % DE CUSTO POR MACROGRUPO (Curva S reconstruida)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")
    ws["A2"] = "% de cada MG no total do projeto, agregado (mediana) por padrao construtivo."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:H2")

    headers = ["Padrao", "Macrogrupo", "N", "min %", "p25 %", "mediana %", "p75 %", "max %", "CV"]
    widths = [14, 38, 6, 10, 10, 12, 10, 10, 8]
    write_headers(ws, 4, headers, widths)
    row = 5
    nf = {i: "0.00" for i in range(4, 9)}
    nf[9] = "0.000"
    dist = data.get("distribuicao_mg_por_padrao", {})
    for padrao in ["alto", "medio-alto", "medio", "economico", "insuficiente", "desconhecido"]:
        mgs = dist.get(padrao, {})
        if not mgs:
            continue
        ordered = sorted(mgs.items(), key=lambda x: -x[1].get("mediana", 0))
        for canon, s in ordered:
            write_row(ws, row, [padrao, canon, s["n"], s.get("min"), s.get("p25"),
                                s.get("mediana"), s.get("p75"), s.get("max"), s.get("cv")], nf)
            row += 1
    if row > 5:
        ws.auto_filter.ref = f"A4:I{row-1}"
    ws.freeze_panes = "A5"


def build_rsm2_por_mg(wb, data):
    ws = wb.create_sheet("RSM2_POR_MG")
    ws.sheet_properties.tabColor = TAB_COLORS["RSM2_POR_MG"]
    ws["A1"] = "R$/m² POR MACROGRUPO x PADRAO"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")

    headers = ["Padrao", "Macrogrupo", "N", "min", "p25", "mediana", "p75", "max", "CV"]
    widths = [14, 38, 6, 12, 12, 12, 12, 12, 8]
    write_headers(ws, 3, headers, widths)
    row = 4
    nf = {i: "R$ #,##0.00" for i in range(4, 9)}
    nf[9] = "0.000"
    rs = data.get("rsm2_por_mg_por_padrao", {})
    for padrao in ["alto", "medio-alto", "medio", "economico", "insuficiente", "desconhecido"]:
        mgs = rs.get(padrao, {})
        if not mgs:
            continue
        ordered = sorted(mgs.items(), key=lambda x: -x[1].get("mediana", 0))
        for canon, s in ordered:
            write_row(ws, row, [padrao, canon, s["n"], s.get("min"), s.get("p25"),
                                s.get("mediana"), s.get("p75"), s.get("max"), s.get("cv")], nf)
            row += 1
    if row > 4:
        ws.auto_filter.ref = f"A3:I{row-1}"
    ws.freeze_panes = "A4"


def build_mo_material(wb, data):
    ws = wb.create_sheet("MO_MATERIAL")
    ws.sheet_properties.tabColor = TAB_COLORS["MO_MATERIAL"]
    ws["A1"] = "SPLIT % MO (Mao-de-Obra) x % Material - Por disciplina x padrao"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:G1")
    ws["A2"] = "MO % alta = oportunidade de empreitada/terceirizacao. Material % alta = foco em negociacao de compra."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:G2")

    headers = ["Padrao", "Disciplina", "N", "% MO (med)", "% Material (med)", "Soma", "Foco"]
    widths = [14, 35, 6, 14, 16, 10, 22]
    write_headers(ws, 4, headers, widths)
    row = 5
    nf = {4: "0.0", 5: "0.0", 6: "0.0"}
    sm = data.get("split_mo_material", {})
    for padrao in ["alto", "medio-alto", "medio", "economico"]:
        discs = sm.get(padrao, {})
        if not discs:
            continue
        for canon, info in sorted(discs.items()):
            mo = info.get("mo_pct", {}) or {}
            mat = info.get("material_pct", {}) or {}
            mo_med = mo.get("mediana") if mo else None
            mat_med = mat.get("mediana") if mat else None
            soma = (mo_med or 0) + (mat_med or 0)
            if mo_med and mat_med:
                if mo_med > mat_med * 1.2:
                    foco = "EMPREITADA (MO alta)"
                elif mat_med > mo_med * 1.5:
                    foco = "COMPRA (material alto)"
                else:
                    foco = "Misto"
            else:
                foco = "-"
            n = (mo.get("n") if mo else 0) or (mat.get("n") if mat else 0)
            write_row(ws, row, [padrao, canon, n, mo_med, mat_med, soma, foco], nf)
            row += 1
    if row > 5:
        ws.auto_filter.ref = f"A4:G{row-1}"
    ws.freeze_panes = "A5"


def build_custo_indireto(wb, data):
    ws = wb.create_sheet("CUSTO_INDIRETO")
    ws.sheet_properties.tabColor = TAB_COLORS["CUSTO_INDIRETO"]
    ws["A1"] = "CUSTO INDIRETO (CI) / TOTAL % por projeto"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"] = "CI inclui gerenciamento, projetos/consultorias, taxas, equipe administrativa, licencas."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")

    headers = ["Padrao", "Projeto", "CI/Total %", "CI total (R$)"]
    widths = [14, 38, 14, 18]
    write_headers(ws, 4, headers, widths)
    row = 5
    nf = {3: "0.00", 4: "R$ #,##0"}
    ci = data.get("custo_indireto_por_padrao", {})
    for padrao in ["alto", "medio-alto", "medio", "economico"]:
        info = ci.get(padrao, {})
        if not info:
            continue
        for amo in sorted(info.get("amostras", []), key=lambda x: -x["pct"]):
            write_row(ws, row, [padrao, amo["slug"], amo["pct"], amo["ci_total"]], nf); row += 1
    if row > 5:
        ws.auto_filter.ref = f"A4:D{row-1}"
    ws.freeze_panes = "A5"


def build_quartis(wb, data):
    ws = wb.create_sheet("QUARTIS_EFICIENCIA")
    ws.sheet_properties.tabColor = TAB_COLORS["QUARTIS_EFICIENCIA"]
    ws["A1"] = "QUARTIS DE EFICIENCIA - R$/m² no mesmo padrao"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")
    ws["A2"] = "TOP quartil = 25% mais eficientes (R$/m² mais baixo). BOTTOM = 25% mais caros. Ambos no mesmo padrao construtivo."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    headers = ["Padrao", "Quartil", "Projeto", "R$/m²", "AC (m²)", "UR"]
    widths = [14, 14, 38, 14, 12, 8]
    write_headers(ws, 4, headers, widths)
    row = 5
    nf = {4: "R$ #,##0", 5: "#,##0", 6: "#,##0"}
    for padrao, info in data.get("quartis_eficiencia", {}).items():
        for p in info.get("top_quartil_eficientes", []):
            c = ws.cell(row, 2, "TOP eficiente"); c.font = Font(size=9, color=GREEN, bold=True, name="Arial"); c.border = THIN
            write_row(ws, row, [padrao, None, p["slug"], p["rsm2"], p["ac"], p["ur"]], nf)
            ws.cell(row, 2, "TOP eficiente").font = Font(size=9, color=GREEN, bold=True, name="Arial")
            row += 1
        for p in info.get("bottom_quartil_caros", []):
            write_row(ws, row, [padrao, "BOTTOM caro", p["slug"], p["rsm2"], p["ac"], p["ur"]], nf)
            ws.cell(row, 2, "BOTTOM caro").font = Font(size=9, color=RED, bold=True, name="Arial")
            row += 1
    if row > 5:
        ws.auto_filter.ref = f"A4:F{row-1}"
    ws.freeze_panes = "A5"


def build_top_projetos(wb, data):
    ws = wb.create_sheet("TOP_PROJETOS")
    ws.sheet_properties.tabColor = TAB_COLORS["TOP_PROJETOS"]
    ws["A1"] = "TOP 10 PROJETOS (filtro: AC >= 1000 m²)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    t = data.get("top_bottom_projetos", {})

    row = 3
    ws.cell(row, 1, "TOP 10 MAIS EFICIENTES (menor R$/m²)").font = Font(bold=True, size=11, color=GREEN, name="Arial")
    row += 1
    write_headers(ws, row, ["Projeto", "Padrao", "R$/m²", "AC (m²)"], [38, 14, 14, 14])
    row += 1
    nf = {3: "R$ #,##0", 4: "#,##0"}
    for p in t.get("top_10_mais_eficientes_rsm2", []):
        write_row(ws, row, [p["slug"], p["padrao"], p["rsm2"], p["ac"]], nf); row += 1

    row += 2
    ws.cell(row, 1, "TOP 10 MAIS CAROS (maior R$/m²)").font = Font(bold=True, size=11, color=RED, name="Arial")
    row += 1
    write_headers(ws, row, ["Projeto", "Padrao", "R$/m²", "AC (m²)"], [38, 14, 14, 14])
    row += 1
    for p in t.get("top_10_mais_caros_rsm2", []):
        write_row(ws, row, [p["slug"], p["padrao"], p["rsm2"], p["ac"]], nf); row += 1

    row += 2
    ws.cell(row, 1, "TOP 10 MAIORES (maior AC)").font = Font(bold=True, size=11, color=ACCENT, name="Arial")
    row += 1
    write_headers(ws, row, ["Projeto", "Padrao", "AC (m²)", "R$/m²"], [38, 14, 14, 14])
    row += 1
    for p in t.get("top_10_maiores_ac", []):
        write_row(ws, row, [p["slug"], p["padrao"], p["ac"], p.get("rsm2")], {3: "#,##0", 4: "R$ #,##0"})
        row += 1


def build_fachada(wb, data):
    ws = wb.create_sheet("FACHADA")
    ws.sheet_properties.tabColor = TAB_COLORS["FACHADA"]
    ws["A1"] = "FACHADA - Envidracada vs Convencional"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    fa = data.get("fachada_analysis", {})
    row = 3
    ws.cell(row, 1, "FACHADA ENVIDRACADA").font = Font(bold=True, size=11, color=ACCENT, name="Arial"); row += 1
    headers = ["Projeto", "Padrao", "AC", "Fach total (R$)", "Fach R$/m² AC", "% pele"]
    widths = [35, 14, 12, 18, 18, 10]
    write_headers(ws, row, headers, widths); row += 1
    nf = {3: "#,##0", 4: "R$ #,##0", 5: "R$ #,##0.00", 6: "0.0"}
    for p in fa.get("com_envidracada", []):
        write_row(ws, row, [p["slug"], p["padrao"], p["ac"], p["fach_total"],
                            p["fach_rsm2"], p["pele_pct"]], nf); row += 1
    if not fa.get("com_envidracada"):
        ws.cell(row, 1, "(nenhum projeto com pele de vidro dominante detectado)").font = Font(italic=True, size=9, color="999999", name="Arial"); row += 1

    row += 2
    ws.cell(row, 1, "FACHADA CONVENCIONAL (reboco+pintura)").font = Font(bold=True, size=11, color=ORANGE, name="Arial"); row += 1
    write_headers(ws, row, headers, widths); row += 1
    for p in fa.get("sem_envidracada", []):
        write_row(ws, row, [p["slug"], p["padrao"], p["ac"], p["fach_total"],
                            p["fach_rsm2"], p["pele_pct"]], nf); row += 1


def build_sist_especiais(wb, data):
    ws = wb.create_sheet("SISTEMAS_ESPECIAIS")
    ws.sheet_properties.tabColor = TAB_COLORS["SISTEMAS_ESPECIAIS"]
    ws["A1"] = "SISTEMAS ESPECIAIS - Elevador, Gerador, Piscina"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")

    sist = data.get("sistemas_especiais", {})

    row = 3
    ws.cell(row, 1, "ELEVADORES - qtd e PU por projeto").font = Font(bold=True, size=11, color=ACCENT, name="Arial"); row += 1
    s = sist.get("elevador_pu_stats")
    if s:
        ws.cell(row, 1, f"Mediana PU: R$ {s['mediana']:,.0f} por elevador  |  n={s['n']}  |  faixa: R$ {s['min']:,.0f} - R$ {s['max']:,.0f}").font = Font(size=9, name="Arial")
        row += 2
    write_headers(ws, row, ["Projeto", "Padrao", "Qtd elevadores", "PU por elevador (R$)", "Valor total (R$)"], [35, 14, 14, 20, 20]); row += 1
    nf = {3: "#,##0", 4: "R$ #,##0", 5: "R$ #,##0"}
    for e in sist.get("elevadores", []):
        write_row(ws, row, [e["slug"], e["padrao"], e["qtd"], e["pu_un"], e["valor"]], nf); row += 1

    row += 2
    ws.cell(row, 1, f"GERADOR - {sist.get('n_com_gerador', 0)} projetos").font = Font(bold=True, size=11, color=PURPLE, name="Arial"); row += 1
    gs = sist.get("gerador_valor_stats")
    if gs:
        ws.cell(row, 1, f"Mediana: R$ {gs['mediana']:,.0f}  |  faixa: R$ {gs['min']:,.0f} - R$ {gs['max']:,.0f}").font = Font(size=9, name="Arial")
        row += 1

    row += 2
    ws.cell(row, 1, f"PISCINA - {sist.get('n_com_piscina', 0)} projetos").font = Font(bold=True, size=11, color=TEAL, name="Arial"); row += 1
    ps = sist.get("piscina_valor_stats")
    if ps:
        ws.cell(row, 1, f"Mediana: R$ {ps['mediana']:,.0f}  |  faixa: R$ {ps['min']:,.0f} - R$ {ps['max']:,.0f}").font = Font(size=9, name="Arial")


def build_instalacoes(wb, data):
    ws = wb.create_sheet("INSTALACOES")
    ws.sheet_properties.tabColor = TAB_COLORS["INSTALACOES"]
    ws["A1"] = "INSTALACOES - Breakdown por disciplina x padrao (R$/m² AC)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")

    headers = ["Padrao", "Tipo", "N", "min", "p25", "mediana", "p75", "max"]
    widths = [14, 22, 6, 12, 12, 14, 12, 12]
    write_headers(ws, 3, headers, widths)
    row = 4
    nf = {i: "R$ #,##0.00" for i in range(4, 9)}
    ib = data.get("instalacoes_breakdown", {})
    for padrao in ["alto", "medio-alto", "medio", "economico"]:
        for tipo, s in ib.get(padrao, {}).items():
            write_row(ws, row, [padrao, tipo, s["n"], s["min"], s["p25"],
                                s["mediana"], s["p75"], s["max"]], nf); row += 1
    if row > 4:
        ws.auto_filter.ref = f"A3:H{row-1}"
    ws.freeze_panes = "A4"


def build_regressao(wb, data):
    ws = wb.create_sheet("REGRESSAO_RSM2")
    ws.sheet_properties.tabColor = TAB_COLORS["REGRESSAO_RSM2"]
    ws["A1"] = "REGRESSAO - R$/m² vs AC e UR"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    regr = data.get("regressao_rsm2", {})

    row = 3
    ws.cell(row, 1, "CORRELACOES DE PEARSON (n >= 2)").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 2
    write_headers(ws, row, ["Variavel 1", "Variavel 2", "r", "Leitura"], [18, 18, 10, 40]); row += 1
    corr_ac = regr.get("correlacao_rsm2_vs_ac")
    corr_ur = regr.get("correlacao_rsm2_vs_ur")
    write_row(ws, row, ["R$/m²", "AC (escala)", corr_ac,
                        regr.get("interpretacao", {}).get("ac_effect", "")], {3: "0.000"}); row += 1
    write_row(ws, row, ["R$/m²", "UR", corr_ur, "-"], {3: "0.000"}); row += 1

    row += 2
    ws.cell(row, 1, "MEDIANA R$/m² POR PADRAO").font = Font(bold=True, size=11, color=DARK, name="Arial"); row += 1
    write_headers(ws, row, ["Padrao", "Mediana R$/m²", "Delta vs alto %"], [18, 18, 18]); row += 1
    med = regr.get("mediana_rsm2_por_padrao", {})
    alto = med.get("alto", 1)
    for pad in ["economico", "medio", "medio-alto", "alto"]:
        v = med.get(pad)
        if v is None:
            continue
        delta = (v / alto - 1) * 100 if alto else 0
        write_row(ws, row, [pad, v, delta], {2: "R$ #,##0", 3: "+0.0;-0.0;0"}); row += 1


def main():
    print("Loading financial data...", flush=True)
    data = _load(AGG_FILE)
    if not data:
        print(f"ERROR: {AGG_FILE} not found")
        return
    print(f"  {data['n_projetos']} projects", flush=True)

    wb = Workbook()
    wb.remove(wb.active)
    print("Building tabs...", flush=True)
    build_leia_me(wb, data); print("  LEIA_ME")
    build_resumo(wb, data); print("  RESUMO")
    build_distribuicao_mg(wb, data); print("  DISTRIBUICAO_MG")
    build_rsm2_por_mg(wb, data); print("  RSM2_POR_MG")
    build_mo_material(wb, data); print("  MO_MATERIAL")
    build_custo_indireto(wb, data); print("  CUSTO_INDIRETO")
    build_quartis(wb, data); print("  QUARTIS_EFICIENCIA")
    build_top_projetos(wb, data); print("  TOP_PROJETOS")
    build_fachada(wb, data); print("  FACHADA")
    build_sist_especiais(wb, data); print("  SISTEMAS_ESPECIAIS")
    build_instalacoes(wb, data); print("  INSTALACOES")
    build_regressao(wb, data); print("  REGRESSAO_RSM2")

    wb.save(str(OUT_XLSX))
    print(f"\nSalvo: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
