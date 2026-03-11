#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-
"""
Gerador de Orçamento Paramétrico - Armínio Tavares (Atualizado)
Incorpora informações das reuniões do projeto
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# Dados do projeto (extraídos do briefing atualizado)
PROJETO = {
    "nome": "Armínio Tavares",
    "cliente": "PLACON Empreendimentos",
    "codigo": "ARQ 24.027",
    "localizacao": "Rua Dr. Armínio Tavares - Centro - Florianópolis/SC",
    "ac": 7996.45,  # m²
    "at": 486.40,   # m²
    "ur": 45,       # unidades
    "np": 16,       # pavimentos
    "subsolos": 2,
    "data": datetime.now().strftime("%d/%m/%Y")
}

# Medianas base (R$/m², CUB dez/23 = R$ 2.752,67)
MEDIANAS_BASE = {
    "Gerenciamento": 391.96,
    "Mov. Terra": 15.90,
    "Infraestrutura": 212.40,
    "Supraestrutura": 693.58,
    "Alvenaria": 145.84,
    "Impermeabilização": 53.18,
    "Instalações": 349.88,
    "Sist. Especiais": 30.49,
    "Climatização": 0.00,
    "Rev. Int. Parede": 182.84,
    "Teto": 75.42,
    "Pisos": 182.42,
    "Pintura": 75.32,
    "Esquadrias": 340.20,
    "Louças e Metais": 98.64,
    "Fachada": 247.02,
    "Complementares": 64.25,
    "Imprevistos": 100.00
}

# Fatores de ajuste baseados no briefing atualizado
FATORES_AJUSTE = {
    "Gerenciamento": 1.00,
    "Mov. Terra": 1.00,
    "Infraestrutura": 1.10,  # 2 subsolos + centro urbano
    "Supraestrutura": 1.08,  # 16 pavimentos
    "Alvenaria": 1.00,
    "Impermeabilização": 1.15,  # 2 subsolos + cortina
    "Instalações": 1.08,  # Barramento blindado + complexidade
    "Sist. Especiais": 1.05,  # SPDA + automação
    "Climatização": 1.00,
    "Rev. Int. Parede": 1.02,  # Forro em dois níveis
    "Teto": 1.05,  # Forro 2,50m e 2,40m (especial)
    "Pisos": 1.00,
    "Pintura": 1.00,
    "Esquadrias": 1.00,
    "Louças e Metais": 1.00,
    "Fachada": 1.00,
    "Complementares": 1.05,  # Logística centro urbano
    "Imprevistos": 1.00
}

# CUB atualizado (mar/2026 estimado)
CUB_BASE = 2752.67  # dez/23
CUB_ATUAL = 3100.00  # mar/2026 (estimado)
FATOR_CUB = CUB_ATUAL / CUB_BASE  # 1.126

def criar_planilha():
    """Cria planilha de orçamento paramétrico"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ORÇAMENTO PARAMÉTRICO"
    
    # Estilos
    titulo_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    titulo_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    subtitulo_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subtitulo_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Configurar largura de colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    
    # CABEÇALHO
    row = 1
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "ORÇAMENTO PARAMÉTRICO - ARMÍNIO TAVARES"
    cell.fill = titulo_fill
    cell.font = titulo_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados do Projeto
    row += 2
    ws[f'A{row}'] = "Projeto:"
    ws[f'B{row}'] = PROJETO['nome']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Cliente:"
    ws[f'B{row}'] = PROJETO['cliente']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Código:"
    ws[f'B{row}'] = PROJETO['codigo']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Localização:"
    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'] = PROJETO['localizacao']
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Data:"
    ws[f'B{row}'] = PROJETO['data']
    ws[f'A{row}'].font = Font(bold=True)
    
    # Parâmetros
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "PARÂMETROS DO PROJETO"
    cell.fill = subtitulo_fill
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    params = [
        ("Área Construída (AC)", f"{PROJETO['ac']:.2f} m²"),
        ("Área Terreno (AT)", f"{PROJETO['at']:.2f} m²"),
        ("Unidades (UR)", f"{PROJETO['ur']} un"),
        ("Pavimentos (NP)", f"{PROJETO['np']} pav"),
        ("Subsolos", f"{PROJETO['subsolos']} níveis"),
    ]
    
    for param, valor in params:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = valor
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # CUB
    row += 1
    ws[f'A{row}'] = "CUB Base (dez/23):"
    ws[f'B{row}'] = f"R$ {CUB_BASE:,.2f}/m²"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "CUB Atual (mar/26):"
    ws[f'B{row}'] = f"R$ {CUB_ATUAL:,.2f}/m²"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Fator CUB:"
    ws[f'B{row}'] = f"{FATOR_CUB:.3f}"
    ws[f'A{row}'].font = Font(bold=True)
    
    # Orçamento por Macrogrupo
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "ORÇAMENTO POR MACROGRUPO"
    cell.fill = subtitulo_fill
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    headers = ["Macrogrupo", "Base R$/m²", "Fator", "R$/m²", "Valor Total", "% Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Calcular custos
    custos = {}
    total_valor = 0
    
    for macro, base in MEDIANAS_BASE.items():
        fator = FATORES_AJUSTE[macro]
        rs_m2 = base * FATOR_CUB * fator
        valor_total = rs_m2 * PROJETO['ac']
        custos[macro] = {
            'base': base,
            'fator': fator,
            'rs_m2': rs_m2,
            'valor': valor_total
        }
        total_valor += valor_total
    
    # Preencher linhas
    row += 1
    start_row = row
    for macro in MEDIANAS_BASE.keys():
        c = custos[macro]
        ws[f'A{row}'] = macro
        ws[f'B{row}'] = c['base']
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        ws[f'C{row}'] = c['fator']
        ws[f'C{row}'].number_format = '0.00'
        ws[f'D{row}'] = c['rs_m2']
        ws[f'D{row}'].number_format = 'R$ #,##0.00'
        ws[f'E{row}'] = c['valor']
        ws[f'E{row}'].number_format = 'R$ #,##0.00'
        ws[f'F{row}'] = c['valor'] / total_valor
        ws[f'F{row}'].number_format = '0.0%'
        
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # TOTAL
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'E{row}'] = total_valor
    ws[f'E{row}'].number_format = 'R$ #,##0.00'
    ws[f'E{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'] = 1.0
    ws[f'F{row}'].number_format = '0.0%'
    ws[f'F{row}'].font = Font(bold=True, size=12)
    
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = header_fill
        ws.cell(row=row, column=col).border = border
    
    # Indicadores
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "INDICADORES"
    cell.fill = subtitulo_fill
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    indicadores = [
        ("Valor Total", f"R$ {total_valor:,.2f}"),
        ("Custo/m²", f"R$ {total_valor/PROJETO['ac']:,.2f}"),
        ("CUBs", f"{(total_valor/PROJETO['ac'])/CUB_ATUAL:.2f}"),
        ("Custo/Unidade", f"R$ {total_valor/PROJETO['ur']:,.2f}"),
    ]
    
    for ind, valor in indicadores:
        ws[f'A{row}'] = ind
        ws[f'B{row}'] = valor
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Observações
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "OBSERVAÇÕES"
    cell.fill = subtitulo_fill
    cell.font = subtitulo_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 1
    obs = [
        "• Orçamento atualizado com base em reuniões do projeto (mar/2026)",
        "• Fatores de ajuste aplicados:",
        "  - Instalações: +8% (barramento blindado)",
        "  - Teto: +5% (forro em dois níveis: 2,50m e 2,40m)",
        "  - Infraestrutura: +10% (2 subsolos + contenção cortina)",
        "  - Supraestrutura: +8% (16 pavimentos)",
        "  - Impermeabilização: +15% (2 subsolos)",
        "• Base de calibração: 58 projetos (dez/2023)",
        "• Precisão estimada: ±20%",
        "• NÃO inclui: terreno, projetos, aprovações, mobiliário",
        "• Unidades já vendidas - limitação para alterações de layout",
    ]
    
    for ob in obs:
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = ob
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Salvar
    output_dir = "/Users/leokock/orcamentos/output"
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{output_dir}/arminio-tavares-orcamento-atualizado-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    wb.save(filename)
    
    return filename, total_valor, total_valor/PROJETO['ac']

if __name__ == "__main__":
    arquivo, total, custo_m2 = criar_planilha()
    print("=" * 70)
    print("ORÇAMENTO PARAMÉTRICO ATUALIZADO - ARMÍNIO TAVARES")
    print("=" * 70)
    print(f"\nValor Total: R$ {total:,.2f}")
    print(f"Custo/m²: R$ {custo_m2:,.2f} ({custo_m2/CUB_ATUAL:.2f} CUBs)")
    print(f"Custo/Unidade: R$ {total/PROJETO['ur']:,.2f}")
    print(f"\nArquivo gerado: {arquivo}")
    print("\nAtualizações aplicadas:")
    print("  • Instalações: +8% (barramento blindado)")
    print("  • Teto: +5% (forro em dois níveis)")
    print("  • Infraestrutura: +10% (contenção cortina)")
    print("  • Impermeabilização: +15% (2 subsolos)")
    print("=" * 70)
