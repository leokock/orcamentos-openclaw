#!/usr/bin/env python3
"""
Gerador de Orçamento Paramétrico V2 — Bottom-Up com Dropdowns Interativos

Gera planilha Excel com:
- Aba BRIEFING: 7 dropdowns que o usuário muda
- Aba INDICES: fórmulas IF() que reagem aos dropdowns
- 18 abas de detalhe: Qtd × PU com fórmulas vivas referenciando INDICES
- PAINEL: KPIs automaticos
- BENCHMARK: comparação com projetos similares
- PREMISSAS: documentação das fontes

Uso:
  python gerar_template_dinamico_v2.py --config projeto.json
  python gerar_template_dinamico_v2.py --ac 13200 --ur 136 --np 24

Baseado na revisão da Patricia + calibração 75 executivos (24/mar/2026).
"""
import argparse
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# === PATHS ===
BASE_DIR = os.path.expanduser("~/orcamentos")
CAL_PATH = os.path.join(BASE_DIR, "base", "calibration-indices.json")

# === STYLES ===
DARK = "2C3E50"
ACCENT = "2980B9"
ORANGE = "E67E22"
GREEN = "27AE60"
GREEN_BG = "E8F5E9"
YELLOW_BG = "FFF8E1"
RED_BG = "FDEDEC"
GRAY_BG = "F5F5F5"
INPUT_BG = "FFF3E0"  # orange claro para células editáveis

THIN = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN

def style_input(cell):
    cell.font = Font(bold=True, size=11, name="Arial", color=ORANGE)
    cell.fill = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
    cell.border = THIN

def style_calc(cell):
    cell.font = Font(size=9, name="Arial")
    cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    cell.border = THIN

def style_label(cell):
    cell.font = Font(bold=True, size=9, name="Arial")
    cell.border = THIN

def hdr(ws, r, cols, widths=None):
    for i, t in enumerate(cols, 1):
        c = ws.cell(r, i, t)
        style_header(c)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def dr(ws, r, data, bold=False, shade=None):
    for i, v in enumerate(data, 1):
        c = ws.cell(r, i, v)
        c.font = Font(bold=bold, size=9, name="Arial")
        c.border = THIN
        if isinstance(v, float): c.number_format = '#,##0.00'
        elif isinstance(v, int): c.number_format = '#,##0'
        if shade: c.fill = PatternFill(start_color=shade, end_color=shade, fill_type="solid")


def gerar_parametrico(P, output_path):
    """Generate the full parametric spreadsheet with interactive dropdowns."""

    with open(CAL_PATH) as f:
        CAL = json.load(f)

    ac = P['ac']
    ur = P['ur']
    np_ = P.get('np', 24)
    npt = P.get('npt', max(1, np_ - 5))
    prazo = P.get('prazo', 30)
    cub = P.get('cub', 3028.45)
    nome = P.get('nome', 'Projeto')

    wb = Workbook()

    # ========================================================================
    # ABA 1: DADOS_PROJETO (inputs fixos)
    # ========================================================================
    ws_dados = wb.active
    ws_dados.title = "DADOS_PROJETO"
    ws_dados.sheet_properties.tabColor = ACCENT

    ws_dados['A1'] = f"{nome.upper()} — ORÇAMENTO PARAMÉTRICO V2"
    ws_dados['A1'].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws_dados.merge_cells('A1:D1')

    dados = [
        ("Projeto", nome, 3), ("Cidade/UF", P.get('cidade', ''), 4),
        ("AC Total (m²)", ac, 5), ("UR", ur, 6),
        ("Pavimentos", np_, 7), ("Pav. Tipo", npt, 8),
        ("Elevadores", P.get('elev', 2), 9), ("Vagas", P.get('vag', 100), 10),
        ("Prazo (meses)", prazo, 11), ("CUB Atual (R$)", cub, 12),
    ]
    for label, val, row in dados:
        ws_dados.cell(row, 1, label)
        style_label(ws_dados.cell(row, 1))
        ws_dados.cell(row, 2, val)
        if isinstance(val, (int, float)):
            style_input(ws_dados.cell(row, 2))
            ws_dados.cell(row, 2).number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
    ws_dados.column_dimensions['A'].width = 20
    ws_dados.column_dimensions['B'].width = 25

    # ========================================================================
    # ABA 2: BRIEFING (dropdowns interativos)
    # ========================================================================
    ws_brief = wb.create_sheet("BRIEFING")
    ws_brief.sheet_properties.tabColor = ORANGE

    ws_brief['A1'] = "BRIEFING — ALTERE OS DROPDOWNS PARA RECALCULAR"
    ws_brief['A1'].font = Font(bold=True, size=13, color=DARK, name="Arial")
    ws_brief.merge_cells('A1:D1')
    ws_brief['A2'] = "Cada alteração muda automaticamente os índices e custos nas abas de detalhe"
    ws_brief['A2'].font = Font(size=9, color="666666", name="Arial")

    # Briefing questions with dropdowns
    # Cell B4-B17 will have dropdowns, C column has the impact description
    questions = [
        (4, "Tipo de Laje", "Convencional", ["Convencional", "Protendida", "Nervurada"],
         "Muda: concreto, aço, forma, escoramento, cordoalha"),
        (5, "Subsolos", "0", ["0", "1", "2", "3"],
         "Muda: contenção, mov.terra, impermeabilização, fundação"),
        (6, "Fundação", "Hélice", ["Hélice", "Tubulão", "Sapata"],
         "Muda: custo infraestrutura, tipo de perfuração"),
        (7, "Padrão Acabamento", "Médio-Alto", ["Médio", "Médio-Alto", "Alto", "Luxo"],
         "Muda: pisos, rev.parede, esquadrias, louças, fachada"),
        (8, "Fachada", "Textura", ["Textura", "Cerâmica", "Pele de vidro", "ACM"],
         "Muda: custo fachada (R$/m²)"),
        (9, "Pressurização", "Não", ["Sim", "Não"],
         "Muda: custo sist.especiais (+R$ 80k)"),
        (10, "Nº Torres", "1", ["1", "2", "3"],
         "Muda: gerador, elevadores, equipamentos"),
        (11, "Gerador Dedicado", "Sim", ["Sim", "Não"],
         "Muda: sist.especiais (R$ 120-350k)"),
        (12, "Entrega", "Completa", ["Completa", "Shell"],
         "Shell desconta ~R$ 200/m² (pisos, pintura, rev.parede, louças, alvenaria)"),
        (13, "Tipologia", "Misto", ["Studios", "1-2 Dormitórios", "3-4 Dormitórios", "Misto"],
         "Muda: pontos elétricos, louças, alvenaria interna"),
        (14, "Pé-Direito", "Padrão (3.00)", ["Baixo (2.80)", "Padrão (3.00)", "Alto (3.20)", "Duplo"],
         "Muda: fachada, alvenaria, estrutura (+2-8%)"),
        (15, "Nº Banheiros/Apto", "2", ["1", "2", "3", "4"],
         "Muda: hidro, louças, impermeab., rev.parede"),
        (16, "Tipo Piso Predominante", "Misto", ["Porcelanato", "Laminado", "Misto"],
         "Muda: PU pisos (porcelanato ~R$ 81 vs laminado ~R$ 65)"),
        (17, "Piscina", "Sim", ["Sim", "Não", "Aquecida"],
         "Muda: sist.especiais (R$ 0 / R$ 220k / R$ 320k)"),
    ]

    hdr(ws_brief, 3, ["Pergunta", "Resposta", "Impacto", ""], [30, 18, 45, 5])

    for row, question, default, options, impact in questions:
        ws_brief.cell(row, 1, question)
        style_label(ws_brief.cell(row, 1))
        ws_brief.cell(row, 2, default)
        style_input(ws_brief.cell(row, 2))
        ws_brief.cell(row, 3, impact)
        ws_brief.cell(row, 3).font = Font(size=8, color="666666", name="Arial")

        # Add dropdown validation
        dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=False)
        dv.error = "Selecione uma opção da lista"
        dv.errorTitle = "Valor inválido"
        ws_brief.add_data_validation(dv)
        dv.add(ws_brief.cell(row, 2))

    # ========================================================================
    # ABA 3: INDICES (fórmulas IF que reagem ao BRIEFING)
    # ========================================================================
    ws_idx = wb.create_sheet("INDICES")
    ws_idx.sheet_properties.tabColor = GREEN

    ws_idx['A1'] = "ÍNDICES CALIBRADOS — REAGEM AO BRIEFING"
    ws_idx['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws_idx.merge_cells('A1:E1')
    ws_idx['A2'] = "NÃO EDITAR — valores calculados automaticamente"
    ws_idx['A2'].font = Font(size=9, color="E74C3C", name="Arial")

    hdr(ws_idx, 4, ["Índice", "Valor", "Unidade", "Fonte", "Fórmula"], [35, 12, 12, 25, 50])

    # Reference cells for briefing answers
    # BRIEFING!B4 = Laje, B5 = Subsolos, B6 = Fundação, B7 = Padrão, B8 = Fachada, B9 = Pressurização, B10 = Torres

    idx_rows = {
        'concreto_m3_m2': 5,
        'aco_kg_m3': 6,
        'cordoalha_kg_m2': 7,
        'forma_m2_m3': 8,
        'escoramento_pu': 9,
        'fck_laje': 10,
        'n_estacas_base': 11,
        'comp_estaca': 12,
        'idx_alvenaria': 13,
        'idx_piso': 14,
        'idx_parede': 15,
        'idx_forro': 16,
        'idx_pintura': 17,
        'idx_fachada': 18,
        'idx_impermeab': 19,
        'fator_padrao': 20,
        'fator_fachada_pu': 21,
        'pressur_valor': 22,
        'gerador_valor': 23,
        'contencao_valor': 24,
        'mov_terra_pu': 25,
    }

    # BRIEFING cells: B4=Laje, B5=Subsolos, B6=Fundação, B7=Padrão, B8=Fachada,
    # B9=Pressurização, B10=Torres, B11=Gerador, B12=Entrega, B13=Tipologia,
    # B14=Pé-Direito, B15=BWC/apto, B16=Tipo Piso, B17=Piscina

    indices_def = [
        # --- ESTRUTURA (reagem a Laje B4) ---
        (5, "Concreto (m³/m² AC)",
         '=IF(BRIEFING!B4="Protendida",0.22,IF(BRIEFING!B4="Nervurada",0.20,0.25))',
         "m³/m²", "Master 7p exec", "IF Laje"),
        (6, "Aço convenc. (kg/m³)",
         '=IF(BRIEFING!B4="Protendida",70,IF(BRIEFING!B4="Nervurada",90,106))',
         "kg/m³", "Master 7p+11p narr", "IF Laje"),
        (7, "Cordoalha (kg/m² laje)",
         '=IF(BRIEFING!B4="Protendida",8,0)',
         "kg/m²", "Master protendida", "IF Laje=Protendida"),
        (8, "Forma (m²/m³)",
         '=IF(BRIEFING!B4="Protendida",6.05,IF(BRIEFING!B4="Nervurada",6.5,7.12))',
         "m²/m³", "Master 7p exec", "IF Laje"),
        (9, "Escoramento (R$/m²)",
         '=IF(BRIEFING!B4="Protendida",50,IF(BRIEFING!B4="Nervurada",35,0))',
         "R$/m²", "PU base 7p", "IF Laje"),
        (10, "fck Laje",
         '=IF(BRIEFING!B4="Protendida",40,30)',
         "MPa", "", "IF Laje"),

        # --- FUNDAÇÃO (reagem a Subsolos B5, Fundação B6) ---
        (11, "Nº estacas (base AC)",
         '=ROUND(DADOS_PROJETO!B5/75,0)*(1+VALUE(BRIEFING!B5)*0.15)',
         "un", "~1 est/75m² + sub", "AC/75 × (1+sub×15%)"),
        (12, "Comp. estaca (m)",
         '=IF(BRIEFING!B6="Hélice",18,IF(BRIEFING!B6="Tubulão",12,0))+VALUE(BRIEFING!B5)*2',
         "m", "Tipo + prof.sub", "IF Fundação + sub×2m"),

        # --- ACABAMENTOS (reagem a Padrão B7, Entrega B12, Tipologia B13, Pé-Direito B14, BWC B15, Piso B16) ---
        (13, "Alvenaria (m²/m² AC)",
         '=2.25*IF(BRIEFING!B12="Shell",0.75,1)*IF(BRIEFING!B14="Duplo",1.15,IF(BRIEFING!B14="Alto (3.20)",1.05,1))',
         "m²/m²", "2,25 × shell × pé-dir", "IF Entrega + Pé-Direito"),
        (14, "Pisos (m²/m² AC)",
         '=1.70*IF(BRIEFING!B12="Shell",0.40,1)',
         "m²/m²", "1,70 × shell", "IF Entrega (shell = 40%)"),
        (15, "Rev. Parede (m²/m² AC)",
         '=2.85*IF(BRIEFING!B12="Shell",0.50,1)*IF(VALUE(BRIEFING!B15)>=2,1.15,IF(VALUE(BRIEFING!B15)>=1,1.0,0.85))',
         "m²/m²", "2,85 × shell × bwc", "IF Entrega + BWC/apto"),
        (16, "Forro (m²/m² AC)",
         '=1.16*IF(BRIEFING!B12="Shell",0.45,1)',
         "m²/m²", "1,16 × shell", "IF Entrega (shell = 45%)"),
        (17, "Pintura (m² sup/m² AC)",
         '=5.65*IF(BRIEFING!B12="Shell",0.40,1)',
         "m²/m²", "5,65 × shell", "IF Entrega (shell = 40%)"),
        (18, "Fachada (m²/m² AC)",
         '=IF(BRIEFING!B8="Pele de vidro",1.8,IF(BRIEFING!B8="ACM",1.7,1.55))*IF(BRIEFING!B14="Duplo",1.12,IF(BRIEFING!B14="Alto (3.20)",1.05,1))',
         "m²/m²", "Tipo fachada × pé-dir", "IF Fachada + Pé-Direito"),
        (19, "Impermeab. (m²/m² AC)",
         '=0.37+VALUE(BRIEFING!B5)*0.05+IF(VALUE(BRIEFING!B15)>=3,0.04,IF(VALUE(BRIEFING!B15)>=2,0.03,0))',
         "m²/m²", "Base + sub + bwc", "IF Sub + BWC"),

        # --- FATORES DE AJUSTE ---
        (20, "Fator padrão acabamento",
         '=IF(BRIEFING!B7="Luxo",1.35,IF(BRIEFING!B7="Alto",1.15,IF(BRIEFING!B7="Médio-Alto",1.0,0.85)))',
         "fator", "", "IF Padrão"),
        (21, "PU fachada (R$/m²)",
         '=IF(BRIEFING!B8="Pele de vidro",350,IF(BRIEFING!B8="ACM",280,IF(BRIEFING!B8="Cerâmica",180,100)))',
         "R$/m²", "PU base + briefing", "IF Fachada tipo"),
        (22, "Pressurização (R$)",
         '=IF(BRIEFING!B9="Sim",80000,0)',
         "R$", "Param. base", "IF Sim"),
        (23, "Gerador (R$)",
         '=IF(BRIEFING!B11="Não",0,IF(BRIEFING!B10="3",350000,IF(BRIEFING!B10="2",250000,180000)))',
         "R$", "IF Gerador Sim + Torres", "IF Gerador + Torres"),
        (24, "Contenção (R$)",
         '=IF(VALUE(BRIEFING!B5)>0,VALUE(BRIEFING!B5)*3.5*120*350,0)',
         "R$", "Sub×prof×perim×R$", "IF Sub>0"),
        (25, "Mov. Terra PU (R$/m²)",
         '=IF(VALUE(BRIEFING!B5)>0,17+VALUE(BRIEFING!B5)*15,17)',
         "R$/m²", "Base + sub×15", "IF Sub>0"),

        # --- NOVOS: Tipologia, Piso, Piscina, Pé-Direito, Entrega ---
        (26, "Fator tipologia louças",
         '=IF(BRIEFING!B13="Studios",0.65,IF(BRIEFING!B13="1-2 Dormitórios",0.85,IF(BRIEFING!B13="3-4 Dormitórios",1.20,1.0)))',
         "fator", "Studios menos louças", "IF Tipologia"),
        (27, "Fator tipologia elétrica",
         '=IF(BRIEFING!B13="Studios",0.80,IF(BRIEFING!B13="3-4 Dormitórios",1.15,1.0))',
         "fator", "Studios menos pontos", "IF Tipologia"),
        (28, "Fator entrega (shell)",
         '=IF(BRIEFING!B12="Shell",0.85,1.0)',
         "fator", "Shell desconta ~15% total", "IF Entrega"),
        (29, "PU piso predominante",
         '=IF(BRIEFING!B16="Porcelanato",81.21,IF(BRIEFING!B16="Laminado",65,73))*INDICES!B20',
         "R$/m²", "Tipo piso × padrão", "IF Tipo Piso × Padrão"),
        (30, "Piscina (R$)",
         '=IF(BRIEFING!B17="Aquecida",320000,IF(BRIEFING!B17="Sim",220000,0))',
         "R$", "Param. base", "IF Piscina"),
        (31, "Fator pé-direito estrutura",
         '=IF(BRIEFING!B14="Duplo",1.12,IF(BRIEFING!B14="Alto (3.20)",1.05,IF(BRIEFING!B14="Baixo (2.80)",0.97,1.0)))',
         "fator", "Pé-dir afeta estrutura", "IF Pé-Direito"),
        (32, "Fator BWC louças",
         '=IF(VALUE(BRIEFING!B15)>=4,1.80,IF(VALUE(BRIEFING!B15)>=3,1.50,IF(VALUE(BRIEFING!B15)>=2,1.20,1.0)))',
         "fator", "Mais bwc = mais louças", "IF BWC/apto"),
    ]

    for row, label, formula, unit, fonte, nota in indices_def:
        ws_idx.cell(row, 1, label)
        style_label(ws_idx.cell(row, 1))
        ws_idx.cell(row, 2).value = formula
        style_calc(ws_idx.cell(row, 2))
        ws_idx.cell(row, 2).number_format = '#,##0.00' if 'R$' not in unit else '#,##0'
        ws_idx.cell(row, 3, unit)
        ws_idx.cell(row, 4, fonte)
        ws_idx.cell(row, 4).font = Font(size=8, color="666666", name="Arial")
        ws_idx.cell(row, 5, nota)
        ws_idx.cell(row, 5).font = Font(size=8, color="999999", name="Arial")

    ws_idx.column_dimensions['A'].width = 30
    ws_idx.column_dimensions['B'].width = 14
    ws_idx.column_dimensions['D'].width = 22
    ws_idx.column_dimensions['E'].width = 30

    # ========================================================================
    # ABA 4: CUSTOS_MACROGRUPO (resumo com fórmulas)
    # ========================================================================
    ws_custos = wb.create_sheet("CUSTOS_MACROGRUPO")
    ws_custos.sheet_properties.tabColor = ORANGE

    ws_custos['A1'] = "CUSTOS POR MACROGRUPO — BOTTOM-UP COM BRIEFING"
    ws_custos['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws_custos.merge_cells('A1:E1')

    hdr(ws_custos, 3, ["Macrogrupo", "R$/m²", "% Total", "Custo Total (R$)", "Aba"], [28, 12, 10, 18, 18])

    MG_ORDER = ['Gerenciamento', 'Mov. Terra', 'Infraestrutura', 'Supraestrutura', 'Alvenaria',
                'Impermeabilização', 'Instalações', 'Sist. Especiais', 'Climatização', 'Rev. Int. Parede',
                'Teto', 'Pisos', 'Pintura', 'Esquadrias', 'Louças e Metais', 'Fachada',
                'Complementares', 'Imprevistos']

    # Each detail tab will have its total in cell G2
    # Custos_macrogrupo references those totals
    for i, mg in enumerate(MG_ORDER):
        row = 4 + i
        tab_name = mg[:31]
        ws_custos.cell(row, 1, mg)
        style_label(ws_custos.cell(row, 1))
        # Total from detail tab
        ws_custos.cell(row, 4).value = f"='{tab_name}'!G2"
        style_calc(ws_custos.cell(row, 4))
        ws_custos.cell(row, 4).number_format = '#,##0'
        # R$/m²
        ws_custos.cell(row, 2).value = f'=D{row}/DADOS_PROJETO!B5'
        ws_custos.cell(row, 2).number_format = '#,##0.00'
        style_calc(ws_custos.cell(row, 2))
        # %
        ws_custos.cell(row, 3).value = f'=D{row}/D{4 + len(MG_ORDER)}'
        ws_custos.cell(row, 3).number_format = '0.0%'
        style_calc(ws_custos.cell(row, 3))
        # Aba link
        ws_custos.cell(row, 5, tab_name)

    # Total row
    total_row = 4 + len(MG_ORDER)
    ws_custos.cell(total_row, 1, "TOTAL")
    ws_custos.cell(total_row, 1).font = Font(bold=True, size=10, name="Arial")
    ws_custos.cell(total_row, 4).value = f'=SUM(D4:D{total_row - 1})'
    ws_custos.cell(total_row, 4).font = Font(bold=True)
    ws_custos.cell(total_row, 4).number_format = '#,##0'
    ws_custos.cell(total_row, 2).value = f'=D{total_row}/DADOS_PROJETO!B5'
    ws_custos.cell(total_row, 2).number_format = '#,##0.00'
    ws_custos.cell(total_row, 2).font = Font(bold=True)

    # ========================================================================
    # ABAS DE DETALHE — cada macrogrupo com fórmulas referenciando INDICES
    # ========================================================================
    # Helper: ref to AC, UR, prazo, CUB
    AC_REF = "DADOS_PROJETO!B5"
    UR_REF = "DADOS_PROJETO!B6"
    NP_REF = "DADOS_PROJETO!B7"
    NPT_REF = "DADOS_PROJETO!B8"
    PRAZO_REF = "DADOS_PROJETO!B11"
    CUB_REF = "DADOS_PROJETO!B12"

    def make_detail_tab(wb, mg_name, items):
        """Create a detail tab. items = list of (desc, desc2, qtd_formula, unit, pu_formula, fonte)"""
        tab = mg_name[:31]
        ws = wb.create_sheet(tab)
        ws.sheet_properties.tabColor = ORANGE

        ws['A1'] = mg_name.upper()
        ws['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")
        ws.merge_cells('A1:E1')
        # G2 = total (referenced by CUSTOS_MACROGRUPO)
        # Will be =SUM of column F
        ws['G1'] = "Total:"
        ws['G1'].font = Font(bold=True, size=9, name="Arial")

        hdr(ws, 3, ["Item", "Descrição", "Qtd", "Un", "PU (R$)", "Total (R$)", "Fonte"],
            [20, 25, 12, 6, 14, 16, 30])

        for i, (desc, desc2, qtd_f, unit, pu_f, fonte) in enumerate(items):
            row = 4 + i
            ws.cell(row, 1, desc)
            ws.cell(row, 2, desc2)
            # Qtd (formula or static)
            if isinstance(qtd_f, str) and qtd_f.startswith('='):
                ws.cell(row, 3).value = qtd_f
                style_calc(ws.cell(row, 3))
            else:
                ws.cell(row, 3, qtd_f)
            ws.cell(row, 3).number_format = '#,##0'
            ws.cell(row, 4, unit)
            # PU (formula or static)
            if isinstance(pu_f, str) and pu_f.startswith('='):
                ws.cell(row, 5).value = pu_f
                style_calc(ws.cell(row, 5))
            else:
                ws.cell(row, 5, pu_f)
            ws.cell(row, 5).number_format = '#,##0.00'
            # Total = Qtd × PU
            ws.cell(row, 6).value = f'=C{row}*E{row}'
            ws.cell(row, 6).number_format = '#,##0'
            style_calc(ws.cell(row, 6))
            # Fonte
            ws.cell(row, 7, fonte)
            ws.cell(row, 7).font = Font(size=7, color="666666", name="Arial")

            for c in range(1, 8):
                ws.cell(row, c).border = THIN

        # Total row
        last_row = 4 + len(items)
        ws.cell(last_row, 1, "TOTAL")
        ws.cell(last_row, 1).font = Font(bold=True)
        ws.cell(last_row, 6).value = f'=SUM(F4:F{last_row - 1})'
        ws.cell(last_row, 6).font = Font(bold=True)
        ws.cell(last_row, 6).number_format = '#,##0'

        # G2 = total (for CUSTOS_MACROGRUPO reference)
        ws['G2'] = f'=F{last_row}'
        ws['G2'].number_format = '#,##0'

        return ws

    # --- SUPRAESTRUTURA ---
    supra_items = [
        ("Concreto Pilares", "fck30", f'=ROUND(INDICES!B5*{AC_REF}*0.22*INDICES!B31,0)', "m³", 590, "Índ×AC×22% × pé-dir (7p)"),
        ("Concreto Vigas", "fck30",
         f'=ROUND(INDICES!B5*{AC_REF}*IF(BRIEFING!B4="Protendida",0.08,0.30),0)', "m³", 590, "IF Laje→peso vigas"),
        ("Concreto Lajes", "fck conforme laje",
         f'=ROUND(INDICES!B5*{AC_REF}*IF(BRIEFING!B4="Protendida",0.60,0.35),0)', "m³",
         '=IF(INDICES!B10=40,690,590)', "IF fck→PU"),
        ("Concreto Escadas", "fck30", f'=ROUND(INDICES!B5*{AC_REF}*0.05,0)', "m³", 590, "Índ×AC×5%"),
        ("Concreto Blocos", "Baldrame fck30", f'=ROUND(INDICES!B5*{AC_REF}*0.05,0)', "m³", 590, "Índ×AC×5%"),
        ("Aço CA-50", "Convencional", f'=ROUND(INDICES!B6*INDICES!B5*{AC_REF},0)', "kg", 8.67, f"Índ aço×conc×AC"),
        ("Cordoalha CP-190", "Protensão", f'=ROUND(INDICES!B7*{AC_REF}*0.65,0)', "kg", 22, "IF Prot→8kg/m² laje"),
        ("Forma", "Compensado plastif.", f'=ROUND(INDICES!B8*INDICES!B5*{AC_REF},0)', "m²", 88.07, "Índ forma×conc×AC"),
        ("Escoramento", "Metálico", f'=ROUND({AC_REF}*0.65,0)', "m²", '=INDICES!B9', "IF Laje→PU escor"),
        ("MO Estrutura", "Empreitada",
         f'=ROUND(SUM(F4:F12)*0.326/0.674/{AC_REF},0)*{AC_REF}/{AC_REF}', "m²",
         f'=ROUND(SUM(F4:F12)*0.326/0.674/{AC_REF},2)', "Split 32,6% (10p)"),
    ]
    # Fix MO: need simpler formula
    supra_items[-1] = ("MO Estrutura", "Empreitada (32,6%)", f'={AC_REF}', "m²",
                        f'=ROUND((SUM(F4:F12)*0.326/0.674)/{AC_REF},2)', "Split 32,6% MO (10p)")
    make_detail_tab(wb, "Supraestrutura", supra_items)

    # --- INFRAESTRUTURA ---
    infra_items = [
        ("Perfuração hélice", "Ø40cm", '=ROUND(INDICES!B11*INDICES!B12,0)', "m", 82, "PU base R$ 82/m (6p)"),
        ("Concreto estacas", "fck30", '=ROUND(INDICES!B11*3.14159*0.04*INDICES!B12,0)', "m³", 632, "PU base R$ 632 (6p)"),
        ("Concreto blocos", "Baldrames", f'=ROUND(0.02*{AC_REF},0)', "m³", 590, "Índ 0,02 m³/m²"),
        ("Aço CA-50", "Fundação", '=ROUND(90*(C5+C6),0)', "kg", 8.77, "PU base R$ 8,77 (7p)"),
        ("Forma", "Blocos/baldrames", f'=ROUND(0.10*{AC_REF},0)', "m²", 45.64, "PU base R$ 45,64 (8p)"),
        ("Arrasamento", "Estacas", '=INDICES!B11', "un", 65, "PU base R$ 65 (2p)"),
        ("Limpeza/remoção", "Solo perfuração", 1, "vb", 35000, "PU base R$ 35k (2p)"),
        ("MO fundação", "Empreitada (35,5%)", 1, "vb", '=ROUND(SUM(F4:F10)*0.355/0.645,0)', "Split 35,5% (7p)"),
        ("Contenção", "Se houver subsolo", 1, "vb", '=INDICES!B24', "IF Sub>0"),
    ]
    make_detail_tab(wb, "Infraestrutura", infra_items)

    # --- MOV. TERRA ---
    make_detail_tab(wb, "Mov. Terra", [
        ("Terraplanagem", "Corte+aterro+bota-fora", f'={AC_REF}', "m²", '=INDICES!B25', "Param. base + IF sub"),
    ])

    # --- GERENCIAMENTO ---
    ger_items = [
        ("Projetos", "Arq+compl+compat", 1, "vb", 650000, "Param. base"),
        ("Consultorias", "ATP+compat BIM", 1, "vb", 180000, "Param. base"),
        ("Ensaios", "Ctrl tecnol.", 1, "vb", 130000, "Param. base"),
        ("Taxas e seguros", "Incorp+licenças", 1, "vb", 380000, "Param. base"),
        ("Engenheiro PJ", "1×prazo", f'={PRAZO_REF}', "mês", 10500, "Master R$ 10,5k (5p)"),
        ("Mestre obras", "1×prazo", f'={PRAZO_REF}', "mês", 9940, "Master R$ 9,9k (4p)"),
        ("Encarregado", "1×prazo", f'={PRAZO_REF}', "mês", 8000, "Narrativo R$ 8k"),
        ("Estagiário", "2×prazo", f'={PRAZO_REF}*2', "mês", 1580, "Master R$ 1,6k (3p)"),
        ("Téc.segurança", "1×prazo", f'={PRAZO_REF}', "mês", 6000, "PJ R$ 6k/mês"),
        ("Almoxarife", "1×prazo", f'={PRAZO_REF}', "mês", 3377, "Master R$ 3,4k (4p)"),
        ("Limpeza obra", "2×prazo", f'={PRAZO_REF}*2', "mês", 2500, "PJ R$ 2,5k/mês"),
        ("Vigilância", "1×prazo", f'={PRAZO_REF}', "mês", 15261, "Narrativo R$ 15,3k"),
        ("EPCs", "Fixo por porte", 1, "vb", f'=IF({AC_REF}<5000,150000,IF({AC_REF}<15000,300000,500000))', "Master 75 exec"),
        ("EPI", "Individual", f'={PRAZO_REF}', "mês", 908, "PU base R$ 908 (8p)"),
        ("Meio ambiente", "PCMAT+PCMSO", 1, "vb", 45000, "Param. base"),
        ("Op. inicial", "Mobilização", 1, "vb", 55000, "Param. base"),
        ("Inst. provisórias", "Canteiro+containers", 1, "vb", 160000, "Param. base"),
        ("Desp. consumo", "Água+en+internet+IPTU", f'={PRAZO_REF}', "mês", 11500, "PU base (9p)"),
        ("Equipamentos", "Grua+cremalheira", 1, "vb", 480000, "Narrativo"),
        ("Comunic.visual", "Placas+tapume", 1, "vb", 47000, "PU base R$ 47k (17p)"),
    ]
    make_detail_tab(wb, "Gerenciamento", ger_items)

    # --- ALVENARIA ---
    make_detail_tab(wb, "Alvenaria", [
        ("Bloco cerâm.", "Vedação 14cm", f'=ROUND(INDICES!B13*{AC_REF}*0.40,0)', "m²", 32.95, "Índ 2,25×40%"),
        ("Drywall ST", "Internas", f'=ROUND(INDICES!B13*{AC_REF}*0.25,0)', "m²", 156, "Índ 2,25×25%"),
        ("Drywall RU", "Molhadas", f'=ROUND(INDICES!B13*{AC_REF}*0.08,0)', "m²", 195, "Índ 2,25×8%"),
        ("Argamassa", "Assentamento", f'=ROUND(INDICES!B13*{AC_REF}*0.40,0)', "m²", 3.80, "PU base"),
        ("MO Alvenaria", "Empreitada", f'=ROUND(INDICES!B13*{AC_REF}*0.40,0)', "m²", 28.50, "Split 35,7% (9p)"),
    ])

    # --- IMPERMEABILIZAÇÃO ---
    make_detail_tab(wb, "Impermeabilização", [
        ("Manta 4mm", "Material", f'=ROUND(INDICES!B19*{AC_REF},0)', "m²", 82.11, "Índ impermeab×AC"),
        ("Arg.polim.", "BWC", f'=ROUND(INDICES!B19*{AC_REF}*0.3,0)', "m²", 8.50, "PU base"),
        ("Regulariz.", "Superfície", f'=ROUND(INDICES!B19*{AC_REF},0)', "m²", 5.57, "PU base (15p)"),
        ("MO imper.", "Empreitada", f'=ROUND(INDICES!B19*{AC_REF},0)', "m²", 68, "Split 56,5% (6p)"),
        ("MO regul.", "Empreitada", f'=ROUND(INDICES!B19*{AC_REF},0)', "m²", 19.50, "PU base (7p)"),
    ])

    # --- INSTALAÇÕES (reage a Tipologia B13 via fator elétrica INDICES!B27) ---
    fcub_formula = f'({CUB_REF}/2752.67)'
    make_detail_tab(wb, "Instalações", [
        ("Elétricas mat.", "IF Tipologia", f'={AC_REF}', "m²", f'=ROUND(130*{fcub_formula}/1.1*0.618*INDICES!B27,2)', "Índ 1,77m/m² × fator tipol."),
        ("Elétricas MO", "Empreitada", f'={AC_REF}', "m²", f'=ROUND(130*{fcub_formula}/1.1*0.382*INDICES!B27,2)', "Split 38,2% × fator tipol."),
        ("Hidro mat.", "IF BWC/apto", f'={AC_REF}', "m²", f'=ROUND(145*{fcub_formula}/1.1*0.66*IF(VALUE(BRIEFING!B15)>=2,1.15,1),2)', "Índ 1,08m/m² × IF BWC"),
        ("Hidro MO", "Empreitada", f'={AC_REF}', "m²", f'=ROUND(145*{fcub_formula}/1.1*0.34*IF(VALUE(BRIEFING!B15)>=2,1.15,1),2)', "Split 34,0% × IF BWC"),
        ("Preventivas", "Mat+MO", f'={AC_REF}', "m²", f'=ROUND(45*{fcub_formula}/1.1,2)', "Param."),
        ("Gás", "Mat+MO", f'={AC_REF}', "m²", f'=ROUND(18*{fcub_formula}/1.1,2)', "Param. base"),
        ("Telecom", "Mat+MO", f'={AC_REF}', "m²", f'=ROUND(15*{fcub_formula}/1.1,2)', "Param. base"),
    ])

    # --- SIST. ESPECIAIS (reage a Gerador B11, Piscina B17, Pressurização B9, Torres B10) ---
    make_detail_tab(wb, "Sist. Especiais", [
        ("Elevador social", "N-1 elevadores", f'=MAX(1,DADOS_PROJETO!B9-1)', "un", 192450, "PU base (5p)"),
        ("Elevador serv.", "1un", 1, "un", 96000, "PU base (5p)"),
        ("Gerador", "IF Gerador+Torres", 1, "vb", '=INDICES!B23', "IF Gerador B11 + Torres B10"),
        ("Automação", "BMS+sensores", 1, "vb", 120000, "Param. base"),
        ("Piscina", "IF Piscina B17", 1, "vb", '=INDICES!B30', "IF Piscina (Sim/Não/Aquecida)"),
        ("CFTV", "Câmeras+DVR", 1, "vb", 120000, "Param. base"),
        ("Pressurização", "IF Pressurização B9", 1, "vb", '=INDICES!B22', "IF Pressurização"),
        ("SPDA", "Para-raios+malha", 1, "vb", 55000, "Param. base"),
        ("Interfonia", "Ctrl acesso", 1, "vb", 130000, "Param. base"),
        ("Bombas", "Recalque+incêndio", 1, "vb", 95000, "Param. base"),
        ("Quadros comando", "Bombas+sist", 1, "vb", 60000, "Param. base"),
    ])

    # --- CLIMATIZAÇÃO ---
    make_detail_tab(wb, "Climatização", [
        ("Split", "1,5/apto", f'=ROUND({UR_REF}*1.5,0)', "un", 3500, "Param. 1,5/apto"),
        ("Exaust.BWC", "4/pav enclaus.", f'=4*{NPT_REF}', "un", 1200, "4 bwc/pav"),
        ("Exaust.churr.", "2/pav+lazer", f'=2*{NPT_REF}+2', "un", 2500, "2/pav+lazer"),
        ("Infra AR", "Dutos+tub", f'={AC_REF}', "m²", 10, "PU base"),
    ])

    # --- REV. INT. PAREDE ---
    make_detail_tab(wb, "Rev. Int. Parede", [
        ("Reboco", "Massa única", f'=ROUND(INDICES!B15*{AC_REF}*0.35,0)', "m²", 7, "Índ 2,85×35%"),
        ("Chapisco", "Rolado+colante", f'=ROUND(INDICES!B15*{AC_REF}*0.35,0)', "m²", 5.50, "PU base"),
        ("Cerâm.BWC", "30×60", f'=ROUND(INDICES!B15*{AC_REF}*0.14,0)', "m²", '=48*INDICES!B20', "PU×fator padrão"),
        ("Cerâm.cozinha", "30×60", f'=ROUND(INDICES!B15*{AC_REF}*0.06,0)', "m²", '=42*INDICES!B20', "PU×fator padrão"),
        ("Porcel.parede", "Comuns", f'=ROUND(INDICES!B15*{AC_REF}*0.04,0)', "m²", '=85*INDICES!B20', "PU×fator padrão"),
        ("Granito bancadas", "Lavat+coz", f'=ROUND({UR_REF}*3.5,0)', "m", '=180*INDICES!B20', "R$ 180×fator padrão"),
        ("Arg.colante", "AC-II/III", f'=ROUND(INDICES!B15*{AC_REF}*0.24,0)', "m²", 12, "PU base"),
        ("Rejunte", "Flexível", f'=ROUND(INDICES!B15*{AC_REF}*0.24,0)', "m²", 5, "PU base"),
        ("MO reboco", "Empreitada", f'=ROUND(INDICES!B15*{AC_REF}*0.35,0)', "m²", 22.50, "Split 54,4%"),
        ("MO cerâmica", "Assentamento", f'=ROUND(INDICES!B15*{AC_REF}*0.24,0)', "m²", 35, "Split 54,4%"),
        ("MO granito", "Instalação", f'=ROUND({UR_REF}*3.5,0)', "m", 35.90, "PU base R$ 36"),
    ])

    # --- TETO ---
    make_detail_tab(wb, "Teto", [
        ("Forro ST", "Acartonado", f'=ROUND(INDICES!B16*{AC_REF}*0.70,0)', "m²", 28, "Índ 1,16×70%"),
        ("Forro RU", "BWC", f'=ROUND(INDICES!B16*{AC_REF}*0.15,0)', "m²", 35, "PU base"),
        ("Perfis", "Estrutura", f'=ROUND(INDICES!B16*{AC_REF}*0.85,0)', "m²", 15, "PU base"),
        ("MO forro", "Empreitada", f'=ROUND(INDICES!B16*{AC_REF},0)', "m²", 25, "Split 79,1%"),
    ])

    # --- PISOS (reage a Entrega B12 via INDICES!B14, Tipo Piso B16 via INDICES!B29, Padrão B7) ---
    make_detail_tab(wb, "Pisos", [
        ("Contrapiso", "Autonivelante", f'=ROUND(INDICES!B14*{AC_REF},0)', "m²", 12.10, "Índ piso×AC (IF shell)"),
        ("Piso principal", "IF Tipo Piso", f'=ROUND(INDICES!B14*{AC_REF}*0.50,0)', "m²", '=INDICES!B29', "PU IF Tipo Piso × Padrão"),
        ("Cimentado", "Garagem 20%", f'=ROUND(INDICES!B14*{AC_REF}*0.20,0)', "m²", 32, "PU base"),
        ("Granito", "Halls 5%", f'=ROUND(INDICES!B14*{AC_REF}*0.05,0)', "m²", '=180*INDICES!B20', "PU×fator padrão"),
        ("Rodapé", "Poliestireno", f'=ROUND(INDICES!B14*{AC_REF}*0.35,0)', "m", 10.77, "PU base (10p)"),
        ("MO contrapiso", "Empreitada", f'=ROUND(INDICES!B14*{AC_REF},0)', "m²", 18, "Split 54%"),
        ("MO pisos", "Assentamento", f'=ROUND(INDICES!B14*{AC_REF}*0.70,0)', "m²", 32, "Split 54%"),
    ])

    # --- PINTURA ---
    make_detail_tab(wb, "Pintura", [
        ("Massa PVA", "Paredes 2dem", f'=ROUND(INDICES!B17*{AC_REF}*0.55,0)', "m²", 4.29, "Índ pintura×55%"),
        ("Acrílica par.", "3 demãos", f'=ROUND(INDICES!B17*{AC_REF}*0.55,0)', "m²", 5.40, "PU base (6p)"),
        ("Acrílica teto", "2 demãos", f'=ROUND(INDICES!B17*{AC_REF}*0.20,0)', "m²", 3.83, "PU base"),
        ("Selador", "Base", f'=ROUND(INDICES!B17*{AC_REF}*0.75,0)', "m²", 2.50, "PU base"),
        ("MO pintura", "Empreitada", f'=ROUND(INDICES!B17*{AC_REF},0)', "m²", 15, "Split 66,2%"),
        ("MO lixamento", "Preparação", f'=ROUND(INDICES!B17*{AC_REF}*0.55,0)', "m²", 8, "Split 66,2%"),
    ])

    # --- ESQUADRIAS (reage a Padrão B7 e Entrega B12) ---
    make_detail_tab(wb, "Esquadrias", [
        ("Esquadrias Al", "Jan+portas", f'={AC_REF}', "m²", f'=280*INDICES!B20*INDICES!B28', "R$280 × padrão × entrega"),
        ("Serralheria", "GC+corrimão", f'={AC_REF}', "m²", f'=45*INDICES!B20', "R$45 × fator padrão"),
    ])

    # --- LOUÇAS E METAIS (reage a Tipologia B13, BWC B15, Padrão B7, Entrega B12) ---
    make_detail_tab(wb, "Louças e Metais", [
        ("Louças+metais", "IF Tipol+BWC+Padrão+Entrega", f'={UR_REF}', "apto",
         f'=2800*INDICES!B20*INDICES!B26*INDICES!B32*INDICES!B28',
         "R$2800 × padrão × tipol × bwc × entrega"),
    ])

    # --- FACHADA ---
    make_detail_tab(wb, "Fachada", [
        ("Revestimento", "Conforme briefing", f'=ROUND(INDICES!B18*{AC_REF},0)', "m²", '=INDICES!B21', "IF tipo fachada"),
        ("MO fachada", "Empreitada", f'=ROUND(INDICES!B18*{AC_REF},0)', "m²", 35, "PU base"),
        ("Balancim", "Fachadeiro", f'=ROUND(INDICES!B18*{AC_REF},0)', "m²", 18, "PU base"),
    ])

    # --- COMPLEMENTARES ---
    make_detail_tab(wb, "Complementares", [
        ("Mobiliário", "Áreas comuns", 1, "vb", 540000, "PU base R$ 540k (5p)"),
        ("Ambientação", "Decoração halls", 1, "vb", f'=300000*INDICES!B20', "R$ 300k×fator padrão"),
        ("Paisagismo", "", 1, "vb", 120000, "Param. base"),
        ("Equip. lazer", "Acad+playground", 1, "vb", 180000, "PU base (8p)"),
        ("Limpeza final", "", f'={AC_REF}', "m²", 12, "PU base R$ 12/m² (10p)"),
        ("Comunic.visual", "Sinalização", 1, "vb", 47000, "PU base R$ 47k (17p)"),
        ("Cobertura", "Telhado+estrutura", f'=ROUND({AC_REF}*0.04,0)', "m²", 95, "PU base"),
        ("Ligações def.", "Água+energia+gás", 1, "vb", 60000, "Param. base"),
        ("Desmobilização", "Pessoal+equip", 1, "vb", 15000, "PU base (6p)"),
        ("Paviment.ext.", "Calçadas", f'=ROUND({AC_REF}*0.06,0)', "m²", 80, "PU base"),
    ])

    # --- IMPREVISTOS (referencia SUM D4:D20, sem incluir a si mesmo D21) ---
    imp_sum_range = f"CUSTOS_MACROGRUPO!D4:D{4 + len(MG_ORDER) - 2}"  # All except Imprevistos
    make_detail_tab(wb, "Imprevistos", [
        ("Imprevistos", "1,5% do subtotal",
         1, "vb",
         f"=ROUND(SUM({imp_sum_range})*0.015,0)",
         "Percentual padrão"),
    ])

    # ========================================================================
    # ABA PAINEL (KPIs com fórmulas)
    # ========================================================================
    ws_painel = wb.create_sheet("PAINEL", 0)  # First tab
    ws_painel.sheet_properties.tabColor = GREEN

    ws_painel['A1'] = f"{nome.upper()} — PARAMÉTRICO V2"
    ws_painel['A1'].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws_painel.merge_cells('A1:C1')
    ws_painel['A2'] = "Bottom-up + Dropdowns interativos | 75 executivos calibrados"
    ws_painel['A2'].font = Font(size=9, color="666666", name="Arial")

    kpis = [
        (4, "Custo Total", f'=CUSTOS_MACROGRUPO!D{4+len(MG_ORDER)}', '#,##0'),
        (5, "R$/m² AC", f'=CUSTOS_MACROGRUPO!D{4+len(MG_ORDER)}/{AC_REF}', '#,##0.00'),
        (6, "CUB Ratio", f'=B5/{CUB_REF}', '0.00'),
        (7, "Custo/UR", f'=CUSTOS_MACROGRUPO!D{4+len(MG_ORDER)}/{UR_REF}', '#,##0'),
        (8, "AC", f'={AC_REF}', '#,##0'),
        (9, "UR", f'={UR_REF}', '#,##0'),
        (10, "CUB", f'={CUB_REF}', '#,##0.00'),
        (11, "Laje", '=BRIEFING!B4', '@'),
        (12, "Subsolos", '=BRIEFING!B5', '@'),
        (13, "Padrão", '=BRIEFING!B7', '@'),
        (14, "Fachada", '=BRIEFING!B8', '@'),
    ]
    for row, label, formula, fmt in kpis:
        ws_painel.cell(row, 1, label)
        ws_painel.cell(row, 1).font = Font(bold=True, size=11, name="Arial")
        ws_painel.cell(row, 2).value = formula
        ws_painel.cell(row, 2).font = Font(size=11, name="Arial")
        ws_painel.cell(row, 2).number_format = fmt
        style_calc(ws_painel.cell(row, 2))
    ws_painel.column_dimensions['A'].width = 18
    ws_painel.column_dimensions['B'].width = 22

    # ========================================================================
    # ABA PREMISSAS
    # ========================================================================
    ws_n = wb.create_sheet("PREMISSAS")
    premissas = [
        f"PARAMÉTRICO V2 — {nome}",
        "18/18 macrogrupos bottom-up | 14 dropdowns interativos",
        "Altere os dropdowns na aba BRIEFING para recalcular automaticamente",
        "",
        "FONTES: calibration-indices.json (75 exec + 79 narrativos + IFC + curva ABC)",
        "PUs: base-pus-cartesian.json (1.504 itens, medianas de 75 executivos)",
        "Splits MO/Material: medianas reais de 38 executivos detalhados",
        "",
        "14 DROPDOWNS ATIVOS:",
        "  1. Laje: Convencional/Protendida/Nervurada → concreto, aço, forma, escoramento",
        "  2. Subsolos: 0-3 → contenção, mov.terra, impermeabilização, fundação",
        "  3. Fundação: Hélice/Tubulão/Sapata → infraestrutura",
        "  4. Padrão: Médio/Médio-Alto/Alto/Luxo → acabamentos, esquadrias, louças (×0.85-1.35)",
        "  5. Fachada: Textura/Cerâmica/Pele de vidro/ACM → PU fachada R$ 100-350/m²",
        "  6. Pressurização: Sim/Não → +R$ 80k",
        "  7. Torres: 1/2/3 → gerador proporcional",
        "  8. Gerador dedicado: Sim/Não → R$ 0-350k",
        "  9. Entrega: Completa/Shell → Shell desconta ~15% (pisos, pintura, rev, louças, esq)",
        "  10. Tipologia: Studios/1-2D/3-4D/Misto → pontos elétricos, louças (×0.65-1.20)",
        "  11. Pé-Direito: Baixo/Padrão/Alto/Duplo → estrutura (×0.97-1.12), fachada",
        "  12. BWC/apto: 1-4 → hidro, louças (×1.0-1.80), rev.parede, impermeab",
        "  13. Tipo Piso: Porcelanato/Laminado/Misto → PU pisos R$ 65-81/m²",
        "  14. Piscina: Sim/Não/Aquecida → R$ 0/220k/320k",
        "",
        "Baseado revisão Patricia (coord. custos) + calibração 75 exec — 24/mar/2026",
    ]
    for i, p in enumerate(premissas, 2):
        ws_n.cell(i, 1, p).font = Font(size=9, name="Arial")
    ws_n.column_dimensions['A'].width = 80

    # Save
    wb.save(output_path)
    print(f"Gerado: {output_path}")
    print(f"Abas: PAINEL + DADOS_PROJETO + BRIEFING (7 dropdowns) + INDICES + CUSTOS_MACROGRUPO + 18 detalhe + PREMISSAS")
    print(f"Altere BRIEFING para recalcular automaticamente!")


# === CLI ===
def main():
    parser = argparse.ArgumentParser(description='Gerar Paramétrico V2 (Bottom-Up + Dropdowns)')
    parser.add_argument('--config', help='JSON com dados do projeto')
    parser.add_argument('--nome', default='Projeto', help='Nome do projeto')
    parser.add_argument('--ac', type=float, help='Área construída (m²)')
    parser.add_argument('--ur', type=int, help='Unidades residenciais')
    parser.add_argument('--np', type=int, default=24, help='Nº pavimentos')
    parser.add_argument('--npt', type=int, help='Nº pavimentos tipo')
    parser.add_argument('--elev', type=int, default=2, help='Nº elevadores')
    parser.add_argument('--vag', type=int, default=100, help='Nº vagas')
    parser.add_argument('--prazo', type=int, default=30, help='Prazo (meses)')
    parser.add_argument('--cub', type=float, default=3028.45, help='CUB atual')
    parser.add_argument('--cidade', default='', help='Cidade/UF')
    parser.add_argument('-o', '--output', help='Arquivo de saída (.xlsx)')
    args = parser.parse_args()

    if args.config:
        with open(args.config) as f:
            P = json.load(f)
    else:
        if not args.ac or not args.ur:
            parser.error("--ac e --ur são obrigatórios (ou use --config)")
        P = {
            'nome': args.nome, 'ac': args.ac, 'ur': args.ur, 'np': args.np,
            'npt': args.npt or max(1, args.np - 5), 'elev': args.elev, 'vag': args.vag,
            'prazo': args.prazo, 'cub': args.cub, 'cidade': args.cidade,
        }

    output = args.output or f"{P.get('nome', 'projeto').replace(' ', '-')}-Parametrico-V2.xlsx"
    gerar_parametrico(P, output)


if __name__ == '__main__':
    main()
