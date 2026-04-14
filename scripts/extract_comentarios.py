#!/usr/bin/env python3
"""Phase 8 — Extração exaustiva de comentários e texto livre dos xlsx.

Extrai:
1. cell.comment.text de todas as células com comentário
2. Células de texto livre (>25 chars) fora das colunas de descrição padrão
3. Notas e observações em abas dedicadas (Observações, Notas, Comentários)

Saída: base/comentarios-completos/[projeto].json
"""
from __future__ import annotations

import json
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
QUEUE = BASE / "gemma-queue.json"
OUT_DIR = BASE / "comentarios-completos"
OUT_DIR.mkdir(parents=True, exist_ok=True)
LOG = BASE / "phase8-extract.log.jsonl"

MIN_TEXT_LEN = 25
SKIP_SHEET_PATTERNS = [
    r"^cap[ai]", r"^gr[áa]fic", r"^[íi]ndice$", r"^sum[áa]rio$",
    r"^instr", r"^legenda", r"^plan\d+$",
]


def should_skip(name: str) -> bool:
    n = name.strip().lower()
    return any(re.search(p, n) for p in SKIP_SHEET_PATTERNS)


def log_event(e: dict) -> None:
    e.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
    with LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(e, ensure_ascii=False) + "\n")


def extract_xlsx(xlsx_path: Path) -> dict:
    out = {
        "arquivo": str(xlsx_path),
        "abas": [],
        "total_comentarios": 0,
        "total_textos_livres": 0,
        "errors": [],
    }
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        out["errors"].append(f"open: {type(e).__name__}: {str(e)[:80]}")
        return out

    for sn in wb.sheetnames:
        if should_skip(sn):
            continue
        try:
            ws = wb[sn]
        except Exception as e:
            out["errors"].append(f"sheet {sn}: {e}")
            continue

        aba_data = {
            "nome": sn,
            "comentarios": [],
            "textos_livres": [],
        }

        seen_textos = set()

        rows_iter = 0
        for row in ws.iter_rows():
            rows_iter += 1
            if rows_iter > 5000:
                break
            for cell in row:
                if cell.comment and cell.comment.text:
                    text = cell.comment.text.strip()
                    if text:
                        aba_data["comentarios"].append({
                            "cell": cell.coordinate,
                            "author": cell.comment.author,
                            "text": text[:500],
                        })
                        out["total_comentarios"] += 1

                if isinstance(cell.value, str):
                    text = cell.value.strip()
                    if len(text) >= MIN_TEXT_LEN:
                        if text.lower() not in seen_textos:
                            seen_textos.add(text.lower())
                            aba_data["textos_livres"].append({
                                "cell": cell.coordinate,
                                "text": text[:500],
                            })
                            out["total_textos_livres"] += 1
                            if len(aba_data["textos_livres"]) > 500:
                                break
            if len(aba_data["textos_livres"]) > 500:
                break

        if aba_data["comentarios"] or aba_data["textos_livres"]:
            out["abas"].append(aba_data)

    try:
        wb.close()
    except Exception:
        pass
    return out


def pick_heaviest_xlsx(sources: dict) -> str | None:
    files = sources.get("xlsx", [])
    if not files:
        return None
    sized = []
    for p in files:
        try:
            sized.append((Path(p).stat().st_size, p))
        except Exception:
            pass
    sized.sort(reverse=True)
    return sized[0][1] if sized else None


def process(project: dict) -> dict:
    slug = project["projeto"]
    summary = {"projeto": slug, "status": "pending"}
    t0 = time.time()

    xlsx = pick_heaviest_xlsx(project.get("sources", {}))
    if not xlsx:
        summary["status"] = "no_xlsx"
        summary["duration_s"] = round(time.time() - t0, 2)
        log_event(summary)
        return summary

    try:
        result = extract_xlsx(Path(xlsx))
        result["projeto"] = slug
        (OUT_DIR / f"{slug}.json").write_text(
            json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        summary["status"] = "done"
        summary["abas"] = len(result["abas"])
        summary["comentarios"] = result["total_comentarios"]
        summary["textos_livres"] = result["total_textos_livres"]
    except Exception as e:
        summary["status"] = "failed"
        summary["error"] = f"{type(e).__name__}: {e}"

    summary["duration_s"] = round(time.time() - t0, 2)
    log_event(summary)
    return summary


def main():
    queue = json.loads(QUEUE.read_text(encoding="utf-8"))
    print(f"processing {len(queue)} projects -> {OUT_DIR}")
    counts = {}
    t0 = time.time()
    for i, p in enumerate(queue, 1):
        s = process(p)
        counts[s["status"]] = counts.get(s["status"], 0) + 1
        if s["status"] == "done":
            print(f"[{i}/{len(queue)}] {p['projeto']:<45} done com={s['comentarios']:>4} txt={s['textos_livres']:>5} {s['duration_s']}s", flush=True)
    print()
    print(f"summary: {counts}")
    print(f"total time: {int(time.time() - t0)}s")


if __name__ == "__main__":
    main()
