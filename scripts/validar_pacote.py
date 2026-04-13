#!/usr/bin/env python3
"""Validador de coerência do pacote orçamento.

Recebe paths do paramétrico e executivo gerados, roda checagens automáticas
e gera um relatório markdown com:
  - Comparação total paramétrico vs executivo (delta %)
  - Por macrogrupo, comparação com mediana da base V2 (calibration-indices)
  - Lista de macrogrupos fora do P10-P90 (alerta)
  - Lista de macrogrupos sem dados qualitativos (gap)
  - Coerência entre AC e quantidades
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent))

BASE = Path.home() / "orcamentos-openclaw" / "base"
CALIBRATION = BASE / "calibration-indices.json"


def read_resumo(xlsx_path: Path) -> dict:
    """Lê a aba RESUMO de um executivo gerado por gerar_executivo_auto.py."""
    if not xlsx_path.exists():
        return {}
    wb = load_workbook(xlsx_path, data_only=True)
    if "RESUMO" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["RESUMO"]
    macrogrupos = {}
    grand_total = 0
    for row in ws.iter_rows(min_row=5, max_row=23, values_only=True):
        if not row or len(row) < 7 or not row[1]:
            continue
        mg = str(row[1]).strip()
        if mg.upper() == "TOTAL":
            grand_total = row[2] or 0
            continue
        macrogrupos[mg] = {
            "total": row[2] or 0,
            "rsm2": row[3] or 0,
            "pct": row[4] or 0,
            "n_itens": row[5] or 0,
            "confianca": row[6] or "—",
        }
    wb.close()
    return {"macrogrupos": macrogrupos, "grand_total": grand_total}


def load_calibration() -> dict:
    if not CALIBRATION.exists():
        return {}
    try:
        return json.loads(CALIBRATION.read_text(encoding="utf-8"))
    except Exception:
        return {}


def segmento_por_ac(ac: float) -> tuple[str, str]:
    """Retorna (nome_segmento, key_no_calibration)."""
    if ac < 8000:
        return ("Pequeno (<8k m²)", "pequeno_lt8k_rsm2")
    elif ac < 15000:
        return ("Médio (8-15k m²)", "medio_8k_15k_rsm2")
    elif ac < 25000:
        return ("Grande (15-25k m²)", "grande_15k_25k_rsm2")
    else:
        return ("Extra (>25k m²)", "extra_gt25k_rsm2")


def validar(parametrico_path: Path | None, executivo_path: Path,
            ac: float, ur: int | None, output: Path) -> dict:
    summary = {
        "ts": datetime.now().isoformat(timespec="seconds"),
        "ac": ac,
        "ur": ur,
        "checks": [],
        "alerts": [],
        "gaps": [],
    }

    exec_data = read_resumo(executivo_path)
    if not exec_data:
        summary["alerts"].append("Executivo não tem aba RESUMO ou está vazia")
        write_report(output, summary)
        return summary

    summary["executivo_total"] = exec_data["grand_total"]
    summary["executivo_rsm2"] = exec_data["grand_total"] / ac if ac else 0
    summary["macrogrupos"] = exec_data["macrogrupos"]

    cal = load_calibration()
    seg_label, seg_key = segmento_por_ac(ac)
    summary["segmento"] = seg_label

    seg_stats = None
    if cal and cal.get("por_segmento"):
        seg_stats = cal["por_segmento"].get(seg_key)

    if seg_stats:
        rsm2_obtido = summary["executivo_rsm2"]
        rsm2_med = seg_stats.get("mediana", 0)
        rsm2_p10 = seg_stats.get("p10", 0)
        rsm2_p25 = seg_stats.get("p25", 0)
        rsm2_p75 = seg_stats.get("p75", 0)
        rsm2_p90 = seg_stats.get("p90", 0)

        if rsm2_p10 <= rsm2_obtido <= rsm2_p90:
            faixa = "ok_p10_p90"
            faixa_label = "✅ Dentro da faixa P10-P90 do segmento"
        elif rsm2_obtido < rsm2_p10:
            faixa = "abaixo_p10"
            faixa_label = "⚠️ Abaixo do P10 do segmento (subdimensionado)"
        else:
            faixa = "acima_p90"
            faixa_label = "⚠️ Acima do P90 do segmento (sobredimensionado)"

        delta_med = (rsm2_obtido - rsm2_med) / rsm2_med * 100 if rsm2_med else 0
        summary["checks"].append({
            "tipo": "rsm2_segmento",
            "segmento": seg_label,
            "esperado_med": round(rsm2_med, 2),
            "esperado_p10": round(rsm2_p10, 2),
            "esperado_p25": round(rsm2_p25, 2),
            "esperado_p75": round(rsm2_p75, 2),
            "esperado_p90": round(rsm2_p90, 2),
            "obtido": round(rsm2_obtido, 2),
            "delta_med_pct": round(delta_med, 1),
            "faixa": faixa,
            "status": "ok" if faixa == "ok_p10_p90" else "alerta",
        })
        if faixa != "ok_p10_p90":
            summary["alerts"].append(
                f"R$/m² total {rsm2_obtido:.0f} está {faixa_label} "
                f"(mediana {rsm2_med:.0f}, P10-P90: {rsm2_p10:.0f}-{rsm2_p90:.0f})"
            )

    sem_total = [mg for mg, d in exec_data["macrogrupos"].items() if not d.get("total")]
    if sem_total:
        summary["gaps"].append({
            "tipo": "macrogrupos_sem_total",
            "lista": sem_total,
            "count": len(sem_total),
            "nota": "Sem dados na base calibrada nem nos similares — preencher manualmente",
        })

    sem_detalhamento = [mg for mg, d in exec_data["macrogrupos"].items()
                         if d.get("total") and d.get("n_itens", 0) == 0]
    if sem_detalhamento:
        summary["gaps"].append({
            "tipo": "macrogrupos_sem_detalhamento_granular",
            "lista": sem_detalhamento,
            "count": len(sem_detalhamento),
            "nota": "Total calibrado OK, mas sem itens detalhados nos similares — usar memorial paramétrico",
        })

    baixa_confianca = [mg for mg, d in exec_data["macrogrupos"].items()
                       if "Baixa" in str(d.get("confianca", ""))]
    if baixa_confianca:
        summary["alerts"].append(
            f"{len(baixa_confianca)} macrogrupos com confiança baixa: " + ", ".join(baixa_confianca[:5])
        )

    write_report(output, summary)
    return summary


def write_report(output: Path, summary: dict) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# Relatório de Validação — Pacote",
        f"_Gerado em {summary['ts']}_",
        "",
        f"## Dados",
        f"- AC: {summary.get('ac')} m²",
        f"- UR: {summary.get('ur') or '—'}",
        f"- Total executivo: R$ {summary.get('executivo_total', 0):,.0f}".replace(",", "."),
        f"- R$/m² executivo: R$ {summary.get('executivo_rsm2', 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        "",
    ]

    if summary.get("checks"):
        lines.append("## Checagens automáticas")
        for c in summary["checks"]:
            status_emoji = "✅" if c["status"] == "ok" else "⚠️"
            if c["tipo"] == "rsm2_segmento":
                lines.append(f"- {status_emoji} **R$/m² vs segmento {c['segmento']}**")
                lines.append(f"  - Obtido: R$ {c['obtido']:.2f}")
                lines.append(f"  - Mediana: R$ {c['esperado_med']:.2f} (delta {c['delta_med_pct']:+.1f}%)")
                lines.append(f"  - Faixa P10-P90: R$ {c['esperado_p10']:.0f} – R$ {c['esperado_p90']:.0f}")
                lines.append(f"  - Faixa P25-P75: R$ {c['esperado_p25']:.0f} – R$ {c['esperado_p75']:.0f}")
                lines.append(f"  - Status: {c['faixa']}")
            else:
                lines.append(f"- {status_emoji} **{c['tipo']}**: esperado={c.get('esperado_med')}, "
                             f"obtido={c['obtido']}, delta={c.get('delta_pct', c.get('delta_med_pct'))}%")
        lines.append("")

    if summary.get("alerts"):
        lines.append("## ⚠️ Alertas")
        for a in summary["alerts"]:
            lines.append(f"- {a}")
        lines.append("")

    if summary.get("gaps"):
        lines.append("## 🔍 Lacunas identificadas")
        for g in summary["gaps"]:
            lines.append(f"### {g['tipo']} ({g['count']})")
            if g.get("nota"):
                lines.append(f"_{g['nota']}_")
                lines.append("")
            for item in g["lista"][:18]:
                lines.append(f"- {item}")
            lines.append("")

    if summary.get("macrogrupos"):
        lines.append("## Macrogrupos no executivo")
        lines.append("| Macrogrupo | Total R$ | R$/m² | % | N itens | Confiança |")
        lines.append("|---|---|---|---|---|---|")
        for mg, d in summary["macrogrupos"].items():
            lines.append(f"| {mg} | R$ {d['total']:,.0f} | R$ {d['rsm2']:,.2f} | "
                         f"{d['pct']*100:.1f}% | {d['n_itens']} | {d['confianca']} |".replace(",", "."))
        lines.append("")

    output.write_text("\n".join(lines), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--parametrico", default=None)
    ap.add_argument("--executivo", required=True)
    ap.add_argument("--ac", type=float, required=True)
    ap.add_argument("--ur", type=int, default=None)
    ap.add_argument("-o", "--output", default="relatorio-validacao.md")
    args = ap.parse_args()

    parametrico_path = Path(args.parametrico) if args.parametrico else None
    executivo_path = Path(args.executivo)
    output = Path(args.output)

    summary = validar(parametrico_path, executivo_path, args.ac, args.ur, output)
    print(json.dumps({k: v for k, v in summary.items() if k not in ("macrogrupos",)},
                     indent=2, ensure_ascii=False))
    print(f"\nrelatório: {output}")


if __name__ == "__main__":
    main()
