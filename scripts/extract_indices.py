#!/usr/bin/env python3
"""
Extrator de índices por subdisciplina dos executivos da Cartesian.
Processa 3 formatos:
  1. Analítico (aba GERENCIAMENTO_EXECUTIVO ou Orçamento Executivo) — subcategorias com R$/vb
  2. Gerenciamento (aba INSTALAÇÕES separada) — subcategorias com R$/m² AC
  3. Resumo (aba Orçamento Resumo) — só macrogrupos
  4. Detalhado (abas individuais por disciplina, ex: west-coast) — itens com quantidades reais

Output: JSON com todos os índices extraídos por projeto.
"""

import os
import sys
import json
import re
import openpyxl
from collections import defaultdict
from pathlib import Path

_ARCHIVE_V1 = Path.home() / "orcamentos-openclaw" / "archive" / "v1-pre-fase19"
BASE_DIR = str(_ARCHIVE_V1 / "executivos")
ENTREGAS_DIR = str(_ARCHIVE_V1 / "entregas")
OUTPUT_FILE = str(_ARCHIVE_V1 / "indices-subdisciplina.json")

def safe_float(val):
    """Convert value to float safely."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '.').replace(' ', ''))
    except (ValueError, TypeError):
        return None

def get_area_construida(wb):
    """Extract AC (área construída) from OBRA or first sheet."""
    # Priority: check DADOS_INICIAIS first (gerenciamento format), then OBRA, then ORÇAMENTO sheets
    priority_sheets = ['DADOS_INICIAIS', 'Dados Gerais', 'OBRA', 'Obra', 'Obra ']
    # Also check ORÇAMENTO_EXECUTIVO sheets where AC is in the header
    orc_sheets = [s for s in wb.sheetnames if 'ORÇAMENTO' in s.upper() or 'EXECUTIVO' in s.upper()]
    other_sheets = [s for s in wb.sheetnames if s not in priority_sheets and s not in orc_sheets]
    search_order = [s for s in priority_sheets if s in wb.sheetnames] + orc_sheets + other_sheets
    
    for sheet_name in search_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=30, max_col=6, values_only=True):
            if not row:
                continue
            # Check all columns for AC keywords
            for i, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    upper = cell.upper()
                    if 'ÁREA TOTAL CONSTRUÍDA' in upper or 'AREA TOTAL CONSTRUIDA' in upper or \
                       'ÁREA CONSTRUÍDA' in upper or 'AREA CONSTRUIDA' in upper:
                        # Find the numeric value in nearby columns
                        for j in range(len(row)):
                            if j != i:
                                val = safe_float(row[j])
                                if val and val > 100 and val < 200000:
                                    return val
    return None

def get_num_unidades(wb):
    """Extract number of units from OBRA sheet."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=30, max_col=6, values_only=True):
            if row and row[1]:
                desc = str(row[1]).upper()
                if 'UNIDADES HABITACIONAIS' in desc or 'NÚMERO DE UNIDADES' in desc:
                    val = safe_float(row[3]) or safe_float(row[2])
                    if val and val > 0:
                        return int(val)
    return None

def extract_analitico(wb, ac):
    """Extract from analítico format (GERENCIAMENTO_EXECUTIVO or Orçamento Executivo sheet)."""
    target_sheets = ['GERENCIAMENTO_EXECUTIVO', 'Orçamento Executivo', 'Orçamento_Executivo',
                     'ORÇAMENTO RESUMO', 'Orçamento']
    ws = None
    for s in target_sheets:
        if s in wb.sheetnames:
            ws = wb[s]
            break
    if not ws:
        return None
    
    result = {
        'format': 'analitico',
        'ac': ac,
        'macrogrupos': {},
        'subdisciplinas': {}
    }
    
    current_macro = None
    current_sub = None
    
    for row in ws.iter_rows(max_row=300, max_col=8, values_only=True):
        if not row or not row[1]:
            continue
        
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip()
        qty = safe_float(row[2])
        unit = str(row[3]).strip() if row[3] else ''
        valor = safe_float(row[4])
        
        # Detect macrogroup level (single digit code like "7")
        if re.match(r'^\d+$', code) and valor:
            current_macro = desc
            result['macrogrupos'][desc] = {
                'valor': valor,
                'rsm2': valor / ac if ac else None
            }
            current_sub = None
        
        # Detect sub-level (X.X format)
        elif re.match(r'^\d+\.\d+$', code) and valor:
            current_sub = desc
            if current_macro:
                if current_macro not in result['subdisciplinas']:
                    result['subdisciplinas'][current_macro] = {}
                result['subdisciplinas'][current_macro][desc] = {
                    'valor': valor,
                    'rsm2': valor / ac if ac else None,
                    'itens': {}
                }
        
        # Detect item level (X.X.X format)
        elif re.match(r'^\d+\.\d+\.\d+$', code) and valor:
            item_data = {
                'valor': valor,
                'rsm2': valor / ac if ac else None,
                'quantidade': qty,
                'unidade': unit
            }
            if current_macro and current_sub:
                if current_macro in result['subdisciplinas'] and current_sub in result['subdisciplinas'][current_macro]:
                    result['subdisciplinas'][current_macro][current_sub]['itens'][desc] = item_data
    
    return result

def extract_gerenciamento_instalacoes(wb, ac):
    """Extract from gerenciamento format (separate INSTALAÇÕES sheet)."""
    inst_sheet = None
    for s in wb.sheetnames:
        if 'instalações' in s.lower() or 'instalacoes' in s.lower():
            inst_sheet = s
            break
    if not inst_sheet:
        return None
    
    ws = wb[inst_sheet]
    result = {
        'format': 'gerenciamento_instalacoes',
        'ac': ac,
        'subdisciplinas': {}
    }
    
    current_section = None
    
    for row in ws.iter_rows(max_row=100, max_col=10, values_only=True):
        if not row:
            continue
        
        # Find section headers
        for cell in row:
            if cell and isinstance(cell, str):
                upper = cell.upper()
                if 'INSTALAÇÕES ELÉTRICAS' in upper and 'TOTAL' not in upper:
                    current_section = 'Instalações Elétricas'
                    if current_section not in result['subdisciplinas']:
                        result['subdisciplinas'][current_section] = {'itens': {}}
                elif 'INSTALAÇÕES HIDRO' in upper or 'INSTALAÇÕES HIDRÁULICAS' in upper:
                    current_section = 'Instalações Hidráulicas'
                    if current_section not in result['subdisciplinas']:
                        result['subdisciplinas'][current_section] = {'itens': {}}
                elif 'INSTALAÇÕES PREVENTIVAS' in upper or 'PREVENTIVAS E GLP' in upper:
                    current_section = 'Instalações Preventivas e GLP'
                    if current_section not in result['subdisciplinas']:
                        result['subdisciplinas'][current_section] = {'itens': {}}
                elif 'TOTAL INSTALAÇÕES' in upper:
                    # Get total value
                    for c in row:
                        v = safe_float(c)
                        if v and v > 10000:
                            result['total_instalacoes'] = v
                            result['total_rsm2'] = v / ac if ac else None
                            break
        
        # Extract items within sections
        if current_section and row[0] and isinstance(row[0], str):
            desc = str(row[0]).strip()
            if desc in ('Descrição', 'TOTAL', '') or 'INSTALAÇÕES' in desc.upper():
                continue
            
            param = safe_float(row[1])
            valor = safe_float(row[3]) or safe_float(row[2])
            
            if valor and valor > 0:
                result['subdisciplinas'][current_section]['itens'][desc] = {
                    'valor': valor,
                    'rsm2': valor / ac if ac else None,
                    'parametro_rsm2': param
                }
    
    # Calculate totals per section
    for section, data in result['subdisciplinas'].items():
        total = sum(item['valor'] for item in data['itens'].values() if item.get('valor'))
        data['total'] = total
        data['rsm2'] = total / ac if ac else None
    
    return result

def extract_resumo(wb, ac):
    """Extract from resumo format — any sheet with macrogroup-level summary."""
    # Exact name matches (priority order)
    target_sheets = [
        'Orçamento Resumo', 'Gerenciamento executivo ', 'Gerenciamento executivo',
        'Resumo', 'RESUMOS', 'GER_EXECUTIVO', 'Gerenciamento_Exec',
    ]
    # Also match sheets containing these patterns
    pattern_matches = ['ORÇAMENTO_EXECUTIVO', 'ORÇAMENTO EXECUTIVO', 'ORÇAMENTO_PRELIMINAR',
                       'ORÇAMENTO_PARAMÉTRICO', 'ORÇAMENTO_PARCIAL']
    
    ws = None
    for s in target_sheets:
        if s in wb.sheetnames:
            ws = wb[s]
            break
    
    if not ws:
        for sheet_name in wb.sheetnames:
            upper = sheet_name.upper().strip()
            for pattern in pattern_matches:
                if pattern in upper:
                    ws = wb[sheet_name]
                    break
            if ws:
                break
    
    if not ws:
        return None
    
    result = {
        'format': 'resumo',
        'ac': ac,
        'macrogrupos': {}
    }
    
    # Also try to extract AC from the sheet header if not already found
    if not ac:
        for row in ws.iter_rows(max_row=5, max_col=10, values_only=True):
            if not row:
                continue
            for i, cell in enumerate(row):
                if cell and isinstance(cell, str) and 'ÁREA CONSTRUÍDA' in cell.upper():
                    # Look for numeric value in nearby cells
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 100 < val < 200000:
                                ac = val
                                result['ac'] = ac
                                break
    
    # Known macrogroup keywords to validate entries
    macro_keywords = [
        'GERENCIAMENTO', 'MOVIMENTAÇÃO', 'INFRAESTRUTURA', 'SUPRAESTRUTURA',
        'ALVENARIA', 'PAREDES', 'INSTALAÇÕES', 'EQUIPAMENTOS', 'SISTEMAS',
        'IMPERMEABILIZAÇÃO', 'REVESTIMENTO', 'ACABAMENTO', 'PINTURA',
        'ESQUADRIA', 'VIDRO', 'FACHADA', 'COBERTURA', 'COMPLEMENTAR',
        'IMPREVISTO', 'FUNDAÇÃO', 'CONTENÇ', 'LOUÇAS', 'METAIS',
        'PISO', 'TETO', 'FORRO', 'SERVIÇOS',
    ]
    
    skip_values = {'Etapa', 'ETAPA', '', 'TOTAL', 'CUB/SC', 'CUB/SP', 'CUB',
                   'Rótulos de Linha', 'Soma de %', 'VALOR TOTAL', 'VALOR ORÇADO'}
    
    for row in ws.iter_rows(max_row=50, max_col=8, values_only=True):
        if not row or not row[0]:
            continue
        
        desc = str(row[0]).strip()
        if not desc or desc in skip_values:
            continue
        
        # Validate it looks like a macrogroup
        upper = desc.upper()
        is_macro = any(kw in upper for kw in macro_keywords)
        if not is_macro:
            continue
        
        valor = safe_float(row[1])
        pct = safe_float(row[2])
        rsm2 = safe_float(row[3])
        
        if valor and valor > 100:
            result['macrogrupos'][desc] = {
                'valor': valor,
                'pct': pct,
                'rsm2': rsm2 or (valor / ac if ac else None)
            }
    
    if not result['macrogrupos']:
        return None
    
    return result

def extract_detalhado(wb, ac):
    """Extract from detailed format with individual discipline sheets (e.g., west-coast)."""
    discipline_sheets = {
        'ELÉTRICA': 'Instalações Elétricas',
        'HIDRÁULICAS': 'Instalações Hidráulicas',
        'HIDROSSANIT': 'Instalações Hidráulicas',
        'ESGOTO E PLUVIAL': 'Esgoto e Pluvial',
        'LOUÇAS E METAIS': 'Louças e Metais',
        'EXAUSTÃO': 'Exaustão',
        'TELECOM': 'Telecom',
        'SPDA': 'SPDA',
        'SUPRAESTRUTURA': 'Supraestrutura',
        'ALVENARIA': 'Alvenaria',
        'REVESTIMENTO': 'Revestimentos',
        'PISOS': 'Pisos',
        'PAVIMENTAÇ': 'Pisos',
        'FORRO': 'Forro',
        'TETO': 'Forro',
        'FACHADA': 'Fachada',
        'IMPER': 'Impermeabilização',
        'ESCAVAÇÕES': 'Escavações',
        'ESTACAS': 'Estacas',
        'BALDRAMES': 'Baldrames',
        'ESCORAMENTO': 'Escoramento',
        'CHURRASQUEIRA': 'Churrasqueira',
        'MOVIMENTAÇÃO': 'Movimentação de Terra',
        'INFRAESTRUTURA': 'Infraestrutura',
        'COMPLEMENTAR': 'Serviços Complementares',
        'SISTEMAS ESPECIAIS': 'Sistemas Especiais',
        'ESQUADRIA': 'Esquadrias',
        'PINTURA': 'Pintura',
        'COBERTURA': 'Cobertura',
        'CANTEIRO': 'Canteiro',
    }
    
    found_sheets = {}
    for sheet_name in wb.sheetnames:
        for key, label in discipline_sheets.items():
            if sheet_name.upper().strip() == key or key in sheet_name.upper():
                found_sheets[label] = sheet_name
    
    if len(found_sheets) < 3:
        return None  # Not enough discipline sheets to consider it "detalhado"
    
    result = {
        'format': 'detalhado',
        'ac': ac,
        'disciplinas': {}
    }
    
    for label, sheet_name in found_sheets.items():
        ws = wb[sheet_name]
        items = {}
        subcategories = {}
        current_subcat = None
        
        for row in ws.iter_rows(max_row=500, max_col=10, values_only=True):
            if not row:
                continue
            
            # Find subcategory headers (have total in col 6 or 7)
            desc = str(row[1]).strip() if row[1] else ''
            if not desc:
                desc = str(row[0]).strip() if row[0] else ''
            
            if not desc:
                continue
            
            qty = safe_float(row[3])
            unit = str(row[4]).strip() if row[4] else ''
            pu = safe_float(row[5])
            total = safe_float(row[6])
            
            # Subcategory header (has total but no qty/unit)
            if total and total > 1000 and not qty and not unit:
                current_subcat = desc
                subcategories[desc] = {
                    'total': total,
                    'rsm2': total / ac if ac else None,
                    'itens': []
                }
            # Individual item
            elif qty and unit and total:
                item = {
                    'descricao': desc,
                    'quantidade': qty,
                    'unidade': unit,
                    'preco_unitario': pu,
                    'total': total
                }
                if current_subcat and current_subcat in subcategories:
                    subcategories[current_subcat]['itens'].append(item)
                items[desc] = item
        
        total_disc = sum(s['total'] for s in subcategories.values())
        result['disciplinas'][label] = {
            'total': total_disc,
            'rsm2': total_disc / ac if ac else None,
            'subcategorias': subcategories,
            'total_itens': len(items)
        }
    
    return result

def extract_sienge_relatorio(wb, ac):
    """Extract from Sienge export format (single 'Relatório' sheet).
    Structure: code in col 0, description in col 1, unit in col 13,
    quantity in col 16, unit price in col 22, total price in col 28.
    Hierarchy: XX = macrogroup, XX.XXX = section, XX.XXX.XXX = subsection,
    XX.XXX.XXX.XXX = item.
    """
    if 'Relatório' not in wb.sheetnames:
        return None
    
    ws = wb['Relatório']
    
    # Extract project name from header
    project_name = None
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if row and row[0] and isinstance(row[0], str):
            if row[0].strip() == 'Obra' and row[4]:
                # Format: "40 - Duo Brava " -> "Duo Brava"
                name = str(row[4]).strip()
                if ' - ' in name:
                    name = name.split(' - ', 1)[1].strip()
                project_name = name
                break
    
    result = {
        'format': 'sienge_relatorio',
        'ac': ac,
        'project_name': project_name,
        'macrogrupos': {},
        'subdisciplinas': {},
    }
    
    current_macro = None
    current_section = None
    
    # Process all rows
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        
        code = str(row[0]).strip() if row[0] else ''
        desc = str(row[1]).strip() if row[1] else ''
        total = safe_float(row[28]) if len(row) > 28 else None
        
        if not code or not desc or code == 'Código':
            continue
        
        # Count code depth by dots
        parts = code.split('.')
        depth = len(parts)
        
        if depth == 1 and re.match(r'^\d+\s*$', code):
            # Top-level macrogroup (e.g., "01", "02")
            current_macro = desc
            current_section = None
            if total and total > 0:
                result['macrogrupos'][desc] = {
                    'valor': total,
                    'rsm2': total / ac if ac else None
                }
                result['subdisciplinas'][desc] = {}
        
        elif depth == 2 and re.match(r'^\d+\.\d+\s*$', code):
            # Section level (e.g., "01.001")
            current_section = desc
            if total and total > 0 and current_macro:
                if current_macro in result['subdisciplinas']:
                    result['subdisciplinas'][current_macro][desc] = {
                        'valor': total,
                        'rsm2': total / ac if ac else None,
                        'itens': {}
                    }
        
        elif depth == 3 and re.match(r'^\d+\.\d+\.\d+\s*$', code):
            # Subsection level (e.g., "01.001.001")
            if total and total > 0 and current_macro and current_section:
                if current_macro in result['subdisciplinas']:
                    sec = result['subdisciplinas'][current_macro].get(current_section)
                    if sec:
                        sec['itens'][desc] = {
                            'valor': total,
                            'rsm2': total / ac if ac else None
                        }
    
    if not result['macrogrupos']:
        return None
    
    return result

def extract_all_macrogroups(wb, ac):
    """Try to extract macrogroup-level data from any sheet that has a summary."""
    # Check for summary-style sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        macros = {}
        for row in ws.iter_rows(max_row=40, max_col=8, values_only=True):
            if not row:
                continue
            
            desc = str(row[0]).strip() if row[0] else ''
            if not desc:
                continue
            
            # Look for typical macrogroup names
            upper = desc.upper()
            is_macro = any(kw in upper for kw in [
                'MOVIMENTAÇÃO', 'INFRAESTRUTURA', 'SUPRAESTRUTURA', 'ALVENARIA',
                'INSTALAÇÕES', 'IMPERMEABILIZAÇÃO', 'REVESTIMENTO', 'PINTURA',
                'ESQUADRIA', 'FACHADA', 'COBERTURA', 'COMPLEMENTAR', 'IMPREVISTO',
                'GERENCIAMENTO', 'FUNDAÇÃO', 'CONTENÇ', 'SISTEMA'
            ])
            
            if is_macro:
                valor = safe_float(row[1])
                rsm2 = safe_float(row[3]) or safe_float(row[2])
                if valor and valor > 100:
                    macros[desc] = {'valor': valor, 'rsm2': rsm2}
        
        if len(macros) >= 5:
            return macros
    
    return None

def process_file(filepath):
    """Process a single executivo file and extract all possible indices."""
    filename = os.path.basename(filepath)
    project_name = filename.replace('.xlsx', '')
    
    # Skip non-relevant files
    skip_patterns = ['prazo', 'calculo', 'curva', 'custos-indiretos', 'levantamentos', 'custos_indiretos']
    if any(p in project_name.lower() for p in skip_patterns):
        return None
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        return {'project': project_name, 'error': str(e)}
    
    ac = get_area_construida(wb)
    num_units = get_num_unidades(wb)
    
    result = {
        'project': project_name,
        'file': filename,
        'ac': ac,
        'num_unidades': num_units,
        'extractions': {}
    }
    
    # Try each extraction method
    analitico = extract_analitico(wb, ac)
    if analitico:
        result['extractions']['analitico'] = analitico
    
    gerenc = extract_gerenciamento_instalacoes(wb, ac)
    if gerenc:
        result['extractions']['gerenciamento'] = gerenc
    
    resumo = extract_resumo(wb, ac)
    if resumo:
        result['extractions']['resumo'] = resumo
        # Update AC from resumo if found there
        if not ac and resumo.get('ac'):
            ac = resumo['ac']
            result['ac'] = ac
    
    detalhado = extract_detalhado(wb, ac)
    if detalhado:
        result['extractions']['detalhado'] = detalhado
    
    sienge = extract_sienge_relatorio(wb, ac)
    if sienge:
        result['extractions']['sienge'] = sienge
    
    wb.close()
    
    if not result['extractions']:
        return None
    
    return result

def load_ac_from_calibration():
    """Load AC values from calibration-data.json as fallback."""
    cal_path = str(_ARCHIVE_V1 / "raiz-mar-2026" / "calibration-data.json")
    ac_map = {}
    try:
        with open(cal_path) as f:
            cal = json.load(f)
        for p in cal:
            name = p['name'].lower().replace(' ', '-')
            if p.get('ac'):
                ac_map[name] = p['ac']
    except Exception:
        pass
    return ac_map

def match_ac_from_calibration(project_name, ac_map):
    """Try to match project name to calibration AC."""
    name = project_name.lower()
    # Direct match
    if name in ac_map:
        return ac_map[name]
    # Try removing suffixes
    for suffix in ['-gerenciamento', '-analitico', '-executivo', '-gerenciamento-executivo',
                   '-gerenciamento-mo', '-gerenciamento-resid', '-gerenciamento-hotel',
                   '-gerenciamento-mall-r02', '-gerenciamento-mall1',
                   '-analitico-cliente', '-analitico-entregavel',
                   '-custos-indiretos', '-h-empreendimentos', '-holze',
                   '-inbrasul', '-indepy-executivo', '-indepy', '-lotisa',
                   '-librahaus', '-kirchner', '-jta', '-hacasa', '-lumis']:
        stripped = name.replace(suffix, '')
        if stripped in ac_map:
            return ac_map[stripped]
    # Partial match
    for cal_name, ac in ac_map.items():
        if cal_name in name or name in cal_name:
            return ac
    return None

def recalc_rsm2(data, ac):
    """Recursively recalculate rsm2 values given AC."""
    if isinstance(data, dict):
        if 'valor' in data and 'rsm2' in data:
            v = data['valor']
            if v and isinstance(v, (int, float)):
                data['rsm2'] = v / ac
        if 'total' in data and ac:
            t = data.get('total')
            if t and isinstance(t, (int, float)):
                data['rsm2'] = t / ac
        for k, v in data.items():
            if isinstance(v, (dict, list)):
                recalc_rsm2(v, ac)
    elif isinstance(data, list):
        for item in data:
            recalc_rsm2(item, ac)

def main():
    ac_map = load_ac_from_calibration()
    
    # Collect xlsx from executivos/ (flat) + entregas/ (recursive)
    files = sorted([
        os.path.join(BASE_DIR, f) 
        for f in os.listdir(BASE_DIR) 
        if f.endswith('.xlsx')
    ])
    
    # Also scan entregas/ recursively (Leo's primary upload folder)
    if os.path.isdir(ENTREGAS_DIR):
        for root, dirs, fnames in os.walk(ENTREGAS_DIR):
            for fname in fnames:
                if fname.endswith('.xlsx'):
                    files.append(os.path.join(root, fname))
        files = sorted(set(files))  # dedupe just in case
    
    print(f"Processing {len(files)} files...")
    
    results = []
    errors = []
    skipped = 0
    
    for f in files:
        try:
            r = process_file(f)
            if r:
                if 'error' in r:
                    errors.append(r)
                else:
                    # Try to fill AC from calibration if missing
                    if not r['ac']:
                        cal_ac = match_ac_from_calibration(r['project'], ac_map)
                        if cal_ac:
                            r['ac'] = cal_ac
                            r['ac_source'] = 'calibration'
                            # Recalculate rsm2 for all extractions
                            for fmt, data in r['extractions'].items():
                                recalc_rsm2(data, cal_ac)
                    
                    results.append(r)
                    formats = list(r['extractions'].keys())
                    src = r.get('ac_source', 'file')
                    print(f"  ✅ {os.path.basename(f)} → AC={r['ac']} ({src}), formats={formats}")
            else:
                skipped += 1
        except Exception as e:
            errors.append({'file': os.path.basename(f), 'error': str(e)})
            print(f"  ❌ {os.path.basename(f)} → {str(e)[:80]}")
    
    print(f"\nDone: {len(results)} extracted, {skipped} skipped, {len(errors)} errors")
    
    # Save results
    output = {
        'total_projects': len(results),
        'extraction_date': '2026-03-11',
        'projects': results,
        'errors': errors
    }
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"Saved to {OUTPUT_FILE}")
    
    # Generate summary
    generate_summary(results)

def generate_summary(results):
    """Generate a human-readable summary of all extracted indices."""
    summary_path = str(_ARCHIVE_V1 / "INDICES-SUBDISCIPLINA-RESUMO.md")
    
    # Collect all subdiscipline indices across projects
    indices = defaultdict(list)  # key: "Macro > Sub > Item", value: list of {project, rsm2, valor}
    
    for r in results:
        ac = r['ac']
        project = r['project']
        
        for fmt, data in r['extractions'].items():
            if fmt == 'analitico':
                for macro, subs in data.get('subdisciplinas', {}).items():
                    for sub, sub_data in subs.items():
                        key = f"{macro} > {sub}"
                        if sub_data.get('rsm2'):
                            indices[key].append({
                                'project': project,
                                'rsm2': sub_data['rsm2'],
                                'valor': sub_data.get('valor')
                            })
                        for item_name, item_data in sub_data.get('itens', {}).items():
                            item_key = f"{macro} > {sub} > {item_name}"
                            if item_data.get('rsm2'):
                                indices[item_key].append({
                                    'project': project,
                                    'rsm2': item_data['rsm2'],
                                    'valor': item_data.get('valor'),
                                    'quantidade': item_data.get('quantidade'),
                                    'unidade': item_data.get('unidade')
                                })
            
            elif fmt == 'gerenciamento':
                for section, sec_data in data.get('subdisciplinas', {}).items():
                    key = f"Instalações > {section}"
                    if sec_data.get('rsm2'):
                        indices[key].append({
                            'project': project,
                            'rsm2': sec_data['rsm2']
                        })
                    for item_name, item_data in sec_data.get('itens', {}).items():
                        item_key = f"Instalações > {section} > {item_name}"
                        if item_data.get('rsm2'):
                            indices[item_key].append({
                                'project': project,
                                'rsm2': item_data['rsm2'],
                                'parametro': item_data.get('parametro_rsm2')
                            })
            
            elif fmt == 'sienge':
                for macro, macro_data in data.get('macrogrupos', {}).items():
                    key = f"Sienge > {macro}"
                    if macro_data.get('rsm2'):
                        indices[key].append({
                            'project': project,
                            'rsm2': macro_data['rsm2'],
                            'valor': macro_data.get('valor')
                        })
                for macro, subs in data.get('subdisciplinas', {}).items():
                    for sub, sub_data in subs.items():
                        key = f"Sienge > {macro} > {sub}"
                        if sub_data.get('rsm2'):
                            indices[key].append({
                                'project': project,
                                'rsm2': sub_data['rsm2'],
                                'valor': sub_data.get('valor')
                            })
            
            elif fmt == 'detalhado':
                for disc, disc_data in data.get('disciplinas', {}).items():
                    key = f"Detalhado > {disc}"
                    if disc_data.get('rsm2'):
                        indices[key].append({
                            'project': project,
                            'rsm2': disc_data['rsm2']
                        })
                    for subcat, subcat_data in disc_data.get('subcategorias', {}).items():
                        subcat_key = f"Detalhado > {disc} > {subcat}"
                        if subcat_data.get('rsm2'):
                            indices[subcat_key].append({
                                'project': project,
                                'rsm2': subcat_data['rsm2']
                            })
    
    # Write summary
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("# Índices por Subdisciplina — Resumo Extraído\n\n")
        f.write(f"**Data:** 2026-03-11\n")
        f.write(f"**Projetos processados:** {len(results)}\n\n")
        f.write("---\n\n")
        
        # Group by first level
        grouped = defaultdict(dict)
        for key, values in sorted(indices.items()):
            parts = key.split(' > ')
            top = parts[0]
            rest = ' > '.join(parts[1:])
            grouped[top][rest] = values
        
        for top_level, items in sorted(grouped.items()):
            f.write(f"## {top_level}\n\n")
            
            for item_name, values in sorted(items.items()):
                rsm2_list = [v['rsm2'] for v in values if v.get('rsm2')]
                if not rsm2_list:
                    continue
                
                n = len(rsm2_list)
                med = sorted(rsm2_list)[n // 2]
                mn = min(rsm2_list)
                mx = max(rsm2_list)
                avg = sum(rsm2_list) / n
                
                f.write(f"### {item_name}\n")
                f.write(f"- *Projetos:* {n}\n")
                f.write(f"- *Mediana:* R$ {med:.2f}/m² AC\n")
                f.write(f"- *Média:* R$ {avg:.2f}/m² AC\n")
                f.write(f"- *Faixa:* R$ {mn:.2f} — R$ {mx:.2f}/m² AC\n")
                
                # List projects
                f.write(f"- *Detalhes:*\n")
                for v in sorted(values, key=lambda x: x.get('rsm2', 0)):
                    extra = ''
                    if v.get('parametro'):
                        extra = f" (param: {v['parametro']})"
                    if v.get('quantidade') and v.get('unidade'):
                        extra += f" | qty: {v['quantidade']:.1f} {v['unidade']}"
                    f.write(f"  - {v['project']}: R$ {v['rsm2']:.2f}/m²{extra}\n")
                f.write("\n")
            
            f.write("---\n\n")
    
    print(f"Summary saved to {summary_path}")

if __name__ == '__main__':
    main()
