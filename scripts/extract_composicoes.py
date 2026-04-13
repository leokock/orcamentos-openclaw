#!/usr/bin/env python3
"""Phase 4a — Extract composition abas (CPU / Insumos / Composições) from xlsx.

Para os ~22 projetos da base que têm abas de composição, lê os xlsx pesados
e extrai os insumos, classificando cada um como material/mao_obra/equipamento.

Saída: base/composicoes-raw/[projeto].json
Schema:
{
  "projeto": "slug",
  "abas": [
    {
      "nome": "CPU",
      "n_linhas": 1730,
      "items": [
        {"descricao": "...", "unidade": "kg", "consumo": 0.5, "pu": 8.50,
         "total": 4.25, "categoria": "material"}
      ]
    }
  ]
}
"""
from __future__ import annotations

import json
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
QUEUE = BASE / "gemma-queue.json"
OUT_DIR = BASE / "composicoes-raw"
OUT_DIR.mkdir(parents=True, exist_ok=True)
LOG = BASE / "phase4a-extract.log.jsonl"

COMP_PATTERNS = [r"\bcpu\b", r"composi", r"insumo"]


HEADER_KEYWORDS = {
    "codigo": ["código", "codigo", "cod.", "cod ", "item", "ref"],
    "descricao": ["descrição", "descricao", "descricao do servico",
                  "discriminação", "discriminacao", "insumo", "material",
                  "serviço", "servico", "componente", "elemento"],
    "unidade": ["unidade", "und", "und.", "un.", "un ", "u.m.", "um"],
    "consumo": ["consumo", "coef", "coeficiente", "cons.", "qtd", "quantidade",
                "quant.", "ind.", "indice", "índice"],
    "pu": ["preço unit", "preco unit", "p. unit", "p.unit", "valor unit",
           "unitário", "unitario", "p.u.", "pu", "custo unit", "vlr unit"],
    "total": ["total", "valor total", "preço total", "preco total",
              "vlr total", "subtotal", "custo total"],
}

MAT_KEYWORDS = ["concreto", "aço", "aco", "bloco", "argamassa", "tinta", "cabo",
                "tubo", "fio", "porcelanato", "cerâmica", "ceramica", "manta",
                "vidro", "alumínio", "aluminio", "madeira", "compensado",
                "cimento", "areia", "brita", "cordoalha", "pino", "parafuso",
                "porca", "arruela", "telha", "fechadura", "dobradiça", "dobradica",
                "torneira", "bacia", "registro", "metal", "metais", "louça", "louca",
                "rejunte", "selador", "massa corrida", "fita", "isolante",
                "impermeabilizante", "verniz", "esmalte", "perfil", "chapa", "barra"]

MO_KEYWORDS = ["mão de obra", "mao de obra", "m.o.", "m.o ", " mo ",
               "encarregado", "pedreiro", "servente", "carpinteiro",
               "armador", "eletricista", "encanador", "pintor", "azulejista",
               "ajudante", "ferreiro", "oficial", "sub-empreitada",
               "subempreitada", "instalador", "operador", "mestre",
               "engenheiro", "almoxarife", "encargo", "execução", "execucao"]

EQUIP_KEYWORDS = ["equipamento", "máquina", "maquina", "ferramenta",
                  "vibrador", "betoneira", "andaime", "escora", "guincho",
                  "elevador de obra", "policorte", "serra", "lixadeira",
                  "compressor", "gerador", "bomba", "balancim",
                  "perfuratriz", "martelete", "esmerilhadeira"]


def normalize(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def is_number(v) -> bool:
    if v is None or v == "":
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", ".").replace(" ", ""))
        return True
    except Exception:
        return False


def to_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except Exception:
        return None


def detect_header(rows):
    for ri, row in enumerate(rows[:30]):
        if not row:
            continue
        col_map = {}
        for ci, cell in enumerate(row):
            n = normalize(cell)
            if not n:
                continue
            for field, kws in HEADER_KEYWORDS.items():
                if field in col_map:
                    continue
                for kw in kws:
                    if kw in n:
                        col_map[field] = ci
                        break
        if "descricao" in col_map and ("consumo" in col_map or "pu" in col_map or "total" in col_map):
            return ri, col_map
    return None, None


def classify(text: str) -> str:
    if not text:
        return "outros"
    t = text.lower()
    for kw in MO_KEYWORDS:
        if kw in t:
            return "mao_obra"
    for kw in EQUIP_KEYWORDS:
        if kw in t:
            return "equipamento"
    for kw in MAT_KEYWORDS:
        if kw in t:
            return "material"
    return "outros"


def is_composition_aba(name: str) -> bool:
    n = normalize(name)
    return any(re.search(p, n) for p in COMP_PATTERNS)


def extract_xlsx(xlsx_path: Path) -> dict:
    out = {
        "arquivo": str(xlsx_path),
        "abas": [],
        "total_items": 0,
        "errors": [],
    }
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        out["errors"].append(f"open: {e}")
        return out

    for sn in wb.sheetnames:
        if not is_composition_aba(sn):
            continue
        try:
            ws = wb[sn]
            rows = []
            for r in ws.iter_rows(values_only=True):
                if any(c is not None and c != "" for c in r):
                    rows.append(list(r))
                if len(rows) > 8000:
                    break
        except Exception as e:
            out["errors"].append(f"sheet {sn}: {e}")
            continue

        if not rows:
            continue

        header_idx, col_map = detect_header(rows)
        if not col_map:
            continue

        items = []
        cat_counts = {"material": 0, "mao_obra": 0, "equipamento": 0, "outros": 0}
        for row in rows[header_idx + 1:]:
            it = {}
            for field, idx in col_map.items():
                if idx < len(row):
                    v = row[idx]
                    if field in ("consumo", "pu", "total"):
                        it[field] = to_number(v)
                    else:
                        it[field] = str(v).strip() if v is not None else None
            desc = it.get("descricao") or ""
            if not desc or len(desc) < 3:
                continue
            has_value = any(it.get(k) is not None and it.get(k) != 0
                            for k in ("consumo", "pu", "total"))
            if not has_value:
                continue
            it["categoria"] = classify(desc)
            cat_counts[it["categoria"]] += 1
            items.append(it)

        if items:
            out["abas"].append({
                "nome": sn,
                "header_idx": header_idx,
                "col_map": col_map,
                "n_items": len(items),
                "cat_counts": cat_counts,
                "items": items,
            })
            out["total_items"] += len(items)

    try:
        wb.close()
    except Exception:
        pass
    return out


def log_event(e: dict) -> None:
    e.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
    with LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(e, ensure_ascii=False) + "\n")


def pick_xlsx(sources: dict) -> str | None:
    files = sources.get("xlsx", [])
    if not files:
        return None
    sized = []
    for p in files:
        try:
            sized.append((Path(p).stat().st_size, p))
        except Exception:
            pass
    sized.sort(reverse=True)
    return sized[0][1] if sized else None


def process_one(project: dict) -> dict:
    slug = project["projeto"]
    summary = {"projeto": slug, "status": "pending"}
    started = time.time()

    xlsx = pick_xlsx(project.get("sources", {}))
    if not xlsx:
        summary["status"] = "no_xlsx"
        summary["duration_s"] = round(time.time() - started, 2)
        log_event(summary)
        return summary

    try:
        result = extract_xlsx(Path(xlsx))
        result["projeto"] = slug
        if result["total_items"] == 0:
            summary["status"] = "no_composition"
            summary["duration_s"] = round(time.time() - started, 2)
            log_event(summary)
            return summary

        out_path = OUT_DIR / f"{slug}.json"
        out_path.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")

        summary["status"] = "done"
        summary["abas"] = len(result["abas"])
        summary["total_items"] = result["total_items"]
        summary["mat"] = sum(a["cat_counts"]["material"] for a in result["abas"])
        summary["mo"] = sum(a["cat_counts"]["mao_obra"] for a in result["abas"])
        summary["eq"] = sum(a["cat_counts"]["equipamento"] for a in result["abas"])
        summary["outros"] = sum(a["cat_counts"]["outros"] for a in result["abas"])
    except Exception as e:
        summary["status"] = "failed"
        summary["error"] = f"{type(e).__name__}: {e}"
        summary["traceback"] = traceback.format_exc()[-300:]

    summary["duration_s"] = round(time.time() - started, 2)
    log_event(summary)
    return summary


def main():
    queue = json.loads(QUEUE.read_text(encoding="utf-8"))
    print(f"processing {len(queue)} projects -> {OUT_DIR}")
    counts = {}
    t0 = time.time()
    for i, p in enumerate(queue, 1):
        s = process_one(p)
        counts[s["status"]] = counts.get(s["status"], 0) + 1
        if s["status"] == "done":
            print(f"[{i}/{len(queue)}] {p['projeto']:<45} done   abas={s['abas']:>2} items={s['total_items']:>5} (mat={s['mat']:>4} mo={s['mo']:>4} eq={s['eq']:>3} other={s['outros']:>4}) {s['duration_s']}s", flush=True)
    print()
    print(f"summary: {counts}")
    print(f"total time: {int(time.time() - t0)}s")


if __name__ == "__main__":
    main()
