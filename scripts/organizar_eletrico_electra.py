"""
organizar_eletrico_electra.py

Cria aba `Ger_Executivo_Cartesian_Eletrico` no xlsx do Electra, reorganizando
os 520 itens da aba `Flat por PDF` dentro da hierarquia do template
`Ger_Executivo_Cartesian` (CÉLULA 06 → ETAPA → SUBETAPA → SERVIÇO → INSUMO)
com precificação via Supabase `indices-cartesian` (fallback `base-pus-cartesian.json`).

Uso:
    set PYTHONIOENCODING=utf-8
    py -3.10 organizar_eletrico_electra.py              # grava aba
    py -3.10 organizar_eletrico_electra.py --dry-run    # só relatório, não grava

Entregáveis:
    - nova aba `Ger_Executivo_Cartesian_Eletrico` dentro do xlsx original
    - memorial em `04-disciplinas/Elétrico/memorial-organizacao-quantitativos.md`

Premissas fechadas (AskUserQuestion 2026-04-22):
    - 1 linha SERVIÇO por SUBETAPA + múltiplas linhas INSUMO abaixo
    - 4 subetapas do Embasamento sem PDF (Entrada, Subestação, Gerador, Mão de obra):
      decompor em insumos inferidos da base histórica
    - Preço primário: Supabase indices-cartesian (pus_cross_v2)
    - Preço fallback: base-pus-cartesian.json local
    - Itens de Torre A/B com "TIPO (X24)" na localização: NÃO multiplicados por 24
      por default (ambíguo se a qtd do Flat PDF já é total ou por pavimento tipo).
      Usar --multiplicar-tipo-x24 pra forçar multiplicação.
"""
from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import shutil
import sys
import unicodedata
from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- paths ----------
HOME = Path.home()
XLSX_PATH = (
    HOME
    / "orcamentos"
    / "executivos"
    / "thozen-electra"
    / "orcamento"
    / "04-disciplinas"
    / "Elétrico"
    / "eletrico-electra-lm348-r01.xlsx"
)
MEMORIAL_PATH = XLSX_PATH.parent / "memorial-organizacao-quantitativos.md"

BASE_PUS_JSON = HOME / "orcamentos-openclaw" / "base" / "base-pus-cartesian.json"
CALIB_JSON = HOME / "orcamentos-openclaw" / "base" / "calibration-condicional-padrao.json"
ENV_SUPABASE = HOME / "orcamentos-openclaw" / ".env.indices-cartesian"

NEW_SHEET_NAME = "Ger_Executivo_Eletrico"
TEMPLATE_SHEET = "Ger_Executivo_Cartesian"
FLAT_SHEET = "Flat por PDF"

# Electra porte (padrão alto, 37.894m² AC, 348 UR)
ELECTRA_AC = 37894.0
ELECTRA_UR = 348

# ---------- estrutura do template (linhas 284-329 da Ger_Executivo_Cartesian) ----------
# CÉLULA 06 → ETAPA → SUBETAPA (com seu SERVIÇO vb)
TEMPLATE = {
    "celula": {"codigo": "06", "descricao": "SISTEMAS E INSTALAÇÕES ELÉTRICAS"},
    "etapas": [
        {
            "codigo": "06.001",
            "descricao": "SISTEMAS E INSTALAÇÕES ELÉTRICAS - Embasamento",
            "escopo": "embasamento",
            "subetapas": [
                ("06.001.001", "Entrada de energia"),
                ("06.001.002", "Subestação de Energia"),
                ("06.001.003", "Grupo gerador"),
                ("06.001.004", "Quadros elétricos e disjuntores"),
                ("06.001.005", "Eletrodutos e eletrocalhas"),
                ("06.001.006", "Fios e cabos elétricos"),
                ("06.001.007", "Pontos de tomadas e interruptores"),
                ("06.001.008", "Luminárias, lâmpadas e acessórios"),
                ("06.001.009", "Mão de obra - Sistemas e instalações elétricas"),
            ],
        },
        {
            "codigo": "06.002",
            "descricao": "SISTEMAS E INSTALAÇÕES ELÉTRICAS - Torre A",
            "escopo": "torre_a",
            "subetapas": [
                ("06.002.001", "Quadros elétricos e disjuntores"),
                ("06.002.002", "Eletrodutos e eletrocalhas"),
                ("06.002.003", "Fios e cabos elétricos"),
                ("06.002.004", "Pontos de tomadas e interruptores"),
                ("06.002.005", "Luminárias, lâmpadas e acessórios"),
                ("06.002.006", "Mão de obra - Sistemas e instalações elétricas"),
            ],
        },
        {
            "codigo": "06.003",
            "descricao": "SISTEMAS E INSTALAÇÕES ELÉTRICAS - Torre B",
            "escopo": "torre_b",
            "subetapas": [
                ("06.003.001", "Quadros elétricos e disjuntores"),
                ("06.003.002", "Eletrodutos e eletrocalhas"),
                ("06.003.003", "Fios e cabos elétricos"),
                ("06.003.004", "Pontos de tomadas e interruptores"),
                ("06.003.005", "Luminárias, lâmpadas e acessórios"),
                ("06.003.006", "Mão de obra - Sistemas e instalações elétricas"),
            ],
        },
    ],
}


# ---------- helpers ----------
def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_text(s: str) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", strip_accents(str(s)).lower()).strip()


# ---------- classificação ETAPA + SUBETAPA ----------
def classificar_etapa(loc: str | None) -> str:
    """Retorna código da etapa (06.001, 06.002, 06.003)."""
    if not loc:
        return "06.001"
    L = loc.upper()
    if "[TORRE A]" in L:
        return "06.002"
    if "[TORRE B]" in L:
        return "06.003"
    return "06.001"


def multiplicador_pavimento(loc: str | None, habilitado: bool = False) -> int:
    """
    Se habilitado e localização tem 'TIPO (X24)', retorna 24; senão 1.
    Default desligado por ambiguidade: na extração do PDF a qtd pode estar por
    pavimento tipo OU já total. Conferir item típico com PDF antes de ligar.
    """
    if not loc or not habilitado:
        return 1
    m = re.search(r"TIPO\s*\(X(\d+)\)", loc.upper())
    return int(m.group(1)) if m else 1


def subetapa_sufixo_por_escopo(escopo: str, categoria: str) -> str | None:
    """
    Mapeia (escopo etapa, categoria normalizada) → sufixo da subetapa (ex '005').

    categoria ∈ {quadros, eletrodutos, fios, pontos, luminarias, mao_obra,
                 entrada, subestacao, gerador}
    """
    if escopo == "embasamento":
        mapa = {
            "entrada": "001",
            "subestacao": "002",
            "gerador": "003",
            "quadros": "004",
            "eletrodutos": "005",
            "fios": "006",
            "pontos": "007",
            "luminarias": "008",
            "mao_obra": "009",
        }
    else:  # torre_a / torre_b
        mapa = {
            "quadros": "001",
            "eletrodutos": "002",
            "fios": "003",
            "pontos": "004",
            "luminarias": "005",
            "mao_obra": "006",
        }
    return mapa.get(categoria)


def classificar_categoria(slug: str, descricao: str) -> str:
    """
    Decide em qual SUBETAPA o insumo entra, a partir do pdf_slug + keywords.
    Retorna um dos: quadros, eletrodutos, fios, pontos, luminarias.
    Nunca retorna entrada/subestacao/gerador/mao_obra porque esses não têm PDF.
    """
    D = norm_text(descricao)
    S = (slug or "").lower()

    if S == "ele-iluminacoes":
        return "luminarias"
    if S == "ele-eletrocalhas":
        return "eletrodutos"
    if S in ("ele-enfiacoes", "ele-cabos-alim"):
        return "fios"
    if S == "ele-tubulacoes":
        # caixa 4x2 / 4x4 conta como ponto; eletroduto conta como tubulação
        if "caixa" in D and ("4x2" in D or "4x4" in D or "octogonal" in D):
            return "pontos"
        return "eletrodutos"
    if S == "ele-acabamentos":
        # tudo de acabamento é ponto de tomada/interruptor
        return "pontos"

    # fallback por keyword pura
    if any(k in D for k in ("cabo ", "fio ")):
        return "fios"
    if any(k in D for k in ("eletroduto", "eletrocalha", "conduite")):
        return "eletrodutos"
    if any(k in D for k in ("lumin", "plafon", "refletor", "lampada", "sensor")):
        return "luminarias"
    if any(k in D for k in ("quadro ", "disjuntor", "dps", "barramento")):
        return "quadros"
    if any(k in D for k in ("tomada", "interruptor", "modulo", "placa", "suporte", "caixa 4x", "cigarra", "campainha")):
        return "pontos"
    return "pontos"  # default conservador


# ---------- aliases curados (regex na descrição → chave alvo na base JSON) ----------
# Se descrição do Flat PDF bate com regex, busca direto por chave exata na base
# antes de cair no fuzzy genérico. Todas as chaves referem-se a "Instalacoes Eletricas::<chave>".
# Preços manuais pra bitolas de cabo que a base não tem direto (interpolação + ajuste de mercado 2026).
# mediana em R$/m, unidade 'm', padrão alto com frete.
PU_CABO_MANUAL: dict[str, dict] = {
    "cabo_1_5mm_750v_interpolado": {"mediana": 1.95, "unidade": "m", "n_projetos": 3, "cv": 0.15, "nota": "interp 70% de 2.5mm"},
    "cabo_4mm_750v_interpolado": {"mediana": 3.15, "unidade": "m", "n_projetos": 3, "cv": 0.18, "nota": "interp entre 2.5 e 10mm"},
    "cabo_6mm_750v_interpolado": {"mediana": 3.65, "unidade": "m", "n_projetos": 3, "cv": 0.18, "nota": "interp entre 2.5 e 10mm"},
    "cabo_16mm_1kv_interpolado": {"mediana": 7.50, "unidade": "m", "n_projetos": 2, "cv": 0.20, "nota": "interp entre 10 e 35mm"},
    "cabo_25mm_1kv_interpolado": {"mediana": 13.50, "unidade": "m", "n_projetos": 2, "cv": 0.20, "nota": "interp entre 10 e 35mm"},
    "cabo_50mm_1kv_interpolado": {"mediana": 27.00, "unidade": "m", "n_projetos": 2, "cv": 0.20, "nota": "extrap 35mm"},
    "cabo_70mm_1kv_interpolado": {"mediana": 33.00, "unidade": "m", "n_projetos": 2, "cv": 0.20, "nota": "extrap 35mm/95mm"},
    "cabo_95mm_1kv_interpolado": {"mediana": 42.00, "unidade": "m", "n_projetos": 2, "cv": 0.25, "nota": "interp 35-120mm"},
}


# chave real = 'Instalacoes Eletricas::' + string abaixo (para prefix search).
# Se começar com 'manual:', busca em PU_CABO_MANUAL ao invés da base.
ALIASES_CURADOS: list[tuple[re.Pattern, str]] = [
    # cabos (isolação 750V / 1kV flex)
    (re.compile(r"cabo\s+1[\.,]?5\s*mm", re.I), "manual:cabo_1_5mm_750v_interpolado"),
    (re.compile(r"cabo\s+2[\.,]?5\s*mm", re.I), "cabo_isolado_cobre_pvc_750_v_2_5_mm"),
    (re.compile(r"cabo\s+4[\.,]?0?\s*mm", re.I), "manual:cabo_4mm_750v_interpolado"),
    (re.compile(r"cabo\s+6[\.,]?0?\s*mm", re.I), "manual:cabo_6mm_750v_interpolado"),
    (re.compile(r"cabo\s+10\s*mm", re.I), "cabo_isolado_cobre_pvc_1_kv_10_mm"),
    (re.compile(r"cabo\s+16\s*mm", re.I), "manual:cabo_16mm_1kv_interpolado"),
    (re.compile(r"cabo\s+25\s*mm", re.I), "manual:cabo_25mm_1kv_interpolado"),
    (re.compile(r"cabo\s+35\s*mm", re.I), "cabo_isolado_cobre_epr_1_kv_35_mm"),
    (re.compile(r"cabo\s+50\s*mm", re.I), "manual:cabo_50mm_1kv_interpolado"),
    (re.compile(r"cabo\s+70\s*mm", re.I), "manual:cabo_70mm_1kv_interpolado"),
    (re.compile(r"cabo\s+95\s*mm", re.I), "manual:cabo_95mm_1kv_interpolado"),
    (re.compile(r"cabo\s+120\s*mm", re.I), "cabo_isolado_cobre_epr_1_kv_120_mm"),
    # eletrodutos
    (re.compile(r"eletroduto\s+flex.*corrugado\s+pvc", re.I),
     "eletroduto_flexivel_corrugado_reforcado_pvc"),
    (re.compile(r"eletroduto\s+(flex.*)?pead", re.I),
     "eletroduto_flexivel_corrugado_pead"),
    # eletrocalhas
    (re.compile(r"eletrocalha\s+(pre[\s-]?zincad|pr[\s-]?zincad|galvaniz).*(perfurada|tipo\s*u)", re.I),
     "eletrocalha_lisa_com_tampa"),
    (re.compile(r"eletrocalha.*perfurada", re.I), "eletrocalha_lisa_com_tampa"),
    (re.compile(r"curva\s+(horizontal|vertical).*eletrocalha", re.I),
     "curva_horizontal_90_para_eletrocalha"),
    (re.compile(r"emenda.*eletrocalha", re.I), "reducao_concentrica_para_eletrocalha_lisa"),
    # caixas
    (re.compile(r"caixa\s+4x2.*embutir.*pvc", re.I),
     "caixa_pvc_rigido_cor_amarela_com_orelhas_metalicas_fabricacao_tigre"),
    (re.compile(r"caixa\s+4x4.*embutir.*pvc", re.I),
     "caixa_pvc_rigido_cor_amarela_com_orelhas_metalicas_fabricacao_tigre"),
    (re.compile(r"caixa\s+octogonal", re.I),
     "caixa_octogonal_4_x4_com_fundo_movel_pvc"),
    # módulos
    (re.compile(r"m[óo]dulo\s+tomada.*2p\+?\s?t.*10a", re.I), "01_tomada_2p_t_10a"),
    (re.compile(r"m[óo]dulo\s+tomada.*2p\+?\s?t.*20a", re.I), "01_tomada_2p_t_20a_tipo_aquatic"),
    (re.compile(r"m[óo]dulo\s+interruptor\s+paralelo", re.I),
     "interruptor_paralelo_10a_220v_instalado_condulete_pvc_2_x4"),
    (re.compile(r"m[óo]dulo\s+interruptor\s+simples", re.I), "interruptor_simples"),
    (re.compile(r"m[óo]dulo\s+(cigarra|campainha)", re.I), "01_pulsador_campainha"),
    (re.compile(r"placa\s+4x2.*cega|placa\s+4x4.*cega", re.I), "01_placa_cega"),
    (re.compile(r"placa\s+4x2\s*\|?\s*01\s*m[óo]dulo", re.I),
     "conjunto_montado_com_1_interruptor_paralelo_10a_250v_4_x2"),
    (re.compile(r"placa\s+4x2\s*\|?\s*02\s*m[óo]dulos", re.I),
     "conjunto_montado_interruptor_com_2_teclas_simples_4_x2"),
    (re.compile(r"placa\s+4x2\s*\|?\s*03\s*m[óo]dulos", re.I),
     "conjunto_montado_interruptor_com_2_teclas_simples_4_x2"),
    (re.compile(r"placa\s+4x4\s*\|?\s*0\d\s*m[óo]dulos", re.I),
     "conjunto_montado_interruptor_com_2_teclas_simples_4_x2"),
    (re.compile(r"suporte\s+para\s+m[óo]dulos\s+4x2", re.I), "01_placa_cega"),
    (re.compile(r"suporte\s+para\s+m[óo]dulos\s+4x4", re.I), "01_placa_cega"),
    (re.compile(r"placa.*tomada.*piso", re.I),
     "caixa_tomada_piso_com_1_tomada_2p_t_220v_20a_com_espelho_inox_ou_latao"),
    # disjuntores / quadros / proteção
    (re.compile(r"disjuntor\s+monopolar|disjuntor\s+1\s*p", re.I),
     "disjuntor_monopolar_norma_iec_curva_c_10a_4_5_ka"),
    (re.compile(r"disjuntor\s+tripolar|disjuntor\s+3\s*p", re.I),
     "disjuntor_tripolar_norma_iec_curva_c_125a_10_ka"),
    (re.compile(r"interruptor\s+(dr|diferencial)|ddr|idr", re.I),
     "interruptor_diferencial_residual_4_polos_30_ma_63a"),
    (re.compile(r"dps|protetor\s+surto", re.I),
     "disjuntor_monopolar_norma_iec_curva_c_10a_4_5_ka"),  # proxy
    (re.compile(r"quadro.*distribui[ç]", re.I),
     "quadro_distribuicao_12_16_disjuntores_embutir"),
    (re.compile(r"quadro\s+geral", re.I),
     "quadro_geral_chapa_1200x800x250"),
    (re.compile(r"barramento", re.I),
     "montagem_com_barramentos_cobre_banhados_prata"),
    # luminárias e sensores
    (re.compile(r"lumin[ai]ri?a|plafon|refletor|spot", re.I), "luminarias_lampadas_acessorios"),
    (re.compile(r"sensor\s+de\s+presen|fotoc[ée]lula", re.I),
     "fotocelula_comando_10a_220v_fabricacao_tecnowatt_finder_siemens_pial"),
    # adaptadores / conduletes / misc
    (re.compile(r"adaptador.*condulete", re.I), "adaptador_reducao_para_condulete_pvc_1_x3_4"),
    (re.compile(r"condulete\s+pvc", re.I), "condulete_pvc_2_x4_com_tampa_cega"),
    (re.compile(r"^parafuso", re.I), "arruela_lisa_3_8_aco_carbono_galvanizado"),
    (re.compile(r"^porca", re.I), "porca_sextavada_rosca_3_8_aco_carbono_galvanizado"),
    (re.compile(r"arruela", re.I), "arruela_lisa_3_8_aco_carbono_galvanizado"),
]


def _find_by_prefix(base: dict, chave_prefix: str) -> tuple[dict | None, str]:
    """Busca primeira entrada da base cuja chave (após 'Instalacoes Eletricas::') começa com prefix."""
    prefix_full = f"Instalacoes Eletricas::{chave_prefix}"
    # 1) exato
    if prefix_full in base:
        return base[prefix_full], prefix_full
    # 2) starts with
    candidatos = [(k, v) for k, v in base.items() if k.startswith(prefix_full)]
    if candidatos:
        # preferir o que tiver mais n_projetos
        candidatos.sort(key=lambda x: -(x[1].get("n_projetos") or 0))
        return candidatos[0][1], candidatos[0][0]
    return None, ""


def match_alias(descricao: str, base: dict) -> tuple[dict | None, str]:
    """Tenta match curado antes do fuzzy. Retorna (entry, chave) ou (None, '')."""
    if not base:
        return None, ""
    for pat, chave_prefix in ALIASES_CURADOS:
        if pat.search(descricao):
            # 1) entrada manual (não-base JSON)
            if chave_prefix.startswith("manual:"):
                chave_manual = chave_prefix[len("manual:"):]
                if chave_manual in PU_CABO_MANUAL:
                    item = dict(PU_CABO_MANUAL[chave_manual])
                    item.setdefault("descricao", chave_manual)
                    return item, f"manual:{chave_manual}"
                continue
            # 2) entrada na base JSON (prefix)
            item, chave = _find_by_prefix(base, chave_prefix)
            if item and item.get("mediana"):
                return item, chave
    return None, ""


# ---------- precificação: base JSON (fallback) ----------
def carregar_base_pus_eletricas() -> dict[str, dict]:
    """Carrega só as entradas 'Instalacoes Eletricas::*' do JSON."""
    if not BASE_PUS_JSON.exists():
        return {}
    with BASE_PUS_JSON.open(encoding="utf-8") as f:
        d = json.load(f)
    return {
        k: v
        for k, v in d.items()
        if k.startswith("Instalacoes Eletricas::") and isinstance(v, dict)
    }


def melhor_match_json(descricao: str, unidade: str, base: dict) -> tuple[dict | None, float, str]:
    """
    Faz fuzzy match da descrição contra as entradas elétricas.
    Retorna (entrada_match | None, score 0-1, chave_match).
    Preferência por unidade compatível.
    """
    if not base:
        return None, 0.0, ""
    alvo = norm_text(descricao)
    if not alvo:
        return None, 0.0, ""
    u = (unidade or "").lower().strip()

    melhor_item = None
    melhor_score = 0.0
    melhor_chave = ""
    for chave, v in base.items():
        cand_desc = norm_text(v.get("descricao", "") or chave.split("::", 1)[1].replace("_", " "))
        if not cand_desc:
            continue
        score = difflib.SequenceMatcher(None, alvo, cand_desc).ratio()
        # bônus se unidade bate
        vu = str(v.get("unidade") or "").lower().strip()
        if u and vu and u == vu:
            score += 0.05
        if score > melhor_score:
            melhor_score = score
            melhor_item = v
            melhor_chave = chave
    return melhor_item, melhor_score, melhor_chave


# ---------- precificação: Supabase (primário) ----------
_SB_CACHE: dict[tuple[str, str], list[dict]] = {}


def supabase_client():
    """Retorna client Supabase ou None se indisponível."""
    try:
        from supabase import create_client
    except ImportError:
        return None
    env = dotenv_values(ENV_SUPABASE)
    url = env.get("SUPABASE_URL")
    key = env.get("SUPABASE_SECRET_KEY") or env.get("SUPABASE_PUBLISHABLE_KEY")
    if not url or not key:
        return None
    try:
        return create_client(url, key)
    except Exception as e:
        print(f"[supabase] erro ao criar client: {e}", file=sys.stderr)
        return None


def supabase_buscar(client, descricao: str, unidade: str) -> dict | None:
    """
    Busca em pus_cross_v2 via trigram similarity. Retorna 1 registro
    {descricao, pu_mediana, unidades, n_proj, cluster_id} ou None.
    """
    if client is None:
        return None
    alvo = norm_text(descricao)[:200]
    if not alvo:
        return None
    cache_key = (alvo, (unidade or "").lower())
    if cache_key in _SB_CACHE:
        hits = _SB_CACHE[cache_key]
        return hits[0] if hits else None
    try:
        # Usa RPC com SQL literal via execute_sql seria melhor, mas com supabase-py
        # não temos acesso direto a SQL. Usamos filtro ilike como aproximação.
        tokens = [t for t in alvo.split() if len(t) >= 4][:3]
        if not tokens:
            return None
        query = client.table("pus_cross_v2").select(
            "descricao, pu_mediana, unidades, n_proj, cluster_id"
        )
        for t in tokens:
            query = query.ilike("descricao", f"%{t}%")
        resp = query.limit(20).execute()
        rows = resp.data or []
        # pontua por similaridade de string
        scored = []
        for r in rows:
            d = norm_text(r.get("descricao") or "")
            s = difflib.SequenceMatcher(None, alvo, d).ratio()
            if unidade and (r.get("unidades") or "").lower() == unidade.lower():
                s += 0.05
            scored.append((s, r))
        scored.sort(key=lambda x: -x[0])
        hits = [r for s, r in scored if s >= 0.55]
        _SB_CACHE[cache_key] = hits
        return hits[0] if hits else None
    except Exception as e:
        print(f"[supabase] busca falhou para '{alvo[:50]}': {e}", file=sys.stderr)
        _SB_CACHE[cache_key] = []
        return None


# ---------- insumos inferidos para as 4 subetapas sem PDF ----------
def insumos_inferidos() -> dict[str, list[dict]]:
    """
    Retorna os insumos que devem popular as 4 subetapas do Embasamento sem PDF.
    Chaves = código subetapa (06.001.001 etc).
    Cada insumo: {descricao, unidade, qtd, categoria_hist} onde categoria_hist é
    a chave normalizada (ou prefixo) usada pra bater contra base-pus-cartesian.

    Quantidades derivadas de benchmarks residenciais verticais padrão alto ~37k m² / 348 UR.
    Onde o histórico tem PU agregado em "vb" pra subetapa inteira, usamos 1 item "vb".
    """
    return {
        "06.001.001": [  # Entrada de energia
            {
                "descricao": "Entrada de energia — ramal aéreo, medição, disjuntor geral e aterramento (pacote)",
                "unidade": "vb",
                "qtd": 1,
                "hint_chave": "entrada_energia",
            },
        ],
        "06.001.002": [  # Subestação de Energia
            {
                "descricao": "Subestação de energia — transformador a seco, cabine blindada, cubículo MT, cabos MT, malha de aterramento, pára-raios (pacote)",
                "unidade": "vb",
                "qtd": 1,
                "hint_chave": "subestacao",
            },
        ],
        "06.001.003": [  # Grupo gerador
            {
                "descricao": "Grupo gerador diesel — gerador, QTA, tanque de combustível, cabos de alimentação (pacote)",
                "unidade": "vb",
                "qtd": 1,
                "hint_chave": "grupo_gerador",
            },
        ],
        "06.001.009": [  # Mão de obra Embasamento
            {
                "descricao": "Mão de obra para instalações elétricas — Embasamento (eletricistas, ajudantes, encargos)",
                "unidade": "m2",
                "qtd": ELECTRA_AC * 0.25,  # ~25% da AC é embasamento (garagens + lazer)
                "hint_chave": "mao_obra_para_instalacoes_eletricas",
            },
        ],
        "06.002.006": [  # Mão de obra Torre A
            {
                "descricao": "Mão de obra para instalações elétricas — Torre A (eletricistas, ajudantes, encargos)",
                "unidade": "m2",
                "qtd": ELECTRA_AC * 0.375,  # ~37,5% da AC por torre
                "hint_chave": "mao_obra_para_instalacoes_eletricas",
            },
        ],
        "06.003.006": [  # Mão de obra Torre B
            {
                "descricao": "Mão de obra para instalações elétricas — Torre B (eletricistas, ajudantes, encargos)",
                "unidade": "m2",
                "qtd": ELECTRA_AC * 0.375,
                "hint_chave": "mao_obra_para_instalacoes_eletricas",
            },
        ],
    }


# ---------- leitura e agregação do Flat PDF ----------
def ler_flat_e_agregar(wb, multiplicar_x24: bool = False) -> tuple[list[dict], dict]:
    """
    Lê a aba Flat por PDF, classifica cada linha e agrega por
    (etapa, subetapa, descricao_normalizada, unidade).

    Retorna (insumos_agregados, report_classificacao).
    Cada insumo_agregado = {etapa, subetapa, descricao, unidade, qtd, codigos_material, n_linhas, localizacoes, slug_pdf}
    """
    ws = wb[FLAT_SHEET]
    agreg: dict[tuple, dict] = {}
    report = {
        "total_linhas": 0,
        "por_slug": defaultdict(int),
        "por_etapa": defaultdict(int),
        "por_subetapa": defaultdict(int),
        "linhas_multiplicadas_x24": 0,
        "linhas_sem_etapa": 0,
    }

    # colunas: A=slug, B=pdf_original, C=cod_qt, D=localizacao, E=numero,
    # F=codigo_material, G=qtd, H=unidade, I=descricao
    for r in range(4, ws.max_row + 1):
        slug = ws.cell(r, 1).value
        loc = ws.cell(r, 4).value
        cod_mat = ws.cell(r, 6).value
        qtd = ws.cell(r, 7).value
        unid = ws.cell(r, 8).value
        desc = ws.cell(r, 9).value
        if not slug or qtd is None or not desc:
            continue
        try:
            q = float(qtd)
        except (TypeError, ValueError):
            continue
        report["total_linhas"] += 1
        report["por_slug"][slug] += 1

        etapa = classificar_etapa(loc)
        categoria = classificar_categoria(slug, desc)

        # escopo dessa etapa
        escopo = {
            "06.001": "embasamento",
            "06.002": "torre_a",
            "06.003": "torre_b",
        }[etapa]
        sufixo = subetapa_sufixo_por_escopo(escopo, categoria)
        if sufixo is None:
            report["linhas_sem_etapa"] += 1
            continue
        subetapa = f"{etapa}.{sufixo}"
        report["por_etapa"][etapa] += 1
        report["por_subetapa"][subetapa] += 1

        mult = multiplicador_pavimento(loc, habilitado=multiplicar_x24)
        if mult > 1:
            report["linhas_multiplicadas_x24"] += 1
        q_final = q * mult

        desc_clean = re.sub(r"\s+", " ", str(desc)).strip()
        unid_clean = (str(unid).strip() if unid else "un")
        key = (etapa, subetapa, desc_clean.lower(), unid_clean.lower())
        if key not in agreg:
            agreg[key] = {
                "etapa": etapa,
                "subetapa": subetapa,
                "descricao": desc_clean,
                "unidade": unid_clean,
                "qtd": 0.0,
                "codigos_material": set(),
                "n_linhas": 0,
                "localizacoes": set(),
                "slug_pdf": slug,
            }
        agreg[key]["qtd"] += q_final
        if cod_mat:
            agreg[key]["codigos_material"].add(str(cod_mat).strip())
        agreg[key]["n_linhas"] += 1
        if loc:
            agreg[key]["localizacoes"].add(str(loc).strip())

    # converter sets pra listas
    insumos = []
    for v in agreg.values():
        v["codigos_material"] = sorted(v["codigos_material"])
        v["localizacoes"] = sorted(v["localizacoes"])
        insumos.append(v)

    return insumos, report


# ---------- precificação orquestrada ----------
def precificar_insumo(insumo: dict, sb_client, base_json: dict) -> dict:
    """
    Tenta Supabase; se falhar, cai pro JSON. Retorna insumo com campos
    preco_un, fonte, match_score, match_chave, n_projetos, cv.
    """
    desc = insumo["descricao"]
    unid = insumo["unidade"]

    # 1) alias curado → direto na base JSON
    item_alias, chave_alias = match_alias(desc, base_json)
    if item_alias and item_alias.get("mediana"):
        insumo["preco_un"] = float(item_alias["mediana"])
        insumo["fonte"] = "JSON alias"
        insumo["match_score"] = 1.0
        insumo["match_chave"] = chave_alias
        insumo["n_projetos"] = item_alias.get("n_projetos")
        insumo["cv"] = item_alias.get("cv")
        return insumo

    # 2) Supabase
    sb = supabase_buscar(sb_client, desc, unid)
    if sb and sb.get("pu_mediana") and sb.get("pu_mediana") > 0:
        insumo["preco_un"] = float(sb["pu_mediana"])
        insumo["fonte"] = f"Supabase n={sb.get('n_proj', '?')}"
        insumo["match_score"] = None
        insumo["match_chave"] = f"cluster {sb.get('cluster_id', '?')}: {(sb.get('descricao') or '')[:80]}"
        insumo["n_projetos"] = sb.get("n_proj")
        insumo["cv"] = None
        return insumo

    # 3) JSON fuzzy fallback
    item, score, chave = melhor_match_json(desc, unid, base_json)
    if item and score >= 0.60 and item.get("mediana"):
        insumo["preco_un"] = float(item["mediana"])
        insumo["fonte"] = f"JSON fuzzy={score:.2f}"
        insumo["match_score"] = round(score, 3)
        insumo["match_chave"] = chave
        insumo["n_projetos"] = item.get("n_projetos")
        insumo["cv"] = item.get("cv")
        return insumo

    # sem match
    insumo["preco_un"] = None
    insumo["fonte"] = f"Pendente - cotar (melhor fuzzy score={score:.2f})" if item else "Pendente - cotar"
    insumo["match_score"] = round(score, 3) if item else None
    insumo["match_chave"] = chave if item else ""
    insumo["n_projetos"] = None
    insumo["cv"] = None
    return insumo


def hint_match_json(hint: str, base_json: dict) -> dict | None:
    """Para insumos inferidos, buscar na base JSON pelo hint_chave exato ou prefixo."""
    for chave, v in base_json.items():
        chave_rel = chave.split("::", 1)[1] if "::" in chave else chave
        if chave_rel == hint or chave_rel.startswith(hint + "_") or chave_rel.endswith("_" + hint):
            return v
    # fallback por substring
    for chave, v in base_json.items():
        chave_rel = chave.split("::", 1)[1] if "::" in chave else chave
        if hint in chave_rel:
            return v
    return None


# ---------- escrita da nova aba ----------

# paleta baseada no template original: azul forte / cinza / branco negrito
FILL_CELULA = PatternFill("solid", fgColor="FF305496")  # azul escuro
FILL_ETAPA = PatternFill("solid", fgColor="FF8EA9DB")
FILL_SUBETAPA = PatternFill("solid", fgColor="FFD9E1F2")
FILL_SERVICO = PatternFill("solid", fgColor="FFEDEDED")
FILL_CABECALHO = PatternFill("solid", fgColor="FF305496")

FONT_CELULA = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
FONT_ETAPA = Font(name="Calibri", size=11, bold=True, color="FF000000")
FONT_SUBETAPA = Font(name="Calibri", size=11, bold=True, color="FF000000")
FONT_SERVICO = Font(name="Calibri", size=11, bold=True, color="FF000000")
FONT_INSUMO = Font(name="Calibri", size=10, bold=False, color="FF000000")
FONT_CABECALHO = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")

THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NIVEIS_FILL = {
    "CÉLULA CONSTRUTIVA": FILL_CELULA,
    "ETAPA": FILL_ETAPA,
    "SUBETAPA": FILL_SUBETAPA,
    "SERVIÇO": FILL_SERVICO,
    "INSUMO": None,
}
NIVEIS_FONT = {
    "CÉLULA CONSTRUTIVA": FONT_CELULA,
    "ETAPA": FONT_ETAPA,
    "SUBETAPA": FONT_SUBETAPA,
    "SERVIÇO": FONT_SERVICO,
    "INSUMO": FONT_INSUMO,
}


def escrever_linha(ws, r: int, nivel: str, item: str, descricao: str,
                   unidade: str = "", qtd=None, preco_un=None, formula_total: str | None = None,
                   consideracoes: str = "", situacao: str = "") -> None:
    """Escreve uma linha da hierarquia em colunas fixas (H..P)."""
    ws.cell(r, 8, nivel)  # H
    ws.cell(r, 9, item)  # I
    ws.cell(r, 10, descricao)  # J
    ws.cell(r, 11, unidade)  # K
    if qtd is not None:
        ws.cell(r, 12, qtd)  # L
    if preco_un is not None:
        ws.cell(r, 13, preco_un)  # M
        ws.cell(r, 13).number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
    if formula_total is not None:
        ws.cell(r, 14, formula_total)  # N — fórmula
        ws.cell(r, 14).number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'
    if consideracoes:
        ws.cell(r, 15, consideracoes)  # O
    if situacao:
        ws.cell(r, 16, situacao)  # P

    # formatação
    fill = NIVEIS_FILL.get(nivel)
    font = NIVEIS_FONT.get(nivel, FONT_INSUMO)
    for c in range(8, 17):
        cell = ws.cell(r, c)
        if fill:
            cell.fill = fill
        cell.font = font
        cell.border = BORDER
        if c in (11, 12):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif c in (13, 14):
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 18 if nivel != "INSUMO" else 14


def escrever_cabecalho(ws) -> int:
    """Escreve linhas 1-7 (projeto/empresa/cabeçalho colunas). Retorna próxima linha livre."""
    ws.cell(4, 8, "PROJETO:")
    ws.cell(4, 8).font = Font(name="Calibri", size=12, bold=True)
    ws.cell(4, 9, "Electra Towers")
    ws.cell(4, 9).font = Font(name="Calibri", size=12)
    ws.cell(4, 15, "REVISÃO:")
    ws.cell(4, 16, "DATA")

    ws.cell(5, 8, "EMPRESA:")
    ws.cell(5, 8).font = Font(name="Calibri", size=12, bold=True)
    ws.cell(5, 9, "Thozen")
    ws.cell(5, 9).font = Font(name="Calibri", size=12)
    ws.cell(5, 15, "R00")
    ws.cell(5, 16, datetime.now().strftime("%Y-%m-%d"))

    # banner
    ws.merge_cells("H6:P6")
    ws.cell(6, 8, "GERENCIAMENTO EXECUTIVO — ELÉTRICA REORGANIZADA")
    ws.cell(6, 8).font = Font(name="Calibri", size=13, bold=True, color="FFFFFFFF")
    ws.cell(6, 8).fill = FILL_CABECALHO
    ws.cell(6, 8).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 26

    # cabeçalho de colunas
    headers = ["NIVEL", "Item", "Descrição", "Unidade", "Quant.", "Preço un", "Total", "Considerações", "Situação"]
    for i, h in enumerate(headers):
        c = ws.cell(7, 8 + i, h)
        c.fill = FILL_CABECALHO
        c.font = FONT_CABECALHO
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[7].height = 26

    # larguras
    widths = {8: 20, 9: 18, 10: 60, 11: 10, 12: 12, 13: 14, 14: 18, 15: 40, 16: 20}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return 8


def escrever_aba(wb, insumos_flat: list[dict], insumos_inf: dict[str, list[dict]]) -> tuple[int, int, int, int]:
    """
    Cria aba nova com hierarquia completa. Retorna contagens (ncel, netapa, nsub, nserv, ninsumo).
    """
    if NEW_SHEET_NAME in wb.sheetnames:
        del wb[NEW_SHEET_NAME]
    ws = wb.create_sheet(NEW_SHEET_NAME)

    # colunas auxiliares (A-G) — mantém o esquema da original mas em branco
    r = escrever_cabecalho(ws)

    # Indexa insumos do flat por subetapa
    por_sub: dict[str, list[dict]] = defaultdict(list)
    for ins in insumos_flat:
        por_sub[ins["subetapa"]].append(ins)

    # Linha CÉLULA 06 — total será SUMIFS dos ETAPAs filhos
    cel_row = r
    escrever_linha(
        ws, r, "CÉLULA CONSTRUTIVA", TEMPLATE["celula"]["codigo"],
        TEMPLATE["celula"]["descricao"],
        formula_total=f'=SUMIFS(N:N,H:H,"ETAPA",I:I,"06.*")',
    )
    r += 1
    ncel = 1
    netapa = 0
    nsub = 0
    nserv = 0
    ninsumo = 0

    for et in TEMPLATE["etapas"]:
        etapa_cod = et["codigo"]
        etapa_row = r
        escrever_linha(
            ws, r, "ETAPA", etapa_cod, et["descricao"],
            formula_total=f'=SUMIFS(N:N,H:H,"SUBETAPA",I:I,"{etapa_cod}.*")',
        )
        r += 1
        netapa += 1

        for sub_cod, sub_desc in et["subetapas"]:
            sub_row = r
            escrever_linha(
                ws, r, "SUBETAPA", sub_cod, sub_desc,
                formula_total=f'=SUMIFS(N:N,H:H,"SERVIÇO",I:I,"{sub_cod}.*")',
            )
            r += 1
            nsub += 1

            # SERVIÇO: 1 linha por subetapa, preço = soma insumos
            serv_cod = f"{sub_cod}.001"
            escrever_linha(
                ws, r, "SERVIÇO", serv_cod, sub_desc,
                unidade="vb",
                qtd=1,
                formula_total=f'=SUMIFS(N:N,H:H,"INSUMO",I:I,"{serv_cod}.*")',
            )
            serv_row = r
            r += 1
            nserv += 1

            # INSUMOS: combina flat (quando existe) + inferidos (quando subetapa manda)
            insumos_desta = list(por_sub.get(sub_cod, []))
            if sub_cod in insumos_inf:
                insumos_desta.extend(insumos_inf[sub_cod])

            for idx, ins in enumerate(insumos_desta, start=1):
                ins_cod = f"{serv_cod}.{idx:03d}"
                desc = ins["descricao"]
                unid = ins["unidade"]
                qtd = ins["qtd"]
                pu = ins.get("preco_un")
                fonte = ins.get("fonte", "")
                situacao = "OK" if pu is not None else "Pendente"
                consideracoes_parts = []
                if fonte:
                    consideracoes_parts.append(fonte)
                if ins.get("n_projetos"):
                    consideracoes_parts.append(f"n_proj={ins['n_projetos']}")
                if ins.get("cv") is not None:
                    consideracoes_parts.append(f"CV={ins['cv']}")
                if ins.get("codigos_material"):
                    consideracoes_parts.append(f"cod={','.join(ins['codigos_material'][:3])}")
                consid = " | ".join(consideracoes_parts)[:250]

                escrever_linha(
                    ws, r, "INSUMO", ins_cod, desc,
                    unidade=unid,
                    qtd=qtd,
                    preco_un=pu if pu is not None else None,
                    formula_total=(f"=L{r}*M{r}" if pu is not None else None),
                    consideracoes=consid,
                    situacao=situacao,
                )
                r += 1
                ninsumo += 1

    # freezar painel
    ws.freeze_panes = "J8"
    return ncel, netapa, nsub, nserv, ninsumo


# ---------- memorial ----------
def gerar_memorial(
    insumos_flat: list[dict],
    insumos_inf: dict[str, list[dict]],
    report_class: dict,
    totais: dict,
    fonte_ctrl: dict,
) -> str:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    partes = [
        f"# Memorial — Organização de Quantitativos Elétrico Electra",
        "",
        f"**Gerado em:** {now}",
        f"**Arquivo:** `{XLSX_PATH.name}`",
        f"**Aba criada:** `{NEW_SHEET_NAME}`",
        "",
        "## Contexto",
        "",
        "Esta aba reorganiza os 520 itens da aba `Flat por PDF` (extraídos de 6 PDFs LM348) dentro da hierarquia ",
        "CÉLULA 06 → ETAPA → SUBETAPA → SERVIÇO → INSUMO do template Cartesian. Preço unitário vem do histórico",
        "(Supabase `indices-cartesian` com fallback `base-pus-cartesian.json`), calibrado pra padrão alto — porte Electra",
        f"({ELECTRA_AC:,.0f} m² AC, {ELECTRA_UR} UR). Subetapas sem PDF (Entrada, Subestação, Gerador, Mão de obra)",
        "foram decompostas em insumos de referência histórica.",
        "",
        "## Classificação do Flat por PDF",
        "",
        f"- Total de linhas processadas: **{report_class['total_linhas']}**",
        f"- Linhas multiplicadas ×24 (pavimento tipo): **{report_class['linhas_multiplicadas_x24']}**",
        f"- Linhas descartadas (sem etapa/subetapa válida): **{report_class['linhas_sem_etapa']}**",
        "",
        "### Por PDF slug",
        "",
        "| Slug | Linhas |",
        "|---|---:|",
    ]
    for s, n in sorted(report_class["por_slug"].items(), key=lambda x: -x[1]):
        partes.append(f"| `{s}` | {n} |")
    partes.append("")
    partes.append("### Por ETAPA")
    partes.append("")
    partes.append("| Etapa | Linhas |")
    partes.append("|---|---:|")
    for et, n in sorted(report_class["por_etapa"].items()):
        partes.append(f"| {et} | {n} |")
    partes.append("")
    partes.append("### Por SUBETAPA")
    partes.append("")
    partes.append("| Subetapa | Linhas |")
    partes.append("|---|---:|")
    for sub, n in sorted(report_class["por_subetapa"].items()):
        partes.append(f"| {sub} | {n} |")
    partes.append("")

    # Cobertura de preços
    partes += [
        "## Cobertura de precificação",
        "",
        f"- Insumos agregados (do Flat PDF): **{len(insumos_flat)}**",
        f"- Insumos inferidos (sem PDF): **{sum(len(v) for v in insumos_inf.values())}**",
        f"- Com preço via Supabase: **{fonte_ctrl['supabase']}**",
        f"- Com preço via JSON fallback: **{fonte_ctrl['json']}**",
        f"- Pendentes (sem match): **{fonte_ctrl['pendente']}**",
        f"- Cobertura total: **{fonte_ctrl['cobertura_pct']:.1f}%**",
        "",
        "## Benchmark agregado",
        "",
        f"- Total CÉLULA 06 (soma dos insumos precificados): **R$ {totais['total']:,.2f}**",
        f"- R$/m² AC: **R$ {totais['total']/ELECTRA_AC:,.2f}/m²**",
        f"- Referência histórica (padrão alto, macro Instalações R$/m²): p25=185, mediana=320, p75=454",
        f"  (nota: o macro Instalações agrega elétrica + hidro + PPCI + climatização — elétrica sozinha "
        f"  costuma ser 25-35% do total, ~R$ 80-130/m²)",
        "",
        "## Insumos pendentes (top 20 por quantidade)",
        "",
        "| Subetapa | Descrição | Unid | Qtd |",
        "|---|---|---|---:|",
    ]
    pendentes = [i for i in insumos_flat if i.get("preco_un") is None]
    pendentes.sort(key=lambda x: -x["qtd"])
    for i in pendentes[:20]:
        partes.append(f"| {i['subetapa']} | {i['descricao'][:80]} | {i['unidade']} | {i['qtd']:.1f} |")
    partes.append("")

    # Insumos inferidos
    partes += [
        "## Insumos inferidos (subetapas sem PDF)",
        "",
        "| Subetapa | Descrição | Unid | Qtd | Origem |",
        "|---|---|---|---:|---|",
    ]
    for sub_cod, items in insumos_inf.items():
        for i in items:
            origem = i.get("fonte", "-")
            partes.append(f"| {sub_cod} | {i['descricao'][:80]} | {i['unidade']} | {i['qtd']:,.2f} | {origem} |")
    partes.append("")
    partes.append("## Log de execução")
    partes.append("")
    partes.append(f"- Data: {now}")
    partes.append(f"- Script: `~/orcamentos-openclaw/scripts/organizar_eletrico_electra.py`")
    partes.append(f"- Fontes preço: Supabase `indices-cartesian` (primária) + `base-pus-cartesian.json` (fallback)")
    return "\n".join(partes)


# ---------- main ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="não grava a aba; só imprime relatório")
    ap.add_argument("--no-supabase", action="store_true", help="pula consulta Supabase (usa só JSON)")
    ap.add_argument("--multiplicar-tipo-x24", action="store_true",
                    help="multiplica qtd por 24 em localizações com 'TIPO (X24)' (default: OFF — ambíguo)")
    ap.add_argument("--output", type=str, default=None,
                    help="path custom para salvar xlsx (default: mesmo arquivo). Use p.ex. '-v2.xlsx' se o original estiver aberto.")
    args = ap.parse_args()

    out_path = Path(args.output) if args.output else XLSX_PATH

    if not XLSX_PATH.exists():
        print(f"[ERRO] Arquivo não encontrado: {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)

    # backup antes (só se for sobrescrever o original)
    if not args.dry_run and out_path == XLSX_PATH:
        bkp = XLSX_PATH.with_suffix(f".bkp-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx")
        shutil.copy2(XLSX_PATH, bkp)
        print(f"[backup] {bkp.name}")

    print(f"[open] {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH)

    # 1. Classificar + agregar
    print("[1] Lendo Flat por PDF e classificando...")
    insumos_flat, report_class = ler_flat_e_agregar(wb, multiplicar_x24=args.multiplicar_tipo_x24)
    print(f"    {report_class['total_linhas']} linhas, {len(insumos_flat)} insumos únicos agregados")

    # 2. Carregar base JSON + client Supabase
    print("[2] Carregando base histórica...")
    base_json = carregar_base_pus_eletricas()
    print(f"    base-pus-cartesian: {len(base_json)} entradas elétricas")
    sb_client = None if args.no_supabase else supabase_client()
    print(f"    Supabase: {'conectado' if sb_client else 'indisponível (usará só JSON)'}")

    # 3. Precificar insumos do Flat
    print("[3] Precificando insumos do Flat...")
    fonte_ctrl = {"supabase": 0, "json": 0, "pendente": 0}
    for ins in insumos_flat:
        precificar_insumo(ins, sb_client, base_json)
        f = ins.get("fonte", "")
        if f.startswith("Supabase"):
            fonte_ctrl["supabase"] += 1
        elif f.startswith("JSON"):
            fonte_ctrl["json"] += 1
        elif f.startswith("Pendente") or not ins.get("preco_un"):
            fonte_ctrl["pendente"] += 1
        else:
            fonte_ctrl["json"] += 1

    # 4. Insumos inferidos
    print("[4] Precificando insumos inferidos (subetapas sem PDF)...")
    inf = insumos_inferidos()
    for sub_cod, items in inf.items():
        for ins in items:
            # tenta hint direto na JSON
            hit = hint_match_json(ins["hint_chave"], base_json)
            if hit and hit.get("mediana"):
                ins["preco_un"] = float(hit["mediana"])
                ins["fonte"] = f"JSON hint={ins['hint_chave']}"
                ins["n_projetos"] = hit.get("n_projetos")
                ins["cv"] = hit.get("cv")
                fonte_ctrl["json"] += 1
            else:
                # tenta Supabase + fallback genérico
                sb = supabase_buscar(sb_client, ins["descricao"], ins["unidade"]) if sb_client else None
                if sb and sb.get("pu_mediana"):
                    ins["preco_un"] = float(sb["pu_mediana"])
                    ins["fonte"] = f"Supabase n={sb.get('n_proj','?')}"
                    fonte_ctrl["supabase"] += 1
                else:
                    item, score, chave = melhor_match_json(ins["descricao"], ins["unidade"], base_json)
                    if item and score >= 0.55 and item.get("mediana"):
                        ins["preco_un"] = float(item["mediana"])
                        ins["fonte"] = f"JSON fuzzy={score:.2f}"
                        ins["n_projetos"] = item.get("n_projetos")
                        ins["cv"] = item.get("cv")
                        fonte_ctrl["json"] += 1
                    else:
                        ins["preco_un"] = None
                        ins["fonte"] = "Pendente - cotar"
                        fonte_ctrl["pendente"] += 1

    # 5. totalizar
    total = 0.0
    for ins in insumos_flat:
        if ins.get("preco_un"):
            total += ins["qtd"] * ins["preco_un"]
    for items in inf.values():
        for ins in items:
            if ins.get("preco_un"):
                total += ins["qtd"] * ins["preco_un"]
    totais = {"total": total}

    # cobertura
    total_insumos = len(insumos_flat) + sum(len(v) for v in inf.values())
    com_preco = total_insumos - fonte_ctrl["pendente"]
    fonte_ctrl["cobertura_pct"] = (com_preco / total_insumos * 100) if total_insumos else 0

    print()
    print(f"[resumo] {len(insumos_flat)} insumos Flat + {sum(len(v) for v in inf.values())} inferidos")
    print(f"         Supabase={fonte_ctrl['supabase']}  JSON={fonte_ctrl['json']}  Pendente={fonte_ctrl['pendente']}")
    print(f"         Cobertura: {fonte_ctrl['cobertura_pct']:.1f}%")
    print(f"         Total CÉLULA 06: R$ {total:,.2f}  ({total/ELECTRA_AC:,.2f}/m² AC)")

    if args.dry_run:
        print("[dry-run] não gravou xlsx nem memorial")
        return

    # 6. escrever aba
    print(f"[6] Escrevendo aba `{NEW_SHEET_NAME}`...")
    counts = escrever_aba(wb, insumos_flat, inf)
    print(f"    CÉLULA={counts[0]}  ETAPAS={counts[1]}  SUBETAPAS={counts[2]}  SERVIÇOS={counts[3]}  INSUMOS={counts[4]}")

    # 7. salvar
    print(f"[7] Salvando xlsx em {out_path.name}...")
    wb.save(out_path)
    print(f"    OK: {out_path}")

    # 8. memorial
    print(f"[8] Gerando memorial...")
    memo = gerar_memorial(insumos_flat, inf, report_class, totais, fonte_ctrl)
    MEMORIAL_PATH.write_text(memo, encoding="utf-8")
    print(f"    OK: {MEMORIAL_PATH}")


if __name__ == "__main__":
    main()
