#!/usr/bin/env python3
"""Phase 20g - Excel da Analise Avancada: clientes, clusters, qualitativa.

Abas:
 1. LEIA_ME
 2. CLIENTES (top 30 clientes com stats)
 3. PROJETOS_POR_CLIENTE (drill-down)
 4. CLUSTERS (4 clusters descobertos)
 5. CLUSTER_MEMBROS (drill-down de cada cluster)
 6. COMPLEXIDADE (abas/itens/mgs por padrao)
 7. OBSERVACOES (categorias + motivos fora da curva)
 8. PADROES_IDENTIFICADOS (itens recorrentes)

Uso: python scripts/gerar_planilha_avancada.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
AGG_FILE = BASE / "analise-avancada-agregada.json"
OUT_XLSX = BASE / "analise-avancada-cartesian.xlsx"

DARK, ACCENT, ORANGE, GREEN, PURPLE, RED, GRAY, TEAL, YELLOW_D = (
    "2C3E50", "2980B9", "E67E22", "27AE60", "8E44AD", "C0392B", "7F8C8D", "16A085", "F39C12"
)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

TAB_COLORS = {
    "LEIA_ME": GREEN, "CLIENTES": ACCENT, "PROJETOS_POR_CLIENTE": ACCENT,
    "CLUSTERS": PURPLE, "CLUSTER_MEMBROS": PURPLE,
    "COMPLEXIDADE": TEAL, "OBSERVACOES": ORANGE,
    "PADROES_IDENTIFICADOS": YELLOW_D,
}


def _load(p):
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def sh(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def wh(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, 1):
        sh(ws.cell(row, i, h), color)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wr(ws, row, values, nf=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if nf and i in nf:
            c.number_format = nf[i]
        elif isinstance(v, float):
            c.number_format = "#,##0.00"
        elif isinstance(v, int):
            c.number_format = "#,##0"


def build_leia_me(wb, data):
    ws = wb.create_sheet("LEIA_ME")
    ws.sheet_properties.tabColor = TAB_COLORS["LEIA_ME"]
    ws["A1"] = "ANALISE AVANCADA - Clientes, Clusters, Qualitativa"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')} . {data['n_projetos']} projetos"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")

    row = 4
    ws.cell(row, 1, "PARA QUE SERVE").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    body = (
        "Analises avancadas cross-projeto: perfil de clientes (quem eles sao, quanto pagam, padrao tipico), "
        "agrupamento natural dos projetos por assinatura de distribuicao de custo (revela tipologias ocultas), "
        "e analise da camada qualitativa (observacoes, padroes identificados, itens fora da curva). "
        "Complementa analise-produto (indicadores fisicos) e analise-financeira (R$, % MG, R$/m2)."
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 1).font = Font(size=9, name="Arial")
    row += 3

    ws.cell(row, 1, "MAPA DAS ABAS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["#", "Aba", "Descricao"], [4, 28, 70])
    row += 1
    abas = [
        ("1", "LEIA_ME", "Este guia"),
        ("2", "CLIENTES", "Top 30 clientes: N projetos, R$/m² medio, padrao dominante"),
        ("3", "PROJETOS_POR_CLIENTE", "Drill-down: lista cada projeto de cada cliente"),
        ("4", "CLUSTERS", "4 tipologias descobertas por K-means nas assinaturas % MG"),
        ("5", "CLUSTER_MEMBROS", "Quais projetos caem em cada cluster"),
        ("6", "COMPLEXIDADE", "Metricas de escala/complexidade por padrao"),
        ("7", "OBSERVACOES", "Observacoes do orcamentista: categorias + motivos de fora da curva"),
        ("8", "PADROES_IDENTIFICADOS", "Itens recorrentes e itens fora da curva nos projetos"),
    ]
    for num, aba, desc in abas:
        wr(ws, row, [num, aba, desc])
        row += 1

    row += 2
    ws.cell(row, 1, "ACHADOS-CHAVE").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    findings = [
        "1. Nova Empreendimentos paga em medias ~R$ 6.100/m² (46% acima da mediana alto) - investigar margem/escopo.",
        "2. Paludo e Santa Maria sao clientes eficientes (R$ 2.781 e R$ 1.604 /m²) - benchmark interno.",
        "3. Cluster_3 (13 projetos) tem GERENCIAMENTO 40% do total - projetos alto-consultoria.",
        "4. Qualitativa revela 544 observacoes: 181 'outro', 169 'premissa', 140 'justificativa' - fonte rica de IA.",
        "5. 'Mao de obra supraestrutura' e 'Mao de obra infraestrutura' aparecem em fora-da-curva - precificacao inconsistente.",
        "6. Padrao alto tem med 1.736 itens, economico so 38 - escala de detalhamento totalmente diferente.",
    ]
    for f in findings:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1, f).font = Font(size=9, name="Arial")
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40


def build_clientes(wb, data):
    ws = wb.create_sheet("CLIENTES")
    ws.sheet_properties.tabColor = TAB_COLORS["CLIENTES"]
    ws["A1"] = "TOP CLIENTES - Perfil agregado"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:I1")
    ws["A2"] = "Ordenado por numero de projetos. R$/m² mediana apenas dos projetos com AC >= 1000 m²."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:I2")

    wh(ws, 4, ["Cliente", "N projetos", "N validos", "R$/m² min", "R$/m² mediana", "R$/m² max",
               "AC total (m²)", "AC medio", "Padrao dominante"],
       [28, 10, 10, 12, 14, 12, 16, 14, 16])
    row = 5
    nf = {4: "R$ #,##0", 5: "R$ #,##0", 6: "R$ #,##0", 7: "#,##0", 8: "#,##0"}
    for c in data.get("analise_por_cliente", [])[:30]:
        wr(ws, row, [
            c["cliente"], c["n_projetos"], c["n_validos_rsm2"],
            c.get("rsm2_min"), c.get("rsm2_mediana"), c.get("rsm2_max"),
            c["ac_total"], c["ac_medio"], c["padrao_dominante"],
        ], nf)
        # Color: Nova Empreendimentos (caro) em vermelho, Paludo/Santa Maria (eficiente) em verde
        if c["cliente"] in ("Nova Empreendimentos", "ALL"):
            ws.cell(row, 5).fill = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
        elif c["cliente"] in ("Paludo", "Santa Maria"):
            ws.cell(row, 5).fill = PatternFill(start_color="E8F8E0", end_color="E8F8E0", fill_type="solid")
        row += 1
    ws.auto_filter.ref = f"A4:I{row-1}"
    ws.freeze_panes = "A5"


def build_projetos_por_cliente(wb, data):
    ws = wb.create_sheet("PROJETOS_POR_CLIENTE")
    ws.sheet_properties.tabColor = TAB_COLORS["PROJETOS_POR_CLIENTE"]
    ws["A1"] = "DRILL-DOWN: Projetos por cliente (top 30)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")

    wh(ws, 3, ["Cliente", "Projeto", "Padrao", "AC (m²)", "R$/m²"], [22, 38, 14, 14, 14])
    row = 4
    nf = {4: "#,##0", 5: "R$ #,##0"}
    for c in data.get("analise_por_cliente", [])[:30]:
        for p in c["projetos"]:
            wr(ws, row, [c["cliente"], p["slug"], p["padrao"], p["ac"] or None, p["rsm2"] or None], nf)
            row += 1
    ws.auto_filter.ref = f"A3:E{row-1}"
    ws.freeze_panes = "A4"


def build_clusters(wb, data):
    ws = wb.create_sheet("CLUSTERS")
    ws.sheet_properties.tabColor = TAB_COLORS["CLUSTERS"]
    ws["A1"] = "CLUSTERS - Tipologias descobertas por K-means em % MG"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Projetos com total valido foram agrupados em 4 clusters baseado na assinatura % de cada MG. "
                "Cada cluster representa um 'tipo' de projeto com distribuicao similar de custos.")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    row = 4
    for name, info in data.get("clusters", {}).items():
        if name == "error":
            continue
        ws.cell(row, 1, f"{name.upper()} — {info['n_projetos']} projetos — padrao dom: {info['padrao_dominante']}").font = Font(bold=True, size=11, color=DARK, name="Arial")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
        ws.cell(row, 1, f"AC mediana: {info['ac_mediano']:,.0f} m²   |   R$/m² mediana: R$ {info['rsm2_mediano']:,.2f}   |   Mix padroes: {info['padroes_mix']}").font = Font(italic=True, size=9, color="666666", name="Arial")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2
        wh(ws, row, ["Top 5 MGs (centroide)", "% do total (mediana)"], [40, 20])
        row += 1
        for mg in info.get("top_mgs_centroide", []):
            wr(ws, row, [mg["mg"], mg["pct"]], {2: "0.00"})
            row += 1
        row += 2


def build_cluster_membros(wb, data):
    ws = wb.create_sheet("CLUSTER_MEMBROS")
    ws.sheet_properties.tabColor = TAB_COLORS["CLUSTER_MEMBROS"]
    ws["A1"] = "CLUSTERS - Membros (quais projetos em cada cluster)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    wh(ws, 3, ["Cluster", "Projeto", "Cliente", "Padrao original", "AC (m²)", "R$/m²"],
       [14, 38, 22, 16, 14, 14])
    row = 4
    nf = {5: "#,##0", 6: "R$ #,##0"}
    for name, info in data.get("clusters", {}).items():
        if name == "error":
            continue
        for m in info.get("membros", []):
            wr(ws, row, [name, m["slug"], m["cliente"], m["padrao_original"],
                         m["ac"] or None, m["rsm2"] or None], nf)
            row += 1
    ws.auto_filter.ref = f"A3:F{row-1}"
    ws.freeze_panes = "A4"


def build_complexidade(wb, data):
    ws = wb.create_sheet("COMPLEXIDADE")
    ws.sheet_properties.tabColor = TAB_COLORS["COMPLEXIDADE"]
    ws["A1"] = "COMPLEXIDADE POR PADRAO - escala do orcamento"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    ws["A2"] = "Mediana de n. abas, n. itens e n. macrogrupos. Reflete o nivel de detalhamento tipico."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")

    wh(ws, 4, ["Padrao", "N abas (med)", "N itens (med)", "N macrogrupos (med)", "N projetos amostra"],
       [16, 14, 16, 22, 20])
    row = 5
    nf = {2: "0.0", 3: "#,##0", 4: "0.0", 5: "#,##0"}
    for pad in ["alto", "medio-alto", "medio", "economico", "insuficiente"]:
        c = data.get("complexidade_por_padrao", {}).get(pad)
        if not c:
            continue
        wr(ws, row, [pad, c.get("n_abas_med"), c.get("n_itens_med"),
                     c.get("n_mgs_med"), c["n_projetos_amostra"]], nf)
        row += 1


def build_observacoes(wb, data):
    ws = wb.create_sheet("OBSERVACOES")
    ws.sheet_properties.tabColor = TAB_COLORS["OBSERVACOES"]
    ws["A1"] = "OBSERVACOES DO ORCAMENTISTA - Categorizadas"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")

    q = data.get("qualitativa", {})
    row = 3
    ws.cell(row, 1, "CATEGORIAS DE OBSERVACOES").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    ws.cell(row, 1, f"Total: {q.get('total_observacoes', 0)} observacoes em 126 projetos").font = Font(italic=True, size=9, color="666666", name="Arial")
    row += 2
    wh(ws, row, ["Categoria", "N observacoes"], [22, 16])
    row += 1
    for cat, n in q.get("categorias_observacoes", {}).items():
        wr(ws, row, [cat, n])
        row += 1

    row += 3
    ws.cell(row, 1, "MOTIVOS DE FORA DA CURVA (amostra de 30)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    ws.cell(row, 1, f"Total: {q.get('total_fora_curva', 0)} instancias de itens fora da curva").font = Font(italic=True, size=9, color="666666", name="Arial")
    row += 2
    wh(ws, row, ["Projeto", "Padrao", "Item", "Motivo"], [28, 14, 38, 60])
    row += 1
    for m in q.get("motivos_fora_curva_amostra", [])[:30]:
        wr(ws, row, [m["slug"], m["padrao"], m["item"], m["motivo"]])
        row += 1
    ws.auto_filter.ref = f"A{row-30 if row>30 else 4}:D{row-1}"


def build_padroes(wb, data):
    ws = wb.create_sheet("PADROES_IDENTIFICADOS")
    ws.sheet_properties.tabColor = TAB_COLORS["PADROES_IDENTIFICADOS"]
    ws["A1"] = "PADROES IDENTIFICADOS + ITENS FORA DA CURVA - Recorrencia"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:C1")

    q = data.get("qualitativa", {})
    row = 3
    ws.cell(row, 1, "PADROES IDENTIFICADOS (itens destacados pelo orcamentista)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Item", "N projetos"], [60, 16])
    row += 1
    for item, n in q.get("padroes_identificados_top", {}).items():
        wr(ws, row, [item, n])
        row += 1

    row += 3
    ws.cell(row, 1, "ITENS FORA DA CURVA (aparecem discrepantes)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Item", "N projetos"], [60, 16])
    row += 1
    for item, n in q.get("itens_fora_curva_recorrentes", {}).items():
        wr(ws, row, [item, n])
        row += 1


def main():
    print("Loading...", flush=True)
    data = _load(AGG_FILE)
    if not data:
        print("ERROR: not found")
        return
    print(f"  {data['n_projetos']} projects", flush=True)

    wb = Workbook()
    wb.remove(wb.active)
    print("Building tabs...", flush=True)
    build_leia_me(wb, data); print("  LEIA_ME")
    build_clientes(wb, data); print("  CLIENTES")
    build_projetos_por_cliente(wb, data); print("  PROJETOS_POR_CLIENTE")
    build_clusters(wb, data); print("  CLUSTERS")
    build_cluster_membros(wb, data); print("  CLUSTER_MEMBROS")
    build_complexidade(wb, data); print("  COMPLEXIDADE")
    build_observacoes(wb, data); print("  OBSERVACOES")
    build_padroes(wb, data); print("  PADROES_IDENTIFICADOS")
    wb.save(str(OUT_XLSX))
    print(f"\nSalvo: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
