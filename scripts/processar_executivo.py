#!/usr/bin/env python3
"""
Processador de orçamentos executivos Cartesian — Camada 2 (PUs detalhados).

Processa 1 xlsx por vez e gera:
  - pus-raw/{slug}-raw.json (dados brutos extraídos)
  - indices-executivo/{slug}.json (índices detalhados + derivados)
  - Atualiza projetos-metadados.json (metadados editáveis)

Reutiliza lógica do extract_indices.py (format detection, AC extraction, discipline mapping).

Uso:
  python processar_executivo.py --inventory              # Listar e classificar arquivos
  python processar_executivo.py --process <xlsx_path>    # Processar 1 arquivo
  python processar_executivo.py --batch                  # Processar todos pendentes
  python processar_executivo.py --batch --limit 5        # Processar N pendentes
"""

import os
import sys
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime

import openpyxl

# === PATHS ===
EXEC_DIR = os.path.expanduser('~/orcamentos/executivos/entregues')
BASE_DIR = os.path.expanduser('~/orcamentos/base')
RAW_DIR = os.path.join(BASE_DIR, 'pus-raw')
INDICES_DIR = os.path.join(BASE_DIR, 'indices-executivo')
METADADOS_PATH = os.path.join(BASE_DIR, 'projetos-metadados.json')
CALIBRATION_PATH = os.path.join(BASE_DIR, 'calibration-data.json')
MANIFEST_PATH = os.path.expanduser(
    '~/orcamentos-openclaw/archive/v1-pre-fase19/raiz-mar-2026/executivos-manifest.json'
)

# === DISCIPLINE MAPPING (from extract_indices.py) ===
DISCIPLINE_SHEETS = {
    'ELÉTRICA': 'Instalacoes Eletricas',
    'ELETRICA': 'Instalacoes Eletricas',
    'HIDRÁULICAS': 'Instalacoes Hidraulicas',
    'HIDRAULICAS': 'Instalacoes Hidraulicas',
    'HIDROSSANIT': 'Instalacoes Hidraulicas',
    'ESGOTO E PLUVIAL': 'Esgoto e Pluvial',
    'LOUÇAS E METAIS': 'Loucas e Metais',
    'LOUÇAS': 'Loucas e Metais',
    'EXAUSTÃO': 'Exaustao',
    'TELECOM': 'Telecom',
    'SPDA': 'SPDA',
    'SUPRAESTRUTURA': 'Supraestrutura',
    'ALVENARIA': 'Alvenaria',
    'REVESTIMENTO': 'Revestimentos',
    'PISOS': 'Pisos',
    'PAVIMENTAÇ': 'Pisos',
    'FORRO': 'Forro',
    'TETO': 'Forro',
    'FACHADA': 'Fachada',
    'IMPER': 'Impermeabilizacao',
    'ESCAVAÇÕES': 'Escavacoes',
    'ESTACAS': 'Estacas',
    'BALDRAMES': 'Baldrames',
    'ESCORAMENTO': 'Escoramento',
    'CHURRASQUEIRA': 'Churrasqueira',
    'MOVIMENTAÇÃO': 'Movimentacao de Terra',
    'INFRAESTRUTURA': 'Infraestrutura',
    'COMPLEMENTAR': 'Servicos Complementares',
    'SISTEMAS ESPECIAIS': 'Sistemas Especiais',
    'ESQUADRIA': 'Esquadrias',
    'PINTURA': 'Pintura',
    'COBERTURA': 'Cobertura',
    'CANTEIRO': 'Canteiro',
    'CLIMATIZAÇÃO': 'Climatizacao',
    'CLIMATIZACAO': 'Climatizacao',
    'GÁS': 'Gas',
    'GAS': 'Gas',
    'PREVENTIV': 'PCI',
    'INCÊNDIO': 'PCI',
    'INCENDIO': 'PCI',
    'CONTENÇ': 'Contencao',
    'CONTENCAO': 'Contencao',
    'FUNDAÇ': 'Fundacoes',
    'FUNDACOES': 'Fundacoes',
    'GERENCIAMENTO': 'Gerenciamento',
}

# Macrogroup mapping: discipline label -> standard macrogroup
DISCIPLINE_TO_MACROGROUP = {
    'Instalacoes Eletricas': 'Instalacoes',
    'Instalacoes Hidraulicas': 'Instalacoes',
    'Esgoto e Pluvial': 'Instalacoes',
    'Loucas e Metais': 'Loucas e Metais',
    'Exaustao': 'Sistemas Especiais',
    'Telecom': 'Sistemas Especiais',
    'SPDA': 'Instalacoes',
    'Supraestrutura': 'Supraestrutura',
    'Alvenaria': 'Alvenaria',
    'Revestimentos': 'Revestimentos',
    'Pisos': 'Pisos',
    'Forro': 'Forro',
    'Fachada': 'Fachada',
    'Impermeabilizacao': 'Impermeabilizacao',
    'Escavacoes': 'Mov. Terra',
    'Estacas': 'Infraestrutura',
    'Baldrames': 'Infraestrutura',
    'Escoramento': 'Infraestrutura',
    'Churrasqueira': 'Complementares',
    'Movimentacao de Terra': 'Mov. Terra',
    'Infraestrutura': 'Infraestrutura',
    'Servicos Complementares': 'Complementares',
    'Sistemas Especiais': 'Sistemas Especiais',
    'Esquadrias': 'Esquadrias',
    'Pintura': 'Pintura',
    'Cobertura': 'Cobertura',
    'Canteiro': 'Gerenciamento',
    'Climatizacao': 'Climatizacao',
    'Gas': 'Instalacoes',
    'PCI': 'Instalacoes',
    'Contencao': 'Infraestrutura',
    'Fundacoes': 'Infraestrutura',
    'Gerenciamento': 'Gerenciamento',
}

# Keywords for MO (mão de obra) detection
MO_KEYWORDS = [
    'mão de obra', 'mao de obra', 'moe', 'empreiteira', 'serviço de mão',
    'servico de mao', 'mão-de-obra', 'mao-de-obra', 'instalação de',
    'instalacao de', 'execução de', 'execucao de', 'assentamento de',
    'aplicação de', 'aplicacao de', 'montagem de',
]

# Macrogroup keywords for resumo extraction
MACRO_KEYWORDS = [
    'GERENCIAMENTO', 'MOVIMENTAÇÃO', 'INFRAESTRUTURA', 'SUPRAESTRUTURA',
    'ALVENARIA', 'PAREDES', 'INSTALAÇÕES', 'EQUIPAMENTOS', 'SISTEMAS',
    'IMPERMEABILIZAÇÃO', 'REVESTIMENTO', 'ACABAMENTO', 'PINTURA',
    'ESQUADRIA', 'VIDRO', 'FACHADA', 'COBERTURA', 'COMPLEMENTAR',
    'IMPREVISTO', 'FUNDAÇÃO', 'CONTENÇ', 'LOUÇAS', 'METAIS',
    'PISO', 'TETO', 'FORRO', 'SERVIÇOS', 'CLIMATIZAÇÃO',
]


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace(',', '.').replace(' ', '').replace('\xa0', '')
        s = re.sub(r'[R$%]', '', s).strip()
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


def normalizar_item(descricao):
    """Normalizar descrição de item para chave canônica."""
    if not descricao:
        return ''
    s = str(descricao).strip()

    # Remove accents
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))

    s = s.lower()

    # Standardize measurements
    s = re.sub(r'(\d+)/(\d+)"', r'\1_\2_pol', s)  # 3/4" -> 3_4_pol
    s = re.sub(r'[øØ∅](\d+)', r'\1', s)  # ø20mm -> 20mm
    s = re.sub(r'dn\s*(\d+)', r'dn\1', s)  # DN 50 -> dn50
    s = re.sub(r'm²', 'm2', s)
    s = re.sub(r'm³', 'm3', s)

    # Remove articles and short prepositions
    s = re.sub(r'\b(de|do|da|dos|das|em|no|na|nos|nas|ao|a|o|e|p/|c/|s/)\b', ' ', s)

    # Collapse whitespace and convert to snake_case
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', '_', s.strip())
    s = re.sub(r'_+', '_', s).strip('_')

    return s[:80]


def normalizar_unidade(unit):
    """Padronizar unidades."""
    if not unit:
        return ''
    u = str(unit).strip().lower()
    mapping = {
        'pç': 'un', 'pc': 'un', 'peça': 'un', 'peca': 'un', 'und': 'un',
        'unid': 'un', 'unid.': 'un', 'un.': 'un',
        'ml': 'm', 'm.l.': 'm', 'ml.': 'm',
        'm²': 'm2', 'mq': 'm2', 'm2.': 'm2',
        'm³': 'm3', 'mc': 'm3', 'm3.': 'm3',
        'kg.': 'kg', 'kgf': 'kg',
        'vb': 'vb', 'verba': 'vb', 'vb.': 'vb', 'gb': 'vb', 'cj': 'vb',
        'conj': 'vb', 'conj.': 'vb',
    }
    return mapping.get(u, u)


def slug_from_filename(filename):
    """Extrair slug do projeto a partir do nome ou path do arquivo."""
    # If path contains client/project structure, use it
    parts = filename.replace(EXEC_DIR, '').strip('/').split('/')
    if len(parts) >= 3:
        # client/project/file.xlsx -> client-project
        name = f"{parts[-3]}-{parts[-2]}"
    elif len(parts) >= 2 and not parts[-1].endswith('/'):
        # client/project/file.xlsx where file is parts[-1]
        name = '-'.join(parts[:-1])
    else:
        name = os.path.basename(filename)
        # Remove prefixo drive- ou entrega-
        name = re.sub(r'^(drive|entrega)-', '', name)
        # Pegar parte antes do "--"
        if '--' in name:
            name = name.split('--')[0]

    # Normalize accents
    name = unicodedata.normalize('NFKD', name)
    name = ''.join(c for c in name if not unicodedata.combining(c))
    # Clean
    name = re.sub(r'[^a-zA-Z0-9\-]', '-', name)
    name = re.sub(r'-+', '-', name).strip('-').lower()
    return name


def is_executivo(filename):
    """Classificar se o arquivo é um orçamento (executivo ou apresentação)."""
    upper = filename.upper()
    # Exclude only pure slides
    if 'PLANILHAS SLIDE' in upper:
        return False
    # Include everything that looks like a budget
    if any(p in upper for p in ['ORÇAMENTO', 'ORCAMENTO', 'EXECUTIVO', 'GERENCIAMENTO',
                                  'ENTREGÁVEL', 'ENTREGAVEL', 'COMPLETO', 'EAP',
                                  'APRESENTAÇÃO', 'APRESENTACAO', 'APRESENTAÇÃO']):
        return True
    return False


def load_metadados():
    if os.path.exists(METADADOS_PATH):
        with open(METADADOS_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_metadados(metadados):
    with open(METADADOS_PATH, 'w', encoding='utf-8') as f:
        json.dump(metadados, f, ensure_ascii=False, indent=2)


def load_manifest():
    if os.path.exists(MANIFEST_PATH):
        with open(MANIFEST_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'projects': []}


def load_calibration_ac():
    ac_map = {}
    if os.path.exists(CALIBRATION_PATH):
        with open(CALIBRATION_PATH, 'r', encoding='utf-8') as f:
            cal = json.load(f)
        for p in cal:
            name = p['name'].lower().replace(' ', '-')
            if p.get('ac'):
                ac_map[name] = p['ac']
    return ac_map


# ============================================================
# EXTRACTION FUNCTIONS
# ============================================================

def extrair_metadados(wb, filename):
    """Extract project metadata from workbook."""
    meta = {
        'arquivo_fonte': os.path.basename(filename),
        'processado_em': datetime.now().strftime('%Y-%m-%d'),
        'status': 'processado',
    }

    # Search for AC
    priority_sheets = ['DADOS_INICIAIS', 'Dados Gerais', 'OBRA', 'Obra', 'Obra ',
                       'CAPA', 'Capa', 'DADOS', 'Dados do Empreendimento']
    orc_sheets = [s for s in wb.sheetnames if 'ORÇAMENTO' in s.upper() or 'EXECUTIVO' in s.upper()]
    search_order = [s for s in priority_sheets if s in wb.sheetnames] + orc_sheets

    for sheet_name in search_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=40, max_col=10, values_only=True):
            if not row:
                continue
            for i, cell in enumerate(row):
                if not cell or not isinstance(cell, str):
                    continue
                upper = cell.upper()

                # AC
                if ('ÁREA' in upper or 'AREA' in upper) and ('CONSTRUÍDA' in upper or 'CONSTRUIDA' in upper):
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 100 < val < 200000:
                                meta['ac'] = val
                                break

                # UR
                if 'UNIDADES' in upper and ('HABITACION' in upper or 'AUTÔNOM' in upper or 'AUTONOM' in upper):
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 0 < val < 5000:
                                meta['ur'] = int(val)
                                break

                # Pavimentos
                if 'PAVIMENTO' in upper and 'TIPO' in upper:
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 0 < val < 100:
                                meta['pavimentos_tipo'] = int(val)
                                break

                if 'PAVIMENTO' in upper and 'TOTAL' in upper and 'TIPO' not in upper:
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 0 < val < 200:
                                meta['pavimentos'] = int(val)
                                break

                # Subsolos
                if 'SUBSOLO' in upper:
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val is not None and 0 <= val < 20:
                                meta['subsolos'] = int(val)
                                break

                # Vagas
                if 'VAGA' in upper and 'GARAGEM' in upper:
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 0 < val < 5000:
                                meta['vagas'] = int(val)
                                break

                # CUB
                if 'CUB' in upper and ('R$' in upper or 'VALOR' in upper or 'BASE' in upper):
                    for j in range(len(row)):
                        if j != i:
                            val = safe_float(row[j])
                            if val and 1000 < val < 10000:
                                meta['cub_base'] = val
                                break

    return meta



# Sheets that are NOT disciplines (main budget/summary sheets)
SKIP_SHEETS = {
    'GERENCIAMENTO EXECUTIVO', 'GERENCIAMENTO_EXECUTIVO', 'ORÇAMENTO EXECUTIVO',
    'ORCAMENTO EXECUTIVO', 'ORÇAMENTO', 'ORCAMENTO', 'ORÇAMENTO RESUMO',
    'EAP', 'EAP ANÁLISE', 'EAP ANALISE', 'RESUMO', 'RESUMOS',
    'DADOS_INICIAIS', 'DADOS GERAIS', 'OBRA', 'CAPA', 'COMPARATIVO',
    'EPCS', 'ENSAIOS', 'RELATÓRIO', 'RELATORIO', 'PROJETOS',
    'CRONOGRAMA', 'CURVA', 'BDI', 'ENCARGOS', 'ARQ', 'ARQUITETURA',
    'PISCINA', 'MEMORIAL', 'PROJETOS E GERENCIAMENTO',
}


def mapear_disciplina(sheet_name):
    """Map sheet name to canonical discipline label."""
    upper = sheet_name.upper().strip()

    # Skip non-discipline sheets
    if upper in SKIP_SHEETS:
        return None
    if any(upper.startswith(s) for s in ['PROJETOS E GERENCIAMENTO', 'DADOS', 'COMPARATIV']):
        return None

    for key, label in DISCIPLINE_SHEETS.items():
        if upper == key or key in upper:
            return label
    return None


def detectar_colunas(ws, max_scan_rows=20):
    """Detect column layout by scanning header rows.
    Returns dict: {'desc': idx, 'qty': idx, 'unit': idx, 'pu': idx, 'total': idx}
    """
    header_keywords = {
        'desc': ['DESCRIÇÃO', 'DESCRICAO', 'SERVIÇO', 'SERVICO', 'ITEM', 'ESPECIFICAÇÃO'],
        'qty': ['QTD', 'QUANTIDADE', 'QUANT'],
        'unit': ['UND', 'UNID', 'UN', 'UNIDADE'],
        'pu': ['P.U.', 'PU', 'PREÇO UNIT', 'PRECO UNIT', 'CUSTO UNIT', 'VALOR UNIT',
               'R$/UNID', 'UNITÁRIO', 'UNITARIO'],
        'total': ['TOTAL', 'VALOR', 'CUSTO TOTAL', 'SUBTOTAL'],
    }

    for row in ws.iter_rows(max_row=max_scan_rows, values_only=True):
        if not row:
            continue
        mapping = {}
        for col_idx, cell_val in enumerate(row):
            text = str(cell_val).upper().strip() if cell_val else ''
            if not text:
                continue
            for field, keywords in header_keywords.items():
                if field not in mapping:
                    for kw in keywords:
                        if kw in text:
                            mapping[field] = col_idx
                            break

        if 'desc' in mapping and 'total' in mapping and len(mapping) >= 3:
            return mapping

    # Fallback: assume standard layout (desc=1, qty=3, unit=4, pu=5, total=6)
    return {'desc': 1, 'qty': 3, 'unit': 4, 'pu': 5, 'total': 6}


def extrair_itens_disciplina(ws, ac=None):
    """Extract items from a discipline worksheet.
    Returns dict with 'itens' list, 'subcategorias' dict, 'total'.
    """
    cols = detectar_colunas(ws)

    result = {
        'itens': [],
        'subcategorias': {},
        'total': 0,
        'warnings': [],
    }

    current_subcat = 'GERAL'
    current_pavimento = 'GERAL'
    items_seen = 0

    # Detect pavimento patterns
    pav_patterns = [
        r'(T[ÉE]RREO)', r'(TIPO)', r'(SUBSOLO)', r'(COBERTURA)',
        r'(\d+[°ºª]\s*(ao|a)\s*\d+[°ºª])', r'(LAZER)', r'(GARAGEM)',
        r'(CASA DE M[ÁA]QUINAS)', r'(BARRILETE)', r'(MEZANINO)',
        r'(PILOTIS)', r'(ÁTICO)', r'(ATICO)',
    ]

    for row in ws.iter_rows(min_row=2, max_row=2000, values_only=True):
        if not row:
            continue

        # Get cell values
        desc_idx = cols.get('desc', 1)
        qty_idx = cols.get('qty', 3)
        unit_idx = cols.get('unit', 4)
        pu_idx = cols.get('pu', 5)
        total_idx = cols.get('total', 6)

        desc = str(row[desc_idx]).strip() if desc_idx < len(row) and row[desc_idx] else ''
        if not desc or desc == 'None':
            # Try column 0
            desc = str(row[0]).strip() if row[0] else ''
            if not desc or desc == 'None':
                continue

        qty = safe_float(row[qty_idx]) if qty_idx < len(row) else None
        unit = str(row[unit_idx]).strip() if unit_idx < len(row) and row[unit_idx] else ''
        pu = safe_float(row[pu_idx]) if pu_idx < len(row) else None
        total = safe_float(row[total_idx]) if total_idx < len(row) else None

        upper_desc = desc.upper()

        # Skip header/footer rows
        if any(skip in upper_desc for skip in ['DESCRIÇÃO', 'TOTAL GERAL', 'SUBTOTAL GERAL',
                                                  'VALOR TOTAL', 'SOMA']):
            continue

        # Detect pavimento
        for pat in pav_patterns:
            m = re.search(pat, upper_desc, re.IGNORECASE)
            if m:
                current_pavimento = desc.strip()
                break

        # Subcategory header: has total but no qty/unit
        if total and total > 500 and not qty and (not unit or unit == 'None'):
            current_subcat = desc
            if current_subcat not in result['subcategorias']:
                result['subcategorias'][current_subcat] = {
                    'total': total,
                    'itens': [],
                }
            continue

        # Individual item: has qty AND unit AND total
        if qty and qty > 0 and unit and unit != 'None' and total and total > 0:
            item = {
                'descricao': desc,
                'quantidade': qty,
                'unidade': normalizar_unidade(unit),
                'pu': pu if pu else (total / qty if qty > 0 else None),
                'total': total,
                'subgrupo': current_subcat,
                'pavimento': current_pavimento,
                'chave_normalizada': normalizar_item(desc),
            }
            result['itens'].append(item)

            if current_subcat in result['subcategorias']:
                result['subcategorias'][current_subcat]['itens'].append(item)

            items_seen += 1

    result['total'] = sum(i['total'] for i in result['itens'] if i.get('total'))
    result['n_itens'] = items_seen

    return result


def extrair_analitico_eap(ws, ac=None):
    """Extract items from analítico/EAP hierarchical format.
    Codes: X = macrogroup, X.X = section, X.X.X = item (with qty, unit, pu, total).
    Columns detected dynamically via header row.
    """
    ANALITICO_DISC_MAP = {
        'GERENCIAMENTO': 'Gerenciamento',
        'MOVIMENTAÇÃO DE TERRA': 'Movimentacao de Terra',
        'INFRAESTRUTURA': 'Infraestrutura',
        'CONTENÇ': 'Contencao',
        'FUNDAÇ': 'Fundacoes',
        'SUPRAESTRUTURA': 'Supraestrutura',
        'ALVENARIA': 'Alvenaria',
        'VEDAÇ': 'Alvenaria',
        'IMPERMEABILIZAÇÃO': 'Impermeabilizacao',
        'INSTALAÇÕES ELÉTRICAS': 'Instalacoes Eletricas',
        'INSTALAÇÕES HIDRO': 'Instalacoes Hidraulicas',
        'HIDROSSANIT': 'Instalacoes Hidraulicas',
        'PREVENTIV': 'PCI',
        'INCÊNDIO': 'PCI',
        'CLIMATIZAÇÃO': 'Climatizacao',
        'REVESTIMENTO': 'Revestimentos',
        'ACABAMENTO': 'Revestimentos',
        'PINTURA': 'Pintura',
        'ESQUADRIA': 'Esquadrias',
        'FACHADA': 'Fachada',
        'COBERTURA': 'Cobertura',
        'LOUÇAS': 'Loucas e Metais',
        'COMPLEMENTAR': 'Servicos Complementares',
        'IMPREVISTO': 'Servicos Complementares',
        'PISO': 'Pisos',
        'TETO': 'Forro',
        'FORRO': 'Forro',
        'SISTEMA': 'Sistemas Especiais',
        'GÁS': 'Gas',
    }

    # Detect columns
    cols = detectar_colunas(ws, max_scan_rows=15)

    disciplinas = {}
    current_disc_label = None
    current_subcat = 'GERAL'

    for row in ws.iter_rows(min_row=2, max_row=3000, values_only=True):
        if not row or not row[0]:
            continue

        code = str(row[0]).strip()
        if not re.match(r'^[\d.]+\s*$', code):
            continue

        desc_idx = cols.get('desc', 1)
        desc = str(row[desc_idx]).strip() if desc_idx < len(row) and row[desc_idx] else ''
        if not desc or desc == 'None':
            continue

        qty_idx = cols.get('qty', 3)
        unit_idx = cols.get('unit', 2)
        pu_idx = cols.get('pu', 4)
        total_idx = cols.get('total', 5)

        qty = safe_float(row[qty_idx]) if qty_idx < len(row) else None
        unit = str(row[unit_idx]).strip() if unit_idx < len(row) and row[unit_idx] else ''
        pu = safe_float(row[pu_idx]) if pu_idx < len(row) else None
        total = safe_float(row[total_idx]) if total_idx < len(row) else None

        parts = code.strip().split('.')
        depth = len(parts)

        # Level 1: macrogroup (single digit)
        if depth == 1:
            upper_desc = desc.upper()
            current_disc_label = None
            for key, label in ANALITICO_DISC_MAP.items():
                if key in upper_desc:
                    current_disc_label = label
                    break
            if current_disc_label and current_disc_label not in disciplinas:
                disciplinas[current_disc_label] = {
                    'itens': [], 'subcategorias': {}, 'total': 0, 'n_itens': 0, 'warnings': [],
                }
            current_subcat = desc

        # Level 2: section
        elif depth == 2:
            current_subcat = desc
            if current_disc_label and current_disc_label in disciplinas:
                if current_subcat not in disciplinas[current_disc_label]['subcategorias']:
                    disciplinas[current_disc_label]['subcategorias'][current_subcat] = {
                        'total': total or 0, 'itens': [],
                    }

        # Level 3+: item (with qty, unit, total)
        elif depth >= 3 and qty and qty > 0 and unit and unit != 'None' and total and total > 0:
            if current_disc_label and current_disc_label in disciplinas:
                item = {
                    'descricao': desc,
                    'quantidade': qty,
                    'unidade': normalizar_unidade(unit),
                    'pu': pu if pu else (total / qty if qty > 0 else None),
                    'total': total,
                    'subgrupo': current_subcat,
                    'pavimento': 'GERAL',
                    'chave_normalizada': normalizar_item(desc),
                }
                disciplinas[current_disc_label]['itens'].append(item)
                disciplinas[current_disc_label]['n_itens'] += 1
                if current_subcat in disciplinas[current_disc_label]['subcategorias']:
                    disciplinas[current_disc_label]['subcategorias'][current_subcat]['itens'].append(item)

    # Calculate totals
    for disc_data in disciplinas.values():
        disc_data['total'] = sum(i['total'] for i in disc_data['itens'] if i.get('total'))
        disc_data['rsm2'] = disc_data['total'] / ac if ac and disc_data['total'] else None

    return disciplinas


def extrair_abc_insumos(ws, ac=None):
    """Extract items from ABC Insumos/Serviços format.
    Columns: Código | Descrição | Un. | Quantidade | Preço unitário | Preço total
    """
    cols = detectar_colunas(ws, max_scan_rows=15)

    itens = []
    for row in ws.iter_rows(min_row=2, max_row=5000, values_only=True):
        if not row:
            continue

        desc_idx = cols.get('desc', 1)
        desc = str(row[desc_idx]).strip() if desc_idx < len(row) and row[desc_idx] else ''
        if not desc or desc == 'None' or len(desc) < 3:
            continue

        qty_idx = cols.get('qty', 3)
        unit_idx = cols.get('unit', 2)
        pu_idx = cols.get('pu', 4)
        total_idx = cols.get('total', 5)

        qty = safe_float(row[qty_idx]) if qty_idx < len(row) else None
        unit = str(row[unit_idx]).strip() if unit_idx < len(row) and row[unit_idx] else ''
        pu = safe_float(row[pu_idx]) if pu_idx < len(row) else None
        total = safe_float(row[total_idx]) if total_idx < len(row) else None

        if qty and qty > 0 and unit and unit != 'None' and total and total > 0:
            itens.append({
                'descricao': desc,
                'quantidade': qty,
                'unidade': normalizar_unidade(unit),
                'pu': pu if pu else (total / qty if qty > 0 else None),
                'total': total,
                'subgrupo': 'ABC',
                'pavimento': 'GERAL',
                'chave_normalizada': normalizar_item(desc),
            })

    if not itens:
        return {}

    # All ABC items go under a single "Insumos" discipline
    total_val = sum(i['total'] for i in itens if i.get('total'))
    return {
        'Insumos ABC': {
            'itens': itens,
            'subcategorias': {},
            'total': total_val,
            'n_itens': len(itens),
            'rsm2': total_val / ac if ac else None,
            'warnings': [],
        }
    }


def extrair_sienge(ws, ac=None, ur=None):
    """Extract items from Sienge 'Relatório' format.
    Structure: code col0, desc col1, unit col13, qty col16, pu col22, total col28.
    Hierarchy: XX = macrogroup, XX.XXX = section, XX.XXX.XXX = subsection, XX.XXX.XXX.XXX = item.
    """
    SIENGE_DISC_MAP = {
        'INFRAESTRUTURA': 'Infraestrutura',
        'SUPRAESTRUTURA': 'Supraestrutura',
        'VEDAÇÕES': 'Alvenaria',
        'VEDAÇ': 'Alvenaria',
        'INSTALAÇÕES ELÉTRICAS': 'Instalacoes Eletricas',
        'INSTALAÇÕES HIDRO': 'Instalacoes Hidraulicas',
        'HIDROSSANIT': 'Instalacoes Hidraulicas',
        'PREVENTIV': 'PCI',
        'GLP': 'PCI',
        'CLIMATIZAÇÃO': 'Climatizacao',
        'EXAUSTÃO': 'Climatizacao',
        'PRESSURIZAÇÃO': 'Climatizacao',
        'AUTOMAÇÃO': 'Sistemas Especiais',
        'TELEFONIA': 'Sistemas Especiais',
        'IMPERMEABILIZAÇÃO': 'Impermeabilizacao',
        'REVESTIMENTOS ARGAMASSADOS': 'Revestimentos',
        'REVESTIMENTOS DE PISO': 'Pisos',
        'REVESTIMENTOS DE PAREDE': 'Revestimentos',
        'REVESTIMENTOS DE TETO': 'Forro',
        'PINTURA': 'Pintura',
        'ESQUADRIAS': 'Esquadrias',
        'FACHADA': 'Fachada',
        'COBERTURA': 'Cobertura',
        'COMPLEMENTAR': 'Servicos Complementares',
        'IMPREVISTO': 'Servicos Complementares',
        'RODAPÉS': 'Revestimentos',
        'LOUÇAS': 'Loucas e Metais',
        'SERVIÇOS INICIAIS': 'Gerenciamento',
        'SERVIÇOS TÉCNICOS': 'Gerenciamento',
        'EQUIPAMENTOS': 'Gerenciamento',
        'MARKETING': 'Gerenciamento',
        'MOVIMENTAÇÃO': 'Movimentacao de Terra',
        'FUNDAÇ': 'Fundacoes',
        'CONTENÇ': 'Contencao',
    }

    disciplinas = {}
    current_macro_desc = None
    current_disc_label = None
    current_subcat = 'GERAL'

    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue

        code = str(row[0]).strip()
        desc = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        if not desc or code == 'Código':
            continue

        parts = code.split('.')
        depth = len(parts)

        unit_val = str(row[13]).strip() if len(row) > 13 and row[13] else ''
        qty = safe_float(row[16]) if len(row) > 16 else None
        pu = safe_float(row[22]) if len(row) > 22 else None
        total = safe_float(row[28]) if len(row) > 28 else None

        # Level 1: macrogroup (XX)
        if depth == 1 and re.match(r'^\d+\s*$', code):
            current_macro_desc = desc
            current_subcat = 'GERAL'
            # Map to discipline
            upper_desc = desc.upper()
            current_disc_label = None
            for key, label in SIENGE_DISC_MAP.items():
                if key in upper_desc:
                    current_disc_label = label
                    break
            if current_disc_label and current_disc_label not in disciplinas:
                disciplinas[current_disc_label] = {
                    'itens': [], 'subcategorias': {}, 'total': 0, 'n_itens': 0, 'warnings': [],
                }

        # Level 2: section (XX.XXX)
        elif depth == 2 and re.match(r'^\d+\.\d+\s*$', code):
            current_subcat = desc
            if current_disc_label and current_disc_label in disciplinas:
                if current_subcat not in disciplinas[current_disc_label]['subcategorias']:
                    disciplinas[current_disc_label]['subcategorias'][current_subcat] = {
                        'total': total or 0, 'itens': [],
                    }

        # Level 4: item (XX.XXX.XXX.XXX) — has qty, unit, pu
        elif depth == 4 and re.match(r'^\d+\.\d+\.\d+\.\d+\s*$', code):
            if current_disc_label and current_disc_label in disciplinas and qty and unit_val and total:
                item = {
                    'descricao': desc,
                    'quantidade': qty,
                    'unidade': normalizar_unidade(unit_val),
                    'pu': pu if pu else (total / qty if qty > 0 else None),
                    'total': total,
                    'subgrupo': current_subcat,
                    'pavimento': 'GERAL',
                    'chave_normalizada': normalizar_item(desc),
                }
                disciplinas[current_disc_label]['itens'].append(item)
                disciplinas[current_disc_label]['n_itens'] += 1
                if current_subcat in disciplinas[current_disc_label]['subcategorias']:
                    disciplinas[current_disc_label]['subcategorias'][current_subcat]['itens'].append(item)

    # Calculate totals
    for disc_label, disc_data in disciplinas.items():
        disc_data['total'] = sum(i['total'] for i in disc_data['itens'] if i.get('total'))
        disc_data['rsm2'] = disc_data['total'] / ac if ac and disc_data['total'] else None

    return disciplinas


def extrair_macrogrupos(wb, ac):
    """Extract macrogroup-level summary. Returns dict of macrogroups."""
    target_sheets = [
        'Orçamento Resumo', 'Gerenciamento executivo ', 'Gerenciamento executivo',
        'Resumo', 'RESUMOS', 'GER_EXECUTIVO', 'Gerenciamento_Exec',
        'GERENCIAMENTO_EXECUTIVO', 'Orçamento Executivo', 'Orçamento_Executivo',
        'Orçamento', 'EAP', 'EAP Análise',
    ]

    ws = None
    for s in target_sheets:
        if s in wb.sheetnames:
            ws = wb[s]
            break
    if not ws:
        for sheet_name in wb.sheetnames:
            upper = sheet_name.upper().strip()
            if any(p in upper for p in ['ORÇAMENTO', 'EXECUTIVO', 'RESUMO', 'EAP', 'GERENCIAMENTO']):
                ws = wb[sheet_name]
                break
    if not ws:
        return {}

    macros = {}
    for row in ws.iter_rows(max_row=60, max_col=10, values_only=True):
        if not row or not row[0]:
            continue
        desc = str(row[0]).strip()
        upper = desc.upper()
        if not any(kw in upper for kw in MACRO_KEYWORDS):
            continue
        valor = safe_float(row[1])
        if not valor or valor < 100:
            # Try other columns
            for c in row[2:6]:
                v = safe_float(c)
                if v and v > 100:
                    valor = v
                    break
        if valor and valor > 100:
            macros[desc] = {
                'valor': valor,
                'rsm2': valor / ac if ac else None,
            }

    return macros


# ============================================================
# DERIVED INDICES
# ============================================================

def calcular_indices_consumo(disciplinas, ac, ur=None):
    """Calculate consumption indices (quantity-based, price-independent)."""
    indices = {}

    for disc_label, disc_data in disciplinas.items():
        disc_indices = {}
        itens = disc_data.get('itens', [])

        # Structural indices
        if disc_label == 'Supraestrutura':
            concreto_m3 = sum(i['quantidade'] for i in itens
                              if 'concreto' in i.get('chave_normalizada', '') and i.get('unidade') == 'm3')
            aco_kg = sum(i['quantidade'] for i in itens
                         if ('aco' in i.get('chave_normalizada', '') or 'armadura' in i.get('chave_normalizada', ''))
                         and i.get('unidade') == 'kg')
            forma_m2 = sum(i['quantidade'] for i in itens
                           if 'forma' in i.get('chave_normalizada', '') and i.get('unidade') == 'm2')

            if concreto_m3 > 0 and ac:
                disc_indices['concreto_m3_por_m2_ac'] = round(concreto_m3 / ac, 4)
            if aco_kg > 0 and concreto_m3 > 0:
                disc_indices['aco_kg_por_m3_concreto'] = round(aco_kg / concreto_m3, 1)
            if forma_m2 > 0 and concreto_m3 > 0:
                disc_indices['forma_m2_por_m3_concreto'] = round(forma_m2 / concreto_m3, 2)

        # Electrical indices
        elif disc_label == 'Instalacoes Eletricas':
            eletroduto_m = sum(i['quantidade'] for i in itens
                               if 'eletroduto' in i.get('chave_normalizada', '') and i.get('unidade') == 'm')
            pontos = sum(i['quantidade'] for i in itens
                         if any(k in i.get('chave_normalizada', '')
                                for k in ['ponto', 'tomada', 'interruptor', 'luminaria']))
            if eletroduto_m > 0 and ac:
                disc_indices['eletroduto_m_por_m2_ac'] = round(eletroduto_m / ac, 2)
            if pontos > 0 and ur:
                disc_indices['pontos_eletricos_por_ur'] = round(pontos / ur, 1)

        # Hydraulic indices
        elif disc_label == 'Instalacoes Hidraulicas':
            tubo_m = sum(i['quantidade'] for i in itens
                         if any(k in i.get('chave_normalizada', '')
                                for k in ['tubo', 'tubulacao', 'cano', 'encanamento'])
                         and i.get('unidade') == 'm')
            pontos_hidro = sum(i['quantidade'] for i in itens
                               if 'ponto' in i.get('chave_normalizada', ''))
            if tubo_m > 0 and ac:
                disc_indices['tubulacao_m_por_m2_ac'] = round(tubo_m / ac, 2)
            if pontos_hidro > 0 and ur:
                disc_indices['pontos_hidro_por_ur'] = round(pontos_hidro / ur, 1)

        if disc_indices:
            indices[disc_label] = disc_indices

    return indices


def calcular_split_mo_material(disciplinas):
    """Calculate MO vs Material split per discipline."""
    splits = {}

    for disc_label, disc_data in disciplinas.items():
        itens = disc_data.get('itens', [])
        total_disc = sum(i['total'] for i in itens if i.get('total'))
        if total_disc <= 0:
            continue

        mo_total = 0
        for item in itens:
            desc_lower = item.get('descricao', '').lower()
            chave = item.get('chave_normalizada', '')
            if any(kw in desc_lower or kw in chave for kw in MO_KEYWORDS):
                mo_total += item.get('total', 0)

        mo_pct = mo_total / total_disc if total_disc > 0 else 0
        splits[disc_label] = {
            'mo_pct': round(mo_pct, 3),
            'material_pct': round(1 - mo_pct, 3),
        }

    return splits


def calcular_curva_abc(disciplinas):
    """Calculate ABC curve (top items by cost concentration)."""
    curva = {}

    for disc_label, disc_data in disciplinas.items():
        itens = disc_data.get('itens', [])
        total_disc = sum(i['total'] for i in itens if i.get('total'))
        if total_disc <= 0 or not itens:
            continue

        sorted_itens = sorted(itens, key=lambda x: x.get('total', 0), reverse=True)
        top_20 = sorted_itens[:20]
        top_20_total = sum(i['total'] for i in top_20 if i.get('total'))

        curva[disc_label] = {
            'top_20_pct': round(top_20_total / total_disc, 3) if total_disc > 0 else 0,
            'n_itens_total': len(itens),
            'itens': [
                {
                    'desc': i['descricao'][:60],
                    'pct': round(i['total'] / total_disc, 3) if total_disc > 0 else 0,
                    'total': i['total'],
                }
                for i in top_20
                if i.get('total', 0) > 0
            ],
        }

    return curva


# ============================================================
# MAIN PROCESSING
# ============================================================

def processar_executivo(xlsx_path, slug=None):
    """Process a single executivo xlsx. Returns (metadados, indices, raw) or raises."""
    filename = os.path.basename(xlsx_path)
    if not slug:
        slug = slug_from_filename(filename)

    print(f"  Processando: {filename} (slug: {slug})")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    print(f"    Abas ({len(sheets)}): {sheets[:10]}{'...' if len(sheets) > 10 else ''}")

    # 1. Metadados
    meta = extrair_metadados(wb, filename)
    meta['slug'] = slug
    ac = meta.get('ac')

    # Fallback AC from calibration
    if not ac:
        ac_map = load_calibration_ac()
        for cal_name, cal_ac in ac_map.items():
            if cal_name in slug or slug in cal_name:
                ac = cal_ac
                meta['ac'] = ac
                meta['ac_source'] = 'calibration'
                break

    if ac:
        print(f"    AC: {ac:.2f} m²")
    else:
        print(f"    ⚠ AC não encontrado")

    ur = meta.get('ur')

    # 2. Macrogrupos
    macrogrupos = extrair_macrogrupos(wb, ac)
    if macrogrupos:
        total_geral = sum(m['valor'] for m in macrogrupos.values())
        print(f"    Macrogrupos: {len(macrogrupos)}, Total: R$ {total_geral:,.2f}")
        meta['total'] = total_geral
        if ac:
            meta['rsm2'] = total_geral / ac

    # 3. Itens detalhados por disciplina
    disciplinas = {}
    total_itens = 0

    # Strategy 1: Sienge format (single "Relatório" or "EAP" sheet with XX.XXX.XXX.XXX codes)
    sienge_sheets = [s for s in sheets if s.strip() in ('Relatório', 'EAP') or 'Planilha Orçamento Sienge' in s]
    if sienge_sheets and len(sheets) <= 5:
        print(f"    Formato Sienge/EAP detectado ({sienge_sheets[0]})")
        for ss in sienge_sheets:
            try:
                # EAP sheets use same code structure but columns at 0-5 instead of Sienge's wide layout
                ws_check = wb[ss]
                # Peek at first data row to decide extractor
                sample_row = None
                for r in ws_check.iter_rows(min_row=8, max_row=25, values_only=True):
                    if r and r[0] and re.match(r'^\d+\s*$', str(r[0]).strip()):
                        sample_row = r
                        break
                if sample_row and len([c for c in sample_row if c is not None]) > 15:
                    # Wide Sienge format (30+ columns)
                    disciplinas = extrair_sienge(ws_check, ac, ur)
                else:
                    # Narrow EAP format (6-8 columns)
                    disciplinas = extrair_analitico_eap(ws_check, ac)
            except Exception as e:
                print(f"    ⚠ Erro Sienge/EAP ({ss}): {e}")
            if disciplinas:
                break

    # Strategy 2: Multi-tab format (individual discipline sheets)
    if not disciplinas:
        for sheet_name in sheets:
            disc_label = mapear_disciplina(sheet_name)
            if disc_label and disc_label not in disciplinas:
                try:
                    disc_data = extrair_itens_disciplina(wb[sheet_name], ac)
                    if disc_data['n_itens'] > 0:
                        disc_data['rsm2'] = disc_data['total'] / ac if ac and disc_data['total'] else None
                        disciplinas[disc_label] = disc_data
                        total_itens += disc_data['n_itens']
                        print(f"    {disc_label}: {disc_data['n_itens']} itens, R$ {disc_data['total']:,.2f}")
                except Exception as e:
                    print(f"    ⚠ Erro em {sheet_name}: {e}")

    # Strategy 3: Analítico/EAP format (hierarchical codes in budget sheet)
    if not disciplinas:
        eap_candidates = [s for s in sheets if any(k in s.upper() for k in
                          ['ORÇAMENTO EXECUTIVO', 'ORCAMENTO EXECUTIVO', 'EAP_ANALITICO',
                           'EAP ANALITICO', 'ORÇAMENTO_EXECUTIVO', 'ORCAMENTO_EXECUTIVO',
                           'OR ANALITICO'])]
        for ss in eap_candidates:
            try:
                print(f"    Tentando analítico em '{ss}'")
                disciplinas = extrair_analitico_eap(wb[ss], ac)
            except Exception as e:
                print(f"    ⚠ Erro analítico ({ss}): {e}")
            if disciplinas:
                break

    # Strategy 4: ABC Insumos (flat item list sorted by cost)
    if not disciplinas:
        abc_candidates = [s for s in sheets if any(k in s.upper() for k in
                          ['ABC_INSUMOS', 'ABC INSUMOS', 'INSUMOS'])]
        for ss in abc_candidates:
            try:
                print(f"    Tentando ABC/Insumos em '{ss}'")
                disciplinas = extrair_abc_insumos(wb[ss], ac)
            except Exception as e:
                print(f"    ⚠ Erro ABC ({ss}): {e}")
            if disciplinas:
                break

    # Count totals from all strategies
    total_itens = sum(d.get('n_itens', 0) for d in disciplinas.values())
    for disc_label, disc_data in disciplinas.items():
        if disc_data.get('n_itens', 0) > 0:
            print(f"    {disc_label}: {disc_data['n_itens']} itens, R$ {disc_data['total']:,.2f}")

    print(f"    Total: {len(disciplinas)} disciplinas, {total_itens} itens")

    # 4. Índices derivados
    indices_consumo = calcular_indices_consumo(disciplinas, ac, ur)
    split_mo = calcular_split_mo_material(disciplinas)
    curva_abc = calcular_curva_abc(disciplinas)

    wb.close()

    # Build outputs
    raw_data = {
        'projeto': slug,
        'arquivo': filename,
        'ac': ac,
        'ur': ur,
        'disciplinas': {},
    }
    for disc_label, disc_data in disciplinas.items():
        raw_data['disciplinas'][disc_label] = {
            'total': disc_data['total'],
            'n_itens': disc_data['n_itens'],
            'itens': disc_data['itens'],
        }

    indices_data = {
        'projeto': slug,
        'ac': ac,
        'ur': ur,
        'total': meta.get('total'),
        'rsm2': meta.get('rsm2'),
        'macrogrupos': macrogrupos,
        'disciplinas': {
            disc: {
                'total': d['total'],
                'rsm2': d.get('rsm2'),
                'n_itens': d['n_itens'],
                'subcategorias': {
                    k: {'total': v['total'], 'n_itens': len(v.get('itens', []))}
                    for k, v in d.get('subcategorias', {}).items()
                },
            }
            for disc, d in disciplinas.items()
        },
        'indices_consumo': indices_consumo,
        'split_mo_material': split_mo,
        'curva_abc': curva_abc,
    }

    return meta, indices_data, raw_data


def salvar_outputs(slug, meta, indices_data, raw_data):
    """Save all outputs for a processed project."""
    # Raw data
    raw_path = os.path.join(RAW_DIR, f'{slug}-raw.json')
    with open(raw_path, 'w', encoding='utf-8') as f:
        json.dump(raw_data, f, ensure_ascii=False, indent=2)

    # Indices
    idx_path = os.path.join(INDICES_DIR, f'{slug}.json')
    with open(idx_path, 'w', encoding='utf-8') as f:
        json.dump(indices_data, f, ensure_ascii=False, indent=2)

    # Update metadados
    metadados = load_metadados()
    # Preserve user edits (cidade, padrao, tipologia)
    existing = metadados.get(slug, {})
    for key in ['cidade', 'estado', 'padrao', 'tipologia', 'nome_completo', 'cliente']:
        if key in existing and key not in meta:
            meta[key] = existing[key]
    metadados[slug] = meta
    save_metadados(metadados)

    print(f"    Salvo: {raw_path}")
    print(f"    Salvo: {idx_path}")


# ============================================================
# CLI COMMANDS
# ============================================================

def cmd_inventory():
    """List and classify all xlsx files."""
    files = sorted([
        os.path.join(root, f)
        for root, dirs, fnames in os.walk(EXEC_DIR)
        for f in fnames
        if f.endswith('.xlsx')
    ])

    metadados = load_metadados()
    processados_files = {m.get('arquivo_fonte') for m in metadados.values() if m.get('status') == 'processado'}
    processados_slugs = {slug for slug, m in metadados.items() if m.get('status') == 'processado'}

    executivos_by_slug = {}
    apresentacoes = []

    for filepath in files:
        filename = os.path.basename(filepath)
        slug = slug_from_filename(filepath)
        already = filename in processados_files or slug in processados_slugs

        if is_executivo(filename):
            entry = {
                'path': filepath,
                'filename': filename,
                'slug': slug,
                'processado': already,
            }
            # Dedup: prefer executivo > apresentação, entrega- > drive-, higher revision
            if slug in executivos_by_slug:
                existing = executivos_by_slug[slug]
                new_upper = filename.upper()
                old_upper = existing['filename'].upper()

                # Score: executivo/orçamento > apresentação
                def _score(fn):
                    s = 0
                    u = fn.upper()
                    if 'EXECUTIVO' in u or 'COMPLETO' in u or 'ENTREGAVEL' in u or 'ENTREGÁVEL' in u:
                        s += 10
                    if 'ORÇAMENTO' in u and 'APRESENTAÇÃO' not in u and 'APRESENTACAO' not in u:
                        s += 5
                    if 'entrega-' in fn:
                        s += 2
                    rev = re.search(r'R(\d+)', fn)
                    if rev:
                        s += int(rev.group(1))
                    return s

                if _score(filename) > _score(existing['filename']):
                    executivos_by_slug[slug] = entry
            else:
                executivos_by_slug[slug] = entry
        else:
            apresentacoes.append(filename)

    executivos = sorted(executivos_by_slug.values(), key=lambda x: x['slug'])

    print(f"\n{'='*70}")
    print(f"INVENTÁRIO DE EXECUTIVOS")
    print(f"{'='*70}")
    print(f"Total arquivos: {len(files)}")
    print(f"Executivos: {len(executivos)}")
    print(f"Apresentações (excluídos): {len(apresentacoes)}")
    print(f"Já processados: {sum(1 for e in executivos if e['processado'])}")
    print(f"Pendentes: {sum(1 for e in executivos if not e['processado'])}")

    print(f"\n--- EXECUTIVOS PENDENTES ---")
    for e in executivos:
        if not e['processado']:
            print(f"  [{e['slug']}] {e['filename']}")

    print(f"\n--- JÁ PROCESSADOS ---")
    for e in executivos:
        if e['processado']:
            print(f"  [{e['slug']}] {e['filename']}")

    print(f"\n--- EXCLUÍDOS (apresentações) ---")
    for f in apresentacoes[:10]:
        print(f"  {f}")
    if len(apresentacoes) > 10:
        print(f"  ... e mais {len(apresentacoes) - 10}")

    return executivos


def cmd_process(xlsx_path):
    """Process a single file."""
    if not os.path.exists(xlsx_path):
        print(f"Arquivo não encontrado: {xlsx_path}")
        return False

    slug = slug_from_filename(xlsx_path)
    try:
        meta, indices_data, raw_data = processar_executivo(xlsx_path, slug)
        salvar_outputs(slug, meta, indices_data, raw_data)
        print(f"  ✅ {slug} processado com sucesso")
        return True
    except Exception as e:
        print(f"  ❌ {slug} erro: {e}")
        # Register error in metadados
        metadados = load_metadados()
        metadados[slug] = {
            'arquivo_fonte': os.path.basename(xlsx_path),
            'processado_em': datetime.now().strftime('%Y-%m-%d'),
            'status': 'erro',
            'erro': str(e),
        }
        save_metadados(metadados)
        return False


def cmd_batch(limit=None):
    """Process all pending files."""
    executivos = cmd_inventory()
    pendentes = [e for e in executivos if not e['processado']]

    if limit:
        pendentes = pendentes[:limit]

    print(f"\n{'='*70}")
    print(f"PROCESSAMENTO BATCH: {len(pendentes)} arquivos")
    print(f"{'='*70}\n")

    success = 0
    errors = 0

    for i, e in enumerate(pendentes, 1):
        print(f"\n[{i}/{len(pendentes)}] {e['filename']}")
        ok = cmd_process(e['path'])
        if ok:
            success += 1
        else:
            errors += 1

        # Save progress every 5
        if i % 5 == 0:
            print(f"\n  >>> Progresso: {success} ok, {errors} erros, {len(pendentes) - i} restantes\n")

    print(f"\n{'='*70}")
    print(f"BATCH COMPLETO: {success} ok, {errors} erros")
    print(f"{'='*70}")

    return success, errors


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Processar executivos Cartesian')
    parser.add_argument('--inventory', action='store_true', help='Listar e classificar arquivos')
    parser.add_argument('--process', type=str, help='Processar 1 arquivo (path)')
    parser.add_argument('--batch', action='store_true', help='Processar todos pendentes')
    parser.add_argument('--limit', type=int, help='Limite de arquivos no batch')

    args = parser.parse_args()

    if args.inventory:
        cmd_inventory()
    elif args.process:
        cmd_process(args.process)
    elif args.batch:
        cmd_batch(args.limit)
    else:
        parser.print_help()


if __name__ == '__main__':
    main()
