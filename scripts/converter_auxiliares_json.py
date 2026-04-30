"""Converte xlsx auxiliares em JSON machine-readable.

- 01-eap/eap.xlsx     → 01-eap/eap.json
  Árvore hierárquica: Unidade Construtiva → Célula Construtiva → Etapa → Subetapa

- 02-composicoes-insumos/insumos.xlsx → 02-composicoes-insumos/insumos-precos.json
  Lista {descricao, und, preco, grupo, sub_grupo, grupo_sienge, observacoes}

Uso:
    py -3.10 -X utf8 ~/orcamentos-openclaw/scripts/converter_auxiliares_json.py
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_ORCAMENTO_DIR = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra\orcamento"
)


def convert_eap(xlsx_path: Path, json_path: Path) -> int:
    """Converte EAP em árvore hierárquica."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = wb["EAP"]

    tree: dict = {"unidades": []}
    current_unidade = None
    current_celula = None
    current_etapa = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        tipo = str(row[0]).strip()
        codigo = row[1] if len(row) > 1 else None
        descricao = row[2] if len(row) > 2 else None

        if tipo == "UNIDADE CONSTRUTIVA":
            current_unidade = {
                "codigo": str(codigo) if codigo is not None else "",
                "descricao": descricao or "",
                "celulas": [],
            }
            tree["unidades"].append(current_unidade)
            current_celula = None
            current_etapa = None
        elif tipo == "CÉLULA CONSTRUTIVA" and current_unidade is not None:
            current_celula = {
                "codigo": str(codigo) if codigo is not None else "",
                "descricao": descricao or "",
                "etapas": [],
            }
            current_unidade["celulas"].append(current_celula)
            current_etapa = None
        elif tipo == "ETAPA" and current_celula is not None:
            current_etapa = {
                "codigo": str(codigo) if codigo is not None else "",
                "descricao": descricao or "",
                "subetapas": [],
            }
            current_celula["etapas"].append(current_etapa)
        elif tipo == "Subetapa" and current_etapa is not None:
            current_etapa["subetapas"].append({
                "codigo": str(codigo) if codigo is not None else "",
                "descricao": descricao or "",
            })

    wb.close()

    # agregados
    n_unidades = len(tree["unidades"])
    n_celulas = sum(len(u["celulas"]) for u in tree["unidades"])
    n_etapas = sum(len(c["etapas"]) for u in tree["unidades"] for c in u["celulas"])
    n_subetapas = sum(
        len(e["subetapas"])
        for u in tree["unidades"]
        for c in u["celulas"]
        for e in c["etapas"]
    )
    tree["_resumo"] = {
        "unidades": n_unidades,
        "celulas": n_celulas,
        "etapas": n_etapas,
        "subetapas": n_subetapas,
    }

    json_path.write_text(json.dumps(tree, ensure_ascii=False, indent=2), encoding="utf-8")
    return n_subetapas


def convert_insumos(xlsx_path: Path, json_path: Path) -> int:
    """Converte insumos em lista plana com preços."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    ws = wb["Insumos"]

    # primeira linha = header
    header = {str(c.value).strip().lower() if c.value else "": c.column_letter for c in ws[1]}
    # mapeamento robustos
    col_desc = ws[1][1].column_letter  # B
    col_und = ws[1][2].column_letter   # C
    col_custo = ws[1][4].column_letter  # E
    col_grupo = ws[1][5].column_letter  # F
    col_sub = ws[1][6].column_letter    # G
    col_obs = ws[1][7].column_letter    # H
    col_sienge = ws[1][8].column_letter  # I

    insumos = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        desc = str(row[1]).strip()
        if not desc or desc in seen:
            continue
        seen.add(desc)

        try:
            preco = float(row[4]) if row[4] is not None else None
        except (TypeError, ValueError):
            preco = None

        insumos.append({
            "descricao": desc,
            "unidade": str(row[2]).strip() if row[2] else None,
            "preco_unitario": preco,
            "grupo": str(row[5]).strip() if row[5] else None,
            "sub_grupo": str(row[6]).strip() if row[6] else None,
            "observacoes": str(row[7]).strip() if row[7] else None,
            "grupo_sienge": str(row[8]).strip() if row[8] else None,
        })

    wb.close()

    out = {
        "fonte": xlsx_path.name,
        "total_insumos": len(insumos),
        "insumos": insumos,
    }
    json_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    return len(insumos)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--orcamento-dir", type=Path, default=DEFAULT_ORCAMENTO_DIR)
    args = parser.parse_args(argv)

    print(f"Orçamento dir: {args.orcamento_dir}", flush=True)
    print()

    # EAP
    eap_xlsx = args.orcamento_dir / "01-eap" / "eap.xlsx"
    eap_json = args.orcamento_dir / "01-eap" / "eap.json"
    if eap_xlsx.exists():
        n = convert_eap(eap_xlsx, eap_json)
        print(f"[01-eap] eap.json escrito ({n} subetapas, {eap_json.stat().st_size/1024:.1f} KB)", flush=True)
    else:
        print(f"[01-eap] eap.xlsx não encontrado, skip", flush=True)

    # Insumos
    insumos_xlsx = args.orcamento_dir / "02-composicoes-insumos" / "insumos.xlsx"
    insumos_json = args.orcamento_dir / "02-composicoes-insumos" / "insumos-precos.json"
    if insumos_xlsx.exists():
        n = convert_insumos(insumos_xlsx, insumos_json)
        print(f"[02-comp] insumos-precos.json escrito ({n} insumos, {insumos_json.stat().st_size/1024:.1f} KB)", flush=True)
    else:
        print(f"[02-comp] insumos.xlsx não encontrado, skip", flush=True)

    return 0


if __name__ == "__main__":
    sys.exit(main())
