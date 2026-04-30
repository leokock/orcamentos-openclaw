"""Gera a aba Visus de uma planilha executiva Cartesian a partir da aba EAP.

Pra cada Subetapa dentro das células construtivas alvo, cria 1 item-filho com
quant=1 e unidade=vb — placeholder que o Leo desdobra depois dentro do AltoQi Visus.

Estrutura gerada (com ponto final em todos os códigos, padrão AltoQi):
    1.           INFRAESTRUTURA
    1.1.         MOVIMENTAÇÃO DE TERRA
    1.1.1.       Corte e aterro
    1.1.1.1.     Corte e aterro              1   vb
    ...

Uso:
    py -3.10 -X utf8 gerar_visus_da_eap.py --xlsx "<path>" [--celulas 1,2,4,5,6,7] [--dry-run]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl


TIPO_UC = "UNIDADE CONSTRUTIVA"
TIPO_CELULA = "CÉLULA CONSTRUTIVA"
TIPO_ETAPA = "ETAPA"
TIPO_SUBETAPA = "Subetapa"

ABA_EAP = "EAP"
ABA_VISUS = "Visus"

HEADER_VISUS = ("EAP", "Descrição", "Quant.", "Unidade")


def ler_eap(ws, celulas_alvo: set[int], uc_alvo: int) -> list[tuple[str, str, str, str, object, object]]:
    """Lê aba EAP e retorna lista de linhas prontas pra escrever na Visus.

    Só considera células dentro da UNIDADE CONSTRUTIVA alvo (uc_alvo).
    Cada item da lista: (tipo, cod, nome_celula, nome_etapa, quant, unidade).
    Tipo ∈ {'celula', 'etapa', 'subetapa', 'item'}.
    """
    linhas_saida: list[tuple] = []

    uc_atual = None
    celula_num = None
    celula_nome = None
    etapa_idx_por_celula: dict[int, int] = {}
    subetapa_idx_por_etapa: dict[tuple[int, int], int] = {}
    etapa_atual = None  # (celula_num, etapa_idx)

    dentro_alvo = False

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        tipo, cod, nome = row[0], row[1], row[2]
        if tipo is None:
            continue

        if tipo == TIPO_UC:
            try:
                uc_atual = int(cod)
            except (TypeError, ValueError):
                uc_atual = None
            dentro_alvo = False
            continue

        if tipo == TIPO_CELULA:
            try:
                celula_num = int(cod)
            except (TypeError, ValueError):
                celula_num = None
                dentro_alvo = False
                continue
            celula_nome = str(nome).strip() if nome else ""
            dentro_alvo = (uc_atual == uc_alvo) and (celula_num in celulas_alvo)
            if dentro_alvo:
                cod_str = f"{celula_num:02d}."
                linhas_saida.append(("celula", cod_str, celula_nome, "", "", ""))
            continue

        if not dentro_alvo:
            continue

        if tipo == TIPO_ETAPA:
            etapa_idx_por_celula[celula_num] = etapa_idx_por_celula.get(celula_num, 0) + 1
            etapa_idx = etapa_idx_por_celula[celula_num]
            etapa_atual = (celula_num, etapa_idx)
            subetapa_idx_por_etapa[etapa_atual] = 0
            etapa_nome = str(nome).strip() if nome else ""
            cod_str = f"{celula_num:02d}.{etapa_idx}."
            linhas_saida.append(("etapa", cod_str, etapa_nome, "", "", ""))
            continue

        if tipo == TIPO_SUBETAPA:
            if etapa_atual is None:
                continue
            subetapa_idx_por_etapa[etapa_atual] += 1
            sub_idx = subetapa_idx_por_etapa[etapa_atual]
            sub_nome = str(nome) if nome else ""
            cod_sub = f"{celula_num:02d}.{etapa_atual[1]}.{sub_idx}."
            cod_item = f"{celula_num:02d}.{etapa_atual[1]}.{sub_idx}.1."
            # Linha subetapa (sem quant/un)
            linhas_saida.append(("subetapa", cod_sub, sub_nome, "", "", ""))
            # Linha item-filho (quant=1, un=vb), nome = mesmo da subetapa
            linhas_saida.append(("item", cod_item, sub_nome, "", 1, "vb"))
            continue

    return linhas_saida


def escrever_visus(ws, linhas: list[tuple]) -> None:
    """Limpa aba Visus (preserva linha 1 = header) e escreve novas linhas."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    # Garante header correto
    for col_idx, valor in enumerate(HEADER_VISUS, start=1):
        ws.cell(row=1, column=col_idx, value=valor)

    for (_tipo, cod, nome, _filler, quant, unid) in linhas:
        descricao = f"{cod} {nome}".strip()
        ws.append((cod, descricao, quant, unid))


def resumo(linhas: list[tuple]) -> dict[str, int]:
    contagem = {"celula": 0, "etapa": 0, "subetapa": 0, "item": 0}
    por_celula: dict[str, int] = {}
    for (tipo, cod, nome, *_rest) in linhas:
        contagem[tipo] = contagem.get(tipo, 0) + 1
        if tipo == "celula":
            por_celula[f"{cod} {nome}"] = 0
        elif tipo == "item":
            # Pega prefixo da célula (ex "1." de "1.1.1.1.")
            prefixo = cod.split(".")[0] + "."
            for chave in por_celula:
                if chave.startswith(prefixo + " "):
                    por_celula[chave] += 1
                    break
    return {"total": contagem, "por_celula_itens": por_celula}


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", required=True, help="Caminho da planilha executiva")
    p.add_argument(
        "--celulas",
        default="1,2,4,5,6,7",
        help="Códigos de células construtivas da EAP (default: 1,2,4,5,6,7)",
    )
    p.add_argument(
        "--uc",
        type=int,
        default=2,
        help="Unidade Construtiva alvo (default: 2 = GERENCIAMENTO EXECUTIVO / obra)",
    )
    p.add_argument("--dry-run", action="store_true", help="Não salva, só imprime resumo")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERRO: {xlsx_path} não existe", file=sys.stderr)
        return 2

    celulas_alvo = {int(x.strip()) for x in args.celulas.split(",") if x.strip()}
    print(f"UC alvo: {args.uc}  |  Células alvo: {sorted(celulas_alvo)}")

    wb = openpyxl.load_workbook(xlsx_path)
    if ABA_EAP not in wb.sheetnames:
        print(f"ERRO: aba '{ABA_EAP}' não encontrada", file=sys.stderr)
        return 2
    if ABA_VISUS not in wb.sheetnames:
        print(f"ERRO: aba '{ABA_VISUS}' não encontrada", file=sys.stderr)
        return 2

    linhas = ler_eap(wb[ABA_EAP], celulas_alvo, args.uc)
    info = resumo(linhas)
    print(f"Linhas geradas: {sum(info['total'].values())}")
    for tipo, n in info["total"].items():
        print(f"  {tipo:10} = {n}")
    print("Itens-filho (vb) por célula:")
    for chave, n in info["por_celula_itens"].items():
        print(f"  {chave:60} → {n}")

    if args.dry_run:
        print("--dry-run: nada salvo.")
        return 0

    escrever_visus(wb[ABA_VISUS], linhas)
    wb.save(xlsx_path)
    print(f"OK: aba '{ABA_VISUS}' atualizada em {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
