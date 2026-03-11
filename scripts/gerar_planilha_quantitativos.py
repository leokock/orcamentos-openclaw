#!/usr/bin/env python3.11
"""
Gerador de Planilha de Extração de Quantitativos por Pavimento
Template padrão Cartesian - aguardando arquivo Ger_Executivo_Cartesian.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os

# ============================================================================
# SERVIÇOS PADRÃO CARTESIAN (template)
# ============================================================================

SERVICOS_PADRAO = [
    # ESTRUTURA
    {"grupo": "ESTRUTURA", "servico": "Concreto Usinado fck 30 MPa", "unidade": "m³"},
    {"grupo": "ESTRUTURA", "servico": "Aço CA-50 - fornecimento e instalação", "unidade": "kg"},
    {"grupo": "ESTRUTURA", "servico": "Forma de madeira compensada resinada", "unidade": "m²"},
    {"grupo": "ESTRUTURA", "servico": "Laje nervurada com cubetas plásticas", "unidade": "m²"},
    
    # ALVENARIA
    {"grupo": "ALVENARIA", "servico": "Alvenaria de vedação bloco cerâmico 14x19x29", "unidade": "m²"},
    {"grupo": "ALVENARIA", "servico": "Alvenaria estrutural bloco 14x19x39", "unidade": "m²"},
    
    # REVESTIMENTOS
    {"grupo": "REVESTIMENTOS", "servico": "Chapisco", "unidade": "m²"},
    {"grupo": "REVESTIMENTOS", "servico": "Emboço interno e=2cm", "unidade": "m²"},
    {"grupo": "REVESTIMENTOS", "servico": "Reboco externo e=2cm", "unidade": "m²"},
    {"grupo": "REVESTIMENTOS", "servico": "Gesso liso interno", "unidade": "m²"},
    {"grupo": "REVESTIMENTOS", "servico": "Massa corrida e pintura acrílica", "unidade": "m²"},
    
    # PISOS
    {"grupo": "PISOS", "servico": "Contrapiso 5cm", "unidade": "m²"},
    {"grupo": "PISOS", "servico": "Porcelanato 60x60 retificado", "unidade": "m²"},
    {"grupo": "PISOS", "servico": "Porcelanato 80x80 retificado", "unidade": "m²"},
    {"grupo": "PISOS", "servico": "Cerâmica 45x45", "unidade": "m²"},
    {"grupo": "PISOS", "servico": "Rodapé porcelanato h=7cm", "unidade": "m"},
    
    # REVESTIMENTOS ESPECIAIS
    {"grupo": "REVESTIMENTOS", "servico": "Revestimento cerâmico parede 30x60", "unidade": "m²"},
    {"grupo": "REVESTIMENTOS", "servico": "Revestimento porcelanato parede 60x120", "unidade": "m²"},
    
    # FORROS
    {"grupo": "FORROS", "servico": "Forro gesso acartonado liso", "unidade": "m²"},
    {"grupo": "FORROS", "servico": "Forro gesso rebaixado", "unidade": "m²"},
    {"grupo": "FORROS", "servico": "Forro mineral removível", "unidade": "m²"},
    
    # ESQUADRIAS
    {"grupo": "ESQUADRIAS", "servico": "Porta de madeira c/ marco 80x210", "unidade": "un"},
    {"grupo": "ESQUADRIAS", "servico": "Janela alumínio anodizado de correr", "unidade": "m²"},
    {"grupo": "ESQUADRIAS", "servico": "Porta corta-fogo PCF-90", "unidade": "un"},
    {"grupo": "ESQUADRIAS", "servico": "Guarda-corpo vidro temperado 10mm", "unidade": "m"},
    
    # IMPERMEABILIZAÇÃO
    {"grupo": "IMPERMEABILIZAÇÃO", "servico": "Impermeabilização manta asfáltica 4mm", "unidade": "m²"},
    {"grupo": "IMPERMEABILIZAÇÃO", "servico": "Impermeabilização cristalizada reservatório", "unidade": "m²"},
    {"grupo": "IMPERMEABILIZAÇÃO", "servico": "Impermeabilização banheiros", "unidade": "m²"},
    
    # INSTALAÇÕES HIDRÁULICAS
    {"grupo": "INSTALAÇÕES", "servico": "Ponto água fria Ø 3/4\"", "unidade": "un"},
    {"grupo": "INSTALAÇÕES", "servico": "Ponto água quente Ø 3/4\"", "unidade": "un"},
    {"grupo": "INSTALAÇÕES", "servico": "Ponto esgoto Ø 100mm", "unidade": "un"},
    {"grupo": "INSTALAÇÕES", "servico": "Prumada água fria Ø 1 1/2\"", "unidade": "m"},
    {"grupo": "INSTALAÇÕES", "servico": "Prumada esgoto Ø 100mm", "unidade": "m"},
    
    # INSTALAÇÕES ELÉTRICAS
    {"grupo": "INSTALAÇÕES", "servico": "Ponto luz", "unidade": "un"},
    {"grupo": "INSTALAÇÕES", "servico": "Ponto tomada 2P+T", "unidade": "un"},
    {"grupo": "INSTALAÇÕES", "servico": "Quadro distribuição embutir", "unidade": "un"},
    {"grupo": "INSTALAÇÕES", "servico": "Prumada elétrica", "unidade": "m"},
    
    # LOUÇAS E METAIS
    {"grupo": "LOUÇAS E METAIS", "servico": "Bacia sanitária c/ caixa acoplada", "unidade": "un"},
    {"grupo": "LOUÇAS E METAIS", "servico": "Lavatório sobrepor", "unidade": "un"},
    {"grupo": "LOUÇAS E METAIS", "servico": "Cuba inox sobrepor", "unidade": "un"},
    {"grupo": "LOUÇAS E METAIS", "servico": "Misturador monocomando", "unidade": "un"},
    {"grupo": "LOUÇAS E METAIS", "servico": "Chuveiro elétrico", "unidade": "un"},
]

# ============================================================================
# PAVIMENTOS DO ARMINIO TAVARES (extraídos do IFC)
# ============================================================================

PAVIMENTOS = [
    "SUBSOLO BAIXO",
    "-01 SUBSOLO",
    "1º PAVIMENTO",
    "2º PAVIMENTO",
    "3º PAVIMENTO",
    "4º PAVIMENTO",
    "5º PAVIMENTO",
    "6º PAVIMENTO",
    "7º PAVIMENTO",
    "8º PAVIMENTO",
    "9º PAVIMENTO",
    "10º PAVIMENTO",
    "11º PAVIMENTO",
    "12º PAVIMENTO",
    "13º PAVIMENTO",
    "14º PAVIMENTO",
    "15º PAVIMENTO",
    "16º PAVIMENTO",
    "BARRILETE",
    "CASA DE MÁQUINAS",
    "RESERVATÓRIO",
    "COBERTURA"
]

# ============================================================================
# GERAR PLANILHA
# ============================================================================

def gerar_planilha_quantitativos():
    wb = Workbook()
    
    # Remover sheet padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========================================================================
    # ABA 1: CAPA
    # ========================================================================
    
    ws_capa = wb.create_sheet("CAPA", 0)
    
    # Header
    ws_capa.merge_cells('A1:D1')
    cell = ws_capa['A1']
    cell.value = "EXTRAÇÃO DE QUANTITATIVOS POR PAVIMENTO"
    cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws_capa.row_dimensions[1].height = 30
    
    # Info
    row = 3
    info = [
        ("Projeto:", "PLACON - ARMÍNIO TAVARES"),
        ("Data:", datetime.now().strftime("%d/%m/%Y")),
        ("Total de Pavimentos:", len(PAVIMENTOS)),
        ("Total de Serviços:", len(SERVICOS_PADRAO)),
        ("", ""),
        ("Instruções:", ""),
        ("1. Preencher quantitativos na aba QUANTITATIVOS", ""),
        ("2. Os totais são calculados automaticamente", ""),
        ("3. Use filtros para visualizar por grupo de serviço", ""),
    ]
    
    for label, value in info:
        ws_capa[f'A{row}'] = label
        ws_capa[f'A{row}'].font = Font(name="Calibri", size=10, bold=True if label else False)
        if value:
            ws_capa[f'B{row}'] = value
            ws_capa[f'B{row}'].font = Font(name="Calibri", size=10)
        row += 1
    
    row += 1
    
    # Lista de pavimentos
    ws_capa[f'A{row}'] = "LISTA DE PAVIMENTOS"
    ws_capa[f'A{row}'].font = Font(name="Calibri", size=11, bold=True)
    ws_capa[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws_capa[f'A{row}'].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    row += 1
    
    for i, pav in enumerate(PAVIMENTOS, 1):
        ws_capa[f'A{row}'] = f"{i}. {pav}"
        ws_capa[f'A{row}'].font = Font(name="Calibri", size=10)
        row += 1
    
    ws_capa.column_dimensions['A'].width = 30
    ws_capa.column_dimensions['B'].width = 40
    
    # ========================================================================
    # ABA 2: QUANTITATIVOS
    # ========================================================================
    
    ws_quant = wb.create_sheet("QUANTITATIVOS", 1)
    
    # Estilos
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalho principal
    col = 1
    headers_fixos = ["GRUPO", "SERVIÇO", "UNIDADE", "TOTAL"]
    
    for header in headers_fixos:
        cell = ws_quant.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
        col += 1
    
    # Cabeçalho dos pavimentos
    for pav in PAVIMENTOS:
        cell = ws_quant.cell(row=1, column=col)
        cell.value = pav
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_thin
        ws_quant.column_dimensions[get_column_letter(col)].width = 12
        col += 1
    
    ws_quant.row_dimensions[1].height = 30
    
    # Linhas de serviços
    row = 2
    grupo_anterior = None
    
    for servico in SERVICOS_PADRAO:
        # Linha de separação de grupo
        if servico['grupo'] != grupo_anterior:
            # Pular linha
            if grupo_anterior is not None:
                row += 1
            
            grupo_anterior = servico['grupo']
        
        # Coluna GRUPO
        ws_quant.cell(row=row, column=1).value = servico['grupo']
        ws_quant.cell(row=row, column=1).font = Font(name="Calibri", size=9, bold=True)
        ws_quant.cell(row=row, column=1).border = border_thin
        ws_quant.cell(row=row, column=1).alignment = Alignment(vertical='top')
        
        # Coluna SERVIÇO
        ws_quant.cell(row=row, column=2).value = servico['servico']
        ws_quant.cell(row=row, column=2).font = Font(name="Calibri", size=9)
        ws_quant.cell(row=row, column=2).border = border_thin
        ws_quant.cell(row=row, column=2).alignment = Alignment(vertical='top', wrap_text=True)
        
        # Coluna UNIDADE
        ws_quant.cell(row=row, column=3).value = servico['unidade']
        ws_quant.cell(row=row, column=3).font = Font(name="Calibri", size=9)
        ws_quant.cell(row=row, column=3).border = border_thin
        ws_quant.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='top')
        
        # Coluna TOTAL (fórmula de soma)
        primeira_col_pav = 5
        ultima_col_pav = 4 + len(PAVIMENTOS)
        formula = f"=SUM({get_column_letter(primeira_col_pav)}{row}:{get_column_letter(ultima_col_pav)}{row})"
        
        ws_quant.cell(row=row, column=4).value = formula
        ws_quant.cell(row=row, column=4).font = Font(name="Calibri", size=9, bold=True)
        ws_quant.cell(row=row, column=4).border = border_thin
        ws_quant.cell(row=row, column=4).number_format = '#,##0.00'
        ws_quant.cell(row=row, column=4).alignment = Alignment(horizontal='right', vertical='top')
        
        # Colunas dos pavimentos (vazias para preenchimento)
        for col_pav in range(primeira_col_pav, ultima_col_pav + 1):
            cell = ws_quant.cell(row=row, column=col_pav)
            cell.font = Font(name="Calibri", size=9)
            cell.border = border_thin
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='top')
        
        row += 1
    
    # Ajustar larguras das colunas fixas
    ws_quant.column_dimensions['A'].width = 18
    ws_quant.column_dimensions['B'].width = 45
    ws_quant.column_dimensions['C'].width = 8
    ws_quant.column_dimensions['D'].width = 12
    
    # Congelar painéis (fixar cabeçalho e primeiras 3 colunas)
    ws_quant.freeze_panes = 'E2'
    
    # Ativar filtros
    ws_quant.auto_filter.ref = f"A1:{get_column_letter(ultima_col_pav)}1"
    
    # ========================================================================
    # SALVAR
    # ========================================================================
    
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    
    filename = f'Extracao-Quantitativos-Arminio-Tavares_{datetime.now().strftime("%Y%m%d")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    print("\n" + "="*80)
    print("✓ Planilha de Extração de Quantitativos gerada com sucesso!")
    print("="*80)
    print(f"\nArquivo: {filepath}")
    print(f"\nEstrutura:")
    print(f"  • {len(PAVIMENTOS)} pavimentos")
    print(f"  • {len(SERVICOS_PADRAO)} serviços padrão")
    print(f"  • Colunas congeladas para fácil navegação")
    print(f"  • Filtros automáticos habilitados")
    print(f"  • Totais calculados automaticamente")
    print("\n⚠️  IMPORTANTE: Esta é uma planilha MODELO com serviços padrão.")
    print("   Para adaptar aos serviços do Ger_Executivo_Cartesian.xlsx,")
    print("   envie o arquivo para ajuste automático.")
    print("="*80)
    
    return filepath

if __name__ == '__main__':
    gerar_planilha_quantitativos()
