"""Atualiza o painel mestre `Estrutura do orçamento.md` com status das 22 disciplinas.

Varre `04-disciplinas/` + `05-extras/`, detecta status (template / memorial / processada),
lê total + pendências dos quantitativos.xlsx processados, e regenera o md.

Status por disciplina:
- ⬜ Template — só extracao.xlsx + memorial stub (< 2 KB)
- 📝 Memorial — memorial.md preenchido (≥ 2 KB) mas sem quantitativos.xlsx
- ✅ Processada — tem quantitativos.xlsx final

Uso:
    py -3.10 -X utf8 ~/orcamentos-openclaw/scripts/atualizar_painel.py
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

ORCAMENTO_DIR = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra\orcamento"
)
PAINEL_PATH = ORCAMENTO_DIR / "Estrutura do orçamento.md"
DISCIPLINAS_DIR = ORCAMENTO_DIR / "04-disciplinas"
EXTRAS_DIR = ORCAMENTO_DIR / "05-extras"

MEMORIAL_STUB_THRESHOLD = 2000  # bytes — acima disso consideramos preenchido


def ler_total_quantitativos(xlsx_path: Path) -> tuple[float | None, dict[str, int]]:
    """Extrai total e contagem de pendências por severidade de um quantitativos.xlsx."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception:
        return None, {}

    total = None
    if "Resumo" in wb.sheetnames:
        ws = wb["Resumo"]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            for i, cell in enumerate(row):
                if isinstance(cell, str) and "TOTAL GERAL" in cell.upper():
                    # procura valor numérico nas células seguintes
                    for j in range(i + 1, len(row)):
                        v = row[j]
                        if isinstance(v, (int, float)):
                            total = float(v)
                            break
                        if isinstance(v, str):
                            m = re.search(r"([\d.,]+)", v.replace("R$", "").strip())
                            if m:
                                num_str = m.group(1).replace(".", "").replace(",", ".")
                                try:
                                    total = float(num_str)
                                    break
                                except ValueError:
                                    pass
                    break
            if total is not None:
                break

    pend_counts = {"Alta": 0, "Média": 0, "Baixa": 0}
    if "Pendências" in wb.sheetnames:
        ws = wb["Pendências"]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row:
                continue
            sev = None
            for cell in row:
                if isinstance(cell, str) and cell.strip() in pend_counts:
                    sev = cell.strip()
                    break
            if sev:
                pend_counts[sev] += 1

    wb.close()
    return total, pend_counts


def detectar_marker(memorial_path: Path) -> str | None:
    """Procura `<!-- STATUS: xxx -->` no topo do memorial."""
    if not memorial_path.exists():
        return None
    try:
        head = memorial_path.read_text(encoding="utf-8")[:2000]
    except Exception:
        return None
    m = re.search(r"<!--\s*STATUS:\s*([\w-]+)\s*-->", head)
    return m.group(1) if m else None


def discovery(folder: Path, kind: str) -> dict:
    """Analisa uma pasta de disciplina e retorna dict de status."""
    memorial = folder / "memorial.md"
    extracao = folder / "extracao.xlsx"
    quantitativos = folder / "quantitativos.xlsx"

    mem_size = memorial.stat().st_size if memorial.exists() else 0
    has_quant = quantitativos.exists()
    mem_filled = mem_size >= MEMORIAL_STUB_THRESHOLD
    marker = detectar_marker(memorial)

    if has_quant:
        status = "processada"
    elif marker == "aguardando-cross-disciplinas":
        status = "aguardando-cross"
    elif mem_filled:
        status = "memorial"
    else:
        status = "template"

    total = None
    pend_counts = {}
    if has_quant:
        total, pend_counts = ler_total_quantitativos(quantitativos)

    return {
        "folder": folder.name,
        "kind": kind,
        "status": status,
        "memorial_size_kb": round(mem_size / 1024, 1),
        "memorial_filled": mem_filled,
        "has_extracao": extracao.exists(),
        "has_quantitativos": has_quant,
        "total_rs": total,
        "pendencias": pend_counts,
        "marker": marker,
    }


def ordem_planejada():
    """Ordem das disciplinas conforme base/pacotes/thozen-electra/mapping-disciplinas.json."""
    mapping_path = (
        Path(__file__).resolve().parent.parent
        / "base" / "pacotes" / "thozen-electra" / "mapping-disciplinas.json"
    )
    mapping = json.loads(mapping_path.read_text(encoding="utf-8"))
    ordem = []
    for d in mapping["disciplinas"]:
        ordem.append((d["order"], d["folder"], "disc"))
    for d in mapping.get("extras", []):
        ordem.append((d["order"], d["folder"], "extra"))
    return ordem


def fmt_brl(v: float | None) -> str:
    if v is None:
        return "—"
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def status_icon(status: str) -> str:
    return {
        "template": "⬜",
        "memorial": "📝",
        "aguardando-cross": "⏸️",
        "processada": "✅",
    }.get(status, "?")


def status_label(status: str) -> str:
    return {
        "template": "Template",
        "memorial": "Memorial",
        "aguardando-cross": "Aguarda outras",
        "processada": "Processada",
    }.get(status, "?")


def render_painel(itens: list[dict]) -> str:
    total_disc = len(itens)
    processadas = sum(1 for x in itens if x["status"] == "processada")
    memoriais = sum(1 for x in itens if x["status"] in ("memorial", "processada"))
    total_geral = sum(x["total_rs"] for x in itens if x["total_rs"] is not None)
    total_pend = {
        "Alta": sum(x["pendencias"].get("Alta", 0) for x in itens),
        "Média": sum(x["pendencias"].get("Média", 0) for x in itens),
        "Baixa": sum(x["pendencias"].get("Baixa", 0) for x in itens),
    }

    # BRT timestamp
    brt = timezone(timedelta(hours=-3))
    now = datetime.now(brt).strftime("%Y-%m-%d %H:%M BRT")

    lines = []
    lines.append("---")
    lines.append("tags: [anchor, electra, orcamento-executivo, painel]")
    lines.append(f"atualizado: \"{now}\"")
    lines.append("---")
    lines.append("")
    lines.append("# Estrutura do Orçamento — Thozen Electra")
    lines.append("")
    lines.append("> Painel mestre de progresso do orçamento executivo.")
    lines.append(f"> Última atualização: **{now}**")
    lines.append(">")
    lines.append("> Atualizar via: `py -3.10 -X utf8 ~/orcamentos-openclaw/scripts/atualizar_painel.py`")
    lines.append("")
    lines.append("## 🚨 Fonte de verdade — Estudo de Projeto / Ata")
    lines.append("")
    lines.append("Antes de fechar QUALQUER disciplina, rodar o checklist de revisão contra as decisões oficiais da Thozen:")
    lines.append("")
    lines.append("- **Extrato narrativo:** [[estudo-de-projeto-ata]]")
    lines.append("- **Dados structured:** `00-projeto/decisoes-estudo-projeto.json` (JSON — abrir direto)")
    lines.append("- **Checklist de revisão por disciplina:** [[REVISAO-POS-ESTUDO]]")
    lines.append("")
    lines.append("Estes arquivos prevalecem sobre qualquer inferência do IFC/DWG quando houver conflito.")
    lines.append("")
    lines.append("## Resumo geral")
    lines.append("")
    lines.append(f"- **Disciplinas totais:** {total_disc}")
    lines.append(f"- **Processadas (quantitativos prontos):** {processadas}/{total_disc} ({processadas/total_disc*100:.1f}%)")
    lines.append(f"- **Memorial preenchido:** {memoriais}/{total_disc} ({memoriais/total_disc*100:.1f}%)")
    lines.append(f"- **Custo acumulado (só processadas):** **{fmt_brl(total_geral)}**")
    lines.append(f"- **Pendências totais:** 🔴 {total_pend['Alta']} Alta · 🟡 {total_pend['Média']} Média · ⚪ {total_pend['Baixa']} Baixa")
    lines.append("")
    lines.append("## Legenda de status")
    lines.append("")
    lines.append("| Ícone | Status | Significado |")
    lines.append("|:---:|---|---|")
    lines.append("| ⬜ | Template | Só `extracao.xlsx` gerada a partir do template — memorial ainda é stub |")
    lines.append("| 📝 | Memorial | Memorial preenchido (análise das fórmulas feita) — falta gerar quantitativos.xlsx |")
    lines.append("| ⏸️ | Aguarda outras | Memorial pronto MAS depende de varredura cross-disciplinas (processar por último) |")
    lines.append("| ✅ | Processada | `quantitativos.xlsx` pronto com valores Electra + pendências sinalizadas |")
    lines.append("")
    lines.append("## Status por disciplina")
    lines.append("")
    lines.append("| # | Disciplina | Status | Memorial | Extração | Quantitativos | Total R$ | Pendências |")
    lines.append("|---:|---|:---:|:---:|:---:|:---:|---:|:---:|")

    for item in itens:
        num = item.get("order", "—")
        folder = item["folder"]
        kind_rel = "05-extras" if item.get("kind") == "extra" else "04-disciplinas"
        # Link pro memorial.md da disciplina (wikilink com alias do folder)
        wiki = f"[[{kind_rel}/{folder}/memorial|{folder}]]"
        icon = status_icon(item["status"])
        label = status_label(item["status"])
        mem_cell = f"✅ ({item['memorial_size_kb']} KB)" if item["memorial_filled"] else "⬜ stub"
        ext_cell = "✅" if item["has_extracao"] else "❌"
        quant_cell = "✅" if item["has_quantitativos"] else "⬜"
        total_cell = fmt_brl(item["total_rs"])
        pend = item.get("pendencias") or {}
        if pend:
            pend_cell = f"🔴 {pend.get('Alta',0)} · 🟡 {pend.get('Média',0)} · ⚪ {pend.get('Baixa',0)}"
        else:
            pend_cell = "—"
        lines.append(
            f"| {num} | {wiki} | {icon} {label} | {mem_cell} | {ext_cell} | {quant_cell} | {total_cell} | {pend_cell} |"
        )

    lines.append("")
    lines.append("## Próximas ações sugeridas")
    lines.append("")
    pendentes = [i for i in itens if i["status"] == "template"]
    em_memorial = [i for i in itens if i["status"] == "memorial"]
    aguardando = [i for i in itens if i["status"] == "aguardando-cross"]
    if em_memorial:
        lines.append("**Prontas pra gerar quantitativos (memorial já pronto):**")
        for i in em_memorial:
            lines.append(f"- {i['folder']}")
        lines.append("")
    if pendentes:
        lines.append("**Aguardando análise das fórmulas (memorial stub):**")
        for i in pendentes[:10]:
            lines.append(f"- {i['folder']}")
        if len(pendentes) > 10:
            lines.append(f"- ... (+{len(pendentes)-10})")
        lines.append("")
    if aguardando:
        lines.append("**⏸️ Aguardando outras disciplinas (processar por último):**")
        for i in aguardando:
            lines.append(f"- {i['folder']} — depende de varredura cross-disciplinas (ver memorial)")
        lines.append("")

    lines.append("## Workflow por disciplina")
    lines.append("")
    lines.append("Para cada disciplina, seguir 3 passos em ordem:")
    lines.append("")
    lines.append("1. **Analisar fórmulas** do `extracao.xlsx` → escrever `memorial.md` detalhado")
    lines.append("   (regras de extração, fontes, dependências, pendências)")
    lines.append("2. **Escrever script gerador** `gerar_quantitativos_{disciplina}.py` que lê `projeto.json`")
    lines.append("   e produz `quantitativos.xlsx` com valores Electra fechados + pendências sinalizadas")
    lines.append("3. **Rodar** e atualizar este painel via `atualizar_painel.py`")
    lines.append("")
    lines.append("## Referências")
    lines.append("")
    lines.append("- **Memorial consolidado (entregável final):** [[memorial-orcamento]] — `.md` + `.docx` gerado automaticamente")
    lines.append("- **Dados-fonte do projeto:** [[projeto]] · [[areas]] · [[pavimentos]] · [[apartamentos]] · [[lazer]] · [[vagas]]")
    lines.append("- **Revisão vs estudo do cliente:** [[REVISAO-POS-ESTUDO]] · [[estudo-de-projeto-ata]]")
    lines.append("- **EAP:** [[eap]] (pasta `01-eap/`)")
    lines.append("- **Composições e insumos:** [[composicoes]] · [[insumos]] (pasta `02-composicoes-insumos/`)")
    lines.append("- **Orçamento resumo:** [[orcamento-resumo]] (pasta `03-orcamento-resumo/`)")
    lines.append("- **README geral:** [[orcamento/README|README]]")
    lines.append("")
    return "\n".join(lines) + "\n"


def main() -> int:
    ordem = ordem_planejada()
    itens = []
    for order, folder, kind in ordem:
        root = EXTRAS_DIR if kind == "extra" else DISCIPLINAS_DIR
        pasta = root / folder
        if not pasta.exists():
            continue
        info = discovery(pasta, kind)
        info["order"] = order
        itens.append(info)

    # ordenar por order
    itens.sort(key=lambda x: x.get("order", 999))

    conteudo = render_painel(itens)
    PAINEL_PATH.write_text(conteudo, encoding="utf-8")

    processadas = sum(1 for i in itens if i["status"] == "processada")
    memoriais = sum(1 for i in itens if i["status"] == "memorial")
    aguardando = sum(1 for i in itens if i["status"] == "aguardando-cross")
    templates = sum(1 for i in itens if i["status"] == "template")
    print(f"✓ Painel atualizado: {PAINEL_PATH}")
    print(f"  Total disciplinas: {len(itens)}")
    print(f"  ✅ Processadas: {processadas}")
    print(f"  📝 Memoriais preenchidos: {memoriais}")
    print(f"  ⏸️ Aguardando cross-disciplinas: {aguardando}")
    print(f"  ⬜ Templates (stub): {templates}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
