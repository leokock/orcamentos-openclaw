import openpyxl
import os
from pathlib import Path

# Paths
_OPENCLAW_MEDIA = str(Path.home() / ".openclaw" / "media" / "inbound")
analitico_path = os.path.join(_OPENCLAW_MEDIA, 'e965de44-ee1c-42c2-8db9-abfdd18c188d.xlsx')
v2_path = os.path.join(_OPENCLAW_MEDIA, '80472290-d61a-4608-922c-e17327fe1b6c.xlsx')
output_dir = str(Path.home() / "orcamentos" / "output")
output_path = os.path.join(output_dir, 'CTN-ALF-SFL-Orcamento_Parametrico_Analitico_R03.xlsx')

os.makedirs(output_dir, exist_ok=True)

# Load wb1 WITHOUT data_only (preserves formulas as strings)
wb1 = openpyxl.load_workbook(analitico_path)
# Load V2 WITH data_only (gets computed values)
wb2 = openpyxl.load_workbook(v2_path, data_only=True)

# ============================================================
# V2 GROUP TOTALS from CUSTOS_MACROGRUPO
# ============================================================
v2 = {}
ws_macro = wb2['CUSTOS_MACROGRUPO']
for row in ws_macro.iter_rows(min_row=4, max_row=21, values_only=True):
    if row[0] and row[3]:
        v2[row[0]] = float(row[3])

print("V2 totais por grupo:")
for k, vv in v2.items():
    print(f"  {k}: R$ {vv:,.2f}")

# ============================================================
# HELPERS
# ============================================================
def scale_rows(ws, rows, scale_factor, col=5):
    for r in rows:
        val = ws.cell(row=r, column=col).value
        if isinstance(val, (int, float)):
            ws.cell(row=r, column=col).value = round(float(val) * scale_factor, 2)

def sum_rows(ws, rows, col=5):
    total = 0.0
    for r in rows:
        val = ws.cell(row=r, column=col).value
        if isinstance(val, (int, float)):
            total += float(val)
    return total

# ============================================================
# 1. UPDATE OBRA SHEET
# ============================================================
ws_obra = wb1['OBRA']
ws_obra['D10'] = 7648.96  # AC
ws_obra['D11'] = 54       # UR
ws_obra['B3'] = 'R03'     # Revisao

if 'EPCs' in wb1.sheetnames:
    wb1['EPCs']['B3'] = 'R03'
if 'CUSTOS_INDIRETOS' in wb1.sheetnames:
    wb1['CUSTOS_INDIRETOS']['B10'] = 'R03'

# ============================================================
# 2. OE SHEET: scale leaf rows per group
# ============================================================
ws_oe = wb1['Orçamento Executivo']

groups = {
    'G2_MovTerra':  {'rows': [7, 8, 10, 12, 13, 14],
                     'target': v2.get('Mov. Terra', 0)},
    'G3_Estrutura': {'rows': [17,18,19,20,21,23,24,26,27,28,29,30,31,33,34,35],
                     'target': v2.get('Infraestrutura', 0) + v2.get('Supraestrutura', 0)},
    'G6_Alvenaria': {'rows': [38,39,40,41,42,43,44,45,46,47],
                     'target': v2.get('Alvenaria', 0)},
    'G7_Inst':      {'rows': [50,51,52,53,55,56,58,59,60,61,62,64,66,67,68],
                     'target': v2.get('Instalações', 0)},
    'G7_Loucas':    {'rows': [63],
                     'target': v2.get('Louças e Metais', 0)},
    'G8_SistEsp':   {'rows': [71,72,73,74,76,77,78,81,83,84],
                     'target': v2.get('Sist. Especiais', 0) + v2.get('Climatização', 0)},
    'G9_Impermeab': {'rows': [87,88,89,90,91,92,93,94],
                     'target': v2.get('Impermeabilização', 0)},
    'G10G11':       {'rows': [97,98,99,100,101,103,104,105,108,110,113,114,116,118],
                     'target': v2.get('Rev. Int. Parede', 0) + v2.get('Pisos', 0)},
    'G12_Teto':     {'rows': [122,123,124],
                     'target': v2.get('Teto', 0)},
    'G13_Pintura':  {'rows': [128,129,130,131,132],
                     'target': v2.get('Pintura', 0)},
    'G14_Esq':      {'rows': [135,136,137,138,140,141,143,144],
                     'target': v2.get('Esquadrias', 0)},
    'G15_Fachada':  {'rows': [147,150,153],
                     'target': v2.get('Fachada', 0)},
    'G16_Compl':    {'rows': [156,157,158,159,160,161],
                     'target': v2.get('Complementares', 0)},
}

print("\n--- Scaling OE groups ---")
for name, g in groups.items():
    current = sum_rows(ws_oe, g['rows'])
    target = g['target']
    if current > 0 and target > 0:
        scale = target / current
        scale_rows(ws_oe, g['rows'], scale)
        print(f"  {name}: R${current:,.0f} -> R${target:,.0f}  (x{scale:.4f})")
    elif target == 0:
        print(f"  {name}: target=0, skipped")
    else:
        print(f"  {name}: current=0, skipped")

# ============================================================
# 3. SCALE CUSTOS_INDIRETOS for Gerenciamento
# ============================================================
ws_ci = wb1['CUSTOS_INDIRETOS']
ci_target = v2.get('Gerenciamento', 0)
# Get current total from data_only version
wb1_do = openpyxl.load_workbook(analitico_path, data_only=True)
ci_current = wb1_do['CUSTOS_INDIRETOS'].cell(row=72, column=6).value
print(f"\n--- Gerenciamento (CUSTOS_INDIRETOS) ---")
print(f"  Atual:  R${ci_current:,.2f}")
print(f"  Target: R${ci_target:,.2f}")
if ci_current and ci_current > 0 and ci_target > 0:
    ci_scale = ci_target / ci_current
    scaled = 0
    for r in ws_ci.iter_rows(min_row=14, max_row=71):
        cell_e = r[4]  # column E = index 4 (0-based)
        if isinstance(cell_e.value, (int, float)):
            cell_e.value = round(float(cell_e.value) * ci_scale, 4)
            scaled += 1
    print(f"  Escalonados {scaled} celulas col E (x{ci_scale:.4f})")

# ============================================================
# 4. SAVE
# ============================================================
wb1.save(output_path)
print(f"\nSalvo em: {output_path}")

# Verify leaf sums
wb_check = openpyxl.load_workbook(output_path, data_only=True)
ws_check = wb_check['Orçamento Executivo']
print("\n--- Verificacao somas folha ---")
checks = [
    ('G2  Mov Terra',    [7,8,10,12,13,14],
     v2.get('Mov. Terra',0)),
    ('G3  Estrutura',    [17,18,19,20,21,23,24,26,27,28,29,30,31,33,34,35],
     v2.get('Infraestrutura',0)+v2.get('Supraestrutura',0)),
    ('G6  Alvenaria',    [38,39,40,41,42,43,44,45,46,47],
     v2.get('Alvenaria',0)),
    ('G7  Inst+Loucas',  [50,51,52,53,55,56,58,59,60,61,62,63,64,66,67,68],
     v2.get('Instalações',0)+v2.get('Louças e Metais',0)),
    ('G8  SistEsp+Clim', [71,72,73,74,76,77,78,81,83,84],
     v2.get('Sist. Especiais',0)+v2.get('Climatização',0)),
    ('G9  Impermeab',    [87,88,89,90,91,92,93,94],
     v2.get('Impermeabilização',0)),
    ('G10+11 RevPiso',   [97,98,99,100,101,103,104,105,108,110,113,114,116,118],
     v2.get('Rev. Int. Parede',0)+v2.get('Pisos',0)),
    ('G12 Teto',         [122,123,124],
     v2.get('Teto',0)),
    ('G13 Pintura',      [128,129,130,131,132],
     v2.get('Pintura',0)),
    ('G14 Esquadrias',   [135,136,137,138,140,141,143,144],
     v2.get('Esquadrias',0)),
    ('G15 Fachada',      [147,150,153],
     v2.get('Fachada',0)),
    ('G16 Complementar', [156,157,158,159,160,161],
     v2.get('Complementares',0)),
]
for label, rows, target in checks:
    s = sum_rows(ws_check, rows)
    diff = s - target
    ok = "OK" if abs(diff) < 1 else f"DIFF {diff:+.2f}"
    print(f"  {label}: R${s:,.2f}  (target R${target:,.2f})  [{ok}]")
