#!/usr/bin/env python3
"""Gera relatório consolidado dos pacotes de uma noite (multi-projeto).

Lê todos os pacotes em base/pacotes/ que tenham state.json + executivo.xlsx
e produz um relatório markdown único comparando todos:
- KPIs lado a lado
- Distribuição de macrogrupos por projeto
- Comparação com base por segmento
- Análise arquitetônica resumida
- Pendências/lacunas
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
PACOTES = BASE / "pacotes"


def fmt_money(v) -> str:
    if v is None or v == 0:
        return "—"
    try:
        return f"R$ {float(v):,.0f}".replace(",", ".")
    except Exception:
        return str(v)


def fmt_money_short(v) -> str:
    if v is None or v == 0:
        return "—"
    try:
        v = float(v)
        if v >= 1_000_000:
            return f"R$ {v/1_000_000:.1f}M"
        elif v >= 1_000:
            return f"R$ {v/1_000:.0f}k"
        return f"R$ {v:.0f}"
    except Exception:
        return str(v)


def load_packet(slug: str) -> dict:
    pasta = PACOTES / slug
    if not (pasta / "state.json").exists():
        return None

    state = json.loads((pasta / "state.json").read_text(encoding="utf-8"))

    arq = {}
    if (pasta / "analise-arquitetura.json").exists():
        arq = json.loads((pasta / "analise-arquitetura.json").read_text(encoding="utf-8"))

    macrogrupos = []
    grand_total = 0
    rsm2 = 0
    exec_xlsx = pasta / f"executivo-{slug}.xlsx"
    if exec_xlsx.exists():
        wb = load_workbook(exec_xlsx, data_only=True)
        if "RESUMO" in wb.sheetnames:
            ws = wb["RESUMO"]
            for r in ws.iter_rows(min_row=5, max_row=24, values_only=True):
                if not r or not r[1]:
                    continue
                if r[1] == "TOTAL":
                    grand_total = r[2] or 0
                    rsm2 = r[3] or 0
                    break
                macrogrupos.append({
                    "nome": r[1], "total": r[2] or 0, "rsm2": r[3] or 0,
                    "pct": r[4] or 0, "n_itens": r[5] or 0,
                    "confianca": r[6] or "", "fonte": r[7] or "",
                    "p10p90": r[8] or "",
                })
        wb.close()

    validacao = {}
    val_md = pasta / f"validacao-{slug}.md"
    if val_md.exists():
        validacao["path"] = str(val_md)
        validacao["content"] = val_md.read_text(encoding="utf-8")

    return {
        "slug": slug,
        "ac": state.get("ac"),
        "ur": state.get("ur"),
        "padrao": state.get("padrao"),
        "grand_total": grand_total,
        "rsm2": rsm2,
        "macrogrupos": macrogrupos,
        "decisoes_arq": arq.get("decisoes_inferidas", {}),
        "categorias_arq": list(arq.get("categorias_detectadas", {}).keys()),
        "validacao": validacao,
    }


def gerar(projetos: list[str], output: Path) -> dict:
    pacotes = [load_packet(slug) for slug in projetos]
    pacotes = [p for p in pacotes if p]

    if not pacotes:
        raise RuntimeError("Nenhum pacote encontrado")

    today = datetime.now().strftime("%Y-%m-%d")
    lines = [
        f"# Relatório Consolidado — Pacotes Noturnos {today}",
        "",
        f"_Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}_",
        "",
        f"Pacotes orçados nesta sessão: **{len(pacotes)}**",
        "",
        "---",
        "",
        "## 📊 Sumário Executivo",
        "",
        "| # | Projeto | AC | UR | Padrão | Total | R$/m² | Status |",
        "|---|---|---|---|---|---|---|---|",
    ]

    grand_total_all = 0
    total_ac = 0
    total_ur = 0
    for i, p in enumerate(pacotes, 1):
        lines.append(
            f"| {i} | **{p['slug']}** | {p['ac']:,.0f} m² | {p.get('ur') or '—'} | "
            f"{p.get('padrao') or '—'} | **{fmt_money(p['grand_total'])}** | "
            f"R$ {p['rsm2']:,.0f} | ✅ |".replace(",", ".")
        )
        grand_total_all += p["grand_total"]
        total_ac += p["ac"] or 0
        total_ur += p.get("ur") or 0

    avg_rsm2 = grand_total_all / total_ac if total_ac else 0
    lines.append(
        f"| | **TOTAL CONSOLIDADO** | **{total_ac:,.0f}** | **{total_ur}** | — | "
        f"**{fmt_money(grand_total_all)}** | **R$ {avg_rsm2:,.0f}** | — |".replace(",", ".")
    )
    lines.append("")
    lines.append("")

    lines.append("## 🏗 Distribuição dos 18 Macrogrupos por Projeto")
    lines.append("")
    lines.append("Comparação lado-a-lado dos R$/m² em cada macrogrupo:")
    lines.append("")
    headers = ["Macrogrupo"] + [f"{p['slug'][:14]}" for p in pacotes]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join("---" for _ in headers) + "|")

    if pacotes[0]["macrogrupos"]:
        for i, mg in enumerate(pacotes[0]["macrogrupos"]):
            row = [mg["nome"]]
            for p in pacotes:
                if i < len(p["macrogrupos"]):
                    rsm2 = p["macrogrupos"][i]["rsm2"]
                    row.append(f"R$ {rsm2:,.0f}".replace(",", "."))
                else:
                    row.append("—")
            lines.append("| " + " | ".join(row) + " |")
    lines.append("")

    lines.append("## 🏊 Análise Arquitetônica (Bloco 0)")
    lines.append("")
    lines.append("Itens de lazer/sistemas detectados via análise multi-camada (IFC + DXF + PDF):")
    lines.append("")

    item_labels = {
        "tem_piscina": "Piscina",
        "tem_ofuro_spa": "Ofurô / SPA",
        "tem_sauna": "Sauna",
        "tem_academia": "Academia",
        "tem_quadra": "Quadra esportiva",
        "tem_salao_festas": "Salão de festas",
        "tem_gourmet": "Espaço gourmet",
        "tem_churrasqueira": "Churrasqueira",
        "tem_playground": "Playground / kids",
        "tem_coworking": "Coworking",
        "tem_pet": "Pet place",
        "tem_bicicletario": "Bicicletário",
        "tem_gerador_dedicado": "Gerador",
    }
    headers = ["Item"] + [p["slug"][:14] for p in pacotes]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("|" + "|".join("---" for _ in headers) + "|")
    for key, label in item_labels.items():
        row = [label]
        for p in pacotes:
            row.append("✓" if p["decisoes_arq"].get(key) else "—")
        lines.append("| " + " | ".join(row) + " |")
    lines.append("")

    lines.append("## 📐 Detalhe por projeto")
    lines.append("")

    for p in pacotes:
        lines.append(f"### {p['slug'].upper()}")
        lines.append("")
        lines.append(f"- **AC:** {p['ac']:,.2f} m²".replace(",", "X").replace(".", ",").replace("X", "."))
        lines.append(f"- **UR:** {p.get('ur') or '—'}")
        lines.append(f"- **Padrão:** {p.get('padrao') or '—'}")
        lines.append(f"- **Total:** {fmt_money(p['grand_total'])}")
        lines.append(f"- **R$/m²:** R$ {p['rsm2']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        lines.append("")

        if p["categorias_arq"]:
            lines.append(f"**Lazer/sistemas detectados ({len(p['categorias_arq'])}):** "
                         + ", ".join(p["categorias_arq"]))
            lines.append("")
        else:
            lines.append("**Lazer/sistemas detectados:** (nenhum — projeto compacto)")
            lines.append("")

        lines.append("**Top 5 macrogrupos por valor:**")
        lines.append("")
        sorted_mgs = sorted(p["macrogrupos"], key=lambda x: -x["total"])[:5]
        lines.append("| # | Macrogrupo | Total | R$/m² | % |")
        lines.append("|---|---|---|---|---|")
        for i, mg in enumerate(sorted_mgs, 1):
            lines.append(f"| {i} | {mg['nome']} | {fmt_money_short(mg['total'])} | "
                         f"R$ {mg['rsm2']:,.0f} | {mg['pct']*100:.1f}% |".replace(",", "."))
        lines.append("")

        if p["validacao"].get("content"):
            seg_line = ""
            for line in p["validacao"]["content"].split("\n"):
                if "segmento" in line.lower() and "rsm2" in line.lower():
                    seg_line = line
                    break
            if seg_line:
                lines.append(f"**Validação:** {seg_line.strip('- *')}")
                lines.append("")

        lines.append("**Arquivos:**")
        lines.append("")
        pasta_rel = f"base/pacotes/{p['slug']}"
        for arq in [
            f"gate-{p['slug']}.xlsx",
            f"gate-{p['slug']}-validado.xlsx",
            f"parametrico-{p['slug']}.xlsx",
            f"parametrico-{p['slug']}.docx",
            f"executivo-{p['slug']}.xlsx",
            f"executivo-{p['slug']}.docx",
            f"validacao-{p['slug']}.md",
            "analise-arquitetura.json",
            "state.json",
        ]:
            lines.append(f"- `{pasta_rel}/{arq}`")
        lines.append("")
        lines.append("---")
        lines.append("")

    lines.append("## 🎯 Próximos Passos")
    lines.append("")
    lines.append("1. **Revisar cada pacote** abrindo `executivo-{slug}.xlsx` na aba RESUMO")
    lines.append("2. **Comparar memorial Word** (`*.docx`) com as premissas que você quer entregar ao cliente")
    lines.append("3. **Validar cobertura** dos macrogrupos no relatório `validacao-{slug}.md`")
    lines.append("4. **Ajustar manualmente** o que precisar (eu re-rodo `gerar_pacote.py --slug X --continue`)")
    lines.append("5. **Copiar para Drive** quando aprovado: `~/orcamentos/parametricos/{slug}/` e `~/orcamentos/executivos/{slug}/`")
    lines.append("")

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines), encoding="utf-8")

    return {
        "output": str(output),
        "n_pacotes": len(pacotes),
        "grand_total": round(grand_total_all, 2),
        "total_ac": round(total_ac, 2),
        "total_ur": total_ur,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--projetos", nargs="+", required=True,
                    help="lista de slugs dos pacotes")
    ap.add_argument("-o", "--output", default=None)
    args = ap.parse_args()

    today = datetime.now().strftime("%Y-%m-%d")
    output = Path(args.output) if args.output else PACOTES / f"relatorio-noturno-{today}.md"
    result = gerar(args.projetos, output)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
