#!/usr/bin/env python3
"""Phase 20j — Ficha de cliente (entregavel comercial).

Cria Excel com uma aba por cliente prioritario contendo:
- Resumo do cliente (N proj, R$/m², AC medio, padrao tipico)
- Projetos com detalhes
- Assinatura de distribuicao % MG
- R$/m² por MG do cliente
- Observacoes qualitativas dos orcamentistas
- Comparacao vs mediana do padrao

Saida: base/fichas-cliente-cartesian.xlsx

Uso: python scripts/gerar_fichas_cliente.py
"""
from __future__ import annotations

import json
import statistics
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
IDX_DIR = BASE / "indices-executivo"
PAD_FILE = BASE / "padroes-classificados-consolidado.json"
AF_FILE = BASE / "analise-financeira-agregada.json"
OUT_XLSX = BASE / "fichas-cliente-cartesian.xlsx"

# Clientes prioritarios (ordem do Leo)
CLIENTES_PRIORITARIOS = [
    ("Nova Empreendimentos", ["nova-empreendimentos"]),
    ("Paludo", ["paludo"]),
    ("Mussi Empreendimentos", ["mussi"]),
    ("Pass-e", ["pass-e"]),
    ("Santa Maria", ["santa-maria"]),
    ("Amalfi", ["amalfi"]),
    ("Chiquetti & Dalvesco", ["chiquetti"]),
    ("F. Nogueira", ["f-nogueira"]),
    # Expandidos na sessao
    ("CK", ["ck-"]),
    ("CN Brava", ["cn-brava"]),
    ("Grandezza", ["grandezza"]),
    ("Inbrasul", ["inbrasul"]),
    ("Muller Empreendimentos", ["muller"]),
    ("Neuhaus", ["neuhaus"]),
]

MG_CANON = [
    ("Gerenciamento", ["gerenciamento", "ger. tec", "ger tec"]),
    ("Movimentacao de Terra", ["mov", "terra", "terraplan"]),
    ("Infraestrutura", ["infraestrutura", "fundac", "contenc", "estaca", "baldram"]),
    ("Supraestrutura", ["supraestrutura", "estrutura de concreto"]),
    ("Alvenaria", ["alvenaria", "vedac", "divisori"]),
    ("Impermeabilizacao", ["impermeab", "tratament"]),
    ("Instalacoes Hidrossanitarias", ["hidros", "hidraul", "drenagem"]),
    ("Instalacoes Eletricas", ["eletric", "telefon", "logic", "comunica", "automaca"]),
    ("Instalacoes Preventivas", ["preventiv", "pci", "spda", "glp"]),
    ("Instalacoes Gerais", ["instalac", "instalações"]),
    ("Climatizacao", ["climat", "exaust", "ar condic"]),
    ("Revestimentos Parede", ["rev. int. parede", "rev.int.parede", "rev int parede",
                              "revestimentos internos em paredes", "reboco interno",
                              "revestimentos internos de parede", "revestimentos e acabamentos internos em parede",
                              "revestimentos argamassados parede", "acabamentos em parede",
                              "acabamentos internos em parede", "revestimentos ceramicos",
                              "rev. int parede", "rev parede"]),
    ("Revestimentos Teto", ["teto", "forro"]),
    ("Pisos e Pavimentacoes", ["piso", "pavimenta", "contrapiso"]),
    ("Pintura Interna", ["pintura interna", "sistemas de pintura interna"]),
    ("Pintura Geral", ["pintura", "pinturas", "sistema de pintura"]),
    ("Esquadrias", ["esquadri", "vidro", "ferragen"]),
    ("Loucas e Metais", ["louc", "metai"]),
    ("Fachada", ["fachada", "pintura externa", "pintura de fachada"]),
    ("Cobertura", ["cobertura"]),
    ("Sistemas Especiais", ["especia", "equipament", "outros sistemas"]),
    ("Servicos Complementares", ["complementa", "imprevisto", "contingenc"]),
]

DARK, ACCENT, ORANGE, GREEN, PURPLE, RED, GRAY, TEAL = (
    "2C3E50", "2980B9", "E67E22", "27AE60", "8E44AD", "C0392B", "7F8C8D", "16A085"
)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
GREEN_FILL = PatternFill(start_color="E8F8E0", end_color="E8F8E0", fill_type="solid")
RED_FILL = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
BLUE_FILL = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")


def norm(s):
    s = str(s or "").lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def canonize(raw):
    r = norm(raw).strip().lstrip("0123456789. ")
    for canon, kws in MG_CANON:
        for kw in kws:
            if kw in r:
                return canon
    return "Outros"


def _load(p):
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def sh(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def wh(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, 1):
        sh(ws.cell(row, i, h), color)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wr(ws, row, values, nf=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if nf and i in nf:
            c.number_format = nf[i]
        elif isinstance(v, float):
            c.number_format = "#,##0.00"
        elif isinstance(v, int):
            c.number_format = "#,##0"


def load_projetos_cliente(prefixes: list) -> list:
    pad = _load(PAD_FILE)
    pad_map = {p["projeto"]: p for p in pad.get("projetos", []) if p.get("projeto")}
    projs = []
    for f in sorted(IDX_DIR.glob("*.json")):
        slug = f.stem
        sn = norm(slug)
        if not any(sn.startswith(p) for p in prefixes):
            continue
        d = _load(f)
        pi = pad_map.get(slug, {})
        ac = d.get("ac") or 0
        rsm2 = d.get("rsm2") or 0
        # Canonize MGs
        mgs_pct = {}
        mgs_val = defaultdict(float)
        total = d.get("total") or 0
        for k, v in (d.get("macrogrupos") or {}).items():
            val = v.get("valor") if isinstance(v, dict) else (v or 0)
            if val and val > 0:
                mgs_val[canonize(k)] += val
        if total:
            mgs_pct = {k: v/total*100 for k, v in mgs_val.items()}
        projs.append({
            "slug": slug,
            "padrao": pi.get("padrao", "desconhecido"),
            "ac": ac, "ur": d.get("ur") or 0,
            "total": total, "rsm2": rsm2,
            "mgs_pct": mgs_pct,
            "mgs_val": dict(mgs_val),
            "qualitative": d.get("qualitative") or {},
        })
    return projs


def build_ficha(wb, cliente_nome: str, projs: list, af: dict):
    aba_nome = cliente_nome.replace(" ", "_")[:31]
    ws = wb.create_sheet(aba_nome)
    ws.sheet_properties.tabColor = ACCENT

    # Cabecalho
    ws["A1"] = f"FICHA DE CLIENTE — {cliente_nome.upper()}"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    if not projs:
        ws["A3"] = "Sem projetos"
        return

    validos = [p for p in projs if p["ac"] >= 1000 and 500 <= (p["rsm2"] or 0) <= 10000]
    rsm2s = [p["rsm2"] for p in validos]
    acs = [p["ac"] for p in projs if p["ac"] > 0]

    # Metricas-chave
    row = 3
    ws.cell(row, 1, "METRICAS-CHAVE").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Metrica", "Valor"], [30, 20])
    row += 1
    med_rsm2 = statistics.median(rsm2s) if rsm2s else None
    med_ac = statistics.median(acs) if acs else None
    kvs = [
        ("N projetos na base", len(projs)),
        ("N projetos validos para R$/m²", len(validos)),
        ("AC total acumulado", f"{sum(acs):,.0f} m²" if acs else "N/D"),
        ("AC mediana", f"{med_ac:,.0f} m²" if med_ac else "N/D"),
        ("R$/m² mediana", f"R$ {med_rsm2:,.2f}" if med_rsm2 else "N/D"),
        ("R$/m² faixa", f"R$ {min(rsm2s):,.0f} - R$ {max(rsm2s):,.0f}" if rsm2s else "N/D"),
        ("Total acumulado (R$)", f"R$ {sum(p['total'] for p in projs):,.0f}"),
    ]
    # Padrao dominante
    from collections import Counter
    pads = Counter(p["padrao"] for p in projs)
    pad_dom = pads.most_common(1)[0][0] if pads else "?"
    kvs.append(("Padrao dominante", pad_dom))
    kvs.append(("Mix de padroes", ", ".join(f"{p}: {n}" for p, n in pads.most_common())))

    for k, v in kvs:
        wr(ws, row, [k, v])
        row += 1

    # Benchmark vs mediana do padrao
    row += 2
    ws.cell(row, 1, "POSICIONAMENTO vs MEDIANA DO PADRAO").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    med_por_pad = af.get("regressao_rsm2", {}).get("mediana_rsm2_por_padrao", {})
    med_pad = med_por_pad.get(pad_dom)
    if med_pad and med_rsm2:
        delta = (med_rsm2 / med_pad - 1) * 100
        wh(ws, row, ["Cliente R$/m²", "Mediana do padrao", "Delta %", "Interpretacao"], [18, 18, 14, 40])
        row += 1
        interp = ""
        if abs(delta) < 5:
            interp = "Alinhado com a mediana do padrao"
        elif delta > 20:
            interp = "MUITO ACIMA — investigar escopo ou margem"
        elif delta > 0:
            interp = "Acima da mediana — cliente premium"
        elif delta > -20:
            interp = "Abaixo da mediana — cliente eficiente"
        else:
            interp = "MUITO ABAIXO — escopo simples ou registro incompleto"
        wr(ws, row, [med_rsm2, med_pad, delta, interp], {1: "R$ #,##0.00", 2: "R$ #,##0.00", 3: "+0.0;-0.0"})
        cell = ws.cell(row, 3)
        if abs(delta) < 5:
            cell.fill = GREEN_FILL
        elif abs(delta) < 20:
            cell.fill = YELLOW_FILL
        else:
            cell.fill = RED_FILL
        row += 1

    # Assinatura de MG (mediana %)
    row += 2
    ws.cell(row, 1, "ASSINATURA DE DISTRIBUICAO % MG (mediana dos projetos do cliente)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    # Acumula todos os MGs
    mg_all = set()
    for p in projs:
        mg_all.update(p["mgs_pct"].keys())
    # Mediana por MG
    mg_med = {}
    for mg in mg_all:
        vals = [p["mgs_pct"].get(mg, 0) for p in projs if p["mgs_pct"].get(mg, 0) > 0]
        if vals:
            mg_med[mg] = statistics.median(vals)
    # Benchmark: mediana do padrao
    bench = af.get("distribuicao_mg_por_padrao", {}).get(pad_dom, {})

    wh(ws, row, ["Macrogrupo", "Cliente %", "Padrao %", "Delta pp", "Leitura"], [28, 12, 12, 14, 30])
    row += 1
    rows_data = []
    for mg in sorted(mg_med.keys(), key=lambda x: -mg_med[x]):
        cli_pct = mg_med[mg]
        pad_pct = (bench.get(mg, {}).get("mediana") or 0) if isinstance(bench.get(mg), dict) else 0
        delta = cli_pct - pad_pct
        leitura = ""
        if abs(delta) < 2:
            leitura = "Em linha com o padrao"
        elif delta > 5:
            leitura = "Cliente ALOCA MAIS neste MG"
        elif delta > 2:
            leitura = "Levemente acima"
        elif delta < -5:
            leitura = "Cliente ALOCA MENOS"
        else:
            leitura = "Levemente abaixo"
        rows_data.append((mg, cli_pct, pad_pct, delta, leitura))
    for mg, cp, pp, d, l in rows_data:
        wr(ws, row, [mg, cp, pp, d, l], {2: "0.00", 3: "0.00", 4: "+0.00;-0.00"})
        if abs(d) > 5:
            ws.cell(row, 4).fill = YELLOW_FILL if d > 0 else BLUE_FILL
        if abs(d) > 15:
            ws.cell(row, 4).fill = RED_FILL
        row += 1

    # Lista de projetos
    row += 2
    ws.cell(row, 1, "PROJETOS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Projeto", "Padrao", "AC (m²)", "UR", "Total (R$)", "R$/m²"], [35, 14, 12, 8, 18, 12])
    row += 1
    for p in sorted(projs, key=lambda x: -x["total"]):
        wr(ws, row, [p["slug"], p["padrao"], p.get("ac") or None,
                     p.get("ur") or None, p["total"] or None, p["rsm2"] or None],
           {3: "#,##0", 4: "#,##0", 5: "R$ #,##0", 6: "R$ #,##0"})
        row += 1

    # Alertas + Revisões + Fora da curva (destacado)
    row += 2
    ws.cell(row, 1, "PONTOS DE ATENCAO (alertas / revisoes / fora da curva)").font = Font(bold=True, size=11, color=RED, name="Arial")
    row += 1
    alertas_total = 0
    revisoes_total = 0
    fc_total = 0
    for p in projs:
        q = p.get("qualitative") or {}
        for obs in q.get("observacoes_orcamentista") or []:
            if isinstance(obs, dict):
                cat = obs.get("categoria", "")
                if cat == "alerta":
                    alertas_total += 1
                elif cat == "revisao":
                    revisoes_total += 1
        fc_total += len(q.get("fora_da_curva") or [])
    ws.cell(row, 1, f"Total: {alertas_total} alertas | {revisoes_total} revisoes | {fc_total} fora-da-curva").font = Font(italic=True, size=9, color="666666", name="Arial")
    row += 2

    if alertas_total + revisoes_total + fc_total > 0:
        wh(ws, row, ["Tipo", "Projeto", "Item / Contexto", "Observacao / Motivo"], [14, 28, 26, 65])
        row += 1
        # Alertas primeiro
        for p in projs:
            q = p.get("qualitative") or {}
            for obs in q.get("observacoes_orcamentista") or []:
                if isinstance(obs, dict) and obs.get("categoria") == "alerta":
                    wr(ws, row, ["ALERTA", p["slug"], obs.get("contexto", "")[:60], obs.get("observacao", "")[:250]])
                    ws.cell(row, 1).font = Font(size=9, bold=True, color=RED, name="Arial")
                    ws.cell(row, 1).fill = RED_FILL
                    row += 1
        # Revisoes
        for p in projs:
            q = p.get("qualitative") or {}
            for obs in q.get("observacoes_orcamentista") or []:
                if isinstance(obs, dict) and obs.get("categoria") == "revisao":
                    wr(ws, row, ["REVISAO", p["slug"], obs.get("contexto", "")[:60], obs.get("observacao", "")[:250]])
                    ws.cell(row, 1).font = Font(size=9, bold=True, color=ORANGE, name="Arial")
                    ws.cell(row, 1).fill = YELLOW_FILL
                    row += 1
        # Fora da curva
        for p in projs:
            q = p.get("qualitative") or {}
            for fc in q.get("fora_da_curva") or []:
                if isinstance(fc, dict):
                    wr(ws, row, ["FORA CURVA", p["slug"], fc.get("item", "")[:60], fc.get("motivo", "")[:250]])
                    ws.cell(row, 1).font = Font(size=9, bold=True, color=PURPLE, name="Arial")
                    row += 1
    else:
        ws.cell(row, 1, "(sem alertas/revisoes/fora-da-curva registrados — bom sinal)").font = Font(italic=True, size=9, color=GREEN, name="Arial")
        row += 1

    # Observacoes qualitativas gerais
    row += 2
    ws.cell(row, 1, "OBSERVACOES GERAIS DO ORCAMENTISTA").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Projeto", "Categoria", "Contexto", "Observacao"], [28, 14, 20, 60])
    row += 1
    for p in projs:
        q = p["qualitative"] or {}
        for obs in (q.get("observacoes_orcamentista") or [])[:10]:
            if isinstance(obs, dict):
                cat = obs.get("categoria", "")
                if cat in ("alerta", "revisao"):
                    continue  # ja mostrado acima
                wr(ws, row, [
                    p["slug"],
                    cat,
                    obs.get("contexto", "")[:40],
                    obs.get("observacao", "")[:200],
                ])
                row += 1


def build_leia_me(wb, dados_gerais):
    ws = wb.create_sheet("LEIA_ME")
    ws.sheet_properties.tabColor = GREEN
    ws["A1"] = "FICHAS DE CLIENTE — Entregavel Comercial"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")

    row = 4
    ws.cell(row, 1, "COMO USAR").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    body = (
        "Uma ficha por cliente prioritario. Cada ficha tem: (1) metricas-chave do cliente, "
        "(2) posicionamento R$/m² vs mediana do padrao dominante, "
        "(3) assinatura de distribuicao % MG comparada com o benchmark do padrao, "
        "(4) lista de projetos, (5) observacoes qualitativas. "
        "Use em reuniao comercial pra entender o perfil do cliente antes de abordar/apresentar proposta."
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 1).font = Font(size=9, name="Arial")
    row += 3

    ws.cell(row, 1, "CLIENTES INCLUIDOS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Cliente", "N projetos", "R$/m² mediana", "Padrao dominante",
                 "Alertas", "Revisoes", "Fora curva"],
       [26, 12, 16, 18, 10, 10, 12])
    row += 1
    for nome, n_proj, med_rsm2, pad, a, r, fc in dados_gerais:
        nf = {3: "R$ #,##0.00"} if isinstance(med_rsm2, (int, float)) else None
        wr(ws, row, [nome, n_proj, med_rsm2 if med_rsm2 is not None else "N/D", pad, a, r, fc], nf)
        # Destaca se tem alerta
        if a > 0:
            ws.cell(row, 5).fill = RED_FILL
            ws.cell(row, 5).font = Font(size=9, bold=True, color=RED, name="Arial")
        if fc > 0:
            ws.cell(row, 7).fill = YELLOW_FILL
        row += 1

    row += 2
    ws.cell(row, 1, "LEGENDA DELTA % (R$/m² vs mediana padrao)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    legend = [
        ("Verde", "Alinhado (|delta| < 5%)", GREEN_FILL),
        ("Amarelo", "Desvio moderado (5-20%)", YELLOW_FILL),
        ("Vermelho", "Desvio forte (>20%)", RED_FILL),
    ]
    for lab, desc, fill in legend:
        c = ws.cell(row, 1, lab)
        c.fill = fill
        c.font = Font(size=9, name="Arial")
        ws.cell(row, 2, desc).font = Font(size=9, name="Arial")
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30


def main():
    print("Loading...", flush=True)
    af = _load(AF_FILE)

    wb = Workbook()
    wb.remove(wb.active)

    dados_gerais = []

    # Monta cada ficha
    for nome, prefixes in CLIENTES_PRIORITARIOS:
        projs = load_projetos_cliente(prefixes)
        if not projs:
            print(f"  {nome}: SEM projetos, skipped", flush=True)
            continue
        build_ficha(wb, nome, projs, af)
        print(f"  {nome}: {len(projs)} projetos — ficha criada", flush=True)

        # Pra leia_me
        validos = [p for p in projs if p["ac"] >= 1000 and 500 <= (p["rsm2"] or 0) <= 10000]
        rsm2s = [p["rsm2"] for p in validos]
        from collections import Counter
        pads = Counter(p["padrao"] for p in projs)
        pad_dom = pads.most_common(1)[0][0] if pads else "?"
        med_rsm2 = statistics.median(rsm2s) if rsm2s else None
        # Counts qualitativos
        a_total = r_total = fc_total = 0
        for p in projs:
            q = p.get("qualitative") or {}
            for obs in q.get("observacoes_orcamentista") or []:
                if isinstance(obs, dict):
                    cat = obs.get("categoria", "")
                    if cat == "alerta": a_total += 1
                    elif cat == "revisao": r_total += 1
            fc_total += len(q.get("fora_da_curva") or [])
        dados_gerais.append((nome, len(projs), med_rsm2, pad_dom, a_total, r_total, fc_total))

    # LEIA_ME (criado por ultimo mas movido pra primeiro)
    build_leia_me(wb, dados_gerais)
    # Move LEIA_ME pra primeira posicao
    wb.move_sheet("LEIA_ME", offset=-len(wb.sheetnames) + 1)

    wb.save(str(OUT_XLSX))
    print(f"\nSalvo: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
