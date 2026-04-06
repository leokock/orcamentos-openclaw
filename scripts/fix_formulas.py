#!/usr/bin/env python3.11
import openpyxl
import sys

arquivo_entrada = sys.argv[1]
arquivo_saida = sys.argv[2]
colunas = [int(c) for c in sys.argv[3:]]

# Carregar com fórmulas
wb = openpyxl.load_workbook(arquivo_entrada)
ws = wb.active

# Carregar com valores
wb_val = openpyxl.load_workbook(arquivo_entrada, data_only=True)
ws_val = wb_val.active

# Substituir fórmulas por valores
for col in colunas:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value and str(cell.value).startswith('='):
            cell.value = ws_val.cell(row=row, column=col).value

wb.save(arquivo_saida)
print(f"✅ Convertido: {arquivo_saida}")
