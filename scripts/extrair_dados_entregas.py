#!/usr/bin/env python3
"""Fase 12 — Extrai dados REAIS das entregas (ground truth).

Ao contrário dos scripts anteriores que inferem cidade/UF via mapa manual,
este script lê a capa dos PDFs de entrega que tem dados reais:
- Empreendimento + cliente
- Data da entrega (revisão)
- Cidade / UF
- Data base (mês de referência do CUB)
- CUB referência (valor + mês)
- Total orçado
- R$/m²
- Distribuição % MG

Pipeline retomável.

Output:
- base/entregas-ground-truth/{slug}.json
- base/entregas-ground-truth-queue.json
"""
from __future__ import annotations

import argparse
import json
import re
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import requests

ENT = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Entregas\Orçamento_executivo")
BASE = Path.home() / "orcamentos-openclaw" / "base"
OUT_DIR = BASE / "entregas-ground-truth"
OUT_DIR.mkdir(parents=True, exist_ok=True)
QUEUE = BASE / "entregas-ground-truth-queue.json"
LOG = BASE / "entregas-ground-truth.log.jsonl"

OLLAMA_URL = "http://localhost:11434/api/generate"
MODEL = "gemma4:e4b"


PROMPT_TEMPLATE = """Você é parser de capa de orçamento executivo brasileiro.

O texto abaixo é da primeira página (capa) de um orçamento. Extraia os dados em JSON.

RETORNE APENAS O JSON (sem texto adicional, use null quando não achar):

{{
  "cliente": "string|null",
  "empreendimento": "string|null",
  "data_entrega_iso": "YYYY-MM-DD|null",
  "cidade": "string|null",
  "uf": "SC|RS|PR|SP|RJ|MG|BA|GO|DF|ES|CE|PE|null",
  "data_base_iso": "YYYY-MM|null (mes/ano de referencia do CUB)",
  "cub_valor": "float|null",
  "cub_mes": "string|null (ex: 'jul/23')",
  "total_rs": "float|null",
  "rsm2": "float|null"
}}

REGRAS:
- Se ver 'ITAJAÍ - SC' ou similar, cidade='Itajaí', uf='SC'
- Data de entrega geralmente no formato 'DD.MM.YYYY' ou 'DD/MM/YYYY'
- Data base geralmente formato 'mês/AA' (jul/23 = 2023-07)
- Total_rs é o R$ TOTAL do orçamento (pode ser próximo a "TOTAL" ou na capa)
- rsm2 é o R$/m² (às vezes mostrado como "R$/m² = ..." ou "VALOR / m²")

TEXTO DA CAPA:
{texto}"""


def norm(s):
    s = str(s or "").lower().replace(" ", "-")
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def listar_entregas():
    """Lista todas as {cliente}/{obra} pairs."""
    if not ENT.exists():
        return []
    out = []
    for cli_dir in sorted(ENT.iterdir()):
        if not cli_dir.is_dir():
            continue
        try:
            for obra_dir in sorted(cli_dir.iterdir()):
                if obra_dir.is_dir():
                    slug = f"{norm(cli_dir.name)}-{norm(obra_dir.name)}"
                    out.append({
                        "cliente": cli_dir.name,
                        "obra": obra_dir.name,
                        "slug": slug,
                        "path": str(obra_dir),
                    })
        except (OSError, PermissionError):
            continue
    return out


def encontrar_arquivo_capa(pasta: Path):
    """Prioriza: PDF apresentação/análises > PDF orçamento > XLSX apresentação > XLSX > XLS."""
    if not pasta.exists():
        return None
    try:
        arquivos = list(pasta.iterdir())
    except (OSError, PermissionError):
        return None

    # 1. PDF com keywords chave
    for kw in ("apresenta", "analise", "análise", "orçamento", "orcamento"):
        for f in arquivos:
            if f.suffix.lower() == ".pdf" and kw in f.name.lower():
                return f

    # 2. Qualquer PDF
    pdfs = [f for f in arquivos if f.suffix.lower() == ".pdf"]
    if pdfs:
        return pdfs[0]

    # 3. XLSX apresentação/analise (mais confiável que xls antigo)
    for kw in ("apresenta", "analise", "análise"):
        for f in arquivos:
            if f.suffix.lower() == ".xlsx" and kw in f.name.lower():
                return f

    # 4. Qualquer XLSX (moderno)
    xlsxs = [f for f in arquivos if f.suffix.lower() == ".xlsx"]
    if xlsxs:
        return xlsxs[0]

    # 5. Último: XLS antigo (pode falhar)
    for f in arquivos:
        if f.suffix.lower() in (".xls", ".xlsb"):
            return f

    return None


def ler_capa(path: Path, max_pages: int = 5) -> str:
    """Lê primeiras páginas do arquivo."""
    if path.suffix.lower() == ".pdf":
        try:
            import pdfplumber
            textos = []
            with pdfplumber.open(str(path)) as pdf:
                # Tenta achar página com conteúdo substancial (ignora slides em branco)
                for i, p in enumerate(pdf.pages[:max_pages]):
                    t = p.extract_text()
                    if t and len(t.strip()) > 20:
                        textos.append(f"=== Page {i+1} ===\n{t}")
                    if sum(len(x) for x in textos) > 3500:
                        break
            return "\n".join(textos)[:5000]
        except Exception as e:
            return f"[erro pdf: {e}]"

    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            textos = []
            for sn in wb.sheetnames[:3]:
                ws = wb[sn]
                textos.append(f"=== Aba: {sn} ===")
                for row in ws.iter_rows(max_row=20, values_only=True):
                    linha = [str(c) for c in row if c is not None and str(c).strip()]
                    if linha:
                        textos.append(" | ".join(linha))
            wb.close()
            return "\n".join(textos)[:5000]
        except Exception as e:
            return f"[erro xlsx: {e}]"

    if path.suffix.lower() in (".xls", ".xlsb"):
        # Fallback via pandas (usa xlrd ou outro engine)
        try:
            import pandas as pd
            textos = []
            xl = pd.ExcelFile(str(path))
            for sn in xl.sheet_names[:3]:
                try:
                    df = pd.read_excel(str(path), sheet_name=sn, nrows=20)
                    textos.append(f"=== Aba: {sn} ===")
                    textos.append(df.to_string(max_cols=10, max_colwidth=30))
                except Exception:
                    continue
            return "\n".join(textos)[:5000]
        except Exception as e:
            return f"[erro xls: {e}]"

    return ""


def call_gemma(prompt: str, timeout: int = 120) -> str:
    r = requests.post(OLLAMA_URL, json={
        "model": MODEL,
        "prompt": prompt,
        "stream": False,
        "format": "json",
        "options": {
            "temperature": 0.05,
            "num_predict": 600,
            "num_ctx": 6144,
        },
    }, timeout=timeout)
    r.raise_for_status()
    return r.json().get("response", "").strip()


def parse_json(raw: str) -> dict:
    raw = re.sub(r"^```(?:json)?\s*", "", raw.strip(), flags=re.MULTILINE)
    raw = re.sub(r"\s*```\s*$", "", raw, flags=re.MULTILINE)
    s, e = raw.find("{"), raw.rfind("}")
    if s < 0 or e <= s:
        raise ValueError(f"JSON nao achado: {raw[:200]}")
    j = raw[s:e + 1]
    try:
        return json.loads(j)
    except json.JSONDecodeError:
        j = re.sub(r",\s*}", "}", j)
        return json.loads(j)


def processar(entrega: dict) -> dict:
    t0 = time.time()
    path = Path(entrega["path"])
    arq = encontrar_arquivo_capa(path)
    if not arq:
        return {"slug": entrega["slug"], "erro": "sem arquivo", "duration_s": round(time.time() - t0, 1)}

    texto = ler_capa(arq)
    if len(texto) < 50 or texto.startswith("[erro"):
        return {"slug": entrega["slug"], "erro": f"texto vazio/erro: {texto[:100]}"}

    prompt = PROMPT_TEMPLATE.format(texto=texto)
    try:
        raw = call_gemma(prompt)
        dados = parse_json(raw)
    except Exception as e:
        return {"slug": entrega["slug"], "erro": f"gemma: {str(e)[:150]}"}

    result = {
        "slug": entrega["slug"],
        "cliente_pasta": entrega["cliente"],
        "obra_pasta": entrega["obra"],
        "arquivo_lido": arq.name,
        "fonte_tipo": arq.suffix.lower(),
        "extracao_em": datetime.now().isoformat(timespec="seconds"),
        "duration_s": round(time.time() - t0, 1),
        "dados": dados,
    }
    (OUT_DIR / f"{entrega['slug']}.json").write_text(
        json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    return result


def log(e: dict):
    e.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
    with LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(e, ensure_ascii=False) + "\n")


def init_queue():
    if QUEUE.exists():
        return json.loads(QUEUE.read_text(encoding="utf-8"))
    entregas = listar_entregas()
    q = {
        "created": datetime.now().isoformat(timespec="seconds"),
        "total_entregas": len(entregas),
        "items": {e["slug"]: {"status": "pending", "cliente": e["cliente"], "obra": e["obra"], "path": e["path"]}
                  for e in entregas},
    }
    QUEUE.write_text(json.dumps(q, indent=2, ensure_ascii=False), encoding="utf-8")
    return q


def save_queue(q):
    QUEUE.write_text(json.dumps(q, indent=2, ensure_ascii=False), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slug", default=None)
    ap.add_argument("--test", action="store_true")
    ap.add_argument("--retry-failed", action="store_true")
    args = ap.parse_args()

    if args.slug:
        # busca por slug específico
        entregas = listar_entregas()
        for e in entregas:
            if e["slug"] == args.slug:
                r = processar(e)
                print(json.dumps(r, indent=2, ensure_ascii=False))
                return
        print(f"slug nao encontrado: {args.slug}")
        return

    if args.test:
        test_slugs = ["amalfi-maiori", "paludo-volo-home", "santa-maria-unimed",
                      "fg-blue-coast", "nova-empreendimentos-evora"]
        entregas = listar_entregas()
        for e in entregas:
            if e["slug"] in test_slugs:
                r = processar(e)
                print(f"\n=== {e['slug']} ===")
                if "erro" in r:
                    print(f"  ERRO: {r['erro']}")
                else:
                    d = r["dados"]
                    print(f"  Cidade: {d.get('cidade')}/{d.get('uf')}  Data base: {d.get('data_base_iso')}  CUB: {d.get('cub_valor')}")
                    print(f"  Total: R$ {d.get('total_rs')}  R$/m²: {d.get('rsm2')}")
        return

    q = init_queue()
    pending = [s for s, v in q["items"].items()
               if v["status"] == "pending" or (args.retry_failed and v["status"] == "failed")]
    print(f"Pendentes: {len(pending)}/{len(q['items'])}", flush=True)
    t_start = time.time()

    for i, slug in enumerate(pending, start=1):
        entrega = {
            "slug": slug,
            "cliente": q["items"][slug].get("cliente"),
            "obra": q["items"][slug].get("obra"),
            "path": q["items"][slug].get("path"),
        }
        elapsed = time.time() - t_start
        eta = (elapsed / i) * (len(pending) - i) if i > 0 else 0
        print(f"[{i}/{len(pending)}] {slug} (elapsed {elapsed/60:.1f}min eta {eta/60:.1f}min)", flush=True)
        q["items"][slug]["status"] = "in_progress"
        save_queue(q)
        try:
            r = processar(entrega)
            if "erro" in r:
                q["items"][slug].update({"status": "failed", "error": r["erro"]})
                log({"slug": slug, "status": "failed", "erro": r["erro"]})
                print(f"  FAIL: {r['erro']}", flush=True)
            else:
                d = r.get("dados", {})
                q["items"][slug].update({
                    "status": "done",
                    "cidade": d.get("cidade"),
                    "uf": d.get("uf"),
                    "data_base_iso": d.get("data_base_iso"),
                    "rsm2": d.get("rsm2"),
                    "duration_s": r.get("duration_s"),
                })
                log({"slug": slug, "status": "done", **d})
                print(f"  OK {r['duration_s']}s | {d.get('cidade')}/{d.get('uf')} base:{d.get('data_base_iso')} R$/m²:{d.get('rsm2')}", flush=True)
        except Exception as e:
            err = str(e)[:200]
            q["items"][slug].update({"status": "failed", "error": err})
            log({"slug": slug, "status": "failed", "erro": err})
            print(f"  CRASH: {err}", flush=True)
        save_queue(q)

    done = sum(1 for v in q["items"].values() if v["status"] == "done")
    failed = sum(1 for v in q["items"].values() if v["status"] == "failed")
    print(f"\nDone: {done}/{len(q['items'])} | Failed: {failed}", flush=True)


if __name__ == "__main__":
    main()
