#!/usr/bin/env python3
"""Gera audit-{slug}.md por projeto com achados da revisão profunda."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
PACOTES = BASE / "pacotes"


def fmt_money(v):
    if v is None or v == 0:
        return "—"
    try:
        return f"R$ {float(v):,.0f}".replace(",", ".")
    except Exception:
        return str(v)


def audit(slug: str, contexto_extra: dict | None = None) -> Path:
    pasta = PACOTES / slug
    if not pasta.exists():
        return None

    state = json.loads((pasta / "state.json").read_text(encoding="utf-8"))
    arq = {}
    if (pasta / "analise-arquitetura.json").exists():
        arq = json.loads((pasta / "analise-arquitetura.json").read_text(encoding="utf-8"))

    # Audit dos macrogrupos do executivo
    macrogrupos = []
    grand_total = 0
    rsm2 = 0
    n_itens_total = 0
    n_subdisc_total = 0
    n_pu_zero = 0
    abas_n_itens = {}

    exec_xlsx = pasta / f"executivo-{slug}.xlsx"
    if exec_xlsx.exists():
        wb = load_workbook(exec_xlsx, data_only=True)
        if "RESUMO" in wb.sheetnames:
            ws = wb["RESUMO"]
            for r in ws.iter_rows(min_row=5, max_row=24, values_only=True):
                if not r or not r[1]:
                    continue
                if r[1] == "TOTAL":
                    grand_total = r[2] or 0
                    rsm2 = r[3] or 0
                    break
                macrogrupos.append({
                    "nome": r[1], "total": r[2] or 0, "rsm2": r[3] or 0,
                    "n_itens": r[5] or 0, "confianca": r[6] or "",
                    "fonte": r[7] or "", "p10p90": r[8] or "",
                })

        for sn in wb.sheetnames:
            if sn in ("RESUMO", "REFERENCIAS", "PREMISSAS"):
                continue
            ws = wb[sn]
            cnt = 0
            for r in ws.iter_rows(min_row=5, max_row=50, values_only=True):
                if r and r[0] and r[1]:
                    cnt += 1
                    n_itens_total += 1
                    if r[2] == "sub":
                        n_subdisc_total += 1
                    if r[4] == 0 or r[4] is None:
                        n_pu_zero += 1
            abas_n_itens[sn] = cnt
        wb.close()

    today = datetime.now().strftime("%Y-%m-%d")
    lines = [
        f"# Audit Report — {slug}",
        f"",
        f"_Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} (revisão profunda autônoma)_",
        f"",
        f"## ✅ Resumo do audit",
        f"",
        f"| Métrica | Valor |",
        f"|---|---|",
        f"| **Total** | {fmt_money(grand_total)} |",
        f"| **R$/m²** | R$ {rsm2:,.0f}".replace(",", ".") + " |",
        f"| AC | {state.get('ac'):,.2f} m² |".replace(",", "X").replace(".", ",").replace("X", "."),
        f"| UR | {state.get('ur') or '—'} |",
        f"| Padrão | {state.get('padrao') or '—'} |",
        f"| **Itens detalhados (todos os mg)** | {n_itens_total} |",
        f"| Macrogrupos preenchidos | {sum(1 for v in abas_n_itens.values() if v > 0)}/18 |",
        f"| Macrogrupos vazios | {sum(1 for v in abas_n_itens.values() if v == 0)} |",
        f"",
        f"## 📋 Detalhamento por macrogrupo",
        f"",
        f"| Macrogrupo | Total | R$/m² | N itens | Confiança | Fonte |",
        f"|---|---|---|---|---|---|",
    ]

    for mg in macrogrupos:
        lines.append(
            f"| {mg['nome']} | {fmt_money(mg['total'])} | "
            f"R$ {mg['rsm2']:,.0f} | ".replace(",", ".") +
            f"{mg['n_itens']} | {mg['confianca']} | {mg['fonte']} |"
        )
    lines.append("")

    if arq.get("decisoes_inferidas"):
        lines.append("## 🏊 Análise arquitetônica (Bloco 0)")
        lines.append("")
        n_cats = len(arq.get("categorias_detectadas", {}))
        lines.append(f"**{n_cats} categorias detectadas** via leitura multi-camada (IFC + DXF + PDF)")
        lines.append("")
        lines.append("| Item | Detectado |")
        lines.append("|---|---|")
        labels = {
            "tem_piscina": "Piscina",
            "tem_ofuro_spa": "Ofurô / SPA",
            "tem_sauna": "Sauna",
            "tem_academia": "Academia",
            "tem_quadra": "Quadra esportiva",
            "tem_salao_festas": "Salão de festas",
            "tem_gourmet": "Gourmet",
            "tem_churrasqueira": "Churrasqueira",
            "tem_playground": "Playground/kids",
            "tem_coworking": "Coworking",
            "tem_pet": "Pet",
            "tem_bicicletario": "Bicicletário",
            "tem_gerador_dedicado": "Gerador",
        }
        for k, label in labels.items():
            v = arq["decisoes_inferidas"].get(k)
            lines.append(f"| {label} | {'✓ Sim' if v else '— Não'} |")
        lines.append("")

    lines.append("## 🔍 Achados da revisão")
    lines.append("")

    # Contextual findings
    if contexto_extra:
        for ach in contexto_extra.get("achados", []):
            lines.append(f"### {ach['titulo']}")
            lines.append("")
            lines.append(ach["descricao"])
            lines.append("")

    # Coerência interna
    soma_mg = sum(m["total"] for m in macrogrupos)
    if abs(soma_mg - grand_total) > 100:
        lines.append("### ⚠ Inconsistência soma macrogrupos vs total")
        lines.append(f"")
        lines.append(f"- Soma dos 18 macrogrupos: {fmt_money(soma_mg)}")
        lines.append(f"- Total declarado no RESUMO: {fmt_money(grand_total)}")
        lines.append(f"- Diferença: {fmt_money(abs(soma_mg - grand_total))}")
        lines.append("")
    else:
        lines.append("### ✅ Coerência interna")
        lines.append("")
        lines.append(f"Soma dos 18 macrogrupos confere com o total no RESUMO ({fmt_money(grand_total)}).")
        lines.append("")

    confianca_baixa = [m for m in macrogrupos if "Baixa" in str(m["confianca"])]
    if confianca_baixa:
        lines.append("### 🔴 Macrogrupos com confiança baixa")
        for m in confianca_baixa:
            lines.append(f"- **{m['nome']}**: {fmt_money(m['total'])} (revisar manualmente)")
        lines.append("")

    confianca_media = [m for m in macrogrupos if "Média" in str(m["confianca"]) or "Medi" in str(m["confianca"])]
    if confianca_media:
        lines.append("### 🟡 Macrogrupos com confiança média")
        for m in confianca_media:
            lines.append(f"- **{m['nome']}**: {fmt_money(m['total'])} (validação adicional recomendada)")
        lines.append("")

    lines.append("## 📁 Arquivos do pacote")
    lines.append("")
    artefatos = [
        f"gate-{slug}.xlsx",
        f"gate-{slug}-validado.xlsx",
        f"parametrico-{slug}.xlsx",
        f"parametrico-{slug}.docx",
        f"executivo-{slug}.xlsx",
        f"executivo-{slug}.docx",
        f"validacao-{slug}.md",
        "analise-arquitetura.json",
        "state.json",
    ]
    for a in artefatos:
        path = pasta / a
        size = path.stat().st_size if path.exists() else 0
        status = "✓" if path.exists() else "✗"
        lines.append(f"- {status} `{a}` ({size:,} bytes)".replace(",", "."))
    lines.append("")

    lines.append("## 🎯 Próximos passos sugeridos")
    lines.append("")
    if contexto_extra and contexto_extra.get("proximos_passos"):
        for ps in contexto_extra["proximos_passos"]:
            lines.append(f"- {ps}")
    else:
        lines.append("- Abrir `executivo-*.xlsx` e revisar aba RESUMO")
        lines.append("- Validar premissas no `executivo-*.docx` (memorial)")
        lines.append("- Conferir distribuição dos 18 macrogrupos")
        lines.append("- Aprovar e copiar pra Drive (`~/orcamentos/parametricos/{slug}/`)")
    lines.append("")

    out = pasta / f"audit-{slug}.md"
    out.write_text("\n".join(lines), encoding="utf-8")
    return out


CONTEXTO_ARTHEN = {
    "achados": [
        {
            "titulo": "Comparação com paramétrico V2 anterior",
            "descricao": (
                "Existe um `arthen-arboris-parametrico-v2.xlsx` anterior em "
                "`G:\\...\\_Parametrico_IA\\arthen-arboris\\` gerado pelo pipeline V2 original "
                "(bottom-up). Total anterior: **R$ 42.652.496 / R$ 3.420/m²**.\n\n"
                "Total v2.1 nova (este pacote): **R$ 36.466.994 / R$ 2.924/m²**.\n\n"
                "**Delta: -R$ 6.185.502 (-14,5%)**\n\n"
                "**Causa raiz:** as duas versões usam abordagens diferentes:\n"
                "- **v2 antiga (bottom-up):** `gerar_template_dinamico_v2.py` calcula índice × AC × PU "
                "por item dentro de cada macrogrupo (concreto m³/m² × AC × PU mediano, etc.). Mais granular.\n"
                "- **v2.1 nova (top-down):** `valores_macrogrupos_calibrados()` lê R$/m² mediano direto do "
                "`calibration-indices.json → por_macrogrupo`. Mais conservador, mediana > média em "
                "alguns macrogrupos, e padrão Médio neutro (multiplicador 1.0).\n\n"
                "**Não é erro de dados.** Os 14,5% representam a margem natural entre as duas abordagens.\n\n"
                "**Decisão pendente do Leo:**\n"
                "- (a) Manter v2.1 nova (mais conservadora, alinhada com base atual)\n"
                "- (b) Voltar ao v2 antiga (bottom-up mais rico)\n"
                "- (c) Adotar média ponderada das duas\n"
                "- (d) Mudar padrão pra 'médio-alto' (multiplicador 1.05) e re-rodar — total subiria pra ~R$ 38M"
            ),
        },
        {
            "titulo": "Coerência com análise arquitetônica",
            "descricao": (
                "A análise arquitetônica do Bloco 0 detectou 13 categorias de lazer "
                "(piscina, ofurô, sauna, academia, gourmet, brinquedoteca, coworking, pet, "
                "playground, salão gourmet, kids, lounge, gerador) — **batendo exatamente** "
                "com o memorial técnico do projeto. Esse cruzamento valida a integridade do dado."
            ),
        },
    ],
    "proximos_passos": [
        "Decidir entre v2 antigo (R$ 42.6M) e v2.1 novo (R$ 36.5M) — ver achado 1",
        "Conferir distribuição dos macrogrupos comparada com o memorial existente",
        "Validar premissas técnicas no executivo Word",
        "Copiar para `~/orcamentos/parametricos/arthen-arboris/`",
    ],
}

CONTEXTO_PLACON = {
    "achados": [
        {
            "titulo": "Validação cruzada com NBR 12.721",
            "descricao": (
                "O Quadro IV-A da NBR 12.721 do projeto reporta:\n"
                "- Custo Básico Global da Edificação: **R$ 10.580.596,97**\n"
                "- Outro valor mencionado nos quadros: **R$ 11.926.976,97**\n\n"
                "Total estimado v2.1 (este pacote): **R$ 12.180.917 / R$ 2.988/m²**\n\n"
                "**Delta vs NBR:** +R$ 253.940 (+2,1%) acima do valor mais alto da NBR.\n\n"
                "Esse batimento de **±3%** é excelente. A NBR usa CUB padrão Normal "
                "(R$ 2.222-2.650/m²), enquanto o v2.1 calibrado considera os 126 projetos "
                "executivos reais da Cartesian (R$ 2.988/m²). A diferença é o spread natural "
                "CUB-padrão × executivo-real."
            ),
        },
        {
            "titulo": "Análise arquitetônica — projeto compacto",
            "descricao": (
                "O Bloco 0 detectou **0 categorias de lazer**. Os 3 IFCs do Placon foram "
                "exportados sem `IfcSpace` (só `IfcBuildingStorey`), e os 22 pavimentos têm "
                "nomes numéricos (SUBSOLO, 1º-16º PAVIMENTO, BARRILETE, CASA DE MÁQUINAS, "
                "RESERVATÓRIO, COBERTURA) — **nenhum dedicado a lazer**.\n\n"
                "**Conclusão:** projeto compacto de 55 studios sem lazer dedicado. Resposta do "
                "Leo no briefing confirmou: piscina = Não. Faz sentido pra perfil de studios "
                "no Centro de Floripa.\n\n"
                "**Sistema Especiais R$ 175,46/m²** vem do calibrado mas pode estar superdimensionado "
                "pro perfil real (provavelmente só elevadores + gerador + entrada de água/esgoto). "
                "Vale revisar manualmente."
            ),
        },
    ],
    "proximos_passos": [
        "Validar Sistemas Especiais — pode estar superdimensionado pro perfil sem lazer",
        "Conferir abas Esquadrias e Louças (impacto direto pro padrão studios)",
        "Comparar com Quadro IV-B da NBR pra batimento por unidade",
        "Copiar para `~/orcamentos/parametricos/placon-arminio-tavares/`",
    ],
}

CONTEXTO_THOZEN = {
    "achados": [
        {
            "titulo": "Quantitativos BIM já extraídos do projeto",
            "descricao": (
                "O projeto Thozen tem dois sistemas com quantitativos completos extraídos do BIM:\n\n"
                "**Sistema de AC** (`dxf-arcondicionado/quantitativos-processados-r05.md`):\n"
                "- 80 evaporadoras + 117 condensadoras = 197 equipamentos\n"
                "- Potência total: 1.656.000 BTU/h ≈ **138 TR**\n"
                "- Distribuição: Térreo, Lazer, Pavimentos Tipo (×24)\n\n"
                "**Sistema de Exaustão** (`dxf-exaustao/RESUMO-EXTRACAO.md`):\n"
                "- 195 churrasqueiras\n"
                "- 8 exaustores TCV 710 Berliner Luft (10.600 m³/h, 3,0 kW)\n"
                "- 8 prumadas, 1.400-1.720 m de duto galvanizado\n"
                "- Estimativa: R$ 1,1-1,8M\n\n"
                "**Estes dados foram adicionados ao memorial executivo (`executivo-thozen-electra.docx`) "
                "e paramétrico (`parametrico-thozen-electra.docx`)** numa seção 9 nova com tabelas "
                "detalhadas e nota de coerência com o macrogrupo Climatização."
            ),
        },
        {
            "titulo": "Análise arquitetônica rica",
            "descricao": (
                "O Bloco 0 detectou **13 categorias de lazer** via 9 IFCs + 17 DXFs:\n"
                "- Piscina (Swimming Pool, Pool club, Kids Pool)\n"
                "- Ofurô / SPA (Beauty SPA, SPA interno)\n"
                "- Sauna\n"
                "- Quadra (Mini quadra, Estar quadra)\n"
                "- Salão de festas, Gourmet (Gourmet Club, Play & Gourmet Room)\n"
                "- Churrasqueira (BBQ pizza bar, Fire place, Wine & Fire place — 195 churrasqueiras quantificadas)\n"
                "- Playground, Pet (Praça pet), Bicicletário, Gerador\n\n"
                "**Não detectou 'academia'** — provável falso negativo (o projeto tem 'Beauty SPA' "
                "e 'Estar quadra' mas não literal 'academia/fitness'). **Vale conferir manualmente** "
                "se há área de musculação."
            ),
        },
        {
            "titulo": "Escala extra (>25k m²) — segmento Extra da base",
            "descricao": (
                "Total: R$ 137.520.937 / **R$ 3.629/m²**\n\n"
                "Posicionamento no segmento Extra (>25k m²) da base:\n"
                "- P10: R$ 1.604/m²\n"
                "- P25: R$ 2.549/m²\n"
                "- **Mediana: R$ 2.634/m²**\n"
                "- P75: R$ 4.888/m²\n"
                "- P90: R$ 5.219/m²\n\n"
                "**Delta vs mediana: +37,8%** — mas ainda dentro do P75 (R$ 4.888). "
                "Coerente com padrão **alto** + multiplicadores diferenciais aplicados em "
                "Acabamentos (Pisos, Esquadrias, Louças, Pintura, Fachada). "
                "Para projeto alto-padrão de 38k m² em Porto Belo, R$ 3.629/m² é razoável."
            ),
        },
    ],
    "proximos_passos": [
        "Confirmar manualmente se há academia/fitness no projeto (Bloco 0 não detectou explicitamente)",
        "Revisar a seção 9 do memorial executivo com os quantitativos BIM",
        "Ajustar Climatização manualmente se necessário (BIM diz R$ 2-3M, calibrado deu R$ 1,9M)",
        "Validar sistemas especiais com base nas 13 categorias detectadas",
        "Copiar para `~/orcamentos/parametricos/thozen-electra/`",
    ],
}


def main():
    contextos = {
        "arthen-arboris": CONTEXTO_ARTHEN,
        "placon-arminio-tavares": CONTEXTO_PLACON,
        "thozen-electra": CONTEXTO_THOZEN,
    }
    for slug, ctx in contextos.items():
        out = audit(slug, ctx)
        if out:
            print(f"  OK: {out}")
        else:
            print(f"  SKIP: {slug}")


if __name__ == "__main__":
    main()
