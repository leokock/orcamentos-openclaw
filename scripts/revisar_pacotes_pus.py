#!/usr/bin/env python3
"""Revisão dos 3 pacotes contra a base de PUs cross-projeto (Fase 10).

Para cada pacote (arthen, placon, thozen), lê o executivo-{slug}.xlsx
e compara os PUs aplicados com a faixa P10-P90 da base cross-projeto
(`itens-pus-agregados.json`).

Identifica:
- Itens com PU dentro da faixa P10-P90 (✅ OK)
- Itens com PU abaixo do P10 (⚠️ subdimensionado)
- Itens com PU acima do P90 (⚠️ superdimensionado)
- Itens sem PU cross-projeto referenciado (sem dado)

Saída: base/pacotes/{slug}/revisao-pus-cross.md
"""
from __future__ import annotations

import json
import re
import unicodedata
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
PACOTES = BASE / "pacotes"
PUS_AGG = BASE / "itens-pus-agregados.json"

STOPWORDS = {"de", "da", "do", "e", "em", "para", "com", "no", "na", "dos", "das",
              "uma", "um", "a", "o", "as", "os", "kg", "m", "m2", "m3", "un"}


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def canonicalize(desc):
    if not desc:
        return ""
    s = strip_accents(desc.lower())
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b\d+([,\.]\d+)?\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def token_hash_key(canon):
    tokens = [t for t in canon.split() if t not in STOPWORDS and len(t) > 2]
    if not tokens:
        return ""
    tokens.sort(key=lambda t: -len(t))
    top = tokens[:4]
    top.sort()
    return "|".join(top)


def build_pus_lookup(pus_data):
    """Build dict key -> pu stats."""
    out = {}
    for p in pus_data.get("pus_agregados", []):
        k = p.get("key")
        if k:
            out[k] = p
    return out


def find_best_match(desc, pus_lookup, pus_list):
    """Find best matching PU cluster by hash key + fuzzy fallback."""
    canon = canonicalize(desc)
    key = token_hash_key(canon)

    if key and key in pus_lookup:
        return pus_lookup[key], "exact_hash"

    if len(desc) < 10:
        return None, None
    return None, None


def inspect_pacote(slug, pus_lookup, pus_list):
    pasta = PACOTES / slug
    xlsx = pasta / f"executivo-{slug}.xlsx"
    if not xlsx.exists():
        return None

    wb = load_workbook(xlsx, data_only=True)
    stats = {
        "slug": slug,
        "ok": 0,
        "below_p10": 0,
        "above_p90": 0,
        "no_data": 0,
        "alerts": [],
        "by_macrogrupo": {},
    }

    for sn in wb.sheetnames:
        if sn in ("RESUMO", "REFERENCIAS", "PREMISSAS", "ANALISE_ARQUITETONICA"):
            continue
        ws = wb[sn]
        mg_stats = {"ok": 0, "below_p10": 0, "above_p90": 0, "no_data": 0}

        for r in ws.iter_rows(min_row=5, max_row=50, values_only=True):
            if not r or not r[0] or not r[1]:
                continue
            desc = str(r[1]).strip()
            pu = r[4] if len(r) > 4 else None
            if not isinstance(pu, (int, float)) or pu <= 0:
                continue

            match, match_type = find_best_match(desc, pus_lookup, pus_list)
            if not match:
                mg_stats["no_data"] += 1
                stats["no_data"] += 1
                continue

            p10 = match.get("pu_p10", 0)
            p90 = match.get("pu_p90", float("inf"))
            p25 = match.get("pu_p25", 0)
            p75 = match.get("pu_p75", float("inf"))
            mediana = match.get("pu_mediana", 0)

            if p10 <= pu <= p90:
                mg_stats["ok"] += 1
                stats["ok"] += 1
            elif pu < p10:
                mg_stats["below_p10"] += 1
                stats["below_p10"] += 1
                delta = (pu - mediana) / mediana * 100 if mediana else 0
                stats["alerts"].append({
                    "macrogrupo": sn,
                    "desc": desc[:70],
                    "pu_usado": round(pu, 2),
                    "pu_mediano_base": round(mediana, 2),
                    "p10_p90": f"{p10:.2f}-{p90:.2f}",
                    "delta_pct": round(delta, 1),
                    "tipo": "subdimensionado",
                    "n_obs_base": match.get("n_observacoes"),
                })
            else:
                mg_stats["above_p90"] += 1
                stats["above_p90"] += 1
                delta = (pu - mediana) / mediana * 100 if mediana else 0
                stats["alerts"].append({
                    "macrogrupo": sn,
                    "desc": desc[:70],
                    "pu_usado": round(pu, 2),
                    "pu_mediano_base": round(mediana, 2),
                    "p10_p90": f"{p10:.2f}-{p90:.2f}",
                    "delta_pct": round(delta, 1),
                    "tipo": "superdimensionado",
                    "n_obs_base": match.get("n_observacoes"),
                })

        stats["by_macrogrupo"][sn] = mg_stats

    wb.close()
    return stats


def render_md(stats):
    slug = stats["slug"]
    total = stats["ok"] + stats["below_p10"] + stats["above_p90"] + stats["no_data"]
    matched = stats["ok"] + stats["below_p10"] + stats["above_p90"]

    lines = [
        f"# Revisão PUs cross-projeto — {slug}",
        "",
        f"_Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}_",
        "",
        "Comparação dos PUs aplicados no executivo com a base de 4.525 PUs",
        "agregados cross-projeto (`itens-pus-agregados.json` — Fase 10).",
        "",
        "## Resumo",
        "",
        f"- Total de itens com PU: {total}",
        f"- **Itens com match na base**: {matched} ({matched/total*100:.0f}%)",
        f"- **Dentro da faixa P10-P90**: {stats['ok']} ({stats['ok']/matched*100:.0f}% dos matched)" if matched else "",
        f"- ⚠️ Abaixo do P10 (subdimensionado): {stats['below_p10']}",
        f"- ⚠️ Acima do P90 (superdimensionado): {stats['above_p90']}",
        f"- Sem dado cross-projeto: {stats['no_data']}",
        "",
    ]

    alerts_below = [a for a in stats["alerts"] if a["tipo"] == "subdimensionado"]
    alerts_above = [a for a in stats["alerts"] if a["tipo"] == "superdimensionado"]

    if alerts_above:
        lines.append("## ⚠️ Itens acima do P90 (top 15)")
        lines.append("")
        alerts_above.sort(key=lambda a: -a["delta_pct"])
        lines.append("| Macrogrupo | Descrição | PU usado | Mediana | Delta | n obs |")
        lines.append("|---|---|---|---|---|---|")
        for a in alerts_above[:15]:
            lines.append(
                f"| {a['macrogrupo']} | {a['desc']} | R$ {a['pu_usado']} | "
                f"R$ {a['pu_mediano_base']} | **+{a['delta_pct']}%** | {a['n_obs_base']} |"
            )
        lines.append("")

    if alerts_below:
        lines.append("## ⚠️ Itens abaixo do P10 (top 15)")
        lines.append("")
        alerts_below.sort(key=lambda a: a["delta_pct"])
        lines.append("| Macrogrupo | Descrição | PU usado | Mediana | Delta | n obs |")
        lines.append("|---|---|---|---|---|---|")
        for a in alerts_below[:15]:
            lines.append(
                f"| {a['macrogrupo']} | {a['desc']} | R$ {a['pu_usado']} | "
                f"R$ {a['pu_mediano_base']} | **{a['delta_pct']}%** | {a['n_obs_base']} |"
            )
        lines.append("")

    lines.append("## Por macrogrupo")
    lines.append("")
    lines.append("| Macrogrupo | OK | <P10 | >P90 | sem dado |")
    lines.append("|---|---|---|---|---|")
    for mg, mg_stats in stats["by_macrogrupo"].items():
        lines.append(
            f"| {mg} | {mg_stats['ok']} | {mg_stats['below_p10']} | "
            f"{mg_stats['above_p90']} | {mg_stats['no_data']} |"
        )
    lines.append("")

    return "\n".join(lines)


def main():
    print("loading pus-agregados (Fase 10)...")
    pus_data = json.loads(PUS_AGG.read_text(encoding="utf-8"))
    pus_list = pus_data.get("pus_agregados", [])
    pus_lookup = build_pus_lookup(pus_data)
    print(f"  {len(pus_list)} PUs agregados disponíveis")

    projetos = ["arthen-arboris", "placon-arminio-tavares", "thozen-electra"]

    print("\ninspecting pacotes...")
    for slug in projetos:
        print(f"\n=== {slug} ===")
        stats = inspect_pacote(slug, pus_lookup, pus_list)
        if not stats:
            print("  no data")
            continue

        total = stats["ok"] + stats["below_p10"] + stats["above_p90"] + stats["no_data"]
        matched = stats["ok"] + stats["below_p10"] + stats["above_p90"]
        print(f"  total items com PU: {total}")
        print(f"  matched na base: {matched} ({matched/total*100:.0f}%)")
        print(f"  dentro P10-P90:  {stats['ok']}")
        print(f"  abaixo P10:      {stats['below_p10']}")
        print(f"  acima P90:       {stats['above_p90']}")
        print(f"  sem dado:        {stats['no_data']}")

        out = PACOTES / slug / f"revisao-pus-cross.md"
        out.write_text(render_md(stats), encoding="utf-8")
        print(f"  saved: {out}")


if __name__ == "__main__":
    main()
