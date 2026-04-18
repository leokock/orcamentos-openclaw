#!/usr/bin/env python3
"""Fase 4 — Excel de benchmarks estratificados (regiao + padrao + tipologia)."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
AGG_FILE = BASE / "analise-enriquecida-agregada.json"
ENR_FILE = BASE / "projetos-enriquecidos.json"
OUT_XLSX = Path.home() / "orcamentos-openclaw" / "analises-cross-projeto" / "benchmarks-estratificados" / "benchmarks-estratificados.xlsx"
OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

DARK, ACCENT, ORANGE, GREEN, PURPLE, RED, TEAL = "2C3E50", "2980B9", "E67E22", "27AE60", "8E44AD", "C0392B", "16A085"
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def sh(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def wh(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, 1):
        sh(ws.cell(row, i, h), color)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wr(ws, row, values, nf=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if nf and i in nf:
            c.number_format = nf[i]


def main():
    d = json.loads(AGG_FILE.read_text(encoding="utf-8"))
    enr = json.loads(ENR_FILE.read_text(encoding="utf-8"))
    combs = d["combinacoes"]

    wb = Workbook(); wb.remove(wb.active)

    # ── LEIA_ME ──
    ws = wb.create_sheet("LEIA_ME"); ws.sheet_properties.tabColor = GREEN
    ws["A1"] = "BENCHMARKS ESTRATIFICADOS — Regiao x Padrao x Tipologia"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')} . {d['n_projetos']} projetos . {len(combs)} combinacoes"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")
    ws["A4"] = "ABAS"
    ws["A4"].font = Font(bold=True, size=11, color=DARK, name="Arial")
    row = 5
    abas_desc = [
        ("REGIAO", "Stats por CUB-regiao (SC-Floripa, BC, Vale Itajai, etc)"),
        ("REGIAO_PADRAO", "Por (regiao, padrao) — n ≥ 3"),
        ("REGIAO_PADRAO_TIPOL", "Por (regiao, padrao, tipologia) — n ≥ 3 — GRANULARIDADE FINA"),
        ("CLIENTE", "Top 15 clientes + perfil"),
        ("MATRIZ_MG", "% MG mediana por combinacao (heatmap visual)"),
        ("INSIGHTS_CHAVE", "Achados-chave que saltam da analise"),
    ]
    for a, desc in abas_desc:
        wr(ws, row, [a, desc], {})
        row += 1
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = [22, 60, 15, 15, 15][col - 1] if col <= 5 else 15

    # ── REGIAO ──
    ws = wb.create_sheet("REGIAO"); ws.sheet_properties.tabColor = ACCENT
    ws["A1"] = "POR CUB-REGIAO"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:J1")
    wh(ws, 3, ["Regiao", "N", "R$/m² min", "R$/m² p25", "R$/m² med", "R$/m² p75", "R$/m² max",
              "AC med", "Total med", "Top clientes"],
       [22, 6, 14, 14, 14, 14, 14, 14, 16, 30])
    row = 4
    nf = {3: "R$ #,##0", 4: "R$ #,##0", 5: "R$ #,##0", 6: "R$ #,##0", 7: "R$ #,##0",
          8: "#,##0", 9: "R$ #,##0"}
    for k, info in combs.items():
        if not k.startswith("regiao::"):
            continue
        r = info.get("rsm2_stats") or {}
        a = info.get("ac_stats") or {}
        t = info.get("total_stats") or {}
        top = ", ".join(f"{c}({n})" for c, n in list(info.get("top_clientes", {}).items())[:3])
        wr(ws, row, [info["nome"], info["n_projetos"],
                     r.get("min"), r.get("p25"), r.get("mediana"), r.get("p75"), r.get("max"),
                     a.get("mediana"), t.get("mediana"), top], nf)
        row += 1
    ws.auto_filter.ref = f"A3:J{row-1}"
    ws.freeze_panes = "A4"

    # ── REGIAO_PADRAO ──
    ws = wb.create_sheet("REGIAO_PADRAO"); ws.sheet_properties.tabColor = PURPLE
    ws["A1"] = "POR (REGIAO, PADRAO)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:J1")
    wh(ws, 3, ["Regiao", "Padrao", "N", "R$/m² min", "R$/m² p25", "R$/m² med", "R$/m² p75", "R$/m² max",
              "AC med", "Top clientes"],
       [20, 14, 6, 12, 12, 14, 12, 12, 14, 30])
    row = 4
    nf = {4: "R$ #,##0", 5: "R$ #,##0", 6: "R$ #,##0", 7: "R$ #,##0", 8: "R$ #,##0",
          9: "#,##0"}
    for k, info in combs.items():
        if not k.startswith("regiao_padrao::"):
            continue
        partes = k.split("::")
        r = info.get("rsm2_stats") or {}
        a = info.get("ac_stats") or {}
        top = ", ".join(f"{c}({n})" for c, n in list(info.get("top_clientes", {}).items())[:3])
        wr(ws, row, [partes[1], partes[2], info["n_projetos"],
                     r.get("min"), r.get("p25"), r.get("mediana"), r.get("p75"), r.get("max"),
                     a.get("mediana"), top], nf)
        row += 1
    ws.auto_filter.ref = f"A3:J{row-1}"
    ws.freeze_panes = "A4"

    # ── REGIAO_PADRAO_TIPOL ──
    ws = wb.create_sheet("REGIAO_PADRAO_TIPOL"); ws.sheet_properties.tabColor = RED
    ws["A1"] = "POR (REGIAO, PADRAO, TIPOLOGIA) — GRANULARIDADE FINA"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:L1")
    ws["A2"] = "Esta aba e o que vai alimentar o SIMULADOR (Fase 10). Pra cada combinacao, o R$/m² esperado."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:L2")
    wh(ws, 4, ["Regiao", "Padrao", "Tipologia canonica", "N", "R$/m² min", "R$/m² p25", "R$/m² med",
              "R$/m² p75", "R$/m² max", "AC med", "UR med", "Top clientes"],
       [18, 14, 32, 6, 12, 12, 14, 12, 12, 12, 10, 25])
    row = 5
    nf = {5: "R$ #,##0", 6: "R$ #,##0", 7: "R$ #,##0", 8: "R$ #,##0", 9: "R$ #,##0",
          10: "#,##0", 11: "#,##0"}
    for k, info in combs.items():
        if not k.startswith("reg_pad_tip::"):
            continue
        partes = k.split("::")
        r = info.get("rsm2_stats") or {}
        a = info.get("ac_stats") or {}
        u = info.get("ur_stats") or {}
        top = ", ".join(f"{c}({n})" for c, n in list(info.get("top_clientes", {}).items())[:3])
        wr(ws, row, [partes[1], partes[2], partes[3], info["n_projetos"],
                     r.get("min"), r.get("p25"), r.get("mediana"), r.get("p75"), r.get("max"),
                     a.get("mediana"), u.get("mediana"), top], nf)
        row += 1
    ws.auto_filter.ref = f"A4:L{row-1}"
    ws.freeze_panes = "A5"

    # ── CLIENTE ──
    ws = wb.create_sheet("CLIENTE"); ws.sheet_properties.tabColor = ORANGE
    ws["A1"] = "POR CLIENTE (top 15)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")
    wh(ws, 3, ["Cliente", "N proj", "R$/m² med", "AC med", "Total med",
              "Concreto m³/m²", "Taxa aço kg/m³", "Observacao"],
       [22, 8, 14, 14, 16, 14, 14, 25])
    row = 4
    nf = {3: "R$ #,##0", 4: "#,##0", 5: "R$ #,##0", 6: "0.000", 7: "0.0"}
    for k, info in combs.items():
        if not k.startswith("cliente::"):
            continue
        cli = k.split("::")[1]
        r = info.get("rsm2_stats") or {}
        a = info.get("ac_stats") or {}
        t = info.get("total_stats") or {}
        c = info.get("concreto_stats") or {}
        ta = info.get("taxa_aco_stats") or {}
        wr(ws, row, [cli, info["n_projetos"],
                     r.get("mediana"), a.get("mediana"), t.get("mediana"),
                     c.get("mediana"), ta.get("mediana"), ""], nf)
        row += 1
    ws.auto_filter.ref = f"A3:H{row-1}"
    ws.freeze_panes = "A4"

    # ── MATRIZ_MG ──
    ws = wb.create_sheet("MATRIZ_MG"); ws.sheet_properties.tabColor = TEAL
    ws["A1"] = "MATRIZ % MG MEDIANA — por (regiao, padrao)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:V1")

    # Lista de MGs presentes em qualquer combinacao regiao+padrao
    mg_set = set()
    for k, info in combs.items():
        if k.startswith("regiao_padrao::"):
            mg_set.update(info.get("mg_pct_med", {}).keys())
    mg_order = sorted(mg_set)

    header = ["Combinacao", "N"] + mg_order
    widths = [30, 6] + [11] * len(mg_order)
    wh(ws, 3, header, widths)
    row = 4
    for k, info in combs.items():
        if not k.startswith("regiao_padrao::"):
            continue
        partes = k.split("::")
        combo_name = f"{partes[1]} | {partes[2]}"
        vals = [combo_name, info["n_projetos"]]
        for mg in mg_order:
            vals.append(info.get("mg_pct_med", {}).get(mg))
        nf = {i: "0.0" for i in range(3, 3 + len(mg_order))}
        wr(ws, row, vals, nf)
        row += 1
    ws.auto_filter.ref = f"A3:{get_column_letter(len(header))}{row-1}"
    ws.freeze_panes = "C4"

    # ── INSIGHTS_CHAVE ──
    ws = wb.create_sheet("INSIGHTS_CHAVE"); ws.sheet_properties.tabColor = RED
    ws["A1"] = "INSIGHTS-CHAVE (descobertas desta análise)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    row = 3
    insights = [
        ("1. DELTA REGIONAL MESMO PADRÃO",
         "SC-Floripa alto-padrão mediana R$ 4.864/m² vs SC-Vale-Itajaí alto R$ 3.784/m² = diferença **22%**. "
         "Mesma tipologia (residencial_vertical_alto). CUB regional é fator dominante."),
        ("2. VALE-ITAJAÍ É A REGIÃO MAIS EFICIENTE",
         "SC-Vale-Itajaí tem mediana 3.295 (n=33) — 12% abaixo de SC-Floripa (3.755, n=35). "
         "Pass-e, Mussi, Brasin, Viva4 estão nesta região."),
        ("3. CAMBORIÚ TEM MAIOR COBERTURA DE ALTO PADRÃO",
         "SC-BC tem 13 alto + 15 médio-alto (total 28). Maior concentração residencial_vertical_alto."),
        ("4. LITORAL NORTE (BOMBINHAS/ITAPEMA/PORTO BELO) É BARATO",
         "SC-Litoral-Norte mediana 3.036/m² — 20% abaixo da média SC. Paludo, Santa Maria."),
        ("5. SÃO PAULO SÓ MÉDIO-ALTO (n=5)",
         "SP-Capital tem 5 projetos, todos médio-alto, R$ 3.028/m². Ajr-spot-one, Indepy, Macom."),
        ("6. RESIDENCIAL_MISTO É MINORIA",
         "Apenas 5/131 projetos classificados como residencial_misto. Arthen-Arboris é exemplo "
         "(comercio+residencial detectado pelo memorial)."),
        ("7. NOVA E ALL SÃO FORA DA CURVA",
         "Nova (n=4) mediana R$ 6.107/m² · ALL (n=2) R$ 7.159/m² — 2× a mediana alto padrão. "
         "Paludo-vs-Nova V2 analisou: escopo expandido + especificação premium, não margem inflada."),
        ("8. SANTA MARIA É OUTLIER DE EFICIÊNCIA",
         "Santa Maria (n=3) R$ 1.604/m² — menos que alto padrão médio. "
         "Projetos grandes em Chapecó (CUB menor) + escopo enxuto."),
        ("9. CAMINHO PRÓXIMO: SIMULADOR DE PRODUTO",
         "Agora que temos (região, padrão, tipologia) pra 14 combinações com n≥3, "
         "podemos alimentar a Fase 10 (simulador de análise de produto novo)."),
    ]
    wh(ws, row, ["Insight", "Descrição"], [30, 100])
    row += 1
    for titulo, desc in insights:
        ws.cell(row, 1, titulo).font = Font(bold=True, size=10, color=DARK, name="Arial")
        ws.cell(row, 1).border = THIN
        ws.cell(row, 2, desc).font = Font(size=9, name="Arial")
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 2).border = THIN
        ws.row_dimensions[row].height = 50
        row += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 110

    wb.save(str(OUT_XLSX))
    print(f"Salvo: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
