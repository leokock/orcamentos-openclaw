#!/usr/bin/env python3
"""Phase 1 — full itemized extraction from heavy executive xlsx files.

Reads each project's heaviest xlsx (orçamento executivo), iterates every sheet
and every row, finds the header dynamically, extracts all line items with
{codigo, descricao, unidade, qtd, pu, total, aba, secao} and captures free-text
observation cells. Saves one rich JSON per project under
base/itens-detalhados/[projeto].json.

This is the foundation for all subsequent Gemma phases — they'll read from
itens-detalhados (compact view) instead of the raw xlsx.
"""
from __future__ import annotations

import json
import re
import sys
import time
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

BASE = Path.home() / "orcamentos-openclaw" / "base"
QUEUE_FILE = BASE / "gemma-queue.json"
OUT_DIR = BASE / "itens-detalhados"
LOG_FILE = BASE / "phase1-extract.log.jsonl"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Header detection — keywords to look for in header rows
HEADER_KEYWORDS = {
    "codigo": ["código", "codigo", "cod.", "cod ", "item", "item.", "ref"],
    "descricao": ["descrição", "descricao", "descricao do servico", "serviço", "servico", "discriminação", "discriminacao", "etapa", "macrogrupo", "disciplina"],
    "unidade": ["unidade", "und", "und.", "un.", "un ", "u.m.", "um"],
    "qtd": ["quantidade", "qtd", "qtde", "quant.", "quant"],
    "pu": ["preço unitário", "preco unitario", "p. unit", "p.unit", "valor unit", "unitário", "unitario", "p.u.", "pu", "custo unit", "valor / m²", "valor/m²", "valor / m2", "valor/m2"],
    "total": ["total", "valor total", "preço total", "preco total", "vlr total", "subtotal", "valor orçado", "valor orcado", "valor"],
    "bdi": ["bdi", "b.d.i.", "lucro"],
}

# How many rows to scan looking for the header
HEADER_SCAN_LIMIT = 30

# Skip sheets that are clearly not item lists
SKIP_SHEET_PATTERNS = [
    r"^cap[ai]",
    r"^gr[áa]fic",
    r"^resumo$",
    r"^[íi]ndice$",
    r"^sum[áa]rio$",
    r"^instr",
    r"^legenda",
    r"^plan\d+$",
]

OBSERVATION_MIN_LEN = 25  # treat cells with >= N chars of text as observations


def normalize(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def is_number(v: Any) -> bool:
    if v is None or v == "":
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", ".").replace(" ", ""))
        return True
    except (ValueError, TypeError):
        return False


def to_number(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def detect_header(rows: list[list[Any]]) -> dict[str, int] | None:
    """Scan first N rows looking for one that contains header keywords.

    Returns {field: column_index} mapping if a header row is found.
    Requires at least: descricao + (qtd or total) to be considered valid.
    """
    for row in rows:
        if not row:
            continue
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            n = normalize(cell)
            if not n:
                continue
            for field, keywords in HEADER_KEYWORDS.items():
                if field in col_map:
                    continue
                for kw in keywords:
                    if n == kw or n.startswith(kw + " ") or kw in n.split():
                        col_map[field] = col_idx
                        break
        if "descricao" in col_map and ("qtd" in col_map or "total" in col_map or "pu" in col_map):
            return col_map
    return None


def extract_row_item(row: list[Any], col_map: dict[str, int]) -> dict | None:
    item: dict = {}
    for field, idx in col_map.items():
        if idx < len(row):
            v = row[idx]
            if field in ("qtd", "pu", "total", "bdi"):
                item[field] = to_number(v)
            else:
                item[field] = str(v).strip() if v is not None else None

    desc = item.get("descricao") or ""
    if not desc:
        return None
    has_value = any(item.get(k) is not None and item.get(k) != 0 for k in ("qtd", "pu", "total"))
    if not has_value and not item.get("codigo"):
        return None
    return item


def extract_observations(rows: list[list[Any]], col_map_cols: set[int]) -> list[str]:
    """Capture long text cells from columns NOT in the header mapping."""
    obs = []
    for row in rows:
        if not row:
            continue
        for col_idx, cell in enumerate(row):
            if col_idx in col_map_cols:
                continue
            if isinstance(cell, str) and len(cell.strip()) >= OBSERVATION_MIN_LEN:
                txt = cell.strip()
                if not any(txt == o for o in obs):
                    obs.append(txt)
    return obs[:50]


def should_skip_sheet(sheet_name: str) -> bool:
    n = normalize(sheet_name)
    return any(re.search(p, n) for p in SKIP_SHEET_PATTERNS)


def extract_xlsx(xlsx_path: Path) -> dict:
    """Extract all line items + observations from a single xlsx."""
    result = {
        "arquivo": str(xlsx_path),
        "abas": [],
        "total_itens": 0,
        "total_observacoes": 0,
        "errors": [],
    }

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        result["errors"].append(f"open: {e}")
        return result

    for sheet_name in wb.sheetnames:
        if should_skip_sheet(sheet_name):
            continue
        try:
            ws = wb[sheet_name]
        except Exception as e:
            result["errors"].append(f"sheet {sheet_name}: {e}")
            continue

        try:
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and c != "" for c in row):
                    all_rows.append(list(row))
                if len(all_rows) > 20000:
                    break
        except Exception as e:
            result["errors"].append(f"iter {sheet_name}: {e}")
            continue

        if not all_rows:
            continue

        col_map = detect_header(all_rows[:HEADER_SCAN_LIMIT])
        if not col_map:
            continue

        header_row_idx = None
        for i, row in enumerate(all_rows[:HEADER_SCAN_LIMIT]):
            if any(normalize(c) for c in row):
                col_map_check = detect_header([row])
                if col_map_check == col_map:
                    header_row_idx = i
                    break

        if header_row_idx is None:
            continue

        secao_atual = None
        itens = []
        for row in all_rows[header_row_idx + 1:]:
            item = extract_row_item(row, col_map)
            if item:
                if secao_atual:
                    item["secao"] = secao_atual
                itens.append(item)
            else:
                non_empty = [c for c in row if c is not None and str(c).strip()]
                text_only = [c for c in non_empty if isinstance(c, str) and len(c.strip()) > 3 and not is_number(c)]
                if 1 <= len(non_empty) <= 2 and len(text_only) >= 1:
                    secao_atual = " ".join(c.strip() for c in text_only)[:120]

        col_map_cols = set(col_map.values())
        observacoes = extract_observations(all_rows[header_row_idx + 1:], col_map_cols)

        if itens or observacoes:
            result["abas"].append({
                "nome": sheet_name,
                "header_row": header_row_idx,
                "col_map": col_map,
                "n_itens": len(itens),
                "itens": itens,
                "observacoes": observacoes,
            })
            result["total_itens"] += len(itens)
            result["total_observacoes"] += len(observacoes)

    try:
        wb.close()
    except Exception:
        pass

    return result


def log_event(event: dict) -> None:
    event["ts"] = datetime.now().isoformat(timespec="seconds")
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")


def pick_xlsx_files(sources: dict[str, list[str]]) -> list[str]:
    """Return all xlsx sorted by priority: Executivo > Comentado > heaviest."""
    xlsx_list = sources.get("xlsx", [])
    if not xlsx_list:
        return []

    def score(path: str) -> tuple:
        name = Path(path).name.lower()
        prio = 0
        if "executivo" in name:
            prio = 3
        elif "comentado" in name:
            prio = 2
        elif "orçamento" in name and "apresenta" not in name:
            prio = 1
        try:
            size = Path(path).stat().st_size
        except Exception:
            size = 0
        return (-prio, -size)

    return sorted(xlsx_list, key=score)


def pick_heaviest_xlsx(sources: dict[str, list[str]]) -> str | None:
    files = pick_xlsx_files(sources)
    return files[0] if files else None


def process_one(project: dict) -> dict:
    slug = project["projeto"]
    out_path = OUT_DIR / f"{slug}.json"
    started = time.time()
    summary = {"projeto": slug, "status": "pending", "duration_s": 0}

    xlsx_files = pick_xlsx_files(project.get("sources", {}))
    if not xlsx_files:
        summary["status"] = "no_xlsx"
        summary["duration_s"] = round(time.time() - started, 2)
        log_event(summary)
        return summary

    merged = {
        "projeto": slug,
        "fontes": xlsx_files,
        "abas": [],
        "total_itens": 0,
        "total_observacoes": 0,
        "errors": [],
    }

    try:
        for xlsx in xlsx_files:
            r = extract_xlsx(Path(xlsx))
            for aba in r["abas"]:
                aba["fonte"] = Path(xlsx).name
                merged["abas"].append(aba)
            merged["total_itens"] += r["total_itens"]
            merged["total_observacoes"] += r["total_observacoes"]
            if r["errors"]:
                merged["errors"].extend([f"{Path(xlsx).name}: {e}" for e in r["errors"]])

        out_path.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
        summary["status"] = "done"
        summary["n_xlsx"] = len(xlsx_files)
        summary["abas"] = len(merged["abas"])
        summary["total_itens"] = merged["total_itens"]
        summary["total_observacoes"] = merged["total_observacoes"]
        if merged["errors"]:
            summary["errors_n"] = len(merged["errors"])
    except Exception as e:
        summary["status"] = "failed"
        summary["error"] = f"{type(e).__name__}: {e}"
        summary["traceback"] = traceback.format_exc()[-500:]

    summary["duration_s"] = round(time.time() - started, 2)
    log_event(summary)
    return summary


def main() -> None:
    only = None
    if len(sys.argv) > 1:
        only = sys.argv[1]

    queue = json.loads(QUEUE_FILE.read_text(encoding="utf-8"))
    if only:
        queue = [q for q in queue if q["projeto"] == only]
        if not queue:
            print(f"project '{only}' not found in queue", file=sys.stderr)
            sys.exit(1)

    print(f"processing {len(queue)} projects -> {OUT_DIR}")
    counts = {"done": 0, "no_xlsx": 0, "failed": 0}
    t0 = time.time()
    for i, p in enumerate(queue, 1):
        s = process_one(p)
        counts[s["status"]] = counts.get(s["status"], 0) + 1
        elapsed = time.time() - t0
        eta = (elapsed / i) * (len(queue) - i) if i > 0 else 0
        msg = f"[{i}/{len(queue)}] {s['projeto']:<40} {s['status']:<8}"
        if s["status"] == "done":
            msg += f" abas={s.get('abas',0):2} itens={s.get('total_itens',0):5} obs={s.get('total_observacoes',0):3}"
        elif s["status"] == "failed":
            msg += f" {s.get('error','')[:60]}"
        msg += f" ({s['duration_s']}s, ETA {int(eta)}s)"
        print(msg, flush=True)

    print()
    print(f"summary: {counts}")
    print(f"total time: {int(time.time() - t0)}s")
    print(f"log: {LOG_FILE}")


if __name__ == "__main__":
    main()
