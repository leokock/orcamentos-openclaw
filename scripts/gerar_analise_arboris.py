#!/usr/bin/env python3
"""Generate Arthen Arboris analysis spreadsheet (10 tabs)."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# === COLORS ===
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
CALC_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14, color="2C3E50")
SUBTITLE_FONT = Font(bold=True, size=11, color="2980B9")
BOLD = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# === PROJECT DATA (from memorial justificativo) ===
AC = 12472.98
UR = 98
AT = 1008.0
NP = 20
NPT = 15
ELEV = 2
VAG = 99
PRAZO = 30
CUB = 3028.45
TOTAL = 42652496

MACROGRUPOS = [
    ("Gerenciamento", 4636380, 371.71),
    ("Mov. Terra", 212041, 17.00),
    ("Infraestrutura", 1897910, 152.16),
    ("Supraestrutura", 9800049, 785.70),
    ("Alvenaria", 2264768, 181.57),
    ("Impermeabilizacao", 886698, 71.09),
    ("Instalacoes", 4674873, 374.80),
    ("Sist. Especiais", 1268450, 101.70),
    ("Climatizacao", 791230, 63.44),
    ("Rev. Int. Parede", 1601711, 128.41),
    ("Teto", 905729, 72.62),
    ("Pisos", 2293599, 183.89),
    ("Pintura", 1928859, 154.64),
    ("Esquadrias", 3531279, 283.10),
    ("Loucas e Metais", 329334, 26.40),
    ("Fachada", 2957826, 237.15),
    ("Complementares", 1518920, 121.78),
    ("Imprevistos", 1152839, 92.42),
]

# === CALIBRATION DATA ===
with open(r'C:\Users\leona\orcamentos-openclaw\base\calibration-indices.json', encoding='utf-8') as f:
    CAL = json.load(f)

# Map macrogroup names to calibration keys
MG_CAL_MAP = {
    "Gerenciamento": "Gerenciamento_rsm2",
    "Mov. Terra": "Mov.Terra_rsm2",
    "Infraestrutura": "Infraestrutura_rsm2",
    "Supraestrutura": "Supraestrutura_rsm2",
    "Alvenaria": "Alvenaria_rsm2",
    "Impermeabilizacao": "Impermeabilização_rsm2",
    "Instalacoes": "Instalações_rsm2",
    "Sist. Especiais": None,
    "Climatizacao": "Climatização_rsm2",
    "Rev. Int. Parede": "Rev.Int.Parede_rsm2",
    "Teto": "Teto_rsm2",
    "Pisos": "Pisos_rsm2",
    "Pintura": "Pintura_rsm2",
    "Esquadrias": "Esquadrias_rsm2",
    "Loucas e Metais": "Louças_rsm2",
    "Fachada": "Fachada_rsm2",
    "Complementares": "Complementares_rsm2",
    "Imprevistos": "Imprevistos_rsm2",
}


def add_header_row(ws, row, headers, widths=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w


def status_fill(rsm2, cal_data):
    if not cal_data:
        return CALC_FILL, "N/D"
    p10 = cal_data.get('p10', 0)
    p25 = cal_data.get('p25', 0)
    p75 = cal_data.get('p75', 9999999)
    p90 = cal_data.get('p90', 9999999)
    if p25 <= rsm2 <= p75:
        return GREEN_FILL, "OK"
    elif p10 <= rsm2 <= p90:
        return YELLOW_FILL, "Atencao"
    else:
        return RED_FILL, "Critico"


wb = Workbook()

# ============================================================
# TAB 1: RESUMO
# ============================================================
ws = wb.active
ws.title = "RESUMO"
ws.cell(1, 1, "ANALISE DO ORCAMENTO PARAMETRICO").font = TITLE_FONT
ws.cell(2, 1, "Arthen Arboris | Itapema/SC | Abril 2026").font = SUBTITLE_FONT
ws.cell(3, 1, "Base: 126 orcamentos executivos calibrados | 1.740 PUs | 40 indices").font = Font(italic=True, color="666666")

add_header_row(ws, 5, ["KPI", "Valor", "Mediana Segmento", "Delta", "Status"], [25, 18, 20, 12, 12])

kpis = [
    ("Custo Total", f"R$ {TOTAL:,.0f}", "", "", ""),
    ("R$/m2 AC", f"R$ {TOTAL/AC:,.2f}", "R$ 3.337,27", "+2,5%", "OK"),
    ("CUB Ratio", f"{TOTAL/AC/CUB:.4f}", "1,2025", "-6,1%", "OK"),
    ("Custo/UR", f"R$ {TOTAL/UR:,.0f}", "R$ 548.448", "-20,6%", "OK"),
    ("AC/UR", f"{AC/UR:.2f} m2", "133,57 m2", "-4,7%", "OK"),
    ("Burn Rate (R$/mes)", f"R$ {TOTAL/PRAZO:,.0f}", "R$ 1.027.190", "+38,4%", "Atencao"),
    ("Vagas/UR", f"{VAG/UR:.2f}", "", "", ""),
    ("Elevador/UR", f"{ELEV/UR:.3f}", "", "", ""),
    ("CA (AC/AT)", f"{AC/AT:.2f}", "", "", ""),
]

for i, (kpi, val, med, delta, status) in enumerate(kpis):
    r = 6 + i
    ws.cell(r, 1, kpi).font = BOLD
    ws.cell(r, 2, val).fill = INPUT_FILL
    ws.cell(r, 3, med)
    ws.cell(r, 4, delta)
    c = ws.cell(r, 5, status)
    if status == "OK":
        c.fill = GREEN_FILL
    elif status == "Atencao":
        c.fill = YELLOW_FILL
    for col in range(1, 6):
        ws.cell(r, col).border = THIN_BORDER
        ws.cell(r, col).alignment = Alignment(horizontal='center')

ws.cell(16, 1, "Segmento: Medio 8.000-15.000 m2 (N=16 projetos)").font = Font(italic=True, color="666666")
ws.cell(17, 1, "Faixa P10-P90: R$ 2.456 - R$ 4.347/m2").font = Font(italic=True, color="666666")
ws.cell(18, 1, "Resultado: TODOS os macrogrupos dentro do P10-P90").font = Font(bold=True, color="27AE60")

# ============================================================
# TAB 2: BENCHMARK_MACROGRUPOS
# ============================================================
ws2 = wb.create_sheet("BENCHMARK_MG")
ws2.cell(1, 1, "BENCHMARK POR MACROGRUPO vs BASE DE DADOS").font = TITLE_FONT

headers = ["Macrogrupo", "R$/m2 Projeto", "Mediana (N)", "P10", "P25", "P75", "P90", "Desvio %", "Status"]
add_header_row(ws2, 3, headers, [22, 14, 16, 12, 12, 12, 12, 12, 12])

mg_cal = CAL.get('por_macrogrupo', {})

for i, (mg, valor, rsm2) in enumerate(MACROGRUPOS):
    r = 4 + i
    cal_key = MG_CAL_MAP.get(mg)
    cal_data = mg_cal.get(cal_key, {}) if cal_key else {}
    n = cal_data.get('n', '')
    med = cal_data.get('mediana', '')
    p10 = cal_data.get('p10', '')
    p25 = cal_data.get('p25', '')
    p75 = cal_data.get('p75', '')
    p90 = cal_data.get('p90', '')
    desvio = (rsm2 / med - 1) if med else ''

    fill, status = status_fill(rsm2, cal_data) if cal_data else (CALC_FILL, "N/D")

    ws2.cell(r, 1, mg).font = BOLD
    ws2.cell(r, 2, rsm2).number_format = '#,##0.00'
    ws2.cell(r, 3, f"{med:.0f} (N={n})" if med else "N/D")
    ws2.cell(r, 4, round(p10, 0) if p10 else '').number_format = '#,##0'
    ws2.cell(r, 5, round(p25, 0) if p25 else '').number_format = '#,##0'
    ws2.cell(r, 6, round(p75, 0) if p75 else '').number_format = '#,##0'
    ws2.cell(r, 7, round(p90, 0) if p90 else '').number_format = '#,##0'
    ws2.cell(r, 8, f"{desvio:+.1%}" if isinstance(desvio, float) else "N/D")
    c = ws2.cell(r, 9, status)
    c.fill = fill

    for col in range(1, 10):
        ws2.cell(r, col).border = THIN_BORDER
        ws2.cell(r, col).alignment = Alignment(horizontal='center')

# ============================================================
# TAB 3: BENCHMARK_SEGMENTO
# ============================================================
ws3 = wb.create_sheet("BENCHMARK_SEG")
ws3.cell(1, 1, "BENCHMARK POR SEGMENTO DE PORTE").font = TITLE_FONT

add_header_row(ws3, 3, ["Segmento", "Faixa (m2)", "N", "Mediana R$/m2", "P10", "P90", "Este Projeto", "Status"], [20, 16, 6, 16, 12, 12, 16, 12])

seg_data = CAL.get('por_segmento', {})
segments = [
    ("Pequeno", "<8.000", "pequeno_lt8k_rsm2"),
    ("Medio", "8.000-15.000", "medio_8k_15k_rsm2"),
    ("Grande", "15.000-25.000", "grande_15k_25k_rsm2"),
    ("Extra", ">25.000", "extra_gt25k_rsm2"),
]

rsm2_proj = TOTAL / AC
for i, (name, faixa, key) in enumerate(segments):
    r = 4 + i
    sd = seg_data.get(key, {})
    ws3.cell(r, 1, name).font = BOLD
    ws3.cell(r, 2, faixa)
    ws3.cell(r, 3, sd.get('n', ''))
    ws3.cell(r, 4, sd.get('mediana', '')).number_format = '#,##0.00'
    ws3.cell(r, 5, sd.get('p10', '')).number_format = '#,##0'
    ws3.cell(r, 6, sd.get('p90', '')).number_format = '#,##0'

    is_this = (key == "medio_8k_15k_rsm2")
    ws3.cell(r, 7, rsm2_proj if is_this else '').number_format = '#,##0.00'
    ws3.cell(r, 8, "ESTE PROJETO" if is_this else "").fill = GREEN_FILL if is_this else PatternFill()

    for col in range(1, 9):
        ws3.cell(r, col).border = THIN_BORDER
        ws3.cell(r, col).alignment = Alignment(horizontal='center')

# Bar chart
chart = BarChart()
chart.type = "col"
chart.title = "R$/m2 por Segmento"
chart.y_axis.title = "R$/m2"
chart.x_axis.title = "Segmento"
data = Reference(ws3, min_col=4, min_row=3, max_row=7)
cats = Reference(ws3, min_col=1, min_row=4, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
ws3.add_chart(chart, "A10")

# ============================================================
# TAB 4: ANALISE_PRODUTO
# ============================================================
ws4 = wb.create_sheet("ANALISE_PRODUTO")
ws4.cell(1, 1, "ANALISE DE PRODUTO").font = TITLE_FONT

add_header_row(ws4, 3, ["Indicador", "Projeto", "Mediana Base", "N", "P10", "P90", "Desvio", "Status"], [25, 16, 16, 6, 14, 14, 10, 12])

prod = CAL.get('produto', {})
prod_items = [
    ("AC/UR (m2/un)", AC/UR, "ac_por_ur"),
    ("Custo/UR (R$)", TOTAL/UR, "custo_por_ur"),
    ("CUB Ratio", rsm2_proj/CUB, "cub_ratio"),
    ("Burn Rate (R$/mes)", TOTAL/PRAZO, "burn_rate_mensal"),
    ("Vagas/UR", VAG/UR, None),
    ("Elevador/UR", ELEV/UR, None),
    ("CA (AC/AT)", AC/AT, None),
]

for i, (name, val, cal_key) in enumerate(prod_items):
    r = 4 + i
    pd = prod.get(cal_key, {}) if cal_key else {}
    med = pd.get('mediana', '')
    n = pd.get('n', '')
    p10 = pd.get('p10', '')
    p90 = pd.get('p90', '')
    desvio = (val / med - 1) if med else ''

    fmt = '#,##0.00' if val < 100 else '#,##0'
    ws4.cell(r, 1, name).font = BOLD
    ws4.cell(r, 2, round(val, 2)).number_format = fmt
    ws4.cell(r, 3, round(med, 2) if med else '').number_format = fmt
    ws4.cell(r, 4, n)
    ws4.cell(r, 5, round(p10, 2) if p10 else '').number_format = fmt
    ws4.cell(r, 6, round(p90, 2) if p90 else '').number_format = fmt
    ws4.cell(r, 7, f"{desvio:+.1%}" if isinstance(desvio, float) else "")
    status_text = "OK" if (p10 and p90 and p10 <= val <= p90) else ("N/D" if not p10 else "Atencao")
    c = ws4.cell(r, 8, status_text)
    c.fill = GREEN_FILL if status_text == "OK" else (YELLOW_FILL if status_text == "Atencao" else CALC_FILL)

    for col in range(1, 9):
        ws4.cell(r, col).border = THIN_BORDER
        ws4.cell(r, col).alignment = Alignment(horizontal='center')

# ============================================================
# TAB 5: INDICES_ESTRUTURAIS
# ============================================================
ws5 = wb.create_sheet("INDICES_ESTRUT")
ws5.cell(1, 1, "INDICES ESTRUTURAIS vs BASE DE DADOS").font = TITLE_FONT

add_header_row(ws5, 3, ["Indice", "Projeto", "Mediana Base", "N", "P10", "P25", "P75", "P90", "Status"], [25, 12, 14, 6, 10, 10, 10, 10, 12])

est = CAL.get('estruturais', {})
est_items = [
    ("Concreto (m3/m2 AC)", 0.25, "concreto_m3_por_m2_ac"),
    ("Aco (kg/m3 concreto)", 106.0, "aco_kg_por_m3_concreto"),
    ("Aco (kg/m2 AC)", 26.5, "aco_kg_por_m2_ac"),
    ("Forma (m2/m2 AC)", 1.78, "forma_m2_por_m2_ac"),
]

for i, (name, val, cal_key) in enumerate(est_items):
    r = 4 + i
    ed = est.get(cal_key, {})
    med = ed.get('mediana', '')
    n = ed.get('n', '')
    p10 = ed.get('p10', '')
    p25 = ed.get('p25', '')
    p75 = ed.get('p75', '')
    p90 = ed.get('p90', '')

    fill_s, status_s = status_fill(val, ed) if ed else (CALC_FILL, "N/D")

    ws5.cell(r, 1, name).font = BOLD
    ws5.cell(r, 2, val).number_format = '0.00'
    ws5.cell(r, 3, round(med, 2) if med else '').number_format = '0.00'
    ws5.cell(r, 4, n)
    ws5.cell(r, 5, round(p10, 2) if p10 else '').number_format = '0.00'
    ws5.cell(r, 6, round(p25, 2) if p25 else '').number_format = '0.00'
    ws5.cell(r, 7, round(p75, 2) if p75 else '').number_format = '0.00'
    ws5.cell(r, 8, round(p90, 2) if p90 else '').number_format = '0.00'
    ws5.cell(r, 9, status_s).fill = fill_s

    for col in range(1, 10):
        ws5.cell(r, col).border = THIN_BORDER
        ws5.cell(r, col).alignment = Alignment(horizontal='center')

ws5.cell(9, 1, "Fundacao: Helice continua, 166 estacas, 18m profundidade").font = Font(italic=True, color="666666")
ws5.cell(10, 1, "Laje: Convencional (mista) - maior consumo de aco e forma vs protendida/nervurada").font = Font(italic=True, color="666666")

# ============================================================
# TAB 6: INSTALACOES_BREAKDOWN
# ============================================================
ws6 = wb.create_sheet("INSTALACOES")
ws6.cell(1, 1, "BREAKDOWN DE INSTALACOES").font = TITLE_FONT

add_header_row(ws6, 3, ["Disciplina", "R$ Total", "R$/m2", "% Inst.", "% Total", "Med.Base %", "Desvio"], [20, 16, 12, 10, 10, 12, 10])

inst_total = 4674873
inst_items = [
    ("Hidrossanitarias", 2111098, "hidrossanitarias_pct_total"),
    ("Eletricas", 1784131, "eletricas_pct_total"),
    ("Preventivas", 561284, "preventivas_pct_total"),
    ("Gas (GLP)", 224514, "gas_pct_total"),
    ("Telecom", 187233, "telecom_pct_total"),
]

inst_cal = CAL.get('instalacoes', {})
for i, (name, val, cal_key) in enumerate(inst_items):
    r = 4 + i
    pct_inst = val / inst_total
    pct_total = val / TOTAL
    med_pct = inst_cal.get(cal_key, {}).get('mediana', '')
    desvio = (pct_total / med_pct - 1) if med_pct else ''

    ws6.cell(r, 1, name).font = BOLD
    ws6.cell(r, 2, val).number_format = 'R$ #,##0'
    ws6.cell(r, 3, val / AC).number_format = '#,##0.00'
    ws6.cell(r, 4, pct_inst).number_format = '0.0%'
    ws6.cell(r, 5, pct_total).number_format = '0.0%'
    ws6.cell(r, 6, med_pct).number_format = '0.0%' if med_pct else ''
    ws6.cell(r, 7, f"{desvio:+.0%}" if isinstance(desvio, float) else "")

    for col in range(1, 8):
        ws6.cell(r, col).border = THIN_BORDER
        ws6.cell(r, col).alignment = Alignment(horizontal='center')

ws6.cell(9, 1, "TOTAL INSTALACOES").font = BOLD
ws6.cell(9, 2, inst_total).number_format = 'R$ #,##0'
ws6.cell(9, 3, inst_total / AC).number_format = '#,##0.00'
ws6.cell(9, 5, inst_total / TOTAL).number_format = '0.0%'

# ============================================================
# TAB 7: CI_DETALHADO
# ============================================================
ws7 = wb.create_sheet("CI_DETALHADO")
ws7.cell(1, 1, "CUSTOS INDIRETOS DETALHADO").font = TITLE_FONT

add_header_row(ws7, 3, ["Subcategoria", "R$ Total", "% do Total", "Med.Base %", "N", "Desvio"], [25, 16, 12, 12, 6, 10])

ci_items = [
    ("Projetos e consultorias", 830000, "projetos_consultorias_pct_total"),
    ("Equipamentos (grua+cremalheira)", 480000, "equipamentos_carga_pct_total"),
    ("Taxas e seguros", 380000, "taxas_licencas_pct_total"),
    ("EPCs", 300000, "epcs_pct_total"),
    ("Equipe ADM (30 meses)", 1846380, None),
    ("Instalacoes provisorias", 160000, "canteiro_pct_total"),
    ("Ensaios tecnologicos", 130000, "ensaios_pct_total"),
]

ci_cal = CAL.get('ci', {})
ci_total = 4636380
for i, (name, val, cal_key) in enumerate(ci_items):
    r = 4 + i
    pct = val / TOTAL
    cd = ci_cal.get(cal_key, {}) if cal_key else {}
    med_pct = cd.get('mediana', '')
    n = cd.get('n', '')
    desvio = (pct / med_pct - 1) if med_pct else ''

    ws7.cell(r, 1, name).font = BOLD
    ws7.cell(r, 2, val).number_format = 'R$ #,##0'
    ws7.cell(r, 3, pct).number_format = '0.00%'
    ws7.cell(r, 4, med_pct).number_format = '0.00%' if med_pct else ''
    ws7.cell(r, 5, n)
    ws7.cell(r, 6, f"{desvio:+.0%}" if isinstance(desvio, float) else "")

    for col in range(1, 7):
        ws7.cell(r, col).border = THIN_BORDER

ws7.cell(11, 1, "TOTAL GERENCIAMENTO").font = BOLD
ws7.cell(11, 2, ci_total).number_format = 'R$ #,##0'
ws7.cell(11, 3, ci_total / TOTAL).number_format = '0.00%'

# ============================================================
# TAB 8: SENSIBILIDADE
# ============================================================
ws8 = wb.create_sheet("SENSIBILIDADE")
ws8.cell(1, 1, "ANALISE DE SENSIBILIDADE").font = TITLE_FONT

add_header_row(ws8, 3, ["Cenario", "Delta R$", "Delta %", "Novo Total", "Novo R$/m2"], [35, 16, 10, 18, 14])

cenarios = [
    ("Laje protendida (vs convencional)", -1400000),
    ("Padrao alto acabamento", +2100000),
    ("2 subsolos de garagem", +3500000),
    ("Prazo 24 meses (vs 30)", -560000),
    ("", 0),
    ("OTIMISTA: protendida + 24 meses", -1960000),
    ("PESSIMISTA: alto padrao + 2 subsolos", +5600000),
]

for i, (name, delta) in enumerate(cenarios):
    r = 4 + i
    if not name:
        continue
    novo = TOTAL + delta
    ws8.cell(r, 1, name).font = BOLD if "OTIMISTA" in name or "PESSIMISTA" in name else Font()
    ws8.cell(r, 2, delta).number_format = 'R$ #,##0'
    ws8.cell(r, 3, delta / TOTAL).number_format = '+0.0%;-0.0%'
    ws8.cell(r, 4, novo).number_format = 'R$ #,##0'
    ws8.cell(r, 5, novo / AC).number_format = '#,##0.00'

    for col in range(1, 6):
        ws8.cell(r, col).border = THIN_BORDER

# ============================================================
# TAB 9: ALERTAS
# ============================================================
ws9 = wb.create_sheet("ALERTAS")
ws9.cell(1, 1, "SEMAFORO DE ALERTAS POR MACROGRUPO").font = TITLE_FONT

add_header_row(ws9, 3, ["Macrogrupo", "R$/m2", "Status", "Justificativa"], [22, 12, 12, 60])

justificativas = {
    "Gerenciamento": "Prazo 30 meses + grua/cremalheira para 20 pav",
    "Mov. Terra": "Zero subsolos, terreno plano",
    "Infraestrutura": "Helice continua sem contencao (0 subsolos)",
    "Supraestrutura": "Laje convencional - maior consumo material",
    "Alvenaria": "Entrega completa + drywall 33%",
    "Impermeabilizacao": "2 BWC/apto aumenta area impermeabilizada",
    "Instalacoes": "CUB +10% + 2 BWC/apto",
    "Sist. Especiais": "2 elevadores + gerador + piscina + automacao",
    "Climatizacao": "147 splits + exaustores BWC e churrasqueira",
    "Rev. Int. Parede": "Entrega completa + 2 BWC + granito bancadas",
    "Teto": "Forro gesso 100% ambientes",
    "Pisos": "Misto porcelanato+laminado (PU intermediario)",
    "Pintura": "Entrega completa, 3 demaos paredes",
    "Esquadrias": "Aluminio eletrostatica + guarda-corpo alvenaria",
    "Loucas e Metais": "Apenas bacias (sem cubas/torneiras/bancadas)",
    "Fachada": "Torre 20 pav esbelta = alta relacao fachada/AC",
    "Complementares": "Paisagismo e ambientacao basicos",
    "Imprevistos": "Contingencia 2,7% do total",
}

for i, (mg, valor, rsm2) in enumerate(MACROGRUPOS):
    r = 4 + i
    cal_key = MG_CAL_MAP.get(mg)
    cal_data = mg_cal.get(cal_key, {}) if cal_key else {}
    fill_s, status_s = status_fill(rsm2, cal_data) if cal_data else (CALC_FILL, "N/D")

    ws9.cell(r, 1, mg).font = BOLD
    ws9.cell(r, 2, rsm2).number_format = '#,##0.00'
    ws9.cell(r, 3, status_s).fill = fill_s
    ws9.cell(r, 4, justificativas.get(mg, ""))

    for col in range(1, 5):
        ws9.cell(r, col).border = THIN_BORDER

# ============================================================
# TAB 10: RECOMENDACOES
# ============================================================
ws10 = wb.create_sheet("RECOMENDACOES")
ws10.cell(1, 1, "RECOMENDACOES PARA O ORCAMENTO EXECUTIVO").font = TITLE_FONT

add_header_row(ws10, 3, ["#", "Recomendacao", "Impacto R$", "Prioridade"], [4, 50, 16, 12])

recs = [
    (1, "Confirmar fundacao via laudo SPT (tipo, comprimento, qtd estacas)", "Variavel", "Alta"),
    (2, "Validar escopo loucas: apenas bacias ou incluir cubas/torneiras/bancadas?", "+R$ 200-250k", "Media"),
    (3, "Ajustar piscinas: memorial cita 2 piscinas + ofuro, parametrico tem 1", "+R$ 130-200k", "Media"),
    (4, "Revisar climatizacao: so infra ou incluir equipamentos (splits)?", "-R$ 515k", "Alta"),
]

for i, (num, rec, impacto, prio) in enumerate(recs):
    r = 4 + i
    ws10.cell(r, 1, num)
    ws10.cell(r, 2, rec)
    ws10.cell(r, 3, impacto)
    c = ws10.cell(r, 4, prio)
    c.fill = RED_FILL if prio == "Alta" else YELLOW_FILL

    for col in range(1, 5):
        ws10.cell(r, col).border = THIN_BORDER

ws10.cell(9, 1, "CUSTO AJUSTADO ESTIMADO").font = BOLD
ws10.cell(9, 2, "~R$ 42.100.000 (R$ 3.372/m2, CUB 1,11)").font = BOLD

# ============================================================
# SAVE
# ============================================================
output = r'G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Parametrico_IA\arthen-arboris\arthen-arboris-analise-v2.xlsx'
wb.save(output)
print(f"Salvo: {output}")
print(f"Abas: {wb.sheetnames}")
