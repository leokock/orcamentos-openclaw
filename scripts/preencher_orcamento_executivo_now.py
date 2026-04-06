#!/usr/bin/env python3.11
"""
Preenche orçamento executivo NOW Residence com valores do paramétrico.
Distribui os macrogrupos na estrutura EAP do template CTN.
"""

import openpyxl
from openpyxl import load_workbook
import sys

def buscar_linha(ws, termo, inicio=13, fim=300):
    """Busca linha por termo na descrição."""
    for i in range(inicio, fim):
        desc = ws.cell(i, 9).value
        if desc and termo.lower() in str(desc).lower():
            return i
    return None

def preencher_servico(ws, linha, un, qtd, pu, obs=''):
    """Preenche um serviço (unidade, quantidade, preço unitário)."""
    if linha is None:
        print(f'  ⚠️  Linha não encontrada: {obs}')
        return
    
    ws.cell(linha, 10).value = un
    ws.cell(linha, 11).value = round(qtd, 2)
    ws.cell(linha, 12).value = round(pu, 2)
    ws.cell(linha, 13).value = f'=K{linha}*L{linha}'
    
    desc = ws.cell(linha, 9).value or ''
    print(f'  L{linha:3}: {desc[:40]:40} | {un:4} | {qtd:>10.2f} | R$ {pu:>12,.2f}')

# Carregar valores do paramétrico
print('📊 Carregando valores do orçamento paramétrico...')
wb_param = load_workbook('projetos/downloads/NOW-Residence-Orcamento-Parametrico.xlsx', data_only=True)
ws_custos = wb_param['CUSTOS_MACROGRUPO']
ws_dados = wb_param['DADOS_PROJETO']
ws_estrutural = wb_param['ESTRUTURAL']

# Extrair valores
V = {}
for i in range(2, 20):
    macro = ws_custos.cell(i, 2).value
    valor = ws_custos.cell(i, 7).value
    if macro and isinstance(valor, (int, float)):
        V[macro] = valor

# Dados do projeto
area = ws_dados['B6'].value
n_unidades = ws_dados['B7'].value + ws_dados['B8'].value
n_pavimentos = ws_dados['B9'].value

# Índices estruturais
vol_concreto_m2 = ws_estrutural['F42'].value or 0.22
taxa_aco_kg_m3 = ws_estrutural['F43'].value or 79.5

print(f'Área: {area:,.2f} m² | Unidades: {n_unidades} | Pavimentos: {n_pavimentos}')
print(f'Concreto: {vol_concreto_m2} m³/m² | Aço: {taxa_aco_kg_m3} kg/m³\n')

# Carregar template
print('📋 Carregando template...')
wb = load_workbook('output/CTN-NOW-ORCAMENTO-EXECUTIVO-R00.xlsx')
ws = wb['ORÇAMENTO']

print('='*90)
print('PREENCHENDO ORÇAMENTO EXECUTIVO NOW RESIDENCE')
print('='*90)

# Helpers locais
def p(termo, un, qtd, pu, obs='', inicio=13, fim=300):
    """Atalho para preencher serviço."""
    linha = buscar_linha(ws, termo, inicio, fim)
    preencher_servico(ws, linha, un, qtd, pu, obs)

# =============================================================================
# 1. GERENCIAMENTO
# =============================================================================
print(f'\n1. GERENCIAMENTO (R$ {V["Gerenciamento"]:,.2f})')

G = V['Gerenciamento']
p('Projeto Arquitetônico', 'vb', 1, G * 0.075, 'Projetos')
p('Projeto Estrutural', 'vb', 1, G * 0.0625, 'Projetos')
p('Projeto Elétrico', 'vb', 1, G * 0.0375, 'Projetos')
p('Projeto Hidrossanitário', 'vb', 1, G * 0.0375, 'Projetos')
p('Projeto Preventivo', 'vb', 1, G * 0.02, 'Projetos')
p('Projeto Climatização', 'vb', 1, G * 0.01, 'Projetos')
p('Projeto Fundações', 'vb', 1, G * 0.0075, 'Projetos')
p('Concreto (CP)', 'vb', 1, G * 0.075, 'Ensaios')
p('Solo/Fundações', 'vb', 1, G * 0.045, 'Ensaios')
p('Engenheiro Civil', 'mês', 48, G * 0.105 / 48, 'Admin')
p('Mestre Geral', 'mês', 48, G * 0.070 / 48, 'Admin')
p('Locação de Conteiner', 'mês', 48, G * 0.0525 / 48, 'Admin')
p('Instalações de canteiro', 'vb', 1, G * 0.070, 'Admin')
p('Equipamento de proteção individual', 'vb', 1, G * 0.035, 'Admin')
p('Locação de elevador cremalheira', 'mês', 30, G * 0.06 / 30, 'Equipamentos')
p('Bandeja de protecao primaria', 'm', 290, G * 0.03 / 290, 'Segurança')
p('Tela fachadeira', 'm²', area * 0.4, G * 0.03 / (area * 0.4), 'Segurança')
p('Alvará', 'vb', 1, G * 0.02, 'Taxas')
p('ARTs/RRTs', 'vb', 1, G * 0.01, 'Taxas')

# =============================================================================
# 2. MOVIMENTAÇÃO DE TERRA
# =============================================================================
print(f'\n2. MOVIMENTAÇÃO DE TERRA (R$ {V["Mov. Terra"]:,.2f})')

MT = V['Mov. Terra']
vol_escav = area * 0.20
p('Escavação', 'm³', vol_escav, MT * 0.60 / vol_escav, 'Escavação', 90, 150)
p('Movimentação de bota-fora', 'm³', vol_escav * 0.5, MT * 0.25 / (vol_escav * 0.5), 'Bota-fora', 90, 150)
p('Reaterro', 'm³', vol_escav * 0.3, MT * 0.15 / (vol_escav * 0.3), 'Reaterro', 90, 150)

# =============================================================================
# 3. INFRAESTRUTURA
# =============================================================================
print(f'\n3. INFRAESTRUTURA (R$ {V["Infraestrutura"]:,.2f})')

INF = V['Infraestrutura']
ml_estacas = 150 * 12
m3_blocos = 150 * 2
ml_baldrames = area * 0.10

# Tentar vários termos para estacas
for termo in ['Estaqueamento', 'Estaca', 'Fundação profunda']:
    linha = buscar_linha(ws, termo, 100, 200)
    if linha:
        preencher_servico(ws, linha, 'm', ml_estacas, INF * 0.60 / ml_estacas, 'Estacas 60%')
        break

p('Bloco de coroamento', 'm³', m3_blocos, INF * 0.25 / m3_blocos, 'Blocos', 100, 200)
p('Viga baldrame', 'm', ml_baldrames, INF * 0.15 / ml_baldrames, 'Baldrames', 100, 200)

# =============================================================================
# 4. SUPRAESTRUTURA
# =============================================================================
print(f'\n4. SUPRAESTRUTURA (R$ {V["Supraestrutura"]:,.2f})')

SUP = V['Supraestrutura']
vol_conc = area * vol_concreto_m2
kg_aco = vol_conc * taxa_aco_kg_m3
m2_formas = vol_conc * 10

p('Concreto 40MPa', 'm³', vol_conc, SUP * 0.18 / vol_conc, 'Concreto 18%', 100, 250)
p('Concreto 30MPa', 'm³', vol_conc * 0.3, SUP * 0.05 / (vol_conc * 0.3), 'Concreto complementar', 100, 250)
p('Armadura', 'kg', kg_aco, SUP * 0.37 / kg_aco, 'Aço 37%', 100, 250)
p('Forma', 'm²', m2_formas, SUP * 0.05 / m2_formas, 'Formas 5%', 100, 250)

# Mão de obra estrutura (40%)
for termo in ['Mão de obra estrutura', 'Concretagem', 'Montagem estrutura']:
    linha = buscar_linha(ws, termo, 100, 250)
    if linha:
        preencher_servico(ws, linha, 'm²', area, SUP * 0.40 / area, 'MO 40%')
        break

# =============================================================================
# 5. ALVENARIA
# =============================================================================
print(f'\n5. ALVENARIA (R$ {V["Alvenaria"]:,.2f})')

ALV = V['Alvenaria']
m2_alv = area * 2.5  # ~2.5m² alvenaria por m² AC

p('Alvenaria com blocos cerâmicos', 'm²', m2_alv, ALV * 0.70 / m2_alv, 'Blocos 70%', 140, 200)
p('Mão de obra vedações', 'm²', m2_alv, ALV * 0.30 / m2_alv, 'MO Alv 30%', 140, 200)

# =============================================================================
# 6. IMPERMEABILIZAÇÃO
# =============================================================================
print(f'\n6. IMPERMEABILIZAÇÃO (R$ {V["Impermeabilização"]:,.2f})')

IMP = V['Impermeabilização']
m2_imp = area * 0.30  # 30% da área (lajes, banheiros, etc.)

p('Impermeabilização', 'm²', m2_imp, IMP / m2_imp, 'Impermeabilização', 140, 200)

# =============================================================================
# 7. INSTALAÇÕES
# =============================================================================
print(f'\n7. INSTALAÇÕES (R$ {V["Instalações"]:,.2f})')

INST = V['Instalações']
p('Instalações hidrossanitárias', 'm²', area, INST * 0.40 / area, 'Hidro 40%', 160, 250)
p('Instalações elétricas', 'm²', area, INST * 0.36 / area, 'Elétrica 36%', 160, 250)
p('Instalações preventivas', 'm²', area, INST * 0.08 / area, 'Preventiva 8%', 160, 250)
p('Instalação de gás', 'un', n_unidades, INST * 0.05 / n_unidades, 'Gás 5%', 160, 250)
p('Cabeamento estruturado', 'm²', area, INST * 0.11 / area, 'Telecom 11%', 160, 250)

# =============================================================================
# 8. SISTEMAS ESPECIAIS
# =============================================================================
print(f'\n8. SISTEMAS ESPECIAIS (R$ {V["Sist. Especiais"]:,.2f})')

ESP = V['Sist. Especiais']
n_elevadores = 3

p('Elevador', 'un', n_elevadores, ESP * 0.60 / n_elevadores, 'Elevadores 60%', 200, 280)
p('Gerador', 'un', 1, ESP * 0.15, 'Gerador 15%', 200, 280)
p('Automação', 'm²', area, ESP * 0.15 / area, 'Automação 15%', 200, 280)
p('Equipamentos de piscina', 'vb', 1, ESP * 0.10, 'Piscina 10%', 200, 280)

# =============================================================================
# 9. CLIMATIZAÇÃO
# =============================================================================
print(f'\n9. CLIMATIZAÇÃO (R$ {V["Climatização"]:,.2f})')

CLIM = V['Climatização']
n_pontos_ac = n_unidades * 2

p('Ar condicionado', 'ponto', n_pontos_ac, CLIM * 0.70 / n_pontos_ac, 'Pontos AC 70%', 200, 280)
p('Exaustão', 'ponto', n_pavimentos * 4, CLIM * 0.20 / (n_pavimentos * 4), 'Exaustão 20%', 200, 280)
p('Pressurização', 'vb', 1, CLIM * 0.10, 'Pressurização 10%', 200, 280)

# =============================================================================
# 10-18. ACABAMENTOS
# =============================================================================
print(f'\n10. REV. INT. PAREDE (R$ {V["Rev. Int. Parede"]:,.2f})')
p('Revestimento interno', 'm²', area * 2.5, V["Rev. Int. Parede"] / (area * 2.5), 'Rev parede', 220, 280)

print(f'\n11. TETO (R$ {V["Teto"]:,.2f})')
p('Forro', 'm²', area * 0.70, V["Teto"] / (area * 0.70), 'Forro', 220, 280)

print(f'\n12. PISOS (R$ {V["Pisos"]:,.2f})')
p('Piso', 'm²', area * 0.85, V["Pisos"] * 0.70 / (area * 0.85), 'Piso cerâmico 70%', 220, 280)
p('Contrapiso', 'm²', area, V["Pisos"] * 0.30 / area, 'Contrapiso 30%', 220, 280)

print(f'\n13. PINTURA (R$ {V["Pintura"]:,.2f})')
p('Pintura interna', 'm²', area * 3, V["Pintura"] * 0.60 / (area * 3), 'Pintura interna 60%', 220, 280)
p('Pintura externa', 'm²', area * 0.4, V["Pintura"] * 0.40 / (area * 0.4), 'Pintura externa 40%', 220, 280)

print(f'\n14. ESQUADRIAS (R$ {V["Esquadrias"]:,.2f})')
p('Esquadria em alumínio', 'm²', area * 0.15, V["Esquadrias"] * 0.80 / (area * 0.15), 'Esquadrias 80%', 260, 280)
p('Portão', 'un', 2, V["Esquadrias"] * 0.10 / 2, 'Portões 10%', 260, 280)
p('Vidros', 'm²', area * 0.12, V["Esquadrias"] * 0.10 / (area * 0.12), 'Vidros 10%', 260, 280)

print(f'\n15. LOUÇAS E METAIS (R$ {V["Louças e Metais"]:,.2f})')
p('Louças e metais', 'un', n_unidades * 8, V["Louças e Metais"] / (n_unidades * 8), 'Louças/metais', 260, 290)

print(f'\n16. FACHADA (R$ {V["Fachada"]:,.2f})')
p('Fachada', 'm²', area * 0.40, V["Fachada"] / (area * 0.40), 'Fachada', 260, 290)

print(f'\n17. COMPLEMENTARES (R$ {V["Complementares"]:,.2f})')
p('Limpeza final', 'm²', area, V["Complementares"] * 0.30 / area, 'Limpeza 30%', 260, 295)
p('Paisagismo', 'm²', area * 0.05, V["Complementares"] * 0.40 / (area * 0.05), 'Paisagismo 40%', 260, 295)
p('Diversos complementares', 'vb', 1, V["Complementares"] * 0.30, 'Complementares 30%', 260, 295)

print(f'\n18. IMPREVISTOS (R$ {V["Imprevistos"]:,.2f})')
p('Imprevistos', 'vb', 1, V["Imprevistos"], 'Imprevistos 2%', 260, 298)

# =============================================================================
# SALVAR
# =============================================================================
print('\n' + '='*90)
print('✅ Salvando arquivo...')
wb.save('output/CTN-NOW-ORCAMENTO-EXECUTIVO-R00-PREENCHIDO.xlsx')
print('✅ Arquivo salvo: output/CTN-NOW-ORCAMENTO-EXECUTIVO-R00-PREENCHIDO.xlsx')
print('='*90)
