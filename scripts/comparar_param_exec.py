#!/usr/bin/env python3
"""Phase 20i — Compara parametrico vs executivo do MESMO slug.

Gera relatorio MD com desvio por macrogrupo, diagnostico global e
classificacao (DENTRO / ATENCAO / FORA DA FAIXA).

Uso:
    python scripts/comparar_param_exec.py --slug arthen-arboris

Pre-requisito:
    - base/pacotes/{slug}/state.json com parametrico_result.grand_total e rsm2
    - base/indices-executivo/{slug}.json com total, rsm2, macrogrupos

Quando o slug e diferente entre as duas bases, passar:
    --slug-param arthen-arboris --slug-exec arthen-arboris-empreendimento-x
"""
from __future__ import annotations

import argparse
import json
import unicodedata
from datetime import datetime
from pathlib import Path

BASE = Path.home() / "orcamentos-openclaw" / "base"
PAR_DIR = BASE / "pacotes"
IDX_DIR = BASE / "indices-executivo"
OUT_DIR = BASE / "comparacoes-param-exec"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Reutiliza MG_CANON compativel
MG_CANON = [
    ("Gerenciamento", ["gerenciamento", "ger. tec", "ger tec", "gerenciament"]),
    ("Movimentacao de Terra", ["mov", "terra", "terraplan", "movimenta"]),
    ("Infraestrutura", ["infraestrutura", "fundac", "contenc", "estaca", "baldram"]),
    ("Supraestrutura", ["supraestrutura", "estrutura de concreto"]),
    ("Alvenaria", ["alvenaria", "vedac", "divisori"]),
    ("Impermeabilizacao", ["impermeab", "tratament"]),
    ("Instalacoes Hidrossanitarias", ["hidros", "hidraul", "drenagem"]),
    ("Instalacoes Eletricas", ["eletric", "telefon", "logic", "comunica", "automaca"]),
    ("Instalacoes Preventivas", ["preventiv", "pci", "spda", "glp"]),
    ("Instalacoes Gerais", ["instalac", "instalações"]),
    ("Climatizacao", ["climat", "exaust", "ar condic"]),
    ("Revestimentos Parede", ["rev. int. parede", "rev.int.parede", "rev int parede",
                              "revestimentos internos em paredes", "reboco interno",
                              "revestimentos internos de parede", "revestimentos e acabamentos internos em parede",
                              "revestimentos argamassados parede", "acabamentos em parede",
                              "acabamentos internos em parede", "revestimentos ceramicos",
                              "rev. int parede", "rev parede"]),
    ("Revestimentos Teto", ["teto", "forro"]),
    ("Pisos e Pavimentacoes", ["piso", "pavimenta", "contrapiso"]),
    ("Pintura Interna", ["pintura interna", "sistemas de pintura interna"]),
    ("Pintura Geral", ["pintura", "pinturas", "sistema de pintura"]),
    ("Esquadrias", ["esquadri", "vidro", "ferragen"]),
    ("Loucas e Metais", ["louc", "metai"]),
    ("Fachada", ["fachada", "pintura externa", "pintura de fachada"]),
    ("Cobertura", ["cobertura"]),
    ("Sistemas Especiais", ["especia", "equipament", "outros sistemas"]),
    ("Servicos Complementares", ["complementa", "imprevisto", "contingenc"]),
]


def norm(s):
    s = str(s or "").lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def canonize(raw: str) -> str:
    r = norm(raw).strip().lstrip("0123456789. ")
    for canon, kws in MG_CANON:
        for kw in kws:
            if kw in r:
                return canon
    return "Outros"


def _load(p: Path) -> dict:
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def extrair_mgs_exec(idx_data: dict) -> dict:
    """Extrai {canon: valor} dos macrogrupos do executivo."""
    mgs = {}
    for k, v in (idx_data.get("macrogrupos") or {}).items():
        val = v.get("valor") if isinstance(v, dict) else (v or 0)
        if val and val > 0:
            c = canonize(k)
            mgs[c] = mgs.get(c, 0) + val
    return mgs


def extrair_mgs_parametrico(state_data: dict, pkg_dir: Path) -> dict:
    """Extrai {canon: valor} do paramétrico.

    Tenta: state.preliminar_result.macrogrupos, depois
    parametrico_v2_config.json, depois Excel parametrico-*.xlsx.
    """
    # Tentativa 1: preliminar_result / executivo_result
    for k in ("preliminar_result", "executivo_result", "parametrico_result"):
        r = state_data.get(k)
        if isinstance(r, dict) and r.get("macrogrupos"):
            mgs = {}
            for mg, v in r["macrogrupos"].items():
                val = v.get("valor") if isinstance(v, dict) else (v or 0)
                if val and val > 0:
                    c = canonize(mg)
                    mgs[c] = mgs.get(c, 0) + val
            if mgs:
                return mgs

    # Tentativa 2: parametrico-v2-config.json
    cfg = pkg_dir / "parametrico-v2-config.json"
    if cfg.exists():
        c = _load(cfg)
        # Estrutura variavel — procura chaves likely
        for key in ("macrogrupos", "mg_breakdown", "custo_macrogrupo"):
            if c.get(key):
                mgs = {}
                for mg, v in c[key].items():
                    val = v.get("valor") if isinstance(v, dict) else (v or 0)
                    if val and val > 0:
                        canon = canonize(mg)
                        mgs[canon] = mgs.get(canon, 0) + val
                if mgs:
                    return mgs

    # Tentativa 3: Excel (fallback)
    param_xlsx = next(pkg_dir.glob("parametrico-*.xlsx"), None)
    if param_xlsx:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(param_xlsx), read_only=True, data_only=True)
            if "CUSTOS_MACROGRUPO" in wb.sheetnames:
                ws = wb["CUSTOS_MACROGRUPO"]
                mgs = {}
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if row and len(row) >= 4 and row[0] and row[3]:
                        mg = str(row[0])
                        try:
                            val = float(row[3])
                        except (TypeError, ValueError):
                            continue
                        if val > 0 and mg.lower() != "total":
                            c = canonize(mg)
                            mgs[c] = mgs.get(c, 0) + val
                wb.close()
                if mgs:
                    return mgs
        except Exception:
            pass

    return {}


def comparar(slug_param: str, slug_exec: str) -> dict:
    pkg = PAR_DIR / slug_param
    idx = IDX_DIR / f"{slug_exec}.json"
    if not pkg.exists():
        return {"erro": f"pacote paramétrico '{slug_param}' nao encontrado"}
    if not idx.exists():
        return {"erro": f"executivo '{slug_exec}' nao encontrado em indices-executivo"}

    state = _load(pkg / "state.json")
    idx_data = _load(idx)

    # Totais
    par_total = None
    par_rsm2 = None
    par_ac = None
    for k in ("preliminar_result", "executivo_result", "parametrico_result"):
        r = state.get(k)
        if isinstance(r, dict) and r.get("grand_total"):
            par_total = r["grand_total"]
            par_rsm2 = r.get("rsm2")
            break
    if state.get("ac") and par_total and not par_rsm2:
        par_ac = state["ac"]
        par_rsm2 = par_total / par_ac
    elif state.get("ac"):
        par_ac = state["ac"]

    exec_total = idx_data.get("total") or 0
    exec_rsm2 = idx_data.get("rsm2") or 0
    exec_ac = idx_data.get("ac") or 0

    # MGs canonizados
    mgs_par = extrair_mgs_parametrico(state, pkg)
    mgs_exec = extrair_mgs_exec(idx_data)

    # Compara por MG
    all_mgs = set(mgs_par.keys()) | set(mgs_exec.keys())
    comparacao = []
    for mg in sorted(all_mgs):
        vp = mgs_par.get(mg, 0)
        ve = mgs_exec.get(mg, 0)
        delta = ve - vp
        delta_pct = (delta / vp * 100) if vp else None
        comparacao.append({
            "macrogrupo": mg,
            "parametrico_r$": round(vp, 2),
            "executivo_r$": round(ve, 2),
            "delta_r$": round(delta, 2),
            "delta_pct": round(delta_pct, 1) if delta_pct is not None else None,
        })

    # Totais
    if par_total and exec_total:
        delta_total = exec_total - par_total
        delta_total_pct = delta_total / par_total * 100
    else:
        delta_total = None
        delta_total_pct = None

    # Diagnostico
    if delta_total_pct is None:
        diag = "DADOS INSUFICIENTES (faltam totais)"
    elif abs(delta_total_pct) < 5:
        diag = "DENTRO da faixa (desvio < 5%)"
    elif abs(delta_total_pct) < 10:
        diag = "ALERTA leve (5-10%)"
    elif abs(delta_total_pct) < 20:
        diag = "ATENCAO (10-20%)"
    else:
        diag = "FORA DA FAIXA (>20%)"

    return {
        "slug_parametrico": slug_param,
        "slug_executivo": slug_exec,
        "data": datetime.now().isoformat(timespec="seconds"),
        "parametrico": {
            "total_r$": par_total,
            "rsm2": par_rsm2,
            "ac": par_ac,
            "padrao": state.get("padrao"),
        },
        "executivo": {
            "total_r$": exec_total,
            "rsm2": exec_rsm2,
            "ac": exec_ac,
        },
        "delta_total_r$": round(delta_total, 2) if delta_total is not None else None,
        "delta_total_pct": round(delta_total_pct, 1) if delta_total_pct is not None else None,
        "diagnostico": diag,
        "comparacao_por_mg": comparacao,
    }


def gerar_md(r: dict, out_path: Path) -> None:
    if "erro" in r:
        out_path.write_text(f"# ERRO\n\n{r['erro']}\n", encoding="utf-8")
        return
    lines = [
        f"# Comparacao Parametrico x Executivo — {r['slug_parametrico']}",
        "",
        f"**Data:** {r['data']}",
        f"**Slug paramétrico:** `{r['slug_parametrico']}`",
        f"**Slug executivo:** `{r['slug_executivo']}`",
        "",
        "## Totais",
        "",
        "| Fonte | Total (R$) | AC (m²) | R$/m² |",
        "|---|---:|---:|---:|",
    ]
    p, e = r["parametrico"], r["executivo"]
    lines.append(f"| Paramétrico | {p['total_r$']:,.0f} | {p['ac']:,.0f} | {p['rsm2']:.2f} |" if p.get('total_r$') else "| Paramétrico | N/D | N/D | N/D |")
    lines.append(f"| Executivo | {e['total_r$']:,.0f} | {e['ac']:,.0f} | {e['rsm2']:.2f} |" if e.get('total_r$') else "| Executivo | N/D | N/D | N/D |")
    if r["delta_total_pct"] is not None:
        sign = "+" if r["delta_total_pct"] > 0 else ""
        lines.append(f"| **Delta** | **{r['delta_total_r$']:+,.0f}** | — | **{sign}{r['delta_total_pct']:.1f}%** |")
    lines.append("")
    lines.append(f"**Diagnostico:** {r['diagnostico']}")
    lines.append("")
    lines.append("## Comparacao por Macrogrupo")
    lines.append("")
    lines.append("| Macrogrupo | Paramétrico (R$) | Executivo (R$) | Delta (R$) | Delta % |")
    lines.append("|---|---:|---:|---:|---:|")
    for c in r["comparacao_por_mg"]:
        dp = f"{c['delta_pct']:+.1f}%" if c["delta_pct"] is not None else "N/D"
        lines.append(f"| {c['macrogrupo']} | {c['parametrico_r$']:,.0f} | {c['executivo_r$']:,.0f} | {c['delta_r$']:+,.0f} | {dp} |")
    lines.append("")
    lines.append("## Leitura")
    lines.append("")
    # Top 3 MGs que cresceram
    growing = [c for c in r["comparacao_por_mg"] if c["delta_pct"] is not None and c["delta_pct"] > 20]
    shrinking = [c for c in r["comparacao_por_mg"] if c["delta_pct"] is not None and c["delta_pct"] < -20]
    growing.sort(key=lambda x: -x["delta_pct"])
    shrinking.sort(key=lambda x: x["delta_pct"])
    if growing:
        lines.append("**MGs com maior alta no executivo vs paramétrico:**")
        for c in growing[:5]:
            lines.append(f"- {c['macrogrupo']}: +{c['delta_pct']:.0f}% (delta R$ {c['delta_r$']:+,.0f})")
        lines.append("")
    if shrinking:
        lines.append("**MGs com maior queda no executivo vs paramétrico:**")
        for c in shrinking[:5]:
            lines.append(f"- {c['macrogrupo']}: {c['delta_pct']:.0f}% (delta R$ {c['delta_r$']:+,.0f})")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slug", required=False, help="slug comum a paramétrico e executivo")
    ap.add_argument("--slug-param", required=False)
    ap.add_argument("--slug-exec", required=False)
    ap.add_argument("--output", required=False, help="caminho do MD de saida")
    args = ap.parse_args()

    slug_param = args.slug_param or args.slug
    slug_exec = args.slug_exec or args.slug
    if not slug_param or not slug_exec:
        print("ERROR: informe --slug ou --slug-param + --slug-exec", flush=True)
        return

    r = comparar(slug_param, slug_exec)
    out_json = OUT_DIR / f"{slug_param}-vs-{slug_exec}.json"
    out_json.write_text(json.dumps(r, indent=2, ensure_ascii=False), encoding="utf-8")

    out_md = Path(args.output) if args.output else OUT_DIR / f"{slug_param}-vs-{slug_exec}.md"
    gerar_md(r, out_md)

    print(f"Salvo: {out_json}", flush=True)
    print(f"Salvo: {out_md}", flush=True)
    if "erro" in r:
        print(f"\n{r['erro']}")
    else:
        print(f"\nDiagnostico: {r['diagnostico']}")
        if r.get("delta_total_pct") is not None:
            print(f"Delta total: {r['delta_total_pct']:+.1f}%")


if __name__ == "__main__":
    main()
