"""
Cadastra insumos e composições da disciplina Hidrossanitário
(extraídos da extracao.xlsx) nos catálogos canônicos Thozen-Electra.

Entrada:
  - extracao.xlsx aba 'HIDROSSANITÁRIO EXTRAÇÃO' (linhas coloridas marcam grupos)
  - Thozen - Electra_INSUMOS.xlsx (538 insumos existentes)
  - Thozen - Electra_COMPOSIÇÕES.xlsx (961 linhas / 539 composições únicas)

Saída:
  - Thozen - Electra_INSUMOS-v2.xlsx (538 + novos, preservando linhas 1-545)
  - Thozen - Electra_COMPOSIÇÕES-v2.xlsx (961 + 8 composições novas)
  - mapeamento-insumos-composicoes.csv (auditoria)
  - relatorio no stdout (sanity check contra SUBTOTAL da extracao)

Decisões fechadas:
  - Unidade das 8 composições = 'vb'
  - Dedup insumos por (descricao + dimensao + unidade) normalizado
  - Mão de obra vira composição com 1 insumo
  - Sem marca nos insumos
  - Preço = coluna G da extracao
"""
from __future__ import annotations

import csv
import os
import re
import shutil
import sys
import unicodedata
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ---------- paths ----------
DRIVE_BASE = Path.home() / "orcamentos" / "executivos" / "thozen-electra" / "orcamento"
EXTRACAO = DRIVE_BASE / "04-disciplinas" / "Hidrossanitário" / "extracao.xlsx"
INSUMOS_SRC = DRIVE_BASE / "07-visus" / "Thozen - Electra_INSUMOS.xlsx"
COMPOS_SRC = DRIVE_BASE / "07-visus" / "Thozen - Electra_COMPOSIÇÕES.xlsx"
INSUMOS_OUT = DRIVE_BASE / "07-visus" / "Thozen - Electra_INSUMOS-v2.xlsx"
COMPOS_OUT = DRIVE_BASE / "07-visus" / "Thozen - Electra_COMPOSIÇÕES-v2.xlsx"
MAPPING_CSV = DRIVE_BASE / "04-disciplinas" / "Hidrossanitário" / "mapeamento-insumos-composicoes.csv"

# ---------- cores (extracao.xlsx) ----------
# ARGB hex — 8 grupos + 1 subtotal conhecidos da análise prévia
GROUP_COLORS = {
    "FF2F5496": "Instalações de água fria",
    "FFED7D31": "Instalações de água quente",
    "FF548235": "Instalações de esgoto e pluviais",
    "FFC00000": "Cisterna",
    "FF00B0F0": "Reservatórios",
    "FFFFC000": "Sistema de bombas",
    "FF7030A0": "Hidrômetros",
    "FF808080": "Mão de obra",
}
SUBTOTAL_COLOR = "FFD6E4F0"

# Grupo → tipo default do insumo
GROUP_TIPO = {
    "Mão de obra": "Mao_obra",  # espelha convenção do catálogo
}
BOMBA_KEYWORDS = ("bomba", "motor", "pressurizador")  # grupo Sistema de bombas


# ---------- helpers ----------
def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_key(descricao: str, dimensao: str, unidade: str) -> str:
    parts = [strip_accents(str(p or "").strip().lower()) for p in (descricao, dimensao, unidade)]
    parts = [re.sub(r"\s+", " ", p) for p in parts]
    return "|".join(parts)


def full_description(descricao: str, dimensao: str) -> str:
    d = (descricao or "").strip()
    dim = (dimensao or "").strip()
    return f"{d} {dim}".strip() if dim else d


def get_fill_color(cell) -> str | None:
    """Retorna ARGB hex do fill da célula ou None."""
    fill = cell.fill
    if not fill or fill.patternType is None:
        return None
    c = fill.start_color
    if c is None:
        return None
    rgb = c.rgb
    if rgb is None or not isinstance(rgb, str):
        return None
    # rgb pode vir como 'FF2F5496' ou apenas '2F5496'
    rgb = rgb.upper()
    if len(rgb) == 6:
        rgb = "FF" + rgb
    return rgb


def classify_tipo(grupo: str, descricao: str) -> str:
    if grupo in GROUP_TIPO:
        return GROUP_TIPO[grupo]
    if grupo == "Sistema de bombas":
        desc_low = strip_accents(descricao.lower())
        if any(k in desc_low for k in BOMBA_KEYWORDS):
            return "Equipamento"
    return "Material"


# ---------- parse extracao ----------
def parse_extracao() -> list[dict]:
    """Retorna lista de dicts {grupo, descricao, dimensao, qtd, unidade, pu, row_num}."""
    wb = load_workbook(EXTRACAO, data_only=True)
    ws = wb["HIDROSSANITÁRIO EXTRAÇÃO"]
    itens: list[dict] = []
    grupo_atual: str | None = None

    for r in range(7, ws.max_row + 1):
        cell_a = ws.cell(r, 1)
        cell_b = ws.cell(r, 2)
        # Detectar cor: testa A primeiro, depois B
        color = get_fill_color(cell_a) or get_fill_color(cell_b)

        if color == SUBTOTAL_COLOR:
            continue
        if color in GROUP_COLORS:
            # Linha de cabeçalho de grupo — começa novo grupo
            grupo_atual = GROUP_COLORS[color]
            continue
        if grupo_atual is None:
            continue

        # É linha de insumo (fundo branco ou sem cor relevante)
        descricao = cell_b.value
        if not descricao or not str(descricao).strip():
            continue
        dimensao = ws.cell(r, 3).value
        qtd = ws.cell(r, 5).value
        unidade = ws.cell(r, 6).value
        pu = ws.cell(r, 7).value

        # Detecta linha "<grupo>, '0 itens'" que é um sub-cabeçalho sem cor (Mão de obra)
        if str(descricao).strip().lower() in {"0 itens", "1 itens", "2 itens"} or (
            dimensao and "itens" in str(dimensao).lower()
        ):
            continue

        # Mão de obra com PU indefinido → cadastra com placeholder qtd=1, pu=0, un='vb'
        if qtd is None and pu is None:
            itens.append(
                {
                    "grupo": grupo_atual,
                    "descricao": str(descricao).strip(),
                    "dimensao": str(dimensao).strip() if dimensao else "",
                    "qtd": 1.0,
                    "unidade": str(unidade).strip() if unidade else "vb",
                    "pu": 0.0,
                    "pu_placeholder": True,
                    "row_num": r,
                }
            )
            continue

        if qtd is None or pu is None or unidade is None:
            print(f"  [WARN] row {r} grupo={grupo_atual!r} pulada (qtd/un/pu faltando): {descricao}")
            continue

        itens.append(
            {
                "grupo": grupo_atual,
                "descricao": str(descricao).strip(),
                "dimensao": str(dimensao).strip() if dimensao else "",
                "qtd": float(qtd),
                "unidade": str(unidade).strip(),
                "pu": float(pu),
                "pu_placeholder": False,
                "row_num": r,
            }
        )
    wb.close()
    return itens


# ---------- load catalogo insumos existentes ----------
def load_insumos_existentes() -> tuple[dict[str, int], int]:
    """Retorna (chave_norm → codigo, max_codigo). Dados reais começam em row 5."""
    wb = load_workbook(INSUMOS_SRC, data_only=True)
    ws = wb.active
    chave_to_cod: dict[str, int] = {}
    codigos: list[int] = []
    for r in range(5, ws.max_row + 1):
        cod = ws.cell(r, 1).value
        desc = ws.cell(r, 2).value
        un = ws.cell(r, 3).value
        if cod is None or desc is None:
            continue
        try:
            cod_int = int(cod)
        except (ValueError, TypeError):
            continue
        codigos.append(cod_int)
        # Indexar pela descrição inteira (inclui dimensão se houve merge) + unidade
        # Extracao tem dimensão separada, mas nos catálogos existentes pode estar na descrição.
        # Por segurança, indexar só desc completo + unid:
        key = norm_key(str(desc), "", str(un or ""))
        if key not in chave_to_cod:
            chave_to_cod[key] = cod_int
    wb.close()
    return chave_to_cod, max(codigos) if codigos else 0


def load_max_composicao() -> int:
    wb = load_workbook(COMPOS_SRC, data_only=True)
    ws = wb.active
    codigos: list[int] = []
    for r in range(5, ws.max_row + 1):
        cod = ws.cell(r, 1).value
        if cod is None:
            continue
        try:
            codigos.append(int(cod))
        except (ValueError, TypeError):
            pass
    wb.close()
    return max(codigos) if codigos else 0


# ---------- dedup + atribuição de códigos ----------
def dedup_and_assign_codes(
    itens: list[dict], existentes: dict[str, int], next_cod: int
) -> tuple[list[dict], list[dict]]:
    """Para cada item da extracao, atribui código (novo ou reaproveitado).
    Retorna (itens_enriquecidos, insumos_novos_unicos)."""
    # Chave única considera descrição completa (desc + dim) + unidade, pra bater com formato do catálogo
    intra_session: dict[str, dict] = {}
    insumos_novos: list[dict] = []

    for it in itens:
        desc_full = full_description(it["descricao"], it["dimensao"])
        key = norm_key(desc_full, "", it["unidade"])
        it["desc_full"] = desc_full
        it["chave_norm"] = key

        if key in existentes:
            it["codigo_insumo"] = existentes[key]
            it["status"] = "existente"
        elif key in intra_session:
            # Mesmo insumo já visto em outro grupo da própria extracao
            it["codigo_insumo"] = intra_session[key]["codigo_insumo"]
            it["status"] = "reaproveitado_intra"
        else:
            it["codigo_insumo"] = next_cod
            it["status"] = "novo"
            fonte = (
                "Extracao Hidro 2026-04-22 (placeholder — PU a definir)"
                if it.get("pu_placeholder")
                else "Extracao Hidro 2026-04-22"
            )
            novo = {
                "codigo": next_cod,
                "descricao": desc_full,
                "unidade": it["unidade"],
                "valor": it["pu"],
                "tipo": classify_tipo(it["grupo"], desc_full),
                "fonte_preco": fonte,
            }
            intra_session[key] = {"codigo_insumo": next_cod}
            insumos_novos.append(novo)
            next_cod += 1

    return itens, insumos_novos


# ---------- escrever INSUMOS-v2 ----------
def write_insumos_v2(insumos_novos: list[dict]) -> None:
    shutil.copy2(INSUMOS_SRC, INSUMOS_OUT)
    wb = load_workbook(INSUMOS_OUT)
    ws = wb.active
    start_row = ws.max_row + 1
    # Achar primeira célula de dados pra copiar estilo
    template_row = 5
    now = datetime.now()

    for i, ins in enumerate(insumos_novos):
        r = start_row + i
        ws.cell(r, 1, ins["codigo"])
        ws.cell(r, 2, ins["descricao"])
        ws.cell(r, 3, ins["unidade"])
        ws.cell(r, 4, ins["valor"])
        ws.cell(r, 5, ins["fonte_preco"])
        ws.cell(r, 6, ins["tipo"])
        ws.cell(r, 7, now)
        ws.cell(r, 8, True)
        # Copia estilo da linha template (col A-H)
        for col in range(1, 9):
            src = ws.cell(template_row, col)
            dst = ws.cell(r, col)
            if src.has_style:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
    wb.save(INSUMOS_OUT)
    wb.close()


# ---------- escrever COMPOSIÇÕES-v2 ----------
def write_composicoes_v2(itens: list[dict], next_cod_comp: int, insumos_novos: list[dict]) -> list[dict]:
    """Gera 8 composições. Retorna lista de composições com código atribuído pra auditoria."""
    shutil.copy2(COMPOS_SRC, COMPOS_OUT)
    wb = load_workbook(COMPOS_OUT)
    ws = wb.active
    template_row = 5

    # Preço dos insumos novos (pra lookup ao calcular PU da composição)
    pu_lookup_novos = {ins["codigo"]: ins["valor"] for ins in insumos_novos}
    # Também precisa do preço dos insumos existentes que vão ser usados
    wb_ins = load_workbook(INSUMOS_SRC, data_only=True)
    ws_ins = wb_ins.active
    pu_lookup_existentes: dict[int, float] = {}
    for r in range(5, ws_ins.max_row + 1):
        cod = ws_ins.cell(r, 1).value
        val = ws_ins.cell(r, 4).value
        if cod is None or val is None:
            continue
        try:
            pu_lookup_existentes[int(cod)] = float(val)
        except (ValueError, TypeError):
            pass
    wb_ins.close()

    def get_pu(cod: int) -> float:
        return pu_lookup_novos.get(cod, pu_lookup_existentes.get(cod, 0.0))

    # Agrupar itens por grupo (mantendo ordem de aparição)
    grupos_ordem: list[str] = []
    grupos_dict: dict[str, list[dict]] = {}
    for it in itens:
        g = it["grupo"]
        if g not in grupos_dict:
            grupos_dict[g] = []
            grupos_ordem.append(g)
        grupos_dict[g].append(it)

    composicoes_meta: list[dict] = []
    r = ws.max_row + 1
    cod_comp = next_cod_comp

    for grupo in grupos_ordem:
        itens_g = grupos_dict[grupo]
        # PU da composição = Σ(qtd × PU_insumo) — como é vb/1, cada insumo contribui qtd*pu direto
        pu_comp = round(sum(it["qtd"] * get_pu(it["codigo_insumo"]) for it in itens_g), 6)

        for it in itens_g:
            ws.cell(r, 1, cod_comp)
            ws.cell(r, 2, grupo)
            ws.cell(r, 3, "vb")
            ws.cell(r, 4, "")  # referencia
            ws.cell(r, 5, "")  # codigo_referencia
            ws.cell(r, 6, pu_comp)  # preço
            ws.cell(r, 7, 1)  # P.E.
            ws.cell(r, 8, 0)  # FIC
            ws.cell(r, 9, 0)  # FIT
            ws.cell(r, 10, "HIDROSSANITÁRIO")
            ws.cell(r, 11, grupo)
            ws.cell(r, 12, str(it["codigo_insumo"]))  # código insumo como string (padrão observado)
            ws.cell(r, 13, it["qtd"])
            ws.cell(r, 14, "Thozen - Electra")
            ws.cell(r, 15, "INSUMO")
            # Copia estilo
            for col in range(1, 16):
                src = ws.cell(template_row, col)
                dst = ws.cell(r, col)
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.alignment = copy(src.alignment)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.number_format = src.number_format
            r += 1

        composicoes_meta.append(
            {
                "codigo": cod_comp,
                "descricao": grupo,
                "unidade": "vb",
                "n_insumos": len(itens_g),
                "pu_total": pu_comp,
            }
        )
        cod_comp += 1

    wb.save(COMPOS_OUT)
    wb.close()
    return composicoes_meta


# ---------- escrever CSV mapeamento ----------
def write_mapping_csv(itens: list[dict]) -> None:
    MAPPING_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(MAPPING_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(
            [
                "grupo",
                "row_extracao",
                "descricao_original",
                "dimensao",
                "desc_final_cadastrada",
                "unidade",
                "qtd",
                "pu",
                "chave_norm",
                "codigo_insumo",
                "status",
            ]
        )
        for it in itens:
            w.writerow(
                [
                    it["grupo"],
                    it["row_num"],
                    it["descricao"],
                    it["dimensao"],
                    it["desc_full"],
                    it["unidade"],
                    it["qtd"],
                    it["pu"],
                    it["chave_norm"],
                    it["codigo_insumo"],
                    it["status"],
                ]
            )


# ---------- sanity check contra SUBTOTAL da extracao ----------
def read_subtotais_extracao() -> dict[str, float]:
    """Lê linhas de SUBTOTAL da extracao (azul claro) → {grupo: valor}."""
    wb = load_workbook(EXTRACAO, data_only=True)
    ws = wb["HIDROSSANITÁRIO EXTRAÇÃO"]
    subtotais: dict[str, float] = {}
    for r in range(7, ws.max_row + 1):
        cell_a = ws.cell(r, 1)
        # SUBTOTAL fica na col A — ex: "SUBTOTAL Sistema de bombas"
        desc_a = str(cell_a.value or "")
        if not desc_a.upper().startswith("SUBTOTAL"):
            continue
        total = ws.cell(r, 8).value  # col H = Custo Total
        if total is None:
            continue
        m = re.match(r"\s*SUBTOTAL\s+(.+)", desc_a, flags=re.IGNORECASE)
        grupo = m.group(1).strip() if m else desc_a.strip()
        try:
            subtotais[grupo] = float(total)
        except (ValueError, TypeError):
            pass
    wb.close()
    return subtotais


# ---------- main ----------
def main():
    print("=" * 80)
    print("CADASTRO HIDROSSANITÁRIO — Thozen-Electra")
    print("=" * 80)

    if not EXTRACAO.exists():
        print(f"[ERR] extracao não encontrada: {EXTRACAO}")
        sys.exit(1)

    print(f"\n[1/6] Parsing extracao.xlsx …")
    itens = parse_extracao()
    print(f"      → {len(itens)} insumos extraídos")
    por_grupo: dict[str, int] = {}
    for it in itens:
        por_grupo[it["grupo"]] = por_grupo.get(it["grupo"], 0) + 1
    for g, n in por_grupo.items():
        print(f"        · {g}: {n}")

    print(f"\n[2/6] Carregando catálogos existentes …")
    existentes, max_cod_ins = load_insumos_existentes()
    max_cod_comp = load_max_composicao()
    print(f"      → {len(existentes)} chaves no catálogo de insumos (max cod: {max_cod_ins})")
    print(f"      → max cod composição: {max_cod_comp}")

    print(f"\n[3/6] Dedup + atribuição de códigos …")
    itens, insumos_novos = dedup_and_assign_codes(itens, existentes, max_cod_ins + 1)
    status_count = {"existente": 0, "reaproveitado_intra": 0, "novo": 0}
    for it in itens:
        status_count[it["status"]] += 1
    print(f"      → {status_count['novo']} insumos novos")
    print(f"      → {status_count['reaproveitado_intra']} reaproveitados intra-extracao")
    print(f"      → {status_count['existente']} já existiam no catálogo")

    print(f"\n[4/6] Escrevendo INSUMOS-v2.xlsx …")
    write_insumos_v2(insumos_novos)
    print(f"      → {INSUMOS_OUT.name} ({len(insumos_novos)} linhas novas)")

    print(f"\n[5/6] Escrevendo COMPOSIÇÕES-v2.xlsx …")
    composicoes = write_composicoes_v2(itens, max_cod_comp + 1, insumos_novos)
    print(f"      → {COMPOS_OUT.name} ({len(composicoes)} composições, {len(itens)} linhas de insumo)")

    print(f"\n[6/6] CSV mapeamento + sanity check …")
    write_mapping_csv(itens)
    print(f"      → {MAPPING_CSV.name}")

    # Sanity: PU composição vs SUBTOTAL extracao
    subtotais_ext = read_subtotais_extracao()
    print("\n" + "=" * 80)
    print("SANITY CHECK — PU composição vs SUBTOTAL extracao")
    print("=" * 80)
    print(f"{'Grupo':<40} {'PU comp':>14} {'Subtotal ext':>14} {'Diff':>10}")
    total_comp = 0.0
    total_ext = 0.0
    for c in composicoes:
        nome = c["descricao"]
        subtotal_ext = subtotais_ext.get(nome, 0.0)
        diff = c["pu_total"] - subtotal_ext
        total_comp += c["pu_total"]
        total_ext += subtotal_ext
        flag = "OK" if abs(diff) < 1 else "WARN"
        print(f"{nome:<40} {c['pu_total']:>14,.2f} {subtotal_ext:>14,.2f} {diff:>10,.2f}  {flag}")
    print("-" * 80)
    print(f"{'TOTAL':<40} {total_comp:>14,.2f} {total_ext:>14,.2f} {total_comp-total_ext:>10,.2f}")
    print()
    print(f"✅ Composições cadastradas: cod {max_cod_comp+1}..{max_cod_comp+len(composicoes)}")
    print(f"✅ Insumos novos:           cod {max_cod_ins+1}..{max_cod_ins+len(insumos_novos)}")


if __name__ == "__main__":
    main()
