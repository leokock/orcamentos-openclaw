#!/usr/bin/env python3
"""Gate de Validação — gera planilha Excel com decisões pra Leo aprovar.

Antes de rodar o executivo automatizado, gera um xlsx curto com:
- Aba GATE: linhas com decisões binárias/lista (dropdowns), defaults vindos
  dos projetos similares
- Aba PROJETOS_REFERENCIA: 5 similares usados como base
- Aba PREMISSAS_PROPOSTAS: premissas técnicas consolidadas
- Aba BDI: encargos observados nos similares

Leo abre o xlsx, ajusta dropdowns/valores na aba GATE, salva com sufixo
`-validado`. O pacote orquestrador lê esse xlsx validado e usa as decisões
no gerar_executivo_auto.

Uso:
  python gerar_gate_validacao.py --slug projeto-novo --ac 15000 --ur 90 \\
                                 --padrao alto -o gate.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
import consulta_similares as cs  # noqa: E402

DARK = "2C3E50"
ORANGE = "E67E22"
GREEN = "27AE60"
INPUT_BG = "FFF3E0"
HEADER_FILL = "2C3E50"
ALT_FILL = "F8F9FA"

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def hdr(ws, row, cols, widths=None):
    for i, t in enumerate(cols, 1):
        c = ws.cell(row, i, t)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def cell_input(ws, r, c, value):
    cell = ws.cell(r, c, value)
    cell.font = Font(bold=True, size=11, color=ORANGE, name="Arial")
    cell.fill = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
    cell.border = THIN
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def cell_label(ws, r, c, value, bold=False):
    cell = ws.cell(r, c, value)
    cell.font = Font(bold=bold, size=10, name="Arial")
    cell.border = THIN
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return cell


def aba_dados_projeto(wb: Workbook, slug: str, ac: float, ur: int | None,
                       padrao: str | None, similares: list[dict]) -> None:
    ws = wb.create_sheet("DADOS_PROJETO", 0)
    ws.sheet_properties.tabColor = DARK

    ws["A1"] = f"GATE DE VALIDAÇÃO — {slug.upper()}"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")

    ws["A2"] = "Use este Excel para validar as premissas e decisões antes do executivo ser gerado automaticamente."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    rows = [
        ("Projeto", slug),
        ("AC alvo (m²)", ac),
        ("UR alvo", ur or ""),
        ("Padrão", padrao or ""),
        ("N projetos similares", len(similares)),
        ("Slugs similares", ", ".join(p["_slug"] for p in similares)),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        cell_label(ws, i, 1, k, bold=True)
        cell_input(ws, i, 2, v)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def aba_referencias(wb: Workbook, similares: list[dict]) -> None:
    ws = wb.create_sheet("PROJETOS_REFERENCIA")
    ws.sheet_properties.tabColor = ORANGE

    ws["A1"] = "PROJETOS DE REFERÊNCIA — 5 mais similares"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:G1")

    hdr(ws, 3, ["#", "Projeto", "AC (m²)", "UR", "R$/m²", "Total R$", "Sub-disc"], [4, 38, 12, 8, 14, 18, 10])

    for i, p in enumerate(similares, start=1):
        cell_label(ws, i + 3, 1, i)
        cell_label(ws, i + 3, 2, p["_slug"])
        cell_label(ws, i + 3, 3, p.get("ac"))
        cell_label(ws, i + 3, 4, p.get("ur"))
        cell_label(ws, i + 3, 5, round(p.get("rsm2") or 0, 2) if p.get("rsm2") else "")
        cell_label(ws, i + 3, 6, p.get("total"))
        n_sub = len((p.get("qualitative") or {}).get("sub_disciplinas") or [])
        cell_label(ws, i + 3, 7, n_sub)


def aba_gate(wb: Workbook, similares: list[dict], padrao: str | None, ac: float) -> None:
    """Aba principal — decisões com dropdowns que Leo edita."""
    ws = wb.create_sheet("GATE")
    ws.sheet_properties.tabColor = GREEN

    ws["A1"] = "GATE — DECISÕES PRA VALIDAR"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    ws["A2"] = "Para cada decisão: ajuste o dropdown (coluna B) ou aceite o default. Coluna C mostra fontes."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    hdr(ws, 4, ["Decisão", "Resposta", "Default sugerido", "Fontes (projetos)", "Impacto", "Categoria"],
        [38, 22, 26, 38, 36, 16])

    decisoes = [
        ("Tipo de Laje", "Convencional", ["Convencional", "Protendida", "Nervurada", "Maciça", "Mista"],
         "Define concreto, aço, forma, escoramento", "Estrutura"),
        ("Tipo de Fundação", "Hélice contínua", ["Hélice contínua", "Tubulão", "Sapata", "Estaca raiz", "Estaca cravada"],
         "Define cota e PU da infraestrutura", "Estrutura"),
        ("Subsolos", "0", ["0", "1", "2", "3"],
         "Cada subsolo +R$ 200-400/m²", "Geometria"),
        ("Padrão de Acabamento", padrao or "Médio-Alto", ["Médio", "Médio-Alto", "Alto", "Luxo"],
         "Define PUs de pisos, esquadrias, louças (impacto 30-40%)", "Acabamento"),
        ("Tipo de Fachada", "Textura projetada", ["Textura projetada", "Cerâmica", "Pastilha", "ACM", "Pele de vidro"],
         "Define R$/m² da fachada", "Fachada"),
        ("Tipo de Esquadria", "Alumínio anodizado", ["Alumínio anodizado", "Alumínio pintado", "Madeira", "PVC"],
         "Define PU médio das esquadrias", "Acabamento"),
        ("Tipo de Piso predominante", "Misto", ["Porcelanato", "Laminado", "Misto"],
         "Porcelanato ~R$ 81/m² vs laminado ~R$ 65/m²", "Acabamento"),
        ("Pintura padrão", "Acrílica + selador", ["Acrílica + selador", "PVA", "Texturizada", "Esmalte"],
         "Define PU da pintura", "Acabamento"),
        ("Climatização", "Infra apenas", ["Infra apenas", "Infra + equipamentos", "Sem"],
         "Equipamentos somam R$ 80-200/m²", "Sistemas"),
        ("Sistema de Pressurização", "Não", ["Sim", "Não"],
         "+R$ 80k se sim", "Sistemas"),
        ("Gerador Dedicado", "Sim", ["Sim", "Não"],
         "+R$ 120-350k se sim", "Sistemas"),
        ("Piscina", "Sim", ["Não", "Sim", "Aquecida"],
         "+R$ 0/220k/320k", "Sistemas"),
        ("Perdas estimadas (concreto)", "13%", ["5%", "8%", "10%", "13%", "15%", "20%"],
         "Premissa de orçamento estrutural", "Premissa"),
        ("Prazo de obra (meses)", "30", ["18", "24", "30", "36", "42", "48"],
         "Define duração do CI/equipe", "Premissa"),
        ("BDI total estimado", "25%", ["20%", "22%", "25%", "27%", "28%", "30%"],
         "Aplicado sobre custos diretos", "Premissa"),
        ("Encargos (sobre MO)", "85%", ["80%", "85%", "90%", "95%", "100%"],
         "Aplicado sobre mão-de-obra direta", "Premissa"),
        ("Tempo de protensão (meses)", "0", ["0", "3", "6", "12"],
         "Apenas se laje protendida", "Premissa"),
        ("Tipologia predominante", "Misto", ["Studios", "1-2 dorms", "3-4 dorms", "Misto"],
         "Define pontos de elétrica/hidro/louças", "Geometria"),
        ("Garagem (vagas/UR)", "1.0", ["0.5", "1.0", "1.5", "2.0"],
         "Define escala de impermeab/piso garagem", "Geometria"),
        ("Tipo de Contenção", "Não", ["Não", "Cortina", "Muro de arrimo", "Solo grampeado", "Tirantes"],
         "Custo significativo se houver subsolo", "Estrutura"),
    ]

    fontes_padrao = ", ".join(p["_slug"][:25] for p in similares[:3])

    for i, (decisao, default, options, impacto, categoria) in enumerate(decisoes, start=5):
        cell_label(ws, i, 1, decisao, bold=True)
        cell_input(ws, i, 2, default)

        dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=False)
        dv.error = "Selecione uma opção da lista"
        dv.errorTitle = "Valor inválido"
        ws.add_data_validation(dv)
        dv.add(ws.cell(i, 2))

        cell_label(ws, i, 3, default)
        cell_label(ws, i, 4, fontes_padrao)
        cell_label(ws, i, 5, impacto)
        cell_label(ws, i, 6, categoria)

    n_decisoes = len(decisoes)

    info_row = 5 + n_decisoes + 2
    ws.cell(info_row, 1, "Após validar, salve este arquivo com sufixo `-validado.xlsx` na mesma pasta.")
    ws.cell(info_row, 1).font = Font(bold=True, size=10, color=DARK, name="Arial")
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=6)


def aba_premissas(wb: Workbook, premissas: list[dict]) -> None:
    ws = wb.create_sheet("PREMISSAS_PROPOSTAS")
    ws.sheet_properties.tabColor = "8E44AD"

    ws["A1"] = "PREMISSAS PROPOSTAS — extraídas de projetos similares"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")

    ws["A2"] = "Use estas premissas como base no memorial. Marque 'usar?' = NÃO se quiser excluir alguma."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")

    hdr(ws, 4, ["Área", "Premissa", "Frequência", "Fontes", "Usar?"],
        [22, 65, 14, 28, 10])

    if not premissas:
        ws.cell(5, 1, "(nenhuma premissa consolidada nos similares)")
        return

    dv_usar = DataValidation(type="list", formula1='"SIM,NÃO"', allow_blank=False)
    dv_usar.error = "Use SIM ou NÃO"
    ws.add_data_validation(dv_usar)

    for i, pr in enumerate(premissas, start=5):
        cell_label(ws, i, 1, pr.get("area", ""), bold=True)
        cell_label(ws, i, 2, pr.get("premissa", ""))
        cell_label(ws, i, 3, f"{pr['freq']}/{len(set().union(*[set(pr['fontes'])])) or pr['freq']}")
        cell_label(ws, i, 4, ", ".join(pr.get("fontes", [])[:3]))
        cell_input(ws, i, 5, "SIM")
        dv_usar.add(ws.cell(i, 5))


def aba_bdi(wb: Workbook, similares: list[dict]) -> None:
    bdi = cs.bdi_encargos_observados(similares)
    if not bdi:
        return

    ws = wb.create_sheet("BDI_ENCARGOS")
    ws.sheet_properties.tabColor = "C0392B"

    ws["A1"] = "BDI / ENCARGOS observados nos projetos de referência"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")

    hdr(ws, 3, ["Componente", "Valor (%)", "Nota", "Fonte"],
        [28, 14, 50, 26])

    for i, b in enumerate(bdi, start=4):
        cell_label(ws, i, 1, b.get("componente", ""), bold=True)
        cell_label(ws, i, 2, b.get("valor_pct", ""))
        cell_label(ws, i, 3, b.get("nota", ""))
        cell_label(ws, i, 4, b.get("fonte", ""))


def aba_subdisciplinas(wb: Workbook, enriq: dict) -> None:
    ws = wb.create_sheet("SUB_DISCIPLINAS")
    ws.sheet_properties.tabColor = "16A085"

    ws["A1"] = "SUB-DISCIPLINAS extraídas dos projetos similares (preview)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")

    ws["A2"] = "Estas sub-disciplinas serão usadas pra detalhar cada macrogrupo no executivo automatizado."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    hdr(ws, 4, ["Macrogrupo", "Sub-disciplina", "Frequência", "Fontes", "Itens exemplo", "Usar?"],
        [22, 26, 12, 26, 60, 8])

    dv_usar = DataValidation(type="list", formula1='"SIM,NÃO"', allow_blank=False)
    ws.add_data_validation(dv_usar)

    row = 5
    for mg, subs in enriq.get("sub_disciplinas_por_mg", {}).items():
        for sd in subs:
            cell_label(ws, row, 1, mg, bold=True)
            cell_label(ws, row, 2, sd["sub"])
            cell_label(ws, row, 3, f"{sd['freq']}/{enriq['n_similares']}")
            cell_label(ws, row, 4, ", ".join(sd.get("fontes", [])[:3]))
            cell_label(ws, row, 5, " | ".join(sd.get("itens_exemplo", [])[:3]))
            cell_input(ws, row, 6, "SIM")
            dv_usar.add(ws.cell(row, 6))
            row += 1


def gerar_gate(slug: str, ac: float, ur: int | None, padrao: str | None, output: str) -> dict:
    similares = cs.projetos_similares(ac=ac, ur=ur, padrao=padrao, n=5)
    enriq = cs.enriquecer_parametrico(similares)
    premissas = cs.premissas_consolidadas(similares)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    aba_dados_projeto(wb, slug, ac, ur, padrao, similares)
    aba_gate(wb, similares, padrao, ac)
    aba_referencias(wb, similares)
    aba_premissas(wb, premissas)
    aba_subdisciplinas(wb, enriq)
    aba_bdi(wb, similares)

    Path(output).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    return {
        "output": output,
        "n_similares": len(similares),
        "similares": [p["_slug"] for p in similares],
        "n_premissas": len(premissas),
        "n_macrogrupos_com_sub": len(enriq.get("sub_disciplinas_por_mg", {})),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slug", required=True, help="slug do projeto novo (ex: novo-projeto-x)")
    ap.add_argument("--ac", type=float, required=True, help="área construída em m²")
    ap.add_argument("--ur", type=int, default=None, help="número de unidades residenciais")
    ap.add_argument("--padrao", default=None, help="padrão (médio/alto/luxo)")
    ap.add_argument("-o", "--output", default=None, help="caminho do xlsx de saída")
    args = ap.parse_args()

    output = args.output or f"gate-{args.slug}.xlsx"
    result = gerar_gate(args.slug, args.ac, args.ur, args.padrao, output)
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
