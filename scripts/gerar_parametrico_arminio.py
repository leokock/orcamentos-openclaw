#!/usr/bin/env python3.11
"""
Gerador de Orçamento Paramétrico - Arminio Tavares
Usando dados extraídos do IFC + base de calibração Cartesian
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import json
import os
from datetime import datetime

# ============================================================================
# 1. CARREGAR DADOS EXTRAÍDOS DO IFC
# ============================================================================

with open('projetos/arminio-tavares/dados_parametrico.json', 'r', encoding='utf-8') as f:
    dados = json.load(f)

# ============================================================================
# 2. CARREGAR BASE DE CALIBRAÇÃO
# ============================================================================

with open('calibration-stats.json', 'r', encoding='utf-8') as f:
    calibration = json.load(f)

# ============================================================================
# 3. PARÂMETROS DO PROJETO
# ============================================================================

# Dados extraídos do IFC
AREA_CONSTRUIDA = dados['AC']
UNIDADES = dados['UR']
PAVIMENTOS_TIPO = dados['NPT']
NUM_SUBSOLOS = dados['NS']
ELEVADORES = dados['ELEV']
VAGAS = dados['VAG']
AREA_TERRENO = dados['AT']

# CUB atualizado (mar/2026)
CUB_ATUAL = 3150.00  # CUB SC mar/2026 (estimado)
CUB_BASE = 2752.67   # CUB dez/23 (base de calibração)
FATOR_CUB = CUB_ATUAL / CUB_BASE

# ============================================================================
# 4. PREMISSAS DE PROJETO (valores padrão razoáveis)
# ============================================================================

PREMISSAS = {
    # Alto impacto
    'tipo_laje': 'Cubetas',                    # Padrão Cartesian
    'padrao_acabamento': 'Alto Padrão',        # Razoável para Placon
    'contencao': 'Cortina de estacas',         # Centro urbano, vizinhos
    'prazo_obra': 30,                          # meses
    'num_subsolos': NUM_SUBSOLOS,
    
    # Médio impacto
    'tipo_fundacao': 'Hélice Contínua',        # Padrão SC
    'tipo_esquadria': 'Alumínio Anodizado',    # Padrão
    'tipo_fachada': 'Textura + Pintura',       # Padrão econômico
    'mo_fachada': 'Equipe Própria',
    'vedacao_interna': 'Alvenaria',
    'nivel_lazer': 'Completo',                 # Piscina + academia
    
    # Baixo impacto
    'piso': 'Porcelanato',
    'forro': 'Gesso Liso',
    'cobertura_habitavel': 'Não',
    'aquecimento': 'Gás Individual',
    'automacao': 'Básico',
    'energia_solar': 'Não',
    'paisagismo': 'Básico',
    'mobiliario': 'Sem',
    'regiao': 'Capital Floripa',
    
    # Infraestrutura
    'gerador': 'Sim',
    'subestacao': 'Não',
    'fotovoltaico': 'Não',
    'infra_carro_eletrico': 'Sim',
    'pressurizacao_escada': 'Sim'
}

# ============================================================================
# 5. FATORES DE AJUSTE BASEADOS NAS PREMISSAS
# ============================================================================

def calcular_fatores():
    """Calcula fatores de ajuste para cada macrogrupo baseado nas premissas"""
    
    fatores = {
        'Gerenciamento': 1.00,
        'Mov. Terra': 1.00,
        'Infraestrutura': 1.00,
        'Supraestrutura': 1.00,
        'Alvenaria': 1.00,
        'Impermeabilização': 1.00,
        'Instalações': 1.00,
        'Sist. Especiais': 1.00,
        'Climatização': 1.00,
        'Rev. Int. Parede': 1.00,
        'Teto': 1.00,
        'Pisos': 1.00,
        'Pintura': 1.00,
        'Esquadrias': 1.00,
        'Louças e Metais': 1.00,
        'Fachada': 1.00,
        'Complementares': 1.00,
        'Imprevistos': 1.00
    }
    
    # Ajustes por tipo de laje
    if PREMISSAS['tipo_laje'] == 'Protendida':
        fatores['Supraestrutura'] *= 1.10
    elif PREMISSAS['tipo_laje'] == 'Maciça':
        fatores['Supraestrutura'] *= 1.15
    
    # Ajustes por padrão de acabamento
    if PREMISSAS['padrao_acabamento'] == 'Super Alto Padrão':
        fatores['Esquadrias'] *= 1.40
        fatores['Pisos'] *= 1.30
        fatores['Rev. Int. Parede'] *= 1.25
        fatores['Louças e Metais'] *= 1.50
        fatores['Fachada'] *= 1.35
    elif PREMISSAS['padrao_acabamento'] == 'Alto Padrão':
        fatores['Esquadrias'] *= 1.20
        fatores['Pisos'] *= 1.15
        fatores['Rev. Int. Parede'] *= 1.10
        fatores['Louças e Metais'] *= 1.25
        fatores['Fachada'] *= 1.15
    elif PREMISSAS['padrao_acabamento'] == 'Econômico':
        fatores['Esquadrias'] *= 0.85
        fatores['Pisos'] *= 0.85
        fatores['Rev. Int. Parede'] *= 0.90
        fatores['Louças e Metais'] *= 0.80
    
    # Ajustes por contenção
    if PREMISSAS['contencao'] != 'Não':
        # Adicionar componente de contenção (~5% do total)
        fatores['Infraestrutura'] *= 1.35
    
    # Ajustes por subsolos
    if NUM_SUBSOLOS >= 2:
        fatores['Mov. Terra'] *= 1.80
        fatores['Infraestrutura'] *= 1.25
        fatores['Impermeabilização'] *= 1.40
    elif NUM_SUBSOLOS == 1:
        fatores['Mov. Terra'] *= 1.30
        fatores['Infraestrutura'] *= 1.15
        fatores['Impermeabilização'] *= 1.20
    
    # Ajustes por fachada
    if PREMISSAS['tipo_fachada'] == 'ACM':
        fatores['Fachada'] *= 1.40
    elif PREMISSAS['tipo_fachada'] == 'Pele de Vidro':
        fatores['Fachada'] *= 1.80
    elif PREMISSAS['tipo_fachada'] == 'Cerâmica/Pastilha':
        fatores['Fachada'] *= 1.25
    
    # Ajustes por lazer
    if PREMISSAS['nivel_lazer'] == 'Premium':
        fatores['Complementares'] *= 1.40
    elif PREMISSAS['nivel_lazer'] == 'Completo':
        fatores['Complementares'] *= 1.20
    
    # Ajustes por infraestrutura especial
    if PREMISSAS['gerador'] == 'Sim':
        fatores['Sist. Especiais'] *= 1.15
    if PREMISSAS['infra_carro_eletrico'] == 'Sim':
        fatores['Instalações'] *= 1.05
    if PREMISSAS['pressurizacao_escada'] == 'Sim':
        fatores['Instalações'] *= 1.08
    
    return fatores

# ============================================================================
# 6. GERAR PLANILHA DE ORÇAMENTO
# ============================================================================

def gerar_planilha():
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento Paramétrico"
    
    # Estilos
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True)
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================================================
    # CABEÇALHO
    # ========================================================================
    
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = "ORÇAMENTO PARAMÉTRICO"
    cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Info do projeto
    row = 3
    info = [
        ("Projeto:", dados['nome_projeto']),
        ("Cliente:", dados['cliente']),
        ("Cidade:", f"{dados['cidade']}, {dados['estado']}"),
        ("Área Construída:", f"{AREA_CONSTRUIDA:,.2f} m²"),
        ("Unidades:", f"{UNIDADES} un"),
        ("Pavimentos Tipo:", f"{PAVIMENTOS_TIPO}"),
        ("Subsolos:", f"{NUM_SUBSOLOS}"),
        ("Data Base:", datetime.now().strftime("%d/%m/%Y")),
        ("CUB SC (mar/2026):", f"R$ {CUB_ATUAL:,.2f}/m²"),
    ]
    
    for label, value in info:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name="Calibri", size=10, bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name="Calibri", size=10)
        row += 1
    
    row += 1
    
    # ========================================================================
    # PREMISSAS
    # ========================================================================
    
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "PREMISSAS DO PROJETO"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    premissas_display = [
        ("Tipo de Laje:", PREMISSAS['tipo_laje']),
        ("Padrão Acabamento:", PREMISSAS['padrao_acabamento']),
        ("Contenção:", PREMISSAS['contencao']),
        ("Tipo de Fundação:", PREMISSAS['tipo_fundacao']),
        ("Tipo de Esquadria:", PREMISSAS['tipo_esquadria']),
        ("Tipo de Fachada:", PREMISSAS['tipo_fachada']),
        ("Nível de Lazer:", PREMISSAS['nivel_lazer']),
        ("Gerador:", PREMISSAS['gerador']),
        ("Infra Carro Elétrico:", PREMISSAS['infra_carro_eletrico']),
        ("Pressurização Escada:", PREMISSAS['pressurizacao_escada']),
        ("Prazo de Obra:", f"{PREMISSAS['prazo_obra']} meses"),
    ]
    
    for label, value in premissas_display:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name="Calibri", size=9, bold=True)
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name="Calibri", size=9)
        ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    row += 1
    
    # ========================================================================
    # ORÇAMENTO POR MACROGRUPO
    # ========================================================================
    
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "COMPOSIÇÃO DE CUSTOS"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Cabeçalho da tabela
    headers = ["Macrogrupo", "R$/m² Base", "Fator", "R$/m² Ajustado", "Custo Total", "% Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_thin
    row += 1
    
    # Calcular fatores
    fatores = calcular_fatores()
    
    # Linhas de macrogrupos
    custo_total = 0
    custos_macrogrupos = {}
    
    for macrogrupo in calibration['categories'].keys():
        if macrogrupo not in ['Louças', 'Contenção', 'Cobertura', 'Decoração']:
            rsm2_base = calibration['categories'][macrogrupo]['rsm2']['median']
            fator = fatores.get(macrogrupo, 1.00)
            rsm2_ajustado = rsm2_base * FATOR_CUB * fator
            custo_macro = rsm2_ajustado * AREA_CONSTRUIDA
            
            custos_macrogrupos[macrogrupo] = custo_macro
            custo_total += custo_macro
            
            # Escrever linha
            ws.cell(row=row, column=1).value = macrogrupo
            ws.cell(row=row, column=1).font = Font(name="Calibri", size=10)
            ws.cell(row=row, column=1).border = border_thin
            
            ws.cell(row=row, column=2).value = rsm2_base
            ws.cell(row=row, column=2).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=2).border = border_thin
            
            ws.cell(row=row, column=3).value = fator
            ws.cell(row=row, column=3).number_format = '0.00'
            ws.cell(row=row, column=3).border = border_thin
            
            ws.cell(row=row, column=4).value = rsm2_ajustado
            ws.cell(row=row, column=4).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=4).border = border_thin
            
            ws.cell(row=row, column=5).value = custo_macro
            ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=5).border = border_thin
            
            ws.cell(row=row, column=6).value = custo_macro / custo_total if custo_total > 0 else 0
            ws.cell(row=row, column=6).number_format = '0.00%'
            ws.cell(row=row, column=6).border = border_thin
            
            row += 1
    
    # Recalcular % após ter custo_total
    for i, macrogrupo in enumerate(custos_macrogrupos.keys()):
        pct = custos_macrogrupos[macrogrupo] / custo_total
        ws.cell(row=row-len(custos_macrogrupos)+i, column=6).value = pct
    
    # Linha de total
    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=1).font = total_font
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).border = border_thin
    
    ws.cell(row=row, column=5).value = custo_total
    ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'
    ws.cell(row=row, column=5).font = total_font
    ws.cell(row=row, column=5).fill = total_fill
    ws.cell(row=row, column=5).border = border_thin
    
    ws.cell(row=row, column=6).value = 1.00
    ws.cell(row=row, column=6).number_format = '0.00%'
    ws.cell(row=row, column=6).font = total_font
    ws.cell(row=row, column=6).fill = total_fill
    ws.cell(row=row, column=6).border = border_thin
    
    row += 2
    
    # ========================================================================
    # RESUMO EXECUTIVO
    # ========================================================================
    
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws[f'A{row}']
    cell.value = "RESUMO EXECUTIVO"
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center')
    row += 1
    
    custo_m2 = custo_total / AREA_CONSTRUIDA
    custo_unidade = custo_total / UNIDADES
    cub_ratio = custo_m2 / CUB_ATUAL
    
    resumo = [
        ("Custo Total da Obra:", f"R$ {custo_total:,.2f}"),
        ("Custo por m²:", f"R$ {custo_m2:,.2f}/m²"),
        ("Custo por Unidade:", f"R$ {custo_unidade:,.2f}"),
        ("CUB Ratio:", f"{cub_ratio:.2f}×"),
        ("", ""),
        ("Observações:", ""),
        ("• Valores em reais (BRL), base mar/2026", ""),
        ("• Inclui BDI e gerenciamento", ""),
        ("• Não inclui terreno, projetos e licenças", ""),
        ("• Prazo estimado: 30 meses", ""),
    ]
    
    for label, value in resumo:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name="Calibri", size=10, bold=True if value else False)
        if value:
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(name="Calibri", size=10)
            ws.merge_cells(f'B{row}:D{row}')
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    
    # Salvar
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    
    filename = f'Parametrico_Arminio-Tavares_{datetime.now().strftime("%Y%m%d")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    
    print("\n" + "="*80)
    print(f"✓ Orçamento paramétrico gerado com sucesso!")
    print("="*80)
    print(f"\nArquivo: {filepath}")
    print(f"\nRESUMO:")
    print(f"  Custo Total: R$ {custo_total:,.2f}")
    print(f"  Custo/m²: R$ {custo_m2:,.2f}")
    print(f"  Custo/Unidade: R$ {custo_unidade:,.2f}")
    print(f"  CUB Ratio: {cub_ratio:.2f}×")
    print("="*80)
    
    return filepath

# ============================================================================
# EXECUTAR
# ============================================================================

if __name__ == '__main__':
    gerar_planilha()
