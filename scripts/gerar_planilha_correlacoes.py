#!/usr/bin/env python3
"""Fase 5 — Excel de correlacoes controladas + PCA + clusters."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
INPUT = BASE / "correlacoes-controladas.json"
OUT = Path.home() / "orcamentos-openclaw" / "analises-cross-projeto" / "correlacoes-controladas" / "correlacoes-controladas.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

DARK = "2C3E50"; ACCENT = "2980B9"; RED = "C0392B"; GREEN = "27AE60"
PURPLE = "8E44AD"; TEAL = "16A085"; YELLOW_D = "F39C12"
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)
GREEN_FILL = PatternFill(start_color="E8F8E0", end_color="E8F8E0", fill_type="solid")
RED_FILL = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")


def sh(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def wh(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, 1):
        sh(ws.cell(row, i, h), color)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wr(ws, row, values, nf=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if nf and i in nf:
            c.number_format = nf[i]


def main():
    d = json.loads(INPUT.read_text(encoding="utf-8"))
    wb = Workbook(); wb.remove(wb.active)

    # === LEIA_ME ===
    ws = wb.create_sheet("LEIA_ME"); ws.sheet_properties.tabColor = GREEN
    ws["A1"] = "CORRELACOES CONTROLADAS + PCA + CLUSTERS (Fase 5)"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')} . {d['n_projetos']} projetos"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")

    row = 4
    ws.cell(row, 1, "METODOLOGIA").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    metodo = [
        "1. Correlacoes simples: Pearson r entre pares de variaveis, com p-value. Filtro |r| > 0.2 e n >= 20.",
        "2. Correlacoes stratificadas: r(X, Y) dentro de cada subgrupo (ex: cub_regiao). Revela se efeito e uniforme ou concentrado.",
        "3. Correlacoes PARCIAIS: r(X, Y | Z) — residualiza X e Y em Z antes de correlacionar. Revela causalidade vs confundimento.",
        "4. PCA: reducao de 25+ indicadores em 5 componentes principais.",
        "5. Clustering Ward hierarquico: alternativa ao K-means, sobre assinatura % MG.",
    ]
    for m in metodo:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row, 1, m).font = Font(size=9, name="Arial")
        row += 1

    row += 2
    ws.cell(row, 1, "ABAS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    abas = [
        ("SIMPLES", "Pearson r + p-value | >0.2, n>=20 | 23 pares"),
        ("STRATIFICADAS", "r(AC, rsm2) dentro de cada regiao — expoe heterogeneidade"),
        ("PARCIAIS", "r(X, Y | Z) — controla confundidores"),
        ("INSIGHTS_CAUSAIS", "Correlacoes que MUDARAM quando controladas (ver abaixo)"),
        ("PCA", "5 componentes principais (67% variancia)"),
        ("CLUSTERS_WARD", "Clustering hierarquico — 4 clusters stavel"),
    ]
    wh(ws, row, ["Aba", "Conteudo"], [24, 80])
    row += 1
    for a, c in abas:
        wr(ws, row, [a, c])
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110

    # === SIMPLES ===
    ws = wb.create_sheet("SIMPLES"); ws.sheet_properties.tabColor = ACCENT
    ws["A1"] = "CORRELACOES SIMPLES (Pearson r, |r|>0.2, n>=20)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    wh(ws, 3, ["Par de variaveis", "r", "p-value", "n", "Interpretacao"], [55, 10, 12, 8, 40])
    row = 4
    simples = sorted(d["correlacoes"]["correlacoes_simples"].items(),
                     key=lambda x: -abs(x[1]["r"]))
    for par, info in simples:
        r = info["r"]
        interp = ""
        if abs(r) > 0.7: interp = "Forte"
        elif abs(r) > 0.5: interp = "Moderada"
        elif abs(r) > 0.3: interp = "Leve"
        else: interp = "Fraca"
        if r > 0: interp += " positiva"
        else: interp += " negativa"
        if info["p"] > 0.05:
            interp += " (NAO significativa p>0.05)"
        wr(ws, row, [par, r, info["p"], info["n"], interp], {2: "0.000", 3: "0.0000"})
        if info["p"] > 0.05:
            ws.cell(row, 3).fill = YELLOW_FILL
        if abs(r) > 0.5:
            ws.cell(row, 2).fill = RED_FILL if r > 0 else GREEN_FILL
        row += 1
    ws.auto_filter.ref = f"A3:E{row-1}"
    ws.freeze_panes = "A4"

    # === STRATIFICADAS ===
    ws = wb.create_sheet("STRATIFICADAS"); ws.sheet_properties.tabColor = PURPLE
    ws["A1"] = "CORRELACOES STRATIFICADAS — r(X,Y) dentro de cada regiao"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")
    ws["A2"] = "Revela se efeito e uniforme (todas regioes similares) ou heterogeneo (varia por regiao)."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")
    wh(ws, 4, ["Par de variaveis", "Regiao", "r", "p-value", "n", "Interpretacao"],
       [32, 22, 10, 12, 6, 30])
    row = 5
    for par, regioes in d["correlacoes"]["correlacoes_stratificadas"].items():
        for regiao, info in regioes.items():
            r = info["r"]
            interp = ""
            if info["n"] < 10:
                interp = "n muito pequeno"
            elif abs(r) < 0.2:
                interp = "Nula / muito fraca"
            elif abs(r) < 0.4:
                interp = "Leve"
            elif abs(r) < 0.6:
                interp = "Moderada"
            else:
                interp = "Forte"
            if r > 0: interp += " +"
            else: interp += " -"
            wr(ws, row, [par, regiao, r, info["p"], info["n"], interp], {3: "0.000", 4: "0.0000"})
            if abs(r) > 0.4:
                ws.cell(row, 3).fill = RED_FILL if r > 0 else GREEN_FILL
            row += 1
    ws.auto_filter.ref = f"A4:F{row-1}"
    ws.freeze_panes = "A5"

    # === PARCIAIS ===
    ws = wb.create_sheet("PARCIAIS"); ws.sheet_properties.tabColor = RED
    ws["A1"] = "CORRELACOES PARCIAIS r(X, Y | Z) — controlam confundidor"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")
    ws["A2"] = "Se r_parcial ~ r_simples, a correlacao e robusta. Se r_parcial ~ 0, o efeito vinha do confundidor."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")
    wh(ws, 4, ["Par + controle", "r_parcial", "r_simples (referencia)", "delta r", "n", "Conclusao"],
       [45, 12, 22, 12, 6, 40])
    row = 5
    for par, info in d["correlacoes"]["correlacoes_parciais"].items():
        rp = info["r_parcial"]
        rs = info.get("r_simples_comparacao")
        delta = info.get("delta_r")
        conclusao = ""
        if rs is not None:
            if abs(rp) < 0.1 and abs(rs) > 0.2:
                conclusao = "CONFUNDIDOR! Controle reduz r quase a zero."
            elif abs(rp - rs) < 0.05:
                conclusao = "Robusta. Controle nao muda."
            elif abs(rp) < abs(rs):
                conclusao = f"Atenuada pelo controle ({abs(rs)-abs(rp):.2f})."
            else:
                conclusao = f"Amplificada pelo controle."
        else:
            conclusao = "Sem r_simples pra comparar."
        wr(ws, row, [par, rp, rs, delta, info["n"], conclusao], {2: "0.000", 3: "0.000", 4: "+0.000;-0.000"})
        if abs(rp) < 0.1 and rs is not None and abs(rs) > 0.2:
            ws.cell(row, 2).fill = RED_FILL
            ws.cell(row, 6).fill = RED_FILL
        row += 1
    ws.auto_filter.ref = f"A4:F{row-1}"
    ws.freeze_panes = "A5"

    # === INSIGHTS_CAUSAIS ===
    ws = wb.create_sheet("INSIGHTS_CAUSAIS"); ws.sheet_properties.tabColor = YELLOW_D
    ws["A1"] = "INSIGHTS-CAUSAIS — O que aprendemos das correlacoes parciais"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:C1")
    row = 3
    insights = [
        ("Economia de escala e real", "r(AC, rsm2 | cub_regiao) = -0.317",
         "Mesmo controlando regiao, projetos maiores tem R$/m² menor. Efeito leve mas real."),
        ("Supraestrutura NAO causa R$/m² alto", "r(supraestrutura%, rsm2 | cub_regiao) ~ 0",
         "Correlacao simples vinha de confundidor regional. Ter muita estrutura nao torna projeto caro."),
        ("Gerenciamento% tampouco causa R$/m² alto", "r(ger%, rsm2 | regiao) = -0.095",
         "Projetos EPCM nao sao caros por causa de gerenciamento — e pelo pacote de escopo completo."),
        ("Taxa aco NAO tem efeito direto no R$/m²", "r(taxa_aco, rsm2 | cub_regiao) = -0.122",
         "Aco kg/m³ e caracteristica da engenharia estrutural, nao preditor de preco."),
        ("AC prediz Total bem (orcamento escala com area)", "r(AC, total | cub_regiao) = +0.529",
         "Relacao forte mesmo controlando regiao. Regra pratica: R$ total ~ AC x R$/m² regional."),
        ("Gerenciamento e oposto de Hidro/Supra", "r(ger%, hidro%) = -0.685",
         "Quando ger concentra custo, hidrossanitaria fica reduzido proporcionalmente — padrao EPCM."),
        ("Acabamentos crescem juntos (PC1 = 28% var)", "PCA componente 1: acabamento",
         "PC1 pega cobertura + porcelanato + taxa aco. 'Densidade de acabamento' e eixo principal."),
        ("Economia de escala HETEROGENEA por regiao", "stratificado",
         "Em Floripa/Vale-Itajai quase zero (r=-0.1 e -0.15). Em Litoral-Norte forte (r=-0.52)."),
    ]
    wh(ws, row, ["Insight", "Evidencia numerica", "Implicacao acionavel"], [40, 35, 55])
    row += 1
    for titulo, num, impl in insights:
        ws.cell(row, 1, titulo).font = Font(bold=True, size=10, color=DARK, name="Arial")
        ws.cell(row, 1).border = THIN
        ws.cell(row, 2, num).font = Font(size=9, name="Consolas")
        ws.cell(row, 2).border = THIN
        ws.cell(row, 3, impl).font = Font(size=9, name="Arial")
        ws.cell(row, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 3).border = THIN
        ws.row_dimensions[row].height = 45
        row += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 65

    # === PCA ===
    ws = wb.create_sheet("PCA"); ws.sheet_properties.tabColor = TEAL
    ws["A1"] = "PCA — 5 componentes principais"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    pca = d.get("pca", {})
    if "erro" not in pca:
        ws["A2"] = f"{pca['n_projetos_usados']} projetos × {pca['n_variaveis']} variaveis. 5 componentes explicam 67% da variancia."
        ws.merge_cells("A2:D2")
        row = 4
        for comp in pca["componentes"]:
            ws.cell(row, 1, f"{comp['componente']} — variancia explicada {comp['variancia_explicada']:.1%} (acum {comp['variancia_acumulada']:.1%})").font = Font(bold=True, size=11, color=DARK, name="Arial")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
            wh(ws, row, ["Variavel", "Loading", "Direcao", "Contribuicao"], [45, 12, 14, 18])
            row += 1
            for tv in comp["top_variaveis"]:
                dir_ = "+" if tv["loading"] > 0 else "-"
                abs_l = abs(tv["loading"])
                contrib = "Alta" if abs_l > 0.3 else "Media" if abs_l > 0.2 else "Baixa"
                wr(ws, row, [tv["variavel"], tv["loading"], dir_, contrib], {2: "+0.000;-0.000"})
                row += 1
            row += 1

    # === CLUSTERS_WARD ===
    ws = wb.create_sheet("CLUSTERS_WARD"); ws.sheet_properties.tabColor = PURPLE
    ws["A1"] = "CLUSTERING HIERARQUICO WARD (alternativa ao K-means)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    clust = d.get("clustering", {})
    if "erro" not in clust:
        ws["A2"] = f"{clust['n_projetos_usados']} projetos com dados MG completos. Ward method sobre {len(clust['colunas_usadas'])} colunas % MG."
        ws.merge_cells("A2:D2")
        row = 4
        ws.cell(row, 1, "4 CLUSTERS:").font = Font(bold=True, size=11, color=DARK, name="Arial")
        row += 2
        wh(ws, row, ["Cluster", "N projetos", "Exemplos (top 5)"], [12, 12, 80])
        row += 1
        for cn, slugs in clust["clusters_4"].items():
            wr(ws, row, [cn, len(slugs), ", ".join(slugs[:5]) + (f"... (+{len(slugs)-5} mais)" if len(slugs) > 5 else "")])
            row += 1

        row += 2
        ws.cell(row, 1, "6 CLUSTERS (granularidade maior):").font = Font(bold=True, size=11, color=DARK, name="Arial")
        row += 2
        wh(ws, row, ["Cluster", "N projetos", "Exemplos (top 5)"], [12, 12, 80])
        row += 1
        for cn, slugs in clust["clusters_6"].items():
            wr(ws, row, [cn, len(slugs), ", ".join(slugs[:5]) + (f"... (+{len(slugs)-5} mais)" if len(slugs) > 5 else "")])
            row += 1

    wb.save(str(OUT))
    print(f"Salvo: {OUT}", flush=True)


if __name__ == "__main__":
    main()
