#!/usr/bin/env python3
"""Phase 20h — Excel consolidado: Cluster 3 deep-dive + Parametrico validacao.

Abas:
 1. LEIA_ME
 2. CLUSTER3_DETALHADO (13 projetos + % gerenciamento)
 3. CLUSTER3_vs_BASE (comparacao % MG)
 4. PARAMETRICO_VALIDACAO (4 pacotes ativos vs nuvem de executivos)
 5. CROSS_REFERENCE_PROTOCOLO (recomendacao operacional)

Uso: python scripts/gerar_planilha_cluster3_parametrico.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
CLUSTER_FILE = BASE / "cluster3-e-parametrico-validacao.json"
AGG_FILE = BASE / "analise-avancada-agregada.json"
AF_FILE = BASE / "analise-financeira-agregada.json"
OUT_XLSX = BASE / "cluster3-e-parametrico.xlsx"

DARK, ACCENT, ORANGE, GREEN, PURPLE, RED, GRAY, TEAL, YELLOW_D = (
    "2C3E50", "2980B9", "E67E22", "27AE60", "8E44AD", "C0392B", "7F8C8D", "16A085", "F39C12"
)
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
GREEN_FILL = PatternFill(start_color="E8F8E0", end_color="E8F8E0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
RED_FILL = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")


def _load(p):
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def sh(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def wh(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, 1):
        sh(ws.cell(row, i, h), color)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def wr(ws, row, values, nf=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if nf and i in nf:
            c.number_format = nf[i]
        elif isinstance(v, float):
            c.number_format = "#,##0.00"
        elif isinstance(v, int):
            c.number_format = "#,##0"


def build_leia_me(wb, data):
    ws = wb.create_sheet("LEIA_ME")
    ws.sheet_properties.tabColor = GREEN
    ws["A1"] = "CLUSTER 3 DEEP-DIVE + VALIDACAO PARAMETRICO"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")

    row = 4
    sections = [
        ("PARTE 1: CLUSTER 3",
         "Os 13 projetos do Cluster 3 descobertos na analise avancada tem em comum ~35% do total em GERENCIAMENTO (vs 12% na mediana da base). Aqui investigamos quais projetos, quais clientes, quanto de gerenciamento por projeto, e comparamos cada MG com o resto da base."),
        ("PARTE 2: VALIDACAO PARAMETRICO",
         "4 pacotes parametricos ativos (arthen-arboris, pacote-piloto, placon-arminio-tavares, thozen-electra). Nenhum tem par direto em indices-executivo. Validamos cada um estatisticamente contra a NUVEM de executivos do mesmo padrao construtivo."),
        ("CONCLUSAO CLUSTER 3",
         "Cluster 3 = segmento de contratacao onde Cartesian concentra gerenciamento (35% vs 12%). Nova Empreendimentos tem 4/4 projetos aqui, sao os mais caros (R$/m² 5-7k). Paludo tem 4 projetos com mesma estrutura mas R$/m² eficientes (2-3k). Assinatura nao e sobre tipo de obra, e sobre MODO DE CONTRATACAO (EPCM/fee-based)."),
        ("CONCLUSAO PARAMETRICO",
         "arthen-arboris (-2.7%) e thozen-electra (-0.8%) estao bem centrados no padrao. pacote-piloto (-17.7% vs alto) pode estar subestimando. placon-arminio-tavares (+17.4% vs medio) pode estar superestimando. A base atual nao permite medir erro real - precisa de pares diretos."),
    ]
    for title, body in sections:
        ws.cell(row, 1, title).font = Font(bold=True, size=11, color=DARK, name="Arial")
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row, 1).font = Font(size=9, name="Arial")
        row += 3

    ws.cell(row, 1, "MAPA DAS ABAS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["#", "Aba", "Descricao"], [4, 35, 70])
    row += 1
    abas = [
        ("1", "LEIA_ME", "Este guia"),
        ("2", "CLUSTER3_DETALHADO", "13 projetos com cliente, AC, total, % gerenciamento"),
        ("3", "CLUSTER3_vs_BASE", "Comparacao % MG: cluster 3 vs resto da base (63 projetos)"),
        ("4", "PARAMETRICO_VALIDACAO", "4 pacotes validados contra mediana do padrao"),
        ("5", "CROSS_REFERENCE_PROTOCOLO", "Recomendacao: como manter paramétrico<->executivo rastreavel"),
    ]
    for n, a, d in abas:
        wr(ws, row, [n, a, d])
        row += 1

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 60


def build_cluster3_detalhado(wb, data):
    ws = wb.create_sheet("CLUSTER3_DETALHADO")
    ws.sheet_properties.tabColor = PURPLE
    ws["A1"] = "CLUSTER 3 — 13 projetos com gerenciamento elevado (>30% do total)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:G1")

    c3 = data.get("cluster3_detalhado", [])
    # Agrupa por cliente
    wh(ws, 3, ["Cliente", "Projeto", "Padrao", "AC (m²)", "Total (R$)", "R$/m²", "Gerenc %"],
       [22, 38, 14, 14, 18, 14, 12])
    row = 4
    nf = {4: "#,##0", 5: "R$ #,##0", 6: "R$ #,##0.00", 7: "0.00"}
    from collections import defaultdict
    agg = defaultdict(list)
    for p in c3:
        agg[p["cliente"]].append(p)

    for cli in sorted(agg.keys(), key=lambda x: -len(agg[x])):
        for p in sorted(agg[cli], key=lambda x: -x["ger_pct"]):
            # Padrão do projeto
            pad = "?"
            for ent in data.get("parametrico_validacao", []):
                pass  # mantemos do outro source
            # Apenas coloca os dados disponíveis
            wr(ws, row, [p.get("cliente"), p["slug"], "", p["ac"], p["total"],
                         p.get("rsm2"), p["ger_pct"]], nf)
            # Color ger_pct
            if p["ger_pct"] >= 50:
                ws.cell(row, 7).fill = RED_FILL
                ws.cell(row, 7).font = Font(size=9, bold=True, color=RED, name="Arial")
            elif p["ger_pct"] >= 40:
                ws.cell(row, 7).fill = YELLOW_FILL
            row += 1

    # Summary
    row += 2
    ws.cell(row, 1, "RESUMO POR CLIENTE").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Cliente", "N projetos", "Ger % med", "R$/m² med", "Total R$ med"],
       [22, 12, 12, 14, 18])
    row += 1
    import statistics
    for cli in sorted(agg.keys(), key=lambda x: -len(agg[x])):
        gs = [p["ger_pct"] for p in agg[cli] if p["ger_pct"]]
        rs = [p.get("rsm2") for p in agg[cli] if p.get("rsm2")]
        ts = [p["total"] for p in agg[cli] if p["total"]]
        wr(ws, row, [
            cli, len(agg[cli]),
            round(statistics.median(gs), 1) if gs else None,
            round(statistics.median(rs), 2) if rs else None,
            round(statistics.median(ts), 0) if ts else None,
        ], {3: "0.0", 4: "R$ #,##0", 5: "R$ #,##0"})
        row += 1
    ws.freeze_panes = "A4"


def build_cluster3_vs_base(wb, data_af):
    ws = wb.create_sheet("CLUSTER3_vs_BASE")
    ws.sheet_properties.tabColor = ORANGE
    ws["A1"] = "CLUSTER 3 vs RESTO DA BASE — Comparacao % MG"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    ws["A2"] = "Mediana de % de cada MG no total. Diferenca = Cluster 3 - Outros. Sinal + = cluster 3 pesa MAIS."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")

    # Dados hardcoded da análise
    rows = [
        ("Gerenciamento",           35.39, 12.38,  23.01),
        ("Supraestrutura",          14.03, 20.02,  -5.99),
        ("Outros",                  10.03, 15.80,  -5.77),
        ("Hidrossanitaria",          2.67,  6.25,  -3.58),
        ("Pisos",                    4.40,  7.94,  -3.54),
        ("Sist especiais",           2.72,  5.43,  -2.71),
        ("Infraestrutura",           3.87,  6.02,  -2.14),
        ("Alvenaria",                2.32,  4.38,  -2.06),
        ("Complementares",           4.74,  6.56,  -1.82),
        ("Pintura interna",          2.35,  3.86,  -1.51),
        ("Esquadrias",               7.94,  8.91,  -0.97),
        ("Fachada",                  3.03,  3.94,  -0.91),
        ("Forro/teto",               1.24,  1.92,  -0.68),
        ("Impermeabilizacao",        1.87,  1.78,   0.10),
    ]
    wh(ws, 4, ["Macrogrupo", "Cluster 3 %", "Outros %", "Diferenca (pp)", "Leitura"],
       [24, 14, 14, 16, 40])
    row = 5
    for mg, c3m, om, diff in rows:
        sign = "+" if diff > 0 else ""
        leitura = ""
        if mg == "Gerenciamento":
            leitura = "Assinatura do cluster — 3x a base"
        elif diff < -5:
            leitura = "Cluster 3 aloca MENOS em custos diretos (compensando pra ger)"
        elif diff > 5:
            leitura = "Cluster 3 aloca MAIS (atipico)"
        elif abs(diff) < 1:
            leitura = "Similar"
        else:
            leitura = "Sublotado" if diff < 0 else "Sobrelotado"
        wr(ws, row, [mg, c3m, om, diff, leitura], {2: "0.00", 3: "0.00", 4: "+0.00;-0.00"})
        if mg == "Gerenciamento":
            ws.cell(row, 4).fill = RED_FILL
            ws.cell(row, 4).font = Font(size=9, bold=True, color=RED, name="Arial")
        elif diff < -3:
            ws.cell(row, 4).fill = YELLOW_FILL
        row += 1
    ws.freeze_panes = "A5"


def build_parametrico_validacao(wb, data):
    ws = wb.create_sheet("PARAMETRICO_VALIDACAO")
    ws.sheet_properties.tabColor = TEAL
    ws["A1"] = "VALIDACAO PARAMETRICO contra NUVEM DE EXECUTIVOS"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")
    ws["A2"] = ("4 pacotes parametricos ativos comparados contra a mediana R$/m² de executivos do mesmo padrao construtivo. "
                "Nenhum pacote tem par executivo direto na base - validacao e ESTATISTICA, nao pontual.")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:H2")

    wh(ws, 4, ["Pacote", "Padrao", "AC (m²)", "Total Parametrico (R$)",
               "R$/m² Parametrico", "Mediana Exec (padrao)", "Desvio %", "Diagnostico"],
       [28, 14, 12, 20, 16, 18, 12, 20])
    row = 5
    nf = {3: "#,##0", 4: "R$ #,##0", 5: "R$ #,##0.00", 6: "R$ #,##0.00", 7: "+0.0;-0.0"}
    for p in data.get("parametrico_validacao", []):
        wr(ws, row, [
            p["pacote"], p["padrao"], p["ac"], p["total_parametrico"],
            p["rsm2_parametrico"], p["rsm2_mediana_padrao"],
            p["desvio_pct"], p["diagnostico"],
        ], nf)
        # Color diagnosis
        diag = p["diagnostico"]
        if diag == "DENTRO da faixa":
            ws.cell(row, 8).fill = GREEN_FILL
            ws.cell(row, 8).font = Font(size=9, bold=True, color=GREEN, name="Arial")
        elif diag == "ATENCAO":
            ws.cell(row, 8).fill = YELLOW_FILL
            ws.cell(row, 8).font = Font(size=9, bold=True, color=ORANGE, name="Arial")
        elif "FORA" in diag:
            ws.cell(row, 8).fill = RED_FILL
            ws.cell(row, 8).font = Font(size=9, bold=True, color=RED, name="Arial")
        row += 1

    row += 2
    ws.cell(row, 1, "INTERPRETACAO DOS DESVIOS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    interps = [
        "arthen-arboris: -2.7% vs mediana medio-alto. Parametrico centrado, bom pra entregar.",
        "thozen-electra: -0.8% vs mediana alto. Parametrico quase exato na mediana.",
        "pacote-piloto: -17.7% vs mediana alto. Ou projeto e simples dentro do alto, ou parametrico subestima.",
        "placon-arminio-tavares: +17.4% vs mediana medio. Amostra de medio e pequena (n=2) - pouca confianca.",
    ]
    for i in interps:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row, 1, i).font = Font(size=9, name="Arial")
        row += 1


def build_protocolo(wb):
    ws = wb.create_sheet("CROSS_REFERENCE_PROTOCOLO")
    ws.sheet_properties.tabColor = ACCENT
    ws["A1"] = "PROTOCOLO DE CROSS-REFERENCE PARAMETRICO <-> EXECUTIVO"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:C1")
    ws["A2"] = "Como habilitar medicao de erro do parametrico no futuro."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:C2")

    row = 4
    ws.cell(row, 1, "PROBLEMA").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row, 1, "Atualmente 0 dos 4 parametricos tem executivo correspondente na base indices-executivo. Razoes: (1) slug diferente entre parametrico e executivo; (2) executivos sao projetos antigos (antes do parametrico existir); (3) parametricos sao projetos novos em andamento.").alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 1).font = Font(size=9, name="Arial")
    row += 3

    ws.cell(row, 1, "PROTOCOLO PROPOSTO").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Etapa", "Acao", "Output esperado"], [24, 46, 40])
    row += 1
    protocolo = [
        ("1. Fechamento parametrico", "Salvar slug canonico no state.json (ex: arthen-arboris)", "state.json com slug estavel"),
        ("2. Inicio executivo", "Reutilizar MESMO slug no indices-executivo do projeto", "arquivo indices-executivo/{slug}.json"),
        ("3. Entrega executivo", "Quando concluir, executar scripts/comparar_param_exec.py --slug X", "relatorio de desvio por MG"),
        ("4. Registro no git", "Commit no orcamentos-openclaw mantendo pares rastreaveis", "rastro de versao"),
        ("5. Calibracao agregada", "A cada 5+ pares novos, rodar recalibracao da base parametrica", "indices-catalogo.xlsx atualizado"),
    ]
    for e, a, o in protocolo:
        wr(ws, row, [e, a, o])
        row += 1

    row += 2
    ws.cell(row, 1, "SCRIPT DE COMPARACAO (a implementar)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row, 1, ("Criar scripts/comparar_param_exec.py que, dado um slug, le parametrico (base/pacotes/{slug}/state.json) "
                     "e executivo (base/indices-executivo/{slug}.json) e gera relatorio MD com desvio por MG, "
                     "itens que vieram a mais/menos, e classifica projeto em 'DENTRO' / 'ATENCAO' / 'FORA DA FAIXA'.")).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row, 1).font = Font(size=9, name="Arial")
    row += 3

    ws.cell(row, 1, "METAS DE COBERTURA").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    wh(ws, row, ["Meta", "Prazo sugerido", "Criterio de sucesso"], [28, 18, 40])
    row += 1
    metas = [
        ("5 pares parametrico-executivo", "12 meses", "Desvio medio < 15%, dp < 20%"),
        ("10 pares", "18 meses", "Calibracao confiavel por padrao"),
        ("20 pares", "24 meses", "Modelo preditivo de erro por MG"),
    ]
    for m, p, c in metas:
        wr(ws, row, [m, p, c])
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 42


def main():
    print("Loading...", flush=True)
    data = _load(CLUSTER_FILE)
    af = _load(AF_FILE)
    if not data:
        print("ERROR: cluster file not found")
        return
    wb = Workbook()
    wb.remove(wb.active)
    print("Building tabs...", flush=True)
    build_leia_me(wb, data); print("  LEIA_ME")
    build_cluster3_detalhado(wb, data); print("  CLUSTER3_DETALHADO")
    build_cluster3_vs_base(wb, af); print("  CLUSTER3_vs_BASE")
    build_parametrico_validacao(wb, data); print("  PARAMETRICO_VALIDACAO")
    build_protocolo(wb); print("  CROSS_REFERENCE_PROTOCOLO")
    wb.save(str(OUT_XLSX))
    print(f"\nSalvo: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
