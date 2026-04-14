#!/usr/bin/env python3
"""Phase 9 — Extração de fórmulas Excel dos xlsx.

Lê xlsx com data_only=False (openpyxl) e captura todas as células com
fórmulas (que começam com '='). Isso revela como cada valor foi calculado:
dependências entre células, constantes hardcoded, multiplicações de índice × AC.

Saída: base/formulas/[projeto].json
Schema:
{
  "abas": [
    {"nome": "SUPRA", "formulas": [
      {"cell": "F10", "formula": "=AC*0.25*C10", "sheet": "SUPRA"}
    ]}
  ],
  "stats": {"n_formulas_total": 1234, "refs_ac": 45, "refs_cross_sheet": 12}
}
"""
from __future__ import annotations

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path.home() / "orcamentos-openclaw" / "base"
QUEUE = BASE / "gemma-queue.json"
OUT_DIR = BASE / "formulas"
OUT_DIR.mkdir(parents=True, exist_ok=True)
LOG = BASE / "phase9-extract.log.jsonl"

SKIP_PATTERNS = [r"^cap[ai]", r"^gr[áa]fic", r"^[íi]ndice$", r"^sum[áa]rio$",
                  r"^instr", r"^legenda"]


def should_skip(name: str) -> bool:
    n = name.strip().lower()
    return any(re.search(p, n) for p in SKIP_PATTERNS)


def log_event(e: dict) -> None:
    e.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
    with LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(e, ensure_ascii=False) + "\n")


def extract_xlsx(xlsx_path: Path) -> dict:
    out = {"arquivo": str(xlsx_path), "abas": [], "errors": []}
    try:
        wb = load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        out["errors"].append(f"open: {type(e).__name__}")
        return out

    stats = {
        "n_formulas_total": 0,
        "cross_sheet_refs": 0,
        "with_absolute_ref": 0,
        "with_ac_ref": 0,
        "with_constant": 0,
    }

    AC_REF_RE = re.compile(r"\bAC\b|\$[A-Z]+\$\d+", re.IGNORECASE)
    CROSS_SHEET_RE = re.compile(r"'?[^!']+'?!")
    ABS_REF_RE = re.compile(r"\$[A-Z]+\$?\d+|\$?[A-Z]+\$\d+")

    for sn in wb.sheetnames:
        if should_skip(sn):
            continue
        try:
            ws = wb[sn]
        except Exception as e:
            out["errors"].append(f"sheet {sn}: {e}")
            continue

        formulas = []
        rows_iter = 0
        for row in ws.iter_rows():
            rows_iter += 1
            if rows_iter > 8000:
                break
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("="):
                    formula = val[:300]
                    formulas.append({"cell": cell.coordinate, "formula": formula})
                    stats["n_formulas_total"] += 1
                    if CROSS_SHEET_RE.search(formula):
                        stats["cross_sheet_refs"] += 1
                    if AC_REF_RE.search(formula):
                        stats["with_ac_ref"] += 1
                    if ABS_REF_RE.search(formula):
                        stats["with_absolute_ref"] += 1
                    if re.search(r"[\*\+\-\/]\s*\d+\.?\d*", formula):
                        stats["with_constant"] += 1
                    if len(formulas) > 2500:
                        break
            if len(formulas) > 2500:
                break

        if formulas:
            out["abas"].append({"nome": sn, "n_formulas": len(formulas),
                                 "formulas_sample": formulas[:100]})

    out["stats"] = stats
    try:
        wb.close()
    except Exception:
        pass
    return out


def pick_xlsx(sources):
    files = sources.get("xlsx", [])
    if not files:
        return None
    sized = []
    for p in files:
        try:
            sized.append((Path(p).stat().st_size, p))
        except Exception:
            pass
    sized.sort(reverse=True)
    return sized[0][1] if sized else None


def process(project):
    slug = project["projeto"]
    summary = {"projeto": slug, "status": "pending"}
    t0 = time.time()
    xlsx = pick_xlsx(project.get("sources", {}))
    if not xlsx:
        summary["status"] = "no_xlsx"
        log_event(summary)
        return summary
    try:
        result = extract_xlsx(Path(xlsx))
        result["projeto"] = slug
        (OUT_DIR / f"{slug}.json").write_text(
            json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        summary["status"] = "done"
        summary["stats"] = result["stats"]
        summary["n_abas"] = len(result["abas"])
    except Exception as e:
        summary["status"] = "failed"
        summary["error"] = f"{type(e).__name__}: {str(e)[:100]}"

    summary["duration_s"] = round(time.time() - t0, 2)
    log_event(summary)
    return summary


def main():
    queue = json.loads(QUEUE.read_text(encoding="utf-8"))
    print(f"processing {len(queue)} projects")
    counts = {}
    t0 = time.time()
    for i, p in enumerate(queue, 1):
        s = process(p)
        counts[s["status"]] = counts.get(s["status"], 0) + 1
        if s["status"] == "done":
            st = s.get("stats", {})
            print(f"[{i}/{len(queue)}] {p['projeto']:<45} f={st.get('n_formulas_total',0):>5} ac_ref={st.get('with_ac_ref',0):>4} x_sheet={st.get('cross_sheet_refs',0):>4} {s['duration_s']}s", flush=True)
    print()
    print(f"summary: {counts}")
    print(f"total time: {int(time.time() - t0)}s")


if __name__ == "__main__":
    main()
