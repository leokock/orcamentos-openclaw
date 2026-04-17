#!/usr/bin/env python3
"""Phase 20c v2 — Gera planilha de Análise de Produto Cartesian.

Abas:
 1. LEIA_ME
 2. RESUMO (números-chave agregados)
 3. PROJETOS
 4. CUSTO_POR_MG
 5. ESTRUTURAL
 6. ELETRICA
 7. HIDRO
 8. ALVENARIA_REVEST
 9. ESQUADRIAS_LOUCAS
10. DIVERSOS
11. MATRIZ_PROJETOS
12. BENCHMARKS (faixas por indicador × padrão)
13. OUTLIERS (projetos fora da faixa)
14. ABC (top itens por macrogrupo)
15. PUS_REFERENCIA (PUs validados)
16. CORRELACOES
17. INSIGHTS (flags por projeto)
18. DIFERENCIADORES

Uso: python scripts/gerar_planilha_analise_produto.py
"""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
AGG_FILE = BASE / "indicadores-produto-agregados.json"
COND_FILE = BASE / "calibration-condicional-padrao.json"
OUT_XLSX = BASE / "analise-produto-cartesian.xlsx"

DARK = "2C3E50"
ACCENT = "2980B9"
ORANGE = "E67E22"
GREEN = "27AE60"
PURPLE = "8E44AD"
RED = "C0392B"
GRAY = "7F8C8D"
TEAL = "16A085"
YELLOW_D = "F39C12"

THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

GREEN_FILL = PatternFill(start_color="E8F8E0", end_color="E8F8E0", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
RED_FILL = PatternFill(start_color="FDECEC", end_color="FDECEC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")

ABA_COLORS = {
    "LEIA_ME": GREEN,
    "RESUMO": DARK,
    "PROJETOS": ACCENT,
    "CUSTO_POR_MG": ORANGE,
    "ESTRUTURAL": PURPLE,
    "ELETRICA": TEAL,
    "HIDRO": ACCENT,
    "ALVENARIA_REVEST": ORANGE,
    "ESQUADRIAS_LOUCAS": PURPLE,
    "DIVERSOS": GRAY,
    "MATRIZ_PROJETOS": RED,
    "BENCHMARKS": GREEN,
    "OUTLIERS": RED,
    "ABC": YELLOW_D,
    "PUS_REFERENCIA": TEAL,
    "CORRELACOES": ACCENT,
    "INSIGHTS": DARK,
    "DIFERENCIADORES": PURPLE,
}

DISCIPLINE_TABS = {
    "ESTRUTURAL": {
        "title": "Indicadores Estruturais",
        "prefixes": ["concreto_", "aco_", "forma_", "escoramento_", "taxa_aco_", "estaca", "movimento_terra", "estucamento"],
    },
    "ELETRICA": {
        "title": "Indicadores Eletricos e Eletronicos",
        "prefixes": ["pontos_iluminacao_", "tomadas_", "quadros_eletricos_", "luminarias_", "eletroduto_",
                     "pontos_eletricos_", "ar_condicionado", "carro_eletrico", "portao_automacao"],
    },
    "HIDRO": {
        "title": "Indicadores Hidrossanitarios",
        "prefixes": ["pontos_agua_", "pontos_esgoto_", "tubulacao_total_", "registros_", "ralos_"],
    },
    "ALVENARIA_REVEST": {
        "title": "Alvenaria, Revestimentos e Pintura",
        "prefixes": ["alvenaria_", "blocos_", "chapisco_", "reboco_", "contrapiso_", "porcelanato_",
                     "revestimento_fachada_", "pintura_", "rodape_", "piso_vinilico_", "seladora_", "fachada_"],
    },
    "ESQUADRIAS_LOUCAS": {
        "title": "Esquadrias e Loucas/Metais",
        "prefixes": ["portas_", "janelas_", "vidros_", "guarda_corpo_", "bacias_", "lavatorios_",
                     "cubas_", "chuveiros_", "contramarco_", "loucas_", "esquadrias_"],
    },
    "DIVERSOS": {
        "title": "Diversos (Forro, Imper, Cobertura, Elevador)",
        "prefixes": ["forro_", "elevadores_", "hidrantes_", "sprinklers_", "cobertura_", "drywall_",
                     "manta_", "cristalizacao_", "piscina", "acabamentos_"],
    },
}


def _load_json(p: Path) -> dict:
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def style_header(cell, color=DARK):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Arial")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def style_subheader(cell, color=ACCENT):
    cell.font = Font(bold=True, color=color, size=10, name="Arial")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_headers(ws, row, headers, widths, color=DARK):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row, i, h)
        style_header(c, color)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(ws, row, values, number_formats=None):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row, i, v)
        c.font = Font(size=9, name="Arial")
        c.border = THIN
        if number_formats and i in number_formats:
            c.number_format = number_formats[i]
        elif isinstance(v, float):
            c.number_format = "#,##0.00"
        elif isinstance(v, int):
            c.number_format = "#,##0"


def prettify_ind(name: str) -> str:
    return name.replace("_por_", "/").replace("_m2_", " m2/").replace("_", " ")


def get_indicators_for_tab(all_names, prefixes):
    return [n for n in all_names if any(n.startswith(p) for p in prefixes)]


# ═══════════════════════════════════════════════════════════════════════
# ABA: LEIA_ME
# ═══════════════════════════════════════════════════════════════════════
def build_leia_me(wb, data):
    ws = wb.create_sheet("LEIA_ME")
    ws.sheet_properties.tabColor = ABA_COLORS["LEIA_ME"]

    ws["A1"] = "ANALISE DE PRODUTO - Indicadores Quantitativos Cartesian"
    ws["A1"].font = Font(bold=True, size=14, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")

    ws["A2"] = f"Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')} . {data['n_projetos_total']} projetos . {data['n_indicadores_unicos']} indicadores"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:E2")

    sections = [
        ("PARA QUE SERVE",
         "Permite avaliar a qualidade do PRODUTO (empreendimento) comparando indicadores "
         "fisicos (un/m, m/m AC, kg/m AC) contra benchmarks de projetos similares da base Cartesian. "
         "Exemplo: 'esse projeto tem 42 pontos de iluminacao por UR - e alto ou baixo pra alto padrao?' "
         "Agora com 49 indicadores, outliers automaticos, curva ABC, PUs validados e correlacoes."),
        ("LEGENDA DE CORES", None),
    ]
    row = 4
    for title, body in sections:
        ws.cell(row, 1, title).font = Font(bold=True, size=11, color=DARK, name="Arial")
        row += 1
        if body:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row, 1, body).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 1).font = Font(size=9, name="Arial")
            row += 2

    legend = [
        ("Verde", "Dentro da faixa normal (p25-p75)", GREEN_FILL),
        ("Amarelo", "Atencao (p10-p25 ou p75-p90)", YELLOW_FILL),
        ("Vermelho", "Fora da faixa (<p10 ou >p90)", RED_FILL),
        ("Azul", "Valor agregado/destaque", BLUE_FILL),
    ]
    for label, desc, fill in legend:
        c = ws.cell(row, 1, label)
        c.fill = fill
        c.font = Font(size=9, name="Arial")
        ws.cell(row, 2, desc).font = Font(size=9, name="Arial")
        row += 1

    row += 1
    ws.cell(row, 1, "MAPA DAS ABAS").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2
    write_headers(ws, row, ["#", "Aba", "Descricao"], [4, 22, 65])
    row += 1
    abas = [
        ("1", "LEIA_ME", "Este guia"),
        ("2", "RESUMO", "Numeros-chave agregados, TL;DR da base"),
        ("3", "PROJETOS", f"{data['n_projetos_total']} projetos com padrao, AC, UR, R$/m2, total"),
        ("4", "CUSTO_POR_MG", "R$/m2 por macrogrupo x padrao (calibracao V2)"),
        ("5", "ESTRUTURAL", "Concreto, aco, forma, estacas, mov terra"),
        ("6", "ELETRICA", "Pontos, tomadas, luminarias, AR, carro eletrico"),
        ("7", "HIDRO", "Pontos agua/esgoto, tubulacao, ralos"),
        ("8", "ALVENARIA_REVEST", "Alvenaria, chapisco/reboco, pisos, pintura, rodape"),
        ("9", "ESQUADRIAS_LOUCAS", "Portas, janelas, vidros, loucas, contramarco"),
        ("10", "DIVERSOS", "Forro, imper, cobertura, elevador, piscina"),
        ("11", "MATRIZ_PROJETOS", "Matriz projeto x todos os indicadores"),
        ("12", "BENCHMARKS", "Faixas (p10/p25/p75/p90) por indicador x padrao"),
        ("13", "OUTLIERS", "Projetos fora da faixa em cada indicador"),
        ("14", "ABC", "Top itens por macrogrupo (Curva ABC cross-projeto)"),
        ("15", "PUS_REFERENCIA", "PUs validados de acabamentos, esquadrias, elevador"),
        ("16", "CORRELACOES", "Correlacoes |r|>0.5 entre indicadores"),
        ("17", "INSIGHTS", "Flags automaticos por projeto (desvios vs mediana padrao)"),
        ("18", "DIFERENCIADORES", "Indicadores que mais diferenciam entre padroes"),
    ]
    for num, aba, desc in abas:
        write_row(ws, row, [num, aba, desc])
        row += 1

    row += 1
    ws.cell(row, 1, "SEGMENTACAO POR PADRAO").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    for pad in data.get("padroes", []):
        n_pad = data.get("indicadores_por_padrao", {}).get(pad, {}).get("n_projetos", 0)
        flag = " (amostra pequena)" if n_pad < 10 else ""
        ws.cell(row, 1, f"- {pad}: {n_pad} projetos{flag}").font = Font(size=9, name="Arial")
        row += 1

    row += 1
    ws.cell(row, 1, "METODOLOGIA v2").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    method = [
        "1. Extracao: classificador local keyword-based sobre itens-detalhados (294k itens, 6k abas unicas).",
        "2. Dedupe cross-aba: descartamos cpu/insumos/canteiro/epcs e deduplica mesmo item em abas diferentes.",
        "3. Merge: sobrescreve com indices-executivo (concreto/aco/forma sao mais confiaveis la).",
        "4. Valida: indicadores /UR so computados quando UR > 0 (evita outliers absurdos).",
        "5. Agregacao: percentis (p10/p25/med/p75/p90), media, DP, CV por padrao.",
        "6. Outliers: identifica projetos > p90 ou < p10.",
        "7. ABC: top itens cross-projeto por macrogrupo (via curva_abc de indices-executivo).",
        "8. Correlacoes: Pearson r entre indicadores com n >= 10.",
    ]
    for m in method:
        ws.cell(row, 1, m).font = Font(size=9, name="Arial")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    row += 1
    ws.cell(row, 1, "COMO REGENERAR").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    cmds = [
        "python -X utf8 scripts/extrair_indicadores_produto.py",
        "python -X utf8 scripts/agregar_indicadores_produto.py",
        "python -X utf8 scripts/gerar_planilha_analise_produto.py",
    ]
    for cmd in cmds:
        c = ws.cell(row, 1, cmd)
        c.font = Font(name="Consolas", size=9)
        c.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40


# ═══════════════════════════════════════════════════════════════════════
# ABA: RESUMO
# ═══════════════════════════════════════════════════════════════════════
def build_resumo(wb, data):
    ws = wb.create_sheet("RESUMO")
    ws.sheet_properties.tabColor = ABA_COLORS["RESUMO"]

    ws["A1"] = "RESUMO EXECUTIVO - Analise de Produto Cartesian"
    ws["A1"].font = Font(bold=True, size=13, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")

    row = 3
    ws.cell(row, 1, "NUMEROS-CHAVE").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 2

    matrix = data.get("projeto_matrix", [])
    ac_total = sum(p.get("ac_m2") or 0 for p in matrix)
    ur_total = sum(p.get("ur") or 0 for p in matrix)
    total_total = sum(p.get("total_r$") or 0 for p in matrix)
    com_ur = sum(1 for p in matrix if (p.get("ur") or 0) > 0)
    com_ac = sum(1 for p in matrix if (p.get("ac_m2") or 0) > 0)

    kvs = [
        ("Projetos na base", f"{data['n_projetos_total']}"),
        ("Indicadores unicos extraidos", f"{data['n_indicadores_unicos']}"),
        ("Area construida total", f"{ac_total:,.0f} m2"),
        ("Unidades residenciais total", f"{ur_total:,}"),
        ("Valor total acumulado (R$)", f"R$ {total_total:,.0f}" if total_total else "N/D"),
        ("Projetos com AC definida", f"{com_ac}/{data['n_projetos_total']}"),
        ("Projetos com UR definida", f"{com_ur}/{data['n_projetos_total']}"),
        ("", ""),
        ("Outliers detectados (indicadores)", f"{len(data.get('outliers_por_indicador', {}))}"),
        ("Correlacoes |r|>0.5", f"{len(data.get('correlacoes', {}))}"),
        ("Macrogrupos na curva ABC", f"{len(data.get('abc_summary', {}))}"),
        ("PUs de referencia validados", f"{len(data.get('pus_stats', {}))}"),
        ("Diferenciadores de padrao (ratio max/min > 1.5)", f"{len(data.get('diferenciadores', []))}"),
    ]
    for k, v in kvs:
        if k:
            c1 = ws.cell(row, 1, k)
            c1.font = Font(bold=True, size=10, name="Arial")
            c2 = ws.cell(row, 2, v)
            c2.font = Font(size=10, name="Arial")
        row += 1

    row += 2
    ws.cell(row, 1, "PROJETOS POR PADRAO").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    write_headers(ws, row, ["Padrao", "N projetos", "% do total", "Indicadores /projeto (med)"], [16, 14, 14, 26])
    row += 1
    total = data['n_projetos_total']
    for pad in data.get("padroes", []):
        pinfo = data.get("indicadores_por_padrao", {}).get(pad, {})
        n = pinfo.get("n_projetos", 0)
        n_ind = len(pinfo.get("indicadores", {}))
        write_row(ws, row, [pad, n, round(n/total*100, 1) if total else 0, n_ind])
        row += 1

    row += 2
    ws.cell(row, 1, "TOP 10 INDICADORES POR COBERTURA").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    write_headers(ws, row, ["Indicador", "N projetos", "Mediana", "CV"], [40, 10, 14, 8])
    row += 1
    globais = data.get("indicadores_globais", {})
    ranked = sorted(globais.items(), key=lambda x: -x[1]["n"])
    for name, s in ranked[:10]:
        write_row(ws, row, [prettify_ind(name), s["n"], s["mediana"], s["cv"]], {4: "0.000"})
        row += 1

    row += 2
    ws.cell(row, 1, "INDICADORES COM MAIOR VARIABILIDADE (CV > 2.0)").font = Font(bold=True, size=11, color=DARK, name="Arial")
    row += 1
    write_headers(ws, row, ["Indicador", "N", "CV", "Mediana", "Max/Min"], [40, 6, 8, 14, 14])
    row += 1
    high_cv = [(n, s) for n, s in globais.items() if s.get("cv", 0) > 2.0]
    high_cv.sort(key=lambda x: -x[1]["cv"])
    for name, s in high_cv[:10]:
        ratio = round(s["max"]/s["min"], 1) if s["min"] > 0 else None
        write_row(ws, row, [prettify_ind(name), s["n"], s["cv"], s["mediana"], ratio], {3: "0.000", 4: "#,##0.0000"})
        row += 1

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = [42, 14, 14, 26][col-1] if col <= 4 else 15


# ═══════════════════════════════════════════════════════════════════════
# ABA: PROJETOS
# ═══════════════════════════════════════════════════════════════════════
def build_projetos(wb, data):
    ws = wb.create_sheet("PROJETOS")
    ws.sheet_properties.tabColor = ABA_COLORS["PROJETOS"]

    headers = ["Projeto", "Padrao", "AC (m2)", "UR", "Total (R$)", "R$/m2"]
    widths = [32, 14, 14, 8, 18, 12]
    write_headers(ws, 1, headers, widths)

    matrix = data.get("projeto_matrix", [])
    nf = {3: "#,##0.00", 4: "#,##0", 5: "#,##0", 6: "#,##0.00"}
    for i, p in enumerate(sorted(matrix, key=lambda x: (x.get("padrao") or "", -(x.get("ac_m2") or 0))), start=2):
        write_row(ws, i, [
            p["slug"],
            p.get("padrao", ""),
            p.get("ac_m2") or None,
            p.get("ur") or None,
            p.get("total_r$") or None,
            p.get("rsm2") or None,
        ], nf)

    ws.auto_filter.ref = f"A1:F{len(matrix) + 1}"
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════
# ABA: CUSTO_POR_MG
# ═══════════════════════════════════════════════════════════════════════
def build_custo_por_mg(wb, data):
    ws = wb.create_sheet("CUSTO_POR_MG")
    ws.sheet_properties.tabColor = ABA_COLORS["CUSTO_POR_MG"]

    cond = _load_json(COND_FILE)
    if not cond:
        ws["A1"] = "Dados de calibracao condicional nao encontrados"
        return

    headers = ["Padrao", "Macrogrupo", "n", "min", "p25", "mediana", "p75", "max", "CV"]
    widths = [14, 35, 6, 12, 12, 12, 12, 12, 8]
    write_headers(ws, 1, headers, widths)

    nf = {i: "#,##0.00" for i in range(4, 9)}
    nf[9] = "0.000"
    row = 2
    por_padrao = cond.get("por_padrao_mg", {})
    for padrao in sorted(por_padrao.keys()):
        mgs = por_padrao[padrao]
        if not isinstance(mgs, dict):
            continue
        for mg_name, s in sorted(mgs.items()):
            if not isinstance(s, dict):
                continue
            write_row(ws, row, [
                padrao, mg_name,
                s.get("n", 0),
                s.get("min"), s.get("p25"), s.get("mediana"),
                s.get("p75"), s.get("max"), s.get("cv"),
            ], nf)
            row += 1

    if row > 2:
        ws.auto_filter.ref = f"A1:I{row - 1}"
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════════
# ABAS: DISCIPLINAS
# ═══════════════════════════════════════════════════════════════════════
def build_discipline_tab(wb, tab_name, tab_config, data):
    ws = wb.create_sheet(tab_name)
    ws.sheet_properties.tabColor = ABA_COLORS.get(tab_name, GRAY)

    ind_names = get_indicators_for_tab(data.get("nomes_indicadores", []), tab_config["prefixes"])
    if not ind_names:
        ws["A1"] = "Nenhum indicador encontrado"
        return

    ws["A1"] = tab_config["title"]
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:J1")

    headers = ["Indicador", "n", "min", "p10", "p25", "mediana", "p75", "p90", "max", "CV"]
    widths = [48, 6, 10, 10, 10, 12, 10, 10, 10, 8]
    nf_stats = {i: "#,##0.0000" for i in range(3, 10)}
    nf_stats[10] = "0.000"

    row = 3
    padroes = data.get("padroes", [])
    for padrao in padroes:
        style_subheader(ws.cell(row, 1, f"> {padrao.upper()}"), ACCENT)
        pad_data = data.get("indicadores_por_padrao", {}).get(padrao, {})
        n_proj = pad_data.get("n_projetos", 0)
        ws.cell(row, 2, f"({n_proj} projetos)").font = Font(italic=True, size=8, color="999999", name="Arial")
        row += 1
        write_headers(ws, row, headers, widths, ACCENT if padrao in ("alto", "medio-alto") else ORANGE)
        row += 1
        pad_indicators = pad_data.get("indicadores", {})
        any_rows = False
        for ind_name in ind_names:
            s = pad_indicators.get(ind_name)
            if not s:
                continue
            any_rows = True
            write_row(ws, row, [
                prettify_ind(ind_name), s.get("n", 0),
                s.get("min"), s.get("p10"), s.get("p25"),
                s.get("mediana"), s.get("p75"), s.get("p90"),
                s.get("max"), s.get("cv"),
            ], nf_stats)
            row += 1
        if not any_rows:
            ws.cell(row, 1, "(sem dados)").font = Font(italic=True, size=8, color="999999", name="Arial")
            row += 1
        row += 1

    row += 1
    style_subheader(ws.cell(row, 1, "> VALORES POR PROJETO"), DARK)
    row += 1

    proj_headers = ["Projeto", "Padrao", "AC (m2)"] + [prettify_ind(n)[:30] for n in ind_names]
    proj_widths = [28, 12, 12] + [14] * len(ind_names)
    write_headers(ws, row, proj_headers, proj_widths, DARK)
    row += 1

    nf_proj = {i: "#,##0.0000" for i in range(4, 4 + len(ind_names))}
    nf_proj[3] = "#,##0.00"

    matrix = data.get("projeto_matrix", [])
    first_row = row
    for p in sorted(matrix, key=lambda x: (x.get("padrao") or "", -(x.get("ac_m2") or 0))):
        vals = [p["slug"], p.get("padrao", ""), p.get("ac_m2")]
        for ind_name in ind_names:
            vals.append(p.get(ind_name))
        write_row(ws, row, vals, nf_proj)
        row += 1
    last_row = row - 1

    globais = data.get("indicadores_globais", {})
    for col_offset, ind_name in enumerate(ind_names):
        col_letter = get_column_letter(4 + col_offset)
        s = globais.get(ind_name)
        if not s:
            continue
        data_range = f"{col_letter}{first_row}:{col_letter}{last_row}"
        p25, p75, p10, p90 = s.get("p25", 0), s.get("p75", 0), s.get("p10", 0), s.get("p90", 0)
        if p25 and p75:
            ws.conditional_formatting.add(data_range, CellIsRule(
                operator="between", formula=[str(p25), str(p75)], fill=GREEN_FILL))
        if p10:
            ws.conditional_formatting.add(data_range, CellIsRule(
                operator="lessThan", formula=[str(p10)], fill=RED_FILL))
        if p90:
            ws.conditional_formatting.add(data_range, CellIsRule(
                operator="greaterThan", formula=[str(p90)], fill=RED_FILL))

    ws.freeze_panes = "A3"


def build_matriz_projetos(wb, data):
    ws = wb.create_sheet("MATRIZ_PROJETOS")
    ws.sheet_properties.tabColor = ABA_COLORS["MATRIZ_PROJETOS"]
    all_names = data.get("nomes_indicadores", [])
    headers = ["Projeto", "Padrao", "AC (m2)", "UR"] + [prettify_ind(n)[:25] for n in all_names]
    widths = [28, 12, 12, 8] + [13] * len(all_names)
    write_headers(ws, 1, headers, widths)
    nf = {i: "#,##0.0000" for i in range(5, 5 + len(all_names))}
    nf[3] = "#,##0.00"
    nf[4] = "#,##0"
    matrix = data.get("projeto_matrix", [])
    for i, p in enumerate(sorted(matrix, key=lambda x: (x.get("padrao") or "", x["slug"])), start=2):
        vals = [p["slug"], p.get("padrao", ""), p.get("ac_m2"), p.get("ur")]
        for ind_name in all_names:
            vals.append(p.get(ind_name))
        write_row(ws, i, vals, nf)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(matrix) + 1}"
    ws.freeze_panes = "C2"


def build_benchmarks(wb, data):
    ws = wb.create_sheet("BENCHMARKS")
    ws.sheet_properties.tabColor = ABA_COLORS["BENCHMARKS"]
    ws["A1"] = "BENCHMARKS - Faixas por Indicador e Padrao"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:J1")
    headers = ["Indicador", "Padrao", "n", "< p10", "p10-p25", "p25-p75", "p75-p90", "> p90", "Mediana", "CV"]
    widths = [45, 14, 6, 14, 18, 18, 18, 14, 14, 8]
    write_headers(ws, 3, headers, widths)
    nf = {i: "#,##0.0000" for i in (4, 9)}
    nf[10] = "0.000"
    row = 4
    all_names = data.get("nomes_indicadores", [])
    padroes = data.get("padroes", [])
    for ind_name in all_names:
        for padrao in padroes:
            s = data.get("indicadores_por_padrao", {}).get(padrao, {}).get("indicadores", {}).get(ind_name)
            if not s or s.get("n", 0) < 3:
                continue
            p10, p25, p75, p90 = s.get("p10"), s.get("p25"), s.get("p75"), s.get("p90")
            write_row(ws, row, [
                prettify_ind(ind_name), padrao, s["n"],
                p10, f"{p10}-{p25}", f"{p25}-{p75}", f"{p75}-{p90}", p90,
                s.get("mediana"), s.get("cv"),
            ], nf)
            row += 1
    if row > 4:
        ws.auto_filter.ref = f"A3:J{row - 1}"
    ws.freeze_panes = "A4"


def build_outliers(wb, data):
    ws = wb.create_sheet("OUTLIERS")
    ws.sheet_properties.tabColor = ABA_COLORS["OUTLIERS"]
    ws["A1"] = "OUTLIERS - Projetos fora da faixa (p10-p90) por indicador"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")
    headers = ["Indicador", "Tipo", "Projeto", "Valor", "Razao vs p10/p90", "Mediana global"]
    widths = [42, 12, 30, 14, 16, 14]
    write_headers(ws, 3, headers, widths)
    nf = {4: "#,##0.0000", 5: "0.00", 6: "#,##0.0000"}
    row = 4
    globais = data.get("indicadores_globais", {})
    for ind_name, o in data.get("outliers_por_indicador", {}).items():
        med = globais.get(ind_name, {}).get("mediana")
        for item in o.get("acima_p90", [])[:5]:
            c = ws.cell(row, 1, prettify_ind(ind_name)); c.font = Font(size=9, name="Arial"); c.border = THIN
            c = ws.cell(row, 2, "acima p90"); c.font = Font(size=9, name="Arial", color=RED); c.border = THIN
            write_row(ws, row, [None, None, item["slug"], item["valor"], item.get("ratio_p90"), med], nf)
            ws.cell(row, 1, prettify_ind(ind_name))
            ws.cell(row, 2, "acima p90").font = Font(size=9, color=RED, name="Arial")
            row += 1
        for item in o.get("abaixo_p10", [])[:5]:
            write_row(ws, row, [prettify_ind(ind_name), "abaixo p10", item["slug"], item["valor"], item.get("ratio_p10"), med], nf)
            ws.cell(row, 2, "abaixo p10").font = Font(size=9, color=ORANGE, name="Arial")
            row += 1
    if row > 4:
        ws.auto_filter.ref = f"A3:F{row - 1}"
    ws.freeze_panes = "A4"


def build_abc(wb, data):
    ws = wb.create_sheet("ABC")
    ws.sheet_properties.tabColor = ABA_COLORS["ABC"]
    ws["A1"] = "CURVA ABC - Top itens por Macrogrupo (agregado cross-projeto)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:E1")
    headers = ["Macrogrupo", "Descricao item", "N projetos", "Total R$ agregado", "% mediana do MG"]
    widths = [35, 60, 12, 18, 14]
    write_headers(ws, 3, headers, widths)
    nf = {3: "#,##0", 4: "R$ #,##0", 5: "0.0%"}
    row = 4
    for mg, items in data.get("abc_summary", {}).items():
        if not items:
            continue
        for item in items[:10]:
            write_row(ws, row, [mg, item["desc"], item["n_projetos"], item["total_r$_agregado"], item.get("pct_med_do_mg", 0)], nf)
            row += 1
    if row > 4:
        ws.auto_filter.ref = f"A3:E{row - 1}"
    ws.freeze_panes = "A4"


def build_pus(wb, data):
    ws = wb.create_sheet("PUS_REFERENCIA")
    ws.sheet_properties.tabColor = ABA_COLORS["PUS_REFERENCIA"]
    ws["A1"] = "PUs DE REFERENCIA - Acabamentos, Esquadrias, Elevador"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")
    headers = ["PU", "n", "min", "p25", "mediana", "p75", "max", "CV"]
    widths = [32, 6, 12, 12, 12, 12, 12, 8]
    write_headers(ws, 3, headers, widths)
    nf = {i: "R$ #,##0.00" for i in range(3, 8)}
    nf[8] = "0.000"
    row = 4
    for pu_name, s in data.get("pus_stats", {}).items():
        write_row(ws, row, [
            pu_name, s.get("n"), s.get("min"), s.get("p25"),
            s.get("mediana"), s.get("p75"), s.get("max"), s.get("cv"),
        ], nf)
        row += 1
    if row > 4:
        ws.auto_filter.ref = f"A3:H{row - 1}"
    ws.freeze_panes = "A4"


def build_correlacoes(wb, data):
    ws = wb.create_sheet("CORRELACOES")
    ws.sheet_properties.tabColor = ABA_COLORS["CORRELACOES"]
    ws["A1"] = "CORRELACOES entre Indicadores (|r| > 0.5, n >= 10)"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:D1")
    ws["A2"] = "r > 0: indicadores sobem juntos | r < 0: um sobe, outro cai | |r| > 0.8: forte"
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:D2")
    headers = ["Par de indicadores", "r", "n", "Interpretacao"]
    widths = [60, 10, 8, 40]
    write_headers(ws, 4, headers, widths)
    row = 5
    for pair, info in data.get("correlacoes", {}).items():
        r = info["r"]
        if abs(r) > 0.8:
            interp = "Forte - indicadores quase determinam um ao outro"
        elif abs(r) > 0.6:
            interp = "Media - relacao consistente"
        else:
            interp = "Leve - relacao perceptivel"
        if r < 0:
            interp += " (inversa)"
        write_row(ws, row, [pair, r, info["n"], interp], {2: "0.000"})
        row += 1
    if row > 5:
        ws.auto_filter.ref = f"A4:D{row - 1}"
    ws.freeze_panes = "A5"


def build_insights(wb, data):
    ws = wb.create_sheet("INSIGHTS")
    ws.sheet_properties.tabColor = ABA_COLORS["INSIGHTS"]
    ws["A1"] = "INSIGHTS AUTOMATICOS - Desvios por projeto vs mediana do padrao"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Para cada projeto, lista indicadores onde esta fora da faixa p10-p90 do seu padrao. "
                "delta_pct mostra quanto o projeto difere da mediana do seu padrao.")
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:F2")

    headers = ["Projeto", "Padrao", "Tipo", "Indicador", "Valor", "delta % vs mediana padrao"]
    widths = [30, 14, 14, 40, 14, 22]
    write_headers(ws, 4, headers, widths)
    row = 5
    for slug, info in data.get("insights_por_projeto", {}).items():
        pad = info.get("padrao", "")
        for f in info.get("top_flags", []):
            tipo_txt = "ACIMA p90" if f["tipo"] == "fora_alto" else "ABAIXO p10"
            write_row(ws, row, [slug, pad, tipo_txt, prettify_ind(f["indicador"]),
                                f["valor"], f.get("delta_pct")],
                     {5: "#,##0.0000", 6: "+0.0;-0.0;0"})
            # Color
            c = ws.cell(row, 3)
            c.font = Font(size=9, color=RED if f["tipo"] == "fora_alto" else ORANGE, bold=True, name="Arial")
            row += 1
    if row > 5:
        ws.auto_filter.ref = f"A4:F{row - 1}"
    ws.freeze_panes = "A5"


def build_diferenciadores(wb, data):
    ws = wb.create_sheet("DIFERENCIADORES")
    ws.sheet_properties.tabColor = ABA_COLORS["DIFERENCIADORES"]
    ws["A1"] = "DIFERENCIADORES DE PADRAO - Indicadores que mais mudam entre padroes"
    ws["A1"].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws.merge_cells("A1:H1")
    ws["A2"] = "Ratio max/min entre padroes (quanto mais alto, mais o indicador distingue padrao construtivo)."
    ws["A2"].font = Font(italic=True, size=9, color="666666", name="Arial")
    ws.merge_cells("A2:H2")

    diff = data.get("diferenciadores", [])
    if not diff:
        return
    # Descobre padrões
    first = diff[0]
    padroes = list(first.get("medianas_por_padrao", {}).keys())
    headers = ["Indicador", "Ratio max/min"] + [f"Mediana {p}" for p in padroes]
    widths = [45, 14] + [14] * len(padroes)
    write_headers(ws, 4, headers, widths)
    row = 5
    nf = {2: "0.00"}
    for i, p in enumerate(padroes, start=3):
        nf[i] = "#,##0.0000"
    for d in diff:
        vals = [prettify_ind(d["indicador"]), d["ratio_max_min"]]
        for p in padroes:
            vals.append(d["medianas_por_padrao"].get(p))
        write_row(ws, row, vals, nf)
        row += 1
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = "A5"


def main():
    print("Loading aggregated data...", flush=True)
    data = _load_json(AGG_FILE)
    if not data:
        print(f"ERROR: {AGG_FILE} not found")
        return

    print(f"  {data['n_projetos_total']} projects, {data['n_indicadores_unicos']} indicators", flush=True)

    wb = Workbook()
    wb.remove(wb.active)

    print("Building tabs...", flush=True)
    build_leia_me(wb, data); print("  LEIA_ME")
    build_resumo(wb, data); print("  RESUMO")
    build_projetos(wb, data); print("  PROJETOS")
    build_custo_por_mg(wb, data); print("  CUSTO_POR_MG")
    for tab_name, tab_config in DISCIPLINE_TABS.items():
        build_discipline_tab(wb, tab_name, tab_config, data)
        print(f"  {tab_name}")
    build_matriz_projetos(wb, data); print("  MATRIZ_PROJETOS")
    build_benchmarks(wb, data); print("  BENCHMARKS")
    build_outliers(wb, data); print("  OUTLIERS")
    build_abc(wb, data); print("  ABC")
    build_pus(wb, data); print("  PUS_REFERENCIA")
    build_correlacoes(wb, data); print("  CORRELACOES")
    build_insights(wb, data); print("  INSIGHTS")
    build_diferenciadores(wb, data); print("  DIFERENCIADORES")

    wb.save(str(OUT_XLSX))
    print(f"\nSalvo: {OUT_XLSX}", flush=True)


if __name__ == "__main__":
    main()
