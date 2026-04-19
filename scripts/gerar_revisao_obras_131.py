#!/usr/bin/env python3
"""Gera planilha REVISAO-OBRAS-131.xlsx a partir de projetos-enriquecidos.

Planilha com 2 abas:
- Revisão Obras: linha por projeto com metadados + amarelo pra campos que precisam revisão
- Resumo: contagens por status (cidade, padrão, tipologia, data_base)

Output:
- analises-cross-projeto/REVISAO-OBRAS-131.xlsx
- ~/orcamentos/parametricos/REVISAO-OBRAS-131.xlsx (cópia)
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = Path.home() / "orcamentos-openclaw" / "base"
ENR = BASE / "projetos-enriquecidos.json"
OUT_REPO = Path.home() / "orcamentos-openclaw" / "analises-cross-projeto" / "REVISAO-OBRAS-131.xlsx"
OUT_DRIVE = Path.home() / "orcamentos" / "parametricos" / "REVISAO-OBRAS-131.xlsx"

AMARELO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
VERDE = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
VERMELHO = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")


def main():
    enr = json.loads(ENR.read_text(encoding="utf-8"))
    ENR_DIR = BASE / "projetos-enriquecidos"

    # Merge individual files (fonte autoritativa por campo)
    for p in enr["projetos"]:
        ind_f = ENR_DIR / f"{p['slug']}.json"
        if ind_f.exists():
            try:
                ind = json.loads(ind_f.read_text(encoding="utf-8"))
            except Exception:
                continue
            # Campos autoritativos do individual
            for k in [
                "cidade", "uf", "cub_regiao", "cidade_fonte",
                "tipologia_canonica", "tipologia_confianca", "tipologia_motivo", "tipologia_fonte",
                "padrao", "padrao_canonico", "padrao_confianca", "padrao_motivo", "padrao_fonte", "padrao_anterior",
                "data_base", "data_entrega", "data_base_fonte", "data_base_confianca", "data_base_motivo",
                "cub_valor_entrega",
            ]:
                v = ind.get(k)
                if v is not None and v != "":
                    p[k] = v

    projs = sorted(enr["projetos"], key=lambda p: p["slug"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Revisão Obras"

    headers = [
        "Slug", "Cliente", "Cidade", "UF", "CUB Região",
        "Padrão", "Tipologia", "Data Base", "Data Entrega",
        "AC (m²)", "UR", "R$/m²", "Total",
        "Fonte cidade", "Conf. tipologia",
        "Fonte padrão", "Fonte data_base", "Motivo tipologia", "Motivo padrão",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    contadores = {
        "cidade_gt": 0, "cidade_manual": 0, "cidade_mapa": 0, "cidade_bug_fix": 0, "cidade_qwen": 0, "cidade_outra": 0,
        "padrao_ok": 0, "padrao_qwen": 0, "padrao_problem": 0,
        "tip_alta": 0, "tip_media": 0, "tip_baixa": 0, "tip_sem": 0, "tip_qwen": 0,
        "db_gt": 0, "db_ind": 0, "db_qwen": 0, "db_sem": 0,
    }

    for p in projs:
        slug = p["slug"]
        cidade = p.get("cidade") or ""
        uf = p.get("uf") or ""
        cub = p.get("cub_regiao") or ""
        padrao = p.get("padrao") or ""
        tip = p.get("tipologia_canonica") or ""
        db = p.get("data_base") or ""
        de = p.get("data_entrega") or ""
        ac = p.get("ac_m2") or ""
        ur = p.get("ur") or ""
        rsm2 = p.get("rsm2") or ""
        total = p.get("total_rs") or ""
        fonte_cid = p.get("cidade_fonte") or ""
        conf_tip = p.get("tipologia_confianca") or ""
        fonte_pad = p.get("padrao_fonte") or ""
        fonte_db = p.get("data_base_fonte") or ""
        motivo_tip = (p.get("tipologia_motivo") or "")[:200]
        motivo_pad = (p.get("padrao_motivo") or "")[:200]

        row = [
            slug, p.get("cliente_inferido") or "",
            cidade, uf, cub,
            padrao, tip, db, de,
            ac, ur, rsm2, total,
            fonte_cid, conf_tip, fonte_pad, fonte_db, motivo_tip, motivo_pad,
        ]
        ws.append(row)
        r = ws.max_row

        # Destacar pendências
        pad_norm = padrao.lower().replace("_", "-")
        if pad_norm in ("null", "desconhecido", "insuficiente", "medio_alto", ""):
            ws.cell(row=r, column=6).fill = VERMELHO
            contadores["padrao_problem"] += 1
        elif fonte_pad == "qwen_revisao":
            ws.cell(row=r, column=6).fill = VERDE
            contadores["padrao_qwen"] += 1
        else:
            contadores["padrao_ok"] += 1

        if conf_tip == "media":
            ws.cell(row=r, column=7).fill = AMARELO
            contadores["tip_media"] += 1
        elif conf_tip == "alta":
            contadores["tip_alta"] += 1
        elif conf_tip == "baixa":
            ws.cell(row=r, column=7).fill = AMARELO
            contadores["tip_baixa"] += 1
        elif not conf_tip:
            ws.cell(row=r, column=7).fill = VERMELHO
            contadores["tip_sem"] += 1
        if p.get("tipologia_fonte") == "qwen_revisao":
            contadores["tip_qwen"] += 1

        if not db:
            ws.cell(row=r, column=8).fill = AMARELO
            contadores["db_sem"] += 1
        elif fonte_db == "ground_truth_entrega":
            ws.cell(row=r, column=8).fill = VERDE
            contadores["db_gt"] += 1
        elif fonte_db == "indices_executivo_pdf_metadata":
            contadores["db_ind"] += 1
        elif fonte_db == "qwen_inferencia":
            ws.cell(row=r, column=8).fill = VERDE
            contadores["db_qwen"] += 1

        # Fonte cidade
        if fonte_cid == "ground_truth_entrega":
            contadores["cidade_gt"] += 1
        elif fonte_cid == "revisao_manual_leo":
            contadores["cidade_manual"] += 1
        elif fonte_cid == "mapa_manual":
            contadores["cidade_mapa"] += 1
        elif fonte_cid == "correcao_bug":
            contadores["cidade_bug_fix"] += 1
        elif fonte_cid == "qwen_inferencia":
            contadores["cidade_qwen"] += 1
        else:
            contadores["cidade_outra"] += 1

    # Larguras
    widths = {
        1: 38, 2: 18, 3: 22, 4: 5, 5: 20, 6: 14, 7: 36, 8: 12, 9: 12,
        10: 12, 11: 6, 12: 12, 13: 16, 14: 24, 15: 14, 16: 18, 17: 32, 18: 40, 19: 40,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    # ===== Aba Resumo =====
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["RESUMO DA REVISÃO"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.append([])

    total = len(projs)
    ws2.append(["Total projetos", total])
    ws2.append([])

    ws2.append(["CIDADE (fonte)", ""])
    ws2.append(["  ground_truth_entrega (capa da entrega)", contadores["cidade_gt"]])
    ws2.append(["  revisao_manual_leo", contadores["cidade_manual"]])
    ws2.append(["  mapa_manual (inferência de cliente)", contadores["cidade_mapa"]])
    ws2.append(["  correcao_bug", contadores["cidade_bug_fix"]])
    ws2.append(["  qwen_inferencia", contadores["cidade_qwen"]])
    ws2.append(["  (outra/vazio)", contadores["cidade_outra"]])
    ws2.append([])

    ws2.append(["PADRÃO", ""])
    ws2.append(["  OK (mapa/inferência)", contadores["padrao_ok"]])
    ws2.append(["  Revisado via Qwen", contadores["padrao_qwen"]])
    ws2.append(["  Problemático (null/desconhecido)", contadores["padrao_problem"]])
    ws2.append([])

    ws2.append(["TIPOLOGIA", ""])
    ws2.append(["  Confiança alta", contadores["tip_alta"]])
    ws2.append(["  Confiança média (amarelo)", contadores["tip_media"]])
    ws2.append(["  Confiança baixa (amarelo)", contadores["tip_baixa"]])
    ws2.append(["  Sem classificação (vermelho)", contadores["tip_sem"]])
    ws2.append(["  — destes, revisados via Qwen", contadores["tip_qwen"]])
    ws2.append([])

    ws2.append(["DATA BASE", ""])
    ws2.append(["  Ground truth entrega", contadores["db_gt"]])
    ws2.append(["  Indices-executivo pdf_metadata", contadores["db_ind"]])
    ws2.append(["  Qwen inferência", contadores["db_qwen"]])
    ws2.append(["  Sem data (amarelo)", contadores["db_sem"]])

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 12

    wb.save(OUT_REPO)
    shutil.copy(OUT_REPO, OUT_DRIVE)
    print(f"Salvo: {OUT_REPO}")
    print(f"Copiado: {OUT_DRIVE}")
    print()
    for k, v in contadores.items():
        print(f"  {k:<30} {v}")


if __name__ == "__main__":
    main()
