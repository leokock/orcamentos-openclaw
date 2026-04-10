#!/usr/bin/env python3
"""
Extract INSTALACOES BREAKDOWN from construction budget xlsx files.
Processes projects 0-41 from _all_projects_mapping.json.
Outputs: updates each slug.json in indices-executivo/ with instalacoes_breakdown and split_mo_material.
"""
import json
import os
import re
import sys
import traceback
from pathlib import Path

import openpyxl

# Paths
BASE_DIR = Path("C:/Users/leona/orcamentos-openclaw/base")
MAPPING_FILE = BASE_DIR / "_all_projects_mapping.json"
INDICES_DIR = BASE_DIR / "indices-executivo"

# Load mapping
with open(MAPPING_FILE, "r", encoding="utf-8") as f:
    all_projects = json.load(f)

# We process projects 0-41
projects = all_projects[0:42]

def normalize(s):
    """Normalize string for comparison: remove accents, lowercase, strip."""
    if s is None:
        return ""
    s = str(s).strip()
    # Common replacements for encoding issues
    replacements = {
        '\ufffd': '', '\x00': '',
        'Ã£': 'a', 'Ã¡': 'a', 'Ã©': 'e', 'Ã­': 'i', 'Ã³': 'o', 'Ãº': 'u',
        'Ã§': 'c', 'Ã': 'A', 'Ê': 'E',
        'á': 'a', 'à': 'a', 'ã': 'a', 'â': 'a',
        'é': 'e', 'ê': 'e', 'è': 'e',
        'í': 'i', 'ì': 'i', 'î': 'i',
        'ó': 'o', 'ô': 'o', 'õ': 'o', 'ò': 'o',
        'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c', 'Ç': 'C',
        'Á': 'A', 'À': 'A', 'Ã': 'A', 'Â': 'A',
        'É': 'E', 'Ê': 'E', 'È': 'E',
        'Í': 'I', 'Ì': 'I', 'Î': 'I',
        'Ó': 'O', 'Ô': 'O', 'Õ': 'O', 'Ò': 'O',
        'Ú': 'U', 'Ù': 'U', 'Û': 'U', 'Ü': 'U',
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s.lower().strip()

def is_numeric(v):
    """Check if value is a number."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(',', '.'))
        return True
    except (ValueError, TypeError):
        return False

def to_float(v):
    """Convert value to float."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return 0.0

def classify_discipline(text):
    """Classify a text into one of the 5 installation disciplines."""
    n = normalize(text)
    # Order matters - check more specific patterns first
    if any(k in n for k in ['preventiv', 'ppci', 'hidrante', 'sprinkler', 'alarme incendio', 'protecao contra incendio', 'spda', 'aterramento', 'para-raio']):
        return 'preventivas'
    if any(k in n for k in ['glp', 'gas', 'gn ', 'g.l.p', 'tubulacao cobre', 'rede primaria', 'rede secundaria', 'medidor']):
        # Check it's gas-specific, not generic
        if 'elet' not in n and 'hidro' not in n:
            return 'gas'
    if any(k in n for k in ['telecom', 'comunicac', 'interfon', 'cftv', 'rede estruturad', 'tv e internet', 'cabeamento', 'fibra optic', 'telefon']):
        return 'telecom'
    if any(k in n for k in ['eletric', 'eletrodut', 'eletrocalh', 'fiacao', 'cabos', 'quadro', 'disjuntor', 'acabamento eletric', 'iluminac', 'entrada de energia']):
        return 'eletricas'
    if any(k in n for k in ['hidrossanit', 'hidrosanitari', 'hidraulic', 'agua fria', 'agua quente', 'esgoto', 'pluvial', 'sanitari', 'loucas e met', 'metais', 'drenag']):
        return 'hidrossanitarias'
    # Broader pattern for section headers
    if 'eletric' in n:
        return 'eletricas'
    if 'hidro' in n or 'hidra' in n:
        return 'hidrossanitarias'
    return None

def is_mo_line(text):
    """Check if line is a 'Mao de Obra' line."""
    n = normalize(text)
    return any(k in n for k in ['mao de obra', 'mão de obra', 'm.o.', 'empreitada', 'mo para instal', 'mo instal'])

def is_total_line(text):
    """Check if line is a TOTAL line."""
    n = normalize(text)
    return n.startswith('total') or n == 'subtotal' or 'total instal' in n

def extract_from_instalacoes_sheet(wb):
    """
    Extract from a dedicated INSTALACOES sheet (Cartesian format).
    These sheets have sections: INSTALACOES ELETRICAS, HIDROSSANITARIAS, PREVENTIVAS E GLP, etc.
    """
    # Find the INSTALACOES sheet
    target = None
    for name in wb.sheetnames:
        n = normalize(name)
        if 'instalac' in n or 'instalac' in n.replace('\ufffd', ''):
            target = name
            break

    if target is None:
        return None, None

    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))

    breakdown = {}
    mo_splits = {}
    current_discipline = None
    current_total = 0.0
    current_mo = 0.0
    current_material = 0.0

    for i, row in enumerate(rows):
        if not any(v is not None for v in row):
            continue

        cell0 = str(row[0]) if row[0] is not None else ""
        cell0_norm = normalize(cell0)

        # Detect section headers
        if 'instalacoes eletricas' in cell0_norm or 'sistemas e instalacoes eletricas' in cell0_norm:
            # Save previous discipline if any
            if current_discipline and current_total > 0:
                breakdown[current_discipline] = current_total
                if current_mo > 0:
                    mo_splits[current_discipline] = {
                        'mo_valor': round(current_mo, 2),
                        'material_valor': round(current_total - current_mo, 2),
                        'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
                    }
            current_discipline = 'eletricas'
            current_total = 0.0
            current_mo = 0.0
            continue

        if any(k in cell0_norm for k in ['instalacoes hidrossanit', 'instalacoes hidrosanitari', 'instalacoes hidraulic']):
            if current_discipline and current_total > 0:
                breakdown[current_discipline] = current_total
                if current_mo > 0:
                    mo_splits[current_discipline] = {
                        'mo_valor': round(current_mo, 2),
                        'material_valor': round(current_total - current_mo, 2),
                        'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
                    }
            current_discipline = 'hidrossanitarias'
            current_total = 0.0
            current_mo = 0.0
            continue

        if any(k in cell0_norm for k in ['instalacoes preventiv', 'instalacoes de protec', 'preventivas e glp']):
            if current_discipline and current_total > 0:
                breakdown[current_discipline] = current_total
                if current_mo > 0:
                    mo_splits[current_discipline] = {
                        'mo_valor': round(current_mo, 2),
                        'material_valor': round(current_total - current_mo, 2),
                        'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
                    }
            current_discipline = 'preventivas'
            current_total = 0.0
            current_mo = 0.0
            continue

        if any(k in cell0_norm for k in ['instalacoes de glp', 'instalacoes de gas', 'instalacoes gas']):
            if current_discipline and current_total > 0:
                breakdown[current_discipline] = current_total
                if current_mo > 0:
                    mo_splits[current_discipline] = {
                        'mo_valor': round(current_mo, 2),
                        'material_valor': round(current_total - current_mo, 2),
                        'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
                    }
            current_discipline = 'gas'
            current_total = 0.0
            current_mo = 0.0
            continue

        if any(k in cell0_norm for k in ['instalacoes de comunicac', 'instalacoes telecom', 'telecomunicac']):
            if current_discipline and current_total > 0:
                breakdown[current_discipline] = current_total
                if current_mo > 0:
                    mo_splits[current_discipline] = {
                        'mo_valor': round(current_mo, 2),
                        'material_valor': round(current_total - current_mo, 2),
                        'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
                    }
            current_discipline = 'telecom'
            current_total = 0.0
            current_mo = 0.0
            continue

        # Look for TOTAL line to capture section total
        if current_discipline and is_total_line(cell0):
            # Find the value - usually in column D (index 3)
            val = None
            for ci in range(1, min(len(row), 10)):
                if is_numeric(row[ci]) and to_float(row[ci]) > 100:
                    val = to_float(row[ci])
                    break
            if val and val > current_total:
                current_total = val
            continue

        # Within a section, check for MO lines
        if current_discipline and is_mo_line(cell0):
            for ci in range(1, min(len(row), 10)):
                if is_numeric(row[ci]) and to_float(row[ci]) > 100:
                    current_mo += to_float(row[ci])
                    break

        # Check for GLP lines within preventivas section
        if current_discipline == 'preventivas':
            if any(k in cell0_norm for k in ['glp', 'gas', 'g.l.p']):
                # This might need to be separated into 'gas' discipline
                pass  # Keep in preventivas for now, handle post-processing

        # Handle "TOTAL INSTALACOES" - final line
        if 'total instal' in cell0_norm:
            # Save last discipline
            if current_discipline and current_total > 0:
                breakdown[current_discipline] = current_total
                if current_mo > 0:
                    mo_splits[current_discipline] = {
                        'mo_valor': round(current_mo, 2),
                        'material_valor': round(current_total - current_mo, 2),
                        'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
                    }
            break

    # Save final discipline if not yet saved
    if current_discipline and current_total > 0 and current_discipline not in breakdown:
        breakdown[current_discipline] = current_total
        if current_mo > 0:
            mo_splits[current_discipline] = {
                'mo_valor': round(current_mo, 2),
                'material_valor': round(current_total - current_mo, 2),
                'mo_pct': round(current_mo / current_total, 3) if current_total > 0 else 0
            }

    # Post-process: if preventivas includes GLP lines, try to separate
    # (in the Adore example, preventivas+GLP are in the same section)

    return breakdown if breakdown else None, mo_splits if mo_splits else None

def extract_from_separate_sheets(wb):
    """
    Extract from projects with separate sheets per discipline.
    E.g., sheets named HIDROSSANITARIO, ELETRICA, PPCI, etc.
    """
    breakdown = {}
    mo_splits = {}

    sheet_map = {}
    for name in wb.sheetnames:
        n = normalize(name)
        if any(k in n for k in ['hidrossanit', 'hidro_', 'hidrosanitari']):
            sheet_map['hidrossanitarias'] = name
        elif any(k in n for k in ['eletric', 'eletrica_']):
            if 'hidro' not in n:
                sheet_map['eletricas'] = name
        elif any(k in n for k in ['preventiv', 'ppci']):
            sheet_map['preventivas'] = name
        elif any(k in n for k in ['gas', 'glp']):
            if 'elet' not in n and 'hidro' not in n:
                sheet_map['gas'] = name
        elif any(k in n for k in ['telecom', 'comunicac']):
            sheet_map['telecom'] = name

    if not sheet_map:
        return None, None

    for discipline, sheet_name in sheet_map.items():
        ws = wb[sheet_name]
        total = 0.0
        mo = 0.0

        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            cell0 = str(row[0]) if row[0] is not None else ""
            cell0_norm = normalize(cell0)

            if is_total_line(cell0):
                for ci in range(1, min(len(row), 10)):
                    if is_numeric(row[ci]) and to_float(row[ci]) > 100:
                        v = to_float(row[ci])
                        if v > total:
                            total = v
                        break

            if is_mo_line(cell0):
                for ci in range(1, min(len(row), 10)):
                    if is_numeric(row[ci]) and to_float(row[ci]) > 100:
                        mo += to_float(row[ci])
                        break

        if total > 0:
            breakdown[discipline] = total
            if mo > 0:
                mo_splits[discipline] = {
                    'mo_valor': round(mo, 2),
                    'material_valor': round(total - mo, 2),
                    'mo_pct': round(mo / total, 3) if total > 0 else 0
                }

    return breakdown if breakdown else None, mo_splits if mo_splits else None

def extract_from_resumo_sheet(wb):
    """
    Extract from Resumo/Ger_Executivo/summary sheets.
    These have one line per discipline with the total value.
    """
    # Find summary sheet
    target = None
    for name in wb.sheetnames:
        n = normalize(name)
        if n in ['resumo', 'ger_executivo', 'gerenciamento executivo', 'gerenciamento', 'apresentacao', 'apresentac']:
            target = name
            break
    if target is None:
        # Try sheets containing 'resumo' or 'apresenta'
        for name in wb.sheetnames:
            n = normalize(name)
            if 'resumo' in n or 'apresenta' in n:
                target = name
                break

    if target is None:
        return None, None

    ws = wb[target]
    breakdown = {}

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        # Check each cell for discipline keywords
        for ci, cell in enumerate(row):
            if cell is None:
                continue
            cell_norm = normalize(str(cell))

            discipline = None
            if any(k in cell_norm for k in ['sistemas e instalacoes eletricas', 'instalacoes eletricas']):
                if 'hidro' not in cell_norm:
                    discipline = 'eletricas'
            elif any(k in cell_norm for k in ['instalacoes hidrossanit', 'sistemas e instalacoes hidrossanit', 'instalacoes hidrosanitari', 'instalacoes hidraulic']):
                discipline = 'hidrossanitarias'
            elif any(k in cell_norm for k in ['instalacoes preventiv', 'preventiv']):
                if 'medicamento' not in cell_norm and 'laudo' not in cell_norm and 'manutencao' not in cell_norm:
                    discipline = 'preventivas'
            elif any(k in cell_norm for k in ['instalacoes de glp', 'instalacoes de gas', 'instalacoes gas', 'instalacoes glp']):
                discipline = 'gas'
            elif any(k in cell_norm for k in ['instalacoes de comunicac', 'instalacoes telecom', 'telecomunicac', 'comunicacoes']):
                if 'visual' not in cell_norm:
                    discipline = 'telecom'

            if discipline:
                # Find the value - look in nearby columns
                val = None
                for vi in range(ci+1, min(len(row), ci+5)):
                    if vi < len(row) and is_numeric(row[vi]) and to_float(row[vi]) > 100:
                        val = to_float(row[vi])
                        break
                # Also check the column before
                if val is None and ci > 0:
                    if is_numeric(row[ci-1]) and to_float(row[ci-1]) > 100:
                        val = to_float(row[ci-1])

                if val:
                    breakdown[discipline] = val
                break

    return breakdown if breakdown else None, None  # No MO split from summary

def extract_from_relatorio_sheet(wb):
    """
    Extract from Sienge 'Relatorio' sheet using hierarchical codes.
    Installation sections typically start with codes like 05, 06, 07...
    """
    target = None
    for name in wb.sheetnames:
        n = normalize(name)
        if 'relat' in n:
            target = name
            break

    if target is None:
        return None, None

    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))

    breakdown = {}
    mo_splits = {}

    # First pass: find the main installation section codes and their totals
    section_codes = {}  # code -> (discipline, total)

    for i, row in enumerate(rows):
        if not any(v is not None for v in row):
            continue

        cell0 = str(row[0]).strip() if row[0] is not None else ""
        cell1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        cell1_norm = normalize(cell1)

        # Match N1 level sections (e.g., "05 ", "06 ")
        code_match = re.match(r'^(\d{2})\s*$', cell0)
        if code_match:
            code = code_match.group(1)
            discipline = None

            if any(k in cell1_norm for k in ['instalacoes eletricas', 'sistemas e instalacoes eletricas']):
                if 'hidro' not in cell1_norm:
                    discipline = 'eletricas'
            elif any(k in cell1_norm for k in ['hidrossanit', 'hidrosanitari', 'hidraulic']):
                discipline = 'hidrossanitarias'
            elif any(k in cell1_norm for k in ['preventiv', 'ppci', 'protecao contra incendio']):
                discipline = 'preventivas'
            elif any(k in cell1_norm for k in ['glp', ' gas', 'instalacoes de gas']):
                discipline = 'gas'
            elif any(k in cell1_norm for k in ['telecom', 'comunicac']):
                if 'visual' not in cell1_norm:
                    discipline = 'telecom'

            if discipline:
                section_codes[code] = discipline

        # Also try N1-level codes like "05" in format "05 " with description
        code_match2 = re.match(r'^(\d{2})\s+$', cell0)
        if code_match2:
            code = code_match2.group(1)
            if code in section_codes:
                # Find total value
                for ci in range(2, min(len(row), 15)):
                    if is_numeric(row[ci]) and to_float(row[ci]) > 100:
                        breakdown[section_codes[code]] = to_float(row[ci])
                        break

    # Second pass: if we found sections but no totals, sum up items
    if section_codes and not breakdown:
        current_discipline = None
        current_total = 0.0
        current_mo = 0.0

        for row in rows:
            if not any(v is not None for v in row):
                continue

            cell0 = str(row[0]).strip() if row[0] is not None else ""
            cell1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

            # Detect section change
            code_match = re.match(r'^(\d{2})\s*$', cell0)
            if code_match:
                code = code_match.group(1)
                if code in section_codes:
                    if current_discipline and current_total > 0:
                        breakdown[current_discipline] = current_total
                        if current_mo > 0:
                            mo_splits[current_discipline] = {
                                'mo_valor': round(current_mo, 2),
                                'material_valor': round(current_total - current_mo, 2),
                                'mo_pct': round(current_mo / current_total, 3)
                            }
                    current_discipline = section_codes[code]
                    current_total = 0.0
                    current_mo = 0.0
                    continue
                elif current_discipline:
                    # New N1 section that's not installation -> save current
                    if current_total > 0:
                        breakdown[current_discipline] = current_total
                        if current_mo > 0:
                            mo_splits[current_discipline] = {
                                'mo_valor': round(current_mo, 2),
                                'material_valor': round(current_total - current_mo, 2),
                                'mo_pct': round(current_mo / current_total, 3)
                            }
                    current_discipline = None
                    current_total = 0.0
                    current_mo = 0.0

            # Within section, accumulate values from leaf items
            if current_discipline:
                # Look for value columns
                for ci in range(2, min(len(row), 15)):
                    if is_numeric(row[ci]) and to_float(row[ci]) > 0:
                        val = to_float(row[ci])
                        if val > current_total:
                            current_total = val
                        break

                if is_mo_line(cell1):
                    for ci in range(2, min(len(row), 15)):
                        if is_numeric(row[ci]) and to_float(row[ci]) > 0:
                            current_mo += to_float(row[ci])
                            break

        # Save last discipline
        if current_discipline and current_total > 0:
            breakdown[current_discipline] = current_total
            if current_mo > 0:
                mo_splits[current_discipline] = {
                    'mo_valor': round(current_mo, 2),
                    'material_valor': round(current_total - current_mo, 2),
                    'mo_pct': round(current_mo / current_total, 3)
                }

    return breakdown if breakdown else None, mo_splits if mo_splits else None

def extract_from_any_sheet(wb):
    """
    Last resort: scan all sheets for installation discipline totals.
    Look for rows with discipline keywords + numeric values.
    """
    breakdown = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue

            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                cell_norm = normalize(str(cell))

                discipline = None
                if any(k in cell_norm for k in ['total instalacoes eletricas', 'total eletric', 'subtotal eletric']):
                    discipline = 'eletricas'
                elif any(k in cell_norm for k in ['total instalacoes hidrossanit', 'total hidrossanit', 'subtotal hidrossanit', 'total hidraulic']):
                    discipline = 'hidrossanitarias'
                elif any(k in cell_norm for k in ['total instalacoes preventiv', 'total preventiv', 'subtotal preventiv', 'total ppci']):
                    discipline = 'preventivas'
                elif any(k in cell_norm for k in ['total instalacoes de glp', 'total glp', 'total gas', 'subtotal glp']):
                    discipline = 'gas'
                elif any(k in cell_norm for k in ['total instalacoes de comunicac', 'total telecom', 'total comunicac', 'subtotal comunicac']):
                    discipline = 'telecom'

                if discipline and discipline not in breakdown:
                    # Find value in nearby columns
                    for vi in range(max(0, ci-2), min(len(row), ci+5)):
                        if vi != ci and is_numeric(row[vi]) and to_float(row[vi]) > 100:
                            breakdown[discipline] = to_float(row[vi])
                            break

    return breakdown if breakdown else None, None

def extract_combined_instalacoes_from_resumo(wb):
    """
    Some projects have 'INSTALACOES ELETRICAS, HIDRAULICAS, GLP E PREVENTIVAS' combined.
    In those cases, also look in deeper sheets for the breakdown.
    But first check if Resumo has individual lines.
    """
    # Find Resumo-like sheet
    target = None
    for name in wb.sheetnames:
        n = normalize(name)
        if 'resumo' in n:
            target = name
            break
    if not target:
        return None, None

    ws = wb[target]
    breakdown = {}
    combined_total = None

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue

        for ci, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            cell_norm = normalize(cell_str)

            # Find value - look in next columns for the first numeric > 100
            def find_val():
                for vi in range(ci+1, min(len(row), ci+6)):
                    if vi < len(row) and is_numeric(row[vi]) and to_float(row[vi]) > 100:
                        return to_float(row[vi])
                # Check col before
                if ci > 0 and is_numeric(row[ci-1]) and to_float(row[ci-1]) > 100:
                    return to_float(row[ci-1])
                return None

            # Individual discipline lines (most detailed)
            if any(k in cell_norm for k in ['sistemas e instalacoes eletricas', 'instalacoes eletricas']):
                if 'hidro' not in cell_norm and 'hidraulic' not in cell_norm:
                    v = find_val()
                    if v:
                        breakdown['eletricas'] = v
            elif any(k in cell_norm for k in ['instalacoes hidrossanit', 'sistemas e instalacoes hidrossanit', 'hidrossanit', 'instalacoes hidraulic']):
                if 'eletric' not in cell_norm:
                    v = find_val()
                    if v:
                        breakdown['hidrossanitarias'] = v
            elif any(k in cell_norm for k in ['instalacoes preventiv']):
                if 'medicamento' not in cell_norm and 'laudo' not in cell_norm and 'manutencao' not in cell_norm:
                    v = find_val()
                    if v:
                        breakdown['preventivas'] = v
            elif any(k in cell_norm for k in ['instalacoes de glp', 'instalacoes glp', 'instalacoes de gas', 'instalacoes gas']):
                v = find_val()
                if v:
                    breakdown['gas'] = v
            elif any(k in cell_norm for k in ['instalacoes de comunicac', 'comunicacoes']):
                if 'visual' not in cell_norm:
                    v = find_val()
                    if v:
                        breakdown['telecom'] = v
            # Combined line
            elif 'instalacoes' in cell_norm and ('eletric' in cell_norm or 'hidraulic' in cell_norm or 'hidro' in cell_norm):
                if 'glp' in cell_norm or 'preventiv' in cell_norm or 'gas' in cell_norm:
                    # This is a combined discipline line
                    v = find_val()
                    if v:
                        combined_total = v

            if cell_norm.startswith('total') and 'instal' in cell_norm:
                break

    # If we have individual lines, return them
    if len(breakdown) >= 2:
        return breakdown, None

    # If only combined, return as-is with a note
    if combined_total and not breakdown:
        breakdown['_combined'] = combined_total
        return breakdown, None

    return breakdown if breakdown else None, None

def process_project(proj):
    """Process a single project and return breakdown + mo_splits."""
    slug = proj['slug']
    path = proj['path']

    # Convert path
    local_path = path

    if not os.path.exists(local_path):
        return slug, None, None, f"File not found: {local_path}"

    try:
        wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)
    except Exception as e:
        return slug, None, None, f"Error opening: {e}"

    breakdown = None
    mo_splits = None

    try:
        # Strategy 1: Dedicated INSTALACOES sheet
        breakdown, mo_splits = extract_from_instalacoes_sheet(wb)
        if breakdown:
            wb.close()
            return slug, breakdown, mo_splits, "instalacoes_sheet"

        # Strategy 2: Separate discipline sheets
        breakdown, mo_splits = extract_from_separate_sheets(wb)
        if breakdown:
            wb.close()
            return slug, breakdown, mo_splits, "separate_sheets"

        # Strategy 3: Resumo/Ger_Executivo with individual lines
        breakdown, mo_splits = extract_combined_instalacoes_from_resumo(wb)
        if breakdown:
            wb.close()
            return slug, breakdown, mo_splits, "resumo_sheet"

        # Strategy 4: Relatorio (Sienge) with hierarchical codes
        breakdown, mo_splits = extract_from_relatorio_sheet(wb)
        if breakdown:
            wb.close()
            return slug, breakdown, mo_splits, "relatorio_sheet"

        # Strategy 5: Any sheet scan
        breakdown, mo_splits = extract_from_any_sheet(wb)
        if breakdown:
            wb.close()
            return slug, breakdown, mo_splits, "any_sheet_scan"

    except Exception as e:
        wb.close()
        return slug, None, None, f"Error processing: {e}\n{traceback.format_exc()}"

    wb.close()
    return slug, None, None, "no_data_found"

# Main processing
results = {}
summary_stats = {
    'total': len(projects),
    'with_breakdown': 0,
    'with_mo_split': 0,
    'by_method': {},
    'errors': [],
    'no_data': []
}

discipline_totals = {
    'eletricas': [], 'hidrossanitarias': [], 'preventivas': [], 'gas': [], 'telecom': []
}
mo_pcts = {
    'eletricas': [], 'hidrossanitarias': [], 'preventivas': [], 'gas': [], 'telecom': []
}

for i, proj in enumerate(projects):
    slug = proj['slug']
    print(f"[{i+1:02d}/{len(projects)}] Processing {slug}...", end=" ", flush=True)

    slug, breakdown, mo_splits, method = process_project(proj)

    if breakdown:
        # Filter out _combined entries for stats
        real_disciplines = {k: v for k, v in breakdown.items() if not k.startswith('_')}

        if real_disciplines:
            summary_stats['with_breakdown'] += 1
            summary_stats['by_method'][method] = summary_stats['by_method'].get(method, 0) + 1

            # Get area from existing JSON
            json_path = INDICES_DIR / f"{slug}.json"
            ac = None
            if json_path.exists():
                with open(json_path, 'r', encoding='utf-8') as f:
                    existing = json.load(f)
                ac = existing.get('ac')

            # Build breakdown with rsm2
            breakdown_with_rsm2 = {}
            for disc, val in real_disciplines.items():
                entry = {'valor': round(val, 2)}
                if ac and ac > 0:
                    entry['rsm2'] = round(val / ac, 2)
                breakdown_with_rsm2[disc] = entry
                discipline_totals[disc].append(val)

            results[slug] = {
                'instalacoes_breakdown': breakdown_with_rsm2,
                'method': method
            }

            if mo_splits:
                summary_stats['with_mo_split'] += 1
                results[slug]['split_mo_material_instalacoes'] = mo_splits
                for disc, split in mo_splits.items():
                    if 'mo_pct' in split:
                        mo_pcts[disc].append(split['mo_pct'])

            print(f"OK ({method}) - {len(real_disciplines)} disciplines")
        else:
            if '_combined' in breakdown:
                print(f"COMBINED ONLY - R$ {breakdown['_combined']:,.0f}")
                results[slug] = {
                    'instalacoes_breakdown': {'_combined': {'valor': round(breakdown['_combined'], 2)}},
                    'method': method
                }
                summary_stats['with_breakdown'] += 1
                summary_stats['by_method'][method] = summary_stats['by_method'].get(method, 0) + 1
            else:
                print(f"NO DATA ({method})")
                summary_stats['no_data'].append(slug)
    else:
        print(f"NO DATA ({method})")
        if 'Error' in method or 'not found' in method:
            summary_stats['errors'].append(f"{slug}: {method}")
        else:
            summary_stats['no_data'].append(slug)

# Now update each project's JSON file
print("\n" + "="*80)
print("UPDATING JSON FILES")
print("="*80)

updated_count = 0
for slug, data in results.items():
    json_path = INDICES_DIR / f"{slug}.json"
    if not json_path.exists():
        print(f"  SKIP {slug} - no JSON file")
        continue

    with open(json_path, 'r', encoding='utf-8') as f:
        existing = json.load(f)

    existing['instalacoes_breakdown'] = data['instalacoes_breakdown']
    if 'split_mo_material_instalacoes' in data:
        existing['split_mo_material_instalacoes'] = data['split_mo_material_instalacoes']

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(existing, f, indent=2, ensure_ascii=False)

    updated_count += 1
    print(f"  Updated {slug}.json")

# Print summary
print("\n" + "="*80)
print("SUMMARY")
print("="*80)
print(f"Total projects processed: {summary_stats['total']}")
print(f"With installation breakdown: {summary_stats['with_breakdown']}")
print(f"With MO/Material split: {summary_stats['with_mo_split']}")
print(f"JSON files updated: {updated_count}")
print(f"\nExtraction methods:")
for method, count in sorted(summary_stats['by_method'].items(), key=lambda x: -x[1]):
    print(f"  {method}: {count}")

print(f"\nDiscipline coverage:")
for disc in ['eletricas', 'hidrossanitarias', 'preventivas', 'gas', 'telecom']:
    vals = discipline_totals[disc]
    if vals:
        avg = sum(vals) / len(vals)
        print(f"  {disc}: {len(vals)} projects, avg R$ {avg:,.0f}")
    else:
        print(f"  {disc}: 0 projects")

print(f"\nAverage MO% per discipline:")
for disc in ['eletricas', 'hidrossanitarias', 'preventivas', 'gas', 'telecom']:
    pcts = mo_pcts[disc]
    if pcts:
        avg = sum(pcts) / len(pcts)
        print(f"  {disc}: avg MO% = {avg:.1%} (n={len(pcts)})")
    else:
        print(f"  {disc}: no MO data")

if summary_stats['errors']:
    print(f"\nErrors ({len(summary_stats['errors'])}):")
    for e in summary_stats['errors']:
        print(f"  {e}")

if summary_stats['no_data']:
    print(f"\nNo installation data found ({len(summary_stats['no_data'])}):")
    for s in summary_stats['no_data']:
        print(f"  {s}")
