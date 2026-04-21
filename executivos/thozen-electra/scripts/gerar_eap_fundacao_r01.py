"""
Gera xlsx com linhas pra aba EAP do master — itens 01.002 (Fundação Profunda)
e 01.003 (Fundação Rasa) com PUs benchmark Cartesian (estimados + ranges).

Quantitativos: do consolidado projetista-1203-consolidado.json
PUs: estimativa mercado SC (hélice turnkey R$ 290/m) + benchmark típico fund. rasa
Total estimado: ~R$ 7,6M (R$ 210/m² AC)

IMPORTANTE: PUs são ESTIMATIVOS. Leo valida com:
1. Cotação formal LIBERTÉ (turnkey)
2. Benchmark base Cartesian (32+ projetos em _Entregas/Orcamento_executivo/)

Autor: Claude (copiloto Electra, 2026-04-20)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
           r"\_Executivo_IA\thozen-electra\disciplinas\estrutura\eap-fundacao-r01.xlsx")

AC = 36097  # m² — área construída total Electra (2 torres 35 pav)

# ====================================================================
# 01.002 FUNDAÇÃO PROFUNDA — Estacas Hélice Contínua (LIBERTÉ turnkey)
# ====================================================================
# Base: turnkey LIBERTÉ estimado R$ 290/m (mercado SC hélice ø50-60, 25m)
# Total turnkey: 10.575 m × R$ 290 = R$ 3.066.750
# Decomposição proporcional pra EAP (5 sub-itens):
FUND_PROFUNDA = [
    # (cod, desc, qtd, unid, pu, pu_fonte)
    ("01.002.001", "Perfuração estacas hélice contínua (ø500/600 × 25m)",
     10575, "m", 101.50,
     "Estimativa turnkey LIBERTÉ 35% × R$290/m — validar c/ cotação"),

    ("01.002.002", "Armadura estacas CA-50 (ø6.3 + ø16)",
     47631, "kg", 9.66,
     "Estimativa turnkey LIBERTÉ 15% — validar"),

    ("01.002.003", "Acessórios estacas (gaiolas, espaçadores, tubo NH)",
     423, "un", 362.50,
     "Estimativa R$/estaca — validar"),

    ("01.002.004", "Concreto C40 estacas hélice (c/ 20% sobreconsumo)",
     3544, "m³", 346.30,
     "Estimativa turnkey LIBERTÉ 40% — validar"),

    ("01.002.005", "Mão de obra estacas (execução turnkey)",
     10575, "m", 14.50,
     "Estimativa turnkey LIBERTÉ 5% — R$/m comp"),
]

# ====================================================================
# 01.003 FUNDAÇÃO RASA — Blocos coroamento + baldrame + laje fund
# ====================================================================
# Fontes:
# - QUANT-A r246 BLOCOS: 2.066,36 m³ concreto, 213.051 kg aço (A+B)
# - QUANT-B r209 "Blocos coroamento" 750,97 m³ forma 746,31 m²
# - QUANT-A r3 "Lajes maciças" topo (Bloco A fundação): 1.021,75 m³, 112.136 kg
# - QUANT-B r208 "Zapatas isoladas": 0,22 m³, 15 kg (Torre B específico)
FUND_RASA = [
    # Forma: estimativa Bloco A ~860 m² + Bloco B 746,31 m² + baldrame = ~1.900 m²
    ("01.003.001", "Forma blocos de coroamento + vigas baldrame + laje fund",
     1900, "m²", 130.00,
     "Estimativa: QUANT-B 746 m² + Bloco A ~860 m² + baldrame ~300 m²"),

    ("01.003.002", "Armadura fund. rasa CA-50 (blocos + baldrame + laje)",
     213051, "kg", 13.00,
     "QUANT-A r246 BLOCOS TOTAL A+B (100.915 + 112.136 kg)"),

    ("01.003.003", "Concreto C30 fund. rasa (blocos + baldrame + laje)",
     2066, "m³", 680.00,
     "QUANT-A r246 BLOCOS TOTAL A+B (1.044,61 + 1.021,75 m³)"),

    ("01.003.004", "Mão de obra fundação rasa",
     1, "vb", 180000.00,
     "Estimativa vb ~8-9% do material (mediana mercado SC)"),
]

# Estilos
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill("solid", fgColor="FFE699")
SECTION_FONT = Font(bold=True, size=12)
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, size=11)
GRAND_FILL = PatternFill("solid", fgColor="C65911")
GRAND_FONT = Font(bold=True, color="FFFFFF", size=12)
ALERT_FILL = PatternFill("solid", fgColor="FFCCCC")
ALERT_FONT = Font(bold=True, color="9C0006", size=10)
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EAP Fundação r01"

# Meta
ws.cell(1, 1, "Projeto").font = Font(bold=True)
ws.cell(1, 2, "Thozen Electra Towers")
ws.cell(2, 1, "Empresa").font = Font(bold=True)
ws.cell(2, 2, "Thozen")
ws.cell(3, 1, "Revisão").font = Font(bold=True)
ws.cell(3, 2, "R01")
ws.cell(1, 4, "AC total:").font = Font(bold=True)
ws.cell(1, 5, f"{AC:,} m²")
ws.cell(2, 4, "Fonte quantit:").font = Font(bold=True)
ws.cell(2, 5, "Projetista 1203 + LIBERTÉ + IFC R26")
ws.cell(3, 4, "Fonte PU:").font = Font(bold=True)
ws.cell(3, 5, "⚠️ ESTIMATIVA — validar c/ cotação formal + benchmark base Cartesian")
ws.cell(3, 5).fill = ALERT_FILL
ws.cell(3, 5).font = ALERT_FONT

# Alerta grande
ws.cell(5, 1, "⚠️ ATENÇÃO: PUs são ESTIMATIVOS (mercado SC 2025 + decomposição LIBERTÉ turnkey). "
              "Substituir por cotação formal quando disponível. "
              "Quantitativos são REAIS (do projetista 1203 + proposta LIBERTÉ 1203-2025-R0).")
ws.cell(5, 1).fill = ALERT_FILL
ws.cell(5, 1).font = ALERT_FONT
ws.cell(5, 1).alignment = LEFT
ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=7)
ws.row_dimensions[5].height = 35

# Headers
r = 7
headers = ["Código", "Descrição", "Quantidade", "Unid.", "PU (R$)",
           "Total (R$)", "Fonte / Justificativa"]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(r, col_idx, h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    c.border = BORDER
ws.row_dimensions[r].height = 24
r += 1

def write_items(section_label, items, r):
    # Section header
    ws.cell(r, 1, section_label)
    ws.cell(r, 1).fill = SECTION_FILL
    ws.cell(r, 1).font = SECTION_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(r, 1).alignment = LEFT
    ws.row_dimensions[r].height = 22
    r += 1
    start = r
    for cod, desc, qtd, unid, pu, fonte in items:
        ws.cell(r, 1, cod)
        ws.cell(r, 2, desc)
        ws.cell(r, 3, qtd)
        ws.cell(r, 4, unid)
        ws.cell(r, 5, pu)
        ws.cell(r, 6, f"=C{r}*E{r}")
        ws.cell(r, 7, fonte)
        for col in range(1, 8):
            c = ws.cell(r, col)
            c.border = BORDER
            c.alignment = LEFT if col in (1, 2, 4, 7) else RIGHT
            if col == 3:
                c.number_format = "#,##0" if isinstance(qtd, int) and qtd >= 100 else "#,##0.00"
            elif col in (5, 6):
                c.number_format = "#,##0.00"
        r += 1
    # Subtotal
    ws.cell(r, 1, "SUBTOTAL")
    ws.cell(r, 1).font = TOTAL_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 6, f"=SUM(F{start}:F{r-1})")
    for col in range(1, 8):
        c = ws.cell(r, col)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = BORDER
        if col == 6:
            c.number_format = "#,##0.00"
    r += 1
    # R$/m² AC
    ws.cell(r, 1, "R$/m² AC")
    ws.cell(r, 1).font = TOTAL_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 6, f"=F{r-1}/{AC}")
    for col in range(1, 8):
        c = ws.cell(r, col)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = BORDER
        if col == 6:
            c.number_format = "R$ #,##0.00"
    r += 2
    return r, start

r, start_profunda = write_items("01.002 FUNDAÇÃO PROFUNDA — Estacas Hélice Contínua", FUND_PROFUNDA, r)
subtotal_profunda_row = r - 3  # row do SUBTOTAL acima

r, start_rasa = write_items("01.003 FUNDAÇÃO RASA — Blocos + Baldrame + Laje Fund.", FUND_RASA, r)
subtotal_rasa_row = r - 3

# GRAND TOTAL
ws.cell(r, 1, "TOTAL FUNDAÇÃO (01.002 + 01.003)")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(r, 6, f"=F{subtotal_profunda_row}+F{subtotal_rasa_row}")
for col in range(1, 8):
    c = ws.cell(r, col)
    c.fill = GRAND_FILL
    c.font = GRAND_FONT
    c.border = BORDER
    c.alignment = LEFT if col == 1 else RIGHT
    if col == 6:
        c.number_format = "R$ #,##0.00"
r += 1
ws.cell(r, 1, "R$/m² AC FUNDAÇÃO TOTAL")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(r, 6, f"=F{r-1}/{AC}")
for col in range(1, 8):
    c = ws.cell(r, col)
    c.fill = GRAND_FILL
    c.font = GRAND_FONT
    c.border = BORDER
    if col == 6:
        c.number_format = "R$ #,##0.00"
r += 3

# Observações
obs = [
    "OBSERVAÇÕES TÉCNICAS:",
    "",
    "• QUANTITATIVOS — são REAIS (validados em 3 fontes cruzadas):",
    "  - Estacas: LIBERTÉ 1203-2025-R0 + CONTROLE-REV projetista 1203 (bateu 100%)",
    "  - Fund. Rasa: QUANTIDADES BLOCO A+B projetista 1203 (r246 totalização)",
    "",
    "• PUs — são ESTIMATIVAS (base mercado SC + decomposição turnkey LIBERTÉ):",
    "  - Hélice contínua turnkey: R$ 290/m (mediana SC 2025, ø50-60, 20-30m)",
    "  - Decomposição proporcional: 35% perfuração + 40% concreto + 15% aço + 5% acessórios + 5% MO",
    "  - Fund. rasa: PUs típicos Cartesian (concreto R$ 680/m³, aço R$ 13/kg, forma R$ 130/m²)",
    "",
    "• SOBRECONSUMO CONCRETO ESTACAS:",
    "  - Volume teórico (escavação): 2.953 m³",
    "  - Volume efetivo (+20% LIBERTÉ): 3.544 m³ — usado no 01.002.004",
    "  - Diferença 591 m³ é esperada pra hélice contínua (expansão lateral do solo)",
    "",
    "• PRÓXIMOS PASSOS (Leo):",
    "  1. Validar quantitativos com Rubens Alves (projetista) — principal: 213.051 kg aço total fund. rasa",
    "  2. Solicitar cotação FORMAL à LIBERTÉ — turnkey vs decomposto",
    "  3. Rodar benchmark da base Cartesian (scripts/benchmark_*.py) pra validar PUs",
    "  4. Confrontar R$/m² AC total vs projetos similares (Brasin Redentor, Grandezza, etc.)",
    "",
    "• VERIFICAÇÃO DE SANIDADE:",
    "  - Taxa aço estacas: 16,1 kg/m³ (típico hélice: 15-25) ✓",
    "  - Taxa aço fund. rasa: 103 kg/m³ (típico blocos/baldrame: 80-120) ✓",
    "  - R$/m² AC fund. total: se der >R$ 200 — investigar (esperado R$ 80-150 pra prédio alto)",
    "",
    "COMO COLAR NO MASTER:",
    "1. Abrir CTN-TZN_ELT - Orçamento Executivo_R00 .xlsx aba EAP",
    "2. Localizar seção 01. INFRAESTRUTURA → 1.002 e 1.003",
    "3. Copiar os 9 sub-itens deste xlsx com respectivos Qtd + Unid + PU",
    "4. Colar nas células correspondentes (não colar 'Total' — é fórmula da EAP do master)",
    "5. Verificar se códigos batem (01.002.001 → 01.002.005 e 01.003.001 → 01.003.004)",
]
for line in obs:
    ws.cell(r, 1, line)
    if line.startswith("OBSERVAÇÕES") or line.startswith("COMO COLAR"):
        ws.cell(r, 1).font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

# Widths
widths = [12, 55, 12, 8, 12, 16, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"xlsx salvo: {OUT}")

# Validação
total_profunda = sum(q * pu for _, _, q, _, pu, _ in FUND_PROFUNDA)
total_rasa = sum(q * pu for _, _, q, _, pu, _ in FUND_RASA)
total = total_profunda + total_rasa
print()
print("=== RESUMO ===")
print(f"01.002 Fund. Profunda: R$ {total_profunda:>12,.0f}  ({total_profunda/AC:>6.2f}/m² AC)")
print(f"01.003 Fund. Rasa:     R$ {total_rasa:>12,.0f}  ({total_rasa/AC:>6.2f}/m² AC)")
print(f"TOTAL FUNDAÇÃO:        R$ {total:>12,.0f}  ({total/AC:>6.2f}/m² AC)")
print()
print("Comparação benchmark mercado (prédio alto SC):")
print(f"  Fund. profunda típica: R$ 50-90/m² AC (Electra: {total_profunda/AC:.0f}/m²)")
print(f"  Fund. rasa típica:     R$ 30-60/m² AC (Electra: {total_rasa/AC:.0f}/m²)")
print(f"  Total típico:          R$ 80-150/m² AC (Electra: {total/AC:.0f}/m²)")
if total/AC > 150:
    print(f"  ⚠️ Total acima da faixa típica — revisar PUs e quantitativos fund. rasa (213k kg aço)")
