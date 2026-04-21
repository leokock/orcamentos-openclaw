#!/usr/bin/env python3
"""
Generate PPCI Electra R03 from R02 + PCI Elétrico data.
Adds ALARME DE INCÊNDIO group and PEE sinalização to each pavimento.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

INPUT = '/Users/leokock/orcamentos/executivos/thozen-electra/entregas/ppci-electra-r02.xlsx'
OUTPUT = '/Users/leokock/orcamentos/executivos/thozen-electra/entregas/ppci-electra-r03.xlsx'

# ─── PCI Elétrico data per pavimento ─────────────────────────────────
# Format: (ilum, sirenes, sinalizacao_pee, acionadores, detectores)
PCI_DATA = {
    # Embasamento pavimentos (repetição=1)
    17: {'uc': 'EMBASAMENTO', 'pav': 'TÉRREO', 'rep': 1,
         'ilum': 41, 'sirene': 13, 'sinal': 11, 'acion': 0, 'detect': 0},
    27: {'uc': 'EMBASAMENTO', 'pav': 'GARAGEM 01', 'rep': 1,
         'ilum': 37, 'sirene': 5, 'sinal': 2, 'acion': 0, 'detect': 0},
    36: {'uc': 'EMBASAMENTO', 'pav': 'GARAGEM 02', 'rep': 1,
         'ilum': 36, 'sirene': 5, 'sinal': 2, 'acion': 0, 'detect': 0},
    45: {'uc': 'EMBASAMENTO', 'pav': 'GARAGEM 03', 'rep': 1,
         'ilum': 36, 'sirene': 5, 'sinal': 2, 'acion': 0, 'detect': 0},
    54: {'uc': 'EMBASAMENTO', 'pav': 'GARAGEM 04', 'rep': 1,
         'ilum': 33, 'sirene': 5, 'sinal': 2, 'acion': 0, 'detect': 0},
    64: {'uc': 'EMBASAMENTO', 'pav': 'GARAGEM 05', 'rep': 1,
         'ilum': 33, 'sirene': 5, 'sinal': 2, 'acion': 0, 'detect': 23},
    74: {'uc': 'EMBASAMENTO', 'pav': 'LAZER', 'rep': 1,
         'ilum': 41, 'sirene': 16, 'sinal': 21, 'acion': 0, 'detect': 0},
    # Torre A
    83: {'uc': 'TORRE A', 'pav': 'PAVIMENTO TIPO', 'rep': 24,
         'ilum': 5, 'sirene': 1, 'sinal': 1, 'acion': 1, 'detect': 0},
    93: {'uc': 'TORRE A', 'pav': 'CASA DE MÁQUINAS', 'rep': 1,
         'ilum': 7, 'sirene': 0, 'sinal': 2, 'acion': 0, 'detect': 0},
    # Torre B
    106: {'uc': 'TORRE B', 'pav': 'PAVIMENTO TIPO', 'rep': 24,
          'ilum': 6, 'sirene': 1, 'sinal': 1, 'acion': 0, 'detect': 0},
    116: {'uc': 'TORRE B', 'pav': 'CASA DE MÁQUINAS', 'rep': 1,
          'ilum': 7, 'sirene': 0, 'sinal': 2, 'acion': 0, 'detect': 0},
}

# PU references
PU_ILUM = 45.00
PU_SIRENE = 85.00
PU_ACION = 120.00
PU_DETECT = 95.00
PU_SINAL = 15.00

# ─── Styles ───────────────────────────────────────────────────────────
FONT_DATA = Font(name='Arial', size=11, bold=False, color='00333333')
FONT_HEADER_PAV = Font(name='Arial', size=11, bold=True, color='00000000')
FONT_SUBTOTAL_PAV = Font(name='Arial', size=11, bold=True, color='00333333')
FONT_SUBTOTAL_UC = Font(name='Arial', size=11, bold=True, color='00000000')
FONT_TOTAL = Font(name='Arial', size=12, bold=True, color='00FFFFFF')
FONT_HEADER_COL = Font(name='Poppins', size=10, bold=True, color='00FFFFFF')
FONT_OBS = Font(name='Arial', size=11, bold=False, color='00333333')

FILL_HEADER_PAV = PatternFill(start_color='FF8DB4E2', end_color='FF8DB4E2', fill_type='solid')
FILL_SUBTOTAL_PAV = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
FILL_SUBTOTAL_UC = PatternFill(start_color='FFB8CCE4', end_color='FFB8CCE4', fill_type='solid')
FILL_TOTAL = PatternFill(start_color='FF1851D6', end_color='FF1851D6', fill_type='solid')
FILL_COL_HEADER = PatternFill(start_color='FF333333', end_color='FF333333', fill_type='solid')

NUMFMT_NUM = '#,##0.00'

# ─── Load R02 ─────────────────────────────────────────────────────────
wb_src = openpyxl.load_workbook(INPUT)
ws_src = wb_src.active

# Read all source data
src_rows = []
for r in range(1, ws_src.max_row + 1):
    row_data = []
    for c in range(1, 13):  # A-L
        cell = ws_src.cell(row=r, column=c)
        row_data.append({
            'value': cell.value,
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
            'border': copy(cell.border),
        })
    src_rows.append((r, row_data))

# ─── Build new rows ──────────────────────────────────────────────────
def make_data_row(uc, pav, grupo, subgrupo, desc, qty, rep, perda, unit, pu):
    """Create a data row dict list (cols A-L)."""
    row_n = 999  # placeholder
    return [
        {'value': uc},
        {'value': pav},
        {'value': grupo},
        {'value': subgrupo},
        {'value': desc},
        {'value': qty},
        {'value': rep},
        {'value': perda},
        {'value': '__FORMULA_I__'},  # placeholder
        {'value': unit},
        {'value': pu},
        {'value': '__FORMULA_L__'},  # placeholder
    ]

def make_alarme_rows(data):
    """Generate ALARME DE INCÊNDIO rows for a pavimento."""
    uc = data['uc']
    pav = data['pav']
    rep = data['rep']
    rows = []
    
    # Sinalização PEE — add to SINALIZAÇÃO group
    if data['sinal'] > 0:
        rows.append(('SINAL_PEE', make_data_row(
            uc, pav, 'SINALIZAÇÃO', 'Sinalização de Abandono de Local (SAL)',
            'Placa de sinalização de emergência fotoluminescente (rota de fuga)',
            data['sinal'], rep, 0, 'pç', PU_SINAL
        )))
    
    # ALARME DE INCÊNDIO group
    if data['ilum'] > 0:
        rows.append(('ALARME', make_data_row(
            uc, pav, 'ALARME DE INCÊNDIO', 'Sistema de Alarme e Detecção de Incêndio',
            'Luminária de emergência autônoma LED',
            data['ilum'], rep, 0, 'pç', PU_ILUM
        )))
    if data['sirene'] > 0:
        rows.append(('ALARME', make_data_row(
            uc, pav, 'ALARME DE INCÊNDIO', 'Sistema de Alarme e Detecção de Incêndio',
            'Sirene/Campainha de alarme de incêndio',
            data['sirene'], rep, 0, 'pç', PU_SIRENE
        )))
    if data['acion'] > 0:
        rows.append(('ALARME', make_data_row(
            uc, pav, 'ALARME DE INCÊNDIO', 'Sistema de Alarme e Detecção de Incêndio',
            'Acionador manual de alarme',
            data['acion'], rep, 0, 'pç', PU_ACION
        )))
    if data['detect'] > 0:
        rows.append(('ALARME', make_data_row(
            uc, pav, 'ALARME DE INCÊNDIO', 'Sistema de Alarme e Detecção de Incêndio',
            'Detector de fumaça fotoelétrico',
            data['detect'], rep, 0, 'pç', PU_DETECT
        )))
    
    return rows

# Build output rows: iterate through source, inserting new rows before subtotals
output_rows = []  # list of (row_type, data)
# row_type: 'src' (from source), 'new_sinal' (PEE sinalização), 'new_alarme' (alarme items)

for orig_row, row_data in src_rows:
    if orig_row in PCI_DATA:
        # This is a subtotal row — insert new items before it
        new_rows = make_alarme_rows(PCI_DATA[orig_row])
        
        # Insert SINAL_PEE rows right before the subtotal (after last SINALIZAÇÃO row)
        # Insert ALARME rows right before the subtotal
        for rtype, rdata in new_rows:
            output_rows.append(('new', rdata))
    
    output_rows.append(('src', orig_row, row_data))

# ─── Write new workbook ──────────────────────────────────────────────
wb_new = openpyxl.Workbook()
ws = wb_new.active
ws.title = 'PPCI'

# Column widths
col_widths = {'A': 18, 'B': 25, 'C': 16, 'D': 40, 'E': 50, 'F': 13, 'G': 12, 'H': 8, 'I': 14, 'J': 10, 'K': 14, 'L': 19}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Track row mapping for subtotals
new_row = 0
# Track which new rows are data rows for subtotal formulas
# We need to track pavimento data ranges for subtotal SUM formulas

# First pass: assign row numbers
row_map = {}  # orig_row -> new_row
data_start = {}  # subtotal_orig_row -> first data row in new sheet
data_end = {}    # subtotal_orig_row -> last data row in new sheet

current_data_start = None
current_subtotal = None

# Figure out pavimento structure
# Pavimento header rows: 7, 18, 28, 37, 46, 55, 65, 76, 84, 94, 99, 107, 117, 122
# Subtotal rows: 17, 27, 36, 45, 54, 64, 74, 83, 93, 97, 106, 116, 120, 125
PAV_HEADERS = {7, 18, 28, 37, 46, 55, 65, 76, 84, 94, 99, 107, 117, 122}
SUBTOTAL_ROWS = {17, 27, 36, 45, 54, 64, 74, 83, 93, 97, 106, 116, 120, 125}
UC_SUBTOTAL_ROWS = {75, 98, 121, 126}
TOTAL_ROW = 128

# Assign row numbers
new_r = 0
row_assignments = []  # (new_row_num, type, orig_row_or_None, data)

for entry in output_rows:
    new_r += 1
    if entry[0] == 'src':
        orig_r = entry[1]
        row_assignments.append((new_r, 'src', orig_r, entry[2]))
        row_map[orig_r] = new_r
    else:
        row_assignments.append((new_r, 'new', None, entry[1]))

# Now figure out data ranges for each subtotal
# For each subtotal, find the data rows between the header row and this subtotal
# in the new numbering

# Map: for each subtotal orig row, what's the pavimento header orig row?
subtotal_to_header = {
    17: 7, 27: 18, 36: 28, 45: 37, 54: 46, 64: 55, 74: 65,
    83: 76, 93: 84, 97: 94, 106: 99, 116: 107, 120: 117, 125: 122
}

# For each subtotal, collect data row numbers in new sheet
subtotal_data_ranges = {}
for sub_orig, hdr_orig in subtotal_to_header.items():
    new_sub = row_map[sub_orig]
    new_hdr = row_map[hdr_orig]
    # Data rows are between header+1 and subtotal-1
    first_data = new_hdr + 1
    last_data = new_sub - 1
    subtotal_data_ranges[sub_orig] = (first_data, last_data)

# UC subtotal formulas need the new subtotal row numbers
# Embasamento: sum of subtotals 17,27,36,45,54,64,74
# Torre A: sum of subtotals 83,93,97
# Torre B: sum of subtotals 106,116,120
# Geral: sum of subtotal 125
UC_SUBTOTAL_PARTS = {
    75: [17, 27, 36, 45, 54, 64, 74],
    98: [83, 93, 97],
    121: [106, 116, 120],
    126: [125],
}

# Total row: sum of UC subtotals
TOTAL_PARTS = [75, 98, 121, 126]

# ─── Write rows ──────────────────────────────────────────────────────
def apply_data_style(cell, col_idx):
    """Apply standard data row style."""
    cell.font = copy(FONT_DATA)
    if col_idx in (6, 9, 11, 12):  # F, I, K, L — numeric
        cell.number_format = NUMFMT_NUM

def apply_subtotal_pav_style(cell):
    """Apply pavimento subtotal style."""
    cell.font = copy(FONT_SUBTOTAL_PAV)
    cell.fill = copy(FILL_SUBTOTAL_PAV)

def apply_header_pav_style(cell):
    """Apply pavimento header style."""
    cell.font = copy(FONT_HEADER_PAV)
    cell.fill = copy(FILL_HEADER_PAV)

for new_r, rtype, orig_r, data in row_assignments:
    if rtype == 'src':
        # Copy from source with original styling
        for c in range(12):
            cell = ws.cell(row=new_r, column=c+1)
            d = data[c]
            
            # Handle special rows
            if orig_r == 4:
                # Update revision
                if c == 3:  # D4
                    cell.value = 'R03 - Jarvis 25/03/2026'
                else:
                    cell.value = d['value']
            elif orig_r == 3:
                # Update origin
                if c == 3:  # D3
                    cell.value = 'CTN-TZN_PCI_EX_000_QTD_R03'
                else:
                    cell.value = d['value']
            elif orig_r in SUBTOTAL_ROWS:
                # Subtotal row — rebuild formula
                if c == 11:  # L column
                    first, last = subtotal_data_ranges[orig_r]
                    cell.value = f'=SUM(L{first}:L{last})'
                elif c == 4:  # E column — subtotal label
                    cell.value = d['value']
                else:
                    cell.value = d['value']
            elif orig_r in UC_SUBTOTAL_ROWS:
                if c == 11:  # L column
                    parts = UC_SUBTOTAL_PARTS[orig_r]
                    refs = '+'.join([f'L{row_map[p]}' for p in parts])
                    cell.value = f'={refs}'
                else:
                    cell.value = d['value']
            elif orig_r == TOTAL_ROW:
                if c == 11:  # L column
                    refs = '+'.join([f'L{row_map[p]}' for p in TOTAL_PARTS])
                    cell.value = f'={refs}'
                else:
                    cell.value = d['value']
            elif orig_r > 7 and orig_r not in PAV_HEADERS and orig_r not in SUBTOTAL_ROWS and orig_r not in UC_SUBTOTAL_ROWS and orig_r != TOTAL_ROW and orig_r <= 128:
                # Data row — rebuild formulas
                if c == 8:  # I column
                    cell.value = f'=(F{new_r}+H{new_r}*F{new_r})*G{new_r}'
                elif c == 11:  # L column
                    cell.value = f'=I{new_r}*K{new_r}'
                else:
                    cell.value = d['value']
            else:
                cell.value = d['value']
            
            # Apply original styling
            cell.font = copy(d['font'])
            if d['fill'].patternType:
                cell.fill = copy(d['fill'])
            cell.alignment = copy(d['alignment'])
            cell.number_format = d['number_format']
            cell.border = copy(d['border'])
    
    else:
        # New row (PCI Elétrico)
        for c in range(12):
            cell = ws.cell(row=new_r, column=c+1)
            d = data[c]
            
            if c == 8:  # I — formula
                cell.value = f'=(F{new_r}+H{new_r}*F{new_r})*G{new_r}'
            elif c == 11:  # L — formula
                cell.value = f'=I{new_r}*K{new_r}'
            else:
                cell.value = d['value']
            
            # Apply data row style
            cell.font = copy(FONT_DATA)
            if c in (5, 8, 10, 11):  # F, I, K, L
                cell.number_format = NUMFMT_NUM

# Add observation about PCI Elétrico
obs_row = row_map.get(130, None)
if obs_row:
    pass  # Already copied

# Find last row and add new observation
last_row = max(nr for nr, _, _, _ in row_assignments)

# Check if obs rows were copied
obs_mapped = row_map.get(135, None)
if obs_mapped:
    new_obs_row = obs_mapped + 1
    ws.cell(row=new_obs_row, column=1).value = '6. Itens de PCI Elétrico (iluminação de emergência, alarme, sinalização PEE) extraídos dos DXFs PEE rev.01'
    ws.cell(row=new_obs_row, column=1).font = copy(FONT_OBS)

# Merge header row A5:L5
new_r5 = row_map[5]
ws.merge_cells(start_row=new_r5, start_column=1, end_row=new_r5, end_column=12)

# Freeze panes
new_r7 = row_map.get(7, 7)
ws.freeze_panes = f'A{new_r7}'

# ─── Save ─────────────────────────────────────────────────────────────
wb_new.save(OUTPUT)
print(f'Saved R03 to {OUTPUT}')

# Print summary
total_new = sum(1 for _, rtype, _, _ in row_assignments if rtype == 'new')
print(f'Added {total_new} new PCI Elétrico rows')
print(f'Total rows: {last_row}')

# Verify data ranges
for sub_orig in sorted(subtotal_data_ranges.keys()):
    first, last = subtotal_data_ranges[sub_orig]
    new_sub = row_map[sub_orig]
    pav = subtotal_to_header[sub_orig]
    pav_name = ''
    for nr, rt, orig, d in row_assignments:
        if rt == 'src' and orig == pav:
            pav_name = f"{d[0]['value']} / {d[1]['value']}"
            break
    print(f'  {pav_name}: data L{first}:L{last}, subtotal L{new_sub}')
