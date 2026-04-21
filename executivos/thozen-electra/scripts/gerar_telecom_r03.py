"""Gera telecomunicacoes-electra-r03.xlsx — suplemento ao R02 com 28 caixas de
passagem telecom (PDF Eletrowatts 07/04/2026), **organizadas por pavimento**.

Estrutura do xlsx (por pavimento — convencao Electra):
- Aba 1 "CAIXAS POR PAVIMENTO": tabela flat com Pavimento | Torre | # | Qtd | Dim | Local | Desc | PU | Total. Subtotais por pavimento.
- Aba 2 "RESUMO POR PAVIMENTO": pivot qtd/total por pavimento × torre
- Aba 3 "RESUMO R03": R02 + suplemento + total consolidado
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra")
PDF_JSON = BASE / "quantitativos" / "listas-materiais" / "telecom" / "telecom.json"
OUT_XLSX = BASE / "disciplinas" / "telefonico" / "telecomunicacoes-electra-r03.xlsx"

sys.path.insert(0, str(BASE / "scripts"))
from normalizar_pavimento import PAVIMENTOS_ORDER, PAVIMENTO_LABEL, ordem_pavimento

R02_TOTAL = 424000
AC_TOTAL_M2 = 36092


def pu_por_dim(dim: str) -> tuple[float, str]:
    m = re.match(r"(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)", dim)
    if not m:
        return 150.0, "fallback"
    w, h, p = int(m.group(1)), int(m.group(2)), int(m.group(3))
    vol = (w * h * p) / 1000
    if vol <= 5:
        return 85.0, "Faixa 1 (<= 5 dm3)"
    if vol <= 15:
        return 140.0, "Faixa 2 (5-15 dm3)"
    if vol <= 30:
        return 280.0, "Faixa 3 (15-30 dm3)"
    return 420.0, "Faixa 4 (> 30 dm3)"


def load_items() -> list[dict]:
    pdf = json.loads(PDF_JSON.read_text(encoding="utf-8"))
    doc = pdf["documentos"][0]
    rows = []
    for it in doc["itens"]:
        pu, fonte_pu = pu_por_dim(it["dimensoes"])
        rows.append({
            "pavimento": it["pavimento"],
            "torre": it["torre"],
            "pavimento_label": it["pavimento_label"],
            "numero": it["numero"],
            "qtd": int(it["qtd"]),
            "unidade": it["unidade"],
            "dimensoes": it["dimensoes"],
            "local": it["especificacao_local"],
            "descricao": it["produto"],
            "pu": pu,
            "fonte_pu": fonte_pu,
        })
    # ordenar por (pavimento, torre, numero)
    rows.sort(key=lambda r: (ordem_pavimento(r["pavimento"]), r["torre"], r["numero"]))
    return rows


def _styles():
    return {
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "header_fill": PatternFill("solid", fgColor="2E75B6"),
        "subtot_fill": PatternFill("solid", fgColor="FFF2CC"),
        "total_fill": PatternFill("solid", fgColor="C6E0B4"),
        "pav_fill": PatternFill("solid", fgColor="DDEBF7"),
        "border": Border(*[Side(border_style="thin", color="888888")] * 4),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
    }


def write_xlsx(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    s = _styles()

    # ============ ABA 1 — CAIXAS POR PAVIMENTO ============
    ws = wb.active
    ws.title = "CAIXAS POR PAVIMENTO"

    ws["A1"] = "THOZEN - RES. ELECTRA TOWERS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Disciplina: 08 Telecomunicacoes - R03 Suplemento (por pavimento)"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = "Revisao: R03 | Data: 2026-04-14 | AC: 36.092 m2"
    ws["A4"] = "Fonte: PDF Eletrowatts 'Caixas e Quadros - Telecom' 07/04/2026 (28 caixas tronco)"
    ws["A4"].font = Font(italic=True, size=9, color="666666")

    headers = ["Pavimento", "Torre", "#", "Qtd", "Unid", "Dimensoes", "Local (original)",
               "Descricao", "PU (R$)", "Total (R$)", "Fonte PU"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]

    row = 7
    current_pav = None
    current_pav_start = row
    pav_subtotal_rows = []

    def fechar_subtotal_pavimento():
        nonlocal row
        if current_pav is None:
            return
        pav_label = PAVIMENTO_LABEL.get(current_pav, current_pav)
        ws.cell(row=row, column=1, value=f"Subtotal — {pav_label}").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1).fill = s["subtot_fill"]
        ws.cell(row=row, column=1).alignment = s["left"]
        ws.cell(row=row, column=10, value=f"=SUM(J{current_pav_start}:J{row-1})").number_format = '#,##0.00'
        ws.cell(row=row, column=10).font = Font(bold=True)
        ws.cell(row=row, column=10).fill = s["subtot_fill"]
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = s["border"]
            if col != 10:
                ws.cell(row=row, column=col).fill = s["subtot_fill"]
        pav_subtotal_rows.append(row)
        row += 1

    for it in rows:
        if it["pavimento"] != current_pav:
            fechar_subtotal_pavimento()
            current_pav = it["pavimento"]
            current_pav_start = row

        ws.cell(row=row, column=1, value=PAVIMENTO_LABEL.get(it["pavimento"], it["pavimento"])).alignment = s["left"]
        ws.cell(row=row, column=2, value=it["torre"]).alignment = s["center"]
        ws.cell(row=row, column=3, value=it["numero"]).alignment = s["center"]
        ws.cell(row=row, column=4, value=it["qtd"]).alignment = s["right"]
        ws.cell(row=row, column=5, value=it["unidade"]).alignment = s["center"]
        ws.cell(row=row, column=6, value=it["dimensoes"]).alignment = s["center"]
        ws.cell(row=row, column=7, value=it["local"]).alignment = s["left"]
        ws.cell(row=row, column=8, value=it["descricao"]).alignment = s["left"]
        ws.cell(row=row, column=9, value=it["pu"]).alignment = s["right"]
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        ws.cell(row=row, column=10, value=f'=D{row}*I{row}').alignment = s["right"]
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=it["fonte_pu"]).alignment = s["left"]
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = s["border"]
        row += 1

    fechar_subtotal_pavimento()

    # Total geral
    row += 1
    ws.cell(row=row, column=1, value="TOTAL SUPLEMENTO (28 caixas por pavimento)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1).fill = s["total_fill"]
    ws.cell(row=row, column=1).alignment = s["left"]
    if pav_subtotal_rows:
        ref = "+".join(f"J{r}" for r in pav_subtotal_rows)
        ws.cell(row=row, column=10, value=f"={ref}").number_format = '#,##0.00'
    ws.cell(row=row, column=10).font = Font(bold=True, size=11)
    ws.cell(row=row, column=10).fill = s["total_fill"]
    total_material_row = row
    row += 1

    # MO 20%
    ws.cell(row=row, column=1, value="MO instalacao (20% sobre material)").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=10, value=f"=J{total_material_row}*0.20").number_format = '#,##0.00'
    ws.cell(row=row, column=10).fill = s["subtot_fill"]
    mo_row = row
    row += 1

    ws.cell(row=row, column=1, value="TOTAL SUPLEMENTO R03 (material + MO)").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=10, value=f"=J{total_material_row}+J{mo_row}").number_format = '#,##0.00'
    ws.cell(row=row, column=10).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="1F4E78")
    total_suplemento_row = row

    widths = [22, 6, 4, 5, 5, 18, 26, 50, 10, 14, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[6].height = 24
    ws.freeze_panes = "A7"

    # ============ ABA 2 — RESUMO POR PAVIMENTO ============
    ws2 = wb.create_sheet("RESUMO POR PAVIMENTO")
    ws2["A1"] = "TELECOM R03 — Distribuicao por Pavimento"
    ws2["A1"].font = Font(bold=True, size=14)

    hdr = ["Pavimento", "Torre A qtd", "Torre A R$", "Torre B qtd", "Torre B R$", "Total qtd", "Total R$"]
    for col, h in enumerate(hdr, start=1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]

    pavimentos_unicos = sorted(set(r["pavimento"] for r in rows), key=ordem_pavimento)
    r2 = 4
    for pav in pavimentos_unicos:
        itens_a = [x for x in rows if x["pavimento"] == pav and x["torre"] == "A"]
        itens_b = [x for x in rows if x["pavimento"] == pav and x["torre"] == "B"]
        qtd_a = sum(x["qtd"] for x in itens_a)
        tot_a = sum(x["qtd"] * x["pu"] for x in itens_a)
        qtd_b = sum(x["qtd"] for x in itens_b)
        tot_b = sum(x["qtd"] * x["pu"] for x in itens_b)
        ws2.cell(row=r2, column=1, value=PAVIMENTO_LABEL.get(pav, pav)).alignment = s["left"]
        ws2.cell(row=r2, column=2, value=qtd_a).alignment = s["right"]
        ws2.cell(row=r2, column=3, value=tot_a).number_format = '#,##0.00'
        ws2.cell(row=r2, column=4, value=qtd_b).alignment = s["right"]
        ws2.cell(row=r2, column=5, value=tot_b).number_format = '#,##0.00'
        ws2.cell(row=r2, column=6, value=qtd_a + qtd_b).font = Font(bold=True)
        ws2.cell(row=r2, column=7, value=tot_a + tot_b).number_format = '#,##0.00'
        ws2.cell(row=r2, column=7).font = Font(bold=True)
        for col in range(1, 8):
            ws2.cell(row=r2, column=col).border = s["border"]
        r2 += 1

    # Linha total
    tot_qtd = sum(r["qtd"] for r in rows)
    tot_mat = sum(r["qtd"] * r["pu"] for r in rows)
    ws2.cell(row=r2, column=1, value="TOTAL").font = Font(bold=True, size=11)
    ws2.cell(row=r2, column=1).fill = s["total_fill"]
    ws2.cell(row=r2, column=6, value=tot_qtd).font = Font(bold=True, size=11)
    ws2.cell(row=r2, column=6).fill = s["total_fill"]
    ws2.cell(row=r2, column=7, value=tot_mat).number_format = '#,##0.00'
    ws2.cell(row=r2, column=7).font = Font(bold=True, size=11)
    ws2.cell(row=r2, column=7).fill = s["total_fill"]
    for col in range(1, 8):
        ws2.cell(row=r2, column=col).border = s["border"]
        if col not in (6, 7):
            ws2.cell(row=r2, column=col).fill = s["total_fill"]

    for i, w in enumerate([28, 12, 14, 12, 14, 12, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ============ ABA 3 — RESUMO R03 ============
    ws3 = wb.create_sheet("RESUMO R03")
    ws3["A1"] = "TELECOM R03 — Consolidado com R02"
    ws3["A1"].font = Font(bold=True, size=14)

    for col, h in enumerate(["Fonte", "Escopo", "Total (R$)", "R$/m2 AC"], start=1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]

    ws3.cell(row=4, column=1, value="R02 (30/mar)").font = Font(bold=True)
    ws3.cell(row=4, column=2, value="IFC (648 caixas) + DWG (222 pontos ativos)").alignment = s["left"]
    ws3.cell(row=4, column=3, value=R02_TOTAL).number_format = '#,##0.00'
    ws3.cell(row=4, column=4, value=R02_TOTAL / AC_TOTAL_M2).number_format = '#,##0.00'

    ws3.cell(row=5, column=1, value="R03 suplemento").font = Font(bold=True)
    ws3.cell(row=5, column=2, value="28 caixas de passagem tronco (PDF) por pavimento").alignment = s["left"]
    ws3.cell(row=5, column=3, value=f"='CAIXAS POR PAVIMENTO'!J{total_suplemento_row}").number_format = '#,##0.00'
    ws3.cell(row=5, column=3).fill = PatternFill("solid", fgColor="FFE699")
    ws3.cell(row=5, column=4, value=f"=C5/{AC_TOTAL_M2}").number_format = '#,##0.00'

    ws3.cell(row=6, column=1, value="TOTAL R03").font = Font(bold=True, size=11)
    ws3.cell(row=6, column=2, value="R02 + Suplemento").alignment = s["left"]
    ws3.cell(row=6, column=3, value="=C4+C5").number_format = '#,##0.00'
    ws3.cell(row=6, column=3).font = Font(bold=True, size=11)
    ws3.cell(row=6, column=3).fill = s["total_fill"]
    ws3.cell(row=6, column=4, value=f"=C6/{AC_TOTAL_M2}").number_format = '#,##0.00'

    for r in range(4, 7):
        for col in range(1, 5):
            ws3.cell(row=r, column=col).border = s["border"]

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 60
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 12

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"gravado: {OUT_XLSX}")


def main():
    rows = load_items()
    print(f"caixas: {len(rows)}")
    write_xlsx(rows)


if __name__ == "__main__":
    main()
