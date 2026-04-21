"""
Extrair projetista 1203 (Rubens Alves) + proposta fundacionista Liberté
para consolidar dados de FUNDAÇÃO (estacas + fundação rasa) do Electra Towers.

Fontes:
- 1203 - QUANTIDADES AÇO E CONCRETO - BLOCO A/B.xlsx  (por pavimento + resumos)
- 1203 - CONTROLE DE REVISÃO - BLOCO A/B (1).xlsx     (aço por bitola + estacas)
- ESTACAS.jpeg (Proposta Comercial LIBERTÉ 1203-2025-R0) - hardcoded abaixo

Outputs:
- disciplinas/estrutura/projetista-1203-consolidado.json
- disciplinas/estrutura/projetista-1203-consolidado.xlsx (5 abas)

Autor: Claude (copiloto Electra, 17-abr-2026)
"""
from __future__ import annotations
import json
import math
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_PROJ = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
    r"\_Projetos_IA\thozen-electra\projetos\01 ESTRUTURA"
)
BASE_EXEC = Path(
    r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
    r"\_Executivo_IA\thozen-electra"
)
OUT_DIR = BASE_EXEC / "disciplinas" / "estrutura"

# ---------------------------------------------------------------------------
# Proposta Comercial LIBERTÉ (empresa de fundações), extraída de ESTACAS.jpeg
# Código 1203-2025-R0. Trata o empreendimento Electra como 1 único contrato
# (sem separação Torre A / Torre B).
# ---------------------------------------------------------------------------
LIBERTE = {
    "fornecedor": "LIBERTÉ Estruturas de Concreto Armado, Fundações e Protendidos",
    "codigo_proposta": "1203-2025-R0",
    "tipo_fundacao": "Estaca Hélice Contínua Monitorada",
    "localizacao": "Rua Rubens Alves esq. Rua Canoinhas e Rua Wilson Belber, "
                   "Balneário Perequê, Porto Belo/SC",
    "perfuracao": [
        {"diametro_mm": 500, "diametro_m": 0.50, "L_med_m": 25,
         "qtd": 17, "L_total_m": 425},
        {"diametro_mm": 600, "diametro_m": 0.60, "L_med_m": 25,
         "qtd": 406, "L_total_m": 10150},
    ],
    "perfuracao_total": {"qtd": 423, "L_total_m": 10575},
    "concreto_C40_m3": {
        "teorico_mais_20pct_sobreconsumo": 3544,
        "fator_sobreconsumo": 0.20,
        "teorico_sem_sobreconsumo_deduzido": 3544 / 1.20,  # 2953.3 m³
    },
    "aco_CA50": [
        {"diametro_mm": 6.3, "tipo": "CA-50", "L_total_m": 24699,
         "n_barras": 2059, "peso_kg": 6175},
        {"diametro_mm": 16.0, "tipo": "CA-50", "L_total_m": 26238,
         "n_barras": 2187, "peso_kg": 41456},
    ],
    "aco_total_kg": 47631,
}


def read_quantidades(block: str) -> dict:
    """Lê 1203 - QUANTIDADES AÇO E CONCRETO - BLOCO [A|B].xlsx"""
    fn = f"1203 - QUANTIDADES AÇO E CONCRETO - BLOCO {block}.xlsx"
    path = BASE_PROJ / fn
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Table 1"]

    def num(v):
        if v is None or v == "-" or (isinstance(v, str) and v.strip() in ("-", "")):
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    out = {
        "fonte": fn,
        "por_pavimento": {},
        "fundacao_rasa_topo": {},  # Só Bloco B trouxe isso explicito
        "resumos": {},
    }

    # ---- Fundação rasa no topo (formato varia entre A e B) ----
    # Bloco A (9 cols): headers em Fôrmas=2, Superfície=3, Volume=5, Aço=7
    # Bloco B (8 cols): headers em Fôrmas=2, Superfície=3, Volume=4, Aço=6
    # Detecta pelo max_column
    if block == "B":
        # Bloco B: r3 = Zapatas isoladas, r4 = Blocos de coroamento
        # Cols no B (quando é "Total obra"): Fôrmas=2, Volume=4, Aço=6 (sem Superfície)
        out["fundacao_rasa_topo"] = {
            "zapatas_isoladas": {
                "forma_m2": num(ws.cell(3, 2).value),
                "volume_m3": num(ws.cell(3, 4).value),
                "aco_kg": num(ws.cell(3, 6).value),
            },
            "blocos_coroamento": {
                "forma_m2": num(ws.cell(4, 2).value),
                "volume_m3": num(ws.cell(4, 4).value),
                "aco_kg": num(ws.cell(4, 6).value),
            },
        }
    else:
        # Bloco A: r3 = Lajes maciças (topo, parece ser laje de fundação)
        # Header row 2 tem "Fôrmas | Superfície | | Volume | | Barras"
        # → Fôrmas=2, Superfície=3, Volume=5, Aço=7
        out["fundacao_rasa_topo"] = {
            "lajes_macicas_fundacao": {
                "forma_m2": num(ws.cell(3, 2).value),
                "volume_m3": num(ws.cell(3, 5).value),
                "aco_kg": num(ws.cell(3, 7).value),
            },
        }

    # ---- Percorre linhas identificando pavimentos (L1-L35) e elementos ----
    # Estratégia: varrer 1 a 250, detectar "L<n>" no início de linha
    # e agregar elementos até o próximo "L<n>" ou EOF.
    current_pav = None
    n_col = ws.max_column
    # offset de coluna varia entre A (9c) e B (8c) — usar headers da row 2
    # A: B=Fôrmas, C=Superfície, E=Volume, G=Barras
    # B: B=Fôrmas, C=Superfície, D=Volume ou E=Volume, ...
    # Vou detectar dinamicamente pela linha 2 ou 8
    # Bloco A: header é "Elemento | Fôrmas | Superfície | | Volume | | Barras | |"
    #          colunas: 1=Elemento, 2=Fôrmas, 3=Superfície, 5=Volume, 7=Barras
    # Bloco B: header é "Elemento | Fôrmas | Superfície | | Volume | | Barras |"
    #          colunas: 1=Elemento, 2=Fôrmas, 3=Superfície, 5=Volume, 7=Barras
    # Ambos idênticos — mapeamento fixo
    cols = {"forma": 2, "superficie": 3, "volume": 5, "aco": 7}

    for i in range(1, ws.max_row + 1):
        col_a = ws.cell(i, 1).value
        a = str(col_a).strip() if col_a else ""

        # detectar cabeçalho de pavimento "L<N> - NOME"
        if a.startswith("L") and " - " in a and len(a) < 30:
            current_pav = a
            out["por_pavimento"][current_pav] = {}
            continue

        # ignorar sub-headers e agregados
        if a in ("Elemento", "Total", "Índices (por m²)", "Nº blocos de l. nervurada"):
            continue
        if "Superfície total" in a:
            continue

        # sair do modo pavimento se chegou em seção "Total obra" ou similar
        if a in ("Total obra", "BLOCO A", "BLOCO B", "RESUMO TOTAL DA OBRA COM FUNDAÇÃO",
                 "TORRE A + B SEM FUNDAÇÃO", "HÉLICE", "BLOCOS", "TOTAL"):
            current_pav = None
            # Continua iteração pra capturar resumos abaixo
            continue

        if current_pav and a:
            # é um elemento dentro do pavimento atual
            out["por_pavimento"][current_pav][a] = {
                "forma_m2": num(ws.cell(i, cols["forma"]).value),
                "superficie_m2": num(ws.cell(i, cols["superficie"]).value),
                "volume_m3": num(ws.cell(i, cols["volume"]).value),
                "aco_kg": num(ws.cell(i, cols["aco"]).value),
            }

    # ---- Resumos finais (rows 207-248 no A, 206-220 no B) ----
    # Total obra Torre A (r214-220), Torre B (r226-232)
    # Resumo HÉLICE + BLOCOS (r243-247) só no arquivo BLOCO A
    if block == "A":
        for i in range(240, min(ws.max_row + 1, 260)):
            label = str(ws.cell(i, 1).value or "").strip()
            if label in ("HÉLICE", "BLOCOS", "TOTAL"):
                out["resumos"][label] = {
                    "volume_m3": num(ws.cell(i, cols["volume"]).value),
                    "aco_kg": num(ws.cell(i, cols["aco"]).value),
                }

    return out


def read_controle_revisao(block: str) -> dict:
    """Lê 1203 - CONTROLE DE REVISÃO - BLOCO [A|B] (1).xlsx aba QUANTIITATIVOS"""
    fn = f"1203 - CONTROLE DE REVISÃO DE PROJETOS E CONSUMO - BLOCO {block} (1).xlsx"
    path = BASE_PROJ / fn
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["QUANTIITATIVOS"]

    # Bitolas (colunas D-L = 4-12): Q92, 5, 6.3, 8, 10, 12.5, 16, 20, 25
    # Concreto = coluna M (13)
    BITOLAS = ["Q92", "5", "6.3", "8", "10", "12.5", "16", "20", "25"]

    def num(v):
        if v is None:
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def parse_element_row(i):
        return {
            "aco_por_bitola": {
                BITOLAS[j]: num(ws.cell(i, 4 + j).value) for j in range(9)
            },
            "aco_total_kg": sum(num(ws.cell(i, 4 + j).value) for j in range(9)),
            "concreto_m3": num(ws.cell(i, 13).value),
        }

    out = {
        "fonte": fn,
        "estacas": {},  # "Somente Estacas"
        "por_pavimento": {},  # L1-Lxx
        "resumos": {
            "aco_fundacao_sem_sobre": {},
            "aco_fundacao_com_5pct_sobre": {},
            "aco_estrutura_sem_sobre": {},
            "aco_estrutura_com_5pct_sobre": {},
            "concreto_estacas_sem_sobre": 0.0,
            "concreto_estacas_com_20pct_sobre": 0.0,
            "concreto_estrutura_sem_sobre": 0.0,
            "concreto_estrutura_com_10pct_sobre": 0.0,
            "area_estrutural_m2": 0.0,
            "espessura_media_m": 0.0,
            "aco_por_volume_kg_m3": 0.0,
            "aco_por_area_kg_m2": 0.0,
        },
    }

    current_pav = None
    for i in range(1, ws.max_row + 1):
        col_b = ws.cell(i, 2).value
        col_c = ws.cell(i, 3).value
        b = str(col_b).strip() if col_b else ""
        c = str(col_c).strip() if col_c else ""

        # "Somente Estacas" (row 6 tipicamente)
        if c == "Somente Estacas":
            out["estacas"] = parse_element_row(i)
            continue

        # Detectar início de pavimento
        if b.startswith("L") and " - " in b:
            current_pav = b
            out["por_pavimento"][current_pav] = {}
            continue

        # Elemento dentro de pavimento
        if current_pav and c and c not in ("Elemento Estrutural", "Totalização"):
            out["por_pavimento"][current_pav][c] = parse_element_row(i)
            continue
        if c == "Totalização" and current_pav:
            out["por_pavimento"][current_pav]["_totalizacao"] = parse_element_row(i)
            continue

        # Resumos finais
        if b.startswith("CONSUMO DE AÇO DA FU"):
            # próximas 2 linhas (sem e com 5% sobre)
            out["resumos"]["aco_fundacao_sem_sobre"] = {
                BITOLAS[j]: num(ws.cell(i + 1, 4 + j).value) for j in range(9)
            }
            out["resumos"]["aco_fundacao_sem_sobre"]["total"] = num(ws.cell(i + 1, 13).value)
            out["resumos"]["aco_fundacao_com_5pct_sobre"] = {
                BITOLAS[j]: num(ws.cell(i + 2, 4 + j).value) for j in range(9)
            }
            out["resumos"]["aco_fundacao_com_5pct_sobre"]["total"] = num(ws.cell(i + 2, 13).value)
        elif b.startswith("CONSUMO DE AÇO DA ES"):
            out["resumos"]["aco_estrutura_sem_sobre"] = {
                BITOLAS[j]: num(ws.cell(i + 1, 4 + j).value) for j in range(9)
            }
            out["resumos"]["aco_estrutura_sem_sobre"]["total"] = num(ws.cell(i + 1, 13).value)
            out["resumos"]["aco_estrutura_com_5pct_sobre"] = {
                BITOLAS[j]: num(ws.cell(i + 2, 4 + j).value) for j in range(9)
            }
            out["resumos"]["aco_estrutura_com_5pct_sobre"]["total"] = num(ws.cell(i + 2, 13).value)
        elif b.startswith("CONSUMO DE CONCRETO"):
            # detectar se é fundação ou estrutura pelo valor da próxima linha
            sem = num(ws.cell(i + 1, 13).value)
            com = num(ws.cell(i + 2, 13).value)
            if sem < 2000:  # fundação
                out["resumos"]["concreto_estacas_sem_sobre"] = sem
                out["resumos"]["concreto_estacas_com_20pct_sobre"] = com
            else:  # estrutura (supra)
                out["resumos"]["concreto_estrutura_sem_sobre"] = sem
                out["resumos"]["concreto_estrutura_com_10pct_sobre"] = com
        elif "Área Estrutural" in b:
            out["resumos"]["area_estrutural_m2"] = num(ws.cell(i, 12).value)
        elif "Espessura média" in b:
            out["resumos"]["espessura_media_m"] = num(ws.cell(i, 12).value)
        elif "Consumo de Aço por V" in b:
            out["resumos"]["aco_por_volume_kg_m3"] = num(ws.cell(i, 12).value)
        elif "Consumo de Aço por Á" in b:
            out["resumos"]["aco_por_area_kg_m2"] = num(ws.cell(i, 12).value)

    return out


def consolidate() -> dict:
    print("Lendo QUANTIDADES BLOCO A...")
    q_a = read_quantidades("A")
    print("Lendo QUANTIDADES BLOCO B...")
    q_b = read_quantidades("B")
    print("Lendo CONTROLE-REV BLOCO A...")
    c_a = read_controle_revisao("A")
    print("Lendo CONTROLE-REV BLOCO B...")
    c_b = read_controle_revisao("B")

    # Validação cruzada: somatório de estacas deve bater com Liberté
    lib_aco = LIBERTE["aco_total_kg"]
    est_a = c_a["estacas"]["aco_total_kg"]
    est_b = c_b["estacas"]["aco_total_kg"]
    soma_estacas = est_a + est_b
    divergencia_aco = abs(soma_estacas - lib_aco) / lib_aco * 100 if lib_aco else 0
    print(f"  Aço estacas Bloco A = {est_a:,.0f} kg")
    print(f"  Aço estacas Bloco B = {est_b:,.0f} kg")
    print(f"  Soma A+B = {soma_estacas:,.0f} kg | LIBERTÉ = {lib_aco:,.0f} kg")
    print(f"  Divergência: {divergencia_aco:.2f}%")

    return {
        "meta": {
            "gerado_em": "2026-04-17",
            "script": "extrair_projetista_1203.py",
            "fontes": {
                "quantidades_bloco_a": q_a["fonte"],
                "quantidades_bloco_b": q_b["fonte"],
                "controle_rev_bloco_a": c_a["fonte"],
                "controle_rev_bloco_b": c_b["fonte"],
                "proposta_liberte": "ESTACAS.jpeg (Proposta 1203-2025-R0)",
            },
        },
        "validacao_cruzada": {
            "aco_estacas_soma_cr": soma_estacas,
            "aco_estacas_liberte": lib_aco,
            "divergencia_pct": round(divergencia_aco, 3),
            "concreto_estacas_soma_cr_sem": (
                c_a["resumos"]["concreto_estacas_sem_sobre"]
                + c_b["resumos"]["concreto_estacas_sem_sobre"]
            ),
            "concreto_estacas_liberte_com_20pct": LIBERTE["concreto_C40_m3"][
                "teorico_mais_20pct_sobreconsumo"
            ],
            "concreto_estacas_liberte_sem_sobre_deduzido": LIBERTE["concreto_C40_m3"][
                "teorico_sem_sobreconsumo_deduzido"
            ],
        },
        "bloco_a": {
            "quantidades": q_a,
            "controle_rev": c_a,
        },
        "bloco_b": {
            "quantidades": q_b,
            "controle_rev": c_b,
        },
        "liberte_fundacao_profunda": LIBERTE,
    }


# ---------------------------------------------------------------------------
# Geração do xlsx humano-legível
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="FFE699")
SECTION_FONT = Font(bold=True, size=12)
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def set_header(ws, row, headers, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
    if widths:
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 30


def set_section(ws, row, text, span):
    ws.cell(row, 1, text)
    ws.cell(row, 1).fill = SECTION_FILL
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).alignment = LEFT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


def fmt_data_row(ws, row, values, totals=False, numfmts=None):
    for j, v in enumerate(values, 1):
        c = ws.cell(row, j, v)
        c.border = BORDER
        if isinstance(v, (int, float)):
            c.alignment = RIGHT
            if numfmts and numfmts.get(j):
                c.number_format = numfmts[j]
        else:
            c.alignment = LEFT
        if totals:
            c.fill = TOTAL_FILL
            c.font = TOTAL_FONT


def build_xlsx(data: dict, out_path: Path):
    wb = openpyxl.Workbook()

    # ---------------- Aba 1: RESUMO ----------------
    ws = wb.active
    ws.title = "RESUMO"
    ws.cell(1, 1, "Projetista 1203 (Rubens Alves) — Fundação Electra consolidado").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(2, 1, f"Gerado: {data['meta']['gerado_em']} | Fonte LIBERTÉ: {LIBERTE['codigo_proposta']}").font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    r = 4
    set_section(ws, r, "1. Fundação Profunda (Estacas Hélice Contínua) — Empreendimento", 8)
    r += 1
    set_header(ws, r, ["Fonte", "Ø (m)", "Qtd (un)", "L médio (m)", "L total (m)",
                        "Concreto (m³)", "Aço ø6.3 (kg)", "Aço ø16 (kg)"],
               widths=[28, 10, 10, 12, 12, 14, 14, 14])
    r += 1

    for lin in LIBERTE["perfuracao"]:
        fmt_data_row(ws, r, [
            "LIBERTÉ (proposta)",
            lin["diametro_m"],
            lin["qtd"],
            lin["L_med_m"],
            lin["L_total_m"],
            round(math.pi * (lin["diametro_m"] / 2) ** 2 * lin["L_total_m"], 2),
            "", "",
        ], numfmts={2: "0.00", 3: "#,##0", 5: "#,##0", 6: "#,##0.00"})
        r += 1
    # Liberté totais + sobreconsumo
    fmt_data_row(ws, r, [
        "LIBERTÉ total (sem sobre)",
        "",
        LIBERTE["perfuracao_total"]["qtd"],
        "",
        LIBERTE["perfuracao_total"]["L_total_m"],
        round(LIBERTE["concreto_C40_m3"]["teorico_sem_sobreconsumo_deduzido"], 2),
        LIBERTE["aco_CA50"][0]["peso_kg"],
        LIBERTE["aco_CA50"][1]["peso_kg"],
    ], totals=True, numfmts={3: "#,##0", 5: "#,##0", 6: "#,##0.00", 7: "#,##0", 8: "#,##0"})
    r += 1
    fmt_data_row(ws, r, [
        "LIBERTÉ total (+20% sobre)", "", "", "", "",
        LIBERTE["concreto_C40_m3"]["teorico_mais_20pct_sobreconsumo"],
        "", "",
    ], totals=True, numfmts={6: "#,##0"})
    r += 2

    # Por bloco via CONTROLE-REV
    set_header(ws, r, ["Fonte", "", "", "", "",
                        "Concreto (m³)", "Aço ø6.3 (kg)", "Aço ø16 (kg)"])
    r += 1
    ba_est = data["bloco_a"]["controle_rev"]["estacas"]
    bb_est = data["bloco_b"]["controle_rev"]["estacas"]
    fmt_data_row(ws, r, [
        "Bloco A (CONTROLE-REV)", "", "", "", "",
        ba_est["concreto_m3"], ba_est["aco_por_bitola"]["6.3"],
        ba_est["aco_por_bitola"]["16"],
    ], numfmts={6: "#,##0.00", 7: "#,##0", 8: "#,##0"})
    r += 1
    fmt_data_row(ws, r, [
        "Bloco B (CONTROLE-REV)", "", "", "", "",
        bb_est["concreto_m3"], bb_est["aco_por_bitola"]["6.3"],
        bb_est["aco_por_bitola"]["16"],
    ], numfmts={6: "#,##0.00", 7: "#,##0", 8: "#,##0"})
    r += 1
    fmt_data_row(ws, r, [
        "Soma A + B", "", "", "", "",
        ba_est["concreto_m3"] + bb_est["concreto_m3"],
        ba_est["aco_por_bitola"]["6.3"] + bb_est["aco_por_bitola"]["6.3"],
        ba_est["aco_por_bitola"]["16"] + bb_est["aco_por_bitola"]["16"],
    ], totals=True, numfmts={6: "#,##0.00", 7: "#,##0", 8: "#,##0"})
    r += 2

    # Validação cruzada
    set_section(ws, r, "Validação cruzada CONTROLE-REV vs LIBERTÉ", 8)
    r += 1
    val = data["validacao_cruzada"]
    fmt_data_row(ws, r, [
        "Aço total estacas", "", "", "", "",
        "", val["aco_estacas_soma_cr"], val["aco_estacas_liberte"],
    ], numfmts={7: "#,##0", 8: "#,##0"})
    ws.cell(r, 6, "CR A+B:").font = Font(bold=True)
    ws.cell(r, 6).alignment = RIGHT
    r += 1
    fmt_data_row(ws, r, [
        "Divergência aço", "", "", "", "",
        "", val["divergencia_pct"], "%",
    ], numfmts={7: "0.000"})
    r += 2

    # Seção Fundação Rasa
    set_section(ws, r, "2. Fundação Rasa / Blocos de Coroamento", 8)
    r += 1
    set_header(ws, r, ["Bloco", "Item", "Fôrma (m²)", "Volume (m³)",
                        "Aço total (kg)", "", "", ""])
    r += 1

    # Bloco A - primeira seção topo (Lajes maciças fundação)
    fra_a = data["bloco_a"]["quantidades"]["fundacao_rasa_topo"]
    if "lajes_macicas_fundacao" in fra_a:
        x = fra_a["lajes_macicas_fundacao"]
        fmt_data_row(ws, r, [
            "A", "Lajes maciças (topo arquivo — Laje Fund?)",
            x["forma_m2"], x["volume_m3"], x["aco_kg"], "", "", "",
        ], numfmts={3: "#,##0.00", 4: "#,##0.00", 5: "#,##0"})
        r += 1

    # Bloco B - zapatas + blocos coroamento
    fra_b = data["bloco_b"]["quantidades"]["fundacao_rasa_topo"]
    for key, label in [("zapatas_isoladas", "Zapatas isoladas"),
                       ("blocos_coroamento", "Blocos de coroamento")]:
        if key in fra_b:
            x = fra_b[key]
            fmt_data_row(ws, r, [
                "B", label, x["forma_m2"], x["volume_m3"], x["aco_kg"],
                "", "", "",
            ], numfmts={3: "#,##0.00", 4: "#,##0.00", 5: "#,##0"})
            r += 1

    # Resumos QUANT-A (soma blocos A+B do resumo)
    r += 1
    resumos_a = data["bloco_a"]["quantidades"]["resumos"]
    if resumos_a:
        set_header(ws, r, ["Resumo QUANT-A (r245-247)", "", "", "Volume (m³)",
                            "Aço (kg)", "", "", ""])
        r += 1
        for label, vals in resumos_a.items():
            fmt_data_row(ws, r, [label, "", "",
                                  vals["volume_m3"], vals["aco_kg"],
                                  "", "", ""],
                         numfmts={4: "#,##0.00", 5: "#,##0"})
            r += 1

    # ---------------- Aba 2-3: Por pavimento Bloco A e B ----------------
    for block_key, label in [("bloco_a", "A"), ("bloco_b", "B")]:
        ws = wb.create_sheet(f"Supra {label} por pav")
        crev = data[block_key]["controle_rev"]
        ws.cell(1, 1, f"Supraestrutura Bloco {label} — aço por bitola + concreto").font = Font(bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        BITOLAS = ["Q92", "5", "6.3", "8", "10", "12.5", "16", "20", "25"]
        set_header(ws, 3, ["Pavimento", "Elemento",
                            *[f"ø{b} (kg)" for b in BITOLAS], "Aço total (kg)",
                            "Concreto (m³)"],
                   widths=[18, 14, *([10] * 9), 14, 14])
        r = 4
        for pav, elems in crev["por_pavimento"].items():
            for elem, vals in elems.items():
                if elem == "_totalizacao":
                    continue
                row_vals = [
                    pav, elem,
                    *[vals["aco_por_bitola"][b] for b in BITOLAS],
                    vals["aco_total_kg"], vals["concreto_m3"],
                ]
                fmt_data_row(ws, r, row_vals,
                             numfmts={**{i: "#,##0" for i in range(3, 13)},
                                      13: "#,##0.00"})
                r += 1
            # totalização do pavimento
            tot = elems.get("_totalizacao")
            if tot:
                fmt_data_row(ws, r, [
                    pav, "TOTAL",
                    *[tot["aco_por_bitola"][b] for b in BITOLAS],
                    tot["aco_total_kg"], tot["concreto_m3"],
                ], totals=True,
                   numfmts={**{i: "#,##0" for i in range(3, 13)}, 13: "#,##0.00"})
                r += 1

    # ---------------- Aba 4: Aço por Bitola consolidado ----------------
    ws = wb.create_sheet("Aço por Bitola")
    ws.cell(1, 1, "Aço CA-50 por bitola — sem sobreconsumo").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    BITOLAS = ["Q92", "5", "6.3", "8", "10", "12.5", "16", "20", "25"]
    set_header(ws, 3, ["Bloco", "Item", *[f"ø{b}" for b in BITOLAS], "Total (kg)"],
               widths=[10, 28, *([9] * 9), 14])
    r = 4
    for block_key, label in [("bloco_a", "A"), ("bloco_b", "B")]:
        crev = data[block_key]["controle_rev"]
        if crev.get("estacas"):
            e = crev["estacas"]
            fmt_data_row(ws, r, [
                label, "Estacas Hélice Contínua",
                *[e["aco_por_bitola"][b] for b in BITOLAS],
                e["aco_total_kg"],
            ], numfmts={i: "#,##0" for i in range(3, 13)})
            r += 1
        res_f = crev["resumos"].get("aco_fundacao_sem_sobre", {})
        if res_f:
            fmt_data_row(ws, r, [
                label, "Fundação (resumo)",
                *[res_f.get(b, 0) for b in BITOLAS],
                res_f.get("total", 0),
            ], totals=True, numfmts={i: "#,##0" for i in range(3, 13)})
            r += 1
        res_e = crev["resumos"].get("aco_estrutura_sem_sobre", {})
        if res_e:
            fmt_data_row(ws, r, [
                label, "Estrutura (supra)",
                *[res_e.get(b, 0) for b in BITOLAS],
                res_e.get("total", 0),
            ], numfmts={i: "#,##0" for i in range(3, 13)})
            r += 1

    # ---------------- Aba 5: Fundação Rasa detalhada ----------------
    ws = wb.create_sheet("Fund. Rasa detalhe")
    ws.cell(1, 1, "Fundação Rasa — detalhe por bloco").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    set_header(ws, 3, ["Bloco", "Item", "Fôrma (m²)", "Volume (m³)",
                        "Aço (kg)", "Fonte"],
               widths=[10, 30, 14, 14, 14, 30])
    r = 4
    # Bloco A
    fra_a = data["bloco_a"]["quantidades"]["fundacao_rasa_topo"]
    if "lajes_macicas_fundacao" in fra_a:
        x = fra_a["lajes_macicas_fundacao"]
        fmt_data_row(ws, r, ["A", "Lajes maciças (fundação/topo arquivo)",
                              x["forma_m2"], x["volume_m3"], x["aco_kg"],
                              "QUANT-A r3"],
                     numfmts={3: "#,##0.00", 4: "#,##0.00", 5: "#,##0"})
        r += 1
    # L1 Blocos A - via CONTROLE-REV
    l1_a = data["bloco_a"]["controle_rev"]["por_pavimento"].get("L1 - Térreo", {}).get("Blocos")
    if l1_a:
        fmt_data_row(ws, r, ["A", "L1 Térreo - Blocos (CR)",
                              "-", l1_a["concreto_m3"], l1_a["aco_total_kg"],
                              "CR-A L1 Blocos"],
                     numfmts={4: "#,##0.00", 5: "#,##0"})
        r += 1
    # Bloco B
    fra_b = data["bloco_b"]["quantidades"]["fundacao_rasa_topo"]
    if "zapatas_isoladas" in fra_b:
        x = fra_b["zapatas_isoladas"]
        fmt_data_row(ws, r, ["B", "Zapatas isoladas",
                              x["forma_m2"], x["volume_m3"], x["aco_kg"],
                              "QUANT-B r3"],
                     numfmts={3: "#,##0.00", 4: "#,##0.00", 5: "#,##0"})
        r += 1
    if "blocos_coroamento" in fra_b:
        x = fra_b["blocos_coroamento"]
        fmt_data_row(ws, r, ["B", "Blocos de coroamento",
                              x["forma_m2"], x["volume_m3"], x["aco_kg"],
                              "QUANT-B r4"],
                     numfmts={3: "#,##0.00", 4: "#,##0.00", 5: "#,##0"})
        r += 1
    l1_b = data["bloco_b"]["controle_rev"]["por_pavimento"].get("L1 - Térreo", {}).get("Blocos")
    if l1_b:
        fmt_data_row(ws, r, ["B", "L1 Térreo - Blocos (CR)",
                              "-", l1_b["concreto_m3"], l1_b["aco_total_kg"],
                              "CR-B L1 Blocos"],
                     numfmts={4: "#,##0.00", 5: "#,##0"})
        r += 1

    wb.save(out_path)
    print(f"xlsx salvo: {out_path}")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = consolidate()

    json_path = OUT_DIR / "projetista-1203-consolidado.json"
    json_path.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str),
                         encoding="utf-8")
    print(f"json salvo: {json_path}")

    xlsx_path = OUT_DIR / "projetista-1203-consolidado.xlsx"
    build_xlsx(data, xlsx_path)

    # resumo pro terminal
    print("\n" + "=" * 60)
    print("RESUMO CONSOLIDADO")
    print("=" * 60)
    est_a = data["bloco_a"]["controle_rev"]["estacas"]
    est_b = data["bloco_b"]["controle_rev"]["estacas"]
    print(f"\nEstacas:")
    print(f"  Bloco A: {est_a['concreto_m3']:>10,.1f} m³ / {est_a['aco_total_kg']:>7,.0f} kg")
    print(f"  Bloco B: {est_b['concreto_m3']:>10,.1f} m³ / {est_b['aco_total_kg']:>7,.0f} kg")
    print(f"  Soma   : {est_a['concreto_m3']+est_b['concreto_m3']:>10,.1f} m³ / {est_a['aco_total_kg']+est_b['aco_total_kg']:>7,.0f} kg")
    print(f"  LIBERTÉ: {LIBERTE['concreto_C40_m3']['teorico_sem_sobreconsumo_deduzido']:>10,.1f} m³ / {LIBERTE['aco_total_kg']:>7,.0f} kg")
    print(f"  LIBERTÉ (+20%): {LIBERTE['concreto_C40_m3']['teorico_mais_20pct_sobreconsumo']:>6,.0f} m³")

    print(f"\nFundação rasa (topo arquivos QUANT):")
    fra_a = data["bloco_a"]["quantidades"]["fundacao_rasa_topo"]
    fra_b = data["bloco_b"]["quantidades"]["fundacao_rasa_topo"]
    if "lajes_macicas_fundacao" in fra_a:
        x = fra_a["lajes_macicas_fundacao"]
        print(f"  A Lajes maciças: vol {x['volume_m3']:,.2f} m³ | aço {x['aco_kg']:,.0f} kg")
    if "zapatas_isoladas" in fra_b:
        x = fra_b["zapatas_isoladas"]
        print(f"  B Zapatas:       vol {x['volume_m3']:,.2f} m³ | aço {x['aco_kg']:,.0f} kg")
    if "blocos_coroamento" in fra_b:
        x = fra_b["blocos_coroamento"]
        print(f"  B Blocos coroam: vol {x['volume_m3']:,.2f} m³ | aço {x['aco_kg']:,.0f} kg")

    print("\nOK.")


if __name__ == "__main__":
    main()
