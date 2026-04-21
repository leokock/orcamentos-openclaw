"""Gera spda-electra-r01.xlsx — primeira entrega formal da disciplina SPDA.

Fontes:
- briefings/spda-r00.json: quantitativos por NBR 5419 (captacao, descidas, equipot,
  aterramento, acessorios)
- quantitativos/listas-materiais/spda/spda.json: PDF Eletrowatts adiciona 12 caixas
  de equipotencializacao (BEL) aluminio 20x20x15cm
- Benchmark Cartesian (27 projetos): mediana R$ 5,14/m2 AC — sanity check

AC Electra (Total Torre A + B): 36.092 m2 (conforme log-execucao sec 17/18).
Benchmark: R$ 5,14/m2 × 36.092 = R$ 185.513. Usar como referencia de total.

Output: disciplinas/spda/spda-electra-r01.xlsx
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra")
BRIEFING = BASE / "briefings" / "spda-r00.json"
PDF_JSON = BASE / "quantitativos" / "listas-materiais" / "spda" / "spda.json"
OUT_XLSX = BASE / "disciplinas" / "spda" / "spda-electra-r01.xlsx"

AC_TOTAL_M2 = 36092  # Torre A + B (bench R$ 5,14/m2)
BENCH_MEDIANA_R_M2 = 5.14

# Localizacao semantica por item SPDA (nao e pavimento estrito pois SPDA e sistema
# vertical que atravessa o predio). Convencao Electra:
#   COBERTURA    — captores e base
#   DESCIDAS     — cabo de cobre nu 50mm2, bracadeiras, buchas, conectores ao longo
#                  das 4 descidas verticais (Terreo -> Cobertura)
#   ANEIS EQUIP  — 5 aneis equipotenciais (distribuidos no prumo — G1/G5/Lazer/Cobertura)
#   ATERRAMENTO  — malha enterrada no Subsolo/Terreo (hastes, cabo 50mm2 malha,
#                  caixas de inspecao, solda exotermica)
#   EQUIP LOCAL  — BEP, barras equipot, cabos 25mm2 (quadros eletricos — provavelmente
#                  Terreo, Casa de Maquinas e quadros de pavimento)
#   ACESSORIOS   — distribuicao geral (conectores, arruelas, fita)
#   BEL APTO     — 12 caixas equipotencializacao (PDF novo) — distribuidas
LOCALIZACAO_ITEM = {
    "07.02.01.001": "COBERTURA",
    "07.02.01.002": "COBERTURA",
    "07.02.01.003": "COBERTURA",
    "07.02.01.004": "COBERTURA",
    "07.02.02.001": "DESCIDAS (Terreo -> Cobertura)",
    "07.02.02.002": "DESCIDAS (Terreo -> Cobertura)",
    "07.02.02.003": "DESCIDAS (Terreo -> Cobertura)",
    "07.02.02.004": "DESCIDAS (topo e base)",
    "07.02.02.005": "DESCIDAS (caixas de inspecao)",
    "07.02.03.001": "ANEIS EQUIPOTENCIAIS (5 niveis)",
    "07.02.03.002": "ANEIS EQUIPOTENCIAIS",
    "07.02.03.003": "ANEIS EQUIPOTENCIAIS",
    "07.02.03.004": "EQUIP LOCAL (quadros QGBT/QGB — Terreo e C.Maq)",
    "07.02.03.005": "EQUIP LOCAL",
    "07.02.03.006": "EQUIP LOCAL (quadros e prumadas)",
    "07.02.04.001": "ATERRAMENTO (subsolo)",
    "07.02.04.002": "ATERRAMENTO (subsolo)",
    "07.02.04.003": "ATERRAMENTO (subsolo)",
    "07.02.04.004": "ATERRAMENTO (subsolo)",
    "07.02.04.005": "ATERRAMENTO (malha enterrada)",
    "07.02.04.006": "ATERRAMENTO (malha)",
    "07.02.04.007": "ATERRAMENTO (malha)",
    "07.02.04.008": "ATERRAMENTO (subsolo — acesso)",
    "07.02.05.001": "ACESSORIOS (distribuicao geral)",
    "07.02.05.002": "ACESSORIOS (distribuicao geral)",
    "07.02.05.003": "ACESSORIOS (distribuicao geral)",
    "07.02.05.004": "ACESSORIOS (descidas)",
    "07.02.03.007": "BEL APTO (12 caixas — distribuicao por BEPs)",
}

# Precos unitarios (R$) — fonte: mercado SC 2026, composicoes SINAPI/ORSE referencia
# Estes PUs sao materiais + instalacao basica. MO vem embutida no indice %.
PU = {
    # Captacao
    "07.02.01.001": {"desc": "Captor Franklin h=0.6m aco inox 304 Ø1/2\"", "pu": 180.0, "fonte": "Mercado SC"},
    "07.02.01.002": {"desc": "Base captor chapa galv 200x200x5mm", "pu": 45.0, "fonte": "SINAPI ref"},
    "07.02.01.003": {"desc": "Chumbador M10x100 galvanizado", "pu": 3.50, "fonte": "Mercado"},
    "07.02.01.004": {"desc": "Conector captor-descida bronze 50mm2", "pu": 28.0, "fonte": "Mercado"},

    # Descidas
    "07.02.02.001": {"desc": "Cabo cobre nu 50mm2 tempera mole NBR 13248", "pu": 65.0, "fonte": "Mercado SC"},
    "07.02.02.002": {"desc": "Bracadeira fixacao cabo 50mm2 galv", "pu": 6.50, "fonte": "Mercado"},
    "07.02.02.003": {"desc": "Bucha/parafuso M8x60 galvanizado", "pu": 2.20, "fonte": "Mercado"},
    "07.02.02.004": {"desc": "Caixa inspecao 300x300x300 alvenaria c/ tampa concr", "pu": 380.0, "fonte": "SINAPI"},
    "07.02.02.005": {"desc": "Conector aparafusado bronze 50mm2", "pu": 22.0, "fonte": "Mercado"},

    # Equipotencializacao
    "07.02.03.001": {"desc": "Cabo cobre nu 35mm2 tempera mole NBR 13248", "pu": 48.0, "fonte": "Mercado"},
    "07.02.03.002": {"desc": "Conector aparafusado bronze 35mm2", "pu": 18.0, "fonte": "Mercado"},
    "07.02.03.003": {"desc": "Bracadeira fixacao cabo 35mm2 galv", "pu": 5.80, "fonte": "Mercado"},
    "07.02.03.004": {"desc": "Barra equipot cobre 20x3mm 300mm", "pu": 95.0, "fonte": "Mercado"},
    "07.02.03.005": {"desc": "Conector barra-cabo bronze 35mm2", "pu": 24.0, "fonte": "Mercado"},
    "07.02.03.006": {"desc": "Cabo cobre flex 25mm2 isol 750V verde", "pu": 28.0, "fonte": "Mercado"},

    # Aterramento
    "07.02.04.001": {"desc": "Haste aterramento cobreada Ø5/8 x 2.4m NBR 13571", "pu": 55.0, "fonte": "Mercado"},
    "07.02.04.002": {"desc": "Conector haste-cabo bronze compressao", "pu": 18.0, "fonte": "Mercado"},
    "07.02.04.003": {"desc": "Caixa inspecao 400x400x400 alvenaria c/ tampa", "pu": 520.0, "fonte": "SINAPI"},
    "07.02.04.004": {"desc": "Solda exotermica cabo 50mm2 + haste 5/8", "pu": 42.0, "fonte": "Mercado"},
    "07.02.04.005": {"desc": "Cabo cobre nu 50mm2 malha aterramento NBR 13248", "pu": 65.0, "fonte": "Mercado SC"},
    "07.02.04.006": {"desc": "Conector aparafusado bronze 50mm2 malha", "pu": 22.0, "fonte": "Mercado"},
    "07.02.04.007": {"desc": "Solda exotermica cabo 50mm2 malha", "pu": 35.0, "fonte": "Mercado"},
    "07.02.04.008": {"desc": "Caixa inspecao 400x400x400 tampa articulada", "pu": 580.0, "fonte": "SINAPI"},

    # Acessorios
    "07.02.05.001": {"desc": "Conector paralelo bronze 50mm2 derivacao", "pu": 18.0, "fonte": "Mercado"},
    "07.02.05.002": {"desc": "Conector cruzado bronze 35-50mm2", "pu": 24.0, "fonte": "Mercado"},
    "07.02.05.003": {"desc": "Arruela pressao bronze M10", "pu": 1.20, "fonte": "Mercado"},
    "07.02.05.004": {"desc": "Fita advertencia SPDA PVC", "pu": 2.80, "fonte": "Mercado"},

    # NOVO - do PDF Eletrowatts 07/04/2026
    "07.02.03.007": {"desc": "Caixa equipotencializacao (BEL) aluminio 20x20x15cm pintura branca", "pu": 165.0, "fonte": "PDF Eletrowatts 07/04"},
}


def load_items() -> list[dict]:
    briefing = json.loads(BRIEFING.read_text(encoding="utf-8"))
    pdf = json.loads(PDF_JSON.read_text(encoding="utf-8"))

    rows = []
    # Itens do briefing (NBR 5419)
    secoes = [
        ("Captacao", briefing["quantitativos"]["captacao"]),
        ("Descidas", briefing["quantitativos"]["descidas"]),
        ("Equipotencializacao", briefing["quantitativos"]["equipotencializacao"]),
        ("Aterramento", briefing["quantitativos"]["aterramento"]),
        ("Acessorios", briefing["quantitativos"]["acessorios"]),
    ]
    for subgrupo, itens in secoes:
        for it in itens:
            cod = it["codigo"]
            pu_info = PU.get(cod, {"desc": it["descricao"], "pu": 0, "fonte": "-"})
            rows.append({
                "subgrupo": subgrupo,
                "codigo": cod,
                "descricao": it["descricao"],
                "especificacao": it["especificacao"],
                "unidade": it["unidade"],
                "qtd": it["quantidade"],
                "pu": pu_info["pu"],
                "fonte_qtd": "NBR 5419 (briefing R00)",
                "fonte_pu": pu_info["fonte"],
                "observacao": it["observacao"],
            })

    # Itens novos do PDF
    pdf_doc = pdf["documentos"][0]
    for it in pdf_doc["itens"]:
        cod = "07.02.03.007"
        rows.append({
            "subgrupo": "Equipotencializacao",
            "codigo": cod,
            "descricao": it["produto"],
            "especificacao": f"Dimensoes {it['dimensoes']}",
            "unidade": it["unidade"],
            "qtd": int(it["qtd"]),
            "pu": PU[cod]["pu"],
            "fonte_qtd": "PDF Eletrowatts 07/04/2026",
            "fonte_pu": PU[cod]["fonte"],
            "observacao": "NOVO — nao constava no briefing R00",
        })

    return rows


def write_xlsx(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SPDA R01"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subtot_fill = PatternFill("solid", fgColor="FFF2CC")
    total_fill = PatternFill("solid", fgColor="C6E0B4")
    new_fill = PatternFill("solid", fgColor="FFE699")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    # Metadata rows
    ws["A1"] = "THOZEN - RES. ELECTRA TOWERS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Disciplina: 11 SPDA - Sistema de Protecao Contra Descargas Atmosfericas"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = "Revisao: R01 | Data: 2026-04-14 | AC Total: 36.092 m2 (Torre A+B)"
    ws["A4"] = "Fontes: briefing SPDA R00 (NBR 5419) + PDF Eletrowatts 07/04/2026 + benchmark Cartesian (27 proj, mediana R$ 5,14/m2)"
    ws["A4"].font = Font(italic=True, size=9, color="666666")

    # Cabecalho da tabela (linha 6)
    headers = ["Subgrupo", "Localizacao no predio", "Codigo", "Descricao", "Especificacao", "Unid", "Qtd",
               "PU (R$)", "Total (R$)", "Fonte Qtd", "Fonte PU", "Observacao"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    # Dados
    row = 7
    current_subgrupo = None
    subgrupo_start = row
    subgrupo_subtotal_rows = []

    for item in rows:
        if item["subgrupo"] != current_subgrupo:
            # Fecha subgrupo anterior
            if current_subgrupo is not None:
                _write_subtotal(ws, row, subgrupo_start, row - 1, current_subgrupo, subtot_fill, border, center, right)
                subgrupo_subtotal_rows.append(row)
                row += 1
            current_subgrupo = item["subgrupo"]
            subgrupo_start = row

        ws.cell(row=row, column=1, value=item["subgrupo"]).alignment = left
        ws.cell(row=row, column=2, value=LOCALIZACAO_ITEM.get(item["codigo"], "-")).alignment = left
        ws.cell(row=row, column=3, value=item["codigo"]).alignment = center
        ws.cell(row=row, column=4, value=item["descricao"]).alignment = left
        ws.cell(row=row, column=5, value=item["especificacao"]).alignment = left
        ws.cell(row=row, column=6, value=item["unidade"]).alignment = center
        ws.cell(row=row, column=7, value=item["qtd"]).alignment = right
        ws.cell(row=row, column=8, value=item["pu"]).alignment = right
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=9, value=f'=G{row}*H{row}').alignment = right
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        ws.cell(row=row, column=10, value=item["fonte_qtd"]).alignment = left
        ws.cell(row=row, column=11, value=item["fonte_pu"]).alignment = left
        ws.cell(row=row, column=12, value=item["observacao"]).alignment = left

        for col in range(1, 13):
            ws.cell(row=row, column=col).border = border

        # Destacar itens NOVOS do PDF
        if "NOVO" in item["observacao"]:
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = new_fill

        row += 1

    # Fecha ultimo subgrupo
    _write_subtotal(ws, row, subgrupo_start, row - 1, current_subgrupo, subtot_fill, border, center, right)
    subgrupo_subtotal_rows.append(row)
    row += 1

    # Linha em branco
    row += 1

    # Total geral (subtotais)
    ws.cell(row=row, column=1, value="TOTAL GERAL SPDA (material)").font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).alignment = left
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ref = "+".join(f"I{r}" for r in subgrupo_subtotal_rows)
    ws.cell(row=row, column=9, value=f"={ref}").number_format = '#,##0.00'
    ws.cell(row=row, column=9).font = Font(bold=True, size=11)
    ws.cell(row=row, column=9).fill = total_fill
    ws.cell(row=row, column=9).alignment = right
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = border
        if col != 9:
            ws.cell(row=row, column=col).fill = total_fill
    total_row = row
    row += 1

    # MO instalacao (30% sobre material — padrao disciplina SPDA)
    row += 1
    ws.cell(row=row, column=1, value="Mao de obra instalacao (30% sobre material)").font = Font(bold=True, size=10)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=9, value=f"=I{total_row}*0.30").number_format = '#,##0.00'
    ws.cell(row=row, column=9).font = Font(bold=True)
    ws.cell(row=row, column=9).fill = subtot_fill
    mo_row = row
    row += 1

    # Total com MO
    ws.cell(row=row, column=1, value="TOTAL SPDA (material + MO)").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=9, value=f"=I{total_row}+I{mo_row}").number_format = '#,##0.00'
    ws.cell(row=row, column=9).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=9).fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = border
    total_mo_row = row
    row += 2

    # Verificacao benchmark
    ws.cell(row=row, column=1, value="VERIFICACAO BENCHMARK").font = Font(bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="R$/m2 AC (nosso)").alignment = left
    ws.cell(row=row, column=9, value=f"=I{total_mo_row}/{AC_TOTAL_M2}").number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value=f"R$/m2 AC (mediana 27 proj Cartesian)").alignment = left
    ws.cell(row=row, column=9, value=BENCH_MEDIANA_R_M2).number_format = '#,##0.00'
    row += 1
    ws.cell(row=row, column=1, value="Area construida total (m2)").alignment = left
    ws.cell(row=row, column=9, value=AC_TOTAL_M2).number_format = '#,##0'
    row += 1
    ws.cell(row=row, column=1, value="Benchmark total (mediana × AC)").alignment = left
    ws.cell(row=row, column=9, value=BENCH_MEDIANA_R_M2 * AC_TOTAL_M2).number_format = '#,##0.00'

    # Larguras de coluna
    widths = [18, 30, 15, 42, 38, 6, 10, 12, 14, 22, 22, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[6].height = 28
    ws.freeze_panes = "A7"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"gravado: {OUT_XLSX}")


def _write_subtotal(ws, row, first, last, subgrupo_label, fill, border, center, right):
    ws.cell(row=row, column=1, value=f"Subtotal {subgrupo_label}").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = center
    for col in range(2, 9):
        ws.cell(row=row, column=col).fill = fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=9, value=f"=SUM(I{first}:I{last})")
    ws.cell(row=row, column=9).number_format = '#,##0.00'
    ws.cell(row=row, column=9).font = Font(bold=True)
    ws.cell(row=row, column=9).fill = fill
    ws.cell(row=row, column=9).alignment = right
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = border
        ws.cell(row=row, column=col).fill = fill


def main():
    rows = load_items()
    print(f"total itens: {len(rows)}")
    write_xlsx(rows)


if __name__ == "__main__":
    main()
