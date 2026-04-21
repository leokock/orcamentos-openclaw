#!/usr/bin/env python3
"""Gera aba 'HIDROSSANITÁRIO v2' como lista geral precificada dos 187 itens LM348."""
import sys, io, re, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra"
INPUT = f"{BASE}/disciplinas/hidraulico/hidraulico-electra-lm348-r01.xlsx"
OUTPUT = f"{BASE}/disciplinas/hidraulico/hidraulico-electra-lm348-r02.xlsx"
AC = 36092

# ============================================================
# PREÇOS UNITÁRIOS
# ============================================================

# Tubos água fria — R$/metro (Belle Ville)
PU_TUBO_AF = {
    '25': 2.57, '32': 7.80, '40': 12.83, '50': 10.47,
    '60': 19.05, '75': 27.48,
}

# Tubos água quente CPVC — R$/metro (Belle Ville)
PU_TUBO_AQ_CPVC = {'15': 17.63, '22': 21.30, '28': 12.88}

# Tubos PPR — R$/metro (Belle Ville)
PU_TUBO_PPR = {'90': 85.00}

# Tubos esgoto — R$/barra (Santa Maria). Converter pra R$/m dividindo por 6.
PU_TUBO_ESG_NORMAL = {
    '40': 5.64/6, '50': 9.60/6, '75': 13.86/6, '100': 15.64/6, '150': 40.01/6,
}
PU_TUBO_ESG_REFORC = {
    '50': 16.81/6, '75': 22.03/6, '100': 38.60/6, '150': 78.45/6,
}

# Conexões esgoto Normal — R$/un (Santa Maria)
PU_CX_ESG_N = {
    ('joelho 90', '40'): 2.00, ('joelho 90', '50'): 2.69, ('joelho 90', '100'): 8.89,
    ('joelho 45', '40'): 2.29, ('joelho 45', '50'): 3.35, ('joelho 45', '75'): 7.93,
    ('joelho 45', '100'): 8.83, ('joelho 45', '150'): 62.93,
    ('juncao simples', '40'): 5.00, ('juncao simples', '50'): 7.58,
    ('juncao simples', '100'): 23.13, ('juncao simples', '150'): 142.78,
    ('luva simples', '40'): 2.00, ('luva simples', '50'): 3.09, ('luva simples', '75'): 5.84,
    ('luva simples', '100'): 6.78, ('luva simples', '150'): 32.81,
    ('bucha reducao', '50'): 2.47,
    ('reducao excentrica', '75'): 9.26, ('reducao excentrica', '100'): 9.26,
    ('curva 90', '50'): 5.80, ('curva 90', '150'): 35.00,
    ('te inspecao', '75'): 18.00, ('te inspecao', '100'): 28.00, ('te inspecao', '150'): 85.00,
}

# Conexões esgoto Reforçada — R$/un (Santa Maria)
PU_CX_ESG_R = {
    ('joelho 90', '50'): 9.44, ('joelho 90', '100'): 33.28, ('joelho 90', '150'): 110.58,
    ('joelho 45', '50'): 7.59, ('joelho 45', '75'): 17.71, ('joelho 45', '100'): 24.67,
    ('joelho 45', '150'): 85.84,
    ('juncao simples', '50'): 15.00, ('juncao simples', '100'): 62.72,
    ('juncao simples', '150'): 142.78,
    ('luva simples', '50'): 9.04, ('luva simples', '75'): 12.00,
    ('luva simples', '100'): 15.52, ('luva simples', '150'): 46.11,
    ('reducao excentrica', '100'): 15.00, ('reducao excentrica', '150'): 35.00,
    ('bucha reducao', '50'): 4.50,
}

# Conexões água fria PVC — R$/un (Belle Ville)
PU_CX_AF = {
    'joelho 90 soldavel': 0.78, 'joelho 90 reducao soldavel': 1.50,
    'joelho 90 roscavel': 2.50,
    'joelho 90 soldavel bucha latao': 4.50, 'joelho 90 soldavel roscavel': 3.50,
    'joelho 45 soldavel': 1.20,
    'te soldavel': 1.55, 'te reducao soldavel': 2.50,
    'te soldavel bucha latao': 6.00, 'te soldavel roscavel': 5.00,
    'uniao soldavel': 4.53,
    'bucha reducao soldavel curta': 1.66, 'bucha reducao soldavel longa': 2.50,
    'registro esfera soldavel': 31.72,
    'adaptador soldavel curto': 4.08,
    'curva transposicao soldavel': 9.80,
    'luva soldavel': 2.50, 'luva soldavel roscavel': 2.50,
}

# Conexões CPVC FlowGuard água quente — R$/un (Belle Ville + estimativas)
PU_CX_AQ = {
    'joelho 90 ff': 3.59, 'joelho 45 ff': 2.25,
    'joelho 90 transicao ff': 5.50,
    'te fff': 4.50, 'te misturador fff': 6.00, 'te reducao fff': 5.00,
    'luva transicao ff': 3.50, 'conector transicao fm': 4.50,
    'curva transposicao ff': 6.00, 'bucha reducao mf': 3.00,
}

# Conexões PPR — R$/un (Belle Ville)
PU_CX_PPR = {
    'joelho 90': 15.00, 'joelho 45': 12.00, 'te': 25.00,
    'adaptador transicao macho': 20.00, 'adaptador transicao femea': 20.00,
}

# Acessórios — R$/un (estimativas mercado)
PU_ACESS = {
    'caixa sifonada 100': 28.00, 'caixa sifonada 150': 42.00,
    'ralo seco': 18.00, 'terminal ventilacao 50': 8.00, 'terminal ventilacao 100': 12.00,
    'anel vedacao oring 50': 1.50, 'anel vedacao oring 75': 2.00,
    'anel vedacao oring 100': 2.50, 'anel vedacao oring 150': 3.50,
    'anel vedacao vaso sanitario': 8.00,
}

# Equipamentos — R$/un (Belle Ville + mercado)
PU_EQUIP = {
    'hidrometro unijato': 350.00, 'hidrometro multijato': 547.30,
    'filtro pluvial vf1': 2800.00, 'chave fluxo': 250.00,
    'skid 2 pressurizadores ca plus': 18500.00,
    'skid 1 pressurizador sp 12': 8500.00,
    'skid 2 motobombas recalque': 45000.00,
    'valvula redutora pressao 42 lp': 1200.00,
    'valvula redutora pressao pilotada 420': 2800.00,
    'valvula controladora bomba 740q': 3500.00,
    'valvula controladora nivel 210': 2500.00,
    'valvula alivio seguranca 430': 1800.00,
    'valvula retencao va 407': 850.00,
    'valvula pe crivo': 180.00,
    'filtro t compact': 350.00,
    'ventosa triplice c30 1': 650.00, 'ventosa triplice c30 2': 950.00,
}

# ETE — R$/un (estimativas mercado)
PU_ETE = {
    'caixa gordura 1.00': 3500.00, 'caixa gordura 0.60': 1800.00,
    'caixa distribuicao': 2500.00,
    'tampa inspecao 0.60': 350.00, 'tampa inspecao 1.0': 550.00,
    'tampa inspecao 0.80': 450.00,
    'tanque anaerobio': 35000.00, 'tanque anoxico': 38000.00,
    'tanque aerobio': 38000.00, 'tanque decantador': 32000.00,
    'tanque desinfeccao': 25000.00, 'tanque descarte lodo': 38000.00,
    'difusor ar': 850.00, 'compressor radial': 12000.00,
    'medidor vazao': 3500.00, 'painel controle': 4500.00,
    'contactora': 800.00, 'timer digital': 250.00,
    'bomba retorno lodo': 5500.00,
}

# ============================================================
# HELPERS
# ============================================================

def norm(s):
    s = str(s).lower().strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    # Remove special chars that break matching
    for ch in ['°', 'º', 'ø', 'Ø', '"', '"', '"', '\'', '(', ')', ',', '.', '-', '/']:
        s = s.replace(ch, ' ')
    s = s.replace('\u00e7','c').replace('\u00e3','a').replace('\u00e1','a').replace('\u00e9','e')
    s = s.replace('\u00ed','i').replace('\u00f3','o').replace('\u00fa','u').replace('\u00f4','o')
    s = s.replace('\u00ea','e').replace('\u00e2','a')
    # Collapse multiple spaces
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def extract_diam(dim_str):
    """Extrai primeiro diâmetro de '25Ø-25Ø' ou 'Ø40mm'."""
    if not dim_str:
        return ''
    m = re.search(r'(\d+)', str(dim_str))
    return m.group(1) if m else ''

def extract_second_diam(dim_str):
    """Extrai segundo diâmetro de '100Ø-50Ø'."""
    if not dim_str:
        return ''
    nums = re.findall(r'(\d+)', str(dim_str))
    return nums[1] if len(nums) >= 2 else ''

def classify_item(slug, desc, dim):
    """Retorna (subgrupo, sistema) para cada item."""
    nd = norm(desc)

    if slug == 'hidro-tubulacoes':
        if 'esgoto' in nd and 'reforcada' in nd:
            return 'TUBULAÇÕES ESGOTO REFORÇADA', 'ESG_R'
        elif 'esgoto' in nd:
            return 'TUBULAÇÕES ESGOTO NORMAL', 'ESG_N'
        elif 'cpvc' in nd or 'flowguard' in nd:
            return 'TUBULAÇÕES ÁGUA QUENTE', 'AQ'
        elif 'ppr' in nd:
            return 'TUBULAÇÕES ÁGUA QUENTE', 'PPR'
        else:
            return 'TUBULAÇÕES ÁGUA FRIA', 'AF'

    elif slug == 'hidro-ete':
        return 'ETE / TRATAMENTO', 'ETE'

    elif slug == 'hidro-valvulas':
        return 'EQUIPAMENTOS', 'EQUIP'

    elif slug == 'hidro-conexoes':
        if 'esgoto' in nd and 'reforcada' in nd:
            return 'CONEXÕES ESGOTO REFORÇADA', 'CX_ESG_R'
        elif 'esgoto' in nd:
            return 'CONEXÕES ESGOTO NORMAL', 'CX_ESG_N'
        elif 'cpvc' in nd or 'flowguard' in nd or 'super cpvc' in nd:
            return 'CONEXÕES ÁGUA QUENTE', 'CX_AQ'
        elif 'ppr' in nd:
            return 'CONEXÕES PPR', 'CX_PPR'
        elif any(x in nd for x in ['caixa sifonada', 'ralo seco', 'terminal ventilacao',
                                    'anel vedacao', 'anel de vedacao']):
            return 'ACESSÓRIOS', 'ACESS'
        elif "caixa d'agua" in nd or 'polietileno' in nd:
            return 'EQUIPAMENTOS', 'EQUIP'
        elif 'adaptador longo' in nd and 'flange' in nd:
            return 'ACESSÓRIOS', 'ACESS'
        else:
            return 'CONEXÕES ÁGUA FRIA', 'CX_AF'

    return 'OUTROS', 'OUTROS'

def tube_multiplier(desc):
    """Retorna fator de conversão barras→metros."""
    nd = norm(desc)
    if 'tubo de 6m' in nd or '6m)' in nd:
        return 6
    elif 'tubo de 3m' in nd or '3m)' in nd:
        return 3
    return 1

def find_pu(sistema, desc, dim):
    """Busca preço unitário. Retorna (pu, fonte) ou (None, 'A PRECIFICAR')."""
    nd = norm(desc)
    d1 = extract_diam(dim)
    d2 = extract_second_diam(dim)

    if sistema == 'AF':
        return PU_TUBO_AF.get(d1, (None,)), 'Belle Ville' if d1 in PU_TUBO_AF else 'A PRECIFICAR'

    if sistema == 'AQ':
        if 'cpvc' in nd or 'flowguard' in nd:
            v = PU_TUBO_AQ_CPVC.get(d1)
            if v: return v, 'Belle Ville'
        if 'ppr' in nd:
            v = PU_TUBO_PPR.get(d1)
            if v: return v, 'Belle Ville'
        return None, 'A PRECIFICAR'

    if sistema == 'PPR':
        v = PU_TUBO_PPR.get(d1)
        if v: return v, 'Belle Ville'
        return None, 'A PRECIFICAR'

    if sistema == 'ESG_N':
        v = PU_TUBO_ESG_NORMAL.get(d1)
        if v: return v, 'Santa Maria'
        return None, 'A PRECIFICAR'

    if sistema == 'ESG_R':
        v = PU_TUBO_ESG_REFORC.get(d1)
        if v: return v, 'Santa Maria'
        return None, 'A PRECIFICAR'

    if sistema == 'CX_ESG_N':
        tipo = _esg_conn_type(nd)
        v = PU_CX_ESG_N.get((tipo, d1))
        if v: return v, 'Santa Maria'
        if d2:
            v = PU_CX_ESG_N.get((tipo, d2))
            if v: return v, 'Santa Maria'
        # Fallback: use price of larger diameter for juncao reducao
        if tipo == 'juncao reducao':
            v = PU_CX_ESG_N.get(('juncao simples', d1))
            if v: return v, 'Santa Maria (est.)'
            v = PU_CX_ESG_N.get(('juncao simples', d2))
            if v: return v * 1.2, 'Estimativa'  # add 20% for reduction fitting
            # Estimate by diameter range
            return 15.00, 'Estimativa'
        return None, 'A PRECIFICAR'

    if sistema == 'CX_ESG_R':
        tipo = _esg_conn_type(nd)
        v = PU_CX_ESG_R.get((tipo, d1))
        if v: return v, 'Santa Maria'
        if d2:
            v = PU_CX_ESG_R.get((tipo, d2))
            if v: return v, 'Santa Maria'
        # Fallback: joelho 90 reforc DN75 ~ 20.00
        if tipo == 'joelho 90' and d1 == '75':
            return 20.00, 'Estimativa'
        if tipo == 'juncao reducao':
            v = PU_CX_ESG_R.get(('juncao simples', d1))
            if v: return v, 'Santa Maria (est.)'
            # By larger diameter
            if d1 == '150': return 142.78, 'Estimativa'
            if d1 == '100': return 62.72, 'Estimativa'
            if d1 == '75': return 25.00, 'Estimativa'
        return None, 'A PRECIFICAR'

    if sistema == 'CX_AF':
        # Try exact-ish key match first
        for key, pu in PU_CX_AF.items():
            if key in nd:
                return pu, 'Belle Ville'
        # Fuzzy fallbacks for common AF connections
        if 'joelho 90' in nd and 'soldavel' in nd and 'bucha latao' in nd:
            return 4.50, 'Belle Ville'
        if 'joelho 90' in nd and 'soldavel' in nd and 'roscavel' in nd:
            return 3.50, 'Belle Ville'
        if 'joelho 90' in nd and 'roscavel' in nd:
            return 2.50, 'Belle Ville'
        if 'joelho 90' in nd and 'reducao' in nd and 'soldavel' in nd:
            return 1.50, 'Belle Ville'
        if 'joelho 90' in nd and 'soldavel' in nd:
            return 0.78, 'Belle Ville'
        if 'joelho 45' in nd and 'soldavel' in nd:
            return 1.20, 'Belle Ville'
        if 'te' in nd and 'reducao' in nd and 'soldavel' in nd:
            return 2.50, 'Belle Ville'
        if 'te' in nd and 'soldavel' in nd and 'bucha latao' in nd:
            return 6.00, 'Belle Ville'
        if 'te' in nd and 'soldavel' in nd and 'roscavel' in nd:
            return 5.00, 'Belle Ville'
        if 'te' in nd and 'soldavel' in nd:
            return 1.55, 'Belle Ville'
        if 'bucha' in nd and 'reducao' in nd and 'longa' in nd:
            return 2.50, 'Belle Ville'
        if 'bucha' in nd and 'reducao' in nd and 'curta' in nd:
            return 1.66, 'Belle Ville'
        if 'bucha' in nd and 'reducao' in nd:
            return 1.66, 'Belle Ville'
        if 'curva' in nd and 'transposicao' in nd:
            return 9.80, 'Belle Ville'
        if 'adaptador' in nd and 'soldavel' in nd:
            return 4.08, 'Belle Ville'
        if 'registro' in nd and 'esfera' in nd:
            return 31.72, 'Belle Ville'
        if 'uniao' in nd and 'soldavel' in nd:
            return 4.53, 'Belle Ville'
        if 'luva' in nd and 'soldavel' in nd:
            return 2.50, 'Belle Ville'
        if 'terminal' in nd and 'ventilacao' in nd:
            d = extract_diam(dim)
            if d == '100': return 12.00, 'Estimativa'
            return 8.00, 'Estimativa'
        return None, 'A PRECIFICAR'

    if sistema == 'CX_AQ':
        if 'joelho 90' in nd and 'transicao' in nd:
            return 5.50, 'Estimativa'
        if 'joelho 90' in nd:
            return 3.59, 'Estimativa'
        if 'joelho 45' in nd:
            return 2.25, 'Estimativa'
        if 'te' in nd and 'misturador' in nd:
            return 6.00, 'Estimativa'
        if 'te' in nd and 'reducao' in nd:
            return 5.00, 'Estimativa'
        if 'te' in nd:
            return 4.50, 'Estimativa'
        if 'luva' in nd and 'transicao' in nd:
            return 3.50, 'Estimativa'
        if 'conector' in nd and 'transicao' in nd:
            return 4.50, 'Estimativa'
        if 'curva' in nd and 'transposicao' in nd:
            return 6.00, 'Estimativa'
        if 'bucha' in nd and 'reducao' in nd:
            return 3.00, 'Estimativa'
        return None, 'A PRECIFICAR'

    if sistema == 'CX_PPR':
        if 'joelho 90' in nd:
            return 15.00, 'Belle Ville'
        if 'joelho 45' in nd:
            return 12.00, 'Belle Ville'
        if 'te' in nd:
            return 25.00, 'Belle Ville'
        if 'adaptador' in nd and 'macho' in nd:
            return 20.00, 'Belle Ville'
        if 'adaptador' in nd and 'femea' in nd:
            return 20.00, 'Belle Ville'
        if 'adaptador' in nd:
            return 20.00, 'Belle Ville'
        return None, 'A PRECIFICAR'

    if sistema == 'ACESS':
        d = extract_diam(dim)
        if 'caixa sifonada' in nd:
            if '150' in nd or d == '150':
                return 42.00, 'Estimativa'
            return 28.00, 'Estimativa'
        if 'ralo seco' in nd:
            return 18.00, 'Estimativa'
        if 'terminal' in nd and 'ventilacao' in nd:
            if d == '100': return 12.00, 'Estimativa'
            return 8.00, 'Estimativa'
        if 'anel' in nd and 'vedacao' in nd and 'vaso' in nd:
            return 8.00, 'Estimativa'
        if 'anel' in nd and 'vedacao' in nd and 'oring' in nd:
            if d == '150': return 3.50, 'Estimativa'
            if d == '100': return 2.50, 'Estimativa'
            if d == '75': return 2.00, 'Estimativa'
            return 1.50, 'Estimativa'
        if 'adaptador longo' in nd and 'flange' in nd:
            return 45.00, 'Estimativa'
        return None, 'A PRECIFICAR'

    if sistema == 'EQUIP':
        if 'hidrometro' in nd and 'unijato' in nd:
            return 350.00, 'Mercado'
        if 'hidrometro' in nd and 'multijato' in nd:
            return 547.30, 'Belle Ville'
        if 'filtro pluvial' in nd or 'vf1' in nd:
            return 2800.00, 'Mercado'
        if 'chave' in nd and 'fluxo' in nd:
            return 250.00, 'Mercado'
        if 'skid' in nd and 'pressurizador' in nd and '02' in nd:
            return 18500.00, 'Mercado'
        if 'skid' in nd and 'pressurizador' in nd:
            return 8500.00, 'Mercado'
        if 'skid' in nd and 'motobomba' in nd:
            return 45000.00, 'Mercado'
        if 'valvula redutora' in nd and '42 lp' in nd:
            return 1200.00, 'Mercado'
        if 'valvula redutora' in nd and 'pilotada' in nd:
            return 2800.00, 'Mercado'
        if 'valvula redutora' in nd:
            return 1200.00, 'Mercado'
        if 'valvula controladora' in nd and 'bomba' in nd:
            return 3500.00, 'Mercado'
        if 'valvula controladora' in nd and 'nivel' in nd:
            return 2500.00, 'Mercado'
        if 'valvula' in nd and 'alivio' in nd:
            return 1800.00, 'Mercado'
        if 'valvula' in nd and 'retencao' in nd:
            return 850.00, 'Mercado'
        if 'valvula' in nd and 'pe' in nd and 'crivo' in nd:
            return 180.00, 'Mercado'
        if 'filtro' in nd and ('t compact' in nd or 'tipo' in nd):
            return 350.00, 'Mercado'
        if 'ventosa' in nd and 'triplice' in nd:
            if 'de 2' in nd or 'c30 de 2' in nd or '2 ' in nd.split('c30')[-1] if 'c30' in nd else False:
                return 950.00, 'Mercado'
            return 650.00, 'Mercado'
        # Fallback
        for key, pu in [('reservatorio polietileno', 4200), ("caixa d agua", 4200)]:
            if key in nd:
                return pu, 'Belle Ville'
        return None, 'A PRECIFICAR'

    if sistema == 'ETE':
        if 'caixa' in nd and 'gordura' in nd:
            if '1 00' in nd or '1 0' in nd.split('x')[0]:
                return 3500.00, 'Estimativa'
            return 1800.00, 'Estimativa'
        if 'caixa' in nd and 'distribuicao' in nd:
            return 2500.00, 'Estimativa'
        if 'tampa' in nd and 'inspecao' in nd:
            if '1 0' in nd:
                return 550.00, 'Estimativa'
            if '0 80' in nd or '0 8' in nd:
                return 450.00, 'Estimativa'
            return 350.00, 'Estimativa'
        if 'tanque' in nd and 'anaerobio' in nd:
            return 35000.00, 'Estimativa'
        if 'tanque' in nd and 'anoxico' in nd:
            return 38000.00, 'Estimativa'
        if 'tanque' in nd and 'aerobio' in nd:
            return 38000.00, 'Estimativa'
        if 'tanque' in nd and 'decantador' in nd:
            return 32000.00, 'Estimativa'
        if 'tanque' in nd and 'desinfeccao' in nd:
            return 25000.00, 'Estimativa'
        if 'tanque' in nd and 'descarte' in nd:
            return 38000.00, 'Estimativa'
        if 'difusor' in nd:
            return 850.00, 'Estimativa'
        if 'compressor' in nd:
            return 12000.00, 'Estimativa'
        if 'medidor' in nd and 'vazao' in nd:
            return 3500.00, 'Estimativa'
        if 'painel' in nd and 'controle' in nd:
            return 4500.00, 'Estimativa'
        if 'contactora' in nd:
            return 800.00, 'Estimativa'
        if 'timer' in nd:
            return 250.00, 'Estimativa'
        if 'bomba' in nd and 'lodo' in nd:
            return 5500.00, 'Estimativa'
        return None, 'A PRECIFICAR'

    return None, 'A PRECIFICAR'

def _esg_conn_type(nd):
    if 'juncao simples com reducao' in nd or 'juncao simples com red' in nd:
        return 'juncao reducao'
    if 'juncao simples' in nd:
        return 'juncao simples'
    if 'te de inspecao' in nd or 'te inspecao' in nd:
        return 'te inspecao'
    if 'joelho 90' in nd:
        return 'joelho 90'
    if 'joelho 45' in nd:
        return 'joelho 45'
    if 'luva simples' in nd:
        return 'luva simples'
    if 'bucha' in nd and 'reducao' in nd:
        return 'bucha reducao'
    if 'reducao excentrica' in nd:
        return 'reducao excentrica'
    if 'curva 90' in nd:
        return 'curva 90'
    return 'outro'

# ============================================================
# MAIN
# ============================================================

def main():
    wb = openpyxl.load_workbook(INPUT)
    flat_ws = wb['Flat por PDF']

    # 1. Ler todos os 187 itens do Flat
    items = []
    for r in range(4, flat_ws.max_row + 1):
        slug = flat_ws.cell(r, 1).value
        if not slug:
            continue
        desc = str(flat_ws.cell(r, 9).value or '')
        dim = str(flat_ws.cell(r, 10).value or '')
        marca = str(flat_ws.cell(r, 11).value or '')
        qty_raw = flat_ws.cell(r, 7).value
        if qty_raw is None:
            continue
        qty_raw = float(qty_raw)

        subgrupo, sistema = classify_item(slug, desc, dim)

        # Converter barras → metros para tubulações
        is_tube = sistema in ('AF', 'AQ', 'PPR', 'ESG_N', 'ESG_R')
        mult = tube_multiplier(desc) if is_tube else 1
        qty_m = qty_raw * mult
        unid = 'm' if is_tube else 'un'

        # Buscar PU
        pu_val, fonte = find_pu(sistema, desc, dim)
        if isinstance(pu_val, tuple):
            pu_val = pu_val[0] if pu_val[0] else None

        items.append({
            'slug': slug, 'desc': desc, 'dim': dim, 'marca': marca,
            'qty_raw': qty_raw, 'qty': qty_m, 'unid': unid,
            'subgrupo': subgrupo, 'sistema': sistema,
            'pu': pu_val, 'fonte': fonte, 'mult': mult,
        })

    # 2. Ordenar por subgrupo (ordem definida) e descrição
    SUBGRUPO_ORDER = [
        'TUBULAÇÕES ÁGUA FRIA', 'TUBULAÇÕES ÁGUA QUENTE',
        'TUBULAÇÕES ESGOTO NORMAL', 'TUBULAÇÕES ESGOTO REFORÇADA',
        'CONEXÕES ÁGUA FRIA', 'CONEXÕES ÁGUA QUENTE', 'CONEXÕES PPR',
        'CONEXÕES ESGOTO NORMAL', 'CONEXÕES ESGOTO REFORÇADA',
        'ACESSÓRIOS', 'EQUIPAMENTOS', 'ETE / TRATAMENTO',
    ]
    order_map = {s: i for i, s in enumerate(SUBGRUPO_ORDER)}
    items.sort(key=lambda x: (order_map.get(x['subgrupo'], 99), x['desc']))

    # 3. Criar nova aba
    ws = wb.create_sheet('HIDROSSANITÁRIO v2', 1)

    # Estilos
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    sec_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sec_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sub_font = Font(name="Calibri", size=10, bold=True)
    data_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1851D6", end_color="1851D6", fill_type="solid")
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    qty_fmt = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
    rs_fmt = '_-"R$"\\ * #,##0.00_-;\\-"R$"\\ * #,##0.00_-;_-"R$"\\ * "-"??_-;_-@_-'

    # Metadata
    ws['B1'] = 'Projeto'; ws['C1'] = 'Electra Towers'
    ws['B2'] = 'Empresa'; ws['C2'] = 'Thozen'
    ws['B3'] = 'Origem'; ws['C3'] = 'CTN-TZN_HID-LM348-R01'
    ws['B4'] = 'Revisão'; ws['C4'] = 'R02 — Lista geral precificada (LM348 PDFs Eletrowatts) — 16/abr/2026'
    ws['A5'] = 'Hidrossanitário v2 — Lista Geral Precificada'
    ws['A5'].font = title_font
    ws['A5'].fill = title_fill
    for c in range(2, 11):
        ws.cell(5, c).fill = title_fill

    # Headers (row 6)
    headers = ['GRUPO/EAP', 'SUBGRUPO', 'Descrição', 'Dimensões', 'Marca',
               'Qtd', 'Unid', 'Custo unit.', 'Custo Total', 'Fonte PU']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(6, ci, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 7
    current_subgrupo = None
    subtotal_start = row
    subtotal_rows = []
    n_precificado = 0
    n_total = 0

    for item in items:
        # Section header quando muda subgrupo
        if item['subgrupo'] != current_subgrupo:
            if current_subgrupo is not None:
                # Subtotal do subgrupo anterior
                ws.cell(row, 2, f'SUBTOTAL {current_subgrupo}')
                ws.cell(row, 2).font = sub_font
                ws.cell(row, 9).number_format = rs_fmt
                ws.cell(row, 9, f'=SUM(I{subtotal_start}:I{row-1})')
                for ci in range(1, 11):
                    ws.cell(row, ci).fill = sub_fill
                    ws.cell(row, ci).font = sub_font
                subtotal_rows.append(row)
                row += 1

            current_subgrupo = item['subgrupo']
            subtotal_start = row + 1

            # Grupo/EAP
            grupo = 'ETE / TRATAMENTO EFLUENTES' if item['sistema'] == 'ETE' else 'INST. HIDROSSANITÁRIAS'
            ws.cell(row, 1, grupo)
            ws.cell(row, 2, current_subgrupo)
            for ci in range(1, 11):
                ws.cell(row, ci).fill = sec_fill
                ws.cell(row, ci).font = sec_font
            row += 1

        # Data row
        grupo = 'ETE / TRATAMENTO EFLUENTES' if item['sistema'] == 'ETE' else 'INST. HIDROSSANITÁRIAS'
        ws.cell(row, 1, grupo)
        ws.cell(row, 2, item['subgrupo'])
        ws.cell(row, 3, item['desc'])
        ws.cell(row, 4, item['dim'] if item['dim'] else '')
        ws.cell(row, 5, item['marca'] if item['marca'] else '')
        ws.cell(row, 6, round(item['qty'], 2))
        ws.cell(row, 6).number_format = qty_fmt
        ws.cell(row, 7, item['unid'])

        if item['pu'] is not None:
            ws.cell(row, 8, round(item['pu'], 2))
            ws.cell(row, 8).number_format = rs_fmt
            ws.cell(row, 9, f'=F{row}*H{row}')  # formula
            ws.cell(row, 9).number_format = rs_fmt
            n_precificado += 1
        else:
            ws.cell(row, 9, 0)
            ws.cell(row, 9).number_format = rs_fmt

        ws.cell(row, 10, item['fonte'])

        for ci in range(1, 11):
            ws.cell(row, ci).font = data_font
            ws.cell(row, ci).border = thin

        n_total += 1
        row += 1

    # Último subtotal
    if current_subgrupo:
        ws.cell(row, 2, f'SUBTOTAL {current_subgrupo}')
        ws.cell(row, 2).font = sub_font
        ws.cell(row, 9, f'=SUM(I{subtotal_start}:I{row-1})')
        ws.cell(row, 9).number_format = rs_fmt
        for ci in range(1, 11):
            ws.cell(row, ci).fill = sub_fill
            ws.cell(row, ci).font = sub_font
        subtotal_rows.append(row)
        row += 2

    # Total geral — soma apenas data rows (pular subtotais usando SUMPRODUCT)
    # Collect subtotal rows to build a formula that sums only subtotals
    ws.cell(row, 2, 'TOTAL GERAL HIDROSSANITÁRIO (LM348)')
    ws.cell(row, 2).font = Font(name="Calibri", size=12, bold=True)
    # Sum all subtotals
    subtotal_refs = '+'.join(f'I{sr}' for sr in subtotal_rows)
    ws.cell(row, 9, f'={subtotal_refs}')
    ws.cell(row, 9).number_format = rs_fmt
    ws.cell(row, 9).font = Font(name="Calibri", size=12, bold=True)

    # Linha de info
    row += 1
    ws.cell(row, 2, f'Itens: {n_total} | Precificados: {n_precificado} | A precificar: {n_total - n_precificado}')
    ws.cell(row, 2).font = Font(name="Calibri", size=9, italic=True)
    row += 1
    ws.cell(row, 2, f'AC Total: {AC:,} m² | Fontes: Belle Ville, Santa Maria, Estimativa, Mercado')
    ws.cell(row, 2).font = Font(name="Calibri", size=9, italic=True)

    # Column widths
    widths = [28, 32, 60, 18, 18, 14, 8, 14, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = 'A7'

    # Salvar
    wb.save(OUTPUT)
    print(f'Arquivo salvo: {OUTPUT}')
    print(f'Itens total: {n_total}')
    print(f'Precificados: {n_precificado}')
    print(f'A precificar: {n_total - n_precificado}')

if __name__ == "__main__":
    main()
