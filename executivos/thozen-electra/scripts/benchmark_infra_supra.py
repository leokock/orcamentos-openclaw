"""
Benchmark R$/m² AC de INFRAESTRUTURA, MOV. TERRA e SUPRAESTRUTURA
na base Cartesian (_Entregas/Orçamento_executivo/*.xlsx).

Objetivo: validar os R$/m² AC da fundação Electra (projetado R$ 226/m² AC infra total)
contra mediana de ~160 projetos entregues.

Output:
- dados/benchmark_infra_supra.json (bruto)
- print estatísticas agregadas

Uso:
    python scripts/benchmark_infra_supra.py
"""
import json
import re
from pathlib import Path
import openpyxl
import statistics

BASE = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
            r"\_Entregas\Orçamento_executivo")
OUT = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
           r"\_Executivo_IA\thozen-electra\dados\benchmark_infra_supra.json")

# Keywords pra identificar macrogrupos na coluna A (flexibilidade pra variação de escrita)
KEYWORDS = {
    "mov_terra": ["MOVIMENTAÇÃO DE TERRA", "MOV. TERRA", "MOVIMENTACAO DE TERRA"],
    "infra": ["INFRAESTRUTURA", "FUNDAÇÃO", "FUNDACOES"],
    "supra": ["SUPRAESTRUTURA", "SUPRA ESTRUTURA"],
    "paredes": ["PAREDES E PAINÉIS", "ALVENARIA"],
}

def find_ac(ws):
    """Busca área construída nas primeiras 10 linhas."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 15: break
        for j, cell in enumerate(row):
            if cell and isinstance(cell, str) and "ÁREA CONSTRUÍDA" in cell.upper():
                # próxima célula ou uma à direita
                for k in range(j+1, min(j+4, len(row))):
                    if isinstance(row[k], (int, float)) and row[k] > 100:
                        return float(row[k])
    return None


def find_macro(ws, keywords):
    """Busca linha com keywords na col A e retorna valor coluna B + R$/m² coluna D."""
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 100: break
        col_a = row[0] if row else None
        if col_a and isinstance(col_a, str):
            upper = col_a.upper().strip()
            for kw in keywords:
                if kw in upper or upper.startswith(kw[:15]):
                    # valor em B, R$/m² em D
                    val = row[1] if len(row) > 1 else None
                    r_m2 = row[3] if len(row) > 3 else None
                    return {
                        "label": col_a.strip(),
                        "valor": float(val) if isinstance(val, (int, float)) else None,
                        "r_m2": float(r_m2) if isinstance(r_m2, (int, float)) else None,
                    }
    return None


def process_xlsx(path):
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {"path": str(path), "error": str(e)}

    # Buscar sheet relevante
    target_sheet = None
    for sn in wb.sheetnames:
        if "ORÇAMENTO_EXECUTIVO" in sn.upper() or "ORCAMENTO_EXECUTIVO" in sn.upper() \
           or "EXECUTIVO" in sn.upper():
            target_sheet = sn
            break
    if not target_sheet:
        target_sheet = wb.sheetnames[0]  # fallback primeiro sheet

    ws = wb[target_sheet]
    ac = find_ac(ws)
    result = {
        "path": str(path.relative_to(BASE)),
        "sheet": target_sheet,
        "ac": ac,
    }
    for key, kws in KEYWORDS.items():
        result[key] = find_macro(ws, kws)

    wb.close()
    return result


def main():
    files = sorted(BASE.rglob("*.xlsx"))
    files = [f for f in files if not f.name.startswith("~$")]
    print(f"Processando {len(files)} xlsx em {BASE.name}...")

    results = []
    errors = 0
    for i, f in enumerate(files, 1):
        if i % 20 == 0:
            print(f"  {i}/{len(files)}...")
        r = process_xlsx(f)
        if "error" in r:
            errors += 1
        results.append(r)

    # Agregar stats
    def extract_r_m2(key):
        """Extrai R$/m² de todos os projetos pra essa categoria."""
        vals = []
        for r in results:
            v = r.get(key)
            if v and isinstance(v, dict) and v.get("r_m2") and v["r_m2"] > 0:
                vals.append(v["r_m2"])
        return vals

    # Calcular infra ajustada (INFRA típico inclui fundação — se não tem INFRA explícita, somar MOV + outros)
    # Pra comparar com Electra que tem Mov (R$14) + Fund.Prof (R$85) + Fund.Rasa (R$127) = R$226
    stats = {}
    for key in KEYWORDS.keys():
        vals = extract_r_m2(key)
        if len(vals) >= 3:
            vals_sorted = sorted(vals)
            stats[key] = {
                "n": len(vals),
                "min": min(vals),
                "p25": statistics.quantiles(vals, n=4)[0] if len(vals) >= 4 else None,
                "mediana": statistics.median(vals),
                "p75": statistics.quantiles(vals, n=4)[2] if len(vals) >= 4 else None,
                "max": max(vals),
                "media": statistics.mean(vals),
                "stdev": statistics.stdev(vals) if len(vals) > 1 else 0,
            }

    # Calcular soma MOV + INFRA (comparável ao total Electra "infraestrutura")
    mov_infra = []
    for r in results:
        m = r.get("mov_terra")
        i = r.get("infra")
        if m and i and isinstance(m, dict) and isinstance(i, dict):
            if m.get("r_m2") and i.get("r_m2"):
                mov_infra.append(m["r_m2"] + i["r_m2"])
    if len(mov_infra) >= 3:
        mov_infra_sorted = sorted(mov_infra)
        stats["mov_mais_infra"] = {
            "n": len(mov_infra),
            "min": min(mov_infra),
            "p25": statistics.quantiles(mov_infra, n=4)[0] if len(mov_infra) >= 4 else None,
            "mediana": statistics.median(mov_infra),
            "p75": statistics.quantiles(mov_infra, n=4)[2] if len(mov_infra) >= 4 else None,
            "max": max(mov_infra),
            "media": statistics.mean(mov_infra),
        }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({
        "base": str(BASE),
        "n_files": len(files),
        "n_errors": errors,
        "stats": stats,
        "results": results,
    }, indent=2, ensure_ascii=False, default=str), encoding="utf-8")

    # Print resumo
    print()
    print("="*70)
    print("BENCHMARK R$/m² AC — base Cartesian")
    print("="*70)
    for key, s in stats.items():
        if s:
            print(f"\n{key.upper()}")
            print(f"  n={s['n']} projetos")
            print(f"  min={s['min']:>8.1f}  P25={s.get('p25', 0) or 0:>8.1f}  "
                  f"MEDIANA={s['mediana']:>8.1f}  P75={s.get('p75', 0) or 0:>8.1f}  max={s['max']:>8.1f}")
            print(f"  média={s.get('media', 0):>8.1f}  desvio={s.get('stdev', 0):>6.1f}")

    # Comparação Electra
    print()
    print("="*70)
    print("COMPARAÇÃO ELECTRA")
    print("="*70)
    electra = {
        "mov_terra": 14.52,   # sessão 10-11/abr
        "infra_total": 85 + 127,  # fund profunda + fund rasa (da seção 29)
        "infra_apenas_fund": 85,  # só fundação profunda
    }
    if "mov_terra" in stats:
        s = stats["mov_terra"]
        pos = "<P25" if electra["mov_terra"] < (s.get("p25") or 999) else \
              "<mediana" if electra["mov_terra"] < s["mediana"] else \
              "<P75" if electra["mov_terra"] < (s.get("p75") or 9999) else ">P75"
        print(f"  Mov. Terra Electra: R$ {electra['mov_terra']}/m² AC  [posicao: {pos} | mediana base R$ {s['mediana']:.0f}/m²]")
    if "infra" in stats:
        s = stats["infra"]
        pos = "<P25" if electra["infra_total"] < (s.get("p25") or 999) else \
              "<mediana" if electra["infra_total"] < s["mediana"] else \
              "<P75" if electra["infra_total"] < (s.get("p75") or 9999) else ">P75"
        print(f"  Infra Electra (Fund.Prof + Rasa) : R$ {electra['infra_total']}/m² AC  [posicao: {pos} | mediana base R$ {s['mediana']:.0f}/m²]")
    if "mov_mais_infra" in stats:
        s = stats["mov_mais_infra"]
        electra_total = electra["mov_terra"] + electra["infra_total"]
        pos = "<P25" if electra_total < (s.get("p25") or 999) else \
              "<mediana" if electra_total < s["mediana"] else \
              "<P75" if electra_total < (s.get("p75") or 9999) else ">P75"
        print(f"  TOTAL Mov+Infra Electra: R$ {electra_total}/m² AC  [posicao: {pos} | mediana base R$ {s['mediana']:.0f}/m²]")

    print(f"\nArquivo salvo: {OUT}")
    print(f"Erros: {errors}/{len(files)}")


if __name__ == "__main__":
    main()
