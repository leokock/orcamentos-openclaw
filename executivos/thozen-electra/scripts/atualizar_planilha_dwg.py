#!/usr/bin/env python3
"""
Script para atualizar a planilha de telecomunicações com dados extraídos dos DWGs
Adiciona novos itens por pavimento preservando formatação e fórmulas
"""

import openpyxl
from openpyxl.styles import Font, Fill, Border, Alignment
from copy import copy
import json
from collections import defaultdict

# Mapeamento DXF → Pavimento
PAVIMENTO_MAP = {
    ('T02', 'T03'): 'TÉRREO',
    ('T04', 'T05'): 'GARAGEM 01',
    ('T06', 'T07'): 'GARAGEM 02',
    ('T08', 'T09'): 'GARAGEM 03',
    ('T10', 'T11'): 'GARAGEM 04',
    ('T12', 'T13'): 'GARAGEM 05',
    ('T14', 'T15'): 'LAZER',
    ('T16', 'T17'): 'PAVIMENTO TIPO',
    ('T18', 'T19'): 'CASA DE MÁQUINAS'
}

def consolidar_dados_por_pavimento(json_path):
    """Consolida dados dos DWGs por pavimento"""
    with open(json_path, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    
    pavimentos = defaultdict(lambda: {
        'eletrodutos': defaultdict(int),
        'caixas_passagem': defaultdict(int),
        'componentes': defaultdict(int)
    })
    
    # Mapear cada chave T0X para o pavimento correspondente
    for (t_a, t_b), pav_nome in PAVIMENTO_MAP.items():
        # Somar dados da Torre A
        if t_a in dados:
            # Eletrodutos
            for spec, qty in dados[t_a].get('eletrodutos_resumo', {}).items():
                # Extrair diâmetro do spec (ex: "1xø1\"" → "ø1\"")
                parts = spec.split('ø')
                if len(parts) == 2:
                    diametro = 'ø' + parts[1]
                    qty_numeric = int(parts[0].split('x')[0])  # Pegar o multiplicador
                    pavimentos[pav_nome]['eletrodutos'][diametro] += qty * qty_numeric
            
            # Caixas de passagem
            for cx in dados[t_a].get('caixas_passagem', []):
                dim = cx.get('dimensao', '')
                if dim:
                    pavimentos[pav_nome]['caixas_passagem'][dim] += 1
            
            # Componentes
            for comp, qty in dados[t_a].get('componentes', {}).items():
                # Mapear componentes para descrições simplificadas
                if 'Cotovelo' in comp:
                    pavimentos[pav_nome]['componentes']['Cotovelo Eletroduto'] += qty
                elif 'Conector Box' in comp or 'Arruela' in comp:
                    pavimentos[pav_nome]['componentes']['Conector Box/Arruela'] += qty
                elif 'Bucha Terminal' in comp:
                    pavimentos[pav_nome]['componentes']['Bucha Terminal'] += qty
                elif 'CFTV' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor CFTV'] += qty
                elif 'UTP' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor UTP'] += qty
                elif 'CCI' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor CCI'] += qty
                elif 'Cordplast' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor Cordplast'] += qty
                elif 'Caixa 4x2 - Interfone' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x2 - Interfone'] += qty
                elif 'Caixa 4x2 - Câmera' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x2 - Câmera'] += qty
                elif 'Caixa 4x2 - Controle de Acesso' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x2 - Controle de Acesso'] += qty
                elif 'Caixa 4x4 - Dados/Telefone' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x4 - Dados/Telefone'] += qty
        
        # Somar dados da Torre B
        if t_b in dados:
            # Mesma lógica para Torre B
            for spec, qty in dados[t_b].get('eletrodutos_resumo', {}).items():
                parts = spec.split('ø')
                if len(parts) == 2:
                    diametro = 'ø' + parts[1]
                    qty_numeric = int(parts[0].split('x')[0])
                    pavimentos[pav_nome]['eletrodutos'][diametro] += qty * qty_numeric
            
            for cx in dados[t_b].get('caixas_passagem', []):
                dim = cx.get('dimensao', '')
                if dim:
                    pavimentos[pav_nome]['caixas_passagem'][dim] += 1
            
            for comp, qty in dados[t_b].get('componentes', {}).items():
                if 'Cotovelo' in comp:
                    pavimentos[pav_nome]['componentes']['Cotovelo Eletroduto'] += qty
                elif 'Conector Box' in comp or 'Arruela' in comp:
                    pavimentos[pav_nome]['componentes']['Conector Box/Arruela'] += qty
                elif 'Bucha Terminal' in comp:
                    pavimentos[pav_nome]['componentes']['Bucha Terminal'] += qty
                elif 'CFTV' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor CFTV'] += qty
                elif 'UTP' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor UTP'] += qty
                elif 'CCI' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor CCI'] += qty
                elif 'Cordplast' in comp:
                    pavimentos[pav_nome]['componentes']['Condutor Cordplast'] += qty
                elif 'Caixa 4x2 - Interfone' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x2 - Interfone'] += qty
                elif 'Caixa 4x2 - Câmera' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x2 - Câmera'] += qty
                elif 'Caixa 4x2 - Controle de Acesso' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x2 - Controle de Acesso'] += qty
                elif 'Caixa 4x4 - Dados/Telefone' in comp:
                    pavimentos[pav_nome]['componentes']['Caixa 4x4 - Dados/Telefone'] += qty
    
    return pavimentos

def encontrar_linha_pavimento(ws, pavimento):
    """Encontra a linha do header do pavimento"""
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 2).value == pavimento and ws.cell(row, 1).value == 'TORRE':
            return row
    return None

def encontrar_proximo_pavimento(ws, linha_inicial):
    """Encontra a linha do próximo pavimento (onde inserir novos dados)"""
    for row in range(linha_inicial + 1, ws.max_row + 1):
        if ws.cell(row, 1).value == 'TORRE' and ws.cell(row, 2).value != ws.cell(linha_inicial, 2).value:
            return row
    return ws.max_row + 1

def copiar_estilo(cell_src, cell_dst):
    """Copia estilo de uma célula para outra"""
    if cell_src.font:
        cell_dst.font = copy(cell_src.font)
    if cell_src.fill:
        cell_dst.fill = copy(cell_src.fill)
    if cell_src.border:
        cell_dst.border = copy(cell_src.border)
    if cell_src.alignment:
        cell_dst.alignment = copy(cell_src.alignment)
    if cell_src.number_format:
        cell_dst.number_format = cell_src.number_format

def inserir_item_planilha(ws, linha_insercao, pav_nome, grupo, subgrupo, descricao, qtd, unidade, linha_template):
    """Insere um novo item na planilha copiando estilo da linha template"""
    ws.insert_rows(linha_insercao, 1)
    
    # Copiar estilos da linha template
    for col in range(1, 13):  # A até L
        cell_src = ws.cell(linha_template, col)
        cell_dst = ws.cell(linha_insercao, col)
        copiar_estilo(cell_src, cell_dst)
    
    # Preencher dados
    ws.cell(linha_insercao, 1).value = f'=A{linha_insercao - 1}'  # Coluna A (Torre)
    ws.cell(linha_insercao, 2).value = f'=B{linha_insercao - 1}'  # Coluna B (Pavimento)
    ws.cell(linha_insercao, 3).value = grupo
    ws.cell(linha_insercao, 4).value = subgrupo
    ws.cell(linha_insercao, 5).value = descricao
    ws.cell(linha_insercao, 6).value = qtd
    ws.cell(linha_insercao, 7).value = 1  # Repetição
    ws.cell(linha_insercao, 8).value = 0  # Perda
    ws.cell(linha_insercao, 9).value = f'=(F{linha_insercao}+H{linha_insercao}*F{linha_insercao})*G{linha_insercao}'  # Qte com perda
    ws.cell(linha_insercao, 10).value = unidade
    ws.cell(linha_insercao, 11).value = 0  # Custo unitário (a preencher depois)
    ws.cell(linha_insercao, 12).value = f'=IF(SUM(F{linha_insercao}:K{linha_insercao})=0,SUMIFS($L{linha_insercao + 1}:$L$535,$B{linha_insercao + 1}:$B$535,$B{linha_insercao},$A{linha_insercao + 1}:$A$535,$A{linha_insercao}),I{linha_insercao}*K{linha_insercao})'
    
    return linha_insercao + 1

def processar_planilha(planilha_path, json_path, output_path):
    """Processa a planilha adicionando novos itens dos DWGs"""
    print("📂 Carregando planilha...")
    wb = openpyxl.load_workbook(planilha_path, data_only=False)
    ws = wb['TELECOMUNICAÇÃO']
    
    print("📊 Consolidando dados dos DWGs...")
    pavimentos = consolidar_dados_por_pavimento(json_path)
    
    print(f"\n✅ Dados consolidados para {len(pavimentos)} pavimentos")
    
    # Processar cada pavimento
    for pav_nome in ['TÉRREO', 'GARAGEM 01', 'GARAGEM 02', 'GARAGEM 03', 'GARAGEM 04', 
                      'GARAGEM 05', 'LAZER', 'PAVIMENTO TIPO', 'CASA DE MÁQUINAS']:
        if pav_nome not in pavimentos:
            continue
        
        print(f"\n🏗️  Processando {pav_nome}...")
        
        linha_pav = encontrar_linha_pavimento(ws, pav_nome)
        if not linha_pav:
            print(f"  ⚠️  Pavimento {pav_nome} não encontrado na planilha")
            continue
        
        linha_template = linha_pav + 1  # Usar primeira linha de dados como template
        linha_insercao = encontrar_proximo_pavimento(ws, linha_pav)
        
        pav_data = pavimentos[pav_nome]
        items_added = 0
        
        # 1. ELETRODUTOS (novos diâmetros)
        for diametro in ['ø1¼"', 'ø1 ¼"', 'ø3"']:  # Normalizar variações
            qty = pav_data['eletrodutos'].get(diametro, 0) + pav_data['eletrodutos'].get(diametro.replace(' ', ''), 0)
            if qty > 0:
                # Normalizar descrição
                diam_clean = diametro.replace(' ', '').replace('ø', '')
                if '1¼' in diam_clean or '1 ¼' in diam_clean:
                    desc = 'PVC Flexível Antichama - NBR 15465 ø1.1/4"'
                    diam_norm = 'ø1.1/4"'
                elif '3"' in diam_clean:
                    desc = 'PVC Flexível Antichama - NBR 15465 ø3"'
                    diam_norm = 'ø3"'
                else:
                    continue
                
                linha_insercao = inserir_item_planilha(
                    ws, linha_insercao, pav_nome,
                    'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'ELETRODUTOS', desc, qty, 'm', linha_template
                )
                items_added += 1
                print(f"  ✓ Adicionado: {desc} - {qty}m")
        
        # 2. CAIXAS DE PASSAGEM (novas dimensões)
        for dim, qty in pav_data['caixas_passagem'].items():
            if dim == '30x30x12cm':
                continue  # Já existe
            desc = f'Cx de passagem de embutir com dimensão de {dim}'
            linha_insercao = inserir_item_planilha(
                ws, linha_insercao, pav_nome,
                'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'CAIXA DE PASSAGEM', desc, qty, 'pç', linha_template
            )
            items_added += 1
            print(f"  ✓ Adicionado: {desc} - {qty} pç")
        
        # 3. COTOVELOS
        if pav_data['componentes'].get('Cotovelo Eletroduto', 0) > 0:
            qty = pav_data['componentes']['Cotovelo Eletroduto']
            linha_insercao = inserir_item_planilha(
                ws, linha_insercao, pav_nome,
                'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'COTOVELOS', 
                'Cotovelo para eletroduto flexível PVC amarelo', qty, 'un', linha_template
            )
            items_added += 1
            print(f"  ✓ Adicionado: Cotovelo - {qty} un")
        
        # 4. CONECTORES E BUCHAS
        if pav_data['componentes'].get('Conector Box/Arruela', 0) > 0:
            qty = pav_data['componentes']['Conector Box/Arruela']
            linha_insercao = inserir_item_planilha(
                ws, linha_insercao, pav_nome,
                'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'CONECTORES E BUCHAS',
                'Conector box e arruela - Alumínio padrão', qty, 'un', linha_template
            )
            items_added += 1
            print(f"  ✓ Adicionado: Conector Box/Arruela - {qty} un")
        
        if pav_data['componentes'].get('Bucha Terminal', 0) > 0:
            qty = pav_data['componentes']['Bucha Terminal']
            linha_insercao = inserir_item_planilha(
                ws, linha_insercao, pav_nome,
                'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'CONECTORES E BUCHAS',
                'Bucha terminal simples - Alumínio padrão', qty, 'un', linha_template
            )
            items_added += 1
            print(f"  ✓ Adicionado: Bucha Terminal - {qty} un")
        
        # 5. CONDUTORES TELECOM
        for condutor in ['CFTV', 'UTP', 'CCI', 'Cordplast']:
            if pav_data['componentes'].get(f'Condutor {condutor}', 0) > 0:
                qty = pav_data['componentes'][f'Condutor {condutor}']
                linha_insercao = inserir_item_planilha(
                    ws, linha_insercao, pav_nome,
                    'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'CONDUTORES TELECOM',
                    f'Condutor {condutor}', qty, 'un', linha_template
                )
                items_added += 1
                print(f"  ✓ Adicionado: Condutor {condutor} - {qty} un")
        
        # 6. PONTOS DE USO
        pontos_uso = {
            'Caixa 4x2 - Interfone': 'Caixa 4x2 - Interfone',
            'Caixa 4x2 - Câmera': 'Caixa 4x2 - Câmera CFTV',
            'Caixa 4x2 - Controle de Acesso': 'Caixa 4x2 - Controle de Acesso',
            'Caixa 4x4 - Dados/Telefone': 'Caixa 4x4 - Dados/Telefone (RJ45/RJ11)'
        }
        
        for comp_key, desc in pontos_uso.items():
            if pav_data['componentes'].get(comp_key, 0) > 0:
                qty = pav_data['componentes'][comp_key]
                linha_insercao = inserir_item_planilha(
                    ws, linha_insercao, pav_nome,
                    'INSTALAÇÕES DE TELECOMUNICAÇÃO', 'PONTOS DE USO',
                    desc, qty, 'un', linha_template
                )
                items_added += 1
                print(f"  ✓ Adicionado: {desc} - {qty} un")
        
        if items_added > 0:
            print(f"  ✅ {items_added} itens adicionados ao {pav_nome}")
    
    # Atualizar revisão
    ws['D4'] = 'R01 - Jarvis/IFC+DWG 24/03/2026'
    
    print(f"\n💾 Salvando planilha: {output_path}")
    wb.save(output_path)
    print("✅ Planilha atualizada com sucesso!")

if __name__ == '__main__':
    planilha_r00 = '/Users/leokock/orcamentos/executivos/thozen-electra/entregas/telecomunicacoes-electra-r00.xlsx'
    json_dados = '/Users/leokock/orcamentos/executivos/thozen-electra/quantitativos/telefonico/dados_brutos_telefonico.json'
    planilha_r01 = '/Users/leokock/orcamentos/executivos/thozen-electra/entregas/telecomunicacoes-electra-r01.xlsx'
    
    processar_planilha(planilha_r00, json_dados, planilha_r01)
