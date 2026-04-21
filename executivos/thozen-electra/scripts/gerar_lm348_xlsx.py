"""Gerador xlsx comum para o lote 348_LM Electra (Fase 4 do pipeline).

Consome os JSONs produzidos por gemma_extract_lm + gemma_rateio_pavimento e
produz um xlsx por disciplina com a organizacao canonica Por Pavimento exigida
pelo feedback_orcamento_por_pavimento. Cada xlsx e autocontido (o lote 348_LM
como incremento) e tem 4 abas:

  1. "Por Pavimento"  -> itens ordenados por TERREO->...->COBERTURA com subtotais
  2. "Flat por PDF"   -> itens na ordem do PDF, com codigo/qtd/descricao/fonte
  3. "Rateio"         -> so se houver distribuicoes (format A ratreadas)
  4. "Resumo"         -> totais, ~ R$/m2 AC, comparativo vs mediana Cartesian

Uso:
    python scripts/gerar_lm348_xlsx.py --disciplina eletrico
    python scripts/gerar_lm348_xlsx.py --all
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPTS_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPTS_DIR))
import normalizar_pavimento as norm

EXEC_DIR = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra")
JSON_DIR = EXEC_DIR / "quantitativos" / "listas-materiais-348"
DISC_DIR = EXEC_DIR / "disciplinas"

AC_TOTAL_M2 = 36092  # Torre A + B

# Meta por disciplina: (subpasta disciplinas/, titulo, codigo orcamentario, mediana benchmark)
DISC_META = {
    "eletrico": {
        "disc_dir": "eletrico",
        "titulo": "09 ELETRICO - Instalacoes Eletricas (complemento LM 348)",
        "bench_mediana_rs_m2": 178.56,  # sistema eletrico R02 Electra
        "out_filename": "eletrico-electra-lm348-r01.xlsx",
    },
    "hidraulico": {
        "disc_dir": "hidraulico",
        "titulo": "05 HIDRAULICO - Instalacoes Hidrossanitarias (complemento LM 348)",
        "bench_mediana_rs_m2": 48.32,
        "out_filename": "hidraulico-electra-lm348-r01.xlsx",
    },
    "telefonico": {
        "disc_dir": "telefonico",
        "titulo": "10 TELEFONICO - Cabeamento Voz/Dados (complemento LM 348)",
        "bench_mediana_rs_m2": 12.83,
        "out_filename": "telecomunicacoes-electra-lm348-r01.xlsx",
    },
    "ppci-civil": {
        "disc_dir": "pci-civil",
        "titulo": "07 PREVENTIVO INCENDIO CIVIL (complemento LM 348)",
        "bench_mediana_rs_m2": 22.50,
        "out_filename": "ppci-civil-electra-lm348-r01.xlsx",
    },
    "ppci-eletrico": {
        "disc_dir": "pci-eletrico",
        "titulo": "08 PREVENTIVO INCENDIO ELETRICO (complemento LM 348)",
        "bench_mediana_rs_m2": 14.80,
        "out_filename": "ppci-eletrico-electra-lm348-r01.xlsx",
    },
    "spda": {
        "disc_dir": "spda",
        "titulo": "11 SPDA - Sistema Protecao Descargas Atmosfericas (complemento LM 348)",
        "bench_mediana_rs_m2": 5.14,
        "out_filename": "spda-electra-lm348-r01.xlsx",
    },
}

# Ordem canonica de pavimentos pra ordenacao e subtotais
PAVIMENTOS_ORDEM = norm.PAVIMENTOS_ORDER  # ["TERREO", "G1", ..., "COBERTURA", "NA"]
PAVIMENTO_LABEL = norm.PAVIMENTO_LABEL


# -------------------- Loaders --------------------

def load_all_for_disciplina(disciplina: str) -> dict:
    """Le todos os {slug}.gemma.json da disciplina, aplica rateios quando
    existirem ({slug}.rateio.json) e retorna um dict com todos os itens
    normalizados.

    Estrutura de retorno:
    {
      "disciplina": "eletrico",
      "pdfs": [{"slug": "ele-acabamentos", "formato": "B", "n_itens": 47, ...}],
      "itens_por_pavimento": [{"pavimento": "TIPO", "torre": "A", "itens": [...]}],
      "itens_flat": [...],
      "rateios_detalhados": [...]
    }
    """
    d_path = JSON_DIR / disciplina
    if not d_path.exists():
        return {"disciplina": disciplina, "pdfs": [], "itens_flat": [],
                "itens_por_pavimento": [], "rateios_detalhados": [],
                "_error": f"pasta nao existe: {d_path}"}

    pdfs_info = []
    itens_flat = []  # cada item normalizado pronto pra aba Flat

    # Estrutura por pavimento: dict[(pav, torre)] = list de itens
    buckets: dict[tuple[str, str], list[dict]] = {}

    rateios_detalhados = []  # para aba Rateio (so format A)

    for gemma_file in sorted(d_path.glob("*.gemma.json")):
        slug = gemma_file.stem.replace(".gemma", "")
        try:
            data = json.loads(gemma_file.read_text(encoding="utf-8"))
        except Exception as e:
            pdfs_info.append({"slug": slug, "error": str(e)})
            continue

        fmt = data.get("formato")
        n_itens = data.get("total_itens", 0)
        pdf_info = {
            "slug": slug,
            "pdf": data.get("pdf_origem"),
            "formato": fmt,
            "n_itens": n_itens,
            "n_docs": len(data.get("documentos", [])),
        }

        # Tenta carregar rateio
        rateio_file = gemma_file.parent / gemma_file.name.replace(".gemma.json", ".rateio.json")
        rateio_data = None
        if rateio_file.exists():
            try:
                rateio_data = json.loads(rateio_file.read_text(encoding="utf-8"))
                pdf_info["has_rateio"] = True
            except Exception:
                pass

        # Processar documentos
        for doc_idx, doc in enumerate(data.get("documentos", [])):
            meta = doc.get("metadata") or {}
            pav_doc = meta.get("pavimento_canonico") or "NA"
            torre_doc = meta.get("torre_canonica") or "NA"

            for it in doc.get("itens", []):
                base_item = {
                    "pdf_slug": slug,
                    "pdf_origem": data.get("pdf_origem"),
                    "doc_codigo": meta.get("codigo"),
                    "doc_localizacao": meta.get("localizacao"),
                    "doc_referente": meta.get("referente"),
                    "numero": it.get("numero"),
                    "codigo": it.get("codigo"),
                    "qtd": it.get("qtd"),
                    "unidade": it.get("unidade"),
                    "descricao": it.get("descricao") or it.get("produto"),
                    "dimensoes": it.get("dimensoes"),
                    "marca": it.get("marca"),
                    "formato": fmt,
                }
                itens_flat.append(base_item)

                # Distribuicao por pavimento
                if fmt == "B":
                    # Format B: pavimento vem do metadata do documento
                    buckets.setdefault((pav_doc, torre_doc), []).append({
                        **base_item, "qtd_pav": it.get("qtd"),
                        "pavimento": pav_doc, "torre": torre_doc,
                        "origem_rateio": "formato B (pavimento no PDF)",
                    })
                else:
                    # Format A: aplicar rateio se existir, senao colocar em NA
                    if rateio_data:
                        r_item = _find_rateio(rateio_data, it.get("numero"))
                        if r_item and r_item.get("distribuicao"):
                            for dist in r_item["distribuicao"]:
                                pav = dist.get("pavimento", "NA")
                                torre = dist.get("torre", "NA")
                                buckets.setdefault((pav, torre), []).append({
                                    **base_item,
                                    "qtd_pav": dist.get("qtd_rateada"),
                                    "qtd_total_original": it.get("qtd"),
                                    "pct_rateio": dist.get("pct"),
                                    "pavimento": pav,
                                    "torre": torre,
                                    "origem_rateio": f"rateio Gemma: {dist.get('justificativa')}",
                                })
                                rateios_detalhados.append({
                                    "pdf_slug": slug,
                                    "numero": it.get("numero"),
                                    "codigo": it.get("codigo"),
                                    "descricao": base_item["descricao"],
                                    "qtd_total": it.get("qtd"),
                                    "pavimento": pav,
                                    "torre": torre,
                                    "qtd_rateada": dist.get("qtd_rateada"),
                                    "pct": dist.get("pct"),
                                    "justificativa": dist.get("justificativa"),
                                })
                            continue
                    # sem rateio: vai pra NA
                    buckets.setdefault(("NA", "NA"), []).append({
                        **base_item, "qtd_pav": it.get("qtd"),
                        "pavimento": "NA", "torre": "NA",
                        "origem_rateio": "format A sem rateio (NAO RATEADO)",
                    })
        pdfs_info.append(pdf_info)

    # Ordenar buckets pela ordem canonica
    def ord_key(k):
        pav, torre = k
        pav_idx = PAVIMENTOS_ORDEM.index(pav) if pav in PAVIMENTOS_ORDEM else 99
        torre_idx = {"A": 0, "B": 1, "AMBAS": 2, "NA": 3}.get(torre, 4)
        return (pav_idx, torre_idx)

    itens_por_pavimento = []
    for k in sorted(buckets.keys(), key=ord_key):
        pav, torre = k
        itens_por_pavimento.append({
            "pavimento": pav,
            "torre": torre,
            "pavimento_label": PAVIMENTO_LABEL.get(pav, pav),
            "itens": buckets[k],
        })

    return {
        "disciplina": disciplina,
        "pdfs": pdfs_info,
        "itens_flat": itens_flat,
        "itens_por_pavimento": itens_por_pavimento,
        "rateios_detalhados": rateios_detalhados,
    }


def _find_rateio(rateio_data: dict, numero) -> dict | None:
    for r in rateio_data.get("rateios", []):
        if r.get("numero") == numero or str(r.get("numero")) == str(numero):
            return r
    return None


# -------------------- XLSX writer --------------------

def _thin_border():
    thin = Side(border_style="thin", color="888888")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_xlsx(disciplina: str, data: dict, out_path: Path) -> None:
    meta = DISC_META[disciplina]
    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    subtot_fill = PatternFill("solid", fgColor="FFF2CC")
    total_fill = PatternFill("solid", fgColor="C6E0B4")
    na_fill = PatternFill("solid", fgColor="FFC7CE")
    border = _thin_border()
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # -------------------- Aba 1: Por Pavimento --------------------
    ws = wb.active
    ws.title = "Por Pavimento"

    ws["A1"] = "THOZEN - RES. ELECTRA TOWERS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = meta["titulo"]
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = f"Revisao: LM348-R01 | Data: {datetime.now().strftime('%Y-%m-%d')} | AC Total: {AC_TOTAL_M2} m2"
    ws["A4"] = f"Fonte: PDFs 348_LM da Eletrowatts, extracao Gemma local + rateio heuristico por sistema"
    ws["A4"].font = Font(italic=True, size=9, color="666666")
    ws["A5"] = f"PDFs processados: {len(data['pdfs'])} | Itens flat: {len(data['itens_flat'])}"
    ws["A5"].font = Font(italic=True, size=9, color="666666")

    headers = ["Pavimento", "Torre", "PDF origem", "Numero", "Codigo", "Qtd", "Unid", "Descricao", "Dimensoes", "Marca", "Origem rateio"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    row = 8
    for grupo in data["itens_por_pavimento"]:
        # Header do grupo pavimento
        ws.cell(row=row, column=1, value=f"{grupo['pavimento_label']} | Torre: {grupo['torre']}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        fill = na_fill if grupo["pavimento"] == "NA" else subtot_fill
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = border
        row += 1

        grupo_start = row
        for it in grupo["itens"]:
            ws.cell(row=row, column=1, value=it["pavimento"]).alignment = center
            ws.cell(row=row, column=2, value=it["torre"]).alignment = center
            ws.cell(row=row, column=3, value=it.get("pdf_slug")).alignment = left
            ws.cell(row=row, column=4, value=it.get("numero")).alignment = center
            ws.cell(row=row, column=5, value=it.get("codigo")).alignment = center
            ws.cell(row=row, column=6, value=it.get("qtd_pav")).alignment = right
            ws.cell(row=row, column=7, value=it.get("unidade")).alignment = center
            ws.cell(row=row, column=8, value=it.get("descricao")).alignment = left
            ws.cell(row=row, column=9, value=it.get("dimensoes")).alignment = left
            ws.cell(row=row, column=10, value=it.get("marca")).alignment = left
            ws.cell(row=row, column=11, value=it.get("origem_rateio")).alignment = left
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Subtotal por grupo (contagem de itens + linha destacada)
        n_itens = row - grupo_start
        ws.cell(row=row, column=1, value=f"Subtotal {grupo['pavimento']} {grupo['torre']}: {n_itens} itens")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = subtot_fill
            ws.cell(row=row, column=col).border = border
        row += 2  # pula linha

    # Larguras col
    widths_p1 = [10, 8, 18, 7, 12, 10, 6, 50, 22, 14, 42]
    for i, w in enumerate(widths_p1, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    # -------------------- Aba 2: Flat por PDF --------------------
    ws2 = wb.create_sheet("Flat por PDF")
    ws2["A1"] = f"Flat (ordem original do PDF) - {len(data['itens_flat'])} itens"
    ws2["A1"].font = Font(bold=True, size=12)
    headers2 = ["PDF slug", "PDF original", "Doc cod QT-MAT", "Doc localizacao", "Numero", "Codigo", "Qtd", "Unid", "Descricao", "Dimensoes", "Marca", "Formato"]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, it in enumerate(data["itens_flat"], start=4):
        ws2.cell(row=i, column=1, value=it.get("pdf_slug")).alignment = left
        ws2.cell(row=i, column=2, value=it.get("pdf_origem")).alignment = left
        ws2.cell(row=i, column=3, value=it.get("doc_codigo")).alignment = center
        ws2.cell(row=i, column=4, value=it.get("doc_localizacao")).alignment = left
        ws2.cell(row=i, column=5, value=it.get("numero")).alignment = center
        ws2.cell(row=i, column=6, value=it.get("codigo")).alignment = center
        ws2.cell(row=i, column=7, value=it.get("qtd")).alignment = right
        ws2.cell(row=i, column=8, value=it.get("unidade")).alignment = center
        ws2.cell(row=i, column=9, value=it.get("descricao")).alignment = left
        ws2.cell(row=i, column=10, value=it.get("dimensoes")).alignment = left
        ws2.cell(row=i, column=11, value=it.get("marca")).alignment = left
        ws2.cell(row=i, column=12, value=it.get("formato")).alignment = center
        for col in range(1, 13):
            ws2.cell(row=i, column=col).border = border

    widths_p2 = [18, 40, 15, 38, 7, 12, 10, 6, 50, 22, 14, 8]
    for i, w in enumerate(widths_p2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[3].height = 22
    ws2.freeze_panes = "A4"

    # -------------------- Aba 3: Rateio --------------------
    ws3 = wb.create_sheet("Rateio")
    ws3["A1"] = f"Rateio por pavimento (format A) - {len(data['rateios_detalhados'])} distribuicoes"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3["A2"] = "Gerado por Gemma usando heuristica tecnica por sistema (ver gemma_rateio_pavimento.py)"
    ws3["A2"].font = Font(italic=True, size=9, color="666666")
    headers3 = ["PDF slug", "Num", "Codigo", "Descricao", "Qtd total", "Pavimento", "Torre", "Qtd rateada", "%", "Justificativa"]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, r in enumerate(data["rateios_detalhados"], start=5):
        ws3.cell(row=i, column=1, value=r.get("pdf_slug")).alignment = left
        ws3.cell(row=i, column=2, value=r.get("numero")).alignment = center
        ws3.cell(row=i, column=3, value=r.get("codigo")).alignment = center
        ws3.cell(row=i, column=4, value=r.get("descricao")).alignment = left
        ws3.cell(row=i, column=5, value=r.get("qtd_total")).alignment = right
        ws3.cell(row=i, column=6, value=r.get("pavimento")).alignment = center
        ws3.cell(row=i, column=7, value=r.get("torre")).alignment = center
        ws3.cell(row=i, column=8, value=r.get("qtd_rateada")).alignment = right
        ws3.cell(row=i, column=9, value=r.get("pct")).number_format = "0.0%"
        ws3.cell(row=i, column=10, value=r.get("justificativa")).alignment = left
        for col in range(1, 11):
            ws3.cell(row=i, column=col).border = border

    widths_p3 = [18, 7, 12, 50, 12, 10, 8, 12, 8, 60]
    for i, w in enumerate(widths_p3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.row_dimensions[4].height = 22
    ws3.freeze_panes = "A5"

    # -------------------- Aba 4: Resumo --------------------
    ws4 = wb.create_sheet("Resumo")
    ws4["A1"] = f"RESUMO - {meta['titulo']}"
    ws4["A1"].font = Font(bold=True, size=14)
    ws4["A2"] = f"Lote LM348 - processado em {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws4["A2"].font = Font(italic=True, size=10, color="666666")

    r = 4
    ws4.cell(row=r, column=1, value="Metrica").font = header_font
    ws4.cell(row=r, column=1).fill = header_fill
    ws4.cell(row=r, column=1).alignment = center
    ws4.cell(row=r, column=2, value="Valor").font = header_font
    ws4.cell(row=r, column=2).fill = header_fill
    ws4.cell(row=r, column=2).alignment = center
    r += 1

    resumo_rows = [
        ("PDFs processados", len(data["pdfs"])),
        ("Total itens extraidos (flat)", len(data["itens_flat"])),
        ("Itens distribuidos por pavimento", sum(len(g["itens"]) for g in data["itens_por_pavimento"])),
        ("Rateios detalhados (format A)", len(data["rateios_detalhados"])),
        ("AC Total Torre A+B (m2)", AC_TOTAL_M2),
        ("Benchmark mediana disciplina (R$/m2)", meta["bench_mediana_rs_m2"]),
        ("Benchmark estimado (R$)", meta["bench_mediana_rs_m2"] * AC_TOTAL_M2),
    ]
    for lbl, val in resumo_rows:
        ws4.cell(row=r, column=1, value=lbl).alignment = left
        ws4.cell(row=r, column=2, value=val).alignment = right
        if isinstance(val, (int, float)) and "R$" in lbl:
            ws4.cell(row=r, column=2).number_format = "#,##0.00"
        r += 1

    # Contagem por pavimento
    r += 1
    ws4.cell(row=r, column=1, value="Distribuicao por pavimento").font = Font(bold=True, size=11)
    r += 1
    ws4.cell(row=r, column=1, value="Pavimento").font = header_font
    ws4.cell(row=r, column=1).fill = header_fill
    ws4.cell(row=r, column=2, value="Torre").font = header_font
    ws4.cell(row=r, column=2).fill = header_fill
    ws4.cell(row=r, column=3, value="Itens").font = header_font
    ws4.cell(row=r, column=3).fill = header_fill
    r += 1
    for grupo in data["itens_por_pavimento"]:
        ws4.cell(row=r, column=1, value=grupo["pavimento"]).alignment = center
        ws4.cell(row=r, column=2, value=grupo["torre"]).alignment = center
        ws4.cell(row=r, column=3, value=len(grupo["itens"])).alignment = right
        if grupo["pavimento"] == "NA":
            for col in range(1, 4):
                ws4.cell(row=r, column=col).fill = na_fill
        r += 1

    # PDFs processados
    r += 2
    ws4.cell(row=r, column=1, value="PDFs processados").font = Font(bold=True, size=11)
    r += 1
    for h_idx, h in enumerate(["Slug", "PDF origem", "Formato", "Docs", "Itens"], start=1):
        ws4.cell(row=r, column=h_idx, value=h).font = header_font
        ws4.cell(row=r, column=h_idx).fill = header_fill
    r += 1
    for pdf in data["pdfs"]:
        ws4.cell(row=r, column=1, value=pdf.get("slug")).alignment = left
        ws4.cell(row=r, column=2, value=pdf.get("pdf")).alignment = left
        ws4.cell(row=r, column=3, value=pdf.get("formato")).alignment = center
        ws4.cell(row=r, column=4, value=pdf.get("n_docs")).alignment = right
        ws4.cell(row=r, column=5, value=pdf.get("n_itens")).alignment = right
        r += 1

    widths_p4 = [28, 48, 10, 8, 10]
    for i, w in enumerate(widths_p4, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Save
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[xlsx] gravado: {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--disciplina", choices=list(DISC_META.keys()))
    ap.add_argument("--all", action="store_true")
    args = ap.parse_args()

    alvos = [args.disciplina] if args.disciplina else list(DISC_META.keys())

    for disc in alvos:
        meta = DISC_META[disc]
        print(f"=== {disc} ===")
        data = load_all_for_disciplina(disc)
        if "_error" in data:
            print(f"  skip: {data['_error']}")
            continue
        if not data["itens_flat"]:
            print(f"  sem itens ainda (Gemma pode estar rodando)")
            continue
        out = DISC_DIR / meta["disc_dir"] / meta["out_filename"]
        write_xlsx(disc, data, out)


if __name__ == "__main__":
    main()
