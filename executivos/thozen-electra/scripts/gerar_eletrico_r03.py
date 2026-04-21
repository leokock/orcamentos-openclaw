"""Gera eletrico-electra-r03.xlsx — suplemento ao R02 com itens de 5 PDFs,
organizado **por pavimento** conforme convencao Electra.

Abas:
  1. RESUMO R03            (pivot qtd/total por pavimento e subgrupo)
  2. QUADROS PRINCIPAIS    (16 — aprovativo, fallback TERREO)
  3. CAIXAS PASSAGEM       (162 — exec + prev, por pavimento)
  4. MATERIAL INTERNO CD   (39 — geral QT-MAT-1941, por documento/pavimento)
  5. DISJUNTORES 18 CDs    (167 — por documento/pavimento)

Ordenacao: por ordem_pavimento Electra, depois Torre (A antes de B), depois item.
Subtotais por pavimento em cada aba tematica.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra")
LM = BASE / "quantitativos" / "listas-materiais" / "eletrico"
OUT_XLSX = BASE / "disciplinas" / "eletrico" / "eletrico-electra-r03.xlsx"

sys.path.insert(0, str(BASE / "scripts"))
from normalizar_pavimento import PAVIMENTOS_ORDER, PAVIMENTO_LABEL, ordem_pavimento

AC_TOTAL_M2 = 36092
R02_TOTAL = 6_444_152

# ------------------------ Tabelas de PU (iguais a versao anterior) ------------------------

PU_QUADROS = [
    (r"QUADRO GERAL.*Q\.?G\.?B\.?T", 28_000, "QGBT — mercado SC"),
    (r"BANCO DE CAPACITORES", 32_000, "Banco capacitores automatico"),
    (r"QM.*BUS\s*WAY", 12_000, "QM Bus Way medicao"),
    (r"QM.*CONVENCIONAL", 8_500, "QM convencional"),
    (r"QM.*MEDIDOR DE DEMANDA", 6_500, "QM MDR"),
    (r"CAIXA PORTA FUS.*VEIS", 4_200, "Caixa porta fusiveis NH"),
    (r"TELEMEDI.*CONCENTRADORA.*MEDI.*LOCAL", 9_800, "Telemedicao com medicao local"),
    (r"TELEMEDI.*CONCENTRADORA", 6_200, "Telemedicao concentradora"),
    (r"BEP.*EQUIPOTENCIALIZA", 3_800, "BEP"),
    (r"[ÁA]RM[AÁ]RIO.*DOCUMENTOS", 1_500, "Armario documentos"),
]


def pu_quadro(desc: str) -> tuple[float, str]:
    for pattern, pu, fonte in PU_QUADROS:
        if re.search(pattern, desc.upper()):
            return pu, fonte
    return 3_000.0, "fallback — revisar"


def pu_caixa_passagem(dim: str) -> tuple[float, str]:
    m = re.match(r"(\d+)\s*[xX]\s*(\d+)\s*[xX]\s*(\d+)", dim)
    if not m:
        return 150.0, "fallback (dim VERIFICAR DIAGRAMA)"
    w, h, p = int(m.group(1)), int(m.group(2)), int(m.group(3))
    vol = (w * h * p) / 1000
    if vol <= 3:
        return 75.0, "<= 3 dm3"
    if vol <= 8:
        return 130.0, "3-8 dm3"
    if vol <= 20:
        return 240.0, "8-20 dm3"
    if vol <= 40:
        return 420.0, "20-40 dm3"
    return 680.0, "> 40 dm3"


PU_MATERIAL = [
    (r"BARRAMENTO NEUTRO.*BASE AZUL", 48.0, "Barr neutro c/ base"),
    (r"BARRAMENTO TERRA.*BASE VERDE", 52.0, "Barr terra c/ base"),
    (r"BARRAMENTO PENTE.*TRIF[AÁ]SICO.*18P", 85.0, "Barr pente 18P 63A"),
    (r"BARRAMENTO PENTE.*TRIF[AÁ]SICO.*12P", 68.0, "Barr pente 12P 80A"),
    (r"TERMINAL PR[EÉ] ISOLADO TIPO TUBO.*DUPLO.*10", 0.90, "Ilhos duplo 10mm²"),
    (r"TERMINAL PR[EÉ] ISOLADO TIPO TUBO.*10", 0.65, "Ilhos 10mm²"),
    (r"TERMINAL PR[EÉ] ISOLADO TIPO TUBO.*6", 0.45, "Ilhos 6mm²"),
    (r"CABO 10MM.*FLEX[IÍ]VEL", 12.50, "Cabo 10mm² flex"),
    (r"CABO 6MM.*FLEX[IÍ]VEL", 8.20, "Cabo 6mm² flex"),
    (r"CABO 4MM.*FLEX[IÍ]VEL", 5.60, "Cabo 4mm² flex"),
    (r"CABO 2[,.]5MM.*FLEX[IÍ]VEL", 3.80, "Cabo 2,5mm² flex"),
    (r"CABO.*1[,.]5MM.*FLEX[IÍ]VEL", 2.60, "Cabo 1,5mm² flex"),
    (r"CABO.*16MM.*FLEX[IÍ]VEL", 19.80, "Cabo 16mm² flex"),
    (r"CABO.*25MM.*FLEX[IÍ]VEL", 31.50, "Cabo 25mm² flex"),
    (r"CABO.*35MM.*FLEX[IÍ]VEL", 44.00, "Cabo 35mm² flex"),
    (r"CABO.*50MM.*FLEX[IÍ]VEL", 62.00, "Cabo 50mm² flex"),
]


def pu_material(desc: str) -> tuple[float, str]:
    for pattern, pu, fonte in PU_MATERIAL:
        if re.search(pattern, desc.upper()):
            return pu, fonte
    return 15.0, "fallback"


PU_DISJUNTORES = [
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*6A", 24.0, "Mini mono 6A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*10A", 24.0, "Mini mono 10A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*16A", 26.0, "Mini mono 16A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*20A", 28.0, "Mini mono 20A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*25A", 30.0, "Mini mono 25A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*32A", 34.0, "Mini mono 32A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*40A", 42.0, "Mini mono 40A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*50A", 52.0, "Mini mono 50A"),
    (r"MINIDISJUNTOR\s*-\s*MONOF[AÁ]SICO.*63A", 68.0, "Mini mono 63A"),
    (r"MINIDISJUNTOR\s*-\s*BIPOLAR", 85.0, "Mini bipolar"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*6A", 120.0, "Mini trif 6A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*10A", 120.0, "Mini trif 10A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*16A", 125.0, "Mini trif 16A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*20A", 130.0, "Mini trif 20A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*25A", 140.0, "Mini trif 25A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*32A.*6KA", 180.0, "Mini trif 32A 6kA"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*32A", 155.0, "Mini trif 32A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*40A", 170.0, "Mini trif 40A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*50A", 190.0, "Mini trif 50A"),
    (r"MINIDISJUNTOR\s*-\s*TRIF[AÁ]SICO.*63A", 220.0, "Mini trif 63A"),
    (r"IDR.*BIPOLAR.*25A", 195.0, "IDR bi 25A"),
    (r"IDR.*BIPOLAR.*40A", 240.0, "IDR bi 40A"),
    (r"IDR.*BIPOLAR.*63A", 320.0, "IDR bi 63A"),
    (r"IDR.*TETRAPOLAR.*40A", 480.0, "IDR tetra 40A"),
    (r"IDR.*TETRAPOLAR.*63A", 560.0, "IDR tetra 63A"),
    (r"IDR.*TETRAPOLAR.*80A", 720.0, "IDR tetra 80A"),
    (r"IDR.*TETRAPOLAR.*100A", 920.0, "IDR tetra 100A"),
    (r"DPS.*MONOPOLAR.*275V", 130.0, "DPS mono 275V"),
    (r"DPS.*TRIPOLAR.*275V", 340.0, "DPS tri 275V"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*70A", 720.0, "Disj cx mold 70A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*100A", 880.0, "Disj cx mold 100A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*125A", 1_050.0, "Disj cx mold 125A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*150A", 1_280.0, "Disj cx mold 150A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*175A", 1_480.0, "Disj cx mold 175A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*200A", 1_680.0, "Disj cx mold 200A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*225A", 1_920.0, "Disj cx mold 225A"),
    (r"DISJUNTOR\s*-\s*TRIF[AÁ]SICO.*CAIXA MOLDADA.*250A", 2_180.0, "Disj cx mold 250A"),
    (r"TEMPORIZADOR.*TIMER.*DIGITAL", 215.0, "Timer digital"),
    (r"CONTATOR\s*10A", 185.0, "Contator 10A tri"),
    (r"CONTATOR\s*25A", 260.0, "Contator 25A tri"),
    (r"CONTATOR\s*40A", 410.0, "Contator 40A tri"),
]


def pu_disjuntor(desc: str) -> tuple[float, str]:
    up = desc.upper()
    for pattern, pu, fonte in PU_DISJUNTORES:
        if re.search(pattern, up):
            return pu, fonte
    return 50.0, "fallback"


# ------------------------ Estilos ------------------------

def styles():
    return {
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "header_fill": PatternFill("solid", fgColor="C00000"),
        "subtot_fill": PatternFill("solid", fgColor="FFF2CC"),
        "total_fill": PatternFill("solid", fgColor="C6E0B4"),
        "warn_fill": PatternFill("solid", fgColor="FFE699"),
        "pav_fill": PatternFill("solid", fgColor="DDEBF7"),
        "border": Border(*[Side(border_style="thin", color="888888")] * 4),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
    }


def write_headers(ws, headers, row, s):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]


def put_meta(ws, disciplina_header, revisao):
    ws["A1"] = "THOZEN - RES. ELECTRA TOWERS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = disciplina_header
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = f"Revisao: {revisao} | Data: 2026-04-14 | AC: 36.092 m2 | Organizado por pavimento"
    ws["A4"] = "Fonte: PDFs Eletrowatts 07/04/2026"
    ws["A4"].font = Font(italic=True, size=9, color="666666")
    ws["A5"] = "Suplemento ao eletrico-r02-completo.xlsx (25/mar/2026)"
    ws["A5"].font = Font(italic=True, size=9, color="C00000")


def _escrever_subtotal(ws, row, first, last, label, total_col_letter, total_cols_merge, s):
    """Escreve linha de subtotal por pavimento."""
    ws.cell(row=row, column=1, value=f"Subtotal — {label}").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols_merge)
    ws.cell(row=row, column=1).fill = s["subtot_fill"]
    ws.cell(row=row, column=1).alignment = s["left"]
    ws.cell(row=row, column=total_cols_merge + 1,
            value=f"=SUM({total_col_letter}{first}:{total_col_letter}{last})").number_format = '#,##0.00'
    ws.cell(row=row, column=total_cols_merge + 1).font = Font(bold=True)
    ws.cell(row=row, column=total_cols_merge + 1).fill = s["subtot_fill"]
    # Borda full row
    for col in range(1, total_cols_merge + 3):
        ws.cell(row=row, column=col).border = s["border"]
        if col != total_cols_merge + 1:
            ws.cell(row=row, column=col).fill = s["subtot_fill"]


# ------------------------ Abas ------------------------

def aba_quadros_principais(wb, s) -> tuple[str, int, float]:
    ws = wb.active
    ws.title = "QUADROS PRINCIPAIS"
    put_meta(ws, "Disciplina: 06 Eletrico - Quadros Principais (aprovativo, fallback Terreo)", "R03")

    headers = ["Pavimento", "Torre", "#", "Qtd", "Unid", "Descricao", "PU (R$)", "Total (R$)", "Fonte PU"]
    write_headers(ws, headers, 7, s)

    aprov = json.loads((LM / "aprovativo.json").read_text(encoding="utf-8"))
    items = aprov["documentos"][0]["itens"]
    items.sort(key=lambda it: (ordem_pavimento(it["pavimento"]), it["torre"], it["numero"]))

    row = 8
    current_pav = None
    pav_start = row
    pav_subtotal_rows = []

    def fechar_pav():
        nonlocal row
        if current_pav is None:
            return
        label = PAVIMENTO_LABEL.get(current_pav, current_pav)
        _escrever_subtotal(ws, row, pav_start, row - 1, label, "H", 7, s)
        pav_subtotal_rows.append(row)
        row += 1

    for it in items:
        prod = it.get("produto") or ""
        pu, fonte = pu_quadro(prod)
        if it["pavimento"] != current_pav:
            fechar_pav()
            current_pav = it["pavimento"]
            pav_start = row

        ws.cell(row=row, column=1, value=PAVIMENTO_LABEL.get(it["pavimento"], it["pavimento"])).alignment = s["left"]
        ws.cell(row=row, column=2, value=it["torre"]).alignment = s["center"]
        ws.cell(row=row, column=3, value=it["numero"]).alignment = s["center"]
        ws.cell(row=row, column=4, value=int(it["qtd"])).alignment = s["right"]
        ws.cell(row=row, column=5, value=it["unidade"]).alignment = s["center"]
        ws.cell(row=row, column=6, value=prod).alignment = s["left"]
        ws.cell(row=row, column=7, value=pu).alignment = s["right"]
        ws.cell(row=row, column=7).number_format = '#,##0.00'
        ws.cell(row=row, column=8, value=f'=D{row}*G{row}').alignment = s["right"]
        ws.cell(row=row, column=8).number_format = '#,##0.00'
        ws.cell(row=row, column=9, value=fonte).alignment = s["left"]
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = s["border"]
        if "fallback" in fonte:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = s["warn_fill"]
        row += 1

    fechar_pav()

    # Total
    row += 1
    ws.cell(row=row, column=1, value="TOTAL QUADROS PRINCIPAIS").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    if pav_subtotal_rows:
        ws.cell(row=row, column=8, value="+".join(f"H{r}" for r in pav_subtotal_rows)).number_format = '#,##0.00'
        ws.cell(row=row, column=8).value = "=" + "+".join(f"H{r}" for r in pav_subtotal_rows)
    ws.cell(row=row, column=8).number_format = '#,##0.00'
    ws.cell(row=row, column=8).font = Font(bold=True, size=11)
    ws.cell(row=row, column=8).fill = s["total_fill"]
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = s["border"]
    total_row = row

    widths = [22, 6, 4, 6, 6, 60, 12, 14, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    # Calcula total pra resumo
    total_calc = sum(int(it["qtd"]) * pu_quadro(it.get("produto") or "")[0] for it in items)
    return ws.title, total_row, total_calc


def aba_caixas_passagem(wb, s) -> tuple[str, int, float]:
    ws = wb.create_sheet("CAIXAS PASSAGEM")
    put_meta(ws, "Disciplina: 06 Eletrico - Caixas de Passagem (Executivo + Preventivo)", "R03")

    headers = ["Pavimento", "Torre", "Fonte", "#", "Qtd", "Unid", "Dimensoes", "Local (orig)",
               "PU (R$)", "Total (R$)", "Fonte PU"]
    write_headers(ws, headers, 7, s)

    executivo = json.loads((LM / "executivo.json").read_text(encoding="utf-8"))
    preventivo = json.loads((LM / "preventivo.json").read_text(encoding="utf-8"))

    items = []
    for src, data in [("Exec", executivo), ("Prev", preventivo)]:
        for it in data["documentos"][0]["itens"]:
            items.append((src, it))
    items.sort(key=lambda x: (ordem_pavimento(x[1]["pavimento"]), x[1]["torre"], x[0], x[1]["numero"]))

    row = 8
    current_pav = None
    pav_start = row
    pav_subtotal_rows = []

    def fechar_pav():
        nonlocal row
        if current_pav is None:
            return
        label = PAVIMENTO_LABEL.get(current_pav, current_pav)
        _escrever_subtotal(ws, row, pav_start, row - 1, label, "J", 9, s)
        pav_subtotal_rows.append(row)
        row += 1

    for src, it in items:
        pu, fonte = pu_caixa_passagem(it["dimensoes"])
        if it["pavimento"] != current_pav:
            fechar_pav()
            current_pav = it["pavimento"]
            pav_start = row

        ws.cell(row=row, column=1, value=PAVIMENTO_LABEL.get(it["pavimento"], it["pavimento"])).alignment = s["left"]
        ws.cell(row=row, column=2, value=it["torre"]).alignment = s["center"]
        ws.cell(row=row, column=3, value=src).alignment = s["center"]
        ws.cell(row=row, column=4, value=it["numero"]).alignment = s["center"]
        ws.cell(row=row, column=5, value=int(it["qtd"])).alignment = s["right"]
        ws.cell(row=row, column=6, value=it["unidade"]).alignment = s["center"]
        ws.cell(row=row, column=7, value=it["dimensoes"]).alignment = s["center"]
        ws.cell(row=row, column=8, value=it["especificacao_local"]).alignment = s["left"]
        ws.cell(row=row, column=9, value=pu).alignment = s["right"]
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        ws.cell(row=row, column=10, value=f'=E{row}*I{row}').alignment = s["right"]
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=fonte).alignment = s["left"]
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = s["border"]
        if "fallback" in fonte:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = s["warn_fill"]
        row += 1

    fechar_pav()

    row += 1
    ws.cell(row=row, column=1, value="TOTAL CAIXAS PASSAGEM").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    if pav_subtotal_rows:
        ws.cell(row=row, column=10, value="=" + "+".join(f"J{r}" for r in pav_subtotal_rows)).number_format = '#,##0.00'
    ws.cell(row=row, column=10).number_format = '#,##0.00'
    ws.cell(row=row, column=10).font = Font(bold=True, size=11)
    ws.cell(row=row, column=10).fill = s["total_fill"]
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = s["border"]
    total_row = row

    widths = [22, 6, 6, 4, 6, 5, 18, 38, 12, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    total_calc = sum(int(it["qtd"]) * pu_caixa_passagem(it["dimensoes"])[0] for _, it in items)
    return ws.title, total_row, total_calc


def aba_material_interno(wb, s) -> tuple[str, int, float]:
    ws = wb.create_sheet("MATERIAL INTERNO CD")
    put_meta(ws, "Disciplina: 06 Eletrico - Material Interno CD (QT-MAT-1940/1941/1950)", "R03")

    headers = ["Pavimento", "Torre", "Doc (referente)", "#", "Codigo", "Qtd", "Unid",
               "Descricao", "PU (R$)", "Total (R$)", "Fonte PU"]
    write_headers(ws, headers, 7, s)

    geral = json.loads((LM / "geral.json").read_text(encoding="utf-8"))

    rows_list = []
    for doc in geral["documentos"]:
        doc_label = (doc["metadata"].get("codigo", "") + " - " +
                     doc["metadata"].get("referente", ""))
        for it in doc["itens"]:
            rows_list.append((doc_label, it))
    rows_list.sort(key=lambda x: (ordem_pavimento(x[1]["pavimento"]), x[1]["torre"], x[0], x[1]["numero"]))

    row = 8
    current_pav = None
    pav_start = row
    pav_subtotal_rows = []

    def fechar_pav():
        nonlocal row
        if current_pav is None:
            return
        label = PAVIMENTO_LABEL.get(current_pav, current_pav)
        _escrever_subtotal(ws, row, pav_start, row - 1, label, "J", 9, s)
        pav_subtotal_rows.append(row)
        row += 1

    for doc_label, it in rows_list:
        pu, fonte = pu_material(it["descricao"])
        if it["pavimento"] != current_pav:
            fechar_pav()
            current_pav = it["pavimento"]
            pav_start = row

        ws.cell(row=row, column=1, value=PAVIMENTO_LABEL.get(it["pavimento"], it["pavimento"])).alignment = s["left"]
        ws.cell(row=row, column=2, value=it["torre"]).alignment = s["center"]
        ws.cell(row=row, column=3, value=doc_label).alignment = s["left"]
        ws.cell(row=row, column=4, value=it["numero"]).alignment = s["center"]
        ws.cell(row=row, column=5, value=it["codigo"]).alignment = s["center"]
        ws.cell(row=row, column=6, value=int(it["qtd"])).alignment = s["right"]
        ws.cell(row=row, column=7, value=it["unidade"]).alignment = s["center"]
        ws.cell(row=row, column=8, value=it["descricao"]).alignment = s["left"]
        ws.cell(row=row, column=9, value=pu).alignment = s["right"]
        ws.cell(row=row, column=9).number_format = '#,##0.00'
        ws.cell(row=row, column=10, value=f'=F{row}*I{row}').alignment = s["right"]
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=fonte).alignment = s["left"]
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = s["border"]
        if "fallback" in fonte:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = s["warn_fill"]
        row += 1

    fechar_pav()

    row += 1
    ws.cell(row=row, column=1, value="TOTAL MATERIAL INTERNO CD").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    if pav_subtotal_rows:
        ws.cell(row=row, column=10, value="=" + "+".join(f"J{r}" for r in pav_subtotal_rows)).number_format = '#,##0.00'
    ws.cell(row=row, column=10).number_format = '#,##0.00'
    ws.cell(row=row, column=10).font = Font(bold=True, size=11)
    ws.cell(row=row, column=10).fill = s["total_fill"]
    for col in range(1, 12):
        ws.cell(row=row, column=col).border = s["border"]
    total_row = row

    widths = [22, 6, 40, 4, 14, 6, 5, 52, 12, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    total_calc = sum(int(it["qtd"]) * pu_material(it["descricao"])[0] for _, it in rows_list)
    return ws.title, total_row, total_calc


def aba_disjuntores(wb, s) -> tuple[str, int, float]:
    ws = wb.create_sheet("DISJUNTORES 18 CDs")
    put_meta(ws, "Disciplina: 06 Eletrico - Disjuntores/IDR/DPS (18 CDs por pavimento)", "R03")

    headers = ["Pavimento", "Torre", "CD (QT-MAT)", "Referente", "#", "Codigo", "Qtd", "Unid",
               "Descricao", "PU (R$)", "Total (R$)", "Fonte PU"]
    write_headers(ws, headers, 7, s)

    disj = json.loads((LM / "disjuntores.json").read_text(encoding="utf-8"))

    rows_list = []
    for doc in disj["documentos"]:
        meta = doc["metadata"]
        cd_code = meta.get("codigo", "")
        cd_ref = meta.get("referente", "")
        for it in doc["itens"]:
            rows_list.append((cd_code, cd_ref, it))
    rows_list.sort(key=lambda x: (ordem_pavimento(x[2]["pavimento"]), x[2]["torre"], x[0], x[2]["numero"]))

    row = 8
    current_pav = None
    pav_start = row
    pav_subtotal_rows = []

    def fechar_pav():
        nonlocal row
        if current_pav is None:
            return
        label = PAVIMENTO_LABEL.get(current_pav, current_pav)
        _escrever_subtotal(ws, row, pav_start, row - 1, label, "K", 10, s)
        pav_subtotal_rows.append(row)
        row += 1

    for cd_code, cd_ref, it in rows_list:
        pu, fonte = pu_disjuntor(it["descricao"])
        if it["pavimento"] != current_pav:
            fechar_pav()
            current_pav = it["pavimento"]
            pav_start = row

        ws.cell(row=row, column=1, value=PAVIMENTO_LABEL.get(it["pavimento"], it["pavimento"])).alignment = s["left"]
        ws.cell(row=row, column=2, value=it["torre"]).alignment = s["center"]
        ws.cell(row=row, column=3, value=cd_code).alignment = s["center"]
        ws.cell(row=row, column=4, value=cd_ref).alignment = s["left"]
        ws.cell(row=row, column=5, value=it["numero"]).alignment = s["center"]
        ws.cell(row=row, column=6, value=it["codigo"]).alignment = s["center"]
        ws.cell(row=row, column=7, value=int(it["qtd"])).alignment = s["right"]
        ws.cell(row=row, column=8, value=it["unidade"]).alignment = s["center"]
        ws.cell(row=row, column=9, value=it["descricao"]).alignment = s["left"]
        ws.cell(row=row, column=10, value=pu).alignment = s["right"]
        ws.cell(row=row, column=10).number_format = '#,##0.00'
        ws.cell(row=row, column=11, value=f'=G{row}*J{row}').alignment = s["right"]
        ws.cell(row=row, column=11).number_format = '#,##0.00'
        ws.cell(row=row, column=12, value=fonte).alignment = s["left"]
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = s["border"]
        if "fallback" in fonte:
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = s["warn_fill"]
        row += 1

    fechar_pav()

    row += 1
    ws.cell(row=row, column=1, value="TOTAL DISJUNTORES").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    if pav_subtotal_rows:
        ws.cell(row=row, column=11, value="=" + "+".join(f"K{r}" for r in pav_subtotal_rows)).number_format = '#,##0.00'
    ws.cell(row=row, column=11).number_format = '#,##0.00'
    ws.cell(row=row, column=11).font = Font(bold=True, size=11)
    ws.cell(row=row, column=11).fill = s["total_fill"]
    for col in range(1, 13):
        ws.cell(row=row, column=col).border = s["border"]
    total_row = row

    widths = [22, 6, 14, 45, 4, 14, 6, 5, 50, 12, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[7].height = 22
    ws.freeze_panes = "A8"

    total_calc = sum(int(it["qtd"]) * pu_disjuntor(it["descricao"])[0] for _, _, it in rows_list)
    return ws.title, total_row, total_calc


def aba_resumo(wb, s, totais: dict) -> None:
    ws = wb.create_sheet("RESUMO R03", 0)
    ws["A1"] = "ELETRICO R03 — Resumo por Pavimento"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Suplemento do PDFs Eletrowatts 07/04/2026 ao R02 existente — organizado por pavimento"
    ws["A2"].font = Font(bold=True, size=11)

    # Seção 1: totais por aba
    write_headers(ws, ["Aba tematica", "Escopo", "Total (R$)", "R$/m2 AC", "Fallbacks"], 4, s)

    row = 5
    for aba, desc, total, fb_count, ncol_total in [
        ("QUADROS PRINCIPAIS", "16 quadros principais (QGBT/QM/BEP/banco cap)",
         totais["quadros"]["total"], totais["quadros"]["fallback"], "H"),
        ("CAIXAS PASSAGEM", "162 caixas de passagem (Exec+Prev) por pavimento",
         totais["caixas"]["total"], totais["caixas"]["fallback"], "J"),
        ("MATERIAL INTERNO CD", "39 itens barramentos/cabos/terminais interno CDs",
         totais["material"]["total"], totais["material"]["fallback"], "J"),
        ("DISJUNTORES 18 CDs", "167 disjuntores/IDR/DPS em 18 CDs",
         totais["disjuntores"]["total"], totais["disjuntores"]["fallback"], "K"),
    ]:
        ws.cell(row=row, column=1, value=aba).alignment = s["left"]
        ws.cell(row=row, column=2, value=desc).alignment = s["left"]
        ws.cell(row=row, column=3, value=total).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=total / AC_TOTAL_M2).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=fb_count).alignment = s["center"]
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = s["border"]
        row += 1

    tot_supl = sum(t["total"] for k, t in totais.items() if k != "pivot_by_pav")
    ws.cell(row=row, column=1, value="TOTAL SUPLEMENTO R03").font = Font(bold=True, size=11)
    ws.cell(row=row, column=2, value="(sem MO adicional — MO embutida nos PUs)").alignment = s["left"]
    ws.cell(row=row, column=3, value=tot_supl).number_format = '#,##0.00'
    ws.cell(row=row, column=3).font = Font(bold=True, size=11)
    ws.cell(row=row, column=3).fill = s["total_fill"]
    ws.cell(row=row, column=4, value=tot_supl / AC_TOTAL_M2).number_format = '#,##0.00'
    ws.cell(row=row, column=4).fill = s["total_fill"]
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = s["border"]
    row += 2

    # Seção 2: pivot por pavimento (total consolidado através das 4 abas)
    ws.cell(row=row, column=1, value="DISTRIBUICAO POR PAVIMENTO (todas as 4 abas)").font = Font(bold=True, size=11)
    row += 1
    write_headers(ws, ["Pavimento", "Quadros R$", "Caixas R$", "Material R$", "Disjuntores R$", "Total R$"], row, s)
    pivot_start_row = row + 1
    row += 1

    pivots = totais["pivot_by_pav"]
    pavs = sorted(pivots.keys(), key=ordem_pavimento)
    for pav in pavs:
        ws.cell(row=row, column=1, value=PAVIMENTO_LABEL.get(pav, pav)).alignment = s["left"]
        p = pivots[pav]
        ws.cell(row=row, column=2, value=p.get("quadros", 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=p.get("caixas", 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=p.get("material", 0)).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=p.get("disjuntores", 0)).number_format = '#,##0.00'
        total_pav = sum(p.values())
        ws.cell(row=row, column=6, value=total_pav).number_format = '#,##0.00'
        ws.cell(row=row, column=6).font = Font(bold=True)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = s["border"]
        row += 1

    # Linha total pivot
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = s["total_fill"]
    for col, key in [(2, "quadros"), (3, "caixas"), (4, "material"), (5, "disjuntores")]:
        v = sum(p.get(key, 0) for p in pivots.values())
        ws.cell(row=row, column=col, value=v).number_format = '#,##0.00'
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = s["total_fill"]
    ws.cell(row=row, column=6, value=tot_supl).number_format = '#,##0.00'
    ws.cell(row=row, column=6).font = Font(bold=True)
    ws.cell(row=row, column=6).fill = s["total_fill"]
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = s["border"]
    row += 2

    # Seção 3: alerta
    ws.cell(row=row, column=1, value="⚠️ R03 NAO SOMA AO R02").font = Font(bold=True, color="C00000")
    ws.cell(row=row, column=2, value="Substitui os itens 'vermelhos' (sem fonte) do R02 pelos itens rastreaveis PDF").alignment = s["left"]
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = s["warn_fill"]
        ws.cell(row=row, column=col).border = s["border"]
    row += 1

    ws.cell(row=row, column=1, value="R02 anterior").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Total R02 (log sec 19)").alignment = s["left"]
    ws.cell(row=row, column=3, value=R02_TOTAL).number_format = '#,##0.00'
    ws.cell(row=row, column=4, value=R02_TOTAL / AC_TOTAL_M2).number_format = '#,##0.00'
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = s["border"]
    row += 2

    # Pendencias
    ws.cell(row=row, column=1, value="PENDENCIAS R02 nao cobertas pelos PDFs").font = Font(bold=True)
    row += 1
    for p in [
        "1. Comprimento medio trechos eletroduto (IFC sem Length)",
        "2. Fator cabos vs eletrodutos",
        "3. Subestacao/gerador/barramento compartilhados?",
        "4. MO eletrica (mediana base R$ 26/m2 vs R02 R$ 170/m2)",
        "5. Cabos de forca metragem",
        "6. Luminarias PU",
        "7. PU medio dos fallback amarelo em CAIXAS PASSAGEM",
    ]:
        ws.cell(row=row, column=2, value=p).alignment = s["left"]
        row += 1

    for i, w in enumerate([24, 60, 16, 12, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def compute_pivot(totais_items: dict) -> dict:
    """Agrega total por (pavimento, aba) -> R$."""
    pivot: dict = {}
    for aba_key, items in totais_items.items():
        for pav, subtotal in items.items():
            pivot.setdefault(pav, {})[aba_key] = pivot.setdefault(pav, {}).get(aba_key, 0) + subtotal
    return pivot


def main():
    wb = openpyxl.Workbook()
    s = styles()

    # Pre-calcula pivots por pavimento
    aprov = json.loads((LM / "aprovativo.json").read_text(encoding="utf-8"))
    exe = json.loads((LM / "executivo.json").read_text(encoding="utf-8"))
    prev = json.loads((LM / "preventivo.json").read_text(encoding="utf-8"))
    ger = json.loads((LM / "geral.json").read_text(encoding="utf-8"))
    disj = json.loads((LM / "disjuntores.json").read_text(encoding="utf-8"))

    def _pivot(rows, pu_fn, desc_key):
        p: dict = {}
        fb = 0
        tot = 0
        for it in rows:
            pu, fonte = pu_fn(it[desc_key])
            v = int(it["qtd"]) * pu
            p[it["pavimento"]] = p.get(it["pavimento"], 0) + v
            tot += v
            if "fallback" in fonte:
                fb += 1
        return p, tot, fb

    p_q, t_q, fb_q = _pivot(aprov["documentos"][0]["itens"],
                             lambda d: pu_quadro(d or ""), "produto")
    p_c, t_c, fb_c = _pivot(exe["documentos"][0]["itens"] + prev["documentos"][0]["itens"],
                             pu_caixa_passagem, "dimensoes")
    g_items = [it for doc in ger["documentos"] for it in doc["itens"]]
    p_m, t_m, fb_m = _pivot(g_items, pu_material, "descricao")
    d_items = [it for doc in disj["documentos"] for it in doc["itens"]]
    p_d, t_d, fb_d = _pivot(d_items, pu_disjuntor, "descricao")

    totais = {
        "quadros": {"total": t_q, "fallback": fb_q, "pivot": p_q},
        "caixas": {"total": t_c, "fallback": fb_c, "pivot": p_c},
        "material": {"total": t_m, "fallback": fb_m, "pivot": p_m},
        "disjuntores": {"total": t_d, "fallback": fb_d, "pivot": p_d},
    }
    pivot_total: dict = {}
    for key, data in totais.items():
        for pav, v in data["pivot"].items():
            pivot_total.setdefault(pav, {})[key] = pivot_total.setdefault(pav, {}).get(key, 0) + v
    totais["pivot_by_pav"] = pivot_total

    # Gera abas
    aba_quadros_principais(wb, s)
    aba_caixas_passagem(wb, s)
    aba_material_interno(wb, s)
    aba_disjuntores(wb, s)
    aba_resumo(wb, s, totais)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"gravado: {OUT_XLSX}")
    print(f"  Quadros:    R$ {t_q:>14,.2f}  fallbacks={fb_q}")
    print(f"  Caixas:     R$ {t_c:>14,.2f}  fallbacks={fb_c}")
    print(f"  Material:   R$ {t_m:>14,.2f}  fallbacks={fb_m}")
    print(f"  Disjuntor:  R$ {t_d:>14,.2f}  fallbacks={fb_d}")
    print(f"  TOTAL R03:  R$ {t_q+t_c+t_m+t_d:>14,.2f}  ({(t_q+t_c+t_m+t_d)/AC_TOTAL_M2:.2f} R$/m²)")


if __name__ == "__main__":
    main()
