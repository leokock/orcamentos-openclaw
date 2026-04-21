#!/usr/bin/env python3.11
"""
Gera telecomunicacoes-electra-r02.xlsx separando Torre A e Torre B.
Base: R01 (torre única) + dados brutos JSON + extrações por DXF.
"""

import json
import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path.home() / "orcamentos/executivos/thozen-electra"
R01_PATH = BASE_DIR / "entregas/telecomunicacoes-electra-r01.xlsx"
R02_PATH = BASE_DIR / "entregas/telecomunicacoes-electra-r02.xlsx"
JSON_PATH = BASE_DIR / "quantitativos/telefonico/dados_brutos_telefonico.json"
EXTRACT_DIR = BASE_DIR / "quantitativos/telefonico"

# DXF -> (pavimento, torre)
DXF_MAP = {
    "T02": ("TÉRREO", "A"),
    "T03": ("TÉRREO", "B"),
    "T04": ("GARAGEM 01", "A"),
    "T05": ("GARAGEM 01", "B"),
    "T06": ("GARAGEM 02", "A"),
    "T07": ("GARAGEM 02", "B"),
    "T08": ("GARAGEM 03", "A"),
    "T09": ("GARAGEM 03", "B"),
    "T10": ("GARAGEM 04", "A"),
    "T11": ("GARAGEM 04", "B"),
    "T12": ("GARAGEM 05", "A"),
    "T13": ("GARAGEM 05", "B"),
    "T14": ("LAZER", "A"),
    "T15": ("LAZER", "B"),
    "T16": ("PAVIMENTO TIPO", "A"),
    "T17": ("PAVIMENTO TIPO", "B"),
    "T18": ("CASA DE MÁQUINAS", "A"),
    "T19": ("CASA DE MÁQUINAS", "B"),
}

PAVIMENTO_ORDER = [
    "TÉRREO", "GARAGEM 01", "GARAGEM 02", "GARAGEM 03",
    "GARAGEM 04", "GARAGEM 05", "LAZER", "PAVIMENTO TIPO",
    "CASA DE MÁQUINAS"
]

# Subgrupo ordering within each pavimento
SUBGRUPO_ORDER = [
    "CAIXA OCTOGONAL",
    "CAIXA 4x2\" E 4x4\"",
    "CAIXA DE PASSAGEM",
    "ELETRODUTOS",
    "MÓDULOS",
    "QUADROS DE DISTRIBUIÇÃO",
    "COTOVELOS",
    "CONECTORES E BUCHAS",
    "CONDUTORES TELECOM",
    "PONTOS DE USO",
]


def load_r01():
    """Load R01 and extract item structure + cost map per pavimento."""
    wb = openpyxl.load_workbook(R01_PATH)
    ws = wb.active

    pavimentos = {}
    current_pav = None

    for row in range(7, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        e_val = ws.cell(row=row, column=5).value
        d_val = ws.cell(row=row, column=4).value
        f_val = ws.cell(row=row, column=6).value
        g_val = ws.cell(row=row, column=7).value
        j_val = ws.cell(row=row, column=10).value
        k_val = ws.cell(row=row, column=11).value

        if b_val is not None and not str(b_val).startswith("="):
            current_pav = b_val
            pavimentos[current_pav] = {"items": []}
        elif e_val is not None and current_pav:
            pavimentos[current_pav]["items"].append({
                "subgrupo": d_val if d_val and not str(d_val).startswith("=") else None,
                "descricao": e_val,
                "qtd": f_val,
                "rep": g_val,
                "unidade": j_val,
                "custo_unit": k_val,
            })

    # Build global cost map
    cost_map = {}
    for pav, info in pavimentos.items():
        for item in info["items"]:
            if item["custo_unit"] and item["custo_unit"] > 0:
                cost_map[item["descricao"]] = item["custo_unit"]

    return wb, pavimentos, cost_map


def parse_extraction(filepath):
    """Parse an extraction markdown file and return structured data."""
    text = filepath.read_text(encoding="utf-8")
    items = {}

    # Parse eletrodutos
    eletrodutos = {}
    in_eletrodutos = False
    for line in text.split("\n"):
        if "Eletrodutos (Diâmetros)" in line:
            in_eletrodutos = True
            continue
        if in_eletrodutos:
            if line.startswith("##") or (line.startswith("*") and "Eletrodutos verticais" in line):
                if line.startswith("##"):
                    in_eletrodutos = False
                continue
            m = re.match(r"\|\s*(.+?)\s*\|\s*(\d+)\s*\|", line)
            if m and "Especificação" not in m.group(1) and "---" not in m.group(1):
                spec = m.group(1).strip()
                qty = int(m.group(2))
                eletrodutos[spec] = qty

    # Parse caixas de passagem
    caixas = {}
    in_caixas = False
    for line in text.split("\n"):
        if "Caixas de Passagem" in line:
            in_caixas = True
            continue
        if in_caixas:
            if line.startswith("##"):
                in_caixas = False
                continue
            m = re.match(r"\|\s*(.+?)\s*\|\s*(\d+)\s*\|", line)
            if m and "Dimensão" not in m.group(1) and "---" not in m.group(1):
                dim = m.group(1).strip()
                qty = int(m.group(2))
                caixas[dim] = qty

    # Parse componentes
    componentes = {}
    in_comp = False
    for line in text.split("\n"):
        if "Componentes (Blocos)" in line:
            in_comp = True
            continue
        if in_comp:
            if line.startswith("##") or line.startswith("*Blocos"):
                if line.startswith("##"):
                    in_comp = False
                continue
            m = re.match(r"\|\s*(.+?)\s*\|\s*(\d+)\s*\|", line)
            if m and "Componente" not in m.group(1) and "---" not in m.group(1):
                comp = m.group(1).strip()
                qty = int(m.group(2))
                componentes[comp] = qty

    # Parse comprimento total
    comprimento = 0
    for line in text.split("\n"):
        m = re.match(r"\*\*Comprimento total estimado:\*\*\s*([\d.]+)", line)
        if m:
            comprimento = float(m.group(1))

    return {
        "eletrodutos": eletrodutos,
        "caixas_passagem": caixas,
        "componentes": componentes,
        "comprimento": comprimento,
    }


def build_tower_data(r01_pavimentos, cost_map):
    """Build per-tower, per-pavimento item lists from extraction files."""
    tower_data = {"A": {}, "B": {}}

    for dxf_key, (pav, torre) in DXF_MAP.items():
        ext_file = EXTRACT_DIR / f"extracao_{dxf_key.lower()}.md"
        if not ext_file.exists():
            print(f"WARNING: {ext_file} not found, skipping")
            continue

        ext = parse_extraction(ext_file)

        # Find the R01 template items for this pavimento
        r01_items = r01_pavimentos.get(pav, {}).get("items", [])

        # Build items for this tower/pavimento using R01 structure as template
        items = []
        used_descriptions = set()

        for r01_item in r01_items:
            desc = r01_item["descricao"]
            subgrupo = r01_item["subgrupo"]
            unidade = r01_item["unidade"]
            pu = cost_map.get(desc, r01_item.get("custo_unit", 0))

            # Determine quantity from extraction data
            qty = get_quantity_for_item(desc, subgrupo, ext, r01_item, torre, pav)

            if qty is not None and qty > 0:
                rep = 24 if pav == "PAVIMENTO TIPO" and r01_item["rep"] == 24 else 1
                # For Tipo, some items in R01 had rep=1 (infrastructure items)
                if pav == "PAVIMENTO TIPO" and r01_item["rep"] == 1:
                    rep = 1

                items.append({
                    "subgrupo": subgrupo,
                    "descricao": desc,
                    "qtd": qty,
                    "rep": rep,
                    "unidade": unidade,
                    "custo_unit": pu if pu and pu > 0 else 0,
                })
                used_descriptions.add(desc)

        tower_data[torre][pav] = items

    return tower_data


def get_quantity_for_item(desc, subgrupo, ext, r01_item, torre, pav):
    """Determine the quantity of an item for a specific tower from extraction data."""

    # Map R01 descriptions to extraction data
    comps = ext["componentes"]
    eletrodutos = ext["eletrodutos"]
    caixas = ext["caixas_passagem"]

    # --- CAIXA OCTOGONAL ---
    if subgrupo == "CAIXA OCTOGONAL":
        if "Placa cega" in desc:
            # Placa cega for octogonal - not directly in extraction
            # Estimate based on ratio from R01
            total_oct = sum(v for k, v in comps.items() if "Rack" in k or "Quadro" in k)
            # Actually, octogonal counts are not directly in the extraction components
            # Use the eletrodutos count as proxy or just split R01 data
            return split_r01_qty(r01_item["qtd"], torre, pav)
        else:
            # Caixa Octogonal - not directly counted in extraction components
            return split_r01_qty(r01_item["qtd"], torre, pav)

    # --- CAIXA 4x2" E 4x4" ---
    # The extraction "Caixa 4x2 - *" counts are point-of-use only, not physical boxes.
    # R01's "Caixa 4x2 de Embutir" includes infrastructure boxes. Use proportional split.
    if subgrupo == 'CAIXA 4x2" E 4x4"':
        return split_r01_qty(r01_item["qtd"], torre, pav)

    # --- CAIXA DE PASSAGEM ---
    if subgrupo == "CAIXA DE PASSAGEM":
        # Match by dimension in description
        for dim, qty in caixas.items():
            # Normalize dimension matching
            dim_clean = dim.replace(" ", "").lower()
            desc_clean = desc.lower().replace(" ", "")
            if dim_clean in desc_clean:
                return qty
        # Not found in this tower's extraction
        return 0

    # --- ELETRODUTOS ---
    # R01 values are in meters. Extraction counts are eletroduto annotations (not meters).
    # Use eletroduto occurrence ratios to split R01 meter values between towers.
    if subgrupo == "ELETRODUTOS":
        # For diameter-specific items, use the occurrence ratio for that diameter
        # between tower A and B for the same pavimento
        return split_eletroduto_by_ratio(desc, r01_item["qtd"], torre, pav, eletrodutos)

    # --- MÓDULOS ---
    if subgrupo == "MÓDULOS":
        return split_r01_qty(r01_item["qtd"], torre, pav)

    # --- QUADROS DE DISTRIBUIÇÃO ---
    if subgrupo == "QUADROS DE DISTRIBUIÇÃO":
        rack_count = comps.get("Rack/Quadro Distribuição", 0)
        if rack_count > 0:
            # R01 typically had specific rack sizes; for splitting we use extraction count
            return split_r01_qty(r01_item["qtd"], torre, pav)
        return split_r01_qty(r01_item["qtd"], torre, pav)

    # --- COTOVELOS ---
    if subgrupo == "COTOVELOS":
        return comps.get("Cotovelo Eletroduto", 0)

    # --- CONECTORES E BUCHAS ---
    if subgrupo == "CONECTORES E BUCHAS":
        if "Conector box" in desc:
            return comps.get("Conector Box/Arruela", 0)
        elif "Bucha terminal" in desc:
            return comps.get("Bucha Terminal", 0)

    # --- CONDUTORES TELECOM ---
    if subgrupo == "CONDUTORES TELECOM":
        if "CFTV" in desc:
            return comps.get("Condutor CFTV", 0)
        elif "UTP" in desc:
            return comps.get("Condutor UTP", 0)
        elif "CCI" in desc:
            return comps.get("Condutor CCI", 0)
        elif "Cordplast" in desc:
            return comps.get("Condutor Cordplast", 0)

    # --- PONTOS DE USO ---
    if subgrupo == "PONTOS DE USO":
        if "Interfone" in desc:
            return comps.get("Caixa 4x2 - Interfone", 0)
        elif "Câmera" in desc:
            return comps.get("Caixa 4x2 - Câmera", comps.get("Caixa 4x2 - Câmera CFTV", 0))
        elif "Controle de Acesso" in desc:
            return comps.get("Caixa 4x2 - Controle de Acesso", 0)
        elif "Dados/Telefone" in desc:
            total = comps.get("Caixa 4x4 - Dados/Telefone", 0) + comps.get("Caixa 4x2 - Dados/Telefone", 0)
            return total

    # Default: split R01 data proportionally
    return split_r01_qty(r01_item["qtd"], torre, pav)


def get_conduit_multiplier(spec):
    """Extract the multiplier from conduit spec like '3xø1"' -> 3."""
    m = re.match(r"(\d+)x", spec)
    if m:
        return int(m.group(1))
    return 1


# Precomputed eletroduto ratios per pavimento per diameter class
ELETRODUTO_RATIOS = {}

def compute_eletroduto_ratios():
    """Compute eletroduto occurrence ratios between towers per pavimento per diameter."""
    global ELETRODUTO_RATIOS
    with open(JSON_PATH) as f:
        data = json.load(f)

    # Diameter classes: 1", 3/4", 1.1/4", 3"
    def classify_spec(spec):
        if '¾' in spec or '3/4' in spec:
            return '3/4'
        elif '1 ¼' in spec or '1¼' in spec:
            return '1.1/4'
        elif 'ø3"' in spec:
            return '3'
        elif 'ø1"' in spec:
            return '1'
        return 'other'

    for pav in PAVIMENTO_ORDER:
        ELETRODUTO_RATIOS[pav] = {}
        # Find DXF keys for tower A and B
        a_key = None
        b_key = None
        for dxf, (p, t) in DXF_MAP.items():
            if p == pav:
                if t == "A":
                    a_key = dxf
                else:
                    b_key = dxf

        if not a_key or not b_key:
            continue

        a_data = data.get(a_key, {}).get("eletrodutos_resumo", {})
        b_data = data.get(b_key, {}).get("eletrodutos_resumo", {})

        # Sum by diameter class
        for diam_class in ['1', '3/4', '1.1/4', '3', 'other']:
            a_count = sum(v for k, v in a_data.items() if classify_spec(k) == diam_class)
            b_count = sum(v for k, v in b_data.items() if classify_spec(k) == diam_class)
            total = a_count + b_count
            if total > 0:
                ELETRODUTO_RATIOS[pav][diam_class] = {"A": a_count / total, "B": b_count / total}
            else:
                ELETRODUTO_RATIOS[pav][diam_class] = {"A": 0.5, "B": 0.5}


def split_eletroduto_by_ratio(desc, qty, torre, pav, eletrodutos_local):
    """Split eletroduto meter quantity using tower occurrence ratios."""
    if qty is None or qty == 0:
        return 0

    # Determine diameter class from description
    if '1.1/4' in desc or '1¼' in desc:
        diam = '1.1/4'
    elif 'ø3/4' in desc or 'ø¾' in desc:
        diam = '3/4'
    elif 'ø3"' in desc:
        diam = '3'
    elif 'ø1"' in desc:
        diam = '1'
    else:
        diam = 'other'

    ratios = ELETRODUTO_RATIOS.get(pav, {}).get(diam, {"A": 0.5, "B": 0.5})
    result = round(qty * ratios[torre])
    return max(result, 0)


# Precomputed split ratios from consolidado (component counts per tower)
# Using the component totals as a proxy for splitting
SPLIT_RATIOS = {}

def compute_split_ratios():
    """Compute split ratios between Tower A and B for each pavimento."""
    global SPLIT_RATIOS

    # From consolidado_telefonico.md - Componentes column (good proxy for overall activity)
    pav_data = {
        "TÉRREO": {"A": 902, "B": 1146},
        "GARAGEM 01": {"A": 213, "B": 289},
        "GARAGEM 02": {"A": 144, "B": 191},
        "GARAGEM 03": {"A": 137, "B": 191},
        "GARAGEM 04": {"A": 146, "B": 157},
        "GARAGEM 05": {"A": 138, "B": 136},
        "LAZER": {"A": 799, "B": 891},
        "PAVIMENTO TIPO": {"A": 676, "B": 687},
        "CASA DE MÁQUINAS": {"A": 113, "B": 76},
    }

    for pav, towers in pav_data.items():
        total = towers["A"] + towers["B"]
        if total > 0:
            SPLIT_RATIOS[pav] = {
                "A": towers["A"] / total,
                "B": towers["B"] / total,
            }
        else:
            SPLIT_RATIOS[pav] = {"A": 0.5, "B": 0.5}


def split_r01_qty(qty, torre, pav):
    """Split R01 quantity between towers using the ratio."""
    if qty is None or qty == 0:
        return 0
    ratio = SPLIT_RATIOS.get(pav, {"A": 0.5, "B": 0.5})
    result = round(qty * ratio[torre])
    return max(result, 0)


def copy_cell_style(src, dst):
    """Copy all style attributes from src cell to dst cell."""
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.number_format:
        dst.number_format = src.number_format


def generate_r02():
    """Generate the R02 spreadsheet."""
    compute_split_ratios()
    compute_eletroduto_ratios()

    r01_wb, r01_pavimentos, cost_map = load_r01()
    r01_ws = r01_wb.active

    tower_data = build_tower_data(r01_pavimentos, cost_map)

    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TELECOMUNICAÇÃO"

    # --- HEADER ROWS 1-4 ---
    # Copy header styles from R01
    for row in range(1, 5):
        for col in range(1, 37):
            src = r01_ws.cell(row=row, column=col)
            dst = ws.cell(row=row, column=col)
            copy_cell_style(src, dst)

    ws["C1"] = "Projeto"
    ws["C1"].font = Font(name="Arial", size=10, bold=True)
    ws["D1"] = "Electra Towers"
    ws["D1"].font = Font(name="Arial", size=10)

    ws["C2"] = "Empresa"
    ws["C2"].font = Font(name="Arial", size=10, bold=True)
    ws["D2"] = "Thozen"
    ws["D2"].font = Font(name="Arial", size=9.5)

    ws["C3"] = "Origem"
    ws["C3"].font = Font(name="Arial", size=10, bold=True)
    ws["D3"] = "CTN-TZN_ELT-TEL-R02"
    ws["D3"].font = Font(name="Arial", size=9.5)

    ws["C4"] = "Revisão"
    ws["C4"].font = Font(name="Arial", size=10, bold=True)
    ws["D4"] = "R02 - Jarvis/IFC+DWG por Torre 30/03/2026"
    ws["D4"].font = Font(name="Arial", size=10)

    # Merge D3:E3
    ws.merge_cells("D3:E3")

    # --- ROW 5 - TITLE ---
    ws["A5"] = "TELECOMUNICAÇÃO"
    ws["A5"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws["A5"].fill = PatternFill(start_color="1851D6", end_color="1851D6", fill_type="solid")
    ws.merge_cells("A5:AJ5")
    # Copy style for all merged cells in row 5
    for col in range(1, 37):
        cell = ws.cell(row=5, column=col)
        cell.fill = PatternFill(start_color="1851D6", end_color="1851D6", fill_type="solid")

    # --- ROW 6 - COLUMN HEADERS ---
    headers_main = {
        1: "TORRE", 2: "PAVIMENTO", 3: "GRUPO/EAP", 4: "SUBGRUPO",
        5: "Descrição", 6: "Quantidade", 7: "Repetição", 8: "Perda",
        9: "Qte com perda", 10: "Unidade", 11: "Custo unitário", 12: "Custo Total"
    }
    headers_pivot = {
        14: "Torre", 15: "EAP",
        16: "TÉRREO", 17: "GARAGEM 01", 18: "GARAGEM 02", 19: "GARAGEM 03",
        20: "GARAGEM 04", 21: "GARAGEM 05", 22: "LAZER",
        23: "PAVIMENTO TIPO", 24: "CASA DE MÁQUINAS", 25: "TOTAL"
    }

    for col, header in {**headers_main, **headers_pivot}.items():
        cell = ws.cell(row=6, column=col)
        cell.value = header
        # Copy style from R01 row 6
        src = r01_ws.cell(row=6, column=col)
        copy_cell_style(src, cell)
        if not cell.font.name:
            cell.font = Font(name="Poppins", size=10, bold=True)

    # --- DATA ROWS ---
    current_row = 7
    # Track pavimento->torre->EAP->subtotal row for pivot
    pivot_data = {}  # (torre, eap) -> {pav: cost}

    # Number format constants
    NUM_FMT = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
    CURR_FMT = '_-"R$"\\ * #,##0.00_-;\\-"R$"\\ * #,##0.00_-;_-"R$"\\ * "-"??_-;_-@_-'

    # Get R01 row 7 and 8 styles as templates
    header_row_style = {}
    for col in range(1, 13):
        src = r01_ws.cell(row=7, column=col)
        header_row_style[col] = {
            "font": copy(src.font),
            "fill": copy(src.fill),
            "alignment": copy(src.alignment),
            "border": copy(src.border),
            "number_format": src.number_format,
        }

    data_row_style = {}
    for col in range(1, 13):
        src = r01_ws.cell(row=8, column=col)
        data_row_style[col] = {
            "font": copy(src.font),
            "fill": copy(src.fill),
            "alignment": copy(src.alignment),
            "border": copy(src.border),
            "number_format": src.number_format,
        }

    max_data_row = 7  # will be updated

    for torre in ["A", "B"]:
        torre_label = f"Torre {torre}"
        for pav in PAVIMENTO_ORDER:
            items = tower_data.get(torre, {}).get(pav, [])
            if not items:
                continue

            # --- Pavimento header row ---
            pav_header_row = current_row
            ws.cell(row=current_row, column=1).value = torre_label
            ws.cell(row=current_row, column=2).value = pav

            # Apply header row style
            for col in range(1, 13):
                cell = ws.cell(row=current_row, column=col)
                style = header_row_style[col]
                cell.font = copy(style["font"])
                cell.fill = copy(style["fill"])
                cell.alignment = copy(style["alignment"])
                cell.border = copy(style["border"])
                cell.number_format = style["number_format"]

            # Pivot area: torre label
            ws.cell(row=current_row, column=14).value = torre_label

            current_row += 1

            # --- Data rows ---
            first_data_row = current_row
            for item in items:
                r = current_row
                ws.cell(row=r, column=1).value = f"=A{pav_header_row}"
                ws.cell(row=r, column=2).value = f"=B{pav_header_row}"
                ws.cell(row=r, column=3).value = "INSTALAÇÕES DE TELECOMUNICAÇÃO"
                ws.cell(row=r, column=4).value = item["subgrupo"]
                ws.cell(row=r, column=5).value = item["descricao"]
                ws.cell(row=r, column=6).value = item["qtd"]
                ws.cell(row=r, column=7).value = item["rep"]
                ws.cell(row=r, column=8).value = 0
                ws.cell(row=r, column=9).value = f"=(F{r}+H{r}*F{r})*G{r}"
                ws.cell(row=r, column=10).value = item["unidade"]
                ws.cell(row=r, column=11).value = item["custo_unit"] if item["custo_unit"] > 0 else 0
                ws.cell(row=r, column=12).value = f"=I{r}*K{r}"

                # Apply data row style
                for col in range(1, 13):
                    cell = ws.cell(row=r, column=col)
                    style = data_row_style[col]
                    cell.font = copy(style["font"])
                    if col == 12:
                        cell.fill = copy(style["fill"])
                    cell.alignment = copy(style["alignment"])
                    cell.border = copy(style["border"])
                    if col in (6, 7, 8, 9, 10):
                        cell.number_format = NUM_FMT
                    elif col == 11:
                        cell.number_format = CURR_FMT
                    elif col == 12:
                        cell.number_format = NUM_FMT

                # Pivot area: torre label
                ws.cell(row=r, column=14).value = torre_label

                current_row += 1

            max_data_row = current_row - 1

            # Set the L formula for the header row (sumifs)
            ws.cell(row=pav_header_row, column=12).value = (
                f"=IF(SUM(F{pav_header_row}:K{pav_header_row})=0,"
                f"SUMIFS($L{pav_header_row+1}:$L$9999,"
                f"$B{pav_header_row+1}:$B$9999,$B{pav_header_row},"
                f"$A{pav_header_row+1}:$A$9999,$A{pav_header_row}),"
                f"I{pav_header_row}*K{pav_header_row})"
            )

    # --- PIVOT TABLE (cols N-Y) ---
    # Build pivot summary
    # Re-read the data to compute costs per pavimento per torre per EAP
    # We'll create a simpler pivot: one row per unique subgrupo per torre
    pivot_row = 7
    pav_col_map = {
        "TÉRREO": 16, "GARAGEM 01": 17, "GARAGEM 02": 18,
        "GARAGEM 03": 19, "GARAGEM 04": 20, "GARAGEM 05": 21,
        "LAZER": 22, "PAVIMENTO TIPO": 23, "CASA DE MÁQUINAS": 24,
    }

    # Collect all unique subgrupos across all towers/pavimentos
    all_subgrupos = set()
    for torre in ["A", "B"]:
        for pav in PAVIMENTO_ORDER:
            for item in tower_data.get(torre, {}).get(pav, []):
                all_subgrupos.add(item["subgrupo"])

    ordered_subgrupos = [s for s in SUBGRUPO_ORDER if s in all_subgrupos]
    # Add any not in the order
    for s in sorted(all_subgrupos):
        if s not in ordered_subgrupos:
            ordered_subgrupos.append(s)

    for torre in ["A", "B"]:
        torre_label = f"Torre {torre}"
        for subgrupo in ordered_subgrupos:
            r = pivot_row
            ws.cell(row=r, column=14).value = torre_label
            ws.cell(row=r, column=15).value = subgrupo

            total = 0
            for pav in PAVIMENTO_ORDER:
                col = pav_col_map[pav]
                pav_cost = 0
                for item in tower_data.get(torre, {}).get(pav, []):
                    if item["subgrupo"] == subgrupo:
                        qty = item["qtd"]
                        rep = item["rep"]
                        pu = item["custo_unit"] if item["custo_unit"] > 0 else 0
                        pav_cost += qty * rep * pu
                if pav_cost > 0:
                    ws.cell(row=r, column=col).value = round(pav_cost, 2)
                    ws.cell(row=r, column=col).number_format = NUM_FMT
                total += pav_cost

            ws.cell(row=r, column=25).value = round(total, 2)
            ws.cell(row=r, column=25).number_format = NUM_FMT

            # Style pivot cells
            for col in range(14, 26):
                cell = ws.cell(row=r, column=col)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            pivot_row += 1

    # --- COLUMN WIDTHS ---
    col_widths = {
        "A": 12, "B": 20, "C": 35, "D": 30, "E": 50,
        "F": 12, "G": 12, "H": 10, "I": 15, "J": 10,
        "K": 15, "L": 15, "N": 12, "O": 30,
        "P": 12, "Q": 14, "R": 14, "S": 14, "T": 14,
        "U": 14, "V": 12, "W": 16, "X": 18, "Y": 12,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save
    wb.save(R02_PATH)
    print(f"R02 saved to: {R02_PATH}")

    # Summary
    total_items = 0
    for torre in ["A", "B"]:
        for pav in PAVIMENTO_ORDER:
            items = tower_data.get(torre, {}).get(pav, [])
            total_items += len(items)
            if items:
                print(f"  Torre {torre} / {pav}: {len(items)} items")

    print(f"\nTotal items: {total_items}")


if __name__ == "__main__":
    generate_r02()
