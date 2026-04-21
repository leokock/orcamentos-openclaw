"""
Gera xlsx de Estacas Electra r01 — linhas limpas (sem template Gessele) pra Leo colar
na aba Estacas do master CTN-TZN_ELT - Orçamento Executivo_R00 .xlsx

Estrutura igual à aba master, 2 linhas agregadas por diâmetro:
- Linha 1: Ø0.50m, 17 un, 25m, C40 (da LIBERTÉ)
- Linha 2: Ø0.60m, 406 un, 25m, C40
Aço distribuído proporcionalmente ao comprimento total.

Autor: Claude (copiloto Electra, 2026-04-20)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
           r"\_Executivo_IA\thozen-electra\disciplinas\estrutura\estacas-electra-r01.xlsx")

# Dados consolidados (LIBERTÉ 1203-2025-R0 + CONTROLE-REV projetista 1203)
DADOS = [
    # (ø_m, qtd, L_med, L_total, aco_6_3, aco_16)
    (0.50, 17,  25, 425,    248,   1666),   # 4.02% do comp → 4.02% do aço
    (0.60, 406, 25, 10150,  5927, 39790),   # 95.98% do comp → 95.98% do aço
]
# Total esperado: 423 estacas, 10575 m, 6175 kg ø6.3, 41456 kg ø16 = 47.631 kg total

# Estilos
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
TOTAL_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estacas r01"

# Meta
ws.cell(1, 1, "Projeto").font = Font(bold=True)
ws.cell(1, 2, "Thozen Electra Towers")
ws.cell(2, 1, "Empresa").font = Font(bold=True)
ws.cell(2, 2, "Thozen")
ws.cell(3, 1, "Revisão").font = Font(bold=True)
ws.cell(3, 2, "R01")
ws.cell(1, 4, "Fonte:")
ws.cell(1, 5, "LIBERTÉ Proposta 1203-2025-R0 + CR projetista 1203 (bloco A + B)")

# Headers (row 5-6) — espelhando aba Estacas do master
headers_row5 = ["Item", "Item", "Projetos", "Descrição", "Ø peça",
                "Quantidade", "C. Apoio", "Comprimento total",
                "Escavação", "Bota-fora", "Concreto", "Tipo",
                "Aço ø6.3", "Aço ø16", "Taxa aço kg/m³"]
headers_row6 = ["", "", "", "", "m", "und.", "m", "m",
                "m³", "m³", "m³", "", "kg", "kg", ""]

for col_idx, (h5, h6) in enumerate(zip(headers_row5, headers_row6), 1):
    c5 = ws.cell(5, col_idx, h5)
    c6 = ws.cell(6, col_idx, h6)
    c5.fill = HEADER_FILL
    c5.font = HEADER_FONT
    c5.alignment = CENTER
    c5.border = BORDER
    c6.fill = HEADER_FILL
    c6.font = HEADER_FONT
    c6.alignment = CENTER
    c6.border = BORDER

ws.row_dimensions[5].height = 22
ws.row_dimensions[6].height = 18

# Row 7: grupo
ws.cell(7, 1, "Empreendimento Electra").font = SUB_FONT
ws.cell(7, 1).fill = SUB_FILL
ws.cell(7, 2, "Estaca Hélice Contínua").font = SUB_FONT
ws.cell(7, 2).fill = SUB_FILL

# Data rows (8+) — com fórmulas iguais à aba master
for i, (diam, qtd, L_med, L_total, aco_6_3, aco_16) in enumerate(DADOS):
    r = 8 + i
    ws.cell(r, 1, f"=A7")  # herda "Empreendimento Electra"
    ws.cell(r, 2, f"=B7")  # herda "Estaca Hélice Contínua"
    ws.cell(r, 3, "LIBERTÉ 1203-2025-R0")
    ws.cell(r, 4, f'=A{r}&" Ø"&E{r}*100&"cm"')  # descricao auto
    ws.cell(r, 5, diam)
    ws.cell(r, 6, qtd)
    ws.cell(r, 7, L_med)  # C. Apoio = L médio por estaca
    ws.cell(r, 8, f'=IF(F{r}="","",F{r}*G{r})')  # Comp total
    ws.cell(r, 9, f'=IF(F{r}="","",PI()*((E{r}/2)^2)*H{r})')  # Escavação
    ws.cell(r, 10, f'=IF(F{r}="","",I{r}*1.3)')  # Bota-fora
    ws.cell(r, 11, f'=IF(F{r}="","",I{r})')  # Concreto = Escavação
    ws.cell(r, 12, "C40")
    ws.cell(r, 13, aco_6_3)
    ws.cell(r, 14, aco_16)
    ws.cell(r, 15, f'=IFERROR((M{r}+N{r})/K{r},"")')

    # Formatação
    for col in range(1, 16):
        c = ws.cell(r, col)
        c.border = BORDER
        if col in (5,):
            c.number_format = "0.00"
        elif col in (6,):
            c.number_format = "#,##0"
        elif col in (7, 8):
            c.number_format = "#,##0.0"
        elif col in (9, 10, 11):
            c.number_format = "#,##0.00"
        elif col in (13, 14):
            c.number_format = "#,##0"
        elif col == 15:
            c.number_format = "0.00"
        c.alignment = RIGHT if col >= 5 else LEFT

# TOTAL row
r = 8 + len(DADOS)
ws.cell(r, 1, "TOTAL EMPREENDIMENTO")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.cell(r, 5, "")
ws.cell(r, 6, f"=SUM(F8:F{r-1})")
ws.cell(r, 7, "")
ws.cell(r, 8, f"=SUM(H8:H{r-1})")
ws.cell(r, 9, f"=SUM(I8:I{r-1})")
ws.cell(r, 10, f"=SUM(J8:J{r-1})")
ws.cell(r, 11, f"=SUM(K8:K{r-1})")
ws.cell(r, 12, "")
ws.cell(r, 13, f"=SUM(M8:M{r-1})")
ws.cell(r, 14, f"=SUM(N8:N{r-1})")
ws.cell(r, 15, f'=IFERROR((M{r}+N{r})/K{r},"")')

for col in range(1, 16):
    c = ws.cell(r, col)
    c.fill = TOTAL_FILL
    c.font = TOTAL_FONT
    c.border = BORDER
    if col == 6:
        c.number_format = "#,##0"
    elif col == 8:
        c.number_format = "#,##0.0"
    elif col in (9, 10, 11):
        c.number_format = "#,##0.00"
    elif col in (13, 14):
        c.number_format = "#,##0"
    elif col == 15:
        c.number_format = "0.00"

# Obs (abaixo)
r += 3
obs = [
    "OBSERVAÇÕES:",
    "",
    f"• Fonte primária: Proposta Comercial LIBERTÉ 1203-2025-R0 (Estaca Hélice Contínua Monitorada)",
    f"• Cross-check: CONTROLE-REV projetista 1203 (soma Bloco A + B = aço 47.631 kg, bate 100% com LIBERTÉ)",
    f"• C. Apoio = comprimento médio por estaca (25m) — NÃO é cota topográfica",
    f"• Concreto = Escavação (simplificação, sem desconto de armação, margem <1%)",
    f"• Sobreconsumo: Liberté prevê +20% (vol efetivo 3.544 m³). Escavação calculada é teórica (2.953 m³).",
    f"  Se precisar registrar sobreconsumo, adicionar linha 'Sobreconsumo estacas +20%' com vol = 591 m³",
    f"• Bitolas reais do projeto: ø6.3mm (6.175 kg) + ø16mm (41.456 kg) — NÃO ø20 (coluna N renomeada)",
    f"• Distribuição aço por ø = proporcional ao comprimento total (4.02% ø500 + 95.98% ø600)",
    f"• Concreto C40 (fck = 40 MPa) conforme Memorial estrutural 1203",
    f"• Geotecnia: Rua Rubens Alves esq. Rua Canoinhas e Rua Wilson Belber, Balneário Perequê, Porto Belo/SC",
    "",
    "COMO COLAR NO MASTER:",
    "1. Abrir CTN-TZN_ELT - Orçamento Executivo_R00 .xlsx aba Estacas",
    "2. DELETAR linha 7 (dados do template Gessele — 198×ø60cm de Gessele Elisabeth, não Electra)",
    "3. Ajustar row 5 coluna N: renomear header de '20' para '16'",
    "4. Copiar rows 7-10 deste xlsx (grupo + 2 linhas + TOTAL)",
    "5. Colar a partir da row 6 do master (mantém formatação do master via Colar Especial → Valores)",
    "6. Fórmulas vão recalcular automaticamente",
]
for line in obs:
    ws.cell(r, 1, line)
    if line == "OBSERVAÇÕES:" or line == "COMO COLAR NO MASTER:":
        ws.cell(r, 1).font = Font(bold=True)
    r += 1

# Widths
widths = [22, 22, 28, 30, 8, 10, 10, 14, 12, 12, 12, 8, 12, 12, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"xlsx salvo: {OUT}")

# Validação final
print()
print("=== VALIDAÇÃO ===")
total_qtd = sum(d[1] for d in DADOS)
total_L = sum(d[1] * d[2] for d in DADOS)
import math
total_vol = sum(math.pi * (d[0]/2)**2 * d[1] * d[2] for d in DADOS)
total_aco = sum(d[4] + d[5] for d in DADOS)
print(f"Total estacas: {total_qtd} un (esperado 423) {'✓' if total_qtd == 423 else '✗'}")
print(f"Total comprimento: {total_L} m (esperado 10.575) {'✓' if total_L == 10575 else '✗'}")
print(f"Volume teórico: {total_vol:.2f} m³ (esperado ~2.953) {'✓' if abs(total_vol - 2953.3) < 5 else '✗'}")
print(f"Aço total: {total_aco} kg (esperado 47.631) {'✓' if total_aco == 47631 else '✗'}")
print(f"Taxa aço: {total_aco/total_vol:.2f} kg/m³ (típico hélice: 15-25)")
