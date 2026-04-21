#!/usr/bin/env python3
"""
Adicionar aba 'Torre A - EAP Memorial' na planilha existente
"""
import os, sys, requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv('.env.sensitive')

SUPABASE_URL = os.getenv('MEMORIAL_SUPABASE_URL')
ANON_KEY = os.getenv('MEMORIAL_SUPABASE_ANON')
EMAIL = os.getenv('MEMORIAL_EMAIL')
PASSWORD = os.getenv('MEMORIAL_PASSWORD')
BUDGET_ID = "3dc996e6-7bdd-4523-80c4-bb391cf7060d"

print("🔐 Autenticando no Memorial...", flush=True)
auth = requests.post(
    f"{SUPABASE_URL}/auth/v1/token?grant_type=password",
    headers={"apikey": ANON_KEY, "Content-Type": "application/json"},
    json={"email": EMAIL, "password": PASSWORD},
    timeout=10
).json()

token = auth['access_token']
headers = {'apikey': ANON_KEY, 'Authorization': f'Bearer {token}'}
print("✅ Autenticado\n", flush=True)

# Buscar EAP do Memorial
print("📋 Buscando EAP do Memorial...", flush=True)
items_response = requests.get(
    f"{SUPABASE_URL}/rest/v1/budget_items?select=*&budget_id=eq.{BUDGET_ID}&order=code,sequence&limit=1000",
    headers=headers,
    timeout=30
)
items = items_response.json()
print(f"✅ {len(items)} itens encontrados\n", flush=True)

# Carregar planilha existente
print("📂 Carregando planilha existente...", flush=True)
wb = load_workbook('executivo/thozen-electra/CTN-TZN_ELT_Orcamento_Executivo_R00.xlsx')
print(f"✅ {len(wb.sheetnames)} abas existentes\n", flush=True)

# Remover aba se já existir
if "Torre A - Memorial" in wb.sheetnames:
    del wb["Torre A - Memorial"]
    print("🗑️ Aba antiga removida\n", flush=True)

# Criar nova aba
print("📄 Criando aba 'Torre A - Memorial'...", flush=True)
ws = wb.create_sheet("Torre A - Memorial")

# Header style
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

headers = [
    "Código Memorial",
    "Descrição",
    "Level",
    "UN",
    "QTD",
    "Preço Unit. (R$)",
    "Total (R$)",
    "CPU Vinculada",
    "Parent ID",
    "Sequence"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(1, col, header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Preencher dados
print(f"📊 Adicionando {len(items)} itens...", flush=True)

row = 2
for item in items:
    ws[f'A{row}'] = item['code']
    ws[f'B{row}'] = item['name'] or item.get('description', '')
    ws[f'C{row}'] = f"N{item['level']}"
    ws[f'D{row}'] = item.get('unit', '')
    ws[f'E{row}'] = item.get('quantity', '')
    ws[f'F{row}'] = item.get('unit_price', '')
    ws[f'G{row}'] = item.get('total_price', 0)
    ws[f'H{row}'] = "Sim" if item.get('composicao_id') else ""
    ws[f'I{row}'] = str(item.get('parent_id', ''))[:8] + "..." if item.get('parent_id') else ""
    ws[f'J{row}'] = item.get('sequence', '')
    
    # Indent por level
    indent = (item['level'] - 1) * 2
    ws[f'B{row}'].alignment = Alignment(indent=indent, wrap_text=True)
    
    # Bold nos níveis superiores
    if item['level'] < 4:
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
    
    # Highlight subetapas com CPU
    if item.get('composicao_id'):
        ws[f'H{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    row += 1

# Adicionar totais
total_row = row + 1
ws[f'F{total_row}'] = "TOTAL:"
ws[f'F{total_row}'].font = Font(bold=True)
ws[f'G{total_row}'] = f"=SUM(G2:G{row-1})"
ws[f'G{total_row}'].font = Font(bold=True)
ws[f'G{total_row}'].number_format = '#,##0.00'

# Ajustar larguras
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 10

# Congelar painéis (linha 1)
ws.freeze_panes = 'A2'

# Salvar
output_path = "output/CTN-TZN_ELT_Orcamento_Executivo_R00_Atualizado.xlsx"
wb.save(output_path)

print(f"\n{'='*70}")
print(f"✅ Planilha atualizada com sucesso!")
print(f"{'='*70}")
print(f"📁 {output_path}")
print(f"\n📊 Nova aba 'Torre A - Memorial' adicionada com {len(items)} itens")
print(f"{'='*70}")
