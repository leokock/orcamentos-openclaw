#!/usr/bin/env python3
"""Gera abas HIDROSSANITÁRIO v3 (lista geral 8 subetapas) e v4 (por torre) no r02."""
import sys, io, re, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento\_Executivo_IA\thozen-electra"
INPUT = f"{BASE}/disciplinas/hidraulico/hidraulico-electra-lm348-r02.xlsx"
OUTPUT = f"{BASE}/disciplinas/hidraulico/hidraulico-electra-lm348-r03.xlsx"
AC = 36092

# ============================================================
# SUBETAPAS DO ORÇAMENTO
# ============================================================
SUBETAPAS = [
    'Instalações de água fria',
    'Instalações de água quente',
    'Instalações de esgoto e pluviais',
    'Cisterna',
    'Hidrômetros',
    'Reservatórios',
    'Sistema de bombas',
    'Mão de obra',
]

# Cores por subetapa (matching screenshot)
SUBETAPA_COLORS = {
    'Instalações de água fria':        'FF2F5496',
    'Instalações de água quente':      'FFED7D31',
    'Instalações de esgoto e pluviais':'FF548235',
    'Cisterna':                        'FFC00000',
    'Hidrômetros':                     'FF7030A0',
    'Reservatórios':                   'FF00B0F0',
    'Sistema de bombas':               'FFFFC000',
    'Mão de obra':                     'FF808080',
}

def norm(s):
    s = str(s).lower().strip()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    for ch in ['°','º','ø','Ø','"','"','"',"'",'(',')',',','.','-','/']:
        s = s.replace(ch, ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def classify_subetapa(sub_v2, desc):
    """Mapeia subgrupo v2 + descrição → subetapa do orçamento."""
    nd = norm(desc)
    sv = str(sub_v2 or '')

    # Hidrômetros
    if 'hidrometro' in nd or 'hidrômetro' in nd.replace(' ',''):
        return 'Hidrômetros'

    # Reservatórios
    if "caixa d agua" in nd or 'polietileno' in nd and ('7500' in nd or 'reservat' in nd):
        return 'Reservatórios'

    # Sistema de bombas (skids, válvulas, bombas, ventosas, filtro T, chave fluxo)
    if any(x in nd for x in ['skid', 'pressurizador', 'motobomba',
                              'valvula redutora', 'valvula controladora',
                              'valvula de alivio', 'valvula de retencao', 'valvula pe',
                              'ventosa', 'filtro tipo', 'filtro t compact',
                              'chave de fluxo', 'chave fluxo']):
        return 'Sistema de bombas'

    # Cisterna (= ETE)
    if sv == 'ETE / TRATAMENTO':
        return 'Cisterna'

    # Esgoto e pluviais
    if 'ESGOTO' in sv:
        return 'Instalações de esgoto e pluviais'
    if any(x in nd for x in ['esgoto', 'serie normal', 'serie reforcada',
                              'caixa sifonada', 'ralo seco', 'terminal de ventilacao',
                              'anel de vedacao', 'anel vedacao',
                              'filtro pluvial', 'juncao', 'te de inspecao']):
        return 'Instalações de esgoto e pluviais'

    # Água quente
    if 'QUENTE' in sv or 'CPVC' in sv or 'PPR' in sv:
        return 'Instalações de água quente'
    if any(x in nd for x in ['cpvc', 'flowguard', 'super cpvc', 'ppr pn25', 'ppr pn 25']):
        return 'Instalações de água quente'

    # Água fria (default para tubulações/conexões PVC soldável)
    if 'FRIA' in sv or 'TUBULAÇÕES ÁGUA' in sv or 'CONEXÕES ÁGUA' in sv:
        return 'Instalações de água fria'
    if 'ACESSÓRIOS' in sv:
        # Adaptadores longos com flanges → AF
        if 'adaptador longo' in nd:
            return 'Instalações de água fria'
        # Anéis de vedação, caixas sifonadas, ralos → esgoto
        return 'Instalações de esgoto e pluviais'

    if 'EQUIPAMENTOS' in sv:
        return 'Sistema de bombas'

    return 'Instalações de água fria'


def classify_torre(pav, torre):
    """Classifica em Embasamento/Torre A/Torre B baseado no pavimento."""
    p = str(pav or '').upper()
    t = str(torre or '').upper()
    if 'TIPO' in p:
        if t == 'A': return 'TORRE A'
        if t == 'B': return 'TORRE B'
        return 'AMBAS_TIPO'
    return 'EMBASAMENTO'


# ============================================================
# ESTILOS
# ============================================================
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=10)
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_font = Font(name="Calibri", size=10, bold=True)
title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="1851D6", end_color="1851D6", fill_type="solid")
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
qty_fmt = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
rs_fmt = '_-"R$"\\ * #,##0.00_-;\\-"R$"\\ * #,##0.00_-;_-"R$"\\ * "-"??_-;_-@_-'

HEADERS_V3 = ['Subetapa', 'Descrição', 'Dimensões', 'Marca', 'Qtd', 'Unid', 'Custo unit.', 'Custo Total', 'Fonte PU']
HEADERS_V4 = ['Etapa', 'Subetapa', 'Descrição', 'Dimensões', 'Marca', 'Qtd', 'Unid', 'Custo unit.', 'Custo Total', 'Fonte PU']
WIDTHS_V3 = [34, 60, 18, 18, 14, 8, 14, 16, 14]
WIDTHS_V4 = [22, 34, 55, 18, 18, 14, 8, 14, 16, 14]


def write_metadata(ws, title, ncols):
    ws.cell(1, 2, 'Projeto'); ws.cell(1, 3, 'Electra Towers')
    ws.cell(2, 2, 'Empresa'); ws.cell(2, 3, 'Thozen')
    ws.cell(3, 2, 'Origem'); ws.cell(3, 3, 'CTN-TZN_HID-LM348-R01')
    ws.cell(4, 2, 'Revisão'); ws.cell(4, 3, 'R03 — 8 subetapas orçamento — 17/abr/2026')
    ws.cell(5, 1, title)
    ws.cell(5, 1).font = title_font
    ws.cell(5, 1).fill = title_fill
    for c in range(2, ncols + 1):
        ws.cell(5, c).fill = title_fill


def write_headers(ws, headers, row=6):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row, ci, h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.load_workbook(INPUT)

    # --- Read v2 data items ---
    ws_v2 = wb['HIDROSSANITÁRIO v2']
    items = []
    for r in range(7, ws_v2.max_row + 1):
        sub = ws_v2.cell(r, 2).value
        desc = ws_v2.cell(r, 3).value
        if not desc or str(sub or '').startswith('SUBTOTAL') or 'TOTAL GERAL' in str(sub or ''):
            continue
        # Skip section headers (no desc but has sub)
        grupo = ws_v2.cell(r, 1).value
        if grupo and not desc:
            continue
        items.append({
            'sub_v2': sub, 'desc': desc,
            'dim': ws_v2.cell(r, 4).value or '',
            'marca': ws_v2.cell(r, 5).value or '',
            'qty': ws_v2.cell(r, 6).value or 0,
            'unid': ws_v2.cell(r, 7).value or 'un',
            'pu': ws_v2.cell(r, 8).value,
            'fonte': ws_v2.cell(r, 10).value or '',
        })

    # Classify each item
    for item in items:
        item['subetapa'] = classify_subetapa(item['sub_v2'], item['desc'])

    print(f'Total items: {len(items)}')
    for se in SUBETAPAS:
        cnt = sum(1 for i in items if i['subetapa'] == se)
        print(f'  {se}: {cnt}')

    # --- Read Por Pavimento for v4 distribution ---
    ws_pav = wb['Por Pavimento']
    pav_items = []
    for r in range(8, ws_pav.max_row + 1):
        pav = ws_pav.cell(r, 1).value
        torre = ws_pav.cell(r, 2).value
        slug = ws_pav.cell(r, 3).value
        if not pav or not slug:
            continue
        num = ws_pav.cell(r, 4).value
        qty = ws_pav.cell(r, 6).value
        desc_pav = ws_pav.cell(r, 8).value or ''
        dim = ws_pav.cell(r, 9).value or ''
        pav_items.append({
            'pav': pav, 'torre': torre, 'slug': slug, 'num': num,
            'qty': qty or 0, 'desc': desc_pav, 'dim': dim,
        })

    # Build lookup: (slug, num) → v2 item data (pu, fonte, unid, subetapa, etc.)
    # Match via desc similarity since slug/num maps to flat
    ws_flat = wb['Flat por PDF']
    flat_lookup = {}
    for r in range(4, ws_flat.max_row + 1):
        slug = ws_flat.cell(r, 1).value
        num = ws_flat.cell(r, 5).value
        desc_flat = ws_flat.cell(r, 9).value or ''
        dim = ws_flat.cell(r, 10).value or ''
        marca = ws_flat.cell(r, 11).value or ''
        if slug and num is not None:
            flat_lookup[(slug, num)] = {'desc': desc_flat, 'dim': dim, 'marca': marca}

    # Match flat items to v2 items by description
    def find_v2_item(desc_flat, dim_flat):
        nd = norm(desc_flat)
        best = None
        for item in items:
            ni = norm(item['desc'])
            if ni == nd or nd in ni or ni in nd:
                if norm(str(item['dim'])) == norm(str(dim_flat)) or not dim_flat:
                    return item
                if best is None:
                    best = item
        return best

    # Aggregate pav_items into (torre_zone, subetapa, desc) → qty
    v4_data = {}
    for pi in pav_items:
        zone = classify_torre(pi['pav'], pi['torre'])
        fl = flat_lookup.get((pi['slug'], pi['num']))
        if not fl:
            continue
        v2_item = find_v2_item(fl['desc'], fl['dim'])
        if not v2_item:
            continue
        subetapa = v2_item['subetapa']
        key = (zone, subetapa, v2_item['desc'], v2_item['dim'])
        if key not in v4_data:
            v4_data[key] = {
                'qty': 0, 'marca': v2_item['marca'], 'unid': v2_item['unid'],
                'pu': v2_item['pu'], 'fonte': v2_item['fonte'],
                'is_tube': v2_item['unid'] == 'm',
            }
        # For tubes, convert bar count to meters
        mult = 1
        if v2_item['unid'] == 'm':
            fdesc = fl['desc']
            if '6m' in fdesc: mult = 6
            elif '3m' in fdesc: mult = 3
        v4_data[key]['qty'] += pi['qty'] * mult

    # Handle AMBAS_TIPO: split 50/50
    keys_to_split = [k for k in v4_data if k[0] == 'AMBAS_TIPO']
    for k in keys_to_split:
        data = v4_data.pop(k)
        for torre in ['TORRE A', 'TORRE B']:
            new_key = (torre,) + k[1:]
            if new_key in v4_data:
                v4_data[new_key]['qty'] += data['qty'] / 2
            else:
                v4_data[new_key] = dict(data)
                v4_data[new_key]['qty'] = data['qty'] / 2

    # ============================================================
    # V3: Lista geral com 8 subetapas
    # ============================================================
    ws3 = wb.create_sheet('HIDROSSANITÁRIO v3')
    write_metadata(ws3, 'Hidrossanitário v3 — 8 Subetapas (Lista Geral)', len(HEADERS_V3))
    write_headers(ws3, HEADERS_V3)

    row = 7
    subtotal_rows_v3 = []
    for se in SUBETAPAS:
        se_items = [i for i in items if i['subetapa'] == se]
        if not se_items and se != 'Mão de obra':
            continue

        # Section header
        color = SUBETAPA_COLORS.get(se, 'FF2F5496')
        sec_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        sec_font_w = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws3.cell(row, 1, se)
        ws3.cell(row, 2, f'{len(se_items)} itens')
        for ci in range(1, len(HEADERS_V3) + 1):
            ws3.cell(row, ci).fill = sec_fill
            ws3.cell(row, ci).font = sec_font_w
        row += 1
        data_start = row

        if se == 'Mão de obra':
            ws3.cell(row, 1, 'Mão de obra')
            ws3.cell(row, 2, 'A definir — benchmark Cartesian: R$ 55,00/m² AC')
            for ci in range(1, len(HEADERS_V3) + 1):
                ws3.cell(row, ci).font = Font(name="Calibri", size=10, italic=True)
                ws3.cell(row, ci).border = thin
            row += 1
        else:
            for item in sorted(se_items, key=lambda x: x['desc']):
                ws3.cell(row, 1, se)
                ws3.cell(row, 2, item['desc'])
                ws3.cell(row, 3, item['dim'])
                ws3.cell(row, 4, item['marca'])
                ws3.cell(row, 5, round(item['qty'], 2))
                ws3.cell(row, 5).number_format = qty_fmt
                ws3.cell(row, 6, item['unid'])
                if item['pu'] is not None:
                    ws3.cell(row, 7, round(item['pu'], 2))
                    ws3.cell(row, 7).number_format = rs_fmt
                    ws3.cell(row, 8, f'=E{row}*G{row}')
                    ws3.cell(row, 8).number_format = rs_fmt
                else:
                    ws3.cell(row, 8, 0).number_format = rs_fmt
                ws3.cell(row, 9, item['fonte'])
                for ci in range(1, len(HEADERS_V3) + 1):
                    ws3.cell(row, ci).font = data_font
                    ws3.cell(row, ci).border = thin
                row += 1

        # Subtotal
        ws3.cell(row, 1, f'SUBTOTAL {se}')
        ws3.cell(row, 1).font = sub_font
        ws3.cell(row, 8, f'=SUM(H{data_start}:H{row-1})')
        ws3.cell(row, 8).number_format = rs_fmt
        for ci in range(1, len(HEADERS_V3) + 1):
            ws3.cell(row, ci).fill = sub_fill
            ws3.cell(row, ci).font = sub_font
        subtotal_rows_v3.append(row)
        row += 1

    # Total geral
    row += 1
    ws3.cell(row, 1, 'TOTAL GERAL — CON 05 HIDROSSANITÁRIO')
    ws3.cell(row, 1).font = Font(name="Calibri", size=12, bold=True)
    refs = '+'.join(f'H{r}' for r in subtotal_rows_v3)
    ws3.cell(row, 8, f'={refs}')
    ws3.cell(row, 8).number_format = rs_fmt
    ws3.cell(row, 8).font = Font(name="Calibri", size=12, bold=True)

    set_widths(ws3, WIDTHS_V3)
    ws3.freeze_panes = 'A7'

    print(f'\nv3: {row} rows, {len(subtotal_rows_v3)} subetapas')

    # ============================================================
    # V4: Por torre (Embasamento / Torre A / Torre B) × 8 subetapas
    # ============================================================
    ws4 = wb.create_sheet('HIDROSSANITÁRIO v4')
    write_metadata(ws4, 'Hidrossanitário v4 — Por Torre × 8 Subetapas', len(HEADERS_V4))
    write_headers(ws4, HEADERS_V4)

    ETAPAS = [
        ('EMBASAMENTO', '5.001'),
        ('TORRE A', '5.002'),
        ('TORRE B', '5.003'),
    ]

    row = 7
    subtotal_rows_v4 = []
    etapa_subtotal_rows = []

    for etapa_name, etapa_cod in ETAPAS:
        # Etapa header
        etapa_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        etapa_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
        label = f'{etapa_cod} — SISTEMAS E INST. HIDROSSANITÁRIAS E DRENAGEM — {etapa_name}'
        ws4.cell(row, 1, label)
        for ci in range(1, len(HEADERS_V4) + 1):
            ws4.cell(row, ci).fill = etapa_fill
            ws4.cell(row, ci).font = etapa_font
        row += 1
        etapa_data_start = row

        for idx, se in enumerate(SUBETAPAS):
            # Get items for this etapa + subetapa
            se_items_v4 = []
            for k, v in v4_data.items():
                zone, subetapa, desc, dim = k
                if zone == etapa_name and subetapa == se:
                    se_items_v4.append({
                        'desc': desc, 'dim': dim, 'marca': v['marca'],
                        'qty': v['qty'], 'unid': v['unid'],
                        'pu': v['pu'], 'fonte': v['fonte'],
                    })

            if not se_items_v4 and se != 'Mão de obra':
                continue

            # Subetapa header
            color = SUBETAPA_COLORS.get(se, 'FF2F5496')
            sec_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sec_font_w = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cod_sub = f'{etapa_cod}.{idx+1:03d}'
            ws4.cell(row, 1, etapa_name)
            ws4.cell(row, 2, f'{cod_sub} — {se}')
            ws4.cell(row, 3, f'{len(se_items_v4)} itens')
            for ci in range(1, len(HEADERS_V4) + 1):
                ws4.cell(row, ci).fill = sec_fill
                ws4.cell(row, ci).font = sec_font_w
            row += 1
            sub_data_start = row

            if se == 'Mão de obra':
                ws4.cell(row, 1, etapa_name)
                ws4.cell(row, 2, 'Mão de obra')
                ws4.cell(row, 3, 'A definir')
                for ci in range(1, len(HEADERS_V4) + 1):
                    ws4.cell(row, ci).font = Font(name="Calibri", size=10, italic=True)
                    ws4.cell(row, ci).border = thin
                row += 1
            else:
                for item in sorted(se_items_v4, key=lambda x: x['desc']):
                    ws4.cell(row, 1, etapa_name)
                    ws4.cell(row, 2, se)
                    ws4.cell(row, 3, item['desc'])
                    ws4.cell(row, 4, item['dim'])
                    ws4.cell(row, 5, item['marca'])
                    ws4.cell(row, 6, round(item['qty'], 2))
                    ws4.cell(row, 6).number_format = qty_fmt
                    ws4.cell(row, 7, item['unid'])
                    if item['pu'] is not None:
                        ws4.cell(row, 8, round(item['pu'], 2))
                        ws4.cell(row, 8).number_format = rs_fmt
                        ws4.cell(row, 9, f'=F{row}*H{row}')
                        ws4.cell(row, 9).number_format = rs_fmt
                    else:
                        ws4.cell(row, 9, 0).number_format = rs_fmt
                    ws4.cell(row, 10, item['fonte'])
                    for ci in range(1, len(HEADERS_V4) + 1):
                        ws4.cell(row, ci).font = data_font
                        ws4.cell(row, ci).border = thin
                    row += 1

            # Subetapa subtotal
            ws4.cell(row, 2, f'Subtotal {se}')
            ws4.cell(row, 2).font = sub_font
            ws4.cell(row, 9, f'=SUM(I{sub_data_start}:I{row-1})')
            ws4.cell(row, 9).number_format = rs_fmt
            for ci in range(1, len(HEADERS_V4) + 1):
                ws4.cell(row, ci).fill = sub_fill
                ws4.cell(row, ci).font = sub_font
            subtotal_rows_v4.append(row)
            row += 1

        # Etapa subtotal
        etapa_subs = [r2 for r2 in subtotal_rows_v4 if r2 >= etapa_data_start]
        etapa_refs = '+'.join(f'I{r2}' for r2 in etapa_subs)
        ws4.cell(row, 1, f'TOTAL {etapa_name}')
        ws4.cell(row, 1).font = Font(name="Calibri", size=11, bold=True)
        ws4.cell(row, 9, f'={etapa_refs}')
        ws4.cell(row, 9).number_format = rs_fmt
        ws4.cell(row, 9).font = Font(name="Calibri", size=11, bold=True)
        for ci in range(1, len(HEADERS_V4) + 1):
            ws4.cell(row, ci).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            ws4.cell(row, ci).font = Font(name="Calibri", size=11, bold=True)
        etapa_subtotal_rows.append(row)
        row += 2

    # Grand total
    refs_grand = '+'.join(f'I{r}' for r in etapa_subtotal_rows)
    ws4.cell(row, 1, 'TOTAL GERAL — CON 05 HIDROSSANITÁRIO')
    ws4.cell(row, 1).font = Font(name="Calibri", size=12, bold=True)
    ws4.cell(row, 9, f'={refs_grand}')
    ws4.cell(row, 9).number_format = rs_fmt
    ws4.cell(row, 9).font = Font(name="Calibri", size=12, bold=True)

    set_widths(ws4, WIDTHS_V4)
    ws4.freeze_panes = 'A7'

    print(f'v4: {row} rows, {len(etapa_subtotal_rows)} etapas')

    # --- v4 distribution summary ---
    for etapa_name, _ in ETAPAS:
        items_in = sum(1 for k in v4_data if k[0] == etapa_name)
        print(f'  {etapa_name}: {items_in} itens')

    wb.save(OUTPUT)
    print(f'\nArquivo salvo: {OUTPUT}')

if __name__ == "__main__":
    main()
