"""
Extrai todas as IfcFooting (blocos + vigas baldrame) do IFC R26 do Electra
com bounding box 3D, classificação por tipologia e dimensões individuais.

Gera xlsx detalhado pra colar na aba "Fund. Rasa | Contenção" do master.
Output: disciplinas/estrutura/fund-rasa-electra-r01-detalhado.xlsx

Autor: Claude (copiloto Electra, 2026-04-20)
"""
import re
import json
from pathlib import Path
from collections import defaultdict
import numpy as np
import ifcopenshell
import ifcopenshell.geom
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IFC = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
           r"\_Projetos_IA\thozen-electra\projetos\01 ESTRUTURA\IFC"
           r"\1203 - THOZEN - RUBENS ALVES - BLOCOS+RAMPAS DE ACESSO - R26.ifc")
OUT = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
           r"\_Executivo_IA\thozen-electra\disciplinas\estrutura"
           r"\fund-rasa-electra-r01-detalhado.xlsx")
JSON_OUT = Path(r"G:\Drives compartilhados\03 CTN Projetos\2. Projetos em Andamento"
                r"\_Executivo_IA\thozen-electra\dados\ifc-r26-fund-rasa.json")


def classify_tipology(name: str) -> str:
    clean = name.replace("IfcFooting - IfcFooting - ", "").replace("IfcFooting - ", "")
    if re.match(r"^Viga", clean, re.I):
        return "Viga baldrame"
    elif re.match(r"^Pra", clean, re.I):
        return "Bloco quadr. retangular (Pra)"
    elif re.match(r"^P\d", clean, re.I):
        return "Bloco pilar"
    elif re.match(r"^B\d", clean, re.I):
        return "Bloco coroamento"
    elif re.match(r"^Sap", clean, re.I):
        return "Sapata"
    elif re.match(r"^Hex", clean, re.I):
        return "Bloco hexagonal"
    return "Outro"


def parse_section(name: str) -> str | None:
    m = re.search(r"\(([0-9x\s]+)\)", name)
    return m.group(1).strip() if m else None


def main():
    print(f"Abrindo IFC ({IFC.stat().st_size // 1024 // 1024} MB)...")
    f = ifcopenshell.open(str(IFC))
    settings = ifcopenshell.geom.settings()
    settings.set(settings.USE_WORLD_COORDS, True)

    footings = f.by_type("IfcFooting")
    print(f"Total IfcFooting: {len(footings)}")

    elements = []
    errors = 0
    for i, fo in enumerate(footings, 1):
        if i % 20 == 0:
            print(f"  {i}/{len(footings)}...")
        name_raw = fo.Name or "(sem nome)"
        name = name_raw.replace("IfcFooting - IfcFooting - ", "").replace("IfcFooting - ", "")
        tipo = classify_tipology(name_raw)
        secao_nominal = parse_section(name) or "-"

        try:
            shape = ifcopenshell.geom.create_shape(settings, fo)
            verts = np.array(shape.geometry.verts).reshape(-1, 3)
            dx = float(verts[:, 0].max() - verts[:, 0].min())
            dy = float(verts[:, 1].max() - verts[:, 1].min())
            dz = float(verts[:, 2].max() - verts[:, 2].min())
            cx = float((verts[:, 0].max() + verts[:, 0].min()) / 2)
            cy = float((verts[:, 1].max() + verts[:, 1].min()) / 2)
            cz = float((verts[:, 2].max() + verts[:, 2].min()) / 2)

            # Ordenar dimensões (menor = largura, maior = comprimento pra vigas)
            dims = sorted([dx, dy, dz], reverse=True)
            comprimento = dims[0]
            altura = dims[1]
            largura = dims[2]

            # Volume aproximado (caixa bounding)
            volume = dx * dy * dz
        except Exception as e:
            errors += 1
            comprimento = altura = largura = volume = None
            cx = cy = cz = None

        elements.append({
            "name": name,
            "tipo": tipo,
            "secao_nominal": secao_nominal,
            "dx": dx if 'dx' in dir() else None,
            "dy": dy if 'dy' in dir() else None,
            "dz": dz if 'dz' in dir() else None,
            "comprimento_m": comprimento,
            "altura_m": altura,
            "largura_m": largura,
            "volume_bbox_m3": volume,
            "centro_x": cx,
            "centro_y": cy,
            "centro_z": cz,
        })

    print(f"Processadas: {len(elements)}, erros: {errors}")

    # Agregar por (tipo, seção)
    grupos = defaultdict(lambda: {"count": 0, "volume_total": 0.0, "comp_total": 0.0, "elementos": []})
    for e in elements:
        key = (e["tipo"], e["secao_nominal"])
        grupos[key]["count"] += 1
        if e["volume_bbox_m3"]:
            grupos[key]["volume_total"] += e["volume_bbox_m3"]
        if e["comprimento_m"]:
            grupos[key]["comp_total"] += e["comprimento_m"]
        grupos[key]["elementos"].append(e["name"])

    # Salvar JSON raw
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps({
        "meta": {
            "ifc": IFC.name,
            "gerado_em": "2026-04-20",
            "n_footings": len(elements),
            "n_errors": errors,
        },
        "elementos": elements,
        "grupos_por_tipo_secao": {f"{t}|{s}": {
            "count": g["count"],
            "volume_total_m3": round(g["volume_total"], 3),
            "comp_total_m": round(g["comp_total"], 3),
            "elementos": g["elementos"],
        } for (t, s), g in grupos.items()},
    }, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"JSON salvo: {JSON_OUT}")

    # ===== Gerar xlsx detalhado =====
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    GROUP_FILL = PatternFill("solid", fgColor="FFE699")
    GROUP_FONT = Font(bold=True, size=11)
    TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
    TOTAL_FONT = Font(bold=True, size=11)
    THIN = Side(style="thin", color="808080")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")

    wb = openpyxl.Workbook()

    # -------- Aba 1: POR ELEMENTO (todos 70) --------
    ws = wb.active
    ws.title = "Por Elemento"

    ws.cell(1, 1, "Fundação Rasa — Detalhe por elemento individual").font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.cell(2, 1, f"Fonte: IFC R26 ({IFC.name}) | {len(elements)} footings extraídos").font = Font(italic=True, size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)

    headers = ["#", "Nome", "Tipo", "Seção nominal", "Larg (m)", "Alt (m)",
               "Comp (m)", "Volume (m³)", "Centro X", "Centro Y", "Centro Z"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(4, col_idx, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    r = 5
    for i, e in enumerate(elements, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, e["name"])
        ws.cell(r, 3, e["tipo"])
        ws.cell(r, 4, e["secao_nominal"])
        ws.cell(r, 5, round(e["largura_m"], 3) if e["largura_m"] else None)
        ws.cell(r, 6, round(e["altura_m"], 3) if e["altura_m"] else None)
        ws.cell(r, 7, round(e["comprimento_m"], 3) if e["comprimento_m"] else None)
        ws.cell(r, 8, round(e["volume_bbox_m3"], 4) if e["volume_bbox_m3"] else None)
        ws.cell(r, 9, round(e["centro_x"], 2) if e["centro_x"] is not None else None)
        ws.cell(r, 10, round(e["centro_y"], 2) if e["centro_y"] is not None else None)
        ws.cell(r, 11, round(e["centro_z"], 2) if e["centro_z"] is not None else None)
        for col in range(1, 12):
            c = ws.cell(r, col)
            c.border = BORDER
            c.alignment = RIGHT if col >= 5 else LEFT
            if col in (5, 6, 7):
                c.number_format = "0.000"
            elif col == 8:
                c.number_format = "0.0000"
            elif col in (9, 10, 11):
                c.number_format = "0.00"
        r += 1

    # Widths Aba 1
    widths = [5, 35, 28, 14, 10, 10, 10, 12, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # -------- Aba 2: AGRUPADO POR TIPO + SEÇÃO --------
    ws2 = wb.create_sheet("Agrupado")
    ws2.cell(1, 1, "Fundação Rasa — Agrupado por tipologia + seção nominal").font = Font(bold=True, size=13)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws2.cell(2, 1, "Este formato é compatível com a aba 'Fund. Rasa | Contenção' do master "
                    "(1 linha por tipologia+seção, fórmulas AUTO-derivadas)").font = Font(italic=True, size=9)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers2 = ["Tipo", "Seção nominal", "Quantidade", "Larg média (m)",
                "Alt média (m)", "Comp total (m)", "Comp médio (m)", "Volume total (m³)"]
    for col_idx, h in enumerate(headers2, 1):
        c = ws2.cell(4, col_idx, h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER

    r = 5
    total_count = 0
    total_vol = 0
    for (tipo, sec), g in sorted(grupos.items()):
        # Calcular larg/alt médias
        grp_els = [e for e in elements if e["tipo"] == tipo and e["secao_nominal"] == sec]
        largs = [e["largura_m"] for e in grp_els if e["largura_m"]]
        alts = [e["altura_m"] for e in grp_els if e["altura_m"]]
        comps = [e["comprimento_m"] for e in grp_els if e["comprimento_m"]]
        larg_media = sum(largs) / len(largs) if largs else None
        alt_media = sum(alts) / len(alts) if alts else None
        comp_total = sum(comps) if comps else 0
        comp_medio = comp_total / len(comps) if comps else None

        ws2.cell(r, 1, tipo)
        ws2.cell(r, 2, sec)
        ws2.cell(r, 3, g["count"])
        ws2.cell(r, 4, round(larg_media, 3) if larg_media else None)
        ws2.cell(r, 5, round(alt_media, 3) if alt_media else None)
        ws2.cell(r, 6, round(comp_total, 2))
        ws2.cell(r, 7, round(comp_medio, 2) if comp_medio else None)
        ws2.cell(r, 8, round(g["volume_total"], 3))
        for col in range(1, 9):
            c = ws2.cell(r, col)
            c.border = BORDER
            c.alignment = RIGHT if col >= 3 else LEFT
            if col == 3:
                c.number_format = "#,##0"
            elif col in (4, 5, 7):
                c.number_format = "0.000"
            elif col == 6:
                c.number_format = "#,##0.0"
            elif col == 8:
                c.number_format = "0.000"
        total_count += g["count"]
        total_vol += g["volume_total"]
        r += 1

    # Total
    ws2.cell(r, 1, "TOTAL")
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws2.cell(r, 3, total_count)
    ws2.cell(r, 8, round(total_vol, 3))
    for col in range(1, 9):
        c = ws2.cell(r, col)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = BORDER
        c.alignment = RIGHT if col >= 3 else LEFT
        if col == 3:
            c.number_format = "#,##0"
        elif col == 8:
            c.number_format = "0.000"
    r += 2

    # Observações
    obs = [
        "OBSERVAÇÕES:",
        "",
        "• Volume é calculado via BOUNDING BOX 3D (aproximação por caixa envolvente)",
        "  - Pode superestimar em elementos com geometria complexa (escalonamento, reentrâncias)",
        "  - Pra concreto real, melhor comparar com o valor projetista: 2.066 m³ (QUANT-A r246)",
        "",
        "• LARG/ALT/COMP são deduzidos do bounding box (menor/meio/maior dimensão)",
        "  - Pra vigas horizontais, geralmente bate com seção nominal do nome",
        "  - Pra blocos cúbicos (Pra 70x70x30), aparece como 0.70m larg, 0.70m alt, 0.30m comp",
        "",
        "• Tipologias Electra identificadas:",
    ]
    # Acrescentar tipologias
    for (tipo, sec), g in sorted(grupos.items()):
        obs.append(f"    {tipo} [{sec}]: {g['count']} elementos")
    obs.extend([
        "",
        "• IfcPile = 0 (estacas NÃO modeladas no IFC — ver scripts/gerar_estacas_r01.py)",
        "",
        "COMO USAR:",
        "1. Aba 'Agrupado' → base pra preencher as linhas da aba Fund. Rasa | Contenção do master",
        "2. Aba 'Por Elemento' → referência detalhada (auditoria, memorial técnico)",
        "3. Comparar volume total BBOX vs projetista (QUANT-A r246 = 2.066 m³):",
        f"   - BBOX IFC R26: {total_vol:.2f} m³",
        f"   - Projetista 1203: 2.066 m³",
        f"   - Diferença: {abs(total_vol - 2066):.2f} m³ ({abs(total_vol - 2066)/2066*100:.1f}%)",
    ])
    for line in obs:
        ws2.cell(r, 1, line)
        if line.startswith("OBSERVAÇÕES") or line.startswith("COMO USAR"):
            ws2.cell(r, 1).font = Font(bold=True)
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1

    # Widths Aba 2
    widths2 = [28, 14, 12, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"xlsx salvo: {OUT}")

    print(f"\n=== RESUMO ===")
    print(f"Total elementos: {total_count}")
    print(f"Volume BBOX total: {total_vol:.2f} m³")
    print(f"Volume projetista: 2.066 m³")
    print(f"Divergência: {abs(total_vol - 2066):.2f} m³ ({abs(total_vol - 2066)/2066*100:.1f}%)")
    print(f"\nAgrupamentos:")
    for (t, s), g in sorted(grupos.items()):
        print(f"  {t} [{s}]: {g['count']} elementos | vol BBOX {g['volume_total']:.2f} m³")


if __name__ == "__main__":
    main()
