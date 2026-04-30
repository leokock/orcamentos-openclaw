#!/usr/bin/env python3.11
"""
Template de Orçamento Paramétrico - Cartesian Engenharia
Gera planilha Excel (.xlsx) com 8 abas, fórmulas nativas Excel,
formatação condicional (semáforos) e data validation (dropdowns).
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from copy import copy
import os

# ── Styles ──────────────────────────────────────────────────────────
THIN = Side(style='thin', color='999999')
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)

SUBHEADER_FILL = PatternFill('solid', fgColor='95A5A6')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10)

INPUT_FILL = PatternFill('solid', fgColor='D6EAF8')
INPUT_FONT = Font(color='000000', size=10)

CALC_FILL = PatternFill('solid', fgColor='E8E8E8')
CALC_FONT = Font(color='000000', size=10)

TITLE_FONT = Font(bold=True, color='2C3E50', size=14)
SECTION_FONT = Font(bold=True, color='2C3E50', size=12)

GREEN_FILL = PatternFill('solid', fgColor='27AE60')
YELLOW_FILL = PatternFill('solid', fgColor='F39C12')
RED_FILL = PatternFill('solid', fgColor='E74C3C')

# Number formats (Brazilian)
FMT_BRL = '#,##0.00'       # R$ with 2 decimals
FMT_INT = '#,##0'           # integers
FMT_PCT = '0.0%'            # percentages
FMT_IDX = '#,##0.000'       # indices (3 decimals)
FMT_TAXA = '#,##0.0'        # taxa aço etc


def style_cell(cell, fill=None, font=None, fmt=None, border=True, align=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = BORDER_THIN
    if align:
        cell.alignment = align


def write_header_row(ws, row, cols, widths=None):
    """Write header row with HEADER style."""
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        style_cell(c, HEADER_FILL, HEADER_FONT,
                   align=Alignment(horizontal='center', vertical='center', wrap_text=True))
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def write_subheader(ws, row, col_start, col_end, text):
    """Write a sub-header spanning columns."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    style_cell(c, SUBHEADER_FILL, SUBHEADER_FONT,
               align=Alignment(horizontal='left', vertical='center'))
    for col in range(col_start + 1, col_end + 1):
        style_cell(ws.cell(row=row, column=col), SUBHEADER_FILL, SUBHEADER_FONT)


def input_cell(ws, row, col, value=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    style_cell(c, INPUT_FILL, INPUT_FONT, fmt)
    return c


def calc_cell(ws, row, col, formula=None, fmt=None):
    c = ws.cell(row=row, column=col, value=formula)
    style_cell(c, CALC_FILL, CALC_FONT, fmt)
    c.protection = openpyxl.styles.Protection(locked=True)
    return c


def label_cell(ws, row, col, value, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    f = Font(bold=bold, color='000000', size=10)
    style_cell(c, font=f, align=Alignment(vertical='center'))
    return c


# ════════════════════════════════════════════════════════════════════
# ABA 1: DADOS_PROJETO
# ════════════════════════════════════════════════════════════════════
def create_dados_projeto(wb):
    ws = wb.active
    ws.title = 'DADOS_PROJETO'
    ws.sheet_properties.tabColor = '2C3E50'

    # Title
    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value='DADOS DO PROJETO')
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical='center')

    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value='Cartesian Engenharia — Orçamento Paramétrico').font = Font(
        italic=True, color='7F8C8D', size=10)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25

    # Fields - two columns layout
    fields_left = [
        ('Nome do Projeto', None, 4),
        ('Código', None, 5),
        ('Cidade / UF', None, 6),
        ('Área Construída - AC (m²)', None, 7),
        ('Unidades Residenciais (UR)', None, 8),
        ('Nº Pavimentos (NP)', None, 9),
        ('Nº Pav. Tipo (NPT)', None, 10),
        ('Nº Pav. Garagem (NPG)', None, 11),
        ('Elevadores', None, 12),
        ('Vagas de Garagem', None, 13),
    ]

    fields_right = [
        ('Prazo (meses)', None, 4),
        ('Data-base', None, 5),
        ('CUB Referência (R$/m²)', None, 6),
        ('Tipo de Laje', None, 7),
        ('Tipo de Fundação', None, 8),
        ('Padrão de Acabamento', None, 9),
    ]

    for label, val, row in fields_left:
        label_cell(ws, row, 1, label, bold=True)
        input_cell(ws, row, 2, val, FMT_BRL if 'R$' in label or 'AC' in label or 'CUB' in label else None)

    for label, val, row in fields_right:
        label_cell(ws, row, 4, label, bold=True)
        input_cell(ws, row, 5, val, FMT_BRL if 'CUB' in label else None)

    # Data Validations (dropdowns)
    dv_laje = DataValidation(
        type='list',
        formula1='"Nervurada,Maciça,Mista,Protendida"',
        allow_blank=True
    )
    dv_laje.prompt = 'Selecione o tipo de laje'
    dv_laje.promptTitle = 'Tipo de Laje'
    ws.add_data_validation(dv_laje)
    dv_laje.add('E7')

    dv_fund = DataValidation(
        type='list',
        formula1='"Hélice Contínua,Estaca Raiz,Sapata,Radier"',
        allow_blank=True
    )
    dv_fund.prompt = 'Selecione o tipo de fundação'
    dv_fund.promptTitle = 'Tipo de Fundação'
    ws.add_data_validation(dv_fund)
    dv_fund.add('E8')

    dv_padrao = DataValidation(
        type='list',
        formula1='"Econômico,Médio,Médio-Alto,Alto,Super Alto"',
        allow_blank=True
    )
    dv_padrao.prompt = 'Selecione o padrão de acabamento'
    dv_padrao.promptTitle = 'Padrão de Acabamento'
    ws.add_data_validation(dv_padrao)
    dv_padrao.add('E9')

    # Summary section
    row = 15
    ws.merge_cells(f'A{row}:B{row}')
    c = ws.cell(row=row, column=1, value='RESUMO GERAL')
    c.font = SECTION_FONT

    row = 16
    label_cell(ws, row, 1, 'Custo Total (R$)', bold=True)
    calc_cell(ws, row, 2, "=CUSTOS_MACROGRUPO!C21", FMT_BRL)

    row = 17
    label_cell(ws, row, 1, 'R$/m² (Custo/AC)', bold=True)
    calc_cell(ws, row, 2, "=IF(B7>0,B16/B7,0)", FMT_BRL)

    row = 18
    label_cell(ws, row, 1, 'CUB Ratio', bold=True)
    calc_cell(ws, row, 2, "=IF(E6>0,B17/E6,0)", FMT_IDX)

    ws.freeze_panes = 'A4'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 2: CUSTOS_MACROGRUPO
# ════════════════════════════════════════════════════════════════════
def create_custos_macrogrupo(wb):
    ws = wb.create_sheet('CUSTOS_MACROGRUPO')
    ws.sheet_properties.tabColor = '2980B9'

    # Title
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='CUSTOS POR MACROGRUPO').font = TITLE_FONT

    # Headers row 2
    # NOTA: O header "Valor Base" deve sempre usar a data-base informada pelo Leo,
    # NÃO a data da calibração (dez/23). Ao gerar pra um projeto, ajustar o header.
    headers = ['#', 'Macrogrupo', 'Valor Base (R$)', 'R$/m²', '%', 'Faixa Min (R$/m²)',
               'Faixa Max (R$/m²)', 'Status']
    widths = [5, 35, 18, 14, 10, 18, 18, 12]
    write_header_row(ws, 2, headers, widths)

    macrogrupos = [
        ('Gerenciamento Técnico/Admin', 80, 200),
        ('Movimentação de Terra', 15, 60),
        ('Infraestrutura', 80, 250),
        ('Supraestrutura', 700, 1500),
        ('Alvenaria', 100, 300),
        ('Impermeabilização', 20, 80),
        ('Instalações (agrupado)', 250, 550),
        ('Sistemas Especiais', 30, 120),
        ('Climatização', 40, 150),
        ('Rev. Internos Parede', 80, 250),
        ('Teto', 30, 100),
        ('Pisos', 80, 250),
        ('Pintura', 30, 100),
        ('Esquadrias', 100, 350),
        ('Louças e Metais', 40, 150),
        ('Fachada', 100, 400),
        ('Complementares', 30, 120),
        ('Imprevistos', 50, 200),
    ]

    ac_ref = 'DADOS_PROJETO!B7'  # AC cell

    for i, (nome, fmin, fmax) in enumerate(macrogrupos, 1):
        r = i + 2  # data starts row 3
        # #
        label_cell(ws, r, 1, i)
        # Macrogrupo name
        label_cell(ws, r, 2, nome)
        # Valor (R$) - INPUT
        input_cell(ws, r, 3, None, FMT_BRL)
        # R$/m² = Valor / AC
        calc_cell(ws, r, 4, f'=IF({ac_ref}>0,C{r}/{ac_ref},0)', FMT_BRL)
        # % = Valor / Total
        calc_cell(ws, r, 5, f'=IF(C$21>0,C{r}/C$21,0)', FMT_PCT)
        # Faixa Min
        input_cell(ws, r, 6, fmin, FMT_BRL)
        # Faixa Max
        input_cell(ws, r, 7, fmax, FMT_BRL)
        # Status (formula: semáforo text)
        status_formula = (
            f'=IF(C{r}=0,"",IF(AND(D{r}>=F{r},D{r}<=G{r}),"✅ OK",'
            f'IF(OR(D{r}<F{r}*0.85,D{r}>G{r}*1.15),"🔴 CRÍTICO","⚠️ ATENÇÃO")))'
        )
        calc_cell(ws, r, 8, status_formula)

    # TOTAL row (row 21 = 2 + 18 + 1 header = row 21)
    total_row = 21
    ws.merge_cells(f'A{total_row}:B{total_row}')
    c = ws.cell(row=total_row, column=1, value='TOTAL')
    style_cell(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))
    style_cell(ws.cell(row=total_row, column=2), HEADER_FILL, HEADER_FONT)

    # Sum Valor
    calc_cell(ws, total_row, 3, '=SUM(C3:C20)', FMT_BRL)
    ws.cell(row=total_row, column=3).fill = HEADER_FILL
    ws.cell(row=total_row, column=3).font = Font(bold=True, color='FFFFFF', size=11)

    # R$/m² total
    calc_cell(ws, total_row, 4, f'=IF({ac_ref}>0,C{total_row}/{ac_ref},0)', FMT_BRL)
    ws.cell(row=total_row, column=4).fill = HEADER_FILL
    ws.cell(row=total_row, column=4).font = Font(bold=True, color='FFFFFF', size=11)

    # % total
    calc_cell(ws, total_row, 5, '=SUM(E3:E20)', FMT_PCT)
    ws.cell(row=total_row, column=5).fill = HEADER_FILL
    ws.cell(row=total_row, column=5).font = Font(bold=True, color='FFFFFF', size=11)

    for col in [6, 7, 8]:
        style_cell(ws.cell(row=total_row, column=col), HEADER_FILL, HEADER_FONT)

    # CUB Ratio row
    r_cub = 23
    label_cell(ws, r_cub, 1, 'CUB Ratio (R$/m² ÷ CUB ref)', bold=True)
    ws.merge_cells(f'A{r_cub}:B{r_cub}')
    calc_cell(ws, r_cub, 3, f'=IF(DADOS_PROJETO!E6>0,D{total_row}/DADOS_PROJETO!E6,0)', FMT_IDX)

    # Conditional formatting for Status column (H3:H20) based on cell content
    # Green: contains "OK"
    ws.conditional_formatting.add('H3:H20',
        FormulaRule(formula=['SEARCH("OK",H3)>0'], fill=PatternFill('solid', fgColor='D5F5E3')))
    # Yellow: contains "ATENÇÃO"
    ws.conditional_formatting.add('H3:H20',
        FormulaRule(formula=['SEARCH("ATENÇÃO",H3)>0'], fill=PatternFill('solid', fgColor='FEF9E7')))
    # Red: contains "CRÍTICO"
    ws.conditional_formatting.add('H3:H20',
        FormulaRule(formula=['SEARCH("CRÍTICO",H3)>0'], fill=PatternFill('solid', fgColor='FADBD8')))

    # Also color the R$/m² column (D3:D20) with semáforo
    for i in range(3, 21):
        r = i
        # Red if outside 15%
        ws.conditional_formatting.add(f'D{r}',
            FormulaRule(formula=[f'OR(D{r}<F{r}*0.85,D{r}>G{r}*1.15)'],
                        fill=PatternFill('solid', fgColor='FADBD8'),
                        font=Font(color='C0392B', bold=True)))
        # Yellow if outside range but within 15%
        ws.conditional_formatting.add(f'D{r}',
            FormulaRule(formula=[f'OR(D{r}<F{r},D{r}>G{r})'],
                        fill=PatternFill('solid', fgColor='FEF9E7'),
                        font=Font(color='E67E22')))

    ws.freeze_panes = 'A3'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 3: ESTRUTURAL
# ════════════════════════════════════════════════════════════════════
def create_estrutural(wb):
    ws = wb.create_sheet('ESTRUTURAL')
    ws.sheet_properties.tabColor = 'E67E22'

    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value='ANÁLISE ESTRUTURAL').font = TITLE_FONT

    widths = [25, 16, 12, 12, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ac_ref = 'DADOS_PROJETO!B7'
    r = 3

    # ── CONCRETO ──
    write_subheader(ws, r, 1, 7, 'CONCRETO')
    r += 1
    headers_conc = ['Elemento', 'Volume (m³)', '% Vol.', 'fck (MPa)', 'PU (R$/m³)', 'Total (R$)', 'R$/m²']
    write_header_row(ws, r, headers_conc)
    r += 1
    conc_start = r

    conc_items = ['Pilares', 'Vigas', 'Lajes', 'Escadas']
    for item in conc_items:
        label_cell(ws, r, 1, item)
        input_cell(ws, r, 2, None, FMT_BRL)   # Volume
        calc_cell(ws, r, 3, f'=IF(B${conc_start+len(conc_items)}>0,B{r}/B${conc_start+len(conc_items)},0)', FMT_PCT)
        input_cell(ws, r, 4, None, FMT_INT)    # fck
        input_cell(ws, r, 5, None, FMT_BRL)    # PU
        calc_cell(ws, r, 6, f'=B{r}*E{r}', FMT_BRL)  # Total
        calc_cell(ws, r, 7, f'=IF({ac_ref}>0,F{r}/{ac_ref},0)', FMT_BRL)  # R$/m²
        r += 1

    conc_end = r - 1
    # Total Concreto
    label_cell(ws, r, 1, 'Total Concreto', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{conc_start}:B{conc_end})', FMT_BRL)
    calc_cell(ws, r, 3, '=1', FMT_PCT)
    ws.cell(row=r, column=4)  # blank
    ws.cell(row=r, column=5)  # blank
    calc_cell(ws, r, 6, f'=SUM(F{conc_start}:F{conc_end})', FMT_BRL)
    calc_cell(ws, r, 7, f'=IF({ac_ref}>0,F{r}/{ac_ref},0)', FMT_BRL)
    conc_total_row = r

    # Fix % formulas to reference total
    for ir in range(conc_start, conc_end + 1):
        calc_cell(ws, ir, 3, f'=IF(B${conc_total_row}>0,B{ir}/B${conc_total_row},0)', FMT_PCT)

    r += 2

    # ── AÇO ──
    write_subheader(ws, r, 1, 7, 'AÇO')
    r += 1
    headers_aco = ['Elemento', 'Peso (kg)', '% Peso', 'Taxa (kg/m³)', '', '', '']
    write_header_row(ws, r, headers_aco)
    r += 1
    aco_start = r

    for idx, item in enumerate(conc_items):
        label_cell(ws, r, 1, item)
        input_cell(ws, r, 2, None, FMT_INT)   # Peso
        # % will be fixed after total
        calc_cell(ws, r, 3, '', FMT_PCT)
        # Taxa = Peso / Volume concreto do mesmo elemento
        conc_vol_row = conc_start + idx
        calc_cell(ws, r, 4, f'=IF(ESTRUTURAL!B{conc_vol_row}>0,B{r}/ESTRUTURAL!B{conc_vol_row},0)', FMT_TAXA)
        r += 1

    aco_end = r - 1
    label_cell(ws, r, 1, 'Total Aço', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{aco_start}:B{aco_end})', FMT_INT)
    calc_cell(ws, r, 3, '=1', FMT_PCT)
    calc_cell(ws, r, 4, f'=IF(B{conc_total_row}>0,B{r}/B{conc_total_row},0)', FMT_TAXA)
    aco_total_row = r

    for ir in range(aco_start, aco_end + 1):
        calc_cell(ws, ir, 3, f'=IF(B${aco_total_row}>0,B{ir}/B${aco_total_row},0)', FMT_PCT)

    r += 2

    # ── FORMA ──
    write_subheader(ws, r, 1, 7, 'FORMA')
    r += 1
    headers_forma = ['Tipo', 'Área (m²)', 'Reutilizações', 'PU (R$/m²)', 'Total (R$)', 'R$/m²', '']
    write_header_row(ws, r, headers_forma)
    r += 1
    forma_start = r

    for item in ['Madeira Serrada', 'Compensado']:
        label_cell(ws, r, 1, item)
        input_cell(ws, r, 2, None, FMT_INT)    # Área
        input_cell(ws, r, 3, None, FMT_INT)    # Reutilizações
        input_cell(ws, r, 4, None, FMT_BRL)    # PU
        calc_cell(ws, r, 5, f'=B{r}*D{r}', FMT_BRL)  # Total
        calc_cell(ws, r, 6, f'=IF({ac_ref}>0,E{r}/{ac_ref},0)', FMT_BRL)
        r += 1

    forma_end = r - 1
    label_cell(ws, r, 1, 'Total Forma', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{forma_start}:B{forma_end})', FMT_INT)
    calc_cell(ws, r, 5, f'=SUM(E{forma_start}:E{forma_end})', FMT_BRL)
    calc_cell(ws, r, 6, f'=IF({ac_ref}>0,E{r}/{ac_ref},0)', FMT_BRL)
    forma_total_row = r

    r += 2

    # ── ÍNDICES CALCULADOS ──
    write_subheader(ws, r, 1, 4, 'ÍNDICES ESTRUTURAIS')
    r += 1
    label_cell(ws, r, 1, 'Concreto / AC (m³/m²)', bold=True)
    calc_cell(ws, r, 2, f'=IF({ac_ref}>0,B{conc_total_row}/{ac_ref},0)', FMT_IDX)
    r += 1
    label_cell(ws, r, 1, 'Taxa Aço Global (kg/m³)', bold=True)
    calc_cell(ws, r, 2, f'=IF(B{conc_total_row}>0,B{aco_total_row}/B{conc_total_row},0)', FMT_TAXA)
    r += 1
    label_cell(ws, r, 1, 'Forma / AC (m²/m²)', bold=True)
    calc_cell(ws, r, 2, f'=IF({ac_ref}>0,B{forma_total_row}/{ac_ref},0)', FMT_IDX)

    ws.freeze_panes = 'A3'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 4: INSTALACOES
# ════════════════════════════════════════════════════════════════════
def create_instalacoes(wb):
    ws = wb.create_sheet('INSTALACOES')
    ws.sheet_properties.tabColor = '8E44AD'

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='INSTALAÇÕES').font = TITLE_FONT

    headers = ['Disciplina', 'Valor (R$)', 'R$/m²', '% Instalações', 'MO (R$/m²)', 'Material (R$/m²)']
    widths = [30, 18, 14, 16, 16, 16]
    write_header_row(ws, 2, headers, widths)

    ac_ref = 'DADOS_PROJETO!B7'
    disciplinas = [
        'Hidrossanitárias',
        'Elétricas',
        'Preventivas (Incêndio)',
        'Gás',
        'Comunicações / Telecom',
    ]

    start_row = 3
    for i, disc in enumerate(disciplinas):
        r = start_row + i
        label_cell(ws, r, 1, disc)
        input_cell(ws, r, 2, None, FMT_BRL)   # Valor
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)  # R$/m²
        calc_cell(ws, r, 4, f'=IF(B${start_row + len(disciplinas)}>0,B{r}/B${start_row + len(disciplinas)},0)', FMT_PCT)
        input_cell(ws, r, 5, None, FMT_BRL)   # MO R$/m²
        calc_cell(ws, r, 6, f'=IF(C{r}>0,C{r}-E{r},0)', FMT_BRL)  # Material = Total - MO

    total_row = start_row + len(disciplinas)
    label_cell(ws, total_row, 1, 'TOTAL INSTALAÇÕES', bold=True)
    calc_cell(ws, total_row, 2, f'=SUM(B{start_row}:B{total_row - 1})', FMT_BRL)
    calc_cell(ws, total_row, 3, f'=IF({ac_ref}>0,B{total_row}/{ac_ref},0)', FMT_BRL)
    calc_cell(ws, total_row, 4, '=1', FMT_PCT)
    # MO total ponderado
    calc_cell(ws, total_row, 5, f'=SUM(E{start_row}:E{total_row - 1})', FMT_BRL)
    calc_cell(ws, total_row, 6, f'=SUM(F{start_row}:F{total_row - 1})', FMT_BRL)

    # Fix % formulas
    for i in range(start_row, total_row):
        calc_cell(ws, i, 4, f'=IF(B${total_row}>0,B{i}/B${total_row},0)', FMT_PCT)

    # Indices
    r = total_row + 2
    write_subheader(ws, r, 1, 4, 'ÍNDICES DE INSTALAÇÕES')
    r += 1
    label_cell(ws, r, 1, 'MO Total / AC (R$/m²)', bold=True)
    calc_cell(ws, r, 2, f'=E{total_row}', FMT_BRL)
    r += 1
    label_cell(ws, r, 1, 'Material Total / AC (R$/m²)', bold=True)
    calc_cell(ws, r, 2, f'=F{total_row}', FMT_BRL)
    r += 1
    label_cell(ws, r, 1, 'Razão MO / Material', bold=True)
    calc_cell(ws, r, 2, f'=IF(F{total_row}>0,E{total_row}/F{total_row},0)', FMT_IDX)

    ws.freeze_panes = 'A3'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 5: ACABAMENTOS
# ════════════════════════════════════════════════════════════════════
def create_acabamentos(wb):
    ws = wb.create_sheet('ACABAMENTOS')
    ws.sheet_properties.tabColor = '16A085'

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='ACABAMENTOS DETALHADOS').font = TITLE_FONT

    widths = [28, 12, 8, 16, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ['Item', 'Qtd', 'Un', 'PU Material (R$)', 'PU MO (R$)',
               'Total Material (R$)', 'Total MO (R$)', 'Total (R$)']

    ac_ref = 'DADOS_PROJETO!B7'

    sections = {
        'REVESTIMENTOS DE PAREDE': [
            ('Chapisco interno', 'm²'),
            ('Reboco interno', 'm²'),
            ('Estucamento / Massa corrida', 'm²'),
            ('Cerâmico parede', 'm²'),
            ('Porcelanato parede', 'm²'),
            ('Pastilha interna', 'm²'),
        ],
        'PISOS': [
            ('Contrapiso', 'm²'),
            ('Porcelanato piso', 'm²'),
            ('Cerâmico piso', 'm²'),
            ('Laminado', 'm²'),
            ('Vinílico', 'm²'),
            ('Rodapé', 'm'),
            ('Soleira / Peitoril', 'm'),
        ],
        'TETO': [
            ('Forro gesso acartonado', 'm²'),
            ('Forro gesso RU', 'm²'),
            ('Forro madeira', 'm²'),
            ('Negativo / Tabica', 'm'),
            ('Sanca', 'm'),
        ],
        'PINTURA': [
            ('Pintura parede (PVA/acrílica)', 'm²'),
            ('Textura', 'm²'),
            ('Pintura teto', 'm²'),
            ('Epóxi', 'm²'),
            ('Verniz', 'm²'),
        ],
        'FACHADA': [
            ('Chapisco externo', 'm²'),
            ('Reboco externo', 'm²'),
            ('Textura externa', 'm²'),
            ('Pastilha fachada', 'm²'),
            ('ACM / Alumínio Composto', 'm²'),
            ('Pele de vidro', 'm²'),
        ],
    }

    r = 3
    grand_total_mat_cells = []
    grand_total_mo_cells = []

    for section_name, items in sections.items():
        write_subheader(ws, r, 1, 8, section_name)
        r += 1
        write_header_row(ws, r, headers)
        r += 1
        sec_start = r

        for item_name, un in items:
            label_cell(ws, r, 1, item_name)
            input_cell(ws, r, 2, None, FMT_INT)     # Qtd
            label_cell(ws, r, 3, un)                  # Un
            input_cell(ws, r, 4, None, FMT_BRL)      # PU Mat
            input_cell(ws, r, 5, None, FMT_BRL)      # PU MO
            calc_cell(ws, r, 6, f'=B{r}*D{r}', FMT_BRL)  # Total Mat
            calc_cell(ws, r, 7, f'=B{r}*E{r}', FMT_BRL)  # Total MO
            calc_cell(ws, r, 8, f'=F{r}+G{r}', FMT_BRL)  # Total
            r += 1

        sec_end = r - 1
        # Subtotal
        label_cell(ws, r, 1, f'Subtotal {section_name.title()}', bold=True)
        calc_cell(ws, r, 6, f'=SUM(F{sec_start}:F{sec_end})', FMT_BRL)
        calc_cell(ws, r, 7, f'=SUM(G{sec_start}:G{sec_end})', FMT_BRL)
        calc_cell(ws, r, 8, f'=F{r}+G{r}', FMT_BRL)
        grand_total_mat_cells.append(f'F{r}')
        grand_total_mo_cells.append(f'G{r}')
        r += 2

    # Grand total
    write_subheader(ws, r, 1, 8, 'TOTAL ACABAMENTOS')
    r += 1
    label_cell(ws, r, 1, 'Total Material', bold=True)
    calc_cell(ws, r, 6, f'={"+".join(grand_total_mat_cells)}', FMT_BRL)
    r += 1
    label_cell(ws, r, 1, 'Total Mão de Obra', bold=True)
    calc_cell(ws, r, 7, f'={"+".join(grand_total_mo_cells)}', FMT_BRL)
    r += 1
    label_cell(ws, r, 1, 'TOTAL GERAL ACABAMENTOS', bold=True)
    calc_cell(ws, r, 8, f'=F{r-2}+G{r-1}', FMT_BRL)
    r += 1
    label_cell(ws, r, 1, 'R$/m² Acabamentos', bold=True)
    calc_cell(ws, r, 8, f'=IF({ac_ref}>0,H{r-1}/{ac_ref},0)', FMT_BRL)

    ws.freeze_panes = 'A3'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 6: CI_DETALHADO
# ════════════════════════════════════════════════════════════════════
def create_ci_detalhado(wb):
    ws = wb.create_sheet('CI_DETALHADO')
    ws.sheet_properties.tabColor = 'C0392B'

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='CUSTOS INDIRETOS DETALHADOS').font = TITLE_FONT

    widths = [30, 14, 14, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ac_ref = 'DADOS_PROJETO!B7'
    r = 3
    subtotal_cells = []

    # ── Projetos e Consultorias ──
    write_subheader(ws, r, 1, 6, 'PROJETOS E CONSULTORIAS')
    r += 1
    headers_proj = ['Disciplina', 'Valor (R$)', 'R$/m²', '', '', '']
    write_header_row(ws, r, headers_proj)
    r += 1
    proj_start = r
    projetos = ['Arquitetônico', 'Estrutural', 'Elétrico', 'Hidrossanitário',
                'Preventivo', 'Climatização', 'Paisagismo', 'Interiores',
                'Fundações', 'Outros Projetos']
    for p in projetos:
        label_cell(ws, r, 1, p)
        input_cell(ws, r, 2, None, FMT_BRL)
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
        r += 1
    proj_end = r - 1
    label_cell(ws, r, 1, 'Subtotal Projetos', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{proj_start}:B{proj_end})', FMT_BRL)
    calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
    subtotal_cells.append(f'B{r}')
    r += 2

    # ── Taxas e Licenças ──
    write_subheader(ws, r, 1, 6, 'TAXAS E LICENÇAS')
    r += 1
    write_header_row(ws, r, ['Descrição', 'Valor (R$)', 'R$/m²', '', '', ''])
    r += 1
    tax_start = r
    taxas = ['Alvará de Construção', 'ARTs / RRTs', 'Corpo de Bombeiros',
             'Habite-se', 'Registro de Incorporação', 'Outras Taxas']
    for t in taxas:
        label_cell(ws, r, 1, t)
        input_cell(ws, r, 2, None, FMT_BRL)
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
        r += 1
    tax_end = r - 1
    label_cell(ws, r, 1, 'Subtotal Taxas', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{tax_start}:B{tax_end})', FMT_BRL)
    calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
    subtotal_cells.append(f'B{r}')
    r += 2

    # ── Equipe ADM ──
    write_subheader(ws, r, 1, 6, 'EQUIPE ADMINISTRATIVA')
    r += 1
    headers_eq = ['Função', 'Qtd', 'Custo/mês (R$)', 'Meses', 'Total (R$)', 'R$/m²']
    write_header_row(ws, r, headers_eq)
    r += 1
    eq_start = r
    equipe = ['Engenheiro Residente', 'Mestre de Obras', 'Almoxarife',
              'Técnico Segurança', 'Administrativo Obra', 'Estagiários']
    for e in equipe:
        label_cell(ws, r, 1, e)
        input_cell(ws, r, 2, None, FMT_INT)    # Qtd
        input_cell(ws, r, 3, None, FMT_BRL)    # Custo/mês
        input_cell(ws, r, 4, None, FMT_INT)    # Meses
        calc_cell(ws, r, 5, f'=B{r}*C{r}*D{r}', FMT_BRL)  # Total
        calc_cell(ws, r, 6, f'=IF({ac_ref}>0,E{r}/{ac_ref},0)', FMT_BRL)
        r += 1
    eq_end = r - 1
    label_cell(ws, r, 1, 'Subtotal Equipe ADM', bold=True)
    calc_cell(ws, r, 5, f'=SUM(E{eq_start}:E{eq_end})', FMT_BRL)
    calc_cell(ws, r, 6, f'=IF({ac_ref}>0,E{r}/{ac_ref},0)', FMT_BRL)
    subtotal_cells.append(f'E{r}')
    r += 2

    # ── EPCs ──
    write_subheader(ws, r, 1, 6, 'EPCs (PROTEÇÃO COLETIVA)')
    r += 1
    write_header_row(ws, r, ['Item', 'Valor (R$)', 'R$/m²', '', '', ''])
    r += 1
    epc_start = r
    epcs = ['Bandeja de Proteção', 'Guarda-corpo', 'Tela de Proteção',
            'Linha de Vida', 'Sinalização', 'Outros EPCs']
    for e in epcs:
        label_cell(ws, r, 1, e)
        input_cell(ws, r, 2, None, FMT_BRL)
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
        r += 1
    epc_end = r - 1
    label_cell(ws, r, 1, 'Subtotal EPCs', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{epc_start}:B{epc_end})', FMT_BRL)
    calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
    subtotal_cells.append(f'B{r}')
    r += 2

    # ── Equipamentos ──
    write_subheader(ws, r, 1, 6, 'EQUIPAMENTOS')
    r += 1
    write_header_row(ws, r, ['Equipamento', 'Valor (R$)', 'R$/m²', '', '', ''])
    r += 1
    equip_start = r
    equips = ['Cremalheira', 'Grua / Guindaste', 'Balancins',
              'Mini-Grua', 'Outros Equipamentos']
    for e in equips:
        label_cell(ws, r, 1, e)
        input_cell(ws, r, 2, None, FMT_BRL)
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
        r += 1
    equip_end = r - 1
    label_cell(ws, r, 1, 'Subtotal Equipamentos', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{equip_start}:B{equip_end})', FMT_BRL)
    calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
    subtotal_cells.append(f'B{r}')
    r += 2

    # ── Ensaios Tecnológicos ──
    write_subheader(ws, r, 1, 6, 'ENSAIOS TECNOLÓGICOS')
    r += 1
    write_header_row(ws, r, ['Descrição', 'Valor (R$)', 'R$/m²', '', '', ''])
    r += 1
    ens_start = r
    ensaios = ['Concreto (corpos de prova)', 'Aço', 'Solo / Fundações',
               'Estanqueidade', 'Outros Ensaios']
    for e in ensaios:
        label_cell(ws, r, 1, e)
        input_cell(ws, r, 2, None, FMT_BRL)
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
        r += 1
    ens_end = r - 1
    label_cell(ws, r, 1, 'Subtotal Ensaios', bold=True)
    calc_cell(ws, r, 2, f'=SUM(B{ens_start}:B{ens_end})', FMT_BRL)
    calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
    subtotal_cells.append(f'B{r}')
    r += 2

    # ── RESUMO CI ──
    write_subheader(ws, r, 1, 6, 'RESUMO CUSTOS INDIRETOS')
    r += 1
    resumo_labels = ['Projetos e Consultorias', 'Taxas e Licenças', 'Equipe ADM',
                     'EPCs', 'Equipamentos', 'Ensaios Tecnológicos']
    for i, (lab, cell_ref) in enumerate(zip(resumo_labels, subtotal_cells)):
        label_cell(ws, r, 1, lab)
        calc_cell(ws, r, 2, f'={cell_ref}', FMT_BRL)
        calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
        r += 1

    label_cell(ws, r, 1, 'TOTAL CUSTOS INDIRETOS', bold=True)
    total_ci_start = r - len(subtotal_cells)
    calc_cell(ws, r, 2, f'=SUM(B{total_ci_start}:B{r-1})', FMT_BRL)
    calc_cell(ws, r, 3, f'=IF({ac_ref}>0,B{r}/{ac_ref},0)', FMT_BRL)
    r += 1
    label_cell(ws, r, 1, '% CI sobre Custo Total', bold=True)
    calc_cell(ws, r, 2, f'=IF(CUSTOS_MACROGRUPO!C21>0,B{r-1}/CUSTOS_MACROGRUPO!C21,0)', FMT_PCT)

    ws.freeze_panes = 'A3'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 7: BENCHMARK
# ════════════════════════════════════════════════════════════════════
def create_benchmark(wb):
    ws = wb.create_sheet('BENCHMARK')
    ws.sheet_properties.tabColor = '27AE60'

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value='BENCHMARK — PROJETOS DE REFERÊNCIA').font = TITLE_FONT

    # Reference projects table
    headers = ['Projeto', 'Cidade', 'AC (m²)', 'NP', 'Prazo (meses)',
               'CUB (R$/m²)', 'R$/m²', 'CUB Ratio']
    widths = [25, 18, 14, 8, 14, 14, 14, 12]
    write_header_row(ws, 3, headers, widths)

    # 10 rows for reference projects
    for i in range(4, 14):
        for col in range(1, 9):
            if col <= 2:
                input_cell(ws, i, col)
            elif col in [3, 5, 6, 7]:
                input_cell(ws, i, col, None, FMT_BRL if col >= 6 else FMT_INT)
            elif col == 4:
                input_cell(ws, i, col, None, FMT_INT)
            elif col == 8:
                calc_cell(ws, i, col, f'=IF(F{i}>0,G{i}/F{i},0)', FMT_IDX)

    # Averages row
    r = 14
    label_cell(ws, r, 1, 'MÉDIA', bold=True)
    for col in [3, 4, 5, 6, 7, 8]:
        fmt = FMT_BRL if col in [6, 7] else (FMT_INT if col in [3, 4, 5] else FMT_IDX)
        calc_cell(ws, r, col, f'=IF(COUNTA({get_column_letter(col)}4:{get_column_letter(col)}13)>0,'
                  f'AVERAGE({get_column_letter(col)}4:{get_column_letter(col)}13),0)', fmt)

    # Comparative by macrogrupo
    r = 17
    ws.merge_cells(f'A{r}:E{r}')
    ws.cell(row=r, column=1, value='COMPARATIVO R$/m² POR MACROGRUPO').font = SECTION_FONT

    r = 19
    comp_headers = ['#', 'Macrogrupo', 'Este Projeto (R$/m²)', 'Média Referências (R$/m²)', 'Desvio (%)']
    comp_widths = [5, 35, 22, 22, 14]
    write_header_row(ws, r, comp_headers, comp_widths)

    macrogrupos = [
        'Gerenciamento Técnico/Admin', 'Movimentação de Terra', 'Infraestrutura',
        'Supraestrutura', 'Alvenaria', 'Impermeabilização', 'Instalações (agrupado)',
        'Sistemas Especiais', 'Climatização', 'Rev. Internos Parede', 'Teto',
        'Pisos', 'Pintura', 'Esquadrias', 'Louças e Metais', 'Fachada',
        'Complementares', 'Imprevistos'
    ]

    for i, nome in enumerate(macrogrupos):
        ri = r + 1 + i
        label_cell(ws, ri, 1, i + 1)
        label_cell(ws, ri, 2, nome)
        # Este Projeto = referência da aba CUSTOS_MACROGRUPO
        macro_row = i + 3  # data starts at row 3 in CUSTOS_MACROGRUPO
        calc_cell(ws, ri, 3, f'=CUSTOS_MACROGRUPO!D{macro_row}', FMT_BRL)
        # Média Referências (INPUT - user fills from their database)
        input_cell(ws, ri, 4, None, FMT_BRL)
        # Desvio
        calc_cell(ws, ri, 5, f'=IF(D{ri}>0,(C{ri}-D{ri})/D{ri},0)', FMT_PCT)

    # Benchmark ranges note
    r_note = r + 1 + len(macrogrupos) + 2
    ws.merge_cells(f'A{r_note}:E{r_note}')
    ws.cell(row=r_note, column=1, value='FAIXAS TÍPICAS — BASE CARTESIAN').font = SECTION_FONT
    notes = [
        'Padrão Alto SC: R$/m² 3.000 – 4.500 | CUB ratio 1,05 – 1,50',
        'Supraestrutura: tipicamente 25% – 35% do custo total',
        'Instalações: tipicamente 8% – 15% do custo total',
        'Custos Indiretos: tipicamente 10% – 20% do custo total',
        'Base de referência: 40 projetos Cartesian Engenharia',
    ]
    for i, note in enumerate(notes):
        ws.cell(row=r_note + 1 + i, column=1, value=note).font = Font(
            italic=True, color='7F8C8D', size=10)

    ws.freeze_panes = 'A4'
    return ws


# ════════════════════════════════════════════════════════════════════
# ABA 8: ALERTAS
# ════════════════════════════════════════════════════════════════════
def create_alertas(wb):
    ws = wb.create_sheet('ALERTAS')
    ws.sheet_properties.tabColor = 'E74C3C'

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='PAINEL DE ALERTAS').font = TITLE_FONT

    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1,
            value='Resumo automático dos macrogrupos fora da faixa de benchmark').font = Font(
                italic=True, color='7F8C8D', size=10)

    widths = [5, 35, 16, 16, 16, 18]
    headers = ['#', 'Macrogrupo', 'R$/m²', 'Faixa Min', 'Faixa Max', 'Situação']
    write_header_row(ws, 4, headers, widths)

    macrogrupos = [
        'Gerenciamento Técnico/Admin', 'Movimentação de Terra', 'Infraestrutura',
        'Supraestrutura', 'Alvenaria', 'Impermeabilização', 'Instalações (agrupado)',
        'Sistemas Especiais', 'Climatização', 'Rev. Internos Parede', 'Teto',
        'Pisos', 'Pintura', 'Esquadrias', 'Louças e Metais', 'Fachada',
        'Complementares', 'Imprevistos'
    ]

    for i, nome in enumerate(macrogrupos):
        r = 5 + i
        macro_row = i + 3
        label_cell(ws, r, 1, i + 1)
        label_cell(ws, r, 2, nome)
        # R$/m² from CUSTOS_MACROGRUPO
        calc_cell(ws, r, 3, f'=CUSTOS_MACROGRUPO!D{macro_row}', FMT_BRL)
        # Faixa Min
        calc_cell(ws, r, 4, f'=CUSTOS_MACROGRUPO!F{macro_row}', FMT_BRL)
        # Faixa Max
        calc_cell(ws, r, 5, f'=CUSTOS_MACROGRUPO!G{macro_row}', FMT_BRL)
        # Situação
        situacao = (
            f'=IF(C{r}=0,"—",'
            f'IF(C{r}>E{r},"⚠️ Acima da Faixa",'
            f'IF(C{r}<D{r},"🔽 Abaixo da Faixa","✅ Dentro da Faixa")))'
        )
        calc_cell(ws, r, 6, situacao)

    # Conditional formatting on Situação column
    alert_range = 'F5:F22'
    ws.conditional_formatting.add(alert_range,
        FormulaRule(formula=['SEARCH("Acima",F5)>0'],
                    fill=PatternFill('solid', fgColor='FADBD8'),
                    font=Font(color='C0392B', bold=True)))
    ws.conditional_formatting.add(alert_range,
        FormulaRule(formula=['SEARCH("Abaixo",F5)>0'],
                    fill=PatternFill('solid', fgColor='FEF9E7'),
                    font=Font(color='E67E22')))
    ws.conditional_formatting.add(alert_range,
        FormulaRule(formula=['SEARCH("Dentro",F5)>0'],
                    fill=PatternFill('solid', fgColor='D5F5E3'),
                    font=Font(color='27AE60')))

    # Also color R$/m² column
    for i in range(5, 23):
        ws.conditional_formatting.add(f'C{i}',
            FormulaRule(formula=[f'C{i}>E{i}'],
                        fill=PatternFill('solid', fgColor='FADBD8')))
        ws.conditional_formatting.add(f'C{i}',
            FormulaRule(formula=[f'C{i}<D{i}'],
                        fill=PatternFill('solid', fgColor='FEF9E7')))

    # Summary counts at bottom
    r = 25
    ws.merge_cells(f'A{r}:F{r}')
    ws.cell(row=r, column=1, value='RESUMO').font = SECTION_FONT

    r += 1
    label_cell(ws, r, 1, '⚠️ Acima da Faixa:', bold=True)
    calc_cell(ws, r, 2, '=COUNTIF(F5:F22,"*Acima*")', FMT_INT)

    r += 1
    label_cell(ws, r, 1, '✅ Dentro da Faixa:', bold=True)
    calc_cell(ws, r, 2, '=COUNTIF(F5:F22,"*Dentro*")', FMT_INT)

    r += 1
    label_cell(ws, r, 1, '🔽 Abaixo da Faixa:', bold=True)
    calc_cell(ws, r, 2, '=COUNTIF(F5:F22,"*Abaixo*")', FMT_INT)

    r += 1
    label_cell(ws, r, 1, '— Sem dados:', bold=True)
    calc_cell(ws, r, 2, '=COUNTIF(F5:F22,"—")', FMT_INT)

    ws.freeze_panes = 'A5'
    return ws


# ════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    print("Criando aba DADOS_PROJETO...")
    create_dados_projeto(wb)

    print("Criando aba CUSTOS_MACROGRUPO...")
    create_custos_macrogrupo(wb)

    print("Criando aba ESTRUTURAL...")
    create_estrutural(wb)

    print("Criando aba INSTALACOES...")
    create_instalacoes(wb)

    print("Criando aba ACABAMENTOS...")
    create_acabamentos(wb)

    print("Criando aba CI_DETALHADO...")
    create_ci_detalhado(wb)

    print("Criando aba BENCHMARK...")
    create_benchmark(wb)

    print("Criando aba ALERTAS...")
    create_alertas(wb)

    output = os.path.expanduser('~/clawd/orcamento-parametrico/template-orcamento-parametrico.xlsx')
    wb.save(output)
    print(f"\n✅ Planilha salva em: {output}")
    print(f"   Abas: {wb.sheetnames}")
    print(f"   Tamanho: {os.path.getsize(output) / 1024:.1f} KB")


if __name__ == '__main__':
    main()
