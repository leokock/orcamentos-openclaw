#!/usr/bin/env python3.11
"""
Merge Cowork features into parametric budget template
Cartesian Engenharia - Mar 2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import shutil
from datetime import datetime
from pathlib import Path

# Style definitions
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
SECTION_FONT = Font(bold=True, color="34495E", size=12)
SECTION_FILL = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)

def apply_header_style(cell):
    """Apply header style to a cell"""
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER

def apply_input_style(cell):
    """Apply input cell style"""
    cell.fill = INPUT_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = THIN_BORDER

def apply_calc_style(cell):
    """Apply calculated cell style"""
    cell.fill = CALC_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = THIN_BORDER

def apply_section_style(cell):
    """Apply section header style"""
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = THIN_BORDER

def create_painel_tab(wb, calibration_data):
    """Create the PAINEL executive dashboard tab"""
    print("Creating PAINEL tab...")
    
    # Create new sheet at index 0
    ws = wb.create_sheet("PAINEL", 0)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = "PAINEL EXECUTIVO — ORÇAMENTO PARAMÉTRICO"
    cell.font = Font(bold=True, size=14, color="2C3E50")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # ===== SECTION 1: RESUMO EXECUTIVO =====
    ws[f'A{row}'] = "1. RESUMO EXECUTIVO"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # KPI Cards
    kpis = [
        ("R$/m² (Custo/AC)", "=DADOS_PROJETO!B18", "#,##0.00"),
        ("Custo Total (R$)", "=DADOS_PROJETO!B17", "#,##0.00"),
        ("R$/UR", "=IF(DADOS_PROJETO!B8>0,DADOS_PROJETO!B17/DADOS_PROJETO!B8,0)", "#,##0.00"),
        ("Área Construída (m²)", "=DADOS_PROJETO!B7", "#,##0.00")
    ]
    
    for kpi_name, formula, num_format in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = num_format
        apply_calc_style(ws[f'B{row}'])
        row += 1
    
    row += 1
    
    # ===== SECTION 2: CUSTO POR MACROGRUPO =====
    ws[f'A{row}'] = "2. CUSTO POR MACROGRUPO"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Headers
    headers = ["Macrogrupo", "R$/m² Param.", "Min Mercado", "Max Mercado", "Status/Ação"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Macrogroups (linking to CUSTOS_MACROGRUPO)
    macrogrupos = [
        ("Gerenciamento", 4, 6, 7),
        ("Mov. Terra", 5, 6, 7),
        ("Infraestrutura", 6, 6, 7),
        ("Supraestrutura", 7, 6, 7),
        ("Alvenaria", 8, 6, 7),
        ("Instalações", 9, 6, 7),
        ("Sist. Especiais", 10, 6, 7),
        ("Impermeabilização", 11, 6, 7),
        ("Rev. Int. Parede", 12, 6, 7),
        ("Teto", 13, 6, 7),
        ("Pisos", 14, 6, 7),
        ("Pintura", 15, 6, 7),
        ("Esquadrias", 16, 6, 7),
        ("Cobertura", 17, 6, 7),
        ("Fachada", 18, 6, 7),
        ("Complementares", 19, 6, 7),
        ("Imprevistos", 20, 6, 7)
    ]
    
    for macro_name, custos_row, min_col, max_col in macrogrupos:
        ws[f'A{row}'] = macro_name
        ws[f'A{row}'].border = THIN_BORDER
        
        # Link to CUSTOS_MACROGRUPO R$/m²
        ws[f'B{row}'] = f"=CUSTOS_MACROGRUPO!D{custos_row}"
        ws[f'B{row}'].number_format = "#,##0.00"
        apply_calc_style(ws[f'B{row}'])
        
        # Min/Max from CUSTOS_MACROGRUPO
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!F{custos_row}"
        ws[f'C{row}'].number_format = "#,##0.00"
        ws[f'C{row}'].border = THIN_BORDER
        
        ws[f'D{row}'] = f"=CUSTOS_MACROGRUPO!G{custos_row}"
        ws[f'D{row}'].number_format = "#,##0.00"
        ws[f'D{row}'].border = THIN_BORDER
        
        # Status with conditional logic
        status_formula = f'=IF(B{row}=0,"—",IF(B{row}>D{row},"✗ Alto - Revisar escopo",IF(B{row}<C{row},"⚠ Baixo - Verificar completude","✓ OK - Dentro da faixa")))'
        ws[f'E{row}'] = status_formula
        ws[f'E{row}'].border = THIN_BORDER
        row += 1
    
    row += 1
    
    # ===== SECTION 3: INDICADORES DE PRODUTO =====
    ws[f'A{row}'] = "3. INDICADORES DE PRODUTO"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Headers
    headers = ["Indicador", "Valor", "Faixa Benchmark", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Product indices
    indices = [
        ("CA (Coef. Aproveitamento)", "=IF(DADOS_PROJETO!B14>0,DADOS_PROJETO!B7/DADOS_PROJETO!B14,0)", "0.80 - 3.50", "0.0"),
        ("AC/UR (m²)", "=IF(DADOS_PROJETO!B8>0,DADOS_PROJETO!B7/DADOS_PROJETO!B8,0)", "60 - 150", "0.00"),
        ("Vagas/UR", "=IF(DADOS_PROJETO!B8>0,DADOS_PROJETO!B13/DADOS_PROJETO!B8,0)", "1.0 - 2.5", "0.00"),
        ("UR/Elevador", "=IF(DADOS_PROJETO!B12>0,DADOS_PROJETO!B8/DADOS_PROJETO!B12,0)", "20 - 40", "0.0"),
        ("Pvtos Tipo/Total", "=IF(DADOS_PROJETO!B9>0,DADOS_PROJETO!B10/DADOS_PROJETO!B9,0)", "0.30 - 0.60", "0.00"),
        ("N° Pavimentos", "=DADOS_PROJETO!B9", "5 - 25", "0"),
        ("R$/m²", "=DADOS_PROJETO!B18", "3500 - 5500", "#,##0.00"),
        ("R$/UR", "=IF(DADOS_PROJETO!B8>0,DADOS_PROJETO!B17/DADOS_PROJETO!B8,0)", "250k - 650k", "#,##0.00")
    ]
    
    for ind_name, formula, benchmark, num_format in indices:
        ws[f'A{row}'] = ind_name
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = num_format
        apply_calc_style(ws[f'B{row}'])
        
        ws[f'C{row}'] = benchmark
        ws[f'C{row}'].border = THIN_BORDER
        
        ws[f'D{row}'] = "—"
        ws[f'D{row}'].border = THIN_BORDER
        row += 1
    
    row += 1
    
    # ===== SECTION 4: DECISÕES DO BRIEFING =====
    ws[f'A{row}'] = "4. DECISÕES DO BRIEFING"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Headers
    headers = ["Decisão", "Escolha", "Impacto", "Alternativa", "Economia Est."]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Key briefing decisions (linking to BRIEFING sheet)
    decisoes = [
        ("Tipo de Laje", "=BRIEFING!B6", "Estrutura + Custo", "Ver ÍNDICES", "Variável"),
        ("Padrão Acabamento", "=BRIEFING!B8", "Acabamentos", "Standard/Alto/Super Alto", "10-30%"),
        ("Fachada", "=BRIEFING!B10", "Estética + Custo", "Simples/Intermediária/Complexa", "15-40%"),
        ("Pé Direito", "=BRIEFING!B12", "Estrutura + Vedação", "2.60m / 2.80m / 3.00m", "5-10%"),
        ("Sistema de Elevadores", "=BRIEFING!B20", "Instalações", "Convencional/Alta Performance", "3-8%")
    ]
    
    for decisao, formula, impacto, alt, economia in decisoes:
        ws[f'A{row}'] = decisao
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = formula
        ws[f'B{row}'].border = THIN_BORDER
        
        ws[f'C{row}'] = impacto
        ws[f'C{row}'].border = THIN_BORDER
        
        ws[f'D{row}'] = alt
        ws[f'D{row}'].border = THIN_BORDER
        
        ws[f'E{row}'] = economia
        ws[f'E{row}'].border = THIN_BORDER
        row += 1
    
    row += 1
    
    # ===== SECTION 5: POSICIONAMENTO =====
    ws[f'A{row}'] = "5. POSICIONAMENTO (Projetos Referência)"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:E{row}')
    row += 1
    
    # Headers
    headers = ["Projeto", "AC (m²)", "R$/m²", "CUB Ratio", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Add calibration projects (top 10 by size)
    sorted_projects = sorted(
        [p for p in calibration_data if p.get('ac') is not None], 
        key=lambda x: x['ac'], 
        reverse=True
    )[:10]
    
    for proj in sorted_projects:
        ws[f'A{row}'] = proj['name'].title()
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = proj['ac']
        ws[f'B{row}'].number_format = "#,##0.00"
        ws[f'B{row}'].border = THIN_BORDER
        
        ws[f'C{row}'] = proj['rsm2']
        ws[f'C{row}'].number_format = "#,##0.00"
        ws[f'C{row}'].border = THIN_BORDER
        
        ws[f'D{row}'] = proj['cub_ratio']
        ws[f'D{row}'].number_format = "0.00"
        ws[f'D{row}'].border = THIN_BORDER
        
        ws[f'E{row}'] = "Referência"
        ws[f'E{row}'].border = THIN_BORDER
        row += 1
    
    print(f"  ✓ PAINEL created with {row} rows")
    return ws

def create_indices_tab(wb):
    """Create the ÍNDICES parametric reference tab"""
    print("Creating ÍNDICES tab...")
    
    # Insert after BRIEFING (position 2, since PAINEL is 0 and BRIEFING is 1)
    ws = wb.create_sheet("ÍNDICES", 2)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "ÍNDICES PARAMÉTRICOS — REFERÊNCIA TÉCNICA"
    cell.font = Font(bold=True, size=14, color="2C3E50")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2
    
    # ===== SECTION 1: ÍNDICES POR TIPO DE LAJE =====
    ws[f'A{row}'] = "ÍNDICES POR TIPO DE LAJE"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ["Parâmetro", "Maciça", "Cubetas", "Cub.Protendida", "Treliçada", "Protendida"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Índices
    laje_indices = [
        ("Jogos de Fôrma", "12-15", "10-12", "9-11", "13-16", "8-10"),
        ("Montagem Fôrma (m²/AC)", "3.2-3.8", "2.8-3.4", "2.5-3.0", "3.5-4.2", "2.3-2.8"),
        ("Concreto (m³/AC)", "0.28-0.35", "0.22-0.28", "0.20-0.26", "0.30-0.38", "0.18-0.24"),
        ("Armadura (kg/m³)", "100-130", "85-110", "75-95", "110-140", "65-85"),
        ("MO Estrutura (R$/AC)", "280-350", "250-320", "230-300", "300-380", "210-280"),
        ("Supraestrutura R$/m²", "750-950", "650-850", "600-800", "800-1000", "550-750")
    ]
    
    for idx_name, *values in laje_indices:
        ws[f'A{row}'] = idx_name
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'A{row}'].font = Font(size=10)
        
        for col_idx, value in enumerate(values, 2):
            cell = ws.cell(row, col_idx)
            cell.value = value
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    row += 1
    
    # ===== SECTION 2: ÍNDICES GERAIS =====
    ws[f'A{row}'] = "ÍNDICES GERAIS"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ["Parâmetro", "Unidade", "Mín", "Típico", "Máx", "Observações"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Índices gerais
    indices_gerais = [
        ("Estacas", "m/m²AC", "0.15", "0.20", "0.30", "Varia por solo"),
        ("Alvenaria", "m²/m²AC", "1.8", "2.2", "2.8", "Inclui shafts"),
        ("Contrapiso", "m²/m²AC", "0.85", "0.95", "1.05", "Sobre laje"),
        ("Forro de Gesso", "m²/m²AC", "0.70", "0.85", "0.95", "Áreas nobres"),
        ("Esquadrias Alumínio", "m²/m²AC", "0.15", "0.20", "0.30", "Varanda aumenta"),
        ("Esquadrias PVC", "m²/m²AC", "0.12", "0.18", "0.25", "Alternativa"),
        ("Louças/Metais", "Jogo/UR", "1.8", "2.2", "2.8", "Padrão médio"),
        ("Ponto Elétrico", "Pts/m²AC", "8", "10", "13", "Inclui lógica"),
        ("Ponto Hidráulico", "Pts/m²AC", "4", "5", "7", "AF+AQ+ES+AG"),
        ("Elevador", "UR/Elev", "20", "28", "40", "Baixa: <25, Alta: >35")
    ]
    
    for param, unit, min_val, typ, max_val, obs in indices_gerais:
        ws[f'A{row}'] = param
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = unit
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'C{row}'] = min_val
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'D{row}'] = typ
        ws[f'D{row}'].border = THIN_BORDER
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        
        ws[f'E{row}'] = max_val
        ws[f'E{row}'].border = THIN_BORDER
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'F{row}'] = obs
        ws[f'F{row}'].border = THIN_BORDER
        ws[f'F{row}'].font = Font(size=9, italic=True)
        row += 1
    
    row += 1
    
    # ===== SECTION 3: VERBAS PARAMÉTRICAS =====
    ws[f'A{row}'] = "VERBAS PARAMÉTRICAS (R$/m² AC)"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ["Item", "Standard", "Alto", "Super Alto", "Base", "Observações"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Verbas
    verbas = [
        ("Pintura Interna", "35-45", "50-65", "75-95", "m²AC", "Látex/Acrílica"),
        ("Pintura Externa", "25-35", "40-55", "60-80", "m²AC", "Acrílica/Textura"),
        ("Rev. Parede Interna", "120-150", "180-230", "280-350", "m²AC", "Gesso/Porcelanato"),
        ("Rev. Parede Externa", "80-110", "130-170", "200-260", "m²AC", "Pastilha/ACM"),
        ("Rodapés", "8-12", "15-20", "25-35", "m²AC", "MDF/Porcelanato"),
        ("Soleiras", "5-8", "10-15", "18-25", "m²AC", "Mármore/Granito"),
        ("Paisagismo", "15-25", "35-50", "65-90", "m²AC", "Externo+Interno"),
        ("Sinalização", "3-5", "8-12", "15-22", "m²AC", "Placas+Numeração"),
        ("Playground", "8-12", "18-25", "35-50", "m²AC", "Se houver área lazer"),
        ("Deck/Pergolado", "40-60", "80-110", "140-180", "m²Deck", "Madeira/Composto")
    ]
    
    for item, std, alto, super_alto, base, obs in verbas:
        ws[f'A{row}'] = item
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = std
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'C{row}'] = alto
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'D{row}'] = super_alto
        ws[f'D{row}'].border = THIN_BORDER
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'E{row}'] = base
        ws[f'E{row}'].border = THIN_BORDER
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'E{row}'].font = Font(size=9)
        
        ws[f'F{row}'] = obs
        ws[f'F{row}'].border = THIN_BORDER
        ws[f'F{row}'].font = Font(size=9, italic=True)
        row += 1
    
    row += 1
    
    # ===== SECTION 4: FATORES DE CUSTO =====
    ws[f'A{row}'] = "FATORES DE CUSTO E AJUSTES"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ["Fator", "Valor Típico", "Faixa", "Aplicação", "Observações"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Fatores
    fatores = [
        ("BDI", "1.10", "1.08 - 1.15", "Todos os custos", "Lucro + impostos + risco"),
        ("Perdas Concreto", "1.05", "1.03 - 1.08", "Volume concreto", "Sobra + geometria"),
        ("Perdas Argamassa", "1.15", "1.10 - 1.25", "Revestimentos", "Espessura irregular"),
        ("Empolamento Terra", "1.25", "1.20 - 1.35", "Movimento terra", "Solo solto vs compactado"),
        ("Fator Dificuldade", "1.00-1.20", "—", "MO difícil acesso", "Encosta, centro urbano"),
        ("Ajuste Localização", "0.90-1.15", "—", "Todos", "Interior vs capital"),
        ("Produtividade MO", "0.85-1.15", "—", "MO", "Equipe experiente"),
        ("Fator Prazo", "1.00-1.30", "—", "MO compressão", "Aceleração > custo")
    ]
    
    for fator, valor, faixa, aplicacao, obs in fatores:
        ws[f'A{row}'] = fator
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'A{row}'].font = Font(bold=True, size=10)
        
        ws[f'B{row}'] = valor
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'B{row}'].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        
        ws[f'C{row}'] = faixa
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'D{row}'] = aplicacao
        ws[f'D{row}'].border = THIN_BORDER
        ws[f'D{row}'].font = Font(size=9)
        
        ws[f'E{row}'] = obs
        ws[f'E{row}'].border = THIN_BORDER
        ws[f'E{row}'].font = Font(size=9, italic=True)
        row += 1
    
    row += 1
    
    # ===== SECTION 5: AJUSTES POR PADRÃO =====
    ws[f'A{row}'] = "AJUSTES POR PADRÃO (Multiplicadores Base Standard)"
    apply_section_style(ws[f'A{row}'])
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    # Headers
    headers = ["Macrogrupo", "Standard", "Alto Padrão", "Super Alto", "Impacto", "Observações"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row, col_idx)
        cell.value = header
        apply_header_style(cell)
    row += 1
    
    # Ajustes
    ajustes_padrao = [
        ("Gerenciamento", "1.00", "1.05", "1.10", "Médio", "Mais controle qualidade"),
        ("Infraestrutura", "1.00", "1.10", "1.25", "Médio", "Fundações especiais"),
        ("Supraestrutura", "1.00", "1.08", "1.15", "Baixo", "Sobrecarga maior"),
        ("Instalações", "1.00", "1.25", "1.60", "Alto", "Automação, CFTV"),
        ("Sist. Especiais", "1.00", "1.40", "2.00", "Alto", "Ar central, gerador"),
        ("Impermeabilização", "1.00", "1.15", "1.30", "Médio", "Mantas importadas"),
        ("Pisos", "1.00", "1.50", "2.20", "Alto", "Porcelanato grande formato"),
        ("Pintura", "1.00", "1.30", "1.60", "Médio", "Textura, efeitos"),
        ("Esquadrias", "1.00", "1.40", "1.90", "Alto", "Alumínio/PVC premium"),
        ("Fachada", "1.00", "1.60", "2.40", "Alto", "ACM, vidro especial"),
        ("Complementares", "1.00", "1.35", "1.80", "Alto", "Acabamentos finos")
    ]
    
    for macro, std, alto, super_alto, impacto, obs in ajustes_padrao:
        ws[f'A{row}'] = macro
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = std
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'B{row}'].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        
        ws[f'C{row}'] = alto
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'D{row}'] = super_alto
        ws[f'D{row}'].border = THIN_BORDER
        ws[f'D{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'E{row}'] = impacto
        ws[f'E{row}'].border = THIN_BORDER
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'E{row}'].font = Font(size=9)
        
        ws[f'F{row}'] = obs
        ws[f'F{row}'].border = THIN_BORDER
        ws[f'F{row}'].font = Font(size=9, italic=True)
        row += 1
    
    print(f"  ✓ ÍNDICES created with {row} rows")
    return ws

def expand_dados_projeto(wb):
    """Expand DADOS_PROJETO with new fields"""
    print("Expanding DADOS_PROJETO...")
    
    ws = wb['DADOS_PROJETO']
    
    # Find the last row with data
    last_row = 19  # Current last row as per inspection
    
    # Add spacing
    row = last_row + 2
    
    # Section header
    ws[f'A{row}'] = "DADOS COMPLEMENTARES"
    ws[f'A{row}'].font = Font(bold=True, size=11, color="34495E")
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # New fields
    new_fields = [
        ("APT - Área Projeção Torre (m²)", 1795),
        ("PPT - Perímetro Projeção Torre (m)", 191.4),
        ("APE - Área Projeção Embasamento (m²)", 2400),
        ("PPE - Perímetro Projeção Embasamento (m)", 208),
        ("NPD - Pav. Tipo Diferenciados", 1),
        ("Área de Lazer (m²)", ""),
        ("BDI - Bonif. Desp. Indiretas", 1.10)
    ]
    
    for label, value in new_fields:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(size=10)
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = value
        apply_input_style(ws[f'B{row}'])
        if isinstance(value, (int, float)) and value != "":
            if "BDI" in label:
                ws[f'B{row}'].number_format = "0.00"
            else:
                ws[f'B{row}'].number_format = "#,##0.00"
        row += 1
    
    print(f"  ✓ DADOS_PROJETO expanded to {row} rows")

def add_briefing_questions(wb):
    """Add new questions to BRIEFING tab"""
    print("Adding questions to BRIEFING...")
    
    ws = wb['BRIEFING']
    
    # Find row 29 (before RESUMO section)
    # We'll insert after row 29
    insert_row = 30
    
    # Insert 5 new rows
    ws.insert_rows(insert_row, 5)
    
    # New questions
    new_questions = [
        ("Gerador?", "Não", "Sist. Especiais"),
        ("Subestação?", "Não", "Sist. Especiais"),
        ("Placas Fotovoltaicas?", "Não", "Sist. Especiais"),
        ("Infra Carro Elétrico?", "Não", "Instalações"),
        ("Pressurização Escada?", "Sim", "Instalações")
    ]
    
    row = insert_row
    for question, default, impact in new_questions:
        ws[f'A{row}'] = question
        ws[f'A{row}'].border = THIN_BORDER
        
        ws[f'B{row}'] = default
        apply_input_style(ws[f'B{row}'])
        
        ws[f'C{row}'] = impact
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'C{row}'].font = Font(size=9, italic=True)
        row += 1
    
    print(f"  ✓ BRIEFING updated with {len(new_questions)} new questions")

def main():
    """Main execution"""
    print("\n" + "="*60)
    print("MERGE COWORK FEATURES INTO PARAMETRIC BUDGET TEMPLATE")
    print("Cartesian Engenharia - Mar 2026")
    print("="*60 + "\n")
    
    # Load calibration data
    print("Loading calibration data...")
    with open('calibration-data.json', 'r') as f:
        calibration_data = json.load(f)
    print(f"  ✓ Loaded {len(calibration_data)} calibrated projects\n")
    
    # Load workbook
    print("Loading template...")
    wb = openpyxl.load_workbook('template-orcamento-parametrico.xlsx')
    print(f"  ✓ Loaded with {len(wb.sheetnames)} sheets\n")
    
    # Execute modifications
    create_painel_tab(wb, calibration_data)
    print()
    
    create_indices_tab(wb)
    print()
    
    expand_dados_projeto(wb)
    print()
    
    add_briefing_questions(wb)
    print()
    
    # Save
    print("Saving modified template...")
    wb.save('template-orcamento-parametrico.xlsx')
    print("  ✓ Saved to template-orcamento-parametrico.xlsx\n")
    
    # Copy to outbound
    print("Copying to outbound...")
    shutil.copy(
        'template-orcamento-parametrico.xlsx',
        str(Path.home() / ".openclaw" / "media" / "outbound" / "template-orcamento-parametrico-v13-merged.xlsx")
    )
    print("  ✓ Copied to outbound/template-orcamento-parametrico-v13-merged.xlsx\n")
    
    # Verification summary
    print("="*60)
    print("VERIFICATION SUMMARY")
    print("="*60)
    wb = openpyxl.load_workbook('template-orcamento-parametrico.xlsx')
    
    print("\nFinal sheet structure:")
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        print(f"  {i}. {sheet_name} - {ws.max_row} rows x {ws.max_column} cols")
    
    print("\nKey formulas verification:")
    
    # PAINEL formulas
    painel = wb['PAINEL']
    print(f"\n  PAINEL!B4 (R$/m²): {painel['B4'].value}")
    print(f"  PAINEL!B5 (Custo Total): {painel['B5'].value}")
    print(f"  PAINEL!B6 (R$/UR): {painel['B6'].value}")
    
    # DADOS_PROJETO new fields
    dados = wb['DADOS_PROJETO']
    print(f"\n  DADOS_PROJETO!A22 (APT label): {dados['A22'].value}")
    print(f"  DADOS_PROJETO!B22 (APT value): {dados['B22'].value}")
    print(f"  DADOS_PROJETO!B28 (BDI value): {dados['B28'].value}")
    
    # BRIEFING new questions
    briefing = wb['BRIEFING']
    print(f"\n  BRIEFING!A30 (Gerador): {briefing['A30'].value}")
    print(f"  BRIEFING!A34 (Pressurização): {briefing['A34'].value}")
    
    wb.close()
    
    print("\n" + "="*60)
    print("✓ MERGE COMPLETED SUCCESSFULLY")
    print("="*60 + "\n")

if __name__ == '__main__':
    main()
