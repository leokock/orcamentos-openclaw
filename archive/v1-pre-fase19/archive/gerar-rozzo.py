#!/usr/bin/env python3.11
"""
Orçamento Paramétrico — Edifício Rozzo (Vitório Demarche)
Rozzo Empreendimentos — Brusque/SC
Data-base: Março/2026
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
import os, json

# ── Constants ──
CUB_BASE = 2752.67       # dez/2023
CUB_ATUAL = 3050.00      # mar/2026 (estimado SC)
CUB_FACTOR = CUB_ATUAL / CUB_BASE  # ~1.108

# ── Project Data ──
PROJECT = {
    "nome": "Edifício Rozzo — Vitório Demarche",
    "codigo": "ROZZO-VD",
    "cidade": "Brusque/SC",
    "incorporadora": "Rozzo Empreendimentos",
    "arquiteto": "Eron Kogikowski (Kogika.Arq)",
    "ac": 14854.30,
    "ur": 115,
    "uc": 3,  # comerciais
    "np": 30,
    "npt": 23,
    "npg": 3,  # G02, G03, G04
    "elev": 3,
    "vag": 140,  # estimado (68 contados em G01/G02, + G03/G04)
    "at": 1169.18,
    "ns": 0,  # sem subsolo
    "prazo": 36,
    "cub": CUB_ATUAL,
    "data_base": "mar/2026",
}

# ── Briefing Answers ──
BRIEFING = {
    "laje": "Cubetas",
    "padrao": "Standard",
    "fundacao": "Hélice Contínua",
    "prazo_meses": 36,
    "contencao": "Não",
    "fachada": "Textura + pintura",
    "gerador": True,
    "pressurizacao": False,
    # Defaults for Standard:
    "esquadria": "Alumínio anodizado",
    "piso": "Porcelanato padrão",
    "vedacao": "Alvenaria",
    "forro": "Gesso liso",
    "mo_fachada": "Empreitada",
    "cobertura_habitavel": "Não",
    "aquecimento": "Gás individual",
    "automacao": "Mínimo",
    "energia": "Sem",
    "lazer": "Completo",  # piscina 111m², academia, 2 gourmets...
    "paisagismo": "Básico",
    "mobiliario": "Básico",
    "regiao": "Interior SC",
    "subestacao": False,
    "fotovoltaicas": False,
    "carro_eletrico": False,
}

# ── Calibration Stats (medians in R$/m² normalized to CUB_BASE) ──
# Source: calibration-stats.json (17 projects)
MEDIANS = {
    "Gerenciamento": 407.07,
    "Mov. Terra": 12.47,
    "Infraestrutura": 230.40,
    "Supraestrutura": 722.66,
    "Alvenaria": 148.70,
    "Impermeabilização": 56.43,
    "Instalações": 366.96,
    "Sist. Especiais": 193.39,
    "Climatização": 0,  # not in calibration stats, estimate
    "Rev. Int. Parede": 171.85,
    "Teto": 68.65,
    "Pisos": 193.69,
    "Pintura": 133.12,
    "Esquadrias": 367.23,
    "Louças e Metais": 0,  # included in Complementares
    "Fachada": 170.21,
    "Complementares": 214.32,
    "Imprevistos": 97.79,
}

# Faixas P10-P90 (approximate from calibration data)
FAIXAS = {
    "Gerenciamento": (200, 550),
    "Mov. Terra": (7, 60),
    "Infraestrutura": (120, 280),
    "Supraestrutura": (450, 950),
    "Alvenaria": (60, 300),
    "Impermeabilização": (36, 86),
    "Instalações": (315, 500),
    "Sist. Especiais": (100, 400),
    "Climatização": (0, 80),
    "Rev. Int. Parede": (100, 240),
    "Teto": (45, 140),
    "Pisos": (125, 350),
    "Pintura": (96, 200),
    "Esquadrias": (200, 700),
    "Louças e Metais": (30, 120),
    "Fachada": (55, 275),
    "Complementares": (100, 500),
    "Imprevistos": (35, 120),
}

# ── Briefing Adjustment Factors ──
# Standard padrão adjustments vs median base (which includes alto/luxo projects)
BRIEFING_FACTORS = {
    "Gerenciamento": 0.90,      # Standard = slightly lower admin
    "Mov. Terra": 1.00,
    "Infraestrutura": 1.00,     # HC = reference
    "Supraestrutura": 0.95,     # Cubetas = reference, but standard slightly less reinforcement
    "Alvenaria": 1.00,
    "Impermeabilização": 0.90,  # Standard waterproofing
    "Instalações": 0.95,        # No pressurization, standard
    "Sist. Especiais": 1.15,    # Gerador adds cost, no pressurization balances
    "Climatização": 0,          # Typically not in paramétrico (separate)
    "Rev. Int. Parede": 0.80,   # Standard finish significantly lower
    "Teto": 0.80,               # Standard ceiling
    "Pisos": 0.75,              # Standard porcelanato vs refined/imported
    "Pintura": 0.90,            # Standard paint
    "Esquadrias": 0.70,         # Standard aluminum anodized (big difference)
    "Louças e Metais": 0,       # Not separate in our calibration
    "Fachada": 0.80,            # Texture only (no ceramic/ACM)
    "Complementares": 0.85,     # Standard complement (basic landscaping)
    "Imprevistos": 1.00,        # Same contingency
}

# ── Styles ──
THIN = Side(style='thin', color='999999')
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='34495E')
SUBHEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='95A5A6')
SECTION_FONT = Font(bold=True, color='FFFFFF', size=10)
INPUT_FILL = PatternFill('solid', fgColor='D6EAF8')
INPUT_FONT = Font(color='000000', size=10)
CALC_FILL = PatternFill('solid', fgColor='E8E8E8')
CALC_FONT = Font(color='000000', size=10)
OK_FILL = PatternFill('solid', fgColor='D5F5E3')
WARN_FILL = PatternFill('solid', fgColor='FEF9E7')
CRIT_FILL = PatternFill('solid', fgColor='FADBD8')
TITLE_FONT = Font(bold=True, color='2C3E50', size=14)
SUBTITLE_FONT = Font(italic=True, color='7F8C8D', size=10)
KPI_FONT = Font(bold=True, color='2C3E50', size=16)
KPI_LABEL = Font(bold=True, color='7F8C8D', size=10)

FMT_BRL = '#,##0.00'
FMT_INT = '#,##0'
FMT_PCT = '0.0%'
FMT_IDX = '#,##0.000'
FMT_MOEDA = 'R$ #,##0.00'
FMT_MOEDA_K = 'R$ #,##0'


def sc(cell, fill=None, font=None, fmt=None, border=True, align=None):
    if fill: cell.fill = fill
    if font: cell.font = font
    if fmt: cell.number_format = fmt
    if border: cell.border = BORDER_THIN
    if align: cell.alignment = align


def calculate_costs():
    """Calculate all macrogroup costs."""
    ac = PROJECT["ac"]
    results = {}
    total = 0

    for mg, median in MEDIANS.items():
        factor = BRIEFING_FACTORS.get(mg, 1.0)
        rsm2_adjusted = median * CUB_FACTOR * factor
        valor = rsm2_adjusted * ac
        results[mg] = {
            "median_base": median,
            "cub_factor": CUB_FACTOR,
            "briefing_factor": factor,
            "rsm2": rsm2_adjusted,
            "valor": valor,
        }
        total += valor

    # Calculate percentages
    for mg in results:
        results[mg]["pct"] = results[mg]["valor"] / total if total > 0 else 0

    return results, total


def create_painel(wb, costs, total):
    """Aba PAINEL — Dashboard executivo."""
    ws = wb.active
    ws.title = 'PAINEL'
    ws.sheet_properties.tabColor = '2C3E50'

    # Column widths
    for col, w in enumerate([3, 22, 18, 3, 22, 18, 3, 22, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws.merge_cells('B2:H2')
    c = ws.cell(row=2, column=2, value='ORÇAMENTO PARAMÉTRICO')
    c.font = Font(bold=True, color='2C3E50', size=18)

    ws.merge_cells('B3:H3')
    c = ws.cell(row=3, column=2, value=f'{PROJECT["nome"]}')
    c.font = Font(bold=True, color='2980B9', size=14)

    ws.merge_cells('B4:H4')
    c = ws.cell(row=4, column=2, value=f'{PROJECT["incorporadora"]} — {PROJECT["cidade"]} | Data-base: {PROJECT["data_base"]}')
    c.font = SUBTITLE_FONT

    # KPIs row
    r = 6
    ac = PROJECT["ac"]
    rsm2 = total / ac
    cub_ratio = rsm2 / CUB_ATUAL

    kpis = [
        ("CUSTO TOTAL", total, FMT_MOEDA_K),
        ("R$/m²", rsm2, FMT_MOEDA),
        ("CUB RATIO", cub_ratio, FMT_IDX),
    ]

    for i, (label, value, fmt) in enumerate(kpis):
        col = 2 + i * 3
        c = ws.cell(row=r, column=col, value=label)
        c.font = KPI_LABEL
        c = ws.cell(row=r+1, column=col, value=value)
        c.font = KPI_FONT
        c.number_format = fmt

    # More KPIs
    r = 9
    kpis2 = [
        ("ÁREA CONSTRUÍDA", ac, '#,##0.00 "m²"'),
        ("UNIDADES", f'{PROJECT["ur"]} res + {PROJECT["uc"]} com = {PROJECT["ur"]+PROJECT["uc"]}', None),
        ("CUSTO/UNIDADE", total / PROJECT["ur"], FMT_MOEDA_K),
    ]
    for i, (label, value, fmt) in enumerate(kpis2):
        col = 2 + i * 3
        c = ws.cell(row=r, column=col, value=label)
        c.font = KPI_LABEL
        c = ws.cell(row=r+1, column=col, value=value)
        c.font = Font(bold=True, color='2C3E50', size=13)
        if fmt:
            c.number_format = fmt

    # Project summary
    r = 13
    ws.merge_cells(f'B{r}:C{r}')
    c = ws.cell(row=r, column=2, value='DADOS DO PROJETO')
    sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))
    sc(ws.cell(row=r, column=3), HEADER_FILL, HEADER_FONT)

    ws.merge_cells(f'E{r}:F{r}')
    c = ws.cell(row=r, column=5, value='BRIEFING')
    sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))
    sc(ws.cell(row=r, column=6), HEADER_FILL, HEADER_FONT)

    data_left = [
        ("AC (m²)", f'{ac:,.2f}'),
        ("Unid. Residenciais", PROJECT["ur"]),
        ("Unid. Comerciais", PROJECT["uc"]),
        ("Pavimentos (NP)", PROJECT["np"]),
        ("Pav. Tipo (NPT)", PROJECT["npt"]),
        ("Pav. Garagem (NPG)", PROJECT["npg"]),
        ("Elevadores", PROJECT["elev"]),
        ("Vagas", PROJECT["vag"]),
        ("Área Terreno (m²)", f'{PROJECT["at"]:,.2f}'),
        ("Prazo (meses)", PROJECT["prazo"]),
        ("CUB (R$/m²)", f'R$ {CUB_ATUAL:,.2f}'),
    ]

    data_right = [
        ("Tipo de Laje", BRIEFING["laje"]),
        ("Fundação", BRIEFING["fundacao"]),
        ("Padrão", BRIEFING["padrao"]),
        ("Contenção", BRIEFING["contencao"]),
        ("Fachada", BRIEFING["fachada"]),
        ("Gerador", "Sim" if BRIEFING["gerador"] else "Não"),
        ("Pressurização", "Sim" if BRIEFING["pressurizacao"] else "Não"),
        ("Lazer", BRIEFING["lazer"]),
        ("Esquadria", BRIEFING["esquadria"]),
        ("Região", BRIEFING["regiao"]),
        ("Data-base", PROJECT["data_base"]),
    ]

    for i, (label, val) in enumerate(data_left):
        ri = r + 1 + i
        c = ws.cell(row=ri, column=2, value=label)
        sc(c, font=Font(bold=True, size=10))
        c = ws.cell(row=ri, column=3, value=val)
        sc(c, INPUT_FILL, INPUT_FONT)

    for i, (label, val) in enumerate(data_right):
        ri = r + 1 + i
        c = ws.cell(row=ri, column=5, value=label)
        sc(c, font=Font(bold=True, size=10))
        c = ws.cell(row=ri, column=6, value=val)
        sc(c, INPUT_FILL, INPUT_FONT)

    # Top 5 macrogroups
    r = 27
    ws.merge_cells(f'B{r}:F{r}')
    c = ws.cell(row=r, column=2, value='TOP 5 MACROGRUPOS (por valor)')
    sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))
    for col in range(3, 7):
        sc(ws.cell(row=r, column=col), HEADER_FILL, HEADER_FONT)

    headers = ['Macrogrupo', 'Valor (R$)', 'R$/m²', '%', 'Status']
    r += 1
    for i, h in enumerate(headers):
        col = 2 + i
        c = ws.cell(row=r, column=col, value=h)
        sc(c, SECTION_FILL, SECTION_FONT, align=Alignment(horizontal='center'))

    # Sort by value
    sorted_costs = sorted(costs.items(), key=lambda x: x[1]["valor"], reverse=True)

    for i, (mg, data) in enumerate(sorted_costs[:5]):
        ri = r + 1 + i
        c = ws.cell(row=ri, column=2, value=mg)
        sc(c, font=Font(size=10))

        c = ws.cell(row=ri, column=3, value=data["valor"])
        sc(c, font=Font(size=10), fmt=FMT_MOEDA_K)

        c = ws.cell(row=ri, column=4, value=data["rsm2"])
        sc(c, font=Font(size=10), fmt=FMT_BRL)

        c = ws.cell(row=ri, column=5, value=data["pct"])
        sc(c, font=Font(size=10), fmt=FMT_PCT)

        fmin, fmax = FAIXAS.get(mg, (0, 99999))
        if fmin <= data["rsm2"] <= fmax:
            status, fill = "✅ OK", OK_FILL
        elif data["rsm2"] > fmax:
            status, fill = "⚠️ Acima", WARN_FILL
        else:
            status, fill = "🔽 Abaixo", WARN_FILL
        c = ws.cell(row=ri, column=6, value=status)
        sc(c, fill, Font(size=10), align=Alignment(horizontal='center'))

    return ws


def create_custos(wb, costs, total):
    """Aba CUSTOS_MACROGRUPO — Detalhamento por macrogrupo."""
    ws = wb.create_sheet('CUSTOS_MACROGRUPO')
    ws.sheet_properties.tabColor = '2980B9'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value='CUSTOS POR MACROGRUPO — ORÇAMENTO PARAMÉTRICO').font = TITLE_FONT

    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1,
        value=f'{PROJECT["nome"]} | AC: {PROJECT["ac"]:,.2f} m² | CUB: R$ {CUB_ATUAL:,.2f} | Base: dez/23 × {CUB_FACTOR:.3f}').font = SUBTITLE_FONT

    headers = ['#', 'Macrogrupo', 'Base (R$/m²)', 'Fator CUB', 'Fator Briefing',
               'R$/m² Ajustado', 'Valor Total (R$)', '%',
               'Faixa Min', 'Faixa Max']
    widths = [5, 30, 16, 12, 14, 16, 20, 10, 14, 14]

    r = 4
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=r, column=i, value=h)
        sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center', wrap_text=True))
        ws.column_dimensions[get_column_letter(i)].width = w

    ac = PROJECT["ac"]
    macros = list(MEDIANS.keys())
    r = 5
    data_start = r

    for i, mg in enumerate(macros):
        data = costs[mg]
        fmin, fmax = FAIXAS.get(mg, (0, 99999))

        c = ws.cell(row=r, column=1, value=i+1)
        sc(c, font=Font(size=10), align=Alignment(horizontal='center'))

        c = ws.cell(row=r, column=2, value=mg)
        sc(c, font=Font(size=10))

        c = ws.cell(row=r, column=3, value=data["median_base"])
        sc(c, CALC_FILL, CALC_FONT, FMT_BRL)

        c = ws.cell(row=r, column=4, value=data["cub_factor"])
        sc(c, CALC_FILL, CALC_FONT, FMT_IDX)

        c = ws.cell(row=r, column=5, value=data["briefing_factor"])
        sc(c, INPUT_FILL, INPUT_FONT, FMT_IDX)

        c = ws.cell(row=r, column=6, value=data["rsm2"])
        sc(c, font=Font(bold=True, size=10), fmt=FMT_BRL)

        c = ws.cell(row=r, column=7, value=data["valor"])
        sc(c, font=Font(bold=True, size=10), fmt=FMT_MOEDA_K)

        c = ws.cell(row=r, column=8, value=data["pct"])
        sc(c, font=Font(size=10), fmt=FMT_PCT)

        c = ws.cell(row=r, column=9, value=fmin)
        sc(c, CALC_FILL, CALC_FONT, FMT_BRL)

        c = ws.cell(row=r, column=10, value=fmax)
        sc(c, CALC_FILL, CALC_FONT, FMT_BRL)

        # Color row based on status
        if data["rsm2"] > 0:
            if fmin <= data["rsm2"] <= fmax:
                fill = OK_FILL
            elif data["rsm2"] > fmax * 1.15 or data["rsm2"] < fmin * 0.85:
                fill = CRIT_FILL
            else:
                fill = WARN_FILL
            ws.cell(row=r, column=6).fill = fill

        r += 1

    data_end = r - 1

    # TOTAL row
    r += 1
    for col in range(1, 11):
        sc(ws.cell(row=r, column=col), HEADER_FILL, HEADER_FONT)

    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value='TOTAL').font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL

    ws.cell(row=r, column=6, value=total/ac).number_format = FMT_BRL
    ws.cell(row=r, column=6).font = Font(bold=True, color='FFFFFF', size=12)
    ws.cell(row=r, column=6).fill = HEADER_FILL

    ws.cell(row=r, column=7, value=total).number_format = FMT_MOEDA_K
    ws.cell(row=r, column=7).font = Font(bold=True, color='FFFFFF', size=12)
    ws.cell(row=r, column=7).fill = HEADER_FILL

    ws.cell(row=r, column=8, value=1.0).number_format = FMT_PCT
    ws.cell(row=r, column=8).font = HEADER_FONT
    ws.cell(row=r, column=8).fill = HEADER_FILL

    # CUB Ratio
    r += 2
    c = ws.cell(row=r, column=2, value='CUB Ratio (R$/m² ÷ CUB)')
    c.font = Font(bold=True, size=11)
    c = ws.cell(row=r, column=6, value=total/ac/CUB_ATUAL)
    c.font = Font(bold=True, color='2980B9', size=14)
    c.number_format = FMT_IDX

    # Interpretation
    r += 1
    ratio = total/ac/CUB_ATUAL
    if ratio < 1.00:
        interp = "🟢 Econômico (abaixo de 1.00)"
    elif ratio < 1.15:
        interp = "🟢 Econômico-Médio (1.00 - 1.15)"
    elif ratio < 1.30:
        interp = "🔵 Médio (1.15 - 1.30)"
    elif ratio < 1.50:
        interp = "🟡 Alto (1.30 - 1.50)"
    else:
        interp = "🔴 Luxo (≥ 1.50)"
    ws.cell(row=r, column=2, value=interp).font = Font(italic=True, size=10)

    ws.freeze_panes = 'C5'
    return ws


def create_briefing_tab(wb):
    """Aba BRIEFING — 25 perguntas respondidas."""
    ws = wb.create_sheet('BRIEFING')
    ws.sheet_properties.tabColor = 'E67E22'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value='BRIEFING DE PROJETO').font = TITLE_FONT

    ws.merge_cells('A2:D2')
    ws.cell(row=2, column=1, value=f'{PROJECT["nome"]}').font = SUBTITLE_FONT

    headers = ['#', 'Pergunta', 'Resposta', 'Impacto']
    r = 4
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))

    questions = [
        ("Q1", "Tipo de Fundação", BRIEFING["fundacao"], "Infraestrutura"),
        ("Q2", "Tipo de Laje", BRIEFING["laje"], "Supraestrutura"),
        ("Q3", "Tem Contenção?", BRIEFING["contencao"], "Infraestrutura"),
        ("Q4", "Nº de Subsolos", str(PROJECT["ns"]), "Mov. Terra + Infra"),
        ("Q5", "Padrão de Acabamento", BRIEFING["padrao"], "Acabamentos + Esquadrias"),
        ("Q6", "Tipo de Esquadria", BRIEFING["esquadria"], "Esquadrias"),
        ("Q7", "Piso Predominante", BRIEFING["piso"], "Pisos"),
        ("Q8", "Vedação Interna", BRIEFING["vedacao"], "Alvenaria"),
        ("Q9", "Tipo de Forro", BRIEFING["forro"], "Teto"),
        ("Q10", "Tipo de Fachada", BRIEFING["fachada"], "Fachada"),
        ("Q11", "MO Fachada", BRIEFING["mo_fachada"], "Fachada"),
        ("Q12", "Cobertura Habitável?", BRIEFING["cobertura_habitavel"], "Complementares"),
        ("Q13", "Aquecimento de Água", BRIEFING["aquecimento"], "Instalações"),
        ("Q14", "Nível de Automação", BRIEFING["automacao"], "Sist. Especiais"),
        ("Q15", "Geração de Energia", BRIEFING["energia"], "Sist. Especiais"),
        ("Q16", "Nível de Lazer", BRIEFING["lazer"], "Complementares"),
        ("Q17", "Paisagismo", BRIEFING["paisagismo"], "Complementares"),
        ("Q18", "Mobiliário Áreas Comuns", BRIEFING["mobiliario"], "Complementares"),
        ("Q19", "Prazo de Obra (meses)", str(BRIEFING["prazo_meses"]), "Gerenciamento"),
        ("Q20", "Região", BRIEFING["regiao"], "Geral"),
        ("Q21", "Gerador?", "Sim" if BRIEFING["gerador"] else "Não", "Sist. Especiais +15%"),
        ("Q22", "Subestação?", "Sim" if BRIEFING["subestacao"] else "Não", "Sist. Especiais +10%"),
        ("Q23", "Placas Fotovoltaicas?", "Sim" if BRIEFING["fotovoltaicas"] else "Não", "Sist. Especiais +10%"),
        ("Q24", "Infra Carro Elétrico?", "Sim" if BRIEFING["carro_eletrico"] else "Não", "Instalações +5%"),
        ("Q25", "Pressurização Escada?", "Sim" if BRIEFING["pressurizacao"] else "Não", "Instalações +8%"),
    ]

    for i, (qid, question, answer, impact) in enumerate(questions):
        ri = r + 1 + i
        ws.cell(row=ri, column=1, value=qid).font = Font(bold=True, size=10)
        sc(ws.cell(row=ri, column=1), align=Alignment(horizontal='center'))
        ws.cell(row=ri, column=2, value=question).font = Font(size=10)
        sc(ws.cell(row=ri, column=2))
        c = ws.cell(row=ri, column=3, value=answer)
        sc(c, INPUT_FILL, Font(bold=True, size=10))
        ws.cell(row=ri, column=4, value=impact).font = Font(italic=True, color='7F8C8D', size=9)
        sc(ws.cell(row=ri, column=4))

    ws.freeze_panes = 'A5'
    return ws


def create_produto(wb, costs, total):
    """Aba PRODUTO — Indicadores do produto imobiliário."""
    ws = wb.create_sheet('PRODUTO')
    ws.sheet_properties.tabColor = '16A085'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25

    ws.merge_cells('B1:E1')
    ws.cell(row=1, column=2, value='INDICADORES DO PRODUTO').font = TITLE_FONT

    ac = PROJECT["ac"]
    ur = PROJECT["ur"]

    indicators = [
        ("Área Total / Unidade (m²/UR)", ac / ur, '#,##0.00 "m²"', "Faixa típica: 90-160 m²/UR"),
        ("Custo / Unidade (R$/UR)", total / ur, FMT_MOEDA_K, "Quanto custa produzir cada apto"),
        ("R$/m² Global", total / ac, FMT_MOEDA, "Custo por m² construído"),
        ("CUB Ratio", total / ac / CUB_ATUAL, FMT_IDX, "Posicionamento vs CUB"),
        ("Vagas / Unidade", PROJECT["vag"] / ur, '0.00', "Faixa típica: 1.0-2.0"),
        ("Elevadores / Unidade", PROJECT["elev"] / ur * 100, '0.0 "%"', "Faixa: 1.5-3.0%"),
        ("Coef. Aproveitamento (AC/AT)", ac / PROJECT["at"], '0.00', f'AC {ac:,.0f} / AT {PROJECT["at"]:,.0f}'),
        ("Pav. Tipo / NP (%)", PROJECT["npt"] / PROJECT["np"] * 100, '0.0"%"', "% do edifício que é tipo"),
        ("Área Média Apto (m²)", 70, '#,##0 "m²"', "~65-75 m² (2 suítes + lavabo)"),
        ("Custo Infra+Supra / Total", (costs["Infraestrutura"]["valor"] + costs["Supraestrutura"]["valor"]) / total, FMT_PCT, "Típico: 30-40%"),
        ("Custo Acabamentos / Total",
         (costs["Rev. Int. Parede"]["valor"] + costs["Teto"]["valor"] +
          costs["Pisos"]["valor"] + costs["Pintura"]["valor"] +
          costs["Esquadrias"]["valor"] + costs["Fachada"]["valor"]) / total,
         FMT_PCT, "Típico: 20-35%"),
    ]

    r = 3
    for i, h in enumerate(['', 'Indicador', 'Valor', 'Unidade', 'Referência'], 1):
        c = ws.cell(row=r, column=i, value=h)
        sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))

    for i, (label, value, fmt, ref) in enumerate(indicators):
        ri = r + 1 + i
        ws.cell(row=ri, column=1, value=i+1).font = Font(size=10)
        ws.cell(row=ri, column=2, value=label).font = Font(bold=True, size=10)
        sc(ws.cell(row=ri, column=2))
        c = ws.cell(row=ri, column=3, value=value)
        sc(c, font=Font(bold=True, color='2C3E50', size=11), fmt=fmt)
        ws.cell(row=ri, column=5, value=ref).font = Font(italic=True, color='7F8C8D', size=9)

    ws.freeze_panes = 'A4'
    return ws


def create_alertas(wb, costs, total):
    """Aba ALERTAS — Semáforo automático."""
    ws = wb.create_sheet('ALERTAS')
    ws.sheet_properties.tabColor = 'E74C3C'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='PAINEL DE ALERTAS').font = TITLE_FONT

    headers = ['#', 'Macrogrupo', 'R$/m²', 'Faixa Min', 'Faixa Max', 'Status']
    r = 3
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))

    ok = warn = crit = below = 0

    for i, (mg, data) in enumerate(costs.items()):
        ri = r + 1 + i
        fmin, fmax = FAIXAS.get(mg, (0, 99999))

        ws.cell(row=ri, column=1, value=i+1).font = Font(size=10)
        ws.cell(row=ri, column=2, value=mg).font = Font(size=10)

        c = ws.cell(row=ri, column=3, value=data["rsm2"])
        sc(c, fmt=FMT_BRL)

        ws.cell(row=ri, column=4, value=fmin).number_format = FMT_BRL
        ws.cell(row=ri, column=5, value=fmax).number_format = FMT_BRL

        if data["rsm2"] == 0:
            status, fill = "— Sem valor", None
        elif fmin <= data["rsm2"] <= fmax:
            status, fill = "✅ Dentro da Faixa", OK_FILL
            ok += 1
        elif data["rsm2"] > fmax:
            status, fill = "⚠️ Acima da Faixa", WARN_FILL
            warn += 1
        else:
            status, fill = "🔽 Abaixo da Faixa", WARN_FILL
            below += 1

        c = ws.cell(row=ri, column=6, value=status)
        if fill:
            c.fill = fill
        sc(c, fill=fill, font=Font(size=10))

    # Summary
    r = r + 1 + len(costs) + 2
    ws.cell(row=r, column=2, value='RESUMO').font = Font(bold=True, size=12, color='2C3E50')
    r += 1
    ws.cell(row=r, column=2, value=f'✅ Dentro da Faixa: {ok}').font = Font(size=10)
    r += 1
    ws.cell(row=r, column=2, value=f'⚠️ Acima da Faixa: {warn}').font = Font(size=10)
    r += 1
    ws.cell(row=r, column=2, value=f'🔽 Abaixo da Faixa: {below}').font = Font(size=10)

    ws.freeze_panes = 'A4'
    return ws


def create_composicao(wb, costs, total):
    """Aba COMPOSIÇÃO — Gráfico-friendly data for pie/bar charts."""
    ws = wb.create_sheet('COMPOSIÇÃO')
    ws.sheet_properties.tabColor = '8E44AD'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16

    ws.merge_cells('A1:D1')
    ws.cell(row=1, column=1, value='COMPOSIÇÃO DE CUSTOS').font = TITLE_FONT

    # Grouped categories for high-level view
    groups = {
        "Estrutura (Infra + Supra)": ["Infraestrutura", "Supraestrutura"],
        "Alvenaria + Vedação": ["Alvenaria"],
        "Instalações": ["Instalações"],
        "Sistemas Especiais": ["Sist. Especiais"],
        "Acabamentos Internos": ["Rev. Int. Parede", "Teto", "Pisos", "Pintura"],
        "Esquadrias": ["Esquadrias"],
        "Fachada": ["Fachada"],
        "Impermeabilização": ["Impermeabilização"],
        "Gerenciamento/CI": ["Gerenciamento"],
        "Complementares": ["Complementares", "Mov. Terra"],
        "Imprevistos": ["Imprevistos"],
    }

    headers = ['Grupo', 'Valor (R$)', '%', 'R$/m²']
    r = 3
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        sc(c, HEADER_FILL, HEADER_FONT, align=Alignment(horizontal='center'))

    ac = PROJECT["ac"]
    for group_name, macros in groups.items():
        r += 1
        group_total = sum(costs[m]["valor"] for m in macros if m in costs)
        ws.cell(row=r, column=1, value=group_name).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=group_total).number_format = FMT_MOEDA_K
        ws.cell(row=r, column=3, value=group_total/total if total > 0 else 0).number_format = FMT_PCT
        ws.cell(row=r, column=4, value=group_total/ac if ac > 0 else 0).number_format = FMT_BRL

    r += 2
    ws.cell(row=r, column=1, value='TOTAL').font = Font(bold=True, size=11)
    ws.cell(row=r, column=2, value=total).number_format = FMT_MOEDA_K
    ws.cell(row=r, column=2).font = Font(bold=True, size=11)

    ws.freeze_panes = 'A4'
    return ws


def create_notas(wb):
    """Aba NOTAS — Premissas e limitações."""
    ws = wb.create_sheet('NOTAS')
    ws.sheet_properties.tabColor = '7F8C8D'

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80

    ws.merge_cells('B1:B1')
    ws.cell(row=1, column=2, value='PREMISSAS E LIMITAÇÕES').font = TITLE_FONT

    notes = [
        "PREMISSAS DO CÁLCULO:",
        f"• Valores base calibrados com 17 projetos reais (Cartesian Engenharia, dez/2023)",
        f"• Atualização monetária via CUB/SC: R$ {CUB_BASE:,.2f} (dez/23) → R$ {CUB_ATUAL:,.2f} (mar/26) = fator {CUB_FACTOR:.3f}",
        f"• Ajustes de briefing aplicados por macrogrupo conforme respostas do cliente",
        f"• Região: Interior SC (fator neutro)",
        "",
        "ESCOPO INCLUÍDO:",
        "• Custos diretos de construção (18 macrogrupos)",
        "• Custos indiretos (gerenciamento, equipe ADM, EPCs, equipamentos)",
        "• Imprevistos (~3% do total)",
        "",
        "ESCOPO NÃO INCLUÍDO:",
        "• Terreno e incorporação",
        "• Projetos e consultorias (normalmente 3-5% adicionais)",
        "• Marketing e vendas",
        "• Taxas cartoriais e licenças",
        "• Ligações definitivas (energia, água, esgoto)",
        "• Mobiliário das unidades",
        "• Climatização (orçamento separado — estimativa: R$ 40-80/m² para split)",
        "• Louças e metais (incluídos em Complementares neste paramétrico)",
        "",
        "LIMITAÇÕES:",
        "• Orçamento paramétrico é estimativa de magnitude (precisão ±15-20%)",
        "• Não substitui orçamento executivo (com quantitativos reais)",
        "• Valores de acabamento podem variar significativamente com especificação exata",
        "• Vagas estimadas (~140) — confirmar com plantas garagem completas",
        "",
        "OBSERVAÇÕES ESPECÍFICAS DO PROJETO:",
        "• Prédio de 30 pavimentos (~99m) — OBRIGATÓRIO gerador (confirmado)",
        "• Pressurização de escada NÃO incluída (informado pelo cliente)",
        "  ⚠️ Verificar norma CBMSC — prédios acima de 30m normalmente exigem",
        "• Padrão Standard com lazer Completo — típico de Brusque/região",
        "• Sem contenção — terreno sem desnível relevante",
        "",
        "PRÓXIMOS PASSOS:",
        "• Validar vagas de garagem (plantas G03/G04)",
        "• Confirmar CUB/SC mar/2026 quando publicado pelo Sinduscon",
        "• Para orçamento executivo: necessário projetos complementares",
        "",
        f"Gerado por: Cartesian Engenharia — {PROJECT['data_base']}",
        f"Base: {17} projetos calibrados | Modelo: Base × CUB × Briefing",
    ]

    for i, note in enumerate(notes):
        ri = 3 + i
        c = ws.cell(row=ri, column=2, value=note)
        if note.endswith(":") or note.startswith("PREMISSAS") or note.startswith("ESCOPO") or note.startswith("LIMITAÇÕES") or note.startswith("OBSERVAÇÕES") or note.startswith("PRÓXIMOS"):
            c.font = Font(bold=True, size=10, color='2C3E50')
        elif note.startswith("  ⚠️"):
            c.font = Font(bold=True, size=10, color='E74C3C')
        else:
            c.font = Font(size=10, color='333333')

    return ws


def main():
    # Calculate
    costs, total = calculate_costs()
    ac = PROJECT["ac"]

    print(f"{'='*60}")
    print(f"ORÇAMENTO PARAMÉTRICO — {PROJECT['nome']}")
    print(f"{'='*60}")
    print(f"AC: {ac:,.2f} m² | CUB: R$ {CUB_ATUAL:,.2f} | Fator CUB: {CUB_FACTOR:.3f}")
    print(f"{'─'*60}")

    for mg, data in costs.items():
        if data["valor"] > 0:
            print(f"  {mg:<25} R$ {data['valor']:>14,.2f}  ({data['rsm2']:>8,.2f} R$/m²)  {data['pct']:>5.1%}")

    print(f"{'─'*60}")
    print(f"  {'TOTAL':<25} R$ {total:>14,.2f}  ({total/ac:>8,.2f} R$/m²)")
    print(f"  CUB Ratio: {total/ac/CUB_ATUAL:.3f}")
    print(f"  Custo/Unidade: R$ {total/PROJECT['ur']:,.2f}")
    print(f"{'='*60}")

    # Generate Excel
    wb = openpyxl.Workbook()

    create_painel(wb, costs, total)
    create_custos(wb, costs, total)
    create_briefing_tab(wb)
    create_produto(wb, costs, total)
    create_composicao(wb, costs, total)
    create_alertas(wb, costs, total)
    create_notas(wb)

    output = os.path.expanduser('~/clawd/orcamento-parametrico/rozzo-vd-parametrico.xlsx')
    wb.save(output)
    print(f"\n✅ Planilha salva: {output}")
    print(f"   Abas: {wb.sheetnames}")
    print(f"   Tamanho: {os.path.getsize(output)/1024:.1f} KB")

    return costs, total


if __name__ == '__main__':
    main()
