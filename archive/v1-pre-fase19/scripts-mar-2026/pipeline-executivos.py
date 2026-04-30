#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline de Processamento de Orçamentos Executivos
Escaneia pasta do Drive, extrai dados, gera manifest para processamento em lotes.

Comandos:
    scan                  Escaneia pasta do Drive, atualiza manifest
    status                Mostra status do manifest
    extract <file>        Extrai dados de um xlsx, output JSON
    extract-all           Extrai dados de todos os pending
    mark-done <name>      Marca projeto como done
    mark-error <name>     Marca projeto como error
    reset <name>          Reseta status para pending
    recalibrate           Recalibra base com todos os projetos done
"""

import json
import os
import sys
import subprocess
from datetime import datetime
from pathlib import Path

# ── Paths ─────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent  # orcamento-parametrico/
WORKSPACE = BASE_DIR.parent  # clawd/
MANIFEST_PATH = BASE_DIR / "executivos-manifest.json"
CALIBRATION_PATH = BASE_DIR / "calibration-data.json"
INDICES_DIR = BASE_DIR / "indices"
DOWNLOAD_DIR = BASE_DIR / "executivos-drive"
DRIVE_JS = WORKSPACE / "google-calendar" / "drive.js"

# Pasta do Drive (Projetos em Andamento > Orçamento_executivo)
DRIVE_ROOT_FOLDER = "1-zXfYpxwpaXHhA9qgazwGhJ2v8hvuFY4"

# Extensões de executivo
EXEC_EXTENSIONS = {'.xlsx', '.xlsb', '.xls'}

# ── Manifest Management ──────────────────────────────────

def load_manifest():
    if MANIFEST_PATH.exists():
        with open(MANIFEST_PATH) as f:
            return json.load(f)
    return {"projects": [], "last_scan": None, "stats": {"total": 0, "pending": 0, "done": 0, "error": 0}}

def save_manifest(manifest):
    manifest["stats"] = {
        "total": len(manifest["projects"]),
        "pending": sum(1 for p in manifest["projects"] if p["status"] == "pending"),
        "processing": sum(1 for p in manifest["projects"] if p["status"] == "processing"),
        "done": sum(1 for p in manifest["projects"] if p["status"] == "done"),
        "error": sum(1 for p in manifest["projects"] if p["status"] == "error"),
        "skipped": sum(1 for p in manifest["projects"] if p["status"] == "skipped"),
    }
    with open(MANIFEST_PATH, 'w') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

def get_calibrated_names():
    """Retorna set de nomes já calibrados"""
    if not CALIBRATION_PATH.exists():
        return set()
    with open(CALIBRATION_PATH) as f:
        data = json.load(f)
    return {p.get("name", "").lower() for p in data}

def get_existing_indices():
    """Retorna set de nomes que já têm arquivo de índices"""
    if not INDICES_DIR.exists():
        return set()
    return {f.stem.replace("-indices", "").lower() for f in INDICES_DIR.glob("*-indices.md")}

# ── Drive Scanning ────────────────────────────────────────

def drive_scan(folder_id):
    """Escaneia pasta do Drive recursivamente via drive.js"""
    result = subprocess.run(
        ["node", str(DRIVE_JS), "scan", folder_id],
        capture_output=True, text=True, cwd=str(WORKSPACE / "google-calendar")
    )
    if result.returncode != 0:
        print(f"❌ Erro ao escanear Drive: {result.stderr}")
        sys.exit(1)
    return json.loads(result.stdout)

def drive_download(file_id, output_path):
    """Baixa arquivo do Drive"""
    result = subprocess.run(
        ["node", str(DRIVE_JS), "download", file_id, str(output_path)],
        capture_output=True, text=True, cwd=str(WORKSPACE / "google-calendar")
    )
    if result.returncode != 0:
        print(f"❌ Erro ao baixar: {result.stderr}")
        return False
    return True

def normalize_name(name):
    """Normaliza nome do projeto para comparação"""
    # Remove prefixo CTN-XXX- e sufixos comuns
    import re
    n = name.lower().strip()
    # Remove extensão
    n = re.sub(r'\.(xlsx|xlsb|xls|pdf|pptx)$', '', n)
    # Remove prefixo CTN-XXX-
    n = re.sub(r'^ctn-[a-z]+-[a-z]+-?\s*', '', n)
    # Remove "apresentação", "orçamento", "executivo", revisões
    n = re.sub(r'(apresenta[çc][aã]o|or[çc]amento|executivo|_executivo)', '', n)
    n = re.sub(r'[-_\s]*r\d+[-_\s]*', '', n)
    n = re.sub(r'[-_\s]+', '-', n).strip('-')
    return n

def slug_name(folder_path):
    """Gera slug a partir do path de pastas (cliente/projeto)"""
    import re
    parts = [p.lower().strip() for p in folder_path if p]
    slug = "-".join(parts)
    slug = re.sub(r'[^a-z0-9-]', '', slug)
    slug = re.sub(r'-+', '-', slug).strip('-')
    return slug

# ── Extraction ────────────────────────────────────────────

def extract_executivo(filepath):
    """Extrai dados principais de um executivo xlsx"""
    import openpyxl
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "file": str(filepath),
        "sheets": wb.sheetnames,
        "ac": None,
        "cub": None,
        "cub_date": None,
        "total": None,
        "rsm2": None,
        "cub_ratio": None,
        "categories": [],
        "raw_data": {},
        "warnings": []
    }
    
    # Procurar aba principal (ORÇAMENTO_EXECUTIVO ou similar)
    main_sheet = None
    for name in wb.sheetnames:
        if 'ORÇAMENTO' in name.upper() or 'ORCAMENTO' in name.upper():
            main_sheet = wb[name]
            break
    
    if not main_sheet:
        main_sheet = wb[wb.sheetnames[0]]
        result["warnings"].append(f"Aba principal não encontrada, usando '{main_sheet.title}'")
    
    # Extrair dados da aba principal
    categories = []
    ac = None
    cub = None
    cub_date = None
    total = None
    
    for row in main_sheet.iter_rows(min_row=1, max_row=min(50, main_sheet.max_row), values_only=False):
        cells = {c.coordinate: c.value for c in row if c.value is not None}
        row_vals = [(c.column, c.value) for c in row if c.value is not None]
        
        if not row_vals:
            continue
            
        first_val = row_vals[0][1] if row_vals else None
        
        # Detectar AC (Área Construída)
        for col, val in row_vals:
            if isinstance(val, str) and 'ÁREA CONSTRUÍDA' in val.upper():
                # AC geralmente está na célula seguinte
                for col2, val2 in row_vals:
                    if isinstance(val2, (int, float)) and val2 > 500:
                        ac = val2
                        break
        
        # Detectar CUB
        if isinstance(first_val, str) and 'CUB' in str(first_val).upper():
            for col, val in row_vals:
                if isinstance(val, (int, float)) and 2000 < val < 5000:
                    cub = val
                if isinstance(val, datetime):
                    cub_date = val.strftime("%Y-%m-%d")
        
        # Detectar categorias (nome + valor + % + R$/m²)
        if isinstance(first_val, str) and len(str(first_val)) > 5:
            name = str(first_val).strip()
            # Pular headers
            if name.upper() in ('ETAPA', 'DESCRIÇÃO', 'ITEM'):
                continue
            
            # Procurar valor numérico grande (> 10000 = provavelmente valor orçado)
            valor = None
            pct = None
            rsm2_cat = None
            
            for col, val in row_vals[1:]:
                if isinstance(val, (int, float)):
                    if val > 50000:  # Valor orçado
                        if valor is None:
                            valor = val
                    elif 0 < val < 1.1:  # Percentual
                        if pct is None:
                            pct = val
                    elif 1 < val < 2000:  # R$/m²
                        if rsm2_cat is None:
                            rsm2_cat = val
            
            if valor is not None and name.upper() != 'TOTAL':
                categories.append({
                    "name": name,
                    "valor": valor,
                    "pct": pct,
                    "rsm2": rsm2_cat
                })
            elif name.upper() == 'TOTAL' and valor:
                total = valor
    
    result["ac"] = ac
    result["cub"] = cub
    result["cub_date"] = cub_date
    result["total"] = total
    result["categories"] = categories
    
    if ac and total:
        result["rsm2"] = round(total / ac, 2)
    if cub and result.get("rsm2"):
        result["cub_ratio"] = round(result["rsm2"] / cub, 2)
    
    # Validações
    if not ac:
        result["warnings"].append("⚠️ AC não encontrada")
    if not cub:
        result["warnings"].append("⚠️ CUB não encontrado")
    if not total:
        result["warnings"].append("⚠️ Total não encontrado")
    if not categories:
        result["warnings"].append("⚠️ Nenhuma categoria encontrada")
    elif len(categories) < 10:
        result["warnings"].append(f"⚠️ Poucas categorias ({len(categories)})")
    
    # Verificar se soma das categorias bate com total
    soma = sum(c["valor"] for c in categories)
    if total and abs(soma - total) / total > 0.05:
        result["warnings"].append(f"⚠️ Soma categorias ({soma:.0f}) difere do total ({total:.0f}) em {abs(soma-total)/total*100:.1f}%")
    
    # Extrair dados complementares de outras abas
    for sheet_name in wb.sheetnames:
        if sheet_name == main_sheet.title:
            continue
        ws = wb[sheet_name]
        # Pegar primeiras 5 linhas pra contexto
        preview = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
            vals = [(c.value, c.coordinate) for c in row if c.value is not None]
            if vals:
                preview.append(vals[:6])
        if preview:
            result["raw_data"][sheet_name] = str(preview)[:500]
    
    return result

# ── Commands ──────────────────────────────────────────────

def cmd_scan():
    """Escaneia Drive e atualiza manifest"""
    print("🔍 Escaneando pasta do Drive...")
    files = drive_scan(DRIVE_ROOT_FOLDER)
    
    calibrated = get_calibrated_names()
    existing_indices = get_existing_indices()
    manifest = load_manifest()
    existing_ids = {p["drive_id"] for p in manifest["projects"]}
    
    # Reconstruir árvore de pastas
    # files tem: id, name, mimeType, depth, isFolder
    folder_stack = []
    new_count = 0
    
    for f in files:
        if f["isFolder"]:
            # Manter stack de pastas por profundidade
            while len(folder_stack) > f["depth"]:
                folder_stack.pop()
            folder_stack.append(f["name"])
            continue
        
        # É arquivo — verificar se é executivo
        ext = os.path.splitext(f["name"])[1].lower()
        if ext not in EXEC_EXTENSIONS:
            continue
        
        # Pular apresentações (geralmente são resumo, não executivo detalhado)
        name_lower = f["name"].lower()
        if "apresenta" in name_lower and "orçamento" not in name_lower:
            continue
        
        if f["id"] in existing_ids:
            continue
        
        # Montar path de pastas
        folder_path = list(folder_stack)
        
        # Gerar nome slug
        project_slug = slug_name(folder_path) if folder_path else normalize_name(f["name"])
        
        # Verificar se já está calibrado
        is_calibrated = any(
            slug_part in calibrated 
            for slug_part in project_slug.split("-") 
            if len(slug_part) > 3
        )
        
        # Verificar se já tem índices
        has_indices = any(
            slug_part in existing_indices
            for slug_part in project_slug.split("-")
            if len(slug_part) > 3
        )
        
        status = "skipped" if (is_calibrated and has_indices) else "pending"
        
        entry = {
            "drive_id": f["id"],
            "filename": f["name"],
            "folder_path": folder_path,
            "project_slug": project_slug,
            "size": f.get("size"),
            "modified": f.get("modifiedTime"),
            "status": status,
            "added_at": datetime.now().isoformat(),
            "processed_at": None,
            "local_path": None,
            "extraction": None,
            "notes": "já na base" if status == "skipped" else None
        }
        
        manifest["projects"].append(entry)
        new_count += 1
        marker = "⏭️" if status == "skipped" else "📄"
        print(f"  {marker} {'/'.join(folder_path)}/{f['name']} → {project_slug} [{status}]")
    
    manifest["last_scan"] = datetime.now().isoformat()
    save_manifest(manifest)
    
    print(f"\n📊 Manifest atualizado:")
    print(f"   Total: {manifest['stats']['total']} | Pending: {manifest['stats']['pending']} | Done: {manifest['stats']['done']} | Skipped: {manifest['stats']['skipped']} | Error: {manifest['stats']['error']}")
    print(f"   Novos encontrados: {new_count}")

def cmd_status():
    """Mostra status do manifest"""
    manifest = load_manifest()
    print(f"📊 Pipeline Status (último scan: {manifest.get('last_scan', 'nunca')})")
    print(f"   Total: {manifest['stats']['total']} | Pending: {manifest['stats']['pending']} | Done: {manifest['stats']['done']} | Skipped: {manifest['stats']['skipped']} | Error: {manifest['stats']['error']}")
    
    for status in ['pending', 'processing', 'error']:
        projects = [p for p in manifest["projects"] if p["status"] == status]
        if projects:
            print(f"\n  [{status.upper()}]:")
            for p in projects:
                path = "/".join(p["folder_path"]) if p["folder_path"] else ""
                print(f"    • {path}/{p['filename']} → {p['project_slug']}")

def cmd_download_pending():
    """Baixa todos os pending que ainda não têm arquivo local"""
    manifest = load_manifest()
    pending = [p for p in manifest["projects"] if p["status"] == "pending" and not p.get("local_path")]
    
    if not pending:
        print("✅ Nenhum arquivo pending pra baixar")
        return
    
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    
    for p in pending:
        output = DOWNLOAD_DIR / f"{p['project_slug']}.xlsx"
        print(f"⬇️  Baixando {p['filename']}...")
        if drive_download(p["drive_id"], output):
            p["local_path"] = str(output)
            print(f"   ✅ Salvo em {output}")
        else:
            p["status"] = "error"
            p["notes"] = "Falha no download"
            print(f"   ❌ Falha")
    
    save_manifest(manifest)

def cmd_extract(slug_or_file):
    """Extrai dados de um executivo"""
    filepath = Path(slug_or_file)
    if not filepath.exists():
        # Tentar achar no download dir
        filepath = DOWNLOAD_DIR / f"{slug_or_file}.xlsx"
    if not filepath.exists():
        print(f"❌ Arquivo não encontrado: {slug_or_file}")
        sys.exit(1)
    
    data = extract_executivo(filepath)
    print(json.dumps(data, indent=2, ensure_ascii=False, default=str))

def cmd_extract_all():
    """Extrai dados de todos os pending, salva no manifest"""
    manifest = load_manifest()
    pending = [p for p in manifest["projects"] if p["status"] == "pending"]
    
    # Primeiro baixar os que faltam
    for p in pending:
        if not p.get("local_path") or not Path(p["local_path"]).exists():
            DOWNLOAD_DIR.mkdir(exist_ok=True)
            output = DOWNLOAD_DIR / f"{p['project_slug']}.xlsx"
            print(f"⬇️  Baixando {p['filename']}...")
            if drive_download(p["drive_id"], output):
                p["local_path"] = str(output)
            else:
                p["status"] = "error"
                p["notes"] = "Falha no download"
                continue
    
    save_manifest(manifest)
    
    # Extrair dados
    for p in [p for p in manifest["projects"] if p["status"] == "pending" and p.get("local_path")]:
        print(f"\n📊 Extraindo {p['project_slug']}...")
        try:
            data = extract_executivo(p["local_path"])
            p["extraction"] = {
                "ac": data["ac"],
                "cub": data["cub"],
                "cub_date": data["cub_date"],
                "total": data["total"],
                "rsm2": data["rsm2"],
                "cub_ratio": data["cub_ratio"],
                "num_categories": len(data["categories"]),
                "categories": data["categories"],
                "sheets": data["sheets"],
                "warnings": data["warnings"],
                "raw_data_keys": list(data.get("raw_data", {}).keys())
            }
            
            if data["warnings"]:
                print(f"   ⚠️  {'; '.join(data['warnings'])}")
            else:
                print(f"   ✅ AC={data['ac']:.0f} | CUB={data['cub']:.2f} | Total={data['total']:.0f} | {len(data['categories'])} categorias")
                
        except Exception as e:
            p["status"] = "error"
            p["notes"] = f"Erro extração: {str(e)}"
            print(f"   ❌ {e}")
    
    save_manifest(manifest)
    print(f"\n📊 Extração completa. Manifest atualizado.")

def cmd_mark(slug, status):
    """Marca projeto com status"""
    manifest = load_manifest()
    found = False
    for p in manifest["projects"]:
        if p["project_slug"] == slug or slug in p["project_slug"]:
            p["status"] = status
            if status == "done":
                p["processed_at"] = datetime.now().isoformat()
            found = True
            print(f"✅ {p['project_slug']} → {status}")
            break
    
    if not found:
        print(f"❌ Projeto '{slug}' não encontrado no manifest")
        sys.exit(1)
    
    save_manifest(manifest)

def cmd_next_batch(batch_size=5):
    """Retorna próximo lote de projetos pending com dados extraídos (JSON)"""
    manifest = load_manifest()
    pending = [p for p in manifest["projects"] if p["status"] == "pending" and p.get("extraction")]
    batch = pending[:batch_size]
    
    if not batch:
        print(json.dumps({"batch": [], "remaining": 0}))
        return
    
    # Marcar como processing
    for p in batch:
        p["status"] = "processing"
    save_manifest(manifest)
    
    output = {
        "batch": [{
            "slug": p["project_slug"],
            "filename": p["filename"],
            "folder_path": p["folder_path"],
            "extraction": p["extraction"],
            "local_path": p["local_path"]
        } for p in batch],
        "remaining": len(pending) - len(batch)
    }
    print(json.dumps(output, indent=2, ensure_ascii=False, default=str))

# ── Main ──────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)
    
    cmd = sys.argv[1]
    
    if cmd == "scan":
        cmd_scan()
    elif cmd == "status":
        cmd_status()
    elif cmd == "download":
        cmd_download_pending()
    elif cmd == "extract":
        if len(sys.argv) < 3:
            print("Usage: pipeline-executivos.py extract <file_or_slug>")
            sys.exit(1)
        cmd_extract(sys.argv[2])
    elif cmd == "extract-all":
        cmd_extract_all()
    elif cmd == "mark-done":
        cmd_mark(sys.argv[2], "done")
    elif cmd == "mark-error":
        cmd_mark(sys.argv[2], "error")
    elif cmd == "reset":
        cmd_mark(sys.argv[2], "pending")
    elif cmd == "next-batch":
        batch_size = int(sys.argv[2]) if len(sys.argv) > 2 else 5
        cmd_next_batch(batch_size)
    else:
        print(f"Comando desconhecido: {cmd}")
        print(__doc__)
        sys.exit(1)
