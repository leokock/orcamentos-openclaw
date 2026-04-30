#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-
"""
Adiciona aba DISTRIBUIÇÃO À planilha studios-bc com detalhamento por tipo de área
(Lazer+Tipos vs Garagem) + ajuste shell delivery
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

# Cores
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
INPUT_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
CALC_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
BOLD = Font(name="Calibri", bold=True, size=11)
NORMAL = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2C3E50")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="2C3E50")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def fmt_brl(ws, cell):
    ws[cell].number_format = '#,##0.00'

def fmt_pct(ws, cell):
    ws[cell].number_format = '0.0%'

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_row(ws, row, cols, fill=None, font=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Dados do projeto
MACROGRUPOS = [
    "Gerenciamento", "Mov. Terra", "Infraestrutura", "Supraestrutura",
    "Alvenaria", "Impermeabilização", "Instalações", "Sist. Especiais",
    "Climatização", "Rev. Int. Parede", "Teto", "Pisos",
    "Pintura", "Esquadrias", "Louças e Metais", "Fachada",
    "Complementares", "Imprevistos"
]

# Custo R$/m² AC por macrogrupo (do script anterior)
custos_m2 = {
    "Gerenciamento": 451.85, "Mov. Terra": 17.69, "Infraestrutura": 218.65,
    "Supraestrutura": 747.74, "Alvenaria": 162.97, "Impermeabilização": 61.86,
    "Instalações": 402.02, "Sist. Especiais": 175.04, "Climatização": 56.04,
    "Rev. Int. Parede": 177.03, "Teto": 67.53, "Pisos": 199.91,
    "Pintura": 142.87, "Esquadrias": 467.84, "Louças e Metais": 28.76,
    "Fachada": 221.33, "Complementares": 157.92, "Imprevistos": 55.44
}

# Peso garagem (fração do custo que incide na garagem vs tipos)
peso_garagem = {
    "Gerenciamento": 0.80, "Mov. Terra": 1.00, "Infraestrutura": 1.00,
    "Supraestrutura": 0.85, "Alvenaria": 0.30, "Impermeabilização": 0.60,
    "Instalações": 0.40, "Sist. Especiais": 0.15, "Climatização": 0.05,
    "Rev. Int. Parede": 0.10, "Teto": 0.05, "Pisos": 0.20,
    "Pintura": 0.30, "Esquadrias": 0.05, "Louças e Metais": 0.00,
    "Fachada": 0.30, "Complementares": 0.20, "Imprevistos": 0.80
}

# Economia shell (% do macrogrupo economizado na fração shell)
economia_shell = {
    "Rev. Int. Parede": 0.60, "Teto": 0.80, "Pisos": 0.80,
    "Pintura": 0.60, "Louças e Metais": 0.70, "Alvenaria": 0.30
}

# Áreas
area_garagem = 6864  # térreo + G1-G3
area_tipos = 13416   # lazer + tipos
ac_total = 20280
area_shell = 8400    # 280 studios × 30m²
frac_shell = area_shell / ac_total

# Abrir planilha existente
xlsx_path = os.path.expanduser("~/clawd/orcamento-parametrico/parametricos/studios-bc-parametrico-v1.xlsx")
wb = load_workbook(xlsx_path)

# Criar aba DISTRIBUIÇÃO
if "DISTRIBUIÇÃO" in wb.sheetnames:
    del wb["DISTRIBUIÇÃO"]

ws = wb.create_sheet("DISTRIBUIÇÃO", 1)  # segunda aba (depois do PAINEL)

# Larguras
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 16
ws.column_dimensions['H'].width = 16

# ===== TÍTULO =====
r = 1
ws.merge_cells('A1:H1')
ws['A1'] = "DISTRIBUIÇÃO DE CUSTOS POR TIPO DE ÁREA"
ws['A1'].font = TITLE_FONT
ws['A1'].alignment = Alignment(horizontal='center')

r = 2
ws.merge_cells('A2:H2')
ws['A2'] = "Studios BC — Viabilidade | Shell Delivery (sem acabamento interno nos studios)"
ws['A2'].font = Font(name="Calibri", italic=True, size=11, color="666666")
ws['A2'].alignment = Alignment(horizontal='center')

# ===== QUADRO RESUMO =====
r = 4
ws.merge_cells(f'A{r}:D{r}')
ws[f'A{r}'] = "QUADRO DE ÁREAS"
ws[f'A{r}'].font = SUBTITLE_FONT

r = 5
headers_qa = ["Setor", "Área (m²)", "% AC", "Observação"]
for i, h in enumerate(headers_qa, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 4)

data_qa = [
    ("Térreo (acesso + garagem)", 1716, 1716/ac_total, "Basicamente garagem"),
    ("G1 a G3 (garagem)", 5148, 5148/ac_total, "Sem subsolo, acima do nível"),
    ("Subtotal Garagem", area_garagem, area_garagem/ac_total, ""),
    ("Lazer + Tipos", area_tipos, area_tipos/ac_total, "280 studios ~30m²"),
    ("AC Total", ac_total, 1.0, ""),
]

for i, (setor, area, pct, obs) in enumerate(data_qa):
    row = r + 1 + i
    ws.cell(row=row, column=1, value=setor)
    ws.cell(row=row, column=2, value=area)
    ws.cell(row=row, column=3, value=pct)
    ws.cell(row=row, column=4, value=obs)
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '0.0%'
    fill = None
    font = NORMAL
    if setor.startswith("Subtotal") or setor == "AC Total":
        fill = CALC_FILL
        font = BOLD
    style_row(ws, row, 4, fill=fill, font=font)

# ===== DETALHAMENTO POR MACROGRUPO =====
r = 13
ws.merge_cells(f'A{r}:H{r}')
ws[f'A{r}'] = "DETALHAMENTO POR MACROGRUPO — ENTREGA COMPLETA (antes do desconto shell)"
ws[f'A{r}'].font = SUBTITLE_FONT

r = 14
headers_det = ["Macrogrupo", "R$/m² AC", "Peso\nGaragem", "R$/m²\nGaragem", "Custo\nGaragem (R$)", "R$/m²\nTipos", "Custo\nTipos (R$)", "% do\nTotal"]
for i, h in enumerate(headers_det, 1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 8)

total_gar = 0
total_tip = 0
total_ac = 0

for i, mg in enumerate(MACROGRUPOS):
    row = r + 1 + i
    c_m2 = custos_m2[mg]
    pg = peso_garagem[mg]
    gar_m2 = c_m2 * pg
    custo_gar = gar_m2 * area_garagem
    
    # Tipos recebe o residual
    custo_total_mg = c_m2 * ac_total
    custo_tip = custo_total_mg - custo_gar
    tip_m2 = custo_tip / area_tipos
    
    total_gar += custo_gar
    total_tip += custo_tip
    total_ac += custo_total_mg
    
    ws.cell(row=row, column=1, value=mg)
    ws.cell(row=row, column=2, value=c_m2)
    ws.cell(row=row, column=3, value=pg)
    ws.cell(row=row, column=4, value=gar_m2)
    ws.cell(row=row, column=5, value=custo_gar)
    ws.cell(row=row, column=6, value=tip_m2)
    ws.cell(row=row, column=7, value=custo_tip)
    ws.cell(row=row, column=8, value=custo_total_mg / total_ac if total_ac > 0 else 0)
    
    for c in [2, 4, 6]:
        ws.cell(row=row, column=c).number_format = '#,##0.00'
    for c in [5, 7]:
        ws.cell(row=row, column=c).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '0%'
    ws.cell(row=row, column=8).number_format = '0.0%'
    
    style_row(ws, row, 8)

# Recalcular % do total
for i, mg in enumerate(MACROGRUPOS):
    row = r + 1 + i
    ws.cell(row=row, column=8, value=custos_m2[mg] * ac_total / total_ac)
    ws.cell(row=row, column=8).number_format = '0.0%'

# Total
row_total = r + 1 + len(MACROGRUPOS)
ws.cell(row=row_total, column=1, value="TOTAL")
ws.cell(row=row_total, column=2, value=sum(custos_m2.values()))
ws.cell(row=row_total, column=4, value=total_gar / area_garagem)
ws.cell(row=row_total, column=5, value=total_gar)
ws.cell(row=row_total, column=6, value=total_tip / area_tipos)
ws.cell(row=row_total, column=7, value=total_tip)
ws.cell(row=row_total, column=8, value=1.0)
for c in [2, 4, 6]:
    ws.cell(row=row_total, column=c).number_format = '#,##0.00'
for c in [5, 7]:
    ws.cell(row=row_total, column=c).number_format = '#,##0'
ws.cell(row=row_total, column=8).number_format = '0.0%'
style_row(ws, row_total, 8, fill=CALC_FILL, font=BOLD)

# ===== AJUSTE SHELL DELIVERY =====
r_shell = row_total + 3
ws.merge_cells(f'A{r_shell}:H{r_shell}')
ws[f'A{r_shell}'] = "AJUSTE SHELL DELIVERY — Desconto por entrega sem acabamento interno nos studios"
ws[f'A{r_shell}'].font = SUBTITLE_FONT

r_sh = r_shell + 1
ws.merge_cells(f'A{r_sh}:H{r_sh}')
ws[f'A{r_sh}'] = f"Área shell: {area_shell:,} m² (280 studios × 30m² = {frac_shell*100:.0f}% da AC) — Sem: layout interno, pisos, forro, pintura interna"
ws[f'A{r_sh}'].font = Font(name="Calibri", italic=True, size=10, color="666666")

r_sh2 = r_sh + 1
headers_sh = ["Macrogrupo", "R$/m² AC\nOriginal", "% Economia\nShell", "Economia\nR$/m² AC", "Justificativa", "", "", ""]
for i, h in enumerate(headers_sh, 1):
    ws.cell(row=r_sh2, column=i, value=h)
style_header(ws, r_sh2, 5)

economia_total = 0
justificativas = {
    "Rev. Int. Parede": "Mantém banheiro + periferia, sem rev. interno apto",
    "Teto": "Sem forro nos studios (mantém banheiro)",
    "Pisos": "Sem pisos nos studios (mantém banheiro)",
    "Pintura": "Sem pintura interna (mantém periferia)",
    "Louças e Metais": "Mantém cubas/louças mínimas banheiro",
    "Alvenaria": "Sem paredes internas de layout"
}

for i, (mg, perc) in enumerate(economia_shell.items()):
    row = r_sh2 + 1 + i
    eco_m2 = custos_m2[mg] * perc * frac_shell
    economia_total += eco_m2
    ws.cell(row=row, column=1, value=mg)
    ws.cell(row=row, column=2, value=custos_m2[mg])
    ws.cell(row=row, column=3, value=perc)
    ws.cell(row=row, column=4, value=eco_m2)
    ws.cell(row=row, column=5, value=justificativas.get(mg, ""))
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3).number_format = '0%'
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    style_row(ws, row, 5, fill=YELLOW_FILL)

row_eco = r_sh2 + 1 + len(economia_shell)
ws.cell(row=row_eco, column=1, value="TOTAL ECONOMIA SHELL")
ws.cell(row=row_eco, column=4, value=economia_total)
ws.cell(row=row_eco, column=4).number_format = '#,##0.00'
style_row(ws, row_eco, 5, fill=CALC_FILL, font=BOLD)

# ===== RESULTADO FINAL =====
r_fin = row_eco + 3
ws.merge_cells(f'A{r_fin}:H{r_fin}')
ws[f'A{r_fin}'] = "RESULTADO FINAL — CUSTO POR TIPO DE ÁREA (COM SHELL)"
ws[f'A{r_fin}'].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
ws[f'A{r_fin}'].fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
ws[f'A{r_fin}'].alignment = Alignment(horizontal='center')
for c in range(2, 9):
    ws.cell(row=r_fin, column=c).fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")

r_f = r_fin + 1
headers_fin = ["Tipo de Área", "Área (m²)", "R$/m²", "Custo Total (R$)", "% do Total", "", "", ""]
for i, h in enumerate(headers_fin, 1):
    ws.cell(row=r_f, column=i, value=h)
style_header(ws, r_f, 5)

# Calcular valores finais com shell
total_m2_shell = sum(custos_m2.values()) - economia_total
custo_total_shell = total_m2_shell * ac_total
custo_gar_final = total_gar  # garagem não muda
custo_tip_final = custo_total_shell - custo_gar_final
gar_m2_final = custo_gar_final / area_garagem
tip_m2_final = custo_tip_final / area_tipos

data_fin = [
    ("Garagem (Térreo + G1-G3)", area_garagem, gar_m2_final, custo_gar_final, custo_gar_final/custo_total_shell),
    ("Lazer + Tipos (shell)", area_tipos, tip_m2_final, custo_tip_final, custo_tip_final/custo_total_shell),
    ("AC TOTAL", ac_total, total_m2_shell, custo_total_shell, 1.0),
]

for i, (tipo, area, rm2, custo, pct) in enumerate(data_fin):
    row = r_f + 1 + i
    ws.cell(row=row, column=1, value=tipo)
    ws.cell(row=row, column=2, value=area)
    ws.cell(row=row, column=3, value=rm2)
    ws.cell(row=row, column=4, value=custo)
    ws.cell(row=row, column=5, value=pct)
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=4).number_format = '#,##0'
    ws.cell(row=row, column=5).number_format = '0.0%'
    fill = GREEN_FILL if i < 2 else CALC_FILL
    font = BOLD if i == 2 else NORMAL
    style_row(ws, row, 5, fill=fill, font=font)

# ===== COMPARAÇÃO COM PREMISSA =====
r_comp = r_f + 1 + len(data_fin) + 2
ws.merge_cells(f'A{r_comp}:H{r_comp}')
ws[f'A{r_comp}'] = "COMPARAÇÃO — PREMISSA CLIENTE vs PARAMÉTRICO"
ws[f'A{r_comp}'].font = SUBTITLE_FONT

r_c = r_comp + 1
headers_comp = ["Tipo de Área", "Premissa\nCliente", "Paramétrico\n(shell)", "Diferença\nR$/m²", "Diferença\n%", "Status", "", ""]
for i, h in enumerate(headers_comp, 1):
    ws.cell(row=r_c, column=i, value=h)
style_header(ws, r_c, 6)

premissas = [
    ("Lazer + Tipos", 4000, tip_m2_final),
    ("Garagem", 2400, gar_m2_final),
]

for i, (tipo, prem, param) in enumerate(premissas):
    row = r_c + 1 + i
    diff = param - prem
    diff_pct = (param / prem) - 1
    
    if diff_pct > 0.10:
        status = "⚠ Premissa ABAIXO do paramétrico"
        fill = RED_FILL
    elif diff_pct < -0.10:
        status = "⚠ Premissa ACIMA do paramétrico"
        fill = YELLOW_FILL
    else:
        status = "✓ Dentro da faixa"
        fill = GREEN_FILL
    
    ws.cell(row=row, column=1, value=tipo)
    ws.cell(row=row, column=2, value=prem)
    ws.cell(row=row, column=3, value=param)
    ws.cell(row=row, column=4, value=diff)
    ws.cell(row=row, column=5, value=diff_pct)
    ws.cell(row=row, column=6, value=status)
    
    for c in [2, 3, 4]:
        ws.cell(row=row, column=c).number_format = '#,##0.00'
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    
    style_row(ws, row, 6, fill=fill)

# Total comparação
row_tc = r_c + 1 + len(premissas)
prem_total = 4000 * area_tipos + 2400 * area_garagem
diff_total = custo_total_shell - prem_total
diff_pct_total = (custo_total_shell / prem_total) - 1

ws.cell(row=row_tc, column=1, value="TOTAL")
ws.cell(row=row_tc, column=2, value=prem_total)
ws.cell(row=row_tc, column=3, value=custo_total_shell)
ws.cell(row=row_tc, column=4, value=diff_total)
ws.cell(row=row_tc, column=5, value=diff_pct_total)
ws.cell(row=row_tc, column=6, value=f"Diferença de {diff_pct_total*100:+.1f}% no total")
ws.cell(row=row_tc, column=2).number_format = '#,##0'
ws.cell(row=row_tc, column=3).number_format = '#,##0'
ws.cell(row=row_tc, column=4).number_format = '#,##0'
ws.cell(row=row_tc, column=5).number_format = '+0.0%;-0.0%'
style_row(ws, row_tc, 6, fill=CALC_FILL, font=BOLD)

# ===== SUGESTÃO DE VALORES =====
r_sug = row_tc + 3
ws.merge_cells(f'A{r_sug}:H{r_sug}')
ws[f'A{r_sug}'] = "SUGESTÃO DE VALORES PARA VIABILIDADE"
ws[f'A{r_sug}'].font = SUBTITLE_FONT

r_s = r_sug + 1
headers_sug = ["Tipo de Área", "Área (m²)", "R$/m²\nSugerido", "Custo Total\nSugerido (R$)", "Observação", "", "", ""]
for i, h in enumerate(headers_sug, 1):
    ws.cell(row=r_s, column=i, value=h)
style_header(ws, r_s, 5)

# Valores sugeridos (arredondados pra viabilidade)
sug_tipos = round(tip_m2_final / 50) * 50  # arredonda pra múltiplo de 50
sug_gar = round(gar_m2_final / 50) * 50

data_sug = [
    ("Lazer + Tipos (shell)", area_tipos, sug_tipos, sug_tipos * area_tipos, "Arredondado do paramétrico"),
    ("Garagem", area_garagem, sug_gar, sug_gar * area_garagem, "Arredondado do paramétrico"),
    ("TOTAL", ac_total, None, sug_tipos * area_tipos + sug_gar * area_garagem, ""),
]

for i, (tipo, area, rm2, custo, obs) in enumerate(data_sug):
    row = r_s + 1 + i
    ws.cell(row=row, column=1, value=tipo)
    ws.cell(row=row, column=2, value=area)
    if rm2: ws.cell(row=row, column=3, value=rm2)
    ws.cell(row=row, column=4, value=custo)
    ws.cell(row=row, column=5, value=obs)
    ws.cell(row=row, column=2).number_format = '#,##0'
    if rm2: ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    fill = GREEN_FILL if i < 2 else CALC_FILL
    font = BOLD if i == 2 else NORMAL
    style_row(ws, row, 5, fill=fill, font=font)

# Total sugerido - calcular R$/m² médio
row_sug_total = r_s + 1 + 2
total_sug = sug_tipos * area_tipos + sug_gar * area_garagem
ws.cell(row=row_sug_total, column=3, value=total_sug / ac_total)
ws.cell(row=row_sug_total, column=3).number_format = '#,##0.00'

# ===== NOTA FINAL =====
r_nota = row_sug_total + 3
ws.merge_cells(f'A{r_nota}:H{r_nota}')
ws[f'A{r_nota}'] = "NOTAS"
ws[f'A{r_nota}'].font = SUBTITLE_FONT

notas = [
    "• Shell delivery: entrega apenas paredes perimetrais externas + paredes banheiro, sem layout interno",
    "• Garagem sem subsolo (G1-G3 acima do nível) — custo menor que garagem subterrânea",
    "• Base: 75 projetos reais calibrados | CUB SC mar/2026: R$ 3.028,45",
    "• Padrão Alto (fachada: tijolo aparente + pele de vidro) | 2 elevadores | Laje cubeta | 36 meses",
    "• Pesos de garagem são estimativas técnicas para viabilidade — orçamento executivo valida",
    f"• Ratio Garagem/Tipos: {gar_m2_final/tip_m2_final*100:.0f}% (referência mercado: 50-65%)",
]

for i, nota in enumerate(notas):
    row = r_nota + 1 + i
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = nota
    ws[f'A{row}'].font = Font(name="Calibri", size=10, color="555555")

# Salva
wb.save(xlsx_path)
print(f"✓ Aba DISTRIBUIÇÃO adicionada com sucesso!")
print(f"  Arquivo: {xlsx_path}")
print(f"\n  Resultado final:")
print(f"    Garagem:     R$ {gar_m2_final:,.2f}/m²")
print(f"    Tipos (shell): R$ {tip_m2_final:,.2f}/m²")
print(f"    Sugestão:    Tipos R$ {sug_tipos:,}/m² | Garagem R$ {sug_gar:,}/m²")
print(f"    Total:       R$ {custo_total_shell:,.2f}")
