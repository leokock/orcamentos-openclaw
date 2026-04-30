#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-
"""
Gerador de Template Dinâmico de Orçamento Paramétrico
Cria planilha Excel com briefing interativo (dropdowns) que atualiza custos automaticamente
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os
import json
from datetime import datetime

# ============================================================================
# DADOS DE CALIBRAÇÃO
# ============================================================================

MACROGRUPOS = [
    "Gerenciamento",
    "Mov. Terra",
    "Infraestrutura",
    "Supraestrutura",
    "Alvenaria",
    "Impermeabilização",
    "Instalações",
    "Sist. Especiais",
    "Climatização",
    "Rev. Int. Parede",
    "Teto",
    "Pisos",
    "Pintura",
    "Esquadrias",
    "Louças e Metais",
    "Fachada",
    "Complementares",
    "Imprevistos"
]

# Medianas calibradas (R$/m², base CUB dez/23 = R$ 2.752,67) — REFERÊNCIA HISTÓRICA
# Recalibrado em 10/mar/2026 com 65 projetos (59→65: +6 novos projetos Viva4 e Amalfi)
MEDIANAS_BASE_DEZ23 = {
    "Gerenciamento": 410.70,
    "Mov. Terra": 16.08,
    "Infraestrutura": 198.74,
    "Supraestrutura": 679.65,
    "Alvenaria": 148.13,
    "Impermeabilização": 56.23,
    "Instalações": 338.34,
    "Sist. Especiais": 162.76,
    "Climatização": 50.94,
    "Rev. Int. Parede": 160.91,
    "Teto": 61.38,
    "Pisos": 181.71,
    "Pintura": 129.86,
    "Esquadrias": 303.74,
    "Louças e Metais": 26.14,
    "Fachada": 128.96,
    "Complementares": 168.87,
    "Imprevistos": 50.39,
}

# Faixas P10-P90 (base CUB dez/23) — REFERÊNCIA HISTÓRICA
# Recalibrado em 11/mar/2026 com 75 projetos (65→75: +10 novos projetos)
FAIXAS_BASE_DEZ23 = {
    "Gerenciamento": (257.68, 768.30),
    "Mov. Terra": (5.75, 75.54),
    "Infraestrutura": (106.84, 312.21),
    "Supraestrutura": (490.87, 843.10),
    "Alvenaria": (105.35, 207.77),
    "Impermeabilização": (35.79, 89.57),
    "Instalações": (256.77, 445.76),
    "Sist. Especiais": (96.38, 240.25),
    "Climatização": (22.03, 116.40),
    "Rev. Int. Parede": (102.61, 227.56),
    "Teto": (34.69, 94.08),
    "Pisos": (133.25, 262.76),
    "Pintura": (91.80, 180.45),
    "Esquadrias": (211.11, 483.43),
    "Louças e Metais": (6.00, 76.56),
    "Fachada": (69.77, 225.41),
    "Complementares": (95.27, 297.73),
    "Imprevistos": (30.09, 103.99),
}

CUB_BASE_HISTORICO = 2752.67

# ============================================================================
# BRIEFING - PERGUNTAS E OPÇÕES
# ============================================================================

BRIEFING = [
    {
        "num": "Q1",
        "pergunta": "Tipo de Fundação",
        "opcoes": ["Hélice Contínua", "Estaca Franki", "Tubulão", "Sapata/Radier", "Estaca raiz"],
        "afeta": ["Infraestrutura"],
        "fatores": {
            "Infraestrutura": {
                "Hélice Contínua": 1.00,
                "Estaca Franki": 0.90,
                "Tubulão": 1.15,
                "Sapata/Radier": 0.75,
                "Estaca raiz": 1.10
            }
        }
    },
    {
        "num": "Q2",
        "pergunta": "Tipo de Laje",
        "opcoes": ["Cubetas", "Protendida", "Maciça", "Mista", "Treliçada"],
        "afeta": ["Supraestrutura"],
        "fatores": {
            "Supraestrutura": {
                "Cubetas": 1.00,
                "Protendida": 1.05,
                "Maciça": 1.10,
                "Mista": 1.03,
                "Treliçada": 0.95
            }
        }
    },
    {
        "num": "Q3",
        "pergunta": "Contenção",
        "opcoes": ["Não", "Cortina de estacas", "Muro de arrimo", "Solo grampeado", "Tirantes"],
        "afeta": ["Infraestrutura"],
        "fatores": {
            "Infraestrutura": {
                "Não": 1.00,
                "Cortina de estacas": 1.40,
                "Muro de arrimo": 1.20,
                "Solo grampeado": 1.30,
                "Tirantes": 1.35
            }
        }
    },
    {
        "num": "Q4",
        "pergunta": "Subsolos",
        "opcoes": ["0", "1", "2", "3+"],
        "afeta": ["Mov. Terra", "Infraestrutura", "Impermeabilização"],
        "fatores": {
            "Mov. Terra": {"0": 1.00, "1": 2.00, "2": 3.50, "3+": 5.00},
            "Infraestrutura": {"0": 1.00, "1": 1.20, "2": 1.40, "3+": 1.60},
            "Impermeabilização": {"0": 1.00, "1": 1.30, "2": 1.60, "3+": 1.90}
        }
    },
    {
        "num": "Q5",
        "pergunta": "Padrão",
        "opcoes": ["Econômico", "Standard", "Alto", "Super Alto", "Luxo"],
        "afeta": ["Rev. Int. Parede", "Teto", "Pisos", "Pintura", "Esquadrias", "Fachada", "Complementares"],
        "fatores": {
            "Rev. Int. Parede": {"Econômico": 0.65, "Standard": 0.80, "Alto": 1.00, "Super Alto": 1.20, "Luxo": 1.45},
            "Teto": {"Econômico": 0.70, "Standard": 0.80, "Alto": 1.00, "Super Alto": 1.20, "Luxo": 1.40},
            "Pisos": {"Econômico": 0.60, "Standard": 0.75, "Alto": 1.00, "Super Alto": 1.30, "Luxo": 1.70},
            "Pintura": {"Econômico": 0.85, "Standard": 0.90, "Alto": 1.00, "Super Alto": 1.10, "Luxo": 1.20},
            "Esquadrias": {"Econômico": 0.55, "Standard": 0.70, "Alto": 1.00, "Super Alto": 1.35, "Luxo": 1.60},
            "Fachada": {"Econômico": 0.80, "Standard": 0.85, "Alto": 1.00, "Super Alto": 1.25, "Luxo": 1.50},
            "Complementares": {"Econômico": 0.70, "Standard": 0.85, "Alto": 1.00, "Super Alto": 1.20, "Luxo": 1.50}
        }
    },
    {
        "num": "Q6",
        "pergunta": "Esquadria",
        "opcoes": ["Alumínio anodizado", "Pintura eletrostática", "PVC", "Alto desempenho"],
        "afeta": ["Esquadrias"],
        "fatores": {
            "Esquadrias": {
                "Alumínio anodizado": 1.00,
                "Pintura eletrostática": 1.15,
                "PVC": 0.90,
                "Alto desempenho": 1.40
            }
        }
    },
    {
        "num": "Q7",
        "pergunta": "Piso",
        "opcoes": ["Porcelanato padrão", "Cerâmica", "Vinílico", "Importado", "Mármore"],
        "afeta": ["Pisos"],
        "fatores": {
            "Pisos": {
                "Porcelanato padrão": 1.00,
                "Cerâmica": 0.75,
                "Vinílico": 0.85,
                "Importado": 1.50,
                "Mármore": 2.00
            }
        }
    },
    {
        "num": "Q8",
        "pergunta": "Vedação",
        "opcoes": ["Alvenaria", "Drywall", "Misto"],
        "afeta": ["Alvenaria"],
        "fatores": {
            "Alvenaria": {
                "Alvenaria": 1.00,
                "Drywall": 1.80,
                "Misto": 1.30
            }
        }
    },
    {
        "num": "Q9",
        "pergunta": "Forro",
        "opcoes": ["Estucamento", "Gesso liso", "Gesso negativo", "Gesso sanca", "Mineral"],
        "afeta": ["Teto"],
        "fatores": {
            "Teto": {
                "Estucamento": 0.80,
                "Gesso liso": 1.00,
                "Gesso negativo": 1.15,
                "Gesso sanca": 1.20,
                "Mineral": 1.10
            }
        }
    },
    {
        "num": "Q10",
        "pergunta": "Fachada",
        "opcoes": ["Textura + pintura", "Cerâmica/Pastilha", "ACM", "Pele de vidro", "Misto"],
        "afeta": ["Fachada"],
        "fatores": {
            "Fachada": {
                "Textura + pintura": 1.00,
                "Cerâmica/Pastilha": 1.40,
                "ACM": 1.80,
                "Pele de vidro": 2.50,
                "Misto": 1.30
            }
        }
    },
    {
        "num": "Q11",
        "pergunta": "MO Fachada",
        "opcoes": ["Equipe própria", "Empreitada"],
        "afeta": ["Fachada"],
        "fatores": {
            "Fachada": {
                "Equipe própria": 1.00,
                "Empreitada": 1.20
            }
        }
    },
    {
        "num": "Q12",
        "pergunta": "Cobertura Habitável",
        "opcoes": ["Não", "Básica", "Completa"],
        "afeta": ["Complementares"],
        "fatores": {
            "Complementares": {
                "Não": 1.00,
                "Básica": 1.10,
                "Completa": 1.25
            }
        }
    },
    {
        "num": "Q13",
        "pergunta": "Aquecimento",
        "opcoes": ["Gás individual", "Central", "Solar", "Bomba calor", "Sem"],
        "afeta": ["Instalações"],
        "fatores": {
            "Instalações": {
                "Gás individual": 1.00,
                "Central": 1.15,
                "Solar": 1.20,
                "Bomba calor": 1.25,
                "Sem": 0.90
            }
        }
    },
    {
        "num": "Q14",
        "pergunta": "Automação",
        "opcoes": ["Mínimo", "Básico", "Completo", "Premium"],
        "afeta": ["Sist. Especiais"],
        "fatores": {
            "Sist. Especiais": {
                "Mínimo": 0.85,
                "Básico": 1.00,
                "Completo": 1.30,
                "Premium": 1.60
            }
        }
    },
    {
        "num": "Q15",
        "pergunta": "Energia",
        "opcoes": ["Sem", "Solar comum", "Solar completo"],
        "afeta": ["Sist. Especiais"],
        "fatores": {
            "Sist. Especiais": {
                "Sem": 1.00,
                "Solar comum": 1.10,
                "Solar completo": 1.20
            }
        }
    },
    {
        "num": "Q16",
        "pergunta": "Lazer",
        "opcoes": ["Básico", "Completo", "Premium"],
        "afeta": ["Complementares"],
        "fatores": {
            "Complementares": {
                "Básico": 0.70,
                "Completo": 1.00,
                "Premium": 1.40
            }
        }
    },
    {
        "num": "Q17",
        "pergunta": "Paisagismo",
        "opcoes": ["Sem", "Básico", "Elaborado", "Premium"],
        "afeta": ["Complementares"],
        "fatores": {
            "Complementares": {
                "Sem": 0.85,
                "Básico": 1.00,
                "Elaborado": 1.20,
                "Premium": 1.50
            }
        }
    },
    {
        "num": "Q18",
        "pergunta": "Mobiliário",
        "opcoes": ["Sem", "Básico", "Completo", "Decorado"],
        "afeta": ["Complementares"],
        "fatores": {
            "Complementares": {
                "Sem": 0.85,
                "Básico": 1.00,
                "Completo": 1.25,
                "Decorado": 1.50
            }
        }
    },
    {
        "num": "Q19",
        "pergunta": "Prazo",
        "opcoes": ["18", "24", "30", "36", "42", "48"],
        "afeta": ["Gerenciamento"],
        "fatores": {
            "Gerenciamento": {
                "18": 0.60,
                "24": 0.75,
                "30": 0.88,
                "36": 1.00,
                "42": 1.15,
                "48": 1.30
            }
        }
    },
    {
        "num": "Q20",
        "pergunta": "Região",
        "opcoes": ["Interior SC", "Litoral SC", "Capital Floripa", "Litoral SP", "Capital SP-RJ"],
        "afeta": ["TODOS"],  # Afeta todos os macrogrupos
        "fatores": {
            "TODOS": {
                "Interior SC": 0.95,
                "Litoral SC": 1.00,
                "Capital Floripa": 1.05,
                "Litoral SP": 1.15,
                "Capital SP-RJ": 1.25
            }
        }
    },
    {
        "num": "Q21",
        "pergunta": "Gerador",
        "opcoes": ["Sim", "Não"],
        "afeta": ["Sist. Especiais"],
        "fatores": {
            "Sist. Especiais": {
                "Sim": 1.15,
                "Não": 1.00
            }
        }
    },
    {
        "num": "Q22",
        "pergunta": "Subestação",
        "opcoes": ["Sim", "Não"],
        "afeta": ["Sist. Especiais"],
        "fatores": {
            "Sist. Especiais": {
                "Sim": 1.10,
                "Não": 1.00
            }
        }
    },
    {
        "num": "Q23",
        "pergunta": "Fotovoltaicas",
        "opcoes": ["Sim", "Não"],
        "afeta": ["Sist. Especiais"],
        "fatores": {
            "Sist. Especiais": {
                "Sim": 1.10,
                "Não": 1.00
            }
        }
    },
    {
        "num": "Q24",
        "pergunta": "Carro Elétrico",
        "opcoes": ["Sim", "Não"],
        "afeta": ["Instalações"],
        "fatores": {
            "Instalações": {
                "Sim": 1.05,
                "Não": 1.00
            }
        }
    },
    {
        "num": "Q25",
        "pergunta": "Pressurização",
        "opcoes": ["Sim", "Não"],
        "afeta": ["Instalações"],
        "fatores": {
            "Instalações": {
                "Sim": 1.08,
                "Não": 1.00
            }
        }
    }
]

# Respostas padrão Rozzo
ROZZO_BRIEFING = {
    "Q1": "Hélice Contínua",
    "Q2": "Cubetas",
    "Q3": "Não",
    "Q4": "0",
    "Q5": "Standard",
    "Q6": "Alumínio anodizado",
    "Q7": "Porcelanato padrão",
    "Q8": "Alvenaria",
    "Q9": "Gesso liso",
    "Q10": "Textura + pintura",
    "Q11": "Empreitada",
    "Q12": "Não",
    "Q13": "Gás individual",
    "Q14": "Mínimo",
    "Q15": "Sem",
    "Q16": "Completo",
    "Q17": "Básico",
    "Q18": "Básico",
    "Q19": "36",
    "Q20": "Interior SC",
    "Q21": "Sim",
    "Q22": "Não",
    "Q23": "Não",
    "Q24": "Não",
    "Q25": "Não"
}

# Dados do Projeto Rozzo
ROZZO_DADOS = {
    "Nome": "Edifício Rozzo — Vitório Demarche",
    "Código": "ROZZO-VD",
    "Cidade": "Brusque/SC",
    "AC": 14854.30,
    "UR": 115,
    "UC": 3,
    "NP": 30,
    "NPT": 23,
    "NPG": 3,
    "ELEV": 3,
    "VAG": 140,
    "AT": 1169.18,
    "NS": 0,
    "Prazo": 36,
    "CUB": 3050.00
}

# ============================================================================
# RECALIBRAÇÃO PARA VALOR PRESENTE
# ============================================================================

def recalibrar_base(cub_atual, calibration_stats_path=None):
    """Recalibra medianas e faixas de dez/2023 para valor presente (CUB Atual)."""
    fator = cub_atual / CUB_BASE_HISTORICO

    medianas = {k: round(v * fator, 2) for k, v in MEDIANAS_BASE_DEZ23.items()}
    faixas = {k: (round(lo * fator, 2), round(hi * fator, 2))
              for k, (lo, hi) in FAIXAS_BASE_DEZ23.items()}

    total_projetos = 58
    if calibration_stats_path and os.path.exists(calibration_stats_path):
        with open(calibration_stats_path, 'r') as f:
            stats = json.load(f)
            total_projetos = stats.get("total_projects", 58)

    MESES_PT = {1:"jan",2:"fev",3:"mar",4:"abr",5:"mai",6:"jun",
                7:"jul",8:"ago",9:"set",10:"out",11:"nov",12:"dez"}
    agora = datetime.now()
    data_base_label = f"{MESES_PT[agora.month]}/{agora.year}"

    return {
        "medianas": medianas,
        "faixas": faixas,
        "fator": fator,
        "total_projetos": total_projetos,
        "data_base_label": data_base_label
    }

# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def criar_estilo_header():
    """Estilo para cabeçalhos"""
    return {
        'font': Font(bold=True, size=11, color="FFFFFF"),
        'fill': PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def criar_estilo_input():
    """Estilo para células de input"""
    return {
        'fill': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        'alignment': Alignment(horizontal="left", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def criar_estilo_calculado():
    """Estilo para células calculadas"""
    return {
        'fill': PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        'alignment': Alignment(horizontal="right", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def aplicar_estilo(cell, estilo_dict):
    """Aplica dicionário de estilo a uma célula"""
    for attr, value in estilo_dict.items():
        setattr(cell, attr, value)

def gerar_formula_ifs(macrogrupo, questao_idx):
    """
    Gera fórmula IF aninhado (compatível com Excel 2016-) para uma célula da matriz FATORES
    
    Args:
        macrogrupo: Nome do macrogrupo (linha)
        questao_idx: Índice da questão (0-24)
    
    Returns:
        String com a fórmula IF aninhado ou "=1" se não houver efeito
    """
    questao = BRIEFING[questao_idx]
    cell_ref = f"BRIEFING!$C${2 + questao_idx}"  # C2 a C26 (row 2 = Q1, row 26 = Q25)
    
    # Verifica se esta questão afeta este macrogrupo
    if "TODOS" in questao["afeta"]:
        # Q20 (Região) afeta todos
        fatores = questao["fatores"]["TODOS"]
    elif macrogrupo not in questao["afeta"]:
        # Não afeta, retorna 1
        return "=1"
    else:
        # Busca fatores específicos
        fatores = questao["fatores"].get(macrogrupo, {})
    
    if not fatores:
        return "=1"
    
    # Constrói fórmula IF aninhado (recursivo)
    # Lógica: IF(cond1, val1, IF(cond2, val2, IF(cond3, val3, fallback)))
    opcoes_list = list(fatores.items())
    
    def construir_if_aninhado(idx):
        """Recursivamente constrói IF aninhado"""
        if idx >= len(opcoes_list):
            # Fallback final
            return "1"
        
        opcao, fator = opcoes_list[idx]
        proximo = construir_if_aninhado(idx + 1)
        return f'IF({cell_ref}="{opcao}",{fator},{proximo})'
    
    formula = f"={construir_if_aninhado(0)}"
    return formula

# ============================================================================
# FUNÇÕES DE CRIAÇÃO DAS ABAS
# ============================================================================

def criar_aba_dados_projeto(wb, dados=None, data_base_label=""):
    """Cria aba DADOS_PROJETO"""
    ws = wb.create_sheet("DADOS_PROJETO")
    
    # Header
    ws['A1'] = "DADOS DO PROJETO"
    ws.merge_cells('A1:B1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    
    # Campos
    campos = [
        ("Nome do Projeto", dados.get("Nome", "") if dados else ""),
        ("Código", dados.get("Código", "") if dados else ""),
        ("Cidade", dados.get("Cidade", "") if dados else ""),
        ("", ""),
        ("Área Construída (m²)", dados.get("AC", "") if dados else ""),
        ("Unidades Residenciais", dados.get("UR", "") if dados else ""),
        ("Unidades Comerciais", dados.get("UC", "") if dados else ""),
        ("Número de Pavimentos", dados.get("NP", "") if dados else ""),
        ("Pavimentos Tipo", dados.get("NPT", "") if dados else ""),
        ("Pavimentos Garagem", dados.get("NPG", "") if dados else ""),
        ("Elevadores", dados.get("ELEV", "") if dados else ""),
        ("Vagas", dados.get("VAG", "") if dados else ""),
        ("Área Terreno (m²)", dados.get("AT", "") if dados else ""),
        ("Número de Subsolos", dados.get("NS", "") if dados else ""),
        ("", ""),
        ("Prazo (meses)", dados.get("Prazo", "") if dados else ""),
        ("", ""),
        ("CUB Atual (R$/m²)", dados.get("CUB", "") if dados else ""),
        (f"CUB Base {data_base_label} (R$/m²)" if data_base_label else "CUB Base (R$/m²)", dados.get("CUB_BASE", "") if dados else ""),
        ("Data-base", dados.get("Data-base", "") if dados else "")
    ]
    
    row = 2
    for label, valor in campos:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = valor
        
        if label:  # Não formatar linhas vazias
            ws[f'A{row}'].font = Font(bold=True)
            if valor != "":
                aplicar_estilo(ws[f'B{row}'], criar_estilo_input())
                # Formato numérico para campos numéricos
                if any(x in label for x in ["Área", "CUB", "Unidades", "Número", "Pavimentos", "Elevadores", "Vagas", "Prazo", "Subsolos"]):
                    ws[f'B{row}'].number_format = '#,##0.00'
        
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    
    # Freeze panes
    ws.freeze_panes = 'A2'

def criar_aba_briefing(wb, respostas=None):
    """Cria aba BRIEFING com dropdowns"""
    ws = wb.create_sheet("BRIEFING")
    ws.sheet_properties.tabColor = "FFA500"  # Laranja
    
    # Header
    headers = ["#", "Pergunta", "Resposta", "Macrogrupos Afetados"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Perguntas
    for idx, q in enumerate(BRIEFING, 0):
        row = idx + 2  # Começa na linha 2
        
        # # (A)
        ws[f'A{row}'] = q["num"]
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        
        # Pergunta (B)
        ws[f'B{row}'] = q["pergunta"]
        ws[f'B{row}'].font = Font(bold=True)
        
        # Resposta (C) - com dropdown
        if respostas and q["num"] in respostas:
            ws[f'C{row}'] = respostas[q["num"]]
        
        aplicar_estilo(ws[f'C{row}'], criar_estilo_input())
        
        # Data Validation
        opcoes_str = ','.join(q["opcoes"])
        dv = DataValidation(type="list", formula1=f'"{opcoes_str}"', allow_blank=False)
        dv.error = 'Selecione uma opção válida'
        dv.errorTitle = 'Entrada Inválida'
        ws.add_data_validation(dv)
        dv.add(ws[f'C{row}'])
        
        # Macrogrupos afetados (D)
        afeta_str = ", ".join(q["afeta"]) if "TODOS" not in q["afeta"] else "Todos os macrogrupos"
        ws[f'D{row}'] = afeta_str
        ws[f'D{row}'].alignment = Alignment(wrap_text=True)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    
    # Freeze panes
    ws.freeze_panes = 'A2'

def criar_aba_fatores(wb):
    """Cria aba FATORES com matriz de fórmulas IF aninhadas"""
    ws = wb.create_sheet("FATORES")
    
    # Header linha 1
    ws['A1'] = "MATRIZ DE FATORES"
    ws.merge_cells('A1:AA1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    
    # Header linha 2 - colunas de perguntas
    ws['A2'] = "Macrogrupo"
    aplicar_estilo(ws['A2'], criar_estilo_header())
    
    for idx, q in enumerate(BRIEFING, 0):
        col = idx + 2  # B=2, C=3, etc
        col_letter = get_column_letter(col)
        cell = ws[f'{col_letter}2']
        cell.value = q["num"]
        aplicar_estilo(cell, criar_estilo_header())
        ws.column_dimensions[col_letter].width = 6
    
    # Última coluna - Fator Briefing Composto
    col_composto = len(BRIEFING) + 2
    col_letter_composto = get_column_letter(col_composto)
    ws[f'{col_letter_composto}2'] = "Fator Briefing"
    aplicar_estilo(ws[f'{col_letter_composto}2'], criar_estilo_header())
    ws.column_dimensions[col_letter_composto].width = 12
    
    # Matriz de fatores
    for mg_idx, macrogrupo in enumerate(MACROGRUPOS, 0):
        row = mg_idx + 3  # Começa na linha 3
        
        # Nome do macrogrupo (A)
        ws[f'A{row}'] = macrogrupo
        ws[f'A{row}'].font = Font(bold=True)
        ws.column_dimensions['A'].width = 20
        
        # Fórmulas IFS para cada questão (B a Z)
        for q_idx in range(25):
            col = q_idx + 2
            col_letter = get_column_letter(col)
            cell = ws[f'{col_letter}{row}']
            
            formula = gerar_formula_ifs(macrogrupo, q_idx)
            cell.value = formula
            cell.number_format = '0.00'
            
            # Formatação condicional por cores (será aplicada manualmente ou via regra)
            # Por enquanto, apenas cinza padrão
            aplicar_estilo(cell, criar_estilo_calculado())
        
        # Fator Briefing Composto (última coluna) = PRODUCT de B a Z
        first_col = get_column_letter(2)  # B
        last_col = get_column_letter(len(BRIEFING) + 1)  # Z ou AA-1
        formula_product = f"=PRODUCT({first_col}{row}:{last_col}{row})"
        
        cell_composto = ws[f'{col_letter_composto}{row}']
        cell_composto.value = formula_product
        cell_composto.number_format = '0.00'
        cell_composto.font = Font(bold=True)
        aplicar_estilo(cell_composto, criar_estilo_calculado())
        cell_composto.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Amarelo claro
    
    # Freeze panes
    ws.freeze_panes = 'B3'

def criar_aba_custos_macrogrupo(wb, recal=None):
    """Cria aba CUSTOS_MACROGRUPO com fórmulas"""
    ws = wb.create_sheet("CUSTOS_MACROGRUPO")
    
    # Header
    headers = ["#", "Macrogrupo", "Base R$/m²", "Fator CUB", "Fator Briefing", 
                "R$/m² Ajustado", "Valor Total", "%", "Faixa Min", "Faixa Max", "Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Dados dos macrogrupos
    for idx, macrogrupo in enumerate(MACROGRUPOS, 0):
        row = idx + 2
        
        # # (A)
        ws[f'A{row}'] = idx + 1
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        
        # Macrogrupo (B)
        ws[f'B{row}'] = macrogrupo
        ws[f'B{row}'].font = Font(bold=True)
        
        # Base R$/m² (C)
        base = recal["medianas"][macrogrupo] if recal else MEDIANAS_BASE_DEZ23[macrogrupo]
        ws[f'C{row}'] = base
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        # Fator CUB (D) = DADOS_PROJETO!B19 / DADOS_PROJETO!B20 (CUB Atual / CUB Base)
        ws[f'D{row}'] = "=DADOS_PROJETO!$B$19/DADOS_PROJETO!$B$20"
        ws[f'D{row}'].number_format = '0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        # Fator Briefing (E) = FATORES última coluna
        col_fator_briefing = get_column_letter(len(BRIEFING) + 2)
        fatores_row = idx + 3  # FATORES começa na linha 3
        ws[f'E{row}'] = f"=FATORES!${col_fator_briefing}${fatores_row}"
        ws[f'E{row}'].number_format = '0.00'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        # R$/m² Ajustado (F) = C * D * E
        ws[f'F{row}'] = f"=C{row}*D{row}*E{row}"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        ws[f'F{row}'].font = Font(bold=True)
        
        # Valor Total (G) = F * AC (DADOS_PROJETO!B6)
        ws[f'G{row}'] = f"=F{row}*DADOS_PROJETO!$B$6"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        ws[f'G{row}'].font = Font(bold=True)
        
        # % (H) = G / SUM(G:G)
        ws[f'H{row}'] = f"=G{row}/SUM($G$2:$G$19)"
        ws[f'H{row}'].number_format = '0.0%'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        # Faixa Min (I)
        faixa = recal["faixas"][macrogrupo] if recal else FAIXAS_BASE_DEZ23[macrogrupo]
        ws[f'I{row}'] = faixa[0]
        ws[f'I{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'I{row}'], criar_estilo_calculado())
        
        # Faixa Max (J)
        ws[f'J{row}'] = faixa[1]
        ws[f'J{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'J{row}'], criar_estilo_calculado())
        
        # Status (K) - Semáforo
        # Verde se F entre I e J, Amarelo se +/-20%, Vermelho caso contrário
        ws[f'K{row}'] = f'=IF(AND(F{row}>=I{row},F{row}<=J{row}),"✓",IF(OR(F{row}<I{row}*0.8,F{row}>J{row}*1.2),"✗","⚠"))'
        ws[f'K{row}'].alignment = Alignment(horizontal="center")
        aplicar_estilo(ws[f'K{row}'], criar_estilo_calculado())
    
    # Linha de totais
    total_row = len(MACROGRUPOS) + 2
    ws[f'A{total_row}'] = ""
    ws[f'B{total_row}'] = "TOTAL"
    ws[f'B{total_row}'].font = Font(bold=True, size=12)
    
    ws[f'G{total_row}'] = f"=SUM(G2:G{total_row-1})"
    ws[f'G{total_row}'].number_format = '#,##0.00'
    ws[f'G{total_row}'].font = Font(bold=True, size=12)
    ws[f'G{total_row}'].fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    ws[f'H{total_row}'] = "100.0%"
    ws[f'H{total_row}'].font = Font(bold=True, size=12)
    ws[f'H{total_row}'].fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 8
    
    # Freeze panes
    ws.freeze_panes = 'C2'

def criar_aba_painel(wb):
    """Cria aba PAINEL com dashboard de KPIs"""
    ws = wb.create_sheet("PAINEL")
    
    # Título
    ws['A1'] = "PAINEL DE CONTROLE"
    ws.merge_cells('A1:D1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # KPIs
    kpis = [
        ("", ""),
        ("Projeto:", "=DADOS_PROJETO!B2"),
        ("Área Construída:", "=DADOS_PROJETO!B6&\" m²\""),
        ("Unidades:", "=DADOS_PROJETO!B7+DADOS_PROJETO!B8"),
        ("", ""),
        ("Custo Total:", "=CUSTOS_MACROGRUPO!G20"),
        ("Custo/m²:", "=CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B6"),
        ("Custo/Unidade:", "=CUSTOS_MACROGRUPO!G20/(DADOS_PROJETO!B7+DADOS_PROJETO!B8)"),
        ("", ""),
        ("CUB Atual:", "=DADOS_PROJETO!B19"),
        ("Fator CUB:", "=DADOS_PROJETO!B19/DADOS_PROJETO!B20"),
        ("", ""),
        ("Top 3 Macrogrupos:", ""),
        ("1º", "=INDEX(CUSTOS_MACROGRUPO!B:B,MATCH(LARGE(CUSTOS_MACROGRUPO!H:H,1),CUSTOS_MACROGRUPO!H:H,0))&\" (\"&TEXT(LARGE(CUSTOS_MACROGRUPO!H:H,1),\"0.0%\")&\")\""),
        ("2º", "=INDEX(CUSTOS_MACROGRUPO!B:B,MATCH(LARGE(CUSTOS_MACROGRUPO!H:H,2),CUSTOS_MACROGRUPO!H:H,0))&\" (\"&TEXT(LARGE(CUSTOS_MACROGRUPO!H:H,2),\"0.0%\")&\")\""),
        ("3º", "=INDEX(CUSTOS_MACROGRUPO!B:B,MATCH(LARGE(CUSTOS_MACROGRUPO!H:H,3),CUSTOS_MACROGRUPO!H:H,0))&\" (\"&TEXT(LARGE(CUSTOS_MACROGRUPO!H:H,3),\"0.0%\")&\")\""),
    ]
    
    for idx, (label, formula) in enumerate(kpis, 3):
        ws[f'A{idx}'] = label
        ws[f'B{idx}'] = formula
        
        if label:
            ws[f'A{idx}'].font = Font(bold=True)
        
        if formula and "=" in formula:
            ws[f'B{idx}'].font = Font(size=11)
            # Formato numérico para valores monetários
            if any(x in label for x in ["Custo", "CUB"]):
                if "/m²" in label or "/Unidade" in label or "CUB" in label:
                    ws[f'B{idx}'].number_format = '#,##0.00'
                elif "Total" in label:
                    ws[f'B{idx}'].number_format = '#,##0.00'
                    ws[f'B{idx}'].font = Font(bold=True, size=13)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    
    # Freeze panes
    ws.freeze_panes = 'A2'

def criar_aba_alertas(wb):
    """Cria aba ALERTAS com semáforo"""
    ws = wb.create_sheet("ALERTAS")
    
    # Título
    ws['A1'] = "ALERTAS E VALIDAÇÕES"
    ws.merge_cells('A1:D1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    
    # Header
    headers = ["Macrogrupo", "R$/m² Calculado", "Faixa Esperada", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(2, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Dados - referências para CUSTOS_MACROGRUPO
    for idx in range(18):
        row = idx + 3
        custos_row = idx + 2
        
        ws[f'A{row}'] = f"=CUSTOS_MACROGRUPO!B{custos_row}"
        ws[f'B{row}'] = f"=CUSTOS_MACROGRUPO!F{custos_row}"
        ws[f'B{row}'].number_format = '#,##0.00'
        
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!I{custos_row}&\" - \"&CUSTOS_MACROGRUPO!J{custos_row}"
        
        ws[f'D{row}'] = f"=CUSTOS_MACROGRUPO!K{custos_row}"
        ws[f'D{row}'].alignment = Alignment(horizontal="center")
        ws[f'D{row}'].font = Font(size=14)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    
    # Freeze panes
    ws.freeze_panes = 'A3'

def criar_aba_notas(wb, recal=None):
    """Cria aba NOTAS com premissas e limitações"""
    ws = wb.create_sheet("NOTAS")
    
    # Título
    ws['A1'] = "PREMISSAS E LIMITAÇÕES"
    ws.merge_cells('A1:B1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    
    notas = [
        "",
        "SOBRE ESTE TEMPLATE:",
        "Este orçamento paramétrico foi desenvolvido pela Cartesian Engenharia para estimativas rápidas de custos de obras residenciais/comerciais.",
        "",
        "BASE DE CALIBRAÇÃO:",
        f"- Mediana calibrada com {recal['total_projetos'] if recal else 18} projetos reais executados entre 2022-2024",
        f"- Valores calibrados a valor presente ({recal['data_base_label']})" if recal and recal['data_base_label'] else "- CUB base: R$ 2.752,67 (dez/2023)",
        "- Faixas P10-P90 representam variação observada nos projetos da base",
        "",
        "COMO USAR:",
        "1. Preencha DADOS_PROJETO com informações do empreendimento",
        "2. Responda as 25 perguntas do BRIEFING (use os dropdowns)",
        "3. Os custos em CUSTOS_MACROGRUPO são calculados automaticamente",
        "4. Verifique ALERTAS para validar se os valores estão dentro das faixas esperadas",
        "5. Use PAINEL para visualização rápida dos principais indicadores",
        "",
        "MODELO DE CÁLCULO:",
        "Valor Final (R$/m²) = Base × Fator CUB × Fator Briefing Composto",
        "",
        "Onde:",
        f"- Base = Mediana calibrada do macrogrupo ({recal['data_base_label']})" if recal and recal['data_base_label'] else "- Base = Mediana calibrada do macrogrupo (dez/2023)",
        "- Fator CUB = CUB Atual / CUB Base (ajuste inflação)",
        "- Fator Briefing Composto = Produto de todos os fatores das 25 perguntas",
        "",
        "LIMITAÇÕES:",
        "- Estimativa preliminar, não substitui orçamento executivo",
        "- Não inclui: terreno, projetos, aprovações, marketing, impostos",
        "- Climatização e Louças/Metais são placeholders (valor 0)",
        "- Faixas podem variar conforme região e características específicas",
        "",
        "ATENÇÃO:",
        "- Semáforo verde: dentro da faixa P10-P90",
        "- Semáforo amarelo: fora da faixa, mas dentro de ±20%",
        "- Semáforo vermelho: >20% fora da faixa (revisar premissas)",
        "",
        "CONTATO:",
        "Cartesian Engenharia",
        "leonardo@cartesianengenharia.com",
        "www.cartesianengenharia.com"
    ]
    
    for idx, nota in enumerate(notas, 2):
        ws[f'A{idx}'] = nota
        
        if nota in ["SOBRE ESTE TEMPLATE:", "BASE DE CALIBRAÇÃO:", "COMO USAR:", 
                    "MODELO DE CÁLCULO:", "LIMITAÇÕES:", "ATENÇÃO:", "CONTATO:"]:
            ws[f'A{idx}'].font = Font(bold=True, size=11)
            ws[f'A{idx}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        ws[f'A{idx}'].alignment = Alignment(wrap_text=True, vertical="top")
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 100
    
    # Freeze panes
    ws.freeze_panes = 'A2'

def criar_aba_indices(wb):
    """Cria aba ÍNDICES com tabelas de referência de índices de consumo"""
    ws = wb.create_sheet("ÍNDICES")
    ws.sheet_properties.tabColor = "F39C12"  # Laranja
    
    # Título
    ws['A1'] = "ÍNDICES DE REFERÊNCIA"
    ws.merge_cells('A1:E1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # ========== SEÇÃO A: Índices Estruturais por Tipo de Laje ==========
    row = 3
    ws[f'A{row}'] = "A. ÍNDICES ESTRUTURAIS POR TIPO DE LAJE"
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    
    row += 1
    headers_laje = ["Parâmetro", "Cubetas", "Protendida", "Maciça", "Treliçada"]
    for col, header in enumerate(headers_laje, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    dados_laje = [
        ("Concreto (m³/m² AC)", 0.22, 0.20, 0.25, 0.19),
        ("Aço (kg/m³)", 90, 110, 85, 80),
        ("Forma (m²/m³)", 10, 9, 11, 12)
    ]
    
    for param, *valores in dados_laje:
        ws[f'A{row}'] = param
        ws[f'A{row}'].font = Font(bold=True)
        for col_idx, val in enumerate(valores, 2):
            cell = ws.cell(row, col_idx, val)
            if "m³/m²" in param or "m²/m³" in param:
                cell.number_format = '0.00'
            else:
                cell.number_format = '0'
            aplicar_estilo(cell, criar_estilo_calculado())
        row += 1
    
    # ========== SEÇÃO B: Índices de Acabamentos ==========
    row += 2
    ws[f'A{row}'] = "B. ÍNDICES DE ACABAMENTOS"
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
    
    row += 1
    headers_acab = ["Item", "Índice (m²/m² AC)", "Referência"]
    for col, header in enumerate(headers_acab, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    dados_acab = [
        ("Chapisco interno", 1.80, "Área de parede interna"),
        ("Reboco interno", 1.80, "Idem"),
        ("Estucamento", 1.50, "Área pintável"),
        ("Cerâmico parede", 0.30, "BWCs + cozinhas"),
        ("Porcelanato parede", 0.15, "BWCs padrão alto"),
        ("Contrapiso", 0.85, "Áreas internas"),
        ("Porcelanato piso", 0.55, "Áreas secas"),
        ("Cerâmico piso", 0.15, "BWCs + A. serviço"),
        ("Rodapé", 0.40, "m/m² AC"),
        ("Forro gesso", 0.70, "Área de teto"),
        ("Pintura interna", 2.20, "Paredes + teto"),
        ("Textura externa", 0.45, "Fachada"),
        ("Chapisco externo", 0.45, "Fachada"),
        ("Reboco externo", 0.45, "Fachada")
    ]
    
    for item, indice, ref in dados_acab:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = indice
        ws[f'B{row}'].number_format = '0.00'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        ws[f'C{row}'] = ref
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # ========== SEÇÃO C: Índices de Instalações ==========
    row += 2
    ws[f'A{row}'] = "C. ÍNDICES DE INSTALAÇÕES"
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    
    row += 1
    headers_inst = ["Disciplina", "R$/m² típico", "% do grupo"]
    for col, header in enumerate(headers_inst, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    dados_inst = [
        ("Hidrossanitárias", 120, 0.33),
        ("Elétricas", 110, 0.30),
        ("Preventivas", 60, 0.16),
        ("Gás", 35, 0.10),
        ("Telecom", 40, 0.11)
    ]
    
    for disc, rsm2, perc in dados_inst:
        ws[f'A{row}'] = disc
        ws[f'B{row}'] = rsm2
        ws[f'B{row}'].number_format = '#,##0'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        ws[f'C{row}'] = perc
        ws[f'C{row}'].number_format = '0%'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        row += 1
    
    # ========== SEÇÃO D: Verbas Paramétricas ==========
    row += 2
    ws[f'A{row}'] = "D. VERBAS PARAMÉTRICAS"
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    
    row += 1
    headers_verbas = ["Item", "R$/m²", "Observação"]
    for col, header in enumerate(headers_verbas, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    dados_verbas = [
        ("Elevadores", "50-80", "Por elevador: R$ 250-400k"),
        ("Gerador", "15-25", "Prédio alto"),
        ("Pressurização", "10-18", "Se aplicável"),
        ("Automação básica", "8-15", "CFTV + interfone")
    ]
    
    for item, rsm2, obs in dados_verbas:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = rsm2
        ws[f'C{row}'] = obs
        ws[f'C{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Freeze panes
    ws.freeze_panes = 'A2'

def criar_aba_estrutural(wb):
    """Cria aba ESTRUTURAL com valores paramétricos conectados ao briefing"""
    ws = wb.create_sheet("ESTRUTURAL")
    ws.sheet_properties.tabColor = "E67E22"
    
    # Título
    ws['A1'] = "DETALHAMENTO ESTRUTURAL"
    ws.merge_cells('A1:H1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # ========== SEÇÃO CONCRETO (45% da Supraestrutura) ==========
    row = 3
    ws[f'A{row}'] = "CONCRETO (45% da Supraestrutura)"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    
    row += 1
    headers_conc = ["Elemento", "Peso %", "Qtd (m³)", "Un", "PU (R$/m³)", "Total (R$)", "R$/m²", "Índice"]
    for col, header in enumerate(headers_conc, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Dados concreto: Pilares, Vigas, Lajes, Escadas
    row += 1
    conc_items = [
        ("Pilares", 0.25, 0.055),
        ("Vigas", 0.30, 0.066),
        ("Lajes", 0.35, 0.077),
        ("Escadas", 0.10, 0.022)
    ]
    
    for item, peso, indice in conc_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Qtd = Indice × AC
        ws[f'C{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = "m³"
        
        # Total = CUSTOS_MACROGRUPO!G5 × 0.45 × Peso
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$5*0.45*B{row}"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        # PU = Total / Qtd
        ws[f'E{row}'] = f"=IF(C{row}>0,F{row}/C{row},0)"
        ws[f'E{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        # R$/m²
        ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        # Índice
        ws[f'H{row}'] = indice
        ws[f'H{row}'].number_format = '0.000'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Concreto
    subtotal_conc_row = row
    ws[f'A{row}'] = "TOTAL CONCRETO"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-4}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-4}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'H{row}'].number_format = '0.000'
    ws[f'H{row}'].font = Font(bold=True)
    
    # ========== SEÇÃO AÇO (30% da Supraestrutura) ==========
    row += 2
    ws[f'A{row}'] = "AÇO (30% da Supraestrutura)"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
    
    row += 1
    headers_aco = ["Elemento", "Peso %", "Peso (kg)", "Un", "Taxa (kg/m³)", "Total (R$)", "R$/m²", ""]
    for col, header in enumerate(headers_aco, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    aco_items = [
        ("Pilares", 0.25, 95),
        ("Vigas", 0.30, 90),
        ("Lajes", 0.35, 88),
        ("Escadas", 0.10, 85)
    ]
    
    for item, peso, taxa in aco_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Peso (kg) = Taxa × Volume de concreto do item
        # Busca o volume na seção concreto acima
        volume_ref_map = {
            "Pilares": subtotal_conc_row - 4,
            "Vigas": subtotal_conc_row - 3,
            "Lajes": subtotal_conc_row - 2,
            "Escadas": subtotal_conc_row - 1
        }
        volume_ref = volume_ref_map[item]
        ws[f'C{row}'] = f"={taxa}*C{volume_ref}"
        ws[f'C{row}'].number_format = '#,##0.0'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = "kg"
        
        ws[f'E{row}'] = taxa
        ws[f'E{row}'].number_format = '0'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        # Total = CUSTOS_MACROGRUPO!G5 × 0.30 × Peso
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$5*0.30*B{row}"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        # R$/m²
        ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Aço
    subtotal_aco_row = row
    ws[f'A{row}'] = "TOTAL AÇO"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-4}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.0'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-4}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    
    # ========== SEÇÃO FORMA (25% da Supraestrutura) ==========
    row += 2
    ws[f'A{row}'] = "FORMA (25% da Supraestrutura)"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    
    row += 1
    headers_forma = ["Tipo", "Peso %", "Área (m²)", "Un", "", "Total (R$)", "R$/m²", ""]
    for col, header in enumerate(headers_forma, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    forma_items = [
        ("Madeira/compensado pilares", 0.30),
        ("Madeira/compensado vigas", 0.35),
        ("Cubetas/forma lajes", 0.35)
    ]
    
    for tipo, peso in forma_items:
        ws[f'A{row}'] = tipo
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Área aproximada (10 m²/m³ de concreto total)
        ws[f'C{row}'] = f"=10*C{subtotal_conc_row}*B{row}"
        ws[f'C{row}'].number_format = '#,##0.0'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = "m²"
        
        # Total = CUSTOS_MACROGRUPO!G5 × 0.25 × Peso
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$5*0.25*B{row}"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        # R$/m²
        ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Forma
    subtotal_forma_row = row
    ws[f'A{row}'] = "TOTAL FORMA"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-3}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.0'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-3}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    
    # ========== SEÇÃO FUNDAÇÃO (85% da Infraestrutura) ==========
    row += 2
    ws[f'A{row}'] = "FUNDAÇÃO (85% da Infraestrutura)"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="5D4037", end_color="5D4037", fill_type="solid")
    
    row += 1
    headers_fund = ["Item", "Peso %", "", "", "", "Total (R$)", "R$/m²", ""]
    for col, header in enumerate(headers_fund, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    fund_items = [
        ("Estacas (HC)", 0.60),
        ("Blocos de coroamento", 0.25),
        ("Vigas baldrame", 0.15)
    ]
    
    for item, peso in fund_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Total = CUSTOS_MACROGRUPO!G4 × 0.85 × Peso
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$4*0.85*B{row}"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        # R$/m²
        ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Fundação
    ws[f'A{row}'] = "TOTAL FUNDAÇÃO"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-3}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    
    # ========== SEÇÃO CONTENÇÃO (15% da Infraestrutura, se houver) ==========
    row += 2
    ws[f'A{row}'] = "CONTENÇÃO (15% da Infraestrutura, se houver)"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
    
    row += 1
    ws[f'A{row}'] = "Contenção"
    
    # Total = IF(BRIEFING Q3 != "Não", CUSTOS_MACROGRUPO!G4 × 0.15, 0)
    ws[f'F{row}'] = f'=IF(BRIEFING!$C$4<>"Não",CUSTOS_MACROGRUPO!$G$4*0.15,0)'
    ws[f'F{row}'].number_format = '#,##0.00'
    aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
    
    ws[f'G{row}'] = f"=F{row}/DADOS_PROJETO!$B$6"
    ws[f'G{row}'].number_format = '#,##0.00'
    aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
    
    # ========== ÍNDICES CALCULADOS ==========
    row += 2
    ws[f'A{row}'] = "ÍNDICES CALCULADOS"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    row += 1
    ws[f'A{row}'] = "Concreto/AC (m³/m²)"
    ws[f'F{row}'] = f"=C{subtotal_conc_row}/DADOS_PROJETO!$B$6"
    ws[f'F{row}'].number_format = '0.000'
    ws[f'F{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Taxa Aço Global (kg/m³)"
    ws[f'F{row}'] = f"=C{subtotal_aco_row}/C{subtotal_conc_row}"
    ws[f'F{row}'].number_format = '0.0'
    ws[f'F{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Forma/AC (m²/m²)"
    ws[f'F{row}'] = f"=C{subtotal_forma_row}/DADOS_PROJETO!$B$6"
    ws[f'F{row}'].number_format = '0.000'
    ws[f'F{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    for i, w in enumerate([28, 12, 14, 8, 14, 16, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = 'A4'

def criar_aba_instalacoes(wb):
    """Cria aba INSTALACOES com valores paramétricos conectados ao briefing"""
    ws = wb.create_sheet("INSTALACOES")
    ws.sheet_properties.tabColor = "8E44AD"
    
    # Título
    ws['A1'] = "DETALHAMENTO DE INSTALAÇÕES"
    ws.merge_cells('A1:G1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # ========== INSTALAÇÕES (CUSTOS_MACROGRUPO!G8) ==========
    row = 3
    ws[f'A{row}'] = "INSTALAÇÕES"
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    
    row += 1
    headers = ["Disciplina", "Peso %", "Valor (R$)", "R$/m²", "MO 55% (R$)", "Material 45% (R$)", ""]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    inst_items = [
        ("Hidrossanitárias", 0.33),
        ("Elétricas", 0.30),
        ("Preventivas", 0.16),
        ("Gás", 0.10),
        ("Telecom", 0.11)
    ]
    
    for disc, peso in inst_items:
        ws[f'A{row}'] = disc
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Valor = CUSTOS_MACROGRUPO!G8 × Peso
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$8*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        # R$/m² = Valor / AC
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        # MO 55%
        ws[f'E{row}'] = f"=C{row}*0.55"
        ws[f'E{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        # Material 45%
        ws[f'F{row}'] = f"=C{row}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Instalações
    subtotal_inst_row = row
    ws[f'A{row}'] = "TOTAL INSTALAÇÕES"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-5}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'] = f"=SUM(E{row-5}:E{row-1})"
    ws[f'E{row}'].number_format = '#,##0.00'
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-5}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    
    # ========== SISTEMAS ESPECIAIS (CUSTOS_MACROGRUPO!G9) ==========
    row += 2
    ws[f'A{row}'] = "SISTEMAS ESPECIAIS"
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    
    row += 1
    headers_se = ["Item", "Peso %", "Valor (R$)", "R$/m²", "", "", ""]
    for col, header in enumerate(headers_se, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    se_items = [
        ("Elevadores", 0.40),
        ("Gerador", 0.20),
        ("Automação/CFTV", 0.20),
        ("Pressuriz./Outros", 0.20)
    ]
    
    for item, peso in se_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Valor = CUSTOS_MACROGRUPO!G9 × Peso
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$9*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        # R$/m²
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Sistemas Especiais
    ws[f'A{row}'] = "TOTAL SIST. ESPECIAIS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-4}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    for i, w in enumerate([30, 12, 16, 14, 16, 16, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = 'A4'

def criar_aba_acabamentos(wb):
    """Cria aba ACABAMENTOS com valores paramétricos conectados ao briefing"""
    ws = wb.create_sheet("ACABAMENTOS")
    ws.sheet_properties.tabColor = "16A085"
    
    # Título
    ws['A1'] = "DETALHAMENTO DE ACABAMENTOS"
    ws.merge_cells('A1:H1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # ========== REVESTIMENTOS DE PAREDE (CUSTOS_MACROGRUPO!G11) ==========
    row = 3
    ws[f'A{row}'] = "REVESTIMENTOS DE PAREDE"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
    
    row += 1
    headers = ["Item", "Qtd", "Un", "PU Material (R$)", "PU MO (R$)", "Total Material (R$)", "Total MO (R$)", "Total (R$)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    rev_items = [
        ("Chapisco interno", 0.12, 1.80),
        ("Reboco interno", 0.35, 1.80),
        ("Estucamento", 0.25, 1.50),
        ("Cerâmico parede", 0.15, 0.30),
        ("Porcelanato parede", 0.13, 0.15)
    ]
    
    for item, peso, indice in rev_items:
        ws[f'A{row}'] = item
        
        # Qtd = Índice × AC
        ws[f'B{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
        ws[f'B{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = "m²"
        
        # Total Material = CUSTOS_MACROGRUPO!G11 × Peso × 0.45
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$11*{peso}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        # Total MO = CUSTOS_MACROGRUPO!G11 × Peso × 0.55
        ws[f'G{row}'] = f"=CUSTOS_MACROGRUPO!$G$11*{peso}*0.55"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        # Total = Material + MO
        ws[f'H{row}'] = f"=F{row}+G{row}"
        ws[f'H{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        # PU Material = Total Material / Qtd
        ws[f'D{row}'] = f"=IF(B{row}>0,F{row}/B{row},0)"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        # PU MO = Total MO / Qtd
        ws[f'E{row}'] = f"=IF(B{row}>0,G{row}/B{row},0)"
        ws[f'E{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Rev. Parede
    subtotal_rev_row = row
    ws[f'A{row}'] = "TOTAL REV. PAREDE"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-5}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G{row-5}:G{row-1})"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=F{row}+G{row}"
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # ========== TETO (CUSTOS_MACROGRUPO!G12) ==========
    row += 2
    ws[f'A{row}'] = "TETO"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    teto_items = [
        ("Forro gesso acartonado", 0.55, 0.50),
        ("Forro gesso UR", 0.30, 0.20),
        ("Negativo/Tabica", 0.15, 0.10)
    ]
    
    for item, peso, indice in teto_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
        ws[f'B{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = "m²" if "Tabica" not in item else "m"
        
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$12*{peso}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        ws[f'G{row}'] = f"=CUSTOS_MACROGRUPO!$G$12*{peso}*0.55"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        ws[f'H{row}'] = f"=F{row}+G{row}"
        ws[f'H{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=IF(B{row}>0,F{row}/B{row},0)"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        ws[f'E{row}'] = f"=IF(B{row}>0,G{row}/B{row},0)"
        ws[f'E{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Teto
    ws[f'A{row}'] = "TOTAL TETO"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-3}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G{row-3}:G{row-1})"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=F{row}+G{row}"
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # ========== PISOS (CUSTOS_MACROGRUPO!G13) ==========
    row += 2
    ws[f'A{row}'] = "PISOS"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    pisos_items = [
        ("Contrapiso", 0.20, 0.85),
        ("Porcelanato piso", 0.45, 0.55),
        ("Cerâmico piso", 0.10, 0.15),
        ("Rodapé", 0.10, 0.40),
        ("Soleira/peitoril", 0.05, 0.08),
        ("Outros", 0.10, 0)
    ]
    
    for item, peso, indice in pisos_items:
        ws[f'A{row}'] = item
        
        if indice > 0:
            ws[f'B{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
            ws[f'B{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = "m²" if "Rodapé" not in item else "m"
        
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$13*{peso}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        ws[f'G{row}'] = f"=CUSTOS_MACROGRUPO!$G$13*{peso}*0.55"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        ws[f'H{row}'] = f"=F{row}+G{row}"
        ws[f'H{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        if indice > 0:
            ws[f'D{row}'] = f"=IF(B{row}>0,F{row}/B{row},0)"
            ws[f'D{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
            
            ws[f'E{row}'] = f"=IF(B{row}>0,G{row}/B{row},0)"
            ws[f'E{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Pisos
    ws[f'A{row}'] = "TOTAL PISOS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-6}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G{row-6}:G{row-1})"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=F{row}+G{row}"
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # ========== PINTURA (CUSTOS_MACROGRUPO!G14) ==========
    row += 2
    ws[f'A{row}'] = "PINTURA"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    pintura_items = [
        ("Pintura parede (PVA/acrílica)", 0.55, 2.20),
        ("Pintura teto", 0.25, 0.70),
        ("Textura/Grafiato", 0.10, 0.15),
        ("Verniz/Epóxi", 0.10, 0.05)
    ]
    
    for item, peso, indice in pintura_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
        ws[f'B{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = "m²"
        
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$14*{peso}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        ws[f'G{row}'] = f"=CUSTOS_MACROGRUPO!$G$14*{peso}*0.55"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        ws[f'H{row}'] = f"=F{row}+G{row}"
        ws[f'H{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=IF(B{row}>0,F{row}/B{row},0)"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        ws[f'E{row}'] = f"=IF(B{row}>0,G{row}/B{row},0)"
        ws[f'E{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Pintura
    ws[f'A{row}'] = "TOTAL PINTURA"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-4}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G{row-4}:G{row-1})"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=F{row}+G{row}"
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # ========== ESQUADRIAS (CUSTOS_MACROGRUPO!G15) ==========
    row += 2
    ws[f'A{row}'] = "ESQUADRIAS"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    esq_items = [
        ("Janelas alumínio", 0.35, 0.12),
        ("Portas internas", 0.25, 0.08),
        ("Porta entrada", 0.10, 0),
        ("Vidros", 0.20, 0.15),
        ("Guarda-corpo", 0.10, 0.03)
    ]
    
    for item, peso, indice in esq_items:
        ws[f'A{row}'] = item
        
        if indice > 0:
            ws[f'B{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
            ws[f'B{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = "m²" if "Porta" not in item else "un"
        
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$15*{peso}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        ws[f'G{row}'] = f"=CUSTOS_MACROGRUPO!$G$15*{peso}*0.55"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        ws[f'H{row}'] = f"=F{row}+G{row}"
        ws[f'H{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        if indice > 0:
            ws[f'D{row}'] = f"=IF(B{row}>0,F{row}/B{row},0)"
            ws[f'D{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
            
            ws[f'E{row}'] = f"=IF(B{row}>0,G{row}/B{row},0)"
            ws[f'E{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Esquadrias
    ws[f'A{row}'] = "TOTAL ESQUADRIAS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-5}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G{row-5}:G{row-1})"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=F{row}+G{row}"
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # ========== FACHADA (CUSTOS_MACROGRUPO!G17) ==========
    row += 2
    ws[f'A{row}'] = "FACHADA"
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    fachada_items = [
        ("Chapisco externo", 0.10, 0.45),
        ("Reboco externo", 0.20, 0.45),
        ("Textura/acabamento", 0.40, 0.45),
        ("Detalhes/frisos", 0.15, 0),
        ("Selantes/juntas", 0.15, 0)
    ]
    
    for item, peso, indice in fachada_items:
        ws[f'A{row}'] = item
        
        if indice > 0:
            ws[f'B{row}'] = f"={indice}*DADOS_PROJETO!$B$6"
            ws[f'B{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = "m²"
        
        ws[f'F{row}'] = f"=CUSTOS_MACROGRUPO!$G$17*{peso}*0.45"
        ws[f'F{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'F{row}'], criar_estilo_calculado())
        
        ws[f'G{row}'] = f"=CUSTOS_MACROGRUPO!$G$17*{peso}*0.55"
        ws[f'G{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'G{row}'], criar_estilo_calculado())
        
        ws[f'H{row}'] = f"=F{row}+G{row}"
        ws[f'H{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'H{row}'], criar_estilo_calculado())
        
        if indice > 0:
            ws[f'D{row}'] = f"=IF(B{row}>0,F{row}/B{row},0)"
            ws[f'D{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
            
            ws[f'E{row}'] = f"=IF(B{row}>0,G{row}/B{row},0)"
            ws[f'E{row}'].number_format = '#,##0.00'
            aplicar_estilo(ws[f'E{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Fachada
    ws[f'A{row}'] = "TOTAL FACHADA"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = f"=SUM(F{row-5}:F{row-1})"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'G{row}'] = f"=SUM(G{row-5}:G{row-1})"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    ws[f'H{row}'] = f"=F{row}+G{row}"
    ws[f'H{row}'].number_format = '#,##0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    for i, w in enumerate([28, 12, 8, 16, 14, 18, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = 'A4'

def criar_aba_ci_detalhado(wb):
    """Cria aba CI_DETALHADO com valores paramétricos conectados ao briefing"""
    ws = wb.create_sheet("CI_DETALHADO")
    ws.sheet_properties.tabColor = "C0392B"
    
    # Título
    ws['A1'] = "CUSTOS INDIRETOS DETALHADOS"
    ws.merge_cells('A1:F1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # ========== PROJETOS E CONSULTORIAS (25% do Gerenciamento) ==========
    row = 3
    ws[f'A{row}'] = "PROJETOS E CONSULTORIAS (25%)"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    
    row += 1
    headers = ["Disciplina", "Peso %", "Valor (R$)", "R$/m²", "", ""]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    proj_items = [
        ("Arquitetônico", 0.25),
        ("Estrutural", 0.20),
        ("Elétrico", 0.12),
        ("Hidrossanitário", 0.12),
        ("Preventivo", 0.08),
        ("Climatização", 0.05),
        ("Paisagismo", 0.05),
        ("Interiores", 0.05),
        ("Fundações", 0.03),
        ("Outros", 0.05)
    ]
    
    for disc, peso in proj_items:
        ws[f'A{row}'] = disc
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Valor = CUSTOS_MACROGRUPO!G2 × 0.25 × Peso
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$2*0.25*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        # R$/m²
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Projetos
    subtotal_proj_row = row
    ws[f'A{row}'] = "TOTAL PROJETOS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-10}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # ========== TAXAS E LICENÇAS (5% do Gerenciamento) ==========
    row += 2
    ws[f'A{row}'] = "TAXAS E LICENÇAS (5%)"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    taxas_items = [
        ("Alvará", 0.25),
        ("ARTs/RRTs", 0.20),
        ("Bombeiros", 0.20),
        ("Habite-se", 0.15),
        ("Registro Incorporação", 0.10),
        ("Outras", 0.10)
    ]
    
    for item, peso in taxas_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$2*0.05*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Taxas
    ws[f'A{row}'] = "TOTAL TAXAS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-6}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # ========== EQUIPE ADM (40% do Gerenciamento) ==========
    row += 2
    ws[f'A{row}'] = "EQUIPE ADM (40%)"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    
    row += 1
    headers_equipe = ["Função", "Peso %", "Valor (R$)", "Custo/mês (R$)", "Qtd", ""]
    for col, header in enumerate(headers_equipe, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    equipe_items = [
        ("Eng. Residente", 0.30, 1),
        ("Mestre de Obras", 0.25, 1),
        ("Almoxarife", 0.12, 1),
        ("Téc. Segurança", 0.12, 1),
        ("Admin. Obra", 0.12, 1),
        ("Estagiários", 0.09, 2)
    ]
    
    for funcao, peso, qtd in equipe_items:
        ws[f'A{row}'] = funcao
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        # Valor = CUSTOS_MACROGRUPO!G2 × 0.40 × Peso
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$2*0.40*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        # Custo/mês = Valor / Prazo
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$17"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        ws[f'E{row}'] = qtd
        
        row += 1
    
    # Subtotal Equipe
    ws[f'A{row}'] = "TOTAL EQUIPE ADM"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-6}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$17"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # ========== EPCs (15% do Gerenciamento) ==========
    row += 2
    ws[f'A{row}'] = "EPCs (15%)"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    epc_items = [
        ("Bandeja proteção", 0.25),
        ("Guarda-corpo", 0.25),
        ("Tela proteção", 0.20),
        ("Linha de vida", 0.15),
        ("Sinalização", 0.10),
        ("Outros", 0.05)
    ]
    
    for item, peso in epc_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$2*0.15*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal EPCs
    ws[f'A{row}'] = "TOTAL EPCs"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-6}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # ========== EQUIPAMENTOS (10% do Gerenciamento) ==========
    row += 2
    ws[f'A{row}'] = "EQUIPAMENTOS (10%)"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    equip_items = [
        ("Cremalheira", 0.40),
        ("Grua/Guindaste", 0.30),
        ("Balancins", 0.15),
        ("Mini-Grua", 0.10),
        ("Outros", 0.05)
    ]
    
    for item, peso in equip_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$2*0.10*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Equipamentos
    ws[f'A{row}'] = "TOTAL EQUIPAMENTOS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-5}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # ========== ENSAIOS (5% do Gerenciamento) ==========
    row += 2
    ws[f'A{row}'] = "ENSAIOS (5%)"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="7F8C8D", end_color="7F8C8D", fill_type="solid")
    
    row += 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    row += 1
    ensaios_items = [
        ("Concreto (CP)", 0.40),
        ("Aço", 0.15),
        ("Solo/Fundações", 0.20),
        ("Estanqueidade", 0.15),
        ("Outros", 0.10)
    ]
    
    for item, peso in ensaios_items:
        ws[f'A{row}'] = item
        ws[f'B{row}'] = peso
        ws[f'B{row}'].number_format = '0%'
        
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!$G$2*0.05*B{row}"
        ws[f'C{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
        ws[f'D{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Subtotal Ensaios
    ws[f'A{row}'] = "TOTAL ENSAIOS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = f"=SUM(C{row-5}:C{row-1})"
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'C{row}'].font = Font(bold=True)
    ws[f'D{row}'] = f"=C{row}/DADOS_PROJETO!$B$6"
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'D{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    for i, w in enumerate([30, 12, 14, 14, 10, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = 'A4'

def criar_aba_benchmark(wb, recal=None):
    """Cria aba BENCHMARK com projetos de referência"""
    ws = wb.create_sheet("BENCHMARK")
    ws.sheet_properties.tabColor = "27AE60"
    
    # Título
    ws['A1'] = "BENCHMARK — PROJETOS DE REFERÊNCIA"
    ws.merge_cells('A1:H1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # Headers
    row = 3
    headers = ["Projeto", "Cidade", "AC (m²)", "NP", "Prazo (meses)", "CUB (R$/m²)", "R$/m² Original", "CUB Ratio", "R$/m² Atual"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Dados dos 8 projetos de referência
    row += 1
    projetos = [
        ("Maison Beach", "Floripa", 12880, 25, 36, 2753, 3291, 1.20),
        ("Catena", "Floripa", 9242, 20, 30, 3009, 3376, 1.12),
        ("Connect", "Itajaí", 13144, 28, 36, 2900, 3150, 1.09),
        ("Dlohn", "Blumenau", 8500, 18, 24, 2800, 3080, 1.10),
        ("Eternity", "Itajaí", 11200, 22, 30, 2850, 3420, 1.20),
        ("Lorenzo", "Itajaí", 9800, 20, 28, 2780, 3190, 1.15),
        ("Monolyt", "Floripa", 18500, 32, 42, 2900, 3480, 1.20),
        ("Cota365", "Floripa", 10500, 24, 30, 2850, 3250, 1.14)
    ]
    
    for projeto, cidade, ac, np, prazo, cub, rsm2, ratio in projetos:
        ws[f'A{row}'] = projeto
        ws[f'B{row}'] = cidade
        
        ws[f'C{row}'] = ac
        ws[f'C{row}'].number_format = '#,##0'
        
        ws[f'D{row}'] = np
        
        ws[f'E{row}'] = prazo
        
        ws[f'F{row}'] = cub
        ws[f'F{row}'].number_format = '#,##0.00'
        
        ws[f'G{row}'] = rsm2
        ws[f'G{row}'].number_format = '#,##0.00'
        
        ws[f'H{row}'] = ratio
        ws[f'H{row}'].number_format = '0.00'
        
        # R$/m² Atual (I) = CUB Ratio × CUB Atual
        ws[f'I{row}'] = f"=H{row}*DADOS_PROJETO!$B$19"
        ws[f'I{row}'].number_format = '#,##0.00'
        aplicar_estilo(ws[f'I{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Linha do projeto atual
    row += 1
    ws[f'A{row}'] = "ESTE PROJETO"
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    ws[f'B{row}'] = "=DADOS_PROJETO!B4"
    
    ws[f'C{row}'] = "=DADOS_PROJETO!B6"
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'C{row}'].font = Font(bold=True)
    
    ws[f'D{row}'] = "=DADOS_PROJETO!B9"
    ws[f'D{row}'].font = Font(bold=True)
    
    ws[f'E{row}'] = "=DADOS_PROJETO!B17"
    ws[f'E{row}'].font = Font(bold=True)
    
    ws[f'F{row}'] = "=DADOS_PROJETO!B19"
    ws[f'F{row}'].number_format = '#,##0.00'
    ws[f'F{row}'].font = Font(bold=True)
    
    ws[f'G{row}'] = "=CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B6"
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    
    ws[f'H{row}'] = "=(CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B6)/DADOS_PROJETO!B19"
    ws[f'H{row}'].number_format = '0.00'
    ws[f'H{row}'].font = Font(bold=True)
    
    # R$/m² Atual (I) = mesmo que G pra este projeto (já está no CUB atual)
    ws[f'I{row}'] = "=CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B6"
    ws[f'I{row}'].number_format = '#,##0.00'
    ws[f'I{row}'].font = Font(bold=True)
    ws[f'I{row}'].fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    # Comparativo por macrogrupo
    row += 2
    ws[f'A{row}'] = "COMPARATIVO POR MACROGRUPO (R$/m²)"
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
    
    row += 1
    headers_comp = ["Macrogrupo", "Média Base", "Este Projeto", "Diferença %"]
    for col, header in enumerate(headers_comp, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Média da base (valores aproximados dos 8 projetos, ajustados para valor presente)
    medias_base_historico = {
        "Gerenciamento": 380,
        "Infraestrutura": 210,
        "Supraestrutura": 680,
        "Instalações": 350,
        "Pisos": 180,
        "Esquadrias": 330,
        "Fachada": 160
    }
    fator = recal["fator"] if recal else 1.0
    medias_base = {k: round(v * fator, 0) for k, v in medias_base_historico.items()}
    
    row += 1
    for macro, media in medias_base.items():
        # Busca a linha do macrogrupo em CUSTOS_MACROGRUPO
        macro_idx = MACROGRUPOS.index(macro)
        custos_row = macro_idx + 2
        
        ws[f'A{row}'] = macro
        
        ws[f'B{row}'] = media
        ws[f'B{row}'].number_format = '#,##0'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        
        ws[f'C{row}'] = f"=CUSTOS_MACROGRUPO!F{custos_row}"
        ws[f'C{row}'].number_format = '#,##0'
        aplicar_estilo(ws[f'C{row}'], criar_estilo_calculado())
        
        ws[f'D{row}'] = f"=(C{row}-B{row})/B{row}"
        ws[f'D{row}'].number_format = '0.0%'
        aplicar_estilo(ws[f'D{row}'], criar_estilo_calculado())
        
        row += 1
    
    # Ajustar larguras
    for i, w in enumerate([25, 18, 14, 10, 14, 14, 14, 12, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.freeze_panes = 'A4'

def criar_aba_analise_produto(wb):
    """Cria aba ANÁLISE_PRODUTO com benchmark do produto"""
    ws = wb.create_sheet("ANÁLISE_PRODUTO")
    ws.sheet_properties.tabColor = "3498DB"
    
    # Título
    ws['A1'] = "ANÁLISE DE PRODUTO"
    ws.merge_cells('A1:D1')
    aplicar_estilo(ws['A1'], criar_estilo_header())
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    
    # Headers
    row = 3
    headers = ["Indicador", "Este Projeto", "Faixa Típica", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        aplicar_estilo(cell, criar_estilo_header())
    
    # Indicadores
    row += 1
    indicadores = [
        ("AC/UR (m²/UR)", "=DADOS_PROJETO!B6/DADOS_PROJETO!B7", "90-160", 
         '=IF(AND(B4>=90,B4<=160),"OK","ATENÇÃO")'),
        ("Custo/UR (R$)", "=CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B7", "—", "—"),
        ("R$/m² (R$/m²)", "=CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B6", "—", "—"),
        ("CUB Ratio", "=(CUSTOS_MACROGRUPO!G20/DADOS_PROJETO!B6)/DADOS_PROJETO!B19", "1.00-1.50", 
         '=IF(AND(B7>=1,B7<=1.5),"OK",IF(B7>1.5,"ALTO","BAIXO"))'),
        ("Vagas/UR", "=DADOS_PROJETO!B13/DADOS_PROJETO!B7", "1.0-2.0", 
         '=IF(AND(B8>=1,B8<=2),"OK","ATENÇÃO")'),
        ("% Tipo (NPT/NP)", "=DADOS_PROJETO!B10/DADOS_PROJETO!B9", "60-80%", 
         '=IF(AND(B9>=0.6,B9<=0.8),"OK","ATENÇÃO")'),
        ("Elevador/UR", "=DADOS_PROJETO!B12/DADOS_PROJETO!B7", "0.015-0.030", 
         '=IF(AND(B10>=0.015,B10<=0.03),"OK","ATENÇÃO")')
    ]
    
    for ind, formula, faixa, status in indicadores:
        ws[f'A{row}'] = ind
        ws[f'A{row}'].font = Font(bold=True)
        
        ws[f'B{row}'] = formula
        if "%" in ind:
            ws[f'B{row}'].number_format = '0.0%'
        elif "R$" in ind:
            ws[f'B{row}'].number_format = '#,##0.00'
        else:
            ws[f'B{row}'].number_format = '0.00'
        aplicar_estilo(ws[f'B{row}'], criar_estilo_calculado())
        ws[f'B{row}'].font = Font(bold=True)
        
        ws[f'C{row}'] = faixa
        ws[f'C{row}'].alignment = Alignment(horizontal="center")
        
        if status != "—":
            ws[f'D{row}'] = status
            ws[f'D{row}'].alignment = Alignment(horizontal="center")
            ws[f'D{row}'].font = Font(bold=True)
        
        row += 1
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    
    ws.freeze_panes = 'A4'

# ============================================================================
# FUNÇÃO PRINCIPAL
# ============================================================================

def gerar_planilha(output_path, dados_projeto=None, briefing_respostas=None, calibration_stats_path=None):
    """
    Gera planilha Excel com template dinâmico

    Args:
        output_path: Caminho do arquivo de saída
        dados_projeto: Dict com dados do projeto (ou None para template vazio)
        briefing_respostas: Dict com respostas do briefing (ou None para vazio)
        calibration_stats_path: Caminho para calibration-stats.json (ou None)
    """
    print(f"Gerando planilha: {output_path}")

    # Recalibrar medianas para valor presente
    cub_atual = dados_projeto.get("CUB") if dados_projeto else None
    if cub_atual:
        recal = recalibrar_base(cub_atual, calibration_stats_path)
        dados_projeto["CUB_BASE"] = cub_atual  # CUB_BASE = CUB_Atual
        dados_projeto["Data-base"] = recal["data_base_label"]
    else:
        recal = {"medianas": dict(MEDIANAS_BASE_DEZ23), "faixas": dict(FAIXAS_BASE_DEZ23),
                 "fator": 1.0, "total_projetos": 58, "data_base_label": ""}

    wb = Workbook()
    
    # Remove sheet padrão
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Cria as abas na ordem especificada
    # Ordem: PAINEL, DADOS_PROJETO, BRIEFING, FATORES, CUSTOS_MACROGRUPO, 
    #        ÍNDICES, ESTRUTURAL, INSTALACOES, ACABAMENTOS, CI_DETALHADO, 
    #        BENCHMARK, ANÁLISE_PRODUTO, ALERTAS, NOTAS
    
    print("  ├─ Criando PAINEL...")
    criar_aba_painel(wb)
    
    print("  ├─ Criando DADOS_PROJETO...")
    criar_aba_dados_projeto(wb, dados_projeto, recal["data_base_label"])
    
    print("  ├─ Criando BRIEFING...")
    criar_aba_briefing(wb, briefing_respostas)
    
    print("  ├─ Criando FATORES...")
    criar_aba_fatores(wb)
    
    print("  ├─ Criando CUSTOS_MACROGRUPO...")
    criar_aba_custos_macrogrupo(wb, recal)
    
    print("  ├─ Criando ÍNDICES...")
    criar_aba_indices(wb)
    
    print("  ├─ Criando ESTRUTURAL...")
    criar_aba_estrutural(wb)
    
    print("  ├─ Criando INSTALACOES...")
    criar_aba_instalacoes(wb)
    
    print("  ├─ Criando ACABAMENTOS...")
    criar_aba_acabamentos(wb)
    
    print("  ├─ Criando CI_DETALHADO...")
    criar_aba_ci_detalhado(wb)
    
    print("  ├─ Criando BENCHMARK...")
    criar_aba_benchmark(wb, recal)
    
    print("  ├─ Criando ANÁLISE_PRODUTO...")
    criar_aba_analise_produto(wb)
    
    print("  ├─ Criando ALERTAS...")
    criar_aba_alertas(wb)
    
    print("  ├─ Criando NOTAS...")
    criar_aba_notas(wb, recal)
    
    # Salva
    print(f"  └─ Salvando em {output_path}...")
    wb.save(output_path)
    print(f"✓ Planilha gerada com sucesso!")

# ============================================================================
# EXECUÇÃO
# ============================================================================

if __name__ == "__main__":
    # Cria diretório se não existir
    output_dir = os.path.expanduser("~/clawd/orcamento-parametrico/parametricos")
    os.makedirs(output_dir, exist_ok=True)
    
    print("=" * 70)
    print("GERADOR DE TEMPLATE DE ORÇAMENTO PARAMÉTRICO V4")
    print("=" * 70)
    print()
    
    calibration_stats_path = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "calibration-stats.json"
    )

    # 1. Gera versão com dados do Rozzo
    rozzo_path = os.path.join(output_dir, "rozzo-vd-parametrico-v4.xlsx")
    print("1. Gerando versão ROZZO (pré-preenchida)...")
    gerar_planilha(rozzo_path, ROZZO_DADOS, ROZZO_BRIEFING, calibration_stats_path)
    print()

    # 2. Gera versão template limpa
    template_path = os.path.join(output_dir, "template-orcamento-parametrico-v4.xlsx")
    print("2. Gerando versão TEMPLATE (limpa)...")
    gerar_planilha(template_path, None, None, calibration_stats_path)
    print()
    
    print("=" * 70)
    print("CONCLUÍDO!")
    print("=" * 70)
    print(f"Arquivos gerados:")
    print(f"  • {rozzo_path}")
    print(f"  • {template_path}")
    print()
