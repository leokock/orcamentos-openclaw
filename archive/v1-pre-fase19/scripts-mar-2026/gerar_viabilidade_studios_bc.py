#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-
"""
Planilha de Viabilidade — Studios BC
VGV R$ 12.000/m² | Custo paramétrico shell delivery
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# Cores
DARK_BLUE = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
GREEN_DARK = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
GREEN_LIGHT = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_LIGHT = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
YELLOW_LIGHT = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
BLUE_LIGHT = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
GRAY_LIGHT = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
ORANGE_LIGHT = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD = Font(name="Calibri", bold=True, size=11)
BOLD_BIG = Font(name="Calibri", bold=True, size=13)
NORMAL = Font(name="Calibri", size=11)
TITLE = Font(name="Calibri", bold=True, size=16, color="2C3E50")
SUBTITLE = Font(name="Calibri", bold=True, size=12, color="2C3E50")
SMALL_GRAY = Font(name="Calibri", size=9, color="888888")
THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

BRL = '#,##0'
BRL2 = '#,##0.00'
PCT = '0.0%'
PCT1 = '0.0%'

def style_header(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DARK_BLUE
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN

def style_cell(ws, row, col, fill=None, font=None, fmt=None, align='center'):
    cell = ws.cell(row=row, column=col)
    if fill: cell.fill = fill
    if font: cell.font = font
    if fmt: cell.number_format = fmt
    cell.border = THIN
    cell.alignment = Alignment(horizontal=align, vertical='center')
    return cell

def write_row(ws, row, data, fill=None, font=None, formats=None):
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=i+1, value=val)
        cell.border = THIN
        if fill: cell.fill = fill
        if font: cell.font = font
        cell.alignment = Alignment(
            horizontal='left' if i == 0 else 'center',
            vertical='center'
        )
        if formats and i < len(formats) and formats[i]:
            cell.number_format = formats[i]

# ============================================================================
# DADOS DO PROJETO
# ============================================================================
# Áreas
area_privativa_total = 280 * 30  # 8.400 m² (área vendável studios)
area_garagem = 6864
area_tipos = 13416
ac_total = 20280
n_unidades = 280
area_media_un = 30

# Venda
vgv_m2 = 12000  # R$/m² área privativa
vgv_total = vgv_m2 * area_privativa_total  # sobre área privativa

# Custos (do paramétrico)
custo_obra_total = 73_328_012
custo_tipos_m2 = 4556
custo_garagem_m2 = 1779
custo_medio_m2 = 3616

# Premissas de viabilidade (mercado BC)
terreno_pct = 0.12  # 12% do VGV (referência BC)
projetos_pct = 0.04  # 4% do custo obra
marketing_pct = 0.04  # 4% do VGV
comissao_pct = 0.05  # 5% do VGV (corretagem)
adm_incorp_pct = 0.03  # 3% do VGV
impostos_pct = 0.037  # 3,7% (RET regime especial)
financ_pct = 0.02  # 2% do VGV (custo financeiro)
contingencia_pct = 0.03  # 3% do custo obra

# Cálculos
terreno = vgv_total * terreno_pct
projetos = custo_obra_total * projetos_pct
marketing = vgv_total * marketing_pct
comissao = vgv_total * comissao_pct
adm_incorp = vgv_total * adm_incorp_pct
impostos = vgv_total * impostos_pct
financ = vgv_total * financ_pct
contingencia = custo_obra_total * contingencia_pct

custo_total = terreno + custo_obra_total + projetos + marketing + comissao + adm_incorp + impostos + financ + contingencia
lucro = vgv_total - custo_total
margem = lucro / vgv_total
markup = lucro / custo_total
roi = lucro / (terreno + custo_obra_total * 0.3)  # capital investido ~30% da obra + terreno

# ============================================================================
# ABA 1: PAINEL DE VIABILIDADE
# ============================================================================
ws = wb.active
ws.title = "VIABILIDADE"

ws.column_dimensions['A'].width = 36
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 22

# Título
ws.merge_cells('A1:F1')
ws['A1'] = "ESTUDO DE VIABILIDADE — STUDIOS BALNEÁRIO CAMBORIÚ"
ws['A1'].font = TITLE
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:F2')
ws['A2'] = "280 studios ~30m² | Padrão Alto | Shell Delivery | CUB SC mar/2026"
ws['A2'].font = SMALL_GRAY
ws['A2'].alignment = Alignment(horizontal='center')

# ===== RECEITA =====
r = 4
ws.merge_cells(f'A{r}:F{r}')
ws[f'A{r}'] = "RECEITA (VGV)"
ws[f'A{r}'].font = WHITE_FONT
ws[f'A{r}'].fill = GREEN_DARK
ws[f'A{r}'].alignment = Alignment(horizontal='center')
for c in range(2,7): ws.cell(row=r, column=c).fill = GREEN_DARK

r = 5
style_header(ws, r, 1, 4)
for i, h in enumerate(["Receita", "Base de Cálculo", "Valor (R$)", "R$/m² AC"], 1):
    ws.cell(row=r, column=i, value=h)

receita_rows = [
    ("Venda Studios (280 un × 30m²)", f"{n_unidades} un × R$ {vgv_m2:,}/m² priv.", vgv_total, vgv_total/ac_total),
    ("Vagas de Garagem (excedentes)", "Incluído no VGV", 0, 0),
]

r = 6
for label, base, val, rm2 in receita_rows:
    write_row(ws, r, [label, base, val, rm2], formats=[None, None, BRL, BRL2])
    if val > 0:
        style_cell(ws, r, 3, fill=GREEN_LIGHT, font=BOLD, fmt=BRL)
        style_cell(ws, r, 4, fill=GREEN_LIGHT, fmt=BRL2)
    r += 1

write_row(ws, r, ["VGV TOTAL", "", vgv_total, vgv_total/ac_total],
          fill=GREEN_LIGHT, font=BOLD, formats=[None, None, BRL, BRL2])
r += 1

# ===== CUSTOS =====
r += 1
ws.merge_cells(f'A{r}:F{r}')
ws[f'A{r}'] = "CUSTOS"
ws[f'A{r}'].font = WHITE_FONT
ws[f'A{r}'].fill = DARK_BLUE
ws[f'A{r}'].alignment = Alignment(horizontal='center')
for c in range(2,7): ws.cell(row=r, column=c).fill = DARK_BLUE

r += 1
style_header(ws, r, 1, 6)
for i, h in enumerate(["Item de Custo", "Base de Cálculo", "Valor (R$)", "% do VGV", "R$/m² AC", "Observação"], 1):
    ws.cell(row=r, column=i, value=h)

custos = [
    ("Terreno", f"{terreno_pct*100:.0f}% do VGV", terreno, terreno/vgv_total, terreno/ac_total, "Referência mercado BC"),
    ("Custo de Obra (paramétrico)", f"R$ {custo_medio_m2:,}/m² AC × {ac_total:,} m²", custo_obra_total, custo_obra_total/vgv_total, custo_medio_m2, "Shell delivery, 75 projetos calibrados"),
    ("   Lazer + Tipos", f"R$ {custo_tipos_m2:,}/m² × {area_tipos:,} m²", custo_tipos_m2 * area_tipos, None, custo_tipos_m2, "Com desconto shell"),
    ("   Garagem", f"R$ {custo_garagem_m2:,}/m² × {area_garagem:,} m²", custo_garagem_m2 * area_garagem, None, custo_garagem_m2, "Sem subsolo"),
    ("Projetos e Aprovações", f"{projetos_pct*100:.0f}% do custo obra", projetos, projetos/vgv_total, projetos/ac_total, "Arq, estrutural, complementares, aprovações"),
    ("Marketing e Vendas", f"{marketing_pct*100:.0f}% do VGV", marketing, marketing/vgv_total, marketing/ac_total, "Stand, material, digital, eventos"),
    ("Comissão de Vendas", f"{comissao_pct*100:.0f}% do VGV", comissao, comissao/vgv_total, comissao/ac_total, "Corretagem"),
    ("Adm. da Incorporação", f"{adm_incorp_pct*100:.0f}% do VGV", adm_incorp, adm_incorp/vgv_total, adm_incorp/ac_total, "Equipe, jurídico, contábil"),
    ("Impostos (RET)", f"{impostos_pct*100:.1f}% do VGV", impostos, impostos/vgv_total, impostos/ac_total, "Regime Especial de Tributação"),
    ("Custo Financeiro", f"{financ_pct*100:.0f}% do VGV", financ, financ/vgv_total, financ/ac_total, "Capital de giro, juros construção"),
    ("Contingência", f"{contingencia_pct*100:.0f}% do custo obra", contingencia, contingencia/vgv_total, contingencia/ac_total, "Margem de segurança"),
]

r += 1
r_custos_start = r
for label, base, val, pct_vgv, rm2, obs in custos:
    is_sub = label.startswith("   ")
    write_row(ws, r, [label, base, val, pct_vgv, rm2, obs],
              formats=[None, None, BRL, PCT if pct_vgv else None, BRL2, None])
    if is_sub:
        style_cell(ws, r, 1, font=Font(name="Calibri", size=10, color="666666"), align='left')
        style_cell(ws, r, 3, fill=GRAY_LIGHT, fmt=BRL)
    r += 1

# Total custos
write_row(ws, r, ["CUSTO TOTAL", "", custo_total, custo_total/vgv_total, custo_total/ac_total, ""],
          fill=RED_LIGHT, font=BOLD, formats=[None, None, BRL, PCT, BRL2, None])
r_custo_total = r
r += 1

# ===== RESULTADO =====
r += 1
ws.merge_cells(f'A{r}:F{r}')
ws[f'A{r}'] = "RESULTADO"
ws[f'A{r}'].font = WHITE_FONT
fill_resultado = GREEN_DARK if margem > 0.15 else DARK_BLUE
ws[f'A{r}'].fill = fill_resultado
ws[f'A{r}'].alignment = Alignment(horizontal='center')
for c in range(2,7): ws.cell(row=r, column=c).fill = fill_resultado

r += 1
style_header(ws, r, 1, 5)
for i, h in enumerate(["Indicador", "Valor", "Referência Mercado", "Status", "Observação"], 1):
    ws.cell(row=r, column=i, value=h)

# Status check
margem_status = "✓ Saudável" if margem > 0.15 else ("⚠ Apertada" if margem > 0.10 else "✗ Crítica")
markup_status = "✓ Bom" if markup > 0.18 else ("⚠ Apertado" if markup > 0.12 else "✗ Baixo")

resultado_rows = [
    ("VGV Total", vgv_total, BRL, "—", "", ""),
    ("(-) Custo Total", custo_total, BRL, "—", "", ""),
    ("(=) LUCRO BRUTO", lucro, BRL, "—", "✓ Positivo" if lucro > 0 else "✗ Negativo", ""),
    ("", None, None, None, None, None),  # spacer
    ("Margem sobre VGV", margem, PCT1, "> 15%", margem_status,
     "Margem líquida antes de IR sobre lucro"),
    ("Markup sobre Custo", markup, PCT1, "> 18%", markup_status, ""),
    ("Custo / VGV", custo_total/vgv_total, PCT1, "< 85%",
     "✓ OK" if custo_total/vgv_total < 0.85 else "⚠ Alto", ""),
    ("Lucro por Unidade", lucro/n_unidades, BRL, "> R$ 50.000", 
     "✓" if lucro/n_unidades > 50000 else "⚠", ""),
    ("Lucro por m² Privativo", lucro/area_privativa_total, BRL2, "> R$ 1.500/m²",
     "✓" if lucro/area_privativa_total > 1500 else "⚠", ""),
    ("VGV / Custo Obra", vgv_total/custo_obra_total, '0.00x', "> 1,3x",
     "✓" if vgv_total/custo_obra_total > 1.3 else "⚠", "Alavancagem sobre custo de obra"),
]

r += 1
for label, val, fmt, ref, status, obs in resultado_rows:
    if val is None:
        r += 1
        continue
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=2, value=val)
    ws.cell(row=r, column=3, value=ref)
    ws.cell(row=r, column=4, value=status)
    ws.cell(row=r, column=5, value=obs)
    
    if fmt: ws.cell(row=r, column=2).number_format = fmt
    
    fill = None
    font_r = NORMAL
    if "LUCRO" in label:
        fill = GREEN_LIGHT if lucro > 0 else RED_LIGHT
        font_r = BOLD
    elif "Margem" in label or "Markup" in label:
        fill = GREEN_LIGHT if "✓" in status else (YELLOW_LIGHT if "⚠" in status else RED_LIGHT)
    
    for c in range(1, 6):
        cell = ws.cell(row=r, column=c)
        cell.border = THIN
        if fill: cell.fill = fill
        if font_r == BOLD: cell.font = BOLD
        cell.alignment = Alignment(horizontal='left' if c == 1 else 'center', vertical='center')
    r += 1

# ===== SENSIBILIDADE =====
r += 2
ws.merge_cells(f'A{r}:F{r}')
ws[f'A{r}'] = "ANÁLISE DE SENSIBILIDADE — VARIAÇÃO DO PREÇO DE VENDA"
ws[f'A{r}'].font = SUBTITLE
ws[f'A{r}'].alignment = Alignment(horizontal='center')

r += 1
style_header(ws, r, 1, 6)
for i, h in enumerate(["Cenário", "R$/m² Venda", "VGV (R$)", "Lucro (R$)", "Margem", "Status"], 1):
    ws.cell(row=r, column=i, value=h)

cenarios = [
    ("Pessimista (-20%)", 9600),
    ("Conservador (-10%)", 10800),
    ("Base (projeção)", 12000),
    ("Otimista (+10%)", 13200),
    ("Aquecido (+20%)", 14400),
]

r += 1
for nome, preco in cenarios:
    vgv_c = preco * area_privativa_total
    # Custos variáveis ajustam com VGV
    custos_var = vgv_c * (marketing_pct + comissao_pct + adm_incorp_pct + impostos_pct + financ_pct + terreno_pct)
    custos_fix = custo_obra_total + projetos + contingencia
    custo_c = custos_var + custos_fix
    lucro_c = vgv_c - custo_c
    margem_c = lucro_c / vgv_c if vgv_c > 0 else 0
    
    status_c = "✓ Viável" if margem_c > 0.15 else ("⚠ Apertado" if margem_c > 0.05 else "✗ Inviável")
    
    fill = GREEN_LIGHT if margem_c > 0.15 else (YELLOW_LIGHT if margem_c > 0.05 else RED_LIGHT)
    if preco == 12000:
        fill = BLUE_LIGHT
    
    write_row(ws, r, [nome, preco, vgv_c, lucro_c, margem_c, status_c],
              formats=[None, BRL, BRL, BRL, PCT, None])
    for c in range(1, 7):
        style_cell(ws, r, c, fill=fill, align='left' if c == 1 else 'center')
    if preco == 12000:
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = BOLD
    r += 1

# ===== COMPOSIÇÃO % VGV (gráfico textual) =====
r += 2
ws.merge_cells(f'A{r}:F{r}')
ws[f'A{r}'] = "COMPOSIÇÃO DO VGV"
ws[f'A{r}'].font = SUBTITLE
ws[f'A{r}'].alignment = Alignment(horizontal='center')

r += 1
style_header(ws, r, 1, 4)
for i, h in enumerate(["Componente", "Valor (R$)", "% do VGV", ""], 1):
    ws.cell(row=r, column=i, value=h)

composicao = [
    ("Terreno", terreno),
    ("Custo de Obra", custo_obra_total),
    ("Projetos e Aprovações", projetos),
    ("Marketing + Vendas", marketing),
    ("Comissão", comissao),
    ("Adm. Incorporação", adm_incorp),
    ("Impostos (RET)", impostos),
    ("Custo Financeiro", financ),
    ("Contingência", contingencia),
    ("LUCRO", lucro),
]

r += 1
for label, val in composicao:
    pct = val / vgv_total
    barra = "█" * int(pct * 50)
    fill = GREEN_LIGHT if label == "LUCRO" else None
    font_c = BOLD if label == "LUCRO" else NORMAL
    write_row(ws, r, [label, val, pct, barra],
              formats=[None, BRL, PCT, None])
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        cell.border = THIN
        if fill: cell.fill = fill
        cell.font = font_c
    r += 1

write_row(ws, r, ["VGV TOTAL", vgv_total, 1.0, ""],
          formats=[None, BRL, PCT, None])
for c in range(1, 5):
    style_cell(ws, r, c, fill=DARK_BLUE, font=WHITE_FONT)

# ===== PREMISSAS =====
r += 3
ws.merge_cells(f'A{r}:F{r}')
ws[f'A{r}'] = "PREMISSAS"
ws[f'A{r}'].font = SUBTITLE

r += 1
premissas = [
    "• VGV calculado sobre área privativa (8.400 m²), não sobre AC total",
    "• Custo de obra: paramétrico calibrado com 75 projetos reais (CUB SC mar/2026)",
    "• Shell delivery: studios entregues sem acabamento interno (apenas paredes perimetrais + banheiro)",
    "• Terreno: 12% do VGV (referência mercado BC para empreendimento vertical)",
    "• Corretagem: 5% (padrão mercado)",
    "• Impostos: RET 3,7% (Regime Especial de Tributação — incorporação)",
    "• Marketing: 4% do VGV (stand de vendas, material, digital, eventos)",
    "• Custo financeiro: 2% do VGV (capital de giro + juros construção)",
    "• NÃO inclui: IR sobre lucro, custo de oportunidade, permuta de terreno",
    "• Valores de venda de garagem excedente não considerados (conservador)",
    "• Análise de sensibilidade considera custos fixos de obra constantes",
]

for p in premissas:
    ws.merge_cells(f'A{r}:F{r}')
    ws[f'A{r}'] = p
    ws[f'A{r}'].font = Font(name="Calibri", size=9.5, color="555555")
    r += 1

# ============================================================================
# ABA 2: FLUXO SIMPLIFICADO
# ============================================================================
ws2 = wb.create_sheet("FLUXO_SIMPLIFICADO")

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20

ws2.merge_cells('A1:D1')
ws2['A1'] = "FLUXO SIMPLIFICADO — ENTRADAS E SAÍDAS"
ws2['A1'].font = TITLE
ws2['A1'].alignment = Alignment(horizontal='center')

r = 3
style_header(ws2, r, 1, 4)
for i, h in enumerate(["Fase", "Período (meses)", "Entrada/Saída", "Valor (R$)"], 1):
    ws2.cell(row=r, column=i, value=h)

fases = [
    ("", "", "SAÍDAS", None),
    ("Aquisição Terreno", "Mês 0", "Saída", terreno),
    ("Projetos e Aprovações", "Mês 0-6", "Saída", projetos),
    ("Obra — Fase 1 (Fundação/Estrutura)", "Mês 1-18", "Saída", custo_obra_total * 0.45),
    ("Obra — Fase 2 (Acabamentos)", "Mês 18-30", "Saída", custo_obra_total * 0.40),
    ("Obra — Fase 3 (Entrega)", "Mês 30-36", "Saída", custo_obra_total * 0.15),
    ("Marketing", "Mês 0-36", "Saída", marketing),
    ("Comissão + Adm + Impostos", "Conforme vendas", "Saída", comissao + adm_incorp + impostos),
    ("Custo Financeiro", "Mês 0-36", "Saída", financ),
    ("Contingência", "Reserva", "Saída", contingencia),
    ("TOTAL SAÍDAS", "", "", custo_total),
    ("", "", "", None),
    ("", "", "ENTRADAS", None),
    ("Vendas Pré-Lançamento (20%)", "Mês -3 a 0", "Entrada", vgv_total * 0.20),
    ("Vendas Construção (50%)", "Mês 1-36", "Entrada", vgv_total * 0.50),
    ("Vendas Pós-Entrega (30%)", "Mês 36-48", "Entrada", vgv_total * 0.30),
    ("TOTAL ENTRADAS (VGV)", "", "", vgv_total),
    ("", "", "", None),
    ("RESULTADO LÍQUIDO", "", "", lucro),
]

r = 4
for label, periodo, tipo, val in fases:
    ws2.cell(row=r, column=1, value=label)
    ws2.cell(row=r, column=2, value=periodo)
    ws2.cell(row=r, column=3, value=tipo)
    if val is not None:
        ws2.cell(row=r, column=4, value=val)
        ws2.cell(row=r, column=4).number_format = BRL
    
    for c in range(1, 5):
        ws2.cell(row=r, column=c).border = THIN
    
    if "TOTAL" in label or "RESULTADO" in label:
        for c in range(1, 5):
            ws2.cell(row=r, column=c).font = BOLD
            if "RESULTADO" in label:
                ws2.cell(row=r, column=c).fill = GREEN_LIGHT
            elif "SAÍDAS" in label:
                ws2.cell(row=r, column=c).fill = RED_LIGHT
            elif "ENTRADAS" in label:
                ws2.cell(row=r, column=c).fill = GREEN_LIGHT
    elif tipo == "SAÍDAS" or tipo == "ENTRADAS":
        for c in range(1, 5):
            ws2.cell(row=r, column=c).fill = DARK_BLUE
            ws2.cell(row=r, column=c).font = WHITE_FONT
    elif tipo == "Saída":
        ws2.cell(row=r, column=4).fill = ORANGE_LIGHT
    elif tipo == "Entrada":
        ws2.cell(row=r, column=4).fill = GREEN_LIGHT
    
    r += 1

# Salvar
output_path = os.path.expanduser("~/clawd/orcamento-parametrico/parametricos/studios-bc-viabilidade.xlsx")
wb.save(output_path)
print(f"✓ Planilha de viabilidade gerada: {output_path}")
print(f"\n  Resumo:")
print(f"    VGV: R$ {vgv_total:,.0f} ({vgv_m2:,}/m² × {area_privativa_total:,} m² priv.)")
print(f"    Custo Total: R$ {custo_total:,.0f}")
print(f"    Lucro: R$ {lucro:,.0f}")
print(f"    Margem: {margem*100:.1f}%")
print(f"    Markup: {markup*100:.1f}%")
print(f"    Lucro/un: R$ {lucro/n_unidades:,.0f}")
