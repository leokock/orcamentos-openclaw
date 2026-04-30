#!/usr/bin/env python3
"""
Preenche o orçamento executivo do Estoril — V2 (limpo e organizado).
Remove abas Visus, cria planilha limpa com formatação consistente.

Uso: python3 fill_estoril_v2.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE_DIR, '..', 'projetos', 'estoril', 'Estoril-Orcamento-Executivo-PREENCHIDO.xlsx')

# ── Styles ──────────────────────────────────────────────
BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
YELLOW = "FFF2CC"
GREEN_LIGHT = "E2EFDA"

title_font = Font(name='Calibri', bold=True, size=14, color=WHITE)
title_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
header_font = Font(name='Calibri', bold=True, size=10, color=BLUE)
header_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
data_font = Font(name='Calibri', size=10)
data_font_bold = Font(name='Calibri', size=10, bold=True)
trace_font = Font(name='Calibri', size=8, italic=True, color="808080")
subtotal_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
total_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
source_fill = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def trace(fonte, rev="", confianca="⭐⭐⭐"):
    parts = [f"[Fonte: {fonte}"]
    if rev:
        parts[0] += f" | Rev: {rev}"
    parts[0] += f" | Confiança: {confianca}]"
    return parts[0]


def write_title(ws, row, title, cols=8):
    """Escreve título da seção."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28


def write_header(ws, row, headers, widths=None):
    """Escreve header de tabela."""
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def write_data_row(ws, row, values, bold=False, fill=None, trace_col=None, trace_val=None):
    """Escreve linha de dados."""
    font = data_font_bold if bold else data_font
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        # Alinhamento: primeira coluna à esquerda, resto centralizado/direita
        if i == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if isinstance(v, float) and v > 100:
                cell.number_format = '#,##0.00'
            elif isinstance(v, float):
                cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    if trace_col and trace_val:
        cell = ws.cell(row=row, column=trace_col, value=trace_val)
        cell.font = trace_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ═══════════════════════════════════════════════════════════
# ABA: CAPA
# ═══════════════════════════════════════════════════════════
def create_capa(wb):
    ws = wb.create_sheet("CAPA", 0)
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    
    ws.merge_cells('B3:D3')
    ws.cell(row=3, column=2, value='ORÇAMENTO EXECUTIVO').font = Font(name='Calibri', bold=True, size=20, color=BLUE)
    
    ws.merge_cells('B5:D5')
    ws.cell(row=5, column=2, value='Residencial Cidade Estoril').font = Font(name='Calibri', bold=True, size=16)
    
    ws.merge_cells('B6:D6')
    ws.cell(row=6, column=2, value='Fonseca Neto | Camboriú/SC').font = Font(name='Calibri', size=14, color="666666")
    
    info = [
        (9, 'CNPJ', '44.070.181/0001-06', trace('ARQ 08 - Carimbo', '', '⭐⭐⭐⭐⭐')),
        (10, 'Endereço', 'Rua Licurana, 721 — Lot. Santa Regina I, Bairro Tabuleiro', trace('ARQ 03, 08', '', '⭐⭐⭐⭐⭐')),
        (11, 'Responsável Técnico', 'Jackson Dias Valentim — CREA/SC 175.540-2', trace('ARQ 01-11 Carimbo', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (12, 'Estrutural', 'Luciano Rosso — CREA SC 085.963-3', trace('01-A/01-B Fundações', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (13, 'Data Projeto', '08/06/2023', trace('Todas as pranchas', '', '⭐⭐⭐⭐⭐')),
        (14, 'Revisão Orçamento', 'R00 — IA (gerado por extração de PDFs)', ''),
        (16, 'Área Total Construída', '14.491,98 m²', trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (17, 'Área Computável', '6.434,32 m²', trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (18, 'Terreno', '1.430,00 m²', trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (19, 'Gabarito', '16 pavimentos (17 níveis)', trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (20, 'Vagas', '138', trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (21, 'Pavimentos Tipo', '10 (Dif + 01-09)', trace('AVAC REV01 confirmado', '', '⭐⭐⭐⭐')),
        (22, 'Elevadores', '2 (sociais)', trace('PPCI P08 - Pav Tipo', 'REV04', '⭐⭐⭐⭐')),
        (23, 'Unidades Residenciais', '~55 (5/pav × 11 pav)', trace('PPCI P08 - medidores gás', '', '⭐⭐⭐')),
    ]
    
    for row, campo, valor, fonte in info:
        ws.cell(row=row, column=2, value=campo).font = data_font_bold
        ws.cell(row=row, column=3, value=valor).font = data_font
        if fonte:
            ws.cell(row=row, column=4, value=fonte).font = trace_font
    
    print("  ✅ CAPA")


# ═══════════════════════════════════════════════════════════
# ABA: QUADRO DE ÁREAS
# ═══════════════════════════════════════════════════════════
def create_quadro_areas(wb):
    ws = wb.create_sheet("Quadro de Áreas")
    
    write_title(ws, 1, 'QUADRO DE ÁREAS — ARQ 01')
    ws.cell(row=2, column=1, value=trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')).font = trace_font
    
    headers = ['Pavimento', 'Á. Computável (m²)', 'Á. Não Comp. (m²)', 'Á. Total (m²)', 'Gabarito', 'Vagas', 'Função', 'Rastreabilidade']
    widths = [22, 18, 18, 18, 12, 10, 22, 50]
    write_header(ws, 4, headers, widths)
    
    SRC = trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')
    
    pavimentos = [
        ('Subsolo', None, 432.46, 432.46, -1, 15, 'Garagem'),
        ('Térreo', None, 1050.46, 1050.46, 1, 11, 'Acesso / Garagem'),
        ('Garagem 01', None, 1082.84, 1082.84, 2, 34, 'Garagem'),
        ('Garagem 02', None, 1155.45, 1155.45, 3, 38, 'Garagem'),
        ('Garagem 03', None, 1155.45, 1155.45, 4, 40, 'Garagem'),
        ('Lazer', None, 1155.45, 1155.45, 5, None, 'Área de lazer'),
        ('Tipo Diferenciado', 557.18, 165.63, 722.81, 6, None, 'Residencial'),
        ('Tipo 01', 557.18, 146.30, 703.48, 7, None, 'Residencial'),
        ('Tipo 02', 557.18, 146.30, 703.48, 8, None, 'Residencial'),
        ('Tipo 03', 557.18, 146.30, 703.48, 9, None, 'Residencial'),
        ('Tipo 04', 557.18, 146.30, 703.48, 10, None, 'Residencial'),
        ('Tipo 05', 557.18, 146.30, 703.48, 11, None, 'Residencial'),
        ('Tipo 06', 557.18, 146.30, 703.48, 12, None, 'Residencial'),
        ('Tipo 07', 557.18, 146.30, 703.48, 13, None, 'Residencial'),
        ('Tipo 08', 557.18, 146.30, 703.48, 14, None, 'Residencial'),
        ('Tipo 09', 557.18, 146.30, 703.48, 15, None, 'Residencial'),
        ('Tipo 10', 557.18, 146.30, 703.48, 16, None, 'Residencial'),
        ('Cobertura', 305.34, 411.92, 717.26, 17, None, 'Cobertura/Barrilete'),
    ]
    
    for i, (pav, ac, anc, at, gab, vagas, funcao) in enumerate(pavimentos):
        row = 5 + i
        write_data_row(ws, row, [pav, ac or '', anc, at, gab, vagas or '', funcao], trace_col=8, trace_val=SRC)
    
    # Total
    row_total = 5 + len(pavimentos)
    write_data_row(ws, row_total, ['TOTAL', 6434.32, 8057.66, 14491.98, '', 138, ''], bold=True, fill=total_fill)
    
    # Parâmetros urbanísticos
    row_urb = row_total + 2
    write_title(ws, row_urb, 'PARÂMETROS URBANÍSTICOS')
    write_header(ws, row_urb+1, ['Parâmetro', 'Valor', '', '', '', '', '', 'Rastreabilidade'])
    
    params = [
        ('Zona Urbana', 'ZU-02'),
        ('Gabarito Permitido', 12),
        ('Taxa Ocupação Permitida', '1.430 m²'),
        ('Taxa Ocupação Utilizada', '701,9 m²'),
        ('Área do Terreno', '1.430,00 m²'),
    ]
    for i, (p, v) in enumerate(params):
        write_data_row(ws, row_urb+2+i, [p, v, '', '', '', '', '', SRC])
    
    print("  ✅ Quadro de Áreas")


# ═══════════════════════════════════════════════════════════
# ABA: MAPA DE PROJETOS
# ═══════════════════════════════════════════════════════════
def create_mapa_projetos(wb):
    ws = wb.create_sheet("Mapa de Projetos")
    
    write_title(ws, 1, 'MAPA DE PROJETOS E DISCIPLINAS', cols=6)
    ws.cell(row=2, column=1, value='Referência de todos os PDFs utilizados na extração de dados').font = trace_font
    
    headers = ['Disciplina', 'Arquivos', 'Principais Pranchas', 'Data/Revisão', 'Confiança', 'Observações']
    widths = [22, 12, 35, 22, 15, 40]
    write_header(ws, 4, headers, widths)
    
    disciplinas = [
        ('Arquitetura', '11 PDFs', 'ARQ 01-11', '08/06/2023', '⭐⭐⭐⭐⭐', 'Quadro de áreas legível, cotas parciais (raster)'),
        ('Fundações', '2 PDFs', '01-A e 01-B', 'Rev 01 — 25/01/2023', '⭐⭐⭐⭐', '01A vetorial (exato), 01B raster (estimado)'),
        ('Estrutural', '8 PDFs', 'Lajes Lazer/G3/G5, Escadas, Rampas', '2023', '⭐⭐⭐⭐', 'Laje protendida h=15cm, vigas h=60cm'),
        ('Hidrossanitário', '3 PDFs', 'SAN P06/P08/P11', 'REV03', '⭐⭐', 'Plantas de execução, sem resumo quantitativo'),
        ('Elétrico', '3 PDFs', 'ELE P31/P32/P39', 'REV05', '⭐⭐', 'Plantas de execução, sem resumo quantitativo'),
        ('PPCI', '5 PDFs', 'INC P08/P11/P12/P13/P15', 'REV04', '⭐⭐⭐⭐', 'SHP, SDAI, PCF, extintores, elevadores'),
        ('Climatização', '3 PDFs + 1 QTV', 'CLI P05/P07 + AVAC REV01', 'REV02 / REV01', '⭐⭐⭐⭐⭐', 'Quantitativo real: 341 splits + 3.704m tubulação'),
        ('SPDA', '1 PDF', 'SPDA P01', 'REV00', '⭐⭐⭐⭐', 'Sistema estrutural, 5 caixas equalização'),
        ('Drenagem', '1 PDF', 'MAC-QT-33490-R2', 'R2', '⭐⭐⭐⭐⭐', 'Quantitativo Maccaferri (MacDrain, MacPipe)'),
        ('ECTAS', '1 PDF', 'MDC Estoril', '—', '⭐⭐⭐⭐⭐', 'MBBR/IFAS, 412 pessoas, 67 m³/dia'),
        ('Comunicação', '19 pranchas', 'Não processado', '—', '⭐', 'Cabeamento, interfone, CFTV, telefonia'),
        ('Automação', 'Obsoleto', '—', '—', '—', 'Pasta "Obsoleto" — sem projeto vigente'),
    ]
    
    for i, d in enumerate(disciplinas):
        row = 5 + i
        write_data_row(ws, row, list(d))
        # Colorir confiança
        conf_cell = ws.cell(row=row, column=5)
        if '⭐⭐⭐⭐⭐' in str(d[4]):
            conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif '⭐⭐⭐⭐' in str(d[4]):
            conf_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        elif '⭐⭐' in str(d[4]):
            conf_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    print("  ✅ Mapa de Projetos")


# ═══════════════════════════════════════════════════════════
# ABA: FUNDAÇÕES (Estacas)
# ═══════════════════════════════════════════════════════════
def create_fundacoes(wb):
    ws = wb.create_sheet("Fundações")
    
    write_title(ws, 1, 'FUNDAÇÃO — ESTACAS HÉLICE CONTÍNUA MONITORADA')
    ws.cell(row=2, column=1, value='Engenheiro: Luciano Rosso — CREA SC 085.963-3 | Data: 26/10/2022 (Rev 01: 25/01/2023)').font = trace_font
    
    # ── PRANCHA 01-A (dados exatos) ──
    write_title(ws, 4, 'PRANCHA 01-A — DADOS EXATOS (PDF vetorial)', cols=9)
    
    headers = ['Diâmetro', 'Quantidade', 'Prof. Máx (m)', 'Metros Perf.', 'Vol. Concreto (m³)', 'Vol. c/ Perdas 20%', 'Aço (kg)', 'Confiança', 'Rastreabilidade']
    widths = [15, 14, 14, 14, 18, 18, 14, 14, 55]
    write_header(ws, 5, headers, widths)
    
    SRC_A = trace('01-A-PROJETO DE FUNDAÇÕES.pdf', 'Rev 01 — 25/01/2023', '⭐⭐⭐⭐⭐')
    
    estacas_a = [
        ('Ø600 mm', 71, 22.0, 1432, None, None, None),
        ('Ø700 mm', 51, 22.0, 1063, None, None, None),
        ('Ø800 mm', 15, 22.0, 330, None, None, None),
    ]
    
    for i, (diam, qtd, prof, metros, vol, vol_p, aco) in enumerate(estacas_a):
        write_data_row(ws, 6+i, [diam, qtd, prof, metros, vol or '', vol_p or '', aco or '', '⭐⭐⭐⭐⭐', SRC_A])
    
    write_data_row(ws, 9, ['SUBTOTAL 01-A', 137, 22.0, 2825, 979.85, 1175.83, 14025.52, '', SRC_A], bold=True, fill=subtotal_fill)
    
    # ── PRANCHA 01-B (estimativa) ──
    write_title(ws, 11, 'PRANCHA 01-B — ESTIMATIVA (PDF raster, contagem visual)', cols=9)
    write_header(ws, 12, headers)
    
    SRC_B = trace('01-B-PROJETO DE FUNDAÇÕES.pdf', 'Rev 01 — 25/01/2023', '⭐⭐⭐ (PDF raster)')
    
    estacas_b = [
        ('Ø600 mm', 36),
        ('Ø700 mm', 23),
        ('Ø800 mm', 6),
    ]
    
    for i, (diam, qtd) in enumerate(estacas_b):
        write_data_row(ws, 13+i, [diam, qtd, '', '', '', '', '', '⭐⭐⭐', SRC_B])
    
    write_data_row(ws, 16, ['SUBTOTAL 01-B', 65, '', '', '', '', '', '', SRC_B], bold=True, fill=subtotal_fill)
    
    # ── TOTAL GERAL ──
    write_data_row(ws, 18, ['TOTAL GERAL', 202, 22.0, '~4.500–5.000', '', '~2.000–2.400', '~28.000', '⭐⭐⭐⭐/⭐⭐⭐',
                            trace('01A (exato) + 01B (proporcional)', 'Rev 01', '⭐⭐⭐⭐')], bold=True, fill=total_fill)
    
    # ── AÇO DAS ESTACAS ──
    write_title(ws, 20, 'AÇO DAS ESTACAS — PRANCHA 01-A', cols=9)
    
    aco_headers = ['Diâmetro Barra', 'Tipo', 'Aplicação', 'Comprimento (m)', 'Peso (kg)', '', '', '', 'Rastreabilidade']
    write_header(ws, 21, aco_headers)
    
    SRC_ACO = trace('01-A FUNDAÇÕES — Quadro Resumo Aço', 'Rev 01', '⭐⭐⭐⭐⭐')
    
    aco = [
        ('Ø6,3 mm', 'CA-50', 'Estribos espiral c/20', 7611.34, 1902.84),
        ('Ø16,0 mm', 'CA-50', 'Longitudinal (Ø500/Ø600)', 3280.20, 5149.91),
        ('Ø20,0 mm', 'CA-50', 'Longitudinal (Ø700/Ø800)', 2811.60, 6972.77),
    ]
    
    for i, (diam, tipo, aplic, comp, peso) in enumerate(aco):
        write_data_row(ws, 22+i, [diam, tipo, aplic, comp, peso, '', '', '', SRC_ACO])
    
    write_data_row(ws, 25, ['TOTAL 01-A', '', '', '', 14025.52, '', '', '', SRC_ACO], bold=True, fill=subtotal_fill)
    write_data_row(ws, 26, ['ESTIMADO TOTAL (01A+01B)', '', 'Proporcional: 14.025 × (202/137)', '', 28000, '', '', '',
                            trace('Proporcional', '', '⭐⭐⭐')], bold=True, fill=total_fill)
    
    # ── ESPECIFICAÇÕES TÉCNICAS ──
    write_title(ws, 28, 'ESPECIFICAÇÕES TÉCNICAS DE CONCRETO', cols=9)
    
    specs = [
        ('fck', '30 MPa'),
        ('Consumo mín. cimento', '≥ 400 kg/m³'),
        ('Slump', '220 ± 20 mm (S220)'),
        ('Fator a/c', '≤ 0,60'),
        ('Agregado', 'Brita 0 / Pedrisco (4,75–12,5 mm)'),
        ('Profundidade máxima', '22,00 m'),
        ('Torque mínimo', '30 tf·m (300 bar)'),
    ]
    
    for i, (param, val) in enumerate(specs):
        write_data_row(ws, 29+i, [param, val, '', '', '', '', '', '', trace('01-A/01-B — Nota técnica', 'Rev 01', '⭐⭐⭐⭐⭐')])
    
    print("  ✅ Fundações")


# ═══════════════════════════════════════════════════════════
# ABA: CLIMATIZAÇÃO
# ═══════════════════════════════════════════════════════════
def create_climatizacao(wb):
    ws = wb.create_sheet("Climatização")
    
    SRC = '1052.094.24_QUANTITATIVO_AVAC_REV01.pdf'
    
    write_title(ws, 1, 'CLIMATIZAÇÃO / AVAC — QUANTITATIVO REAL', cols=8)
    ws.cell(row=2, column=1, value=f'Fonte: {SRC} | Confiança: ⭐⭐⭐⭐⭐ (dados do quantitativo do projetista)').font = trace_font
    
    # ── RESUMO GERAL ──
    headers = ['Equipamento', 'Capacidade', 'Quantidade', 'Unidade', '', '', '', 'Rastreabilidade']
    widths = [30, 18, 14, 10, 10, 10, 10, 55]
    write_header(ws, 4, headers, widths)
    
    resumo = [
        ('Split 9.000 BTU/h', '9.000 BTU/h', 208, 'conj'),
        ('Split 12.000 BTU/h', '12.000 BTU/h', 29, 'conj'),
        ('Split 18.000 BTU/h', '18.000 BTU/h', 90, 'conj'),
        ('Split 24.000 BTU/h', '24.000 BTU/h', 4, 'conj'),
        ('Split 36.000 BTU/h', '36.000 BTU/h', 10, 'conj'),
    ]
    
    for i, (equip, cap, qtd, un) in enumerate(resumo):
        write_data_row(ws, 5+i, [equip, cap, qtd, un, '', '', '', trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')])
    
    write_data_row(ws, 10, ['TOTAL SPLITS', '', 341, 'conj', '', '', '', trace(SRC + ' (soma)', 'REV01', '⭐⭐⭐⭐⭐')], bold=True, fill=total_fill)
    write_data_row(ws, 11, ['Caixas Polares (cobertura)', '', 10, 'un', '', '', '', trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')])
    write_data_row(ws, 12, ['Tubulação Frigorígena Total', '', 3704.21, 'm (par)', '', '', '', trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')], bold=True, fill=total_fill)
    
    # ── DISTRIBUIÇÃO POR PAVIMENTO ──
    write_title(ws, 14, 'DISTRIBUIÇÃO POR PAVIMENTO', cols=8)
    
    pav_headers = ['Pavimento', 'Multiplicador', '9k BTU', '12k BTU', '18k BTU', '24k BTU', 'Tubulação (m)', 'Rastreabilidade']
    write_header(ws, 15, pav_headers)
    
    pavimentos = [
        ('Térreo', 1, 5, 0, 4, 0, 113.60),
        ('Lazer', 1, 6, 3, 5, 0, 278.36),
        ('Diferenciado', 1, 13, 3, 8, 4, 293.30),
        ('Tipo (×10 pavimentos)', 10, 184, 23, 73, 0, 2854.40),
        ('Cobertura', 1, 0, 0, 0, 0, 164.55),
    ]
    
    for i, (pav, mult, s9, s12, s18, s24, tub) in enumerate(pavimentos):
        write_data_row(ws, 16+i, [pav, mult, s9, s12, s18, s24, tub, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')])
    
    write_data_row(ws, 21, ['TOTAL', '', 208, 29, 90, 4, 3704.21, ''], bold=True, fill=total_fill)
    
    # ── TUBULAÇÃO POR DIÂMETRO ──
    write_title(ws, 23, 'TUBULAÇÃO FRIGORÍGENA POR DIÂMETRO (estimado)', cols=8)
    tub_headers = ['Tipo', 'Diâmetro', 'Metragem Est. (m)', '', '', '', '', 'Rastreabilidade']
    write_header(ws, 24, tub_headers)
    
    tubulacao = [
        ('Linha líquido', 'Ø1/4"', 3448),
        ('Linha líquido', 'Ø3/8"', 256),
        ('Linha sucção', 'Ø3/8"', 2315),
        ('Linha sucção', 'Ø1/2"', 1036),
        ('Linha sucção', 'Ø5/8"', 256),
    ]
    
    for i, (tipo, diam, metr) in enumerate(tubulacao):
        write_data_row(ws, 25+i, [tipo, diam, metr, '', '', '', '', trace(SRC + ' (estimativa)', 'REV01', '⭐⭐⭐⭐')])
    
    print("  ✅ Climatização")


# ═══════════════════════════════════════════════════════════
# ABA: PPCI
# ═══════════════════════════════════════════════════════════
def create_ppci(wb):
    ws = wb.create_sheet("PPCI")
    
    write_title(ws, 1, 'INSTALAÇÕES PREVENTIVAS CONTRA INCÊNDIO', cols=7)
    ws.cell(row=2, column=1, value='Fonte: Pranchas INC P08/P11/P12/P13/P15 REV04').font = trace_font
    
    headers = ['Sistema', 'Item', 'Quantidade', 'Unidade', 'Observação', 'Confiança', 'Rastreabilidade']
    widths = [18, 35, 14, 10, 30, 14, 55]
    write_header(ws, 4, headers, widths)
    
    SRC_DET = '1109.020.25_INC_P15_DET_EXE_REV04.pdf'
    SRC_TIP = '1109.020.25_INC_P08_TIP_EXE_REV04.pdf'
    
    items = [
        ('SHP', 'Bomba principal incêndio', 1, 'un', 'Casa de bombas', '⭐⭐⭐⭐', SRC_DET),
        ('SHP', 'Bomba jockey', 1, 'un', 'Casa de bombas', '⭐⭐⭐⭐', SRC_DET),
        ('SHP', 'Hidrante tipo mangotinho', None, 'un', 'Quantitativo pendente', '⭐⭐⭐', SRC_DET),
        ('SDAI', 'Central de alarme', 1, 'un', '', '⭐⭐⭐⭐', SRC_DET),
        ('SDAI', 'Detector de fumaça óptico', None, 'un', 'Quantitativo pendente', '⭐⭐⭐', SRC_DET),
        ('SDAI', 'Sirene audiovisual', None, 'un', 'Quantitativo pendente', '⭐⭐⭐', SRC_DET),
        ('SDAI', 'Acionador manual', None, 'un', 'Quantitativo pendente', '⭐⭐⭐', SRC_DET),
        ('Extintores', 'PQS 6kg + CO2 6kg', '~60', 'un', 'Contagem visual nas plantas', '⭐⭐⭐', 'PPCI P11/P12/P13'),
        ('PCF', 'Porta corta-fogo P-60/P-90', '~25', 'un', 'Escadas de emergência', '⭐⭐⭐', 'PPCI P11/P12/P13'),
        ('Iluminação', 'Bloco autônomo de emergência', None, 'un', 'Quantitativo pendente', '⭐⭐⭐', 'PPCI P08-P13'),
        ('Sinalização', 'Placas fotoluminescentes', None, 'vb', '', '⭐⭐⭐', 'PPCI P08-P13'),
        ('Elevadores', 'Elevador social (cabine 120×241cm)', 2, 'un', '2 elevadores sociais', '⭐⭐⭐⭐', SRC_TIP),
    ]
    
    for i, (sist, item, qtd, un, obs, conf, src) in enumerate(items):
        write_data_row(ws, 5+i, [sist, item, qtd or 'A confirmar', un, obs, conf, trace(src, 'REV04', conf)])
    
    print("  ✅ PPCI")


# ═══════════════════════════════════════════════════════════
# ABA: SPDA
# ═══════════════════════════════════════════════════════════
def create_spda(wb):
    ws = wb.create_sheet("SPDA")
    
    SRC = '1109.020.25_SPDA_P01_EXE_REV00.pdf'
    
    write_title(ws, 1, 'SPDA — PROTEÇÃO CONTRA DESCARGAS ATMOSFÉRICAS', cols=6)
    ws.cell(row=2, column=1, value=f'Fonte: {SRC} | Norma: ABNT NBR 5419-3').font = trace_font
    
    headers = ['Componente', 'Especificação', 'Quantidade', 'Unidade', 'Confiança', 'Rastreabilidade']
    widths = [22, 40, 14, 10, 14, 55]
    write_header(ws, 4, headers, widths)
    
    items = [
        ('Tipo sistema', 'Estrutural (armadura do concreto como descida)', '', '', '⭐⭐⭐⭐⭐'),
        ('Resistência máx.', '0,2 Ω', '', '', '⭐⭐⭐⭐⭐'),
        ('Captação', 'Barra chata alumínio 70mm²', '', 'm', '⭐⭐⭐⭐'),
        ('Descidas', 'Naturais (armadura pilares) + barra aço galv. 50mm²', '', 'vb', '⭐⭐⭐⭐'),
        ('Aterramento', 'Cabo aço galv. 80mm² + fundações', '', 'vb', '⭐⭐⭐⭐'),
        ('Caixas equalização', '5 pavimentos (2º, 6º, 9º, 12º, 15º)', 5, 'un', '⭐⭐⭐⭐⭐'),
        ('Interlig. Elevador', 'Malha', 1, 'vb', '⭐⭐⭐⭐'),
        ('Interlig. SDAI', 'Equalização', 1, 'vb', '⭐⭐⭐⭐'),
        ('Interlig. Gás', 'Equalização', 1, 'vb', '⭐⭐⭐⭐'),
        ('Interlig. SHP', 'Equalização', 1, 'vb', '⭐⭐⭐⭐'),
    ]
    
    for i, (comp, spec, qtd, un, conf) in enumerate(items):
        write_data_row(ws, 5+i, [comp, spec, qtd or '', un, conf, trace(SRC, 'REV00', conf)])
    
    print("  ✅ SPDA")


# ═══════════════════════════════════════════════════════════
# ABA: ECTAS
# ═══════════════════════════════════════════════════════════
def create_ectas(wb):
    ws = wb.create_sheet("ECTAS")
    
    write_title(ws, 1, 'ECTAS — ESTAÇÃO DE TRATAMENTO DE ESGOTO', cols=7)
    ws.cell(row=2, column=1, value='Fonte: MDC Estoril (Memorial Descritivo e Cálculo) | Sistema MBBR/IFAS').font = trace_font
    
    SRC = trace('MDC Estoril — Memorial Descritivo ECTAS', '', '⭐⭐⭐⭐⭐')
    
    # Dados de projeto
    headers = ['Parâmetro', 'Valor', 'Unidade', '', '', '', 'Rastreabilidade']
    widths = [28, 18, 14, 10, 10, 10, 55]
    write_header(ws, 4, headers, widths)
    
    dados = [
        ('População residencial', 412, 'pessoas'),
        ('População comercial', 22, 'pessoas'),
        ('Vazão diária total', 67.02, 'm³/dia'),
        ('Vazão horária média', 2.79, 'm³/h'),
        ('Vazão máxima (pico)', 5.03, 'm³/h'),
        ('Carga orgânica', 21.15, 'kgDBO/dia'),
        ('DBO entrada', 316, 'mg/L'),
        ('DBO saída projetada', '< 40', 'mg/L'),
        ('Eficiência DBO', '85–95%', ''),
        ('Consumo hipoclorito', 252, 'L/mês'),
    ]
    
    for i, (param, val, un) in enumerate(dados):
        write_data_row(ws, 5+i, [param, val, un, '', '', '', SRC])
    
    # Unidades de tratamento
    write_title(ws, 16, 'UNIDADES DE TRATAMENTO', cols=7)
    
    un_headers = ['Unidade', 'Quantidade', 'Volume Útil (m³)', 'Dimensões', 'Material', '', 'Rastreabilidade']
    write_header(ws, 17, un_headers)
    
    unidades = [
        ('Gradeamento', 1, 0.42, '0,70 × 1,00 × 0,80 m', '—'),
        ('Retentor de Sólidos', 1, 4.07, 'Ø2,00 m octogonal', 'Fibra de vidro'),
        ('Reator MBBR/IFAS', 1, 14.58, 'Ø2,00 × 5,12 m octogonal', 'Fibra de vidro'),
        ('Decantador Lamelar', 1, 2.86, 'Ø1,55 × 2,00 m octogonal', 'Fibra de vidro'),
        ('Tanque Desinfecção', 1, 1.85, 'Ø1,55 × 1,29 m octogonal', 'Fibra de vidro'),
    ]
    
    for i, (un, qtd, vol, dim, mat) in enumerate(unidades):
        write_data_row(ws, 18+i, [un, qtd, vol, dim, mat, '', SRC])
    
    write_data_row(ws, 23, ['VOLUME TOTAL', '', 23.78, '', '', '', ''], bold=True, fill=total_fill)
    
    # Equipamentos
    write_title(ws, 25, 'EQUIPAMENTOS', cols=7)
    eq_headers = ['Equipamento', 'Marca/Modelo', 'Quantidade', 'Potência', '', '', 'Rastreabilidade']
    write_header(ws, 26, eq_headers)
    
    equips = [
        ('Soprador de ar', 'GARDNER 2BH7 420', 1, '2,00 kW'),
        ('Difusores bolha grossa', 'B&F Dias 105mm', 12, '—'),
        ('Bomba dosadora', 'EMEC FCE 0505', 1, '12 W'),
        ('Mídias MBBR', 'SETAC PP Ø36×10mm', '5,77 m³', '—'),
    ]
    
    for i, (eq, modelo, qtd, pot) in enumerate(equips):
        write_data_row(ws, 27+i, [eq, modelo, qtd, pot, '', '', SRC])
    
    print("  ✅ ECTAS")


# ═══════════════════════════════════════════════════════════
# ABA: DRENAGEM
# ═══════════════════════════════════════════════════════════
def create_drenagem(wb):
    ws = wb.create_sheet("Drenagem")
    
    SRC = trace('MAC-QT-33490-R2 (Maccaferri)', 'R2', '⭐⭐⭐⭐⭐')
    
    write_title(ws, 1, 'DRENAGEM — QUANTITATIVO MACCAFERRI', cols=6)
    ws.cell(row=2, column=1, value='Fonte: MAC-QT-33490-R2 | Confiança: ⭐⭐⭐⭐⭐ (quantitativo do fornecedor)').font = trace_font
    
    headers = ['Material', 'Especificação', 'Quantidade', 'Unidade', 'Confiança', 'Rastreabilidade']
    widths = [25, 35, 14, 10, 14, 55]
    write_header(ws, 4, headers, widths)
    
    items = [
        ('MacDrain® FP', 'Colchão drenante', 320, 'm²', '⭐⭐⭐⭐⭐'),
        ('MacPipe® Ø100mm', 'Tubo perfurado drenante', 150, 'm', '⭐⭐⭐⭐⭐'),
        ('Mactex® H26.2', 'Geotêxtil não-tecido 130g/m²', 230, 'm²', '⭐⭐⭐⭐⭐'),
    ]
    
    for i, (mat, spec, qtd, un, conf) in enumerate(items):
        write_data_row(ws, 5+i, [mat, spec, qtd, un, conf, SRC])
    
    print("  ✅ Drenagem")


# ═══════════════════════════════════════════════════════════
# ABA: GÁS (GLP)
# ═══════════════════════════════════════════════════════════
def create_gas(wb):
    ws = wb.create_sheet("Gás GLP")
    
    SRC = trace('1109.020.25_INC_P08_TIP_EXE_REV04.pdf', 'REV04', '⭐⭐⭐⭐⭐')
    
    write_title(ws, 1, 'INSTALAÇÕES DE GÁS (GLP)', cols=6)
    ws.cell(row=2, column=1, value='Fonte: PPCI P08 (Pavimento Tipo) — dados exatos por pavimento').font = trace_font
    
    headers = ['Item', 'Especificação', 'Qtd/Pavimento', 'Pavimentos', 'Total', 'Rastreabilidade']
    widths = [25, 30, 14, 14, 14, 55]
    write_header(ws, 4, headers, widths)
    
    items = [
        ('Medidores de gás', 'G2,5', 5, 10, 50),
        ('Churrasqueiras', '484 kcal/min cada', 4, 10, 40),
        ('Apartamentos atendidos', 'Final 01-05', 5, 10, 50),
    ]
    
    for i, (item, spec, qtd_pav, pavs, total) in enumerate(items):
        write_data_row(ws, 5+i, [item, spec, qtd_pav, pavs, total, SRC])
    
    # Tubulação
    write_title(ws, 9, 'TUBULAÇÃO PEX MULTICAMADAS', cols=6)
    tub_headers = ['Material', 'Diâmetro Ext.', 'Diâmetro Int.', 'Referência', '', 'Rastreabilidade']
    write_header(ws, 10, tub_headers)
    
    tubos = [
        ('PEX Multicamadas', '16 mm', '12,4 mm (3/8")', ''),
        ('PEX Multicamadas', '20 mm', '16 mm (1/2")', ''),
        ('PEX Multicamadas', '25 mm', '20 mm (3/4")', ''),
        ('PEX Multicamadas', '32 mm', '25 mm (1")', ''),
    ]
    
    for i, (mat, de, di, ref) in enumerate(tubos):
        write_data_row(ws, 11+i, [mat, de, di, ref, '', SRC])
    
    print("  ✅ Gás GLP")


# ═══════════════════════════════════════════════════════════
# ABA: INSTALAÇÕES (Resumo)
# ═══════════════════════════════════════════════════════════
def create_instalacoes_resumo(wb):
    ws = wb.create_sheet("Instalações (Resumo)")
    
    write_title(ws, 1, 'RESUMO DE INSTALAÇÕES — DADOS PARCIAIS', cols=7)
    ws.cell(row=2, column=1, value='⚠️ PDFs disponíveis são plantas de execução, não resumos quantitativos. Estimativas via índices.').font = Font(name='Calibri', size=10, italic=True, color="CC0000")
    
    headers = ['Disciplina', 'Dados Disponíveis', 'Índice R$/m²', 'Estimativa (14.492 m²)', 'Confiança', 'Fonte Principal', 'Observação']
    widths = [22, 22, 14, 20, 14, 30, 35]
    write_header(ws, 4, headers, widths)
    
    items = [
        ('Climatização', 'Quantitativo real', 'R$ 80–120', 'R$ 1,16M – 1,74M', '⭐⭐⭐⭐⭐', 'AVAC REV01', '341 splits + 3.704m tubo'),
        ('SPDA', 'Especificações', 'R$ 8–15', 'R$ 116k – 217k', '⭐⭐⭐⭐', 'SPDA P01', 'Sistema estrutural'),
        ('Drenagem', 'Quantitativo real', '—', 'R$ 50k – 80k', '⭐⭐⭐⭐⭐', 'Maccaferri MAC-QT', 'Material MacDrain/MacPipe'),
        ('ECTAS', 'Projeto completo', '—', 'R$ 200k – 350k', '⭐⭐⭐⭐⭐', 'MDC Estoril', 'MBBR/IFAS, 67 m³/dia'),
        ('Elétrico', 'Parcial (plantas)', '—', '—', '⭐⭐', 'ELE P31/P32/P39', 'Precisa quantitativo'),
        ('Hidrossanitário', 'Parcial (plantas)', '—', '—', '⭐⭐', 'SAN P06/P08/P11', 'Precisa quantitativo'),
        ('Comunicação/Telecom', 'Identificado', 'R$ 15–25', 'R$ 217k – 362k', '⭐', '19 pranchas', 'Não processado'),
        ('Automação', 'Inexistente', 'R$ 10–20', 'R$ 145k – 290k', '—', 'Pasta "Obsoleto"', 'Sem projeto vigente'),
        ('Gás GLP', 'Dados parciais', '—', '—', '⭐⭐⭐⭐⭐', 'PPCI P08', '50 medidores, 40 churrasqueiras'),
    ]
    
    for i, item in enumerate(items):
        write_data_row(ws, 5+i, list(item))
        # Colorir confiança
        conf_cell = ws.cell(row=5+i, column=5)
        conf = item[4]
        if '⭐⭐⭐⭐⭐' in conf:
            conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif '⭐⭐⭐⭐' in conf:
            conf_cell.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        elif '⭐⭐' in conf:
            conf_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Total estimado
    row_total = 5 + len(items) + 1
    write_data_row(ws, row_total, ['TOTAL ESPECÍFICOS', '', '', 'R$ 1,89M – 3,04M', '', '', 'Estimativa consolidada'], bold=True, fill=total_fill)
    
    print("  ✅ Instalações (Resumo)")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
def main():
    print("📋 Criando planilha limpa do zero...\n")
    
    wb = openpyxl.Workbook()
    # Remove a aba default
    wb.remove(wb.active)
    
    create_capa(wb)
    create_quadro_areas(wb)
    create_mapa_projetos(wb)
    create_fundacoes(wb)
    create_climatizacao(wb)
    create_ppci(wb)
    create_spda(wb)
    create_ectas(wb)
    create_drenagem(wb)
    create_gas(wb)
    create_instalacoes_resumo(wb)
    
    # Definir cores das abas
    tab_colors = {
        'CAPA': '1F4E79',
        'Quadro de Áreas': '4472C4',
        'Mapa de Projetos': '4472C4',
        'Fundações': 'ED7D31',
        'Climatização': '70AD47',
        'PPCI': 'FF0000',
        'SPDA': 'FFC000',
        'ECTAS': '5B9BD5',
        'Drenagem': '548235',
        'Gás GLP': 'BF8F00',
        'Instalações (Resumo)': '7030A0',
    }
    
    for name, color in tab_colors.items():
        if name in wb.sheetnames:
            wb[name].sheet_properties.tabColor = color
    
    print(f"\n💾 Salvando: {OUTPUT}")
    wb.save(OUTPUT)
    
    size_kb = os.path.getsize(OUTPUT) / 1024
    print(f"✅ Planilha criada! ({size_kb:.0f} KB, {len(wb.sheetnames)} abas)")
    
    print("\n📊 Abas:")
    for name in wb.sheetnames:
        print(f"   📑 {name}")


if __name__ == '__main__':
    main()
