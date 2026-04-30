#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Módulo Grupo B - Gerenciamento
Gera 7 abas de gerenciamento para o orçamento paramétrico

Abas geradas:
1. EPCs - Equipamentos de proteção coletiva
2. CANTEIRO - Instalações de canteiro
3. Controle Tecnologico - Controle de concreto
4. Ensaios - Ensaios de controle
5. Cont.Tecnol. - Detalhamento controle tecnológico
6. Ger_Tec e Adm - Gerenciamento técnico e administrativo
7. Ger_Executivo - Orçamento executivo completo (CORE)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ============================================================================
# ESTILOS PADRÃO
# ============================================================================

YELLOW_TAB = "FFFFFF00"  # Yellow tab color for EPCs, CANTEIRO, etc.
THEME_TAB = None  # Theme color (default) for Ger_Executivo

HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ============================================================================
# ABA 1: EPCs - Equipamentos de Proteção Coletiva
# ============================================================================

def create_epcs(wb, projeto_data):
    """Cria aba EPCs - Equipamentos de proteção coletiva"""
    ws = wb.create_sheet("EPCs")
    ws.sheet_properties.tabColor = YELLOW_TAB
    
    # Header rows (R1-R4)
    ws['A1'] = 'Obra'
    ws['B1'] = projeto_data.get('nome_obra', '')
    
    ws['A2'] = 'Empresa'
    ws['B2'] = projeto_data.get('empresa', '')
    
    ws['A3'] = 'Revisão'
    ws['B3'] = projeto_data.get('revisao', 'R00')
    ws['K3'] = 'Guia pavimentos'
    
    ws['A4'] = 'Equipamentos de proteção coletiva'
    ws['K4'] = 'Pavimento'
    ws['L4'] = 'Área'
    ws['M4'] = 'Perímetro'
    ws['N4'] = 'Pé Direito'
    
    # Apply header styling
    for cell in ['A4', 'K4', 'L4', 'M4', 'N4']:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    # Section: Bandejas de proteção (R5-R6)
    ws['A5'] = 'Bandejas de proteção'
    ws['A5'].font = Font(bold=True)
    
    # Section: Guarda-corpo de vãos (R14)
    ws['A14'] = 'Guarda - corpo de vãos'
    ws['A14'].font = Font(bold=True)
    
    # Section: Guarda-corpo de desforma (R20)
    ws['A20'] = 'Guarda - corpo de desforma'
    ws['A20'].font = Font(bold=True)
    
    # Section: Fechamento removível (R27)
    ws['A27'] = 'Fechamento removível de vãos em madeira '
    ws['A27'].font = Font(bold=True)
    
    # Section: Tela fachadeira (R38)
    ws['A38'] = 'Tela fachadeira'
    ws['A38'].font = Font(bold=True)
    
    # Section: SLQA (R44)
    ws['A44'] = 'SLQA'
    ws['A44'].font = Font(bold=True)
    
    # Section: Tapume (R52)
    ws['A52'] = 'Tapume'
    ws['A52'].font = Font(bold=True)
    
    # Section: Linha de Vida (R59)
    ws['A59'] = 'Linha de Vida'
    ws['A59'].font = Font(bold=True)
    
    # Section: Cobertura para Proteção de Pedestres (R65)
    ws['A65'] = 'Cobertura para Proteção de Pedestres'
    ws['A65'].font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 12
    
    return wb


# ============================================================================
# ABA 2: CANTEIRO - Instalações de Canteiro
# ============================================================================

def create_canteiro(wb, projeto_data):
    """Cria aba CANTEIRO - Instalações de canteiro"""
    ws = wb.create_sheet("CANTEIRO")
    ws.sheet_properties.tabColor = YELLOW_TAB
    
    # Header rows (R1-R5)
    ws['A1'] = 'Obra'
    ws['B1'] = projeto_data.get('nome_obra', '')
    
    ws['A2'] = 'Empresa'
    ws['B2'] = projeto_data.get('empresa', '')
    
    ws['A3'] = 'Revisão'
    ws['B3'] = projeto_data.get('revisao', 'R00')
    ws['O3'] = 'Número de funcionários'
    ws['S3'] = 'Produtividade média (h/m2)'
    
    ws['A4'] = 'Instalações de canteiro'
    ws['O4'] = 'Área construída'
    ws['P4'] = projeto_data.get('area_total', 0)
    ws['S4'] = 'A.C/prazo'
    
    ws['A5'] = 'Instalações de canteiro'
    ws['O5'] = 'Prazo da obra'
    ws['P5'] = projeto_data.get('prazo_meses', 0)
    ws['Q5'] = 'meses'
    ws['S5'] = 'Pm'
    ws['U5'] = 'h/m2 '
    
    # Section headers
    ws['A7'] = 'Descrição'
    ws['B7'] = ws['C7'] = ws['D7'] = ''  # Merged in practice
    ws['O7'] = 'Pavimentos'
    
    # Section: Mobiliário de canteiro (R18)
    ws['A18'] = 'Mobiliário de canteiro'
    ws['A18'].font = Font(bold=True)
    
    ws['O19'] = 'Área de vestiário'
    ws['P19'] = 1.5
    ws['Q19'] = 'm²/funcionários'
    
    ws['B20'] = 'Descrição'
    ws['C20'] = 'Unidade'
    ws['D20'] = 'Quantidade'
    ws['E20'] = 'Custo unitário'
    ws['F20'] = 'Custo total'
    
    # Apply header styling
    for cell in ['B7', 'C7', 'D7', 'B20', 'C20', 'D20', 'E20', 'F20']:
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 12
    ws.column_dimensions['S'].width = 15
    
    return wb


# ============================================================================
# ABA 3: Controle Tecnologico - Controle de Concreto
# ============================================================================

def create_controle_tecnologico(wb, projeto_data):
    """Cria aba Controle Tecnologico - Controle de concreto"""
    ws = wb.create_sheet("Controle Tecnologico")
    ws.sheet_properties.tabColor = YELLOW_TAB
    
    # Header row (R6)
    ws['A6'] = 'Item'
    ws['B6'] = 'Descrição/Especificação'
    ws['C6'] = 'Unid.'
    ws['D6'] = 'V. Unit. (R$)'
    ws['E6'] = 'Qtde.'
    ws['F6'] = 'V. Total\n(R$)'
    ws['G6'] = 'Observações'
    
    # Apply header styling
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = f'{col}6'
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(wrap_text=True)
    
    # Section title (R7)
    ws['B7'] = 'CONTROLE TECNOLÓGICO  - CONCRETO'
    ws['B7'].font = Font(bold=True)
    
    # Data rows (R8-R11)
    ws['A8'] = 1
    ws['B8'] = 'Laudo e ART'
    ws['C8'] = 'vb'
    ws['D8'] = 1450
    ws['E8'] = 1
    ws['F8'] = '=D8*E8'
    
    ws['A9'] = 2
    ws['B9'] = 'Controle Tecnologico - Coleta materiais da obra'
    ws['C9'] = 'unid.'
    ws['D9'] = 20
    ws['E9'] = 0  # To be filled
    ws['F9'] = '=D9*E9'
    ws['G9'] = '4 CPs por caminhão betoneira de 8m³ '
    
    ws['A10'] = 3
    ws['B10'] = 'Controle Tecnológico - Ensaio de concreto endurecido/grouth - compressão simples / Detalhe: Concreto'
    ws['C10'] = 'cps'
    ws['D10'] = 12.5
    ws['E10'] = 0  # To be filled
    ws['F10'] = '=D10*E10'
    ws['G10'] = '4 CPs por caminhão betoneira de 8m³ '
    
    ws['A11'] = 4
    ws['B11'] = 'Determinação do módulo de deformação tangente inicial (Eci) (módulo de elasticidade), Incluindo Avaliação Estática dos Resultados (NBR 8522 e NBR 12655)'
    ws['C11'] = 'ensaio'
    ws['D11'] = 480
    ws['E11'] = 0  # To be filled
    ws['F11'] = '=D11*E11'
    ws['G11'] = 'A cada 5 pav. (e quando houver troca de fornecedor)'
    
    # Subtotal (R12)
    ws['C12'] = 'SUBTOTAL:'
    ws['C12'].font = Font(bold=True)
    ws['F12'] = '=SUM(F8:F11)'
    
    # VALOR/M3 (R14)
    ws['C14'] = 'VALOR/M3'
    ws['C14'].font = Font(bold=True)
    ws['F14'] = '=IF(E5>0,F12/E5,0)'
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    
    return wb


# ============================================================================
# ABA 4: Ensaios - Ensaios de Controle
# ============================================================================

def create_ensaios(wb, projeto_data):
    """Cria aba Ensaios - Ensaios de controle"""
    ws = wb.create_sheet("Ensaios ")  # Note: space after name to match reference
    ws.sheet_properties.tabColor = YELLOW_TAB
    
    # Header rows (R1-R4)
    ws['C1'] = 'Projeto'
    ws['D1'] = projeto_data.get('nome_obra', '')
    
    ws['C2'] = 'Empresa'
    ws['D2'] = projeto_data.get('empresa', '')
    
    ws['C3'] = 'Revisão'
    ws['D3'] = projeto_data.get('revisao', 'R00')
    
    ws['A4'] = 'ENSAIOS'
    ws['A4'].font = Font(bold=True, size=14)
    
    # Column headers (R5)
    ws['A5'] = 'Item'
    ws['B5'] = 'Descrição/Especificação'
    ws['C5'] = 'Unid.'
    ws['D5'] = 'V. Unit. (R$)'
    ws['E5'] = 'Qtde.'
    ws['F5'] = 'V. Total\n(R$)'
    ws['G5'] = 'Observações'
    
    # Apply header styling
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = f'{col}5'
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(wrap_text=True)
    
    # Section 1: CONTROLE TECNOLÓGICO - CONCRETO (R6)
    ws['B6'] = 'CONTROLE TECNOLÓGICO  - CONCRETO'
    ws['B6'].font = Font(bold=True)
    ws['F6'] = '=SUM(F7:F12)'
    
    ws['A7'] = 1
    ws['B7'] = 'Laudo e ART'
    ws['C7'] = 'vb'
    ws['D7'] = 1450
    ws['E7'] = 1
    ws['F7'] = '=D7*E7'
    
    # Section 2: ACUSTICA (R13)
    ws['B13'] = 'ACUSTICA - ENSAIO DE RUIDOS ( PISO + FACHADA)'
    ws['B13'].font = Font(bold=True)
    ws['F13'] = '=SUM(F14:F15)'
    
    # Section 3: ENSAIO PARA CAIXILHOS (R16)
    ws['B16'] = 'ENSAIO PARA CAIXILHOS'
    ws['B16'].font = Font(bold=True)
    ws['F16'] = '=SUM(F17:F19)'
    
    # Section 4: ENSAIO PARA GUARDA-CORPO (R20)
    ws['B20'] = 'ENSAIO PARA GUARDA-CORPO'
    ws['B20'].font = Font(bold=True)
    ws['F20'] = '=F21'
    
    # Section 5: ENSAIOS PARA ARGAMASSA DE REVESTIMENTO (R22)
    ws['B22'] = 'ENSAIOS PARA ARGAMASSA DE REVESTIMENTO (FACHADA)'
    ws['B22'].font = Font(bold=True)
    ws['F22'] = '=F23'
    
    # Section 6: ENSAIO PARA REVESTIMENTOS CERÂMICOS (R24)
    ws['B24'] = 'ENSAIO PARA REVESTIMENTOS CERÂMICOS'
    ws['B24'].font = Font(bold=True)
    ws['F24'] = '=F25'
    
    # Section 7: ENSAIO PARA ANCORAGEM (R26)
    ws['B26'] = 'ENSAIO PARA ANCORAGEM (GANHOS DA COBERTURA)'
    ws['B26'].font = Font(bold=True)
    ws['F26'] = '=SUM(F27:F28)'
    
    # Section 8: ENSAIOS PARA ALVENARIA ESTRUTURAL (R29)
    ws['B29'] = 'ENSAIOS PARA ALVENARIA ESTRUTURAL'
    ws['B29'].font = Font(bold=True)
    ws['F29'] = '=SUM(F30:F35)'
    
    # SUBTOTAL (R36)
    ws['C36'] = 'SUBTOTAL:'
    ws['C36'].font = Font(bold=True, size=12)
    ws['F36'] = '=SUM(F6,F13,F16,F20,F22,F24,F26,F29)'
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50
    
    return wb


# ============================================================================
# ABA 5: Cont.Tecnol. - Detalhamento Controle Tecnológico
# ============================================================================

def create_cont_tecnol(wb, projeto_data):
    """Cria aba Cont.Tecnol. - Detalhamento controle tecnológico"""
    ws = wb.create_sheet("Cont.Tecnol.")
    ws.sheet_properties.tabColor = YELLOW_TAB
    
    # Reference to Controle Tecnologico (R5)
    ws['E5'] = 'COND.01'
    
    # Column headers (R6)
    ws['A6'] = 'Item'
    ws['B6'] = 'Descrição/Especificação'
    ws['C6'] = 'Unid.'
    ws['D6'] = 'V. Unit. (R$)'
    ws['E6'] = 'Qtde.'
    ws['F6'] = 'V. Total\n(R$)'
    ws['G6'] = 'Observações'
    
    # Apply header styling
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = f'{col}6'
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
        ws[cell].alignment = Alignment(wrap_text=True)
    
    # Section: FUNDAÇÃO PROFUNDA (R5)
    ws['A5'] = 'FUNDAÇÃO PROFUNDA'
    ws['A5'].font = Font(bold=True)
    ws['D5'] = 'm³'
    
    # Section title (R7)
    ws['B7'] = 'CONTROLE TECNOLÓGICO  - CONCRETO'
    ws['B7'].font = Font(bold=True)
    
    # Data rows (similar structure to Controle Tecnologico)
    ws['A8'] = 1
    ws['B8'] = 'Laudo e ART'
    ws['C8'] = 'vb'
    ws['D8'] = 1450
    ws['E8'] = 1
    ws['F8'] = '=D8*E8'
    
    ws['A9'] = 2
    ws['B9'] = 'Controle Tecnologico - Coleta materiais da obra'
    ws['C9'] = 'unid.'
    ws['D9'] = 20
    ws['E9'] = 0  # To be filled
    ws['F9'] = '=D9*E9'
    ws['G9'] = '4 CPs por caminhão betoneira de 8m³ '
    
    ws['A10'] = 3
    ws['B10'] = 'Controle Tecnológico - Ensaio de concreto endurecido/grouth - compressão simples / Detalhe: Concreto'
    ws['C10'] = 'cps'
    ws['D10'] = 12.5
    ws['E10'] = 0  # To be filled
    ws['F10'] = '=D10*E10'
    ws['G10'] = '4 CPs por caminhão betoneira de 8m³ '
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    
    return wb


# ============================================================================
# ABA 6: Ger_Tec e Adm - Gerenciamento Técnico e Administrativo
# ============================================================================

def create_ger_tec_adm(wb, projeto_data):
    """Cria aba Ger_Tec e Adm - Gerenciamento técnico e administrativo"""
    ws = wb.create_sheet("Ger_Tec e Adm")
    
    # Header rows (R4-R7)
    ws['G4'] = 'PROJETO:'
    ws['I4'] = projeto_data.get('nome_obra', '')
    ws['O4'] = 'REVISÃO:'
    ws['P4'] = 'DATA'
    
    ws['G5'] = 'EMPRESA:'
    ws['I5'] = projeto_data.get('empresa', '')
    ws['O5'] = projeto_data.get('revisao', 'R00')
    ws['P5'] = datetime.now()
    ws['P5'].number_format = 'DD/MM/YYYY'
    
    ws['A6'] = 'Unidade Construtiva'
    ws['B6'] = 'Célula Construtiva'
    ws['C6'] = 'Etapa'
    ws['D6'] = 'Subetapa'
    ws['E6'] = 'Serviço'
    ws['G6'] = 'GERENCIAMENTO TÉCNICO E ADMINISTRATIVO'
    ws['G6'].font = Font(bold=True, size=12)
    
    # Column headers (R7)
    headers = ['A', 'B', 'C', 'D', 'E', '', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    header_values = [0, 0, 0, 0, 0, '', 'Atalho', 'NIVEL', 'Item', 'Descrição', 'Unidade', 'Quant. ', 'Preço un', 'Total', 'Considerações', 'Situação']
    
    for col_letter, value in zip(headers, header_values):
        if col_letter:
            cell = f'{col_letter}7'
            ws[cell] = value
            ws[cell].font = HEADER_FONT
            ws[cell].fill = HEADER_FILL
    
    # Row 8: CÉLULA CONSTRUTIVA - SERVIÇOS TÉCNICOS
    ws['A8'] = 0
    ws['B8'] = 1
    ws['C8'] = 0
    ws['D8'] = 0
    ws['E8'] = 0
    ws['H8'] = 'CÉLULA CONSTRUTIVA'
    ws['I8'] = '01'
    ws['J8'] = 'SERVIÇOS TÉCNICOS'
    ws['N8'] = '=SUMIF(B:B,B8,N:N)-N8'  # Sum all services in this cell
    ws['H8'].font = Font(bold=True)
    
    # Row 9: ETAPA - SERVIÇOS TÉCNICOS
    ws['A9'] = 0
    ws['B9'] = 1
    ws['C9'] = 1
    ws['D9'] = 0
    ws['E9'] = 0
    ws['H9'] = 'ETAPA'
    ws['I9'] = '01.001'
    ws['J9'] = 'SERVIÇOS TÉCNICOS'
    ws['N9'] = '=SUMIF(C:C,C9,N:N)-N9'
    ws['H9'].font = Font(bold=True)
    
    # Row 10: SUBETAPA - Projetos
    ws['A10'] = 0
    ws['B10'] = 1
    ws['C10'] = 1
    ws['D10'] = 1
    ws['E10'] = 0
    ws['F10'] = 12
    ws['H10'] = 'SUBETAPA'
    ws['I10'] = '01.001.001'
    ws['J10'] = 'Projetos'
    ws['N10'] = '=SUMIF(D:D,D10,N:N)-N10'
    ws['H10'].font = Font(bold=True)
    
    # Example services (R11-R20)
    services = [
        ('01.001.001.001', 'Projeto arquitetônico legal', 'vb', 1, 150000),
        ('01.001.001.002', 'Projeto arquitetônico executivo', 'vb', 1, 50000),
        ('01.001.001.003', 'Projeto estrutural', 'vb', 1, 145500),
        ('01.001.001.004', 'Projeto de instalações elétricas', 'vb', 1, 27000),
        ('01.001.001.005', 'Projeto hidrossanitário/drenagem/pluvial', 'vb', 1, 32047.37),
        ('01.001.001.006', 'Projeto preventivo (PPCI) e Gás', 'vb', 1, 33458.75),
        ('01.001.001.007', 'Projeto de climatização', 'vb', 1, 9018.78),
        ('01.001.001.008', 'Projeto de gás', 'vb', 1, 5725.98),
        ('01.001.001.009', 'Projeto de Paisagismo', 'vb', 1, 20781.69),
        ('01.001.001.010', 'Projeto arquitetura de interiores', 'vb', 1, 56000),
    ]
    
    for idx, (item, desc, unit, qty, price) in enumerate(services, start=11):
        row = idx
        ws[f'A{row}'] = 0
        ws[f'B{row}'] = 1
        ws[f'C{row}'] = 1
        ws[f'D{row}'] = 1
        ws[f'E{row}'] = idx - 10  # Service number
        ws[f'H{row}'] = 'SERVIÇO'
        ws[f'I{row}'] = item
        ws[f'J{row}'] = desc
        ws[f'K{row}'] = unit
        ws[f'L{row}'] = qty
        ws[f'M{row}'] = price
        ws[f'N{row}'] = f'=L{row}*M{row}'
        ws[f'P{row}'] = 'Finalizado'
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 30
    ws.column_dimensions['P'].width = 15
    
    return wb


# ============================================================================
# ABA 7: Ger_Executivo - Orçamento Executivo Completo (CORE)
# ============================================================================

# Estrutura de 21 Células Construtivas
CELULAS_CONSTRUTIVAS = [
    ('01', 'SERVIÇOS PRELIMINARES'),
    ('02', 'FUNDAÇÕES E CONTENÇÕES'),
    ('03', 'SUPRAESTRUTURA'),
    ('04', 'ALVENARIA E VEDAÇÕES'),
    ('05', 'INSTALAÇÕES HIDROSSANITÁRIAS'),
    ('06', 'INSTALAÇÕES ELÉTRICAS'),
    ('07', 'INSTALAÇÕES PREVENTIVAS E GLP'),
    ('08', 'INSTALAÇÕES CLIMATIZAÇÃO E EXAUSTÃO'),
    ('09', 'INSTALAÇÕES SPDA'),
    ('10', 'IMPERMEABILIZAÇÃO'),
    ('11', 'REVESTIMENTOS ARGAMASSADOS'),
    ('12', 'ACABAMENTOS DE TETO'),
    ('13', 'ACABAMENTOS DE PISO E PAREDE'),
    ('14', 'PINTURA INTERNA'),
    ('15', 'PINTURA EXTERNA'),
    ('16', 'ESQUADRIAS'),
    ('17', 'COBERTURA'),
    ('18', 'EQUIPAMENTOS E SISTEMAS ESPECIAIS'),
    ('19', 'REVESTIMENTOS DE FACHADA'),
    ('20', 'SERVIÇOS COMPLEMENTARES E FINAIS'),
    ('21', 'IMPREVISTOS E CONTINGÊNCIAS'),
]


def create_ger_executivo(wb, projeto_data):
    """
    Cria aba Ger_Executivo - Orçamento executivo completo (CORE)
    
    Estrutura hierárquica:
    - CÉLULA CONSTRUTIVA (21 células)
      - ETAPA (162 etapas totais)
        - SUBETAPA
          - SERVIÇO (com fórmulas =L*M para total)
    
    Total: ~2803 linhas
    """
    ws = wb.create_sheet("Ger_Executivo")
    # Theme tab color (default blue)
    
    # Header rows (R4-R7)
    ws['G4'] = 'PROJETO:'
    ws['I4'] = projeto_data.get('nome_obra', '')
    ws['O4'] = 'REVISÃO:'
    ws['P4'] = 'DATA'
    
    ws['G5'] = 'EMPRESA:'
    ws['I5'] = projeto_data.get('empresa', '')
    ws['O5'] = projeto_data.get('revisao', 'R00')
    ws['P5'] = datetime.now()
    ws['P5'].number_format = 'DD/MM/YYYY'
    
    ws['A6'] = 'Unidade Construtiva'
    ws['B6'] = 'Célula Construtiva'
    ws['C6'] = 'Etapa'
    ws['D6'] = 'Subetapa'
    ws['E6'] = 'Serviço'
    ws['G6'] = 'GERENCIAMENTO EXECUTIVO'
    ws['G6'].font = Font(bold=True, size=14)
    
    # Column headers (R7)
    headers_r7 = {
        'A7': 0, 'B7': 0, 'C7': 0, 'D7': 0, 'E7': 0,
        'G7': 'Atalho', 'H7': 'NIVEL', 'I7': 'Item', 'J7': 'Descrição',
        'K7': 'Unidade', 'L7': 'Quant. ', 'M7': 'Preço un', 'N7': 'Total',
        'O7': 'Considerações', 'P7': 'Situação'
    }
    
    for cell, value in headers_r7.items():
        ws[cell] = value
        ws[cell].font = HEADER_FONT
        ws[cell].fill = HEADER_FILL
    
    current_row = 8
    
    # Gerar estrutura de todas as 21 células construtivas
    for cell_idx, (cell_num, cell_name) in enumerate(CELULAS_CONSTRUTIVAS, start=1):
        # CÉLULA CONSTRUTIVA
        ws[f'A{current_row}'] = 0
        ws[f'B{current_row}'] = cell_idx
        ws[f'C{current_row}'] = 0
        ws[f'D{current_row}'] = 0
        ws[f'E{current_row}'] = 0
        ws[f'H{current_row}'] = 'CÉLULA CONSTRUTIVA'
        ws[f'I{current_row}'] = cell_num
        ws[f'J{current_row}'] = cell_name
        ws[f'N{current_row}'] = f'=SUMIF(B:B,B{current_row},N:N)-N{current_row}'
        ws[f'H{current_row}'].font = Font(bold=True, size=12)
        
        current_row += 1
        
        # Exemplo de estrutura para CÉLULA 01 - SERVIÇOS PRELIMINARES
        if cell_num == '01':
            # ETAPA 01.001 - PREPARAÇÃO DO TERRENO
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 1
            ws[f'D{current_row}'] = 0
            ws[f'E{current_row}'] = 0
            ws[f'H{current_row}'] = 'ETAPA'
            ws[f'I{current_row}'] = '01.001'
            ws[f'J{current_row}'] = 'PREPARAÇÃO DO TERRENO'
            ws[f'N{current_row}'] = f'=SUMIF(C:C,C{current_row},N:N)-N{current_row}'
            ws[f'H{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # SUBETAPA 01.001.001 - Limpeza do terreno
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 1
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 0
            ws[f'H{current_row}'] = 'SUBETAPA'
            ws[f'I{current_row}'] = '01.001.001'
            ws[f'J{current_row}'] = 'Limpeza do terreno e demolições'
            ws[f'N{current_row}'] = f'=SUMIF(D:D,D{current_row},N:N)-N{current_row}'
            ws[f'H{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # SERVIÇO 01.001.001.001
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 1
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 1
            ws[f'H{current_row}'] = 'SERVIÇO'
            ws[f'I{current_row}'] = '01.001.001.001'
            ws[f'J{current_row}'] = 'Limpeza manual de terreno'
            ws[f'K{current_row}'] = 'm2'
            ws[f'L{current_row}'] = 0  # To be filled
            ws[f'M{current_row}'] = 0  # To be filled
            ws[f'N{current_row}'] = f'=L{current_row}*M{current_row}'
            ws[f'P{current_row}'] = 'Pendente'
            current_row += 1
            
            # ETAPA 01.002 - MOVIMENTAÇÃO DE TERRA
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 2
            ws[f'D{current_row}'] = 0
            ws[f'E{current_row}'] = 0
            ws[f'H{current_row}'] = 'ETAPA'
            ws[f'I{current_row}'] = '01.002'
            ws[f'J{current_row}'] = 'MOVIMENTAÇÃO DE TERRA'
            ws[f'N{current_row}'] = f'=SUMIF(C:C,C{current_row},N:N)-N{current_row}'
            ws[f'H{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # SUBETAPA 01.002.001 - Escavações
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 2
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 0
            ws[f'H{current_row}'] = 'SUBETAPA'
            ws[f'I{current_row}'] = '01.002.001'
            ws[f'J{current_row}'] = 'Escavações'
            ws[f'N{current_row}'] = f'=SUMIF(D:D,D{current_row},N:N)-N{current_row}'
            ws[f'H{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # SERVIÇO 01.002.001.001
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 2
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 1
            ws[f'H{current_row}'] = 'SERVIÇO'
            ws[f'I{current_row}'] = '01.002.001.001'
            ws[f'J{current_row}'] = 'Escavação mecânica de valas'
            ws[f'K{current_row}'] = 'm3'
            ws[f'L{current_row}'] = 0
            ws[f'M{current_row}'] = 0
            ws[f'N{current_row}'] = f'=L{current_row}*M{current_row}'
            ws[f'P{current_row}'] = 'Pendente'
            current_row += 1
            
            # ETAPA 01.003 - LOCAÇÃO DA OBRA
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 3
            ws[f'D{current_row}'] = 0
            ws[f'E{current_row}'] = 1
            ws[f'H{current_row}'] = 'ETAPA'
            ws[f'I{current_row}'] = '01.003'
            ws[f'J{current_row}'] = 'LOCAÇÃO DA OBRA'
            ws[f'N{current_row}'] = f'=SUMIF(C:C,C{current_row},N:N)-N{current_row}'
            ws[f'H{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # SUBETAPA 01.003.001
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 3
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 0
            ws[f'H{current_row}'] = 'SUBETAPA'
            ws[f'I{current_row}'] = '01.003.001'
            ws[f'J{current_row}'] = 'Locação de obra'
            ws[f'N{current_row}'] = f'=SUMIF(D:D,D{current_row},N:N)-N{current_row}'
            ws[f'H{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # SERVIÇO 01.003.001.001
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 3
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 1
            ws[f'H{current_row}'] = 'SERVIÇO'
            ws[f'I{current_row}'] = '01.003.001.001'
            ws[f'J{current_row}'] = 'Locação da obra'
            ws[f'K{current_row}'] = 'm'
            ws[f'L{current_row}'] = 0
            ws[f'M{current_row}'] = 0
            ws[f'N{current_row}'] = f'=L{current_row}*M{current_row}'
            ws[f'P{current_row}'] = 'Pendente'
            current_row += 1
            
            # SERVIÇO 01.003.001.002
            ws[f'A{current_row}'] = 0
            ws[f'B{current_row}'] = cell_idx
            ws[f'C{current_row}'] = 3
            ws[f'D{current_row}'] = 1
            ws[f'E{current_row}'] = 2
            ws[f'H{current_row}'] = 'SERVIÇO'
            ws[f'I{current_row}'] = '01.003.001.002'
            ws[f'J{current_row}'] = 'Locação topográfica da obra'
            ws[f'K{current_row}'] = 'vb'
            ws[f'L{current_row}'] = 1
            ws[f'M{current_row}'] = 40000
            ws[f'N{current_row}'] = f'=L{current_row}*M{current_row}'
            ws[f'O{current_row}'] = 'Custo previsto'
            ws[f'P{current_row}'] = 'Finalizado'
            current_row += 1
        
        # Para outras células, apenas criar a linha da célula (estrutura completa seria muito extensa)
        else:
            # Placeholder para as demais células
            # Em produção, cada célula teria sua estrutura completa de etapas/subetapas/serviços
            pass
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 60
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 35
    ws.column_dimensions['P'].width = 15
    
    return wb


# ============================================================================
# FUNÇÃO PRINCIPAL
# ============================================================================

def create_grupo_b(wb, projeto_data):
    """
    Cria todas as 7 abas do Grupo B - Gerenciamento
    
    Args:
        wb: Workbook openpyxl
        projeto_data: dict com dados do projeto (nome_obra, empresa, revisao, etc.)
    
    Returns:
        wb: Workbook atualizado com as 7 abas
    """
    
    # Aba 1: EPCs
    wb = create_epcs(wb, projeto_data)
    
    # Aba 2: CANTEIRO
    wb = create_canteiro(wb, projeto_data)
    
    # Aba 3: Controle Tecnologico
    wb = create_controle_tecnologico(wb, projeto_data)
    
    # Aba 4: Ensaios
    wb = create_ensaios(wb, projeto_data)
    
    # Aba 5: Cont.Tecnol.
    wb = create_cont_tecnol(wb, projeto_data)
    
    # Aba 6: Ger_Tec e Adm
    wb = create_ger_tec_adm(wb, projeto_data)
    
    # Aba 7: Ger_Executivo (CORE)
    wb = create_ger_executivo(wb, projeto_data)
    
    return wb


# ============================================================================
# TESTE DO MÓDULO (standalone)
# ============================================================================

if __name__ == '__main__':
    # Teste standalone
    from openpyxl import Workbook
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    projeto_data = {
        'nome_obra': 'Teste Grupo B',
        'empresa': 'Cartesian Engenharia',
        'revisao': 'R00',
        'area_total': 6601.28,
        'prazo_meses': 28
    }
    
    wb = create_grupo_b(wb, projeto_data)
    
    output_path = '/tmp/teste_grupo_b.xlsx'
    wb.save(output_path)
    print(f'✅ Módulo Grupo B criado com sucesso: {output_path}')
    print(f'📋 Abas criadas: {wb.sheetnames}')
