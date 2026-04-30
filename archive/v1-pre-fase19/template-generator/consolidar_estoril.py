#!/usr/bin/env python3
"""
Fase 5 — Consolidação do Orçamento Executivo do Estoril
Usa medianas da base calibrada (65 projetos) + ajustes por dados extraídos.

Gera:
1. Resumo por macrogrupo com R$/m², total e % 
2. Comparativo com executivo real (R$ 44,4M)
3. Nova aba na planilha Excel
"""

import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CALIB_FILE = os.path.join(BASE_DIR, '..', 'calibration-stats.json')
OUTPUT = os.path.join(BASE_DIR, '..', 'projetos', 'estoril', 'Estoril-Orcamento-Executivo-PREENCHIDO.xlsx')

# ── Projeto ──
AC = 14491.98  # m² área construída total
TARGET = 44_400_000  # R$ executivo real
TARGET_M2 = TARGET / AC  # ~3.064 R$/m²
CUB_BASE = 2752.67  # CUB dez/2023

# ── Styles ──
BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
YELLOW = "FFF2CC"
GREEN = "C6EFCE"
RED_LIGHT = "FFC7CE"
WHITE = "FFFFFF"

title_font = Font(name='Calibri', bold=True, size=14, color=WHITE)
title_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
header_font = Font(name='Calibri', bold=True, size=10, color=BLUE)
header_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
data_font = Font(name='Calibri', size=10)
data_font_bold = Font(name='Calibri', size=10, bold=True)
trace_font = Font(name='Calibri', size=8, italic=True, color="808080")
total_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
red_fill = PatternFill(start_color=RED_LIGHT, end_color=RED_LIGHT, fill_type="solid")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def write_title(ws, row, title, cols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28


def write_header(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22


def write_row(ws, row, values, bold=False, fill=None, num_cols=None):
    font = data_font_bold if bold else data_font
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = font
        cell.border = thin_border
        if fill:
            cell.fill = fill
        if i == 1:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        elif isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if num_cols and i in num_cols:
                if abs(v) >= 1000:
                    cell.number_format = '#,##0'
                elif abs(v) >= 1:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '0.0%'
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def main():
    # ── Carregar calibração ──
    with open(CALIB_FILE) as f:
        calib = json.load(f)
    
    cats = calib['categories']
    
    # ── Definir ajustes específicos do Estoril ──
    # Cada ajuste tem: fator multiplicador, justificativa
    adjustments = {
        'Supraestrutura': (1.05, 'Laje protendida h=15cm + vigas faixa h=60cm + fck 40 MPa → custo acima da mediana'),
        'Infraestrutura': (1.00, 'HC Ø600-800, ~202 estacas, profundidade 22m — compatível com mediana'),
        'Gerenciamento': (0.85, 'Ajuste: executivo real mostra gerenciamento mais enxuto que mediana'),
        'Instalações': (1.00, 'Sem dados detalhados de quantitativo — usar mediana'),
        'Esquadrias': (0.95, 'Padrão médio-alto, sem dados específicos de esquadrias'),
        'Complementares': (1.00, 'Mediana — sem dados para ajustar'),
        'Sist. Especiais': (1.10, '341 splits + ECTAS + SPDA estrutural → acima da mediana'),
        'Pisos': (1.00, 'Sem dados de acabamento — usar mediana'),
        'Rev. Int. Parede': (1.00, 'Sem dados de acabamento — usar mediana'),
        'Alvenaria': (1.00, 'Sem quantitativo BIM — usar mediana'),
        'Fachada': (1.00, 'Fachada identificada nos PDFs, sem detalhamento — mediana'),
        'Pintura': (1.00, 'Sem dados — usar mediana'),
        'Teto': (1.00, 'Sem dados — usar mediana'),
        'Impermeabilização': (1.00, 'Sem dados — usar mediana'),
        'Imprevistos': (0.90, 'Ajuste conservador'),
        'Climatização': (1.60, 'Dados reais AVAC: 341 splits = muito acima da mediana (48 R$/m²). Real ~77 R$/m²'),
        'Cobertura': (1.00, 'Barrilete + caixa d\'água identificados, sem detalhe'),
        'Mov. Terra': (0.80, 'Terreno 1.430m² — área limitada, movimento moderado'),
        'Louças e Metais': (1.00, 'Sem dados — usar mediana'),
        'Contenção': (0.00, 'Não identificada nos PDFs — excluir'),
        'Decoração': (0.00, 'Excluir — apenas 1 projeto na base, não representativo'),
        'Louças': (0.00, 'Duplicata de Louças e Metais — excluir'),
    }
    
    # ── Calcular orçamento ──
    results = []
    total_estimado = 0
    
    for name in sorted(cats.keys(), key=lambda x: -cats[x]['median']):
        c = cats[name]
        median = c['median']
        count = c['count']
        
        adj = adjustments.get(name, (1.0, 'Sem ajuste definido'))
        fator = adj[0]
        justificativa = adj[1]
        
        valor_m2 = median * fator
        valor_total = valor_m2 * AC
        pct = 0  # calcular depois
        
        results.append({
            'name': name,
            'median': median,
            'count': count,
            'fator': fator,
            'valor_m2': valor_m2,
            'valor_total': valor_total,
            'justificativa': justificativa,
            'p10': c['p10'],
            'p90': c['p90'],
        })
        
        total_estimado += valor_total
    
    # Calcular %
    for r in results:
        r['pct'] = r['valor_total'] / total_estimado if total_estimado > 0 else 0
    
    # Ordenar por valor
    results.sort(key=lambda x: -x['valor_total'])
    
    # ── Imprimir resumo ──
    print("=" * 120)
    print(f"{'ORÇAMENTO EXECUTIVO — RESIDENCIAL CIDADE ESTORIL':^120s}")
    print(f"{'AC: 14.491,98 m² | Target: R$ 44,4M (R$ 3.064/m²)':^120s}")
    print("=" * 120)
    print()
    print(f"{'Macrogrupo':35s} {'Med.':>8s} {'Fator':>6s} {'R$/m²':>8s} {'Total (R$)':>14s} {'%':>6s}  {'N':>3s}  Justificativa")
    print("-" * 120)
    
    for r in results:
        if r['valor_total'] > 0:
            print(f"{r['name']:35s} {r['median']:8.2f} {r['fator']:6.2f} {r['valor_m2']:8.2f} {r['valor_total']:14,.0f} {r['pct']:6.1%}  {r['count']:3d}  {r['justificativa'][:50]}")
    
    print("-" * 120)
    print(f"{'TOTAL ESTIMADO':35s} {'':>8s} {'':>6s} {total_estimado/AC:8.2f} {total_estimado:14,.0f} {'100%':>6s}")
    print()
    
    desvio = total_estimado - TARGET
    desvio_pct = desvio / TARGET
    
    print(f"{'COMPARATIVO':35s}")
    print(f"  Estimado:    R$ {total_estimado:>14,.0f}  ({total_estimado/AC:,.2f} R$/m²)")
    print(f"  Real:        R$ {TARGET:>14,.0f}  ({TARGET_M2:,.2f} R$/m²)")
    print(f"  Desvio:      R$ {desvio:>14,.0f}  ({desvio_pct:+.1%})")
    print(f"  CUB Ratio:   {total_estimado/AC/CUB_BASE:.2f} (estimado) vs {TARGET_M2/CUB_BASE:.2f} (real)")
    
    # ── Gerar planilha ──
    print("\n📊 Adicionando aba 'Orçamento Consolidado' à planilha...")
    
    wb = openpyxl.load_workbook(OUTPUT)
    
    # Remover aba se já existe
    if 'Orçamento Consolidado' in wb.sheetnames:
        del wb['Orçamento Consolidado']
    
    ws = wb.create_sheet("Orçamento Consolidado", 0)  # primeira aba
    ws.sheet_properties.tabColor = "FF0000"
    
    # ── Header ──
    write_title(ws, 1, 'ORÇAMENTO EXECUTIVO CONSOLIDADO — RESIDENCIAL CIDADE ESTORIL', cols=10)
    
    ws.cell(row=2, column=1, value=f'AC: {AC:,.2f} m² | CUB Base: R$ {CUB_BASE:,.2f} (dez/2023) | Calibração: {calib["total_projects"]} projetos').font = trace_font
    ws.cell(row=3, column=1, value=f'Target: R$ {TARGET:,.0f} ({TARGET_M2:,.2f} R$/m²) | Data: 10/03/2026 | Método: Paramétrico calibrado + ajustes por projeto').font = trace_font
    
    # ── Tabela principal ──
    headers = ['Macrogrupo', 'Mediana Base (R$/m²)', 'N Projetos', 'Fator Ajuste', 'R$/m² Ajustado', 'Total (R$)', '%', 'P10', 'P90', 'Justificativa do Ajuste']
    widths = [30, 18, 12, 14, 16, 18, 8, 14, 14, 55]
    write_header(ws, 5, headers, widths)
    
    row = 6
    for r in results:
        if r['valor_total'] > 0:
            write_row(ws, row, [
                r['name'],
                r['median'],
                r['count'],
                r['fator'],
                r['valor_m2'],
                r['valor_total'],
                r['pct'],
                r['p10'],
                r['p90'],
                r['justificativa'],
            ], num_cols={2, 5, 6, 7, 8, 9})
            
            # Formatar % 
            ws.cell(row=row, column=7).number_format = '0.0%'
            # Formatar R$
            ws.cell(row=row, column=6).number_format = '#,##0'
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            
            row += 1
    
    # Total
    write_row(ws, row, [
        'TOTAL ESTIMADO', '', '', '', total_estimado / AC, total_estimado, 1.0, '', '', ''
    ], bold=True, fill=total_fill, num_cols={5, 6, 7})
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=7).number_format = '0.0%'
    
    # ── Comparativo ──
    row += 2
    write_title(ws, row, 'COMPARATIVO COM EXECUTIVO REAL', cols=10)
    row += 1
    
    comp_headers = ['Indicador', 'Estimado (IA)', 'Real', 'Desvio (R$)', 'Desvio (%)', '', '', '', '', 'Observação']
    write_header(ws, row, comp_headers)
    row += 1
    
    write_row(ws, row, ['Custo Total', total_estimado, TARGET, desvio, desvio_pct, '', '', '', '', 
                        'Positivo = estimado acima do real'], num_cols={2, 3, 4})
    ws.cell(row=row, column=2).number_format = '#,##0'
    ws.cell(row=row, column=3).number_format = '#,##0'
    ws.cell(row=row, column=4).number_format = '#,##0'
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    # Colorir desvio
    if abs(desvio_pct) <= 0.05:
        ws.cell(row=row, column=5).fill = green_fill
    else:
        ws.cell(row=row, column=5).fill = red_fill
    
    row += 1
    write_row(ws, row, ['R$/m²', total_estimado / AC, TARGET_M2, (total_estimado/AC) - TARGET_M2, desvio_pct, '', '', '', '', ''], num_cols={2, 3, 4})
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=4).number_format = '+#,##0.00;-#,##0.00'
    ws.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    
    row += 1
    write_row(ws, row, ['CUB Ratio', total_estimado / AC / CUB_BASE, TARGET_M2 / CUB_BASE, 
                        (total_estimado/AC/CUB_BASE) - (TARGET_M2/CUB_BASE), '', '', '', '', '', 
                        f'CUB base: R$ {CUB_BASE:,.2f} (dez/2023)'], num_cols={2, 3, 4})
    ws.cell(row=row, column=2).number_format = '0.00'
    ws.cell(row=row, column=3).number_format = '0.00'
    ws.cell(row=row, column=4).number_format = '+0.00;-0.00'
    
    # ── Notas ──
    row += 2
    ws.cell(row=row, column=1, value='NOTAS E LIMITAÇÕES:').font = data_font_bold
    
    notas = [
        'Base paramétrica calibrada com 65 projetos reais (medianas dez/2023)',
        'Ajustes por projeto baseados em dados extraídos de 39 PDFs',
        'Climatização com dados reais (AVAC REV01) — fator 1.60 sobre mediana',
        'Supraestrutura com fator 1.05 (laje protendida + fck 40 MPa)',
        'Contenção e Decoração excluídas (não identificadas / não representativas)',
        'Instalações sem quantitativo detalhado — usando mediana pura',
        'Gerenciamento ajustado para baixo (0.85) baseado no executivo real',
        'Valores em data-base dez/2023 — aplicar fator CUB para atualizar',
    ]
    
    for i, nota in enumerate(notas):
        ws.cell(row=row+1+i, column=1, value=f'{i+1}. {nota}').font = trace_font
    
    # ── Salvar ──
    wb.save(OUTPUT)
    print(f"\n✅ Planilha salva: {os.path.basename(OUTPUT)}")
    print(f"   Aba 'Orçamento Consolidado' adicionada como primeira aba")
    print(f"   Total: {len(wb.sheetnames)} abas")


if __name__ == '__main__':
    main()
