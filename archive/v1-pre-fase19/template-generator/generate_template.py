#!/usr/bin/env python3
"""
Gerador de Template de Orçamento Executivo — Cartesian Engenharia

Abordagem: Copia a planilha real do Estoril como base e limpa/generaliza
os dados projeto-específicos, mantendo toda a estrutura, formatação e fórmulas.

Uso:
    python3 generate_template.py --source planilha_real.xlsx --output template.xlsx
    python3 generate_template.py --source planilha_real.xlsx --output preenchido.xlsx --fill dados.json

Autor: Jarvis (OpenClaw)
Data: 10/03/2026
"""

import argparse
import copy
import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl não encontrado. Instale: pip install openpyxl")
    sys.exit(1)


# ============================================================
# CONFIGURAÇÃO: Abas e seus tipos de limpeza
# ============================================================

# Tipo de aba determina como limpar os dados:
# - "header_only": manter só headers, limpar dados (quantitativos)
# - "structure": manter estrutura hierárquica, limpar valores (Ger_Executivo)
# - "reference": manter como referência (CPU, Insumos — banco de dados)
# - "overview": limpar dados do projeto, manter estrutura
# - "keep": manter como está (template de exemplo)

TAB_CONFIG = {
    # Grupo A — Overview
    'PROJETOS': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    'Obra': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3', 'D1', 'D2', 'D3']},
    'Obra ': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3', 'D1', 'D2', 'D3']},
    'EAP Análise': {'tipo': 'keep'},  # Fórmulas que referenciam Ger_Executivo
    
    # Grupo B — Gerenciamento
    'EPCs': {'tipo': 'overview', 'clear_cells': ['B1', 'B2', 'B3']},
    'CANTEIRO': {'tipo': 'overview', 'clear_cells': ['B1', 'B2', 'B3']},
    'Controle Tecnologico': {'tipo': 'header_only', 'header_row': 6},
    'Ensaios': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3', 'D1', 'D2', 'D3']},
    'Ensaios ': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3', 'D1', 'D2', 'D3']},
    'Cont.Tecnol.': {'tipo': 'header_only', 'header_row': 6},
    'Ger_Tec e Adm': {'tipo': 'structure', 'header_row': 7},
    'Ger_Executivo': {'tipo': 'structure', 'header_row': 7},
    
    # Grupo C — Referência
    'CPU': {'tipo': 'reference'},
    'Insumos': {'tipo': 'reference'},
    
    # Grupo D — Quantitativos BIM/Visus
    'CHURRASQUEIRAS E SHAFTS': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    'Visus - Alvenaria': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Rev. e Acab. de Parede': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Piso': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Soleiras': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Teto': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Telhado': {'tipo': 'header_only', 'header_row': 12},
    'Rufos, Pintura, garagem, pergol': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Calha': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Impermeabilização': {'tipo': 'header_only', 'header_row': 12},
    
    # Grupo E — Estrutural
    'Estacas': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    'Aço_Estacas': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    'Fundação Rasa': {'tipo': 'overview', 'clear_cells': ['E1', 'E2', 'E3']},
    'Supraestrutura': {'tipo': 'overview', 'clear_cells': ['E1', 'E2', 'E3']},
    'Aço_Infraestrutura': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    
    # Grupo F — Instalações
    'INSTALAÇÕES': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    'HIDROSSANITÁRIO': {'tipo': 'header_only', 'header_row': 6},
    'ELÉTRICO': {'tipo': 'header_only', 'header_row': 6},
    'TELEFONE': {'tipo': 'header_only', 'header_row': 6},
    'PCI': {'tipo': 'header_only', 'header_row': 6},
    'SPDA': {'tipo': 'header_only', 'header_row': 6},
    'CLIMATIZAÇÃO': {'tipo': 'header_only', 'header_row': 6},
    
    # Grupo G — Acabamentos
    'ESQUADRIAS': {'tipo': 'overview', 'clear_cells': ['B1', 'B2', 'B3']},
    'Visus - Guarda Corpo e Corrimão': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Churrasqueiras': {'tipo': 'header_only', 'header_row': 12},
    'Visus - Infraestrutura': {'tipo': 'header_only', 'header_row': 12},
    'LOUÇAS E METAIS': {'tipo': 'overview', 'clear_cells': ['E1', 'E2', 'E3']},
    'LOUÇAS E METAIS RESUMO': {'tipo': 'overview', 'clear_cells': ['D1', 'D2', 'D3']},
    'Piscinas': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
    'MOBILIÁRIO': {'tipo': 'overview', 'clear_cells': ['C1', 'C2', 'C3']},
}

# Campos que contêm nome do projeto na planilha
PROJECT_MARKERS = ['Estoril', 'ESTORIL', 'estoril']
COMPANY_MARKERS = ['Fonseca Neto', 'FONSECA', 'Fonseca', 'fonseca']
REVISION_MARKERS = ['R00 - Letícia', 'R00', 'R00 - Letícia - 29/10/2025']


def clear_project_specific_data(ws, config, projeto_placeholder='{PROJETO}', 
                                  empresa_placeholder='{EMPRESA}',
                                  revisao_placeholder='{REVISÃO}'):
    """Limpa dados específicos do projeto, mantendo estrutura."""
    
    tipo = config.get('tipo', 'keep')
    
    if tipo == 'keep':
        # Só substituir nome do projeto/empresa
        _replace_markers(ws, projeto_placeholder, empresa_placeholder, revisao_placeholder)
        return
    
    if tipo == 'overview':
        # Limpar células específicas e substituir marcadores
        _replace_markers(ws, projeto_placeholder, empresa_placeholder, revisao_placeholder)
        return
    
    if tipo == 'header_only':
        # Manter headers, limpar dados abaixo
        header_row = config.get('header_row', 6)
        for row in range(header_row + 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = None
        _replace_markers(ws, projeto_placeholder, empresa_placeholder, revisao_placeholder)
        return
    
    if tipo == 'structure':
        # Manter hierarquia (CÉLULA/ETAPA/SUBETAPA headers), limpar valores de SERVIÇO
        header_row = config.get('header_row', 7)
        for row in range(header_row + 1, ws.max_row + 1):
            nivel = ws.cell(row=row, column=8).value  # Col H = NIVEL
            if nivel == 'SERVIÇO':
                # Limpar quantidade, preço, total dos serviços
                ws.cell(row=row, column=12).value = None  # Quant
                ws.cell(row=row, column=13).value = None  # Preço un
                ws.cell(row=row, column=14).value = None  # Total
                ws.cell(row=row, column=16).value = None  # Situação
            elif nivel in ('CÉLULA CONSTRUTIVA', 'ETAPA', 'SUBETAPA'):
                # Limpar totais (serão recalculados)
                ws.cell(row=row, column=14).value = None  # Total
        _replace_markers(ws, projeto_placeholder, empresa_placeholder, revisao_placeholder)
        return
    
    if tipo == 'reference':
        # Manter como está — é banco de dados de referência
        return


def _replace_markers(ws, projeto_ph, empresa_ph, revisao_ph):
    """Substituir marcadores de projeto/empresa/revisão em toda a aba."""
    for row in range(1, min(ws.max_row + 1, 10)):  # Só primeiras 10 linhas (cabeçalho)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):
                val = cell.value
                for marker in PROJECT_MARKERS:
                    val = val.replace(marker, projeto_ph)
                for marker in COMPANY_MARKERS:
                    val = val.replace(marker, empresa_ph)
                for marker in REVISION_MARKERS:
                    val = val.replace(marker, revisao_ph)
                cell.value = val


def fill_project_data(ws, sheet_name, projeto_data):
    """Preencher dados do projeto no template."""
    nome = projeto_data.get('nome', '')
    empresa = projeto_data.get('empresa', '')
    revisao = projeto_data.get('revisao', 'R00')
    
    for row in range(1, min(ws.max_row + 1, 10)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):
                cell.value = cell.value.replace('{PROJETO}', nome)
                cell.value = cell.value.replace('{EMPRESA}', empresa)
                cell.value = cell.value.replace('{REVISÃO}', revisao)


def generate_template(source_path, output_path, projeto_data=None):
    """
    Gera template a partir da planilha real.
    
    Args:
        source_path: Caminho da planilha modelo
        output_path: Caminho de saída
        projeto_data: dict opcional para preencher dados do projeto
    """
    print(f"📂 Carregando planilha modelo: {source_path}")
    wb = openpyxl.load_workbook(source_path)
    
    print(f"📋 {len(wb.sheetnames)} abas encontradas")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        config = TAB_CONFIG.get(sheet_name.strip(), {'tipo': 'keep'})
        
        print(f"  ✏️  {sheet_name:40s} → {config['tipo']}")
        clear_project_specific_data(ws, config)
        
        if projeto_data:
            fill_project_data(ws, sheet_name, projeto_data)
    
    print(f"\n💾 Salvando template: {output_path}")
    wb.save(output_path)
    print(f"✅ Template gerado com sucesso!")
    print(f"   {len(wb.sheetnames)} abas preservadas")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Gera template de orçamento executivo')
    parser.add_argument('--source', required=True, help='Planilha modelo (xlsx)')
    parser.add_argument('--output', required=True, help='Arquivo de saída (xlsx)')
    parser.add_argument('--projeto', help='Nome do projeto')
    parser.add_argument('--empresa', help='Nome da empresa/cliente')
    parser.add_argument('--revisao', default='R00', help='Revisão (default: R00)')
    parser.add_argument('--fill', help='JSON com dados do projeto para preenchimento')
    
    args = parser.parse_args()
    
    projeto_data = None
    if args.fill:
        with open(args.fill) as f:
            projeto_data = json.load(f)
    elif args.projeto:
        projeto_data = {
            'nome': args.projeto,
            'empresa': args.empresa or '',
            'revisao': args.revisao,
        }
    
    generate_template(args.source, args.output, projeto_data)


if __name__ == '__main__':
    main()
