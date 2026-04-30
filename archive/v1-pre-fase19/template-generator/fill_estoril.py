#!/usr/bin/env python3
"""
Preenche o template do orçamento executivo com dados extraídos do Estoril.
Cada valor inclui rastreabilidade (fonte, versão, confiança).

Uso:
    python3 fill_estoril.py

Autor: Jarvis (OpenClaw)
Data: 10/03/2026
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(BASE_DIR, '..', 'projetos', 'estoril', 'TEMPLATE-Orcamento-Executivo.xlsx')
OUTPUT = os.path.join(BASE_DIR, '..', 'projetos', 'estoril', 'Estoril-Orcamento-Executivo-PREENCHIDO.xlsx')

# Styles
trace_font = Font(italic=True, size=8, color="808080")
source_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def trace(fonte, rev="", confianca="⭐⭐⭐"):
    """Gera string de rastreabilidade."""
    parts = [f"[Fonte: {fonte}"]
    if rev:
        parts[0] += f" | Rev: {rev}"
    parts[0] += f" | Confiança: {confianca}]"
    return parts[0]


def safe_write(ws, row, col, value, font=None):
    """Escreve numa célula, pulando se for merged."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return  # Skip merged cells
    cell.value = value
    if font:
        cell.font = font



def safe_font(ws, row, col, font):
    """Set font safely, skipping merged cells."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.font = font

def fill_obra(wb):
    """Preenche aba 'Obra ' com dados do projeto."""
    ws = wb['Obra ']
    
    # Dados do projeto
    safe_write(ws, 1, 4, 'Estoril')
    safe_write(ws, 2, 4, 'Fonseca Neto')
    safe_write(ws, 3, 4, 'R00 - IA')
    
    # Quadro de áreas — usar linhas existentes da planilha
    # A aba Obra original tem merge cells complexas, vamos escrever apenas em células livres
    
    # Escrever dados em área segura (abaixo dos merges da planilha original)
    # A planilha original tem dados a partir da row 6-7
    safe_write(ws, 7, 5, 48)  # meses de obra
    
    areas_data = [
        # (row, col_desc, col_unid, col_qtd, col_trace, descricao, unidade, quantidade, fonte)
        (8, 2, 3, 5, 7, 'Área do terreno', 'm²', 1430.00, trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (9, 2, 3, 5, 7, 'Área total construída', 'm²', 14491.98, trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (10, 2, 3, 5, 7, 'Área computável', 'm²', 6434.32, trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (11, 2, 3, 5, 7, 'Área não computável', 'm²', 8057.66, trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (12, 2, 3, 5, 7, 'Vagas de garagem', 'un', 138, trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (13, 2, 3, 5, 7, 'Gabarito (pavimentos)', 'un', 16, trace('ARQ 01 - Quadro de áreas', '08/06/2023', '⭐⭐⭐⭐⭐')),
        (14, 2, 3, 5, 7, 'Pavimentos tipo', 'un', 10, trace('Quantitativo AVAC REV01 + PPCI P08', '', '⭐⭐⭐⭐')),
        (15, 2, 3, 5, 7, 'Unidades residenciais (est.)', 'un', 55, trace('PPCI P08 - 5 med gás/pav × 11 pav', '', '⭐⭐⭐')),
        (16, 2, 3, 5, 7, 'Elevadores sociais', 'un', 2, trace('PPCI P08 - Pavimento Tipo', '', '⭐⭐⭐⭐')),
        (17, 2, 3, 5, 7, 'Pé-direito tipo', 'm', 3.06, trace('PPCI P08 - Cotas (28,44 - 25,38)', '', '⭐⭐⭐⭐')),
    ]
    
    for row, cd, cu, cq, ct, desc, unid, qtd, fonte in areas_data:
        safe_write(ws, row, cd, desc)
        if unid:
            safe_write(ws, row, cu, unid)
        safe_write(ws, row, cq, qtd)
        safe_write(ws, row, ct, fonte, font=trace_font)
    
    print("  ✅ Obra preenchida")


def fill_estacas(wb):
    """Preenche aba 'Estacas' com dados das fundações."""
    ws = wb['Estacas']
    
    # Header
    safe_write(ws, 1, 3, 'Estoril')
    safe_write(ws, 2, 3, 'Fonseca Neto')
    safe_write(ws, 3, 3, 'R00 - IA (extraído de PDFs)')
    
    # Estacas data from 02-estrutural.md
    # Prancha 01A (137 estacas detalhadas)
    estacas_01a = [
        # (row, tipo, quant, c_topo, c_apoio, diametro_m, h)
        # Ø600mm — 71 estacas
        (8, 'HC Ø600', 71, None, None, 0.60, None),
        # Ø700mm — 51 estacas
        (9, 'HC Ø700', 51, None, None, 0.70, None),
        # Ø800mm — 15 estacas
        (10, 'HC Ø800', 15, None, None, 0.80, None),
    ]
    
    # Totals
    safe_write(ws, 7, 1, 'PRANCHA 01-A')
    safe_font(ws, 7, 1, Font(bold=True))
    
    for row, tipo, quant, ct, ca, diam, h in estacas_01a:
        safe_write(ws, row, 2, tipo)
        safe_write(ws, row, 3, quant)
        safe_write(ws, row, 6, diam)
        # Rastreabilidade
        safe_write(ws, row, 16, trace(
            '01-A-PROJETO DE FUNDAÇÕES.pdf',
            'Rev 01 — 25/01/2023',
            '⭐⭐⭐⭐'
        ))
        safe_font(ws, row, 16, trace_font)
    
    # Subtotal 01A
    safe_write(ws, 11, 1, 'Subtotal 01A')
    safe_font(ws, 11, 1, Font(bold=True))
    safe_write(ws, 11, 3, 137)
    safe_write(ws, 11, 8, 2825)  # metros perfurados
    safe_write(ws, 11, 9, 1175.83)  # volume concreto c/ perdas
    safe_write(ws, 11, 16, trace('01-A — Quadro resumo perfuração + volume', 'Rev 01', '⭐⭐⭐⭐⭐'))
    safe_font(ws, 11, 16, trace_font)
    
    # Prancha 01B (65 estacas estimadas)
    safe_write(ws, 13, 1, 'PRANCHA 01-B (resolução limitada — estimativa)')
    safe_font(ws, 13, 1, Font(bold=True))
    
    estacas_01b = [
        (14, 'HC Ø600', 36, '⭐⭐⭐'),
        (15, 'HC Ø700', 23, '⭐⭐⭐'),
        (16, 'HC Ø800', 6, '⭐⭐⭐'),
    ]
    
    for row, tipo, quant, conf in estacas_01b:
        safe_write(ws, row, 2, tipo)
        safe_write(ws, row, 3, quant)
        safe_write(ws, row, 16, trace(
            '01-B-PROJETO DE FUNDAÇÕES.pdf',
            'Rev 01 — 25/01/2023',
            conf + ' (PDF raster, contagem visual)'
        ))
        safe_font(ws, row, 16, trace_font)
    
    # Subtotal 01B
    safe_write(ws, 17, 1, 'Subtotal 01B (estimado)')
    safe_font(ws, 17, 1, Font(bold=True))
    safe_write(ws, 17, 3, 65)
    
    # TOTAL GERAL
    safe_write(ws, 19, 1, 'TOTAL GERAL ESTACAS')
    safe_font(ws, 19, 1, Font(bold=True, size=11))
    safe_write(ws, 19, 3, 202)
    safe_write(ws, 19, 8, '~4.500–5.000')  # metros perfurados estimados
    safe_write(ws, 19, 9, '~2.000–2.400')  # concreto estimado
    safe_write(ws, 19, 16, trace(
        '01-A + 01-B FUNDAÇÕES (01A exato, 01B proporcional)',
        'Rev 01',
        '⭐⭐⭐⭐ (01A) / ⭐⭐⭐ (01B)'
    ))
    safe_font(ws, 19, 16, trace_font)
    
    # Dados complementares
    safe_write(ws, 21, 1, 'DADOS COMPLEMENTARES')
    safe_font(ws, 21, 1, Font(bold=True))
    
    complementares = [
        (22, 'fck estacas', '30 MPa', trace('01-B — Nota técnica', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (23, 'Consumo mín. cimento', '≥ 400 kg/m³', trace('01-B — Nota técnica', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (24, 'Slump', '220 ± 20 mm (S220)', trace('01-B — Nota técnica', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (25, 'Fator a/c', '≤ 0,60', trace('01-B — Nota técnica', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (26, 'Profundidade máx.', '22,00 m', trace('01-A/01-B — Quadro estacas', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (27, 'Torque mínimo', '30 tf·m (300 bar)', trace('01-B — Nota técnica', 'Rev 01', '⭐⭐⭐⭐⭐')),
        (28, 'Responsável', 'Luciano Rosso — CREA SC 085.963-3', trace('01-A/01-B — Carimbo', '', '⭐⭐⭐⭐⭐')),
        (29, 'Data projeto', '26/10/2022 (Rev 01: 25/01/2023)', trace('01-A — Carimbo', '', '⭐⭐⭐⭐⭐')),
    ]
    
    for row, desc, valor, fonte in complementares:
        safe_write(ws, row, 2, desc)
        safe_write(ws, row, 3, valor)
        safe_write(ws, row, 16, fonte)
        safe_font(ws, row, 16, trace_font)
    
    print("  ✅ Estacas preenchida")


def fill_aco_estacas(wb):
    """Preenche aba 'Aço_Estacas'."""
    ws = wb['Aço_Estacas']
    
    safe_write(ws, 1, 3, 'Estoril')
    safe_write(ws, 2, 3, 'Fonseca Neto')
    safe_write(ws, 3, 3, 'R00 - IA')
    
    # Dados da 01-A
    safe_write(ws, 7, 1, 'PRANCHA 01-A')
    safe_write(ws, 7, 2, 'ESTACAS')
    
    # Armadura por diâmetro
    aco_data = [
        (8, 'Ø6,3 (estribos espiral)', 7611.34, 1902.84),
        (9, 'Ø16,0 (long. Ø500/Ø600)', 3280.20, 5149.91),
        (10, 'Ø20,0 (long. Ø700/Ø800)', 2811.60, 6972.77),
    ]
    
    for row, desc, comp_m, peso_kg in aco_data:
        safe_write(ws, row, 2, desc)
        safe_write(ws, row, 10, comp_m)  # comprimento
        safe_write(ws, row, 11, peso_kg)  # peso
        safe_write(ws, row, 13, trace(
            '01-A-PROJETO DE FUNDAÇÕES.pdf — Quadro Aço',
            'Rev 01 — 25/01/2023',
            '⭐⭐⭐⭐⭐'
        ))
        safe_font(ws, row, 13, trace_font)
    
    # Total
    safe_write(ws, 11, 1, 'TOTAL 01-A')
    safe_font(ws, 11, 1, Font(bold=True))
    safe_write(ws, 11, 11, 14025.52)
    safe_write(ws, 11, 13, '14.025,52 kg = ~14 ton')
    
    safe_write(ws, 12, 1, 'TOTAL ESTIMADO (01A+01B)')
    safe_font(ws, 12, 1, Font(bold=True))
    safe_write(ws, 12, 11, 28000)
    safe_write(ws, 12, 13, trace(
        'Proporcional: 14.025 × (202/137) ≈ 28 ton',
        '',
        '⭐⭐⭐ (01B estimado por proporção)'
    ))
    safe_font(ws, 12, 13, trace_font)
    
    print("  ✅ Aço_Estacas preenchida")


def fill_instalacoes_resumo(wb):
    """Preenche aba 'INSTALAÇÕES' (resumo)."""
    ws = wb['INSTALAÇÕES']
    
    safe_write(ws, 1, 3, 'Estoril')
    safe_write(ws, 2, 3, 'Fonseca Neto')
    safe_write(ws, 3, 3, 'R00 - IA')
    
    # Dados do resumo
    instalacoes = [
        (7, 'SISTEMAS E INSTALAÇÕES ELÉTRICAS', 'vb', 1, None, 'Índice paramétrico', trace('Extração elétrica parcial + índices 65 projetos', '', '⭐⭐⭐')),
        (8, 'Luminárias e lâmpadas', 'vb', 1, None, 'Cotação', trace('Não quantificado em detalhe — usar cotação', '', '⭐⭐')),
        (15, 'SISTEMAS E INSTALAÇÕES HIDROSSANITÁRIAS', 'vb', 1, None, 'Quantitativo parcial', trace('SAN P06/P08/P11 + índices', '', '⭐⭐⭐')),
        (25, 'INSTALAÇÕES PREVENTIVAS', '', '', None, '', ''),
        (26, 'SHP (Hidrante/Mangotinho)', 'vb', 1, None, 'PPCI DET', trace('1109.020.25_INC_P15_DET_EXE_REV04.pdf', 'REV04', '⭐⭐⭐⭐')),
        (27, 'SDAI (Detecção e Alarme)', 'vb', 1, None, 'PPCI DET', trace('1109.020.25_INC_P15_DET_EXE_REV04.pdf', 'REV04', '⭐⭐⭐⭐')),
        (28, 'Extintores', 'un', 60, None, 'Estimado (PPCI)', trace('PPCI P08/P11/P12/P13 — contagem visual', '', '⭐⭐⭐')),
        (29, 'PCF (Portas Corta-Fogo)', 'un', 25, None, 'Estimado (PPCI)', trace('PPCI P11/P12/P13 — identificadas nos escapes', '', '⭐⭐⭐')),
        (32, 'INSTALAÇÕES GLP', '', '', None, '', ''),
        (33, 'Medidores G2,5', 'un', 50, None, 'Contagem exata', trace('PPCI P08 — 5 medidores/pav × 10 pav tipo', 'REV04', '⭐⭐⭐⭐⭐')),
        (34, 'Churrasqueiras gás', 'un', 40, None, 'Contagem exata', trace('PPCI P08 — 4 churr/pav × 10 pav tipo', 'REV04', '⭐⭐⭐⭐⭐')),
        (35, 'Tubulação PEX multicamadas', 'vb', 1, None, 'Projeto gás', trace('PPCI P08 — diâmetros Ø16-32mm', 'REV04', '⭐⭐⭐⭐')),
    ]
    
    for row, desc, unid, qtd, valor, consideracao, fonte in instalacoes:
        safe_write(ws, row, 1, desc)
        if unid:
            safe_write(ws, row, 2, unid)
        if qtd:
            safe_write(ws, row, 3, qtd)
        if consideracao:
            safe_write(ws, row, 6, consideracao)
        if fonte:
            safe_write(ws, row, 8, fonte)
            safe_font(ws, row, 8, trace_font)
    
    print("  ✅ INSTALAÇÕES preenchida")


def fill_climatizacao(wb):
    """Preenche aba 'CLIMATIZAÇÃO' — dados reais do quantitativo AVAC."""
    ws = wb['CLIMATIZAÇÃO']
    
    # Headers do projeto
    safe_write(ws, 1, 4, 'Estoril')
    safe_write(ws, 2, 4, 'Fonseca Neto')
    safe_write(ws, 3, 4, '1052.094.24_QUANTITATIVO_AVAC_REV01.pdf')
    safe_write(ws, 4, 4, 'R00 - IA (dados exatos do quantitativo)')
    
    SRC = '1052.094.24_QUANTITATIVO_AVAC_REV01.pdf'
    
    # Dados por pavimento do AVAC
    clima_data = [
        # (row, pavimento, grupo, subgrupo, descricao, qtd, perda, qtd_perda, unidade, custo_unit, rastreabilidade)
        (7, 'CLI-Terreo', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 9.000 BTU/h', 5, 0, 5, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (8, 'CLI-Terreo', 'CLIMATIZAÇÃO', 'Splits', 'Evaporadora 9.000 BTU/h', 5, 0, 5, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (9, 'CLI-Terreo', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 18.000 BTU/h', 4, 0, 4, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (10, 'CLI-Terreo', 'CLIMATIZAÇÃO', 'Splits', 'Evaporadora 18.000 BTU/h', 4, 0, 4, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (11, 'CLI-Terreo', 'CLIMATIZAÇÃO', 'Tubulação', 'Tubulação frigorígena (par)', 113.60, 0, 113.60, 'm', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        
        (13, 'CLI-Lazer', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 9.000 BTU/h', 6, 0, 6, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (14, 'CLI-Lazer', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 12.000 BTU/h', 3, 0, 3, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (15, 'CLI-Lazer', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 18.000 BTU/h', 5, 0, 5, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (16, 'CLI-Lazer', 'CLIMATIZAÇÃO', 'Tubulação', 'Tubulação frigorígena (par)', 278.36, 0, 278.36, 'm', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        
        (18, 'CLI-Diferenciado', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 9.000 BTU/h', 13, 0, 13, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (19, 'CLI-Diferenciado', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 12.000 BTU/h', 3, 0, 3, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (20, 'CLI-Diferenciado', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 18.000 BTU/h', 8, 0, 8, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (21, 'CLI-Diferenciado', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 24.000 BTU/h', 4, 0, 4, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (22, 'CLI-Diferenciado', 'CLIMATIZAÇÃO', 'Tubulação', 'Tubulação frigorígena (par)', 293.30, 0, 293.30, 'm', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        
        (24, 'CLI-Tipo (×10)', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 9.000 BTU/h', 184, 0, 184, 'un', None, trace(SRC + ' × 10 pavimentos', 'REV01', '⭐⭐⭐⭐⭐')),
        (25, 'CLI-Tipo (×10)', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 12.000 BTU/h', 23, 0, 23, 'un', None, trace(SRC + ' × 10 pavimentos', 'REV01', '⭐⭐⭐⭐⭐')),
        (26, 'CLI-Tipo (×10)', 'CLIMATIZAÇÃO', 'Splits', 'Condensadora 18.000 BTU/h', 73, 0, 73, 'un', None, trace(SRC + ' × 10 pavimentos', 'REV01', '⭐⭐⭐⭐⭐')),
        (27, 'CLI-Tipo (×10)', 'CLIMATIZAÇÃO', 'Tubulação', 'Tubulação frigorígena (par)', 2854.40, 0, 2854.40, 'm', None, trace(SRC + ' × 10 pavimentos', 'REV01', '⭐⭐⭐⭐⭐')),
        
        (29, 'CLI-Cobertura', 'CLIMATIZAÇÃO', 'Caixas', 'Caixa Polar (cobertura)', 10, 0, 10, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (30, 'CLI-Cobertura', 'CLIMATIZAÇÃO', 'Tubulação', 'Tubulação frigorígena (par)', 164.55, 0, 164.55, 'm', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
    ]
    
    # Resumo totais
    totais = [
        (32, '', 'TOTAL', '', 'Condensadoras 9.000 BTU/h', 208, '', 208, 'un', None, trace(SRC + ' (soma todos pavimentos)', 'REV01', '⭐⭐⭐⭐⭐')),
        (33, '', 'TOTAL', '', 'Condensadoras 12.000 BTU/h', 29, '', 29, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (34, '', 'TOTAL', '', 'Condensadoras 18.000 BTU/h', 90, '', 90, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (35, '', 'TOTAL', '', 'Condensadoras 24.000 BTU/h', 4, '', 4, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (36, '', 'TOTAL', '', 'Condensadoras 36.000 BTU/h', 10, '', 10, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (37, '', 'TOTAL', '', 'TOTAL CONDENSADORAS', 341, '', 341, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (38, '', 'TOTAL', '', 'Tubulação frigorígena TOTAL', 3704.21, '', 3704.21, 'm', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
        (39, '', 'TOTAL', '', 'Caixas Polares', 10, '', 10, 'un', None, trace(SRC, 'REV01', '⭐⭐⭐⭐⭐')),
    ]
    
    all_data = clima_data + totais
    
    for entry in all_data:
        row = entry[0]
        safe_write(ws, row, 1, entry[1])  # pavimento
        safe_write(ws, row, 2, entry[2])  # grupo
        safe_write(ws, row, 3, entry[3])  # subgrupo
        safe_write(ws, row, 4, entry[4])  # descricao
        safe_write(ws, row, 5, entry[5])  # qtd
        if entry[6] != '' and entry[6] != 0:
            safe_write(ws, row, 6, entry[6])  # perda
        safe_write(ws, row, 7, entry[7])  # qtd com perda
        safe_write(ws, row, 8, entry[8])  # unidade
        # Col 10 = rastreabilidade
        safe_write(ws, row, 10, entry[10])
        safe_font(ws, row, 10, trace_font)
    
    # Bold totais
    for row in range(32, 40):
        safe_font(ws, row, 4, Font(bold=True))
    
    print("  ✅ CLIMATIZAÇÃO preenchida (dados reais AVAC)")


def fill_spda(wb):
    """Preenche aba 'SPDA'."""
    ws = wb['SPDA']
    
    safe_write(ws, 1, 4, 'Estoril')
    safe_write(ws, 2, 4, 'Fonseca Neto')
    safe_write(ws, 3, 4, 'QUANTITATIVO_SPDA_REV00.pdf')
    safe_write(ws, 4, 4, 'R00 - IA')
    
    SRC = '1109.020.25_SPDA_P01_EXE_REV00.pdf'
    
    spda_data = [
        (7, 'SPDA', 'Captação', 'Barra chata alumínio 70mm²', None, None, None, 'm', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (8, 'SPDA', 'Descidas', 'Naturais (armadura pilares) + barra aço galv. 50mm²', None, None, None, 'vb', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (9, 'SPDA', 'Aterramento', 'Cabo aço galv. 80mm² + fundações', None, None, None, 'vb', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (10, 'SPDA', 'Equalização', 'Caixas de equalização', 5, None, 5, 'un', None, trace(SRC + ' — 5 pavimentos (2º, 6º, 9º, 12º, 15º)', 'REV00', '⭐⭐⭐⭐⭐')),
        (11, 'SPDA', 'Interligações', 'Elevador (malha)', 1, None, 1, 'vb', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (12, 'SPDA', 'Interligações', 'SDAI (equalização)', 1, None, 1, 'vb', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (13, 'SPDA', 'Interligações', 'Gás (equalização)', 1, None, 1, 'vb', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (14, 'SPDA', 'Interligações', 'SHP (equalização)', 1, None, 1, 'vb', None, trace(SRC, 'REV00', '⭐⭐⭐⭐')),
        (15, 'SPDA', 'Tipo', 'Estrutural (usa armadura do concreto como descida)', None, None, None, '', None, trace(SRC + ' — NBR 5419-3', 'REV00', '⭐⭐⭐⭐⭐')),
        (16, 'SPDA', 'Resistência', 'Máx. admitida: 0,2 Ω', None, None, None, '', None, trace(SRC, 'REV00', '⭐⭐⭐⭐⭐')),
    ]
    
    for entry in spda_data:
        row = entry[0]
        safe_write(ws, row, 1, entry[1])
        safe_write(ws, row, 2, entry[2])
        safe_write(ws, row, 3, entry[3])  # subgrupo → descrição
        safe_write(ws, row, 4, entry[4])  # qtd
        safe_write(ws, row, 7, entry[6])  # qtd com perda
        safe_write(ws, row, 8, entry[7])  # unidade
        safe_write(ws, row, 10, entry[9])  # rastreabilidade
        safe_font(ws, row, 10, trace_font)
    
    print("  ✅ SPDA preenchida")


def fill_projetos(wb):
    """Preenche aba 'PROJETOS'."""
    ws = wb['PROJETOS']
    
    safe_write(ws, 1, 2, 'Projeto')
    safe_write(ws, 1, 3, 'Estoril')
    safe_write(ws, 2, 2, 'Empresa')
    safe_write(ws, 2, 3, 'Fonseca Neto')
    safe_write(ws, 3, 2, 'Revisão')
    safe_write(ws, 3, 3, 'R00 - IA (gerado por extração de PDFs)')
    
    # Mapa de projetos
    safe_write(ws, 5, 1, 'INFRAESTRUTURA')
    safe_write(ws, 6, 1, 'PAVIMENTO')
    safe_write(ws, 6, 2, 'DISCIPLINA')
    
    disciplinas = [
        (7, 'ARQUITETURA', '11 PDFs', 'ARQ 01-11, data 08/06/2023'),
        (8, 'FUNDAÇÕES', '2 PDFs', '01-A/01-B, Rev 01 25/01/2023'),
        (9, 'ESTRUTURAL', '8 PDFs', 'Lajes, escadas, rampas, detalhes'),
        (10, 'HIDROSSANITÁRIO', '3 PDFs', 'SAN P06/P08/P11 REV03'),
        (11, 'ELÉTRICO', '3 PDFs', 'ELE P31/P32/P39 REV05'),
        (12, 'PPCI', '5 PDFs', 'INC P08/P11/P12/P13/P15 REV04'),
        (13, 'CLIMATIZAÇÃO', '3 PDFs + 1 QTV', 'CLI P05/P07 REV02 + AVAC REV01'),
        (14, 'SPDA', '1 PDF', 'SPDA P01 REV00'),
        (15, 'DRENAGEM', '1 PDF', 'MAC-QT-33490-R2 (Maccaferri)'),
        (16, 'ECTAS', '1 PDF', 'MDC Estoril (MBBR/IFAS)'),
        (17, 'COMUNICAÇÃO', '19 pranchas', 'Não processado em detalhe'),
        (18, 'AUTOMAÇÃO', 'Obsoleto', 'Pasta "Obsoleto" — sem projeto vigente'),
    ]
    
    for row, disc, qtd, obs in disciplinas:
        safe_write(ws, row, 2, disc)
        safe_write(ws, row, 3, qtd)
        safe_write(ws, row, 4, obs)
        safe_write(ws, row, 5, trace('Google Drive folder', '', '⭐⭐⭐⭐⭐'))
        safe_font(ws, row, 5, trace_font)
    
    print("  ✅ PROJETOS preenchida")


def fill_pci(wb):
    """Preenche aba 'PCI' com dados do PPCI."""
    ws = wb['PCI']
    
    safe_write(ws, 1, 4, 'Estoril')
    safe_write(ws, 2, 4, 'Fonseca Neto')
    safe_write(ws, 3, 4, '1109.020.25_QUANTITATIVO_INC_REV01.pdf')
    safe_write(ws, 4, 4, 'R00 - IA')
    
    SRC_DET = '1109.020.25_INC_P15_DET_EXE_REV04.pdf'
    SRC_TIP = '1109.020.25_INC_P08_TIP_EXE_REV04.pdf'
    
    pci_items = [
        (7, 'PCI', 'SHP', 'Hidrante tipo mangotinho', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐⭐')),
        (8, 'PCI', 'SHP', 'Mangotinho semi-rígido Ø25mm × 30m', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐⭐')),
        (9, 'PCI', 'SHP', 'Registro globo angular 45° Ø40mm', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐⭐')),
        (10, 'PCI', 'SHP', 'Abrigo mangotinho 60×40×17cm', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐⭐')),
        (11, 'PCI', 'Bombas', 'Bomba principal incêndio', 1, None, 1, trace(SRC_DET + ' — Casa de bombas', 'REV04', '⭐⭐⭐⭐')),
        (12, 'PCI', 'Bombas', 'Bomba jockey', 1, None, 1, trace(SRC_DET + ' — Casa de bombas', 'REV04', '⭐⭐⭐⭐')),
        (13, 'PCI', 'SDAI', 'Central de alarme', 1, None, 1, trace(SRC_DET + ' — Detalhamento SDAI', 'REV04', '⭐⭐⭐')),
        (14, 'PCI', 'SDAI', 'Detector de fumaça óptico', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐')),
        (15, 'PCI', 'SDAI', 'Sirene audiovisual', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐')),
        (16, 'PCI', 'SDAI', 'Acionador manual', 'un', None, None, trace(SRC_DET, 'REV04', '⭐⭐⭐')),
        (17, 'PCI', 'Extintores', 'PQS 6kg', 'un', None, None, trace('PPCI P11/P12/P13 — contagem visual', 'REV04', '⭐⭐⭐')),
        (18, 'PCI', 'Extintores', 'CO2 6kg', 'un', None, None, trace('PPCI P11/P12/P13 — contagem visual', 'REV04', '⭐⭐⭐')),
        (19, 'PCI', 'PCF', 'Porta corta-fogo P-60/P-90', 'un', 25, None, trace('PPCI P11/P12/P13 — escadas de emergência', 'REV04', '⭐⭐⭐')),
        (20, 'PCI', 'Iluminação', 'Bloco autônomo de emergência', 'un', None, None, trace('PPCI P08/P11/P12/P13', 'REV04', '⭐⭐⭐')),
        (21, 'PCI', 'Sinalização', 'Placas fotoluminescentes', 'vb', 1, None, trace('PPCI — identificadas nas plantas', 'REV04', '⭐⭐⭐')),
        (22, 'PCI', 'Elevadores', '2 elevadores sociais (cabine 120×241cm)', 2, None, 2, trace(SRC_TIP + ' — Planta pavimento tipo', 'REV04', '⭐⭐⭐⭐')),
    ]
    
    for entry in pci_items:
        row = entry[0]
        safe_write(ws, row, 1, entry[1])
        safe_write(ws, row, 2, entry[2])
        safe_write(ws, row, 3, entry[3])
        safe_write(ws, row, 4, entry[4] if isinstance(entry[4], str) else entry[4])
        if entry[5] is not None and not isinstance(entry[5], str):
            safe_write(ws, row, 5, entry[5])
        safe_write(ws, row, 10, entry[7])
        safe_font(ws, row, 10, trace_font)
    
    print("  ✅ PCI preenchida")


def main():
    print("📂 Carregando template...")
    wb = openpyxl.load_workbook(TEMPLATE)
    
    print(f"📋 {len(wb.sheetnames)} abas\n")
    print("🔧 Preenchendo com dados extraídos + rastreabilidade:\n")
    
    fill_projetos(wb)
    fill_obra(wb)
    fill_estacas(wb)
    fill_aco_estacas(wb)
    fill_instalacoes_resumo(wb)
    fill_climatizacao(wb)
    fill_spda(wb)
    fill_pci(wb)
    
    print(f"\n💾 Salvando: {OUTPUT}")
    wb.save(OUTPUT)
    print("✅ Planilha preenchida com rastreabilidade!")
    print(f"   Arquivo: {os.path.basename(OUTPUT)}")
    
    # Stats
    print("\n📊 Abas preenchidas:")
    print("   ✅ PROJETOS — mapa de disciplinas e PDFs")
    print("   ✅ Obra — quadro de áreas completo")
    print("   ✅ Estacas — 202 estacas HC (01A exato, 01B estimado)")
    print("   ✅ Aço_Estacas — armadura 14 ton (01A) + 28 ton (estimado total)")
    print("   ✅ INSTALAÇÕES — resumo geral")
    print("   ✅ CLIMATIZAÇÃO — 341 splits + 3.704m tubulação (dados reais AVAC)")
    print("   ✅ SPDA — sistema estrutural, 5 caixas equalização")
    print("   ✅ PCI — SHP, SDAI, extintores, PCF, elevadores")
    print("\n⏳ Abas pendentes (dados insuficientes para preencher):")
    print("   ⏳ HIDROSSANITÁRIO — precisa quantitativo detalhado")
    print("   ⏳ ELÉTRICO — precisa quantitativo detalhado")
    print("   ⏳ TELEFONE — precisa quantitativo detalhado")
    print("   ⏳ Ger_Executivo — precisa serviços individuais com CPU")
    print("   ⏳ Visus tabs — precisa exportação BIM")
    print("   ⏳ Supraestrutura — dados parciais (só pav. Lazer)")


if __name__ == '__main__':
    main()
