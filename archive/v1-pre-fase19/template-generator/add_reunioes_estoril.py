#!/usr/bin/env python3
"""
Adiciona aba 'Reuniões' à planilha do Estoril com dados extraídos do MongoDB.
Inclui: decisões de projeto, especificações confirmadas, prazos.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUTPUT = str(Path.home() / "clawd" / "orcamento-parametrico" / "projetos" / "estoril" / "Estoril-Orcamento-Executivo-PREENCHIDO.xlsx")

BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
YELLOW = "FFF2CC"
LIGHT_GREEN = "E2EFDA"
LIGHT_ORANGE = "FCE4D6"
WHITE = "FFFFFF"

title_font = Font(name='Calibri', bold=True, size=14, color=WHITE)
title_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
header_font = Font(name='Calibri', bold=True, size=10, color=BLUE)
header_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
date_font = Font(name='Calibri', bold=True, size=11, color=BLUE)
date_fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
item_font = Font(name='Calibri', size=10)
item_font_bold = Font(name='Calibri', size=10, bold=True)
spec_fill = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
action_fill = PatternFill(start_color=LIGHT_ORANGE, end_color=LIGHT_ORANGE, fill_type="solid")
trace_font = Font(name='Calibri', size=8, italic=True, color="808080")

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def write_title(ws, row, title, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28


def write_date_header(ws, row, text, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = date_font
    cell.fill = date_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 24


def write_section(ws, row, section_name, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=section_name)
    cell.font = item_font_bold
    cell.fill = header_fill
    ws.row_dimensions[row].height = 20


def write_item(ws, row, category, item, spec, impact, source, fill=None):
    vals = [category, item, spec, impact, source, '']
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = item_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if fill:
            cell.fill = fill


def main():
    wb = openpyxl.load_workbook(OUTPUT)
    
    # Remover se já existe
    if 'Reuniões' in wb.sheetnames:
        del wb['Reuniões']
    
    # Criar após "Orçamento Consolidado" (posição 1)
    ws = wb.create_sheet("Reuniões", 1)
    ws.sheet_properties.tabColor = "4472C4"
    
    # Larguras
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    
    row = 1
    write_title(ws, row, 'INFORMAÇÕES DE REUNIÕES — RESIDENCIAL CIDADE ESTORIL')
    row += 1
    ws.cell(row=row, column=1, value='Fonte: MongoDB Atlas — transcricoes.reunioes | Obra ID: 693ad2dae3d07ec6453328c4 | 4 reuniões processadas').font = trace_font
    row += 2
    
    # ═══════════════════════════════════════════
    # SEÇÃO 1: ESPECIFICAÇÕES CONFIRMADAS (consolidado de todas as reuniões)
    # ═══════════════════════════════════════════
    write_title(ws, row, 'ESPECIFICAÇÕES CONFIRMADAS EM REUNIÕES', cols=6)
    row += 1
    
    headers = ['Categoria', 'Item', 'Especificação', 'Impacto no Orçamento', 'Fonte (Reunião)', 'Status']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    row += 1
    
    # Dados da reunião 10/dez/2025 — Apresentação do modelo
    specs = [
        # (categoria, item, spec, impacto, fonte)
        ('Cobertura', 'Telhado', 'Telha fibrocimento + estrutura de madeira', 'Padrão médio — compatível com mediana', '10/dez/2025'),
        ('Cobertura', 'Calha e rufos', 'Aço galvanizado', 'Padrão médio', '10/dez/2025'),
        ('Fachada', 'Textura', 'Textura acrílica projetada (áreas com pintura)', 'Custo menor que cerâmica', '10/dez/2025'),
        ('Fachada', 'ACM', 'Ripado e parte amadeirada em ACM', 'Custo elevado — impacta macrogrupo Fachada', '10/dez/2025'),
        ('Fachada', 'Fundo (áreas ACM)', 'Parede pintada de preto', 'Mínimo', '10/dez/2025'),
        ('Esquadrias', 'Peitoris de janela', 'Granito São Gabriel polido', 'Padrão médio-alto', '10/dez/2025'),
        ('Esquadrias', 'Guarda-corpo sacadas/cobertura', 'Vidro + estrutura metálica ("asa de avião")', 'Custo acima da mediana', '10/dez/2025'),
        ('Esquadrias', 'Guarda-corpo (demais)', 'Metálico', 'Padrão', '10/dez/2025'),
        ('Pisos', 'Piso apartamentos (dormitórios)', 'Vinílico', 'Custo médio', '10/dez/2025'),
        ('Pisos', 'Piso apartamentos (demais áreas)', 'Cerâmica 72×70cm', 'Formato grande — custo acima', '10/dez/2025'),
        ('Pisos', 'Piso garagem', 'Lapidado (demarcação perímetro tinta)', 'Econômico', '10/dez/2025'),
        ('Pisos', 'Calçadas e rampas externas', 'Concreto estampado', 'Padrão médio', '10/dez/2025'),
        ('Pisos', 'Áreas técnicas e escadas', 'Piso pintado', 'Econômico', '10/dez/2025'),
        ('Pisos', 'Contrapiso', 'Espessura média 5cm (geral)', 'Padronizado', '10/dez/2025'),
        ('Pisos', 'Soleiras', 'NÃO serão utilizadas (transição por contrapiso)', 'Economia', '10/dez/2025'),
        ('Pisos', 'Deck cobertura', '2 opções: madeira ecológica OU piso amadeirado', 'A definir', '10/dez/2025'),
        ('Teto', 'Apartamentos e áreas nobres', 'Forro de gesso mineral', 'Padrão médio', '10/dez/2025'),
        ('Teto', 'Garagens e áreas técnicas', 'Regularização sobre laje (sem gesso)', 'Econômico', '10/dez/2025'),
        ('Pintura', 'Áreas técnicas', 'Fundo preparador à base de solvente', 'Padrão', '10/dez/2025'),
        ('Pintura', 'Áreas nobres', 'Massa acrílica + pintura', 'Padrão médio-alto', '10/dez/2025'),
        ('Impermeabilização', 'Áreas externas (lazer, floreiras)', 'Manta asfáltica', 'Padrão', '10/dez/2025'),
        ('Impermeabilização', 'Áreas molhadas internas', 'Argamassa polimérica + tela', 'Econômico vs manta', '10/dez/2025'),
        ('Impermeabilização', 'Box banheiro', 'Impermeabilização completa (parede inteira)', 'Maior área = custo acima', '10/dez/2025'),
        ('Impermeabilização', 'Demais paredes banheiro', 'Até 1,50m de altura', 'Padrão', '10/dez/2025'),
        ('Lazer', 'Pergolado', 'Alumínio + cobertura de vidro (não policarbonato)', 'Custo acima de policarbonato', '10/dez/2025'),
        ('Lazer', 'Floreiras fachada', 'Metálicas', 'Custo médio', '10/dez/2025'),
        ('Lazer', 'Floreiras lazer', 'Alvenaria', 'Econômico', '10/dez/2025'),
        ('Alteração futura', 'Cobertura Torre B → apartamentos', 'NÃO considerada no orçamento atual', 'Mantida cobertura original', '10/dez/2025'),
    ]
    
    for s in specs:
        write_item(ws, row, s[0], s[1], s[2], s[3], s[4], fill=spec_fill)
        row += 1
    
    row += 2
    
    # ═══════════════════════════════════════════
    # SEÇÃO 2: CRONOGRAMA E PRAZOS
    # ═══════════════════════════════════════════
    write_title(ws, row, 'CRONOGRAMA E PRAZOS DEFINIDOS EM REUNIÕES', cols=6)
    row += 1
    
    headers2 = ['Área', 'Atividade', 'Prazo/Decisão', 'Impacto', 'Fonte (Reunião)', 'Status']
    for i, h in enumerate(headers2, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    row += 1
    
    prazos = [
        ('Geral', 'Meta entrega obra', 'Jul/2027 (c/ gordura) — set/2027 (prazo oficial)', 'Prazo total ~18 meses', '30/jan/2026'),
        ('Estrutura', 'Ciclo por laje', '16 dias úteis / 15 dias (última ref.)', 'Define ritmo de supraestrutura', '28/jan/2026'),
        ('Estrutura', 'Conclusão estrutura', 'Junho/2026', 'Libera acabamentos', '30/jan/2026'),
        ('Acabamentos', 'Ciclo por pavimento', '13 dias (reboco → contrapiso → forro → pintura)', 'Ritmo constante — 1 pav a cada 13 dias', '04/mar/2026'),
        ('Alvenaria', 'Separação Seporex', 'Seporex (dutos) separado da convencional (cerâmico)', 'Equipes diferentes, ritmo diferente', '04/mar/2026'),
        ('Elevadores', 'Kit obra Torre B', '22 dias após entrega material', 'Logística vertical', '04/mar/2026'),
        ('Elevadores', 'Definitivos', 'Jan/2027 (9 meses antes da entrega)', 'Instalação longa', '30/jan/2026'),
        ('Garagens', 'G3 (canteiro) e Térreo', 'Últimos 6 meses da obra', 'Área livre como canteiro até final', '04/mar/2026'),
        ('Garagens', 'G4 e G5', 'Final 2026 / início 2027', 'Execução com equipe própria', '04/mar/2026'),
        ('Fachada', 'Reboco embasamento', 'Abril/2026', 'Acabamento da fachada só no final', '04/mar/2026'),
        ('Fachada', 'Esquadrias torres', 'Após reboco externo, por pano de fachada', 'Não por pavimento — otimiza fluxo', '28/jan/2026'),
        ('Lazer', 'Alvenaria lazer', 'Após conclusão alvenaria Tipo 3', 'Vínculo redefinido (era Tipo 8)', '04/mar/2026'),
        ('Impermeabilização', 'Cisterna', 'Refeita ao final, após elevador Torre A', 'Atividade final', '04/mar/2026'),
    ]
    
    for p in prazos:
        write_item(ws, row, p[0], p[1], p[2], p[3], p[4], fill=action_fill)
        row += 1
    
    row += 2
    
    # ═══════════════════════════════════════════
    # SEÇÃO 3: STATUS DA OBRA (jan-mar 2026)
    # ═══════════════════════════════════════════
    write_title(ws, row, 'STATUS DA OBRA (JAN-MAR 2026)', cols=6)
    row += 1
    
    headers3 = ['Sistema', 'Informação', 'Status', 'Observação', 'Fonte', '']
    for i, h in enumerate(headers3, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    row += 1
    
    status = [
        ('Estrutura', 'Torre A — 9º pav tipo', 'Em execução', 'Próxima concretagem 13/fev', '28/jan/2026'),
        ('Estrutura', 'Torre B — 8º pav tipo', 'Em execução', '', '28/jan/2026'),
        ('Alvenaria', '2º pav tipo (jan/2026)', 'Em execução', 'Atraso em relação à estrutura', '28/jan/2026'),
        ('Impermeabilização', 'Cisterna', 'Precisa ser refeita', 'Será última atividade', '04/mar/2026'),
        ('Instalações', 'Furos nas lajes', 'PROBLEMA — furos faltantes', 'Impacta avanço hidráulico', '30/jan/2026'),
        ('Garagens', 'Rampas', 'Comprimento incorreto — demolição parcial', 'Retrabalho necessário', '30/jan/2026'),
        ('Planejamento', 'Cronograma anterior', 'Descartado — refeito do zero', 'Novo planejamento no Agilim', '28/jan/2026'),
        ('Torres', 'Configuração', '2 torres gêmeas (A e B)', 'Mesma planta, execução sequenciada', '04/mar/2026'),
    ]
    
    for s in status:
        write_item(ws, row, s[0], s[1], s[2], s[3], s[4])
        row += 1
    
    row += 2
    
    # ═══════════════════════════════════════════
    # SEÇÃO 4: LISTA DE REUNIÕES
    # ═══════════════════════════════════════════
    write_title(ws, row, 'REUNIÕES PROCESSADAS', cols=6)
    row += 1
    
    reunioes = [
        ('04/03/2026', 'CTN & Fonseca - Alinhamentos Planejamento', 'Validação do planejamento no Agilim, produtividade equipes, sequenciamento acabamentos'),
        ('30/01/2026', 'CTN & Fonseca - Alinhamentos Diagrama de Rede', 'Sequenciamento garagens, elevadores, prazo final, rampas, subestação'),
        ('28/01/2026', 'CTN & Fonseca - Alinhamentos Diagrama de Rede', 'Novo cronograma do zero, ciclo de laje, status da obra, sequência acabamentos'),
        ('10/12/2025', 'Apresentação do Modelo', 'Especificações de acabamento, materiais, métodos construtivos — base para orçamento'),
    ]
    
    for r in reunioes:
        write_item(ws, row, r[0], r[1], r[2], '', '')
        row += 1
    
    # ── Salvar ──
    wb.save(OUTPUT)
    print(f"✅ Aba 'Reuniões' adicionada: {len(specs)} especificações, {len(prazos)} prazos, {len(status)} status")
    print(f"   Total: {len(wb.sheetnames)} abas")
    for i, name in enumerate(wb.sheetnames):
        print(f"   {i+1}. {name}")


if __name__ == '__main__':
    main()
