#!/usr/bin/env python3
"""
Consolidador de Orçamento Executivo - Parador AG7
Lê todos os JSONs das disciplinas e gera Excel executivo
"""

import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Paths
DISCIPLINAS_DIR = Path("~/orcamentos/projetos/parador-ag7/disciplinas").expanduser()
OUTPUT_EXCEL = Path("~/orcamentos/projetos/parador-ag7/ORCAMENTO-EXECUTIVO-PARADOR-AG7.xlsx").expanduser()

# Disciplinas principais (ordem de apresentação)
DISCIPLINAS = [
    ("estrutura.json", "01-Estrutura"),
    ("ancoragem.json", "02-Ancoragem"),
    ("vedacoes.json", "03-Vedações"),
    ("hidro-agua-fria.json", "04-Hidro Água Fria"),
    ("hidro-agua-quente.json", "05-Hidro Água Quente"),
    ("hidro-esgoto-subsolo.json", "06-Hidro Esg Subsolo"),
    ("hidro-esgoto-terreo.json", "07-Hidro Esg Térreo"),
    ("hidro-esgoto-tipo.json", "08-Hidro Esg Tipo"),
    ("hidro-pluvial.json", "09-Hidro Pluvial"),
    ("impermeab-subsolo.json", "10-Impermeab Subsolo"),
    ("impermeab-rooftop.json", "11-Impermeab Rooftop"),
    ("eletrico.json", "12-Elétrico"),
    ("telecomunicacoes.json", "13-Telecom"),
    ("pci.json", "14-PCI"),
    ("gas-completo.json", "15-Gás"),
    ("paisagismo-completo.json", "16-Paisagismo"),
    ("esquadrias.json", "17-Esquadrias"),
    ("piscinas.json", "18-Piscinas"),
    ("arq-piso-subsolo.json", "19-Arq Piso Sub"),
    ("arq-pisos-forros.json", "20-Arq Pisos Forros"),
]

def load_json(filepath):
    """Carrega JSON com tratamento de erros"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠️ Erro ao ler {filepath.name}: {e}")
        return None

def processar_disciplina(ws, data, disciplina_nome):
    """Processa dados de disciplina e escreve na worksheet"""
    
    # Headers
    headers = ['Categoria', 'Item/Descrição', 'Quantidade', 'Unidade', 'Observações']
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Verificar se tem categorias
    if not data or 'categorias' not in data:
        ws.append(['SEM DADOS', 'Disciplina não processada ou vazia', '', '', ''])
        return
    
    # Processar cada categoria
    row_num = 2
    for cat in data.get('categorias', []):
        cat_nome = cat.get('nome', 'Categoria sem nome')
        items = cat.get('items', [])
        
        if not items:
            ws.append([cat_nome, '(vazio)', '', '', ''])
            row_num += 1
            continue
        
        for item in items:
            # Tentar extrair descrição (vários formatos possíveis)
            descricao = (
                item.get('descricao') or 
                item.get('tipo') or 
                item.get('nome') or 
                item.get('item') or
                item.get('diametro', '')
            )
            
            quantidade = item.get('quantidade') or item.get('area') or item.get('comprimento') or ''
            unidade = item.get('unidade', '')
            obs = item.get('observacao', '') or item.get('localizacao', '')
            
            ws.append([cat_nome, str(descricao), quantidade, unidade, obs])
            row_num += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    
    # Freeze primeira linha
    ws.freeze_panes = 'A2'

def criar_aba_resumo(wb):
    """Cria aba de resumo executivo"""
    ws = wb.create_sheet("00-RESUMO", 0)
    
    resumo_data = [
        ["ORÇAMENTO EXECUTIVO - PARADOR AG7"],
        ["AG7 Incorporadora - Santa Catarina S.A."],
        [""],
        ["Data de Extração", "11/03/2026"],
        ["Método", "Extração automática de projetos executivos"],
        ["Status", "Base para orçamentação (70-80% dos quantitativos)"],
        [""],
        ["⚠️ OBSERVAÇÕES IMPORTANTES"],
        [""],
        ["1. ELÉTRICO: Apenas 29% dos arquivos processados - solicitar planilha do projetista"],
        ["2. PCI: Quantitativos estimados - validar com projeto executivo"],
        ["3. PAISAGISMO: Faltam espécies vegetais (5 PDFs bloqueados)"],
        ["4. Quantitativos são aproximados - prever margem de segurança 10-15%"],
        [""],
        ["DISCIPLINAS PROCESSADAS"],
        [""],
        ["Nº", "Disciplina", "Status"],
        ["01", "Estrutura", "✅ Completa"],
        ["02", "Ancoragem", "✅ Completa"],
        ["03", "Vedações", "✅ Completa"],
        ["04-09", "Hidrossanitário", "✅ Completo"],
        ["10-11", "Impermeabilização", "✅ Completa"],
        ["12", "Elétrico", "⚠️ Parcial (29%)"],
        ["13", "Telecomunicações", "✅ Completa"],
        ["14", "PCI", "⚠️ Estimativas"],
        ["15", "Gás", "✅ Completa"],
        ["16", "Paisagismo", "⚠️ Parcial (falta lista espécies)"],
        ["17", "Esquadrias", "✅ Completa"],
        ["18", "Piscinas", "✅ Completa"],
        ["19-20", "Arquitetura", "✅ Completa"],
    ]
    
    for row in resumo_data:
        ws.append(row)
    
    # Estilo do título
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws['A2'].font = Font(italic=True, size=12)
    ws['A8'].font = Font(bold=True, size=12, color="FF0000")
    ws['A15'].font = Font(bold=True, size=12)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30

def main():
    print("🚀 Iniciando consolidação de orçamento executivo Parador AG7...")
    
    # Criar workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove sheet padrão
    
    # Criar aba de resumo
    print("📋 Criando aba de resumo...")
    criar_aba_resumo(wb)
    
    # Processar cada disciplina
    total_linhas = 0
    for json_file, sheet_name in DISCIPLINAS:
        filepath = DISCIPLINAS_DIR / json_file
        
        if not filepath.exists():
            print(f"⚠️ {json_file} não encontrado - pulando...")
            continue
        
        print(f"📊 Processando {json_file}...")
        data = load_json(filepath)
        
        if data is None:
            continue
        
        # Criar aba no Excel
        ws = wb.create_sheet(sheet_name)
        
        # Processar e escrever dados
        processar_disciplina(ws, data, sheet_name)
        
        linhas = ws.max_row - 1  # -1 pra não contar o header
        total_linhas += linhas
        print(f"   ✅ {linhas} linhas exportadas")
    
    # Salvar Excel
    print(f"\n💾 Salvando Excel: {OUTPUT_EXCEL}")
    wb.save(OUTPUT_EXCEL)
    
    print(f"\n✅ Consolidação completa!")
    print(f"📁 Arquivo gerado: {OUTPUT_EXCEL}")
    print(f"📊 {len(wb.sheetnames)} abas criadas")
    print(f"📈 {total_linhas} linhas totais de quantitativos")

if __name__ == "__main__":
    main()
