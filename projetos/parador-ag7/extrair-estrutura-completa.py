#!/usr/bin/env python3
"""
Extração COMPLETA de quantitativos de estrutura da planilha de orçamento
"""

import openpyxl
import json

wb = openpyxl.load_workbook('orcamento-r01.xlsx', read_only=True, data_only=True)
sheet = wb['Resumo Estrutura']

dados = {
    "projeto": "Ícaro Parador - AG7",
    "fonte": "Planilha CTN-AG7_PRD-Orcamento-Executivo-R01.xlsx",
    "aba": "Resumo Estrutura",
    "data_extracao": "2026-03-11",
    "totais": {
        "concreto_m3": 0,
        "forma_m2": 0,
        "aco_kg": 0
    },
    "por_disciplina": {}
}

# Mapear colunas
COL_ITEM = 1
COL_EAP = 2
COL_DISCIPLINA = 3
COL_PAVIMENTO = 4
COL_ETAPA = 5
COL_CONCRETO = 9
COL_FORMA = 10
COL_ACO = 15

# Linha 4 tem os cabeçalhos
# Começar da linha 5
for row in range(5, sheet.max_row + 1):
    item = sheet.cell(row, COL_ITEM).value
    eap = sheet.cell(row, COL_EAP).value
    disciplina = sheet.cell(row, COL_DISCIPLINA).value
    pavimento = sheet.cell(row, COL_PAVIMENTO).value
    etapa = sheet.cell(row, COL_ETAPA).value
    
    concreto = sheet.cell(row, COL_CONCRETO).value
    forma = sheet.cell(row, COL_FORMA).value
    aco = sheet.cell(row, COL_ACO).value
    
    # Se tem quantitativo
    if concreto or forma or aco:
        # Criar chave da disciplina
        disc_key = f"{disciplina or 'Geral'} - {pavimento or 'Geral'}"
        
        if disc_key not in dados["por_disciplina"]:
            dados["por_disciplina"][disc_key] = {
                "concreto_m3": 0,
                "forma_m2": 0,
                "aco_kg": 0,
                "items": []
            }
        
        # Adicionar item
        item_data = {
            "etapa": etapa or "Sem especificação",
            "concreto_m3": float(concreto) if concreto else 0,
            "forma_m2": float(forma) if forma else 0,
            "aco_kg": float(aco) if aco else 0
        }
        
        dados["por_disciplina"][disc_key]["items"].append(item_data)
        
        # Somar totais
        if concreto:
            dados["por_disciplina"][disc_key]["concreto_m3"] += float(concreto)
            dados["totais"]["concreto_m3"] += float(concreto)
        if forma:
            dados["por_disciplina"][disc_key]["forma_m2"] += float(forma)
            dados["totais"]["forma_m2"] += float(forma)
        if aco:
            dados["por_disciplina"][disc_key]["aco_kg"] += float(aco)
            dados["totais"]["aco_kg"] += float(aco)

# Salvar JSON
with open('disciplinas/estrutura-completa.json', 'w', encoding='utf-8') as f:
    json.dump(dados, f, indent=2, ensure_ascii=False)

# Imprimir resumo
print("=== QUANTITATIVOS EXTRAÍDOS - ESTRUTURA COMPLETA ===\n")
print(f"Fonte: {dados['fonte']}")
print(f"Aba: {dados['aba']}\n")

print("=== TOTAIS GERAIS ===")
print(f"Concreto: {dados['totais']['concreto_m3']:,.2f} m³")
print(f"Forma: {dados['totais']['forma_m2']:,.2f} m²")
print(f"Aço: {dados['totais']['aco_kg']:,.2f} kg ({dados['totais']['aco_kg']/1000:,.2f} toneladas)\n")

print("=== POR DISCIPLINA/PAVIMENTO ===")
for disc, values in sorted(dados["por_disciplina"].items()):
    print(f"\n{disc}:")
    print(f"  Concreto: {values['concreto_m3']:,.2f} m³")
    print(f"  Forma: {values['forma_m2']:,.2f} m²")
    print(f"  Aço: {values['aco_kg']:,.2f} kg")
    print(f"  Items: {len(values['items'])}")

print(f"\n✅ JSON salvo: disciplinas/estrutura-completa.json")
