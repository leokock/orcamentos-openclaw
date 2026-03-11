#!/usr/bin/env python3
"""
Análise rápida da planilha de orçamento
Extrai etapas/disciplinas e conta quantitativos
"""

import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('orcamento-r01.xlsx', read_only=True, data_only=True)
    sheet = wb['CPU']
    
    etapas = {}
    etapa_atual = None
    
    print("=== ANÁLISE DA PLANILHA DE ORÇAMENTO ===\n")
    print(f"Total de linhas: {sheet.max_row:,}")
    print(f"Total de colunas: {sheet.max_column}\n")
    
    print("=== ETAPAS/DISCIPLINAS (primeiras 100 linhas) ===\n")
    
    for row in range(2, min(102, sheet.max_row + 1)):
        tipo = sheet.cell(row, 2).value
        descricao = sheet.cell(row, 4).value
        qtd = sheet.cell(row, 6).value
        
        if tipo == "Etapa" and descricao:
            etapa_atual = descricao
            if etapa_atual not in etapas:
                etapas[etapa_atual] = 0
            print(f"{len(etapas):2}. {descricao}")
        
        # Contar insumos/serviços com quantidade
        if tipo in ["Insumo", "Serviço"] and qtd and etapa_atual:
            etapas[etapa_atual] += 1
    
    print(f"\n=== RESUMO ===")
    print(f"Total de etapas identificadas (primeiras 100 linhas): {len(etapas)}")
    print("\nItens por etapa:")
    for etapa, count in list(etapas.items())[:10]:
        print(f"  - {etapa}: {count} itens")
    
    # Verificar se tem aba de Estrutura
    print(f"\n=== ABAS COM QUANTITATIVOS ===")
    for name in wb.sheetnames:
        if any(keyword in name.lower() for keyword in ['estrutura', 'estaca', 'fund', 'cpu', 'arquitetura']):
            sheet_temp = wb[name]
            print(f"  - {name}: {sheet_temp.max_row:,} linhas")

except Exception as e:
    print(f"Erro: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
