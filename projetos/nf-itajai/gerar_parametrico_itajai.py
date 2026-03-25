#!/usr/bin/env python3.11
"""
Gerador de Orçamento Paramétrico - Projeto Itajaí/SC
Cartesian Engenharia - 2026-03-20
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===============================================
# DADOS DO PROJETO
# ===============================================

PROJETO = {
    'nome': 'Residencial Itajaí - Uso Misto',
    'localizacao': 'Itajaí/SC',
    'data': '2026-03-20',
    'responsavel': 'Laura Heidrich / Leo Kock',
    
    # Variáveis do briefing
    'ac': 39472,  # m²
    'ur': 332,    # unidades
    'np': 28,     # pavimentos (estimado)
    'npt': 20,    # pavimentos tipo (estimado)
    'ns': 1,      # subsolos
    'pd': 2.80,   # pé-direito médio (m)
    'at': None,   # área do terreno (faltante)
    
    # Composição
    'studios': 194,
    'aptos_1s1d': 99,
    'aptos_2d': 20,
    'lofts': 19,
    
    # Premissas
    'padrao': 'Alto',
    'uso': 'Misto (Residencial + Comercial)',
    'garagem_niveis': 5,
    'subsolo': True,
    'contencao': True,
    'fundacao': 'Estacas pré-moldadas (assumido)',
    'estrutura': 'Concreto Armado',
    'lazer_comercial': 1000,  # m²
    'lazer_residencial': 300,  # m²
}

# Carregar base de calibração
with open('parametrico/calibration-stats.json', 'r') as f:
    STATS = json.load(f)

# Benchmark: Costao (projeto mais similar)
BENCHMARK = {
    'nome': 'Costão',
    'ac': 38429,
    'rsm2_total': 4238.92,
    'infra_rsm2': 263.46,
    'supra_rsm2': 698.98,
    'infra_pct': 6.22,
    'supra_pct': 16.49,
}

# ===============================================
# ESTIMATIVAS DE CUSTO
# ===============================================

def calcular_custos():
    """Calcula custos por macrogrupo com ajustes para o projeto."""
    
    # Medianas da base
    medianas = {cat: data['median'] for cat, data in STATS['categories'].items()}
    
    # Ajustes específicos para Itajaí (alto padrão + uso misto)
    # Baseado em Costão + Scenarium + Colline
    
    custos = {
        # INFRAESTRUTURA — ajustar para cima (subsolo + lençol freático em Itajaí)
        'Mov. Terra': medianas['Mov. Terra'] * 1.8,  # mais escavação (5 níveis garagem)
        'Contenção': 150.00,  # assumido (subsolo urbano)
        'Infraestrutura': 280.00,  # acima da mediana (complexidade costeira)
        
        # SUPRAESTRUTURA — alto padrão
        'Supraestrutura': 720.00,  # próximo de Costão/Colline
        
        # ACABAMENTOS — alto padrão + uso misto
        'Alvenaria': medianas['Alvenaria'] * 1.1,
        'Esquadrias': medianas['Esquadrias'] * 1.3,  # alto padrão
        'Fachada': medianas['Fachada'] * 1.4,  # complexa (uso misto)
        'Pisos': medianas['Pisos'] * 1.2,
        'Rev. Int. Parede': medianas['Rev. Int. Parede'] * 1.1,
        'Pintura': medianas['Pintura'],
        'Teto': medianas['Teto'],
        'Cobertura': medianas['Cobertura'] if 'Cobertura' in medianas else 20,
        
        # INSTALAÇÕES
        'Instalações': medianas['Instalações'] * 1.15,  # uso misto
        'Climatização': medianas['Climatização'] * 1.2,  # alto padrão
        'Sist. Especiais': medianas['Sist. Especiais'] * 1.3,  # automação
        
        # COMPLEMENTARES
        'Impermeabilização': medianas['Impermeabilização'] * 1.3,  # subsolo + lençol
        'Louças e Metais': medianas['Louças e Metais'] * 1.5,  # alto padrão
        'Complementares': medianas['Complementares'] * 1.2,
        'Gerenciamento': medianas['Gerenciamento'],
        'Imprevistos': medianas['Imprevistos'],
    }
    
    return custos

# ===============================================
# GERAR PLANILHA
# ===============================================

def criar_planilha():
    """Cria planilha Excel com orçamento paramétrico."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PAINEL"
    
    # Estilos
    titulo = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    subtitulo = Font(name='Calibri', size=12, bold=True)
    header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    normal = Font(name='Calibri', size=10)
    
    fill_titulo = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    fill_destaque = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    fill_alerta = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    fill_ok = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # TÍTULO
    ws.merge_cells('A1:G1')
    cell = ws['A1']
    cell.value = f"ORÇAMENTO PARAMÉTRICO — {PROJETO['nome'].upper()}"
    cell.font = titulo
    cell.fill = fill_titulo
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # INFO DO PROJETO
    linha = 3
    ws[f'A{linha}'] = 'DADOS DO PROJETO'
    ws[f'A{linha}'].font = subtitulo
    linha += 1
    
    info = [
        ('Localização:', PROJETO['localizacao']),
        ('Data:', PROJETO['data']),
        ('Responsável:', PROJETO['responsavel']),
        ('', ''),
        ('Área Construída (AC):', f"{PROJETO['ac']:,.2f} m²"),
        ('Unidades (UR):', f"{PROJETO['ur']}"),
        ('Pavimentos (NP):', f"{PROJETO['np']} (estimado)"),
        ('Pavimentos Tipo (NPT):', f"{PROJETO['npt']} (estimado)"),
        ('Subsolos (NS):', f"{PROJETO['ns']}"),
        ('Níveis de Garagem:', f"{PROJETO['garagem_niveis']}"),
        ('Pé-direito médio:', f"{PROJETO['pd']:.2f} m"),
        ('', ''),
        ('Padrão:', PROJETO['padrao']),
        ('Uso:', PROJETO['uso']),
        ('Fundação (premissa):', PROJETO['fundacao']),
        ('Estrutura:', PROJETO['estrutura']),
    ]
    
    for label, valor in info:
        ws[f'A{linha}'] = label
        ws[f'B{linha}'] = valor
        ws[f'A{linha}'].font = Font(bold=True if label else False, size=10)
        linha += 1
    
    # COMPOSIÇÃO DAS UNIDADES
    linha += 1
    ws[f'A{linha}'] = 'COMPOSIÇÃO DAS UNIDADES'
    ws[f'A{linha}'].font = subtitulo
    linha += 1
    
    ws[f'A{linha}'] = 'Tipologia'
    ws[f'B{linha}'] = 'Quantidade'
    ws[f'C{linha}'] = '%'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{linha}'].font = header
        ws[f'{col}{linha}'].fill = fill_header
        ws[f'{col}{linha}'].border = borda
    linha += 1
    
    unidades = [
        ('Studios', PROJETO['studios']),
        ('1 Suíte + 1 Dormitório', PROJETO['aptos_1s1d']),
        ('2 Dormitórios', PROJETO['aptos_2d']),
        ('Lofts', PROJETO['lofts']),
    ]
    
    for tipo, qtd in unidades:
        ws[f'A{linha}'] = tipo
        ws[f'B{linha}'] = qtd
        ws[f'C{linha}'] = f"{qtd/PROJETO['ur']*100:.1f}%"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{linha}'].border = borda
        linha += 1
    
    # Total
    ws[f'A{linha}'] = 'TOTAL'
    ws[f'B{linha}'] = PROJETO['ur']
    ws[f'C{linha}'] = '100,0%'
    for col in ['A', 'B', 'C']:
        ws[f'{col}{linha}'].font = Font(bold=True)
        ws[f'{col}{linha}'].border = borda
        ws[f'{col}{linha}'].fill = fill_destaque
    
    # CUSTOS POR MACROGRUPO
    linha += 3
    ws[f'A{linha}'] = 'CUSTOS POR MACROGRUPO (R$/m²)'
    ws[f'A{linha}'].font = subtitulo
    linha += 1
    
    ws[f'A{linha}'] = 'Macrogrupo'
    ws[f'B{linha}'] = 'R$/m²'
    ws[f'C{linha}'] = 'Total (R$)'
    ws[f'D{linha}'] = '%'
    ws[f'E{linha}'] = 'Mediana Base'
    ws[f'F{linha}'] = 'Status'
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{linha}'].font = header
        ws[f'{col}{linha}'].fill = fill_header
        ws[f'{col}{linha}'].border = borda
    linha += 1
    
    custos = calcular_custos()
    
    # Ordenar por valor decrescente
    custos_ordenados = sorted(custos.items(), key=lambda x: x[1], reverse=True)
    
    custo_total_rsm2 = sum(custos.values())
    custo_total_projeto = custo_total_rsm2 * PROJETO['ac']
    
    linha_infra_start = None
    linha_supra = None
    
    for macro, rsm2 in custos_ordenados:
        # Destacar Infraestrutura e Supraestrutura
        if macro in ['Infraestrutura', 'Supraestrutura', 'Mov. Terra', 'Contenção']:
            ws[f'A{linha}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            if macro == 'Infraestrutura' and linha_infra_start is None:
                linha_infra_start = linha
            if macro == 'Supraestrutura':
                linha_supra = linha
        
        ws[f'A{linha}'] = macro
        ws[f'B{linha}'] = rsm2
        ws[f'B{linha}'].number_format = 'R$ #,##0.00'
        ws[f'C{linha}'] = rsm2 * PROJETO['ac']
        ws[f'C{linha}'].number_format = 'R$ #,##0.00'
        ws[f'D{linha}'] = rsm2 / custo_total_rsm2
        ws[f'D{linha}'].number_format = '0.00%'
        
        # Mediana da base
        mediana_base = STATS['categories'].get(macro, {}).get('median', 0)
        ws[f'E{linha}'] = mediana_base if mediana_base > 0 else '-'
        if mediana_base > 0:
            ws[f'E{linha}'].number_format = 'R$ #,##0.00'
        
        # Status: acima/abaixo da mediana
        if mediana_base > 0:
            diff_pct = (rsm2 / mediana_base - 1) * 100
            if diff_pct > 20:
                ws[f'F{linha}'] = f'↑ {diff_pct:.0f}%'
                ws[f'F{linha}'].fill = fill_alerta
            elif diff_pct < -20:
                ws[f'F{linha}'] = f'↓ {diff_pct:.0f}%'
                ws[f'F{linha}'].fill = fill_ok
            else:
                ws[f'F{linha}'] = f'{diff_pct:+.0f}%'
        else:
            ws[f'F{linha}'] = '-'
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{linha}'].border = borda
        
        linha += 1
    
    # Total
    ws[f'A{linha}'] = 'TOTAL GERAL'
    ws[f'B{linha}'] = custo_total_rsm2
    ws[f'B{linha}'].number_format = 'R$ #,##0.00'
    ws[f'C{linha}'] = custo_total_projeto
    ws[f'C{linha}'].number_format = 'R$ #,##0.00'
    ws[f'D{linha}'] = 1.00
    ws[f'D{linha}'].number_format = '0.00%'
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{linha}'].font = Font(bold=True, size=11)
        ws[f'{col}{linha}'].border = borda
        ws[f'{col}{linha}'].fill = fill_destaque
    
    # DESTAQUE: INFRAESTRUTURA + SUPRAESTRUTURA
    linha += 3
    ws[f'A{linha}'] = 'FOCO: INFRAESTRUTURA + SUPRAESTRUTURA'
    ws[f'A{linha}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    ws[f'A{linha}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    ws.merge_cells(f'A{linha}:F{linha}')
    ws[f'A{linha}'].alignment = Alignment(horizontal='center')
    linha += 1
    
    # Calcular totais de infra+supra
    infra_rsm2 = custos.get('Infraestrutura', 0) + custos.get('Mov. Terra', 0) + custos.get('Contenção', 0)
    supra_rsm2 = custos.get('Supraestrutura', 0)
    infra_supra_rsm2 = infra_rsm2 + supra_rsm2
    
    infra_total = infra_rsm2 * PROJETO['ac']
    supra_total = supra_rsm2 * PROJETO['ac']
    infra_supra_total = infra_supra_rsm2 * PROJETO['ac']
    
    ws[f'A{linha}'] = 'Item'
    ws[f'B{linha}'] = 'R$/m²'
    ws[f'C{linha}'] = 'Total (R$)'
    ws[f'D{linha}'] = '% do Total'
    ws[f'E{linha}'] = 'Benchmark (Costão)'
    ws[f'F{linha}'] = 'Variação'
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{linha}'].font = header
        ws[f'{col}{linha}'].fill = fill_header
        ws[f'{col}{linha}'].border = borda
    linha += 1
    
    # Infraestrutura
    ws[f'A{linha}'] = 'INFRAESTRUTURA (Mov.Terra + Contenção + Infra)'
    ws[f'B{linha}'] = infra_rsm2
    ws[f'B{linha}'].number_format = 'R$ #,##0.00'
    ws[f'C{linha}'] = infra_total
    ws[f'C{linha}'].number_format = 'R$ #,##0.00'
    ws[f'D{linha}'] = infra_rsm2 / custo_total_rsm2
    ws[f'D{linha}'].number_format = '0.00%'
    ws[f'E{linha}'] = BENCHMARK['infra_rsm2']
    ws[f'E{linha}'].number_format = 'R$ #,##0.00'
    diff_infra = (infra_rsm2 / BENCHMARK['infra_rsm2'] - 1) * 100
    ws[f'F{linha}'] = f'{diff_infra:+.1f}%'
    if diff_infra > 10:
        ws[f'F{linha}'].fill = fill_alerta
    elif diff_infra < -10:
        ws[f'F{linha}'].fill = fill_ok
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{linha}'].border = borda
    linha += 1
    
    # Supraestrutura
    ws[f'A{linha}'] = 'SUPRAESTRUTURA'
    ws[f'B{linha}'] = supra_rsm2
    ws[f'B{linha}'].number_format = 'R$ #,##0.00'
    ws[f'C{linha}'] = supra_total
    ws[f'C{linha}'].number_format = 'R$ #,##0.00'
    ws[f'D{linha}'] = supra_rsm2 / custo_total_rsm2
    ws[f'D{linha}'].number_format = '0.00%'
    ws[f'E{linha}'] = BENCHMARK['supra_rsm2']
    ws[f'E{linha}'].number_format = 'R$ #,##0.00'
    diff_supra = (supra_rsm2 / BENCHMARK['supra_rsm2'] - 1) * 100
    ws[f'F{linha}'] = f'{diff_supra:+.1f}%'
    if diff_supra > 10:
        ws[f'F{linha}'].fill = fill_alerta
    elif diff_supra < -10:
        ws[f'F{linha}'].fill = fill_ok
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{linha}'].border = borda
    linha += 1
    
    # Total Infra+Supra
    ws[f'A{linha}'] = 'TOTAL INFRA + SUPRA'
    ws[f'B{linha}'] = infra_supra_rsm2
    ws[f'B{linha}'].number_format = 'R$ #,##0.00'
    ws[f'C{linha}'] = infra_supra_total
    ws[f'C{linha}'].number_format = 'R$ #,##0.00'
    ws[f'D{linha}'] = infra_supra_rsm2 / custo_total_rsm2
    ws[f'D{linha}'].number_format = '0.00%'
    bench_is = BENCHMARK['infra_rsm2'] + BENCHMARK['supra_rsm2']
    ws[f'E{linha}'] = bench_is
    ws[f'E{linha}'].number_format = 'R$ #,##0.00'
    diff_is = (infra_supra_rsm2 / bench_is - 1) * 100
    ws[f'F{linha}'] = f'{diff_is:+.1f}%'
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{linha}'].font = Font(bold=True, size=11)
        ws[f'{col}{linha}'].border = borda
        ws[f'{col}{linha}'].fill = fill_destaque
    
    # ALERTAS E DADOS FALTANTES
    linha += 3
    ws[f'A{linha}'] = 'ALERTAS E DADOS FALTANTES'
    ws[f'A{linha}'].font = subtitulo
    ws[f'A{linha}'].fill = fill_alerta
    ws.merge_cells(f'A{linha}:F{linha}')
    linha += 1
    
    alertas = [
        '⚠️ CRÍTICO: Relatório de Sondagem SPT não fornecido',
        '⚠️ CRÍTICO: Número exato de pavimentos tipo (estimado: 20)',
        '⚠️ CRÍTICO: Pé-direito confirmado (assumido: 2,80m)',
        '⚠️ Área do terreno (AT) não informada',
        '⚠️ Projeto estrutural não fornecido',
        '⚠️ Especificações de acabamento não detalhadas',
        '',
        '📌 PREMISSAS ASSUMIDAS:',
        '  • Fundação: Estacas pré-moldadas de concreto',
        '  • Solo: Argiloso/Misto (típico de Itajaí)',
        '  • Contenção: Cortina atirantada (subsolo)',
        '  • Estrutura: Concreto armado convencional',
        '  • Lajes: 12-15 cm (maciças ou nervuradas)',
        '',
        '💡 RECOMENDAÇÃO:',
        '  Solicitar ao cliente os dados faltantes para refinar a estimativa.',
        '  Precisão atual: ±20-30% (nível de viabilidade).',
    ]
    
    for alerta in alertas:
        ws[f'A{linha}'] = alerta
        ws[f'A{linha}'].font = Font(size=9, italic=True)
        ws.merge_cells(f'A{linha}:F{linha}')
        linha += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    
    # Salvar
    output_file = f"output/itajai-parametrico-{PROJETO['data']}.xlsx"
    wb.save(output_file)
    print(f"\n✓ Planilha gerada: {output_file}")
    print(f"\n📊 RESUMO:")
    print(f"   Custo Total: R$ {custo_total_projeto:,.2f}")
    print(f"   Custo/m²:    R$ {custo_total_rsm2:,.2f}/m²")
    print(f"   Infra+Supra: R$ {infra_supra_rsm2:,.2f}/m² ({infra_supra_rsm2/custo_total_rsm2*100:.1f}% do total)")
    
    return output_file

# ===============================================
# MAIN
# ===============================================

if __name__ == '__main__':
    print("=" * 70)
    print("GERADOR DE ORÇAMENTO PARAMÉTRICO — PROJETO ITAJAÍ/SC")
    print("=" * 70)
    print(f"\nProjeto: {PROJETO['nome']}")
    print(f"Localização: {PROJETO['localizacao']}")
    print(f"AC: {PROJETO['ac']:,.0f} m² | UR: {PROJETO['ur']} | NP: {PROJETO['np']} (estimado)")
    print("\nGerando planilha...")
    
    arquivo = criar_planilha()
    
    print("\n" + "=" * 70)
    print("CONCLUÍDO!")
    print("=" * 70)
    print(f"\nArquivo: {arquivo}")
    print("\n💡 Próximos passos:")
    print("   1. Revisar premissas com o cliente")
    print("   2. Solicitar dados faltantes (sondagem, projeto estrutural)")
    print("   3. Refinar estimativa com dados reais")
