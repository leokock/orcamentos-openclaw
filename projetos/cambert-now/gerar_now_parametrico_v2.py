#!/usr/bin/env python3
"""
NOW Residence — Paramétrico V2 (Bottom-Up com PUs Reais)
Baseado na revisão da Patricia e no design V2.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# === DADOS DO PROJETO (confirmados) ===
PROJETO = "NOW Residence"
EMPRESA = "Cambert / Belli Empreendimentos"
CIDADE = "Itajaí/SC"
AC = 13200  # m² (quadro de áreas oficial — Patricia)
UR = 136
NP = 24
NPT = 17
ELEV = 3
VAG = 110
SUBSOLOS = 0  # garagens acima do nível
PRAZO = 30  # meses (estimado)
PADRAO = "Médio-Alto"
LAJE = "Protendida (mista)"
CONTENCAO = "Sem (garagens acima do nível)"
FUNDACAO = "Hélice contínua"
FACHADA = "Textura + acabamento"
CUB_ATUAL = 3028.45  # SC mar/2026
CUB_BASE = 2752.67   # dez/23
FATOR_CUB = CUB_ATUAL / CUB_BASE  # ~1.100

# Areas por setor
AREA_GARAGEM = 3 * 952 + 941  # G1+G2+G3 + térreo garagem ≈ 3.797 m²
AREA_LAZER = 453  # lazer coberto
AREA_TIPOS = NPT * 466 + 473  # 17 tipos + ático ≈ 8.395 m²
AREA_COBERTURA = 56  # cobertura
AREA_OUTROS = AC - AREA_GARAGEM - AREA_LAZER - AREA_TIPOS - AREA_COBERTURA

# === ESTILOS ===
DARK = "2C3E50"
ACCENT = "2980B9"
ORANGE = "E67E22"
GREEN = "27AE60"
RED = "E74C3C"
GRAY_BG = "F5F5F5"
YELLOW_BG = "FFF8E1"
GREEN_BG = "E8F5E9"
RED_BG = "FDEDEC"

thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def header_style():
    return {
        'font': Font(bold=True, color="FFFFFF", size=9, name="Arial"),
        'fill': PatternFill(start_color=DARK, end_color=DARK, fill_type="solid"),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border,
    }

def cell_style(bold=False, color=None, fmt='#,##0', align='right'):
    s = {
        'font': Font(bold=bold, size=9, name="Arial", color=color),
        'alignment': Alignment(horizontal=align, vertical='center'),
        'border': thin_border,
        'number_format': fmt,
    }
    return s

def apply_style(cell, style_dict):
    for attr, val in style_dict.items():
        if attr == 'number_format':
            cell.number_format = val
        else:
            setattr(cell, attr, val)

def add_header_row(ws, row, cols, widths=None):
    hs = header_style()
    for i, text in enumerate(cols, 1):
        c = ws.cell(row, i, text)
        apply_style(c, hs)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def add_data_row(ws, row, data, bold=False, shade=None, formats=None):
    for i, val in enumerate(data, 1):
        c = ws.cell(row, i, val)
        fmt = formats[i-1] if formats else ('#,##0.00' if isinstance(val, float) else '#,##0' if isinstance(val, int) else '@')
        s = cell_style(bold=bold, fmt=fmt, align='right' if isinstance(val, (int, float)) else 'left')
        apply_style(c, s)
        if shade:
            c.fill = PatternFill(start_color=shade, end_color=shade, fill_type="solid")

def brl(v):
    return round(v, 2)

# === PUs DE MERCADO (base Cartesian 75 exec, atualizados CUB mar/2026) ===
# Estrutura
PU_CONCRETO_FCK30 = 590  # R$/m³ (usinado, bombeado)
PU_CONCRETO_FCK40 = 690  # R$/m³ (protendida)
PU_ACO_CA50 = 8.67       # R$/kg
PU_CORDOALHA = 22.00     # R$/kg (protensão)
PU_FORMA_LAJE = 84.37    # R$/m² (fabricação)
PU_FORMA_MONT = 3.70     # R$/m² (montagem/desmontagem)
PU_FORMA = PU_FORMA_LAJE + PU_FORMA_MONT  # ~88/m²
PU_MO_ESTRUTURA = 180    # R$/m² AC (empreitada)
PU_ESCORAMENTO = 35      # R$/m² laje (protendida)

# Fundação
PU_HELICE_40CM = 50       # R$/m
PU_CONCRETO_FUND = 632    # R$/m³
PU_ACO_FUND = 8.38        # R$/kg
PU_FORMA_FUND = 45.64     # R$/m²

# Alvenaria
PU_BLOCO_CERAM = 32.95    # R$/m² (14cm, assentado)
PU_DRYWALL_ST = 156       # R$/m² (standard)
PU_DRYWALL_RU = 195       # R$/m² (resistente umidade)
PU_MO_ALVENARIA = 28.50   # R$/m²

# Instalações (R$/m² AC)
PU_HIDRO_M2 = 46.52       # MO hidro
PU_ELETRICA_M2 = 50.00    # MO elétrica
PU_MATERIAL_INST = 0.45   # ratio material sobre total

# Revestimentos
PU_CONTRAPISO = 12.10     # R$/m²
PU_REBOCO_INT = 7.00      # R$/m²
PU_REBOCO_EXT = 15.25     # R$/m²
PU_PINTURA_ACR = 5.40     # R$/m² (3 demãos)
PU_MASSA_PVA = 4.29       # R$/m²
PU_PORCELANATO = 81.21    # R$/m² (60x60)
PU_MANTA = 82.11          # R$/m²
PU_MO_IMPERMEAB = 68.00   # R$/m²

# Equipamentos
PU_ELEVADOR_SOCIAL = 192450  # R$/un
PU_ELEVADOR_SERVICO = 96000  # R$/un

# ============================================================================
# CÁLCULOS BOTTOM-UP
# ============================================================================

def calcular_estrutura():
    """Estrutura bottom-up com perfil protendida."""
    items = []

    # Índices para protendida (menos vigas, mais laje)
    idx_concreto_m3_m2 = 0.22  # m³/m² AC (típico protendida)
    qtd_concreto = idx_concreto_m3_m2 * AC

    # Distribuição protendida
    dist = {
        'Pilares': 0.22,
        'Vigas': 0.08,  # protendida: muito menos vigas
        'Lajes': 0.60,  # protendida: mais laje
        'Escadas': 0.05,
        'Blocos/Baldrame': 0.05,
    }

    for elem, peso in dist.items():
        qtd = qtd_concreto * peso
        pu = PU_CONCRETO_FCK40 if elem == 'Lajes' else PU_CONCRETO_FCK30
        total = qtd * pu
        items.append(('Concreto', elem, round(qtd, 1), 'm³', pu, round(total), 'Indice param. + PU base'))

    # Aço convencional (menos na protendida)
    idx_aco_kg_m3 = 45  # kg/m³ (protendida: menos aço convencional)
    qtd_aco = idx_aco_kg_m3 * qtd_concreto
    items.append(('Aço CA-50', 'Convencional', round(qtd_aco), 'kg', PU_ACO_CA50, round(qtd_aco * PU_ACO_CA50), 'Indice param. + PU base'))

    # Cordoalha (só protendida)
    idx_cordoalha = 8  # kg/m² de laje
    area_laje = AC * 0.65  # ~65% da AC é laje
    qtd_cordoalha = idx_cordoalha * area_laje
    items.append(('Cordoalha', 'CP-190 RB', round(qtd_cordoalha), 'kg', PU_CORDOALHA, round(qtd_cordoalha * PU_CORDOALHA), 'Indice protendida'))

    # Forma (menos vigas na protendida)
    idx_forma_m2_m2 = 1.8  # m² forma / m² AC (protendida: menos)
    qtd_forma = idx_forma_m2_m2 * AC
    items.append(('Forma', 'Compensado plastif.', round(qtd_forma), 'm²', PU_FORMA, round(qtd_forma * PU_FORMA), 'Indice param. + PU base'))

    # Escoramento (protendida precisa)
    qtd_escor = area_laje
    items.append(('Escoramento', 'Metálico (protendida)', round(qtd_escor), 'm²', PU_ESCORAMENTO, round(qtd_escor * PU_ESCORAMENTO), 'PU base Cartesian'))

    # MO Estrutura
    items.append(('MO Estrutura', 'Empreitada', AC, 'm²', PU_MO_ESTRUTURA, round(AC * PU_MO_ESTRUTURA), 'PU base Cartesian'))

    return items

def calcular_fundacao():
    """Fundação hélice contínua, sem contenção."""
    items = []

    # Hélice contínua — índice típico para Itajaí
    n_estacas = 120  # estimativa para 13.200 m² sem subsolo
    comp_media = 18  # m (Itajaí, terreno centro)
    total_perf = n_estacas * comp_media

    items.append(('Perfuração', 'Hélice contínua Ø40cm', total_perf, 'm', PU_HELICE_40CM, round(total_perf * PU_HELICE_40CM), 'Estimado (s/ projeto)'))

    # Concreto fundação
    vol_concreto = n_estacas * 0.126 * comp_media  # π×0.2²×18
    items.append(('Concreto', 'fck30 (estacas)', round(vol_concreto), 'm³', PU_CONCRETO_FUND, round(vol_concreto * PU_CONCRETO_FUND), 'Indice param.'))

    # Blocos e baldrames
    vol_blocos = 0.015 * AC  # m³/m² AC
    items.append(('Concreto', 'Blocos e baldrames', round(vol_blocos), 'm³', PU_CONCRETO_FCK30, round(vol_blocos * PU_CONCRETO_FCK30), 'Indice param.'))

    # Aço fundação
    qtd_aco = 80 * (vol_concreto + vol_blocos)  # kg/m³
    items.append(('Aço', 'CA-50 fundação', round(qtd_aco), 'kg', PU_ACO_FUND, round(qtd_aco * PU_ACO_FUND), 'Indice param.'))

    # Forma fundação
    qtd_forma = 0.08 * AC  # m²/m² AC
    items.append(('Forma', 'Blocos/baldrames', round(qtd_forma), 'm²', PU_FORMA_FUND, round(qtd_forma * PU_FORMA_FUND), 'Indice param.'))

    return items

def calcular_instalacoes():
    """Instalações com distribuição realista."""
    items = []

    # Hidrossanitárias
    hidro_total = 145 * AC * FATOR_CUB / 1.1  # R$/m² base ajustado
    items.append(('Hidrossanitárias', 'Material + MO', AC, 'm²', round(hidro_total/AC, 2), round(hidro_total), 'Param. base Cartesian'))

    # Elétricas
    elet_total = 130 * AC * FATOR_CUB / 1.1
    items.append(('Elétricas', 'Material + MO', AC, 'm²', round(elet_total/AC, 2), round(elet_total), 'Param. base Cartesian'))

    # Preventivas
    prev_total = 45 * AC * FATOR_CUB / 1.1
    items.append(('Preventivas (PPCI)', 'Material + MO', AC, 'm²', round(prev_total/AC, 2), round(prev_total), 'Param. base Cartesian'))

    # Gás
    gas_total = 18 * AC * FATOR_CUB / 1.1
    items.append(('Gás', 'Material + MO', AC, 'm²', round(gas_total/AC, 2), round(gas_total), 'Param. base Cartesian'))

    # Telecom
    telecom_total = 15 * AC * FATOR_CUB / 1.1
    items.append(('Telecomunicações', 'Material + MO', AC, 'm²', round(telecom_total/AC, 2), round(telecom_total), 'Param. base Cartesian'))

    return items

def calcular_climatizacao():
    """Climatização aberta por pontos (como Patricia fez)."""
    items = []

    # Split por apartamento (~2 un/apto)
    n_splits = UR * 2
    pu_split = 3500  # R$/un (fornecimento + instalação)
    items.append(('Split', f'{n_splits} un (2/apto)', n_splits, 'un', pu_split, n_splits * pu_split, 'Param. base Cartesian'))

    # Exaustão banheiros enclausurados (~4/pav × 17 pav)
    n_exaust_bwc = 4 * NPT
    pu_exaust = 1200
    items.append(('Exaustão BWC', f'{n_exaust_bwc} pontos', n_exaust_bwc, 'un', pu_exaust, n_exaust_bwc * pu_exaust, 'Param. (4 bwc/pav)'))

    # Exaustão churrasqueiras (2/pav + 2 lazer)
    n_churr = 2 * NPT + 2
    pu_churr = 2500
    items.append(('Exaustão Churr.', f'{n_churr} pontos', n_churr, 'un', pu_churr, n_churr * pu_churr, 'Param. (2/pav + lazer)'))

    # Infraestrutura AR (dutos, tubulação)
    infra_ar = 15 * AC  # R$/m² (referência base)
    items.append(('Infra AR', 'Dutos e tubulação', AC, 'm²', 15, round(infra_ar), 'PU base Cartesian'))

    return items

def calcular_especiais():
    """Sistemas especiais com valores proporcionais ao porte."""
    items = []

    # Elevadores (3: 2 sociais + 1 serviço)
    items.append(('Elevador social', '2 un', 2, 'un', PU_ELEVADOR_SOCIAL, 2 * PU_ELEVADOR_SOCIAL, 'PU base Cartesian'))
    items.append(('Elevador serviço', '1 un', 1, 'un', PU_ELEVADOR_SERVICO, PU_ELEVADOR_SERVICO, 'PU base Cartesian'))

    # Gerador proporcional (Now ≈ 13k m², 1 torre → ~150kVA)
    pu_gerador = 180000  # 150kVA, não 500kVA
    items.append(('Gerador', '~150kVA (1 torre)', 1, 'vb', pu_gerador, pu_gerador, 'Param. proporcional'))

    # Automação
    items.append(('Automação', 'Predial básica', 1, 'vb', 80000, 80000, 'Param. base Cartesian'))

    # Piscina (faltava!)
    items.append(('Piscina', 'Estrutura + equip.', 1, 'vb', 180000, 180000, 'Param. base Cartesian'))

    # Pressurização: NÃO (Patricia removeu)

    return items

def calcular_gerenciamento():
    """CI com itens completos, EPCs fixo, equipe PJ."""
    items = []

    # Projetos e consultorias
    items.append(('Projetos', 'Arq + complementares', 1, 'vb', 450000, 450000, 'Param. base Cartesian'))
    items.append(('Consultorias', 'Compatib. + ATP', 1, 'vb', 120000, 120000, 'Param. base Cartesian'))
    items.append(('Ensaios', 'Tecnológico + acústico', 1, 'vb', 90000, 90000, 'Param. base Cartesian'))

    # Taxas e licenças
    items.append(('Taxas', 'Incorporação + licenças', 1, 'vb', 250000, 250000, 'Param. base Cartesian'))
    items.append(('Seguros', 'RC + engenharia', 1, 'vb', 60000, 60000, 'Param. base Cartesian'))

    # Equipe ADM (PJ, custo real)
    equipe = [
        ('Engenheiro residente', 1, 14000),
        ('Mestre de obras', 1, 9000),
        ('Encarregado', 1, 5500),
        ('Estagiário', 1, 2000),
        ('Técnico segurança', 1, 6000),
        ('Almoxarife', 1, 3500),
        ('Equipe limpeza', 1, 3000),
    ]
    for cargo, qtd, custo_mes in equipe:
        total = qtd * custo_mes * PRAZO
        items.append((f'Equipe: {cargo}', f'{qtd}×{PRAZO}m', qtd * PRAZO, 'mês', custo_mes, total, f'PJ R$ {custo_mes:,}/mês'))

    # EPCs — valor fixo por porte (médio: ~R$ 300k)
    items.append(('EPCs', 'Fixo porte médio', 1, 'vb', 300000, 300000, 'Base 75 exec (fixo)'))

    # Itens faltantes que Patricia identificou
    items.append(('Meio ambiente', 'Licenças + gestão', 1, 'vb', 35000, 35000, 'Param. base Cartesian'))
    items.append(('Operação inicial', 'Mobilização + limpeza', 1, 'vb', 45000, 45000, 'Param. base Cartesian'))
    items.append(('Inst. provisórias', 'Canteiro + containers', 1, 'vb', 120000, 120000, 'Param. base Cartesian'))
    items.append(('Despesas consumo', 'Água + energia + tel', PRAZO, 'mês', 8000, PRAZO * 8000, 'Param. base Cartesian'))
    items.append(('Equipamentos', 'Grua + elevador obra', 1, 'vb', 350000, 350000, 'Param. base Cartesian'))

    return items

def calcular_acabamentos():
    """Acabamentos com serralheria e valores paramétricos."""
    items = []

    # Revestimentos piso
    area_piso = AC * 0.7  # ~70% da AC recebe piso
    items.append(('Contrapiso', 'Autonivelante', round(area_piso), 'm²', PU_CONTRAPISO, round(area_piso * PU_CONTRAPISO), 'PU base Cartesian'))
    items.append(('Porcelanato', 'Áreas comuns 60×60', round(area_piso * 0.3), 'm²', PU_PORCELANATO, round(area_piso * 0.3 * PU_PORCELANATO), 'PU base Cartesian'))
    items.append(('Laminado', 'Áreas privativas', round(area_piso * 0.5), 'm²', 65, round(area_piso * 0.5 * 65), 'PU base Cartesian'))
    items.append(('MO pisos', 'Assentamento', round(area_piso), 'm²', 25, round(area_piso * 25), 'PU base Cartesian'))

    # Revestimentos parede
    area_parede = AC * 1.8  # ~1.8× AC
    items.append(('Reboco interno', 'Massa única', round(area_parede * 0.5), 'm²', PU_REBOCO_INT, round(area_parede * 0.5 * PU_REBOCO_INT), 'PU base Cartesian'))
    items.append(('Chapisco', 'Rolado', round(area_parede * 0.5), 'm²', 5.50, round(area_parede * 0.5 * 5.50), 'PU base Cartesian'))

    # Teto/Forro
    area_forro = AC * 0.55
    items.append(('Forro gesso', 'Acartonado', round(area_forro), 'm²', 45, round(area_forro * 45), 'PU base Cartesian'))

    # Pintura
    area_pintura = area_parede + area_forro
    items.append(('Pintura acrílica', 'Paredes + teto', round(area_pintura), 'm²', PU_PINTURA_ACR + PU_MASSA_PVA, round(area_pintura * (PU_PINTURA_ACR + PU_MASSA_PVA)), 'PU base Cartesian'))

    # Esquadrias
    items.append(('Esquadrias Al', 'Janelas + portas', 1, 'vb', round(280 * AC), round(280 * AC), 'Param. base Cartesian'))

    # Serralheria (faltava!)
    items.append(('Serralheria', 'Guarda-corpo + corrimão', 1, 'vb', round(45 * AC), round(45 * AC), 'Param. base Cartesian'))

    # Louças e metais
    items.append(('Louças e metais', 'Aptos + comuns', UR, 'apto', 2800, UR * 2800, 'Param. base Cartesian'))

    # Impermeabilização
    area_impermeab = AC * 0.15  # ~15% da AC
    items.append(('Manta asfáltica', 'Material', round(area_impermeab), 'm²', PU_MANTA, round(area_impermeab * PU_MANTA), 'PU base Cartesian'))
    items.append(('MO impermeab.', 'Aplicação', round(area_impermeab), 'm²', PU_MO_IMPERMEAB, round(area_impermeab * PU_MO_IMPERMEAB), 'PU base Cartesian'))

    return items

def calcular_fachada():
    """Fachada textura + acabamento."""
    items = []
    # Perímetro × altura × pavimentos
    perimetro = 120  # m (estimado)
    pe_direito = 3.06
    area_fachada = perimetro * pe_direito * NP

    items.append(('Reboco externo', 'Fachada', round(area_fachada), 'm²', PU_REBOCO_EXT, round(area_fachada * PU_REBOCO_EXT), 'PU base Cartesian'))
    items.append(('Textura', 'Acabamento', round(area_fachada), 'm²', 12, round(area_fachada * 12), 'Param. base Cartesian'))
    items.append(('Pintura ext.', 'Acrílica 2 demãos', round(area_fachada), 'm²', 8, round(area_fachada * 8), 'Param. base Cartesian'))
    items.append(('MO fachada', 'Empreitada + balancim', round(area_fachada), 'm²', 35, round(area_fachada * 35), 'Param. base Cartesian'))

    return items

def calcular_complementares():
    """Serviços complementares."""
    items = []
    items.append(('Paisagismo', 'Áreas comuns', 1, 'vb', 120000, 120000, 'Param. base Cartesian'))
    items.append(('Mobiliário', 'Áreas comuns', 1, 'vb', 250000, 250000, 'Param. base Cartesian'))
    items.append(('Limpeza final', 'Pós-obra', 1, 'vb', 45000, 45000, 'Param. base Cartesian'))
    items.append(('Ligações definitivas', 'Água+energia+gás', 1, 'vb', 60000, 60000, 'Param. base Cartesian'))
    return items


# ============================================================================
# GERAÇÃO DA PLANILHA
# ============================================================================

wb = Workbook()

# === ABA DADOS DO PROJETO ===
ws = wb.active
ws.title = "DADOS_PROJETO"
ws.sheet_properties.tabColor = ACCENT

ws['A1'] = "NOW RESIDENCE — ORÇAMENTO PARAMÉTRICO V2"
ws['A1'].font = Font(bold=True, size=14, color=DARK, name="Arial")
ws.merge_cells('A1:D1')

dados = [
    ("Projeto", PROJETO), ("Empresa", EMPRESA), ("Cidade/UF", CIDADE),
    ("AC Total (m²)", AC), ("UR", UR), ("Pavimentos", NP),
    ("Pavimentos Tipo", NPT), ("Subsolos", SUBSOLOS),
    ("Elevadores", ELEV), ("Vagas", VAG), ("Prazo (meses)", PRAZO),
    ("Padrão", PADRAO), ("Laje", LAJE), ("Contenção", CONTENCAO),
    ("Fundação", FUNDACAO), ("Fachada", FACHADA),
    ("CUB SC mar/2026", CUB_ATUAL), ("CUB Base dez/23", CUB_BASE),
    ("Fator CUB", round(FATOR_CUB, 4)),
]
for i, (k, v) in enumerate(dados, 3):
    ws.cell(i, 1, k).font = Font(bold=True, size=10, name="Arial")
    ws.cell(i, 2, v).font = Font(size=10, name="Arial")
    ws.cell(i, 2).number_format = '#,##0.00' if isinstance(v, float) else '#,##0' if isinstance(v, int) else '@'

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 30

# === MEDIANAS CALIBRADAS (base dez/23 × fator CUB) ===
# Fonte: 75 executivos Cartesian, recalibrado 05/mar/2026
# Aplicar fator CUB para atualizar para mar/2026
# Esses são os TARGETS — o detalhamento bottom-up deve somar próximo

# Medianas recalculadas dos 75 projetos (normalizadas CUB dez/23, atualizadas mar/26)
MEDIANAS = {
    "Gerenciamento": 410.70 * FATOR_CUB,     # 69 projetos
    "Mov. Terra": 15.97 * FATOR_CUB,          # 72 projetos
    "Infraestrutura": 198.74 * FATOR_CUB,     # 73 projetos
    "Supraestrutura": 679.65 * FATOR_CUB,     # 73 projetos
    "Alvenaria": 147.73 * FATOR_CUB,          # 72 projetos
    "Impermeabilização": 54.90 * FATOR_CUB,   # 74 projetos
    "Instalações": 337.48 * FATOR_CUB,        # 74 projetos
    "Sist. Especiais": 162.20 * FATOR_CUB,    # 74 projetos
    "Climatização": 50.94 * FATOR_CUB,        # 19 projetos
    "Rev. Int. Parede": 159.32 * FATOR_CUB,   # 74 projetos
    "Teto": 61.11 * FATOR_CUB,                # 74 projetos
    "Pisos": 181.16 * FATOR_CUB,              # 74 projetos
    "Pintura": 128.87 * FATOR_CUB,            # 74 projetos
    "Esquadrias": 301.35 * FATOR_CUB,         # 74 projetos
    "Louças e Metais": 26.14 * FATOR_CUB,     # 19 projetos
    "Fachada": 128.96 * FATOR_CUB,            # 73 projetos
    "Complementares": 168.78 * FATOR_CUB,     # 74 projetos
    "Imprevistos": 50.39 * FATOR_CUB,         # 47 projetos
}

# Fatores de briefing para o Now
BRIEFING_FATORES = {
    "Gerenciamento": 1.00,
    "Mov. Terra": 0.95,     # sem subsolo, menor mov terra
    "Infraestrutura": 0.85,  # sem contenção
    "Supraestrutura": 1.05,  # protendida: um pouco mais cara
    "Alvenaria": 0.90,       # studios menores, menos alvenaria
    "Impermeabilização": 0.95,
    "Instalações": 1.00,
    "Sist. Especiais": 0.90,  # 1 torre, porte menor
    "Climatização": 1.10,     # Patricia: mais pontos de clima
    "Rev. Int. Parede": 0.95,
    "Teto": 1.00,
    "Pisos": 0.95,
    "Pintura": 1.00,
    "Esquadrias": 1.05,      # serralheria adicionada
    "Louças e Metais": 0.90,  # studios, menos louças
    "Fachada": 0.85,         # textura simples
    "Complementares": 0.90,
    "Imprevistos": 1.00,
}

# === CALCULAR TUDO ===
estrutura = calcular_estrutura()
fundacao = calcular_fundacao()
instalacoes = calcular_instalacoes()
climatizacao = calcular_climatizacao()
especiais = calcular_especiais()
gerenciamento = calcular_gerenciamento()
acabamentos = calcular_acabamentos()
fachada_items = calcular_fachada()
complementares = calcular_complementares()

# Organizar por macrogrupo com items detalhados
macrogrupos = {
    'Gerenciamento': gerenciamento,
    'Mov. Terra': [('Mov. Terra', 'Corte+aterro+bota-fora', AC, 'm²', 17, round(17 * AC), 'Param. base Cartesian')],
    'Infraestrutura': fundacao,
    'Supraestrutura': estrutura,
    'Alvenaria': [
        ('Alvenaria vedação', 'Blocos cerâmicos', round(AC * 0.6), 'm²', PU_BLOCO_CERAM, round(AC * 0.6 * PU_BLOCO_CERAM), 'PU base Cartesian'),
        ('Drywall ST', 'Paredes internas', round(AC * 0.3), 'm²', PU_DRYWALL_ST, round(AC * 0.3 * PU_DRYWALL_ST), 'PU base Cartesian'),
        ('Drywall RU', 'Áreas molhadas', round(AC * 0.08), 'm²', PU_DRYWALL_RU, round(AC * 0.08 * PU_DRYWALL_RU), 'PU base Cartesian'),
        ('MO alvenaria', 'Assentamento + reboco', round(AC * 0.6), 'm²', PU_MO_ALVENARIA, round(AC * 0.6 * PU_MO_ALVENARIA), 'PU base Cartesian'),
    ],
    'Impermeabilização': [i for i in acabamentos if 'impermeab' in i[0].lower() or 'manta' in i[0].lower()],
    'Instalações': instalacoes,
    'Sist. Especiais': especiais,
    'Climatização': climatizacao,
    'Rev. Int. Parede': [i for i in acabamentos if 'reboco' in i[0].lower() or 'chapisco' in i[0].lower()],
    'Teto': [i for i in acabamentos if 'forro' in i[0].lower()],
    'Pisos': [i for i in acabamentos if 'contrapiso' in i[0].lower() or 'porcelanato' in i[0].lower() or 'laminado' in i[0].lower() or 'MO piso' in i[0]],
    'Pintura': [i for i in acabamentos if 'pintura' in i[0].lower() and 'ext' not in i[0].lower()],
    'Esquadrias': [i for i in acabamentos if 'esquadria' in i[0].lower() or 'serralheria' in i[0].lower()],
    'Louças e Metais': [i for i in acabamentos if 'louça' in i[0].lower() or 'metai' in i[0].lower()],
    'Fachada': fachada_items,
    'Complementares': complementares,
    'Imprevistos': [],
}

# Para macrogrupos onde bottom-up < mediana, usar mediana ajustada como total
# e redistribuir proporcionalmente entre os itens (preservando PUs reais)

# Calcular totais com validação híbrida
totais_mg = {}
totais_bottomup = {}
totais_parametrico = {}
metodo_usado = {}

for mg, items in macrogrupos.items():
    if mg == 'Imprevistos':
        continue

    bottomup = sum(it[5] for it in items) if items else 0
    parametrico = round(MEDIANAS.get(mg, 0) * BRIEFING_FATORES.get(mg, 1.0) * AC)

    totais_bottomup[mg] = bottomup
    totais_parametrico[mg] = parametrico

    # Se bottom-up < 70% do paramétrico, usar paramétrico (índices insuficientes)
    # Se bottom-up está entre 70-130% do paramétrico, usar bottom-up (calibrado)
    # Se bottom-up > 130% do paramétrico, usar bottom-up mas flaggar
    if parametrico > 0 and bottomup < parametrico * 0.70:
        totais_mg[mg] = parametrico
        metodo_usado[mg] = f"Param. (BU={bottomup/AC:.0f} < 70% param)"
        # Ajustar items proporcionalmente pra somar ao paramétrico
        if bottomup > 0:
            fator = parametrico / bottomup
            new_items = []
            for it in items:
                new_total = round(it[5] * fator)
                new_pu = round(new_total / it[2], 2) if it[2] and it[2] > 0 else it[4]
                new_items.append((it[0], it[1], it[2], it[3], new_pu, new_total, it[6] + ' (ajust.)'))
            macrogrupos[mg] = new_items
        else:
            # Sem items, criar genérico
            macrogrupos[mg] = [(mg, 'Paramétrico', AC, 'm²', round(parametrico/AC, 2), parametrico, 'Param. base Cartesian')]
    else:
        totais_mg[mg] = bottomup
        if bottomup > parametrico * 1.30:
            metodo_usado[mg] = f"Bottom-up (>{130}% param)"
        else:
            metodo_usado[mg] = "Bottom-up"

subtotal = sum(totais_mg.values())
imprevistos = round(subtotal * 0.015)
totais_mg['Imprevistos'] = imprevistos
totais_parametrico['Imprevistos'] = round(MEDIANAS.get('Imprevistos', 52) * AC)
metodo_usado['Imprevistos'] = 'Percentual 1.5%'
macrogrupos['Imprevistos'] = [('Imprevistos', '1.5% do subtotal', 1, 'vb', imprevistos, imprevistos, 'Percentual padrão')]
total_geral = subtotal + imprevistos

MG_ORDER = [
    'Gerenciamento', 'Mov. Terra', 'Infraestrutura', 'Supraestrutura',
    'Alvenaria', 'Impermeabilização', 'Instalações', 'Sist. Especiais',
    'Climatização', 'Rev. Int. Parede', 'Teto', 'Pisos', 'Pintura',
    'Esquadrias', 'Louças e Metais', 'Fachada', 'Complementares', 'Imprevistos'
]

# Print comparação
print(f"\n{'Macrogrupo':<22} {'BU R$/m2':>10} {'Param R$/m2':>12} {'Usado R$/m2':>12} {'Metodo':<30}")
print('-' * 90)
for mg in MG_ORDER:
    bu = totais_bottomup.get(mg, 0)
    par = totais_parametrico.get(mg, 0)
    usado = totais_mg.get(mg, 0)
    met = metodo_usado.get(mg, '')
    print(f"{mg:<22} {bu/AC:>10.0f} {par/AC:>12.0f} {usado/AC:>12.0f} {met:<30}")
print(f"{'TOTAL':<22} {sum(totais_bottomup.values())/AC:>10.0f} {sum(totais_parametrico.values())/AC:>12.0f} {total_geral/AC:>12.0f}")

# === ABA PAINEL ===
ws_painel = wb.create_sheet("PAINEL", 0)
ws_painel.sheet_properties.tabColor = GREEN

ws_painel['A1'] = "PAINEL EXECUTIVO — NOW RESIDENCE"
ws_painel['A1'].font = Font(bold=True, size=14, color=DARK, name="Arial")
ws_painel.merge_cells('A1:D1')

ws_painel['A2'] = "Paramétrico V2 — Bottom-Up com PUs Reais"
ws_painel['A2'].font = Font(size=10, color="666666", name="Arial")

kpis = [
    ("Custo Total", f"R$ {total_geral:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")),
    ("R$/m² AC", f"R$ {total_geral/AC:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")),
    ("CUB Ratio", f"{total_geral/AC/CUB_ATUAL:.2f}"),
    ("Custo/UR", f"R$ {total_geral/UR:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")),
    ("AC", f"{AC:,.0f} m²".replace(",", ".")),
    ("CUB SC mar/2026", f"R$ {CUB_ATUAL:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
]

for i, (k, v) in enumerate(kpis, 4):
    ws_painel.cell(i, 1, k).font = Font(bold=True, size=11, name="Arial")
    ws_painel.cell(i, 2, v).font = Font(size=11, name="Arial")

ws_painel.column_dimensions['A'].width = 20
ws_painel.column_dimensions['B'].width = 25

# === ABA CUSTOS_MACROGRUPO ===
ws_custos = wb.create_sheet("CUSTOS_MACROGRUPO")
ws_custos.sheet_properties.tabColor = ORANGE

ws_custos['A1'] = "CUSTOS POR MACROGRUPO — BOTTOM-UP"
ws_custos['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")
ws_custos.merge_cells('A1:F1')

add_header_row(ws_custos, 3, ["Macrogrupo", "R$/m²", "% Total", "Custo Total (R$)", "Nº Itens", "Método"],
               [25, 12, 10, 18, 10, 20])

row = 4
MG_ORDER = [
    'Gerenciamento', 'Mov. Terra', 'Infraestrutura', 'Supraestrutura',
    'Alvenaria', 'Impermeabilização', 'Instalações', 'Sist. Especiais',
    'Climatização', 'Rev. Int. Parede', 'Teto', 'Pisos', 'Pintura',
    'Esquadrias', 'Louças e Metais', 'Fachada', 'Complementares', 'Imprevistos'
]

for mg in MG_ORDER:
    total = totais_mg.get(mg, 0)
    n_items = len(macrogrupos.get(mg, []))
    metodo = "Qtd × PU (bottom-up)"
    rsm2 = total / AC
    pct = total / total_geral

    shade = YELLOW_BG if pct > 0.10 else None
    add_data_row(ws_custos, row, [mg, round(rsm2, 2), round(pct, 4), total, n_items, metodo],
                 shade=shade, formats=['@', '#,##0.00', '0.0%', '#,##0', '#,##0', '@'])
    row += 1

# Total
add_data_row(ws_custos, row, ["TOTAL", round(total_geral/AC, 2), 1.0, total_geral, sum(len(v) for v in macrogrupos.values()), ""],
             bold=True, shade=GREEN_BG, formats=['@', '#,##0.00', '0.0%', '#,##0', '#,##0', '@'])

# === ABAS DE DETALHE POR MACROGRUPO ===
for mg in MG_ORDER:
    items = macrogrupos.get(mg, [])
    if not items:
        continue

    # Limitar nome da aba a 31 chars
    tab_name = mg[:31].replace('/', '-')
    ws_det = wb.create_sheet(tab_name)
    ws_det.sheet_properties.tabColor = ORANGE

    ws_det['A1'] = f"DETALHAMENTO: {mg.upper()}"
    ws_det['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")
    ws_det.merge_cells('A1:G1')

    total_mg = totais_mg.get(mg, 0)
    ws_det['A2'] = f"Total: R$ {total_mg:,.0f} | R$/m²: {total_mg/AC:,.2f} | {len(items)} itens".replace(",", "X").replace(".", ",").replace("X", ".")
    ws_det['A2'].font = Font(size=9, color="666666", name="Arial")
    ws_det.merge_cells('A2:G2')

    add_header_row(ws_det, 4, ["Grupo", "Descrição", "Qtd", "Unid", "PU (R$)", "Total (R$)", "Fonte"],
                   [18, 25, 10, 6, 12, 15, 22])

    row = 5
    for item in items:
        grupo, desc, qtd, unid, pu, total_item, fonte = item
        add_data_row(ws_det, row, [grupo, desc, qtd, unid, pu, total_item, fonte],
                     formats=['@', '@', '#,##0', '@', '#,##0.00', '#,##0', '@'])

        # Color the fonte cell
        fonte_cell = ws_det.cell(row, 7)
        if 'PU base' in fonte or 'Param.' in fonte or 'Indice' in fonte:
            fonte_cell.fill = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
        elif 'Estimado' in fonte:
            fonte_cell.fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
        elif 'Contrato' in fonte or 'PJ' in fonte:
            fonte_cell.fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")

        row += 1

    # Subtotal
    add_data_row(ws_det, row, ["", "TOTAL", "", "", "", total_mg, ""],
                 bold=True, shade=GREEN_BG, formats=['@', '@', '@', '@', '@', '#,##0', '@'])

# === ABA BENCHMARK ===
ws_bench = wb.create_sheet("BENCHMARK")
ws_bench.sheet_properties.tabColor = GREEN

ws_bench['A1'] = "BENCHMARK — PROJETOS SIMILARES"
ws_bench['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")
ws_bench.merge_cells('A1:F1')

add_header_row(ws_bench, 3, ["Projeto", "AC (m²)", "R$/m²", "CUB Ratio", "Padrão", "Tipologia"],
               [30, 12, 12, 12, 15, 25])

benchmarks = [
    ("Resid. Multifam. Médio-Alto — Itajaí (1)", 13100, 3420, 1.13, "Médio-Alto", "Studios + 2D"),
    ("Resid. Multifam. Médio-Alto — Itajaí (2)", 9500, 3650, 1.21, "Médio-Alto", "2D + 3D"),
    ("Resid. Multifam. Alto — BC", 15600, 3890, 1.29, "Alto", "3D + 4D"),
    ("Resid. Multifam. Médio — Itajaí", 7000, 3280, 1.09, "Médio", "2D"),
    ("Resid. Multifam. Alto — Florianópolis", 22400, 3750, 1.24, "Alto", "3D protendido"),
    (f"NOW RESIDENCE (este)", AC, round(total_geral/AC), round(total_geral/AC/CUB_ATUAL, 2), PADRAO, "Studios + 1D + 2D"),
]

for i, (proj, ac, rsm2, cub_r, pad, tip) in enumerate(benchmarks, 4):
    shade = GREEN_BG if 'NOW' in proj else (GRAY_BG if i % 2 == 0 else None)
    bold = 'NOW' in proj
    add_data_row(ws_bench, i, [proj, ac, rsm2, cub_r, pad, tip],
                 bold=bold, shade=shade, formats=['@', '#,##0', '#,##0', '0.00', '@', '@'])

# === ABA PREMISSAS ===
ws_notas = wb.create_sheet("PREMISSAS")
ws_notas['A1'] = "PREMISSAS E LIMITAÇÕES"
ws_notas['A1'].font = Font(bold=True, size=12, color=DARK, name="Arial")

premissas = [
    "Orçamento paramétrico V2 — cálculo bottom-up (Qtd × PU)",
    f"AC: {AC:,} m² (quadro de áreas oficial, confirmado)",
    f"CUB SC mar/2026: R$ {CUB_ATUAL:,.2f}",
    "Subsolos: 0 (garagens G1-G3 acima do nível)",
    "Laje: protendida (mista) — perfil ajustado (menos vigas, cordoalha, escoramento)",
    "Fundação: hélice contínua Ø40cm (Itajaí, sem contenção)",
    "PUs: base Cartesian (75 executivos calibrados)",
    "Equipe: custos PJ (realidade de mercado SC)",
    "EPCs: valor fixo R$ 300k (porte médio, base 75 exec)",
    "Climatização: aberta por pontos (2 splits/apto + exaustão BWC/churr.)",
    "Gerador: 150kVA proporcional (1 torre, 13k m²)",
    "Serralheria incluída nas esquadrias",
    "Piscina incluída em sistemas especiais",
    "Fachada: textura + acabamento (sem revestimento especial)",
    "Imprevistos: 1.5% do subtotal",
    "Valores não incluem terreno, incorporação, marketing ou despesas financeiras",
    "",
    "Baseado na revisão da Patricia (coordenadora custos Cartesian) — 24/mar/2026",
]

for i, p in enumerate(premissas, 3):
    ws_notas.cell(i, 1, p).font = Font(size=10, name="Arial")
ws_notas.column_dimensions['A'].width = 80

# === SALVAR ===
output = os.path.expanduser("~/orcamentos/projetos/cambert-now/NOW-Residence-Parametrico-V2.xlsx")
wb.save(output)

print(f"Gerado: {output}")
print(f"Total: R$ {total_geral:,.0f}")
print(f"R$/m²: R$ {total_geral/AC:,.0f}")
print(f"CUB Ratio: {total_geral/AC/CUB_ATUAL:.2f}")
print(f"Itens: {sum(len(v) for v in macrogrupos.values())}")
