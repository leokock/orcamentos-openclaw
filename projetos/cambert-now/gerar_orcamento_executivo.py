#!/usr/bin/env python3.11
"""
Gerador de Orçamento Executivo - NOW Residence
Cartesian Engenharia - Mar/2026
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from datetime import datetime
import os

# Configurações globais
PROJETO = "NOW Residence"
DATA_BASE = "mar/2026"
AC_TOTAL = 13054.0
UNIDADES = 136
CUB_SC = 3150.00
CUSTO_TOTAL_PARAMETRICO = 58359617.47

# Cores
COR_HEADER = "2C3E50"
COR_TEXTO_BRANCO = "FFFFFF"
COR_ZEBRA = "ECF0F1"

# Estilos de borda
BORDA_FINA = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def formatar_moeda(valor):
    """Retorna fórmula de formato moeda"""
    return valor

def criar_aba_resumo(wb):
    """Aba 1: Resumo Executivo"""
    ws = wb.active
    ws.title = "RESUMO EXECUTIVO"
    
    # Dados do empreendimento
    dados = [
        ["ORÇAMENTO EXECUTIVO - NOW RESIDENCE", ""],
        ["Data-base:", DATA_BASE],
        ["", ""],
        ["DADOS DO EMPREENDIMENTO", ""],
        ["Endereço:", "Rua Luiz Berlim, 123, Centro, Itajaí/SC"],
        ["Proprietário:", "Belli Empreendimentos Ltda."],
        ["CNPJ:", "76.323.427/0001-92"],
        ["", ""],
        ["ÁREAS E UNIDADES", ""],
        ["Área Construída Total:", f"{AC_TOTAL:,.2f} m²"],
        ["Unidades Residenciais:", f"{UNIDADES} aptos"],
        ["Unidades Comerciais:", "2 salas"],
        ["Pavimentos:", "24 (3 garagens + térreo + lazer + 17 tipos + ático + cobertura)"],
        ["Elevadores:", "3"],
        ["Vagas de Garagem:", "110"],
        ["", ""],
        ["INDICADORES DE CUSTO", "VALOR"],
        ["Custo Total Estimado:", "=B19"],
        ["Custo por m²:", "=B18/B10"],
        ["Custo por Unidade:", "=B18/B11"],
        ["CUB-SC (mar/2026):", f"R$ {CUB_SC:,.2f}/m²"],
        ["CUB Ratio:", "=B19/B21"],
    ]
    
    for i, linha in enumerate(dados, start=1):
        ws[f"A{i}"] = linha[0]
        if linha[1]:
            ws[f"B{i}"] = linha[1]
    
    # Formatação header principal
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=14, color=COR_TEXTO_BRANCO)
    ws['A1'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers de seção
    for row in [4, 9, 17]:
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].font = Font(bold=True, size=12, color=COR_TEXTO_BRANCO)
        ws[f'A{row}'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Resumo por macrogrupo (linha 24 em diante)
    ws['A24'] = "CUSTO POR MACROGRUPO"
    ws.merge_cells('A24:E24')
    ws['A24'].font = Font(bold=True, size=12, color=COR_TEXTO_BRANCO)
    ws['A24'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Headers da tabela
    headers = ["#", "Macrogrupo", "Valor Total (R$)", "% do Total", "R$/m²"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=25, column=col, value=header)
        cell.font = Font(bold=True, color=COR_TEXTO_BRANCO)
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados dos macrogrupos (dados do paramétrico)
    macrogrupos = [
        ["1", "Supraestrutura", 10660385.93, "=C26/$C$44", "=C26/$B$10"],
        ["2", "Gerenciamento", 6135118.92, "=C27/$C$44", "=C27/$B$10"],
        ["3", "Infraestrutura", 5985244.38, "=C28/$C$44", "=C28/$B$10"],
        ["4", "Instalações", 5458587.54, "=C29/$C$44", "=C29/$B$10"],
        ["5", "Complementares", 4730035.31, "=C30/$C$44", "=C30/$B$10"],
        ["6", "Esquadrias", 4537309.32, "=C31/$C$44", "=C31/$B$10"],
        ["7", "Alvenaria", 2876618.60, "=C32/$C$44", "=C32/$B$10"],
        ["8", "Sistemas Especiais", 2796003.62, "=C33/$C$44", "=C33/$B$10"],
        ["9", "Pisos", 2714448.76, "=C34/$C$44", "=C34/$B$10"],
        ["10", "Fachada", 2504292.41, "=C35/$C$44", "=C35/$B$10"],
        ["11", "Rev. Int. Parede", 2403763.56, "=C36/$C$44", "=C36/$B$10"],
        ["12", "Pintura", 1939824.40, "=C37/$C$44", "=C37/$B$10"],
        ["13", "Impermeabilização", 1596047.31, "=C38/$C$44", "=C38/$B$10"],
        ["14", "Movimentação de Terra", 1200968.00, "=C39/$C$44", "=C39/$B$10"],
        ["15", "Teto", 916912.96, "=C40/$C$44", "=C40/$B$10"],
        ["16", "Climatização", 760917.66, "=C41/$C$44", "=C41/$B$10"],
        ["17", "Imprevistos", 752693.64, "=C42/$C$44", "=C42/$B$10"],
        ["18", "Louças e Metais", 390445.14, "=C43/$C$44", "=C43/$B$10"],
    ]
    
    for i, dados_macro in enumerate(macrogrupos, start=26):
        for col, valor in enumerate(dados_macro, start=1):
            ws.cell(row=i, column=col, value=valor)
    
    # Total
    ws['A44'] = "TOTAL"
    ws['A44'].font = Font(bold=True)
    ws['C44'] = "=SUM(C26:C43)"
    ws['D44'] = "=SUM(D26:D43)"
    ws['E44'] = "=SUM(E26:E43)"
    
    # Link com orçamento detalhado
    ws['B18'] = "=SUBTOTAL(9,'ORÇAMENTO DETALHADO'!H:H)"
    ws['B19'] = "='ORÇAMENTO DETALHADO'!H$2000"  # Placeholder - será ajustado
    
    # Formatos numéricos
    for row in range(26, 45):
        ws[f"C{row}"].number_format = 'R$ #,##0.00'
        ws[f"D{row}"].number_format = '0.0%'
        ws[f"E{row}"].number_format = 'R$ #,##0.00'
    
    ws['B18'].number_format = 'R$ #,##0.00'
    ws['B19'].number_format = 'R$ #,##0.00'
    ws['B20'].number_format = 'R$ #,##0.00'
    ws['B22'].number_format = '0.00'
    
    # Zebra striping
    for row in range(26, 44):
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=COR_ZEBRA, end_color=COR_ZEBRA, fill_type="solid")
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # Congelar painéis
    ws.freeze_panes = 'A25'
    
    # Adicionar gráfico (pizza)
    try:
        chart = PieChart()
        labels = Reference(ws, min_col=2, min_row=26, max_row=43)
        data = Reference(ws, min_col=3, min_row=25, max_row=43)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.title = "Distribuição de Custos por Macrogrupo"
        chart.width = 15
        chart.height = 12
        ws.add_chart(chart, "G4")
    except:
        pass  # Caso openpyxl não suporte charts no ambiente
    
    return ws

def criar_aba_detalhado(wb):
    """Aba 2: Orçamento Detalhado"""
    ws = wb.create_sheet("ORÇAMENTO DETALHADO")
    
    # Headers
    headers = ["Código", "Descrição", "Unidade", "Quantidade", "Preço Unit. (R$)", "Total Material (R$)", "Total MO (R$)", "Total (R$)", "% do Total"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color=COR_TEXTO_BRANCO)
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDA_FINA
    
    # Estrutura hierárquica com dados realistas
    linha_atual = 2
    total_geral_row = None
    
    # 18 Macrogrupos com itens detalhados
    orcamento = [
        # 1. MOVIMENTAÇÃO DE TERRA
        {
            "codigo": "01", "descricao": "MOVIMENTAÇÃO DE TERRA", "nivel": 1,
            "subitens": [
                {"codigo": "01.01", "descricao": "Serviços Preliminares", "nivel": 2, "subitens": [
                    {"codigo": "01.01.001", "descricao": "Locação da obra (topografia)", "un": "m²", "qtd": 13054.0, "pu": 3.50},
                    {"codigo": "01.01.002", "descricao": "Placa de obra em aço galvanizado", "un": "un", "qtd": 1.0, "pu": 2500.00},
                    {"codigo": "01.01.003", "descricao": "Ligações provisórias (água, luz, esgoto)", "un": "vb", "qtd": 1.0, "pu": 8500.00},
                ]},
                {"codigo": "01.02", "descricao": "Demolição e Remoção", "nivel": 2, "subitens": [
                    {"codigo": "01.02.001", "descricao": "Demolição de edificação existente", "un": "m²", "qtd": 450.0, "pu": 85.00},
                    {"codigo": "01.02.002", "descricao": "Remoção de entulho (caçamba)", "un": "m³", "qtd": 180.0, "pu": 120.00},
                ]},
                {"codigo": "01.03", "descricao": "Escavação", "nivel": 2, "subitens": [
                    {"codigo": "01.03.001", "descricao": "Escavação mecânica (3 subsolos - ~10m profundidade)", "un": "m³", "qtd": 8500.0, "pu": 42.00},
                    {"codigo": "01.03.002", "descricao": "Escavação manual (complementar)", "un": "m³", "qtd": 320.0, "pu": 75.00},
                    {"codigo": "01.03.003", "descricao": "Transporte de terra excedente (bota-fora)", "un": "m³", "qtd": 7200.0, "pu": 35.00},
                ]},
            ]
        },
        # 2. INFRAESTRUTURA (FUNDAÇÕES E CONTENÇÃO)
        {
            "codigo": "02", "descricao": "INFRAESTRUTURA", "nivel": 1,
            "subitens": [
                {"codigo": "02.01", "descricao": "Contenção", "nivel": 2, "subitens": [
                    {"codigo": "02.01.001", "descricao": "Cortina de estacas (hélice contínua Ø600mm)", "un": "m", "qtd": 1200.0, "pu": 850.00},
                    {"codigo": "02.01.002", "descricao": "Tirantes (Q=50tf)", "un": "un", "qtd": 180.0, "pu": 3200.00},
                    {"codigo": "02.01.003", "descricao": "Viga de coroamento (cortina)", "un": "m", "qtd": 280.0, "pu": 450.00},
                ]},
                {"codigo": "02.02", "descricao": "Fundação Profunda", "nivel": 2, "subitens": [
                    {"codigo": "02.02.001", "descricao": "Estaca Franki Ø520mm (prof. média 18m)", "un": "m", "qtd": 2800.0, "pu": 320.00},
                    {"codigo": "02.02.002", "descricao": "Bloco de coroamento (concreto 30MPa)", "un": "m³", "qtd": 380.0, "pu": 950.00},
                    {"codigo": "02.02.003", "descricao": "Baldrame (concreto 25MPa)", "un": "m³", "qtd": 120.0, "pu": 680.00},
                ]},
                {"codigo": "02.03", "descricao": "Radier e Laje de Fundo", "nivel": 2, "subitens": [
                    {"codigo": "02.03.001", "descricao": "Lastro de concreto magro e=5cm", "un": "m²", "qtd": 2800.0, "pu": 28.00},
                    {"codigo": "02.03.002", "descricao": "Laje de fundo e=25cm (concreto 30MPa)", "un": "m³", "qtd": 700.0, "pu": 850.00},
                    {"codigo": "02.03.003", "descricao": "Impermeabilização de laje de fundo (manta 4mm)", "un": "m²", "qtd": 2800.0, "pu": 95.00},
                ]},
            ]
        },
        # 3. SUPRAESTRUTURA
        {
            "codigo": "03", "descricao": "SUPRAESTRUTURA", "nivel": 1,
            "subitens": [
                {"codigo": "03.01", "descricao": "Concreto", "nivel": 2, "subitens": [
                    {"codigo": "03.01.001", "descricao": "Concreto 30MPa (pilares e vigas)", "un": "m³", "qtd": 1400.0, "pu": 580.00},
                    {"codigo": "03.01.002", "descricao": "Concreto 25MPa (lajes)", "un": "m³", "qtd": 950.0, "pu": 520.00},
                    {"codigo": "03.01.003", "descricao": "Lançamento e adensamento (bomba)", "un": "m³", "qtd": 2350.0, "pu": 42.00},
                ]},
                {"codigo": "03.02", "descricao": "Formas", "nivel": 2, "subitens": [
                    {"codigo": "03.02.001", "descricao": "Forma de pilares (compensado resinado)", "un": "m²", "qtd": 3800.0, "pu": 95.00},
                    {"codigo": "03.02.002", "descricao": "Forma de vigas (compensado resinado)", "un": "m²", "qtd": 4200.0, "pu": 85.00},
                    {"codigo": "03.02.003", "descricao": "Forma de lajes (compensado resinado)", "un": "m²", "qtd": 7650.0, "pu": 75.00},
                ]},
                {"codigo": "03.03", "descricao": "Armadura", "nivel": 2, "subitens": [
                    {"codigo": "03.03.001", "descricao": "Aço CA-50 (fornecimento + corte + dobra)", "un": "kg", "qtd": 180000.0, "pu": 8.50},
                    {"codigo": "03.03.002", "descricao": "Montagem de armadura (pilares e vigas)", "un": "kg", "qtd": 180000.0, "pu": 3.20},
                    {"codigo": "03.03.003", "descricao": "Tela soldada Q-196 (lajes)", "un": "kg", "qtd": 20000.0, "pu": 9.80},
                ]},
                {"codigo": "03.04", "descricao": "Protensão", "nivel": 2, "subitens": [
                    {"codigo": "03.04.001", "descricao": "Cordoalha CP-190 RB 12,7mm", "un": "kg", "qtd": 8500.0, "pu": 18.50},
                    {"codigo": "03.04.002", "descricao": "Protensão de lajes (serviço)", "un": "m²", "qtd": 4200.0, "pu": 35.00},
                ]},
            ]
        },
        # 4. ALVENARIA
        {
            "codigo": "04", "descricao": "ALVENARIA", "nivel": 1,
            "subitens": [
                {"codigo": "04.01", "descricao": "Alvenaria Externa", "nivel": 2, "subitens": [
                    {"codigo": "04.01.001", "descricao": "Bloco cerâmico 14x19x29cm (vedação externa)", "un": "m²", "qtd": 5200.0, "pu": 85.00},
                    {"codigo": "04.01.002", "descricao": "Vergas e contravergas (concreto 20MPa)", "un": "m", "qtd": 2400.0, "pu": 45.00},
                    {"codigo": "04.01.003", "descricao": "Encunhamento com argamassa expansiva", "un": "m", "qtd": 1800.0, "pu": 28.00},
                ]},
                {"codigo": "04.02", "descricao": "Vedação Interna (Drywall)", "nivel": 2, "subitens": [
                    {"codigo": "04.02.001", "descricao": "Drywall standard 12,5mm (paredes internas)", "un": "m²", "qtd": 18500.0, "pu": 62.00},
                    {"codigo": "04.02.002", "descricao": "Drywall RU 12,5mm (áreas molhadas)", "un": "m²", "qtd": 3800.0, "pu": 75.00},
                    {"codigo": "04.02.003", "descricao": "Isolamento acústico (lã de rocha 50mm)", "un": "m²", "qtd": 8200.0, "pu": 35.00},
                ]},
                {"codigo": "04.03", "descricao": "Alvenaria Estrutural (Áreas Técnicas)", "nivel": 2, "subitens": [
                    {"codigo": "04.03.001", "descricao": "Bloco sical (escadas - IN 18 CBMSC)", "un": "m²", "qtd": 850.0, "pu": 125.00},
                    {"codigo": "04.03.002", "descricao": "Bloco de concreto (shafts)", "un": "m²", "qtd": 1200.0, "pu": 95.00},
                ]},
            ]
        },
        # 5. IMPERMEABILIZAÇÃO
        {
            "codigo": "05", "descricao": "IMPERMEABILIZAÇÃO", "nivel": 1,
            "subitens": [
                {"codigo": "05.01", "descricao": "Áreas Molhadas (Banheiros e Cozinhas)", "nivel": 2, "subitens": [
                    {"codigo": "05.01.001", "descricao": "Manta líquida acrílica (banheiros)", "un": "m²", "qtd": 4200.0, "pu": 48.00},
                    {"codigo": "05.01.002", "descricao": "Manta líquida acrílica (cozinhas/serviço)", "un": "m²", "qtd": 2800.0, "pu": 38.00},
                    {"codigo": "05.01.003", "descricao": "Rodapé impermeável h=30cm", "un": "m", "qtd": 3500.0, "pu": 22.00},
                ]},
                {"codigo": "05.02", "descricao": "Áreas Externas e Lazer", "nivel": 2, "subitens": [
                    {"codigo": "05.02.001", "descricao": "Impermeabilização de piscina (manta asfáltica 4mm)", "un": "m²", "qtd": 85.0, "pu": 180.00},
                    {"codigo": "05.02.002", "descricao": "Impermeabilização de deck molhado", "un": "m²", "qtd": 250.0, "pu": 95.00},
                    {"codigo": "05.02.003", "descricao": "Impermeabilização de jardineiras de fachada", "un": "m²", "qtd": 320.0, "pu": 120.00},
                ]},
                {"codigo": "05.03", "descricao": "Subsolo e Contenção", "nivel": 2, "subitens": [
                    {"codigo": "05.03.001", "descricao": "Manta asfáltica 4mm (paredes G1-G2-G3)", "un": "m²", "qtd": 3500.0, "pu": 85.00},
                    {"codigo": "05.03.002", "descricao": "Proteção mecânica (argamassa)", "un": "m²", "qtd": 3500.0, "pu": 28.00},
                ]},
            ]
        },
        # 6. REVESTIMENTO INTERNO DE PAREDE
        {
            "codigo": "06", "descricao": "REVESTIMENTO INTERNO DE PAREDE", "nivel": 1,
            "subitens": [
                {"codigo": "06.01", "descricao": "Áreas Secas (Apartamentos)", "nivel": 2, "subitens": [
                    {"codigo": "06.01.001", "descricao": "Emassamento de drywall (massa + lixamento)", "un": "m²", "qtd": 18500.0, "pu": 18.00},
                    {"codigo": "06.01.002", "descricao": "Chapisco + reboco (alvenaria aparente)", "un": "m²", "qtd": 2400.0, "pu": 35.00},
                ]},
                {"codigo": "06.02", "descricao": "Áreas Molhadas (Banheiros e Cozinhas)", "nivel": 2, "subitens": [
                    {"codigo": "06.02.001", "descricao": "Revestimento cerâmico 30x60cm (banheiros)", "un": "m²", "qtd": 3800.0, "pu": 95.00},
                    {"codigo": "06.02.002", "descricao": "Revestimento cerâmico 30x60cm (cozinhas h=1,50m)", "un": "m²", "qtd": 2200.0, "pu": 85.00},
                    {"codigo": "06.02.003", "descricao": "Rejunte epóxi (áreas molhadas)", "un": "m²", "qtd": 6000.0, "pu": 12.00},
                ]},
                {"codigo": "06.03", "descricao": "Áreas Comuns", "nivel": 2, "subitens": [
                    {"codigo": "06.03.001", "descricao": "Grafiato acrílico (halls e corredores)", "un": "m²", "qtd": 1800.0, "pu": 42.00},
                    {"codigo": "06.03.002", "descricao": "Porcelanato 60x60cm (halls de elevador)", "un": "m²", "qtd": 420.0, "pu": 110.00},
                ]},
            ]
        },
        # 7. TETO/FORRO
        {
            "codigo": "07", "descricao": "TETO / FORRO", "nivel": 1,
            "subitens": [
                {"codigo": "07.01", "descricao": "Apartamentos", "nivel": 2, "subitens": [
                    {"codigo": "07.01.001", "descricao": "Emassamento de laje (massa + lixamento)", "un": "m²", "qtd": 9500.0, "pu": 22.00},
                    {"codigo": "07.01.002", "descricao": "Rebaixo em gesso acartonado", "un": "m²", "qtd": 3200.0, "pu": 85.00},
                    {"codigo": "07.01.003", "descricao": "Sanca de gesso (perímetro)", "un": "m", "qtd": 2800.0, "pu": 45.00},
                ]},
                {"codigo": "07.02", "descricao": "Áreas Comuns", "nivel": 2, "subitens": [
                    {"codigo": "07.02.001", "descricao": "Forro de gesso liso (halls e lazer)", "un": "m²", "qtd": 1200.0, "pu": 95.00},
                    {"codigo": "07.02.002", "descricao": "Forro de gesso acartonado (escadas)", "un": "m²", "qtd": 380.0, "pu": 75.00},
                ]},
            ]
        },
        # 8. PISOS
        {
            "codigo": "08", "descricao": "PISOS", "nivel": 1,
            "subitens": [
                {"codigo": "08.01", "descricao": "Apartamentos - Áreas Secas", "nivel": 2, "subitens": [
                    {"codigo": "08.01.001", "descricao": "Contrapiso e=5cm", "un": "m²", "qtd": 11500.0, "pu": 35.00},
                    {"codigo": "08.01.002", "descricao": "Piso laminado de madeira 8mm", "un": "m²", "qtd": 9800.0, "pu": 95.00},
                    {"codigo": "08.01.003", "descricao": "Rodapé de madeira 7cm", "un": "m", "qtd": 8500.0, "pu": 18.00},
                ]},
                {"codigo": "08.02", "descricao": "Apartamentos - Áreas Molhadas", "nivel": 2, "subitens": [
                    {"codigo": "08.02.001", "descricao": "Porcelanato 60x60cm antiderrapante", "un": "m²", "qtd": 4200.0, "pu": 85.00},
                    {"codigo": "08.02.002", "descricao": "Rodapé de porcelanato 7cm", "un": "m", "qtd": 3200.0, "pu": 15.00},
                ]},
                {"codigo": "08.03", "descricao": "Áreas Comuns", "nivel": 2, "subitens": [
                    {"codigo": "08.03.001", "descricao": "Porcelanato retificado 60x120cm (halls/lazer)", "un": "m²", "qtd": 1500.0, "pu": 125.00},
                    {"codigo": "08.03.002", "descricao": "Granito polido (escadas)", "un": "m²", "qtd": 450.0, "pu": 280.00},
                    {"codigo": "08.03.003", "descricao": "Piso de concreto alisado (garagens)", "un": "m²", "qtd": 2800.0, "pu": 45.00},
                ]},
                {"codigo": "08.04", "descricao": "Deck e Áreas Externas", "nivel": 2, "subitens": [
                    {"codigo": "08.04.001", "descricao": "Deck de madeira Cumaru 10cm", "un": "m²", "qtd": 180.0, "pu": 320.00},
                    {"codigo": "08.04.002", "descricao": "Porcelanato antiderrapante tipo madeira", "un": "m²", "qtd": 280.0, "pu": 110.00},
                ]},
            ]
        },
        # 9. PINTURA
        {
            "codigo": "09", "descricao": "PINTURA", "nivel": 1,
            "subitens": [
                {"codigo": "09.01", "descricao": "Pintura Interna (Apartamentos)", "nivel": 2, "subitens": [
                    {"codigo": "09.01.001", "descricao": "Pintura acrílica fosca (paredes e tetos)", "un": "m²", "qtd": 28000.0, "pu": 22.00},
                    {"codigo": "09.01.002", "descricao": "Pintura acrílica semi-brilho (áreas molhadas)", "un": "m²", "qtd": 6000.0, "pu": 28.00},
                    {"codigo": "09.01.003", "descricao": "Preparação de superfície (massa + lixamento)", "un": "m²", "qtd": 34000.0, "pu": 12.00},
                ]},
                {"codigo": "09.02", "descricao": "Pintura Externa (Fachada)", "nivel": 2, "subitens": [
                    {"codigo": "09.02.001", "descricao": "Pintura acrílica elástica para fachadas", "un": "m²", "qtd": 2800.0, "pu": 38.00},
                    {"codigo": "09.02.002", "descricao": "Selador acrílico + massa acrílica", "un": "m²", "qtd": 2800.0, "pu": 28.00},
                ]},
                {"codigo": "09.03", "descricao": "Pintura Áreas Comuns", "nivel": 2, "subitens": [
                    {"codigo": "09.03.001", "descricao": "Pintura acrílica lavável (halls/lazer)", "un": "m²", "qtd": 2500.0, "pu": 32.00},
                    {"codigo": "09.03.002", "descricao": "Esmalte sintético (elementos metálicos)", "un": "m²", "qtd": 850.0, "pu": 45.00},
                ]},
            ]
        },
        # 10. ESQUADRIAS
        {
            "codigo": "10", "descricao": "ESQUADRIAS", "nivel": 1,
            "subitens": [
                {"codigo": "10.01", "descricao": "Esquadrias de Alumínio", "nivel": 2, "subitens": [
                    {"codigo": "10.01.001", "descricao": "Porta de correr 4 folhas (sacadas)", "un": "un", "qtd": 200.0, "pu": 2800.00},
                    {"codigo": "10.01.002", "descricao": "Janela de correr 2 folhas (dormitórios)", "un": "un", "qtd": 350.0, "pu": 950.00},
                    {"codigo": "10.01.003", "descricao": "Janela maxim-ar (banheiros/cozinhas)", "un": "un", "qtd": 280.0, "pu": 480.00},
                ]},
                {"codigo": "10.02", "descricao": "Vidros", "nivel": 2, "subitens": [
                    {"codigo": "10.02.001", "descricao": "Vidro laminado 6mm (portas)", "un": "m²", "qtd": 850.0, "pu": 180.00},
                    {"codigo": "10.02.002", "descricao": "Vidro temperado 8mm (janelas)", "un": "m²", "qtd": 1200.0, "pu": 150.00},
                ]},
                {"codigo": "10.03", "descricao": "Guarda-Corpo", "nivel": 2, "subitens": [
                    {"codigo": "10.03.001", "descricao": "Guarda-corpo em vidro laminado 10mm + inox", "un": "m", "qtd": 1800.0, "pu": 580.00},
                ]},
                {"codigo": "10.04", "descricao": "Portas de Madeira", "nivel": 2, "subitens": [
                    {"codigo": "10.04.001", "descricao": "Porta de entrada blindada (apartamentos)", "un": "un", "qtd": 136.0, "pu": 3500.00},
                    {"codigo": "10.04.002", "descricao": "Porta interna de madeira 80x210cm", "un": "un", "qtd": 450.0, "pu": 680.00},
                    {"codigo": "10.04.003", "descricao": "Porta corta-fogo 90 min (escadas)", "un": "un", "qtd": 24.0, "pu": 2800.00},
                ]},
            ]
        },
        # 11. FACHADA
        {
            "codigo": "11", "descricao": "FACHADA", "nivel": 1,
            "subitens": [
                {"codigo": "11.01", "descricao": "Revestimentos de Fachada", "nivel": 2, "subitens": [
                    {"codigo": "11.01.001", "descricao": "Tijolinho aparente (placas cerâmicas)", "un": "m²", "qtd": 890.0, "pu": 280.00},
                    {"codigo": "11.01.002", "descricao": "Ripado de madeira Cumaru", "un": "m²", "qtd": 475.0, "pu": 420.00},
                    {"codigo": "11.01.003", "descricao": "Peitoril em concreto aparente", "un": "m", "qtd": 800.0, "pu": 180.00},
                ]},
                {"codigo": "11.02", "descricao": "Sistema de Fixação", "nivel": 2, "subitens": [
                    {"codigo": "11.02.001", "descricao": "Estrutura metálica para ripado (perfis aço galv.)", "un": "m²", "qtd": 475.0, "pu": 120.00},
                    {"codigo": "11.02.002", "descricao": "Inserts metálicos chumbados", "un": "un", "qtd": 1200.0, "pu": 45.00},
                ]},
                {"codigo": "11.03", "descricao": "Tratamento de Juntas e Detalhes", "nivel": 2, "subitens": [
                    {"codigo": "11.03.001", "descricao": "Selante elástico (juntas de dilatação)", "un": "m", "qtd": 850.0, "pu": 35.00},
                    {"codigo": "11.03.002", "descricao": "Pingadeiras em alumínio", "un": "m", "qtd": 950.0, "pu": 55.00},
                ]},
            ]
        },
        # 12. INSTALAÇÕES ELÉTRICAS
        {
            "codigo": "12", "descricao": "INSTALAÇÕES ELÉTRICAS", "nivel": 1,
            "subitens": [
                {"codigo": "12.01", "descricao": "Entrada de Energia e Medição", "nivel": 2, "subitens": [
                    {"codigo": "12.01.001", "descricao": "Quadro Geral de Baixa Tensão (QGBT)", "un": "un", "qtd": 1.0, "pu": 45000.00},
                    {"codigo": "12.01.002", "descricao": "Caixa de medição concentrada (136 medidores)", "un": "un", "qtd": 1.0, "pu": 85000.00},
                    {"codigo": "12.01.003", "descricao": "Barramento blindado (busway)", "un": "m", "qtd": 80.0, "pu": 1200.00},
                ]},
                {"codigo": "12.02", "descricao": "Quadros de Distribuição", "nivel": 2, "subitens": [
                    {"codigo": "12.02.001", "descricao": "Quadros de distribuição por pavimento (QD)", "un": "un", "qtd": 45.0, "pu": 3500.00},
                    {"codigo": "12.02.002", "descricao": "Quadros de força (elevadores, bombas)", "un": "un", "qtd": 5.0, "pu": 4800.00},
                ]},
                {"codigo": "12.03", "descricao": "Pontos Elétricos", "nivel": 2, "subitens": [
                    {"codigo": "12.03.001", "descricao": "Ponto de iluminação no teto", "un": "pt", "qtd": 2700.0, "pu": 85.00},
                    {"codigo": "12.03.002", "descricao": "Tomada 2P+T 10A", "un": "pt", "qtd": 3260.0, "pu": 65.00},
                    {"codigo": "12.03.003", "descricao": "Tomada 20A", "un": "pt", "qtd": 506.0, "pu": 95.00},
                    {"codigo": "12.03.004", "descricao": "Interruptor simples/paralelo", "un": "pt", "qtd": 1576.0, "pu": 45.00},
                    {"codigo": "12.03.005", "descricao": "Ponto chuveiro 220V", "un": "pt", "qtd": 340.0, "pu": 120.00},
                    {"codigo": "12.03.006", "descricao": "Ponto ar-condicionado", "un": "pt", "qtd": 340.0, "pu": 180.00},
                ]},
                {"codigo": "12.04", "descricao": "SPDA e Aterramento", "nivel": 2, "subitens": [
                    {"codigo": "12.04.001", "descricao": "Sistema de proteção contra descargas (SPDA)", "un": "un", "qtd": 1.0, "pu": 28000.00},
                    {"codigo": "12.04.002", "descricao": "Malha de aterramento", "un": "un", "qtd": 1.0, "pu": 18000.00},
                ]},
            ]
        },
        # 13. INSTALAÇÕES HIDROSSANITÁRIAS
        {
            "codigo": "13", "descricao": "INSTALAÇÕES HIDROSSANITÁRIAS", "nivel": 1,
            "subitens": [
                {"codigo": "13.01", "descricao": "Água Fria", "nivel": 2, "subitens": [
                    {"codigo": "13.01.001", "descricao": "Reservatório inferior (cisterna 100m³)", "un": "un", "qtd": 1.0, "pu": 85000.00},
                    {"codigo": "13.01.002", "descricao": "Reservatório superior (barrilete 100m³)", "un": "un", "qtd": 1.0, "pu": 75000.00},
                    {"codigo": "13.01.003", "descricao": "Conjunto moto-bomba de recalque", "un": "un", "qtd": 2.0, "pu": 18000.00},
                    {"codigo": "13.01.004", "descricao": "Conjunto moto-bomba de pressurização", "un": "un", "qtd": 2.0, "pu": 22000.00},
                    {"codigo": "13.01.005", "descricao": "Tubulação PVC soldável (prumadas e ramais)", "un": "m", "qtd": 3500.0, "pu": 25.00},
                ]},
                {"codigo": "13.02", "descricao": "Esgoto Sanitário", "nivel": 2, "subitens": [
                    {"codigo": "13.02.001", "descricao": "Tubulação PVC Ø100mm (esgoto)", "un": "m", "qtd": 3200.0, "pu": 28.00},
                    {"codigo": "13.02.002", "descricao": "Tubulação PVC Ø75-100mm (gordura)", "un": "m", "qtd": 800.0, "pu": 32.00},
                    {"codigo": "13.02.003", "descricao": "Tubulação PVC Ø100mm (ventilação)", "un": "m", "qtd": 1000.0, "pu": 22.00},
                    {"codigo": "13.02.004", "descricao": "Caixa de gordura 728L", "un": "un", "qtd": 1.0, "pu": 3500.00},
                    {"codigo": "13.02.005", "descricao": "Caixas sifonadas", "un": "un", "qtd": 150.0, "pu": 85.00},
                ]},
                {"codigo": "13.03", "descricao": "Águas Pluviais", "nivel": 2, "subitens": [
                    {"codigo": "13.03.001", "descricao": "Tubulação PVC Ø150mm (pluvial)", "un": "m", "qtd": 1500.0, "pu": 42.00},
                    {"codigo": "13.03.002", "descricao": "Caixas de areia", "un": "un", "qtd": 8.0, "pu": 850.00},
                ]},
                {"codigo": "13.04", "descricao": "Instalação de Gás", "nivel": 2, "subitens": [
                    {"codigo": "13.04.001", "descricao": "Abrigo de gás GLP (central)", "un": "un", "qtd": 1.0, "pu": 12000.00},
                    {"codigo": "13.04.002", "descricao": "Tubulação de cobre (prumadas e ramais)", "un": "m", "qtd": 1800.0, "pu": 45.00},
                ]},
            ]
        },
        # 14. SISTEMAS ESPECIAIS
        {
            "codigo": "14", "descricao": "SISTEMAS ESPECIAIS", "nivel": 1,
            "subitens": [
                {"codigo": "14.01", "descricao": "Elevadores", "nivel": 2, "subitens": [
                    {"codigo": "14.01.001", "descricao": "Elevador social (8 pessoas, 24 paradas)", "un": "un", "qtd": 2.0, "pu": 380000.00},
                    {"codigo": "14.01.002", "descricao": "Elevador de emergência (13 pessoas)", "un": "un", "qtd": 1.0, "pu": 480000.00},
                ]},
                {"codigo": "14.02", "descricao": "Prevenção e Combate a Incêndio", "nivel": 2, "subitens": [
                    {"codigo": "14.02.001", "descricao": "Sistema de hidrantes e mangotinhos", "un": "un", "qtd": 1.0, "pu": 185000.00},
                    {"codigo": "14.02.002", "descricao": "Reserva técnica de incêndio 50m³", "un": "un", "qtd": 1.0, "pu": 45000.00},
                    {"codigo": "14.02.003", "descricao": "Conjunto moto-bomba de incêndio (elétrico + diesel)", "un": "un", "qtd": 1.0, "pu": 95000.00},
                    {"codigo": "14.02.004", "descricao": "Sistema de detecção e alarme de incêndio", "un": "un", "qtd": 1.0, "pu": 125000.00},
                    {"codigo": "14.02.005", "descricao": "Extintores de incêndio", "un": "un", "qtd": 85.0, "pu": 280.00},
                    {"codigo": "14.02.006", "descricao": "Sinalização de emergência (placas fotoluminescentes)", "un": "un", "qtd": 180.0, "pu": 120.00},
                    {"codigo": "14.02.007", "descricao": "Iluminação de emergência", "un": "un", "qtd": 220.0, "pu": 380.00},
                ]},
                {"codigo": "14.03", "descricao": "Telecomunicações", "nivel": 2, "subitens": [
                    {"codigo": "14.03.001", "descricao": "Sistema de interfonia/vídeo porteiro", "un": "un", "qtd": 1.0, "pu": 85000.00},
                    {"codigo": "14.03.002", "descricao": "CFTV (câmeras IP + NVR)", "un": "un", "qtd": 1.0, "pu": 65000.00"},
                    {"codigo": "14.03.003", "descricao": "Cabeamento estruturado Cat 6", "un": "pt", "qtd": 550.0, "pu": 280.00},
                ]},
            ]
        },
        # 15. COMPLEMENTARES
        {
            "codigo": "15", "descricao": "COMPLEMENTARES", "nivel": 1,
            "subitens": [
                {"codigo": "15.01", "descricao": "Piscina", "nivel": 2, "subitens": [
                    {"codigo": "15.01.001", "descricao": "Estrutura de piscina (concreto armado)", "un": "m³", "qtd": 28.0, "pu": 1200.00},
                    {"codigo": "15.01.002", "descricao": "Revestimento de pastilha de vidro", "un": "m²", "qtd": 85.0, "pu": 180.00},
                    {"codigo": "15.01.003", "descricao": "Sistema de filtração e tratamento", "un": "un", "qtd": 1.0, "pu": 45000.00},
                ]},
                {"codigo": "15.02", "descricao": "Lazer - Academia e Sauna", "nivel": 2, "subitens": [
                    {"codigo": "15.02.001", "descricao": "Equipamentos de academia (musculação + cardio)", "un": "vb", "qtd": 1.0, "pu": 85000.00},
                    {"codigo": "15.02.002", "descricao": "Sauna seca (cabine madeira + aquecedor)", "un": "un", "qtd": 1.0, "pu": 22000.00},
                ]},
                {"codigo": "15.03", "descricao": "Playground e Pet Place", "nivel": 2, "subitens": [
                    {"codigo": "15.03.001", "descricao": "Brinquedos de playground (conjunto completo)", "un": "vb", "qtd": 1.0, "pu": 38000.00},
                    {"codigo": "15.03.002", "descricao": "Piso emborrachado (playground)", "un": "m²", "qtd": 75.0, "pu": 180.00},
                    {"codigo": "15.03.003", "descricao": "Equipamentos pet place", "un": "vb", "qtd": 1.0, "pu": 12000.00},
                ]},
                {"codigo": "15.04", "descricao": "Paisagismo", "nivel": 2, "subitens": [
                    {"codigo": "15.04.001", "descricao": "Grama natural (plantio + irrigação)", "un": "m²", "qtd": 350.0, "pu": 45.00},
                    {"codigo": "15.04.002", "descricao": "Plantas ornamentais (mudas + plantio)", "un": "un", "qtd": 180.0, "pu": 120.00},
                    {"codigo": "15.04.003", "descricao": "Sistema de irrigação automatizado", "un": "un", "qtd": 1.0, "pu": 18000.00},
                ]},
                {"codigo": "15.05", "descricao": "Mobiliário Áreas Comuns", "nivel": 2, "subitens": [
                    {"codigo": "15.05.001", "descricao": "Móveis de lazer (mesas, cadeiras, espreguiçadeiras)", "un": "vb", "qtd": 1.0, "pu": 45000.00},
                    {"codigo": "15.05.002", "descricao": "Churrasqueira pré-moldada", "un": "un", "qtd": 2.0, "pu": 3500.00},
                ]},
            ]
        },
        # 16. LOUÇAS E METAIS
        {
            "codigo": "16", "descricao": "LOUÇAS E METAIS", "nivel": 1,
            "subitens": [
                {"codigo": "16.01", "descricao": "Louças Sanitárias", "nivel": 2, "subitens": [
                    {"codigo": "16.01.001", "descricao": "Vaso sanitário com caixa acoplada", "un": "un", "qtd": 340.0, "pu": 580.00},
                    {"codigo": "16.01.002", "descricao": "Lavatório de sobrepor ou suspenso", "un": "un", "qtd": 340.0, "pu": 420.00},
                ]},
                {"codigo": "16.02", "descricao": "Metais Sanitários", "nivel": 2, "subitens": [
                    {"codigo": "16.02.001", "descricao": "Misturador monocomando para chuveiro", "un": "un", "qtd": 340.0, "pu": 380.00},
                    {"codigo": "16.02.002", "descricao": "Misturador monocomando para lavatório", "un": "un", "qtd": 340.0, "pu": 320.00},
                    {"codigo": "16.02.003", "descricao": "Misturador monocomando para pia (cozinha)", "un": "un", "qtd": 136.0, "pu": 420.00},
                    {"codigo": "16.02.004", "descricao": "Ducha higiênica com registro", "un": "un", "qtd": 340.0, "pu": 180.00},
                ]},
                {"codigo": "16.03", "descricao": "Acessórios", "nivel": 2, "subitens": [
                    {"codigo": "16.03.001", "descricao": "Kit de acessórios (papeleira, cabide, saboneteira)", "un": "cj", "qtd": 340.0, "pu": 150.00},
                ]},
            ]
        },
        # 17. CLIMATIZAÇÃO
        {
            "codigo": "17", "descricao": "CLIMATIZAÇÃO", "nivel": 1,
            "subitens": [
                {"codigo": "17.01", "descricao": "Pré-Instalação Ar-Condicionado", "nivel": 2, "subitens": [
                    {"codigo": "17.01.001", "descricao": "Infraestrutura elétrica (ponto 220V)", "un": "pt", "qtd": 340.0, "pu": 180.00},
                    {"codigo": "17.01.002", "descricao": "Dreno de ar-condicionado (tubulação PVC Ø50mm)", "un": "m", "qtd": 1200.0, "pu": 22.00},
                    {"codigo": "17.01.003", "descricao": "Suporte para condensadora", "un": "un", "qtd": 340.0, "pu": 180.00},
                ]},
                {"codigo": "17.02", "descricao": "Climatização Áreas Comuns", "nivel": 2, "subitens": [
                    {"codigo": "17.02.001", "descricao": "Ar-condicionado split 18.000 BTUs (lazer)", "un": "un", "qtd": 8.0, "pu": 3500.00},
                ]},
            ]
        },
        # 18. GERENCIAMENTO E IMPREVISTOS
        {
            "codigo": "18", "descricao": "GERENCIAMENTO E IMPREVISTOS", "nivel": 1,
            "subitens": [
                {"codigo": "18.01", "descricao": "Administração da Obra", "nivel": 2, "subitens": [
                    {"codigo": "18.01.001", "descricao": "Engenheiro residente (36 meses)", "un": "mês", "qtd": 36.0, "pu": 18000.00},
                    {"codigo": "18.01.002", "descricao": "Mestres de obra (2 profissionais x 36 meses)", "un": "mês", "qtd": 72.0, "pu": 8000.00},
                    {"codigo": "18.01.003", "descricao": "Equipe administrativa (36 meses)", "un": "mês", "qtd": 36.0, "pu": 12000.00},
                ]},
                {"codigo": "18.02", "descricao": "Canteiro de Obras", "nivel": 2, "subitens": [
                    {"codigo": "18.02.001", "descricao": "Instalação de canteiro (containers)", "un": "vb", "qtd": 1.0, "pu": 85000.00},
                    {"codigo": "18.02.002", "descricao": "Manutenção de canteiro (36 meses)", "un": "mês", "qtd": 36.0, "pu": 3500.00},
                    {"codigo": "18.02.003", "descricao": "Guincho cremalheira (36 meses aluguel)", "un": "mês", "qtd": 36.0, "pu": 12000.00},
                ]},
                {"codigo": "18.03", "descricao": "Equipamentos e Ferramentas", "nivel": 2, "subitens": [
                    {"codigo": "18.03.001", "descricao": "Locação de equipamentos diversos", "un": "mês", "qtd": 36.0, "pu": 8500.00},
                    {"codigo": "18.03.002", "descricao": "Ferramentas e EPI", "un": "mês", "qtd": 36.0, "pu": 4500.00},
                ]},
                {"codigo": "18.04", "descricao": "Segurança do Trabalho", "nivel": 2, "subitens": [
                    {"codigo": "18.04.001", "descricao": "PCMAT + ASO + Treinamentos", "un": "vb", "qtd": 1.0, "pu": 28000.00},
                    {"codigo": "18.04.002", "descricao": "Técnico de segurança (36 meses)", "un": "mês", "qtd": 36.0, "pu": 6500.00},
                ]},
                {"codigo": "18.05", "descricao": "Projetos e Consultorias", "nivel": 2, "subitens": [
                    {"codigo": "18.05.001", "descricao": "Projetos complementares e detalhamentos", "un": "vb", "qtd": 1.0, "pu": 180000.00},
                    {"codigo": "18.05.002", "descricao": "Consultorias técnicas especializadas", "un": "vb", "qtd": 1.0, "pu": 85000.00},
                ]},
                {"codigo": "18.06", "descricao": "Ensaios e Controle Tecnológico", "nivel": 2, "subitens": [
                    {"codigo": "18.06.001", "descricao": "Ensaios de concreto (slump test, corpos de prova)", "un": "vb", "qtd": 1.0, "pu": 42000.00},
                    {"codigo": "18.06.002", "descricao": "Ensaios de solo e fundações", "un": "vb", "qtd": 1.0, "pu": 28000.00},
                ]},
                {"codigo": "18.07", "descricao": "Aprovações e Licenças", "nivel": 2, "subitens": [
                    {"codigo": "18.07.001", "descricao": "Alvará de construção + taxas municipais", "un": "vb", "qtd": 1.0, "pu": 35000.00},
                    {"codigo": "18.07.002", "descricao": "Habite-se + vistoria Corpo de Bombeiros", "un": "vb", "qtd": 1.0, "pu": 18000.00},
                ]},
                {"codigo": "18.08", "descricao": "Imprevistos", "nivel": 2, "subitens": [
                    {"codigo": "18.08.001", "descricao": "Reserva técnica para imprevistos (1,3% do total)", "un": "vb", "qtd": 1.0, "pu": 752693.64},
                ]},
            ]
        },
    ]
    
    # Processar orçamento hierárquico
    def processar_itens(itens, nivel_pai=""):
        nonlocal linha_atual
        
        for item in itens:
            cod = item["codigo"]
            desc = item["descricao"]
            nivel = item.get("nivel", 4)
            
            # Indentação conforme nível
            espacos = "  " * (nivel - 1)
            desc_formatada = f"{espacos}{desc}"
            
            # Escrever linha
            ws.cell(row=linha_atual, column=1, value=cod)
            ws.cell(row=linha_atual, column=2, value=desc_formatada)
            
            # Se tem quantidade (item folha)
            if "qtd" in item:
                un = item.get("un", "")
                qtd = item["qtd"]
                pu = item["pu"]
                
                ws.cell(row=linha_atual, column=3, value=un)
                ws.cell(row=linha_atual, column=4, value=qtd)
                ws.cell(row=linha_atual, column=5, value=pu)
                
                # Fórmulas
                ws.cell(row=linha_atual, column=6, value=f"=D{linha_atual}*E{linha_atual}*0.40")  # 40% material
                ws.cell(row=linha_atual, column=7, value=f"=D{linha_atual}*E{linha_atual}*0.60")  # 60% MO
                ws.cell(row=linha_atual, column=8, value=f"=F{linha_atual}+G{linha_atual}")
                
                # % do total (será calculado depois)
                ws.cell(row=linha_atual, column=9, value=f"=H{linha_atual}/$H$2000")
                
                # Formatos numéricos
                ws[f"D{linha_atual}"].number_format = '#,##0.00'
                ws[f"E{linha_atual}"].number_format = 'R$ #,##0.00'
                ws[f"F{linha_atual}"].number_format = 'R$ #,##0.00'
                ws[f"G{linha_atual}"].number_format = 'R$ #,##0.00'
                ws[f"H{linha_atual}"].number_format = 'R$ #,##0.00'
                ws[f"I{linha_atual}"].number_format = '0.00%'
                
            else:
                # Item de agrupamento - formatar bold e subtotal
                ws[f"A{linha_atual}"].font = Font(bold=True, size=10 if nivel == 1 else 9)
                ws[f"B{linha_atual}"].font = Font(bold=True, size=10 if nivel == 1 else 9)
                
                # Fundo cinza para N1
                if nivel == 1:
                    for col in range(1, 10):
                        ws.cell(row=linha_atual, column=col).fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
            
            # Borda
            for col in range(1, 10):
                ws.cell(row=linha_atual, column=col).border = BORDA_FINA
            
            # Zebra para itens folha (nível 4)
            if "qtd" in item and linha_atual % 2 == 0:
                for col in range(1, 10):
                    if not ws.cell(row=linha_atual, column=col).fill.start_color.index:
                        ws.cell(row=linha_atual, column=col).fill = PatternFill(start_color=COR_ZEBRA, end_color=COR_ZEBRA, fill_type="solid")
            
            linha_atual += 1
            
            # Processar subitens recursivamente
            if "subitens" in item:
                processar_itens(item["subitens"], cod)
    
    # Processar todos os macrogrupos
    processar_itens(orcamento)
    
    # Linha de total geral
    total_geral_row = linha_atual
    ws.merge_cells(f'A{total_geral_row}:C{total_geral_row}')
    ws[f"A{total_geral_row}"] = "TOTAL GERAL"
    ws[f"A{total_geral_row}"].font = Font(bold=True, size=12, color=COR_TEXTO_BRANCO)
    ws[f"A{total_geral_row}"].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Subtotais
    primeira_linha_dados = 2
    ultima_linha_dados = linha_atual - 1
    
    ws[f"F{total_geral_row}"] = f"=SUBTOTAL(9,F{primeira_linha_dados}:F{ultima_linha_dados})"
    ws[f"G{total_geral_row}"] = f"=SUBTOTAL(9,G{primeira_linha_dados}:G{ultima_linha_dados})"
    ws[f"H{total_geral_row}"] = f"=F{total_geral_row}+G{total_geral_row}"
    ws[f"I{total_geral_row}"] = "100,00%"
    
    for col in ['F', 'G', 'H']:
        ws[f"{col}{total_geral_row}"].number_format = 'R$ #,##0.00'
        ws[f"{col}{total_geral_row}"].font = Font(bold=True, color=COR_TEXTO_BRANCO)
        ws[f"{col}{total_geral_row}"].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    ws[f"I{total_geral_row}"].font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws[f"I{total_geral_row}"].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 12
    
    # Congelar painéis
    ws.freeze_panes = 'A2'
    
    # Atualizar referência no resumo executivo (B19)
    try:
        resumo = wb["RESUMO EXECUTIVO"]
        resumo['B19'] = f"='ORÇAMENTO DETALHADO'!H{total_geral_row}"
    except:
        pass
    
    return ws

def criar_aba_curva_abc(wb):
    """Aba 3: Curva ABC"""
    ws = wb.create_sheet("CURVA ABC")
    
    # Headers
    headers = ["#", "Código", "Descrição", "Valor (R$)", "% Individual", "% Acumulado", "Classe"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color=COR_TEXTO_BRANCO)
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDA_FINA
    
    # Dados da curva ABC (exemplo com os principais itens)
    # Na implementação real, isso viria do processamento dos dados do orçamento detalhado
    itens_abc = [
        ["1", "03.03.001", "Aço CA-50 (fornecimento + corte + dobra)", 1530000.00],
        ["2", "03.01.001", "Concreto 30MPa (pilares e vigas)", 812000.00],
        ["3", "14.01.001", "Elevador social (8 pessoas, 24 paradas)", 760000.00],
        ["4", "10.03.001", "Guarda-corpo em vidro laminado 10mm + inox", 1044000.00],
        ["5", "02.01.002", "Tirantes (Q=50tf)", 576000.00],
        ["6", "09.01.001", "Pintura acrílica fosca (paredes e tetos)", 616000.00],
        ["7", "12.01.002", "Caixa de medição concentrada (136 medidores)", 85000.00],
        ["8", "08.01.002", "Piso laminado de madeira 8mm", 931000.00],
        ["9", "03.02.003", "Forma de lajes (compensado resinado)", 573750.00],
        ["10", "02.01.001", "Cortina de estacas (hélice contínua Ø600mm)", 1020000.00],
    ]
    
    linha = 2
    for item in itens_abc:
        for col, valor in enumerate(item, start=1):
            ws.cell(row=linha, column=col, value=valor)
        
        # Fórmulas
        ws.cell(row=linha, column=5, value=f"=D{linha}/$D$12")  # % individual
        ws.cell(row=linha, column=6, value=f"=E{linha}+F{linha-1}" if linha > 2 else f"=E{linha}")  # % acumulado
        
        # Classificação ABC
        ws.cell(row=linha, column=7, value=f'=IF(F{linha}<=0.70,"A",IF(F{linha}<=0.90,"B","C"))')
        
        # Formatos
        ws[f"D{linha}"].number_format = 'R$ #,##0.00'
        ws[f"E{linha}"].number_format = '0.00%'
        ws[f"F{linha}"].number_format = '0.00%'
        
        # Zebra
        if linha % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=linha, column=col).fill = PatternFill(start_color=COR_ZEBRA, end_color=COR_ZEBRA, fill_type="solid")
        
        # Borda
        for col in range(1, 8):
            ws.cell(row=linha, column=col).border = BORDA_FINA
        
        linha += 1
    
    # Total
    total_row = linha
    ws.merge_cells(f'A{total_row}:C{total_row}')
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws[f"A{total_row}"].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    ws[f"D{total_row}"] = f"=SUM(D2:D{linha-1})"
    ws[f"D{total_row}"].number_format = 'R$ #,##0.00'
    ws[f"D{total_row}"].font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws[f"D{total_row}"].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Ajustar colunas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10
    
    ws.freeze_panes = 'A2'
    
    return ws

def criar_aba_cronograma(wb):
    """Aba 4: Cronograma Físico-Financeiro"""
    ws = wb.create_sheet("CRONOGRAMA")
    
    # Headers
    ws['A1'] = "Macrogrupo"
    ws['A1'].font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws['A1'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Meses (1 a 36)
    for mes in range(1, 37):
        col = mes + 1
        cell = ws.cell(row=1, column=col, value=f"M{mes}")
        cell.font = Font(bold=True, color=COR_TEXTO_BRANCO)
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Total e % acumulado
    ws.cell(row=1, column=38, value="Total (R$)")
    ws.cell(row=1, column=38).font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws.cell(row=1, column=38).fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Macrogrupos com distribuição mensal simplificada
    macrogrupos_cronograma = [
        "Mov. Terra",
        "Infraestrutura",
        "Supraestrutura",
        "Alvenaria",
        "Impermeabilização",
        "Rev. Int. Parede",
        "Teto/Forro",
        "Pisos",
        "Pintura",
        "Esquadrias",
        "Fachada",
        "Instalações Elétricas",
        "Instalações Hidro",
        "Sistemas Especiais",
        "Complementares",
        "Louças e Metais",
        "Climatização",
        "Gerenciamento",
    ]
    
    linha = 2
    for macro in macrogrupos_cronograma:
        ws.cell(row=linha, column=1, value=macro)
        ws.cell(row=linha, column=1).font = Font(bold=True)
        
        # Distribuição mensal simplificada (exemplo genérico)
        # Na prática, isso viria de um cronograma detalhado
        for mes in range(1, 37):
            col = mes + 1
            # Fórmula placeholder - na prática, distribuir conforme cronograma real
            ws.cell(row=linha, column=col, value=0)  # Será preenchido manualmente
            ws.cell(row=linha, column=col).number_format = 'R$ #,##0.00'
        
        # Total da linha
        ws.cell(row=linha, column=38, value=f"=SUM(B{linha}:AK{linha})")
        ws.cell(row=linha, column=38).number_format = 'R$ #,##0.00'
        
        linha += 1
    
    # Linha de totais mensais
    total_row = linha
    ws.cell(row=total_row, column=1, value="TOTAL MENSAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws.cell(row=total_row, column=1).fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    for mes in range(1, 37):
        col = mes + 1
        ws.cell(row=total_row, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{linha-1})")
        ws.cell(row=total_row, column=col).number_format = 'R$ #,##0.00'
        ws.cell(row=total_row, column=col).font = Font(bold=True, color=COR_TEXTO_BRANCO)
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    ws.cell(row=total_row, column=38, value=f"=SUM(B{total_row}:AK{total_row})")
    ws.cell(row=total_row, column=38).number_format = 'R$ #,##0.00'
    ws.cell(row=total_row, column=38).font = Font(bold=True, color=COR_TEXTO_BRANCO)
    ws.cell(row=total_row, column=38).fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Linha de % acumulado
    acum_row = total_row + 1
    ws.cell(row=acum_row, column=1, value="% ACUMULADO")
    ws.cell(row=acum_row, column=1).font = Font(bold=True)
    
    for mes in range(1, 37):
        col = mes + 1
        if mes == 1:
            ws.cell(row=acum_row, column=col, value=f"={get_column_letter(col)}{total_row}/$AL${total_row}")
        else:
            ws.cell(row=acum_row, column=col, value=f"={get_column_letter(col-1)}{acum_row}+{get_column_letter(col)}{total_row}/$AL${total_row}")
        ws.cell(row=acum_row, column=col).number_format = '0.0%'
        ws.cell(row=acum_row, column=col).font = Font(bold=True)
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 20
    for col in range(2, 39):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    ws.freeze_panes = 'B2'
    
    return ws

def criar_aba_bdi(wb):
    """Aba 5: BDI"""
    ws = wb.create_sheet("BDI")
    
    # Título
    ws['A1'] = "COMPOSIÇÃO DO BDI"
    ws.merge_cells('A1:C1')
    ws['A1'].font = Font(bold=True, size=14, color=COR_TEXTO_BRANCO)
    ws['A1'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    ws['A3'] = "Item"
    ws['B3'] = "Descrição"
    ws['C3'] = "%"
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = Font(bold=True, color=COR_TEXTO_BRANCO)
        ws[f'{col}3'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    # Dados do BDI
    bdi_itens = [
        ["1", "Administração Central", 8.50],
        ["2", "Garantias e Seguros", 1.20],
        ["3", "Riscos e Imprevistos", 1.50],
        ["4", "Lucro", 10.00],
        ["5", "Despesas Financeiras", 1.80],
        ["6", "ISS (5% sobre faturamento)", 5.00],
        ["7", "COFINS (3% sobre faturamento)", 3.00],
        ["8", "PIS (0,65% sobre faturamento)", 0.65],
        ["9", "CPRB (4,5% sobre faturamento)", 4.50],
    ]
    
    linha = 4
    for item in bdi_itens:
        ws.cell(row=linha, column=1, value=item[0])
        ws.cell(row=linha, column=2, value=item[1])
        ws.cell(row=linha, column=3, value=item[2])
        ws.cell(row=linha, column=3).number_format = '0.00%'
        
        # Zebra
        if linha % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=linha, column=col).fill = PatternFill(start_color=COR_ZEBRA, end_color=COR_ZEBRA, fill_type="solid")
        
        linha += 1
    
    # Total BDI (fórmula composta)
    total_row = linha + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value="BDI Composto")
    ws.cell(row=total_row, column=3, value="=PRODUCT(1+C4:C12)-1")
    
    for col in range(1, 4):
        ws.cell(row=total_row, column=col).font = Font(bold=True, color=COR_TEXTO_BRANCO)
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
    
    ws.cell(row=total_row, column=3).number_format = '0.00%'
    
    # Observações
    obs_row = total_row + 3
    ws.cell(row=obs_row, column=1, value="OBSERVAÇÕES:")
    ws.cell(row=obs_row, column=1).font = Font(bold=True)
    ws.cell(row=obs_row + 1, column=1, value="• BDI calculado pelo método composto (multiplicativo)")
    ws.cell(row=obs_row + 2, column=1, value="• Impostos: ISS + COFINS + PIS + CPRB")
    ws.cell(row=obs_row + 3, column=1, value="• Aplicável sobre custo direto da obra")
    ws.cell(row=obs_row + 4, column=1, value="• Referência: Acórdão TCU nº 2622/2013")
    
    # Ajustar colunas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    
    return ws

def main():
    """Função principal"""
    print("=" * 70)
    print("GERANDO ORÇAMENTO EXECUTIVO - NOW RESIDENCE")
    print("=" * 70)
    print()
    
    # Criar workbook
    wb = Workbook()
    
    # Criar abas
    print("[1/5] Criando aba RESUMO EXECUTIVO...")
    criar_aba_resumo(wb)
    
    print("[2/5] Criando aba ORÇAMENTO DETALHADO...")
    criar_aba_detalhado(wb)
    
    print("[3/5] Criando aba CURVA ABC...")
    criar_aba_curva_abc(wb)
    
    print("[4/5] Criando aba CRONOGRAMA...")
    criar_aba_cronograma(wb)
    
    print("[5/5] Criando aba BDI...")
    criar_aba_bdi(wb)
    
    # Salvar
    output_path = "/Users/leokock/orcamentos/projetos/cambert-now/orcamento-executivo-now.xlsx"
    print()
    print(f"Salvando planilha em: {output_path}")
    wb.save(output_path)
    
    print()
    print("=" * 70)
    print("✅ ORÇAMENTO EXECUTIVO GERADO COM SUCESSO!")
    print("=" * 70)
    print()
    print(f"📊 Arquivo: {output_path}")
    print(f"📁 Tamanho: ~{os.path.getsize(output_path) / 1024:.1f} KB")
    print()
    print("Abas criadas:")
    print("  1. RESUMO EXECUTIVO")
    print("  2. ORÇAMENTO DETALHADO (18 macrogrupos hierárquicos)")
    print("  3. CURVA ABC")
    print("  4. CRONOGRAMA")
    print("  5. BDI")
    print()
    print("⚠️  ATENÇÃO:")
    print("  • Revisar valores unitários (baseados em SINAPI SC estimado)")
    print("  • Completar cronograma físico-financeiro (distribuição mensal)")
    print("  • Validar quantitativos com projetos executivos")
    print()

if __name__ == "__main__":
    main()
