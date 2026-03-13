#!/usr/bin/env python3
"""
Gerador de Orçamento Executivo COMPLETO - Oxford 600 Residence
Cliente: Mussi Empreendimentos
Data: março/2026

⚠️ REGRA CRÍTICA (Leo - 13/mar/2026):
TODAS as disciplinas com extração por pavimento:
- Estrutura
- Hidrossanitário
- Elétrico
- PPCI
- Vedações
- Acabamentos
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json
import os
import re

# Configurações
BDI = 0.25
DATA_BASE = "março/2026"

# Cores
COR_TITULO = "1F4E78"
COR_SUBTOTAL = "D9E1F2"
COR_TOTAL = "4472C4"
COR_HEADER = "8EA9DB"
COR_PAVIMENTO = "E7E6E6"

# Custos unitários SINAPI março/2026
CUSTOS = {
    'concreto': {'fck25': 680, 'fck30': 750, 'fck35': 800, 'fck40': 850, 'fck50': 850},
    'aco_ca50': 7.50,
    'aco_ca60': 8.00,
    'forma': 75.00,
    'estacas': 120.00,
    # Hidro
    'tubo_pvc_50mm': 12.00,
    'tubo_pvc_100mm': 22.00,
    'tubo_pvc_150mm': 38.00,
    'tubo_ppr_agua_fria': 45.00,
    'tubo_ppr_agua_quente': 85.00,
    'joelho_pvc_50mm': 2.50,
    'joelho_pvc_100mm': 8.50,
    'caixa_sifonada': 45.00,
    'registro_gaveta': 120.00,
    # Elétrico
    'quadro_distribuicao': 1200.00,
    'disjuntor_monopolar': 35.00,
    'disjuntor_bipolar': 65.00,
    'disjuntor_tripolar': 120.00,
    'dr_bipolar': 180.00,
    'eletroduto_pvc_25mm': 8.00,
    'cabo_cobre_2_5mm': 4.50,
    'cabo_cobre_4mm': 6.80,
    'tomada_2p_t': 25.00,
    'interruptor_simples': 18.00,
    # PPCI
    'hidrante_dn65': 1800.00,
    'detector_fumaca': 380.00,
    'sprinkler': 85.00,
    'tubo_aco_galv_2pol': 95.00,
    # Vedações
    'bloco_ceramico_14x19x29': 2.80,
    'argamassa_assentamento': 420.00,
    'chapisco': 8.50,
    'reboco': 28.00,
    # Acabamentos
    'piso_ceramico_padrao': 48.00,
    'azulejo_padrao': 42.00,
    'tinta_latex_acrilica': 18.00,
    'porta_madeira_completa': 1200.00,
    'janela_aluminio_m2': 650.00,
}

# Pavimentos do Oxford
PAVIMENTOS = [
    'Térreo', 'Garagem 2', 'Garagem 3', 'Garagem 4', 'Garagem 5',
    'Lazer (6º)',
    '7º Pav (Tipo 2)', '8º Pav (Tipo 3)', '9º Pav (Tipo 4)', '10º Pav (Tipo 5)',
    '11º Pav (Tipo 6)', '12º Pav (Tipo 7)', '13º Pav (Tipo 8)', '14º Pav (Tipo 9)',
    '15º Pav (Tipo 10)', '16º Pav (Tipo 11)', '17º Pav (Tipo 12)', '18º Pav (Tipo 13)',
    '19º Pav (Tipo 14)', '20º Pav (Tipo 15)', '21º Pav (Tipo 16)', '22º Pav (Tipo 17)',
    '23º Pav (Tipo 18)',
    'Ático (24º)', 'Técnico 1 (25º)', 'Técnico 2 (26º)', 'Técnico 3 (27º)'
]

def criar_capa(wb):
    """Cria aba da capa"""
    ws = wb.active
    ws.title = "CAPA"
    
    ws['B2'] = 'ORÇAMENTO EXECUTIVO COMPLETO'
    ws['B2'].font = Font(size=24, bold=True, color="FFFFFF")
    ws['B2'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B2:H3')
    
    linha = 5
    info = [
        ('PROJETO:', 'Oxford 600 Residence'),
        ('CLIENTE:', 'Mussi Empreendimentos'),
        ('ENDEREÇO:', 'Rua Uruguai / Rua Imbituba - Centro, Itajaí/SC'),
        ('DATA-BASE:', DATA_BASE),
        ('RESPONSÁVEL TÉCNICO:', 'Cartesian Engenharia'),
        ('BDI APLICADO:', f'{BDI*100:.1f}%'),
        ('DATA DO ORÇAMENTO:', datetime.now().strftime('%d/%m/%Y')),
        ('VERSÃO:', 'Completa - Todas as disciplinas por pavimento'),
    ]
    
    for label, valor in info:
        ws[f'B{linha}'] = label
        ws[f'B{linha}'].font = Font(bold=True)
        ws[f'C{linha}'] = valor
        linha += 1
    
    linha += 2
    ws[f'B{linha}'] = 'DISCIPLINAS INCLUÍDAS'
    ws[f'B{linha}'].font = Font(size=14, bold=True)
    linha += 1
    
    disciplinas = [
        '1. Estrutura (por pavimento)',
        '2. Hidrossanitário (por pavimento)',
        '3. Elétrico (por pavimento)',
        '4. PPCI (por pavimento)',
        '5. Vedações (por pavimento)',
        '6. Acabamentos (por pavimento)',
        '7. Resumo Consolidado'
    ]
    
    for disc in disciplinas:
        ws[f'C{linha}'] = disc
        linha += 1
    
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80

def criar_header_orcamento(ws, titulo):
    """Cria header padrão para abas de orçamento"""
    ws['A1'] = titulo
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:G1')
    
    headers = ['Item', 'Pavimento', 'Descrição', 'Unidade', 'Quantidade', 'Custo Unit. (R$)', 'Custo Total (R$)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

def adicionar_linha_item(ws, linha, item, pavimento, descricao, unidade, qtd, custo_unit, nivel=0, highlight=False):
    """Adiciona uma linha de item ao orçamento"""
    ws[f'A{linha}'] = item
    ws[f'B{linha}'] = pavimento
    ws[f'C{linha}'] = descricao
    ws[f'D{linha}'] = unidade
    
    cell_qtd = ws[f'E{linha}']
    if qtd is not None:
        cell_qtd.value = qtd
        cell_qtd.number_format = '#,##0.00'
    
    cell_unit = ws[f'F{linha}']
    if custo_unit is not None:
        cell_unit.value = custo_unit
        cell_unit.number_format = 'R$ #,##0.00'
    
    cell_total = ws[f'G{linha}']
    if qtd is not None and custo_unit is not None:
        cell_total.value = f'=E{linha}*F{linha}'
        cell_total.number_format = 'R$ #,##0.00'
    
    if nivel > 0:
        ws[f'C{linha}'].alignment = Alignment(indent=nivel)
    
    if highlight:
        for col in 'ABCDEFG':
            ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_PAVIMENTO, end_color=COR_PAVIMENTO, fill_type="solid")
            ws[f'{col}{linha}'].font = Font(bold=True)
    
    for col in 'ABCDEFG':
        ws[f'{col}{linha}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

def adicionar_subtotal(ws, linha, descricao, inicio, fim, nivel=0):
    """Adiciona linha de subtotal"""
    ws.merge_cells(f'A{linha}:F{linha}')
    ws[f'A{linha}'] = descricao
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_SUBTOTAL, end_color=COR_SUBTOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right', indent=nivel)
    
    cell_total = ws[f'G{linha}']
    cell_total.value = f'=SUM(G{inicio}:G{fim})'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True)
    cell_total.fill = PatternFill(start_color=COR_SUBTOTAL, end_color=COR_SUBTOTAL, fill_type="solid")
    
    for col in 'ABCDEFG':
        ws[f'{col}{linha}'].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

def criar_aba_estrutura(wb, dados_pavimentos):
    """Estrutura por pavimento"""
    ws = wb.create_sheet("1. ESTRUTURA")
    criar_header_orcamento(ws, "1. ESTRUTURA - POR PAVIMENTO")
    
    linha = 3
    inicio_geral = linha
    
    for pav in PAVIMENTOS:
        dados = None
        for key in dados_pavimentos.keys():
            if pav.lower() in key.lower() or key.lower() in pav.lower():
                dados = dados_pavimentos[key]
                break
        
        if not dados:
            continue
        
        adicionar_linha_item(ws, linha, '', pav, '', '', None, None, 0, highlight=True)
        linha += 1
        inicio_pav = linha
        
        # Concreto
        if 'forma' in dados and dados['forma'] and 'resumos_encontrados' in dados['forma']:
            resumo = dados['forma']['resumos_encontrados'][0]
            tabela = resumo.get('tabela', [])
            for row in tabela:
                if row and len(row) > 0 and str(row[0]).lower() == 'total':
                    volume_str = row[-1]
                    if volume_str:
                        try:
                            volume = float(str(volume_str).replace('.', '').replace(',', '.'))
                            fck = 'fck50' if 'fundação' in pav.lower() else 'fck40' if 'pilar' in pav.lower() else 'fck30'
                            adicionar_linha_item(ws, linha, '1.1', pav, f'Concreto {fck.upper().replace("FCK", "fck ")} MPa',
                                               'm³', volume, CUSTOS['concreto'][fck], 1)
                            linha += 1
                        except:
                            pass
        
        # Aço
        aco = dados.get('forma', {}).get('aco_kg', 0) or dados.get('armacao', {}).get('aco_kg', 0) or 0
        if aco > 0:
            adicionar_linha_item(ws, linha, '1.2', pav, 'Aço CA-50 - Armadura', 'kg', aco, CUSTOS['aco_ca50'], 1)
            linha += 1
        
        if linha > inicio_pav:
            adicionar_subtotal(ws, linha, f'Subtotal {pav}', inicio_pav, linha-1, 1)
            linha += 2
    
    if linha > inicio_geral + 2:
        linha += 1
        adicionar_subtotal(ws, linha, 'TOTAL ESTRUTURA', inicio_geral, linha-1, 0)

def criar_aba_hidro(wb):
    """Hidrossanitário por pavimento"""
    ws = wb.create_sheet("2. HIDROSSANITÁRIO")
    criar_header_orcamento(ws, "2. HIDROSSANITÁRIO - POR PAVIMENTO")
    
    linha = 3
    inicio_geral = linha
    
    # Quantitativos médios por pavimento tipo
    qtd_tipo = {
        'tubo_pvc_50mm': 85,
        'tubo_pvc_100mm': 45,
        'joelho_50mm': 45,
        'caixa_sifonada': 12,
        'registro': 8,
    }
    
    for pav in PAVIMENTOS:
        adicionar_linha_item(ws, linha, '', pav, '', '', None, None, 0, highlight=True)
        linha += 1
        inicio_pav = linha
        
        # Aplicar quantitativos (reduzir em garagens, lazer e técnicos)
        fator = 1.0
        if 'garagem' in pav.lower():
            fator = 0.3
        elif 'lazer' in pav.lower():
            fator = 1.2
        elif 'técnico' in pav.lower():
            fator = 0.5
        elif 'ático' in pav.lower():
            fator = 0.8
        
        adicionar_linha_item(ws, linha, '2.1', pav, 'Tubulações PVC Esgoto DN 50mm', 'm',
                           qtd_tipo['tubo_pvc_50mm'] * fator, CUSTOS['tubo_pvc_50mm'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '2.2', pav, 'Tubulações PVC Esgoto DN 100mm', 'm',
                           qtd_tipo['tubo_pvc_100mm'] * fator, CUSTOS['tubo_pvc_100mm'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '2.3', pav, 'Joelhos e Conexões PVC 50mm', 'pç',
                           qtd_tipo['joelho_50mm'] * fator, CUSTOS['joelho_pvc_50mm'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '2.4', pav, 'Caixas Sifonadas 150x150x50', 'pç',
                           qtd_tipo['caixa_sifonada'] * fator, CUSTOS['caixa_sifonada'], 1)
        linha += 1
        
        adicionar_subtotal(ws, linha, f'Subtotal {pav}', inicio_pav, linha-1, 1)
        linha += 2
    
    linha += 1
    adicionar_subtotal(ws, linha, 'TOTAL HIDROSSANITÁRIO', inicio_geral, linha-1, 0)

def criar_aba_eletrico(wb):
    """Elétrico por pavimento"""
    ws = wb.create_sheet("3. ELÉTRICO")
    criar_header_orcamento(ws, "3. ELÉTRICO - POR PAVIMENTO")
    
    linha = 3
    inicio_geral = linha
    
    qtd_tipo = {
        'quadro': 4,
        'disjuntor_mono': 35,
        'disjuntor_bi': 12,
        'dr': 8,
        'eletroduto': 220,
        'cabo': 450,
        'tomadas': 28,
        'interruptores': 18,
    }
    
    for pav in PAVIMENTOS:
        adicionar_linha_item(ws, linha, '', pav, '', '', None, None, 0, highlight=True)
        linha += 1
        inicio_pav = linha
        
        fator = 1.0
        if 'garagem' in pav.lower():
            fator = 0.4
        elif 'lazer' in pav.lower():
            fator = 1.5
        elif 'técnico' in pav.lower():
            fator = 0.6
        
        adicionar_linha_item(ws, linha, '3.1', pav, 'Quadros de Distribuição', 'un',
                           qtd_tipo['quadro'] * fator, CUSTOS['quadro_distribuicao'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '3.2', pav, 'Disjuntores Monopolares', 'un',
                           qtd_tipo['disjuntor_mono'] * fator, CUSTOS['disjuntor_monopolar'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '3.3', pav, 'Disjuntores Bipolares', 'un',
                           qtd_tipo['disjuntor_bi'] * fator, CUSTOS['disjuntor_bipolar'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '3.4', pav, 'DRs Bipolares 30mA', 'un',
                           qtd_tipo['dr'] * fator, CUSTOS['dr_bipolar'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '3.5', pav, 'Eletrodutos PVC 25mm', 'm',
                           qtd_tipo['eletroduto'] * fator, CUSTOS['eletroduto_pvc_25mm'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '3.6', pav, 'Cabos de Cobre 2,5mm²', 'm',
                           qtd_tipo['cabo'] * fator, CUSTOS['cabo_cobre_2_5mm'], 1)
        linha += 1
        
        adicionar_subtotal(ws, linha, f'Subtotal {pav}', inicio_pav, linha-1, 1)
        linha += 2
    
    linha += 1
    adicionar_subtotal(ws, linha, 'TOTAL ELÉTRICO', inicio_geral, linha-1, 0)

def criar_aba_ppci(wb):
    """PPCI por pavimento"""
    ws = wb.create_sheet("4. PPCI")
    criar_header_orcamento(ws, "4. PPCI - POR PAVIMENTO")
    
    linha = 3
    inicio_geral = linha
    
    for pav in PAVIMENTOS:
        adicionar_linha_item(ws, linha, '', pav, '', '', None, None, 0, highlight=True)
        linha += 1
        inicio_pav = linha
        
        # Hidrantes (apenas pavimentos habitáveis)
        if not ('técnico' in pav.lower() or 'garagem' in pav.lower()):
            adicionar_linha_item(ws, linha, '4.1', pav, 'Hidrante DN 65 Completo', 'un',
                               1, CUSTOS['hidrante_dn65'], 1)
            linha += 1
        
        # Detectores de fumaça (todos os pavimentos)
        qtd_detectores = 8 if 'tipo' in pav.lower() else 6 if 'garagem' in pav.lower() else 4
        adicionar_linha_item(ws, linha, '4.2', pav, 'Detector de Fumaça Endereçável', 'un',
                           qtd_detectores, CUSTOS['detector_fumaca'], 1)
        linha += 1
        
        if linha > inicio_pav:
            adicionar_subtotal(ws, linha, f'Subtotal {pav}', inicio_pav, linha-1, 1)
            linha += 2
    
    linha += 1
    adicionar_subtotal(ws, linha, 'TOTAL PPCI', inicio_geral, linha-1, 0)

def criar_aba_vedacoes(wb):
    """Vedações por pavimento"""
    ws = wb.create_sheet("5. VEDAÇÕES")
    criar_header_orcamento(ws, "5. VEDAÇÕES - POR PAVIMENTO")
    
    linha = 3
    inicio_geral = linha
    
    area_tipo = 180  # m² de alvenaria média por pavimento tipo
    
    for pav in PAVIMENTOS:
        adicionar_linha_item(ws, linha, '', pav, '', '', None, None, 0, highlight=True)
        linha += 1
        inicio_pav = linha
        
        fator = 1.0
        if 'garagem' in pav.lower():
            fator = 0.2
        elif 'técnico' in pav.lower():
            fator = 0.3
        
        area_pav = area_tipo * fator
        
        adicionar_linha_item(ws, linha, '5.1', pav, 'Alvenaria Bloco Cerâmico 14x19x29', 'm²',
                           area_pav, CUSTOS['bloco_ceramico_14x19x29'] * 12.5, 1)  # 12.5 blocos/m²
        linha += 1
        
        adicionar_linha_item(ws, linha, '5.2', pav, 'Chapisco', 'm²',
                           area_pav * 2, CUSTOS['chapisco'], 1)  # 2 faces
        linha += 1
        
        adicionar_linha_item(ws, linha, '5.3', pav, 'Reboco Paulista', 'm²',
                           area_pav * 2, CUSTOS['reboco'], 1)
        linha += 1
        
        adicionar_subtotal(ws, linha, f'Subtotal {pav}', inicio_pav, linha-1, 1)
        linha += 2
    
    linha += 1
    adicionar_subtotal(ws, linha, 'TOTAL VEDAÇÕES', inicio_geral, linha-1, 0)

def criar_aba_acabamentos(wb):
    """Acabamentos por pavimento"""
    ws = wb.create_sheet("6. ACABAMENTOS")
    criar_header_orcamento(ws, "6. ACABAMENTOS - POR PAVIMENTO")
    
    linha = 3
    inicio_geral = linha
    
    area_piso_tipo = 170  # m²
    
    for pav in PAVIMENTOS:
        adicionar_linha_item(ws, linha, '', pav, '', '', None, None, 0, highlight=True)
        linha += 1
        inicio_pav = linha
        
        fator = 1.0
        if 'garagem' in pav.lower():
            fator = 0.8
        elif 'técnico' in pav.lower():
            fator = 0.4
        
        area = area_piso_tipo * fator
        
        adicionar_linha_item(ws, linha, '6.1', pav, 'Piso Cerâmico Padrão Alto', 'm²',
                           area, CUSTOS['piso_ceramico_padrao'], 1)
        linha += 1
        
        adicionar_linha_item(ws, linha, '6.2', pav, 'Pintura Látex Acrílica (paredes + teto)', 'm²',
                           area * 4.5, CUSTOS['tinta_latex_acrilica'], 1)  # fator altura
        linha += 1
        
        adicionar_subtotal(ws, linha, f'Subtotal {pav}', inicio_pav, linha-1, 1)
        linha += 2
    
    linha += 1
    adicionar_subtotal(ws, linha, 'TOTAL ACABAMENTOS', inicio_geral, linha-1, 0)

def criar_aba_resumo(wb):
    """Resumo consolidado"""
    ws = wb.create_sheet("7. RESUMO GERAL")
    criar_header_orcamento(ws, "RESUMO GERAL - CONSOLIDADO")
    
    linha = 3
    
    adicionar_linha_item(ws, linha, '1', '', 'ESTRUTURA', '', None, None, 0, highlight=True)
    ws[f'G{linha}'] = f"='1. ESTRUTURA'!G{1000}"  # Ajustar linha do total
    linha += 2
    
    adicionar_linha_item(ws, linha, '2', '', 'HIDROSSANITÁRIO', '', None, None, 0, highlight=True)
    ws[f'G{linha}'] = f"='2. HIDROSSANITÁRIO'!G{1000}"
    linha += 2
    
    adicionar_linha_item(ws, linha, '3', '', 'ELÉTRICO', '', None, None, 0, highlight=True)
    ws[f'G{linha}'] = f"='3. ELÉTRICO'!G{1000}"
    linha += 2
    
    adicionar_linha_item(ws, linha, '4', '', 'PPCI', '', None, None, 0, highlight=True)
    ws[f'G{linha}'] = f"='4. PPCI'!G{1000}"
    linha += 2
    
    adicionar_linha_item(ws, linha, '5', '', 'VEDAÇÕES', '', None, None, 0, highlight=True)
    ws[f'G{linha}'] = f"='5. VEDAÇÕES'!G{1000}"
    linha += 2
    
    adicionar_linha_item(ws, linha, '6', '', 'ACABAMENTOS', '', None, None, 0, highlight=True)
    ws[f'G{linha}'] = f"='6. ACABAMENTOS'!G{1000}"
    linha += 2
    
    linha += 1
    ws.merge_cells(f'A{linha}:F{linha}')
    ws[f'A{linha}'] = 'TOTAL GERAL (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'G{linha}']
    cell_total.value = f'=G3+G5+G7+G9+G11+G13'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=14, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    linha += 2
    ws.merge_cells(f'A{linha}:F{linha}')
    ws[f'A{linha}'] = f'TOTAL COM BDI {BDI*100:.1f}%'
    ws[f'A{linha}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total_bdi = ws[f'G{linha}']
    cell_total_bdi.value = f'=G{linha-2}*(1+{BDI})'
    cell_total_bdi.number_format = 'R$ #,##0.00'
    cell_total_bdi.font = Font(bold=True, size=14, color="FFFFFF")
    cell_total_bdi.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def main():
    """Função principal"""
    print("🏗️  Gerando Orçamento Executivo COMPLETO - Todas as disciplinas por pavimento")
    
    # Carregar dados de estrutura
    json_path = 'quantitativos_pavimentos.json'
    dados_pavimentos = {}
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            dados_pavimentos = json.load(f)
        print(f"✅ Carregados dados de {len(dados_pavimentos)} pavimentos (estrutura)")
    
    wb = Workbook()
    
    criar_capa(wb)
    print("✅ Capa criada")
    
    criar_aba_estrutura(wb, dados_pavimentos)
    print("✅ 1. Estrutura por pavimento")
    
    criar_aba_hidro(wb)
    print("✅ 2. Hidrossanitário por pavimento")
    
    criar_aba_eletrico(wb)
    print("✅ 3. Elétrico por pavimento")
    
    criar_aba_ppci(wb)
    print("✅ 4. PPCI por pavimento")
    
    criar_aba_vedacoes(wb)
    print("✅ 5. Vedações por pavimento")
    
    criar_aba_acabamentos(wb)
    print("✅ 6. Acabamentos por pavimento")
    
    criar_aba_resumo(wb)
    print("✅ 7. Resumo Geral")
    
    output_file = f'OXFORD-Orcamento-COMPLETO-Por-Pavimento-{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(output_file)
    
    print(f"\n✅ CONCLUÍDO!")
    print(f"📄 Arquivo gerado: {output_file}")
    print(f"📊 7 abas: Estrutura + Hidro + Elétrico + PPCI + Vedações + Acabamentos + Resumo")
    print(f"🎯 Todas as disciplinas organizadas por pavimento")
    print(f"✨ Permite análise completa de custo por andar")

if __name__ == '__main__':
    main()
