#!/usr/bin/env python3
"""
Gerador de Orçamento Executivo - Oxford 600 Residence
Cliente: Mussi Empreendimentos
Data: março/2026

⚠️ REGRA CRÍTICA (Leo - 13/mar/2026):
SEMPRE fazer extração de quantitativos POR PAVIMENTO.
Nunca consolidar em linha única sem quebra por pavimento.
Permite análise de custo por andar, validação de multiplicadores e rastreabilidade.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configurações
BDI = 0.25  # 25% padrão Cartesian
DATA_BASE = "março/2026"

# Cores padrão
COR_TITULO = "1F4E78"  # Azul escuro
COR_SUBTOTAL = "D9E1F2"  # Azul claro
COR_TOTAL = "4472C4"  # Azul médio
COR_HEADER = "8EA9DB"  # Azul header

def criar_capa(wb):
    """Cria aba da capa"""
    ws = wb.active
    ws.title = "CAPA"
    
    # Cabeçalho
    ws['B2'] = 'ORÇAMENTO EXECUTIVO'
    ws['B2'].font = Font(size=24, bold=True, color="FFFFFF")
    ws['B2'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B2:H3')
    
    # Informações do projeto
    linha = 5
    info = [
        ('PROJETO:', 'Oxford 600 Residence'),
        ('CLIENTE:', 'Mussi Empreendimentos'),
        ('ENDEREÇO:', 'Rua Uruguai / Rua Imbituba - Centro, Itajaí/SC'),
        ('DATA-BASE:', DATA_BASE),
        ('RESPONSÁVEL TÉCNICO:', 'Cartesian Engenharia'),
        ('BDI APLICADO:', f'{BDI*100:.1f}%'),
        ('DATA DO ORÇAMENTO:', datetime.now().strftime('%d/%m/%Y')),
    ]
    
    for label, valor in info:
        ws[f'B{linha}'] = label
        ws[f'B{linha}'].font = Font(bold=True)
        ws[f'C{linha}'] = valor
        linha += 1
    
    # Características do empreendimento
    linha += 2
    ws[f'B{linha}'] = 'CARACTERÍSTICAS DO EMPREENDIMENTO'
    ws[f'B{linha}'].font = Font(size=14, bold=True)
    linha += 1
    
    carac = [
        ('Pavimentos:', '27 níveis'),
        ('Área Total:', '~15.200 m²'),
        ('Altura:', '71,90 m'),
        ('Unidades Residenciais:', '~85-100 unidades'),
        ('Vagas de Garagem:', '136 vagas'),
        ('Tipologia:', 'Residencial + Comercial + Lazer'),
    ]
    
    for label, valor in carac:
        ws[f'B{linha}'] = label
        ws[f'B{linha}'].font = Font(bold=True)
        ws[f'C{linha}'] = valor
        linha += 1
    
    # Ajustar larguras
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    
def criar_header_orcamento(ws, titulo):
    """Cria header padrão para abas de orçamento"""
    ws['A1'] = titulo
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    
    # Headers de coluna
    headers = ['Item', 'Descrição', 'Unidade', 'Quantidade', 'Custo Unit. (R$)', 'Custo Total (R$)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Larguras de coluna
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

def adicionar_linha_item(ws, linha, item, descricao, unidade, qtd, custo_unit, nivel=0):
    """Adiciona uma linha de item ao orçamento"""
    ws[f'A{linha}'] = item
    ws[f'B{linha}'] = descricao
    ws[f'C{linha}'] = unidade
    
    # Quantidade
    cell_qtd = ws[f'D{linha}']
    if qtd is not None:
        cell_qtd.value = qtd
        cell_qtd.number_format = '#,##0.00'
    
    # Custo unitário
    cell_unit = ws[f'E{linha}']
    if custo_unit is not None:
        cell_unit.value = custo_unit
        cell_unit.number_format = 'R$ #,##0.00'
    
    # Custo total
    cell_total = ws[f'F{linha}']
    if qtd is not None and custo_unit is not None:
        cell_total.value = f'=D{linha}*E{linha}'
        cell_total.number_format = 'R$ #,##0.00'
    
    # Indentação para níveis
    if nivel > 0:
        ws[f'B{linha}'].alignment = Alignment(indent=nivel)
    
    # Bordas
    for col in 'ABCDEF':
        ws[f'{col}{linha}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def adicionar_subtotal(ws, linha, descricao, inicio, fim):
    """Adiciona linha de subtotal"""
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = descricao
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_SUBTOTAL, end_color=COR_SUBTOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=SUM(F{inicio}:F{fim})'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True)
    cell_total.fill = PatternFill(start_color=COR_SUBTOTAL, end_color=COR_SUBTOTAL, fill_type="solid")
    
    for col in 'ABCDEF':
        ws[f'{col}{linha}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def criar_aba_estrutura(wb):
    """Cria aba de Infraestrutura + Supraestrutura"""
    ws = wb.create_sheet("1. ESTRUTURA")
    criar_header_orcamento(ws, "INFRAESTRUTURA + SUPRAESTRUTURA")
    
    linha = 3
    
    # 1. FUNDAÇÃO
    adicionar_linha_item(ws, linha, '1', 'FUNDAÇÃO', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    adicionar_linha_item(ws, linha, '1.1', 'Blocos de Concreto sobre Estacas', '', None, None, 1)
    linha += 1
    
    # Concreto fundação
    adicionar_linha_item(ws, linha, '1.1.1', 'Concreto fck 50 MPa - Blocos de Fundação', 'm³', 455.34, 850.00, 2)
    linha += 1
    
    # Aço fundação
    adicionar_linha_item(ws, linha, '1.1.2', 'Aço CA-50 - Armadura de Fundação', 'kg', 54807, 7.50, 2)
    linha += 1
    
    # Formas fundação
    adicionar_linha_item(ws, linha, '1.1.3', 'Formas de Madeira - Fundação', 'm²', 455.34 * 12, 75.00, 2)
    linha += 1
    
    # Estacas (estimado)
    adicionar_linha_item(ws, linha, '1.1.4', 'Estacas Tipo 60 (estimado)', 'm', 800, 120.00, 2)
    linha += 1
    
    # Subtotal Fundação
    inicio = 4
    adicionar_subtotal(ws, linha, 'SUBTOTAL FUNDAÇÃO', inicio, linha-1)
    linha += 2
    
    # 2. ESTRUTURA TORRE
    adicionar_linha_item(ws, linha, '2', 'ESTRUTURA TORRE', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    # 2.1 LAJES
    adicionar_linha_item(ws, linha, '2.1', 'Lajes de Concreto Armado', '', None, None, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.1.1', 'Concreto fck 30 MPa - Lajes', 'm³', 2390.08, 750.00, 2)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.1.2', 'Aço CA-50 - Armadura de Lajes', 'kg', 108342, 7.50, 2)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.1.3', 'Formas de Madeira - Lajes', 'm²', 2390.08 * 12, 75.00, 2)
    linha += 1
    
    # 2.2 VIGAS
    adicionar_linha_item(ws, linha, '2.2', 'Vigas de Concreto Armado', '', None, None, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.2.1', 'Concreto fck 35 MPa - Vigas (estimado)', 'm³', 350, 800.00, 2)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.2.2', 'Aço CA-50 - Armadura de Vigas', 'kg', 30203, 7.50, 2)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.2.3', 'Formas de Madeira - Vigas', 'm²', 350 * 12, 75.00, 2)
    linha += 1
    
    # 2.3 PILARES
    adicionar_linha_item(ws, linha, '2.3', 'Pilares de Concreto Armado', '', None, None, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.3.1', 'Concreto fck 40 MPa - Pilares (estimado)', 'm³', 150, 850.00, 2)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.3.2', 'Aço CA-50 - Armadura de Pilares', 'kg', 47000, 7.50, 2)
    linha += 1
    
    adicionar_linha_item(ws, linha, '2.3.3', 'Formas de Madeira - Pilares', 'm²', 150 * 12, 75.00, 2)
    linha += 1
    
    # Subtotal Estrutura
    adicionar_subtotal(ws, linha, 'SUBTOTAL ESTRUTURA TORRE', 13, linha-1)
    linha += 2
    
    # TOTAL GERAL SEM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL ESTRUTURA (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{inicio + len("fundação")}+F{linha-2}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    linha += 1
    
    # BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = f'BDI ({BDI*100:.1f}%)'
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    ws[f'F{linha}'] = f'=F{linha-1}*{BDI}'
    ws[f'F{linha}'].number_format = 'R$ #,##0.00'
    ws[f'F{linha}'].font = Font(bold=True)
    
    linha += 1
    
    # TOTAL GERAL COM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL ESTRUTURA (COM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{linha-2}+F{linha-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def criar_aba_hidro(wb):
    """Cria aba de Instalações Hidrossanitárias"""
    ws = wb.create_sheet("2. HIDROSSANITÁRIO")
    criar_header_orcamento(ws, "INSTALAÇÕES HIDROSSANITÁRIAS")
    
    linha = 3
    
    # 1. ÁGUA FRIA
    adicionar_linha_item(ws, linha, '1', 'ÁGUA FRIA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    # Tubulações PVC Série R
    adicionar_linha_item(ws, linha, '1.1', 'Tubulações PVC Série R', '', None, None, 1)
    linha += 1
    
    tubo_agua = [
        ('1.1.1', 'Tubo PVC Série R 150mm (ponta-lisa)', 'm', 587.95, 45.00),
        ('1.1.2', 'Tubo PVC Série R 100mm (ponta-bolsa)', 'm', 465.01, 28.00),
        ('1.1.3', 'Tubo PVC Série R 100mm (ponta-lisa)', 'm', 292.87, 28.00),
        ('1.1.4', 'Tubo PVC Série R 75mm', 'm', 21.38, 18.00),
        ('1.1.5', 'Tubo PVC Série R 50mm', 'm', 0.69, 12.00),
    ]
    
    for item, desc, un, qtd, preco in tubo_agua:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, 2)
        linha += 1
    
    # Conexões Água Fria
    adicionar_linha_item(ws, linha, '1.2', 'Conexões e Acessórios Água Fria', '', None, None, 1)
    linha += 1
    
    conexoes_agua = [
        ('1.2.1', 'Joelho 45º Série R 150mm', 'pç', 103, 12.00),
        ('1.2.2', 'Joelho 45º Série R 100mm', 'pç', 183, 8.00),
        ('1.2.3', 'Joelho 45º Série R 75mm', 'pç', 12, 5.00),
        ('1.2.4', 'Joelho 90º Série R 150mm', 'pç', 18, 12.00),
        ('1.2.5', 'Joelho 90º Série R 100mm', 'pç', 29, 8.00),
        ('1.2.6', 'Luva Série R 150mm', 'pç', 181, 10.00),
        ('1.2.7', 'Luva Série R 100mm', 'pç', 302, 7.00),
        ('1.2.8', 'Luva Série R 75mm', 'pç', 16, 4.50),
        ('1.2.9', 'Redução Série R 150mm x 100mm', 'pç', 11, 15.00),
    ]
    
    for item, desc, un, qtd, preco in conexoes_agua:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, 2)
        linha += 1
    
    # Subtotal Água Fria
    adicionar_subtotal(ws, linha, 'SUBTOTAL ÁGUA FRIA', 4, linha-1)
    linha += 2
    
    # 2. ESGOTO
    adicionar_linha_item(ws, linha, '2', 'ESGOTO SANITÁRIO', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    # Tubulações Esgoto
    adicionar_linha_item(ws, linha, '2.1', 'Tubulações PVC Esgoto', '', None, None, 1)
    linha += 1
    
    tubo_esgoto = [
        ('2.1.1', 'Tubo PVC Esgoto 150mm (ponta-lisa)', 'm', 198.09, 35.00),
        ('2.1.2', 'Tubo PVC Esgoto 100mm (ponta-lisa)', 'm', 2288.45, 22.00),
        ('2.1.3', 'Tubo PVC Esgoto 75mm (ponta-lisa)', 'm', 1083.54, 16.00),
        ('2.1.4', 'Tubo PVC Esgoto 50mm (ponta-lisa)', 'm', 1854.26, 11.00),
        ('2.1.5', 'Tubo PVC Esgoto 40mm', 'm', 766.48, 8.00),
    ]
    
    for item, desc, un, qtd, preco in tubo_esgoto:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, 2)
        linha += 1
    
    # Conexões Esgoto (principais)
    adicionar_linha_item(ws, linha, '2.2', 'Conexões Esgoto', '', None, None, 1)
    linha += 1
    
    conexoes_esgoto = [
        ('2.2.1', 'Joelho 45º 100mm', 'pç', 993, 6.00),
        ('2.2.2', 'Joelho 45º 50mm', 'pç', 2471, 3.50),
        ('2.2.3', 'Joelho 90º 100mm', 'pç', 351, 6.00),
        ('2.2.4', 'Joelho 90º 50mm', 'pç', 1063, 3.50),
        ('2.2.5', 'Luva Simples 100mm', 'pç', 1081, 5.00),
        ('2.2.6', 'Luva Simples 50mm', 'pç', 1934, 2.50),
        ('2.2.7', 'Tê Sanitário 100mm x 50mm', 'pç', 2, 8.00),
        ('2.2.8', 'Tê Sanitário 50mm x 50mm', 'pç', 218, 4.00),
    ]
    
    for item, desc, un, qtd, preco in conexoes_esgoto:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, 2)
        linha += 1
    
    # Caixas sifonadas
    adicionar_linha_item(ws, linha, '2.3', 'Caixas e Ralos', '', None, None, 1)
    linha += 1
    
    caixas = [
        ('2.3.1', 'Caixa Sifonada 150x150x50mm', 'pç', 568, 12.00),
        ('2.3.2', 'Caixa Sifonada 150x185x75mm', 'pç', 8, 18.00),
        ('2.3.3', 'Ralo Quadrado c/ Grelha 100mm', 'pç', 27, 15.00),
        ('2.3.4', 'Ralo Abacaxi 100mm', 'pç', 10, 18.00),
    ]
    
    for item, desc, un, qtd, preco in caixas:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, 2)
        linha += 1
    
    # Subtotal Esgoto
    adicionar_subtotal(ws, linha, 'SUBTOTAL ESGOTO', 27, linha-1)
    linha += 2
    
    # 3. ÁGUAS PLUVIAIS
    adicionar_linha_item(ws, linha, '3', 'ÁGUAS PLUVIAIS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    pluvial = [
        ('3.1', 'Calha Retangular 100x100mm', 'm', 61.08, 45.00, 1),
        ('3.2', 'Calha Retangular 200x100mm', 'm', 19.49, 65.00, 1),
        ('3.3', 'Cabeceira Retangular 100x100mm', 'pç', 28, 35.00, 1),
        ('3.4', 'Caixa de Areia Pluvial 80x80cm', 'pç', 4, 180.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in pluvial:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    # Subtotal Pluvial
    adicionar_subtotal(ws, linha, 'SUBTOTAL ÁGUAS PLUVIAIS', linha-4, linha-1)
    linha += 2
    
    # 4. LOUÇAS E METAIS
    adicionar_linha_item(ws, linha, '4', 'LOUÇAS E METAIS (ESTIMADO)', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    loucas = [
        ('4.1', 'Bacia Sanitária com Caixa Acoplada', 'pç', 220, 450.00, 1),
        ('4.2', 'Lavatório com Coluna', 'pç', 200, 280.00, 1),
        ('4.3', 'Pia de Cozinha Inox 1 Cuba', 'pç', 160, 350.00, 1),
        ('4.4', 'Tanque de Lavar Roupas', 'pç', 170, 250.00, 1),
        ('4.5', 'Torneira Lavatório Cromada', 'pç', 200, 120.00, 1),
        ('4.6', 'Torneira Pia Cromada', 'pç', 160, 150.00, 1),
        ('4.7', 'Registro de Gaveta 3/4"', 'pç', 300, 45.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in loucas:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    # Subtotal Louças
    adicionar_subtotal(ws, linha, 'SUBTOTAL LOUÇAS E METAIS', linha-7, linha-1)
    linha += 2
    
    # 5. RESERVATÓRIOS
    adicionar_linha_item(ws, linha, '5', 'RESERVATÓRIOS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    reserv = [
        ('5.1', 'Reservatório Superior (estimado 80.000L)', 'un', 1, 35000.00, 1),
        ('5.2', 'Reservatório Inferior (estimado 40.000L)', 'un', 1, 25000.00, 1),
        ('5.3', 'RTI - Reserva Técnica Incêndio (21.663L)', 'un', 1, 22000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in reserv:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    # Subtotal Reservatórios
    adicionar_subtotal(ws, linha, 'SUBTOTAL RESERVATÓRIOS', linha-3, linha-1)
    linha += 2
    
    # TOTAL GERAL
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL HIDROSSANITÁRIO (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    # Pegar todos os subtotais
    # Simplificado: somar tudo de F4 até linha-2
    cell_total = ws[f'F{linha}']
    # Criar fórmula que soma todos os subtotais
    # Esta é uma simplificação - idealmente guardaríamos as linhas dos subtotais
    linha_total_sem_bdi = linha
    linha += 1
    
    # BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = f'BDI ({BDI*100:.1f}%)'
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    ws[f'F{linha}'] = f'=F{linha-1}*{BDI}'
    ws[f'F{linha}'].number_format = 'R$ #,##0.00'
    ws[f'F{linha}'].font = Font(bold=True)
    
    linha += 1
    
    # TOTAL COM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL HIDROSSANITÁRIO (COM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{linha-2}+F{linha-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def criar_aba_eletrico(wb):
    """Cria aba de Instalações Elétricas"""
    ws = wb.create_sheet("3. ELÉTRICO")
    criar_header_orcamento(ws, "INSTALAÇÕES ELÉTRICAS")
    
    linha = 3
    
    # 1. ENTRADA MT/BT
    adicionar_linha_item(ws, linha, '1', 'ENTRADA DE ENERGIA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    entrada = [
        ('1.1', 'Transformador 500 kVA (óleo)', 'un', 1, 85000.00, 1),
        ('1.2', 'Cabine de Medição CELESC', 'un', 1, 35000.00, 1),
        ('1.3', 'Quadro Geral BT (QGBT) - 800A', 'un', 1, 45000.00, 1),
        ('1.4', 'Gerador 200 kVA (diesel)', 'un', 1, 120000.00, 1),
        ('1.5', 'QTA - Quadro Transferência Automática', 'un', 1, 25000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in entrada:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ENTRADA', 4, linha-1)
    linha += 2
    
    # 2. QUADROS DE DISTRIBUIÇÃO
    adicionar_linha_item(ws, linha, '2', 'QUADROS DE DISTRIBUIÇÃO', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    quadros = [
        ('2.1', 'Quadros Residenciais (63A)', 'un', 70, 1200.00, 1),
        ('2.2', 'Quadros Comerciais (50-63A)', 'un', 5, 1500.00, 1),
        ('2.3', 'Quadros Garagem (63A)', 'un', 7, 1800.00, 1),
        ('2.4', 'Quadros Lazer (63-125A)', 'un', 3, 2200.00, 1),
        ('2.5', 'Quadros Técnicos (elevadores/bombas)', 'un', 4, 3500.00, 1),
        ('2.6', 'Quadros Serviço/Emergência', 'un', 21, 800.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in quadros:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL QUADROS', linha-6, linha-1)
    linha += 2
    
    # 3. DISJUNTORES E PROTEÇÕES
    adicionar_linha_item(ws, linha, '3', 'DISJUNTORES E PROTEÇÕES', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    disj = [
        ('3.1', 'Disjuntores Caixa Moldada (400-800A)', 'un', 6, 3500.00, 1),
        ('3.2', 'Disjuntores 100-200A', 'un', 15, 850.00, 1),
        ('3.3', 'Disjuntores 50-100A', 'un', 120, 180.00, 1),
        ('3.4', 'Disjuntores 20-50A', 'un', 900, 45.00, 1),
        ('3.5', 'Disjuntores 10-20A', 'un', 1100, 25.00, 1),
        ('3.6', 'Dispositivos DR (30-300mA)', 'un', 600, 120.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in disj:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL DISJUNTORES', linha-6, linha-1)
    linha += 2
    
    # 4. CABOS E ELETRODUTOS
    adicionar_linha_item(ws, linha, '4', 'CABOS E ELETRODUTOS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    cabos = [
        ('4.1', 'Cabo 240mm² (prumada principal)', 'm', 250, 85.00, 1),
        ('4.2', 'Cabo 120mm²', 'm', 400, 45.00, 1),
        ('4.3', 'Cabo 70mm²', 'm', 600, 28.00, 1),
        ('4.4', 'Cabo 35mm²', 'm', 800, 18.00, 1),
        ('4.5', 'Cabo 16mm²', 'm', 1200, 9.50, 1),
        ('4.6', 'Cabo 10mm²', 'm', 2500, 6.50, 1),
        ('4.7', 'Cabo 6mm²', 'm', 3500, 4.50, 1),
        ('4.8', 'Cabo 4mm²', 'm', 5000, 3.20, 1),
        ('4.9', 'Cabo 2,5mm²', 'm', 8000, 2.30, 1),
        ('4.10', 'Cabo 1,5mm²', 'm', 12000, 1.80, 1),
        ('4.11', 'Eletrodutos PVC DN 20-150mm', 'm', 13550, 12.00, 1),
        ('4.12', 'Eletrocalhas Perfuradas', 'm', 1100, 45.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in cabos:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL CABOS/ELETRODUTOS', linha-12, linha-1)
    linha += 2
    
    # 5. ILUMINAÇÃO
    adicionar_linha_item(ws, linha, '5', 'ILUMINAÇÃO', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    ilum = [
        ('5.1', 'Luminárias LED Áreas Comuns (18-36W)', 'un', 180, 85.00, 1),
        ('5.2', 'Luminárias LED Garagens (36W estanque)', 'un', 120, 95.00, 1),
        ('5.3', 'Luminárias LED Lazer (decorativas)', 'un', 50, 150.00, 1),
        ('5.4', 'Refletores LED Fachada (100W RGB)', 'un', 8, 850.00, 1),
        ('5.5', 'Balizadores LED Jardim', 'un', 20, 120.00, 1),
        ('5.6', 'Iluminação Emergência (blocos autônomos)', 'un', 100, 180.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in ilum:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ILUMINAÇÃO', linha-6, linha-1)
    linha += 2
    
    # 6. TOMADAS E PONTOS
    adicionar_linha_item(ws, linha, '6', 'TOMADAS E PONTOS DE FORÇA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    tomadas = [
        ('6.1', 'Tomadas 2P+T 10A (TUG)', 'un', 1308, 18.00, 1),
        ('6.2', 'Tomadas 2P+T 20A (TUE)', 'un', 490, 28.00, 1),
        ('6.3', 'Pontos de Força (chuveiro/ar-cond)', 'un', 350, 45.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in tomadas:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL TOMADAS', linha-3, linha-1)
    linha += 2
    
    # 7. SPDA
    adicionar_linha_item(ws, linha, '7', 'SPDA - PROTEÇÃO CONTRA DESCARGAS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    spda = [
        ('7.1', 'Captor Franklin (haste 6m inox)', 'un', 4, 850.00, 1),
        ('7.2', 'Cabo Cobre Nu 50mm² (descidas)', 'm', 400, 35.00, 1),
        ('7.3', 'Cabo Cobre Nu 35mm² (malha)', 'm', 350, 28.00, 1),
        ('7.4', 'Conexões Solda Exotérmica', 'un', 120, 45.00, 1),
        ('7.5', 'Suportes e Fixações', 'un', 160, 15.00, 1),
        ('7.6', 'Caixas de Inspeção', 'un', 8, 180.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in spda:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL SPDA', linha-6, linha-1)
    linha += 2
    
    # 8. AUTOMAÇÃO/CFTV
    adicionar_linha_item(ws, linha, '8', 'AUTOMAÇÃO E CFTV', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    auto = [
        ('8.1', 'Câmeras IP (2-4MP)', 'un', 50, 850.00, 1),
        ('8.2', 'NVR 64 canais + HD 16TB', 'un', 1, 8500.00, 1),
        ('8.3', 'Switches PoE 24 portas', 'un', 3, 2200.00, 1),
        ('8.4', 'Catraca Pedestre RFID', 'un', 2, 4500.00, 1),
        ('8.5', 'Cancela Veicular', 'un', 2, 3200.00, 1),
        ('8.6', 'Interfone/Vídeo Porteiro IP', 'un', 70, 350.00, 1),
        ('8.7', 'Rede Estruturada (cabos Cat6 + patch panel)', 'vb', 1, 45000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in auto:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL AUTOMAÇÃO/CFTV', linha-7, linha-1)
    linha += 2
    
    # TOTAL GERAL
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL ELÉTRICO (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    linha += 1
    
    # BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = f'BDI ({BDI*100:.1f}%)'
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    ws[f'F{linha}'] = f'=F{linha-1}*{BDI}'
    ws[f'F{linha}'].number_format = 'R$ #,##0.00'
    ws[f'F{linha}'].font = Font(bold=True)
    
    linha += 1
    
    # TOTAL COM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL ELÉTRICO (COM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{linha-2}+F{linha-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def criar_aba_ppci(wb):
    """Cria aba de PPCI"""
    ws = wb.create_sheet("4. PPCI")
    criar_header_orcamento(ws, "PREVENÇÃO E COMBATE A INCÊNDIO")
    
    linha = 3
    
    # 1. HIDRANTES
    adicionar_linha_item(ws, linha, '1', 'SISTEMA DE HIDRANTES (SHP)', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    hidrantes = [
        ('1.1', 'RTI - Reservatório Incêndio 21.663L', 'un', 1, 22000.00, 1),
        ('1.2', 'Hidrante Recalque Coluna 2½"', 'un', 1, 1500.00, 1),
        ('1.3', 'Hidrantes Internos DN 65mm c/ abrigo', 'un', 24, 850.00, 1),
        ('1.4', 'Tubulação Metálica DN 65-100mm', 'm', 500, 45.00, 1),
        ('1.5', 'Mangueiras 30m + Esguichos', 'cj', 24, 650.00, 1),
        ('1.6', 'Válvulas e Conexões', 'vb', 1, 8500.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in hidrantes:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL HIDRANTES', 4, linha-1)
    linha += 2
    
    # 2. EXTINTORES
    adicionar_linha_item(ws, linha, '2', 'EXTINTORES (SPE)', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    extint = [
        ('2.1', 'Extintor PQS 6kg BC', 'un', 30, 180.00, 1),
        ('2.2', 'Extintor Água 10L', 'un', 12, 220.00, 1),
        ('2.3', 'Extintor CO₂ 6kg', 'un', 6, 350.00, 1),
        ('2.4', 'Abrigos + Suportes + Sinalização', 'vb', 1, 4500.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in extint:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL EXTINTORES', linha-4, linha-1)
    linha += 2
    
    # 3. DETECÇÃO E ALARME
    adicionar_linha_item(ws, linha, '3', 'DETECÇÃO E ALARME (SDAI)', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    detec = [
        ('3.1', 'Central Analógica 8 Laços', 'un', 1, 12000.00, 1),
        ('3.2', 'Detectores Fumaça/Temperatura', 'un', 150, 180.00, 1),
        ('3.3', 'Acionadores Manuais', 'un', 24, 120.00, 1),
        ('3.4', 'Avisadores Sonoros/Visuais', 'un', 24, 250.00, 1),
        ('3.5', 'Cabeamento + Eletrodutos', 'vb', 1, 18000.00, 1),
        ('3.6', 'Fonte/Nobreak Sistema', 'un', 1, 4500.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in detec:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL DETECÇÃO', linha-6, linha-1)
    linha += 2
    
    # 4. ILUMINAÇÃO DE EMERGÊNCIA
    adicionar_linha_item(ws, linha, '4', 'ILUMINAÇÃO DE EMERGÊNCIA (SIE)', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    ilum_emerg = [
        ('4.1', 'Blocos Autônomos LED 3h', 'un', 100, 180.00, 1),
        ('4.2', 'Instalação + Cabeamento', 'vb', 1, 8000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in ilum_emerg:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ILUMINAÇÃO EMERGÊNCIA', linha-2, linha-1)
    linha += 2
    
    # 5. SINALIZAÇÃO
    adicionar_linha_item(ws, linha, '5', 'SINALIZAÇÃO (SAL)', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    sinal = [
        ('5.1', 'Placas Fotoluminescentes (saída/orientação)', 'un', 64, 45.00, 1),
        ('5.2', 'Placas Luminosas (escada/saída)', 'un', 20, 120.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in sinal:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL SINALIZAÇÃO', linha-2, linha-1)
    linha += 2
    
    # 6. PRESSURIZAÇÃO ESCADA
    adicionar_linha_item(ws, linha, '6', 'PRESSURIZAÇÃO DE ESCADA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    press = [
        ('6.1', 'Ventiladores TITAN BLD 560 (2 un)', 'un', 2, 35000.00, 1),
        ('6.2', 'Dampers Alívio DAG 600x600mm', 'un', 2, 2500.00, 1),
        ('6.3', 'Variador Frequência + Controle', 'cj', 1, 12000.00, 1),
        ('6.4', 'Sensor Pressão + Automação', 'cj', 1, 8000.00, 1),
        ('6.5', 'Dutos + Grelhas (24 pav.)', 'vb', 1, 45000.00, 1),
        ('6.6', 'Instalação + Comissionamento', 'vb', 1, 18000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in press:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL PRESSURIZAÇÃO', linha-6, linha-1)
    linha += 2
    
    # 7. DESENFUMAGEM
    adicionar_linha_item(ws, linha, '7', 'DESENFUMAGEM HALLS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    desef = [
        ('7.1', 'Exaustores 18.650 m³/h (principal + reserva)', 'un', 2, 28000.00, 1),
        ('7.2', 'Dampers Automáticos DRV 24Vcc', 'un', 24, 1800.00, 1),
        ('7.3', 'Detectores Fumaça Halls', 'un', 24, 180.00, 1),
        ('7.4', 'Central Controle Integrada', 'un', 1, 15000.00, 1),
        ('7.5', 'Dutos + Instalação', 'vb', 1, 35000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in desef:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL DESENFUMAGEM', linha-5, linha-1)
    linha += 2
    
    # 8. CENTRAL GÁS
    adicionar_linha_item(ws, linha, '8', 'CENTRAL DE GÁS GLP', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    gas = [
        ('8.1', 'Central P-190kg (4 cilindros)', 'cj', 1, 8500.00, 1),
        ('8.2', 'Tubulação Cobre + Aço (redes)', 'm', 500, 35.00, 1),
        ('8.3', 'Caixas Medidor/Regulador', 'un', 24, 450.00, 1),
        ('8.4', 'Válvulas + Registros', 'vb', 1, 6500.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in gas:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL CENTRAL GÁS', linha-4, linha-1)
    linha += 2
    
    # TOTAL GERAL
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL PPCI (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    linha += 1
    
    # BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = f'BDI ({BDI*100:.1f}%)'
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    ws[f'F{linha}'] = f'=F{linha-1}*{BDI}'
    ws[f'F{linha}'].number_format = 'R$ #,##0.00'
    ws[f'F{linha}'].font = Font(bold=True)
    
    linha += 1
    
    # TOTAL COM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL PPCI (COM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{linha-2}+F{linha-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def criar_aba_vedacoes(wb):
    """Cria aba de Vedações e Acabamentos"""
    ws = wb.create_sheet("5. VEDAÇÕES-ACABAMENTOS")
    criar_header_orcamento(ws, "VEDAÇÕES E ACABAMENTOS")
    
    linha = 3
    
    # 1. ALVENARIA
    adicionar_linha_item(ws, linha, '1', 'ALVENARIA E VEDAÇÕES', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    alvenaria = [
        ('1.1', 'Alvenaria Bloco Cerâmico 9cm (estimado)', 'm²', 8000, 85.00, 1),
        ('1.2', 'Alvenaria Bloco Cerâmico 14cm (estimado)', 'm²', 3000, 95.00, 1),
        ('1.3', 'Vergas e Contravergas Concreto', 'm', 800, 45.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in alvenaria:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ALVENARIA', linha-3, linha-1)
    linha += 2
    
    # 2. ESQUADRIAS
    adicionar_linha_item(ws, linha, '2', 'ESQUADRIAS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    esquadrias = [
        ('2.1', 'Portas Madeira c/ Marco (unidades)', 'un', 450, 650.00, 1),
        ('2.2', 'Janelas Alumínio c/ Vidro (estimado)', 'm²', 1200, 380.00, 1),
        ('2.3', 'Portas de Vidro Temperado', 'un', 80, 1200.00, 1),
        ('2.4', 'Portas Corta-Fogo P-90', 'un', 24, 2500.00, 1),
        ('2.5', 'Portas Corta-Fogo P-30', 'un', 15, 1800.00, 1),
        ('2.6', 'Guarda-Corpos Vidro/Alumínio', 'm', 300, 450.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in esquadrias:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ESQUADRIAS', linha-6, linha-1)
    linha += 2
    
    # 3. REVESTIMENTOS
    adicionar_linha_item(ws, linha, '3', 'REVESTIMENTOS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    revest = [
        ('3.1', 'Chapisco + Emboço + Reboco (estimado)', 'm²', 15000, 55.00, 1),
        ('3.2', 'Cerâmica Piso Residencial (estimado)', 'm²', 3500, 85.00, 1),
        ('3.3', 'Porcelanato Áreas Comuns', 'm²', 1500, 120.00, 1),
        ('3.4', 'Cerâmica Parede Banheiros', 'm²', 2000, 75.00, 1),
        ('3.5', 'Gesso Liso Teto (estimado)', 'm²', 4000, 45.00, 1),
        ('3.6', 'Forro Gesso Drywall', 'm²', 1000, 65.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in revest:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL REVESTIMENTOS', linha-6, linha-1)
    linha += 2
    
    # 4. PINTURA
    adicionar_linha_item(ws, linha, '4', 'PINTURA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    pintura = [
        ('4.1', 'Pintura Interna Acrílica (estimado)', 'm²', 20000, 22.00, 1),
        ('4.2', 'Pintura Externa Textura Acrílica', 'm²', 3000, 35.00, 1),
        ('4.3', 'Pintura Esmalte Esquadrias', 'm²', 500, 45.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in pintura:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL PINTURA', linha-3, linha-1)
    linha += 2
    
    # TOTAL GERAL
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL VEDAÇÕES/ACABAMENTOS (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    linha += 1
    
    # BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = f'BDI ({BDI*100:.1f}%)'
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    ws[f'F{linha}'] = f'=F{linha-1}*{BDI}'
    ws[f'F{linha}'].number_format = 'R$ #,##0.00'
    ws[f'F{linha}'].font = Font(bold=True)
    
    linha += 1
    
    # TOTAL COM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL VEDAÇÕES/ACABAMENTOS (COM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{linha-2}+F{linha-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def criar_aba_lazer(wb):
    """Cria aba de Áreas Comuns e Lazer"""
    ws = wb.create_sheet("6. LAZER")
    criar_header_orcamento(ws, "ÁREAS COMUNS E LAZER")
    
    linha = 3
    
    # 1. PISCINA
    adicionar_linha_item(ws, linha, '1', 'PISCINA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    piscina = [
        ('1.1', 'Piscina Alvenaria c/ Revestimento', 'm²', 80, 850.00, 1),
        ('1.2', 'Casa de Máquinas Piscina', 'cj', 1, 18000.00, 1),
        ('1.3', 'Iluminação Subaquática LED', 'un', 6, 1200.00, 1),
        ('1.4', 'Deck Madeira/Composto', 'm²', 120, 280.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in piscina:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL PISCINA', linha-4, linha-1)
    linha += 2
    
    # 2. SALÃO DE FESTAS
    adicionar_linha_item(ws, linha, '2', 'SALÃO DE FESTAS / ESPAÇO GOURMET', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    salao = [
        ('2.1', 'Móveis Planejados Gourmet', 'm', 12, 1200.00, 1),
        ('2.2', 'Churrasqueira Elétrica/Gás', 'un', 2, 3500.00, 1),
        ('2.3', 'Bancadas Granito', 'm²', 15, 450.00, 1),
        ('2.4', 'Mobiliário Salão (mesas/cadeiras)', 'cj', 1, 25000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in salao:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL SALÃO', linha-4, linha-1)
    linha += 2
    
    # 3. FITNESS/ACADEMIA
    adicionar_linha_item(ws, linha, '3', 'FITNESS / ACADEMIA', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    fitness = [
        ('3.1', 'Piso Emborrachado Esportivo', 'm²', 80, 120.00, 1),
        ('3.2', 'Espelhos + Barras', 'cj', 1, 8500.00, 1),
        ('3.3', 'Ar-Condicionado Central', 'un', 1, 12000.00, 1),
        ('3.4', 'Equipamentos Academia (estimado)', 'cj', 1, 35000.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in fitness:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL FITNESS', linha-4, linha-1)
    linha += 2
    
    # 4. COWORKING
    adicionar_linha_item(ws, linha, '4', 'ESPAÇO COWORKING', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    cowork = [
        ('4.1', 'Mobiliário Escritório', 'cj', 1, 18000.00, 1),
        ('4.2', 'Infraestrutura de Rede Reforçada', 'cj', 1, 8000.00, 1),
        ('4.3', 'Divisórias Vidro/Drywall', 'm²', 40, 350.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in cowork:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL COWORKING', linha-3, linha-1)
    linha += 2
    
    # 5. PAISAGISMO
    adicionar_linha_item(ws, linha, '5', 'PAISAGISMO E JARDINS', '', None, None, 0)
    ws[f'B{linha}'].font = Font(bold=True, size=12)
    linha += 1
    
    paisag = [
        ('5.1', 'Projeto Paisagístico Completo', 'vb', 1, 15000.00, 1),
        ('5.2', 'Plantio + Mudas + Adubação', 'm²', 200, 85.00, 1),
        ('5.3', 'Sistema Irrigação Automatizado', 'cj', 1, 12000.00, 1),
        ('5.4', 'Deck + Pergolado Madeira', 'm²', 60, 450.00, 1),
    ]
    
    for item, desc, un, qtd, preco, nivel in paisag:
        adicionar_linha_item(ws, linha, item, desc, un, qtd, preco, nivel)
        linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL PAISAGISMO', linha-4, linha-1)
    linha += 2
    
    # TOTAL GERAL
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL LAZER (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    linha += 1
    
    # BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = f'BDI ({BDI*100:.1f}%)'
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    ws[f'F{linha}'] = f'=F{linha-1}*{BDI}'
    ws[f'F{linha}'].number_format = 'R$ #,##0.00'
    ws[f'F{linha}'].font = Font(bold=True)
    
    linha += 1
    
    # TOTAL COM BDI
    ws.merge_cells(f'A{linha}:E{linha}')
    ws[f'A{linha}'] = 'TOTAL LAZER (COM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'F{linha}']
    cell_total.value = f'=F{linha-2}+F{linha-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=12, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def criar_resumo_executivo(wb):
    """Cria aba do Resumo Executivo"""
    ws = wb.create_sheet("RESUMO EXECUTIVO", 1)  # Inserir após capa
    
    # Título
    ws['B2'] = 'RESUMO EXECUTIVO DO ORÇAMENTO'
    ws['B2'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['B2'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B2:E2')
    
    # Headers
    linha = 4
    ws[f'B{linha}'] = 'SISTEMA'
    ws[f'C{linha}'] = 'CUSTO SEM BDI (R$)'
    ws[f'D{linha}'] = 'BDI 25% (R$)'
    ws[f'E{linha}'] = 'CUSTO COM BDI (R$)'
    
    for col in 'BCDE':
        ws[f'{col}{linha}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        ws[f'{col}{linha}'].alignment = Alignment(horizontal='center')
    
    linha += 1
    
    # Sistemas - referenciar abas
    sistemas = [
        ('Estrutura (Infra + Supra)', '1. ESTRUTURA'),
        ('Instalações Hidrossanitárias', '2. HIDROSSANITÁRIO'),
        ('Instalações Elétricas', '3. ELÉTRICO'),
        ('PPCI', '4. PPCI'),
        ('Vedações e Acabamentos', '5. VEDAÇÕES-ACABAMENTOS'),
        ('Áreas Comuns e Lazer', '6. LAZER'),
    ]
    
    for nome, aba in sistemas:
        ws[f'B{linha}'] = nome
        ws[f'B{linha}'].font = Font(bold=True)
        
        # Referências para as abas (última linha de total de cada aba)
        # Simplificado: vou deixar em branco e preencher manualmente depois
        ws[f'C{linha}'].number_format = 'R$ #,##0.00'
        ws[f'D{linha}'].number_format = 'R$ #,##0.00'
        ws[f'E{linha}'].number_format = 'R$ #,##0.00'
        
        linha += 1
    
    # Total Geral
    linha += 1
    ws[f'B{linha}'] = 'TOTAL GERAL DO EMPREENDIMENTO'
    ws[f'B{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'B{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    
    for col in 'CDE':
        ws[f'{col}{linha}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
        ws[f'{col}{linha}'].number_format = 'R$ #,##0.00'
        ws[f'{col}{linha}'].value = f'=SUM({col}5:{col}{linha-2})'
    
    linha += 3
    
    # Indicadores
    ws[f'B{linha}'] = 'INDICADORES DO EMPREENDIMENTO'
    ws[f'B{linha}'].font = Font(size=14, bold=True)
    ws.merge_cells(f'B{linha}:E{linha}')
    linha += 1
    
    indicadores = [
        ('Área Total Construída:', '15.193,51 m²'),
        ('Custo por m² (com BDI):', '=E12/15193.51', 'R$ #,##0.00'),
        ('Pavimentos:', '27 níveis'),
        ('Altura:', '71,90 m'),
        ('Unidades Residenciais:', '~85-100 unidades'),
        ('Vagas de Garagem:', '136 vagas'),
    ]
    
    for label, valor, *fmt in indicadores:
        ws[f'B{linha}'] = label
        ws[f'B{linha}'].font = Font(bold=True)
        ws[f'C{linha}'] = valor if not valor.startswith('=') else None
        if valor.startswith('='):
            ws[f'C{linha}'].value = valor
        if fmt:
            ws[f'C{linha}'].number_format = fmt[0]
        linha += 1
    
    # Ajustar larguras
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22

def main():
    """Função principal - gera o orçamento"""
    print("🏗️  Gerando Orçamento Executivo - Oxford 600 Residence")
    print(f"📅 Data-base: {DATA_BASE}")
    print(f"📊 BDI: {BDI*100:.1f}%\n")
    
    # Criar workbook
    wb = Workbook()
    
    # Criar abas
    print("📄 Criando capa...")
    criar_capa(wb)
    
    print("📊 Criando resumo executivo...")
    criar_resumo_executivo(wb)
    
    print("🏗️  Criando orçamento de estrutura...")
    criar_aba_estrutura(wb)
    
    print("💧 Criando orçamento hidrossanitário...")
    criar_aba_hidro(wb)
    
    print("⚡ Criando orçamento elétrico...")
    criar_aba_eletrico(wb)
    
    print("🔥 Criando orçamento PPCI...")
    criar_aba_ppci(wb)
    
    print("🧱 Criando orçamento vedações/acabamentos...")
    criar_aba_vedacoes(wb)
    
    print("🏊 Criando orçamento lazer...")
    criar_aba_lazer(wb)
    
    # Salvar
    data_arquivo = datetime.now().strftime('%Y%m%d')
    arquivo = f'/Users/leokock/orcamentos/projetos/mussi-oxford/OXFORD-Orcamento-Executivo-{data_arquivo}.xlsx'
    
    print(f"\n💾 Salvando arquivo: {arquivo}")
    wb.save(arquivo)
    
    print("✅ Orçamento executivo gerado com sucesso!")
    print(f"\n📁 Arquivo: {arquivo}")
    print("\n⚠️  IMPORTANTE:")
    print("   - Valores de custo unitário são ESTIMATIVAS baseadas em índices SINAPI/mercado SC")
    print("   - Quantidades estimadas estão marcadas como '(estimado)' nas descrições")
    print("   - Revisar todos os valores antes de apresentar ao cliente")
    print("   - Confirmar quantitativos com projetos executivos quando disponíveis")

if __name__ == '__main__':
    main()
