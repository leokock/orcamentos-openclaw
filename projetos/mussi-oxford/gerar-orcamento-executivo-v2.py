#!/usr/bin/env python3
"""
Gerador de Orçamento Executivo - Oxford 600 Residence (v2)
Cliente: Mussi Empreendimentos
Data: março/2026

⚠️ VERSÃO ATUALIZADA (13/mar/2026):
SEMPRE faz extração de quantitativos POR PAVIMENTO.
Permite análise de custo por andar, validação de multiplicadores e rastreabilidade.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os

# Configurações
BDI = 0.25  # 25% padrão Cartesian
DATA_BASE = "março/2026"

# Cores padrão
COR_TITULO = "1F4E78"  # Azul escuro
COR_SUBTOTAL = "D9E1F2"  # Azul claro
COR_TOTAL = "4472C4"  # Azul médio
COR_HEADER = "8EA9DB"  # Azul header
COR_PAVIMENTO = "E7E6E6"  # Cinza claro pra separar pavimentos

# Custos unitários SINAPI março/2026 (estimados)
CUSTOS = {
    'concreto': {
        'fck25': 680.00,
        'fck30': 750.00,
        'fck35': 800.00,
        'fck40': 850.00,
        'fck50': 850.00,
    },
    'aco_ca50': 7.50,
    'aco_ca60': 8.00,
    'forma': 75.00,
    'estacas': 120.00,
}

def criar_capa(wb):
    """Cria aba da capa"""
    ws = wb.active
    ws.title = "CAPA"
    
    # Cabeçalho
    ws['B2'] = 'ORÇAMENTO EXECUTIVO'
    ws['B2'].font = Font(size=24, bold=True, color="FFFFFF")
    ws['B2'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('B2:H3')
    
    # Informações do projeto
    linha = 5
    info = [
        ('PROJETO:', 'Oxford 600 Residence'),
        ('CLIENTE:', 'Mussi Empreendimentos'),
        ('ENDEREÇO:', 'Rua Uruguai / Rua Imbituba - Centro, Itajaí/SC'),
        ('DATA-BASE:', DATA_BASE),
        ('RESPONSÁVEL TÉCNICO:', 'Cartesian Engenharia'),
        ('BDI APLICADO:', f'{BDI*100:.1f}%'),
        ('DATA DO ORÇAMENTO:', datetime.now().strftime('%d/%m/%Y')),
        ('VERSÃO:', 'v2 - Com divisão por pavimento'),
    ]
    
    for label, valor in info:
        ws[f'B{linha}'] = label
        ws[f'B{linha}'].font = Font(bold=True)
        ws[f'C{linha}'] = valor
        linha += 1
    
    # Características do empreendimento
    linha += 2
    ws[f'B{linha}'] = 'CARACTERÍSTICAS DO EMPREENDIMENTO'
    ws[f'B{linha}'].font = Font(size=14, bold=True)
    linha += 1
    
    carac = [
        ('Pavimentos:', '27 níveis (Térreo + 4 Garagens + Lazer + 17 Tipos + Ático + 3 Técnicos)'),
        ('Altura:', '71,90 m'),
        ('Unidades Residenciais:', '~85-100 unidades'),
        ('Vagas de Garagem:', '136 vagas'),
        ('Lazer:', '432,88 m² (piscina, fitness, coworking, 9 ambientes)'),
    ]
    
    for label, valor in carac:
        ws[f'B{linha}'] = label
        ws[f'B{linha}'].font = Font(bold=True)
        ws[f'C{linha}'] = valor
        linha += 1
    
    # Ajustar larguras
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80

def criar_header_orcamento(ws, titulo):
    """Cria header padrão para abas de orçamento"""
    ws['A1'] = titulo
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color=COR_TITULO, end_color=COR_TITULO, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:G1')
    
    # Headers de coluna
    headers = ['Item', 'Pavimento', 'Descrição', 'Unidade', 'Quantidade', 'Custo Unit. (R$)', 'Custo Total (R$)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Larguras de coluna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16

def adicionar_linha_item(ws, linha, item, pavimento, descricao, unidade, qtd, custo_unit, nivel=0, highlight=False):
    """Adiciona uma linha de item ao orçamento"""
    ws[f'A{linha}'] = item
    ws[f'B{linha}'] = pavimento
    ws[f'C{linha}'] = descricao
    ws[f'D{linha}'] = unidade
    
    # Quantidade
    cell_qtd = ws[f'E{linha}']
    if qtd is not None:
        cell_qtd.value = qtd
        cell_qtd.number_format = '#,##0.00'
    
    # Custo unitário
    cell_unit = ws[f'F{linha}']
    if custo_unit is not None:
        cell_unit.value = custo_unit
        cell_unit.number_format = 'R$ #,##0.00'
    
    # Custo total
    cell_total = ws[f'G{linha}']
    if qtd is not None and custo_unit is not None:
        cell_total.value = f'=E{linha}*F{linha}'
        cell_total.number_format = 'R$ #,##0.00'
    
    # Indentação para níveis
    if nivel > 0:
        ws[f'C{linha}'].alignment = Alignment(indent=nivel)
    
    # Highlight para cabeçalhos de pavimento
    if highlight:
        for col in 'ABCDEFG':
            ws[f'{col}{linha}'].fill = PatternFill(start_color=COR_PAVIMENTO, end_color=COR_PAVIMENTO, fill_type="solid")
            ws[f'{col}{linha}'].font = Font(bold=True)
    
    # Bordas
    for col in 'ABCDEFG':
        ws[f'{col}{linha}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def adicionar_subtotal(ws, linha, descricao, inicio, fim, nivel=0):
    """Adiciona linha de subtotal"""
    ws.merge_cells(f'A{linha}:F{linha}')
    ws[f'A{linha}'] = descricao
    ws[f'A{linha}'].font = Font(bold=True)
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_SUBTOTAL, end_color=COR_SUBTOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right', indent=nivel)
    
    cell_total = ws[f'G{linha}']
    cell_total.value = f'=SUM(G{inicio}:G{fim})'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True)
    cell_total.fill = PatternFill(start_color=COR_SUBTOTAL, end_color=COR_SUBTOTAL, fill_type="solid")
    
    for col in 'ABCDEFG':
        ws[f'{col}{linha}'].border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

def processar_estrutura_por_pavimento(wb, dados_pavimentos):
    """Cria aba de estrutura organizada por pavimento"""
    ws = wb.create_sheet("1. ESTRUTURA (POR PAV)")
    criar_header_orcamento(ws, "INFRAESTRUTURA + SUPRAESTRUTURA - DIVISÃO POR PAVIMENTO")
    
    linha = 3
    linha_inicio_geral = linha
    
    # Ordenar pavimentos (fundação, térreo, garagens, tipos, ático, técnicos)
    ordem_pavimentos = [
        'Fundação', 'Térreo',
        'Garagem 2', 'Garagem 3', 'Garagem 4', 'Garagem 5',
        'Lazer (6º)',
        '7º Pavimento (Tipo 2)', '8º Pavimento (Tipo 3)', '9º Pavimento (Tipo 4)',
        '10º Pavimento (Tipo 5)', '11º Pavimento (Tipo 6)', '12º Pavimento (Tipo 7)',
        '13º Pavimento (Tipo 8)', '14º Pavimento (Tipo 9)', '15º Pavimento (Tipo 10)',
        '16º Pavimento (Tipo 11)', '17º Pavimento (Tipo 12)', '18º Pavimento (Tipo 13)',
        '19º Pavimento (Tipo 14)', '20º Pavimento (Tipo 15)', '21º Pavimento (Tipo 16)',
        '22º Pavimento (Tipo 17)', '23º Pavimento (Tipo 18)',
        'Ático (24º)', 'Técnico 1 (25º)', 'Técnico 2 (26º)', 'Técnico 3 (27º)'
    ]
    
    for pav_nome in ordem_pavimentos:
        # Buscar dados do pavimento no JSON
        dados_pav = None
        for key in dados_pavimentos.keys():
            if pav_nome.lower() in key.lower() or key.lower() in pav_nome.lower():
                dados_pav = dados_pavimentos[key]
                break
        
        if not dados_pav:
            continue
        
        # Cabeçalho do pavimento
        adicionar_linha_item(ws, linha, '', pav_nome, '', '', None, None, 0, highlight=True)
        linha += 1
        linha_inicio_pav = linha
        
        # Processar dados de forma (concreto + área)
        if 'forma' in dados_pav and dados_pav['forma']:
            forma = dados_pav['forma']
            
            # Concreto (extrair do resumo)
            if 'resumos_encontrados' in forma and forma['resumos_encontrados']:
                resumo = forma['resumos_encontrados'][0]
                tabela = resumo.get('tabela', [])
                
                # Buscar linha de totais
                for row in tabela:
                    if row and len(row) > 0 and str(row[0]).lower() == 'total':
                        # Volume está na última coluna
                        volume_str = row[-1] if len(row) > 0 else None
                        if volume_str:
                            try:
                                volume = float(str(volume_str).replace('.', '').replace(',', '.'))
                                
                                # Determinar fck (simplificado)
                                fck = 'fck30'  # padrão
                                if 'fundação' in pav_nome.lower():
                                    fck = 'fck50'
                                elif 'pilar' in pav_nome.lower():
                                    fck = 'fck40'
                                
                                adicionar_linha_item(
                                    ws, linha, '1.1', pav_nome,
                                    f'Concreto {fck.upper().replace("FCK", "fck ")} MPa',
                                    'm³', volume, CUSTOS['concreto'][fck], 1
                                )
                                linha += 1
                            except:
                                pass
            
            # Aço (de forma ou armação)
            aco_kg = forma.get('aco_kg', 0) or 0
            if aco_kg > 0:
                adicionar_linha_item(
                    ws, linha, '1.2', pav_nome,
                    'Aço CA-50 - Armadura',
                    'kg', aco_kg, CUSTOS['aco_ca50'], 1
                )
                linha += 1
        
        # Processar dados de armação (complementar)
        if 'armacao' in dados_pav and dados_pav['armacao']:
            armacao = dados_pav['armacao']
            aco_armacao = armacao.get('aco_kg', 0) or 0
            
            # Só adicionar se não tiver sido contabilizado em 'forma'
            if aco_armacao > 0 and not (dados_pav.get('forma', {}).get('aco_kg', 0)):
                adicionar_linha_item(
                    ws, linha, '1.2', pav_nome,
                    'Aço CA-50 - Armadura',
                    'kg', aco_armacao, CUSTOS['aco_ca50'], 1
                )
                linha += 1
        
        # Subtotal do pavimento
        if linha > linha_inicio_pav:
            adicionar_subtotal(ws, linha, f'Subtotal {pav_nome}', linha_inicio_pav, linha-1, 1)
            linha += 2
    
    # Total geral
    if linha > linha_inicio_geral + 2:
        linha += 1
        adicionar_subtotal(ws, linha, 'TOTAL ESTRUTURA', linha_inicio_geral, linha-1, 0)

def criar_aba_resumo_consolidado(wb, dados_pavimentos):
    """Cria aba de resumo consolidado (sem divisão por pavimento)"""
    ws = wb.create_sheet("2. RESUMO CONSOLIDADO")
    criar_header_orcamento(ws, "RESUMO CONSOLIDADO - TODOS OS SISTEMAS")
    
    linha = 3
    
    # Calcular totais de estrutura
    total_concreto = 0
    total_aco = 0
    
    for pav_nome, dados_pav in dados_pavimentos.items():
        if 'forma' in dados_pav and dados_pav['forma']:
            forma = dados_pav['forma']
            
            # Somar concreto
            if 'resumos_encontrados' in forma and forma['resumos_encontrados']:
                resumo = forma['resumos_encontrados'][0]
                tabela = resumo.get('tabela', [])
                for row in tabela:
                    if row and len(row) > 0 and str(row[0]).lower() == 'total':
                        volume_str = row[-1] if len(row) > 0 else None
                        if volume_str:
                            try:
                                volume = float(str(volume_str).replace('.', '').replace(',', '.'))
                                total_concreto += volume
                            except:
                                pass
            
            # Somar aço
            aco_kg = forma.get('aco_kg', 0) or 0
            total_aco += aco_kg
        
        if 'armacao' in dados_pav and dados_pav['armacao']:
            armacao = dados_pav['armacao']
            aco_armacao = armacao.get('aco_kg', 0) or 0
            if not (dados_pav.get('forma', {}).get('aco_kg', 0)):
                total_aco += aco_armacao
    
    # 1. ESTRUTURA
    adicionar_linha_item(ws, linha, '1', '', 'ESTRUTURA', '', None, None, 0, highlight=True)
    linha += 1
    inicio_estrutura = linha
    
    adicionar_linha_item(ws, linha, '1.1', '', 'Concreto Total (todos os fck)', 'm³', 
                         total_concreto, 750.00, 1)  # média
    linha += 1
    
    adicionar_linha_item(ws, linha, '1.2', '', 'Aço CA-50 Total', 'kg', 
                         total_aco, CUSTOS['aco_ca50'], 1)
    linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ESTRUTURA', inicio_estrutura, linha-1)
    linha += 2
    
    # 2. HIDROSSANITÁRIO (dados consolidados)
    adicionar_linha_item(ws, linha, '2', '', 'HIDROSSANITÁRIO', '', None, None, 0, highlight=True)
    linha += 1
    inicio_hidro = linha
    
    # Água fria
    adicionar_linha_item(ws, linha, '2.1', '', 'Tubulações Água Fria PVC/PPR', 'm', 3600, 45.00, 1)
    linha += 1
    
    # Água quente
    adicionar_linha_item(ws, linha, '2.2', '', 'Tubulações Água Quente PPR', 'm', 1800, 85.00, 1)
    linha += 1
    
    # Esgoto
    adicionar_linha_item(ws, linha, '2.3', '', 'Tubulações Esgoto PVC', 'm', 3000, 38.00, 1)
    linha += 1
    
    # Águas pluviais
    adicionar_linha_item(ws, linha, '2.4', '', 'Tubulações Águas Pluviais PVC', 'm', 800, 42.00, 1)
    linha += 1
    
    # Louças e metais
    adicionar_linha_item(ws, linha, '2.5', '', 'Louças e Metais Sanitários', 'cj', 85, 2500.00, 1)
    linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL HIDROSSANITÁRIO', inicio_hidro, linha-1)
    linha += 2
    
    # 3. ELÉTRICO
    adicionar_linha_item(ws, linha, '3', '', 'ELÉTRICO', '', None, None, 0, highlight=True)
    linha += 1
    inicio_eletrico = linha
    
    adicionar_linha_item(ws, linha, '3.1', '', 'Quadros de Distribuição', 'un', 110, 1200.00, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '3.2', '', 'Disjuntores e DRs', 'un', 2820, 85.00, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '3.3', '', 'Eletrodutos e Caixas', 'm', 8500, 22.00, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '3.4', '', 'Cabos e Condutores', 'm', 25000, 8.50, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '3.5', '', 'Transformador 500 kVA', 'un', 1, 85000.00, 1)
    linha += 1
    
    adicionar_linha_item(ws, linha, '3.6', '', 'Gerador 200 kVA', 'un', 1, 120000.00, 1)
    linha += 1
    
    adicionar_subtotal(ws, linha, 'SUBTOTAL ELÉTRICO', inicio_eletrico, linha-1)
    linha += 2
    
    # TOTAL GERAL
    linha += 1
    ws.merge_cells(f'A{linha}:F{linha}')
    ws[f'A{linha}'] = 'TOTAL GERAL (SEM BDI)'
    ws[f'A{linha}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{linha}'].fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")
    ws[f'A{linha}'].alignment = Alignment(horizontal='right')
    
    cell_total = ws[f'G{linha}']
    cell_total.value = f'=G{inicio_estrutura-1}+G{inicio_hidro-1}+G{inicio_eletrico-1}'
    cell_total.number_format = 'R$ #,##0.00'
    cell_total.font = Font(bold=True, size=14, color="FFFFFF")
    cell_total.fill = PatternFill(start_color=COR_TOTAL, end_color=COR_TOTAL, fill_type="solid")

def main():
    """Função principal"""
    print("🏗️  Gerando Orçamento Executivo Oxford - v2 (com divisão por pavimento)")
    
    # Carregar dados de quantitativos por pavimento
    json_path = 'quantitativos_pavimentos.json'
    if not os.path.exists(json_path):
        print(f"❌ Arquivo {json_path} não encontrado!")
        return
    
    with open(json_path, 'r', encoding='utf-8') as f:
        dados_pavimentos = json.load(f)
    
    print(f"✅ Carregados dados de {len(dados_pavimentos)} pavimentos")
    
    # Criar workbook
    wb = Workbook()
    
    # Capa
    criar_capa(wb)
    print("✅ Capa criada")
    
    # Estrutura por pavimento
    processar_estrutura_por_pavimento(wb, dados_pavimentos)
    print("✅ Estrutura por pavimento processada")
    
    # Resumo consolidado
    criar_aba_resumo_consolidado(wb, dados_pavimentos)
    print("✅ Resumo consolidado criado")
    
    # Salvar
    output_file = f'OXFORD-Orcamento-Executivo-v2-{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(output_file)
    
    print(f"\n✅ CONCLUÍDO!")
    print(f"📄 Arquivo gerado: {output_file}")
    print(f"📊 Estrutura organizada por pavimento")
    print(f"🎯 Permite análise de custo por andar e validação de multiplicadores")

if __name__ == '__main__':
    main()
